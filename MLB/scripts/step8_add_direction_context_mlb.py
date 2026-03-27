#!/usr/bin/env python3
"""
step8_add_direction_context_mlb.py  (MLB Pipeline)

Mirrors NBA step8. Reads step7_mlb_ranked.xlsx, adds direction,
outputs CSV + clean formatted XLSX with pitcher/hitter split tabs.

Run:
  py -3.14 step8_add_direction_context_mlb.py \
    --input step7_mlb_ranked.xlsx \
    --output step8_mlb_direction.csv
"""

from __future__ import annotations

import argparse
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


def _norm_pick_type(x: str) -> str:
    t = str(x or "").strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
PITCHER_TAB_COLOR = "8B0000"
HITTER_TAB_COLOR  = "003366"
HEADER_COLOR      = "1C1C1C"
DIR_OVER          = "C8F7C5"
DIR_UNDER         = "F7C5C5"

PITCHER_PROPS = {
    "strikeouts", "pitching_outs", "innings_pitched",
    "hits_allowed", "earned_runs", "walks_allowed", "batters_faced",
}


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name: str, data: pd.DataFrame, tab_color: str = HEADER_COLOR) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    headers = list(data.columns)

    # Keep workbook valid even when slate is empty.
    if not headers:
        ws.cell(row=1, column=1, value="No columns")
        return

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            col_name    = headers[ci - 1]
            display_val = None if pd.isna(val) else val
            cell        = ws.cell(row=ri, column=ci, value=display_val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

            if col_name == "Tier":
                tier_bg, tier_fg = TIER_COLORS.get(str(display_val), ("333333", "FFFFFF"))
                cell.fill = PatternFill("solid", start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name="Arial", size=9)
            elif col_name == "Direction":
                bg = DIR_OVER if display_val == "OVER" else DIR_UNDER if display_val == "UNDER" else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, name="Arial", size=9)
            elif col_name == "Player Type":
                bg = "FFE8E8" if str(display_val).lower() == "pitcher" else "E8F4FF"
                cell.fill = PatternFill("solid", start_color=bg)
            else:
                cell.fill = PatternFill("solid", start_color="F9F9F9" if ri % 2 == 0 else "FFFFFF")

    col_widths = {
        "Tier": 6, "Rank Score": 10, "Player": 22, "Pos": 6, "Player Type": 10,
        "Team": 10, "Opp": 10, "Game Time": 10,
        "Prop": 20, "Pick Type": 10, "Line": 7,
        "Direction": 9, "Edge": 7, "Projection": 10,
        "Hit Rate (5g)": 12, "Last 5 Avg": 10, "Season Avg": 10,
        "L5 Over": 8, "L5 Under": 8,
        "Def Rank": 9, "Def Tier": 10,
        "Min Tier": 9, "Bat Order": 10, "Pitcher Role": 12,
        "Void Reason": 20,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    df2 = df2.where(pd.notna(df2), None)
    df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%-I:%M %p")

    keep = [
        "tier", "rank_score",
        "player", "pos", "player_type_norm", "team", "opp_team", "game_time",
        "prop_type", "pick_type", "line",
        "final_bet_direction",
        "edge", "projection",
        "line_hit_rate_over_ou_5",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "OVERALL_DEF_RANK", "DEF_TIER",
        "minutes_tier", "batting_order_tier", "pitcher_role",
        "void_reason",
    ]
    keep  = [c for c in keep if c in df2.columns]
    clean = df2[keep].copy()

    for col in ["rank_score", "edge", "projection", "line_hit_rate_over_ou_5"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(2)
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "player": "Player", "pos": "Pos", "player_type_norm": "Player Type",
        "team": "Team", "opp_team": "Opp", "game_time": "Game Time",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "final_bet_direction": "Direction",
        "edge": "Edge", "projection": "Projection",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier",
        "minutes_tier": "Min Tier", "batting_order_tier": "Bat Order",
        "pitcher_role": "Pitcher Role",
        "void_reason": "Void Reason",
    }
    clean = clean.rename(columns=rename)
    clean = clean.where(pd.notna(clean), None)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "ALL", clean, HEADER_COLOR)

    for tier in ["A", "B", "C", "D"]:
        subset = clean[clean["Tier"] == tier].copy()
        if len(subset):
            tier_bg = TIER_COLORS.get(tier, ("333333",))[0]
            write_sheet(wb, f"Tier {tier}", subset, tier_bg)

    # Pitcher / Hitter split tabs
    if "Player Type" in clean.columns:
        pitchers = clean[clean["Player Type"].astype(str).str.lower() == "pitcher"].copy()
        hitters  = clean[clean["Player Type"].astype(str).str.lower() == "hitter"].copy()
        if len(pitchers): write_sheet(wb, "Pitchers", pitchers, PITCHER_TAB_COLOR)
        if len(hitters):  write_sheet(wb, "Hitters",  hitters,  HITTER_TAB_COLOR)

    wb.save(xlsx_path)
    print(f"📊 Clean XLSX saved → {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="MLB/scripts/step7_mlb_ranked.xlsx")
    ap.add_argument("--sheet",  default="ALL")
    ap.add_argument("--output", default="MLB/scripts/step8_mlb_direction.csv")
    ap.add_argument("--xlsx",   default="MLB/scripts/step8_mlb_direction_clean.xlsx")
    args = ap.parse_args()

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    df  = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")
    if df.empty:
        raise SystemExit("❌ [PropOracle-MLB-S8] Empty input from step7; aborting.")
    out = df.copy()

    if "edge" not in out.columns:
        proj = pd.to_numeric(out.get("projection", ""), errors="coerce")
        line = pd.to_numeric(out.get("line",        ""), errors="coerce")
        out["edge"] = proj - line

    edge     = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = edge.abs()

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    forced    = pick_type.isin(["Goblin", "Demon"])

    model_dir  = np.where(edge >= 0, "OVER", "UNDER")
    final_dir  = model_dir.copy()
    reason     = np.where(abs_edge < 0.03, "MODEL_TIEBREAK_DIFF<0.03", "")
    final_dir  = np.where(forced, "OVER", final_dir)
    reason     = np.where(forced, "FORCED_OVER_ONLY_GOB_DEM", reason)

    out["model_dir"]           = model_dir
    out["final_bet_direction"] = final_dir
    out["final_dir_reason"]    = reason

    if out.empty:
        raise SystemExit("❌ [PropOracle-MLB-S8] Empty output after direction step; aborting.")

    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"✅ Saved → {args.output}")
    print("final_bet_direction:", pd.Series(final_dir).value_counts().to_dict())
    if "tier" in out.columns:
        print("tier:", out["tier"].value_counts().to_dict())

    xlsx_path = args.xlsx if args.xlsx else args.output.replace(".csv", "_clean.xlsx")
    build_clean_xlsx(out, xlsx_path)


if __name__ == "__main__":
    main()
