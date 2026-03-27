#!/usr/bin/env python3
"""
step9_build_tickets_mlb.py  (MLB Pipeline)

Mirrors NBA/Soccer step9. Reads step8_mlb_direction_clean.xlsx (Tier A)
and builds optimized PrizePicks tickets.

MLB note: min_hit_rate default is 0.65 (between NBA's 0.8 and soccer's 0.6).
Pitcher Ks are the most reliable MLB prop — pipeline will surface these first.

Run:
  py -3.14 step9_build_tickets_mlb.py \
    --input step8_mlb_direction_clean.xlsx \
    --output mlb_best_tickets.xlsx
"""

from __future__ import annotations

import argparse
from itertools import combinations
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Payout tables (identical to NBA/Soccer) ───────────────────────────────────
POWER_PLAY_BASE = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 37.5}
FLEX_PLAY_BASE  = {
    2: {2: 3.0},
    3: {3: 3.0,  2: 0.5},
    4: {4: 6.0,  3: 1.5},
    5: {5: 10.0, 4: 2.0,  3: 0.4},
    6: {6: 25.0, 5: 2.0,  4: 0.4},
}
GOBLIN_POWER_MOD = {1: 0.900, 2: 0.800, 3: 0.767, 4: 0.667}
GOBLIN_FLEX_MOD  = {1: 0.833, 2: 0.750, 3: 0.700, 4: 0.600}
DEMON_POWER_MOD  = {1: 1.333, 2: 1.833, 3: 3.833}
DEMON_FLEX_MOD   = {1: 1.200, 2: 1.600, 3: 3.000}

COLORS = {
    "Standard": "2874A6", "Goblin": "6C3483",
    "Demon": "C0392B",    "Best Mix": "1E8449",
}
HDR_COLOR = "1C1C1C"
DIR_OVER  = "C8F7C5"
DIR_UNDER = "F7C5C5"
TIER_COLORS = {
    "Elite":     "D5F5E3",
    "Above Avg": "EBF5FB",
    "Avg":       "FDFEFE",
    "Weak":      "FDEDEC",
}


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _norm_pick_type(x) -> str:
    t = str(x).strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


def calc_ticket_payout(ticket: list, n_legs: int, play_type: str = "power") -> dict:
    base_top  = POWER_PLAY_BASE.get(n_legs, 37.5) if play_type == "power" else FLEX_PLAY_BASE.get(n_legs, {}).get(n_legs, 25.0)
    power_mod = flex_mod = 1.0
    for row in ticket:
        pt  = str(row.get("Pick Type", "Standard")).strip().lower()
        dev = int(row.get("deviation_level", row.get("Deviation Level", 1)) or 1)
        dev = max(1, min(dev, 4))
        if "gob" in pt:
            power_mod *= GOBLIN_POWER_MOD.get(dev, 0.840)
            flex_mod  *= GOBLIN_FLEX_MOD.get(dev,  0.800)
        elif "dem" in pt:
            power_mod *= DEMON_POWER_MOD.get(dev, 1.627)
            flex_mod  *= DEMON_FLEX_MOD.get(dev,  1.600)
    if play_type == "power":
        top = round(base_top * power_mod, 2)
        return {"top_payout": top, "stake_to_win_100": round(100/top, 2) if top else 0, "play_type": "Power Play"}
    else:
        flex_base = FLEX_PLAY_BASE.get(n_legs, {})
        partials  = {nc: round(mult * flex_mod if nc == n_legs else mult, 2) for nc, mult in flex_base.items()}
        top = partials.get(n_legs, 25.0)
        return {"top_payout": top, "stake_to_win_100": round(100/top, 2) if top else 0, "play_type": "Flex Play", "partials": partials}


def _game_key(row) -> frozenset:
    return frozenset([str(row["Team"]), str(row["Opp"])])


def _ticket_valid(combo) -> bool:
    players = [r["Player"] for r in combo]
    if len(players) != len(set(players)): return False
    games = [_game_key(r) for r in combo]
    return len(games) == len(set(games))


def _ticket_score(combo) -> float:
    scores = []
    for r in combo:
        v = pd.to_numeric(pd.Series([r["Rank Score"]]), errors="coerce").iloc[0]
        if not pd.isna(v): scores.append(float(v))
    return float(np.mean(scores)) if scores else 0.0


def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets: int,
                  used_legs: set, min_hit_rate: float = 0.0) -> list:
    if min_hit_rate > 0.0:
        pool = pool[pd.to_numeric(pool["Hit Rate (5g)"], errors="coerce") >= min_hit_rate].copy()
    pool       = pool.sort_values("Rank Score", ascending=False).copy()
    K          = min(len(pool), 25)
    candidates = [row for _, row in pool.head(K).iterrows()]
    valid:list = []
    seen: set  = set()

    for combo in combinations(candidates, n_legs):
        leg_keys = [(r["Player"], r["Prop"], r["Line"]) for r in combo]
        if any(lk in used_legs for lk in leg_keys): continue
        if not _ticket_valid(combo): continue
        tkey = frozenset((r["Player"], r["Prop"], str(r["Line"])) for r in combo)
        if tkey in seen: continue
        seen.add(tkey)
        valid.append((combo, _ticket_score(combo)))

    valid.sort(key=lambda x: x[1], reverse=True)
    tickets = []
    for combo, _ in valid:
        if len(tickets) >= max_tickets: break
        leg_keys = [(r["Player"], r["Prop"], r["Line"]) for r in combo]
        if any(lk in used_legs for lk in leg_keys): continue
        used_legs.update(leg_keys)
        tickets.append(list(combo))
    return tickets


def write_tickets_sheet(wb, sheet_name: str, tickets: list, n_legs: int, pick_label: str) -> None:
    if not tickets: return
    ws     = wb.create_sheet(sheet_name)
    tab_bg = COLORS.get(pick_label, "1E8449")
    cols   = ["#", "Player", "Team", "Opp", "Player Type", "Prop", "Pick Type", "Line",
              "Direction", "Edge", "Hit Rate (5g)", "Last 5 Avg", "Season Avg",
              "L5 Over", "L5 Under", "Rank Score", "Def Tier"]
    current_row = 1

    for t_idx, ticket in enumerate(tickets, 1):
        ticket_dicts = [dict(r) for r in ticket]
        pp_info      = calc_ticket_payout(ticket_dicts, n_legs, "power")
        flex_info    = calc_ticket_payout(ticket_dicts, n_legs, "flex")
        avg_score    = _ticket_score(ticket)
        avg_hr_vals  = [float(r["Hit Rate (5g)"]) for r in ticket
                        if not pd.isna(pd.to_numeric(pd.Series([r["Hit Rate (5g)"]]), errors="coerce").iloc[0])]
        avg_hr       = float(np.mean(avg_hr_vals)) if avg_hr_vals else 0.0

        header_val = (
            f"Ticket #{t_idx} | {n_legs}-Leg {pick_label} | "
            f"Power: {pp_info['top_payout']}x (${pp_info['stake_to_win_100']:.0f} to win $100) | "
            f"Flex: {flex_info['top_payout']}x | "
            f"Avg Hit Rate: {avg_hr:.0%} | Est Win: {avg_hr**n_legs:.0%} | Score: {avg_score:.2f}"
        )
        c = ws.cell(row=current_row, column=1, value=header_val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=tab_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(cols))
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=HDR_COLOR)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = thin_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for leg_i, row in enumerate(ticket, 1):
            row_bg = "F8F9FA" if leg_i % 2 == 0 else "FFFFFF"
            def_bg = TIER_COLORS.get(str(row.get("Def Tier", "")), "FFFFFF")
            l5o = row.get("L5 Over", ""); l5u = row.get("L5 Under", "")
            l5o = "" if pd.isna(l5o) else int(l5o)
            l5u = "" if pd.isna(l5u) else int(l5u)

            vals = [
                leg_i, row["Player"], row["Team"], row["Opp"],
                row.get("Player Type", ""), row["Prop"], row["Pick Type"], row["Line"],
                row["Direction"], round(float(row["Edge"]), 1),
                row["Hit Rate (5g)"],
                round(float(row["Last 5 Avg"]), 1), round(float(row["Season Avg"]), 1),
                l5o, l5u, round(float(row["Rank Score"]), 2), row.get("Def Tier", ""),
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()
                col_name = cols[ci - 1]
                if col_name == "Direction":
                    c.fill = PatternFill("solid", start_color=DIR_OVER if val == "OVER" else DIR_UNDER)
                    c.font = Font(bold=True, name="Arial", size=9)
                elif col_name == "Player Type":
                    bg = "FFE8E8" if str(val).lower() == "pitcher" else "E8F4FF"
                    c.fill = PatternFill("solid", start_color=bg)
                elif col_name == "Def Tier":
                    c.fill = PatternFill("solid", start_color=def_bg)
                else:
                    c.fill = PatternFill("solid", start_color=row_bg)
            current_row += 1
        current_row += 1


def write_summary_sheet(wb, all_ticket_sets: list) -> None:
    ws      = wb.create_sheet("SUMMARY", 0)
    headers = ["Sheet", "Ticket #", "Legs", "Pick Type", "Power Payout",
               "$ to win $100", "Flex Payout", "Avg Hit Rate", "Est Win %",
               "Avg Rank Score", "Players"]
    col_w   = [25, 10, 6, 12, 13, 13, 12, 13, 12, 15, 65]

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill = PatternFill("solid", start_color=HDR_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for sheet_name, tickets, n_legs, pick_label in all_ticket_sets:
        for t_idx, ticket in enumerate(tickets, 1):
            avg_score   = sum(float(r["Rank Score"]) for r in ticket) / len(ticket)
            avg_hr_vals = [float(r["Hit Rate (5g)"]) for r in ticket
                           if not pd.isna(pd.to_numeric(pd.Series([r["Hit Rate (5g)"]]), errors="coerce").iloc[0])]
            avg_hr       = float(np.mean(avg_hr_vals)) if avg_hr_vals else 0.0
            ticket_dicts = [dict(r) for r in ticket]
            pp_info      = calc_ticket_payout(ticket_dicts, n_legs, "power")
            flex_info    = calc_ticket_payout(ticket_dicts, n_legs, "flex")
            players      = " | ".join(f"{r['Player']} {r['Direction']} {r['Prop']} {r['Line']}" for r in ticket)
            vals = [sheet_name, t_idx, n_legs, pick_label,
                    f"{pp_info['top_payout']}x", f"${pp_info['stake_to_win_100']:.0f}",
                    f"{flex_info['top_payout']}x",
                    f"{avg_hr:.0%}", f"{avg_hr**n_legs:.0%}",
                    round(avg_score, 2), players]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border = thin_border()
                c.fill = PatternFill("solid", start_color="F8F9FA" if row % 2 == 0 else "FFFFFF")
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",        default="MLB/scripts/step8_mlb_direction_clean.xlsx")
    ap.add_argument("--sheet",        default="Tier A")
    ap.add_argument("--output",       default="MLB/scripts/mlb_best_tickets.xlsx")
    ap.add_argument("--min_hit_rate", type=float, default=0.65)
    ap.add_argument("--max_tickets",  type=int,   default=10)
    ap.add_argument("--legs",         default="2,3,4")
    args = ap.parse_args()

    leg_counts = [int(x.strip()) for x in args.legs.split(",")]

    tier_priority = ["Tier A", "Tier B", "Tier C", "ALL"]
    requested_sheet = str(args.sheet).strip()
    start_idx = tier_priority.index(requested_sheet) if requested_sheet in tier_priority else 0

    print(f"→ Loading: {args.input} (requested sheet={requested_sheet})")
    df = None
    used_sheet = None
    last_err: Exception | None = None
    for candidate in tier_priority[start_idx:]:
        try:
            df = pd.read_excel(args.input, sheet_name=candidate)
            used_sheet = candidate
            if candidate != requested_sheet:
                print(f"⚠️ WARNING: Requested sheet '{requested_sheet}' not found; falling back to '{candidate}'")
            break
        except ValueError as e:
            last_err = e
            continue

    if df is None or used_sheet is None:
        raise last_err or RuntimeError("Failed to load any tier sheets")

    # Make fallback obvious from filename itself.
    if used_sheet != requested_sheet:
        out_path = Path(args.output)
        used_sanitized = str(used_sheet).replace(" ", "_")
        args.output = str(out_path.with_name(out_path.stem + f"_fallback_{used_sanitized}" + out_path.suffix))
        print(f"  Output renamed to: {args.output}")
    df["Pick Type"] = df["Pick Type"].astype(str).apply(_norm_pick_type)

    df = df[pd.to_numeric(df["Hit Rate (5g)"], errors="coerce") >= args.min_hit_rate].copy()
    print(f"  Props with hit rate >= {args.min_hit_rate}: {len(df)}")

    df = df.sort_values("Rank Score", ascending=False).drop_duplicates(
        subset=["Player", "Prop", "Line", "Direction"]
    )

    std_pool = df[df["Pick Type"] == "Standard"].drop_duplicates(subset=["Player"])
    gob_pool = df[df["Pick Type"] == "Goblin"].drop_duplicates(subset=["Player"])
    all_pool = df.sort_values("Rank Score", ascending=False).drop_duplicates(subset=["Player"])

    pools = [("Standard", std_pool), ("Goblin", gob_pool), ("Best Mix", all_pool)]

    wb = Workbook()
    wb.remove(wb.active)
    all_ticket_sets = []

    for pick_label, pool in pools:
        if len(pool) < 2:
            print(f"  ⚠ Skipping {pick_label} — not enough ({len(pool)})")
            continue
        used_legs: set = set()
        for n_legs in leg_counts:
            if len(pool) < n_legs: continue
            tickets    = build_tickets(pool.copy(), n_legs, args.max_tickets, used_legs, args.min_hit_rate)
            sheet_name = f"{pick_label} {n_legs}-Leg"
            write_tickets_sheet(wb, sheet_name, tickets, n_legs, pick_label)
            all_ticket_sets.append((sheet_name, tickets, n_legs, pick_label))
            print(f"  {sheet_name}: {len(tickets)} tickets")

    if len(std_pool) >= 1 and len(gob_pool) >= 1:
        fill_pool = pd.concat([
            std_pool.assign(_pref=0), gob_pool.assign(_pref=1)
        ]).sort_values(["_pref", "Rank Score"], ascending=[True, False]).drop("_pref", axis=1)
        used_legs_fill: set = set()
        for n_legs in leg_counts:
            if n_legs < 3 or len(fill_pool) < n_legs: continue
            tickets = build_tickets(fill_pool.copy(), n_legs, args.max_tickets, used_legs_fill, args.min_hit_rate)
            if tickets:
                sheet_name = f"Std+Gob Fill {n_legs}-Leg"
                write_tickets_sheet(wb, sheet_name, tickets, n_legs, "Best Mix")
                all_ticket_sets.append((sheet_name, tickets, n_legs, "Best Mix"))
                print(f"  {sheet_name}: {len(tickets)} tickets")

    write_summary_sheet(wb, all_ticket_sets)
    wb.save(args.output)
    print(f"\n✅ Saved → {args.output}")
    print(f"\nPrizePicks Power Play payout reference:")
    print(f"  {'Legs':<6} {'Power':>8} {'Flex':>8}  {'$ to win $100':>15}")
    for n in [2, 3, 4, 5, 6]:
        pp    = POWER_PLAY_BASE.get(n, "-")
        fl    = FLEX_PLAY_BASE.get(n, {}).get(n, "-")
        stake = round(100/pp, 2) if isinstance(pp, (int, float)) else "-"
        print(f"  {n}-leg  {str(pp)+'x':>8} {str(fl)+'x':>8}  ${stake:>6}")


if __name__ == "__main__":
    main()
