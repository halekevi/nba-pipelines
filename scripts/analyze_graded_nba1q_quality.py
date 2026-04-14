#!/usr/bin/env python3
"""Aggregate graded NBA 1Q workbooks: void survival, hit rates, and line shapes.

Scans ``outputs/**`` and ``ui_runner/graded_slate/**`` for ``graded_nba1q*.xlsx``,
reads the ``Box Raw`` sheet, and writes CSVs for:

- Per-slate file summary (void %, Demon/Tier-B void share, decided counts)
- ``prop × bet_direction × slate_date`` with n, void rate, hit rate (among decided),
  and line quantiles (among all rows with numeric lines)

Hit rate is **only** among rows with result in ``HIT``/``MISS`` (excludes VOID/PUSH
from the denominator). Void rate is ``VOID / all rows`` so you can spot
survivorship vs grading coverage.

Example::

    python scripts/analyze_graded_nba1q_quality.py --out-dir data/reports/nba1q_graded
"""
from __future__ import annotations

import argparse
import re
from pathlib import Path

import numpy as np
import pandas as pd

DEMON_PICK_TYPE = "demon"


def _repo_root() -> Path:
    return Path(__file__).resolve().parent.parent


def _slate_date_from_path(p: Path) -> str | None:
    for part in p.parts:
        m = re.fullmatch(r"(\d{4}-\d{2}-\d{2})", part)
        if m:
            return m.group(1)
    m = re.search(r"(20\d{2}-\d{2}-\d{2})", p.name)
    return m.group(1) if m else None


def _stem_base(stem: str) -> str:
    s = stem.lower()
    return s[: -len("_mlbackfill")] if s.endswith("_mlbackfill") else s


def _count_box_raw_rows(path: Path) -> int:
    """Number of data rows on ``Box Raw`` (0 if missing sheet or unreadable)."""
    try:
        from openpyxl import load_workbook

        wb = load_workbook(path, read_only=True, data_only=True)
        try:
            if "Box Raw" not in wb.sheetnames:
                return 0
            ws = wb["Box Raw"]
            return sum(1 for _ in ws.iter_rows(min_row=2))
        finally:
            wb.close()
    except Exception:
        return 0


def _path_source_rank(p: Path) -> int:
    """When Box Raw row counts tie, prefer outputs/<date>/ over ui_runner/graded_slate copies."""
    parts_l = {x.lower() for x in p.parts}
    score = 0
    if "outputs" in parts_l:
        score += 2
    if "graded_slate" in parts_l:
        score -= 1
    return score


def _pick_best_path_by_box_raw(paths: list[Path]) -> Path | None:
    if not paths:
        return None
    return max(
        paths,
        key=lambda p: (_count_box_raw_rows(p), _path_source_rank(p), str(p.resolve())),
    )


def _pick_preferred(canonical: Path | None, mlbackfill: Path | None) -> Path | None:
    """Prefer non-empty canonical; fall back to mlbackfill only when canonical is empty/missing."""
    if canonical is None:
        return mlbackfill
    n_canonical = _count_box_raw_rows(canonical)
    if n_canonical == 0 and mlbackfill is not None:
        return mlbackfill
    return canonical


def discover_nba1q_workbooks(roots: list[Path]) -> list[Path]:
    """One workbook per slate stem: prefer the file with more Box Raw rows over name-based mlbackfill bias."""
    raw: list[Path] = []
    seen: set[str] = set()
    for root in roots:
        if not root.exists():
            continue
        for p in sorted(root.rglob("graded_nba1q*.xlsx")):
            if "combined_tickets_graded" in p.name.lower():
                continue
            key = str(p.resolve())
            if key in seen:
                continue
            seen.add(key)
            raw.append(p)
    by_stem: dict[str, list[Path]] = {}
    for p in raw:
        stem_key = _stem_base(p.stem)
        by_stem.setdefault(stem_key, []).append(p)

    best: dict[str, Path] = {}
    for stem_key, plist in by_stem.items():
        canon_list = [p for p in plist if not p.stem.lower().endswith("_mlbackfill")]
        ml_list = [p for p in plist if p.stem.lower().endswith("_mlbackfill")]
        canonical = _pick_best_path_by_box_raw(canon_list)
        mlbackfill = _pick_best_path_by_box_raw(ml_list)
        chosen = _pick_preferred(canonical, mlbackfill)
        if chosen is not None:
            best[stem_key] = chosen
    return sorted(best.values(), key=lambda x: str(x))


def _norm_result(s: pd.Series) -> pd.Series:
    return s.astype(str).str.strip().str.upper()


def _direction_series(df: pd.DataFrame) -> pd.Series:
    for c in ("bet_direction", "direction", "final_bet_direction"):
        if c in df.columns:
            d = df[c].astype(str).str.strip().str.upper()
            d = d.mask(d.isin(["", "NAN", "NONE", "NAT"]))
            return d
    return pd.Series(np.nan, index=df.index)


def _prop_series(df: pd.DataFrame) -> pd.Series:
    for c in ("prop_type_norm", "prop_type", "prop"):
        if c in df.columns:
            p = df[c].astype(str).str.strip()
            p = p.mask(p.str.lower().isin(["nan", "none", ""]))
            return p
    return pd.Series("(missing)", index=df.index)


def load_box_raw(path: Path) -> tuple[pd.DataFrame | None, str | None]:
    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        return None, f"open_error: {e}"
    if "Box Raw" not in xl.sheet_names:
        return None, "missing_Box_Raw"
    df = pd.read_excel(path, sheet_name="Box Raw")
    if df is None or len(df) == 0:
        return None, "empty"
    return df, None


def summarize_file(path: Path, df: pd.DataFrame, slate_date: str | None) -> dict:
    ru = _norm_result(df["result"]) if "result" in df.columns else pd.Series([], dtype=object)
    n = len(df)
    void_n = int(ru.eq("VOID").sum())
    push_n = int(ru.eq("PUSH").sum())
    pend_n = int(ru.eq("PENDING").sum())
    hit_n = int(ru.eq("HIT").sum())
    miss_n = int(ru.eq("MISS").sum())
    decided = hit_n + miss_n
    tier_l = df["tier"].astype(str).str.strip().str.upper() if "tier" in df.columns else pd.Series([""] * n)
    b_mask = tier_l.eq("B")
    if "pick_type" in df.columns:
        pt_l = df["pick_type"].astype(str).str.strip().str.lower()
        demon_mask = pt_l.eq(DEMON_PICK_TYPE)
    else:
        demon_mask = tier_l.eq("DEMON")
    non_demon_mask = ~demon_mask
    demon_rows = int(demon_mask.sum())
    demon_void = int((demon_mask & ru.eq("VOID")).sum()) if demon_rows else 0
    non_demon_rows = int(non_demon_mask.sum())
    non_demon_void = int((non_demon_mask & ru.eq("VOID")).sum()) if non_demon_rows else 0
    non_demon_decided = int((non_demon_mask & ru.isin(["HIT", "MISS"])).sum()) if non_demon_rows else 0
    tier_b_rows = int(b_mask.sum())
    tier_b_void = int((b_mask & ru.eq("VOID")).sum()) if tier_b_rows else 0
    return {
        "source_file": path.name,
        "source_path": str(path.resolve()),
        "slate_date": slate_date or "",
        "n_rows": n,
        "n_void": void_n,
        "n_push": push_n,
        "n_pending": pend_n,
        "n_hit": hit_n,
        "n_miss": miss_n,
        "n_decided": decided,
        "void_rate": void_n / n if n else np.nan,
        "n_decided_ex_demon": non_demon_decided,
        "void_rate_ex_demon": non_demon_void / non_demon_rows if non_demon_rows else np.nan,
        "demon_rows": demon_rows,
        "demon_void_rows": demon_void,
        "demon_void_rate": demon_void / demon_rows if demon_rows else np.nan,
        "tier_b_rows": tier_b_rows,
        "tier_b_void_rows": tier_b_void,
        "tier_b_void_rate": tier_b_void / tier_b_rows if tier_b_rows else np.nan,
    }


def enrich_rows(df: pd.DataFrame, slate_date: str | None, source_file: str) -> pd.DataFrame:
    out = df.copy()
    out["_slate_date"] = slate_date or ""
    out["_source_file"] = source_file
    out["result_u"] = _norm_result(out["result"]) if "result" in out.columns else ""
    out["_prop"] = _prop_series(out)
    out["_dir"] = _direction_series(out)
    out["_line"] = pd.to_numeric(out["line"], errors="coerce") if "line" in out.columns else np.nan
    out["_is_decided"] = out["result_u"].isin(["HIT", "MISS"])
    out["_is_void"] = out["result_u"].eq("VOID")
    out["_is_hit"] = out["result_u"].eq("HIT")
    if "pick_type" in out.columns:
        out["_is_demon"] = out["pick_type"].astype(str).str.strip().str.lower().eq(DEMON_PICK_TYPE)
    else:
        out["_is_demon"] = False
    out["_is_void_ex_demon"] = out["_is_void"] & (~out["_is_demon"])
    out["_is_decided_ex_demon"] = out["_is_decided"] & (~out["_is_demon"])
    return out


def aggregate_prop_dir_date(all_rows: pd.DataFrame) -> pd.DataFrame:
    if all_rows.empty:
        return all_rows
    g = all_rows.groupby(["_slate_date", "_prop", "_dir"], dropna=False)
    agg = g.agg(
        n_rows=("_is_void", "size"),
        n_void=("_is_void", "sum"),
        n_rows_ex_demon=("_is_demon", lambda s: int((~s).sum())),
        n_void_ex_demon=("_is_void_ex_demon", "sum"),
        n_decided=("_is_decided", "sum"),
        n_decided_ex_demon=("_is_decided_ex_demon", "sum"),
        n_hit=("_is_hit", "sum"),
        line_min=("_line", "min"),
        line_q25=("_line", lambda s: float(s.quantile(0.25)) if s.notna().any() else np.nan),
        line_median=("_line", "median"),
        line_q75=("_line", lambda s: float(s.quantile(0.75)) if s.notna().any() else np.nan),
        line_max=("_line", "max"),
        line_mean=("_line", "mean"),
        n_line_present=("_line", lambda s: int(s.notna().sum())),
    ).reset_index()
    agg.rename(columns={"_slate_date": "slate_date", "_prop": "prop", "_dir": "bet_direction"}, inplace=True)
    agg["void_rate"] = agg["n_void"] / agg["n_rows"].replace(0, np.nan)
    agg["void_rate_ex_demon"] = agg["n_void_ex_demon"] / agg["n_rows_ex_demon"].replace(0, np.nan)
    agg["hit_rate"] = np.where(agg["n_decided"] > 0, agg["n_hit"] / agg["n_decided"], np.nan)
    agg["decided_rate"] = agg["n_decided"] / agg["n_rows"].replace(0, np.nan)
    return agg.sort_values(["slate_date", "prop", "bet_direction"])


def aggregate_prop_dir_pooled(all_rows: pd.DataFrame) -> pd.DataFrame:
    if all_rows.empty:
        return all_rows
    g = all_rows.groupby(["_prop", "_dir"], dropna=False)
    agg = g.agg(
        n_rows=("_is_void", "size"),
        n_void=("_is_void", "sum"),
        n_rows_ex_demon=("_is_demon", lambda s: int((~s).sum())),
        n_void_ex_demon=("_is_void_ex_demon", "sum"),
        n_decided=("_is_decided", "sum"),
        n_decided_ex_demon=("_is_decided_ex_demon", "sum"),
        n_hit=("_is_hit", "sum"),
        line_min=("_line", "min"),
        line_q25=("_line", lambda s: float(s.quantile(0.25)) if s.notna().any() else np.nan),
        line_median=("_line", "median"),
        line_q75=("_line", lambda s: float(s.quantile(0.75)) if s.notna().any() else np.nan),
        line_max=("_line", "max"),
        line_mean=("_line", "mean"),
        n_line_present=("_line", lambda s: int(s.notna().sum())),
    ).reset_index()
    agg.rename(columns={"_prop": "prop", "_dir": "bet_direction"}, inplace=True)
    agg["void_rate"] = agg["n_void"] / agg["n_rows"].replace(0, np.nan)
    agg["void_rate_ex_demon"] = agg["n_void_ex_demon"] / agg["n_rows_ex_demon"].replace(0, np.nan)
    agg["hit_rate"] = np.where(agg["n_decided"] > 0, agg["n_hit"] / agg["n_decided"], np.nan)
    agg["decided_rate"] = agg["n_decided"] / agg["n_rows"].replace(0, np.nan)
    return agg.sort_values(["prop", "bet_direction"])


def main() -> None:
    ap = argparse.ArgumentParser(description="Summarize graded NBA1Q xlsx files (voids + prop×direction×date).")
    ap.add_argument(
        "--out-dir",
        type=Path,
        default=None,
        help="Output directory (default: <repo>/data/reports/nba1q_graded)",
    )
    ap.add_argument(
        "--roots",
        type=Path,
        nargs="*",
        default=None,
        help="Extra roots to scan (default: repo outputs + ui_runner/graded_slate)",
    )
    args = ap.parse_args()
    root = _repo_root()
    out_dir = args.out_dir or (root / "data" / "reports" / "nba1q_graded")
    out_dir.mkdir(parents=True, exist_ok=True)
    roots = list(args.roots) if args.roots else [root / "outputs", root / "ui_runner" / "graded_slate"]

    paths = discover_nba1q_workbooks(roots)
    file_rows: list[dict] = []
    all_chunks: list[pd.DataFrame] = []

    for p in paths:
        slate = _slate_date_from_path(p)
        df, err = load_box_raw(p)
        if df is None:
            file_rows.append(
                {
                    "source_file": p.name,
                    "source_path": str(p.resolve()),
                    "slate_date": slate or "",
                    "error": err or "unknown",
                }
            )
            continue
        file_rows.append(summarize_file(p, df, slate))
        all_chunks.append(enrich_rows(df, slate, p.name))

    summary_df = pd.DataFrame(file_rows)
    summary_path = out_dir / "nba1q_file_summary.csv"
    summary_df.to_csv(summary_path, index=False)

    if not all_chunks:
        print("No graded_nba1q workbooks found under:", ", ".join(str(r) for r in roots))
        print("Wrote:", summary_path)
        return

    stacked = pd.concat(all_chunks, ignore_index=True)
    by_date = aggregate_prop_dir_date(stacked)
    by_date_path = out_dir / "nba1q_prop_direction_by_date.csv"
    by_date.to_csv(by_date_path, index=False)

    pooled = aggregate_prop_dir_pooled(stacked)
    pooled_path = out_dir / "nba1q_prop_direction_pooled.csv"
    pooled.to_csv(pooled_path, index=False)

    rows_total = len(stacked)
    void_total = int(stacked["_is_void"].sum())
    non_demon_total = int((~stacked["_is_demon"]).sum())
    void_non_demon = int(stacked["_is_void_ex_demon"].sum())
    decided_total = int(stacked["_is_decided"].sum())
    decided_non_demon = int(stacked["_is_decided_ex_demon"].sum())
    void_rate_all = (void_total / rows_total) if rows_total else np.nan
    void_rate_ex_demon = (void_non_demon / non_demon_total) if non_demon_total else np.nan

    print(f"Workbooks found: {len(paths)}")
    print(f"Total Box Raw rows stacked: {len(stacked)}")
    print(
        f"Void rate (all rows): {void_rate_all:.4f} | decided={decided_total}"
    )
    print(
        "Void rate (ex-Demon, gate metric): "
        f"{void_rate_ex_demon:.4f} | decided_ex_demon={decided_non_demon}"
    )
    print(f"Wrote: {summary_path}")
    print(f"Wrote: {by_date_path}")
    print(f"Wrote: {pooled_path}")


if __name__ == "__main__":
    main()
