#!/usr/bin/env python3
"""
Full Slate Grader v4 — NBA/CBB with full breakdowns:
- Summary: Overall, By Pick Type (Standard→OVER/UNDER), By Tier (→pick types→OVER/UNDER),
           By Def Tier, By Minutes Tier, By Player Role — all with OVER/UNDER splits
- Sheets: Box Raw, By Pick Type, By Tier, Prop Type x Direction,
          By Direction, By Edge Bucket, By Minutes Tier, By Def Tier,
          By Player Role, By Shot Role, Void Reasons
"""
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import argparse
import os
import re
import sys
from datetime import datetime
from pathlib import Path

_SCRIPTS_DIR = Path(__file__).resolve().parent.parent
_REPO_ROOT = Path(__file__).resolve().parent.parent.parent
if str(_SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS_DIR))
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from player_name_norm import fold_player_name  # noqa: E402
from utils.slate_fields import (  # noqa: E402
    first_numeric_in_slate_row,
    first_over_under_in_slate_row,
)
from utils.group_rank_tier import assign_tier_column  # noqa: E402


def _def_rank_bucket(x):
    try:
        r = float(x)
    except Exception:
        return "UNK"
    if r <= 5:   return "01-05"
    if r <= 10:  return "06-10"
    if r <= 20:  return "11-20"
    if r <= 25:  return "21-25"
    return "26-30"

C = {
    'hit':'27AE60','miss':'E74C3C','push':'F39C12','void':'95A5A6',
    'hdr':'1C1C1C','hdr2':'1A5276','hdr3':'1E8449','hdr4':'7D6608',
    'hdr5':'922B21','hdr6':'6C3483','hdr7':'117A65','hdr8':'1A5276',
    'alt':'F2F3F4','white':'FFFFFF',
    'tier_a':'D5F5E3','tier_b':'D6EAF8','tier_c':'FEF9E7','tier_d':'FDEDEC',
    'over':'D6EAF8','under':'FDEBD0',
    'goblin':'E8D5F5','demon':'FDEDEC','standard':'F2F3F4',
}
DEF_TIER_ORDER    = ['Elite', 'Above Avg', 'Avg', 'Below Avg', 'Weak']
MINUTES_TIER_ORDER= ['HIGH','MEDIUM','LOW','UNKNOWN']
USAGE_ROLE_ORDER  = ['PRIMARY','SECONDARY','SUPPORT','UNKNOWN']
SHOT_ROLE_ORDER   = ['HIGH_VOL','MID_VOL','LOW_VOL','UNKNOWN']
TIER_ORDER        = ['A','B','C','D']

def bdr(color='CCCCCC'):
    s = Side(style='thin',color=color)
    return Border(left=s,right=s,top=s,bottom=s)

def hc(ws,r,c,v,bg=None,fc='FFFFFF',bold=True,sz=9,align='center'):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=bold,color=fc,name='Arial',size=sz)
    if bg: cell.fill=PatternFill('solid',start_color=bg)
    cell.alignment=Alignment(horizontal=align,vertical='center')
    cell.border=bdr()
    return cell

def dc(ws,r,c,v,bg=None,bold=False,sz=9,align='center',fmt=None,fc='000000'):
    cell=ws.cell(row=r,column=c,value=v)
    cell.font=Font(bold=bold,name='Arial',size=sz,color=fc)
    cell.fill=PatternFill('solid',start_color=bg or C['white'])
    cell.alignment=Alignment(horizontal=align,vertical='center')
    cell.border=bdr()
    if fmt: cell.number_format=fmt
    return cell

def res_bg(r):
    return {'HIT':C['hit'],'MISS':C['miss'],'PUSH':C['push'],'VOID':C['void']}.get(str(r).upper(),'DDDDDD')

def hr_bg(v):
    if v is None or (isinstance(v,float) and np.isnan(v)): return 'DDDDDD'
    if v>=0.65: return C['hit']
    if v>=0.50: return C['push']
    return C['miss']

def tier_bg(t):
    return {'A':C['tier_a'],'B':C['tier_b'],'C':C['tier_c'],'D':C['tier_d']}.get(str(t).upper(),C['white'])

def pct_cell(ws,r,c,val):
    nan=val is None or (isinstance(val,float) and np.isnan(val))
    bg=hr_bg(val) if not nan else 'DDDDDD'
    cell=dc(ws,r,c,val if not nan else '',bg=bg,bold=True)
    if not nan:
        cell.number_format='0.0%'
        cell.font=Font(bold=True,name='Arial',size=9,color='FFFFFF')
    return cell

def hit_rate(sub):
    # Demons are excluded from hit-rate grading (data-collection only).
    # They still appear in Box Raw but are NOT counted in any hit-rate summary.
    graded = sub[sub['pick_type'] != 'Demon'] if 'pick_type' in sub.columns else sub
    dec = graded[graded['result'].isin(['HIT', 'MISS'])]
    h = (dec['result'] == 'HIT').sum()
    # Voids = non-decided graded rows + all Demon rows (they don't count toward the rate)
    demon_count = int((sub['pick_type'] == 'Demon').sum()) if 'pick_type' in sub.columns else 0
    v = int(graded['result'].isin(['VOID', 'PUSH']).sum()) + demon_count
    return (h / len(dec) if len(dec) else np.nan), int(h), int(len(dec) - h), v, int(len(dec))

def drc(df):
    return 'bet_direction' if 'bet_direction' in df.columns else 'final_bet_direction'

def sw(ws,widths):
    for ci,w in enumerate(widths,1):
        ws.column_dimensions[get_column_letter(ci)].width=w

def sheet_hdr8(ws,col1,bg,widths=None):
    sw(ws,widths or [24,10,8,10,8,8,8,12])
    for ci,h in enumerate([col1,'Direction','Total','Decided','Hits','Misses','Voids','Hit Rate'],1):
        hc(ws,1,ci,h,bg=bg)
    ws.row_dimensions[1].height=20
    ws.freeze_panes='A2'

def write_dir_subrows(ws,ri,df,label,bg,ncols=8):
    d=drc(df)
    hr_a,h_a,m_a,v_a,dec_a=hit_rate(df)
    row_bg=bg or (C['alt'] if ri%2==0 else C['white'])
    dc(ws,ri,1,label,bg=row_bg,bold=True,align='left')
    dc(ws,ri,2,'ALL',bg=row_bg,bold=True)
    dc(ws,ri,3,len(df),bg=row_bg); dc(ws,ri,4,dec_a,bg=row_bg)
    dc(ws,ri,5,h_a,bg=row_bg); dc(ws,ri,6,m_a,bg=row_bg); dc(ws,ri,7,v_a,bg=row_bg)
    pct_cell(ws,ri,8,hr_a); ri+=1
    for direction in ['OVER','UNDER']:
        dsub=df[df[d].str.upper()==direction]
        if len(dsub)==0: continue
        hr_d,h_d,m_d,v_d,dec_d=hit_rate(dsub)
        dbg=C['over'] if direction=='OVER' else C['under']
        dc(ws,ri,1,'',bg=dbg); dc(ws,ri,2,direction,bg=dbg,bold=True)
        dc(ws,ri,3,len(dsub),bg=dbg); dc(ws,ri,4,dec_d,bg=dbg)
        dc(ws,ri,5,h_d,bg=dbg); dc(ws,ri,6,m_d,bg=dbg); dc(ws,ri,7,v_d,bg=dbg)
        pct_cell(ws,ri,8,hr_d); ri+=1
    return ri

# ── Grade ─────────────────────────────────────────────────────────────────────
def _row_first_numeric(row, keys):
    """
    First non-null numeric among named columns (pandas Series or mapping).
    Delegates to utils.slate_fields for Series; dict-like rows use .get chain.
    """
    if isinstance(row, pd.Series):
        return first_numeric_in_slate_row(row, tuple(keys))
    for k in keys:
        if not hasattr(row, "get"):
            break
        x = pd.to_numeric(row.get(k, np.nan), errors="coerce")
        if pd.notna(x):
            return float(x)
    return np.nan


def _row_bet_direction(row) -> str:
    keys = (
        "bet_direction",
        "final_bet_direction",
        "Direction",
        "Dir",
        "direction",
        "recommended_side",
    )
    if isinstance(row, pd.Series):
        v = first_over_under_in_slate_row(row, keys)
        return v if v else "OVER"
    if hasattr(row, "get"):
        for k in keys:
            val = row.get(k)
            s = str(val if val is not None else "").strip().upper()
            if s in ("OVER", "UNDER"):
                return s
    return "OVER"


def grade(row, actual):
    """Return (result, void_reason_or_None, margin). Margin is NaN when ungraded."""
    from grading.leg_grade_utils import slate_grade_row

    return slate_grade_row(actual, row)

# Quarter milestone props: actuals come from ESPN PBP in fetch_actuals.parse_nba_quarter_milestone_rows.
PROP_NORM_MAP={
    # short code aliases (common in CBB pipeline files)
    'pts':'points',
    'reb':'rebounds',
    'ast':'assists',
    'stl':'steals',
    'blk':'blocked shots',
    'tov':'turnovers',
    'fg3m':'3-pt made',
    'pr':'pts+rebs',
    'pa':'pts+asts',
    'ra':'rebs+asts',
    'stocks':'blks+stls',
    'pts+rebs+asts':'pts+rebs+asts','pra':'pts+rebs+asts',
    'pts+rebs':'pts+rebs','pts+asts':'pts+asts','rebs+asts':'rebs+asts','blks+stls':'blks+stls',
    'blocked shots':'blocked shots','blocks':'blocked shots',
    'steals':'steals','turnovers':'turnovers',
    # FG Made / Attempted
    'fg made':'fg made','field goals made':'fg made','fgm':'fg made',
    'fg attempted':'fg attempted','field goals attempted':'fg attempted','fga':'fg attempted',
    # 3-PT
    '3-pt made':'3-pt made','3pt made':'3-pt made','3pm':'3-pt made',
    '3 pt made':'3-pt made','3-point made':'3-pt made','3 point made':'3-pt made',
    'three pointers made':'3-pt made','three pointer made':'3-pt made',
    '3-pt attempted':'3-pt attempted','3pt attempted':'3-pt attempted','3pa':'3-pt attempted',
    '3 pt attempted':'3-pt attempted','3-point attempted':'3-pt attempted','3 point attempted':'3-pt attempted',
    'three pointers attempted':'3-pt attempted','three pointer attempted':'3-pt attempted',
    # Two Pointers
    'two pointers made':'two pointers made','2-pt made':'two pointers made',
    '2pt made':'two pointers made','two pointer made':'two pointers made',
    'two pointers attempted':'two pointers attempted','2-pt attempted':'two pointers attempted',
    '2pt attempted':'two pointers attempted','two pointer attempted':'two pointers attempted',
    # Free Throws
    'free throws made':'free throws made','ftm':'free throws made',
    'free throw made':'free throws made',
    'free throws attempted':'free throws attempted','fta':'free throws attempted',
    'free throw attempted':'free throws attempted',
    # Other
    'offensive rebounds':'offensive rebounds','defensive rebounds':'defensive rebounds',
    'personal fouls':'personal fouls','fantasy score':'fantasy score',
    'quarters with 3+ points':'quarters with 3+ points',
    'quarters with 5+ points':'quarters with 5+ points',
    # Milestone yes/no (actuals from box: 1.0 / 0.0 vs typical 0.5 line)
    'double double':'double double','triple double':'triple double',
    'double-double':'double double','triple-double':'triple double',
    'dd':'double double','td':'triple double',
    # Combo props — display names from PP slate files
    'points (combo)':'points',
    'assists (combo)':'assists',
    'rebounds (combo)':'rebounds',
    '3-pt made (combo)':'3-pt made',
    '3-pt attempted (combo)':'3-pt attempted',
    'two pointers made (combo)':'two pointers made',
    'two pointers attempted (combo)':'two pointers attempted',
    'free throws made (combo)':'free throws made',
    'free throws attempted (combo)':'free throws attempted',
    '3pt made (combo)':'3-pt made',
    '3pt attempted (combo)':'3-pt attempted',
    # Alternate display-name capitalizations seen in slate exports
    'field goal attempts':'fg attempted',
    'field goals':'fg made',
    '3-pointers made':'3-pt made',
    '3-pointers attempted':'3-pt attempted',
    '3 pointers made':'3-pt made',
    '3 pointers attempted':'3-pt attempted',
    # NBA step7 prop_norm short codes (align with fetch_actuals Title Case → lower)
    'fg3m':'3-pt made',
    'fg3a':'3-pt attempted',
    'fg2m':'two pointers made',
    'fg2a':'two pointers attempted',
    'fgm':'fg made',
    'fga':'fg attempted',
    'ftm':'free throws made',
    'fta':'free throws attempted',
    'oreb':'offensive rebounds',
    'orebs':'offensive rebounds',
    'dreb':'defensive rebounds',
    'drebs':'defensive rebounds',
    'pf':'personal fouls',
    'personalfouls':'personal fouls',
    'fantasy':'fantasy score',
    'fantasyscore':'fantasy score',
    'fs':'fantasy score',
    '2ptm':'two pointers made',
    '2pta':'two pointers attempted',
    '2pt made':'two pointers made',
    '2pt attempted':'two pointers attempted',
}
def norm_prop_key(p) -> str:
    s = str(p).lower().strip()
    # normalize common suffixes
    s = s.replace("(combo)", "").strip()
    # collapse whitespace
    s = " ".join(s.split())
    return PROP_NORM_MAP.get(s, s)


def norm_player_key(p) -> str:
    return fold_player_name(p)


def norm_team_key(team) -> str:
    return str(team or "").strip().upper()


def _norm_game_date(v) -> str:
    s = str(v or "").strip()[:10]
    if s.lower() in ("", "nan", "none", "<na>"):
        return ""
    return s


def _prop_key_variants(canon: str) -> set[str]:
    """Canonical prop plus PROP_NORM_MAP aliases that normalize to the same stat."""
    keys: set[str] = {canon}
    for short, mapped in PROP_NORM_MAP.items():
        ns = norm_prop_key(short)
        nl = norm_prop_key(mapped)
        if nl == canon or ns == canon:
            keys.add(ns)
            keys.add(nl)
    return keys


def _expand_lookup_keys(
    player: str,
    prop: str,
    *,
    team: str = "",
    game_date: str = "",
) -> tuple[list[str], list[str]]:
    """Return (team_scoped_keys, player_prop_keys) for actuals resolution."""
    p0 = norm_player_key(player)
    canon = norm_prop_key(prop)
    if not p0 or not canon:
        return [], []
    t0 = norm_team_key(team)
    d0 = _norm_game_date(game_date)
    team_keys: list[str] = []
    player_keys: list[str] = []
    for pv in sorted(_prop_key_variants(canon)):
        if t0:
            if d0:
                team_keys.append(f"{p0}|{t0}|{pv}|{d0}")
            team_keys.append(f"{p0}|{t0}|{pv}")
        if d0:
            player_keys.append(f"{p0}|{pv}|{d0}")
        player_keys.append(f"{p0}|{pv}")
    return team_keys, player_keys


class ActualsLookup:
    """NBA/CBB actuals indexes: prefer player|team|prop; player|prop only when unique."""

    __slots__ = (
        "by_player_team_prop",
        "by_player_prop",
        "ambiguous_player_prop",
        "actuals_date",
    )

    def __init__(self) -> None:
        self.by_player_team_prop: dict[str, float] = {}
        self.by_player_prop: dict[str, float] = {}
        self.ambiguous_player_prop: set[str] = set()
        self.actuals_date: str = ""


# Columns used to build ``player|prop`` keys when joining slate rows to actuals.
_PROP_RESOLVE_COLS: tuple[str, ...] = (
    "prop_type_norm",
    "stat_norm",
    "prop_norm",
    "prop_type",
    "Prop",
)


def _alias_cols(df):
    for alias,canon in [('DEF_TIER','def_tier'),('minutes_tier','minutes_tier'),
                         ('shot_role','shot_role'),('usage_role','usage_role')]:
        if alias in df.columns and canon not in df.columns:
            df[canon]=df[alias]
    return df


def _coalesce_line_from_line_score(df):
    """Fill NaN line values from line_score (NBA/CBB exports often split these)."""
    if "line_score" not in df.columns:
        return df
    ls = pd.to_numeric(df["line_score"], errors="coerce")
    if "line" not in df.columns:
        df["line"] = ls
    else:
        ln = pd.to_numeric(df["line"], errors="coerce")
        df["line"] = ln.where(ln.notna(), ls)
    return df


def _coalesce_line_from_projection(df):
    """When book line is blank, use numeric projection / model line (step7 exports)."""
    for col in ("projection", "Projection", "proj", "model_line", "consensus_line"):
        if col not in df.columns:
            continue
        pv = pd.to_numeric(df[col], errors="coerce")
        if pv.notna().sum() == 0:
            continue
        if "line" not in df.columns:
            df["line"] = pv
        else:
            ln = pd.to_numeric(df["line"], errors="coerce")
            df["line"] = ln.where(ln.notna(), pv)
    return df


def load_nba(path: str, sport_code: str = "NBA") -> pd.DataFrame:
    xls = pd.ExcelFile(path, engine="openpyxl")
    preferred = ["ALL", "Full Slate", "Box Raw", "Props", "Sheet1"]
    sheet = next((s for s in preferred if s in xls.sheet_names), xls.sheet_names[0])
    if sheet != "ALL":
        print(f"⚠️ NBA: sheet 'ALL' not found in {os.path.basename(path)}. Using sheet='{sheet}'.")

    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    # --- Normalize column names from PipelineA output ---
    df.columns = [str(c).strip() for c in df.columns]

    if sheet == "Full Slate" and "Sport" in df.columns:
        su = df["Sport"].astype(str).str.strip().str.upper()
        code = str(sport_code or "NBA").strip().upper()
        df = df.loc[su == code].copy()

    df = df.rename(columns={
    "Player": "player",
    "Prop": "prop_type_norm",
    "Pick Type": "pick_type",
    "Line": "line",
    "Direction": "bet_direction",
    "Team": "team",
    "Opp": "opp_team",
    "Pos": "pos",
    "Tier": "tier",
    "Game Time": "game_time",
    "Game Date": "game_date",
    "Rank Score": "rank_score",
    "Edge": "edge",
    "Projection": "projection",
    "Void Reason": "void_reason",

    # ADD THESE 👇
    "Def Rank": "def_rank",
    "Def Tier": "def_tier",
    "Min Tier": "minutes_tier",
    "Shot Role": "shot_role",
    "Usage Role": "usage_role",
    "ML Prob": "ml_prob",
    "ml_prob": "ml_prob",
    "ML Edge": "ml_edge",
    "ml_edge": "ml_edge",
    "Edge Score": "edge_score",
    "edge_score": "edge_score",
    "Blended Score": "blended_score",
    "blended_score": "blended_score",
    "Consistency Grade": "consistency_grade",
    "Top3 Rank": "team_top3_rank",
    "Bottom3 Rank": "team_bottom3_rank",
    "Def Boost Hist": "def_boost_hist",
    "Top3 Weak Over": "top3_weak_overperformer",
    "Top3 Elite Fade": "top3_elite_fader",
    "Top3 Def Context": "top3_def_context",
    "Top3 Under Context": "top3_under_context",
    "L5 Over": "l5_over",
    "L5 Under": "l5_under",
    "last5_over": "l5_over",
    "last5_under": "l5_under",
    "Hit Rate (5g)": "hit_rate",
    "Hit Rate": "hit_rate",
    "last5_hit_rate": "hit_rate",
})
    # fallback compatibility
    if "prop_type_norm" not in df.columns:
        df["prop_type_norm"] = df.get("prop_type", "")

    if "bet_direction" not in df.columns and "final_bet_direction" in df.columns:
        df["bet_direction"] = df["final_bet_direction"]

    if "ml_prob" in df.columns:
        df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")
    if "ml_edge" in df.columns:
        df["ml_edge"] = pd.to_numeric(df["ml_edge"], errors="coerce")
    elif "ml_prob" in df.columns:
        df["ml_edge"] = df["ml_prob"] - 0.5

    _coalesce_line_from_line_score(df)
    _coalesce_line_from_projection(df)

    # Hard fail if player still missing
    if "player" not in df.columns:
        raise KeyError(f"NBA slate missing 'player' column. Found columns: {list(df.columns)}")
    # Diagnostic key (apply_actuals rebuilds with team/date for resolution).
    df["player_key"] = _slate_player_key_series(df)
    return df


def _slate_player_key_series(df: pd.DataFrame) -> pd.Series:
    pk = df["player"].astype(str).apply(norm_player_key)
    pp = df["prop_type_norm"].fillna("").astype(str).apply(norm_prop_key)
    if "team" in df.columns:
        tk = df["team"].map(norm_team_key)
        base = pk + "|" + tk + "|" + pp
    else:
        base = pk + "|" + pp
    if "game_date" in df.columns:
        gd = df["game_date"].map(_norm_game_date)
        has_d = gd.astype(str).str.len() > 0
        return base.where(~has_d, base + "|" + gd)
    return base


def _strict_slate_date_env_exit(message: str) -> None:
    """When PROPORACLE_GRADER_STRICT_SLATE_DATE is set, exit non-zero instead of producing an empty graded slate."""
    v = (os.environ.get("PROPORACLE_GRADER_STRICT_SLATE_DATE") or "").strip().lower()
    if v not in ("1", "true", "yes", "on"):
        return
    print(f"ERROR: {message}", file=sys.stderr)
    sys.exit(2)


def filter_nba_slate_by_grade_date(df: pd.DataFrame, date_str: str) -> pd.DataFrame:
    """
    Keep only rows whose game date matches date_str (YYYY-MM-DD).
    Stops NBA1H/NBA1Q/full NBA from grading the wrong slate when only
    NBA\\step8_*_clean.xlsx (latest pipeline) is available.
    """
    if not date_str or not len(df):
        return df
    ds = str(date_str).strip()[:10]
    if "game_date" in df.columns:
        gd = df["game_date"].astype(str).str.strip().str[:10]
        nonempty = gd.ne("") & gd.ne("nan") & gd.ne("None") & gd.ne("<NA>")
        if nonempty.any():
            mask = gd.eq(ds) & nonempty
            if mask.any():
                print(f"  INFO: Slate date filter ({date_str}) [game_date]: kept {int(mask.sum())}/{len(df)} rows")
                return df.loc[mask].copy()
            sample_dates = sorted(
                {str(gd_i)[:10] for gd_i in gd.loc[nonempty].tolist() if str(gd_i).strip() not in ("", "nan", "None", "<NA>")}
            )[:16]
            print(
                f"  WARN: Date filter {date_str} on 'game_date': 0 rows match -- "
                f"graded workbook will be empty (slate is for other day(s))."
            )
            if sample_dates:
                print(f"  INFO: Non-empty game_date values on this slate (sample): {sample_dates}")
            print(
                "  HINT: Use run_grader.ps1 -Date <ET slate calendar date>, "
                "or place that day's step8 workbook under outputs\\<date>\\ for extraction."
            )
            _strict_slate_date_env_exit(
                f"Strict slate date (PROPORACLE_GRADER_STRICT_SLATE_DATE): "
                f"no rows with game_date={date_str}. Slate carries other dates (see INFO above). "
                "Unset the env var to allow an empty graded workbook."
            )
            return df.iloc[0:0].copy()

    col = None
    for c in ("game_time", "game_start", "fetched_at"):
        if c in df.columns:
            col = c
            break
    if not col:
        return df
    ts = pd.to_datetime(df[col], errors="coerce")
    try:
        target = pd.to_datetime(date_str).date()
    except Exception:
        return df
    # PrizePicks / ticket exports often omit year (MM/DD HH:MM) — match month+day only
    mask = (ts.dt.month == target.month) & (ts.dt.day == target.day) & ts.notna()
    if not mask.any():
        if ts.notna().sum() == 0:
            print(
                f"  WARN: No parseable dates in '{col}'; cannot filter to {date_str} -- using all {len(df)} rows."
            )
            return df
        print(
            f"  WARN: Date filter {date_str} on '{col}': 0 rows match -- "
            f"graded workbook will be empty (slate is for other day(s))."
        )
        print(
            "  HINT: Use run_grader.ps1 -Date <game-day> matching Game Time in the slate, "
            "or place that day's step8 workbook under outputs\\<date>\\ for extraction."
        )
        _strict_slate_date_env_exit(
            f"Strict slate date (PROPORACLE_GRADER_STRICT_SLATE_DATE): "
            f"no '{col}' timestamps match calendar day {date_str}."
        )
        return df.loc[mask].copy()
    print(f"  INFO: Slate date filter ({date_str}): kept {int(mask.sum())}/{len(df)} rows")
    return df.loc[mask].copy()


def load_cbb(path: str) -> pd.DataFrame:
    # CBB graded files sometimes have 'ALL', but sometimes just Sheet1
    xls = pd.ExcelFile(path, engine="openpyxl")
    sheet = "ALL" if "ALL" in xls.sheet_names else xls.sheet_names[0]
    if sheet != "ALL":
        print(f"⚠️ CBB: sheet 'ALL' not found in {os.path.basename(path)}. Using sheet='{sheet}'.")
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
    df.columns = [str(c).strip() for c in df.columns]

    # ── FIX: rename CBB pipeline column names to grader-expected names ──────────
    df = df.rename(columns={
        "Player":              "player",
        "Line":                "line",
        "Void Reason":         "void_reason",
        "opp_team_abbr":       "opp_team",
        "team_abbr":           "team",
        "prop_norm":           "prop_type_norm",   # FIX 2: use normalized prop, not verbose prop_type
        "final_bet_direction": "bet_direction",
        "model_dir_5":         "model_dir_5",
        "opp_def_tier":        "def_tier",          # prefer opp_def_tier if def_tier blank
        "OVERALL_DEF_RANK":    "def_rank",
        "rank_score":          "rank_score",
        "edge":                "edge",
        "ML Prob":             "ml_prob",
        "ml_prob":             "ml_prob",
        "ML Edge":             "ml_edge",
        "ml_edge":             "ml_edge",
        "Edge Score":          "edge_score",
        "edge_score":          "edge_score",
        "Blended Score":       "blended_score",
        "blended_score":       "blended_score",
    })
    # Guard against duplicate column names after renames (e.g. bet_direction).
    if df.columns.duplicated().any():
        df = df.loc[:, ~df.columns.duplicated()]

    df = _alias_cols(df)

    # FIX 1: bet_direction — CBB step3 has no direction column yet; default to OVER
    # but warn so the user knows actuals grading requires a direction column.
    if "bet_direction" not in df.columns:
        dir_candidates = [c for c in df.columns if "direction" in c.lower() or "dir" in c.lower()]
        if dir_candidates:
            df["bet_direction"] = df[dir_candidates[0]].astype(str).str.upper().str.strip()
            print(f"  ℹ️  CBB: using '{dir_candidates[0]}' as bet_direction")
        else:
            print("  ⚠️  CBB: no direction column found — all props will default to OVER. "
                  "Pass a slate file that includes final_bet_direction for correct grading.")
            df["bet_direction"] = "OVER"

    # FIX 2: prop_type_norm — use prop_norm (already normalized) not verbose prop_type
    if "prop_type_norm" not in df.columns:
        if "prop_norm" in df.columns:
            df["prop_type_norm"] = df["prop_norm"]
        else:
            df["prop_type_norm"] = df.get("prop_type", "")

    # FIX 3: tier — may not exist in early pipeline steps; add empty column so
    # downstream breakdown sheets don't silently skip all rows.
    if "tier" not in df.columns:
        print("  ⚠️  CBB: no 'tier' column found — By Tier breakdowns will be empty. "
              "Pass step6_ranked_cbb.xlsx for full tier grading.")
        df["tier"] = ""

    if "ml_prob" in df.columns:
        df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")
    if "ml_edge" in df.columns:
        df["ml_edge"] = pd.to_numeric(df["ml_edge"], errors="coerce")
    elif "ml_prob" in df.columns:
        df["ml_edge"] = df["ml_prob"] - 0.5
    for c in ("edge_score", "blended_score"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    _coalesce_line_from_line_score(df)
    _coalesce_line_from_projection(df)

    df["player_key"] = _slate_player_key_series(df)
    return df


def _actuals_row_game_date(arow, file_date: str) -> str:
    for col in ("game_date", "date", "Game Date"):
        if col not in arow.index:
            continue
        d = _norm_game_date(arow.get(col))
        if d:
            return d
    return file_date


def _build_actuals_lookup(act: pd.DataFrame, actuals_path: str = "") -> ActualsLookup:
    """Index actuals by player|team|prop; player|prop only when unambiguous.

    Period and full-game actuals CSVs use PrizePicks-style labels (``Points``,
    ``3-PT Made``) while some slates use short codes (``pts``, ``fg3m``). Register
    every alias that normalizes to the same canonical prop so rows are not voided
    as ``NO_ACTUAL`` when the game was played and the stat exists under a variant
    label.

    When the same player|prop appears under multiple teams or dates, the
    player|prop fallback is dropped so grading does not silently pick the wrong box score.
    """
    lookup = ActualsLookup()
    lookup.actuals_date = _extract_date_from_actuals_filename(actuals_path)
    if act is None or len(act) == 0:
        return lookup
    if "player" not in act.columns or "actual" not in act.columns:
        return lookup
    prop_col = "prop_type" if "prop_type" in act.columns else ("Prop" if "Prop" in act.columns else None)
    if not prop_col:
        return lookup

    for _, arow in act.iterrows():
        val = pd.to_numeric(arow.get("actual"), errors="coerce")
        if pd.isna(val):
            continue
        p0 = norm_player_key(str(arow.get("player", "") or ""))
        if not p0:
            continue
        raw_prop = str(arow.get(prop_col, "") or "").strip()
        if not raw_prop or raw_prop.lower() in ("nan", "none"):
            continue
        t0 = norm_team_key(arow.get("team", ""))
        apath_l = str(actuals_path or "").lower()
        if "nba" in apath_l and "actuals" in apath_l:
            from espn_injuries import canon_team_abbr

            t0 = canon_team_abbr("NBA", arow.get("team", "")) or t0
        row_date = _actuals_row_game_date(arow, lookup.actuals_date)
        team_keys, player_keys = _expand_lookup_keys(
            p0, raw_prop, team=t0, game_date=row_date
        )
        fval = float(val)
        for k in team_keys:
            lookup.by_player_team_prop[k] = fval
        for k in player_keys:
            if k in lookup.by_player_prop and lookup.by_player_prop[k] != fval:
                lookup.ambiguous_player_prop.add(k)
            else:
                lookup.by_player_prop[k] = fval

    for k in lookup.ambiguous_player_prop:
        lookup.by_player_prop.pop(k, None)

    return lookup


def _props_from_row(row) -> list[str]:
    seen: set[str] = set()
    out: list[str] = []
    for col in _PROP_RESOLVE_COLS:
        if col not in row.index:
            continue
        cell = row.get(col)
        if cell is None or (isinstance(cell, float) and pd.isna(cell)):
            continue
        raw = str(cell).strip()
        if not raw or raw.lower() in ("nan", "none"):
            continue
        nk = norm_prop_key(raw)
        if nk and nk not in seen:
            seen.add(nk)
            out.append(nk)
    return out


def _primary_prop_from_row(row) -> str:
    """First non-empty prop column (same priority order as _resolve_actual)."""
    for col in _PROP_RESOLVE_COLS:
        if col not in row.index:
            continue
        cell = row.get(col)
        if cell is None or (isinstance(cell, float) and pd.isna(cell)):
            continue
        raw = str(cell).strip()
        if not raw or raw.lower() in ("nan", "none"):
            continue
        nk = norm_prop_key(raw)
        if nk:
            return nk
    return ""


def _row_prop_matches_lookup(lookup: ActualsLookup, srow, prop_label: str) -> bool:
    """True when team- or player-scoped keys for this prop resolve in the lookup."""
    team_keys, player_keys = _expand_lookup_keys(
        str(srow.get("player", "") or ""),
        prop_label,
        team=norm_team_key(srow.get("team", "")),
        game_date=_norm_game_date(srow.get("game_date", "")) or lookup.actuals_date,
    )
    if any(k in lookup.by_player_team_prop for k in team_keys):
        return True
    return any(
        k in lookup.by_player_prop and k not in lookup.ambiguous_player_prop
        for k in player_keys
    )


def _resolve_actual(lookup: ActualsLookup, row) -> float:
    """Resolve actual: team+prop first, then unique player+prop; respect game_date."""
    row_date = _norm_game_date(row.get("game_date", ""))
    if lookup.actuals_date and row_date and row_date != lookup.actuals_date:
        return np.nan

    player = str(row.get("player", "") or "")
    team = norm_team_key(row.get("team", ""))
    game_date = row_date or lookup.actuals_date

    for prop_label in _props_from_row(row):
        team_keys, player_keys = _expand_lookup_keys(
            player, prop_label, team=team, game_date=game_date
        )
        for k in team_keys:
            a = lookup.by_player_team_prop.get(k, np.nan)
            if pd.notna(a):
                return float(a)
        for k in player_keys:
            if k in lookup.ambiguous_player_prop:
                continue
            a = lookup.by_player_prop.get(k, np.nan)
            if pd.notna(a):
                return float(a)
    return np.nan


def _resolve_actual_for_player(
    lookup: ActualsLookup, player_str: str, row
) -> float:
    """Resolve actual for one combo leg using the same prop columns as ``_resolve_actual``."""
    row_date = _norm_game_date(row.get("game_date", ""))
    if lookup.actuals_date and row_date and row_date != lookup.actuals_date:
        return np.nan

    team = norm_team_key(row.get("team", ""))
    game_date = row_date or lookup.actuals_date
    for prop_label in _props_from_row(row):
        team_keys, player_keys = _expand_lookup_keys(
            player_str, prop_label, team=team, game_date=game_date
        )
        for k in team_keys:
            a = lookup.by_player_team_prop.get(k, np.nan)
            if pd.notna(a):
                return float(a)
        for k in player_keys:
            if k in lookup.ambiguous_player_prop:
                continue
            a = lookup.by_player_prop.get(k, np.nan)
            if pd.notna(a):
                return float(a)
    return np.nan


def _split_combo_players(player_str: str) -> list[str]:
    s = str(player_str)
    # common separators in PP combo names
    for sep in [" + ", "+", " & ", "&", " / ", "/", " vs ", " and "]:
        if sep in s:
            parts = [p.strip() for p in s.split(sep) if p.strip()]
            if len(parts) >= 2:
                return parts
    return []


_INJURY_STATUSES_DNP: frozenset[str] = frozenset(("Out", "Day-To-Day", "Injured Reserve"))


def _injury_status_marks_dnp(st: object) -> bool:
    s = str(st or "").strip()
    if not s:
        return False
    if s in _INJURY_STATUSES_DNP:
        return True
    return s.upper().startswith("DNP")


def _extract_date_from_actuals_filename(name: str) -> str:
    """Parse YYYY-MM-DD from the actuals file basename (not parent folders)."""
    base = Path(name).name
    m = re.search(r"(\d{4}-\d{2}-\d{2})", base)
    return m.group(1) if m else ""


def _load_nba_injury_dnp_keys(actuals_path: str) -> frozenset[tuple[str, str]]:
    """(norm_player_key, team_upper) for NBA injury/DNP rows (Out, DTD, IR, or DNP-* from box score)."""
    p = Path(actuals_path)
    ds = _extract_date_from_actuals_filename(p.name)
    if not ds:
        return frozenset()
    inj_path = p.parent / f"injuries_nba_{ds}.csv"
    if not inj_path.is_file():
        return frozenset()
    try:
        inj = pd.read_csv(inj_path)
    except Exception:
        return frozenset()
    keys: set[tuple[str, str]] = set()
    for _, r in inj.iterrows():
        if str(r.get("sport", "")).strip().upper() != "NBA":
            continue
        typ = str(r.get("injury_type", "")).strip().upper()
        st = str(r.get("injury_status", "")).strip()
        if typ != "DNP" and not _injury_status_marks_dnp(st):
            continue
        pl = norm_player_key(str(r.get("player", "") or ""))
        from espn_injuries import canon_team_abbr

        tm = canon_team_abbr("NBA", r.get("team", "")) or str(
            r.get("team", "") or ""
        ).strip().upper()
        if pl and tm:
            keys.add((pl, tm))
    return frozenset(keys)


def _injury_confirms_dnp_for_row(
    player_str: object, team_u: str, dnp_keys: frozenset[tuple[str, str]]
) -> bool:
    if not dnp_keys or not team_u:
        return False
    parts = _split_combo_players(str(player_str or ""))
    targets = parts if len(parts) >= 2 else [str(player_str or "").strip()]
    for raw in targets:
        if not raw:
            continue
        pl = norm_player_key(raw)
        if not pl:
            continue
        if (pl, team_u) in dnp_keys:
            return True
        hits = [k for k in dnp_keys if k[0] == pl]
        if len(hits) == 1:
            return True
    return False


def apply_actuals(df, actuals_path):
    act = pd.read_csv(actuals_path)
    apath = str(actuals_path or "")
    lookup = _build_actuals_lookup(act, apath)
    dnp_keys: frozenset[tuple[str, str]] = frozenset()
    if "nba" in apath.lower() and "actuals" in apath.lower():
        dnp_keys = _load_nba_injury_dnp_keys(apath)

    df["player_key"] = _slate_player_key_series(df)

    if lookup.actuals_date and "game_date" in df.columns:
        gd = df["game_date"].map(_norm_game_date)
        nonempty = gd.astype(str).str.len() > 0
        if nonempty.any():
            mismatch = nonempty & (gd != lookup.actuals_date)
            n_mm = int(mismatch.sum())
            if n_mm:
                print(
                    f"  WARN: {n_mm} slate row(s) have game_date != actuals file date "
                    f"({lookup.actuals_date}); those rows will not match actuals."
                )

    if lookup.ambiguous_player_prop:
        print(
            f"  INFO: {len(lookup.ambiguous_player_prop)} player|prop key(s) omitted from "
            "fallback (multiple team/date actuals); use team on slate rows."
        )

    # Diagnostic: prop types with no actuals match on any slate row (primary prop column only).
    # Do not scan every alias column per row — secondary labels (stat_norm, prop_norm, …)
    # often differ in spelling while grading still resolves via the primary column.
    canon_seen: set[str] = set()
    canon_matched: set[str] = set()
    for _, srow in df.iterrows():
        prop_label = _primary_prop_from_row(srow)
        if not prop_label:
            continue
        canon_seen.add(prop_label)
        if _row_prop_matches_lookup(lookup, srow, prop_label):
            canon_matched.add(prop_label)
    unmatched_props = canon_seen - canon_matched
    if unmatched_props:
        print(f"  ⚠️  Prop types in slate with NO actuals matches: {sorted(unmatched_props)}")

    # Upstream void_reason (NBA step7, etc.) marks eligibility / strategy filters
    # (BLOCKED_STD_OVER_LOW_HR, FORCED_OVER_NEG_EDGE, DROPPED_NEG_EDGE_GOBDEM, NO_PROJECTION_OR_LINE, …).
    # It must NOT force VOID when we have a real line + box-score actual — otherwise
    # Prop Evaluation and archives show Actual filled but Result VOID and Margin empty.

    results, void_reasons, margins, actuals_out = [], [], [], []
    for _, row in df.iterrows():
        actual = _resolve_actual(lookup, row)

        # Single-player combo props: sum final component box-score rows (same game snapshot).
        if pd.isna(actual):
            from grading.leg_grade_utils import is_nba_combo_prop, sum_nba_combo_from_actuals_df

            prop_label = _primary_prop_from_row(row)
            if prop_label and is_nba_combo_prop(prop_label):
                combo_val = sum_nba_combo_from_actuals_df(
                    act,
                    str(row.get("player", "") or ""),
                    str(row.get("team", "") or ""),
                    prop_label,
                )
                if combo_val is not None:
                    actual = float(combo_val)

        # Multi-player combo props (PlayerA + PlayerB): sum each player's stat.
        if pd.isna(actual):
            parts = _split_combo_players(row.get("player", ""))
            if parts:
                vals = []
                ok = True
                for p in parts:
                    v = _resolve_actual_for_player(lookup, p, row)
                    if pd.isna(v):
                        ok = False
                        break
                    vals.append(float(v))
                if ok and vals:
                    actual = float(sum(vals))

        actuals_out.append(actual)

        void_r = row.get("void_reason", np.nan)
        void_r_str = str(void_r).strip() if pd.notna(void_r) else ""
        upstream = void_r_str if void_r_str not in ("", "nan") else ""

        r, vr, m = grade(row, actual)
        decided = r in ("HIT", "MISS", "PUSH", "NEAR_LINE")

        if decided:
            results.append(r)
            margins.append(m)
            parts = [upstream] if upstream else []
            if vr:
                parts.append(str(vr))
            void_reasons.append("; ".join(parts))
        else:
            results.append("VOID")
            margins.append(np.nan)
            tail = str(vr or "").strip()
            team_u = str(row.get("team", "") or "").strip().upper()
            if "nba" in apath.lower() and "actuals" in apath.lower():
                from espn_injuries import canon_team_abbr

                team_u = canon_team_abbr("NBA", row.get("team", "")) or team_u
            if tail == "NO_ACTUAL" and dnp_keys and _injury_confirms_dnp_for_row(
                row.get("player", ""), team_u, dnp_keys
            ):
                tail = "INJURY_REPORT_DNP"
            if upstream and tail:
                void_reasons.append(f"{upstream}; {tail}")
            elif upstream:
                void_reasons.append(upstream)
            else:
                void_reasons.append(tail or "NO_ACTUAL")

    df["actual"] = actuals_out
    df["result"] = results
    df["void_reason_grade"] = void_reasons
    df["margin"] = margins
    df["result_sign"] = df["result"].map({"HIT": 1, "MISS": -1, "VOID": 0, "PUSH": 0})
    return df

def breakdown(df,group_col):
    rows=[]
    for key,sub in df.groupby(group_col,dropna=False):
        hr,h,m,v,dec=hit_rate(sub)
        rows.append({group_col:key,'total':len(sub),'hit':h,'miss':m,'void':v,'decided':dec,
                     'hit_rate':round(hr,4) if not np.isnan(hr) else np.nan})
    if not rows:
        return pd.DataFrame(columns=[group_col,'total','hit','miss','void','decided','hit_rate'])
    return pd.DataFrame(rows).sort_values('total',ascending=False)

def write_flat_breakdown(wb,df_b,sheet_name,group_col,bg_hdr):
    ws=wb.create_sheet(sheet_name)
    sw(ws,[28,8,8,8,8,10,12])
    for ci,col in enumerate([group_col,'total','hit','miss','void','decided','hit_rate'],1):
        hc(ws,1,ci,col,bg=bg_hdr)
    ws.row_dimensions[1].height=20; ws.freeze_panes='A2'
    for ri,row in enumerate(df_b.itertuples(),2):
        bg=C['alt'] if ri%2==0 else C['white']
        hr=getattr(row,'hit_rate',np.nan)
        try: gval=getattr(row,group_col)
        except: gval=getattr(row,group_col.replace(' ','_').replace('-','_').replace('+','_'),'')
        dc(ws,ri,1,gval,bg=bg,align='left')
        dc(ws,ri,2,getattr(row,'total',0),bg=bg); dc(ws,ri,3,getattr(row,'hit',0),bg=bg)
        dc(ws,ri,4,getattr(row,'miss',0),bg=bg); dc(ws,ri,5,getattr(row,'void',0),bg=bg)
        dc(ws,ri,6,getattr(row,'decided',0),bg=bg); pct_cell(ws,ri,7,hr)

# ── By Pick Type sheet ────────────────────────────────────────────────────────
def write_pick_type_sheet(wb,df):
    ws=wb.create_sheet('By Pick Type')
    sheet_hdr8(ws,'Pick Type',C['hdr3'],[20,10,8,10,8,8,8,12])
    ri=2
    for pt in ['Goblin','Demon','Standard']:
        sub=df[df['pick_type']==pt]
        if len(sub)==0: continue
        pt_bg={'Goblin':C['goblin'],'Demon':C['demon'],'Standard':C['standard']}.get(pt,C['white'])
        if pt=='Standard':
            ri=write_dir_subrows(ws,ri,sub,pt,pt_bg)
        else:
            hr_a,h_a,m_a,v_a,dec_a=hit_rate(sub)
            dc(ws,ri,1,pt,bg=pt_bg,bold=True,align='left'); dc(ws,ri,2,'OVER',bg=pt_bg,bold=True)
            dc(ws,ri,3,len(sub),bg=pt_bg); dc(ws,ri,4,dec_a,bg=pt_bg)
            dc(ws,ri,5,h_a,bg=pt_bg); dc(ws,ri,6,m_a,bg=pt_bg); dc(ws,ri,7,v_a,bg=pt_bg)
            pct_cell(ws,ri,8,hr_a); ri+=1

# ── By Tier sheet ─────────────────────────────────────────────────────────────
def write_tier_sheet(wb,df):
    ws=wb.create_sheet('By Tier')
    sw(ws,[8,12,10,8,10,8,8,8,12])
    for ci,h in enumerate(['Tier','Pick Type','Direction','Total','Decided','Hits','Misses','Voids','Hit Rate'],1):
        hc(ws,1,ci,h,bg=C['hdr4'])
    ws.row_dimensions[1].height=20; ws.freeze_panes='A2'
    d=drc(df); ri=2

    def r9(ws,ri,c1,c2,c3,total,dec,h,m,v,hr_val,bg):
        dc(ws,ri,1,c1,bg=bg,bold=True,align='center')
        dc(ws,ri,2,c2,bg=bg,bold=True,align='left')
        dc(ws,ri,3,c3,bg=bg,bold=True)
        dc(ws,ri,4,total,bg=bg); dc(ws,ri,5,dec,bg=bg)
        dc(ws,ri,6,h,bg=bg); dc(ws,ri,7,m,bg=bg); dc(ws,ri,8,v,bg=bg)
        pct_cell(ws,ri,9,hr_val); return ri+1

    for t in TIER_ORDER:
        tsub=df[df['tier'].astype(str).str.upper()==t]
        if len(tsub)==0: continue
        tbg=tier_bg(t)
        hr_a,h_a,m_a,v_a,dec_a=hit_rate(tsub)
        ri=r9(ws,ri,f'Tier {t}','ALL','ALL',len(tsub),dec_a,h_a,m_a,v_a,hr_a,tbg)
        for pt in ['Goblin','Demon','Standard']:
            ptsub=tsub[tsub['pick_type']==pt] if 'pick_type' in tsub.columns else pd.DataFrame()
            if len(ptsub)==0: continue
            pt_bg={'Goblin':C['goblin'],'Demon':C['demon'],'Standard':C['standard']}.get(pt,C['alt'])
            hr_p,h_p,m_p,v_p,dec_p=hit_rate(ptsub)
            if pt=='Standard':
                ri=r9(ws,ri,'',pt,'ALL',len(ptsub),dec_p,h_p,m_p,v_p,hr_p,pt_bg)
                for direction in ['OVER','UNDER']:
                    dsub=ptsub[ptsub[d].str.upper()==direction]
                    if len(dsub)==0: continue
                    hr_d,h_d,m_d,v_d,dec_d=hit_rate(dsub)
                    dbg=C['over'] if direction=='OVER' else C['under']
                    ri=r9(ws,ri,'','',direction,len(dsub),dec_d,h_d,m_d,v_d,hr_d,dbg)
            else:
                ri=r9(ws,ri,'',pt,'OVER',len(ptsub),dec_p,h_p,m_p,v_p,hr_p,pt_bg)

# ── Prop Type x Direction sheet ───────────────────────────────────────────────
def write_prop_direction_sheet(wb,df):
    ws=wb.create_sheet('Prop Type x Direction')
    sheet_hdr8(ws,'Prop Type',C['hdr2'],[28,10,8,10,8,8,8,12])
    d=drc(df)
    pt_col='prop_type_norm' if 'prop_type_norm' in df.columns else 'prop_type'
    prop_order=(df[df['result'].isin(['HIT','MISS'])].groupby(pt_col).size()
                .sort_values(ascending=False).index.tolist())
    prop_order+=[p for p in df[pt_col].unique() if p not in prop_order]
    ri=2
    for prop in prop_order:
        psub=df[df[pt_col]==prop]
        ri=write_dir_subrows(ws,ri,psub,prop,C['alt'] if ri%2==0 else C['white'])

# ── Tier-with-direction generic sheet ────────────────────────────────────────
def write_tier_dir_sheet(wb,df,sheet_name,tier_col,tier_order,bg_hdr):
    if tier_col not in df.columns: return
    ws=wb.create_sheet(sheet_name)
    sheet_hdr8(ws,sheet_name.replace('By ',''),bg_hdr)
    ri=2
    tiers=[t for t in tier_order if t in df[tier_col].dropna().unique()]
    tiers+=[t for t in df[tier_col].dropna().unique() if t not in tiers]
    for t in tiers:
        sub=df[df[tier_col]==t]
        ri=write_dir_subrows(ws,ri,sub,t,C['alt'] if ri%2==0 else C['white'])

# ── By Def Rank bucket sheet (NEW) ────────────────────────────────────────────
def write_def_rank_bucket_sheet(wb, df):
    if 'def_rank' not in df.columns:
        return
    tmp = df.copy()
    tmp['def_rank_bucket'] = tmp['def_rank'].apply(_def_rank_bucket)

    ws = wb.create_sheet('By Def Rank')
    sheet_hdr8(ws, 'Def Rank Bucket', C['hdr5'])
    ri = 2
    for b in ['01-05','06-10','11-20','21-25','26-30','UNK']:
        sub = tmp[tmp['def_rank_bucket'] == b]
        if len(sub) == 0:
            continue
        ri = write_dir_subrows(ws, ri, sub, b, C['alt'] if ri % 2 == 0 else C['white'])

# ── Performance Matrix sheet ─────────────────────────────────────────────────
def write_performance_matrix(wb, df):
    """Pick Type × Tier × Direction matrix with hit rates.
    Standard: all tiers, OVER + UNDER.  Goblin/Demon: all tiers, OVER only.
    Cells with 0 decided props show '—'.
    """
    ws = wb.create_sheet('Performance Matrix')
    d = drc(df)

    col_headers = ['Pick Type', 'Tier', 'Direction', 'Total', 'Decided', 'Hits', 'Misses', 'Hit Rate']
    widths      = [14, 7, 11, 8, 9, 7, 8, 12]
    sw(ws, widths)
    for ci, h in enumerate(col_headers, 1):
        hc(ws, 1, ci, h, bg=C['hdr'])
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    def _matrix_row(ri, pick_type, tier, direction, sub):
        pt_bg = {'Goblin': C['goblin'], 'Demon': C['demon'], 'Standard': C['standard']}.get(pick_type, C['white'])
        dir_bg = C['over'] if direction == 'OVER' else (C['under'] if direction == 'UNDER' else pt_bg)
        bg = dir_bg

        if direction in ('OVER', 'UNDER'):
            sub2 = sub[sub[d].str.upper() == direction]
        else:
            sub2 = sub

        # Direct count for all pick types (hit_rate() excludes Demons; matrix should show them)
        decided_sub = sub2[sub2['result'].isin(['HIT', 'MISS'])]
        dec_v = len(decided_sub)
        h_v = int((decided_sub['result'] == 'HIT').sum())
        m_v = int((decided_sub['result'] == 'MISS').sum())
        hr_v = h_v / dec_v if dec_v > 0 else np.nan
        total_v = len(sub2)

        dc(ws, ri, 1, pick_type, bg=bg, bold=True, align='left')
        dc(ws, ri, 2, f'Tier {tier}', bg=bg, bold=True)
        dc(ws, ri, 3, direction, bg=bg, bold=True)
        dc(ws, ri, 4, total_v, bg=bg)
        dc(ws, ri, 5, dec_v, bg=bg)

        if dec_v == 0:
            dc(ws, ri, 6, '—', bg='DDDDDD')
            dc(ws, ri, 7, '—', bg='DDDDDD')
            dc(ws, ri, 8, '—', bg='DDDDDD')
        else:
            dc(ws, ri, 6, h_v, bg=bg)
            dc(ws, ri, 7, m_v, bg=bg)
            pct_cell(ws, ri, 8, hr_v)
        return ri + 1

    ri = 2
    # Header row for Standard
    if 'pick_type' in df.columns:
        hc(ws, ri, 1, '⭐ STANDARD', bg=C['hdr2'])
        for ci in range(2, 9): dc(ws, ri, ci, '', bg=C['hdr2'])
        ri += 1
        std = df[df['pick_type'] == 'Standard'] if 'pick_type' in df.columns else df
        for t in TIER_ORDER:
            tsub = std[std['tier'].astype(str).str.upper() == t] if 'tier' in std.columns else std
            for direction in ('OVER', 'UNDER'):
                ri = _matrix_row(ri, 'Standard', t, direction, tsub)

        ri += 1  # spacer
        hc(ws, ri, 1, '🎃 GOBLIN', bg=C['hdr3'])
        for ci in range(2, 9): dc(ws, ri, ci, '', bg=C['hdr3'])
        ri += 1
        gob = df[df['pick_type'] == 'Goblin'] if 'pick_type' in df.columns else pd.DataFrame()
        for t in TIER_ORDER:
            tsub = gob[gob['tier'].astype(str).str.upper() == t] if 'tier' in gob.columns else gob
            ri = _matrix_row(ri, 'Goblin', t, 'OVER', tsub)

        ri += 1  # spacer
        hc(ws, ri, 1, '😈 DEMON', bg=C['hdr5'])
        for ci in range(2, 9): dc(ws, ri, ci, '', bg=C['hdr5'])
        ri += 1
        dem = df[df['pick_type'] == 'Demon'] if 'pick_type' in df.columns else pd.DataFrame()
        for t in TIER_ORDER:
            tsub = dem[dem['tier'].astype(str).str.upper() == t] if 'tier' in dem.columns else dem
            ri = _matrix_row(ri, 'Demon', t, 'OVER', tsub)

    # Total row (direct count — includes all pick types)
    ri += 1
    dec_df = df[df['result'].isin(['HIT', 'MISS'])]
    dec_tot = len(dec_df)
    h_tot = int((dec_df['result'] == 'HIT').sum())
    m_tot = int((dec_df['result'] == 'MISS').sum())
    hr_tot = h_tot / dec_tot if dec_tot > 0 else np.nan
    hc(ws, ri, 1, 'TOTAL', bg=C['hdr'])
    dc(ws, ri, 2, '', bg=C['hdr'])
    dc(ws, ri, 3, 'ALL', bg=C['hdr'], bold=True)
    dc(ws, ri, 4, len(df), bg=C['hdr'], bold=True)
    dc(ws, ri, 5, dec_tot, bg=C['hdr'], bold=True)
    dc(ws, ri, 6, h_tot, bg=C['hdr'], bold=True)
    dc(ws, ri, 7, m_tot, bg=C['hdr'], bold=True)
    pct_cell(ws, ri, 8, hr_tot)


# ── Def Tier × Performance cross-tab sheet ────────────────────────────────────
def write_def_tier_crosstab(wb, df):
    """Cross-tab: Def Tier rows vs (Pick Type × Rank Tier) columns.
    Column groups: Goblin OVER A-D | Demon OVER A-D | Std OVER A-D | Std UNDER A-D
    Each cell: "XX% (n)" where n=decided, colored by hit rate. '—' when no decided.
    """
    if 'def_tier' not in df.columns or 'tier' not in df.columns:
        return
    d = drc(df)

    ws = wb.create_sheet('Def Tier x Performance')

    # Column definitions: (group_label, pick_type, tier, direction)
    GROUPS = [
        ('Goblin OVER', 'Goblin', None, 'OVER'),
        ('Demon OVER',  'Demon',  None, 'OVER'),
        ('Std OVER',    'Standard', None, 'OVER'),
        ('Std UNDER',   'Standard', None, 'UNDER'),
    ]
    GROUP_BGS = [C['goblin'], C['demon'], C['over'], C['under']]
    G_BGS = [C['hdr3'], C['hdr5'], C['hdr2'], C['hdr4']]

    # col layout: col1 = Def Tier label; then 4 groups × 4 tiers = 16 data cols
    col1_w = 14
    data_w = 13
    num_groups = len(GROUPS)
    tiers = TIER_ORDER

    # Set column widths
    ws.column_dimensions['A'].width = col1_w
    for ci in range(2, 2 + num_groups * len(tiers)):
        ws.column_dimensions[get_column_letter(ci)].width = data_w

    # Row 1: Group headers (merged across 4 tiers each)
    ws.merge_cells(start_row=1, start_column=1, end_row=2, end_column=1)
    c = ws.cell(row=1, column=1, value='Def Tier')
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=C['hdr'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    c.border = bdr()
    ws.row_dimensions[1].height = 18
    ws.row_dimensions[2].height = 18

    for gi, (glabel, pt, _, direction) in enumerate(GROUPS):
        start_col = 2 + gi * len(tiers)
        end_col = start_col + len(tiers) - 1
        ws.merge_cells(start_row=1, start_column=start_col, end_row=1, end_column=end_col)
        c = ws.cell(row=1, column=start_col, value=glabel)
        c.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
        c.fill = PatternFill('solid', start_color=G_BGS[gi])
        c.alignment = Alignment(horizontal='center', vertical='center')
        c.border = bdr()

    # Row 2: Tier sub-headers (skip col 1 — it's part of the A1:A2 merge)
    for gi in range(num_groups):
        for ti, t in enumerate(tiers):
            col = 2 + gi * len(tiers) + ti
            c = ws.cell(row=2, column=col, value=f'Tier {t}')
            c.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
            c.fill = PatternFill('solid', start_color=G_BGS[gi])
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border = bdr()

    ws.freeze_panes = 'A3'

    def _cell_val(sub, pt, direction, tier):
        """Return (hr, decided) for the given filter intersection. Direct count — includes Demon."""
        s = sub
        if pt and 'pick_type' in s.columns:
            s = s[s['pick_type'] == pt]
        if direction and d in s.columns:
            s = s[s[d].str.upper() == direction]
        if tier and 'tier' in s.columns:
            s = s[s['tier'].astype(str).str.upper() == tier]
        dec_s = s[s['result'].isin(['HIT', 'MISS'])]
        dec_v = len(dec_s)
        h_v = int((dec_s['result'] == 'HIT').sum())
        hr_v = h_v / dec_v if dec_v > 0 else np.nan
        return hr_v, dec_v

    def _fmt(hr_v, dec_v):
        if dec_v == 0:
            return '—', 'DDDDDD'
        pct_str = f"{hr_v:.0%}" if not (isinstance(hr_v, float) and np.isnan(hr_v)) else '?'
        return f"{pct_str}\n({dec_v})", hr_bg(hr_v)

    # Data rows
    all_def_tiers = [t for t in DEF_TIER_ORDER if t in df['def_tier'].dropna().unique()]
    other_tiers   = [t for t in df['def_tier'].dropna().unique() if t not in DEF_TIER_ORDER]
    row_tiers = all_def_tiers + other_tiers

    for ri_off, dt in enumerate(row_tiers):
        ri = 3 + ri_off
        ws.row_dimensions[ri].height = 28
        sub = df[df['def_tier'] == dt]
        row_bg = C['alt'] if ri_off % 2 == 0 else C['white']
        c = ws.cell(row=ri, column=1, value=dt)
        c.font = Font(bold=True, name='Arial', size=9)
        c.fill = PatternFill('solid', start_color=row_bg)
        c.alignment = Alignment(horizontal='left', vertical='center')
        c.border = bdr()

        for gi, (glabel, pt, _, direction) in enumerate(GROUPS):
            for ti, t in enumerate(tiers):
                col = 2 + gi * len(tiers) + ti
                hr_v, dec_v = _cell_val(sub, pt, direction, t)
                val, bg = _fmt(hr_v, dec_v)
                cell = ws.cell(row=ri, column=col, value=val)
                cell.font = Font(bold=(dec_v > 0), name='Arial', size=9,
                                 color='FFFFFF' if dec_v > 0 else '999999')
                cell.fill = PatternFill('solid', start_color=bg)
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = bdr()

    # Total row
    ri_tot = 3 + len(row_tiers)
    ws.row_dimensions[ri_tot].height = 28
    c = ws.cell(row=ri_tot, column=1, value='Total')
    c.font = Font(bold=True, color='FFFFFF', name='Arial', size=9)
    c.fill = PatternFill('solid', start_color=C['hdr'])
    c.alignment = Alignment(horizontal='left', vertical='center')
    c.border = bdr()

    for gi, (glabel, pt, _, direction) in enumerate(GROUPS):
        for ti, t in enumerate(tiers):
            col = 2 + gi * len(tiers) + ti
            hr_v, dec_v = _cell_val(df, pt, direction, t)
            val, bg = _fmt(hr_v, dec_v)
            cell = ws.cell(row=ri_tot, column=col, value=val)
            cell.font = Font(bold=True, name='Arial', size=9, color='FFFFFF' if dec_v > 0 else '999999')
            cell.fill = PatternFill('solid', start_color=bg)
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
            cell.border = bdr()


# ── Box Raw sheet ─────────────────────────────────────────────────────────────
def write_raw(wb,df):
    ws=wb.create_sheet('Box Raw')
    desired=['pp_projection_id','ticket_id','player','team','opp_team','prop_type_norm','pick_type','line',
             'bet_direction','tier','def_tier','minutes_tier','shot_role','usage_role',
             'consistency_grade','team_top3_rank','team_bottom3_rank','def_boost_hist',
             'top3_weak_overperformer','top3_elite_fader','top3_def_context','top3_under_context',
             'edge','abs_edge','last5_hit_rate','hit_rate','last5_avg','season_avg',
             'last5_over','last5_under','l5_over','l5_under','strat_hit_rate','strat_n',
             'projection','rank_score','ml_prob','ml_edge',
             'edge_score','blended_score',
             'actual','result','margin','void_reason_grade']
    cols=[c for c in desired if c in df.columns]
    widths={'pp_projection_id':14,'ticket_id':28,'player':22,'team':6,'opp_team':6,'prop_type_norm':20,'pick_type':10,
            'line':7,'bet_direction':10,'tier':5,'def_tier':10,'minutes_tier':12,
            'shot_role':10,'usage_role':10,'edge':8,'abs_edge':8,
            'last5_hit_rate':13,'last5_avg':10,'season_avg':12,
            'last5_over':9,'last5_under':10,'projection':12,'rank_score':12,
            'ml_prob':10,'ml_edge':10,'edge_score':11,'blended_score':12,
            'actual':9,'result':8,'margin':8,'void_reason_grade':22}
    for ci,col in enumerate(cols,1):
        ws.column_dimensions[get_column_letter(ci)].width=widths.get(col,12)
        hc(ws,1,ci,col,bg=C['hdr'])
    ws.row_dimensions[1].height=20; ws.freeze_panes='A2'
    for ri,row in enumerate(df[cols].itertuples(),2):
        bg=C['alt'] if ri%2==0 else C['white']
        res=str(getattr(row,'result','')).upper()
        for ci,col in enumerate(cols,1):
            val=getattr(row,col,'')
            c_bg=res_bg(res) if col=='result' else (tier_bg(val) if col=='tier' else bg)
            cell=dc(ws,ri,ci,val,bg=c_bg,align='left' if col=='player' else 'center')
            if col=='result' and res in ('HIT','MISS','PUSH','VOID'):
                cell.font=Font(bold=True,name='Arial',size=9,color='FFFFFF')
    ws.auto_filter.ref=f"A1:{get_column_letter(len(cols))}1"

# ── Summary dashboard ─────────────────────────────────────────────────────────
def write_dashboard(wb,df,sport,date_str):
    ws=wb.create_sheet('Summary',0)
    ws.column_dimensions['A'].width=22
    ws.column_dimensions['B'].width=12
    for ci in range(3,10): ws.column_dimensions[get_column_letter(ci)].width=11
    ws.merge_cells('A1:I1')
    c=ws['A1']
    c.value=f"{sport} SLATE GRADE  |  {date_str}  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font=Font(bold=True,name='Arial',size=12,color='FFFFFF')
    c.fill=PatternFill('solid',start_color=C['hdr'])
    c.alignment=Alignment(horizontal='center',vertical='center')
    ws.row_dimensions[1].height=28
    d=drc(df)
    # Headline OVERALL includes all tiers A–D (same rows as Box Raw).
    if 'tier' in df.columns:
        df_headline = df.copy()
        headline_label = 'Full Slate (A–D)'
    else:
        df_headline = df
        headline_label = 'Full Slate'

    def sec_hdr(row,label,color):
        hc(ws,row,1,label,bg=color)
        for ci,h in enumerate(['Direction','Total','Decided','Hits','Misses','Voids','Hit Rate'],2):
            hc(ws,row,ci,h,bg=color)
        return row+1

    def simple_row(row,label,sub,bg=None,bold=True):
        hr2,h2,m2,v2,dec2=hit_rate(sub)
        bg=bg or (C['alt'] if row%2==0 else C['white'])
        dc(ws,row,1,label,bold=bold,align='left')
        dc(ws,row,2,'ALL',bg=bg,bold=True); dc(ws,row,3,len(sub),bg=bg)
        dc(ws,row,4,dec2,bg=bg); dc(ws,row,5,int(h2),bg=bg)
        dc(ws,row,6,int(m2),bg=bg); dc(ws,row,7,int(v2),bg=bg)
        pct_cell(ws,row,8,hr2); return row+1

    def dir_rows(row,sub):
        for direction in ['OVER','UNDER']:
            dsub=sub[sub[d].str.upper()==direction]
            if len(dsub)==0: continue
            hr_d,h_d,m_d,v_d,dec_d=hit_rate(dsub)
            dbg=C['over'] if direction=='OVER' else C['under']
            dc(ws,row,1,'',bg=dbg); dc(ws,row,2,direction,bg=dbg,bold=True)
            dc(ws,row,3,len(dsub),bg=dbg); dc(ws,row,4,dec_d,bg=dbg)
            dc(ws,row,5,int(h_d),bg=dbg); dc(ws,row,6,int(m_d),bg=dbg)
            dc(ws,row,7,int(v_d),bg=dbg); pct_cell(ws,row,8,hr_d); row+=1
        return row

    # OVERALL
    row=2
    hc(ws,row,1,'OVERALL',bg=C['hdr2'])
    for ci,h in enumerate(['Direction','Total Props','Decided','Hits','Misses','Voids','Hit Rate'],2):
        hc(ws,row,ci,h,bg=C['hdr2'])
    row+=1
    hr,hits,misses,voids,decided_n = hit_rate(df_headline)
    dc(ws,row,1,headline_label,bold=True,align='left'); dc(ws,row,2,'ALL')
    dc(ws,row,3,len(df_headline)); dc(ws,row,4,decided_n)
    dc(ws,row,5,int(hits)); dc(ws,row,6,int(misses)); dc(ws,row,7,int(voids))
    pct_cell(ws,row,8,hr); row+=1

    # BY PICK TYPE
    if 'pick_type' in df.columns:
        row+=1; row=sec_hdr(row,'BY PICK TYPE',C['hdr3'])
        for pt in ['Goblin','Demon','Standard']:
            sub=df[df['pick_type']==pt]
            if len(sub)==0: continue
            pt_bg={'Goblin':C['goblin'],'Demon':C['demon'],'Standard':C['standard']}.get(pt)
            row=simple_row(row,pt,sub,bg=pt_bg)
            if pt=='Standard': row=dir_rows(row,sub)

    # BY TIER
    if 'tier' in df.columns:
        row+=1; row=sec_hdr(row,'BY TIER',C['hdr4'])
        for t in TIER_ORDER:
            sub=df[df['tier'].astype(str).str.upper()==t]
            if len(sub)==0: continue
            row=simple_row(row,f'Tier {t}',sub,bg=tier_bg(t))
            row=dir_rows(row,sub)

    # BY DEF TIER
    if 'def_tier' in df.columns:
        row+=1; row=sec_hdr(row,'BY OPP DEF TIER',C['hdr5'])
        for dt in [t for t in DEF_TIER_ORDER if t in df['def_tier'].dropna().unique()]:
            sub=df[df['def_tier']==dt]
            row=simple_row(row,dt,sub); row=dir_rows(row,sub)

    # BY DEF RANK BUCKET (NEW)
    if 'def_rank' in df.columns:
        row+=1; row=sec_hdr(row,'BY OPP DEF RANK BUCKET',C['hdr5'])
        tmp = df.copy()
        tmp['def_rank_bucket'] = tmp['def_rank'].apply(_def_rank_bucket)
        for b in ['01-05','06-10','11-20','21-25','26-30','UNK']:
            sub = tmp[tmp['def_rank_bucket'] == b]
            if len(sub) == 0: 
                continue
            row = simple_row(row, b, sub)
            row = dir_rows(row, sub)

    # BY MINUTES TIER
    if 'minutes_tier' in df.columns:
        row+=1; row=sec_hdr(row,'BY MINUTES TIER',C['hdr6'])
        for mt in [t for t in MINUTES_TIER_ORDER if t in df['minutes_tier'].dropna().unique()]:
            sub=df[df['minutes_tier']==mt]
            row=simple_row(row,mt,sub); row=dir_rows(row,sub)

    # BY PLAYER ROLE
    if 'usage_role' in df.columns:
        row+=1; row=sec_hdr(row,'BY PLAYER ROLE',C['hdr7'])
        for role in [r for r in USAGE_ROLE_ORDER if r in df['usage_role'].dropna().unique()]:
            sub=df[df['usage_role']==role]
            row=simple_row(row,role,sub); row=dir_rows(row,sub)

# ── Tier normalization for historical graded slates ───────────────────────────
def _recompute_standard_tiers_from_directional_ml_prob(df: pd.DataFrame, sport: str) -> pd.DataFrame:
    """
    Recompute Standard tier labels from directional ml_prob when available.

    Older graded slates can carry pre-fix tier assignments where Standard UNDER
    rows were bucketed from non-directional probabilities. This keeps grading UI
    and matrix outputs aligned with current tiering logic.
    """
    if "tier" not in df.columns or "ml_prob" not in df.columns or "pick_type" not in df.columns:
        return df
    dcol = drc(df)
    if dcol not in df.columns:
        return df

    try:
        recalculated = assign_tier_column(df.copy(), sport=str(sport or "").lower())
    except Exception as e:
        print(f"[tier-recompute] WARN: skipped ({type(e).__name__}: {e})")
        return df

    pt = df["pick_type"].astype(str).str.strip().str.lower()
    dr = df[dcol].astype(str).str.strip().str.upper()
    ml = pd.to_numeric(df["ml_prob"], errors="coerce")
    mask = pt.eq("standard") & dr.isin(["OVER", "UNDER"]) & ml.notna()
    if not mask.any():
        return df

    old_tier = df.loc[mask, "tier"].astype(str).str.strip().str.upper()
    new_tier = recalculated.loc[mask].astype(str).str.strip().str.upper()
    changed = int((old_tier != new_tier).sum())
    if changed:
        df.loc[mask, "tier"] = new_tier.values
        print(f"[tier-recompute] Standard tiers updated: {changed} rows")
    return df


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap=argparse.ArgumentParser()
    ap.add_argument('--sport', default='NBA', choices=['NBA', 'CBB', 'WNBA'])
    ap.add_argument('--slate',default='')
    ap.add_argument('--actuals',default='')
    ap.add_argument('--output',default='')
    ap.add_argument('--template',action='store_true')
    ap.add_argument('--date',default=datetime.now().strftime('%Y-%m-%d'))
    args=ap.parse_args()

    if args.template:
        pd.DataFrame(columns=['player','prop_type','actual']).to_csv(f'actuals_{args.sport.lower()}.csv',index=False)
        print(f'Saved actuals_{args.sport.lower()}.csv'); return

    if not args.slate: print('ERROR: --slate required.'); return
    if not args.output: args.output=f'{args.sport.lower()}_graded_{args.date}.xlsx'

    print(f'Loading {args.sport} slate...')
    if args.sport == "CBB":
        df = load_cbb(args.slate)
    else:
        df = load_nba(args.slate, "WNBA" if args.sport == "WNBA" else "NBA")
    if args.sport in ("NBA", "WNBA"):
        df = filter_nba_slate_by_grade_date(df, args.date)
    print(f'  {len(df)} props loaded')

    try:
        tpl = _REPO_ROOT / "ui_runner" / "templates"
        from ticket_leg_index import attach_ticket_ids_to_dataframe  # noqa: WPS433

        df = attach_ticket_ids_to_dataframe(
            df,
            live_json=tpl / "tickets_latest.json",
            shadow_json=tpl / "shadow_tickets_latest.json",
        )
    except Exception as exc:
        if "ticket_id" not in df.columns:
            df["ticket_id"] = None
        print(f"  [ticket_id] attach skipped: {exc}")

    if args.actuals:
        print(f'Applying actuals from {args.actuals}...')
        df=apply_actuals(df,args.actuals)
        decided=df[df['result'].isin(['HIT','MISS'])]
        hits=(decided['result']=='HIT').sum()
        print(f'  Graded: {len(decided)} decided — {hits} HIT / {len(decided)-hits} MISS')
    else:
        df['result']='PENDING'; df['void_reason_grade']=''; df['margin']=np.nan
        df['actual']=np.nan; df['result_sign']=0
        print('  No actuals — PENDING slate')

    df = _recompute_standard_tiers_from_directional_ml_prob(df, args.sport)

    if "pp_projection_id" not in df.columns and "projection_id" in df.columns:
        df["pp_projection_id"] = df["projection_id"]

    wb=Workbook(); wb.remove(wb.active)
    write_dashboard(wb,df,args.sport,args.date)
    write_raw(wb,df)
    write_pick_type_sheet(wb,df)
    write_tier_sheet(wb,df)
    write_prop_direction_sheet(wb,df)
    write_flat_breakdown(wb,breakdown(df,drc(df)),'By Direction',drc(df),C['hdr5'])
    if 'abs_edge_bucket' in df.columns:
        write_flat_breakdown(wb,breakdown(df,'abs_edge_bucket'),'By Edge Bucket','abs_edge_bucket',C['hdr'])
    write_tier_dir_sheet(wb,df,'By Minutes Tier','minutes_tier',MINUTES_TIER_ORDER,C['hdr6'])
    write_tier_dir_sheet(wb,df,'By Def Tier','def_tier',DEF_TIER_ORDER,C['hdr5'])
    write_def_rank_bucket_sheet(wb, df)
    write_tier_dir_sheet(wb,df,'By Player Role','usage_role',USAGE_ROLE_ORDER,C['hdr7'])
    write_performance_matrix(wb, df)
    write_def_tier_crosstab(wb, df)
    write_tier_dir_sheet(wb,df,'By Shot Role','shot_role',SHOT_ROLE_ORDER,C['hdr8'])

    vr_col='void_reason' if 'void_reason' in df.columns else 'void_reason_grade'
    vr_df=df[df[vr_col].notna()&(df[vr_col].astype(str).str.strip()!='')]
    if len(vr_df): write_flat_breakdown(wb,breakdown(vr_df,vr_col),'Void Reasons',vr_col,C['hdr'])

    wb.save(args.output)
    print(f'\nSaved -> {args.output}')
    print('Sheets:',wb.sheetnames)

if __name__=='__main__':
    main()