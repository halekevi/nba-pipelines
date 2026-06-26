"""
build_grades_html.py
====================
Reads  nba_graded_{date}.xlsx  and/or  cbb_graded_{date}.xlsx
produced by run_grader.ps1 and emits a styled slate_eval_{date}.html
into the same folder as this script (or wherever --out points).

Usage
-----
  # auto-detect yesterday's files:
  py -3.14 build_grades_html.py

  # explicit date:
  py -3.14 build_grades_html.py --date 2026-02-26

  # explicit file paths:
  py -3.14 build_grades_html.py --nba path\\to\\nba_graded_2026-02-26.xlsx
  py -3.14 build_grades_html.py --nba path\\nba.xlsx --cbb path\\cbb.xlsx

  # custom output location:
  py -3.14 build_grades_html.py --date 2026-02-26 --out ui_runner\\templates\\
"""

from __future__ import annotations

import argparse
import html as html_lib
import json
import numpy as np
import re
import sys
from collections import defaultdict
from datetime import datetime, timedelta
from pathlib import Path
from typing import Any

PROJECT_ROOT = Path(__file__).resolve().parents[2]
SCRIPTS_DIR = Path(__file__).resolve().parent.parent
if str(SCRIPTS_DIR) not in sys.path:
    sys.path.insert(0, str(SCRIPTS_DIR))
if str(PROJECT_ROOT) not in sys.path:
    sys.path.insert(0, str(PROJECT_ROOT))

from grading.leg_grade_utils import is_unplayable_for_grading  # noqa: E402

import pandas as pd
from utils.group_rank_tier import apply_prop_tier_adjustments, assign_tier_column

# ── openpyxl guard ────────────────────────────────────────────────────────────
try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed.  Run:  pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).resolve().parent
OUTPUTS_DIR = SCRIPT_DIR / "outputs"
ROOT_DIR    = SCRIPT_DIR.parent.parent  # scripts/grading -> scripts -> project root
# Two levels up from scripts/grading → project root → outputs/{date}
ROOT_DIR    = SCRIPT_DIR.parent.parent


# ══════════════════════════════════════════════════════════════════════════════
#  UTILITY
# ══════════════════════════════════════════════════════════════════════════════

def h(v: Any) -> str:
    return html_lib.escape(str(v) if v is not None else "")

def pct(v: Any) -> str:
    """Return '55.2%' from 0.552 or 55.2 or '55.2%'."""
    if v is None: return "—"
    s = str(v).strip().rstrip("%")
    try:
        f = float(s)
        if f <= 1.0: f *= 100
        return f"{f:.1f}%"
    except (ValueError, TypeError):
        return str(v)

def pct_f(v: Any) -> float:
    """Return float 0-100."""
    if v is None: return 0.0
    s = str(v).strip().rstrip("%")
    try:
        f = float(s)
        return f * 100 if f <= 1.0 else f
    except (ValueError, TypeError):
        return 0.0

def rate_color(f: float) -> str:
    """CSS color variable for a hit-rate float 0-100."""
    if f >= 65: return "var(--green)"
    if f >= 58: return "#5ef598"
    if f >= 50: return "var(--gold)"
    return "var(--red)"

def _stacked_bar_inner(hit_pct: float, miss_pct: float) -> str:
    """Green = hits share, red = misses share; always spans 100% when either segment exists."""
    parts: list[str] = []
    if hit_pct > 0:
        parts.append(
            f'<div class="rate-bar-fill rate-bar-fill--hit" '
            f'style="width:{hit_pct:.1f}%;background:var(--green)"></div>'
        )
    if miss_pct > 0:
        parts.append(
            f'<div class="rate-bar-fill rate-bar-fill--miss" '
            f'style="width:{miss_pct:.1f}%;background:var(--red)"></div>'
        )
    if not parts:
        return '<div class="rate-bar-fill rate-bar-fill--empty" style="width:100%"></div>'
    return "".join(parts)


def rate_bar_html(f: float, *, hits: int | None = None, misses: int | None = None) -> str:
    decided = 0
    if hits is not None and misses is not None:
        decided = int(hits) + int(misses)
        if decided > 0:
            hit_pct = int(hits) / decided * 100.0
            miss_pct = int(misses) / decided * 100.0
            f = hit_pct
            inner = _stacked_bar_inner(hit_pct, miss_pct)
        else:
            inner = _stacked_bar_inner(0.0, 0.0)
    else:
        hit_pct = min(max(float(f), 0.0), 100.0)
        col = rate_color(f)
        inner = f'<div class="rate-bar-fill" style="width:{hit_pct:.1f}%;background:{col}"></div>'
    txt = "var(--red)" if f < 50 else ("var(--gold)" if f < 58 else "var(--green)")
    return (
        f'<div class="rate-cell">'
        f'<div class="rate-bar-bg">{inner}</div>'
        f'<div class="rate-num" style="color:{txt}">{f:.1f}%</div>'
        f'</div>'
    )


def stacked_rate_bar_bg(hits: int, misses: int) -> str:
    decided = int(hits) + int(misses)
    if decided <= 0:
        inner = '<div class="rate-bar-fill rate-bar-fill--empty" style="width:100%"></div>'
    else:
        inner = _stacked_bar_inner(
            int(hits) / decided * 100.0,
            int(misses) / decided * 100.0,
        )
    return f'<div class="rate-bar-bg">{inner}</div>'

def safe_int(v: Any) -> int:
    try: return int(float(str(v).replace(",","")))
    except: return 0

def fmt_num(n: int) -> str:
    return f"{n:,}"


# ══════════════════════════════════════════════════════════════════════════════
#  XLSX READING
# ══════════════════════════════════════════════════════════════════════════════

def read_sheet(ws) -> list[dict]:
    """Read first sheet with header row → list of dicts."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [str(c).strip() if c is not None else f"_c{i}" for i, c in enumerate(rows[0])]
    out = []
    for row in rows[1:]:
        if any(v is not None for v in row):
            out.append(dict(zip(headers, row)))
    return out


def _sheet_header_has_margin(ws) -> bool:
    row0 = next(ws.iter_rows(min_row=1, max_row=1, values_only=True), None)
    if not row0:
        return False
    return any(str(c or "").strip().lower() == "margin" for c in row0)


# Rollup / dashboard sheets from slate_grader — not per-prop rows (loading them yields 0-prop sport sections).
_ROLLUP_SHEET_NAMES = frozenset({
    "summary",
    "by pick type",
    "by tier",
    "prop type x direction",
    "by direction",
    "by minutes tier",
    "by def tier",
    "by def rank",
    "by player role",
    "by opp def tier",
    "performance matrix",
    "def tier x performance",
    "by shot role",
})


def _is_rollup_sheet(name: str) -> bool:
    return str(name or "").strip().lower() in _ROLLUP_SHEET_NAMES


def _is_prop_level_row(r: dict) -> bool:
    """True when a row is a graded prop (not a Summary / Performance Matrix aggregate line)."""
    player = _cell_str(r.get("player") or r.get("Player"))
    if player:
        low = player.lower()
        if any(
            tok in low
            for tok in (
                "by pick",
                "by tier",
                "by direction",
                "by def",
                "by opp",
                "by shot",
                "by minutes",
                "overall",
                "full slate",
                "performance matrix",
                "slate grade",
            )
        ):
            return False
        return True
    prop = _cell_str(
        r.get("Prop Type")
        or r.get("prop_type_norm")
        or r.get("prop_type")
        or r.get("Prop")
    )
    if not prop:
        return False
    if _cell_str(r.get("Result") or r.get("result") or r.get("Grade") or r.get("grade")):
        return True
    if r.get("actual") is not None and str(r.get("actual")).strip() not in ("", "nan"):
        return True
    return False


def _filter_prop_level_rows(rows: list[dict]) -> list[dict]:
    out: list[dict] = []
    for r in rows:
        if not _is_prop_level_row(r):
            continue
        vr = r.get("void_reason") or r.get("Void Reason") or r.get("void_reason_grade")
        if is_unplayable_for_grading(vr):
            continue
        out.append(r)
    return out


def load_graded(path: Path, sport: str = "") -> list[dict]:
    """
    Load graded workbook rows for Prop Evaluation / HTML.

    When **Box Raw** exists and includes a **margin** column (NBA/MLB slate_grader
    layout), load only that sheet so per-prop margin and actuals are present.
    Otherwise load every sheet (NHL/Soccer/simple workbooks).
    """
    wb = openpyxl.load_workbook(path, read_only=True, data_only=True)
    rows: list[dict] = []
    try:
        if "Box Raw" in wb.sheetnames and _sheet_header_has_margin(wb["Box Raw"]):
            br_rows = read_sheet(wb["Box Raw"])
            if len(br_rows) > 0:
                rows = br_rows
            # Empty Box Raw: no per-prop grades yet — do not load Summary/rollup sheets (they show as 0-prop NBA).
        else:
            # Tennis (and similar) writes the same rows to **graded** and **Box Raw** with no margin column.
            # Loading both doubles counts and breaks Slate Evaluation rollups.
            if (
                "graded" in wb.sheetnames
                and "Box Raw" in wb.sheetnames
                and not _sheet_header_has_margin(wb["Box Raw"])
            ):
                g_rows = read_sheet(wb["graded"])
                br_rows = read_sheet(wb["Box Raw"])
                if len(g_rows) == len(br_rows) and len(g_rows) > 0:
                    rows = g_rows
                else:
                    for shname in wb.sheetnames:
                        if _is_rollup_sheet(shname):
                            continue
                        rows.extend(read_sheet(wb[shname]))
            else:
                for shname in wb.sheetnames:
                    if _is_rollup_sheet(shname):
                        continue
                    rows.extend(read_sheet(wb[shname]))
    finally:
        wb.close()
    # Normalize common column name variants
    normalized = []
    for r in rows:
        nr: dict[str, Any] = {}
        for k, v in r.items():
            nk = k.strip()
            # common aliases
            nk = re.sub(r"pick[\s_-]*type", "Pick Type", nk, flags=re.I)
            nk = re.sub(r"\bhit[\s_]rate\b", "Hit Rate", nk, flags=re.I)
            nk = re.sub(r"\bdef[\s_]?tier\b", "Def Tier", nk, flags=re.I)
            nk = re.sub(r"\bprop[\s_]type\b", "Prop Type", nk, flags=re.I)
            nk = re.sub(r"\bmin(utes)?\b", "Minutes", nk, flags=re.I)
            # Normalize result/grade column
            if nk.lower() == "result":              nk = "Result"
            if nk.lower() == "grade":               nk = "Result"
            if nk.lower() == "direction":           nk = "Direction"
            if nk.lower() == "void_reason_grade":   nk = "void_reason"
            if nk.lower() == "bet_direction":       nk = "Direction"
            if nk.lower() == "prop_type_norm":      nk = "Prop Type"
            if nk.lower() == "pick_type":           nk = "Pick Type"
            if nk.lower() == "tier":                nk = "Tier"
            if nk.lower() in ("ml_prob", "ml prob"): nk = "ML Prob"
            if nk.lower() in ("deviation level", "dev level", "deviation_level"): nk = "deviation_level"
            nr[nk] = v
        normalized.append(nr)
    # Recompute display tiers for historical files:
    # 1) Standard rows from directional ml_prob
    # 2) Prop-type targeted adjustments from existing tier labels
    if normalized:
        sport_key = str(sport or "").strip().lower()
        ndf = pd.DataFrame(normalized)
        pick_series = ndf.get("Pick Type", ndf.get("pick_type", pd.Series("Standard", index=ndf.index)))
        dir_series = ndf.get("Direction", ndf.get("Dir", ndf.get("direction", pd.Series("OVER", index=ndf.index))))
        ml_series = pd.to_numeric(ndf.get("ML Prob", ndf.get("ml_prob", pd.Series(np.nan, index=ndf.index))), errors="coerce")
        line_series = pd.to_numeric(ndf.get("Line", ndf.get("line", ndf.get("line_score", pd.Series(np.nan, index=ndf.index)))), errors="coerce")
        std_line_series = pd.to_numeric(ndf.get("Standard Line", ndf.get("standard_line", pd.Series(np.nan, index=ndf.index))), errors="coerce")
        prop_series = ndf.get(
            "Prop Type",
            ndf.get("prop_type_norm", ndf.get("prop_type", ndf.get("Prop", pd.Series("", index=ndf.index)))),
        )
        work = pd.DataFrame(
            {
                "pick_type": pick_series,
                "bet_direction": dir_series,
                "ml_prob": ml_series,
                "line": line_series,
                "standard_line": std_line_series,
                "prop_type_norm": prop_series,
            }
        )
        tier_series = ndf.get("Tier", pd.Series("D", index=ndf.index)).astype(str)

        # Standard recompute only when ml_prob exists (keeps legacy Goblin/Demon
        # distance-based tiers intact in historical graded workbooks).
        std_mask = pick_series.astype(str).str.strip().str.lower().eq("standard") & ml_series.notna()
        if std_mask.any():
            try:
                recalculated = assign_tier_column(work, sport=sport_key).astype(str)
                tier_series.loc[std_mask] = recalculated.loc[std_mask].values
            except Exception as e:
                print(f"  WARN: standard tier recompute skipped for {path.name} ({type(e).__name__}: {e})")

        # Apply targeted prop-type tier shifts on top of existing/recomputed tiers.
        try:
            tier_series = apply_prop_tier_adjustments(
                tier_series,
                pick_series.astype(str),
                dir_series.astype(str),
                prop_series.astype(str),
                sport=sport_key,
            )
        except Exception as e:
            print(f"  WARN: prop tier adjustments skipped for {path.name} ({type(e).__name__}: {e})")

        ndf["Tier"] = tier_series.astype(str).values
        # Pick Type was only used for tier recompute; matrix / breakdowns need it on each row.
        if sport_key == "soccer":
            def _soccer_pick_type_row(row: pd.Series) -> str:
                raw = str(row.get("Pick Type", "") or "").strip().lower()
                if raw in ("goblin", "demon"):
                    return raw.title()
                return _derive_soccer_pick_type(
                    str(row.get("Tier", "") or ""),
                    str(row.get("Edge", row.get("edge", "")) or ""),
                    str(row.get("ML Prob", row.get("ml_prob", "")) or ""),
                    str(row.get("Hit Rate", row.get("hit_rate", "")) or ""),
                    str(row.get("Blended Score", row.get("blended_score", "")) or ""),
                )

            ndf["Pick Type"] = ndf.apply(_soccer_pick_type_row, axis=1)
        else:
            ps = pick_series.astype(str).str.strip()
            ps = ps.replace({"": "Standard", "nan": "Standard", "None": "Standard"})
            ndf["Pick Type"] = ps.values
        normalized = ndf.to_dict(orient="records")
    return _filter_prop_level_rows(normalized)



# ══════════════════════════════════════════════════════════════════════════════
#  GRADED PROPS JSON (Grades UI /api/graded-props)
# ══════════════════════════════════════════════════════════════════════════════

def _cell_str(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.lower() in ("none", "nan", ""):
        return ""
    return s


def _derive_soccer_pick_type(
    tier: str,
    edge_s: str,
    ml_prob_s: str,
    hit_rate_s: str,
    blended_s: str,
) -> str:
    """
    When Soccer step8 omits pick_type, infer Goblin / Standard / Demon from tier + scores
    (same spirit as scripts.nhl_soccer_grader.load_slate).
    """
    tier_u = str(tier or "").strip().upper()

    def _to_float(v) -> float | None:
        if v is None:
            return None
        s = str(v).strip().replace("%", "")
        if s.lower() in ("", "nan", "none"):
            return None
        try:
            x = float(s)
            return x / 100.0 if x > 1.0 else x
        except (TypeError, ValueError):
            return None

    edge = _to_float(edge_s)
    ml_p = _to_float(ml_prob_s)
    hr = _to_float(hit_rate_s)
    bs = _to_float(blended_s)
    p_candidates = [x for x in (ml_p, hr, bs) if x is not None]
    if p_candidates:
        p = max(0.01, min(0.99, float(p_candidates[0])))
        if p >= 0.70:
            return "Goblin"
        if p >= 0.55:
            return "Standard"
    if tier_u == "A" and edge is not None and edge >= 0.48:
        return "Goblin"
    if tier_u in ("A", "B"):
        return "Standard"
    return "Demon"


def _norm_result_display(row: dict) -> str:
    r = _cell_str(row.get("Result") or row.get("Grade")).upper()
    if r in ("HIT", "WIN", "1", "TRUE", "YES", "W"):
        return "HIT"
    if r in ("MISS", "LOSS", "0", "FALSE", "NO", "L"):
        return "MISS"
    if r in ("VOID", "PUSH", "N/A"):
        return r
    if r:
        return r
    return "—"


def prop_row_for_api(
    row: dict,
    sport: str,
    *,
    live_ticket_keys: set[tuple[str, ...]] | None = None,
    shadow_ticket_keys: set[tuple[str, ...]] | None = None,
    live_id_map: dict[tuple[str, ...], str] | None = None,
    shadow_id_map: dict[tuple[str, ...], str] | None = None,
) -> dict[str, str] | None:
    """One flat dict per prop row for the Prop Evaluation tab."""
    if is_unplayable_for_grading(
        row.get("void_reason") or row.get("Void Reason") or row.get("void_reason_grade")
    ):
        return None

    def _pick(*keys: str) -> str:
        for k in keys:
            v = _cell_str(row.get(k))
            if v:
                return v
        return ""

    def _pick_scalar(*keys: str) -> str:
        for k in keys:
            v = row.get(k)
            if v is None:
                continue
            if isinstance(v, float) and pd.isna(v):
                continue
            if isinstance(v, bool):
                return "1" if v else "0"
            if isinstance(v, (int, float)):
                return str(v)
            s = _cell_str(v)
            if s:
                return s
        return ""

    player = _pick("Player", "player", "Name", "name")
    if not player:
        return None
    team = _pick("Team", "team", "team_abbr")
    prop = _pick("Prop Type", "prop type", "prop_type_norm", "prop_type", "Prop", "prop", "Pick", "pick")
    direction = _pick(
        "Dir",
        "dir",
        "Direction",
        "direction",
        "bet_direction",
        "final_bet_direction",
        "recommended_side",
    ).upper()
    line = _pick("Line", "line", "Game Line", "game line", "O/U", "OU Line", "Pick Line")
    pick_type = _pick("Pick Type", "pick type", "pick_type")
    tier = _pick("Tier", "tier")
    hit_rate_for_pick = _pick(
        "Hit Rate",
        "hit_rate",
        "Hit Rate (5g)",
        "Hit Rate (10g)",
        "composite_hr",
    )
    blended_for_pick = _pick("Blended Score", "blended_score", "Blended")
    # slate_grader Box Raw uses snake_case opp_team; legacy sheets use Opp / Opp Team.
    opp_team = _pick(
        "opp_team",
        "Opp Team",
        "opp team",
        "Opp",
        "opp",
        "Opponent",
        "opponent",
        "opp_team_abbr",
        "pp_opp_team",
    )
    edge = _pick("Edge", "edge", "Edge Score", "edge score")
    ml_prob = _pick("ML Prob", "ml prob", "ml_prob")
    actual_value = _pick("Actual", "actual", "actual_value")
    margin = _pick("Margin", "margin")
    def_tier = _pick("Def Tier", "def_tier", "Defense Tier", "defense_tier", "Opp Def Tier", "opp_def_tier")
    consistency_grade = _pick("Consistency Grade", "consistency_grade")
    team_top3_rank = _pick("Top3 Rank", "team_top3_rank", "top3_rank")
    team_bottom3_rank = _pick("Bottom3 Rank", "team_bottom3_rank", "bottom3_rank")
    top3_weak_over = _pick("Top3 Weak Over", "top3_weak_overperformer")
    top3_elite_fade = _pick("Top3 Elite Fade", "top3_elite_fader")
    l5_over = _pick("L5 Over", "l5_over", "last5_over", "over_L5_raw")
    l5_under = _pick("L5 Under", "l5_under", "last5_under", "under_L5_raw")
    l10_over = _pick("L10 Over", "l10_over", "line_hits_over_10", "over_L10", "over_L10_raw")
    l10_under = _pick("L10 Under", "l10_under", "line_hits_under_10", "under_L10", "under_L10_raw")
    l10_games_played = _pick("l10_games_played", "line_games_played_10", "Games (10g)", "sample_L10")
    l10_streak = _pick("l10_streak", "L10 Streak")
    strat_hit_rate = _pick("strat_hit_rate")
    strat_n = _pick("strat_n")
    hit_rate_l5 = _pick("hit_rate_l5", "Hit Rate L5")
    hit_rate_l10 = _pick("hit_rate_l10", "Hit Rate L10")
    player_hr_hist = _pick("player_hr_historical", "Player HR Hist")
    opp_hr_hist = _pick("opp_hr_historical", "Opp HR Hist")
    sport_maturity = _pick("sport_signal_maturity", "Sport Maturity")
    confidence_tier = _pick("confidence_tier", "Confidence Tier")
    confidence_score = _pick("confidence_score", "Confidence Score")
    confidence_note = _pick("confidence_note", "Confidence Note")
    h2h_bucket = _pick("H2H Tier", "h2h_tier", "H2H Bucket", "h2h_bucket", "Head To Head Bucket", "head_to_head_bucket")
    minutes_tier = _pick("Minutes Tier", "minutes_tier", "Min Tier", "min_tier", "Minutes Bucket", "minutes_bucket")
    role_tier = _pick("Role Tier", "role_tier", "Player Role", "player_role", "Usage Role", "usage_role", "Team Role", "team_role")
    usage_vacuum = _pick_scalar("usage_vacuum", "Usage Vacuum")
    deviation_level_raw = _pick_scalar("deviation_level", "Deviation Level", "Dev Level")
    deviation_level_val: int | None = None
    if deviation_level_raw:
        try:
            dv = int(float(deviation_level_raw))
            if dv > 0:
                deviation_level_val = dv
        except (ValueError, TypeError):
            pass
    team_star_out = _pick_scalar("team_star_out", "Team Star Out")
    key_facilitator_out = _pick_scalar("key_facilitator_out", "Key Facilitator Out")
    injury_boost_candidate = _pick_scalar("injury_boost_candidate", "Injury Boost Candidate", "Injury Boost")
    usage_pct = _pick_scalar("usage_pct", "Usage %", "Usage Pct")
    usage_tier = _pick("usage_tier", "Usage Tier")
    game_total_bucket = _pick("Game Total Bucket", "game_total_bucket", "OU Bucket", "ou_bucket", "Over Under Bucket", "over_under_bucket", "Total Bucket", "total_bucket")
    game_total = _pick("Game O/U", "game_ou", "Game Total", "game_total", "O/U Total", "ou_total")
    h2h_raw = _pick("H2H", "h2h", "Head To Head", "head_to_head")
    actual_source = _pick("actual_source", "Actual Source", "actual source")
    actual_source_conflict = _pick("actual_source_conflict", "Actual Source Conflict", "actual source conflict")
    void_reason = _pick(
        "void_reason",
        "Void Reason",
        "void reason",
        "reason",
        "Reason",
        "status_note",
        "notes",
    )
    proj_id = _pick("pp_projection_id", "projection_id", "PP Projection ID", "Projection ID")
    result = _norm_result_display(row)
    sport_up = str(sport or "").strip().upper()
    vr_up = str(void_reason or "").strip().upper()
    # Soccer VOID rows must carry explicit data quality tags for validator/reporting.
    if result == "VOID" and sport_up == "SOCCER":
        if "DNP" in vr_up:
            void_reason = "DNP"
        elif "NO_DATA" in vr_up or "NO ACTUAL" in vr_up:
            void_reason = "NO_DATA"
        elif not vr_up:
            void_reason = "NO_DATA"
    # Normalize pick_type labels so downstream analytics don't split by casing/spelling.
    pt_raw = str(pick_type or "").strip().lower()
    if pt_raw in ("goblin",):
        pick_type = "Goblin"
    elif pt_raw in ("demon",):
        pick_type = "Demon"
    elif pt_raw in ("standard", "std"):
        pick_type = "Standard"
    elif pt_raw in ("—", "-", "–", "", "nan", "none", "null"):
        pick_type = "—"
        if sport_up == "SOCCER" and tier:
            inferred = _derive_soccer_pick_type(
                tier, edge, ml_prob, hit_rate_for_pick, blended_for_pick
            )
            if inferred:
                pick_type = inferred

    on_live = False
    on_shadow = False
    ticket_id: str | None = None
    try:
        import sys
        _scripts = ROOT_DIR / "scripts"
        if str(_scripts) not in sys.path:
            sys.path.insert(0, str(_scripts))
        from ticket_leg_index import (  # noqa: WPS433
            prop_matches_ticket_keys,
            resolve_ticket_id_for_row,
        )

        if live_ticket_keys:
            on_live = prop_matches_ticket_keys(row, sport, live_ticket_keys)
        if shadow_ticket_keys:
            on_shadow = prop_matches_ticket_keys(row, sport, shadow_ticket_keys)
        ticket_id = resolve_ticket_id_for_row(
            row,
            sport,
            live_id_map=live_id_map,
            shadow_id_map=shadow_id_map,
        )
    except Exception:
        on_live = on_shadow = False
        ticket_id = None

    hit_val: int | None = None
    if result == "HIT":
        hit_val = 1
    elif result == "MISS":
        hit_val = 0

    return {
        "sport": sport,
        "player": player,
        "team": team or "—",
        "opp_team": opp_team or "",
        "prop": prop or "—",
        "line": line or "—",
        "direction": direction or "—",
        "pick_type": pick_type or "—",
        "tier": tier or "",
        "edge": edge or "",
        "ml_prob": ml_prob or "",
        "actual_value": actual_value or "",
        "margin": margin or "",
        "def_tier": def_tier or "",
        "consistency_grade": consistency_grade or "",
        "team_top3_rank": team_top3_rank or "",
        "team_bottom3_rank": team_bottom3_rank or "",
        "top3_weak_overperformer": top3_weak_over or "",
        "top3_elite_fader": top3_elite_fade or "",
        "l5_over": l5_over or "",
        "l5_under": l5_under or "",
        "l10_over": l10_over or "",
        "l10_under": l10_under or "",
        "l10_games_played": l10_games_played or "",
        "l10_streak": l10_streak or "",
        "hit_rate": hit_rate_for_pick or "",
        "strat_hit_rate": strat_hit_rate or "",
        "strat_n": strat_n or "",
        "hit_rate_l5": hit_rate_l5 or "",
        "hit_rate_l10": hit_rate_l10 or "",
        "player_hr_historical": player_hr_hist or "",
        "opp_hr_historical": opp_hr_hist or "",
        "sport_signal_maturity": sport_maturity or "",
        "confidence_tier": confidence_tier or "",
        "confidence_score": confidence_score or "",
        "confidence_note": confidence_note or "",
        "h2h_bucket": h2h_bucket or "",
        "minutes_tier": minutes_tier or "",
        "role_tier": role_tier or "",
        "usage_vacuum": usage_vacuum,
        "deviation_level": deviation_level_val,
        "team_star_out": team_star_out,
        "key_facilitator_out": key_facilitator_out,
        "injury_boost_candidate": injury_boost_candidate,
        "usage_pct": usage_pct,
        "usage_tier": usage_tier or "",
        "game_total_bucket": game_total_bucket or "",
        "game_total": game_total or "",
        "h2h_raw": h2h_raw or "",
        "actual_source": actual_source or "",
        "actual_source_conflict": actual_source_conflict or "",
        "over_under": direction or "—",
        "void_reason": void_reason or "",
        "result": result,
        "pp_projection_id": proj_id or "",
        "ticket_id": ticket_id,
        "on_ticket": on_live,
        "on_shadow_ticket": on_shadow,
        "hit": hit_val,
        "graded_at": _pick("graded_at", "Graded At", "date", "Date") or "",
    }


def export_graded_props_json(
    date_str: str,
    out_dir: Path,
    bundles: list[tuple[str, list[dict]]],
) -> Path:
    """Write graded_props_{date}.json for the Grades UI.

    Operational note: copying an old graded_props_*.json without re-running
    build_grades_html (after slate_grader) can resurrect rows excluded by
    is_unplayable_for_grading — always regenerate from graded workbooks.
    """
    live_keys: set[tuple[str, ...]] = set()
    shadow_keys: set[tuple[str, ...]] = set()
    live_id_map: dict[tuple[str, ...], str] = {}
    shadow_id_map: dict[tuple[str, ...], str] = {}
    try:
        import sys
        _scripts = ROOT_DIR / "scripts"
        if str(_scripts) not in sys.path:
            sys.path.insert(0, str(_scripts))
        from ticket_leg_index import (  # noqa: WPS433
            load_leg_key_to_ticket_id,
            load_ticket_leg_keys,
        )

        tpl = ROOT_DIR / "ui_runner" / "templates"
        live_keys = load_ticket_leg_keys(tpl / "tickets_latest.json")
        shadow_keys = load_ticket_leg_keys(tpl / "shadow_tickets_latest.json")
        live_id_map = load_leg_key_to_ticket_id(tpl / "tickets_latest.json")
        shadow_id_map = load_leg_key_to_ticket_id(tpl / "shadow_tickets_latest.json")
    except Exception:
        pass

    props: list[dict[str, str]] = []
    for sport, rows in bundles:
        for row in rows:
            p = prop_row_for_api(
                row,
                sport,
                live_ticket_keys=live_keys,
                shadow_ticket_keys=shadow_keys,
                live_id_map=live_id_map,
                shadow_id_map=shadow_id_map,
            )
            if p:
                props.append(p)
    payload = {
        "date": date_str,
        "count": len(props),
        "props": props,
        "prop_breakdown_rows": build_prop_breakdown_rows(props),
    }
    out = out_dir / f"graded_props_{date_str}.json"
    out.write_text(json.dumps(payload, ensure_ascii=False, indent=2), encoding="utf-8")
    return out


def export_analysis_tabs_xlsx(
    date_str: str,
    out_dir: Path,
    rows: list[dict],
) -> Path:
    """
    Export core analysis tables used on Grades page into workbook tabs:
      - Performance Matrix (Pick Type × Tier × Direction)
      - DEF TIER BREAKDOWN (combined grid with % + decided counts)
    """
    wb = openpyxl.Workbook()

    # Sheet 1: Performance Matrix
    ws = wb.active
    ws.title = "Performance Matrix"
    ws.append(["Pick Type", "Tier", "Direction", "Decided", "Hits", "Misses", "Hit Rate %"])

    agg = build_pick_tier_direction_agg(rows)
    for pt, tier, direction in _canonical_pick_tier_direction_keys():
        v = agg.get((pt, tier, direction), {"hits": 0, "misses": 0, "decided": 0})
        d = int(v.get("decided", 0) or 0)
        h_ = int(v.get("hits", 0) or 0)
        m_ = int(v.get("misses", 0) or 0)
        hr = round((h_ / d * 100.0), 1) if d > 0 else None
        ws.append([pt, f"TIER {tier}", direction, d, h_, m_, hr])

    # Sheet 2: DEF TIER BREAKDOWN
    ws2 = wb.create_sheet("DEF TIER BREAKDOWN")
    groups: tuple[tuple[str, str, str], ...] = (
        ("Goblin OVER", "Goblin", "OVER"),
        ("Demon OVER", "Demon", "OVER"),
        ("Std OVER", "Standard", "OVER"),
        ("Std UNDER", "Standard", "UNDER"),
    )
    ranks = ("A", "B", "C", "D")
    canon_defs = ("Elite", "Above Avg", "Avg", "Below Avg", "Weak", "Total")

    hdr: list[str] = ["Def Tier"]
    for glabel, _pick, _direction in groups:
        for rk in ranks:
            hdr.append(f"{glabel} {rk} %")
            hdr.append(f"{glabel} {rk} Decided")
    ws2.append(hdr)

    for def_label in canon_defs:
        canon = None if def_label == "Total" else def_label
        row_out: list[Any] = [def_label]
        for _glabel, pick, direction in groups:
            for rk in ranks:
                cell_rows = _rows_for_def_subgrid_cell(rows, canon, pick, direction, rk)
                st = overall_stats(cell_rows)
                d = int(st.get("decided", 0) or 0)
                hr = round(float(st.get("hit_rate", 0.0) or 0.0), 1) if d > 0 else None
                row_out.append(hr)
                row_out.append(d)
        ws2.append(row_out)

    out = out_dir / f"graded_analysis_tabs_{date_str}.xlsx"
    wb.save(out)
    return out


def _norm_def_label(v: Any) -> str:
    t = _cell_str(v).lower()
    if not t:
        return "avg"
    t = t.replace("🟢", "").replace("🟡", "").replace("🔴", "").strip()
    t = t.replace("_", " ")
    if "elite" in t:
        return "elite"
    if "above" in t:
        return "above_avg"
    if t in ("avg", "average"):
        return "avg"
    if "below" in t:
        return "below_avg"
    if "weak" in t:
        return "weak"
    return "avg"


def _map_pt_label(pick_type: Any, direction: Any) -> str | None:
    pt = _cell_str(pick_type).lower()
    dcan = normalize_bet_direction(direction)
    d = dcan or _cell_str(direction).upper()
    is_over = d == "OVER"
    is_under = d == "UNDER"
    if "goblin" in pt:
        return "Goblin"
    if "demon" in pt:
        return "Demon"
    if "standard" in pt:
        if is_over:
            return "Standard Over"
        if is_under:
            return "Standard Under"
        return None
    if is_over:
        return "Over"
    if is_under:
        return "Under"
    return None


def build_prop_breakdown_rows(props: list[dict[str, str]]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, int]] = defaultdict(lambda: {"decided": 0, "hits": 0})
    for p in props:
        result = _cell_str(p.get("result")).upper()
        if result not in ("HIT", "MISS"):
            continue
        pt_label = _map_pt_label(p.get("pick_type"), p.get("direction") or p.get("over_under"))
        if not pt_label:
            continue
        tier = _cell_str(p.get("tier")).upper()
        if tier not in ("A", "B", "C", "D"):
            continue
        prop = _cell_str(p.get("prop")) or "Unknown"
        def_label = _norm_def_label(p.get("def_tier"))
        key = (prop, pt_label, tier, def_label)
        grouped[key]["decided"] += 1
        if result == "HIT":
            grouped[key]["hits"] += 1
    out: list[dict[str, Any]] = []
    for (prop, pt, tier, def_label), v in grouped.items():
        out.append(
            {
                "prop": prop,
                "pt": pt,
                "tier": tier,
                "def": def_label,
                "decided": int(v["decided"]),
                "hits": int(v["hits"]),
            }
        )
    out.sort(key=lambda x: (str(x["pt"]), str(x["tier"]), str(x["prop"])))
    return out

def agg_rows(rows: list[dict], key_col: str) -> dict[str, dict]:
    """Group rows by key_col and sum Hits, Misses, Decided, Voids."""
    groups: dict[str, dict[str, int]] = defaultdict(lambda: {"hits":0,"misses":0,"decided":0,"voids":0})
    for r in rows:
        k = str(r.get(key_col, "") or "").strip()
        if not k or k.lower() in ("none","nan",""): continue
        result = str(r.get("Result","") or r.get("Grade","") or "").strip().upper()
        void   = result in ("VOID","PUSH","N/A","")
        if void:
            groups[k]["voids"] += 1
        else:
            groups[k]["decided"] += 1
            if result in ("HIT","WIN","1","TRUE","YES","W"):
                groups[k]["hits"] += 1
            elif result in ("MISS","LOSS","0","FALSE","NO","L"):
                groups[k]["misses"] += 1
            else:
                # Season/pick Hit Rate is not a same-day outcome — skip ungraded rows.
                pass
    return dict(groups)


def agg_by_precomputed(rows: list[dict], key_col: str) -> list[dict]:
    """
    When rows already have Decided / Hits / Misses columns (pre-aggregated),
    just group and sum them.
    """
    groups: dict[str, dict] = defaultdict(lambda: {"hits":0,"misses":0,"decided":0,"voids":0})
    for r in rows:
        k = str(r.get(key_col, "") or "").strip()
        if not k or k.lower() in ("none","nan",""): continue
        groups[k]["hits"]    += safe_int(r.get("Hits",   r.get("hits",   0)))
        groups[k]["misses"]  += safe_int(r.get("Misses", r.get("misses", 0)))
        groups[k]["decided"] += safe_int(r.get("Decided",r.get("decided",0)))
        groups[k]["voids"]   += safe_int(r.get("Voids",  r.get("voids",  0)))
    result = []
    for k, v in groups.items():
        d = v["decided"]
        h_ = v["hits"]
        result.append({
            "key": k,
            "hits": h_,
            "misses": v["misses"],
            "decided": d,
            "voids": v["voids"],
            "hit_rate": (h_/d*100) if d > 0 else 0.0
        })
    return result


def build_agg_from_rows(rows: list[dict], key_col: str) -> list[dict]:
    """
    Detect whether rows are pre-aggregated (have Decided/Hits cols) or
    raw (have a Result col). Returns list of aggregated dicts with hit_rate.
    """
    has_decided = any(r.get("Decided") is not None or r.get("decided") is not None for r in rows[:20])
    if has_decided:
        return agg_by_precomputed(rows, key_col)
    else:
        raw = agg_rows(rows, key_col)
        result = []
        for k, v in raw.items():
            d = v["decided"]
            h_ = v["hits"]
            result.append({
                "key": k,
                "hits": h_,
                "misses": v["misses"],
                "decided": d,
                "voids": v["voids"],
                "hit_rate": (h_/d*100) if d > 0 else 0.0
            })
        return result


def overall_stats(rows: list[dict]) -> dict:
    """Compute overall summary from graded rows."""
    total = len(rows)
    agg = build_agg_from_rows(rows, "_all_")  # dummy — we compute directly
    # Direct computation
    hits = misses = voids = 0
    for r in rows:
        result = str(r.get("Result","") or r.get("Grade","") or "").strip().upper()
        if result in ("VOID","PUSH","N/A",""): voids += 1
        elif result in ("HIT","WIN","1","TRUE","YES","W"): hits += 1
        elif result in ("MISS","LOSS","0","FALSE","NO","L"): misses += 1
    decided = hits + misses
    # Fallback: pre-aggregated sheet (rows already have Decided/Hits)
    if decided == 0 and total > 0:
        has_d = any(r.get("Decided") is not None for r in rows[:10])
        if has_d:
            hits    = sum(safe_int(r.get("Hits",   0)) for r in rows)
            misses  = sum(safe_int(r.get("Misses", 0)) for r in rows)
            decided = sum(safe_int(r.get("Decided",0)) for r in rows)
            voids   = sum(safe_int(r.get("Voids",  0)) for r in rows)
            total   = decided + voids
    hit_rate = (hits/decided*100) if decided > 0 else 0.0
    return {"total":total,"hits":hits,"misses":misses,"decided":decided,"voids":voids,"hit_rate":hit_rate}


def tier_a_stats(rows: list[dict]) -> dict:
    tier_rows = [r for r in rows if str(r.get("Tier","") or "").strip().upper() == "A"]
    return overall_stats(tier_rows)

def pick_type_stats(rows: list[dict], pick_type: str) -> dict:
    pt = pick_type.lower()
    filtered = [r for r in rows if pt in str(r.get("Pick Type","") or "").lower()]
    return overall_stats(filtered)


def rows_for_pick_type(sub_rows: list[dict], pick_type: str) -> list[dict]:
    pt = pick_type.lower()
    return [r for r in sub_rows if pt in str(r.get("Pick Type", "") or "").lower()]


def over_under_lines_html(sub_rows: list[dict]) -> str:
    """▲ OVER / ▼ UNDER hit rates for any row subset (pick type bucket, tier bucket, def tier slice, etc.)."""
    over = overall_stats(
        [r for r in sub_rows if normalize_bet_direction(row_bet_direction(r)) == "OVER"]
    )
    under = overall_stats(
        [r for r in sub_rows if normalize_bet_direction(row_bet_direction(r)) == "UNDER"]
    )
    inner = ""
    if over["decided"] > 0:
        col = "var(--green)" if over["hit_rate"] >= 55 else "var(--red)"
        inner += (
            f'<div class="sub-dir"><span style="color:var(--cyan);font-size:13px">▲ OVER</span> '
            f'<span style="color:{col}">{pct(over["hit_rate"])}</span> ({fmt_num(over["decided"])} dec)</div>'
        )
    if under["decided"] > 0:
        col = "var(--green)" if under["hit_rate"] >= 55 else "var(--red)"
        inner += (
            f'<div class="sub-dir"><span style="color:var(--gold);font-size:13px">▼ UNDER</span> '
            f'<span style="color:{col}">{pct(under["hit_rate"])}</span> ({fmt_num(under["decided"])} dec)</div>'
        )
    if not inner:
        return ""
    return (
        f'<div class="ou-breakdown" style="font-size:13px;color:var(--muted2);margin-top:5px;line-height:1.45">{inner}</div>'
    )


def _norm_pick_type_matrix(v: object) -> str | None:
    s = str(v or "").strip().lower()
    if "goblin" in s:
        return "Goblin"
    if "demon" in s:
        return "Demon"
    if "standard" in s:
        return "Standard"
    return None


def normalize_def_tier_label(raw: object) -> str | None:
    """
    Map a Def Tier cell to canonical buckets: Elite, Above Avg, Avg, Below Avg, Weak.
    Returns None if the value cannot be mapped (row excluded from tier rows, still eligible for Total).
    """
    s = (
        str(raw or "")
        .lower()
        .replace("🟢", "")
        .replace("🟡", "")
        .replace("🔴", "")
        .strip()
    )
    if not s:
        return None
    if "below" in s and "avg" in s:
        return "Below Avg"
    if "above" in s and "avg" in s:
        return "Above Avg"
    if "elite" in s:
        return "Elite"
    if "weak" in s:
        return "Weak"
    if s in ("avg", "average"):
        return "Avg"
    return None


def normalize_bet_direction(raw: object) -> str | None:
    t = str(raw or "").strip().upper()
    for ch in ("\u25b2", "\u25bc", "\u2191", "\u2193", "▲", "▼"):
        t = t.replace(ch, "")
    t = re.sub(r"\s+", " ", t).strip()
    if "UNDER" in t or t == "LESS":
        return "UNDER"
    if "OVER" in t or t == "MORE":
        return "OVER"
    return None


def row_bet_direction(r: dict) -> object:
    """Graded workbooks use Dir/Direction (most sports) or direction (tennis)."""
    return r.get("Dir") or r.get("Direction") or r.get("direction")


def _row_line_key(r: dict) -> str:
    for k in ("Line", "line", "line_score", "standard_line"):
        v = r.get(k)
        if v is None:
            continue
        s = str(v).strip()
        if s and s.lower() not in ("none", "nan"):
            return s
    return ""


def _row_grade_outcome(r: dict) -> str:
    """Normalized per-prop grade: HIT, MISS, or empty."""
    result = str(
        r.get("Result", "")
        or r.get("Grade", "")
        or r.get("result", "")
        or r.get("void_reason_grade", "")
        or ""
    ).strip().upper()
    if result in ("HIT", "WIN", "1", "TRUE", "YES", "W"):
        return "HIT"
    if result in ("MISS", "LOSS", "0", "FALSE", "NO", "L"):
        return "MISS"
    return ""


def _rows_for_def_subgrid_cell(
    rows: list[dict],
    canon_def: str | None,
    pick: str,
    direction: str,
    rank: str,
) -> list[dict]:
    """Filter props for one COMBINED subgrid cell. canon_def None = no def-tier filter (Total row)."""
    out: list[dict] = []
    for r in rows:
        if canon_def is not None:
            if normalize_def_tier_label(r.get("Def Tier")) != canon_def:
                continue
        pt = _norm_pick_type_matrix(r.get("Pick Type"))
        if pt != pick:
            continue
        bd = normalize_bet_direction(row_bet_direction(r))
        if bd != direction:
            continue
        rk = str(r.get("Tier", "") or "").strip().upper()
        if rk != rank:
            continue
        out.append(r)
    return out


def _subgrid_pct_class(hit_rate: float) -> str:
    if hit_rate >= 65.0:
        return "def-pct-hit"
    if hit_rate >= 50.0:
        return "def-pct-mid"
    return "def-pct-miss"


def _def_tier_combined_subgrid_table(rows: list[dict], min_decided: int = 10) -> tuple[str, int]:
    """
    DEF TIER BREAKDOWN — pick type × rank tier (16 data cols + label col).
    Rows: five canonical def tiers + Total. Two-row header.
    Returns (html_fragment, count of cells with decided >= min_decided among 5×16 tier cells).
    """
    canon_defs = ("Elite", "Above Avg", "Avg", "Below Avg", "Weak")
    groups: tuple[tuple[str, str, str], ...] = (
        ("Goblin OVER", "Goblin", "OVER"),
        ("Demon OVER", "Demon", "OVER"),
        ("Std OVER", "Standard", "OVER"),
        ("Std UNDER", "Standard", "UNDER"),
    )
    ranks = ("A", "B", "C", "D")

    br_sec = "var(--color-border-secondary, rgba(255,255,255,0.18))"
    br_ter = "var(--color-border-tertiary, rgba(255,255,255,0.10))"

    def _cell_td(stats: dict, col_idx: int) -> str:
        """col_idx 1..16 — determines group boundary borders."""
        within = (col_idx - 1) % 4
        bl = f"border-left:1.5px solid {br_sec}" if within == 0 else f"border-left:0.5px solid {br_ter}"
        base = (
            f'padding:6px 8px;text-align:center;font-size:12px;vertical-align:middle;{bl}'
        )
        d = int(stats.get("decided", 0) or 0)
        if d <= 0:
            return f'<td class="def-subgrid-cell def-subgrid-empty" style="{base}">—</td>'
        hr = float(stats.get("hit_rate", 0.0))
        pct_cls = _subgrid_pct_class(hr)
        pct_s = f"{hr:.0f}%"
        return (
            f'<td class="def-subgrid-cell" style="{base}">'
            f'<span class="def-subgrid-pct {pct_cls}">{pct_s}</span>'
            f'<span class="def-subgrid-n">({fmt_num(d)})</span>'
            f"</td>"
        )

    n_ge = 0
    body_rows = ""
    for def_label in canon_defs:
        row_cells = ""
        col_idx = 1
        row_decided = 0
        for _gh, pick, direction in groups:
            for rank in ranks:
                cell_rows = _rows_for_def_subgrid_cell(rows, def_label, pick, direction, rank)
                st = overall_stats(cell_rows)
                row_decided += int(st.get("decided", 0) or 0)
                if int(st.get("decided", 0)) >= int(min_decided):
                    n_ge += 1
                row_cells += _cell_td(st, col_idx)
                col_idx += 1
        if row_decided <= 0:
            continue
        body_rows += (
            f'<tr class="def-subgrid-data-row">'
            f'<td class="def-subgrid-label" style="padding:6px 8px;text-align:left;border-left:none">{def_label}</td>'
            f"{row_cells}</tr>"
        )

    total_cells = ""
    tcol_idx = 1
    for _gh, pick, direction in groups:
        for rank in ranks:
            cell_rows = _rows_for_def_subgrid_cell(rows, None, pick, direction, rank)
            st = overall_stats(cell_rows)
            total_cells += _cell_td(st, tcol_idx)
            tcol_idx += 1

    body_rows += (
        f'<tr class="def-subgrid-total-row">'
        f'<td class="def-subgrid-label def-subgrid-label-total" style="padding:6px 8px;text-align:left;border-left:none">Total</td>'
        f"{total_cells}</tr>"
    )

    hdr2 = ""
    for gi, (_glabel, _p, _d) in enumerate(groups):
        for ri, rank in enumerate(ranks):
            col_idx = gi * 4 + ri + 1
            group_idx = (col_idx - 1) // 4
            within = (col_idx - 1) % 4
            bl = f"border-left:1.5px solid {br_sec}" if within == 0 else f"border-left:0.5px solid {br_ter}"
            hdr2 += (
                f'<th class="def-subgrid-rank-hdr" style="padding:4px 6px;{bl}">{rank}</th>'
            )

    hdr1_cells = ""
    for glabel, _p, _d in groups:
        hdr1_cells += (
            f'<th colspan="4" class="def-subgrid-group-hdr" style="padding:6px 8px;'
            f"text-align:center;border-left:1.5px solid {br_sec}\">{h(glabel)}</th>"
        )

    table_html = f"""<div class="table-wrap">
<table class="def-tier-subgrid-table" style="min-width:920px;border-collapse:collapse;width:100%">
  <thead>
    <tr>
      <th rowspan="2" class="def-subgrid-corner-hdr" style="padding:6px 8px;text-align:left;border-left:none;vertical-align:bottom">Def Tier</th>
      {hdr1_cells}
    </tr>
    <tr>
      {hdr2}
    </tr>
  </thead>
  <tbody>
    {body_rows}
  </tbody>
</table>
</div>"""
    return table_html, n_ge


def build_pick_tier_direction_agg(rows: list[dict]) -> dict[tuple[str, str, str], dict[str, int]]:
    """
    Accumulate decided / hits / misses per (pick type, tier A-D, direction).
    Goblin/Demon: OVER only; Standard: OVER + UNDER.
    """
    agg: dict[tuple[str, str, str], dict[str, int]] = {}
    for r in rows:
        pt = _norm_pick_type_matrix(r.get("Pick Type"))
        if not pt:
            continue
        tier = str(r.get("Tier", "") or "").strip().upper()
        if tier not in ("A", "B", "C", "D"):
            continue
        direction = normalize_bet_direction(row_bet_direction(r))
        if direction not in ("OVER", "UNDER"):
            continue
        if pt in ("Goblin", "Demon") and direction != "OVER":
            continue

        key = (pt, tier, direction)
        if key not in agg:
            agg[key] = {"hits": 0, "misses": 0, "decided": 0}

        result = str(r.get("Result", "") or r.get("Grade", "") or "").strip().upper()
        if result in ("HIT", "WIN", "1", "TRUE", "YES", "W"):
            agg[key]["hits"] += 1
            agg[key]["decided"] += 1
        elif result in ("MISS", "LOSS", "0", "FALSE", "NO", "L"):
            agg[key]["misses"] += 1
            agg[key]["decided"] += 1
        else:
            d = safe_int(r.get("Decided", r.get("decided", 0)))
            h_ = safe_int(r.get("Hits", r.get("hits", 0)))
            m_ = safe_int(r.get("Misses", r.get("misses", 0)))
            if d > 0 or h_ > 0 or m_ > 0:
                agg[key]["decided"] += d
                agg[key]["hits"] += h_
                agg[key]["misses"] += m_
    return agg


def _canonical_pick_tier_direction_keys() -> list[tuple[str, str, str]]:
    """Fixed display order: Standard (A-D × O/U), Goblin (A-D OVER), Demon (A-D OVER)."""
    keys: list[tuple[str, str, str]] = []
    for tier in ("A", "B", "C", "D"):
        keys.append(("Standard", tier, "OVER"))
        keys.append(("Standard", tier, "UNDER"))
    for tier in ("A", "B", "C", "D"):
        keys.append(("Goblin", tier, "OVER"))
    for tier in ("A", "B", "C", "D"):
        keys.append(("Demon", tier, "OVER"))
    return keys


def pick_tier_direction_matrix_html(rows: list[dict], min_decided: int = 10) -> str:
    agg = build_pick_tier_direction_agg(rows)
    keys = _canonical_pick_tier_direction_keys()
    n_ge = sum(1 for k in keys if int(agg.get(k, {}).get("decided", 0)) >= int(min_decided))
    summary = (
        f"Full grid: Standard A–D (OVER/UNDER), Goblin & Demon A–D (OVER). "
        f"{n_ge} of {len(keys)} cells have ≥ {int(min_decided)} decided; others still listed with counts. "
        f"Note: Standard Tier B may show no decided props when ml_prob is compressed below tier thresholds "
        f"(e.g., B cut ≥ 0.65) — “—” cells are expected until calibration / a retrain restores spread."
    )

    body_rows = ""
    for pt, tier, direction in keys:
        v = agg.get((pt, tier, direction), {"hits": 0, "misses": 0, "decided": 0})
        d = int(v["decided"])
        h_ = int(v["hits"])
        m_ = int(v["misses"])
        pt_chip = {
            "Goblin": '<span class="chip chip-goblin">🎃\u00a0Goblin</span>',
            "Demon": '<span class="chip chip-demon">😈\u00a0Demon</span>',
            "Standard": '<span class="chip chip-std">⭐\u00a0Standard</span>',
        }.get(pt, h(pt))
        tier_u = str(tier).upper()
        tier_chip_cls = {"A": "chip-a", "B": "chip-b", "C": "chip-c"}.get(tier_u, "chip-d")
        dir_html = (
            '<span style="color:var(--green);font-size:13px">▲ OVER</span>'
            if direction == "OVER"
            else '<span style="color:var(--cyan);font-size:13px">▼ UNDER</span>'
        )
        if d <= 0:
            body_rows += f"""<tr class="matrix-empty">
          <td>{pt_chip}</td>
          <td><span class="chip {tier_chip_cls}">TIER\u00a0{tier_u}</span></td>
          <td>{dir_html}</td>
          <td class="right mono muted">—</td>
          <td class="right mono muted">—</td>
          <td class="right mono muted">—</td>
          <td class="mono right muted">—</td>
          <td><span class="muted">—</span></td>
        </tr>"""
            continue
        hr = (h_ / d * 100.0) if d > 0 else 0.0
        row_cls = "matrix-hit" if hr >= 60.0 else ("matrix-miss" if hr < 50.0 else "matrix-warn")
        if d < int(min_decided):
            row_cls += " matrix-sparse"
        bar_html = stacked_rate_bar_bg(h_, m_)
        body_rows += f"""<tr class="{row_cls}">
          <td>{pt_chip}</td>
          <td><span class="chip {tier_chip_cls}">TIER\u00a0{tier_u}</span></td>
          <td>{dir_html}</td>
          <td class="right mono">{fmt_num(d)}</td>
          <td class="right mono pos">{fmt_num(h_)}</td>
          <td class="right mono neg">{fmt_num(m_)}</td>
          <td class="mono right">{pct(hr)}</td>
          <td>{bar_html}</td>
        </tr>"""

    return f"""<details class="matrix-collapsible">
      <summary>Performance Matrix — Pick Type × Tier × Direction</summary>
      <div class="matrix-body">
        <div class="matrix-summary">{summary}</div>
        <div class="table-wrap"><table class="table-sortable">
          <thead><tr>
            <th data-sort-key="pick" title="Sort">PICK TYPE</th>
            <th data-sort-key="tier" title="Sort">TIER</th>
            <th data-sort-key="dir" title="Sort">DIRECTION</th>
            <th class="right" data-sort-key="decided" title="Sort">DECIDED</th>
            <th class="right" data-sort-key="hits" title="Sort">HITS</th>
            <th class="right" data-sort-key="misses" title="Sort">MISSES</th>
            <th data-sort-key="rate" title="Sort">HIT RATE</th>
            <th data-sort-key="bar" title="Sort">BAR</th>
          </tr></thead>
          <tbody>{body_rows}</tbody>
        </table></div>
      </div>
    </details>"""


# ══════════════════════════════════════════════════════════════════════════════
#  HTML FRAGMENT BUILDERS
# ══════════════════════════════════════════════════════════════════════════════

def pick_type_row(label: str, icon: str, agg: dict, extra_html: str = "") -> str:
    d   = agg["decided"]
    h_  = agg["hits"]
    m   = agg["misses"]
    hr  = agg["hit_rate"]
    return f"""<tr>
      <td><span class="chip chip-{label.lower()}">{icon}\u00a0{label}</span>{extra_html}</td>
      <td class="right mono">{fmt_num(d)}</td>
      <td class="right mono pos">{fmt_num(h_)}</td>
      <td class="right mono neg">{fmt_num(m)}</td>
      <td>{rate_bar_html(hr, hits=h_, misses=m)}</td>
    </tr>"""


def tier_row(tier: str, agg: dict, extra: str = "") -> str:
    d   = agg["decided"]
    h_  = agg["hits"]
    m   = agg.get("misses", max(0, d - h_))
    hr  = agg["hit_rate"]
    row_cls = "player-hit" if hr >= 55 else ("player-miss" if hr < 48 else "player-warn")
    chip_cls = {"A":"chip-a","B":"chip-b","C":"chip-c"}.get(tier.upper(), "chip-d")
    over_pct = pct(agg.get("over_rate", None))
    under_pct = pct(agg.get("under_rate", None))
    dir_html = ""
    if over_pct != "—":
        col = "var(--green)" if pct_f(over_pct) >= 55 else "var(--red)"
        dir_html += f'<span style="color:var(--cyan);font-size:12px;margin-right:8px">▲ </span><span style="color:{col};font-size:12px;margin-right:8px">{over_pct}</span>'
    if under_pct != "—":
        col = "var(--green)" if pct_f(under_pct) >= 55 else "var(--red)"
        dir_html += f'<span style="color:var(--gold);font-size:12px">▼ </span><span style="color:{col};font-size:12px">{under_pct}</span>'
    tier_lbl = f"TIER {tier.upper()}"
    return f"""<tr class="{row_cls}">
      <td><span class="chip {chip_cls}">{tier_lbl}</span>
      <div style="font-size:12px;color:var(--muted2);margin-top:4px">{dir_html}{extra}</div></td>
      <td class="right mono">{fmt_num(d)}</td>
      <td class="right mono pos">{fmt_num(h_)}</td>
      <td>{rate_bar_html(hr, hits=h_, misses=m)}</td>
    </tr>"""


def prop_type_table(data: list[dict], label: str, min_decided: int = 10) -> str:
    filtered = [d for d in data if d["decided"] >= min_decided]
    filtered.sort(key=lambda x: x["hit_rate"], reverse=True)
    if not filtered:
        return f'<div class="muted-note">No prop types with ≥{min_decided} decided.</div>'
    rows_html = ""
    for d in filtered:
        row_cls = "player-hit" if d["hit_rate"] >= 55 else ("player-miss" if d["hit_rate"] < 48 else "player-warn")
        rows_html += f"""<tr class="{row_cls}">
          <td class="mono">{h(d['key'])}</td>
          <td class="right mono">{fmt_num(d['decided'])}</td>
          <td class="right mono pos">{fmt_num(d['hits'])}</td>
          <td>{rate_bar_html(d['hit_rate'], hits=d['hits'], misses=d.get('misses', max(0, d['decided'] - d['hits'])))}</td>
        </tr>"""
    return f"""<div class="table-wrap"><table>
      <thead><tr><th>{label}</th><th class="right">DECIDED</th><th class="right">HITS</th><th>HIT RATE</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>"""


def player_table(rows: list[dict], top: bool, min_decided: int = 5, limit: int = 8) -> str:
    """
    Same-day player record: dedupe each (player, prop, direction, line) once, then roll up per player.
    Avoids merging different lines (e.g. Points 2.5 vs 11.5) into fake 100% / 50% buckets.
    """
    deduped: dict[tuple[str, str, str, str], dict] = {}
    for r in rows:
        player = str(r.get("Player") or r.get("player") or "").strip()
        team = str(r.get("Team") or r.get("team") or r.get("Sport") or "").strip()
        prop = str(
            r.get("Prop Type")
            or r.get("prop_type_norm")
            or r.get("Prop")
            or ""
        ).strip() or "Unknown Prop"
        side_raw = normalize_bet_direction(row_bet_direction(r))
        side = side_raw if side_raw in ("OVER", "UNDER") else "—"
        line = _row_line_key(r)
        if not player or player.lower() in ("none", "nan", ""):
            continue
        outcome = _row_grade_outcome(r)
        if not outcome:
            continue
        key = (player, prop, side, line)
        if key in deduped:
            continue
        deduped[key] = {"team": team, "prop": prop, "side": side, "line": line, "outcome": outcome}

    player_data: dict[str, dict] = {}
    for (player, _prop, _side, _line), v in deduped.items():
        bucket = player_data.setdefault(
            player,
            {"team": v["team"], "hits": 0, "misses": 0, "decided": 0, "props": set()},
        )
        if v["team"]:
            bucket["team"] = v["team"]
        bucket["decided"] += 1
        bucket["props"].add(v["prop"])
        if v["outcome"] == "HIT":
            bucket["hits"] += 1
        else:
            bucket["misses"] += 1

    candidates = []
    for name, v in player_data.items():
        d = int(v["decided"])
        if d < min_decided:
            continue
        hr = v["hits"] / d * 100.0
        n_props = len(v["props"])
        line_lbl = f"{n_props} prop{'s' if n_props != 1 else ''}"
        candidates.append({
            "name": name,
            "team": v["team"],
            "line_lbl": line_lbl,
            "hits": v["hits"],
            "misses": v["misses"],
            "decided": d,
            "hit_rate": hr,
        })

    if not candidates:
        return (
            f'<div class="muted-note">No players with ≥{min_decided} distinct graded props on this slate.</div>'
        )

    if top:
        candidates.sort(key=lambda x: (x["hit_rate"], x["decided"]), reverse=True)
    else:
        candidates.sort(key=lambda x: (x["hit_rate"], -x["decided"]))
    candidates = candidates[:limit]

    rows_html = ""
    for c in candidates:
        rows_html += f"""<tr>
          <td><strong>{h(c['name'])}</strong></td>
          <td class="mono muted">{h(c['team'])}</td>
          <td class="mono muted">{h(c['line_lbl'])}</td>
          <td class="right mono">{fmt_num(c['decided'])}</td>
          <td class="right mono pos">{fmt_num(c['hits'])}</td>
          <td class="right mono neg">{fmt_num(c['misses'])}</td>
          <td>{rate_bar_html(c['hit_rate'], hits=c['hits'], misses=c['misses'])}</td>
        </tr>"""
    return f"""<div class="table-wrap"><table>
      <thead><tr><th>PLAYER</th><th>TEAM</th><th>PROPS</th><th class="right">DEC</th><th class="right">H</th><th class="right">M</th><th>RATE</th></tr></thead>
      <tbody>{rows_html}</tbody>
    </table></div>"""


def def_tier_table(rows: list[dict], min_decided: int = 10) -> str:
    dt_agg = build_agg_from_rows(rows, "Def Tier")
    if not dt_agg:
        return ""
    combined_subgrid, n_sub_ge = _def_tier_combined_subgrid_table(rows, min_decided=min_decided)
    # Same shell as Performance Matrix (pick_tier_direction_matrix_html): matrix-collapsible + matrix-body + matrix-summary + table-wrap inside subgrid.
    intro = (
        f"Rows = opponent def tier. Columns = pick type × rank tier (A–D). "
        f"Goblin/Demon = OVER only. Standard split OVER/UNDER. "
        f"{n_sub_ge} of 80 opponent-tier × column cells have ≥ {int(min_decided)} decided; others still list rates when 1 ≤ n &lt; {int(min_decided)}. "
        f"<strong>—</strong> = no decided props in that bucket. "
        f"Note: Standard Tier B may be structurally sparse when ml_prob is compressed below tier thresholds — “—” is expected until calibration / a retrain restores spread."
    )
    return f"""<details class="matrix-collapsible">
      <summary>DEF TIER BREAKDOWN — PICK TYPE × RANK TIER (COMBINED)</summary>
      <div class="matrix-body">
        <div class="matrix-summary">{intro}</div>
        {combined_subgrid}
      </div>
    </details>"""


def _build_prop_breakdown_rows_from_rows(rows: list[dict]) -> list[dict[str, Any]]:
    grouped: dict[tuple[str, str, str, str], dict[str, int]] = defaultdict(lambda: {"decided": 0, "hits": 0})
    for r in rows:
        result = str(r.get("Result", "") or r.get("Grade", "") or "").strip().upper()
        if result in ("HIT", "WIN", "1", "TRUE", "YES", "W"):
            norm_result = "HIT"
        elif result in ("MISS", "LOSS", "0", "FALSE", "NO", "L"):
            norm_result = "MISS"
        else:
            continue
        prop = _cell_str(r.get("Prop Type") or r.get("Prop")) or "Unknown"
        pt_label = _map_pt_label(r.get("Pick Type"), row_bet_direction(r))
        if not pt_label:
            continue
        tier = _cell_str(r.get("Tier")).upper()
        if tier not in ("A", "B", "C", "D"):
            continue
        def_label = _norm_def_label(r.get("Def Tier"))
        key = (prop, pt_label, tier, def_label)
        grouped[key]["decided"] += 1
        if norm_result == "HIT":
            grouped[key]["hits"] += 1
    out: list[dict[str, Any]] = []
    for (prop, pt, tier, def_label), v in grouped.items():
        out.append(
            {
                "prop": prop,
                "pt": pt,
                "tier": tier,
                "def": def_label,
                "decided": int(v["decided"]),
                "hits": int(v["hits"]),
            }
        )
    return out


def prop_breakdown_widget(rows: list[dict]) -> str:
    payload = _build_prop_breakdown_rows_from_rows(rows)
    payload_json = h(json.dumps(payload, ensure_ascii=False))
    return f"""<details class="matrix-collapsible prop-breakdown-widget">
      <summary>PROP TYPE BREAKDOWNS</summary>
      <div class="matrix-body">
        <div class="matrix-summary">Interactive Best/Worst and heatmap views across pick type, tier, and prop type. Filters apply to both tabs.</div>
        <div class="pbw-root" data-prop-breakdown-rows='{payload_json}'></div>
      </div>
    </details>"""


# ══════════════════════════════════════════════════════════════════════════════
#  SPORT SECTION BUILDER
# ══════════════════════════════════════════════════════════════════════════════

def build_sport_section(rows: list[dict], sport: str, icon: str) -> str:
    rows = _filter_prop_level_rows(rows)
    if not rows:
        return ""

    stats = overall_stats(rows)
    if stats["decided"] <= 0 and stats["total"] <= 0:
        return ""

    if sport.strip().upper() == "MLB" and (not (icon or "").strip() or (icon or "").strip().upper() == "MLB"):
        icon = "⚾"

    total_label = fmt_num(stats["total"]) if stats["total"] > 0 else fmt_num(stats["decided"] + stats["voids"])
    # Apply the pick-type x tier analysis uniformly across all sport sections.
    matrix_section = pick_tier_direction_matrix_html(rows, min_decided=10)

    # ── Def Tier ───────────────────────────────────────────────────────────────
    def_section = def_tier_table(rows)

    # ── Prop Type Breakdowns (interactive widget) ─────────────────────────────
    prop_section = prop_breakdown_widget(rows)

    # ── Player Leaderboards ────────────────────────────────────────────────────
    top_players   = player_table(rows, top=True,  min_decided=3, limit=8)
    worst_players = player_table(rows, top=False, min_decided=3, limit=8)

    player_section = f"""<div class="two-col">
      <div>
        <div class="section-label">🏆 TOP PLAYERS (SAME-DAY RECORD)</div>
        {top_players}
      </div>
      <div>
        <div class="section-label">💀 COLD PLAYERS (SAME-DAY RECORD)</div>
        {worst_players}
      </div>
    </div>"""

    return f"""<details class="sport-section sport-collapsible">
    <summary>
      <div class="sport-header">
        <div class="sport-label">{icon} {sport}</div>
        <div class="sport-header-line"></div>
        <div class="sport-meta-count">{total_label} TOTAL PROPS</div>
      </div>
    </summary>
    <div class="sport-section-body">
      {matrix_section}
      {def_section}
      {prop_section}
      {player_section}
    </div>
  </details>"""


# ══════════════════════════════════════════════════════════════════════════════
#  TAKEAWAYS / INSIGHTS
# ══════════════════════════════════════════════════════════════════════════════

def _takeaway_sport_snippets(
    bundles: list[tuple[str, list[dict]]],
) -> str:
    """One-line per sport hit rate for the takeaway summary (only sports with decided props)."""
    parts: list[str] = []
    for label, rows in bundles:
        if not rows:
            continue
        st = overall_stats(rows)
        if st["decided"] <= 0:
            continue
        parts.append(f"<strong>{h(label)}</strong>: {pct(st['hit_rate'])} ({fmt_num(st['decided'])} dec)")
    return (" · ".join(parts)) if parts else ""


def build_takeaways(
    nba_rows: list[dict],
    cbb_rows: list[dict],
    nhl_rows: list[dict] | None = None,
    soccer_rows: list[dict] | None = None,
    mlb_rows: list[dict] | None = None,
    wnba_rows: list[dict] | None = None,
    tennis_rows: list[dict] | None = None,
) -> str:
    """
    Insight cards at the bottom of slate_eval. Uses **all loaded sports** so Railway /grades
    shows combined takeaways (not NBA-only) when MLB/NHL/Soccer are graded without NBA.
    """
    nhl_rows = nhl_rows or []
    soccer_rows = soccer_rows or []
    mlb_rows = mlb_rows or []
    wnba_rows = wnba_rows or []
    tennis_rows = tennis_rows or []

    all_rows: list[dict] = (
        list(nba_rows)
        + list(cbb_rows)
        + list(nhl_rows)
        + list(soccer_rows)
        + list(mlb_rows)
        + list(wnba_rows)
        + list(tennis_rows)
    )

    insights: list[str] = []
    alerts: list[str] = []

    def add_insight(icon: str, title: str, body: str) -> None:
        insights.append(f"""<div class="insight-card">
      <div class="insight-icon">{icon}</div>
      <div class="insight-title">{h(title)}</div>
      <div class="insight-body">{body}</div>
    </div>""")

    if not all_rows:
        add_insight("📊", "No Data", "No graded props found for this date.")
        insights_html = "\n".join(insights)
        return f"""<div class="sport-section">
    <div class="sport-header">
      <div class="sport-label">📋 TAKEAWAYS</div>
      <div class="sport-header-line"></div>
    </div>
    <div class="insight-grid">{insights_html}</div>
  </div>"""

    all_stats = overall_stats(all_rows)
    bundles = [
        ("NBA", nba_rows),
        ("CBB", cbb_rows),
        ("NHL", nhl_rows),
        ("Soccer", soccer_rows),
        ("MLB", mlb_rows),
        ("WNBA", wnba_rows),
        ("Tennis", tennis_rows),
    ]
    sport_line = _takeaway_sport_snippets(bundles)

    # ── Combined overall summary (all sports) ─────────────────────────────
    add_insight(
        "📋",
        "Overall Slate Summary (all sports)",
        f"Combined: <strong>{pct(all_stats['hit_rate'])}</strong> on "
        f"{fmt_num(all_stats['decided'])} decided props."
        + (f"<br/><span style=\"opacity:0.92\">{sport_line}</span>" if sport_line else ""),
    )

    # Tier A + Goblin — aggregated across every sport in the slate
    ta_all = tier_a_stats(all_rows)
    gb_all = pick_type_stats(all_rows, "goblin")
    if ta_all["decided"] > 0 or gb_all["decided"] > 0:
        ta_str = f"Tier A: <strong>{pct(ta_all['hit_rate'])}</strong> ({fmt_num(ta_all['decided'])} dec)." if ta_all["decided"] > 0 else ""
        gb_str = f" Goblin: <strong>{pct(gb_all['hit_rate'])}</strong> ({fmt_num(gb_all['decided'])} dec)." if gb_all["decided"] > 0 else ""
        body = (ta_str + " " + gb_str).strip()
        if ta_all["hit_rate"] >= 60 or (gb_all["decided"] > 0 and gb_all["hit_rate"] >= 58):
            add_insight("✅", "Tier A & Goblin (all sports)", body)
        else:
            add_insight("⚠️", "Tier A & Goblin (all sports)", body)

    # Demon — all sports
    dem_all = pick_type_stats(all_rows, "demon")
    if dem_all["decided"] > 0:
        demon_str = f"Demons (all sports): <strong>{pct(dem_all['hit_rate'])}</strong> on {fmt_num(dem_all['decided'])} decided."
        if dem_all["hit_rate"] < 45:
            add_insight(
                "🚨",
                "Demon Line Performance (all sports)",
                demon_str + " Demon hit rate is well below breakeven — monitor before including in slips.",
            )
            alerts.append(
                f'<div class="alert alert-red"><div class="alert-title">🚨 Demon lines (all sports) — '
                f'{pct(dem_all["hit_rate"])} on {fmt_num(dem_all["decided"])} decided</div>'
                f"Demon hit rate is well below breakeven. Exclude from slips until further notice.</div>"
            )
        else:
            add_insight(
                "📊",
                "Demon Line Performance (all sports)",
                demon_str + " Monitor demon performance before including in slips.",
            )

    # Over vs Under — all sports (matches ALL SPORTS matrix scope)
    over_all = overall_stats(
        [r for r in all_rows if normalize_bet_direction(row_bet_direction(r)) == "OVER"]
    )
    under_all = overall_stats(
        [r for r in all_rows if normalize_bet_direction(row_bet_direction(r)) == "UNDER"]
    )
    if over_all["decided"] > 0 and under_all["decided"] > 0:
        add_insight(
            "📈",
            "Over vs Under (all sports)",
            f"OVERs: <strong>{pct(over_all['hit_rate'])}</strong> ({fmt_num(over_all['decided'])} dec). "
            f"UNDERs: <strong>{pct(under_all['hit_rate'])}</strong> ({fmt_num(under_all['decided'])} dec).",
        )

    insights_html = "\n".join(insights)
    alerts_html = "\n".join(alerts)

    return f"""<div class="sport-section">
    <div class="sport-header">
      <div class="sport-label">📋 TAKEAWAYS</div>
      <div class="sport-header-line"></div>
    </div>
    {alerts_html}
    <div class="insight-grid">{insights_html}</div>
  </div>"""


def sport_label(s: str) -> str:
    return {
        "NBA": "🏀 NBA",
        "CBB": "🎓 CBB",
        "NHL": "🏒 NHL",
        "MLB": "⚾ MLB",
        "SOCCER": "⚽ Soccer",
        "WNBA": "🏀 WNBA",
        "TENNIS": "🎾 Tennis",
    }.get(s.upper(), s)


# ══════════════════════════════════════════════════════════════════════════════
#  FILE DISCOVERY
# ══════════════════════════════════════════════════════════════════════════════

def find_graded_file(sport: str, date_str: str) -> Path | None:
    """Search common locations for graded_* workbooks (WNBA may be wnba_graded_{date}.xlsx)."""
    sport_l = sport.lower()
    if sport_l == "wnba":
        patterns = [f"graded_wnba_{date_str}.xlsx", f"wnba_graded_{date_str}.xlsx"]
    else:
        patterns = [f"graded_{sport_l}_{date_str}.xlsx"]
    search_dirs = [
        SCRIPT_DIR,
        SCRIPT_DIR / "outputs",
        SCRIPT_DIR / "outputs" / date_str,
        OUTPUTS_DIR,
        OUTPUTS_DIR / date_str,
        Path.cwd(),
        Path.cwd() / "outputs",
        Path.cwd() / "outputs" / date_str,
        # project root outputs/{date}/ — e.g. PropOracle/outputs/2026-03-06/
        ROOT_DIR / "outputs" / date_str,
        ROOT_DIR / "outputs",
        ROOT_DIR,
    ]
    for pattern in patterns:
        for d in search_dirs:
            p = d / pattern
            if p.exists():
                return p
    # glob fallback
    for d in search_dirs:
        if d.exists():
            matches = list(d.glob(f"*{sport_l}*graded*{date_str}*.xlsx"))
            if matches:
                return matches[0]
            if sport_l == "wnba":
                matches2 = list(d.glob(f"wnba*graded*{date_str}*.xlsx"))
                if matches2:
                    return matches2[0]
    return None


def _tennis_step8_search_paths(bundle_dir: Path, match_date: str, bundle_date: str) -> list[Path]:
    """Mirror run_grader.ps1 Get-TennisStep8SearchPaths (path-based, not mtime)."""
    tennis_dir = bundle_dir / "tennis"
    paths: list[Path] = [
        tennis_dir / "step8_tennis_direction_clean.xlsx",
        tennis_dir / "step8_tennis_direction.csv",
        bundle_dir / f"step8_tennis_direction_clean_{match_date}.xlsx",
        bundle_dir / f"step8_tennis_direction_clean_{bundle_date}.xlsx",
    ]
    if tennis_dir.is_dir():
        paths.extend(sorted(tennis_dir.glob("step8_*.csv")))
        paths.extend(sorted(tennis_dir.glob("step8_*.xlsx")))
    return paths


def find_tennis_graded_file(date_str: str) -> Path | None:
    """
    Resolve graded_tennis workbook for a pipeline grade date.

    When step8 lives in outputs/(grade_date - 1) (tomorrow-fetch), the grader
    writes to outputs/(grade_date + 1)/graded_tennis_{match_day}.xlsx — same
    rule as run_grader.ps1, keyed off step8 bundle location (not file mtimes).
    """
    date_str = str(date_str or "")[:10]
    try:
        grade_dt = datetime.strptime(date_str, "%Y-%m-%d").date()
    except ValueError:
        return find_graded_file("tennis", date_str)

    bundle_date = (grade_dt - timedelta(days=1)).strftime("%Y-%m-%d")
    offset_bundle = ROOT_DIR / "outputs" / bundle_date
    grade_bundle = ROOT_DIR / "outputs" / date_str
    sports_out = ROOT_DIR / "Sports" / "Tennis" / "outputs"

    def _first_existing(candidates: list[Path]) -> Path | None:
        for p in candidates:
            if p.exists():
                return p
        return None

    offset_step8 = _first_existing(_tennis_step8_search_paths(offset_bundle, date_str, bundle_date))
    if offset_step8 is not None:
        match_day = (grade_dt + timedelta(days=1)).strftime("%Y-%m-%d")
        offset_graded = ROOT_DIR / "outputs" / match_day / f"graded_tennis_{match_day}.xlsx"
        if offset_graded.exists():
            return offset_graded
        print(
            f"  WARNING: Tennis step8 from bundle {bundle_date} (offset) but "
            f"no graded workbook at {offset_graded.name} under outputs/{match_day}/"
        )
        return None

    same_step8 = _first_existing(_tennis_step8_search_paths(grade_bundle, date_str, date_str))
    if same_step8 is not None:
        same_graded = grade_bundle / f"graded_tennis_{date_str}.xlsx"
        if same_graded.exists():
            return same_graded

    for static in (
        sports_out / "step8_tennis_direction_clean.xlsx",
        sports_out / "step8_tennis_direction.csv",
    ):
        if static.exists():
            break
    else:
        static = None

    if static is not None:
        return find_graded_file("tennis", date_str)

    return find_graded_file("tennis", date_str)


def load_merged_nba_graded_rows(date_str: str) -> list[dict]:
    """Merge NBA + NBA1Q + NBA1H rows (single list) for slate HTML that shows one NBA section."""
    rows: list[dict] = []
    for key in ("nba", "nba1q", "nba1h"):
        p = find_graded_file(key, date_str)
        if p:
            rows.extend(load_graded(p))
    return rows


def nba_family_bundles_for_json(date_str: str) -> list[tuple[str, list[dict]]]:
    """Separate bundles so graded_props JSON carries NBA vs NBA1Q vs NBA1H sport labels."""
    bundles: list[tuple[str, list[dict]]] = []
    for key, label in (("nba", "NBA"), ("nba1q", "NBA1Q"), ("nba1h", "NBA1H")):
        p = find_graded_file(key, date_str)
        if p:
            bundles.append((label, load_graded(p)))
    return bundles


# ══════════════════════════════════════════════════════════════════════════════
#  FULL HTML
# ══════════════════════════════════════════════════════════════════════════════

NAV_HTML = ""

CSS = """
:root{
  --glass:rgba(255,255,255,0.03);
  --glass-bd:rgba(255,255,255,0.08);
  --bg:transparent;
  --text:rgba(232,236,255,0.95);
  --muted:#94a3b8;
  --muted2:rgba(255,255,255,0.48);
  --gold:#f0a500;
  --gold2:#d4a017;
  --cyan:#00e5ff;
  --green:#39ff6e;
  --amber:#e8b84a;
  --red:#ff4d4d;
  --purple:#c4a5ff;
  --pending:#666;
}
*{box-sizing:border-box;margin:0;padding:0}
html{min-height:0;height:auto;max-height:none;overflow-x:hidden;overflow-y:visible;scrollbar-gutter:stable;-webkit-overflow-scrolling:touch;background:#0a0a14}
body{font-family:'Inter',sans-serif;background:#0a0a14;color:var(--text);min-height:0;height:auto;max-height:none;margin:0;overflow-x:hidden;overflow-y:visible;padding-bottom:max(10px, env(safe-area-inset-bottom, 0px));font-size:clamp(14px,1.02vw,16px);line-height:1.45}
h1,h2,h3,h4,h5,h6{font-family:'Bebas Neue',sans-serif}
header,.main{position:relative;z-index:1}

header{background:transparent;border:none;border-radius:0;padding:12px 20px 0;display:flex;flex-direction:column;align-items:stretch;gap:0;box-shadow:none}
.slate-header-top{display:flex;align-items:center;justify-content:space-between;flex-wrap:wrap;gap:12px;width:100%}
.grades-hub-toolbar-host{flex:1 1 100%;align-self:stretch;width:100%;max-width:100%;min-height:0;box-sizing:border-box}
.logo{display:flex;align-items:center;gap:14px}
.logo-icon{width:120px;height:120px;object-fit:contain;display:block;filter:drop-shadow(0 0 8px rgba(212,160,23,0.45))}
@media(max-width:768px){.logo-icon{width:80px;height:80px}}
.logo-title{font-family:'Bebas Neue',sans-serif;font-size:28px;letter-spacing:3px;background:linear-gradient(to bottom,#f0a500,#d4a017,#f7e08a);-webkit-background-clip:text;background-clip:text;color:transparent}
.logo-sub{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted);letter-spacing:2.2px;margin-top:2px}
.date-badge{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted2);background:var(--glass);backdrop-filter:blur(20px);-webkit-backdrop-filter:blur(20px);
border:1px solid var(--glass-bd);border-radius:999px;padding:8px 14px;letter-spacing:1.5px}
.main{max-width:none;width:100%;margin:0;padding:24px 20px;box-sizing:border-box}
.sport-header{display:flex;align-items:center;gap:14px;margin-bottom:22px;flex-wrap:wrap;min-width:0}
.sport-label{font-family:'Bebas Neue',sans-serif;font-size:32px;letter-spacing:4px;line-height:1;color:var(--gold);text-shadow:0 0 28px rgba(240,165,0,.18)}
.sport-header-line{flex:1;min-width:80px;height:1px;background:rgba(255,255,255,0.08)}
.sport-meta-count{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted2)}
.sport-section{margin-bottom:16px;width:100%;max-width:100%;box-sizing:border-box}
.sport-section:last-child{margin-bottom:0}
.sport-collapsible{border:1px solid var(--glass-bd);border-radius:14px;padding:12px 14px;background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);box-shadow:0 4px 24px rgba(0,0,0,.18)}
.sport-collapsible>summary{list-style:none;cursor:pointer}
.sport-collapsible>summary::-webkit-details-marker{display:none}
.sport-collapsible>summary::marker{display:none;content:''}
.sport-collapsible>summary .sport-header{margin-bottom:0}
.sport-collapsible>summary .sport-label::before{content:'▸ ';color:var(--gold)}
.sport-collapsible[open]>summary .sport-label::before{content:'▾ '}
.sport-collapsible .sport-section-body{padding-top:14px}
.matrix-collapsible{margin:6px 0 20px;border:1px solid var(--glass-bd);border-radius:14px;background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);box-shadow:0 4px 24px rgba(0,0,0,.18);overflow:hidden}
.matrix-collapsible>summary{list-style:none;cursor:pointer;padding:12px 14px;font-family:'Bebas Neue',sans-serif;font-size:clamp(14px,1.08vw,16px);letter-spacing:2px;color:var(--muted);border-bottom:1px solid rgba(255,255,255,0.06);display:flex;align-items:center;gap:8px}
.matrix-collapsible>summary::-webkit-details-marker{display:none}
.matrix-collapsible>summary::before{content:'▸';color:var(--gold);transition:transform .15s ease}
.matrix-collapsible[open]>summary::before{transform:rotate(90deg)}
.matrix-body{padding:12px}
.matrix-summary{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted2);margin:0 0 10px}
.matrix-collapsible tr.matrix-hit td:first-child{border-left:3px solid var(--green)}
.matrix-collapsible tr.matrix-miss td:first-child{border-left:3px solid var(--red)}
.matrix-collapsible tr.matrix-warn td:first-child{border-left:3px solid var(--gold)}
.matrix-collapsible tr.matrix-sparse td:first-child{border-left:3px dashed rgba(255,255,255,0.12)}
.pbw-controls{display:flex;gap:10px;flex-wrap:wrap;margin:6px 0 14px}
.pbw-control{display:flex;align-items:center;gap:8px;background:rgba(0,0,0,0.2);border:1px solid var(--glass-bd);border-radius:10px;padding:6px 10px}
.pbw-control label{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted2)}
.pbw-control select{background:#0f1324;color:var(--text);border:1px solid var(--glass-bd);border-radius:8px;padding:6px 8px;font-size:12px}
.pbw-tabs{display:flex;gap:8px;margin-bottom:12px}
.pbw-tab-btn{background:rgba(255,255,255,0.04);border:1px solid var(--glass-bd);border-radius:10px;padding:8px 12px;color:var(--muted);font-family:'Bebas Neue',sans-serif;letter-spacing:1.5px;cursor:pointer}
.pbw-tab-btn.active{color:var(--text);border-color:rgba(255,255,255,0.25)}
.pbw-tab-panel{display:none}
.pbw-tab-panel.active{display:block}
.pbw-pt-header{font-family:'Bebas Neue',sans-serif;font-size:16px;letter-spacing:2px;color:var(--muted);padding-bottom:8px;border-bottom:1px solid rgba(255,255,255,0.1);margin:16px 0 10px;text-transform:uppercase}
.pbw-tier-title{font-family:'Bebas Neue',sans-serif;font-size:12px;color:var(--muted2);letter-spacing:1.6px;margin:8px 0}
.pbw-tier-grid{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:12px;margin-bottom:14px}
.pbw-panel-title{font-family:'Bebas Neue',sans-serif;font-size:14px;letter-spacing:1.4px;margin-bottom:6px}
.pbw-panel-title.best{color:#0F6E56}
.pbw-panel-title.worst{color:#993C1D}
.pbw-mini-bar{height:6px;background:rgba(255,255,255,.08);border-radius:4px;overflow:hidden}
.pbw-mini-bar-fill{height:100%}
.pbw-heatmap-wrap{overflow:scroll;scrollbar-gutter:stable;border:1px solid var(--glass-bd);border-radius:12px}
.pbw-heatmap{width:100%;border-collapse:separate;border-spacing:0}
.pbw-heatmap th,.pbw-heatmap td{padding:10px;border-bottom:1px solid rgba(255,255,255,0.06);text-align:center}
.pbw-heatmap th:first-child,.pbw-heatmap td:first-child{position:sticky;left:0;background:#0f1324;text-align:left;z-index:1}
.pbw-cell-rate{font-family:'Bebas Neue',sans-serif;font-size:15px;letter-spacing:1px}
.pbw-cell-dec{font-size:11px;color:rgba(0,0,0,0.65)}
.def-tier-subgrid-table{font-size:13px}
.def-tier-subgrid-table thead th{font-family:'Bebas Neue',sans-serif;letter-spacing:1.6px;color:rgba(255,255,255,0.82)}
.def-subgrid-corner-hdr{font-size:12px!important;font-weight:600!important;color:rgba(255,255,255,0.88)!important}
.def-subgrid-group-hdr{font-size:12px!important;font-weight:600!important;color:rgba(255,255,255,0.8)!important}
.def-subgrid-rank-hdr{font-size:11px!important;font-weight:600!important;color:rgba(255,255,255,0.72)!important}
.def-tier-subgrid-table tbody td{font-family:'Inter',sans-serif}
.def-subgrid-label{color:var(--text);font-weight:600;font-size:13px}
.def-subgrid-label-total{font-weight:700}
.def-subgrid-pct{font-weight:700;font-size:15px;letter-spacing:0.02em;line-height:1.15}
.def-pct-hit{color:#5ef598}
.def-pct-mid{color:#f5c842}
.def-pct-miss{color:#ff7070}
.def-subgrid-n{font-size:11px;color:rgba(255,255,255,0.62);display:block;margin-top:2px;font-weight:500}
.def-subgrid-empty{color:rgba(255,255,255,0.38)!important}
.def-subgrid-total-row td{background:transparent}
.def-tier-subgrid-table tbody tr:nth-child(odd) td{background:transparent}
.def-tier-subgrid-table tbody tr.def-subgrid-data-row:hover td{background:rgba(255,255,255,0.03)}
.section-label{font-family:'Bebas Neue',sans-serif;font-size:clamp(14px,1.08vw,16px);color:var(--muted);letter-spacing:2.5px;display:flex;align-items:center;gap:10px;margin-bottom:16px}
.section-label::after{content:'';flex:1;height:1px;background:rgba(255,255,255,0.08)}
.stat-grid{display:grid;gap:14px;margin-bottom:24px;width:100%;max-width:100%;box-sizing:border-box}
.stat-grid-4{grid-template-columns:repeat(4,1fr)}
.stat-grid-2{grid-template-columns:repeat(2,1fr)}
.stat-card{background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);
border:1px solid var(--glass-bd);border-radius:14px;padding:16px 18px;position:relative;overflow:hidden;transition:border-color .2s,box-shadow .2s;
box-shadow:0 4px 24px rgba(0,0,0,.22)}
.stat-card:hover{border-color:rgba(255,255,255,0.12);box-shadow:0 8px 32px rgba(0,0,0,.28)}
.stat-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px}
.stat-card.green::before{background:linear-gradient(90deg,var(--green),transparent)}
.stat-card.blue::before{background:linear-gradient(90deg,var(--cyan),transparent)}
.stat-card.amber::before{background:linear-gradient(90deg,var(--gold),transparent)}
.stat-card.red::before{background:linear-gradient(90deg,var(--red),transparent)}
.stat-card.purple::before{background:linear-gradient(90deg,var(--purple),transparent)}
.stat-label{font-family:'Bebas Neue',sans-serif;font-size:11px;color:var(--muted);letter-spacing:2.2px;margin-bottom:8px}
.stat-val{font-family:'Bebas Neue',sans-serif;font-size:36px;letter-spacing:2px;line-height:1}
.stat-sub{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted2);margin-top:5px}
.stat-sub strong{font-weight:700}
.table-wrap{background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);
border:1px solid var(--glass-bd);border-radius:14px;overflow:hidden;margin-bottom:20px;box-shadow:0 4px 24px rgba(0,0,0,.18)}
table{width:100%;border-collapse:collapse;font-size:clamp(13px,1.05vw,15px);font-family:'Inter',sans-serif}
th{font-family:'Bebas Neue',sans-serif;font-size:clamp(14px,1.15vw,16px);letter-spacing:2px;color:var(--muted);padding:10px 12px;text-align:left;
background:rgba(0,0,0,0.22);backdrop-filter:blur(12px);border-bottom:1px solid var(--glass-bd);white-space:nowrap}
th.right{text-align:right}
td{padding:10px 12px;border-bottom:1px solid rgba(255,255,255,.06);vertical-align:middle}
tr:last-child td{border-bottom:none}
.matrix-collapsible tr:hover td,.table-wrap table.table-sortable tr:hover td{background:rgba(255,255,255,.03)}
.table-wrap table:not(.table-sortable) tr:hover td{background:transparent}
td.right{text-align:right}td.mono{font-family:'Inter',sans-serif;font-size:clamp(13px,1.05vw,15px)}
.rate-cell{display:flex;align-items:center;gap:10px}
.rate-bar-bg{display:flex;flex:1;height:6px;background:rgba(255,255,255,.08);border-radius:3px;overflow:hidden}
.rate-bar-fill{height:100%;flex-shrink:0;transition:width .4s}
.rate-bar-fill:only-child{border-radius:3px}
.rate-bar-fill:first-child:not(:last-child){border-radius:3px 0 0 3px}
.rate-bar-fill:last-child:not(:first-child){border-radius:0 3px 3px 0}
.rate-bar-fill--hit{background:var(--green)}
.rate-bar-fill--miss{background:var(--red)}
.rate-bar-fill--empty{background:rgba(255,255,255,.06)}
.rate-num{font-family:'Inter',sans-serif;font-size:clamp(13px,1.05vw,15px);min-width:48px;text-align:right;flex-shrink:0}
.chip{display:inline-flex;align-items:center;flex-shrink:0;min-width:max-content;border-radius:8px;padding:3px 10px;font-size:12px;font-weight:700;font-family:'Bebas Neue',sans-serif;letter-spacing:.35px;
background:rgba(255,255,255,0.04);backdrop-filter:blur(12px);border:1px solid var(--glass-bd);box-sizing:border-box;vertical-align:middle;white-space:nowrap!important;word-break:normal!important;overflow-wrap:normal!important;hyphens:none!important;max-width:none}
.chip-a,.chip-b,.chip-c,.chip-d,.chip-goblin,.chip-demon,.chip-std{display:inline-flex;align-items:center;flex-shrink:0;min-width:max-content;white-space:nowrap!important;word-break:normal!important;overflow-wrap:normal!important;hyphens:none!important}
.chip-a{background:rgba(57,255,110,.08);color:var(--green);border-color:rgba(57,255,110,.28)}
.chip-b{background:rgba(0,229,255,.08);color:var(--cyan);border-color:rgba(0,229,255,.28)}
.chip-c{background:rgba(240,165,0,.08);color:var(--gold);border-color:rgba(240,165,0,.3)}
.chip-d{background:rgba(255,255,255,.04);color:var(--muted);border-color:var(--glass-bd)}
.chip-goblin{background:rgba(196,165,255,.10);color:var(--purple);border-color:rgba(196,165,255,.32)}
.chip-demon{background:rgba(255,77,77,.10);color:var(--red);border-color:rgba(255,77,77,.32)}
.chip-std{background:rgba(0,229,255,.08);color:var(--cyan);border-color:rgba(0,229,255,.25)}
.two-col{display:grid;grid-template-columns:repeat(2,minmax(0,1fr));gap:20px;margin-bottom:20px;width:100%;max-width:100%;box-sizing:border-box}
.three-col{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:12px;margin-bottom:20px;width:100%;max-width:100%;box-sizing:border-box}
.insight-grid{display:grid;grid-template-columns:repeat(3,minmax(0,1fr));gap:14px;margin-bottom:24px;width:100%;max-width:100%;box-sizing:border-box;min-width:0}
.sport-section:last-child .insight-grid{margin-bottom:0}
/* Last card is alone on its row when count is 1,4,7,… — span full width instead of a narrow left column */
.insight-grid>.insight-card:last-child:nth-child(3n+1){grid-column:1 / -1}
.insight-card{background:var(--glass);backdrop-filter:blur(20px);border:1px solid var(--glass-bd);border-radius:12px;padding:14px 16px;box-shadow:0 4px 20px rgba(0,0,0,.15);min-width:0;box-sizing:border-box}
.insight-icon{font-size:22px;margin-bottom:8px}
.insight-title{font-weight:700;font-size:14px;margin-bottom:6px;font-family:'Bebas Neue',sans-serif;letter-spacing:1px;color:var(--gold)}
.insight-body{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted2);line-height:1.6}
.insight-body strong{color:var(--text)}
tr.player-hit td:first-child{border-left:3px solid var(--green)}
tr.player-miss td:first-child{border-left:3px solid var(--red)}
tr.player-warn td:first-child{border-left:3px solid var(--gold)}
.pos{color:var(--green);font-weight:700}.neg{color:var(--red);font-weight:700}.neu{color:var(--muted2)}
.alert{border-radius:12px;padding:14px 18px;margin-bottom:20px;border:1px solid;font-size:13px;line-height:1.6;backdrop-filter:blur(16px);max-width:100%;box-sizing:border-box}
.alert-red{background:rgba(255,77,77,.08);border-color:rgba(255,77,77,.35)}
.alert-green{background:rgba(57,255,110,.08);border-color:rgba(57,255,110,.32)}
.alert-amber{background:rgba(240,165,0,.08);border-color:rgba(240,165,0,.32)}
.alert-title{font-family:'Bebas Neue',sans-serif;font-weight:700;font-size:13px;letter-spacing:1px;margin-bottom:4px}
.sub-dir{display:inline-block;margin-right:8px}
.muted-note{font-family:'Inter',sans-serif;font-size:12px;color:var(--muted);padding:14px;text-align:center}
td.muted{color:var(--muted2)}
@media(max-width:768px){
.stat-grid{justify-items:start;justify-content:start}
.stat-grid-4,.stat-grid-2{grid-template-columns:repeat(auto-fill,minmax(min(100%,9rem),max-content))}
.two-col,.three-col{grid-template-columns:1fr}
.pbw-tier-grid{grid-template-columns:1fr}
.insight-grid{display:grid;grid-template-columns:1fr;gap:14px;width:100%;max-width:100%}
.insight-grid>.insight-card:last-child:nth-child(3n+1){grid-column:auto}
.insight-card{width:100%;max-width:100%;min-width:0}
.stat-card{width:fit-content;max-width:100%}
.alert{width:fit-content;max-width:100%}
.rate-num{min-width:min(48px,max-content)}
body{font-size:clamp(14px,3.2vw,16px)}
table{font-size:clamp(12px,3.2vw,14px)}
td.mono{font-size:clamp(12px,3.2vw,14px)}
.rate-num{font-size:clamp(12px,3.2vw,14px)}
th{font-size:clamp(12px,3.4vw,14px);padding:8px 8px}
td{padding:8px 8px}
.section-label{font-size:clamp(12px,3.2vw,14px)}
.sport-label{font-size:28px}
.logo-title{font-size:24px}
.two-col.pick-tier-split{grid-template-columns:1fr!important;gap:18px!important}
}
/* Touch / hub iframe: BY PICK TYPE + BY TIER stack when viewport math is wrong (wide iframe on a phone). */
@media(pointer:coarse){
.two-col.pick-tier-split{grid-template-columns:1fr!important;gap:18px!important}
}
/* v20260429mobilecol — ≤960px: tables/chips only (no blanket .two-col stack — desktop keeps 2-col). */
@media(max-width:960px){
.table-wrap{overflow-x:auto;-webkit-overflow-scrolling:touch}
.table-wrap table{min-width:0;width:100%;max-width:100%}
.two-col.pick-tier-split .table-wrap table{min-width:0}
.sport-header{flex-wrap:wrap;gap:6px}
.sport-meta-count{font-size:clamp(10px,2.5vw,12px)}
.chip,.chip-a,.chip-b,.chip-c,.chip-d,.chip-goblin,.chip-demon,.chip-std{font-size:clamp(10px,2.6vw,12px);padding:4px 8px;letter-spacing:.06em;display:inline-flex!important;align-items:center!important;flex-shrink:0!important;min-width:max-content!important;white-space:nowrap!important;word-break:normal!important;overflow-wrap:normal!important;hyphens:none!important;max-width:none!important}
.table-wrap .sub-dir{display:block;margin:5px 0 0;font-size:clamp(11px,2.8vw,13px);line-height:1.35}
.ou-breakdown{font-size:clamp(11px,2.8vw,13px)!important;line-height:1.4!important}
.two-col.pick-tier-split td.mono{font-size:clamp(11px,2.8vw,13px)}
.two-col.pick-tier-split td:first-child,.two-col.pick-tier-split th:first-child{min-width:0}
.two-col.pick-tier-split .table-wrap td{vertical-align:top}
.two-col.pick-tier-split .table-wrap th{white-space:normal;word-break:normal;line-height:1.2}
}
th[data-sort-key]{cursor:pointer;user-select:none;position:relative;padding-right:1.35em}
th[data-sort-key]:hover{color:var(--text)}
th[data-sort-key]:focus-visible{outline:2px solid var(--cyan);outline-offset:2px;border-radius:4px}
th.sort-active.sort-asc::after{content:'\\25B2';font-size:0.55em;opacity:0.9;position:absolute;right:6px;top:50%;transform:translateY(-50%);letter-spacing:0}
th.sort-active.sort-desc::after{content:'\\25BC';font-size:0.55em;opacity:0.9;position:absolute;right:6px;top:50%;transform:translateY(-50%);letter-spacing:0}
"""


TABLE_SORT_JS = """
<script>
(function () {
  function pickOrder(tr, c) {
    var t = (tr.cells[c].textContent || "").toLowerCase();
    if (t.indexOf("goblin") >= 0) return 2;
    if (t.indexOf("demon") >= 0) return 3;
    if (t.indexOf("standard") >= 0) return 1;
    return 0;
  }
  function tierOrder(tr, c) {
    var m = (tr.cells[c].textContent || "").match(/TIER\\s*([ABCD])/i);
    if (!m) return 0;
    return "ABCD".indexOf(m[1].toUpperCase()) + 1;
  }
  function dirOrder(tr, c) {
    var t = tr.cells[c].textContent || "";
    if (/OVER/i.test(t)) return 1;
    if (/UNDER/i.test(t)) return 2;
    return 0;
  }
  function numCell(tr, c) {
    var t = (tr.cells[c].textContent || "").trim().replace(/,/g, "");
    if (t === "\\u2014" || t === "-" || t === "") return null;
    var n = parseInt(t, 10);
    return isNaN(n) ? null : n;
  }
  function pctCell(tr, c) {
    var t = (tr.cells[c].textContent || "").trim();
    if (t === "\\u2014" || t === "-" || t === "") return null;
    var m = t.match(/([\\d.]+)\\s*%/);
    if (m) return parseFloat(m[1]);
    return null;
  }
  function barPct(tr, c) {
    var td = tr.cells[c];
    var hitFill = td.querySelector(".rate-bar-fill--hit");
    if (hitFill && hitFill.style && hitFill.style.width) {
      var hw = parseFloat(hitFill.style.width);
      if (!isNaN(hw)) return hw;
    }
    var fill = td.querySelector(".rate-bar-fill");
    if (fill && fill.style && fill.style.width) {
      var w = parseFloat(fill.style.width);
      if (!isNaN(w)) return w;
    }
    return pctCell(tr, c);
  }
  function hydrateStackedRateBars(table) {
    var ths = table.querySelectorAll("thead th");
    if (!ths.length) return;
    var headers = Array.prototype.map.call(ths, function (th) {
      return (th.textContent || "").trim().toUpperCase();
    });
    var hi = headers.indexOf("H");
    if (hi < 0) hi = headers.indexOf("HITS");
    var mi = headers.indexOf("M");
    if (mi < 0) mi = headers.indexOf("MISSES");
    if (hi < 0 || mi < 0) return;
    var rateIdx = -1;
    for (var i = 0; i < headers.length; i++) {
      if (headers[i] === "RATE" || headers[i].indexOf("HIT RATE") >= 0) {
        rateIdx = i;
        break;
      }
    }
    var barIdx = headers.indexOf("BAR");
    table.querySelectorAll("tbody tr").forEach(function (tr) {
      var h = parseInt((tr.cells[hi].textContent || "").replace(/,/g, ""), 10) || 0;
      var m = parseInt((tr.cells[mi].textContent || "").replace(/,/g, ""), 10) || 0;
      var tot = h + m;
      if (tot <= 0) return;
      var hitPct = h / tot * 100;
      var missPct = m / tot * 100;
      var cell = null;
      if (rateIdx >= 0 && tr.cells[rateIdx]) {
        cell = tr.cells[rateIdx];
      } else if (barIdx >= 0 && tr.cells[barIdx]) {
        cell = tr.cells[barIdx];
      }
      if (!cell) return;
      var bg = cell.querySelector(".rate-bar-bg");
      if (!bg) return;
      var segs = "";
      if (hitPct > 0) {
        segs += '<div class="rate-bar-fill rate-bar-fill--hit" style="width:' + hitPct.toFixed(1) + '%;background:var(--green)"></div>';
      }
      if (missPct > 0) {
        segs += '<div class="rate-bar-fill rate-bar-fill--miss" style="width:' + missPct.toFixed(1) + '%;background:var(--red)"></div>';
      }
      bg.style.display = "flex";
      bg.innerHTML = segs;
    });
  }
  var extractors = {
    pick: pickOrder,
    tier: tierOrder,
    dir: dirOrder,
    decided: numCell,
    hits: numCell,
    misses: numCell,
    rate: pctCell,
    bar: barPct
  };
  function cmp(a, b, key, col, dir) {
    var ex = extractors[key];
    if (!ex) return 0;
    var va = ex(a, col);
    var vb = ex(b, col);
    if (va === null && vb === null) return 0;
    if (va === null) return 1;
    if (vb === null) return -1;
    if (va < vb) return -dir;
    if (va > vb) return dir;
    return 0;
  }
  function initTable(table) {
    hydrateStackedRateBars(table);
    var ths = table.querySelectorAll("thead th[data-sort-key]");
    ths.forEach(function (th) {
      th.setAttribute("tabindex", "0");
      th.setAttribute("role", "button");
      var col = th.cellIndex;
      var key = th.getAttribute("data-sort-key");
      function applySort() {
        var prevCol = parseInt(table.getAttribute("data-sort-col") || "-1", 10);
        var prevDir = parseInt(table.getAttribute("data-sort-dir") || "1", 10);
        var dir = prevCol === col ? -prevDir : 1;
        table.setAttribute("data-sort-col", String(col));
        table.setAttribute("data-sort-dir", String(dir));
        table.querySelectorAll("thead th.sort-active").forEach(function (x) {
          x.classList.remove("sort-active", "sort-asc", "sort-desc");
        });
        th.classList.add("sort-active", dir === 1 ? "sort-asc" : "sort-desc");
        var tbody = table.querySelector("tbody");
        if (!tbody) return;
        var rows = Array.prototype.slice.call(tbody.querySelectorAll("tr"));
        rows.sort(function (r1, r2) { return cmp(r1, r2, key, col, dir); });
        rows.forEach(function (r) { tbody.appendChild(r); });
      }
      th.addEventListener("click", applySort);
      th.addEventListener("keydown", function (e) {
        if (e.key === "Enter" || e.key === " ") {
          e.preventDefault();
          applySort();
        }
      });
    });
  }
  document.querySelectorAll("table.table-sortable").forEach(initTable);
  document.querySelectorAll("table").forEach(function (table) {
    if (!table.classList.contains("table-sortable")) hydrateStackedRateBars(table);
  });
})();
</script>
"""


PROP_BREAKDOWN_WIDGET_JS = """
<script>
(function () {
  const PT_ORDER = ["Standard Over","Standard Under","Goblin","Demon","Over","Under"];
  const TIER_ORDER = ["A","B","C","D"];
  const HEAT_COLS = ["A","B","C","D","Overall"];
  function pct(hits, decided) { return decided > 0 ? (hits / decided * 100) : null; }
  function esc(s){ return String(s ?? "").replace(/[&<>"']/g, c => ({'&':'&amp;','<':'&lt;','>':'&gt;','"':'&quot;',"'":'&#39;'}[c])); }
  function parseRows(root){
    try { return JSON.parse(root.getAttribute("data-prop-breakdown-rows") || "[]"); }
    catch { return []; }
  }
  function cellColor(v){
    if (v == null) return "#1b1f2f";
    const stops = [[0,"#F09595"],[28,"#FAC775"],[42,"#D3D1C7"],[55,"#9FE1CB"],[68,"#1D9E75"],[100,"#085041"]];
    if (v <= stops[0][0]) return stops[0][1];
    for (let i=1;i<stops.length;i++){
      const [x2,c2] = stops[i], [x1,c1] = stops[i-1];
      if (v <= x2){
        const t = (v-x1)/(x2-x1);
        const a = [c1.slice(1,3),c1.slice(3,5),c1.slice(5,7)].map(h=>parseInt(h,16));
        const b = [c2.slice(1,3),c2.slice(3,5),c2.slice(5,7)].map(h=>parseInt(h,16));
        const m = a.map((n,j)=>Math.round(n + (b[j]-n)*t));
        return "#" + m.map(n=>n.toString(16).padStart(2,"0")).join("");
      }
    }
    return "#085041";
  }
  function textColor(v){ return (v != null && (v < 30 || v > 65)) ? "#fff" : "#141414"; }
  function requestParentResize(){
    try { window.parent.postMessage({ type: "resizeRequest" }, "*"); } catch (_e) {}
  }
  function filterRows(rows, minDecided, tierFilter, ptFilter){
    return rows.filter(r => (r.decided || 0) >= minDecided)
      .filter(r => tierFilter === "ALL" || r.tier === tierFilter)
      .filter(r => ptFilter === "ALL" || r.pt === ptFilter);
  }
  function render(root){
    const rows = parseRows(root);
    if (!rows.length){ root.innerHTML = '<div class="muted-note">No prop breakdown rows available.</div>'; return; }
    root.innerHTML = `
      <div class="pbw-controls">
        <div class="pbw-control"><label>Min decided</label><select data-k="minDec"><option value="5">≥5</option><option value="10">≥10</option><option value="20">≥20</option></select></div>
        <div class="pbw-control"><label>Show tiers</label><select data-k="tier"><option value="ALL">All tiers</option><option value="A">Tier A</option><option value="B">Tier B</option><option value="C">Tier C</option><option value="D">Tier D</option></select></div>
        <div class="pbw-control"><label>Show pick types</label><select data-k="pt"><option value="ALL">All</option>${PT_ORDER.map(x=>`<option value="${esc(x)}">${esc(x)}</option>`).join("")}</select></div>
      </div>
      <div class="pbw-tabs">
        <button class="pbw-tab-btn active" data-tab="bw">Best / Worst</button>
        <button class="pbw-tab-btn" data-tab="hm">Heatmap</button>
      </div>
      <div class="pbw-tab-panel active" data-tab-panel="bw"></div>
      <div class="pbw-tab-panel" data-tab-panel="hm"></div>
    `;
    const controls = {
      minDec: root.querySelector('select[data-k="minDec"]'),
      tier: root.querySelector('select[data-k="tier"]'),
      pt: root.querySelector('select[data-k="pt"]'),
    };
    function redraw(){
      const minDec = parseInt(controls.minDec.value, 10) || 5;
      const tierFilter = controls.tier.value;
      const ptFilter = controls.pt.value;
      const filtered = filterRows(rows, minDec, tierFilter, ptFilter);
      const bwPanel = root.querySelector('[data-tab-panel="bw"]');
      const hmPanel = root.querySelector('[data-tab-panel="hm"]');
      let bwHtml = "";
      const ptList = ptFilter === "ALL" ? PT_ORDER : [ptFilter];
      ptList.forEach(pt => {
        const ptRows = filtered.filter(r => r.pt === pt);
        if (!ptRows.length) return;
        bwHtml += `<div class="pbw-pt-header">${esc(pt)}</div>`;
        TIER_ORDER.forEach(tier => {
          if (tierFilter !== "ALL" && tier !== tierFilter) return;
          const tierRows = ptRows.filter(r => r.tier === tier);
          if (!tierRows.length) return;
          const byProp = {};
          tierRows.forEach(r => {
            const p = String(r.prop || "Unknown");
            if (!byProp[p]) byProp[p] = {prop:p,decided:0,hits:0};
            byProp[p].decided += Number(r.decided || 0);
            byProp[p].hits += Number(r.hits || 0);
          });
          const props = Object.values(byProp).map(x => ({...x, rate: pct(x.hits, x.decided)}))
            .filter(x => x.decided >= minDec)
            .sort((a,b) => (b.rate - a.rate) || (b.decided - a.decided));
          if (!props.length) return;
          const best = props.slice(0,5);
          const worst = [...props].sort((a,b)=>(a.rate-b.rate) || (b.decided-a.decided)).slice(0,5);
          function panel(title, cls, rows, barColor){
            return `<div><div class="pbw-panel-title ${cls}">${title}</div><div class="table-wrap"><table><thead><tr><th>PROP</th><th class="right">DEC</th><th>HIT%</th></tr></thead><tbody>${
              rows.map(r=>`<tr><td class="mono">${esc(r.prop)}</td><td class="right mono">${r.decided}</td><td><div class="rate-cell"><div class="rate-num" style="min-width:52px">${(r.rate||0).toFixed(1)}%</div><div class="pbw-mini-bar"><div class="pbw-mini-bar-fill" style="width:${Math.max(0,Math.min(100,r.rate||0)).toFixed(1)}%;background:${barColor}"></div></div></div></td></tr>`).join("")
            }</tbody></table></div></div>`;
          }
          bwHtml += `<div class="pbw-tier-title">TIER ${tier}</div><div class="pbw-tier-grid">${
            panel("Best Props","best",best,"#0F6E56") + panel("Worst Props","worst",worst,"#c35736")
          }</div>`;
        });
      });
      bwPanel.innerHTML = bwHtml || '<div class="muted-note">No qualifying rows for current filters.</div>';

      const byPropTier = {};
      filtered.forEach(r => {
        const p = String(r.prop || "Unknown");
        if (!byPropTier[p]) byPropTier[p] = {A:{d:0,h:0},B:{d:0,h:0},C:{d:0,h:0},D:{d:0,h:0},Overall:{d:0,h:0}};
        const t = String(r.tier || "");
        if (TIER_ORDER.includes(t)){
          byPropTier[p][t].d += Number(r.decided || 0);
          byPropTier[p][t].h += Number(r.hits || 0);
          byPropTier[p].Overall.d += Number(r.decided || 0);
          byPropTier[p].Overall.h += Number(r.hits || 0);
        }
      });
      const propRows = Object.entries(byPropTier)
        .map(([prop, d]) => ({prop, d, overallRate: pct(d.Overall.h, d.Overall.d)}))
        .filter(x => x.d.Overall.d >= minDec)
        .sort((a,b)=>(b.overallRate||-1)-(a.overallRate||-1));
      let hmHtml = `<div class="pbw-heatmap-wrap"><table class="pbw-heatmap"><thead><tr><th>PROP</th>${HEAT_COLS.map(c=>`<th>${c}</th>`).join("")}</tr></thead><tbody>`;
      propRows.forEach(r => {
        hmHtml += `<tr><td class="mono">${esc(r.prop)}</td>`;
        HEAT_COLS.forEach(c => {
          const d = r.d[c];
          const rate = pct(d.h, d.d);
          if (!d.d){
            hmHtml += `<td style="background:#141929;color:#80869a">—</td>`;
          } else {
            hmHtml += `<td style="background:${cellColor(rate)};color:${textColor(rate)}"><div class="pbw-cell-rate">${rate.toFixed(1)}%</div><div class="pbw-cell-dec">n=${d.d}</div></td>`;
          }
        });
        hmHtml += `</tr>`;
      });
      hmHtml += `</tbody></table></div>`;
      hmPanel.innerHTML = propRows.length ? hmHtml : '<div class="muted-note">No heatmap rows for current filters.</div>';
      requestParentResize();
    }
    Object.values(controls).forEach(el => el.addEventListener("change", redraw));
    root.querySelectorAll(".pbw-tab-btn").forEach(btn => {
      btn.addEventListener("click", function(){
        root.querySelectorAll(".pbw-tab-btn").forEach(x=>x.classList.toggle("active", x===btn));
        root.querySelectorAll(".pbw-tab-panel").forEach(p=>p.classList.toggle("active", p.getAttribute("data-tab-panel") === btn.getAttribute("data-tab")));
        requestParentResize();
      });
    });
    redraw();
  }
  document.querySelectorAll(".pbw-root").forEach(render);
  requestParentResize();
})();
</script>
"""


def build_html(date_str: str, nba_rows: list[dict], cbb_rows: list[dict],
               nba_path: Path | None, cbb_path: Path | None,
               nhl_rows: list[dict] | None = None,
               soccer_rows: list[dict] | None = None,
               mlb_rows: list[dict] | None = None,
               wnba_rows: list[dict] | None = None,
               tennis_rows: list[dict] | None = None,
               nhl_path: Path | None = None,
               soccer_path: Path | None = None,
               mlb_path: Path | None = None) -> str:
    try:
        d = datetime.strptime(date_str, "%Y-%m-%d")
        display_date = d.strftime("%b %d, %Y").upper()
    except ValueError:
        display_date = date_str.upper()

    nhl_rows    = nhl_rows    or []
    soccer_rows = soccer_rows or []
    mlb_rows    = mlb_rows    or []
    wnba_rows   = wnba_rows   or []
    tennis_rows = tennis_rows or []
    all_rows = (
        list(nba_rows)
        + list(cbb_rows)
        + list(nhl_rows)
        + list(soccer_rows)
        + list(mlb_rows)
        + list(wnba_rows)
        + list(tennis_rows)
    )
    all_section = build_sport_section(all_rows, "ALL SPORTS", "🌐") if all_rows else ""
    nba_section    = build_sport_section(nba_rows,    "NBA",    "🏀") if nba_rows    else ""
    cbb_section    = build_sport_section(cbb_rows,    "CBB",    "🎓") if cbb_rows    else ""
    nhl_section    = build_sport_section(nhl_rows,    "NHL",    "🏒") if nhl_rows    else ""
    soccer_section = build_sport_section(soccer_rows, "Soccer", "⚽") if soccer_rows else ""
    mlb_section    = build_sport_section(mlb_rows,    "MLB",    "⚾") if mlb_rows    else ""
    wnba_section   = build_sport_section(wnba_rows,   "WNBA",   "🏀") if wnba_rows   else ""
    tennis_section = build_sport_section(tennis_rows, "Tennis", "🎾") if tennis_rows else ""
    takeaways = build_takeaways(
        nba_rows,
        cbb_rows,
        nhl_rows=nhl_rows,
        soccer_rows=soccer_rows,
        mlb_rows=mlb_rows,
        wnba_rows=wnba_rows,
        tennis_rows=tennis_rows,
    )

    if not (
        nba_section
        or cbb_section
        or nhl_section
        or soccer_section
        or mlb_section
        or wnba_section
        or tennis_section
    ):
        body_content = """<div style="text-align:center;padding:60px 20px;font-family:'Inter',sans-serif">
          <div style="font-size:32px;margin-bottom:16px">📭</div>
          <div style="font-size:18px;color:rgba(255,255,255,0.55)">No graded data found for this date.</div>
          <div style="font-size:13px;color:rgba(255,255,255,0.4);margin-top:8px">
            Run <code style="color:var(--cyan)">run_grader.ps1 --date {date_str}</code> to generate grades.
          </div>
        </div>""".replace("{date_str}", date_str)
    else:
        body_content = (
            all_section
            + nba_section
            + cbb_section
            + nhl_section
            + soccer_section
            + mlb_section
            + wnba_section
            + tennis_section
            + takeaways
        )

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, viewport-fit=cover"/>
<meta name="theme-color" content="#0a0a14"/>
<meta name="apple-mobile-web-app-capable" content="yes"/>
<meta name="apple-mobile-web-app-status-bar-style" content="black-translucent"/>
<title>Slate Eval — {h(display_date)}</title>
<link rel="preconnect" href="https://fonts.googleapis.com"/>
<link rel="preconnect" href="https://fonts.gstatic.com" crossorigin/>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&family=Inter:wght@400;500;600;700&display=swap" rel="stylesheet"/>
<link rel="stylesheet" href="/static/global-scrollbar.css?v=20260517platform"/>
<link rel="stylesheet" href="/static/light-theme-dim-overrides.css?v=20260419perf2"/>
<link rel="stylesheet" href="/static/proporacle-mobile-schema.css?v=20260430schemapage"/>
<style>{CSS}</style>
</head>
<body>
<header>
  <div class="slate-header-top">
    <div class="date-badge">📅 {h(display_date)}</div>
  </div>
  <div id="proporacle-grades-toolbar-host" class="grades-hub-toolbar-host"></div>
</header>
<div class="main">
{body_content}
</div>
{TABLE_SORT_JS}
{PROP_BREAKDOWN_WIDGET_JS}
</body>
</html>"""


# ══════════════════════════════════════════════════════════════════════════════
#  MAIN
# ══════════════════════════════════════════════════════════════════════════════

def main() -> None:
    parser = argparse.ArgumentParser(
        description="Generate slate_eval_{date}.html from nba/cbb graded xlsx files."
    )
    parser.add_argument("--date", type=str, default="",
                        help="Date string YYYY-MM-DD (default: yesterday)")
    parser.add_argument("--nba",  type=str, default="",
                        help="Path to nba_graded_*.xlsx")
    parser.add_argument("--cbb",  type=str, default="",
                        help="Path to cbb_graded_*.xlsx")
    parser.add_argument("--nhl",  type=str, default="",
                        help="Path to nhl_graded_*.xlsx")
    parser.add_argument("--soccer", type=str, default="",
                        help="Path to soccer_graded_*.xlsx")
    parser.add_argument("--mlb", type=str, default="",
                        help="Path to graded_mlb_*.xlsx")
    parser.add_argument("--wnba", type=str, default="",
                        help="Path to graded_wnba_*.xlsx or wnba_graded_*.xlsx")
    parser.add_argument("--tennis", type=str, default="",
                        help="Path to graded_tennis_*.xlsx")
    parser.add_argument("--out",  type=str, default="",
                        help="Output path or directory (default: next to this script)")
    parser.add_argument(
        "--allow-empty",
        action="store_true",
        help="Write slate_eval even when no graded xlsx files exist (empty / no-data UI).",
    )
    args = parser.parse_args()

    # Resolve date
    if args.date:
        date_str = args.date.strip()
    else:
        date_str = (datetime.now() - timedelta(days=1)).strftime("%Y-%m-%d")
    print(f"  Date: {date_str}")

    # Resolve file paths
    nba_path: Path | None = None
    cbb_path: Path | None = None

    if args.nba:
        nba_path = Path(args.nba).resolve()
        if not nba_path.exists():
            print(f"  WARNING: NBA file not found: {nba_path}")
            nba_path = None
    else:
        nba_path = find_graded_file("nba", date_str)
        if nba_path:
            print(f"  Auto-detected NBA: {nba_path}")
        else:
            print(f"  WARNING: Could not auto-detect nba_graded_{date_str}.xlsx")

    if args.cbb:
        cbb_path = Path(args.cbb).resolve()
        if not cbb_path.exists():
            print(f"  WARNING: CBB file not found: {cbb_path}")
            cbb_path = None
    else:
        cbb_path = find_graded_file("cbb", date_str)
        if cbb_path:
            print(f"  Auto-detected CBB: {cbb_path}")
        else:
            print(f"  WARNING: Could not auto-detect cbb_graded_{date_str}.xlsx")

    nhl_path: Path | None = None
    if args.nhl:
        nhl_path = Path(args.nhl).resolve()
        if not nhl_path.exists():
            print(f"  WARNING: NHL file not found: {nhl_path}")
            nhl_path = None
    else:
        nhl_path = find_graded_file("nhl", date_str)
        if nhl_path:
            print(f"  Auto-detected NHL: {nhl_path}")

    nba1q_path: Path | None = find_graded_file("nba1q", date_str)
    if nba1q_path:
        print(f"  Auto-detected NBA1Q: {nba1q_path}")
    nba1h_path: Path | None = find_graded_file("nba1h", date_str)
    if nba1h_path:
        print(f"  Auto-detected NBA1H: {nba1h_path}")

    soccer_path: Path | None = None
    if args.soccer:
        soccer_path = Path(args.soccer).resolve()
        if not soccer_path.exists():
            print(f"  WARNING: Soccer file not found: {soccer_path}")
            soccer_path = None
    else:
        soccer_path = find_graded_file("soccer", date_str)
        if soccer_path:
            print(f"  Auto-detected Soccer: {soccer_path}")

    mlb_path: Path | None = None
    if args.mlb:
        mlb_path = Path(args.mlb).resolve()
        if not mlb_path.exists():
            print(f"  WARNING: MLB file not found: {mlb_path}")
            mlb_path = None
    else:
        mlb_path = find_graded_file("mlb", date_str)
        if mlb_path:
            print(f"  Auto-detected MLB: {mlb_path}")

    wnba_path: Path | None = None
    if args.wnba:
        wnba_path = Path(args.wnba).resolve()
        if not wnba_path.exists():
            print(f"  WARNING: WNBA file not found: {wnba_path}")
            wnba_path = None
    else:
        wnba_path = find_graded_file("wnba", date_str)
        if wnba_path:
            print(f"  Auto-detected WNBA: {wnba_path}")

    tennis_path: Path | None = None
    if args.tennis:
        tennis_path = Path(args.tennis).resolve()
        if not tennis_path.exists():
            print(f"  WARNING: Tennis file not found: {tennis_path}")
            tennis_path = None
    else:
        tennis_path = find_tennis_graded_file(date_str)
        if tennis_path:
            print(f"  Auto-detected Tennis: {tennis_path}")

    if not (
        nba_path
        or nba1q_path
        or nba1h_path
        or cbb_path
        or nhl_path
        or soccer_path
        or mlb_path
        or wnba_path
        or tennis_path
    ):
        if args.allow_empty:
            print("  NOTE: No graded files; emitting empty slate eval (--allow-empty).")
        else:
            print(
                "  ERROR: No graded files found. Specify --nba/--cbb/--nhl/--soccer/--mlb/--wnba/--tennis."
            )
            sys.exit(1)

    # Load rows
    nba_rows: list[dict] = []
    cbb_rows: list[dict] = []
    if nba_path:
        print(f"  Loading NBA: {nba_path.name} ...", end="", flush=True)
        nba_rows = load_graded(nba_path, "nba")
        print(f" {len(nba_rows):,} rows")
    nba1q_rows: list[dict] = []
    if nba1q_path:
        print(f"  Loading NBA1Q: {nba1q_path.name} ...", end="", flush=True)
        nba1q_rows = load_graded(nba1q_path, "nba1q")
        print(f" {len(nba1q_rows):,} rows")
    nba1h_rows: list[dict] = []
    if nba1h_path:
        print(f"  Loading NBA1H: {nba1h_path.name} ...", end="", flush=True)
        nba1h_rows = load_graded(nba1h_path, "nba1h")
        print(f" {len(nba1h_rows):,} rows")
    nba_rows_merged = [*nba_rows, *nba1q_rows, *nba1h_rows]
    if cbb_path:
        print(f"  Loading CBB: {cbb_path.name} ...", end="", flush=True)
        cbb_rows = load_graded(cbb_path, "cbb")
        print(f" {len(cbb_rows):,} rows")

    nhl_rows: list[dict] = []
    if nhl_path:
        print(f"  Loading NHL: {nhl_path.name} ...", end="", flush=True)
        nhl_rows = load_graded(nhl_path, "nhl")
        print(f" {len(nhl_rows):,} rows")

    soccer_rows: list[dict] = []
    if soccer_path:
        print(f"  Loading Soccer: {soccer_path.name} ...", end="", flush=True)
        soccer_rows = load_graded(soccer_path, "soccer")
        print(f" {len(soccer_rows):,} rows")

    mlb_rows: list[dict] = []
    if mlb_path:
        print(f"  Loading MLB: {mlb_path.name} ...", end="", flush=True)
        mlb_rows = load_graded(mlb_path, "mlb")
        print(f" {len(mlb_rows):,} rows")

    wnba_rows: list[dict] = []
    if wnba_path:
        print(f"  Loading WNBA: {wnba_path.name} ...", end="", flush=True)
        wnba_rows = load_graded(wnba_path, "wnba")
        print(f" {len(wnba_rows):,} rows")

    tennis_rows: list[dict] = []
    if tennis_path:
        print(f"  Loading Tennis: {tennis_path.name} ...", end="", flush=True)
        tennis_rows = load_graded(tennis_path, "tennis")
        print(f" {len(tennis_rows):,} rows")

    # Build HTML (use merged NBA so 1Q/1H rows appear in Slate Evaluation, not only in JSON)
    print("  Building HTML ...", end="", flush=True)
    html = build_html(
        date_str,
        nba_rows_merged,
        cbb_rows,
        nba_path,
        cbb_path,
        nhl_rows=nhl_rows,
        soccer_rows=soccer_rows,
        mlb_rows=mlb_rows,
        wnba_rows=wnba_rows,
        tennis_rows=tennis_rows,
        nhl_path=nhl_path,
        soccer_path=soccer_path,
        mlb_path=mlb_path,
    )
    print(f" {len(html):,} bytes")

    # Resolve output path
    out_name = f"slate_eval_{date_str}.html"
    if args.out:
        out_p = Path(args.out).resolve()
        if out_p.is_dir() or args.out.endswith(("\\","/")):
            out_p = out_p / out_name
        else:
            out_p = out_p  # treat as full file path
    else:
        out_p = ROOT_DIR / "ui_runner" / "templates" / out_name

    out_p.parent.mkdir(parents=True, exist_ok=True)
    out_p.write_text(html, encoding="utf-8")
    print(f"  Saved  -> {out_p}")

    json_bundles: list[tuple[str, list[dict]]] = []
    if nba_rows:
        json_bundles.append(("NBA", nba_rows))
    if nba1q_rows:
        json_bundles.append(("NBA1Q", nba1q_rows))
    if nba1h_rows:
        json_bundles.append(("NBA1H", nba1h_rows))
    if cbb_rows:
        json_bundles.append(("CBB", cbb_rows))
    if nhl_rows:
        json_bundles.append(("NHL", nhl_rows))
    if soccer_rows:
        json_bundles.append(("Soccer", soccer_rows))
    if mlb_rows:
        json_bundles.append(("MLB", mlb_rows))
    if wnba_rows:
        json_bundles.append(("WNBA", wnba_rows))
    if tennis_rows:
        json_bundles.append(("Tennis", tennis_rows))
    json_p = export_graded_props_json(date_str, out_p.parent, json_bundles)
    print(f"  Saved  -> {json_p}")
    all_rows = [
        *nba_rows_merged,
        *cbb_rows,
        *nhl_rows,
        *soccer_rows,
        *mlb_rows,
        *wnba_rows,
        *tennis_rows,
    ]
    tabs_p = export_analysis_tabs_xlsx(date_str, out_p.parent, all_rows)
    print(f"  Saved  -> {tabs_p}")
    print("  Done.")


if __name__ == "__main__":
    main()
