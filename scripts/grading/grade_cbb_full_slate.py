#!/usr/bin/env python3
"""
grade_cbb_full_slate.py (v4 — pivot fix + 3pm fix + NBA-style formatting)

Upgrades:
- Robust actuals matching using (event_id, athlete_id) when available
- Falls back through multiple match strategies (ID, date+ID, team+name, name)
- Handles duplicate actuals rows by choosing max MIN within a match bucket
- Adds match_method + void_reason columns
- More tolerant prop normalization + actual stat column name variants
- Pivots actuals long→wide so stat_from_row() can read PTS/REB/3PM etc. as columns
- Fixes UNSUPPORTED_PROP on 3-PT Made (3pm)
- NBA-style Excel formatting on all output sheets
- Backward compatible: accepts optional --date (ignored)

Example:
  py -3.14 grade_cbb_full_slate.py ^
    --slate outputs\\2026-02-27\\cbb_slate_extracted_2026-02-27.csv ^
    --actuals outputs\\2026-02-27\\cbb_actuals_2026-02-27.csv ^
    --out outputs\\2026-02-27\\cbb_graded_2026-02-27.xlsx
"""

from __future__ import annotations

import argparse
import re
import unicodedata
from typing import Optional, Tuple, Dict, Any, List

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ──────────────────────────────────────────────────────────────────────────────
# Normalization helpers
# ──────────────────────────────────────────────────────────────────────────────
def _norm_name(s: Any) -> str:
    s = unicodedata.normalize("NFKD", str(s or "")).encode("ascii", "ignore").decode("ascii")
    s = s.lower().strip()
    s = re.sub(r"[^a-z0-9 ]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _clean_id(s: Any) -> str:
    x = str(s or "").strip()
    if x.lower() in ("", "none", "nan", "--"):
        return ""
    # keep digits and separators for combos
    x = x.replace(",", "").replace(" ", "")
    return x


def _split_combo_ids(s: str) -> List[str]:
    # supports "123|456" or "123/456" or "123,456"
    s = _clean_id(s)
    if not s:
        return []
    parts = re.split(r"[|/,]+", s)
    parts = [p for p in parts if p.isdigit()]
    # stable sort to avoid duplicates like "456|123"
    return sorted(list(dict.fromkeys(parts)))


def to_float(x) -> float:
    try:
        s = str(x).strip()
        if s.lower() in ("", "none", "nan", "--"):
            return np.nan
        return float(s)
    except Exception:
        return np.nan


def _first_existing_col(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
    for c in candidates:
        if c in df.columns:
            return c
    return None


# ──────────────────────────────────────────────────────────────────────────────
# Prop normalization + stat extraction
# ──────────────────────────────────────────────────────────────────────────────
_PROP_MAP = {
    # singles
    "pts": "pts", "points": "pts",
    "reb": "reb", "rebs": "reb", "rebounds": "reb",
    "ast": "ast", "assists": "ast",
    "stl": "stl", "steals": "stl",
    "blk": "blk", "blocks": "blk",
    "tov": "tov", "to": "tov", "turnovers": "tov",
    "3pm": "3pm", "3pt": "3pm", "3ptm": "3pm", "3ptmade": "3pm", "3-pt made": "3pm", "3s": "3pm",

    # combos
    "pr": "pr", "pts+reb": "pr", "pts+rebs": "pr",
    "pa": "pa", "pts+ast": "pa", "pts+asts": "pa",
    "ra": "ra", "reb+ast": "ra", "rebs+asts": "ra",
    "pra": "pra", "pts+rebs+asts": "pra", "pts+reb+ast": "pra", "pts+reb+asts": "pra",
    "stocks": "stocks", "stl+blk": "stocks", "blks+stls": "stocks",
}


def norm_prop(p: Any) -> str:
    x = str(p or "").strip().lower()
    x = x.replace(" ", "")
    if "fantasy" in x:
        return "fantasy"
    return _PROP_MAP.get(x, str(p or "").strip().lower())


def _get_stat(actuals_row: pd.Series, keys: List[str]) -> float:
    for k in keys:
        if k in actuals_row.index:
            v = to_float(actuals_row.get(k))
            if not np.isnan(v):
                return v
    return np.nan


def stat_from_row(actuals_row: pd.Series, prop_norm: str) -> float:
    p = norm_prop(prop_norm)

    # tolerate multiple naming conventions in actuals
    pts = _get_stat(actuals_row, ["PTS", "points", "Points"])
    reb = _get_stat(actuals_row, ["REB", "reb", "Rebounds"])
    ast = _get_stat(actuals_row, ["AST", "ast", "Assists"])
    stl = _get_stat(actuals_row, ["STL", "stl", "Steals"])
    blk = _get_stat(actuals_row, ["BLK", "blk", "Blocks"])
    tov = _get_stat(actuals_row, ["TO", "tov", "Turnovers"])
    pm3 = _get_stat(actuals_row, ["3PM", "3PT", "3pm", "3PTM", "3pt_made", "3-PT Made", "FG3M", "3FGM"])

    if p == "pts": return pts
    if p == "reb": return reb
    if p == "ast": return ast
    if p == "stl": return stl
    if p == "blk": return blk
    if p == "tov": return tov
    if p == "3pm": return pm3

    if p == "pr":     return pts + reb
    if p == "pa":     return pts + ast
    if p == "ra":     return reb + ast
    if p == "pra":    return pts + reb + ast
    if p == "stocks": return stl + blk
    if p == "fantasy":
        # same formula you’re using now
        return pts + 1.2*reb + 1.5*ast + 3*stl + 3*blk - tov

    return np.nan


# ──────────────────────────────────────────────────────────────────────────────
# Grading
# ──────────────────────────────────────────────────────────────────────────────
def grade_row(actual_value: float, line: float, dir_played: str) -> str:
    if np.isnan(actual_value) or np.isnan(line) or not dir_played:
        return "VOID"
    d = dir_played.strip().upper()
    if abs(actual_value - line) < 1e-9:
        return "PUSH"
    if d == "OVER":
        return "HIT" if actual_value > line else "MISS"
    if d == "UNDER":
        return "HIT" if actual_value < line else "MISS"
    return "VOID"


def _choose_max_min(df: pd.DataFrame, min_col: str) -> pd.Series:
    if df.empty:
        return pd.Series(dtype=object)
    if min_col not in df.columns:
        return df.iloc[0]
    mins = pd.to_numeric(df[min_col], errors="coerce").fillna(-1)
    return df.loc[mins.idxmax()]


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--slate",   required=True, help="Slate CSV (step5b or step3b output)")
    ap.add_argument("--actuals", required=True, help="Actuals CSV from ESPN fetcher")
    ap.add_argument("--out",     required=True, help="Output .xlsx")
    ap.add_argument("--date",    required=False, help="Optional (ignored) — kept for backward compat")
    args = ap.parse_args()

    slate   = pd.read_csv(args.slate,   dtype=str).fillna("")
    actuals_raw = pd.read_csv(args.actuals, dtype=str).fillna("")

    # ── Pivot actuals from long to wide format ────────────────────────────────
    # fetch_actuals.py outputs one row per player+prop_type (long format).
    # stat_from_row() needs stat columns (PTS, REB, 3PM …) on each row (wide).
    # Strategy: if the actuals already have raw stat columns attached (new fetch_actuals),
    # use them directly. Otherwise pivot prop_type→actual to build wide columns.
    RAW_STAT_COLS = ["PTS","REB","AST","STL","BLK","TO","FGM","FGA",
                     "3PM","3PT","3PA","FTM","FTA","2PM","2PA","OREB","DREB","PF","MIN"]
    has_raw_cols = any(c in actuals_raw.columns for c in RAW_STAT_COLS)

    if has_raw_cols:
        # New fetch_actuals — raw stat columns already present, just use as-is
        actuals = actuals_raw.copy()
        print(f"  Actuals: wide format detected ({len(actuals)} rows, raw stat cols present)")
    else:
        # Old fetch_actuals — long format only. Pivot prop_type→value per player+team.
        # Map prop_type labels to canonical stat column names
        PROP_TO_STAT = {
            "Points": "PTS", "Rebounds": "REB", "Assists": "AST",
            "Steals": "STL", "Blocked Shots": "BLK", "Turnovers": "TO",
            "FG Made": "FGM", "FG Attempted": "FGA",
            "3-PT Made": "3PM", "3-PT Attempted": "3PA",
            "Free Throws Made": "FTM", "Free Throws Attempted": "FTA",
            "Two Pointers Made": "2PM", "Two Pointers Attempted": "2PA",
            "Offensive Rebounds": "OREB", "Defensive Rebounds": "DREB",
            "Personal Fouls": "PF",
        }
        id_cols = [c for c in ["player", "team", "espn_athlete_id", "MIN"] if c in actuals_raw.columns]
        pivot_rows = []
        for (player, team), grp in actuals_raw.groupby(["player", "team"], sort=False):
            row = {"player": player, "team": team}
            for _, r in grp.iterrows():
                pt = str(r.get("prop_type", "")).strip()
                av = r.get("actual", "")
                stat_col = PROP_TO_STAT.get(pt)
                if stat_col:
                    try:
                        row[stat_col] = float(av)
                    except (ValueError, TypeError):
                        pass
                # carry through MIN and espn_athlete_id if available
                for extra in ["MIN", "espn_athlete_id"]:
                    if extra in r.index and extra not in row:
                        row[extra] = r[extra]
            # also add 3PT alias
            if "3PM" in row:
                row["3PT"] = row["3PM"]
            pivot_rows.append(row)
        actuals = pd.DataFrame(pivot_rows).fillna("")
        print(f"  Actuals: long format detected — pivoted to {len(actuals)} player rows with stat columns")

    # Required slate columns
    for req in ("prop_norm", "line"):
        if req not in slate.columns:
            raise RuntimeError(f"Slate missing required column: {req}")

    # Normalize/ensure player_norm + espn_athlete_id
    if "espn_athlete_id" not in slate.columns:
        slate["espn_athlete_id"] = ""
    if "player_norm" not in slate.columns:
        if "player" in slate.columns:
            slate["player_norm"] = slate["player"].apply(_norm_name)
        else:
            raise RuntimeError("Slate missing both 'player' and 'player_norm'.")

    slate["espn_athlete_id"] = slate["espn_athlete_id"].apply(_clean_id)

    # Resolve direction column
    dir_col = _first_existing_col(
        slate,
        ["final_bet_direction", "model_dir_5", "bet_direction", "model_dir", "model_direction", "Direction"]
    )
    slate["_dir"] = slate[dir_col].astype(str).str.upper().str.strip() if dir_col else ""

    # Defensive tier/rank columns
    def_tier_col = _first_existing_col(slate, ["opp_def_tier", "def_tier", "Def Tier"])
    def_rank_col = _first_existing_col(slate, ["opp_def_rank", "OVERALL_DEF_RANK", "opp_def_adj_de"])

    # Pick type / tier
    pick_type_col = _first_existing_col(slate, ["pick_type", "Pick Type"])
    tier_col      = _first_existing_col(slate, ["tier", "Tier"])

    # Event/date/team columns — detect AFTER pivot so column names are correct
    slate_event_col  = _first_existing_col(slate,   ["event_id", "game_id", "espn_event_id"])
    actual_event_col = _first_existing_col(actuals, ["event_id", "game_id", "espn_event_id"])

    slate_date_col   = _first_existing_col(slate,   ["game_date", "date", "Game Date"])
    actual_date_col  = _first_existing_col(actuals, ["game_date", "date", "Game Date"])

    slate_team_col   = _first_existing_col(slate,   ["team", "Team", "team_abbr", "team_id"])
    actual_team_col  = _first_existing_col(actuals, ["team", "team_abbr", "team_id"])

    # Actuals: ensure name_norm + mins
    name_col = "player" if "player" in actuals.columns else ("player_name" if "player_name" in actuals.columns else None)
    actuals["_name_norm"] = actuals[name_col].apply(_norm_name) if name_col else ""
    if "MIN" in actuals.columns:
        actuals["_min_num"] = actuals["MIN"].apply(to_float)
    else:
        actuals["_min_num"] = np.nan

    # Build indices (as dataframes for flexible filtering)
    # Keep ALL rows (do NOT drop_duplicates globally by athlete_id — that’s a key robustness fix)
    # We choose max-min per match bucket during lookup.

    def lookup_actual(row: pd.Series) -> Tuple[Optional[pd.Series], str]:
        """Return (actual_row, method)."""
        aid_raw = str(row.get("espn_athlete_id", "")).strip()
        combo_ids = _split_combo_ids(aid_raw)
        pnorm = str(row.get("player_norm", "")).strip()
        teamv = str(row.get(slate_team_col, "")).strip() if slate_team_col else ""
        ev    = str(row.get(slate_event_col, "")).strip() if slate_event_col else ""
        dv    = str(row.get(slate_date_col, "")).strip() if slate_date_col else ""

        # Helper: filter actuals
        def f(df: pd.DataFrame) -> pd.DataFrame:
            out = df
            if actual_event_col and ev:
                out = out[out[actual_event_col].astype(str).str.strip() == ev]
            if actual_date_col and dv:
                out = out[out[actual_date_col].astype(str).str.strip() == dv]
            if actual_team_col and teamv:
                out = out[out[actual_team_col].astype(str).str.strip() == teamv]
            return out

        # Combo: sum components if possible (prefer same event/date when available)
        if len(combo_ids) >= 2:
            sub = f(actuals)
            found = []
            for cid in combo_ids:
                sub_id = sub[sub["espn_athlete_id"].astype(str).str.strip() == cid] if "espn_athlete_id" in actuals.columns else pd.DataFrame()
                if sub_id.empty:
                    # fallback without event/date/team filter
                    sub_id = actuals[actuals.get("espn_athlete_id", "").astype(str).str.strip() == cid] if "espn_athlete_id" in actuals.columns else pd.DataFrame()
                if not sub_id.empty:
                    found.append(_choose_max_min(sub_id, "_min_num"))
                else:
                    return (None, "COMBO_MISSING_COMPONENT")
            # We return a synthetic Series with summed stats by reusing first row and overwriting stat cols
            base = found[0].copy()
            # sum the primary stat columns used by stat_from_row
            for col in ["PTS", "REB", "AST", "STL", "BLK", "TO", "3PT", "3PM"]:
                if col in base.index:
                    base[col] = sum(to_float(r.get(col)) for r in found)
            base["_combo"] = "|".join(combo_ids)
            return (base, "COMBO_ID")

        # 1) event_id + athlete_id
        if aid_raw and aid_raw.isdigit() and actual_event_col and ev and "espn_athlete_id" in actuals.columns:
            sub = actuals[
                (actuals[actual_event_col].astype(str).str.strip() == ev) &
                (actuals["espn_athlete_id"].astype(str).str.strip() == aid_raw)
            ]
            if not sub.empty:
                return (_choose_max_min(sub, "_min_num"), "EVENT+ID")

        # 2) date + athlete_id
        if aid_raw and aid_raw.isdigit() and actual_date_col and dv and "espn_athlete_id" in actuals.columns:
            sub = actuals[
                (actuals[actual_date_col].astype(str).str.strip() == dv) &
                (actuals["espn_athlete_id"].astype(str).str.strip() == aid_raw)
            ]
            if not sub.empty:
                return (_choose_max_min(sub, "_min_num"), "DATE+ID")

        # 3) athlete_id (fallback)
        if aid_raw and aid_raw.isdigit() and "espn_athlete_id" in actuals.columns:
            sub = actuals[actuals["espn_athlete_id"].astype(str).str.strip() == aid_raw]
            if not sub.empty:
                return (_choose_max_min(sub, "_min_num"), "ID")

        # 4) name + team
        if pnorm and actual_team_col and teamv:
            sub = actuals[(actuals["_name_norm"] == pnorm) & (actuals[actual_team_col].astype(str).str.strip() == teamv)]
            if not sub.empty:
                return (_choose_max_min(sub, "_min_num"), "NAME+TEAM")

        # 5) name only
        if pnorm:
            sub = actuals[actuals["_name_norm"] == pnorm]
            if not sub.empty:
                return (_choose_max_min(sub, "_min_num"), "NAME")

        return (None, "NO_MATCH")

    # Grade each row
    actual_values = []
    actual_statuses = []
    match_methods = []
    void_reasons = []

    for _, r in slate.iterrows():
        line = to_float(r.get("line"))
        d = str(r.get("_dir", "")).strip().upper()
        prop = str(r.get("prop_norm", "")).strip()

        # Pre-void reasons
        if np.isnan(line):
            actual_values.append(np.nan)
            actual_statuses.append("MISSING_LINE")
            match_methods.append("")
            void_reasons.append("MISSING_LINE")
            continue
        if not d:
            # still compute actual_value to debug, but grade will VOID
            pass

        arow, method = lookup_actual(r)
        match_methods.append(method)

        if arow is None:
            actual_values.append(np.nan)
            actual_statuses.append("NO_ACTUAL_FOUND")
            void_reasons.append("NO_ACTUAL_MATCH")
            continue

        av = stat_from_row(arow, prop)
        actual_values.append(av)

        if np.isnan(av):
            actual_statuses.append("UNSUPPORTED_PROP")
            void_reasons.append("UNSUPPORTED_PROP")
        else:
            actual_statuses.append("OK")
            if not d:
                void_reasons.append("MISSING_DIR")
            else:
                void_reasons.append("")

    out = slate.copy()
    out["line_num"]       = out["line"].apply(to_float)
    out["actual_value"]   = actual_values
    out["dir_played"]     = out["_dir"]
    out["match_method"]   = match_methods
    out["actual_status"]  = actual_statuses
    out["void_reason"]    = void_reasons

    out["result"] = [
        grade_row(av, ln, d)
        for av, ln, d in zip(out["actual_value"], out["line_num"], out["dir_played"])
    ]

    out["diff"] = out["actual_value"] - out["line_num"]
    out.drop(columns=["_dir"], inplace=True, errors="ignore")

    # Console summary — exclude Demons from hit rate (data only)
    if pick_type_col:
        decided = out[(out["result"].isin(["HIT", "MISS"])) & (out[pick_type_col].astype(str) != "Demon")]
    else:
        decided = out[out["result"].isin(["HIT", "MISS"])]
    print("\n" + "="*55)
    print(" CBB GRADED SUMMARY (ROBUST)")
    print("="*55)
    print(f" Total rows   : {len(out)}")
    print(f" Actual OK    : {(out['actual_status']=='OK').sum()}")
    print(f" HIT          : {(out['result']=='HIT').sum()}")
    print(f" MISS         : {(out['result']=='MISS').sum()}")
    print(f" PUSH         : {(out['result']=='PUSH').sum()}")
    print(f" VOID         : {(out['result']=='VOID').sum()}")

    if len(decided) > 0:
        overall_hr = (decided["result"] == "HIT").sum() / len(decided)
        print(f" Hit Rate     : {overall_hr:.1%} ({(decided['result']=='HIT').sum()}/{len(decided)} decided)")

    # Void reasons
    vr = (out[out["result"] == "VOID"]["void_reason"].replace("", "UNSPECIFIED").value_counts())
    if not vr.empty:
        print("\n VOID REASONS")
        for k, v in vr.head(12).items():
            print(f"  - {k}: {v}")

    # Match methods
    mm = out["match_method"].replace("", "NONE").value_counts()
    if not mm.empty:
        print("\n MATCH METHODS")
        for k, v in mm.items():
            print(f"  - {k}: {v}")
    print("="*55 + "\n")

    # Build Excel sheets
    def build_summary_block(df: pd.DataFrame, label_col: str, label_vals: list, title: str) -> pd.DataFrame:
        # Exclude Demons from grading — data only, hard-to-hit by design
        if pick_type_col and label_col != pick_type_col:
            df_grade = df[df[pick_type_col].astype(str) != "Demon"]
        else:
            df_grade = df
        decided2 = df_grade[df_grade["result"].isin(["HIT", "MISS"])]
        rows = []
        for val in label_vals:
            if title == "Pick Type" and str(val) == "Demon":
                # Demon row: show total counts but no hit rate
                demon_sub = df[df[label_col].astype(str) == "Demon"]
                voids = int(demon_sub["result"].isin(["VOID", "PUSH"]).sum())
                rows.append({
                    "Group": title,
                    "Label": "Demon",
                    "Hits": "—",
                    "Misses": "—",
                    "Voids": voids,
                    "Decided": "—",
                    "Hit Rate": "EXCL",
                })
                continue
            sub = decided2[decided2[label_col].astype(str) == str(val)]
            hits   = (sub["result"] == "HIT").sum()
            misses = (sub["result"] == "MISS").sum()
            voids  = (df[df[label_col].astype(str) == str(val)]["result"].isin(["VOID", "PUSH"]).sum())
            total  = hits + misses
            hr     = hits / total if total > 0 else np.nan
            rows.append({
                "Group": title,
                "Label": val if str(val) else "UNMAPPED",
                "Hits": hits,
                "Misses": misses,
                "Voids": voids,
                "Decided": total,
                "Hit Rate": round(hr, 4) if not np.isnan(hr) else "",
            })
        return pd.DataFrame(rows)

    def build_crosstab(df: pd.DataFrame, row_col: str, col_col: str, row_vals: list, col_vals: list) -> pd.DataFrame:
        decided2 = df[df["result"].isin(["HIT", "MISS"])]
        records = []
        for rv in row_vals:
            row_data = {"": rv if str(rv) else "UNMAPPED"}
            for cv in col_vals:
                sub = decided2[(decided2[row_col].astype(str) == str(rv)) &
                               (decided2[col_col].astype(str) == str(cv))]
                hits = (sub["result"] == "HIT").sum()
                misses = (sub["result"] == "MISS").sum()
                total = hits + misses
                hr = f"{hits/total:.1%}" if total > 0 else "—"
                row_data[f"{cv} H/M"] = f"{hits}/{misses}"
                row_data[f"{cv} HR"]  = hr
            records.append(row_data)
        return pd.DataFrame(records)

    summary_blocks = []

    overall_rows = []
    for result_val in ["HIT", "MISS", "PUSH", "VOID"]:
        overall_rows.append({"Group": "OVERALL", "Label": result_val, "Count": (out["result"] == result_val).sum()})
    overall_df = pd.DataFrame(overall_rows)

    # Add void reasons + match methods sheets
    void_df = out[out["result"] == "VOID"].copy()
    void_reason_df = void_df.assign(void_reason=void_df["void_reason"].replace("", "UNSPECIFIED")) \
                            .groupby("void_reason", dropna=False).size().reset_index(name="Count") \
                            .sort_values("Count", ascending=False)

    match_method_df = out.assign(match_method=out["match_method"].replace("", "NONE")) \
                         .groupby("match_method", dropna=False).size().reset_index(name="Count") \
                         .sort_values("Count", ascending=False)

    # By Def Tier
    if def_tier_col:
        tier_order = ["Elite", "Above Avg", "Avg", "Weak", ""]
        tier_block = build_summary_block(out, def_tier_col, tier_order, "Def Tier")
        summary_blocks.append(("By Def Tier", tier_block))

        if pick_type_col:
            pick_types = ["Goblin", "Demon", "Standard"]
            summary_blocks.append(("Def Tier x Pick Type", build_crosstab(out, def_tier_col, pick_type_col, tier_order, pick_types)))

        if tier_col:
            model_tiers = ["A", "B", "C", "D"]
            summary_blocks.append(("Def Tier x Model Tier", build_crosstab(out, def_tier_col, tier_col, tier_order, model_tiers)))

    if pick_type_col:
        summary_blocks.append(("By Pick Type", build_summary_block(out, pick_type_col, ["Goblin", "Demon", "Standard"], "Pick Type")))

    if tier_col:
        summary_blocks.append(("By Model Tier", build_summary_block(out, tier_col, ["A", "B", "C", "D"], "Model Tier")))

    # ── Extract structured data from summary_blocks ──────────────────────────
    def _get_block(name):
        for sn, df in summary_blocks:
            if sn == name:
                return df
        return pd.DataFrame()

    by_pt_df   = _get_block("By Pick Type")
    by_tier_df = _get_block("By Model Tier")
    by_def_df  = _get_block("By Def Tier")
    dxpt_df    = _get_block("Def Tier x Pick Type")
    dxtier_df  = _get_block("Def Tier x Model Tier")

    # ── PropOracle color palette (matches NBA graded file exactly) ───────────────
    _C = {
        "hdr_dark":   "FF1C1C1C",
        "hdr_blue":   "FF1A5276",
        "hdr_green":  "FF1E8449",
        "hdr_gold":   "FF7D6608",
        "hdr_teal":   "FF117A65",
        "hdr_red":    "FFC0392B",
        "white":      "FFFFFFFF",
        "black":      "FF000000",
        "row_alt":    "FFF2F3F4",
        "goblin":     "FFE8D5F5",
        "demon":      "FFFDEDEC",
        "standard":   "FFF2F3F4",
        "over":       "FFD6EAF8",
        "under":      "FFFDEBD0",
        "tier_a":     "FFD5F5E3",
        "tier_b":     "FFD6EAF8",
        "tier_c":     "FFFEF9E7",
        "tier_d":     "FFF5CBA7",
        "def_elite":  "FFD5F5E3",
        "def_above":  "FFEAF4FB",
        "def_avg":    "FFFEF9E7",
        "def_weak":   "FFFDEDEC",
        "hit_good":   "FFF39C12",
        "hit_bad":    "FFE74C3C",
        "result_hit": "FF27AE60",
        "result_miss":"FFE74C3C",
        "result_void":"FF717D7E",
    }
    _DEF_COLORS  = {"Elite":_C["def_elite"],"Above Avg":_C["def_above"],"Avg":_C["def_avg"],"Weak":_C["def_weak"]}
    _TIER_COLORS = {"A":_C["tier_a"],"B":_C["tier_b"],"C":_C["tier_c"],"D":_C["tier_d"]}
    _PT_COLORS   = {"Goblin":_C["goblin"],"Demon":_C["demon"],"Standard":_C["standard"]}

    def _thin_border():
        s = Side(style="thin", color="DDDDDD")
        return Border(left=s, right=s, top=s, bottom=s)

    def _cs(ws, row, col, value, bold=False, fc="FF000000", bg=None, align="center", size=9):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold, color=fc, name="Arial", size=size)
        if bg:
            c.fill = PatternFill("solid", start_color=bg)
        c.alignment = Alignment(horizontal=align, vertical="center")
        c.border = _thin_border()
        return c

    def _sec_hdr(ws, row, labels, bg):
        for ci, lbl in enumerate(labels, 1):
            _cs(ws, row, ci, lbl, bold=True, fc=_C["white"], bg=bg)
        ws.row_dimensions[row].height = 18

    def _hr_color(v):
        try:
            return _C["hit_good"] if float(v) >= 0.50 else _C["hit_bad"]
        except Exception:
            return _C["hit_good"]

    def _fmt_hr(v):
        try:
            return f"{float(v):.1%}"
        except Exception:
            return str(v)

    def _fmt_int(v):
        try:
            return int(float(v))
        except Exception:
            return v

    def _safe_int(v):
        try:
            return int(float(v))
        except Exception:
            return 0

    def _write_data_sheet(ws, df, col_widths_map):
        ws.freeze_panes = "A2"
        headers = list(df.columns)
        for ci, h in enumerate(headers, 1):
            c = ws.cell(row=1, column=ci, value=h)
            c.font = Font(bold=True, color="FFFFFFFF", name="Arial", size=9)
            c.fill = PatternFill("solid", start_color=_C["hdr_dark"])
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border = _thin_border()
            ws.column_dimensions[get_column_letter(ci)].width = col_widths_map.get(h, 12)
        ws.row_dimensions[1].height = 20

        for ri, row_vals in enumerate(df.itertuples(index=False), 2):
            row_bg = _C["row_alt"] if ri % 2 == 0 else _C["white"]
            for ci, val in enumerate(row_vals, 1):
                col_name = headers[ci - 1]
                display  = "" if (val is None or (isinstance(val, float) and np.isnan(val)) or str(val) == "nan") else val
                c = ws.cell(row=ri, column=ci, value=display)
                c.font      = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = _thin_border()
                if col_name == "tier":
                    bg = _TIER_COLORS.get(str(display), row_bg)
                    c.fill = PatternFill("solid", start_color=bg)
                    c.font = Font(bold=True, name="Arial", size=9)
                elif col_name == "pick_type":
                    c.fill = PatternFill("solid", start_color=_PT_COLORS.get(str(display), row_bg))
                elif col_name in ("bet_direction", "dir_played", "_dir"):
                    bg = _C["over"] if str(display) == "OVER" else _C["under"] if str(display) == "UNDER" else row_bg
                    c.fill = PatternFill("solid", start_color=bg)
                    c.font = Font(bold=True, name="Arial", size=9)
                elif col_name == "result":
                    if str(display) == "HIT":
                        c.fill = PatternFill("solid", start_color=_C["result_hit"])
                        c.font = Font(bold=True, color=_C["white"], name="Arial", size=9)
                    elif str(display) == "MISS":
                        c.fill = PatternFill("solid", start_color=_C["result_miss"])
                        c.font = Font(bold=True, color=_C["white"], name="Arial", size=9)
                    elif str(display) == "VOID":
                        c.fill = PatternFill("solid", start_color=_C["result_void"])
                        c.font = Font(bold=True, color=_C["white"], name="Arial", size=9)
                    else:
                        c.fill = PatternFill("solid", start_color=row_bg)
                elif col_name == "opp_def_tier":
                    c.fill = PatternFill("solid", start_color=_DEF_COLORS.get(str(display), row_bg))
                else:
                    c.fill = PatternFill("solid", start_color=row_bg)
        ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

    import datetime as _dt_mod, re as _re
    date_match = _re.search(r"(\d{4}-\d{2}-\d{2})", args.out)
    grade_date = date_match.group(1) if date_match else str(_dt_mod.date.today())
    gen_ts     = _dt_mod.datetime.now().strftime("%Y-%m-%d %H:%M")

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary ───────────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet("Summary")
    ws_sum.freeze_panes = "A3"
    ws_sum.merge_cells(start_row=1, start_column=1, end_row=1, end_column=8)
    _cs(ws_sum, 1, 1, f"CBB SLATE GRADE  |  {grade_date}  |  Generated {gen_ts}",
        bold=True, fc=_C["white"], bg=_C["hdr_dark"], align="center", size=12)
    ws_sum.row_dimensions[1].height = 22

    HDR8 = ["LABEL","Direction","Total Props","Decided","Hits","Misses","Voids","Hit Rate"]
    # Exclude Demons from overall hit rate — data only
    _out_grade = out[out[pick_type_col].astype(str) != "Demon"] if pick_type_col else out
    hits_ov   = int((_out_grade["result"] == "HIT").sum())
    misses_ov = int((_out_grade["result"] == "MISS").sum())
    voids_ov  = int((out["result"].isin(["VOID","PUSH"])).sum())  # voids from full slate
    dec_ov    = hits_ov + misses_ov
    total_ov  = len(out)
    hr_ov     = hits_ov / dec_ov if dec_ov > 0 else 0.0

    r = 2
    _sec_hdr(ws_sum, r, ["OVERALL"] + HDR8[1:], _C["hdr_blue"])
    r += 1
    _cs(ws_sum, r, 1, "Full Slate", bold=True, bg=_C["white"], align="left")
    _cs(ws_sum, r, 2, "ALL",        bg=_C["white"])
    _cs(ws_sum, r, 3, total_ov,     bg=_C["white"])
    _cs(ws_sum, r, 4, dec_ov,       bg=_C["white"])
    _cs(ws_sum, r, 5, hits_ov,      bg=_C["white"])
    _cs(ws_sum, r, 6, misses_ov,    bg=_C["white"])
    _cs(ws_sum, r, 7, voids_ov,     bg=_C["white"])
    _cs(ws_sum, r, 8, _fmt_hr(hr_ov), bold=True, fc=_C["white"], bg=_hr_color(hr_ov))

    r += 2
    _sec_hdr(ws_sum, r, ["BY PICK TYPE"] + HDR8[1:], _C["hdr_green"])
    if not by_pt_df.empty:
        for _, rd in by_pt_df.iterrows():
            r += 1
            label = str(rd.get("Label", ""))
            bg    = _PT_COLORS.get(label, _C["white"])
            hr    = float(rd["Hit Rate"]) if str(rd.get("Hit Rate","")).replace(".","").isdigit() else 0.0
            _cs(ws_sum, r, 1, label, bold=True, bg=_C["white"], align="left")
            _cs(ws_sum, r, 2, "ALL",  bold=True, bg=bg)
            _cs(ws_sum, r, 3, _fmt_int(_safe_int(rd.get("Hits",0))+_safe_int(rd.get("Misses",0))+_safe_int(rd.get("Voids",0))), bg=bg)
            _cs(ws_sum, r, 4, _fmt_int(rd.get("Decided",0)), bg=bg)
            _cs(ws_sum, r, 5, _fmt_int(rd.get("Hits",0)),    bg=bg)
            _cs(ws_sum, r, 6, _fmt_int(rd.get("Misses",0)),  bg=bg)
            _cs(ws_sum, r, 7, _fmt_int(rd.get("Voids",0)),   bg=bg)
            _cs(ws_sum, r, 8, _fmt_hr(hr), bold=True, fc=_C["white"], bg=_hr_color(hr))

    r += 2
    _sec_hdr(ws_sum, r, ["BY MODEL TIER"] + HDR8[1:], _C["hdr_gold"])
    if not by_tier_df.empty:
        for _, rd in by_tier_df.iterrows():
            r += 1
            label = str(rd.get("Label", ""))
            bg    = _TIER_COLORS.get(label, _C["white"])
            hr    = float(rd["Hit Rate"]) if str(rd.get("Hit Rate","")).replace(".","").isdigit() else 0.0
            _cs(ws_sum, r, 1, f"Tier {label}", bold=True, bg=_C["white"], align="left")
            _cs(ws_sum, r, 2, "ALL", bold=True, bg=bg)
            _cs(ws_sum, r, 3, _fmt_int(_safe_int(rd.get("Hits",0))+_safe_int(rd.get("Misses",0))+_safe_int(rd.get("Voids",0))), bg=bg)
            _cs(ws_sum, r, 4, _fmt_int(rd.get("Decided",0)), bg=bg)
            _cs(ws_sum, r, 5, _fmt_int(rd.get("Hits",0)),    bg=bg)
            _cs(ws_sum, r, 6, _fmt_int(rd.get("Misses",0)),  bg=bg)
            _cs(ws_sum, r, 7, _fmt_int(rd.get("Voids",0)),   bg=bg)
            _cs(ws_sum, r, 8, _fmt_hr(hr), bold=True, fc=_C["white"], bg=_hr_color(hr))

    r += 2
    _sec_hdr(ws_sum, r, ["BY DEF TIER"] + HDR8[1:], _C["hdr_teal"])
    if not by_def_df.empty:
        for _, rd in by_def_df.iterrows():
            r += 1
            label = str(rd.get("Label", ""))
            bg    = _DEF_COLORS.get(label, _C["white"])
            hr    = float(rd["Hit Rate"]) if str(rd.get("Hit Rate","")).replace(".","").isdigit() else 0.0
            _cs(ws_sum, r, 1, label, bold=True, bg=_C["white"], align="left")
            _cs(ws_sum, r, 2, "ALL", bold=True, bg=bg)
            _cs(ws_sum, r, 3, _fmt_int(_safe_int(rd.get("Hits",0))+_safe_int(rd.get("Misses",0))+_safe_int(rd.get("Voids",0))), bg=bg)
            _cs(ws_sum, r, 4, _fmt_int(rd.get("Decided",0)), bg=bg)
            _cs(ws_sum, r, 5, _fmt_int(rd.get("Hits",0)),    bg=bg)
            _cs(ws_sum, r, 6, _fmt_int(rd.get("Misses",0)),  bg=bg)
            _cs(ws_sum, r, 7, _fmt_int(rd.get("Voids",0)),   bg=bg)
            _cs(ws_sum, r, 8, _fmt_hr(hr), bold=True, fc=_C["white"], bg=_hr_color(hr))

    for col, w in zip("ABCDEFGH", [18,11,12,10,8,8,8,10]):
        ws_sum.column_dimensions[col].width = w

    # ── Box Raw ───────────────────────────────────────────────────────────────
    _col_w = {
        "Sport":9,"tier":7,"rank_score":11,"player":22,"team":8,"opp":8,
        "prop_label":20,"prop_norm":14,"pick_type":10,"line":7,
        "bet_direction":10,"edge":7,"proj":8,"hit_rate":10,
        "ml_prob":10,"ml_edge":10,"edge_score":11,"blended_score":12,
        "L5 Avg":9,"Szn Avg":9,"L5 Over":8,"L5 Under":9,
        "opp_def_tier":11,"game_time":14,"espn_athlete_id":16,
        "player_norm":20,"line_num":9,"actual_value":11,"dir_played":10,
        "match_method":13,"actual_status":14,"void_reason":18,"result":8,"diff":7,
    }
    ws_box = wb.create_sheet("Box Raw")
    _write_data_sheet(ws_box, out, _col_w)

    ws_dec = wb.create_sheet("Decided Only")
    _write_data_sheet(ws_dec, decided, _col_w)

    # ── By Pick Type ──────────────────────────────────────────────────────────
    ws_pt = wb.create_sheet("By Pick Type")
    ws_pt.freeze_panes = "A2"
    _sec_hdr(ws_pt, 1, ["Pick Type","Hits","Misses","Voids","Decided","Hit Rate"], _C["hdr_green"])
    for col, w in zip("ABCDEF", [14,11,11,11,11,11]): ws_pt.column_dimensions[col].width = w
    if not by_pt_df.empty:
        for ri, (_, rd) in enumerate(by_pt_df.iterrows(), 2):
            label = str(rd.get("Label",""))
            bg = _PT_COLORS.get(label, _C["white"])
            hr = float(rd["Hit Rate"]) if str(rd.get("Hit Rate","")).replace(".","").isdigit() else 0.0
            _cs(ws_pt, ri, 1, label, bold=True, bg=bg, align="left")
            _cs(ws_pt, ri, 2, _fmt_int(rd.get("Hits",0)),    bg=bg)
            _cs(ws_pt, ri, 3, _fmt_int(rd.get("Misses",0)),  bg=bg)
            _cs(ws_pt, ri, 4, _fmt_int(rd.get("Voids",0)),   bg=bg)
            _cs(ws_pt, ri, 5, _fmt_int(rd.get("Decided",0)), bg=bg)
            _cs(ws_pt, ri, 6, _fmt_hr(hr), bold=True, fc=_C["white"], bg=_hr_color(hr))

    # ── By Model Tier ─────────────────────────────────────────────────────────
    ws_mt = wb.create_sheet("By Model Tier")
    ws_mt.freeze_panes = "A2"
    _sec_hdr(ws_mt, 1, ["Model Tier","Hits","Misses","Voids","Decided","Hit Rate"], _C["hdr_gold"])
    for col, w in zip("ABCDEF", [14,11,11,11,11,11]): ws_mt.column_dimensions[col].width = w
    if not by_tier_df.empty:
        for ri, (_, rd) in enumerate(by_tier_df.iterrows(), 2):
            label = str(rd.get("Label",""))
            bg = _TIER_COLORS.get(label, _C["white"])
            hr = float(rd["Hit Rate"]) if str(rd.get("Hit Rate","")).replace(".","").isdigit() else 0.0
            _cs(ws_mt, ri, 1, f"Tier {label}", bold=True, bg=bg, align="left")
            _cs(ws_mt, ri, 2, _fmt_int(rd.get("Hits",0)),    bg=bg)
            _cs(ws_mt, ri, 3, _fmt_int(rd.get("Misses",0)),  bg=bg)
            _cs(ws_mt, ri, 4, _fmt_int(rd.get("Voids",0)),   bg=bg)
            _cs(ws_mt, ri, 5, _fmt_int(rd.get("Decided",0)), bg=bg)
            _cs(ws_mt, ri, 6, _fmt_hr(hr), bold=True, fc=_C["white"], bg=_hr_color(hr))

    # ── By Def Tier ───────────────────────────────────────────────────────────
    ws_dt = wb.create_sheet("By Def Tier")
    ws_dt.freeze_panes = "A2"
    _sec_hdr(ws_dt, 1, ["Def Tier","Hits","Misses","Voids","Decided","Hit Rate"], _C["hdr_teal"])
    for col, w in zip("ABCDEF", [14,11,11,11,11,11]): ws_dt.column_dimensions[col].width = w
    if not by_def_df.empty:
        for ri, (_, rd) in enumerate(by_def_df.iterrows(), 2):
            label = str(rd.get("Label",""))
            bg = _DEF_COLORS.get(label, _C["white"])
            hr = float(rd["Hit Rate"]) if str(rd.get("Hit Rate","")).replace(".","").isdigit() else 0.0
            _cs(ws_dt, ri, 1, label, bold=True, bg=bg, align="left")
            _cs(ws_dt, ri, 2, _fmt_int(rd.get("Hits",0)),    bg=bg)
            _cs(ws_dt, ri, 3, _fmt_int(rd.get("Misses",0)),  bg=bg)
            _cs(ws_dt, ri, 4, _fmt_int(rd.get("Voids",0)),   bg=bg)
            _cs(ws_dt, ri, 5, _fmt_int(rd.get("Decided",0)), bg=bg)
            _cs(ws_dt, ri, 6, _fmt_hr(hr), bold=True, fc=_C["white"], bg=_hr_color(hr))

    # ── Def Tier x Pick Type ──────────────────────────────────────────────────
    ws_dxpt = wb.create_sheet("Def Tier x Pick Type")
    ws_dxpt.freeze_panes = "A2"
    dxpt_hdrs = ["Def Tier","Goblin H/M","Goblin HR","Demon H/M","Demon HR","Standard H/M","Standard HR"]
    _sec_hdr(ws_dxpt, 1, dxpt_hdrs, _C["hdr_teal"])
    for col, w in zip("ABCDEFG", [14,13,11,13,11,14,11]): ws_dxpt.column_dimensions[col].width = w
    if not dxpt_df.empty:
        for ri, (_, rd) in enumerate(dxpt_df.iterrows(), 2):
            def_label = str(rd.iloc[0])
            bg = _DEF_COLORS.get(def_label, _C["white"])
            _cs(ws_dxpt, ri, 1, def_label, bold=True, bg=bg, align="left")
            for ci, col_h in enumerate(dxpt_hdrs[1:], 2):
                v = rd.get(col_h, "")
                if "HR" in col_h:
                    try:
                        hr_v = float(str(v).replace("%",""))/100
                        _cs(ws_dxpt, ri, ci, str(v), bold=True, fc=_C["white"], bg=_hr_color(hr_v))
                    except Exception:
                        _cs(ws_dxpt, ri, ci, str(v), bg=bg)
                else:
                    _cs(ws_dxpt, ri, ci, str(v), bg=bg)

    # ── Def Tier x Model Tier ─────────────────────────────────────────────────
    ws_dxt = wb.create_sheet("Def Tier x Model Tier")
    ws_dxt.freeze_panes = "A2"
    dxt_hdrs = ["Def Tier","A H/M","A HR","B H/M","B HR","C H/M","C HR","D H/M","D HR"]
    _sec_hdr(ws_dxt, 1, dxt_hdrs, _C["hdr_teal"])
    for col, w in zip("ABCDEFGHI", [14,11,9,11,9,11,9,11,9]): ws_dxt.column_dimensions[col].width = w
    if not dxtier_df.empty:
        for ri, (_, rd) in enumerate(dxtier_df.iterrows(), 2):
            def_label = str(rd.iloc[0])
            bg = _DEF_COLORS.get(def_label, _C["white"])
            _cs(ws_dxt, ri, 1, def_label, bold=True, bg=bg, align="left")
            for ci, col_h in enumerate(dxt_hdrs[1:], 2):
                v = rd.get(col_h, "")
                if "HR" in col_h:
                    try:
                        hr_v = float(str(v).replace("%",""))/100
                        _cs(ws_dxt, ri, ci, str(v), bold=True, fc=_C["white"], bg=_hr_color(hr_v))
                    except Exception:
                        _cs(ws_dxt, ri, ci, str(v), bg=bg)
                else:
                    _cs(ws_dxt, ri, ci, str(v), bg=bg)

    # ── Void Reasons ──────────────────────────────────────────────────────────
    ws_vr = wb.create_sheet("Void Reasons")
    ws_vr.freeze_panes = "A2"
    _sec_hdr(ws_vr, 1, ["Void Reason","Count"], _C["hdr_red"])
    ws_vr.column_dimensions["A"].width = 24
    ws_vr.column_dimensions["B"].width = 12
    for ri, (_, rd) in enumerate(void_reason_df.iterrows(), 2):
        row_bg = _C["row_alt"] if ri % 2 == 0 else _C["white"]
        _cs(ws_vr, ri, 1, str(rd["void_reason"]), bg=row_bg, align="left")
        _cs(ws_vr, ri, 2, _fmt_int(rd["Count"]),  bg=row_bg)

    # ── Match Methods ─────────────────────────────────────────────────────────
    ws_mm = wb.create_sheet("Match Methods")
    ws_mm.freeze_panes = "A2"
    _sec_hdr(ws_mm, 1, ["Match Method","Count"], _C["hdr_dark"])
    ws_mm.column_dimensions["A"].width = 20
    ws_mm.column_dimensions["B"].width = 12
    for ri, (_, rd) in enumerate(match_method_df.iterrows(), 2):
        row_bg = _C["row_alt"] if ri % 2 == 0 else _C["white"]
        _cs(ws_mm, ri, 1, str(rd["match_method"]), bg=row_bg, align="left")
        _cs(ws_mm, ri, 2, _fmt_int(rd["Count"]),   bg=row_bg)

    wb.save(args.out)
    print(f"✅ Saved: {args.out}")
    print("   Sheets: Summary | Box Raw | Decided Only | By Pick Type | By Model Tier")
    print("           By Def Tier | Def Tier x Pick Type | Def Tier x Model Tier | Void Reasons | Match Methods")


if __name__ == "__main__":
    main()