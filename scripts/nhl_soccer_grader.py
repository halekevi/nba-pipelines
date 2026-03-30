"""
nhl_soccer_grader.py
====================
Grades NHL and Soccer step8 slates against actuals CSVs.
Outputs graded_nhl_DATE.xlsx and/or graded_soccer_DATE.xlsx
in the same format as build_grade_report.py expects.

Usage:
    py -3 nhl_soccer_grader.py --sport NHL --date 2026-03-06 --slate NHL/step8.xlsx
        --actuals "outputs\2026-03-06\actuals_nhl_2026-03-06.csv" \
        --output-dir "outputs\2026-03-06"

    py -3 nhl_soccer_grader.py --sport Soccer --date 2026-03-06 --slate Soccer/step8.xlsx
        --actuals "outputs\2026-03-06\actuals_soccer_2026-03-06.csv" \
        --output-dir "outputs\2026-03-06"
"""
from __future__ import annotations
import argparse, sys, re
from datetime import datetime
from pathlib import Path

try:
    import pandas as pd
    import numpy as np
except ImportError:
    print("ERROR: pip install pandas openpyxl"); sys.exit(1)

# ── Column maps: step8 slate → canonical graded output ────────────────────────
# These match what build_grade_report.py's normalize() function expects

NHL_SLATE_MAP = {
    "player":        "player",
    "team":          "team",
    "opp":           "opp_team",
    "tier":          "tier",
    "def_tier":      "def_tier",
    "direction":     "bet_direction",
    "line":          "line",
    "prop_display":  "prop_type_norm",
    "prop_type":     "prop_type_raw",
    "edge":          "edge",
    "prop_score":    "rank_score",
    "ml_prob":       "ml_prob",
    "ML Prob":       "ml_prob",
    "ml_edge":       "ml_edge",
    "ML Edge":       "ml_edge",
    "Edge Score":    "edge_score",
    "edge_score":    "edge_score",
    "Blended Score": "blended_score",
    "blended_score": "blended_score",
    "composite_hr":  "hit_rate_raw",
    "player_role":   "player_role",
    "position_group":"position_group",
    "scoring_tier":  "scoring_tier",
    "pp_tier":       "pp_tier",
    "toi_avg_L10":   "toi_avg",
}

SOCCER_SLATE_MAP = {
    # ── lowercase / original ──────────────────────────────────────────────────
    "player":              "player",
    "team":                "team",
    "opp_team":            "opp_team",
    "tier":                "tier",
    "DEF_TIER":            "def_tier",
    "def_tier":            "def_tier",
    "line":                "line",
    "prop_type":           "prop_type_norm",
    "prop_norm":           "prop_type_raw",
    "edge_adj":            "edge",
    "edge":                "edge",
    "rank_score":          "rank_score",
    "ml_prob":             "ml_prob",
    "ml_edge":             "ml_edge",
    "line_hit_rate":       "hit_rate_raw",
    "pick_type":           "pick_type",
    "league":              "league",
    "position_group":      "position_group",
    "minutes_tier":        "minutes_tier",
    "projection":          "projection",
    "direction":           "bet_direction",
    "final_bet_direction": "bet_direction",
    "espn_player_id":      "espn_player_id",   # ID-based matching
    "ESPN ID":             "espn_player_id",   # renamed form from step8 clean xlsx
    # ── Title-case / capitalized (what the soccer slate actually has) ─────────
    "Player":              "player",
    "Team":                "team",
    "Opp":                 "opp_team",
    "Opp Team":            "opp_team",
    "Opponent":            "opp_team",
    "Tier":                "tier",
    "Def Tier":            "def_tier",
    "Def_Tier":            "def_tier",
    "Line":                "line",
    "Prop":                "prop_type_norm",
    "Prop Type":           "prop_type_norm",
    "Prop_Type":           "prop_type_norm",
    "Edge":                "edge",
    "Edge Adj":            "edge",
    "Rank Score":          "rank_score",
    "ML Prob":             "ml_prob",
    "ML Edge":             "ml_edge",
    "ml_prob":             "ml_prob",
    "ml_edge":             "ml_edge",
    "Edge Score":          "edge_score",
    "edge_score":          "edge_score",
    "Blended Score":       "blended_score",
    "blended_score":       "blended_score",
    "Hit Rate":            "hit_rate_raw",
    "Hit Rate (5g)":       "hit_rate_raw",
    "Pick Type":           "pick_type",
    "League":              "league",
    "Position Group":      "position_group",
    "Position":            "position_group",
    "Minutes Tier":        "minutes_tier",
    "Min Tier":            "minutes_tier",
    "Projection":          "projection",
    "Direction":           "bet_direction",
    "Final Bet Direction": "bet_direction",
}
MLB_SLATE_MAP = {
    # Title Case (what step8_mlb_direction_clean.xlsx actually has)
    "Player":           "player",
    "Team":             "team",
    "Opp":              "opp_team",
    "Tier":             "tier",
    "Def Rank":         "def_rank",
    "Line":             "line",
    "Prop":             "prop_type_norm",
    "Direction":        "bet_direction",
    "Edge":             "edge",
    "Rank Score":       "rank_score",
    "ML Prob":          "ml_prob",
    "ML Edge":          "ml_edge",
    "Edge Score":       "edge_score",
    "Blended Score":    "blended_score",
    "Hit Rate (5g)":    "hit_rate_raw",
    "Last 5 Avg":       "avg_L5",
    "Season Avg":       "avg_season",
    "L5 Over":          "over_L5_raw",
    "L5 Under":         "under_L5_raw",
    "Pick Type":        "pick_type",
    "Min Tier":         "minutes_tier",
    "Projection":       "projection",
    "Pos":              "position_group",
    "Player Type":      "player_type",
    "Bat Order":        "bat_order",
    "Pitcher Role":     "pitcher_role",
    "Game Time":        "game_time",
    "Void Reason":      "void_reason_slate",
}

# Actuals CSVs: what columns to look for player name and stat value


ACTUALS_PLAYER_COLS  = ["player","player_name","name","Player","athlete_name"]
ACTUALS_VALUE_COLS   = ["actual","value","stat","result_value","actual_value",
                        "stat_value","fantasy_points","actual_stat"]
ACTUALS_PROP_COLS    = ["prop","prop_type","stat_type","prop_norm","market"]
ACTUALS_TEAM_COLS    = ["team","team_abbr","Team"]

# ── Helpers ────────────────────────────────────────────────────────────────────
def _find_col(df, candidates):
    for c in candidates:
        if c in df.columns: return c
    return None

def _norm_name(s):
    """Lowercase, strip accents roughly, collapse whitespace."""
    s = str(s).lower().strip()
    s = re.sub(r"[àáâãäå]","a", s)
    s = re.sub(r"[èéêë]","e", s)
    s = re.sub(r"[ìíîï]","i", s)
    s = re.sub(r"[òóôõö]","o", s)
    s = re.sub(r"[ùúûü]","u", s)
    s = re.sub(r"[ýÿ]","y", s)
    s = re.sub(r"[ñ]","n", s)
    s = re.sub(r"[ç]","c", s)
    s = s = ' '.join(s.split())
    return s

def _norm_prop(s):
    return re.sub(r"[^a-z0-9]","", str(s).lower())

def load_slate(path: Path, sport: str, grade_date: str = None) -> pd.DataFrame:
    sport = sport.upper()
    # Try reading — handle xlsx and csv
    # Sniff: some .xlsx files are actually CSVs
    def _is_csv(p):
        try:
            with open(p, "rb") as f:
                header = f.read(8)
            # Real xlsx starts with PK (zip), CSV starts with text
            return header[:2] != b"PK"
        except: return False

    is_csv = path.suffix.lower() == ".csv" or _is_csv(path)

    if not is_csv:
        try:
            xf = pd.ExcelFile(path)
            sheet = next((s for s in xf.sheet_names if "all" in s.lower()), xf.sheet_names[0])
            print(f"  Reading sheet '{sheet}' from {path.name}")
            df = pd.read_excel(path, sheet_name=sheet)
        except Exception:
            is_csv = True  # fallback to CSV

    if is_csv:
        print(f"  Reading as CSV: {path.name}")
        for enc in ["utf-8","latin-1","cp1252"]:
            try:
                df = pd.read_csv(path, encoding=enc, low_memory=False)
                break
            except Exception:
                continue
        else:
            print(f"ERROR: could not read {path}"); sys.exit(1)

    df.columns = [c.strip() for c in df.columns]

    col_map = NHL_SLATE_MAP if sport == "NHL" else MLB_SLATE_MAP if sport == "MLB" else SOCCER_SLATE_MAP

    if sport in ("SOCCER", "MLB"):
        original_cols = list(df.columns)
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})
        mapped   = [f"{k}->{v}" for k, v in col_map.items() if k in original_cols]
        unmapped = [c for c in original_cols if c not in col_map]
        print(f"  [ColMap] Renamed: {mapped[:12]}")
        if unmapped:
            print(f"  [ColMap] Unmapped (kept): {unmapped[:15]}")
        for required in ("player", "prop_type_norm", "line", "bet_direction"):
            if required not in df.columns:
                print(f"  [ColMap] WARNING: '{required}' missing after rename!")
                print(f"  [ColMap] All cols after rename: {list(df.columns)}")
    else:
        df = df.rename(columns={k: v for k, v in col_map.items() if k in df.columns})

    # Ensure pick_type exists (NHL step8 doesn't have it — derive from edge/tier)
    if "pick_type" not in df.columns:
        if "tier" in df.columns:
            def _pick_type(row):
                tier = str(row.get("tier","")).upper()
                edge = float(row.get("edge", 0.5)) if pd.notna(row.get("edge")) else 0.5
                if tier == "A" and edge >= 0.48: return "goblin"
                if tier in ("A","B"):            return "standard"
                return "demon"
            df["pick_type"] = df.apply(_pick_type, axis=1)
        else:
            df["pick_type"] = "standard"
    else:
        # Normalize existing pick_type to lowercase for consistent matching
        df["pick_type"] = df["pick_type"].astype(str).str.lower()

    # Normalize bet_direction to OVER/UNDER
    if sport == "SOCCER":
        # Soccer CSV has both bet_direction and final_bet_direction — prefer final
        src_col = "final_bet_direction" if "final_bet_direction" in df.columns else "bet_direction"
        if src_col in df.columns:
            df["bet_direction"] = [str(x).upper().strip() for x in df[src_col]]
    elif "bet_direction" in df.columns:
        df["bet_direction"] = [str(x).upper().strip() for x in df["bet_direction"]]

    # Normalize hit_rate to float 0-1
    if "hit_rate_raw" in df.columns:
        def _pct_to_f(v):
            try:
                s = str(v).replace("%","").strip()
                f = float(s)
                return f/100 if f > 1 else f
            except: return float("nan")
        df["hit_rate"] = df["hit_rate_raw"].apply(_pct_to_f)

    if "ml_prob" in df.columns:
        df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")
    if "ml_edge" in df.columns:
        df["ml_edge"] = pd.to_numeric(df["ml_edge"], errors="coerce")
    elif "ml_prob" in df.columns:
        df["ml_edge"] = df["ml_prob"] - 0.5
    for c in ("edge_score", "blended_score"):
        if c in df.columns:
            df[c] = pd.to_numeric(df[c], errors="coerce")

    df["Sport"] = sport

    # ── Soccer: filter to only rows whose game time has passed ────────────────
    # When grading a specific past date, use end-of-that-day as the cutoff
    if sport == "SOCCER":
        game_time_col = next((c for c in df.columns if c.lower() in
                              ("game time", "game_time", "gametime", "kickoff",
                               "start_time", "starttime", "start time")), None)
        if game_time_col:
            sample = str(df[game_time_col].dropna().iloc[0]) if len(df[game_time_col].dropna()) > 0 else ''
            has_date = ('/' in sample or '-' in sample) and len(sample) > 8
            if not has_date:
                print(f"  [DateFilter] '{game_time_col}' is time-only — skipping filter, keeping all {len(df)} rows")
            else:
                if grade_date:
                    try:
                        cutoff = pd.Timestamp(grade_date, tz='UTC') + pd.Timedelta(days=1)
                    except Exception:
                        cutoff = pd.Timestamp.now('UTC')
                else:
                    cutoff = pd.Timestamp.now('UTC')
                try:
                    gts_raw = pd.to_datetime(df[game_time_col], errors="coerce")
                    if hasattr(gts_raw.dtype, 'tz') and gts_raw.dtype.tz is not None:
                        gts = gts_raw.dt.tz_convert('UTC')
                    else:
                        try:
                            gts = gts_raw.dt.tz_localize('UTC')
                        except TypeError:
                            gts = pd.to_datetime(df[game_time_col], utc=True, errors="coerce")
                    before_cutoff = gts <= cutoff
                    n_total = len(df)
                    df = df[before_cutoff].copy()
                    n_kept = len(df)
                    print(f"  [DateFilter] Kept {n_kept}/{n_total} rows with game_time <= {cutoff.date()} "
                          f"(dropped {n_total - n_kept} future rows)")
                    if n_kept == 0:
                        print(f"  [DateFilter] WARNING: 0 rows remain after date filter — "
                              f"all games on this slate are in the future")
                except Exception as exc:
                    print(f"  [DateFilter] Could not parse '{game_time_col}': {exc} — skipping filter")

    return df

def load_actuals(path: Path) -> pd.DataFrame:
    if not path.exists():
        print(f"  WARNING: actuals not found: {path}")
        return pd.DataFrame()

    for enc in ["utf-8","latin-1","cp1252"]:
        try:
            df = pd.read_csv(path, encoding=enc, low_memory=False)
            break
        except Exception:
            continue
    else:
        print(f"  WARNING: could not read actuals {path}")
        return pd.DataFrame()

    df.columns = [c.strip() for c in df.columns]
    print(f"  Actuals: {len(df)} rows, cols: {list(df.columns)}")
    return df

def grade(slate: pd.DataFrame, actuals: pd.DataFrame, sport: str) -> pd.DataFrame:
    """Match slate props to actuals and assign HIT/MISS/VOID."""
    slate = slate.copy()
    slate["actual"] = float("nan")
    slate["result"] = "VOID"
    slate["void_reason_grade"] = "NO_ACTUAL"
    slate["margin"] = float("nan")

    if actuals.empty:
        print("  WARNING: no actuals — all props marked VOID")
        return slate

    # Find key columns in actuals
    p_col   = _find_col(actuals, ACTUALS_PLAYER_COLS)
    v_col   = _find_col(actuals, ACTUALS_VALUE_COLS)
    pr_col  = _find_col(actuals, ACTUALS_PROP_COLS)
    t_col   = _find_col(actuals, ACTUALS_TEAM_COLS)

    if not p_col or not v_col:
        print(f"  WARNING: actuals missing player ({p_col}) or value ({v_col}) column")
        print(f"  Actuals columns: {list(actuals.columns)}")
        return slate

    actuals = actuals.copy()
    actuals["_name"] = actuals[p_col].apply(_norm_name)
    actuals["_val"]  = pd.to_numeric(actuals[v_col], errors="coerce")
    if pr_col:
        actuals["_prop"] = actuals[pr_col].apply(_norm_prop)

    # ── Build TWO lookup indexes ──────────────────────────────────────────────
    # 1) by espn_player_id  (primary for soccer — bulletproof)
    # 2) by normalised name (fallback for NHL or when ID missing)
    act_by_id:   dict[str, list] = {}
    act_by_name: dict[str, list] = {}
    has_id_col = "espn_player_id" in actuals.columns

    for _, row in actuals.iterrows():
        # ID index
        if has_id_col:
            eid = str(row.get("espn_player_id", "")).strip()
            if eid and eid not in ("", "nan", "None"):
                if eid not in act_by_id:
                    act_by_id[eid] = []
                act_by_id[eid].append(row)
        # Name index
        name = row["_name"]
        if name not in act_by_name:
            act_by_name[name] = []
        act_by_name[name].append(row)

    hits = misses = voids = 0

    # ── SOCCER DIAGNOSTIC ────────────────────────────────────────────────────
    if sport == "SOCCER" and len(slate) > 0:
        sample_slate = list(slate["player"].apply(_norm_name).unique())[:5]
        sample_acts  = list(act_by_name.keys())[:5]
        name_overlap = sum(1 for n in slate["player"].apply(_norm_name) if n in act_by_name)
        id_overlap = 0
        if has_id_col and "espn_player_id" in slate.columns:
            id_overlap = sum(
                1 for eid in slate["espn_player_id"].astype(str).str.strip()
                if eid and eid not in ("", "nan", "None") and eid in act_by_id
            )
        prop_col_s = ("prop_type_norm" if "prop_type_norm" in slate.columns
                      else "prop_type_raw" if "prop_type_raw" in slate.columns else None)
        sample_props = list(slate[prop_col_s].apply(_norm_prop).unique())[:8] if prop_col_s else []
        print(f"  [DIAG] Slate name sample (normed): {sample_slate}")
        print(f"  [DIAG] Actuals name sample (normed): {sample_acts}")
        print(f"  [DIAG] Name matches: {name_overlap}/{len(slate)} | "
              f"ID matches: {id_overlap}/{len(slate)}")
        print(f"  [DIAG] Slate prop norms: {sample_props}")
        print(f"  [DIAG] Actuals prop norms: "
              f"{sorted(set(actuals['_prop'].tolist()))[:10] if pr_col else 'no prop col'}")

    # ── Soccer prop alias table ───────────────────────────────────────────────
    SOCCER_PROP_ALIASES = {
        # Goalkeeper saves
        "goaliesaves":          ["goalkeepersaves", "goaliesaves", "saves", "gksaves"],
        "goalkeepersaves":      ["goalkeepersaves", "goaliesaves", "saves", "gksaves"],
        "saves":                ["goalkeepersaves", "goaliesaves", "saves", "gksaves"],
        "gksaves":              ["goalkeepersaves", "goaliesaves", "saves", "gksaves"],
        # Goalkeeper saves combo — same stat, different PrizePicks label
        "goaliesavescombo":     ["goalkeepersaves", "goaliesaves", "saves", "gksaves"],
        "gksavescombo":         ["goalkeepersaves", "goaliesaves", "saves", "gksaves"],
        # Shots on target — must NOT fall back to "shots" (different stat)
        "shotsontarget":        ["shotsontarget", "sot"],
        "sot":                  ["shotsontarget", "sot"],
        "shotstarget":          ["shotsontarget", "sot"],
        # Total shots
        "shots":                ["shots", "totalshots"],
        "totalshots":           ["shots", "totalshots"],
        # Fouls
        "fouls":                ["fouls", "foulscommitted"],
        "foulscommitted":       ["fouls", "foulscommitted"],
        # Yellow cards
        "yellowcards":          ["yellowcards", "yellow", "yc"],
        "yellow":               ["yellowcards", "yellow", "yc"],
        "yc":                   ["yellowcards", "yellow", "yc"],
        # Assists / goal assists
        "assists":              ["assists", "goalassist", "goalassists"],
        "goalassist":           ["assists", "goalassist", "goalassists"],
        "goalassists":          ["assists", "goalassist", "goalassists"],
        # Goals
        "goals":                ["goals", "goal"],
        "goal":                 ["goals", "goal"],
        # Passes attempted
        "passesattempted":      ["passesattempted", "passes"],
        "passes":               ["passesattempted", "passes"],
        "pa":                   ["passesattempted", "passes"],
        # Key passes
        "keypasses":            ["keypasses", "keypass"],
        "keypass":              ["keypasses", "keypass"],
        "kp":                   ["keypasses", "keypass"],
        # Tackles
        "tackles":              ["tackles", "totaltackle"],
        "totaltackle":          ["tackles", "totaltackle"],
        "tk":                   ["tackles", "totaltackle"],
        # Fantasy
        "fantasypoints":        ["fantasypoints", "fantasy"],
        "fantasyscore":         ["fantasypoints", "fantasy"],
    }

    def _find_prop_match(candidates, sprop, sport):
        """Try exact prop match, then alias match. Returns matched row or None."""
        if not pr_col or not candidates:
            return None
        # Exact
        exact = [r for r in candidates if r.get("_prop","") == sprop]
        if exact:
            return exact[0]
        # Alias (soccer only)
        if sport == "SOCCER":
            aliases = SOCCER_PROP_ALIASES.get(sprop, [sprop])
            alias_matches = [r for r in candidates if r.get("_prop","") in aliases]
            if alias_matches:
                return alias_matches[0]
        return None

    for idx, srow in slate.iterrows():
        sname = _norm_name(srow.get("player",""))
        sprop = _norm_prop(srow.get("prop_type_norm", srow.get("prop_type_raw","")))
        sline = srow.get("line")
        sdir  = str(srow.get("bet_direction","")).upper()
        s_eid = str(srow.get("espn_player_id","")).strip() if "espn_player_id" in srow.index else ""
        s_eid_valid = s_eid and s_eid not in ("", "nan", "None")

        matched = None

        # ── PASS 1: ESPN player ID match (soccer, bulletproof) ────────────────
        if sport == "SOCCER" and s_eid_valid and s_eid in act_by_id:
            candidates = act_by_id[s_eid]
            matched = _find_prop_match(candidates, sprop, sport)
            # If still no prop match but only one candidate, use it
            if matched is None and len(candidates) == 1:
                matched = candidates[0]

        # ── PASS 2: name match fallback ───────────────────────────────────────
        if matched is None:
            if sname not in act_by_name:
                voids += 1
                continue
            candidates = act_by_name[sname]
            matched = _find_prop_match(candidates, sprop, sport)
            if matched is None and len(candidates) == 1:
                matched = candidates[0]
            if matched is None and candidates:
                matched = candidates[0]

        if matched is None or pd.isna(matched["_val"]):
            voids += 1
            continue

        actual_val = float(matched["_val"])
        try:
            line_val = float(sline)
        except (TypeError, ValueError):
            voids += 1
            slate.at[idx, "void_reason_grade"] = "NO_LINE"
            continue

        slate.at[idx, "actual"] = actual_val
        margin = actual_val - line_val
        slate.at[idx, "margin"] = margin

        # Grade
        if margin == 0:
            slate.at[idx, "result"] = "VOID"
            slate.at[idx, "void_reason_grade"] = "PUSH"
            voids += 1
        elif (sdir == "OVER" and margin > 0) or (sdir == "UNDER" and margin < 0):
            slate.at[idx, "result"] = "HIT"
            slate.at[idx, "void_reason_grade"] = ""
            hits += 1
        else:
            slate.at[idx, "result"] = "MISS"
            slate.at[idx, "void_reason_grade"] = ""
            misses += 1

    total = len(slate)
    dec   = hits + misses
    rate  = f"{hits/dec*100:.1f}%" if dec else "—"
    print(f"  Graded: {total:,} props → HIT:{hits} MISS:{misses} VOID:{voids} | Hit rate: {rate}")
    return slate

def save_graded(df: pd.DataFrame, out_path: Path, sport: str, date_str: str):
    """Save NBA-style multi-sheet graded Excel with full formatting and Demon exclusion."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    import numpy as np

    # ── Colour palette (matches NBA graded file) ─────────────────────────────
    C = {
        'hit':'27AE60','miss':'E74C3C','push':'F39C12','void':'95A5A6',
        'hdr':'1C1C1C','hdr2':'1A5276','hdr3':'1E8449','hdr4':'7D6608',
        'hdr5':'922B21','hdr6':'6C3483','hdr7':'117A65','hdr8':'1A5276',
        'alt':'F2F3F4','white':'FFFFFF',
        'tier_a':'D5F5E3','tier_b':'D6EAF8','tier_c':'FEF9E7','tier_d':'FDEDEC',
        'over':'D6EAF8','under':'FDEBD0',
        'goblin':'E8D5F5','demon':'FDEDEC','standard':'F2F3F4',
    }
    DEF_TIER_ORDER     = ['Elite','Above Avg','Avg','Weak']
    MINUTES_TIER_ORDER = ['HIGH','MEDIUM','LOW','UNKNOWN']
    TIER_ORDER         = ['A','B','C','D']

    def _bdr(color='CCCCCC'):
        s = Side(style='thin', color=color)
        return Border(left=s, right=s, top=s, bottom=s)

    def _hc(ws, r, c, v, bg=None, fc='FFFFFF', bold=True, sz=9, align='center'):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, color=fc, name='Arial', size=sz)
        if bg: cell.fill = PatternFill('solid', start_color=bg)
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = _bdr()
        return cell

    def _dc(ws, r, c, v, bg=None, bold=False, sz=9, align='center', fmt=None, fc='000000'):
        cell = ws.cell(row=r, column=c, value=v)
        cell.font = Font(bold=bold, name='Arial', size=sz, color=fc)
        cell.fill = PatternFill('solid', start_color=bg or C['white'])
        cell.alignment = Alignment(horizontal=align, vertical='center')
        cell.border = _bdr()
        if fmt: cell.number_format = fmt
        return cell

    def _res_bg(r):
        return {'HIT':C['hit'],'MISS':C['miss'],'PUSH':C['push'],'VOID':C['void']}.get(str(r).upper(),'DDDDDD')

    def _hr_bg(v):
        if v is None or (isinstance(v, float) and np.isnan(v)): return 'DDDDDD'
        if v >= 0.65: return C['hit']
        if v >= 0.50: return C['push']
        return C['miss']

    def _tier_bg(t):
        return {'A':C['tier_a'],'B':C['tier_b'],'C':C['tier_c'],'D':C['tier_d']}.get(str(t).upper(), C['white'])

    def _pct_cell(ws, r, c, val):
        nan = val is None or (isinstance(val, float) and np.isnan(val))
        bg = _hr_bg(val) if not nan else 'DDDDDD'
        cell = _dc(ws, r, c, val if not nan else '', bg=bg, bold=True)
        if not nan:
            cell.number_format = '0.0%'
            cell.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        return cell

    def _sw(ws, widths):
        for ci, w in enumerate(widths, 1):
            ws.column_dimensions[get_column_letter(ci)].width = w

    def _hit_rate(sub):
        """Exclude Demons from hit-rate. They appear in Box Raw but not in grading stats."""
        if 'pick_type' in sub.columns:
            pt_lower = sub['pick_type'].astype(str).str.lower()
            graded = sub[pt_lower != 'demon']
            demon_count = int((pt_lower == 'demon').sum())
        else:
            graded = sub
            demon_count = 0
        dec = graded[graded['result'].isin(['HIT', 'MISS'])]
        h = (dec['result'] == 'HIT').sum()
        v = int(graded['result'].isin(['VOID', 'PUSH']).sum()) + demon_count
        return (h / len(dec) if len(dec) else np.nan), int(h), int(len(dec) - h), v, int(len(dec))

    def _drc(d):
        return 'bet_direction' if 'bet_direction' in d.columns else 'final_bet_direction'

    def _sheet_hdr8(ws, col1, bg, widths=None):
        _sw(ws, widths or [24,10,8,10,8,8,8,12])
        for ci, h in enumerate([col1,'Direction','Total','Decided','Hits','Misses','Voids','Hit Rate'], 1):
            _hc(ws, 1, ci, h, bg=bg)
        ws.row_dimensions[1].height = 20
        ws.freeze_panes = 'A2'

    def _dir_subrows(ws, ri, sub, label, bg):
        d = _drc(sub)
        hr_a, h_a, m_a, v_a, dec_a = _hit_rate(sub)
        row_bg = bg or (C['alt'] if ri % 2 == 0 else C['white'])
        _dc(ws,ri,1,label,bg=row_bg,bold=True,align='left')
        _dc(ws,ri,2,'ALL',bg=row_bg,bold=True)
        _dc(ws,ri,3,len(sub),bg=row_bg); _dc(ws,ri,4,dec_a,bg=row_bg)
        _dc(ws,ri,5,h_a,bg=row_bg); _dc(ws,ri,6,m_a,bg=row_bg); _dc(ws,ri,7,v_a,bg=row_bg)
        _pct_cell(ws,ri,8,hr_a); ri += 1
        for direction in ['OVER','UNDER']:
            dsub = sub[sub[d].str.upper() == direction] if d in sub.columns else pd.DataFrame()
            if len(dsub) == 0: continue
            hr_d, h_d, m_d, v_d, dec_d = _hit_rate(dsub)
            dbg = C['over'] if direction == 'OVER' else C['under']
            _dc(ws,ri,1,'',bg=dbg); _dc(ws,ri,2,direction,bg=dbg,bold=True)
            _dc(ws,ri,3,len(dsub),bg=dbg); _dc(ws,ri,4,dec_d,bg=dbg)
            _dc(ws,ri,5,h_d,bg=dbg); _dc(ws,ri,6,m_d,bg=dbg); _dc(ws,ri,7,v_d,bg=dbg)
            _pct_cell(ws,ri,8,hr_d); ri += 1
        return ri

    wb = Workbook()
    wb.remove(wb.active)

    # ── Summary sheet ─────────────────────────────────────────────────────────
    ws_sum = wb.create_sheet('Summary', 0)
    ws_sum.column_dimensions['A'].width = 22
    ws_sum.column_dimensions['B'].width = 12
    for ci in range(3, 10):
        ws_sum.column_dimensions[get_column_letter(ci)].width = 11
    ws_sum.merge_cells('A1:I1')
    c = ws_sum['A1']
    c.value = f"{sport} SLATE GRADE  |  {date_str}  |  Generated {datetime.now():%Y-%m-%d %H:%M}"
    c.font = Font(bold=True, name='Arial', size=12, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=C['hdr'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws_sum.row_dimensions[1].height = 28

    d_col = _drc(df)

    def _sec_hdr(ws, row, label, color):
        _hc(ws, row, 1, label, bg=color)
        for ci, h in enumerate(['Direction','Total','Decided','Hits','Misses','Voids','Hit Rate'], 2):
            _hc(ws, row, ci, h, bg=color)
        ws.row_dimensions[row].height = 20
        return row + 1

    def _simple_row(ws, row, label, sub, bg=None, bold=True):
        hr2, h2, m2, v2, dec2 = _hit_rate(sub)
        bg = bg or (C['alt'] if row % 2 == 0 else C['white'])
        _dc(ws, row, 1, label, bold=bold, align='left')
        _dc(ws, row, 2, 'ALL', bg=bg, bold=True); _dc(ws, row, 3, len(sub), bg=bg)
        _dc(ws, row, 4, dec2, bg=bg); _dc(ws, row, 5, int(h2), bg=bg)
        _dc(ws, row, 6, int(m2), bg=bg); _dc(ws, row, 7, int(v2), bg=bg)
        _pct_cell(ws, row, 8, hr2)
        return row + 1

    def _dir_rows(ws, row, sub):
        if d_col not in sub.columns: return row
        for direction in ['OVER', 'UNDER']:
            dsub = sub[sub[d_col].str.upper() == direction]
            if len(dsub) == 0: continue
            hr_d, h_d, m_d, v_d, dec_d = _hit_rate(dsub)
            dbg = C['over'] if direction == 'OVER' else C['under']
            _dc(ws,row,1,'',bg=dbg); _dc(ws,row,2,direction,bg=dbg,bold=True)
            _dc(ws,row,3,len(dsub),bg=dbg); _dc(ws,row,4,dec_d,bg=dbg)
            _dc(ws,row,5,int(h_d),bg=dbg); _dc(ws,row,6,int(m_d),bg=dbg)
            _dc(ws,row,7,int(v_d),bg=dbg); _pct_cell(ws,row,8,hr_d); row += 1
        return row

    # OVERALL
    row = 2
    _hc(ws_sum, row, 1, 'OVERALL', bg=C['hdr2'])
    for ci, h in enumerate(['Direction','Total Props','Decided','Hits','Misses','Voids','Hit Rate'], 2):
        _hc(ws_sum, row, ci, h, bg=C['hdr2'])
    row += 1
    hr_ov, h_ov, m_ov, v_ov, dec_ov = _hit_rate(df)
    _dc(ws_sum,row,1,'Full Slate',bold=True,align='left'); _dc(ws_sum,row,2,'ALL')
    _dc(ws_sum,row,3,len(df)); _dc(ws_sum,row,4,dec_ov)
    _dc(ws_sum,row,5,int(h_ov)); _dc(ws_sum,row,6,int(m_ov)); _dc(ws_sum,row,7,int(v_ov))
    _pct_cell(ws_sum,row,8,hr_ov); row += 1

    # BY PICK TYPE
    if 'pick_type' in df.columns:
        pt_lower = df['pick_type'].astype(str).str.lower()
        row += 1; row = _sec_hdr(ws_sum, row, 'BY PICK TYPE', C['hdr3'])
        for pt in ['Goblin', 'Demon', 'Standard']:
            sub = df[pt_lower == pt.lower()]
            if len(sub) == 0: continue
            pt_bg = {'Goblin':C['goblin'],'Demon':C['demon'],'Standard':C['standard']}.get(pt)
            if pt == 'Demon':
                # Demon: show total count but mark as excluded from grading
                _dc(ws_sum,row,1,f"{pt} (excl. from grading)",bold=True,align='left')
                _dc(ws_sum,row,2,'OVER',bg=pt_bg,bold=True)
                _dc(ws_sum,row,3,len(sub),bg=pt_bg); _dc(ws_sum,row,4,'—',bg=pt_bg)
                _dc(ws_sum,row,5,'—',bg=pt_bg); _dc(ws_sum,row,6,'—',bg=pt_bg)
                _dc(ws_sum,row,7,len(sub),bg=pt_bg)
                _dc(ws_sum,row,8,'EXCL',bg='DDDDDD',bold=True); row += 1
            elif pt == 'Standard':
                row = _simple_row(ws_sum, row, pt, sub, bg=pt_bg)
                row = _dir_rows(ws_sum, row, sub)
            else:  # Goblin
                hr_g, h_g, m_g, v_g, dec_g = _hit_rate(sub)
                _dc(ws_sum,row,1,pt,bold=True,align='left'); _dc(ws_sum,row,2,'OVER',bg=pt_bg,bold=True)
                _dc(ws_sum,row,3,len(sub),bg=pt_bg); _dc(ws_sum,row,4,dec_g,bg=pt_bg)
                _dc(ws_sum,row,5,h_g,bg=pt_bg); _dc(ws_sum,row,6,m_g,bg=pt_bg); _dc(ws_sum,row,7,v_g,bg=pt_bg)
                _pct_cell(ws_sum,row,8,hr_g); row += 1

    # BY TIER
    if 'tier' in df.columns:
        row += 1; row = _sec_hdr(ws_sum, row, 'BY TIER', C['hdr4'])
        for t in TIER_ORDER:
            sub = df[df['tier'].astype(str).str.upper() == t]
            if len(sub) == 0: continue
            row = _simple_row(ws_sum, row, f'Tier {t}', sub, bg=_tier_bg(t))
            row = _dir_rows(ws_sum, row, sub)

    # BY DEF TIER
    if 'def_tier' in df.columns:
        row += 1; row = _sec_hdr(ws_sum, row, 'BY OPP DEF TIER', C['hdr5'])
        for dt in [t for t in DEF_TIER_ORDER if t in df['def_tier'].dropna().unique()]:
            sub = df[df['def_tier'] == dt]
            row = _simple_row(ws_sum, row, dt, sub)
            row = _dir_rows(ws_sum, row, sub)

    # BY MINUTES TIER (Soccer / NHL)
    if 'minutes_tier' in df.columns:
        row += 1; row = _sec_hdr(ws_sum, row, 'BY MINUTES TIER', C['hdr6'])
        for mt in [t for t in MINUTES_TIER_ORDER if t in df['minutes_tier'].dropna().unique()]:
            sub = df[df['minutes_tier'] == mt]
            row = _simple_row(ws_sum, row, mt, sub)
            row = _dir_rows(ws_sum, row, sub)

    # ── Box Raw ───────────────────────────────────────────────────────────────
    ws_raw = wb.create_sheet('Box Raw')
    desired = ['player','team','opp_team','prop_type_norm','pick_type','line',
               'bet_direction','tier','def_tier','minutes_tier','position_group',
               'edge','hit_rate','projection','rank_score','ml_prob','ml_edge',
               'edge_score','blended_score',
               'actual','result','margin','void_reason_grade']
    cols = [c for c in desired if c in df.columns]
    widths_map = {'player':22,'team':6,'opp_team':6,'prop_type_norm':20,'pick_type':10,
                  'line':7,'bet_direction':10,'tier':5,'def_tier':10,'minutes_tier':12,
                  'position_group':14,'edge':8,'hit_rate':10,'projection':12,'rank_score':12,
                  'ml_prob':10,'ml_edge':10,'edge_score':11,'blended_score':12,
                  'actual':9,'result':8,'margin':8,'void_reason_grade':22}
    for ci, col in enumerate(cols, 1):
        ws_raw.column_dimensions[get_column_letter(ci)].width = widths_map.get(col, 12)
        _hc(ws_raw, 1, ci, col, bg=C['hdr'])
    ws_raw.row_dimensions[1].height = 20
    ws_raw.freeze_panes = 'A2'
    for ri, row_data in enumerate(df[cols].itertuples(), 2):
        bg = C['alt'] if ri % 2 == 0 else C['white']
        res = str(getattr(row_data, 'result', '')).upper()
        for ci, col in enumerate(cols, 1):
            val = getattr(row_data, col, '')
            c_bg = _res_bg(res) if col == 'result' else (_tier_bg(val) if col == 'tier' else bg)
            cell = _dc(ws_raw, ri, ci, val, bg=c_bg, align='left' if col == 'player' else 'center')
            if col == 'result' and res in ('HIT','MISS','PUSH','VOID'):
                cell.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
    ws_raw.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"

    # ── By Pick Type ──────────────────────────────────────────────────────────
    ws_pt = wb.create_sheet('By Pick Type')
    _sheet_hdr8(ws_pt, 'Pick Type', C['hdr3'], [20,10,8,10,8,8,8,12])
    ri = 2
    pt_lower_df = df['pick_type'].astype(str).str.lower() if 'pick_type' in df.columns else None
    for pt in ['Goblin', 'Demon', 'Standard']:
        sub = df[pt_lower_df == pt.lower()] if pt_lower_df is not None else pd.DataFrame()
        if len(sub) == 0: continue
        pt_bg = {'Goblin':C['goblin'],'Demon':C['demon'],'Standard':C['standard']}.get(pt, C['white'])
        if pt == 'Demon':
            _dc(ws_pt,ri,1,f"{pt} (excl.)",bg=pt_bg,bold=True,align='left')
            _dc(ws_pt,ri,2,'OVER',bg=pt_bg,bold=True)
            _dc(ws_pt,ri,3,len(sub),bg=pt_bg); _dc(ws_pt,ri,4,'—',bg=pt_bg)
            _dc(ws_pt,ri,5,'—',bg=pt_bg); _dc(ws_pt,ri,6,'—',bg=pt_bg)
            _dc(ws_pt,ri,7,len(sub),bg=pt_bg)
            _dc(ws_pt,ri,8,'EXCL',bg='DDDDDD',bold=True); ri += 1
        elif pt == 'Standard':
            ri = _dir_subrows(ws_pt, ri, sub, pt, pt_bg)
        else:
            hr_g, h_g, m_g, v_g, dec_g = _hit_rate(sub)
            _dc(ws_pt,ri,1,pt,bg=pt_bg,bold=True,align='left'); _dc(ws_pt,ri,2,'OVER',bg=pt_bg,bold=True)
            _dc(ws_pt,ri,3,len(sub),bg=pt_bg); _dc(ws_pt,ri,4,dec_g,bg=pt_bg)
            _dc(ws_pt,ri,5,h_g,bg=pt_bg); _dc(ws_pt,ri,6,m_g,bg=pt_bg); _dc(ws_pt,ri,7,v_g,bg=pt_bg)
            _pct_cell(ws_pt,ri,8,hr_g); ri += 1

    # ── By Tier ───────────────────────────────────────────────────────────────
    if 'tier' in df.columns:
        ws_tier = wb.create_sheet('By Tier')
        _sheet_hdr8(ws_tier, 'Tier', C['hdr4'])
        ri = 2
        for t in TIER_ORDER:
            sub = df[df['tier'].astype(str).str.upper() == t]
            if len(sub) == 0: continue
            ri = _dir_subrows(ws_tier, ri, sub, f'Tier {t}', _tier_bg(t))

    # ── Prop Type x Direction ─────────────────────────────────────────────────
    pt_col = 'prop_type_norm' if 'prop_type_norm' in df.columns else None
    if pt_col:
        ws_prop = wb.create_sheet('Prop Type x Direction')
        _sheet_hdr8(ws_prop, 'Prop Type', C['hdr2'], [28,10,8,10,8,8,8,12])
        prop_order = (df[df['result'].isin(['HIT','MISS'])].groupby(pt_col).size()
                      .sort_values(ascending=False).index.tolist())
        prop_order += [p for p in df[pt_col].unique() if p not in prop_order]
        ri = 2
        for prop in prop_order:
            psub = df[df[pt_col] == prop]
            ri = _dir_subrows(ws_prop, ri, psub, prop, C['alt'] if ri % 2 == 0 else C['white'])

    # ── By Direction ──────────────────────────────────────────────────────────
    ws_dir = wb.create_sheet('By Direction')
    _sheet_hdr8(ws_dir, 'Direction', C['hdr5'])
    ri = 2
    if d_col in df.columns:
        for direction in ['OVER', 'UNDER']:
            dsub = df[df[d_col].str.upper() == direction]
            if len(dsub) == 0: continue
            hr_d, h_d, m_d, v_d, dec_d = _hit_rate(dsub)
            dbg = C['over'] if direction == 'OVER' else C['under']
            _dc(ws_dir,ri,1,direction,bg=dbg,bold=True,align='left'); _dc(ws_dir,ri,2,direction,bg=dbg,bold=True)
            _dc(ws_dir,ri,3,len(dsub),bg=dbg); _dc(ws_dir,ri,4,dec_d,bg=dbg)
            _dc(ws_dir,ri,5,h_d,bg=dbg); _dc(ws_dir,ri,6,m_d,bg=dbg); _dc(ws_dir,ri,7,v_d,bg=dbg)
            _pct_cell(ws_dir,ri,8,hr_d); ri += 1

    # ── By Def Tier ───────────────────────────────────────────────────────────
    if 'def_tier' in df.columns:
        ws_def = wb.create_sheet('By Def Tier')
        _sheet_hdr8(ws_def, 'Def Tier', C['hdr5'])
        ri = 2
        for dt in [t for t in DEF_TIER_ORDER if t in df['def_tier'].dropna().unique()]:
            sub = df[df['def_tier'] == dt]
            ri = _dir_subrows(ws_def, ri, sub, dt, C['alt'] if ri % 2 == 0 else C['white'])

    # ── By Minutes Tier ───────────────────────────────────────────────────────
    if 'minutes_tier' in df.columns:
        ws_mt = wb.create_sheet('By Minutes Tier')
        _sheet_hdr8(ws_mt, 'Minutes Tier', C['hdr6'])
        ri = 2
        for mt in [t for t in MINUTES_TIER_ORDER if t in df['minutes_tier'].dropna().unique()]:
            sub = df[df['minutes_tier'] == mt]
            ri = _dir_subrows(ws_mt, ri, sub, mt, C['alt'] if ri % 2 == 0 else C['white'])

    # ── By Position Group (Soccer) / Player Role (NHL) ────────────────────────
    for role_col, sheet_name, hdr_bg in [
        ('position_group', 'By Position', C['hdr7']),
        ('player_role',    'By Player Role', C['hdr7']),
    ]:
        if role_col in df.columns:
            ws_role = wb.create_sheet(sheet_name)
            _sheet_hdr8(ws_role, sheet_name.replace('By ',''), hdr_bg)
            ri = 2
            for role in sorted(df[role_col].dropna().unique()):
                sub = df[df[role_col] == role]
                ri = _dir_subrows(ws_role, ri, sub, str(role), C['alt'] if ri % 2 == 0 else C['white'])

    # ── Void Reasons ──────────────────────────────────────────────────────────
    vr_col = 'void_reason_grade' if 'void_reason_grade' in df.columns else None
    if vr_col:
        vr_df = df[df[vr_col].astype(str).str.strip().str.len() > 0]
        if len(vr_df) > 0:
            ws_vr = wb.create_sheet('Void Reasons')
            _sw(ws_vr, [28, 10])
            _hc(ws_vr, 1, 1, 'Void Reason', bg=C['hdr'])
            _hc(ws_vr, 1, 2, 'Count', bg=C['hdr'])
            ws_vr.freeze_panes = 'A2'
            vr_counts = vr_df[vr_col].value_counts()
            for ri, (reason, cnt) in enumerate(vr_counts.items(), 2):
                bg = C['alt'] if ri % 2 == 0 else C['white']
                _dc(ws_vr, ri, 1, str(reason), bg=bg, align='left')
                _dc(ws_vr, ri, 2, int(cnt), bg=bg)

    wb.save(out_path)
    print(f"  Saved → {out_path}  ({out_path.stat().st_size:,} bytes)")
    print(f"  Sheets: {wb.sheetnames}")

# ── CLI ────────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser(description="Grade NHL/Soccer slates against actuals")
    ap.add_argument("--sport",      required=True, choices=['NHL','Soccer','SOCCER','nhl','soccer','MLB','mlb'])
    ap.add_argument("--date",       required=True)
    ap.add_argument("--slate",      required=True)
    ap.add_argument("--actuals",    required=True)
    ap.add_argument("--output-dir", required=True)
    args = ap.parse_args()

    sport    = args.sport.upper()
    date_str = args.date
    slate_p  = Path(args.slate)
    act_p    = Path(args.actuals)
    out_dir  = Path(args.output_dir)
    out_dir.mkdir(parents=True, exist_ok=True)

    sport_lower = sport.lower()
    out_path = out_dir / f"graded_{sport_lower}_{date_str}.xlsx"

    print(f"\n  [{sport} GRADER]  {date_str}")
    print(f"  Slate:   {slate_p}")
    print(f"  Actuals: {act_p}")
    print(f"  Output:  {out_path}")

    slate   = load_slate(slate_p, sport, grade_date=date_str)
    actuals = load_actuals(act_p)

    print(f"  Slate rows: {len(slate):,}")

    graded = grade(slate, actuals, sport)
    save_graded(graded, out_path, sport, date_str)
    print(f"  Done.\n")

if __name__ == "__main__":
    main()
