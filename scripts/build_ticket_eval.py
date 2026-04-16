#!/usr/bin/env python3
"""
Build ticket_eval_{date}.html for Grades → Ticket evaluation (graded legs, actuals, KPI bar).

Reads ticket JSON (or combined_slate_tickets_{date}.xlsx) and sport step8/graded workbooks,
matches legs to actuals, writes self-contained HTML. Graded overlays merge ``outputs/<d>/`` and
``ui_runner/graded_slate/<d>/`` for the slate ``--date`` **and** each unique leg ``game_time``
calendar day (so evening slates with tomorrow's games pick up graded_* for the game day).

The /tickets page is rendered from tickets_latest.json (today's built slips), not from this file.

P(Win) and EV at entry in the HTML come only from the empirical ``payout`` object on each ticket (same
run as ``tickets_latest.json``). Combined-slate Excel has multipliers on the banner only; archive
per-date ticket JSON if you need those fields historically (``tickets_latest`` is overwritten each run).

Also writes ticket_eval_slate_latest.json — same legs as window.SLATE_DATA in the HTML — so the home
page /api/slate-sport slate cards can match the ticket eval builder (when SLATE_SPORT_SOURCE=auto).

Run after combined_slate_tickets.py --write-web, then this script to refresh graded HTML for that date.

Railway / git deploy: ``outputs/`` is not in the repo. To grade MLB (and other sports) on the live site,
drop ``graded_<sport>_<date>.xlsx`` under ``ui_runner/graded_slate/<date>/`` (same names as under
``outputs/<date>/``), then run this script and commit ``ticket_eval_<date>.html``.
"""
from __future__ import annotations

import argparse
import html
import json
import logging
import math
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Sequence

import pandas as pd
from dateutil.parser import parse as _parse_datetime_guess

REPO_ROOT = Path(__file__).resolve().parent.parent
_log = logging.getLogger(__name__)
_SCRIPTS = REPO_ROOT / "scripts"
if str(_SCRIPTS) not in sys.path:
    sys.path.insert(0, str(_SCRIPTS))
from player_name_norm import fold_player_name as _fold_player_name  # noqa: E402
TEMPLATES_DIR = REPO_ROOT / "ui_runner" / "templates"
TICKET_EVAL_SLATE_JSON = TEMPLATES_DIR / "ticket_eval_slate_latest.json"

# Ticket source search order: combined_slate_tickets_{date}.xlsx only.
DATED_TICKET_JSON = "combined_slate_tickets_{date}.json"
FALLBACK_TICKET_JSON = TEMPLATES_DIR / "tickets_latest.json"
ALLOWED_TICKET_SPORTS = {"NBA", "CBB", "NHL", "SOCCER", "MLB", "WCBB"}

_XLSX_HDR_TO_LEG_FIELD: dict[str, str] = {
    "player": "player",
    "team": "team",
    "opp": "opp",
    "prop": "prop_type",
    "pick type": "pick_type",
    "line": "line",
    "dir": "direction",
    "edge": "edge",
    "hit rate": "hit_rate",
    "l5 avg": "l5_avg",
    "szn avg": "season_avg",
    "sport": "sport",
}

DEF_COL_ALIASES = ["Def Tier", "OPP_DEF_TIER", "opp_def_tier", "def_tier", "DEF_TIER", "Defense Tier"]
PICK_COL_ALIASES = ["Pick Type", "PICK_TYPE", "pick_type", "PickType"]
TIER_COL_ALIASES = ["Tier", "TIER", "tier"]
PROP_COL_ALIASES = ["Prop", "PROP", "prop", "PROP_TYPE", "Prop Type"]
LINE_COL_ALIASES = ["Line", "LINE", "line"]
DIR_COL_ALIASES = ["Dir", "DIR", "dir", "Direction", "DIRECTION"]

# Slate workbooks per sport bucket (first existing file wins within that bucket).
# Ticket legs with sport NBA1H / NBA1Q / WCBB must match rows from these files, not full-game NBA/CBB only.
SPORT_XLSX_CANDIDATES: dict[str, list[Path]] = {
    "NBA": [
        REPO_ROOT / "NBA" / "step8_all_direction_clean.xlsx",
        REPO_ROOT / "NBA" / "data" / "outputs" / "step8_all_direction_clean.xlsx",
    ],
    "NBA1H": [
        REPO_ROOT / "NBA" / "step8_nba1h_direction_clean.xlsx",
    ],
    "NBA1Q": [
        REPO_ROOT / "NBA" / "step8_nba1q_direction_clean.xlsx",
    ],
    "CBB": [
        REPO_ROOT / "CBB" / "step6_ranked_cbb.xlsx",
    ],
    "WCBB": [
        REPO_ROOT / "CBB" / "step6_ranked_wcbb.xlsx",
    ],
    "NHL": [
        REPO_ROOT / "NHL" / "outputs" / "step8_nhl_direction_clean.xlsx",
        REPO_ROOT / "NHL" / "step8_nhl_direction_clean.xlsx",
    ],
    "SOCCER": [
        REPO_ROOT / "Soccer" / "step8_soccer_direction_clean.xlsx",
        REPO_ROOT / "Soccer" / "outputs" / "step8_soccer_direction_clean.xlsx",
    ],
    "MLB": [
        REPO_ROOT / "MLB" / "step8_mlb_direction_clean.xlsx",
    ],
}


# Shared shell CSS versions — keep aligned with ui_runner/templates (e.g. indexGrades.html).
_TICKET_EVAL_SHELL_CSS_VER = "20260411"
_TICKET_EVAL_NAV_UNIFIED_VER = "20260408"
_TICKET_EVAL_NAV_MOBILE_VER = "20260411"
_TICKET_EVAL_MOBILE_WIDTH_VER = "20260413c"
_TICKET_EVAL_NAV_DATETIME_VER = "20260408"
_TICKET_EVAL_NAV_CHROME_VER = "20260412"


def _render_site_nav_grades_active() -> str:
    """Render _site_nav.html with Grades active (static HTML artifact; no Flask Jinja)."""
    path = TEMPLATES_DIR / "_site_nav.html"
    if not path.is_file():
        return (
            '<nav class="snav glass-card" role="navigation" aria-label="Main">'
            '<a class="snav-brand" href="/">PropOracle</a>'
            '<ul class="snav-links nav-links"><li><a href="/grades" class="active">Grades</a></li></ul></nav>'
        )
    raw = path.read_text(encoding="utf-8")
    lines = raw.splitlines()
    if lines and lines[0].lstrip().startswith("{#"):
        lines = lines[1:]
    raw = "\n".join(lines).lstrip()
    raw = re.sub(
        r"\{%\s*set\s+_na\s*=\s*nav_active\|default\('home'\)\s*%\}\s*\n?",
        "",
        raw,
        count=1,
    )
    raw = re.sub(
        r"\{%\s*set\s+_pill\s*=\s*nav_pill_suffix\|default\('LIVE'\)\s*%\}\s*\n?",
        "",
        raw,
        count=1,
    )
    raw = raw.replace("{{ _pill }}", "LIVE")

    def _active_repl(m: re.Match[str]) -> str:
        return "active" if m.group(1) == "grades" else ""

    raw = re.sub(
        r"\{\{\s*'active'\s*if\s*_na\s*==\s*'(\w+)'\s*else\s*''\s*\}\}",
        _active_repl,
        raw,
    )
    return raw.strip()


# Ticket-eval layout only; colors come from proporacle-page-shell + mobile-content-width.
_TICKET_EVAL_PAGE_WRAP_CSS = r"""
body.ticket-eval-page{
  font-family:'Inter',system-ui,sans-serif !important;
  color:var(--text, rgba(255,255,255,0.92));
}
body.ticket-eval-page{
  --glass: var(--card-bg, rgba(26,26,46,0.92));
  --glass-bd: var(--border, rgba(255,255,255,0.08));
  --gold: var(--accent, #d4af37);
  --gold2: #d4a017;
  --pending: #888;
}
:is([data-theme='light'], html.light-theme) .ticket-bucket{
  background:rgba(255,255,255,0.52);
  border-color:rgba(0,0,0,0.1);
  box-shadow:0 2px 16px rgba(0,0,0,0.06);
}
:is([data-theme='light'], html.light-theme) .ticket-bucket > summary{color:#b8860b;}
:is([data-theme='light'], html.light-theme) .ticket-bucket-meta{color:rgba(0,0,0,0.45);}
:is([data-theme='light'], html.light-theme) .ticket-bucket.sb-default{border-left-color:rgba(0,0,0,0.15);}
"""


def _dated_candidates(date_str: str) -> dict[str, list[Path]]:
    """
    Returns a copy of SPORT_XLSX_CANDIDATES with dated archive paths prepended
    for each sport bucket. Dated paths follow the naming convention used by
    run_pipeline.ps1 archive step. Only paths that actually exist are prepended.
    """
    dated_dir = REPO_ROOT / "outputs" / date_str
    dated_map = {
        "NBA": dated_dir / f"step8_nba_direction_clean_{date_str}.xlsx",
        "NBA1H": dated_dir / f"step8_nba1h_direction_clean_{date_str}.xlsx",
        "NBA1Q": dated_dir / f"step8_nba1q_direction_clean_{date_str}.xlsx",
        "CBB": dated_dir / f"step6_ranked_cbb_{date_str}.xlsx",
        "WCBB": dated_dir / f"step6_ranked_wcbb_{date_str}.xlsx",
        "NHL": dated_dir / f"step8_nhl_direction_clean_{date_str}.xlsx",
        "SOCCER": dated_dir / f"step8_soccer_direction_clean_{date_str}.xlsx",
        "MLB": dated_dir / f"step8_mlb_direction_clean_{date_str}.xlsx",
    }
    result: dict[str, list[Path]] = {}
    for bucket, live_paths in SPORT_XLSX_CANDIDATES.items():
        dated_path = dated_map.get(bucket)
        if dated_path and dated_path.is_file():
            result[bucket] = [dated_path] + list(live_paths)
        else:
            result[bucket] = list(live_paths)
    return result


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _norm_player_name(s: str) -> str:
    """Shared with MLB grader + slate_grader (scripts/player_name_norm.py)."""
    if s is None or (isinstance(s, float) and pd.isna(s)):
        return ""
    return _fold_player_name(s)


def _canon_player(row: dict[str, Any]) -> str:
    for k in ("player", "athlete", "name"):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _canon_prop(row: dict[str, Any]) -> str:
    for k in (
        "prop_type",
        "prop type",
        "prop_type_norm",
        "prop_norm",
        "prop_label",
        "prop",
        "prop_display",
        "stat_type",
    ):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _canon_direction(row: dict[str, Any]) -> str:
    for k in ("direction", "bet_direction", "final_bet_direction", "pick direction"):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip().upper()
    return ""


def _canon_line(row: dict[str, Any]) -> float | None:
    for k in ("line", "line_num"):
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        try:
            return float(v)
        except (TypeError, ValueError):
            continue
    return None


def _canon_actual(row: dict[str, Any]) -> float | None:
    for k in (
        "actual",
        "actual_value",
        "act",
        "result_value",
        "stat_actual",
        "final_stat",
        "box",
        "box_score",
        "game_stat",
        "stat",
        "final",
    ):
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if isinstance(v, str) and not v.strip():
            continue
        try:
            return float(v)
        except (TypeError, ValueError):
            continue
    return None


def _cell_looks_like_grade_outcome(s: str) -> bool:
    """
    True if a workbook cell is probably HIT/MISS/etc., not a numeric game stat.
    Prevents columns named 'result' that hold 14.0 from forcing the wrong path.
    """
    u = str(s).strip().upper()
    if not u or u in (".", "-", "—"):
        return False
    if u in (
        "HIT",
        "WIN",
        "W",
        "MISS",
        "LOSS",
        "L",
        "VOID",
        "PUSH",
        "PENDING",
        "N/A",
        "NA",
        "TBD",
        "OPEN",
        "TRUE",
        "FALSE",
        "YES",
        "NO",
        "0",
        "1",
    ):
        return True
    if re.fullmatch(r"-?\d+\.?\d*", u):
        return False
    if len(u) <= 16 and re.fullmatch(r"[A-Z][A-Z0-9_/-]*", u):
        return True
    return False


def _canon_grade_raw(row: dict[str, Any]) -> str:
    for k in ("grade", "leg_result", "outcome", "result"):
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip().upper()
        if s and _cell_looks_like_grade_outcome(s):
            return s
    return ""


def _canon_void_note(row: dict[str, Any]) -> str:
    """Grader diagnostic (e.g. POSTPONED, NO_ACTUAL) from workbook row."""
    for k in (
        "void_reason_grade",
        "void reason grade",
        "void_reason",
        "void reason",
    ):
        if k not in row:
            continue
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip()
        if s:
            return s
    return ""


def _normalize_workbook_rows(path: Path) -> list[dict[str, Any]]:
    """Load all sheets; normalize headers to lowercase single-space keys."""
    xl = pd.ExcelFile(path)
    out: list[dict[str, Any]] = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        if df.empty:
            continue
        df.columns = [_norm_header(c) for c in df.columns]
        # Skip styled title/summary sheets — their first column header is a long
        # title string (e.g. "CBB SLATE GRADE  |  2026-03-27  |  Generated ...").
        # Real data sheets have short, clean column names.
        first_col = df.columns[0] if len(df.columns) else ""
        if len(first_col) > 40 or "|" in first_col:
            continue
        out.extend(df.to_dict(orient="records"))
    return out


def _finite_line_actual(actual: float | None, line: float | None) -> bool:
    if actual is None or line is None:
        return False
    try:
        a, ln = float(actual), float(line)
    except (TypeError, ValueError):
        return False
    if isinstance(a, float) and (math.isnan(a) or math.isinf(a)):
        return False
    if isinstance(ln, float) and (math.isnan(ln) or math.isinf(ln)):
        return False
    return True


def _leg_grade(
    actual: float | None,
    line: float | None,
    direction: str,
    grade_col: str,
) -> str:
    """
    Prefer numeric grading (same margin rules as scripts/nhl_soccer_grader) when actual+line exist,
    so stale VOID from pre-grade slates does not mask a real box score.
    """
    g = (grade_col or "").strip().upper()
    d = str(direction or "").strip().upper()

    if _finite_line_actual(actual, line) and d in ("OVER", "UNDER"):
        a = float(actual)  # type: ignore[arg-type]
        ln = float(line)  # type: ignore[arg-type]
        margin = a - ln
        if margin == 0:
            return "VOID"
        if (d == "OVER" and margin > 0) or (d == "UNDER" and margin < 0):
            return "HIT"
        return "MISS"

    if g in ("HIT", "WIN", "W", "1", "TRUE", "YES"):
        return "HIT"
    if g in ("MISS", "LOSS", "L", "0", "FALSE", "NO"):
        return "MISS"
    if g in ("VOID", "PUSH", "N/A", "NA"):
        return "VOID"
    return "UNGRADED"


def _ticket_is_flex_play_structure(group_name: str, n_legs: int) -> bool:
    """PrizePicks-style flex slips (3+ legs): one miss can still cash at the flex multiplier."""
    if n_legs < 3:
        return False
    return "flex" in str(group_name or "").strip().lower()


def _safe_float_ticket(x: Any, default: float | None = None) -> float | None:
    if x is None or x == "":
        return default
    try:
        v = float(x)
        if isinstance(v, float) and math.isnan(v):
            return default
        return v
    except (TypeError, ValueError):
        return default


def _ticket_eval_money_outcome(group_name: str, leg_grades: list[str], ticket: dict[str, Any]) -> dict[str, Any]:
    """
    Empirical payout model vs graded leg outcomes for ticket_eval HTML.

    ``ticket`` may include ``payout`` from tickets_latest.json (full empirical block). Tickets loaded
    from ``combined_slate_tickets_*.xlsx`` only have ``power_payout`` / ``flex_payout`` from the sheet
    banner — we fall back to those so Predicted / Actual are not N/A and Power wins are not $0.
    """
    pay = ticket.get("payout")
    payd: dict[str, Any] = pay if isinstance(pay, dict) else {}

    if any(g == "UNGRADED" for g in leg_grades):
        return {"pending": True}

    n = len(leg_grades)
    h = sum(1 for g in leg_grades if g == "HIT")
    m = sum(1 for g in leg_grades if g == "MISS")
    v = sum(1 for g in leg_grades if g == "VOID")
    all_hit = bool(n) and all(g == "HIT" for g in leg_grades)

    if n > 0 and v == n:
        return {
            "pending": False,
            "result": "VOID",
            "result_display": "NO CONTEST",
            "result_emoji": "○",
            "result_css": "void",
            "result_detail": f"All {n} legs void — no graded contest against the board.",
            "actual_payout": 0.0,
            "predicted_payout": None,
            "predicted_ev": None,
            "predicted_p_win": None,
            "recommendation_at_entry": "",
            "entry_10_return": 0.0,
            "net_10": 0.0,
            "omit_payout_block": True,
        }

    paid = _ticket_pays_money(group_name, leg_grades)

    flex = _ticket_is_flex_play_structure(group_name, n)
    banner_pow = _safe_float_ticket(ticket.get("power_payout")) or 0.0
    banner_flex = _safe_float_ticket(ticket.get("flex_payout")) or 0.0

    min_x = _safe_float_ticket(payd.get("payout")) or _safe_float_ticket(payd.get("min_guarantee"))
    if min_x is None or min_x <= 0:
        # Excel-combined path: flex cash floor lives on flex_payout; power min tier matches banner power.
        if flex and banner_flex > 0:
            min_x = float(banner_flex)
        elif banner_pow > 0:
            min_x = float(banner_pow)
        else:
            min_x = 0.0

    sweep_x = _safe_float_ticket(payd.get("sweep_payout")) or _safe_float_ticket(payd.get("first_place"))
    if sweep_x is None or sweep_x <= 0:
        sweep_x = float(banner_pow) if banner_pow > 0 else 0.0

    if not paid:
        result = "LOSS"
        emoji = "❌"
        css = "loss"
        actual = 0.0
        if m == 0 and v > 0:
            result = "VOID_LOSS"
            emoji = "⚠"
            css = "void_loss"
    elif flex:
        if all_hit:
            result = "SWEEP"
            emoji = "🏆"
            css = "sweep"
            actual = float(sweep_x)
        else:
            result = "MIN GUARANTEE"
            emoji = "🛡️"
            css = "min_guarantee"
            actual = float(min_x)
    else:
        # Power Play: all legs correct pays the full board (sweep) multiplier on PrizePicks.
        result = "WIN"
        emoji = "✅"
        css = "win"
        pay_win = float(sweep_x) if sweep_x > 0 else float(min_x)
        actual = pay_win

    pred_pay = _safe_float_ticket(payd.get("payout")) or _safe_float_ticket(payd.get("min_guarantee"))
    if pred_pay is None or pred_pay <= 0:
        if flex and banner_flex > 0:
            pred_pay = float(banner_flex)
        elif banner_pow > 0:
            pred_pay = float(banner_pow)
        else:
            pred_pay = None
    # XLSX-only tickets: flex sweep pays the power-board multiplier; show that as predicted, not the 1-miss floor.
    if not payd and flex and paid and all_hit and banner_pow > 0:
        pred_pay = float(banner_pow)
    pred_ev = _safe_float_ticket(payd.get("ev"))
    pred_p = _safe_float_ticket(payd.get("p_all_win"))
    rec = str(payd.get("recommendation") or "").strip()

    gross_10 = round(10.0 * actual, 2)
    net_10 = round(gross_10 - 10.0, 2)

    if (
        not payd
        and paid
        and actual <= 0
        and result in ("WIN", "SWEEP", "MIN GUARANTEE")
    ):
        _log.warning(
            "[ticket_eval] Graded outcome %s but payout multiplier is 0 without empirical JSON; "
            "check banner multipliers. group=%r ticket_no=%r power_payout=%r flex_payout=%r n_legs=%s",
            result,
            group_name,
            ticket.get("ticket_no"),
            ticket.get("power_payout"),
            ticket.get("flex_payout"),
            n,
        )

    h_show = h
    m_show = m
    if result == "MIN GUARANTEE" and n:
        detail = f"{result} ({h_show}/{n} correct)"
    elif result == "SWEEP" and n:
        detail = f"{result} — all {n} legs correct"
    elif result == "WIN" and n:
        detail = f"{result} — all {n} legs correct"
    elif result == "VOID_LOSS" and n:
        detail = f"Voided legs prevented payout ({h_show} hit, {m_show} miss, {v} void)"
    elif result == "LOSS" and n:
        detail = f"{result} ({h_show} hit, {m_show} miss)"
    else:
        detail = result

    out: dict[str, Any] = {
        "pending": False,
        "result": result,
        "result_emoji": emoji,
        "result_css": css,
        "result_detail": detail,
        "actual_payout": actual,
        "predicted_payout": pred_pay,
        "predicted_ev": pred_ev,
        "predicted_p_win": pred_p,
        "recommendation_at_entry": rec,
        "entry_10_return": gross_10,
        "net_10": net_10,
    }
    if result == "VOID_LOSS":
        out["result_display"] = "VOID / NO ACTION"
    return out


def _fmt_pay_cell(v: float | None, suffix: str = "x") -> str:
    if v is None:
        return "N/A"
    try:
        return f"{float(v):.2f}{suffix}"
    except (TypeError, ValueError):
        return "N/A"


def _fmt_pct_cell(v: float | None) -> str:
    if v is None:
        return "N/A"
    try:
        x = float(v)
        if x <= 1.0:
            return f"{x * 100:.1f}%"
        return f"{x:.1f}%"
    except (TypeError, ValueError):
        return "N/A"


def _append_grade_history(record: dict[str, Any]) -> None:
    """Append (or replace same-date) run summary to data/grade_history.json."""
    path = REPO_ROOT / "data" / "grade_history.json"
    path.parent.mkdir(parents=True, exist_ok=True)
    runs: list[Any] = []
    if path.is_file():
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
            if isinstance(raw, list):
                runs = list(raw)
            elif isinstance(raw, dict) and isinstance(raw.get("runs"), list):
                runs = list(raw["runs"])
        except (OSError, json.JSONDecodeError):
            runs = []
    ds = str(record.get("date") or "")[:10]
    runs = [r for r in runs if not (isinstance(r, dict) and str(r.get("date", ""))[:10] == ds)]
    runs.append(record)
    path.write_text(json.dumps(runs, indent=2, ensure_ascii=False) + "\n", encoding="utf-8")


def _ticket_pays_money(group_name: str, leg_grades: list[str]) -> bool:
    """
    Cash outcome: power = every leg HIT; flex (sheet title contains 'Flex', 3+ legs) = at most one
    MISS and at least n-1 HITs (e.g. 2/3 or 3/4). VOID legs are neither HIT nor MISS.
    Caller must ensure no UNGRADED legs.
    """
    if not leg_grades or any(g == "UNGRADED" for g in leg_grades):
        return False
    n = len(leg_grades)
    h = sum(1 for g in leg_grades if g == "HIT")
    m = sum(1 for g in leg_grades if g == "MISS")
    if _ticket_is_flex_play_structure(group_name, n):
        return m <= 1 and h >= n - 1
    return all(g == "HIT" for g in leg_grades)


def _pick_type_tier(pick_type: str) -> str:
    p = (pick_type or "").strip().lower()
    if "goblin" in p:
        return "G"
    if "demon" in p:
        return "D"
    if "standard" in p:
        return "S"
    return (pick_type[:1].upper() if pick_type else "?")


def _sport_key(sport: str) -> str:
    """Normalize for display / CSS (keep variant labels visible)."""
    s = (sport or "").strip().upper().replace(" ", "")
    if s in ("SOC", "MLS", "EPL"):
        return "SOCCER"
    return s


def _ticket_group_bucket(group_name: str) -> str:
    """Bucket for collapsible UI: sport prefix from sheet title, or Cross-sport for XSPORT*."""
    n = (group_name or "").strip()
    if not n:
        return "Other"
    if n.upper().startswith("XSPORT"):
        return "Cross-sport"
    return n.split()[0] or "Other"


def _ticket_bucket_skin_class(bucket: str) -> str:
    if bucket == "Cross-sport":
        return "sb-xsport"
    sk = _sport_key(bucket)
    return {
        "NBA": "sb-nba",
        "NBA1H": "sb-nba1h",
        "NBA1Q": "sb-nba1q",
        "CBB": "sb-cbb",
        "WCBB": "sb-wcbb",
        "NHL": "sb-nhl",
        "SOCCER": "sb-soccer",
        "MLB": "sb-mlb",
    }.get(sk, "sb-default")


def _bucket_ticket_groups(groups: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    order: list[str] = []
    m: dict[str, list[dict[str, Any]]] = {}
    for g in groups:
        key = _ticket_group_bucket(str(g.get("group_name") or ""))
        if key not in m:
            m[key] = []
            order.append(key)
        m[key].append(g)
    return [(k, m[k]) for k in order]


def _leg_match_buckets(sport: str) -> list[str]:
    """
    Order matters: try variant-specific slate first, then parent sport fallback.
    """
    s = (sport or "").strip().upper().replace(" ", "").replace("-", "")
    if s in ("NBA1H", "NBA_1H"):
        return ["NBA1H", "NBA"]
    if s in ("NBA1Q", "NBA_1Q"):
        return ["NBA1Q", "NBA"]
    if s == "WCBB":
        return ["WCBB", "CBB"]
    if s in ("SOC", "MLS", "EPL"):
        return ["SOCCER"]
    if s in ("NBA", "WNBA"):
        return ["NBA"]
    if s == "CBB":
        return ["CBB"]
    if s == "NHL":
        return ["NHL"]
    if s == "SOCCER":
        return ["SOCCER"]
    if s == "MLB":
        return ["MLB"]
    return [s, "NBA", "CBB"]


def _ingest_workbook_rows_into_index(
    rows: list[dict[str, Any]],
    triple: dict[tuple[str, str, str], dict],
    pair_buckets: dict[tuple[str, str], list[dict]],
) -> None:
    for raw in rows:
        pl = _norm_player_name(_canon_player(raw))
        pt = _prop_match_key_from_display(str(_canon_prop(raw) or ""))
        dr = _canon_direction(raw)
        if not pl or not pt:
            continue
        row = {
            "player_lower": pl,
            "prop_lower": pt,
            "direction": dr,
            "line": _canon_line(raw),
            "actual": _canon_actual(raw),
            "grade_raw": _canon_grade_raw(raw),
            "void_note": _canon_void_note(raw),
        }
        # Include line in the triple key so multiple lines for the same
        # player+prop+direction (e.g. Cameron Boozer reb 7.5 OVER vs 10.5 OVER)
        # don't collide and overwrite each other.
        line_val = row["line"]
        line_key = round(float(line_val), 2) if line_val is not None else None
        key3 = (pl, pt, dr, line_key)
        triple[key3] = row
        pair_buckets.setdefault((pl, pt), []).append(row)


def _graded_xlsx_in_dir(d: Path) -> list[Path]:
    """Graded sport workbooks in a directory (outputs date folder or ui_runner graded_slate bundle)."""
    if not d.is_dir():
        return []
    found: set[Path] = set()
    for pat in ("graded_*.xlsx", "*_graded_*.xlsx"):
        for p in d.glob(pat):
            if not p.is_file():
                continue
            low = p.name.lower()
            if "combined_tickets_graded" in low:
                continue
            found.add(p)
    return sorted(found, key=lambda x: x.name.lower())


def _graded_xlsx_in_outputs_date(arg_date: str) -> list[Path]:
    """Graded slates next to daily artifacts: outputs/YYYY-MM-DD/*.xlsx."""
    return _graded_xlsx_in_dir(REPO_ROOT / "outputs" / arg_date)


def _graded_slate_bundle_dir(arg_date: str) -> Path:
    """Committed graded workbooks for deploy hosts without outputs/ (e.g. Railway)."""
    return REPO_ROOT / "ui_runner" / "graded_slate" / arg_date


def find_graded_workbook_path(sport_slug: str, date_str: str) -> Path | None:
    """
    Resolve ``graded_<sport>_<date>.xlsx``: prefer ``ui_runner/graded_slate/<date>/`` (git/Railway),
    then ``outputs/<date>/`` (local). ``sport_slug`` is the filename token, e.g. ``nhl``, ``nba``, ``soccer``.
    """
    fn = f"graded_{sport_slug}_{date_str}.xlsx"
    deploy = REPO_ROOT / "ui_runner" / "graded_slate" / date_str / fn
    if deploy.is_file():
        return deploy
    local = REPO_ROOT / "outputs" / date_str / fn
    if local.is_file():
        return local
    return None


def _merge_strict_graded_date_workbooks(
    out: dict[str, tuple[dict, dict]],
    graded_dir: Path,
    arg_date: str,
) -> None:
    """
    Overlay rows from graded_<sport>_<date>.xlsx (Box Raw sheet) with full-row actual/grade parsing.
    Same file names as in outputs/<date>/.
    """
    if not graded_dir.is_dir():
        return
    sport_to_bucket = {
        "nba": "NBA",
        "cbb": "CBB",
        "nhl": "NHL",
        "soccer": "SOCCER",
        "wcbb": "WCBB",
        "mlb": "MLB",
        "nba1h": "NBA1H",
        "nba1q": "NBA1Q",
    }
    for graded_file in sorted(graded_dir.glob(f"graded_*_{arg_date}.xlsx")):
        m = re.match(r"^graded_(.+)_(\d{4}-\d{2}-\d{2})$", graded_file.stem)
        if not m:
            continue
        sport_tag = m.group(1).lower()
        bucket = sport_to_bucket.get(sport_tag)
        if bucket is None:
            continue
        try:
            xl = pd.ExcelFile(graded_file, engine="openpyxl")
            sheet_priority = ["Box Raw", "Props", "Graded"]
            sheet_to_use = next(
                (s for s in sheet_priority if s in xl.sheet_names),
                xl.sheet_names[0],
            )
            gdf = xl.parse(sheet_to_use)
            gdf.columns = [str(c).lower().strip() for c in gdf.columns]
            pcol = next((c for c in gdf.columns if c == "player"), None)
            propcol = next(
                (c for c in gdf.columns if c in ("prop", "prop_type", "prop_type_norm", "stat", "stat_type")),
                None,
            )
            if pcol is None or propcol is None:
                continue
            trip, pairs = out.get(bucket, ({}, {}))
            for _, grow in gdf.iterrows():
                raw = {c: grow[c] for c in gdf.columns}
                pl = _norm_player_name(_canon_player(raw))
                pt = _prop_match_key_from_display(str(_canon_prop(raw) or ""))
                dr = _canon_direction(raw)
                if not pl or not pt:
                    continue
                row_out = {
                    "player_lower": pl,
                    "prop_lower": pt,
                    "direction": dr,
                    "line": _canon_line(raw),
                    "actual": _canon_actual(raw),
                    "grade_raw": _canon_grade_raw(raw),
                    "void_note": _canon_void_note(raw),
                }
                line_val = row_out["line"]
                line_key = round(float(line_val), 2) if line_val is not None else None
                key3 = (pl, pt, dr, line_key)
                trip[key3] = row_out
                pairs.setdefault((pl, pt), []).append(row_out)
            out[bucket] = (trip, pairs)
        except Exception:
            continue


def _sport_buckets_for_graded_filename(path: Path) -> list[str]:
    """
    Map a graded workbook name to one or more SPORT_XLSX_CANDIDATES keys.
    Unknown names return [] (skipped).
    """
    n = path.name.lower()
    s = path.stem.lower()
    if "mlb" in n:
        return ["MLB"]
    if "nhl" in n:
        return ["NHL"]
    if "soccer" in n or s.startswith("soccer_graded"):
        return ["SOCCER"]
    if "wcbb" in n or "wcbb" in s:
        return ["WCBB"]
    if "cbb" in n or "cbb" in s or "ncaab" in n:
        return ["CBB"]
    if "nba1h" in n or "nba_1h" in n:
        return ["NBA1H"]
    if "nba1q" in n or "nba_1q" in n:
        return ["NBA1Q"]
    if "nba" in n:
        return ["NBA"]
    return []


def _merge_graded_workbooks_into_indices(
    indices: dict[str, tuple[dict[tuple[str, str, str], dict], dict[tuple[str, str], list[dict]]]],
    graded_paths: list[Path],
) -> int:
    """
    Overlay rows from dated graded exports so Actual / Result columns populate ticket eval.
    Returns number of workbook files successfully merged.
    """
    merged = 0
    for path in graded_paths:
        buckets = _sport_buckets_for_graded_filename(path)
        if not buckets:
            continue
        try:
            rows = _normalize_workbook_rows(path)
        except Exception:
            continue
        if not rows:
            continue
        for bkt in buckets:
            trip, pairs = indices.get(bkt, ({}, {}))
            _ingest_workbook_rows_into_index(rows, trip, pairs)
            indices[bkt] = (trip, pairs)
        merged += 1
    return merged


def _reference_datetime_for_parse(slate_date: str) -> datetime:
    """Default for dateutil when parsing leg game_time strings missing year (e.g. '04/14 3:00 PM')."""
    s = str(slate_date or "").strip()[:10]
    if re.match(r"^\d{4}-\d{2}-\d{2}$", s):
        try:
            y, m, d_ = int(s[:4]), int(s[5:7]), int(s[8:10])
            return datetime(y, m, d_, 12, 0, 0)
        except ValueError:
            pass
    return datetime.now(tz=timezone.utc).replace(tzinfo=None)


def _game_dates_from_ticket_payload(payload: dict[str, Any], slate_date: str) -> set[str]:
    """Unique YYYY-MM-DD calendar dates from leg game_time / start_time (ISO or MDY-style strings)."""
    dates: set[str] = set()
    ref = _reference_datetime_for_parse(slate_date)
    for group in payload.get("groups") or []:
        for ticket in group.get("tickets") or []:
            for leg in ticket.get("legs") or []:
                gt = leg.get("game_time") or leg.get("start_time") or ""
                if not str(gt).strip():
                    continue
                try:
                    dt = _parse_datetime_guess(str(gt).strip(), default=ref)
                    dates.add(dt.date().isoformat())
                except (ValueError, TypeError, OverflowError):
                    continue
    return dates


# combined_slate_tickets_*.xlsx often omits Game Time; evening pipeline slips still target the next
# calendar day for these books. Used only when no leg game_time / start_time was parsed.
_TEAM_SPORTS_INFER_NEXT_DAY_GRADED: frozenset[str] = frozenset(
    {"NBA", "NBA1H", "NBA1Q", "MLB", "NHL", "SOCCER", "CBB", "WCBB"}
)


def _payload_has_infer_next_day_sport_legs(payload: dict[str, Any]) -> bool:
    for group in payload.get("groups") or []:
        for ticket in group.get("tickets") or []:
            for leg in ticket.get("legs") or []:
                sp = str(leg.get("sport") or "").strip().upper()
                if sp in _TEAM_SPORTS_INFER_NEXT_DAY_GRADED:
                    return True
    return False


def _inferred_next_calendar_game_date(slate_date: str) -> str | None:
    try:
        y, m, d_ = int(slate_date[:4]), int(slate_date[5:7]), int(slate_date[8:10])
        return (date(y, m, d_) + timedelta(days=1)).isoformat()
    except (ValueError, IndexError, TypeError):
        return None


def resolve_ticket_eval_graded_merge_dates(
    slate_date: str,
    payload: dict[str, Any],
    extra_iso_dates: Sequence[str] | None = None,
) -> tuple[list[str], list[str]]:
    """
    Ordered list of calendar folders whose graded_*.xlsx files are merged into ticket eval,
    and leg-side game dates for logging (parsed from legs, or inferred when xlsx has no game_time).

    Slate date is merged first, then each leg game date (sorted), then optional --game-date extras.
    Later merges overwrite matching keys so game-day graded exports win over slate-day rows.
    """
    parsed = _game_dates_from_ticket_payload(payload, slate_date)
    from_legs: set[str] = set(parsed)
    if not from_legs and _payload_has_infer_next_day_sport_legs(payload):
        nxt = _inferred_next_calendar_game_date(slate_date)
        if nxt:
            from_legs.add(nxt)
    out: list[str] = []
    sd = str(slate_date).strip()
    if re.match(r"^\d{4}-\d{2}-\d{2}$", sd):
        out.append(sd)
    for d in sorted(from_legs):
        if d not in out:
            out.append(d)
    for raw in extra_iso_dates or ():
        ds = str(raw).strip()
        if re.match(r"^\d{4}-\d{2}-\d{2}$", ds) and ds not in out:
            out.append(ds)
    leg_dates_for_log = sorted(from_legs)
    return out, leg_dates_for_log


def _merge_graded_workbooks_for_date(
    out: dict[str, tuple[dict, dict]],
    gd: str,
) -> None:
    """Merge graded workbooks from ui_runner/graded_slate/{gd}/ and outputs/{gd}/ (no-op if folders missing)."""
    d = str(gd).strip()
    if not d or not re.match(r"^\d{4}-\d{2}-\d{2}$", d):
        return
    bundle_dir = _graded_slate_bundle_dir(d)
    _merge_graded_workbooks_into_indices(out, _graded_xlsx_in_dir(bundle_dir))
    _merge_strict_graded_date_workbooks(out, bundle_dir, d)
    outputs_dir = REPO_ROOT / "outputs" / d
    _merge_graded_workbooks_into_indices(out, _graded_xlsx_in_dir(outputs_dir))
    _merge_strict_graded_date_workbooks(out, outputs_dir, d)


def _load_actuals_indices(
    sport_candidates: dict[str, list[Path]],
    graded_merge_dates: list[str],
) -> dict[str, tuple[dict[tuple[str, str, str], dict], dict[tuple[str, str], list[dict]]]]:
    """Per sport-bucket indices (NBA1H separate from NBA, etc.)."""
    out: dict[str, tuple[dict, dict]] = {}
    for bucket, paths in sport_candidates.items():
        triple: dict[tuple[str, str, str], dict] = {}
        pair_buckets: dict[tuple[str, str], list[dict]] = {}
        src = next((p for p in paths if p.is_file()), None)
        if not src:
            out[bucket] = (triple, pair_buckets)
            continue
        try:
            rows = _normalize_workbook_rows(src)
        except Exception:
            out[bucket] = (triple, pair_buckets)
            continue
        _ingest_workbook_rows_into_index(rows, triple, pair_buckets)
        out[bucket] = (triple, pair_buckets)

    for gd in graded_merge_dates:
        if gd and re.match(r"^\d{4}-\d{2}-\d{2}$", str(gd).strip()):
            _merge_graded_workbooks_for_date(out, str(gd).strip())
    return out


# CBB graded workbook (Box Raw sheet) uses abbreviated prop_type_norm values.
# Map them to the full names used in ticket JSON legs so matching works.
_PROP_TYPE_ALIASES: dict[str, str] = {
    # abbreviated → full (ticket JSON uses full names)
    "pts":                   "points",
    "reb":                   "rebounds",
    "ast":                   "assists",
    "blk":                   "blocked shots",
    "stl":                   "steals",
    "to":                    "turnovers",
    "pr":                    "pts+rebs",
    "pa":                    "pts+asts",
    "ra":                    "rebs+asts",
    "pra":                   "pts+rebs+asts",
    # also normalise common full-name variants so reverse lookups work
    "pts+rebs+asts":         "pts+rebs+asts",
    "pts+rebs":              "pts+rebs",
    "pts+asts":              "pts+asts",
    "rebs+asts":             "rebs+asts",
    "points+rebounds+assists": "pts+rebs+asts",
    "points+rebounds":       "pts+rebs",
    "points+assists":        "pts+asts",
    "rebounds+assists":      "rebs+asts",
    # fantasy / combo labels seen in actuals CSV
    "fantasy score":         "fantasy score",
    "pts+rebs+asts":         "pts+rebs+asts",
}

def _norm_prop_type(raw: str) -> str:
    """Lowercase + strip, then apply alias map so abbreviated and full names match."""
    s = raw.strip().lower()
    return _PROP_TYPE_ALIASES.get(s, s)


def _prop_match_key_from_display(raw: str) -> str:
    """
    Stable prop key for index lookups: alias map (CBB etc.) then alphanumeric fold like
    nhl_soccer_grader._norm_prop so 'Total Bases' and 'totalbases' both become 'totalbases'.
    """
    if raw is None or not str(raw).strip():
        return ""
    return re.sub(r"[^a-z0-9]", "", _norm_prop_type(str(raw).strip()))


def _match_leg_in_index(
    leg: dict[str, Any],
    triple: dict[tuple[str, str, str], dict],
    pair_buckets: dict[tuple[str, str], list[dict]],
) -> dict | None:
    pl = _norm_player_name(leg.get("player") or "")
    pt = _prop_match_key_from_display(str(leg.get("prop_type") or ""))
    dr = str(leg.get("direction") or "").strip().upper()
    if not pl or not pt:
        return None

    try:
        leg_line = float(leg.get("line"))
    except (TypeError, ValueError):
        leg_line = None

    # Try exact 4-tuple (player, prop, direction, line) — most specific.
    if leg_line is not None:
        line_key = round(leg_line, 2)
        hit = triple.get((pl, pt, dr, line_key))
        if hit:
            return hit

    # Try 3-tuple without line (legacy / no-line rows).
    hit = triple.get((pl, pt, dr, None))
    if hit:
        return hit

    cands = pair_buckets.get((pl, pt))
    if not cands:
        return None

    # Pass 1: direction + line exact match (most precise).
    if leg_line is not None:
        for r in cands:
            if r["direction"] == dr and r.get("line") is not None:
                if abs(float(r["line"]) - leg_line) < 0.01:
                    return r

    # Pass 2: direction match only.
    for r in cands:
        if r["direction"] == dr:
            return r

    if len(cands) == 1:
        return cands[0]
    for r in cands:
        if r["direction"] == dr or not r["direction"]:
            return r
    return cands[0]


def _match_leg_to_row_multi(
    leg: dict[str, Any],
    indices: dict[str, tuple[dict, dict]],
) -> dict | None:
    for bkt in _leg_match_buckets(str(leg.get("sport") or "")):
        trip, pairs = indices.get(bkt, ({}, {}))
        row = _match_leg_in_index(leg, trip, pairs)
        if row:
            return row
    return None


def _graded_outputs_dir(arg_date: str) -> Path:
    return REPO_ROOT / "outputs" / arg_date


def _debug_list_outputs_graded(arg_date: str) -> list[Path]:
    d = _graded_outputs_dir(arg_date)
    if not d.is_dir():
        return []
    return sorted(d.glob("graded_*.xlsx"))


def _debug_sheet_headers(path: Path, max_sheets: int = 3) -> list[tuple[str, list[str]]]:
    """Per sheet: (sheet_name, normalized column names)."""
    out: list[tuple[str, list[str]]] = []
    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        return [("<<read_error>>", [str(e)])]
    for i, sh in enumerate(xl.sheet_names):
        if i >= max_sheets:
            out.append(("...", [f"(+{len(xl.sheet_names) - max_sheets} more sheets)"]))
            break
        try:
            df = pd.read_excel(path, sheet_name=sh, nrows=0)
            cols = [_norm_header(c) for c in df.columns]
        except Exception as e:
            cols = [f"<<{e}>>"]
        out.append((sh, cols))
    return out


def debug_report(
    arg_date: str,
    payload: dict[str, Any],
    tpath: Path,
    sport_candidates: dict[str, list[Path]],
    graded_merge_dates: list[str],
) -> None:
    """Print why legs may not match (JSON date vs CLI, xlsx paths, headers, sample legs)."""
    print("\n=== build_ticket_eval.py --debug ===\n")
    print(f"CLI --date:     {arg_date}")
    print(f"Ticket source:  {tpath}")
    print(f"Payload \"date\": {payload.get('date')!r}")
    leg_game_dates = sorted(_game_dates_from_ticket_payload(payload, arg_date))
    print(f"Leg game_time dates: {leg_game_dates if leg_game_dates else '(none parsed)'}")
    print(f"Graded merge order:  {graded_merge_dates}")
    if str(payload.get("date") or "").strip() != arg_date:
        print(
            "  ! Mismatch: ticket payload date differs from --date; legs are still matched against"
            " STATIC pipeline workbooks (see below), not per-date outputs unless we add that."
        )
    for gd in graded_merge_dates:
        out_dir = _graded_outputs_dir(gd)
        og = _debug_list_outputs_graded(gd)
        print(f"\noutputs/{gd}/ graded_*.xlsx:")
        if not out_dir.is_dir():
            print(f"  (folder missing: {out_dir})")
        elif not og:
            print("  (none found)")
        else:
            for p in og:
                print(f"  - {p.relative_to(REPO_ROOT)}")
    print("\nWorkbooks used for matching (first existing path per sport; NOT date-specific today):")
    for sport, paths in sport_candidates.items():
        src = next((p for p in paths if p.is_file()), None)
        if not src:
            print(f"  {sport}: (no file at any candidate path)")
            for p in paths:
                print(f"       tried: {p.relative_to(REPO_ROOT)}")
            continue
        print(f"  {sport}: {src.relative_to(REPO_ROOT)}")
        for sh, cols in _debug_sheet_headers(src):
            preview = cols[:24]
            extra = f" ...(+{len(cols) - 24})" if len(cols) > 24 else ""
            print(f"       sheet {sh!r}: {preview}{extra}")

    for gd in graded_merge_dates:
        bundle_dir = _graded_slate_bundle_dir(gd)
        bpaths = _graded_xlsx_in_dir(bundle_dir)
        print(f"\nui_runner/graded_slate/{gd}/ (optional; for Railway when outputs/ is absent):")
        if not bpaths:
            print("  (none — run scripts/run_grader.ps1; it copies graded_*.xlsx here, or copy manually)")
        else:
            for p in bpaths:
                bk = ", ".join(_sport_buckets_for_graded_filename(p)) or "?"
                print(f"  - {p.relative_to(REPO_ROOT)}  -> buckets [{bk}]")

    print("\nfind_graded_workbook_path() (deploy path first, then outputs/):")
    for gd in graded_merge_dates:
        print(f"  --- date {gd} ---")
        for slug in ("nba", "nhl", "mlb", "soccer", "cbb", "wcbb"):
            p = find_graded_workbook_path(slug, gd)
            label = f"graded_{slug}_{gd}.xlsx"
            if p:
                print(f"  {label} -> {p.relative_to(REPO_ROOT)}")
            else:
                print(f"  {label} -> (not found)")

    indices = _load_actuals_indices(sport_candidates, graded_merge_dates)
    print("\nGraded workbook(s) merged into indices (per date in merge order; later overwrites keys):")
    any_g = False
    for gd in graded_merge_dates:
        gpaths = _graded_xlsx_in_outputs_date(gd)
        if not gpaths:
            continue
        any_g = True
        print(f"  outputs/{gd}/:")
        for p in gpaths:
            bk = ", ".join(_sport_buckets_for_graded_filename(p)) or "?"
            print(f"    - {p.relative_to(REPO_ROOT)}  -> buckets [{bk}]")
    if not any_g:
        print("  (none under outputs/<merge-date>/ — add graded_nba_<date>.xlsx, etc.)")
    total_triples = sum(len(t) for t, _ in indices.values())
    total_pairs = sum(len(p) for _, p in indices.values())
    print(f"\nIndex (all buckets): {total_triples:,} triple-keys, {total_pairs:,} player+prop buckets (sum per sport)")
    for bkt, (tr, pr) in indices.items():
        if tr or pr:
            print(f"  {bkt}: {len(tr):,} triples, {len(pr):,} pair-buckets")

    groups = payload.get("groups") or []
    legs_sample: list[dict[str, Any]] = []
    for g in groups:
        for t in g.get("tickets") or []:
            for leg in t.get("legs") or []:
                legs_sample.append(leg)
                if len(legs_sample) >= 8:
                    break
            if len(legs_sample) >= 8:
                break
        if len(legs_sample) >= 8:
            break

    print("\nSample legs (match against index above):")
    for i, leg in enumerate(legs_sample, 1):
        pl = _norm_player_name(leg.get("player") or "")
        pt = str(leg.get("prop_type") or "").strip().lower()
        pk = _prop_match_key_from_display(str(leg.get("prop_type") or ""))
        dr = str(leg.get("direction") or "").strip().upper()
        row = _match_leg_to_row_multi(leg, indices)
        st = "MATCH" if row else "NO MATCH -> UNGRADED"
        sp = str(leg.get("sport") or "")
        bk = " -> ".join(_leg_match_buckets(sp))
        print(
            f"  {i}. sport={sp!r} buckets=[{bk}] player={pl!r} prop_key={pk!r} prop_type={pt!r} direction={dr!r} -> {st}"
        )
        if row:
            print(
                f"      actual={row.get('actual')!r} line={row.get('line')!r} "
                f"grade_raw={row.get('grade_raw')!r} dir_in_row={row.get('direction')!r}"
            )
    total = sum(len(t.get("legs") or []) for g in groups for t in g.get("tickets") or [])
    print(f"\nTotal legs in JSON: {total}")
    print(
        "\nNote: Base rows come from SPORT_XLSX_CANDIDATES (pre-game step8 slates) for the slate --date."
        f"\n      Graded overlays: merge order {graded_merge_dates} (ui_runner/graded_slate/<d>/ then outputs/<d>/ per d)."
    )
    print("=== end debug ===\n")


def find_ticket_json(arg_date: str) -> Path | None:
    """Resolve ticket file from combined_slate_tickets_{date}.xlsx only."""
    px = REPO_ROOT / f"combined_slate_tickets_{arg_date}.xlsx"
    if px.is_file():
        return px
    # Daily pipeline writes combined tickets under outputs/YYYY-MM-DD/ (not always copied to root).
    out_dir = REPO_ROOT / "outputs" / arg_date
    p_out = out_dir / f"combined_slate_tickets_{arg_date}.xlsx"
    if p_out.is_file():
        return p_out
    p_out_strict = out_dir / f"combined_slate_tickets_{arg_date}.strict.xlsx"
    if p_out_strict.is_file():
        return p_out_strict
    return None


def _player_initials(name: str) -> str:
    parts = str(name or "").strip().split()
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:1].upper()
    return (parts[0][:1] + parts[-1][:1]).upper()


def resolve_col(df: pd.DataFrame, aliases: list[str]) -> str:
    norm = {str(c).strip().lower(): str(c) for c in df.columns}
    for a in aliases:
        key = str(a).strip().lower()
        if key in norm:
            return norm[key]
    raise KeyError(f"None of {aliases} found in columns: {list(df.columns)}")


def _clean_team_abbr(s: str) -> str:
    s = str(s or "").strip()
    if not s:
        return ""
    return re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()


def _parse_ticket_banner(s: str) -> tuple[float, float, int]:
    """Parse power/flex multipliers from combined ticket sheet banner row (col A)."""
    m_no = re.search(r"Ticket\s*#?\s*(\d+)", s, re.I)
    ticket_no = int(m_no.group(1)) if m_no else 1

    def _first_float(patterns: list[str]) -> float:
        for pat in patterns:
            m = re.search(pat, s, re.I)
            if not m:
                continue
            try:
                return float(m.group(1))
            except ValueError:
                continue
        return 0.0

    # Workbooks use ``×`` (U+00D7) or ``x``; some rows use ``PWR`` / ``FLEX`` instead of ``Power:`` / ``Flex:``.
    x_or_times = r"[x×]"
    power = _first_float(
        [
            rf"Power:\s*([\d.]+)\s*{x_or_times}?",
            rf"\bPWR\s*([\d.]+)\s*{x_or_times}?",
        ]
    )
    # Require a times marker for ``FLEX`` so we do not treat ``Flex 3-Leg`` as a multiplier.
    flex = _first_float(
        [
            rf"Flex:\s*([\d.]+)\s*{x_or_times}?",
            rf"\bFLEX\s*([\d.]+)\s*{x_or_times}",
        ]
    )
    return power, flex, ticket_no


def _ticket_header_colmap(row: tuple[Any, ...]) -> dict[int, str]:
    out: dict[int, str] = {}
    headers = [str(c or "").strip() for c in row]
    hdf = pd.DataFrame(columns=headers)
    header_idx = {str(c).strip().lower(): i for i, c in enumerate(headers)}

    required_aliases = {
        "pick_type": PICK_COL_ALIASES,
        "prop_type": PROP_COL_ALIASES,
        "line": LINE_COL_ALIASES,
        "direction": DIR_COL_ALIASES,
    }
    optional_aliases = {
        "tier": TIER_COL_ALIASES,
        "def_tier": DEF_COL_ALIASES,
    }

    for field_name, aliases in required_aliases.items():
        try:
            resolved = resolve_col(hdf, aliases)
            idx = header_idx.get(str(resolved).strip().lower())
            if idx is not None:
                out[idx] = field_name
        except KeyError:
            pass
    for field_name, aliases in optional_aliases.items():
        try:
            resolved = resolve_col(hdf, aliases)
            idx = header_idx.get(str(resolved).strip().lower())
            if idx is not None:
                out[idx] = field_name
        except KeyError:
            pass

    for i, cell in enumerate(row):
        key = _norm_header(cell)
        field = _XLSX_HDR_TO_LEG_FIELD.get(key)
        if field:
            out[i] = field
    return out


def _row_has_values(row: tuple[Any, ...]) -> bool:
    return any(str(c or "").strip() for c in row)


def _coerce_hit_rate_cell(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        f = float(v)
    else:
        s = str(v).strip().rstrip("%")
        try:
            f = float(s)
        except (TypeError, ValueError):
            return None
    if f > 1.0:
        f = f / 100.0
    return f


def _coerce_line_cell(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _coerce_edge_cell(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return 0.0


def _leg_from_xlsx_row(row: tuple[Any, ...], colmap: dict[int, str]) -> dict[str, Any] | None:
    leg: dict[str, Any] = {}
    for ci, field in colmap.items():
        if ci >= len(row):
            continue
        val = row[ci]
        if field == "player":
            leg["player"] = str(val or "").strip()
        elif field == "team":
            leg["team"] = _clean_team_abbr(str(val or ""))
        elif field == "opp":
            leg["opp"] = _clean_team_abbr(str(val or ""))
        elif field == "prop_type":
            leg["prop_type"] = str(val or "").strip()
        elif field == "pick_type":
            leg["pick_type"] = str(val or "").strip()
        elif field == "line":
            leg["line"] = _coerce_line_cell(val)
        elif field == "direction":
            leg["direction"] = str(val or "").strip().upper()
        elif field == "edge":
            leg["edge"] = _coerce_edge_cell(val)
        elif field == "hit_rate":
            leg["hit_rate"] = _coerce_hit_rate_cell(val)
        elif field == "l5_avg":
            x = _coerce_line_cell(val)
            leg["l5_avg"] = x
        elif field == "season_avg":
            x = _coerce_line_cell(val)
            leg["season_avg"] = x
        elif field == "sport":
            leg["sport"] = str(val or "").strip().upper()
        elif field == "game_time":
            leg["game_time"] = str(val or "").strip()
    if not leg.get("player"):
        return None
    leg["data_warning"] = "LIMITED_Q1_HISTORY" if str(leg.get("sport") or "").upper() == "NBA1Q" else None
    leg["initials"] = _player_initials(str(leg.get("player") or ""))
    return leg


def _skip_xlsx_ticket_sheet(sheet_name: str) -> bool:
    n = sheet_name.strip().lower()
    if n == "summary":
        return True
    if "slate" in n:
        return True
    return False


def _parse_ticket_sheet(ws: Any) -> list[dict[str, Any]]:
    current: dict[str, Any] | None = None
    colmap: dict[int, str] = {}
    expect_header = False
    tickets: list[dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        r0 = row[0] if row else None
        s0 = str(r0 or "").strip()
        s_lower = s0.lower()
        has_ticket_no = "ticket #" in s_lower
        has_payout_hint = bool(
            re.search(r"power:\s*[\d.]+", s0, re.I)
            or re.search(r"flex:\s*[\d.]+", s0, re.I)
            or re.search(r"\bpwr\s*[\d.]+", s0, re.I)
            or re.search(r"\bflex\s*[\d.]+\s*[x×]", s0, re.I)
        )
        is_banner = bool(s0 and has_ticket_no and has_payout_hint)
        if is_banner:
            if current is not None and current.get("legs"):
                tickets.append(current)
            pow_v, flex_v, tno = _parse_ticket_banner(s0)
            if pow_v <= 0 and flex_v <= 0:
                sh = str(getattr(ws, "title", "") or "?")
                _log.warning(
                    "[ticket_eval] Banner matched as ticket header but parsed power_payout=%s flex_payout=%s. "
                    "If PrizePicks changes labels (e.g. PP instead of PWR), update _parse_ticket_banner. "
                    "sheet=%r col_a=%r",
                    pow_v,
                    flex_v,
                    sh,
                    (s0[:240] + "…") if len(s0) > 240 else s0,
                )
            current = {
                "ticket_no": tno,
                "power_payout": pow_v,
                "flex_payout": flex_v,
                "legs": [],
            }
            expect_header = True
            continue

        if expect_header:
            colmap = _ticket_header_colmap(row)
            expect_header = False
            continue

        if current is None or not colmap:
            continue

        if not _row_has_values(row):
            continue

        leg = _leg_from_xlsx_row(row, colmap)
        if leg:
            current["legs"].append(leg)
            current["has_data_warning"] = bool(current.get("has_data_warning")) or bool(leg.get("data_warning"))

    if current is not None and current.get("legs"):
        tickets.append(current)

    return tickets


def _load_tickets_from_xlsx(path: Path, arg_date: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required to read combined_slate_tickets_*.xlsx; "
            "install with: pip install openpyxl"
        ) from e

    groups: list[dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if _skip_xlsx_ticket_sheet(sheet_name):
                continue
            ws = wb[sheet_name]
            tix = _parse_ticket_sheet(ws)
            if tix:
                groups.append({"group_name": sheet_name, "tickets": tix})
    finally:
        wb.close()

    return {"date": arg_date, "groups": groups}


def _load_tickets(path: Path, arg_date: str) -> dict[str, Any]:
    if path.suffix.lower() == ".xlsx":
        return _load_tickets_from_xlsx(path, arg_date)
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _group_is_allowed(group_name: str) -> bool:
    n = str(group_name or "").strip()
    m = re.match(
        r"^([A-Za-z0-9]+)\s+((?:Power Play|Flex|Standard|Pwr Std|Goblin)\s+\d-Leg)$",
        n,
        flags=re.I,
    )
    if not m:
        return False
    sport = m.group(1).strip().upper()
    return sport in ALLOWED_TICKET_SPORTS


def _leg_allowed_for_render(group_name: str, leg: dict[str, Any]) -> bool:
    return _leg_drop_reason(group_name, leg) is None


def _leg_drop_reason(group_name: str, leg: dict[str, Any]) -> str | None:
    prop = str(leg.get("prop_type") or "").strip().lower()
    direction = str(leg.get("direction") or "").strip().upper()
    line = leg.get("line")
    try:
        line_f = float(line) if line is not None else None
    except (TypeError, ValueError):
        line_f = None

    gl = str(group_name or "").strip().lower()
    is_power = ("power play" in gl) or ("pwr std" in gl) or ("goblin" in gl)

    if prop == "rebounds" and direction == "OVER" and (line_f is None or line_f < 2.5):
        return "rebounds_over_min_line"
    if prop == "points" and direction == "OVER" and (line_f is None or line_f < 8.0):
        return "points_over_min_line"
    if prop == "steals" and is_power:
        return "steals_not_allowed_in_power_play"
    return None


def _filter_payload_groups(payload: dict[str, Any], debug: bool = False) -> dict[str, Any]:
    out_groups: list[dict[str, Any]] = []
    for g in payload.get("groups") or []:
        gname = str(g.get("group_name") or "Group")
        if not _group_is_allowed(gname):
            continue
        gl = gname.strip().lower()
        min_legs = 0
        m_leg = re.search(r"\b(\d+)-leg\b", gl)
        if m_leg:
            try:
                min_legs = int(m_leg.group(1))
            except (TypeError, ValueError):
                min_legs = 0
        filtered_tickets: list[dict[str, Any]] = []
        for t in g.get("tickets") or []:
            legs: list[dict[str, Any]] = []
            for leg in (t.get("legs") or []):
                reason = _leg_drop_reason(gname, leg)
                if reason is not None:
                    if debug:
                        print(
                            f"[DEBUG] Drop leg in {gname}: "
                            f"{leg.get('player')} | {leg.get('prop_type')} {leg.get('line')} "
                            f"{leg.get('direction')} | reason={reason}",
                            flush=True,
                        )
                    continue
                legs.append(leg)
            if not legs:
                continue
            if min_legs and len(legs) < min_legs:
                print(
                    f"[WARN] Dropping partial ticket in {gname}: "
                    f"{len(legs)} legs < required {min_legs}",
                    flush=True,
                )
                continue
            t2 = dict(t)
            t2["legs"] = legs
            filtered_tickets.append(t2)
        if filtered_tickets:
            out_groups.append({"group_name": gname, "tickets": filtered_tickets})
    return {"date": payload.get("date"), "groups": out_groups}


def _fmt_num(x: Any) -> str:
    if x is None:
        return "—"
    if isinstance(x, (int, float)):
        if isinstance(x, float) and x == int(x):
            return str(int(x))
        return f"{x:g}"
    return html.escape(str(x))


def _sport_bucket_slate(sport: Any) -> str:
    """Lowercase sport key for slate_latest.json-style buckets (matches home /api/slate-sport)."""
    x = str(sport or "").strip().upper().replace(" ", "")
    aliases = {
        "NBA": "nba",
        "NBA1H": "nba1h",
        "NBA1Q": "nba1q",
        "CBB": "cbb",
        "WCBB": "wcbb",
        "NHL": "nhl",
        "SOCCER": "soccer",
        "MLB": "mlb",
    }
    return aliases.get(x, x.lower() if x else "unknown")


def _leg_to_slate_explorer_row(leg: dict[str, Any]) -> dict[str, Any]:
    """Map ticket leg (same shape as window.SLATE_DATA) to index.html slate table row."""
    edge = leg.get("edge")
    try:
        edge_f = float(edge) if edge is not None else None
    except (TypeError, ValueError):
        edge_f = None
    hr = leg.get("hit_rate")
    try:
        hr_f = float(hr) if hr is not None and hr != "" else None
    except (TypeError, ValueError):
        hr_f = None
    rs = leg.get("rank_score")
    try:
        rs_f = float(rs) if rs is not None else None
    except (TypeError, ValueError):
        rs_f = None
    if rs_f is None and edge_f is not None:
        rs_f = abs(edge_f)
    direction = str(leg.get("direction") or leg.get("dir") or "").strip().upper()
    return {
        "tier": leg.get("tier"),
        "rank_score": rs_f,
        "player": leg.get("player") or "",
        "team": leg.get("team") or "",
        "opp": leg.get("opp") or "",
        "prop": leg.get("prop_type") or leg.get("prop") or "",
        "pick_type": leg.get("pick_type") or "",
        "line": leg.get("line"),
        "dir": direction,
        "edge": edge_f,
        "hit_rate": hr_f,
        "l5_avg": leg.get("l5_avg"),
        "l5_over": leg.get("l5_over"),
        "l5_under": leg.get("l5_under"),
        "game_time": str(leg.get("game_time") or ""),
    }


def _manual_props_to_ticket_eval_slate_payload(
    manual_props: list[dict[str, Any]], slate_date: str
) -> dict[str, Any]:
    sports: dict[str, list[dict[str, Any]]] = {}
    for leg in manual_props:
        if not isinstance(leg, dict):
            continue
        key = _sport_bucket_slate(leg.get("sport"))
        sports.setdefault(key, []).append(_leg_to_slate_explorer_row(leg))
    return {
        "date": slate_date[:10],
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "source": "ticket_eval",
        "sports": sports,
    }


def write_ticket_eval_slate_json(manual_props: list[dict[str, Any]], slate_date: str) -> Path:
    """Same props embedded as window.SLATE_DATA — written for /api/slate-sport on the home UI."""
    payload = _manual_props_to_ticket_eval_slate_payload(manual_props, slate_date)
    TICKET_EVAL_SLATE_JSON.parent.mkdir(parents=True, exist_ok=True)
    TICKET_EVAL_SLATE_JSON.write_text(
        json.dumps(payload, ensure_ascii=False, separators=(",", ":"), default=str),
        encoding="utf-8",
    )
    return TICKET_EVAL_SLATE_JSON


def _ticket_grade_payout_html(oc: dict[str, Any], esc) -> str:
    """HTML block: predicted vs actual payout (under graded ticket legs)."""
    if oc.get("pending"):
        return (
            '<div class="ticket-grade-payout">'
            '<div class="grade-payout-pending">Payout model: awaiting all leg grades</div>'
            "</div>"
        )
    if oc.get("omit_payout_block"):
        r = str(oc.get("result") or "")
        css = str(oc.get("result_css") or "void")
        emoji = str(oc.get("result_emoji") or "")
        rd = esc(str(oc.get("result_display") or r))
        detail = esc(str(oc.get("result_detail") or r))
        return (
            '<div class="ticket-grade-payout">'
            f'<div class="grade-result grade-result-{esc(css)}">{esc(emoji)} {rd}</div>'
            f'<div class="grade-result-sub">{detail}</div>'
            '<div class="grade-entry-line">No payout table — slip did not resolve against the board.</div>'
            "</div>"
        )
    r = str(oc.get("result") or "")
    r_head = esc(str(oc.get("result_display") or r))
    css = str(oc.get("result_css") or "loss")
    emoji = str(oc.get("result_emoji") or "")
    pred_pay = oc.get("predicted_payout")
    pred_ev = oc.get("predicted_ev")
    pred_p = oc.get("predicted_p_win")
    rec = str(oc.get("recommendation_at_entry") or "")
    act = float(oc.get("actual_payout") or 0.0)
    gross = float(oc.get("entry_10_return") or 0.0)
    ev_line = (
        f"{float(pred_ev):.2f} — {esc(rec)}" if pred_ev is not None else f"N/A — {esc(rec) if rec else 'N/A'}"
    )
    pwin_pred = _fmt_pct_cell(pred_p)
    is_money_loss = r in ("LOSS", "VOID_LOSS")
    pwin_act = "❌" if is_money_loss else "✅"
    ev_act = "-$10.00 on $10" if is_money_loss else f"+${gross:.2f} on $10"
    entry_line = "Lost $10.00" if is_money_loss else f"Won ${gross:.2f}"
    detail = esc(str(oc.get("result_detail") or r))
    return (
        '<div class="ticket-grade-payout">'
        f'<div class="grade-result grade-result-{esc(css)}">{esc(emoji)} {r_head}</div>'
        '<div class="grade-result-sub">' + detail + "</div>"
        '<table class="grade-payout-table">'
        "<tr><th></th><th>Predicted</th><th>Actual</th></tr>"
        f"<tr><td>Payout</td><td>{_fmt_pay_cell(pred_pay)}</td><td>{_fmt_pay_cell(act)}</td></tr>"
        f"<tr><td>P(Win)</td><td>{esc(pwin_pred)}</td><td>{pwin_act}</td></tr>"
        f"<tr><td>EV at Entry</td><td>{ev_line}</td><td>{esc(ev_act)}</td></tr>"
        "</table>"
        f'<div class="grade-entry-line">$10 entry → {entry_line}</div>'
        "</div>"
    )


def _build_html(
    payload: dict[str, Any],
    arg_date: str,
    sport_candidates: dict[str, list[Path]],
    graded_merge_dates: list[str],
) -> tuple[str, dict[str, Any] | None]:
    groups = payload.get("groups") or []
    indices = _load_actuals_indices(sport_candidates, graded_merge_dates)

    all_legs: list[tuple[dict, dict | None, str]] = []
    tickets_flat: list[dict] = []

    for g in groups:
        gname = str(g.get("group_name") or "Group")
        for t in g.get("tickets") or []:
            t["_group_name"] = gname
            tickets_flat.append(t)
            for leg in t.get("legs") or []:
                row = _match_leg_to_row_multi(leg, indices)
                line = leg.get("line")
                try:
                    line_f = float(line) if line is not None else None
                except (TypeError, ValueError):
                    line_f = None
                direction = str(leg.get("direction") or "").strip().upper()
                actual = row["actual"] if row else None
                graw = row["grade_raw"] if row else ""
                if row and row.get("line") is not None and line_f is None:
                    line_f = row["line"]
                grade = _leg_grade(actual, line_f, direction, graw)
                all_legs.append((leg, row, grade))

    total_legs = len(all_legs)
    hits = sum(1 for _, _, g in all_legs if g == "HIT")
    misses = sum(1 for _, _, g in all_legs if g == "MISS")
    voids = sum(1 for _, _, g in all_legs if g == "VOID")
    ungraded = sum(1 for _, _, g in all_legs if g == "UNGRADED")

    # Manual Ticket Builder uses the per-leg model fields from tickets_latest.json.
    # We flatten all legs into a single array and embed it into the HTML as window.SLATE_DATA.
    manual_props: list[dict[str, Any]] = [leg for leg, _, _ in all_legs if isinstance(leg, dict)]
    slate_date_str = str(payload.get("date") or arg_date)[:10]
    try:
        te_path = write_ticket_eval_slate_json(manual_props, slate_date_str)
        print(f"  {te_path.name} -> home slate cards (/api/slate-sport when SLATE_SPORT_SOURCE=auto)")
    except OSError as e:
        print(f"  WARN: could not write ticket_eval_slate_latest.json: {e}")
    manual_props_json = json.dumps(manual_props, ensure_ascii=False)
    # Avoid prematurely terminating the <script> tag if any string contains "</".
    manual_props_json = manual_props_json.replace("</", "<\\/")

    outcome_map: dict[tuple[str, Any], dict[str, Any]] = {}
    perfect = 0
    money_wins = 0
    money_losses = 0
    for t in tickets_flat:
        gname = str(t.get("_group_name") or "Group")
        legs = t.get("legs") or []
        gs: list[str] = []
        for leg in legs:
            row = _match_leg_to_row_multi(leg, indices)
            try:
                lf = float(leg.get("line"))
            except (TypeError, ValueError):
                lf = None
            d = str(leg.get("direction") or "").strip().upper()
            act = row["actual"] if row else None
            gr = row["grade_raw"] if row else ""
            if row and row.get("line") is not None and lf is None:
                lf = row["line"]
            g = _leg_grade(act, lf, d, gr)
            gs.append(g)
        tno = t.get("ticket_no", "?")
        t["_leg_grades_cache"] = gs
        if not gs:
            outcome_map[(gname, tno)] = {"pending": True}
            continue
        outcome_map[(gname, tno)] = _ticket_eval_money_outcome(gname, gs, t)
        if all(x == "HIT" for x in gs):
            perfect += 1
        if any(x == "UNGRADED" for x in gs):
            continue
        if all(x == "VOID" for x in gs):
            continue
        if _ticket_pays_money(gname, gs):
            money_wins += 1
        else:
            money_losses += 1

    ticket_decided = money_wins + money_losses
    ticket_pct = (100.0 * money_wins / ticket_decided) if ticket_decided else 0.0

    pay_summary_rows: list[dict[str, Any]] = []
    for t in tickets_flat:
        gs = t.get("_leg_grades_cache") or []
        if not gs or any(x == "UNGRADED" for x in gs) or all(x == "VOID" for x in gs):
            continue
        gname = str(t.get("_group_name") or "Group")
        oc = outcome_map.get((gname, t.get("ticket_no", "?")), {})
        if oc.get("pending"):
            continue
        pay_summary_rows.append(oc)

    n_pay = len(pay_summary_rows)
    wins_ct = sum(1 for oc in pay_summary_rows if oc.get("result") in ("WIN", "SWEEP"))
    guar_ct = sum(1 for oc in pay_summary_rows if oc.get("result") == "MIN GUARANTEE")
    loss_ct = sum(1 for oc in pay_summary_rows if oc.get("result") in ("LOSS", "VOID_LOSS"))
    total_net_10 = sum(float(oc.get("net_10") or 0.0) for oc in pay_summary_rows)
    evs = [float(oc["predicted_ev"]) for oc in pay_summary_rows if oc.get("predicted_ev") is not None]
    avg_ev = sum(evs) / len(evs) if evs else None

    def _rec_bucket(rec: str) -> str:
        u = (rec or "").strip().upper()
        if u == "STRONG":
            return "STRONG"
        if u == "OK":
            return "OK"
        if u == "MARGINAL":
            return "MARGINAL"
        if u == "SKIP":
            return "SKIP"
        return "SKIP"

    buck: dict[str, dict[str, int]] = {
        "STRONG": {"count": 0, "wins": 0},
        "OK": {"count": 0, "wins": 0},
        "MARGINAL": {"count": 0, "wins": 0},
        "SKIP": {"count": 0, "wins": 0},
    }
    for oc in pay_summary_rows:
        bk = _rec_bucket(str(oc.get("recommendation_at_entry") or ""))
        buck[bk]["count"] += 1
        if oc.get("result") not in ("LOSS", "VOID_LOSS"):
            buck[bk]["wins"] += 1

    win_rate_pay = (wins_ct + guar_ct) / n_pay if n_pay else 0.0
    net_per = total_net_10 / n_pay if n_pay else 0.0
    roi_pct = (100.0 * total_net_10 / (10 * n_pay)) if n_pay else 0.0

    history_record: dict[str, Any] | None = None
    if n_pay:
        history_record = {
            "date": str(payload.get("date") or arg_date)[:10],
            "n_tickets": n_pay,
            "wins": wins_ct,
            "guarantees": guar_ct,
            "losses": loss_ct,
            "win_rate": round(win_rate_pay, 4),
            "avg_ev_predicted": round(avg_ev, 4) if avg_ev is not None else None,
            "net_per_10": round(net_per, 2),
            "roi_pct": round(roi_pct, 2),
            "strong_tickets": dict(buck["STRONG"]),
            "ok_tickets": dict(buck["OK"]),
            "marginal_tickets": dict(buck["MARGINAL"]),
            "skip_tickets": dict(buck["SKIP"]),
        }
    else:
        history_record = None

    # ── HTML
    esc = html.escape
    json_date = esc(str(payload.get("date") or arg_date))
    ungraded_note_html = (
        f" <strong>UNGRADED ({ungraded} legs):</strong> the matched graded row still has no actual or HIT/MISS—often unfinished games, voids, "
        f"or grading before final stats. Re-run the grader after games; on hosts without <code>outputs/{json_date}/</code>, "
        f"commit workbooks under <code>ui_runner/graded_slate/{json_date}/</code> before rebuilding this page. "
        if ungraded > 0
        else ""
    )

    if n_pay:
        total_staked = 10 * n_pay
        win_rate_pct = 100.0 * win_rate_pay
        avg_ev_s = f"{avg_ev:.2f}" if avg_ev is not None else "N/A"
        net_abs = abs(total_net_10)
        net_word = "profit" if total_net_10 >= 0 else "loss"
        net_sign = "+" if total_net_10 >= 0 else "−"
        grade_eval_summary_html = (
            '<div class="grade-eval-summary">'
            f'<div class="grade-eval-summary-line1">Date: {json_date} · {n_pay} tickets graded</div>'
            '<div class="grade-eval-summary-line2">'
            f'<span>✅ Wins: {wins_ct}</span>'
            f'<span>🛡️ Guarantees: {guar_ct}</span>'
            f'<span>❌ Losses: {loss_ct}</span>'
            "</div>"
            '<div class="grade-eval-summary-line3">'
            f"Win rate (W+🛡️): {win_rate_pct:.0f}%"
            f" · Avg EV at entry: {avg_ev_s}"
            "</div>"
            f'<div class="grade-eval-summary-line4">Total: {net_sign}${net_abs:.2f} net on ${total_staked:.0f} staked '
            f"(${10}/ticket flat) — {net_word}</div>"
            "</div>"
        )
    else:
        grade_eval_summary_html = (
            '<div class="grade-eval-summary grade-eval-summary-empty">'
            f'<div>Date: {json_date} · No fully graded tickets for payout summary (ungraded, all-void, or none).</div>'
            "</div>"
        )

    sport_colors_css = """
.sport-nba{background:rgba(212,160,23,.12);color:#f0a500;border:1px solid rgba(212,160,23,.35);}
.sport-nba1h{background:rgba(255,155,86,.12);color:#ffb27d;border:1px solid rgba(255,155,86,.32);}
.sport-nba1q{background:rgba(255,214,102,.12);color:#ffd87a;border:1px solid rgba(255,214,102,.32);}
.sport-cbb{background:rgba(0,229,255,.10);color:#00e5ff;border:1px solid rgba(0,229,255,.32);}
.sport-wcbb{background:rgba(127,199,217,.10);color:#9fd8e8;border:1px solid rgba(127,199,217,.32);}
.sport-nhl{background:rgba(186,130,255,.12);color:#c4a5ff;border:1px solid rgba(186,130,255,.38);}
.sport-soccer{background:rgba(240,165,0,.10);color:#e8b84a;border:1px solid rgba(240,165,0,.34);}
.sport-mlb{background:rgba(255,121,121,.12);color:#ff9a9a;border:1px solid rgba(255,121,121,.32);}
.sport-default{background:rgba(255,255,255,.04);color:#888;border:1px solid rgba(255,255,255,.1);}
"""

    parts: list[str] = [
        "<!DOCTYPE html>",
        '<html lang="en" data-theme="dark">',
        "<head>",
        '<meta charset="UTF-8"/>',
        '<meta name="viewport" content="width=device-width, initial-scale=1.0, maximum-scale=1.0, viewport-fit=cover"/>',
        "<script>(function(){try{var t=localStorage.getItem('proporacle-theme')||localStorage.getItem('theme');"
        "if(t==='light'){document.documentElement.setAttribute('data-theme','light');"
        "document.documentElement.classList.add('light-theme');}}catch(e){}})();</script>",
        '<meta name="theme-color" content="#050505"/>',
        f"<title>Ticket Eval — {json_date}</title>",
        '<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&family=Inter:wght@400;500;600;700;800&display=swap" rel="stylesheet"/>',
        f'<link rel="stylesheet" href="/static/proporacle-page-shell.css?v={_TICKET_EVAL_SHELL_CSS_VER}"/>',
        f'<link rel="stylesheet" href="/static/site-nav-unified.css?v={_TICKET_EVAL_NAV_UNIFIED_VER}"/>',
        f'<link rel="stylesheet" href="/static/nav-mobile-shared.css?v={_TICKET_EVAL_NAV_MOBILE_VER}"/>',
        f'<link rel="stylesheet" href="/static/mobile-content-width.css?v={_TICKET_EVAL_MOBILE_WIDTH_VER}"/>',
        f'<link rel="stylesheet" href="/static/site-nav-datetime.css?v={_TICKET_EVAL_NAV_DATETIME_VER}"/>',
        "<style>",
        _TICKET_EVAL_PAGE_WRAP_CSS,
        "h1,h2,h3,h4,h5,h6{font-family:'Bebas Neue',sans-serif;letter-spacing:3px;}",
        ".bebas{font-family:'Bebas Neue',sans-serif;letter-spacing:3px;}",
        ".stats-bar{position:-webkit-sticky;position:sticky;top:72px;z-index:175;margin:0 auto 8px;width:100%;max-width:min(1520px,96vw);"
        "padding:14px clamp(16px,2.5vw,28px);transition:top .28s ease,box-shadow .22s ease;"
        "background:linear-gradient(180deg,rgba(10,10,18,.92),rgba(8,8,14,.88)),var(--glass);"
        "backdrop-filter:blur(24px) saturate(180%);-webkit-backdrop-filter:blur(24px) saturate(180%);"
        "border:1px solid var(--glass-bd);border-radius:18px;box-shadow:0 8px 32px rgba(0,0,0,.35);}"
        "body.ticket-eval-page .grade-eval-summary:has(+ .stats-bar){margin-bottom:0;padding-bottom:8px;border-bottom:none;"
        "border-radius:16px 16px 0 0;box-shadow:none;}"
        "body.ticket-eval-page .grade-eval-summary:has(+ .stats-bar)+.stats-bar{margin-top:-1px;border-top:none;"
        "border-radius:0 0 18px 18px;padding-top:11px;box-shadow:0 8px 32px rgba(0,0,0,.35);}",
        ".sum-row{display:flex;flex-wrap:wrap;gap:18px 36px;align-items:center;justify-content:center;}",
        ".sum-item{display:flex;flex-direction:column;align-items:center;gap:4px;min-width:88px;}",
        ".sum-val{font-family:'Inter',sans-serif;font-size:clamp(22px,2.6vw,30px);font-weight:700;color:var(--gold);text-shadow:0 0 20px rgba(240,165,0,.25);}",
        ".sum-val.green{color:var(--green);text-shadow:0 0 14px rgba(57,255,110,.35);}",
        ".sum-val.red{color:var(--red);text-shadow:0 0 14px rgba(255,77,77,.35);}",
        ".sum-val.pend{color:var(--pending);text-shadow:none;}",
        ".sum-val.void{color:var(--gold2);text-shadow:none;}",
        ".sum-val-sm{font-size:clamp(18px,2.1vw,24px)!important;}",
        ".sum-lab{font-family:'Bebas Neue',sans-serif;font-size:11px;letter-spacing:2.2px;color:var(--muted);text-align:center;line-height:1.2;max-width:11em;}",
        ".wrap.ticket-eval-main{width:100%;max-width:min(1520px,96vw);margin:0 auto;padding:2px clamp(14px,2.5vw,32px) 0;}",
        ".ticket-sections-wrap{padding-top:2px;}",
        "details.ticket-bucket{margin:0 0 10px;}",
        ".ticket-sections-wrap > details.ticket-bucket:first-child{margin-top:0;}",
        ".sec{margin-top:22px;}",
        ".ticket-sections-wrap > section.sec:first-of-type,.ticket-bucket-body > section.sec:first-of-type{margin-top:8px;}",
        ".sec-head{font-family:'Bebas Neue',sans-serif;font-size:clamp(30px,3.2vw,40px);color:var(--gold);margin-bottom:8px;padding-bottom:14px;"
        "border-bottom:1px solid var(--glass-bd);letter-spacing:3px;text-shadow:0 0 24px rgba(240,165,0,.2);}",
        ".ticket-card{background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);"
        "border:1px solid var(--glass-bd);border-radius:14px;margin-bottom:22px;overflow:hidden;"
        "box-shadow:0 8px 32px rgba(0,0,0,.35);}",
        ".ticket-card.card-warning{border-color:rgba(240,165,0,.55);box-shadow:0 0 0 1px rgba(240,165,0,.22),0 8px 32px rgba(0,0,0,.35);}",
        ".ticket-warning-note{font-family:'Inter',sans-serif;font-size:11px;letter-spacing:.4px;color:#ffd87a;"
        "background:rgba(240,165,0,.08);border-top:1px solid rgba(240,165,0,.3);padding:8px 14px;}",
        ".ticket-card.all-hit{background:rgba(57,255,110,0.06);border-color:rgba(57,255,110,.42);"
        "box-shadow:0 0 28px rgba(57,255,110,.14),0 8px 32px rgba(0,0,0,.3);}",
        ".ticket-card.card-missed{background:rgba(255,77,77,0.06);border:1px solid rgba(255,77,77,0.35);"
        "box-shadow:0 0 24px rgba(255,77,77,0.12),0 0 1px rgba(255,77,77,0.4),0 8px 32px rgba(0,0,0,.28);position:relative;}",
        ".ticket-card.card-missed::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:#ff4d4d;"
        "box-shadow:0 0 14px rgba(255,77,77,0.45);z-index:2;border-radius:14px 14px 0 0;pointer-events:none;}",
        ".thdr{display:flex;flex-wrap:wrap;gap:12px 20px;align-items:center;padding:18px clamp(14px,2vw,24px);border-bottom:1px solid var(--glass-bd);"
        "background:rgba(0,0,0,.18);backdrop-filter:blur(12px);}",
        ".thdr .tn{font-size:clamp(24px,2.8vw,32px);font-family:'Bebas Neue',sans-serif;letter-spacing:2px;color:var(--gold);}",
        ".thdr .tg{font-family:'Inter',sans-serif;font-size:clamp(12px,1.35vw,15px);color:var(--muted);letter-spacing:0.5px;line-height:1.35;}",
        ".payout{font-family:'Inter',sans-serif;font-size:clamp(13px,1.4vw,16px);color:var(--cyan);}",
        ".banner{font-family:'Bebas Neue',sans-serif;font-size:clamp(11px,1.2vw,13px);letter-spacing:2px;padding:8px 18px;border-radius:999px;font-weight:700;"
        "background:rgba(255,255,255,0.04);backdrop-filter:blur(20px);border:1px solid var(--glass-bd);}",
        ".banner.hit{color:var(--green);border-color:rgba(57,255,110,.45);box-shadow:0 0 16px rgba(57,255,110,.15);}",
        ".banner.miss{color:var(--red);border-color:rgba(255,77,77,.5);box-shadow:0 0 16px rgba(255,77,77,.12);}",
        ".banner.pend{color:var(--pending);border-color:rgba(255,255,255,.12);}",
        ".banner.void{color:var(--gold2);border-color:rgba(240,165,0,.35);}",
        ".legrow{font-family:'Inter',sans-serif;display:grid;"
        "grid-template-columns:56px 92px minmax(120px,1fr) 44px minmax(240px,1.45fr) minmax(108px,1fr) minmax(96px,1fr) minmax(76px,.85fr);gap:12px;"
        "align-items:center;padding:14px clamp(14px,2vw,22px);font-size:clamp(13px,1.45vw,16px);line-height:1.35;"
        "border-bottom:1px solid rgba(255,255,255,.06);border-left:3px solid transparent;}",
        ".leg-head{font-family:'Bebas Neue',sans-serif;display:grid;"
        "grid-template-columns:56px 92px minmax(120px,1fr) 44px minmax(240px,1.45fr) minmax(108px,1fr) minmax(96px,1fr) minmax(76px,.85fr);gap:12px;"
        "align-items:end;padding:6px clamp(14px,2vw,22px) 8px;font-size:10px;letter-spacing:2px;color:var(--muted);"
        "border-bottom:1px solid rgba(255,255,255,.1);background:rgba(0,0,0,.12);}",
        ".leg-head-lab{line-height:1.2;}",
        ".legrow:last-child{border-bottom:none;}",
        ".legrow.leg-hit{background:rgba(57,255,110,0.04);border-left-color:var(--green);}",
        ".legrow.leg-miss{background:rgba(255,77,77,0.10);border-left:4px solid #ff4d4d;"
        "box-shadow:0 0 0 1px rgba(255,77,77,0.4),inset 0 0 20px rgba(255,77,77,0.06);}",
        ".legrow.leg-miss .pl-miss{color:#ff4d4d;font-weight:700;text-shadow:0 0 8px rgba(255,77,77,0.35);}",
        ".legrow.leg-miss .pl-line{display:flex;align-items:center;flex-wrap:wrap;gap:8px;}",
        ".miss-tag{font-family:'Bebas Neue',sans-serif;display:inline-flex;align-items:center;"
        "background:rgba(255,77,77,0.15);border:1px solid #ff4d4d;color:#ff4d4d;font-size:9px;"
        "letter-spacing:2px;padding:2px 8px;border-radius:20px;line-height:1;vertical-align:middle;}",
        ".legrow.leg-miss .badge.miss{width:44px;height:44px;min-width:44px;border-radius:12px;display:flex;align-items:center;"
        "justify-content:center;font-size:clamp(22px,2.5vw,28px);line-height:1;background:rgba(255,77,77,0.25);"
        "border:2px solid #ff4d4d;box-shadow:0 0 12px rgba(255,77,77,0.6);color:#ff4d4d;text-shadow:none;}",
        ".legrow.leg-miss .leg-extra.val-miss{color:#ff4d4d;font-weight:700;}",
        ".legrow.leg-miss .miss-leg-cell{color:#ff5c5c!important;font-weight:700;}",
        ".legrow.leg-miss .leg-prop-col.miss-leg-cell > div:first-child{color:#ff7a7a!important;font-weight:800;}",
        ".legrow.leg-miss .leg-prop-col .meta-muted{color:rgba(255,170,170,.95)!important;font-weight:600;}",
        ".legrow.leg-miss .miss-leg-cell .dir-over,.legrow.leg-miss .miss-leg-cell .dir-under{color:#ffc9c9!important;font-weight:800;}",
        ".legrow.leg-miss > div:nth-child(2) .pill{box-shadow:0 0 0 1px rgba(255,90,90,.55),0 0 12px rgba(255,60,60,.2);}",
        ".legrow.leg-pend{background:transparent;border-left-color:transparent;}",
        ".legrow.leg-pend .pl-pend,.legrow.leg-pend .meta-muted{color:var(--pending)!important;}",
        ".legrow.leg-pend .pill{background:rgba(255,255,255,0.04)!important;border-color:rgba(255,255,255,0.1)!important;color:var(--pending)!important;}",
        ".legrow.leg-void{background:rgba(240,165,0,0.04);border-left-color:rgba(240,165,0,.55);}",
        ".legrow.leg-void .pl-void,.legrow.leg-void .meta-muted{color:var(--gold2)!important;}",
        ".legrow.leg-void .pill{background:rgba(240,165,0,0.07)!important;border-color:rgba(240,165,0,0.28)!important;color:var(--gold2)!important;}",
        ".badge{font-size:clamp(28px,3.2vw,36px);line-height:1;text-align:center;}",
        ".badge.hit{color:var(--green);text-shadow:0 0 14px rgba(57,255,110,.6);}",
        ".badge.miss{color:var(--red);text-shadow:0 0 14px rgba(255,77,77,.55);}",
        ".badge.pend{color:var(--pending);text-shadow:none;}",
        ".badge.void{color:var(--gold2);text-shadow:none;}",
        ".pill{font-family:'Bebas Neue',sans-serif;font-size:clamp(10px,1.1vw,12px);letter-spacing:1.2px;padding:5px 12px;border-radius:999px;text-transform:uppercase;}",
        sport_colors_css,
        ".tier{font-family:'Bebas Neue',sans-serif;width:32px;height:32px;border-radius:10px;display:flex;align-items:center;justify-content:center;"
        "font-weight:800;font-size:clamp(13px,1.4vw,15px);letter-spacing:0;background:rgba(255,255,255,0.05);color:var(--gold);"
        "border:1px solid var(--glass-bd);backdrop-filter:blur(12px);box-shadow:inset 0 1px 0 rgba(255,255,255,.06);}",
        ".pl-hit{color:var(--green);text-shadow:0 0 8px rgba(57,255,110,.4);}",
        ".pl-miss{color:var(--red);}",
        ".pl-pend{color:var(--pending);}",
        ".pl-void{color:var(--gold2);}",
        ".dir-over{color:var(--cyan);font-weight:700;}",
        ".dir-under{color:var(--gold);font-weight:700;}",
        ".meta-muted{font-family:'Inter',sans-serif;color:var(--muted);font-size:clamp(11px,1.2vw,13px);margin-top:3px;}",
        ".slate-kicker{font-family:'Inter',sans-serif;font-size:clamp(11px,1.2vw,13px);letter-spacing:3px;color:var(--muted);margin:0 0 6px;}",
        ".pl-hit,.pl-pend{font-size:1em;font-weight:600;}",
        ".warning-chip{display:inline-flex;align-items:center;margin-left:8px;padding:2px 8px;border-radius:999px;"
        "border:1px solid rgba(240,165,0,.4);background:rgba(240,165,0,.12);color:#ffd87a;"
        "font-family:'Inter',sans-serif;font-size:10px;letter-spacing:.6px;cursor:help;}",
        ".grade-eval-summary{max-width:min(1520px,96vw);margin:0 auto 8px;padding:10px 16px;border-radius:16px;"
        "border:1px solid var(--glass-bd);background:rgba(0,0,0,.22);font-family:'Inter',sans-serif;font-size:13px;"
        "line-height:1.55;color:var(--text);}",
        ".grade-eval-summary-empty{color:var(--muted);font-size:12px;}",
        ".grade-eval-summary-line1{font-weight:700;color:var(--gold);margin-bottom:6px;}",
        ".grade-eval-summary-line2{display:flex;flex-wrap:wrap;gap:12px 22px;margin-bottom:4px;}",
        ".grade-eval-summary-line3{color:var(--muted);margin-bottom:4px;}",
        ".grade-eval-summary-line4{color:var(--cyan);font-weight:600;}",
        ".ticket-grade-payout{margin:0;padding:14px clamp(14px,2vw,22px) 16px;border-top:1px solid var(--glass-bd);"
        "background:rgba(0,0,0,.12);}",
        ".grade-result{font-family:'Bebas Neue',sans-serif;letter-spacing:2px;font-size:clamp(16px,1.8vw,20px);"
        "padding:10px 14px;border-radius:10px;margin-bottom:6px;display:inline-block;}",
        ".grade-result-sub{font-size:11px;color:var(--muted);margin:-2px 0 10px;font-family:'Inter',sans-serif;}",
        ".grade-result-sweep{background:gold;color:#000;}",
        ".grade-result-win{background:#00ff88;color:#000;}",
        ".grade-result-min_guarantee{background:#88ccff;color:#000;}",
        ".grade-result-loss{background:#ff4444;color:#fff;}",
        ".grade-result-void{background:rgba(140,140,150,.35);color:#f2f2f4;border:1px solid rgba(255,255,255,.12);}",
        ".grade-result-void_loss{background:rgba(240,165,0,.2);color:#ffe6a8;border:1px solid rgba(240,165,0,.45);}",
        ".grade-payout-table{width:100%;border-collapse:collapse;font-size:clamp(12px,1.25vw,14px);margin-bottom:10px;}",
        ".grade-payout-table th,.grade-payout-table td{padding:6px 8px;border-bottom:1px solid rgba(255,255,255,.08);text-align:left;}",
        ".grade-payout-table th{color:var(--muted);font-weight:600;font-size:11px;letter-spacing:1px;}",
        ".grade-entry-line{font-size:12px;color:var(--gold2);font-family:'Inter',sans-serif;}",
        ".grade-payout-pending{font-size:12px;color:var(--pending);padding:8px 0;}",
        ".grade-ticket-result{font-family:'Inter',sans-serif;font-size:clamp(11px,1.2vw,13px);letter-spacing:.3px;}",
        ".grade-ticket-result.won{color:var(--green);}",
        ".grade-ticket-result.lost{color:var(--red);}",
        ".grade-ticket-result.void-slip{color:var(--gold2);}",
        ".grade-ticket-result.void-loss{color:#ffd87a;}",
        ".grade-ticket-result-label{opacity:.85;font-weight:600;}",
        "html.ticket-eval-embed-grades body{padding-top:0!important;padding-bottom:36px!important;}"
        "html.ticket-eval-embed-grades .snav,html.ticket-eval-embed-grades #mobile-menu{display:none!important;}"
        "html.ticket-eval-embed-grades body > .grade-eval-summary{margin-top:0!important;}"
        "html.ticket-eval-embed-grades .stats-bar{top:0!important;}"
        "html.ticket-eval-embed-grades .grade-eval-summary:has(+ .stats-bar)+.stats-bar{top:0!important;margin-top:-1px!important;border-top:none!important;padding-top:9px!important;}",
        "</style>",
        '<script>try{if(window.self!==window.top)document.documentElement.classList.add("ticket-eval-embed-grades");}catch(e){}</script>',
        "</head>",
        '<body class="ticket-eval-page">',
        _render_site_nav_grades_active(),
        grade_eval_summary_html,
        '<div class="stats-bar">',
        '<div class="sum-row">',
        f'<div class="sum-item"><div class="sum-val">{ticket_pct:.1f}%</div><div class="sum-lab">TICKET HIT RATE</div></div>',
        f'<div class="sum-item"><div class="sum-val green">{hits}</div><div class="sum-lab">HITS</div></div>',
        f'<div class="sum-item"><div class="sum-val red">{misses}</div><div class="sum-lab">MISSES</div></div>',
        f'<div class="sum-item"><div class="sum-val void">{voids}</div><div class="sum-lab">VOID/PUSH</div></div>',
        f'<div class="sum-item"><div class="sum-val pend">{ungraded}</div><div class="sum-lab">UNGRADED</div></div>',
        f'<div class="sum-item"><div class="sum-val">{perfect}</div><div class="sum-lab">PERFECT TICKETS</div></div>',
        f'<div class="sum-item"><div class="sum-val green">{money_wins}</div><div class="sum-lab">PAID TIX</div></div>',
        f'<div class="sum-item"><div class="sum-val red">{money_losses}</div><div class="sum-lab">NO PAYOUT</div></div>',
        f'<div class="sum-item"><div class="sum-val sum-val-sm">{total_net_10:+.2f}</div><div class="sum-lab">NET @ $10/TKT</div></div>'
        if n_pay
        else '<div class="sum-item"><div class="sum-val sum-val-sm pend">—</div><div class="sum-lab">NET @ $10/TKT</div></div>',
        f'<div class="sum-item"><div class="sum-val sum-val-sm">{roi_pct:.1f}%</div><div class="sum-lab">ROI (FLAT $10)</div></div>'
        if n_pay
        else '<div class="sum-item"><div class="sum-val sum-val-sm pend">—</div><div class="sum-lab">ROI (FLAT $10)</div></div>',
        f'<div class="sum-item"><div class="sum-val sum-val-sm">{total_legs}</div><div class="sum-lab">TOTAL LEGS</div></div>',
        "</div></div>",
        '<div class="wrap ticket-eval-main">',
        f'<p class="slate-kicker">SLATE DATE · {json_date}</p>',
        '<p class="meta-muted" style="margin:4px 0 8px;line-height:1.5">'
        "Each leg: <strong>Line</strong> + side · <strong>Actual</strong> (box-score stat; — if none exists yet, e.g. rainout/postponed) · "
        f"<strong>Edge</strong> (model edge, not the result). Graded exports: <code>outputs/{json_date}/graded_*.xlsx</code>. "
        "<strong>Ticket hit rate</strong> = paid ÷ (paid + no payout) among fully graded tickets (all-void slips excluded). "
        "Sheets with <strong>Flex</strong> in the title (3+ legs) use flex cash rules: at least n−1 hits and at most one miss."
        f"{ungraded_note_html}</p>",
    ]
    parts.append('<div class="ticket-sections-wrap">')

    bucketed = _bucket_ticket_groups(groups)
    use_buckets = len(bucketed) > 1

    for bi, (bkey, grplist) in enumerate(bucketed):
        ntix = sum(len(x.get("tickets") or []) for x in grplist)
        nsec = len(grplist)
        if use_buckets:
            skin = _ticket_bucket_skin_class(bkey)
            open_attr = " open" if bi == 0 else ""
            parts.append(f'<details class="ticket-bucket {skin}"{open_attr}>')
            parts.append(
                f'<summary><span>{esc(bkey)}</span>'
                f'<span class="ticket-bucket-meta">{nsec} sections · {ntix} tickets</span></summary>'
            )
            parts.append('<div class="ticket-bucket-body">')

        for g in grplist:
            gname = str(g.get("group_name") or "Group")
            is_nba1q_group = gname.strip().upper().startswith("NBA1Q ")
            parts.append(f'<section class="sec"><h2 class="sec-head bebas">{esc(gname)}</h2>')
            for t in g.get("tickets") or []:
                tno = t.get("ticket_no", "?")
                pp = t.get("power_payout")
                fp = t.get("flex_payout")
                legs = t.get("legs") or []
                leg_grades = list(t.get("_leg_grades_cache") or [])
                oc = outcome_map.get((gname, tno), {})

                h = leg_grades.count("HIT")
                m = leg_grades.count("MISS")
                pnd = leg_grades.count("UNGRADED")
                vct = leg_grades.count("VOID")
                n = len(leg_grades)

                if pnd > 0:
                    banner_cls, banner_txt = "pend", "UNGRADED"
                elif vct > 0 and m == 0:
                    banner_cls, banner_txt = "void", "VOID/PUSH"
                elif m == 0 and n > 0:
                    banner_cls, banner_txt = "hit", "ALL HIT"
                else:
                    banner_cls, banner_txt = "miss", f"MISSED {m}"

                card_cls = "ticket-card"
                if banner_txt == "ALL HIT":
                    card_cls += " all-hit"
                elif banner_cls == "miss":
                    card_cls += " card-missed"
                has_data_warning = bool(t.get("has_data_warning")) or any(
                    bool((leg or {}).get("data_warning")) for leg in legs
                )
                if is_nba1q_group:
                    has_data_warning = True
                if has_data_warning:
                    card_cls += " card-warning"

                parts.append(f'<article class="{card_cls}">')
                parts.append('<div class="thdr">')
                parts.append(f'<span class="tn bebas">#{esc(str(tno))}</span>')
                parts.append(f'<span class="tg">{esc(gname)}</span>')
                parts.append(f'<span class="tg">{h}✓ {m}✗ / {n}</span>')
                parts.append(f'<span class="payout">PWR {_fmt_num(pp)}× · FLEX {_fmt_num(fp)}×</span>')
                if not oc.get("pending") and oc.get("result"):
                    rtxt = str(oc.get("result") or "")
                    rdisp = str(oc.get("result_display") or rtxt)
                    rem = str(oc.get("result_emoji") or "")
                    if rtxt == "VOID":
                        res_cls = "tg grade-ticket-result void-slip"
                        parts.append(
                            f'<span class="{res_cls}">RESULT: {esc(rem)} NO CONTEST'
                            f' <span class="grade-ticket-result-label">({esc(rdisp)})</span></span>'
                        )
                    elif rtxt == "VOID_LOSS":
                        res_cls = "tg grade-ticket-result void-loss"
                        parts.append(
                            f'<span class="{res_cls}">RESULT: {esc(rem)} VOID / NO ACTION'
                            f' <span class="grade-ticket-result-label">(VOID_LOSS)</span></span>'
                        )
                    else:
                        won = rtxt != "LOSS"
                        res_cls = "tg grade-ticket-result won" if won else "tg grade-ticket-result lost"
                        parts.append(
                            f'<span class="{res_cls}">RESULT: {esc(rem)} '
                            f'{"✅ WON" if won else "❌ LOSS"}'
                            f' <span class="grade-ticket-result-label">({esc(rdisp)})</span></span>'
                        )
                parts.append(f'<span class="banner {banner_cls}">{esc(banner_txt)}</span>')
                parts.append("</div>")
                parts.append(
                    '<div class="leg-head" aria-hidden="true">'
                    "<div></div><div></div>"
                    '<div class="leg-head-lab">PLAYER</div>'
                    '<div class="leg-head-lab">T</div>'
                    '<div class="leg-head-lab">PROP</div>'
                    '<div class="leg-head-lab">LINE</div>'
                    '<div class="leg-head-lab">ACTUAL</div>'
                    '<div class="leg-head-lab">EDGE</div>'
                    "</div>"
                )

                for leg, lg in zip(legs, leg_grades):
                    row = _match_leg_to_row_multi(leg, indices)
                    try:
                        lf = float(leg.get("line"))
                    except (TypeError, ValueError):
                        lf = None
                    d = str(leg.get("direction") or "").strip().upper()
                    act = row["actual"] if row else None
                    gr = row["grade_raw"] if row else ""
                    if row and row.get("line") is not None and lf is None:
                        lf = row["line"]

                    if lg == "HIT":
                        bcls, plcls = "hit", "pl-hit"
                    elif lg == "MISS":
                        bcls, plcls = "miss", "pl-miss"
                    elif lg == "VOID":
                        bcls, plcls = "void", "pl-void"
                    else:
                        bcls, plcls = "pend", "pl-pend"

                    sk = _sport_key(str(leg.get("sport") or ""))
                    sp_class = {
                        "NBA": "sport-nba",
                        "NBA1H": "sport-nba1h",
                        "NBA1Q": "sport-nba1q",
                        "CBB": "sport-cbb",
                        "WCBB": "sport-wcbb",
                        "NHL": "sport-nhl",
                        "SOCCER": "sport-soccer",
                        "MLB": "sport-mlb",
                    }.get(sk, "sport-default")

                    tier = _pick_type_tier(str(leg.get("pick_type") or ""))
                    team = esc(str(leg.get("team") or ""))
                    opp = esc(str(leg.get("opp") or ""))
                    ptype = esc(str(leg.get("prop_type") or ""))
                    player = esc(str(leg.get("player") or ""))
                    edge = leg.get("edge")
                    dir_cls = "dir-over" if d == "OVER" else "dir-under" if d == "UNDER" else ""

                    if lg == "HIT":
                        row_cls = "legrow leg-hit"
                    elif lg == "MISS":
                        row_cls = "legrow leg-miss"
                    elif lg == "VOID":
                        row_cls = "legrow leg-void"
                    else:
                        row_cls = "legrow leg-pend"
                    sym = "✓" if lg == "HIT" else "✗" if lg == "MISS" else "○" if lg == "VOID" else "·"

                    miss_cell = " miss-leg-cell" if lg == "MISS" else ""

                    if lg == "MISS":
                        warning = leg.get("data_warning") or ("LIMITED_Q1_HISTORY" if is_nba1q_group else None)
                        warning_chip = (
                            '<span class="warning-chip" title="NBA1Q stats based on limited Q1 history — use with caution">⚠ Limited Data</span>'
                            if warning else ""
                        )
                        pl_html = (
                            f'<div class="{plcls} pl-line{miss_cell}">'
                            f'<span class="pl-name">{player}{warning_chip}</span>'
                            '<span class="miss-tag" aria-label="Missed leg">MISSED</span></div>'
                        )
                    else:
                        warning = leg.get("data_warning") or ("LIMITED_Q1_HISTORY" if is_nba1q_group else None)
                        warning_chip = (
                            '<span class="warning-chip" title="NBA1Q stats based on limited Q1 history — use with caution">⚠ Limited Data</span>'
                            if warning else ""
                        )
                        void_note = (row.get("void_note") or "").strip() if row else ""
                        void_chip = ""
                        if lg == "VOID" and void_note:
                            vn_esc = esc(void_note)
                            short = (
                                "Postponed"
                                if void_note.upper().startswith("POSTPONED")
                                else void_note
                            )
                            if len(short) > 28:
                                short = short[:25] + "…"
                            void_chip = (
                                f'<span class="warning-chip" title="{vn_esc}">'
                                f"{esc(short)}</span>"
                            )
                        pl_html = f'<div class="{plcls}">{player}{warning_chip}{void_chip}</div>'

                    if lg == "MISS":
                        act_div_cls = "leg-extra val-miss"
                    elif lg == "HIT":
                        act_div_cls = "leg-extra pl-hit"
                    elif lg == "VOID":
                        act_div_cls = "leg-extra pl-void"
                    else:
                        act_div_cls = "leg-extra pl-pend"

                    parts.append(f'<div class="{row_cls}">')
                    parts.append(f'<div class="badge {bcls}">{sym}</div>')
                    parts.append(f'<div><span class="pill {sp_class}">{esc(sk)}</span></div>')
                    parts.append(pl_html)
                    parts.append(f'<div class="tier{miss_cell}">{esc(tier)}</div>')
                    parts.append(
                        f'<div class="leg-prop-col{miss_cell}"><div>{ptype}</div>'
                        f'<div class="meta-muted">{team} vs {opp}</div></div>'
                    )
                    parts.append(
                        f'<div class="leg-extra{miss_cell}">{_fmt_num(lf)} <span class="{dir_cls}">{esc(d)}</span></div>'
                    )
                    parts.append(f'<div class="{act_div_cls}{miss_cell}">{_fmt_num(act)}</div>')
                    parts.append(f'<div class="leg-extra{miss_cell}">{_fmt_num(edge)}</div>')
                    parts.append("</div>")

                parts.append(_ticket_grade_payout_html(oc, esc))
                parts.append("</article>")
                if has_data_warning:
                    parts.append('<div class="ticket-warning-note">⚠ NBA1Q stats based on limited Q1 history - use with caution</div>')
            parts.append("</section>")

        if use_buckets:
            parts.append("</div></details>")

    parts.append("</div>")
    # ──────────────────────────────────────────────────────────────────────────
    # Manual Ticket Builder (appended at the bottom of the tickets page)
    # ──────────────────────────────────────────────────────────────────────────
    parts.append(
        f"""
<style>
  /* Manual Ticket Builder additions (kept isolated so existing UI stays intact) */
  .manual-tb .sport-header {{
    display:flex;align-items:center;gap:12px;
    margin:0 auto 14px;cursor:pointer;user-select:none;
  }}
  .manual-tb .sport-header-line {{
    flex:1;height:1px;background:linear-gradient(90deg,var(--bd2),transparent);
  }}
  .manual-tb .manual-mode-wrap {{
    display:flex;align-items:center;gap:8px;flex-wrap:wrap;
  }}
  .manual-tb .manual-mode-btn {{
    font-size:10px;letter-spacing:1px;
    padding:6px 10px;border-radius:20px;
    cursor:pointer;color:var(--muted);
    border:1px solid var(--border);
    background:var(--bg2);
    transition:all .15s;
    font-family:'Inter',sans-serif;
    text-transform:uppercase;
  }}
  .manual-tb .manual-mode-btn:hover {{
    color:var(--text);
    border-color:var(--bd2);
  }}
  .manual-tb .manual-mode-btn.active {{
    color:var(--accent);
    border-color:var(--accent);
    background:rgba(200,255,0,.06);
  }}

  .manual-tb .manual-grid {{
    display:flex;gap:16px;align-items:flex-start;flex-wrap:wrap;
  }}
  .manual-tb .manual-left, .manual-tb .manual-right {{
    flex:1;min-width:320px;
  }}
  @media(max-width:768px) {{
    .manual-tb .manual-left, .manual-tb .manual-right {{ min-width:100%; }}
  }}

  .manual-tb .manual-filter-row {{
    display:flex;gap:12px;flex-wrap:wrap;align-items:center;margin-bottom:10px;
  }}
  .manual-tb .manual-filter-group {{
    display:flex;gap:8px;flex-wrap:wrap;align-items:center;
  }}
  .manual-tb .manual-slider-row {{
    margin-bottom:10px;
    padding:10px 12px;border:1px solid var(--glass-bd);
    border-radius:14px;background:rgba(255,255,255,0.02);
    display:flex;align-items:center;gap:12px;flex-wrap:wrap;
  }}
  .manual-tb .manual-slider-row label {{
    font-family:'Inter',sans-serif;
    color:var(--muted);font-size:12px;
  }}
  .manual-tb input[type="range"] {{
    width:min(320px,80vw);
  }}

  .manual-tb .chip {{
    font-family:'Bebas Neue',sans-serif;
    display:inline-flex;align-items:center;justify-content:center;
    border-radius:8px;padding:4px 10px;
    font-size:11px;font-weight:700;letter-spacing:.5px;
    border:1px solid transparent;
    user-select:none;
    background:rgba(255,255,255,0.04);
    color:var(--muted);
  }}
  .manual-tb .chip-goblin {{ background:rgba(196,165,255,.10);color:var(--purple);border-color:rgba(196,165,255,.32); }}
  .manual-tb .chip-std    {{ background:rgba(0,229,255,.08);color:var(--cyan);border-color:rgba(0,229,255,.25); }}
  .manual-tb .chip-dup    {{ background:rgba(255,77,77,.10);color:var(--red);border-color:rgba(255,77,77,.32); }}

  .manual-tb .manual-prop-list {{
    max-height:520px;overflow-y:auto;
    border:1px solid var(--glass-bd);
    background:rgba(255,255,255,0.02);
    border-radius:16px;padding:12px;
  }}
  .manual-tb .manual-prop-row {{
    display:flex;align-items:center;gap:10px;flex-wrap:wrap;
    border:1px solid rgba(255,255,255,0.06);
    border-radius:14px;padding:10px 10px;
    margin-bottom:10px;background:rgba(255,255,255,0.02);
    cursor:pointer;
  }}
  .manual-tb .manual-prop-row:hover {{ border-color:rgba(255,255,255,0.12); }}
  .manual-tb .manual-prop-row.dup {{
    opacity:.55;cursor:not-allowed;
    border-color:rgba(255,77,77,.35);
  }}
  .manual-tb .manual-prop-main {{
    min-width:240px;flex:1;
  }}
  .manual-tb .manual-prop-name {{
    font-family:'Inter',sans-serif;font-size:13px;color:var(--text);
  }}
  .manual-tb .manual-prop-sub {{
    font-family:'Inter',sans-serif;font-size:11px;color:var(--muted);margin-top:2px;
  }}
  .manual-tb .manual-prop-metric {{
    font-family:'Inter',sans-serif;font-size:12px;color:var(--muted);
  }}
  .manual-tb .manual-prop-rank {{
    font-family:'Inter',sans-serif;font-size:12px;color:var(--cyan);
  }}

  /* rbar (reused for hit-rate visualization) */
  .rbar{{display:flex;align-items:center;gap:6px;min-width:96px;}}
  .rbar-bg{{flex:1;height:6px;background:rgba(255,255,255,.06);border-radius:4px;overflow:hidden;}}
  .rbar-fill{{height:100%;border-radius:4px;}}
  .rbar-num{{font-size:10px;width:52px;text-align:right;flex-shrink:0;color:var(--muted2);}}

  .manual-tb .manual-active-card {{
    border:1px solid var(--glass-bd);
    border-radius:16px;background:rgba(255,255,255,0.02);
    padding:12px;
  }}
  .manual-tb .manual-stats-row {{
    display:flex;gap:14px;flex-wrap:wrap;align-items:center;justify-content:flex-start;margin-bottom:10px;
  }}
  .manual-tb .manual-stat {{
    font-family:'Inter',sans-serif;font-size:12px;color:var(--muted);
  }}
  .manual-tb .manual-stat strong {{
    color:var(--text);font-family:'Inter',sans-serif;
  }}
  .manual-tb .manual-warning {{
    font-family:'Inter',sans-serif;font-size:12px;color:var(--amber);
    border:1px solid rgba(200,160,60,.35);background:rgba(200,160,60,.08);
    padding:6px 10px;border-radius:999px;display:flex;align-items:center;gap:8px;
  }}

  .manual-tb .manual-active-legs {{
    display:flex;flex-direction:column;gap:10px;margin-bottom:12px;
  }}
  .manual-tb .manual-active-leg {{
    display:flex;align-items:center;gap:10px;flex-wrap:wrap;
    border:1px solid rgba(255,255,255,0.06);border-radius:14px;padding:10px;
    background:rgba(255,255,255,0.02);
  }}
  .manual-tb .manual-remove-btn {{
    margin-left:auto;
    border:1px solid rgba(255,255,255,0.12);
    background:rgba(255,255,255,0.04);
    color:var(--muted);
    font-family:'Inter',sans-serif;
    border-radius:10px;
    padding:6px 10px;cursor:pointer;
    transition:all .15s;
  }}
  .manual-tb .manual-remove-btn:hover {{
    color:var(--text);border-color:rgba(255,255,255,0.22);
  }}

  .manual-tb .manual-actions {{
    display:flex;gap:10px;align-items:center;flex-wrap:wrap;margin-bottom:12px;
  }}
  .manual-tb .manual-action-btn {{
    border:1px solid rgba(255,255,255,0.14);
    background:rgba(255,255,255,0.05);
    color:var(--text);
    font-family:'Inter',sans-serif;
    border-radius:14px;
    padding:10px 14px;cursor:pointer;
    transition:all .15s;
  }}
  .manual-tb .manual-action-btn:hover {{
    border-color:rgba(255,255,255,0.24);
    transform:translateY(-1px);
  }}
  .manual-tb .manual-action-btn.primary {{
    border-color:rgba(212,175,55,.45);
    background:rgba(212,175,55,.08);
    color:var(--gold);
  }}
  .manual-tb .manual-action-btn:disabled {{
    opacity:.45;cursor:not-allowed;transform:none;
  }}

  .manual-tb .manual-locked-list {{
    display:flex;flex-direction:column;gap:14px;
  }}
  .manual-tb .manual-locked-card {{
    border:1px solid var(--glass-bd);
    border-radius:16px;background:rgba(255,255,255,0.02);
    overflow:hidden;
  }}
  .manual-tb .manual-locked-hdr {{
    display:flex;align-items:center;gap:10px;flex-wrap:wrap;
    padding:12px 14px;border-bottom:1px solid rgba(255,255,255,0.06);
  }}
  .manual-tb .manual-locked-legs {{
    padding:12px 14px;display:flex;flex-direction:column;gap:8px;
  }}
  .manual-tb .manual-locked-leg {{
    display:flex;align-items:center;gap:8px;flex-wrap:wrap;
    font-family:'Inter',sans-serif;font-size:12px;color:var(--muted);
  }}
  .manual-tb .manual-locked-leg .player {{
    color:var(--text);
    font-size:13px;
  }}
  .manual-tb .manual-locked-x {{
    margin-left:auto;
    border:1px solid rgba(255,255,255,0.12);
    background:rgba(255,255,255,0.04);
    color:var(--muted);
    border-radius:10px;
    padding:6px 10px;cursor:pointer;
  }}
</style>

<details class="ticket-bucket sb-default manual-tb" open>
  <summary>
    <span>🎯 Manual Ticket Builder</span>
    <span class="manual-mode-wrap">
      <button class="manual-mode-btn active" data-manual-mode="MIXED" type="button">MIXED</button>
      <button class="manual-mode-btn" data-manual-mode="GOBLIN" type="button">GOBLIN ONLY</button>
      <button class="manual-mode-btn" data-manual-mode="STANDARD" type="button">STANDARD ONLY</button>
    </span>
  </summary>

  <div class="ticket-bucket-body">
    <div class="manual-grid">
      <div class="manual-left">
        <div class="manual-filter-row">
          <div class="manual-filter-group">
            <button class="pill sport-nba manual-sport-pill manual-sport-btn" data-manual-sport="NBA" type="button">NBA</button>
            <button class="pill sport-cbb manual-sport-pill manual-sport-btn" data-manual-sport="CBB" type="button">CBB</button>
            <button class="pill sport-nhl manual-sport-pill manual-sport-btn" data-manual-sport="NHL" type="button">NHL</button>
            <button class="pill sport-soccer manual-sport-pill manual-sport-btn" data-manual-sport="SOCCER" type="button">SOC</button>
          </div>
        </div>

        <div class="manual-filter-row" style="margin-top:-4px;">
          <div class="manual-filter-group">
            <button class="chip chip-goblin manual-pick-type-btn" data-manual-pick-type="Goblin" type="button">🎃 Goblin</button>
            <button class="chip chip-std manual-pick-type-btn" data-manual-pick-type="Standard" type="button">⭐ Standard</button>
          </div>
        </div>

        <div class="manual-slider-row">
          <label for="manual-hit-rate-slider">Min hit-rate: <strong id="manual-hit-rate-label">70%</strong></label>
          <input id="manual-hit-rate-slider" type="range" min="60" max="100" value="70" step="1" />
        </div>

        <div class="manual-prop-list" id="manual-prop-list" aria-live="polite"></div>
      </div>

      <div class="manual-right">
        <div class="manual-active-card">
          <div class="manual-stats-row">
            <div class="manual-stat">Legs: <strong id="manual-leg-count">0</strong></div>
            <div class="manual-stat">Avg hit-rate: <strong id="manual-avg-hit-rate">0.0%</strong></div>
            <div class="manual-stat">Est. win prob: <strong id="manual-win-prob">0.0%</strong></div>
            <div class="manual-stat">Avg edge: <strong id="manual-avg-edge">0.00</strong></div>
            <div id="manual-corr-warning" style="margin-left:auto;display:none;" class="manual-warning">⚠ Correlation: same team + game</div>
          </div>

          <div class="manual-active-legs" id="manual-active-legs"></div>

          <div class="manual-actions">
            <button id="manual-auto-build" class="manual-action-btn" type="button">⚡ Auto-Build</button>
            <button id="manual-clear" class="manual-action-btn" type="button">Clear</button>
            <button id="manual-lock" class="manual-action-btn primary" type="button" disabled>Lock Ticket</button>
          </div>

          <div class="sec" style="margin-top:0;">
            <h2 class="sec-head bebas" style="margin-bottom:10px;font-size:clamp(24px,2.8vw,34px);">Locked Tickets</h2>
            <div class="manual-locked-list" id="manual-locked-list"></div>
          </div>
        </div>
      </div>
    </div>
  </div>
</details>

<script>
  window.SLATE_DATA = {manual_props_json};
</script>

<script>
(function(){{
  const $ = (sel, root=document) => root.querySelector(sel);
  const $$ = (sel, root=document) => Array.from(root.querySelectorAll(sel));

  const props = Array.isArray(window.SLATE_DATA) ? window.SLATE_DATA : [];

  const state = {{
    mode: 'MIXED',
    selectedSports: new Set(['NBA','CBB','NHL','SOCCER']),
    // In MIXED, both pick types are enabled; in single-mode, pick one.
    selectedPickTypes: new Set(['Goblin','Standard']),
    minHitRate: 0.70,
    activeLegs: [],
    lockedTickets: [],
    // Auto-build targets current leg count target default 3.
    targetLegs: 3,
    maxLegs: 6,
    // Cache for re-render efficiency
    activePlayers: new Set()
  }};

  function legKey(leg){{
    const lt = [leg.sport, leg.player, leg.team, leg.opp, leg.prop_type, leg.pick_type, leg.direction, leg.line, leg.game_time];
    return lt.map(x => (x===null||x===undefined)? '' : String(x)).join('||');
  }}

  function gameKeyNorm(leg){{
    const team = leg.team || '';
    const opp = leg.opp || '';
    const teams = [team, opp].sort().join('-');
    const gt = leg.game_time || '';
    return `${{leg.sport}}|${{teams}}|${{gt}}`;
  }}

  function corrKeyTeamGame(leg){{
    // Correlation warning: two legs share the same team+game (same sport/team/game-time).
    return `${{leg.sport}}|${{leg.team}}|${{gameKeyNorm(leg)}}`;
  }}

  function formatPct(x){{
    const v = Number(x);
    if (!isFinite(v)) return '0.0%';
    return `${{(v*100).toFixed(1)}}%`;
  }}

  function fmt2(x){{
    const v = Number(x);
    if (!isFinite(v)) return '0.00';
    return v.toFixed(2);
  }}

  function hitColor(pct){{
    // pct is 0..100
    if (pct >= 65) return 'var(--green)';
    if (pct >= 50) return 'var(--amber)';
    return 'var(--red)';
  }}

  function rbarHTML(hitRate){{
    const hr = (hitRate===null||hitRate===undefined)? 0 : Number(hitRate);
    const pct = Math.max(0, Math.min(100, hr*100));
    const bc = hitColor(pct);
    const tc = bc;
    return `
      <div class="rbar" aria-label="Hit rate ${{pct.toFixed(1)}}%">
        <div class="rbar-bg"><div class="rbar-fill" style="width:${{pct.toFixed(1)}}%;background:${{bc}}"></div></div>
        <span class="rbar-num" style="color:${{tc}}">${{pct.toFixed(1)}}%</span>
      </div>`;
  }}

  function filteredProps(){{
    const m = state.mode;
    // Mode gate (pick types)
    const allowedPickTypes = state.selectedPickTypes;
    const allowedSports = state.selectedSports;
    const minHr = state.minHitRate;

    return props
      .filter(p => allowedSports.has(p.sport))
      .filter(p => allowedPickTypes.has(p.pick_type))
      .filter(p => (p.hit_rate===null||p.hit_rate===undefined) ? false : Number(p.hit_rate) >= minHr)
      .sort((a,b) => {{
        const ha = Number(a.hit_rate||0), hb = Number(b.hit_rate||0);
        if (hb !== ha) return hb - ha;
        const ra = Number(a.rank_score||-1), rb = Number(b.rank_score||-1);
        return rb - ra;
      }});
  }}

  function renderPropList(){{
    const list = $('#manual-prop-list');
    if (!list) return;
    const fp = filteredProps();

    const sportShort = (s) => s === 'SOCCER' ? 'SOC' : s;
    const modePick = state.mode;

    list.innerHTML = '';
    if (!fp.length) {{
      list.innerHTML = `<div style="color:var(--muted);font-family:'Inter',sans-serif;font-size:12px;padding:8px;">No props match filters.</div>`;
      return;
    }}

    for (const leg of fp) {{
      const dup = state.activePlayers.has(leg.player);
      const pickChip = leg.pick_type === 'Goblin'
        ? `<span class="chip chip-goblin">${{leg.pick_type}}</span>`
        : `<span class="chip chip-std">${{leg.pick_type}}</span>`;

      const dupChip = dup ? `<span class="chip chip-dup">DUP</span>` : '';
      const dirCls = leg.direction === 'OVER' ? 'dir-over' : leg.direction === 'UNDER' ? 'dir-under' : '';

      const lineStr = (leg.line===null||leg.line===undefined) ? '—' : String(leg.line);
      const edgeStr = fmt2(leg.edge);
      const rankStr = (leg.rank_score===null||leg.rank_score===undefined) ? '—' : Number(leg.rank_score).toFixed(2);
      const hitBar = rbarHTML(leg.hit_rate);

      const row = document.createElement('div');
      row.className = 'manual-prop-row' + (dup ? ' dup' : '');
      row.setAttribute('role','button');
      row.setAttribute('tabindex','0');
      row.setAttribute('aria-disabled', dup ? 'true' : 'false');

      row.innerHTML = `
        <div class="manual-prop-main">
          <div class="manual-prop-name">${{leg.player}}</div>
          <div class="manual-prop-sub">${{leg.team}} · ${{leg.prop_type}}</div>
        </div>
        <div class="manual-prop-metric">
          <div style="color:var(--text);font-family:'Inter',sans-serif;font-size:12px;">${{lineStr}} <span class="${{dirCls}}">${{leg.direction || ''}}</span></div>
          <div style="margin-top:4px;">${{pickChip}} ${{dupChip}}</div>
        </div>
        <div class="manual-prop-metric" style="min-width:120px;">${{hitBar}}</div>
        <div class="manual-prop-metric">Edge: <span style="color:var(--text);font-size:12px;">${{edgeStr}}</span></div>
        <div class="manual-prop-rank">Rank: <span>${{rankStr}}</span></div>
      `;

      const key = legKey(leg);
      row.dataset.legKey = key;
      row.addEventListener('click', () => {{
        if (dup) return;
        if (state.activeLegs.length >= state.maxLegs) return;
        addLegToActive(leg);
      }});
      row.addEventListener('keydown', (e) => {{
        if (e.key === 'Enter' || e.key === ' ') row.click();
      }});

      list.appendChild(row);
    }}
  }}

  function computeCorrelation(){{
    if (state.activeLegs.length < 2) return false;
    const counts = new Map();
    for (const leg of state.activeLegs) {{
      const k = corrKeyTeamGame(leg);
      counts.set(k, (counts.get(k) || 0) + 1);
    }}
    for (const [,c] of counts) if (c > 1) return true;
    return false;
  }}

  function updateActiveStats(){{
    const n = state.activeLegs.length;
    $('#manual-leg-count').textContent = String(n);

    if (n === 0) {{
      $('#manual-avg-hit-rate').textContent = '0.0%';
      $('#manual-win-prob').textContent = '0.0%';
      $('#manual-avg-edge').textContent = '0.00';
      $('#manual-corr-warning').style.display = 'none';
      $('#manual-lock').disabled = true;
      return;
    }}

    const avgHit = state.activeLegs.reduce((s,l)=>s+Number(l.hit_rate||0),0) / n;
    const winProb = state.activeLegs.reduce((p,l)=>p*(Number(l.hit_rate||0)),1);
    const avgEdge = state.activeLegs.reduce((s,l)=>s+Number(l.edge||0),0) / n;

    $('#manual-avg-hit-rate').textContent = `${{(avgHit*100).toFixed(1)}}%`;
    $('#manual-win-prob').textContent = `${{(winProb*100).toFixed(1)}}%`;
    $('#manual-avg-edge').textContent = fmt2(avgEdge);
    $('#manual-corr-warning').style.display = computeCorrelation() ? '' : 'none';

    $('#manual-lock').disabled = n < 2;
  }}

  function renderActiveLegs(){{
    const cont = $('#manual-active-legs');
    if (!cont) return;
    cont.innerHTML = '';
    for (const leg of state.activeLegs) {{
      const dirCls = leg.direction === 'OVER' ? 'dir-over' : leg.direction === 'UNDER' ? 'dir-under' : '';
      const spClass = leg.sport === 'NBA' ? 'sport-nba'
        : leg.sport === 'CBB' ? 'sport-cbb'
        : leg.sport === 'NHL' ? 'sport-nhl'
        : leg.sport === 'SOCCER' ? 'sport-soccer'
        : 'sport-default';

      const tierTxt = leg.pick_type === 'Goblin' ? 'G' : 'S';
      const key = legKey(leg);

      const row = document.createElement('div');
      row.className = 'manual-active-leg';
      row.innerHTML = `
        <span class="pill ${{spClass}}">${{leg.sport}}</span>
        <span class="player" style="font-family:'Inter',sans-serif;color:var(--text);font-size:13px;">${{leg.player}}</span>
        <span style="color:var(--muted);font-size:12px;">${{leg.prop_type}}</span>
        <span style="margin-left:auto;color:var(--text);font-size:12px;font-family:'Inter',sans-serif;">
          ${{leg.line}} <span class="${{dirCls}}">${{leg.direction}}</span>
        </span>
        <span style="display:flex;align-items:center;gap:8px;">
          <span class="chip ${{leg.pick_type==='Goblin' ? 'chip-goblin' : 'chip-std'}}">${{leg.pick_type}}</span>
        </span>
        <button type="button" class="manual-remove-btn" data-remove-leg="${{key}}">✕</button>
      `;
      cont.appendChild(row);
    }}

    // Wire remove buttons after DOM creation
    $$('button.manual-remove-btn', cont).forEach(btn => {{
      btn.addEventListener('click', () => {{
        const k = btn.dataset.removeLeg;
        state.activeLegs = state.activeLegs.filter(l => legKey(l) !== k);
        state.activePlayers = new Set(state.activeLegs.map(l => l.player));
        renderActiveLegs();
        renderLockedTicketsHeaderOnly();
        renderPropList();
        updateActiveStats();
      }});
    }});
  }}

  function renderLockedTicketsHeaderOnly(){{
    // placeholder hook (kept for future extension)
    // We intentionally keep it simple to avoid interfering with existing logic.
  }}

  function renderLockedTickets(){{
    const cont = $('#manual-locked-list');
    if (!cont) return;
    cont.innerHTML = '';
    state.lockedTickets.forEach((t, idx) => {{
      const winProbPct = (t.winProb*100).toFixed(1);
      const avgEdge = t.avgEdge;
      const card = document.createElement('div');
      card.className = 'manual-locked-card';
      card.innerHTML = `
        <div class="manual-locked-hdr">
          <span class="tn bebas">Manual #${{idx+1}}</span>
          <span class="tg" style="color:var(--muted);font-family:'Inter',sans-serif;font-size:12px;">${{t.legs.length}} legs</span>
          <span class="payout" style="margin-left:0;">Est. win prob: ${{winProbPct}}%</span>
          <span class="payout" style="margin-left:12px;">Avg edge: ${{fmt2(avgEdge)}}</span>
          <button type="button" class="manual-locked-x" data-remove-ticket="${{idx}}">✕</button>
        </div>
        <div class="manual-locked-legs">
          ${{t.legs.map(l => {{
            const dirCls = l.direction === 'OVER' ? 'dir-over' : l.direction === 'UNDER' ? 'dir-under' : '';
            const spClass = l.sport === 'NBA' ? 'sport-nba'
              : l.sport === 'CBB' ? 'sport-cbb'
              : l.sport === 'NHL' ? 'sport-nhl'
              : l.sport === 'SOCCER' ? 'sport-soccer'
              : 'sport-default';
            return `
              <div class="manual-locked-leg">
                <span class="pill ${{spClass}}">${{l.sport}}</span>
                <span class="player">${{l.player}}</span>
                <span style="color:var(--muted);">${{l.prop_type}}</span>
                <span style="color:var(--text);">${{l.line}} <span class="${{dirCls}}">${{l.direction}}</span></span>
                <span class="chip ${{l.pick_type==='Goblin' ? 'chip-goblin' : 'chip-std'}}">${{l.pick_type}}</span>
              </div>
            `;
          }}).join('')}}
        </div>
      `;
      cont.appendChild(card);
    }});

    // remove wiring
    $$('button.manual-locked-x', cont).forEach(btn => {{
      btn.addEventListener('click', () => {{
        const i = Number(btn.dataset.removeTicket);
        state.lockedTickets.splice(i, 1);
        renderLockedTickets();
      }});
    }});
  }}

  function addLegToActive(leg){{
    const player = leg.player;
    if (!player) return;
    if (state.activePlayers.has(player)) return;
    if (state.activeLegs.length >= state.maxLegs) return;

    state.activeLegs.push(leg);
    state.activePlayers.add(player);

    renderActiveLegs();
    renderPropList();
    updateActiveStats();
  }}

  function clearActive(){{
    state.activeLegs = [];
    state.activePlayers = new Set();
    renderActiveLegs();
    renderPropList();
    updateActiveStats();
  }}

  function lockActive(){{
    if (state.activeLegs.length < 2) return;
    const winProb = state.activeLegs.reduce((p,l)=>p*(Number(l.hit_rate||0)),1);
    const avgEdge = state.activeLegs.reduce((s,l)=>s+Number(l.edge||0),0) / state.activeLegs.length;

    state.lockedTickets.push({{
      legs: state.activeLegs.slice(),
      winProb: winProb,
      avgEdge: avgEdge
    }});

    clearActive();
    renderLockedTickets();
  }}

  function applyModeToPickTypes(){{
    const mode = state.mode;
    if (mode === 'GOBLIN') {{
      state.selectedPickTypes = new Set(['Goblin']);
    }} else if (mode === 'STANDARD') {{
      state.selectedPickTypes = new Set(['Standard']);
    }} else {{
      state.selectedPickTypes = new Set(['Goblin','Standard']);
    }}

    // Update pick-type button visibility/disabled
    for (const btn of $$('.manual-pick-type-btn')) {{
      const pt = btn.dataset.manualPickType;
      const show = state.selectedPickTypes.has(pt);
      btn.style.display = show ? '' : 'none';
    }}
  }}

  function setMode(newMode){{
    state.mode = newMode;
    const modeBtns = $$('.manual-mode-btn');
    modeBtns.forEach(b => b.classList.toggle('active', b.dataset.manualMode === newMode));
    applyModeToPickTypes();
    renderPropList();
    updateActiveStats();
  }}

  function applySportButtonUI(){{
    for (const btn of $$('.manual-sport-btn')) {{
      const s = btn.dataset.manualSport;
      const active = state.selectedSports.has(s);
      // active state uses a stronger border via inline style since we don't own sport CSS here.
      btn.style.opacity = active ? '1' : '.55';
      btn.style.filter = active ? 'none' : 'grayscale(0.2)';
    }}
  }}

  function toggleSport(s){{
    if (state.selectedSports.has(s)) state.selectedSports.delete(s);
    else state.selectedSports.add(s);
    // Keep at least one sport selected for a usable picker.
    if (state.selectedSports.size === 0) state.selectedSports = new Set(['NBA','CBB','NHL','SOCCER']);
    applySportButtonUI();
    renderPropList();
  }}

  function autoBuild(){{
    const activeCount = state.activeLegs.length;
    if (activeCount >= state.maxLegs) return;

    // Target leg count for auto-build is the current target (default 3).
    const needed = Math.max(0, Math.min(state.maxLegs, state.targetLegs) - activeCount);
    if (needed <= 0) return;

    const candidates = filteredProps()
      .filter(c => !state.activePlayers.has(c.player));

    if (!candidates.length) return;

    // Sort by hit_rate desc then rank_score desc
    candidates.sort((a,b) => {{
      const ha = Number(a.hit_rate||0), hb = Number(b.hit_rate||0);
      if (hb !== ha) return hb - ha;
      const ra = Number(a.rank_score||-1), rb = Number(b.rank_score||-1);
      return rb - ra;
    }});

    const chosen = [];
    const usedGames = new Set();

    // Pass 1: one per game where possible
    for (const c of candidates) {{
      if (chosen.length >= needed) break;
      const gk = gameKeyNorm(c);
      if (usedGames.has(gk)) continue;
      usedGames.add(gk);
      chosen.push(c);
    }}

    // Pass 2: fill remaining even if same game
    for (const c of candidates) {{
      if (chosen.length >= needed) break;
      const already = chosen.some(x => x.player === c.player);
      if (already) continue;
      chosen.push(c);
    }}

    for (const c of chosen.slice(0, needed)) addLegToActive(c);
  }}

  // ── Wire UI ─────────────────────────────────────────────────────────────────
  for (const btn of $$('.manual-mode-btn')) {{
    btn.addEventListener('click', () => setMode(btn.dataset.manualMode));
  }}
  for (const btn of $$('.manual-sport-btn')) {{
    btn.addEventListener('click', () => toggleSport(btn.dataset.manualSport));
  }}
  for (const btn of $$('.manual-pick-type-btn')) {{
    btn.addEventListener('click', () => {{
      // In single-mode, pick-type buttons are hidden; only MIXED makes them clickable.
      if (state.mode !== 'MIXED') return;
      const pt = btn.dataset.manualPickType;
      if (state.selectedPickTypes.has(pt)) state.selectedPickTypes.delete(pt);
      else state.selectedPickTypes.add(pt);
      if (state.selectedPickTypes.size === 0) state.selectedPickTypes = new Set(['Goblin','Standard']);
      applyModeToPickTypes();
      renderPropList();
    }});
  }}

  const slider = $('#manual-hit-rate-slider');
  slider.addEventListener('input', () => {{
    const v = Number(slider.value);
    $('#manual-hit-rate-label').textContent = `${{v}}%`;
    state.minHitRate = v / 100;
    renderPropList();
  }});

  $('#manual-auto-build').addEventListener('click', autoBuild);
  $('#manual-clear').addEventListener('click', clearActive);
  $('#manual-lock').addEventListener('click', lockActive);

  // Initial UI
  applyModeToPickTypes();
  applySportButtonUI();
  renderActiveLegs();
  renderLockedTickets();
  updateActiveStats();
  renderPropList();
}})();
</script>
        """
    )

    parts.append("</div>")
    parts.append(
        f'<script src="/static/site-nav-chrome.js?v={_TICKET_EVAL_NAV_CHROME_VER}" defer></script>'
    )
    parts.append("</body></html>")
    return "\n".join(parts), history_record


def main() -> int:
    logging.basicConfig(
        level=logging.WARNING,
        format="%(asctime)s %(levelname)s %(name)s: %(message)s",
        datefmt="%Y-%m-%d %H:%M:%S",
    )
    ap = argparse.ArgumentParser(description="Build ticket_eval HTML for Grades UI.")
    ap.add_argument(
        "--date",
        default="",
        help="Slate date YYYY-MM-DD (default: yesterday local)",
    )
    ap.add_argument(
        "--debug",
        action="store_true",
        help="Print ticket JSON path, payload date, outputs/graded files, Excel headers, sample leg matches; then build.",
    )
    ap.add_argument(
        "--game-date",
        action="append",
        default=None,
        metavar="YYYY-MM-DD",
        help="Extra calendar date(s) to merge graded_*.xlsx from (repeatable). Leg game_time dates are auto-detected.",
    )
    ap.add_argument(
        "--graded",
        default="",
        help="Ignored; retained for run_grader.ps1 compatibility with build_ticket_eval_html.py.",
    )
    ap.add_argument(
        "--out",
        default="",
        help="Ignored; output is always ui_runner/templates/ticket_eval_{--date}.html.",
    )
    args = ap.parse_args()
    if args.date:
        arg_date = args.date.strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", arg_date):
            print("ERROR: --date must be YYYY-MM-DD")
            return 1
    else:
        arg_date = (date.today() - timedelta(days=1)).isoformat()

    sport_candidates = _dated_candidates(arg_date)

    tpath = find_ticket_json(arg_date)
    if not tpath:
        print(
            "ERROR: No ticket file found (combined_slate_tickets_{date}.xlsx)."
        )
        return 1

    try:
        payload = _load_tickets(tpath, arg_date)
        payload = _filter_payload_groups(payload, debug=bool(args.debug))
    except Exception as e:
        print(f"ERROR: Failed to read ticket file: {e}")
        return 1

    extra_game_dates: list[str] = list(args.game_date) if args.game_date else []
    graded_merge_dates, leg_game_dates_for_log = resolve_ticket_eval_graded_merge_dates(
        arg_date, payload, extra_game_dates
    )
    parsed_leg_dates = sorted(_game_dates_from_ticket_payload(payload, arg_date))
    infer_note = ""
    if leg_game_dates_for_log and not parsed_leg_dates:
        infer_note = " (inferred next day; ticket file has no leg game_time column)"
    print(
        f"[TICKET EVAL] Slate date: {arg_date}; leg game_time dates: "
        f"{leg_game_dates_for_log or '[]'}{infer_note}; "
        f"graded workbook merge order: {graded_merge_dates}"
    )

    if args.debug:
        debug_report(arg_date, payload, tpath, sport_candidates, graded_merge_dates)

    html_out, hist = _build_html(payload, arg_date, sport_candidates, graded_merge_dates)
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    dated_name = f"ticket_eval_{arg_date}.html"
    out_dated = TEMPLATES_DIR / dated_name
    try:
        out_dated.write_text(html_out, encoding="utf-8")
    except OSError as e:
        print(f"ERROR: Write failed: {e}")
        return 1

    if hist:
        try:
            _append_grade_history(hist)
            print(f"  Appended grade history -> data/grade_history.json ({hist.get('date')})")
        except OSError as e:
            print(f"  WARN: could not append grade_history.json: {e}")

    print(f"Wrote {out_dated}")
    print("  (Serve /tickets from tickets_latest.json; graded view: Grades → Ticket evaluation.)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
