from __future__ import annotations

import argparse
import csv
from dataclasses import dataclass
from pathlib import Path
from typing import Any

import openpyxl


@dataclass(frozen=True)
class LegKey:
    sport: str
    player: str
    prop: str
    direction: str
    pick_type: str


def _norm(raw: Any) -> str:
    return " ".join(str(raw or "").strip().lower().split())


def _to_float(raw: Any) -> float | None:
    try:
        if raw is None or str(raw).strip() == "":
            return None
        return float(raw)
    except (TypeError, ValueError):
        return None


def _read_full_slate(path: Path) -> dict[LegKey, dict[str, Any]]:
    wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
    try:
        if "Full Slate" not in wb.sheetnames:
            raise ValueError(f"'Full Slate' sheet missing in {path}")
        ws = wb["Full Slate"]
        rows = ws.iter_rows(values_only=True)
        headers = [str(x).strip() if x is not None else "" for x in next(rows)]
        idx = {h: i for i, h in enumerate(headers)}
        required = ["Sport", "Player", "Prop", "Dir", "Pick Type", "Line", "Standard Line"]
        missing = [c for c in required if c not in idx]
        if missing:
            raise ValueError(f"Missing columns in {path}: {missing}")

        out: dict[LegKey, dict[str, Any]] = {}
        for r in rows:
            sport = _norm(r[idx["Sport"]])
            player = _norm(r[idx["Player"]])
            prop = _norm(r[idx["Prop"]])
            direction = _norm(r[idx["Dir"]]).upper()
            pick_type = _norm(r[idx["Pick Type"]]).title()
            if not sport or not player or not prop:
                continue
            key = LegKey(sport=sport, player=player, prop=prop, direction=direction, pick_type=pick_type)
            out[key] = {
                "line": _to_float(r[idx["Line"]]),
                "standard_line": _to_float(r[idx["Standard Line"]]),
                "line_discount_vs_standard": _to_float(r[idx.get("Line Discount Vs Standard", -1)]) if "Line Discount Vs Standard" in idx else None,
            }
        return out
    finally:
        wb.close()


def _write_csv(path: Path, rows: list[dict[str, Any]], fieldnames: list[str]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        w.writeheader()
        for row in rows:
            w.writerow(row)


def main() -> int:
    ap = argparse.ArgumentParser(description="Compare two combined_slate_tickets Full Slate sheets.")
    ap.add_argument("--old", required=True, help="Older combined_slate_tickets_*.xlsx")
    ap.add_argument("--new", required=True, help="Newer combined_slate_tickets_*.xlsx")
    ap.add_argument("--outdir", required=True, help="Output directory for compare report CSVs")
    args = ap.parse_args()

    old_path = Path(args.old).expanduser().resolve()
    new_path = Path(args.new).expanduser().resolve()
    outdir = Path(args.outdir).expanduser().resolve()
    outdir.mkdir(parents=True, exist_ok=True)

    old = _read_full_slate(old_path)
    new = _read_full_slate(new_path)

    old_keys = set(old.keys())
    new_keys = set(new.keys())
    both = old_keys & new_keys

    added = sorted(new_keys - old_keys, key=lambda k: (k.sport, k.player, k.prop, k.direction, k.pick_type))
    removed = sorted(old_keys - new_keys, key=lambda k: (k.sport, k.player, k.prop, k.direction, k.pick_type))

    moved_all: list[dict[str, Any]] = []
    moved_standard: list[dict[str, Any]] = []
    for k in sorted(both, key=lambda x: (x.sport, x.player, x.prop, x.direction, x.pick_type)):
        a = old[k]
        b = new[k]
        old_line = a.get("line")
        new_line = b.get("line")
        old_std = a.get("standard_line")
        new_std = b.get("standard_line")
        old_disc = a.get("line_discount_vs_standard")
        new_disc = b.get("line_discount_vs_standard")
        if old_line != new_line or old_std != new_std or old_disc != new_disc:
            row = {
                "sport": k.sport,
                "player": k.player,
                "prop": k.prop,
                "direction": k.direction,
                "pick_type": k.pick_type,
                "old_line": old_line,
                "new_line": new_line,
                "line_delta": (new_line - old_line) if (old_line is not None and new_line is not None) else "",
                "old_standard_line": old_std,
                "new_standard_line": new_std,
                "standard_line_delta": (new_std - old_std) if (old_std is not None and new_std is not None) else "",
                "old_line_discount_vs_standard": old_disc,
                "new_line_discount_vs_standard": new_disc,
            }
            moved_all.append(row)
            if k.pick_type == "Standard":
                moved_standard.append(row)

    _write_csv(
        outdir / "line_moves_all.csv",
        moved_all,
        [
            "sport",
            "player",
            "prop",
            "direction",
            "pick_type",
            "old_line",
            "new_line",
            "line_delta",
            "old_standard_line",
            "new_standard_line",
            "standard_line_delta",
            "old_line_discount_vs_standard",
            "new_line_discount_vs_standard",
        ],
    )
    _write_csv(
        outdir / "line_moves_standard_only.csv",
        moved_standard,
        [
            "sport",
            "player",
            "prop",
            "direction",
            "pick_type",
            "old_line",
            "new_line",
            "line_delta",
            "old_standard_line",
            "new_standard_line",
            "standard_line_delta",
            "old_line_discount_vs_standard",
            "new_line_discount_vs_standard",
        ],
    )
    _write_csv(
        outdir / "added_props.csv",
        [
            {
                "sport": k.sport,
                "player": k.player,
                "prop": k.prop,
                "direction": k.direction,
                "pick_type": k.pick_type,
            }
            for k in added
        ],
        ["sport", "player", "prop", "direction", "pick_type"],
    )
    _write_csv(
        outdir / "removed_props.csv",
        [
            {
                "sport": k.sport,
                "player": k.player,
                "prop": k.prop,
                "direction": k.direction,
                "pick_type": k.pick_type,
            }
            for k in removed
        ],
        ["sport", "player", "prop", "direction", "pick_type"],
    )
    _write_csv(
        outdir / "summary.csv",
        [
            {
                "old_file": str(old_path),
                "new_file": str(new_path),
                "old_rows": len(old),
                "new_rows": len(new),
                "shared_rows": len(both),
                "added_rows": len(added),
                "removed_rows": len(removed),
                "moved_rows_any_line": len(moved_all),
                "moved_rows_standard_only": len(moved_standard),
            }
        ],
        [
            "old_file",
            "new_file",
            "old_rows",
            "new_rows",
            "shared_rows",
            "added_rows",
            "removed_rows",
            "moved_rows_any_line",
            "moved_rows_standard_only",
        ],
    )

    print(f"[compare] wrote report to: {outdir}")
    print(f"[compare] moved_rows_any_line={len(moved_all)} moved_rows_standard_only={len(moved_standard)}")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())

