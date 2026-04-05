#!/usr/bin/env python3
r"""
combined_ticket_grader_UPDATED.py
================================
Full analytics grader for the output of combined_slate_tickets.py, with **dynamic payout modifiers**
for Goblin/Demon legs based on "distance from line" buckets (dev levels).

Key upgrade vs v1:
- Supports leg types like:
    Standard
    Goblin -1 / -2 / -3   (more discounted => lower payout)
    Demon  +1 / +2 / +3   (more juiced     => higher payout)
- If your ticket workbook only has pick_type = Goblin/Demon (no dev), we default to:
    Goblin -> Goblin -1
    Demon  -> Demon +1
- Payout multipliers are computed like your payout_calculator.jsx:
    Power: POWER_BASE[n] * Π(leg_modifier_power)
    Flex : FLEX_BASE[n][hits] * Π(leg_modifier_flex)

Outputs:
- SUMMARY (key metrics)
- ANALYSIS_INSIGHTS (readable takeaways from the slate)
- TICKET_RESULTS (one row per ticket per mode)
- LEG_RESULTS (one row per leg, with HIT/MISS/PUSH/VOID/NO_ACTUAL)
- LEG_BY_* breakdowns (prop, sport, direction, pick type, sheet)
- TICKET_DEEP_DIVE (per-ticket leg mix + outcomes)
- Analytics tabs per mode (ROI by sheet/legs/sports/pick_types)
- ML_* (optional): RandomForest leg hit model, feature importance, filter simulation
  (exploratory — trained on the same graded slate; use for patterns, not live edge claims)
- ML_CALIBRATION sheet: buckets slate ml_prob vs realized hit rate (find miscalibration before retraining).
- --export-graded-legs-csv: append-friendly training export (stack slates, retrain step8 / leg ML offline).
- TICKET_BACKTEST / TICKET_OBJ_DECILES + --export-graded-tickets-csv: ticket-level paid vs modeled objective (matches /tickets UI).
- Stack CSVs and run: py scripts/backtest_ticket_objectives.py training/graded_tickets_all.csv

Usage (PowerShell):
  py -3.14 .\\combined_ticket_grader_UPDATED.py `
    --tickets .\combined_slate_tickets_2026-02-21.xlsx `
    --nba_actuals ".\grades\actuals_nba_2026-02-21.csv" `
    --cbb_actuals ".\grades\actuals_cbb_2026-02-21.csv" `
    --mode both --stake 20

Optional config override:
  py -3.14 .\\combined_ticket_grader_UPDATED.py ... --payouts_json .\grades\payouts_2026-02-21.json

JSON schema (example):
{
  "power_base": { "2":3.0, "3":6.0, "4":10.0, "5":20.0, "6":37.5 },
  "flex_base": {
    "2": { "2":3.0 },
    "3": { "3":3.0, "2":1.0 },
    "4": { "4":6.0, "3":1.5 },
    "5": { "5":10.0, "4":2.0, "3":0.4 },
    "6": { "6":25.0, "5":2.0, "4":0.4 }
  },
  "mods": {
    "goblin_power": { "1":0.84, "2":0.747, "3":0.707 },
    "goblin_flex":  { "1":0.80, "2":0.720, "3":0.600 },
    "demon_power":  { "1":1.627,"2":2.40,  "3":2.72  },
    "demon_flex":   { "1":1.60, "2":1.520, "3":1.560 }
  }
}
"""

from __future__ import annotations

import argparse
import csv
from datetime import datetime, timezone
import json
import logging
import re
import sys
import unicodedata
from pathlib import Path
from typing import Any, Dict, List, Optional, Set, Tuple

import numpy as np
import pandas as pd
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

ROOT = Path(__file__).resolve().parent.parent
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from utils.goblin_demon_multiplier import (  # noqa: E402
    leg_delta_pct as gd_leg_delta_pct,
    leg_factor as gd_leg_factor,
    leg_payout_method as gd_leg_payout_method,
    load_params as gd_load_params,
    ticket_multiplier as gd_ticket_multiplier,
)

_log = logging.getLogger("combined_ticket_grader")


# -----------------------------
# Defaults
# -----------------------------
POWER_BASE = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 37.5}
FLEX_BASE = {
    2: {2: 3.0},
    3: {3: 3.0, 2: 1.0},
    4: {4: 6.0, 3: 1.5},
    5: {5: 10.0, 4: 2.0, 3: 0.4},
    6: {6: 25.0, 5: 2.0, 4: 0.4},
}

# Modifiers by deviation bucket (dev=1 closest to standard, dev=3 furthest)
GOBLIN_POWER = {1: 0.840, 2: 0.747, 3: 0.707}
GOBLIN_FLEX  = {1: 0.800, 2: 0.720, 3: 0.600}
DEMON_POWER  = {1: 1.627, 2: 2.400, 3: 2.720}
DEMON_FLEX   = {1: 1.600, 2: 1.520, 3: 1.560}


# -----------------------------
# Normalization helpers
# -----------------------------
def strip_norm(s: str) -> str:
    s = "" if s is None else str(s)
    s = unicodedata.normalize("NFKD", s)
    s = "".join(ch for ch in s if not unicodedata.combining(ch))
    s = s.strip().lower()
    s = re.sub(r"\s+", " ", s)
    return s


def player_norm(s: str) -> str:
    """
    Normalize player names for cross-source matching.
    Examples: "Jabari Smith" ~= "Jabari Smith Jr."
    """
    p = strip_norm(s)
    p = p.replace(".", " ")
    p = re.sub(r"\s+", " ", p).strip()
    parts = [x for x in p.split(" ") if x]
    suffixes = {"jr", "sr", "ii", "iii", "iv", "v"}
    parts = [x for x in parts if x not in suffixes]
    return " ".join(parts)


def pick_category_from_cell(pick_type: str) -> str:
    s = strip_norm(pick_type or "")
    if "goblin" in s:
        return "Goblin"
    if "demon" in s:
        return "Demon"
    return "Standard"


def nhl_player_aliases(player: str) -> List[str]:
    """
    Generate NHL-friendly name aliases for matching ticket names against
    API sources that often abbreviate first names (e.g., "Timo Meier" vs "T. Meier").
    """
    p = strip_norm(player)
    if not p:
        return []
    parts = [x for x in p.split(" ") if x]
    aliases = [p]
    if len(parts) >= 2:
        first = parts[0]
        last = " ".join(parts[1:])
        aliases.append(f"{first[:1]}. {last}".strip())
        aliases.append(f"{first[:1]} {last}".strip())
    # preserve order, unique only
    seen = set()
    out = []
    for a in aliases:
        if a and a not in seen:
            seen.add(a)
            out.append(a)
    return out


def prop_norm_from_label(prop: str) -> str:
    p = strip_norm(prop)
    # NHL normalization (ticket labels vs actuals CSV wording differ)
    if "shots on goal" in p or p == "shots_on_goal":
        return "shots_on_goal"
    if "faceoff" in p and "won" in p:
        return "faceoffs_won"

    # Soccer normalization (ticket tokens vs actuals CSV wording differ)
    if "goalie saves" in p or "goalkeeper saves" in p:
        return "goalie saves"
    if "shots on target (combo)" in p:
        return "shots on target"
    if "shots on target" in p:
        return "shots on target"
    if p == "shots":
        # Soccer actuals appear to use a generic "shots" label.
        return "shots on target"

    if "pts+reb+ast" in p or p == "pra":
        return "pra"
    if "pts+reb" in p or p == "pr":
        return "pr"
    if "pts+ast" in p or p == "pa":
        return "pa"
    if "reb+ast" in p or p == "ra":
        return "ra"
    if "points" in p or p == "pts":
        return "points"
    if "rebounds" in p or p == "reb":
        return "rebounds"
    if "assists" in p or p == "ast":
        return "assists"
    if "turnover" in p or p == "tov":
        return "turnovers"
    if "blocked" in p or p == "blk":
        return "blocks"
    if "steal" in p or p == "stl":
        return "steals"
    if "fantasy" in p:
        return "fantasy"
    if any(x in p for x in ["3pm", "3-pointers made", "3 pointers made", "3pt made", "threes made"]):
        return "3pm"
    if any(x in p for x in ["3pa", "3-point attempts", "3 pointers attempted", "3pt att", "threes attempted"]):
        return "3pa"
    if "field goal attempts" in p or p == "fga":
        return "fga"
    if "field goals made" in p or p == "fgm":
        return "fgm"
    if "free throw attempts" in p or p == "fta":
        return "fta"
    if "free throws made" in p or p == "ftm":
        return "ftm"
    return p


def prop_norm_from_actual(prop_type: str) -> str:
    return prop_norm_from_label(prop_type)


# -----------------------------
# Ticket parsing
# -----------------------------
TICKET_RE = re.compile(r"ticket\s*#\s*(\d+)", re.IGNORECASE)

def derive_leg_type(pick_type_cell: str) -> str:
    """
    Convert workbook pick_type into a dev-bucketed leg type.
    Accepts:
      - "Standard"
      - "Goblin" -> "Goblin -1" (default)
      - "Demon"  -> "Demon +1"  (default)
      - "Goblin -2" / "Demon +3" (pass through)
    """
    s = (pick_type_cell or "").strip()
    if not s:
        return "Standard"
    s_norm = strip_norm(s)
    if "goblin" in s_norm:
        # if dev already included like "-2"
        m = re.search(r"-(\d+)", s_norm)
        dev = int(m.group(1)) if m else 1
        dev = max(1, min(3, dev))
        return f"Goblin -{dev}"
    if "demon" in s_norm:
        m = re.search(r"\+(\d+)", s_norm)
        dev = int(m.group(1)) if m else 1
        dev = max(1, min(3, dev))
        return f"Demon +{dev}"
    return "Standard"


def parse_ticket_sheet(tickets_xlsx: Path, sheet_name: str) -> pd.DataFrame:
    df = pd.read_excel(tickets_xlsx, sheet_name=sheet_name, dtype=object)
    if df.empty:
        return pd.DataFrame()

    # If ticket header is in the column name (not a data row), inject it as row 0
    first_col = df.columns[0]
    if isinstance(first_col, str) and TICKET_RE.search(first_col):
        header_row = pd.DataFrame([[first_col] + [None] * (len(df.columns) - 1)], columns=df.columns)
        df = pd.concat([header_row, df], ignore_index=True)
    df = df.copy()
    df.rename(columns={first_col: "ticket_header"}, inplace=True)

    # Column positions in Excel can change across writers/exports, so we key off
    # header names instead of hard-coded indices. This fixes missing `sport`
    # (and therefore missing NBA/Soccer aggregates).
    def _norm_col(c) -> str:
        return re.sub(r"\s+", " ", str(c)).strip().lower()

    col_idx = {_norm_col(c): i for i, c in enumerate(df.columns)}

    # Expected headers from combined_slate_tickets.py
    idx_player = col_idx.get("player", 1)
    idx_team = col_idx.get("team", 2)
    idx_prop = col_idx.get("prop", 4)
    idx_line = col_idx.get("line", 6)
    idx_dir = col_idx.get("dir", 7)
    idx_pick_type = col_idx.get("pick type", 5)
    idx_def_tier = col_idx.get("def tier", 15)
    idx_sport = col_idx.get("sport", 22)
    idx_ml_prob = col_idx.get("ml prob", None)

    rows = []
    i = 0
    while i < len(df):
        hdr = df.at[i, "ticket_header"]
        ticket_no = None
        if isinstance(hdr, str):
            m = TICKET_RE.search(hdr)
            if m:
                ticket_no = int(m.group(1))
        if ticket_no is None:
            i += 1
            continue

        # Find header row (2nd column == "Player")
        j = i + 1
        while j < len(df):
            c1 = df.iloc[j, idx_player] if df.shape[1] > idx_player else None
            if isinstance(c1, str) and strip_norm(c1) == "player":
                break
            nxt = df.at[j, "ticket_header"]
            if isinstance(nxt, str) and TICKET_RE.search(nxt):
                break
            j += 1

        if j >= len(df) or not (
            isinstance(df.iloc[j, idx_player], str) and strip_norm(df.iloc[j, idx_player]) == "player"
        ):
            i += 1
            continue

        # Map label row -> column index (Sport moved to col 25+ when H2H/CV columns exist; old default 22 was B2B).
        hdr_map: dict[str, int] = {}
        for ci in range(df.shape[1]):
            cell = df.iloc[j, ci]
            if isinstance(cell, str) and strip_norm(cell):
                hdr_map[strip_norm(cell)] = ci

        def _col(*names: str, fallback: int) -> int:
            for n in names:
                k0 = strip_norm(n)
                if k0 in hdr_map:
                    return hdr_map[k0]
            return fallback

        idx_player = _col("player", fallback=idx_player)
        idx_team = _col("team", fallback=idx_team)
        idx_prop = _col("prop", fallback=idx_prop)
        idx_line = _col("line", fallback=idx_line)
        idx_dir = _col("dir", fallback=idx_dir)
        idx_pick_type = _col("pick type", fallback=idx_pick_type)
        idx_def_tier = _col("def tier", fallback=idx_def_tier)
        idx_sport = _col("sport", fallback=idx_sport)
        _mp = _col("ml prob", fallback=-1)
        if _mp >= 0:
            idx_ml_prob = _mp
        idx_std_line = _col("standard line", "std line", fallback=-1)

        k = j + 1
        leg_no = 0
        while k < len(df):
            player = df.iloc[k, idx_player] if df.shape[1] > idx_player else None
            if pd.isna(player) or strip_norm(player) == "":
                break
            if isinstance(player, str) and strip_norm(player) == "player":
                k += 1
                continue

            team = df.iloc[k, idx_team] if df.shape[1] > idx_team else ""
            prop = df.iloc[k, idx_prop] if df.shape[1] > idx_prop else ""
            line = df.iloc[k, idx_line] if df.shape[1] > idx_line else np.nan
            direction = df.iloc[k, idx_dir] if df.shape[1] > idx_dir else ""
            pick_type = df.iloc[k, idx_pick_type] if df.shape[1] > idx_pick_type else ""
            tier = df.iloc[k, idx_def_tier] if df.shape[1] > idx_def_tier else ""
            sport = df.iloc[k, idx_sport] if df.shape[1] > idx_sport else ""
            ml_prob = df.iloc[k, idx_ml_prob] if (idx_ml_prob is not None and df.shape[1] > idx_ml_prob) else np.nan
            std_line = np.nan
            if idx_std_line >= 0 and df.shape[1] > idx_std_line:
                std_line = pd.to_numeric(df.iloc[k, idx_std_line], errors="coerce")

            line_num = pd.to_numeric(line, errors="coerce")
            if pd.isna(line_num):
                k += 1
                continue

            leg_no += 1
            leg_type = derive_leg_type("" if pd.isna(pick_type) else str(pick_type))

            rows.append({
                "sheet": sheet_name,
                "ticket_no": ticket_no,
                "leg_no": leg_no,
                "player": str(player),
                "team": str(team) if not pd.isna(team) else "",
                "prop": str(prop) if not pd.isna(prop) else "",
                "prop_norm": prop_norm_from_label(prop),
                "line": float(line_num),
                "dir": str(direction).strip().upper() if not pd.isna(direction) else "",
                "sport": str(sport).strip().upper() if not pd.isna(sport) else "",
                "pick_type": str(pick_type) if not pd.isna(pick_type) else "",
                "leg_type": leg_type,
                "tier": str(tier) if not pd.isna(tier) else "",
                "ml_prob": pd.to_numeric(ml_prob, errors="coerce"),
                "standard_line": float(std_line) if pd.notna(std_line) else np.nan,
            })
            k += 1

        i = k + 1

    return pd.DataFrame(rows)


def parse_tickets_from_combined_json(path: Path) -> pd.DataFrame:
    """
    Load legs from combined_slate_tickets_YYYY-MM-DD.json (same payload as tickets_latest.json).
    Used when the dated .xlsx is missing (e.g. OneDrive copy failed) but the JSON snapshot exists.
    """
    payload = json.loads(path.read_text(encoding="utf-8"))
    rows: list[dict[str, Any]] = []
    for grp in payload.get("groups") or []:
        sheet = str(grp.get("group_name") or "Tickets")
        for t in grp.get("tickets") or []:
            try:
                ticket_no = int(t.get("ticket_no"))
            except (TypeError, ValueError):
                continue
            leg_no = 0
            for leg in t.get("legs") or []:
                leg_no += 1
                prop = str(leg.get("prop_type") or "")
                line_num = pd.to_numeric(leg.get("line"), errors="coerce")
                if pd.isna(line_num):
                    continue
                pick_type = str(leg.get("pick_type") or "")
                tier = leg.get("min_tier") or leg.get("tier") or ""
                std_ln = pd.to_numeric(leg.get("standard_line"), errors="coerce")
                rows.append(
                    {
                        "sheet": sheet,
                        "ticket_no": ticket_no,
                        "leg_no": leg_no,
                        "player": str(leg.get("player") or ""),
                        "team": str(leg.get("team") or ""),
                        "prop": prop,
                        "prop_norm": prop_norm_from_label(prop),
                        "line": float(line_num),
                        "dir": str(leg.get("direction") or "").strip().upper(),
                        "sport": str(leg.get("sport") or "").strip().upper(),
                        "pick_type": pick_type,
                        "leg_type": derive_leg_type(pick_type),
                        "tier": str(tier) if tier is not None else "",
                        "ml_prob": pd.to_numeric(leg.get("ml_prob"), errors="coerce"),
                        "standard_line": float(std_ln) if pd.notna(std_ln) else np.nan,
                    }
                )
    return pd.DataFrame(rows)


# -----------------------------
# Actuals loading + lookup
# -----------------------------
def prep_actuals(csv_path: Path, sport_label: str) -> pd.DataFrame:
    df = pd.read_csv(csv_path, dtype=str).fillna("")
    required = {"player", "team", "prop_type", "actual"}
    missing = required - set(df.columns)
    if missing:
        raise RuntimeError(f"{sport_label} actuals missing columns: {sorted(missing)}. Found: {list(df.columns)}")

    df["actual"] = pd.to_numeric(df["actual"], errors="coerce")
    df = df.dropna(subset=["actual"]).copy()

    df["player_norm"] = df["player"].map(player_norm)
    df["team_norm"] = df["team"].map(strip_norm)
    df["prop_norm"] = df["prop_type"].map(prop_norm_from_actual)
    return df


def _append_grade_latency_row(
    repo_root: Path, date_str: str, sport: str, actuals_path: Path, grade_ts: datetime
) -> None:
    out_dir = repo_root / "outputs"
    out_dir.mkdir(parents=True, exist_ok=True)
    log_path = out_dir / "grade_latency_log.csv"
    try:
        age_min = round((grade_ts.timestamp() - actuals_path.stat().st_mtime) / 60.0, 1)
    except Exception:
        age_min = ""

    write_header = not log_path.exists()
    with log_path.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(
            f,
            fieldnames=["date", "sport", "grade_timestamp", "actuals_file_age_minutes"],
        )
        if write_header:
            w.writeheader()
        w.writerow(
            {
                "date": date_str,
                "sport": sport,
                "grade_timestamp": grade_ts.isoformat(),
                "actuals_file_age_minutes": age_min,
            }
        )


def build_leg_contribution_report(graded_tickets: list[dict]) -> pd.DataFrame:
    records: list[dict] = []
    for t in graded_tickets:
        # Ticket-level result in this grader is payout_status (WIN/CASH/LOSE/NO_ACTUAL/REFUND)
        if str(t.get("result", "")).upper() not in {"LOSS", "LOSE"}:
            continue
        for leg in (t.get("legs") or []):
            if str(leg.get("leg_result", "")).upper() != "MISS":
                continue
            records.append(
                {
                    "sport": leg.get("sport"),
                    "prop_type": leg.get("prop_type"),
                    "pick_type": leg.get("pick_type"),
                    "ml_prob": leg.get("ml_prob"),
                    "tickets_killed": 1,
                }
            )
    if not records:
        return pd.DataFrame(columns=["sport", "prop_type", "pick_type", "tickets_killed", "avg_ml_prob"])
    df = pd.DataFrame(records)
    df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")
    out = (
        df.groupby(["sport", "prop_type", "pick_type"], dropna=False)
        .agg(tickets_killed=("tickets_killed", "sum"), avg_ml_prob=("ml_prob", "mean"))
        .sort_values(["tickets_killed", "avg_ml_prob"], ascending=[False, False])
        .reset_index()
    )
    out["avg_ml_prob"] = out["avg_ml_prob"].round(4)
    return out


def build_lookup(act: pd.DataFrame):
    by_player_prop: Dict[Tuple[str, str], List[dict]] = {}
    by_player_team_prop: Dict[Tuple[str, str, str], List[dict]] = {}
    for _, r in act.iterrows():
        key1 = (r["player_norm"], r["prop_norm"])
        key2 = (r["player_norm"], r["team_norm"], r["prop_norm"])
        by_player_prop.setdefault(key1, []).append(r.to_dict())
        by_player_team_prop.setdefault(key2, []).append(r.to_dict())
    return by_player_prop, by_player_team_prop


def _load_ticket_json_leg_probs(repo_root: Path, tickets_xlsx: Path) -> dict[tuple[str, int, int], float]:
    """
    Best-effort loader for ml_prob/leg_prob_used from a JSON snapshot produced by combined_slate_tickets.py.
    Returns mapping: (sheet_name/group_name, ticket_no, leg_no) -> ml_prob.
    """
    stem = tickets_xlsx.stem
    m = re.search(r"(\d{4}-\d{2}-\d{2})", stem)
    date_str = m.group(1) if m else ""
    candidates: list[Path] = []
    if date_str:
        candidates.append(repo_root / "outputs" / date_str / f"combined_slate_tickets_{date_str}.json")
        candidates.append(repo_root / f"combined_slate_tickets_{date_str}.json")
    candidates.append(tickets_xlsx.with_suffix(".json"))

    json_path = next((p for p in candidates if p.exists()), None)
    if json_path is None:
        return {}
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return {}

    out: dict[tuple[str, int, int], float] = {}
    for g in (payload.get("groups") or []):
        sheet = str(g.get("group_name") or "")
        for t in (g.get("tickets") or []):
            try:
                tno = int(t.get("ticket_no"))
            except Exception:
                continue
            for i, leg in enumerate((t.get("legs") or []), start=1):
                v = leg.get("ml_prob", None)
                if v is None:
                    v = leg.get("leg_prob_used", None)
                try:
                    fv = float(v)
                except Exception:
                    continue
                if 0.0 < fv < 1.0:
                    out[(sheet, tno, i)] = fv
    return out


def _load_ticket_json_ticket_objectives(repo_root: Path, tickets_xlsx: Path) -> dict[tuple[str, int], float]:
    """(sheet / group_name, ticket_no) -> ticket_objective_score from combined_slate JSON when present."""
    stem = tickets_xlsx.stem
    m = re.search(r"(\d{4}-\d{2}-\d{2})", stem)
    date_str = m.group(1) if m else ""
    candidates: list[Path] = []
    if date_str:
        candidates.append(repo_root / "outputs" / date_str / f"combined_slate_tickets_{date_str}.json")
        candidates.append(repo_root / f"combined_slate_tickets_{date_str}.json")
    candidates.append(tickets_xlsx.with_suffix(".json"))
    json_path = next((p for p in candidates if p.exists()), None)
    if json_path is None:
        return {}
    try:
        payload = json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return {}
    out: dict[tuple[str, int], float] = {}
    for g in (payload.get("groups") or []):
        sheet = str(g.get("group_name") or "")
        for t in (g.get("tickets") or []):
            try:
                tno = int(t.get("ticket_no"))
            except Exception:
                continue
            v = t.get("ticket_objective_score")
            if v is None:
                continue
            try:
                out[(sheet, tno)] = float(v)
            except (TypeError, ValueError):
                pass
    return out


def _load_combined_slate_payload(repo_root: Path, tickets_path: Path) -> Optional[dict]:
    stem = tickets_path.stem
    m = re.search(r"(\d{4}-\d{2}-\d{2})", stem)
    date_str = m.group(1) if m else ""
    candidates: list[Path] = []
    if tickets_path.suffix.lower() == ".json":
        candidates.append(tickets_path)
    if date_str:
        candidates.append(repo_root / "outputs" / date_str / f"combined_slate_tickets_{date_str}.json")
        candidates.append(repo_root / f"combined_slate_tickets_{date_str}.json")
    candidates.append(tickets_path.with_suffix(".json"))
    json_path = next((p for p in candidates if p.exists()), None)
    if json_path is None:
        return None
    try:
        return json.loads(json_path.read_text(encoding="utf-8"))
    except Exception:
        return None


def _standard_line_lookup_from_payload(payload: Optional[dict]) -> dict[tuple[str, int, int], float]:
    out: dict[tuple[str, int, int], float] = {}
    if not payload:
        return out
    for g in (payload.get("groups") or []):
        sheet = str(g.get("group_name") or "")
        for t in (g.get("tickets") or []):
            try:
                tno = int(t.get("ticket_no"))
            except Exception:
                continue
            for i, leg in enumerate((t.get("legs") or []), start=1):
                v = leg.get("standard_line")
                if v is None:
                    continue
                try:
                    fv = float(v)
                except (TypeError, ValueError):
                    continue
                if np.isfinite(fv):
                    out[(sheet, tno, i)] = fv
    return out


def merge_standard_lines_into_legs_df(
    legs_df: pd.DataFrame, lookup: dict[tuple[str, int, int], float]
) -> pd.DataFrame:
    if "standard_line" not in legs_df.columns:
        legs_df = legs_df.copy()
        legs_df["standard_line"] = np.nan
    sl_out: list[float] = []
    for _, r in legs_df.iterrows():
        k = (str(r.get("sheet") or ""), int(r.get("ticket_no")), int(r.get("leg_no")))
        if k in lookup:
            sl_out.append(float(lookup[k]))
        else:
            v = r.get("standard_line")
            try:
                fv = float(v)
                sl_out.append(fv if np.isfinite(fv) else np.nan)
            except (TypeError, ValueError):
                sl_out.append(np.nan)
    out_df = legs_df.copy()
    out_df["standard_line"] = sl_out
    return out_df


def _import_combined_slate_ticket_math():
    import sys

    p = str(ROOT / "scripts")
    if p not in sys.path:
        sys.path.insert(0, p)
    import combined_slate_tickets as cst

    return cst


def _leg_series_for_slate_prob(row: pd.Series) -> pd.Series:
    d = row.to_dict()
    pt = d.get("prop_type")
    if pt is None or (isinstance(pt, float) and np.isnan(pt)) or str(pt).strip() == "":
        d["prop_type"] = d.get("prop_norm") or d.get("prop") or ""
    return pd.Series(d)


def _canonical_ticket_leg_grade(leg_result: str) -> str:
    x = str(leg_result or "").upper().strip()
    if x == "HIT":
        return "HIT"
    if x == "MISS":
        return "MISS"
    if x in ("VOID", "PUSH"):
        return "VOID"
    return "UNGRADED"


def _empirical_ticket_paid_ui(sheet: str, grades: list[str]) -> tuple[Optional[bool], bool]:
    """
    Align with build_ticket_eval /tickets: flex 3+ sheet -> flex cash; else all legs HIT.
    Returns (paid_or_none, include_in_ticket_rate).
    """
    if not grades:
        return None, False
    if any(g == "UNGRADED" for g in grades):
        return None, False
    if all(g == "VOID" for g in grades):
        return None, False
    n = len(grades)
    h = sum(1 for g in grades if g == "HIT")
    m = sum(1 for g in grades if g == "MISS")
    is_flex = n >= 3 and "flex" in str(sheet or "").lower()
    if is_flex:
        paid = m <= 1 and h >= n - 1
    else:
        paid = all(g == "HIT" for g in grades)
    return paid, True


def build_ticket_backtest_dataframe(
    legs_df: pd.DataFrame,
    tickets_xlsx: Path,
    grade_date: str,
    source_workbook: str,
) -> pd.DataFrame:
    """One row per ticket_id: empirical paid (UI rules), modeled objective (same math as generator)."""
    cst = _import_combined_slate_ticket_math()
    json_obj = _load_ticket_json_ticket_objectives(ROOT, tickets_xlsx)
    rows_out: list[dict[str, Any]] = []
    for tid, g in legs_df.groupby("ticket_id"):
        g = g.sort_values("leg_no")
        sheet = str(g.iloc[0]["sheet"])
        tno = int(g.iloc[0]["ticket_no"])
        leg_dicts: list[dict[str, Any]] = []
        grades: list[str] = []
        for _, r in g.iterrows():
            ser = _leg_series_for_slate_prob(r)
            leg_dicts.append(ser.to_dict())
            grades.append(_canonical_ticket_leg_grade(str(r.get("leg_result") or "")))
        paid, incl = _empirical_ticket_paid_ui(sheet, grades)
        leg_probs = [cst._resolve_leg_prob(pd.Series(d)) for d in leg_dicts]
        n = len(leg_dicts)
        penalty = cst._correlation_penalty(leg_dicts)
        ep = cst.win_prob(leg_probs, n) * penalty
        fc = cst.flex_cash_prob(leg_probs) * penalty if n >= 3 else ep
        is_flex = n >= 3 and "flex" in sheet.lower()
        obj = float(fc if is_flex else ep)
        jo = json_obj.get((sheet, tno))
        if jo is None:
            jo = np.nan
        rows_out.append(
            {
                "grade_date": grade_date,
                "source_workbook": source_workbook,
                "ticket_id": tid,
                "sheet": sheet,
                "ticket_no": tno,
                "n_legs": n,
                "empirical_ticket_paid": (1.0 if paid else 0.0) if paid is not None else np.nan,
                "include_in_ticket_rate": incl,
                "modeled_ticket_objective": round(obj, 4),
                "modeled_power_prob": round(float(ep), 4),
                "modeled_flex_cash": round(float(fc), 4),
                "json_ticket_objective_score": jo,
                "slate_is_flex": is_flex,
                "n_hit": sum(1 for x in grades if x == "HIT"),
                "n_miss": sum(1 for x in grades if x == "MISS"),
                "n_void": sum(1 for x in grades if x == "VOID"),
                "n_ungraded": sum(1 for x in grades if x == "UNGRADED"),
            }
        )
    return pd.DataFrame(rows_out)


def enrich_ticket_backtest_with_payouts(ticket_bt: pd.DataFrame, ticket_results: pd.DataFrame) -> pd.DataFrame:
    if ticket_bt.empty or ticket_results.empty or "mode" not in ticket_results.columns:
        return ticket_bt
    out = ticket_bt.copy()
    for mode in ticket_results["mode"].dropna().unique():
        m = str(mode).strip().lower()
        sub = ticket_results[ticket_results["mode"] == mode][
            ["ticket_id", "profit", "is_cash", "payout_status", "payout_method"]
        ].copy()
        sub = sub.rename(
            columns={
                "profit": f"profit_{m}",
                "is_cash": f"is_cash_{m}",
                "payout_status": f"payout_status_{m}",
                "payout_method": f"payout_method_{m}",
            }
        )
        out = out.merge(sub, on="ticket_id", how="left")

    # Left merge: tickets present in only one mode get NaN for the other mode's columns.
    # Treat missing payout_method_* as conservative flat_fallback (matches missing delta path).
    _pm_cols = [c for c in ("payout_method_power", "payout_method_flex") if c in out.columns]
    for _c in _pm_cols:
        out[_c] = out[_c].fillna("flat_fallback")

    _allowed_pm = {"curve", "flat_fallback"}
    _bad_nan = [_c for _c in _pm_cols if out[_c].isna().any()]
    _bad_val = [_c for _c in _pm_cols if not out[_c].astype(str).str.strip().isin(_allowed_pm).all()]
    assert not _bad_nan and not _bad_val, (
        "enrich_ticket_backtest_with_payouts: payout_method_* QA failed after fillna — "
        f"NaN in columns={_bad_nan}, disallowed/blank values in columns={_bad_val}"
    )

    return out


def build_ticket_objective_decile_summary(ticket_bt: pd.DataFrame) -> pd.DataFrame:
    sub = ticket_bt[ticket_bt["include_in_ticket_rate"].astype(bool)].copy()
    sub = sub[pd.to_numeric(sub["empirical_ticket_paid"], errors="coerce").notna()].copy()
    if len(sub) < 5:
        return pd.DataFrame([{"note": f"only {len(sub)} decidable tickets for decile summary"}])
    mo = pd.to_numeric(sub["modeled_ticket_objective"], errors="coerce")
    sub = sub.loc[mo.notna()].copy()
    sub["modeled_ticket_objective"] = mo.loc[sub.index]
    nq = min(10, max(3, len(sub) // 3))
    try:
        sub["bin"] = pd.qcut(sub["modeled_ticket_objective"], q=nq, duplicates="drop")
    except (ValueError, TypeError):
        return pd.DataFrame([{"note": "qcut failed — need more spread in modeled_ticket_objective"}])
    g = (
        sub.groupby("bin", observed=True)
        .agg(
            n=("empirical_ticket_paid", "count"),
            mean_model_obj=("modeled_ticket_objective", "mean"),
            empirical_paid_rate=("empirical_ticket_paid", "mean"),
        )
        .reset_index()
    )
    g["empirical_paid_rate"] = g["empirical_paid_rate"].round(4)
    g["mean_model_obj"] = g["mean_model_obj"].round(4)
    return g


def export_graded_tickets_for_ml_training(ticket_bt: pd.DataFrame, out_csv: Path, *, append: bool = False) -> None:
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    if append and out_csv.exists():
        ticket_bt.to_csv(out_csv, mode="a", header=False, index=False)
    else:
        ticket_bt.to_csv(out_csv, index=False)


def lookup_actual(sport: str, player: str, team: str, prop_norm: str,
                  nba_lpt, nba_lp,
                  cbb_lpt, cbb_lp,
                  nba1h_lpt=None, nba1h_lp=None,
                  nba1q_lpt=None, nba1q_lp=None,
                  nhl_lpt=None, nhl_lp=None,
                  soccer_lpt=None, soccer_lp=None) -> float:
    sport = (sport or "").upper()
    if sport == "WCBB":
        sport = "CBB"
    player_n = player_norm(player)
    team_n = strip_norm(team)

    if sport == "NBA1H":
        if nba1h_lpt is not None and nba1h_lp is not None:
            key2 = (player_n, team_n, prop_norm)
            if key2 in nba1h_lpt:
                return float(nba1h_lpt[key2][0]["actual"])
            key1 = (player_n, prop_norm)
            if key1 in nba1h_lp:
                return float(nba1h_lp[key1][0]["actual"])
        # Fallback to full-game NBA when period actuals are unavailable.
        sport = "NBA"

    if sport == "NBA1Q":
        if nba1q_lpt is not None and nba1q_lp is not None:
            key2 = (player_n, team_n, prop_norm)
            if key2 in nba1q_lpt:
                return float(nba1q_lpt[key2][0]["actual"])
            key1 = (player_n, prop_norm)
            if key1 in nba1q_lp:
                return float(nba1q_lp[key1][0]["actual"])
        # Fallback to full-game NBA when period actuals are unavailable.
        sport = "NBA"

    if sport == "NBA":
        key2 = (player_n, team_n, prop_norm)
        if key2 in nba_lpt:
            return float(nba_lpt[key2][0]["actual"])
        key1 = (player_n, prop_norm)
        if key1 in nba_lp:
            return float(nba_lp[key1][0]["actual"])
        return np.nan

    if sport == "CBB":
        # FIX 5: For CBB, team in ticket is an abbreviation (e.g. "COLO") but actuals
        # may use a different format. Try player+prop first (most reliable), then
        # team-keyed as secondary. This is the reverse of NBA priority for CBB.
        key1 = (player_n, prop_norm)
        if key1 in cbb_lp:
            return float(cbb_lp[key1][0]["actual"])
        key2 = (player_n, team_n, prop_norm)
        if key2 in cbb_lpt:
            return float(cbb_lpt[key2][0]["actual"])
        return np.nan

    if sport == "NHL":
        if nhl_lpt is None or nhl_lp is None:
            return np.nan
        # NHL feeds often abbreviate first names in actuals (e.g. "T. Meier"),
        # while tickets contain full names (e.g. "Timo Meier"), so try aliases.
        for pn in nhl_player_aliases(player):
            key2 = (pn, team_n, prop_norm)
            if key2 in nhl_lpt:
                return float(nhl_lpt[key2][0]["actual"])
            key1 = (pn, prop_norm)
            if key1 in nhl_lp:
                return float(nhl_lp[key1][0]["actual"])
        return np.nan

    if sport == "SOCCER":
        if soccer_lpt is None or soccer_lp is None:
            return np.nan
        key2 = (player_n, team_n, prop_norm)
        if key2 in soccer_lpt:
            return float(soccer_lpt[key2][0]["actual"])
        key1 = (player_n, prop_norm)
        if key1 in soccer_lp:
            return float(soccer_lp[key1][0]["actual"])
        return np.nan

    return np.nan


# -----------------------------
# Grading + payout modifiers
# -----------------------------
def grade_leg(dir_: str, line: float, actual: float) -> str:
    if pd.isna(actual):
        return "NO_ACTUAL"
    if abs(actual - line) < 1e-9:
        return "PUSH"
    d = (dir_ or "").upper()
    if d == "OVER":
        return "HIT" if actual > line else "MISS"
    if d == "UNDER":
        return "HIT" if actual < line else "MISS"
    return "UNKNOWN_DIR"


def leg_modifiers(leg_types: List[str]) -> Tuple[float, float]:
    """
    Returns (power_mod, flex_mod) computed as product of per-leg modifiers.
    """
    power_mod = 1.0
    flex_mod = 1.0
    for lt in leg_types:
        s = strip_norm(lt)
        if s.startswith("goblin"):
            m = re.search(r"-(\d+)", s)
            dev = int(m.group(1)) if m else 1
            dev = max(1, min(3, dev))
            power_mod *= float(GOBLIN_POWER.get(dev, 0.84))
            flex_mod  *= float(GOBLIN_FLEX.get(dev, 0.80))
        elif s.startswith("demon"):
            m = re.search(r"\+(\d+)", s)
            dev = int(m.group(1)) if m else 1
            dev = max(1, min(3, dev))
            power_mod *= float(DEMON_POWER.get(dev, 1.627))
            flex_mod  *= float(DEMON_FLEX.get(dev, 1.60))
        else:
            # Standard
            pass
    return power_mod, flex_mod


def compute_ticket_payout(stake: float, mode: str,
                          legs: int, hits: int, misses: int, pushes: int, no_actual: int,
                          voids: int,
                          power_mod: float, flex_mod: float,
                          power_mult_override: Optional[float] = None,
                          flex_mult_override: Optional[float] = None) -> Tuple[float, str, float]:
    """
    Returns (payout_amount, payout_status, applied_multiplier).
    payout_amount includes stake (total returned). profit = payout - stake.
    Optional overrides replace base*mod with an explicit multiplier (curve-based estimate).
    """
    if no_actual > 0:
        return np.nan, "NO_ACTUAL", np.nan

    # Voided / DNP legs drop off the ticket (same bookkeeping as push for leg count).
    effective_legs = legs - pushes - int(voids or 0)
    if effective_legs <= 1:
        return stake, "REFUND", 1.0

    if mode == "power":
        if misses == 0 and int(hits or 0) == effective_legs:
            base = float(POWER_BASE.get(effective_legs, 0.0))
            if power_mult_override is not None:
                mult = float(round(float(power_mult_override), 4))
            else:
                mult = float(round(base * power_mod, 4))
            payout = stake * mult
            return payout, "WIN" if mult > 0 else "WIN_NO_MULT", mult
        return 0.0, "LOSE", 0.0

    if mode == "flex":
        base_table = FLEX_BASE.get(effective_legs, {})
        base = float(base_table.get(hits, 0.0))
        if flex_mult_override is not None:
            mult = float(round(float(flex_mult_override), 4)) if base > 0 else 0.0
        else:
            mult = float(round(base * flex_mod, 4)) if base > 0 else 0.0
        if mult == 0.0:
            return 0.0, "LOSE", 0.0
        return stake * mult, "CASH", mult

    raise ValueError(f"Unknown mode: {mode}")


# -----------------------------
# Analytics helpers
# -----------------------------
def pivot_roi(df: pd.DataFrame, group_cols: List[str], prefix: str) -> pd.DataFrame:
    g = (df
         .groupby(group_cols, dropna=False, as_index=False)
         .agg(
            tickets=("ticket_id", "nunique"),
            staked=("stake", "sum"),
            payout=("payout", "sum"),
            profit=("profit", "sum"),
            win_rate=("is_win", "mean"),
            cash_rate=("is_cash", "mean"),
            no_actual=("no_actual", "sum"),
         ))
    g["roi"] = np.where(g["staked"] > 0, g["profit"] / g["staked"], np.nan)
    g["win_rate"] = g["win_rate"].round(4)
    g["cash_rate"] = g["cash_rate"].round(4)
    g["roi"] = g["roi"].round(4)
    g = g.sort_values(["profit", "roi"], ascending=False).reset_index(drop=True)
    g.insert(0, "view", prefix)
    return g


def build_summary_kv(overall: dict) -> pd.DataFrame:
    return pd.DataFrame([{"metric": k, "value": v} for k, v in overall.items()])


# --- Extra breakdown tables -------------------------------------------------

def _leg_segment_hit_rate(legs_df: pd.DataFrame, group_cols: List[str]) -> pd.DataFrame:
    b = legs_df[legs_df["leg_result"].isin(["HIT", "MISS"])].copy()
    if b.empty:
        return pd.DataFrame(columns=group_cols + ["graded_legs", "hits", "leg_hit_rate"])
    b["is_hit"] = (b["leg_result"] == "HIT").astype(int)
    g = (
        b.groupby(group_cols, dropna=False, as_index=False)
        .agg(graded_legs=("is_hit", "count"), hits=("is_hit", "sum"), leg_hit_rate=("is_hit", "mean"))
    )
    g["leg_hit_rate"] = g["leg_hit_rate"].round(4)
    g["misses"] = g["graded_legs"] - g["hits"]
    return g.sort_values("graded_legs", ascending=False).reset_index(drop=True)


def build_leg_breakdown_tables(legs_df: pd.DataFrame) -> Dict[str, pd.DataFrame]:
    df = legs_df.copy()
    df["pick_cat"] = df["pick_type"].astype(str).map(pick_category_from_cell)
    out: Dict[str, pd.DataFrame] = {
        "LEG_BY_PROP_SPORT": _leg_segment_hit_rate(df, ["sport", "prop_norm"]),
        "LEG_BY_PROP": _leg_segment_hit_rate(df, ["prop_norm"]),
        "LEG_BY_SPORT": _leg_segment_hit_rate(df, ["sport"]),
        "LEG_BY_DIR": _leg_segment_hit_rate(df, ["dir"]),
        "LEG_BY_PICK_CAT": _leg_segment_hit_rate(df, ["pick_cat"]),
        "LEG_BY_SHEET": _leg_segment_hit_rate(df, ["sheet"]),
        "LEG_BY_SPORT_PICK": _leg_segment_hit_rate(df, ["sport", "pick_cat"]),
    }
    noa = legs_df["leg_result"].eq("NO_ACTUAL").groupby(legs_df["sport"]).sum().reset_index(name="no_actual_legs")
    noa = noa.sort_values("no_actual_legs", ascending=False)
    out["LEG_NO_ACTUAL_BY_SPORT"] = noa
    return out


def build_ticket_deep_dive(ticket_results: pd.DataFrame, legs_df: pd.DataFrame) -> pd.DataFrame:
    """One row per ticket_id + mode with compact leg outcome string and mix stats."""
    leg_order = legs_df.sort_values(["ticket_id", "leg_no"])
    pieces = []
    for tid, g in leg_order.groupby("ticket_id"):
        props = "|".join(g["prop_norm"].astype(str).head(6).tolist())
        if len(g) > 6:
            props += "|..."
        res = "".join(
            {"HIT": "W", "MISS": "L", "PUSH": "P", "VOID": "V", "NO_ACTUAL": "?"}.get(x, "?")
            for x in g["leg_result"].tolist()
        )
        pieces.append(
            {
                "ticket_id": tid,
                "sheet": g["sheet"].iloc[0],
                "ticket_no": int(g["ticket_no"].iloc[0]) if pd.notna(g["ticket_no"].iloc[0]) else "",
                "n_legs": len(g),
                "n_hit": int((g["leg_result"] == "HIT").sum()),
                "n_miss": int((g["leg_result"] == "MISS").sum()),
                "n_push": int((g["leg_result"] == "PUSH").sum()),
                "n_void": int((g["leg_result"] == "VOID").sum()),
                "n_no_actual": int((g["leg_result"] == "NO_ACTUAL").sum()),
                "sports_mix": ",".join(sorted({str(s) for s in g["sport"].tolist() if str(s).strip()})),
                "pick_mix": ",".join(sorted({pick_category_from_cell(x) for x in g["pick_type"].tolist()})),
                "result_chain": res,
                "props_sample": props,
            }
        )
    deep = pd.DataFrame(pieces)
    if deep.empty:
        return deep
    tr = ticket_results.copy()
    deep_m = deep.drop(columns=[c for c in ("sheet", "ticket_no") if c in deep.columns], errors="ignore")
    return tr.merge(deep_m, on="ticket_id", how="left")


def build_analysis_insights(
    overall: dict,
    ticket_results: pd.DataFrame,
    leg_breakdowns: Dict[str, pd.DataFrame],
) -> pd.DataFrame:
    rows: List[Dict[str, str]] = []
    for mode in ("power", "flex"):
        roi_k = f"{mode}_roi"
        if roi_k not in overall:
            continue
        rows.append(
            {
                "topic": f"{mode.upper()} summary",
                "detail": (
                    f"ROI {overall.get(roi_k, '')} | profit ${overall.get(mode + '_profit', '')} "
                    f"on ${overall.get(mode + '_staked', '')} staked | "
                    f"win/cash rate {overall.get(mode + '_win_rate', '')} / {overall.get(mode + '_cash_rate', '')}"
                ),
            }
        )

    sub = ticket_results[ticket_results["payout_status"] != "NO_ACTUAL"].copy()
    if not sub.empty:
        best = (
            sub.groupby(["mode", "sheet"], as_index=False)["profit"]
            .sum()
            .sort_values(["mode", "profit"], ascending=[True, False])
        )
        for mode in best["mode"].unique():
            top = best[best["mode"] == mode].head(3)
            for _, r in top.iterrows():
                rows.append(
                    {
                        "topic": f"Top ticket sheets ({mode})",
                        "detail": f"{r['sheet']}: total profit ${float(r['profit']):.2f}",
                    }
                )

    by_prop = leg_breakdowns.get("LEG_BY_PROP_SPORT", pd.DataFrame())
    if by_prop is not None and not by_prop.empty and "graded_legs" in by_prop.columns:
        enough = by_prop[by_prop["graded_legs"] >= 8].copy()
        if not enough.empty:
            worst = enough.nsmallest(3, "leg_hit_rate")
            best = enough.nlargest(3, "leg_hit_rate")
            for _, r in worst.iterrows():
                rows.append(
                    {
                        "topic": "Leg segments to review (low hit%, n>=8)",
                        "detail": f"{r['sport']} {r['prop_norm']}: hit rate {float(r['leg_hit_rate']):.1%} "
                        f"({int(r['hits'])}/{int(r['graded_legs'])})",
                    }
                )
            for _, r in best.iterrows():
                rows.append(
                    {
                        "topic": "Leg segments that hit (n>=8)",
                        "detail": f"{r['sport']} {r['prop_norm']}: hit rate {float(r['leg_hit_rate']):.1%}",
                    }
                )

    rows.append(
        {
            "topic": "How to use ML tabs",
            "detail": (
                "ML_* uses RandomForest on graded legs (HIT vs MISS) for sport/prop/direction/pick mix. "
                "ML_CALIBRATION compares bucketed ml_prob to realized hit rate (fix over/under-confidence in step8). "
                "ML_FILTER_SIM_* compares ROI if you only kept tickets with above-median RF leg strength — "
                "same-day only; stack --export-graded-legs-csv across dates before changing production rules."
            ),
        }
    )
    return pd.DataFrame(rows)


# --- Excel styling ----------------------------------------------------------

_HDR_FILL = PatternFill(start_color="1C2833", end_color="1C2833", fill_type="solid")
_HDR_FONT = Font(color="FFFFFF", bold=True, size=10)
_THIN = Side(style="thin", color="CCCCCC")
_CELL_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)
_HIT_FILL = PatternFill(start_color="27AE60", end_color="27AE60", fill_type="solid")
_MISS_FILL = PatternFill(start_color="C0392B", end_color="C0392B", fill_type="solid")
_PUSH_FILL = PatternFill(start_color="F39C12", end_color="F39C12", fill_type="solid")
_NA_FILL = PatternFill(start_color="BDC3C7", end_color="BDC3C7", fill_type="solid")
_VOID_FILL = PatternFill(start_color="AED6F1", end_color="AED6F1", fill_type="solid")
_PROFIT_POS = PatternFill(start_color="D5F5E3", end_color="D5F5E3", fill_type="solid")
_PROFIT_NEG = PatternFill(start_color="FADBD8", end_color="FADBD8", fill_type="solid")


def _header_cell(ws, row: int, col: int):
    cell = ws.cell(row=row, column=col)
    cell.fill = _HDR_FILL
    cell.font = _HDR_FONT
    cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    cell.border = _CELL_BORDER


def _col_index_by_header(ws, name: str, header_row: int = 1) -> Optional[int]:
    for c in range(1, ws.max_column + 1):
        v = ws.cell(row=header_row, column=c).value
        if v is not None and str(v).strip() == str(name).strip():
            return c
    return None


def apply_graded_workbook_styles(wb) -> None:
    """Bold headers, freeze panes, borders, column widths, conditional leg/profit colors."""
    for ws in wb.worksheets:
        if ws.max_row < 1 or ws.max_column < 1:
            continue
        ws.freeze_panes = "A2"
        ws.row_dimensions[1].height = 28

        for c in range(1, ws.max_column + 1):
            _header_cell(ws, 1, c)
            letter = get_column_letter(c)
            maxlen = len(str(ws.cell(1, c).value or ""))
            for r in range(2, min(ws.max_row + 1, 500)):
                maxlen = max(maxlen, len(str(ws.cell(r, c).value or "")))
            ws.column_dimensions[letter].width = float(min(max(maxlen + 2, 9), 52))

        for r in range(2, ws.max_row + 1):
            for c in range(1, ws.max_column + 1):
                ws.cell(r, c).border = _CELL_BORDER
                ws.cell(r, c).alignment = Alignment(vertical="center", wrap_text=False)

        title = ws.title
        if title == "TICKET_RESULTS":
            pc = _col_index_by_header(ws, "profit")
            if pc:
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(r, pc)
                    try:
                        v = float(cell.value)
                    except (TypeError, ValueError):
                        continue
                    cell.number_format = "#,##0.00"
                    if v > 0:
                        cell.fill = _PROFIT_POS
                    elif v < 0:
                        cell.fill = _PROFIT_NEG
        elif title == "LEG_RESULTS":
            lc = _col_index_by_header(ws, "leg_result")
            if lc:
                for r in range(2, ws.max_row + 1):
                    cell = ws.cell(r, lc)
                    lr = str(cell.value or "").upper().strip()
                    if lr == "HIT":
                        cell.fill = _HIT_FILL
                        cell.font = Font(color="FFFFFF", bold=True, size=10)
                    elif lr == "MISS":
                        cell.fill = _MISS_FILL
                        cell.font = Font(color="FFFFFF", bold=True, size=10)
                    elif lr == "PUSH":
                        cell.fill = _PUSH_FILL
                        cell.font = Font(color="000000", bold=True, size=10)
                    elif lr == "NO_ACTUAL":
                        cell.fill = _NA_FILL
        elif title == "ANALYSIS_INSIGHTS":
            for r in range(2, ws.max_row + 1):
                ws.cell(r, 1).alignment = Alignment(wrap_text=True, vertical="top")
                ws.cell(r, 2).alignment = Alignment(wrap_text=True, vertical="top")
            ws.column_dimensions["A"].width = 28
            ws.column_dimensions["B"].width = 90


# --- ML training export (feedback loop) --------------------------------------

def export_graded_legs_for_ml_training(
    legs_df: pd.DataFrame,
    out_csv: Path,
    *,
    grade_date: str,
    source_workbook: str,
    append: bool = False,
    ticket_bt: Optional[pd.DataFrame] = None,
) -> None:
    """
    Write graded legs (+ outcomes) for offline model improvement.
    Stack daily files into one CSV or Parquet, then retrain the pipeline that produces ml_prob on step8.
    When ticket_bt is set, merges ticket-level paid / modeled objective onto each leg row.
    """
    export = legs_df.copy()
    if ticket_bt is not None and not ticket_bt.empty:
        merge_cols = [
            "ticket_id",
            "empirical_ticket_paid",
            "include_in_ticket_rate",
            "modeled_ticket_objective",
            "modeled_power_prob",
            "modeled_flex_cash",
            "slate_is_flex",
            "json_ticket_objective_score",
        ]
        use = [c for c in merge_cols if c in ticket_bt.columns]
        export = export.merge(
            ticket_bt[use].drop_duplicates(subset=["ticket_id"]),
            on="ticket_id",
            how="left",
        )
    export.insert(0, "grade_date", str(grade_date))
    export.insert(1, "source_workbook", str(source_workbook))
    out_csv = Path(out_csv)
    out_csv.parent.mkdir(parents=True, exist_ok=True)
    if append and out_csv.exists():
        export.to_csv(out_csv, mode="a", header=False, index=False)
    else:
        export.to_csv(out_csv, index=False)


def _ml_prob_calibration_bins(legs_df: pd.DataFrame) -> pd.DataFrame:
    """Empirical hit rate by quantile of slate ml_prob (HIT/MISS legs only)."""
    sub = legs_df[legs_df["leg_result"].isin(["HIT", "MISS"])].copy()
    if sub.empty:
        return pd.DataFrame([{"note": "no HIT/MISS legs"}])
    mp = pd.to_numeric(sub.get("ml_prob"), errors="coerce")
    sub = sub.assign(_mp=mp)
    sub = sub[sub["_mp"].notna() & (sub["_mp"] > 0.0) & (sub["_mp"] < 1.0)]
    if len(sub) < 15:
        return pd.DataFrame(
            [{"note": f"only {len(sub)} legs with ml_prob in (0,1); need ~15+ for bins"}]
        )
    sub["hit"] = (sub["leg_result"] == "HIT").astype(int)
    nq = min(10, max(3, len(sub) // 5))
    try:
        sub["bin"] = pd.qcut(sub["_mp"], q=nq, duplicates="drop")
    except ValueError:
        return pd.DataFrame([{"note": "could not bin ml_prob (try more graded legs)"}])
    g = (
        sub.groupby("bin", observed=True)
        .agg(n=("hit", "count"), pred_mean=("_mp", "mean"), hit_rate=("hit", "mean"))
        .reset_index()
    )
    g["calibration_error"] = (g["hit_rate"] - g["pred_mean"]).round(4)
    g["pred_mean"] = g["pred_mean"].round(4)
    g["hit_rate"] = g["hit_rate"].round(4)
    return g


# --- ML (exploratory) -------------------------------------------------------

def run_ml_profit_layers(
    legs_df: pd.DataFrame,
    ticket_results: pd.DataFrame,
    min_graded_legs: int = 40,
    random_state: int = 42,
) -> Dict[str, Any]:
    """
    Train a leg-level hit classifier and simulate ticket filtering by mean predicted leg hit prob.
    In-sample on this workbook only — for discovery, not production edge.
    """
    out: Dict[str, Any] = {"sheets": {}, "meta": []}
    try:
        from sklearn.ensemble import RandomForestClassifier
        from sklearn.model_selection import cross_val_score
    except ImportError:
        out["sheets"]["ML_SKIPPED"] = pd.DataFrame(
            [{"reason": "scikit-learn not installed (pip install scikit-learn)"}]
        )
        return out

    train = legs_df[legs_df["leg_result"].isin(["HIT", "MISS"])].copy()
    if len(train) < min_graded_legs:
        out["sheets"]["ML_SKIPPED"] = pd.DataFrame(
            [{"reason": f"Only {len(train)} graded legs (need {min_graded_legs}+ for stable ML)"}]
        )
        return out

    train["pick_cat"] = train["pick_type"].astype(str).map(pick_category_from_cell)
    top_props = train["prop_norm"].value_counts().head(14).index.tolist()
    train["prop_bucket"] = train["prop_norm"].where(train["prop_norm"].isin(top_props), "other")

    feat_cols = ["sport", "prop_bucket", "dir", "pick_cat"]
    X = pd.get_dummies(train[feat_cols].fillna(""), drop_first=False)
    y = (train["leg_result"] == "HIT").astype(int).values

    n_splits = min(5, max(2, len(train) // 25))
    rf = RandomForestClassifier(
        n_estimators=120,
        max_depth=10,
        min_samples_leaf=4,
        random_state=random_state,
        class_weight="balanced_subsample",
        n_jobs=1,
    )
    rf.fit(X.values, y)
    cv_acc = cross_val_score(rf, X.values, y, cv=n_splits, scoring="accuracy")
    try:
        cv_auc = cross_val_score(rf, X.values, y, cv=n_splits, scoring="roc_auc")
    except ValueError:
        cv_auc = np.array([np.nan])

    imp = pd.DataFrame({"feature": X.columns, "importance": rf.feature_importances_})
    imp = imp.sort_values("importance", ascending=False).reset_index(drop=True)
    imp["importance_pct"] = (imp["importance"] / imp["importance"].sum() * 100).round(2)
    out["sheets"]["ML_FEATURE_IMPORTANCE"] = imp

    out["sheets"]["ML_MODEL_META"] = pd.DataFrame(
        [
            {"metric": "cv_accuracy_mean", "value": round(float(np.mean(cv_acc)), 4)},
            {"metric": "cv_accuracy_std", "value": round(float(np.std(cv_acc)), 4)},
            {"metric": "cv_roc_auc_mean", "value": round(float(np.mean(cv_auc)), 4)},
            {"metric": "graded_legs", "value": len(train)},
            {"metric": "n_splits", "value": n_splits},
            {
                "metric": "note",
                "value": "Trained on this slate only; use rolling history for real policy changes.",
            },
            {
                "metric": "feedback_loop",
                "value": "Use --export-graded-legs-csv to stack labels; tune step8 ml_prob using ML_CALIBRATION sheet.",
            },
        ]
    )

    # Predict per leg, aggregate to ticket
    full = legs_df.copy()
    full["pick_cat"] = full["pick_type"].astype(str).map(pick_category_from_cell)
    full["prop_bucket"] = full["prop_norm"].where(full["prop_norm"].isin(top_props), "other")
    X_full = pd.get_dummies(full[feat_cols].fillna(""), drop_first=False)
    X_full = X_full.reindex(columns=X.columns, fill_value=0)
    full["ml_p_hit"] = rf.predict_proba(X_full.values)[:, 1]

    ticket_strength = full.groupby("ticket_id", as_index=False).agg(
        ml_avg_leg_hit_prob=("ml_p_hit", "mean"),
        ml_min_leg_hit_prob=("ml_p_hit", "min"),
    )
    tr_enriched = ticket_results.merge(ticket_strength, on="ticket_id", how="left")

    sim_rows = []
    for mode in tr_enriched["mode"].dropna().unique():
        m = tr_enriched[tr_enriched["mode"] == mode].copy()
        el = m[m["payout_status"] != "NO_ACTUAL"]
        if el.empty or el["ml_avg_leg_hit_prob"].isna().all():
            continue
        med = el["ml_avg_leg_hit_prob"].median()
        base_profit = float(el["profit"].sum())
        base_stake = float(el["stake"].sum())
        filt = el[el["ml_avg_leg_hit_prob"] >= med]
        sim_rows.append(
            {
                "mode": mode,
                "filter": f"ml_avg_leg_hit_prob >= median ({med:.3f})",
                "tickets_all": int(len(el)),
                "tickets_kept": int(len(filt)),
                "staked_all": base_stake,
                "staked_filtered": float(filt["stake"].sum()),
                "profit_all": base_profit,
                "profit_filtered": float(filt["profit"].sum()),
                "roi_all": round(base_profit / base_stake, 4) if base_stake > 0 else 0.0,
                "roi_filtered": round(float(filt["profit"].sum()) / float(filt["stake"].sum()), 4)
                if float(filt["stake"].sum()) > 0
                else 0.0,
            }
        )
    out["sheets"]["ML_FILTER_SIM"] = pd.DataFrame(sim_rows)

    # Actionable-ish: which one-hot features associate with higher hit rate in raw data
    lift_rows = []
    base_rate = float(y.mean())
    for col in X.columns:
        mask = X[col].values.astype(bool)
        if mask.sum() < 8:
            continue
        hr = float(y[mask].mean())
        lift_rows.append(
            {
                "segment": col,
                "n_legs": int(mask.sum()),
                "hit_rate": round(hr, 4),
                "lift_vs_baseline": round((hr - base_rate) / max(base_rate, 0.01), 4),
            }
        )
    lift_df = pd.DataFrame(lift_rows).sort_values("lift_vs_baseline", ascending=False)
    out["sheets"]["ML_SEGMENT_LIFT"] = lift_df.head(40).reset_index(drop=True)

    return out


# -----------------------------
# Main
# -----------------------------
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--tickets",
        required=True,
        help="combined_slate_tickets_YYYY-MM-DD.xlsx or .json (JSON when xlsx missing)",
    )
    ap.add_argument("--nba_actuals", required=True, help="actuals_nba_YYYY-MM-DD.csv")
    ap.add_argument("--nba1h_actuals", default="", help="actuals_nba1h_YYYY-MM-DD.csv (optional)")
    ap.add_argument("--nba1q_actuals", default="", help="actuals_nba1q_YYYY-MM-DD.csv (optional)")
    ap.add_argument("--cbb_actuals", required=True, help="actuals_cbb_YYYY-MM-DD.csv")
    ap.add_argument("--nhl_actuals", default="", help="actuals_nhl_YYYY-MM-DD.csv (optional, for NHL legs)")
    ap.add_argument("--soccer_actuals", default="", help="actuals_soccer_YYYY-MM-DD.csv (optional, for Soccer legs)")
    ap.add_argument("--nba_injuries", default="", help="injuries_nba CSV (optional; defaults next to nba actuals)")
    ap.add_argument("--cbb_injuries", default="", help="injuries_cbb CSV (optional)")
    ap.add_argument("--nhl_injuries", default="", help="injuries_nhl CSV (optional)")
    ap.add_argument("--soccer_injuries", default="", help="injuries_soccer CSV (optional; manual / future fetch)")
    ap.add_argument("--out", default="", help="Output graded workbook (default: <tickets>_GRADED.xlsx)")
    ap.add_argument("--mode", choices=["power", "flex", "both"], default="both")
    ap.add_argument("--stake", type=float, default=20.0)
    ap.add_argument("--payouts_json", default="", help="Optional JSON override for base payouts + modifiers")
    ap.add_argument("--no-ml", action="store_true", help="Skip ML analysis sheets (faster)")
    ap.add_argument("--ml-min-legs", type=int, default=40, help="Min graded legs to run ML (default 40)")
    ap.add_argument(
        "--export-graded-legs-csv",
        default="",
        help="Write LEG_RESULTS-style rows+outcomes for offline retraining (path to .csv)",
    )
    ap.add_argument(
        "--append-graded-legs-csv",
        action="store_true",
        help="Append to --export-graded-legs-csv instead of overwriting (same columns required)",
    )
    ap.add_argument(
        "--export-graded-tickets-csv",
        default="",
        help="Ticket-level TICKET_BACKTEST rows for stacking / backtest_ticket_objectives.py",
    )
    ap.add_argument(
        "--append-graded-tickets-csv",
        action="store_true",
        help="Append to --export-graded-tickets-csv instead of overwriting",
    )
    args = ap.parse_args()

    global POWER_BASE, FLEX_BASE, GOBLIN_POWER, GOBLIN_FLEX, DEMON_POWER, DEMON_FLEX

    if args.payouts_json:
        cfg = json.loads(Path(args.payouts_json).read_text(encoding="utf-8"))
        if "power_base" in cfg:
            POWER_BASE = {int(k): float(v) for k, v in cfg["power_base"].items()}
        if "flex_base" in cfg:
            FLEX_BASE = {int(n): {int(k): float(v) for k, v in tab.items()} for n, tab in cfg["flex_base"].items()}
        mods = cfg.get("mods", {})
        if "goblin_power" in mods:
            GOBLIN_POWER = {int(k): float(v) for k, v in mods["goblin_power"].items()}
        if "goblin_flex" in mods:
            GOBLIN_FLEX = {int(k): float(v) for k, v in mods["goblin_flex"].items()}
        if "demon_power" in mods:
            DEMON_POWER = {int(k): float(v) for k, v in mods["demon_power"].items()}
        if "demon_flex" in mods:
            DEMON_FLEX = {int(k): float(v) for k, v in mods["demon_flex"].items()}

    tickets_path = Path(args.tickets)
    nba_csv = Path(args.nba_actuals)
    cbb_csv = Path(args.cbb_actuals)
    if args.out:
        out_xlsx = Path(args.out)
    else:
        out_xlsx = tickets_path.with_name(tickets_path.stem + "_GRADED.xlsx")

    # actuals + lookups
    nba_act = prep_actuals(nba_csv, "NBA")
    cbb_act = prep_actuals(cbb_csv, "CBB")
    nba_lp, nba_lpt = build_lookup(nba_act)
    cbb_lp, cbb_lpt = build_lookup(cbb_act)
    nba1h_lp = nba1h_lpt = None
    nba1q_lp = nba1q_lpt = None
    if args.nba1h_actuals:
        nba1h_csv = Path(args.nba1h_actuals)
        nba1h_act = prep_actuals(nba1h_csv, "NBA1H")
        nba1h_lp, nba1h_lpt = build_lookup(nba1h_act)
    if args.nba1q_actuals:
        nba1q_csv = Path(args.nba1q_actuals)
        nba1q_act = prep_actuals(nba1q_csv, "NBA1Q")
        nba1q_lp, nba1q_lpt = build_lookup(nba1q_act)

    nhl_lp = nhl_lpt = None
    soccer_lp = soccer_lpt = None
    if args.nhl_actuals:
        nhl_csv = Path(args.nhl_actuals)
        nhl_act = prep_actuals(nhl_csv, "NHL")
        nhl_lp, nhl_lpt = build_lookup(nhl_act)
    if args.soccer_actuals:
        soccer_csv = Path(args.soccer_actuals)
        soccer_act = prep_actuals(soccer_csv, "SOCCER")
        soccer_lp, soccer_lpt = build_lookup(soccer_act)

    # Grade latency tracker (one row per actuals file)
    grade_ts = datetime.now(timezone.utc)
    m = re.search(r"_(\d{4}-\d{2}-\d{2})", str(nba_csv.name))
    grade_date = m.group(1) if m else grade_ts.strftime("%Y-%m-%d")
    _append_grade_latency_row(ROOT, grade_date, "NBA", nba_csv, grade_ts)
    _append_grade_latency_row(ROOT, grade_date, "CBB", cbb_csv, grade_ts)
    if args.nba1h_actuals:
        _append_grade_latency_row(ROOT, grade_date, "NBA1H", Path(args.nba1h_actuals), grade_ts)
    if args.nba1q_actuals:
        _append_grade_latency_row(ROOT, grade_date, "NBA1Q", Path(args.nba1q_actuals), grade_ts)
    if args.nhl_actuals:
        _append_grade_latency_row(ROOT, grade_date, "NHL", Path(args.nhl_actuals), grade_ts)
    if args.soccer_actuals:
        _append_grade_latency_row(ROOT, grade_date, "SOCCER", Path(args.soccer_actuals), grade_ts)

    from espn_injuries import (  # noqa: E402
        canon_team_abbr as _inj_canon_team,
        injuries_csv_path_for_actuals,
        load_injury_void_keys,
    )

    def _inj_csv(explicit: str, act_csv: Path, sport: str) -> Path:
        if explicit and str(explicit).strip():
            return Path(str(explicit).strip())
        return injuries_csv_path_for_actuals(act_csv, sport)

    nba_void = load_injury_void_keys(_inj_csv(args.nba_injuries, nba_csv, "NBA"), "NBA")
    cbb_void = load_injury_void_keys(_inj_csv(args.cbb_injuries, cbb_csv, "CBB"), "CBB")
    nhl_void: Set[Tuple[str, str]] = set()
    if args.nhl_actuals:
        nhl_void = load_injury_void_keys(_inj_csv(args.nhl_injuries, Path(args.nhl_actuals), "NHL"), "NHL")
    soccer_void: Set[Tuple[str, str]] = set()
    if args.soccer_actuals:
        soccer_void = load_injury_void_keys(
            _inj_csv(args.soccer_injuries, Path(args.soccer_actuals), "SOCCER"), "SOCCER"
        )

    # ticket sheets (workbook) or JSON snapshot (same legs as --write-web output)
    tickets_xlsx = tickets_path
    if tickets_path.suffix.lower() == ".json":
        legs_df = parse_tickets_from_combined_json(tickets_path)
        if legs_df.empty:
            raise RuntimeError("No ticket legs parsed from JSON. Check combined_slate_tickets payload.")
    else:
        xls = pd.ExcelFile(tickets_path)
        ticket_sheets = [s for s in xls.sheet_names if re.search(r"\b\d-?Leg\b", s, re.IGNORECASE)]

        leg_frames = []
        for s in ticket_sheets:
            legs = parse_ticket_sheet(tickets_path, s)
            if not legs.empty:
                leg_frames.append(legs)
        if not leg_frames:
            raise RuntimeError("No ticket legs parsed. Check sheet format.")

        legs_df = pd.concat(leg_frames, ignore_index=True)
    legs_df["ticket_id"] = legs_df["sheet"].astype(str) + " | " + legs_df["ticket_no"].astype(str)

    # grade legs
    legs_df["actual"] = legs_df.apply(
        lambda r: lookup_actual(
            r["sport"], r["player"], r["team"], r["prop_norm"],
            nba_lpt, nba_lp,
            cbb_lpt, cbb_lp,
            nba1h_lpt=nba1h_lpt, nba1h_lp=nba1h_lp,
            nba1q_lpt=nba1q_lpt, nba1q_lp=nba1q_lp,
            nhl_lpt=nhl_lpt, nhl_lp=nhl_lp,
            soccer_lpt=soccer_lpt, soccer_lp=soccer_lp
        ),
        axis=1,
    )
    legs_df["leg_result"] = legs_df.apply(lambda r: grade_leg(r["dir"], r["line"], r["actual"]), axis=1)

    def _injury_void_leg(row: pd.Series) -> str:
        if row["leg_result"] != "NO_ACTUAL":
            return row["leg_result"]
        sp = str(row["sport"] or "").upper().strip()
        pl = strip_norm(row["player"])
        if sp in ("NBA", "NBA1H", "NBA1Q"):
            tm = _inj_canon_team("NBA", row["team"])
            if pl and tm and (pl, tm) in nba_void:
                return "VOID"
        elif sp in ("CBB", "WCBB"):
            tm = _inj_canon_team("CBB", row["team"])
            if pl and tm and (pl, tm) in cbb_void:
                return "VOID"
        elif sp == "NHL":
            tm = _inj_canon_team("NHL", row["team"])
            if pl and tm and (pl, tm) in nhl_void:
                return "VOID"
        elif sp == "SOCCER":
            tm = _inj_canon_team("SOCCER", row["team"])
            if pl and tm and (pl, tm) in soccer_void:
                return "VOID"
        return row["leg_result"]

    legs_df["leg_result"] = legs_df.apply(_injury_void_leg, axis=1)

    # Enrich leg-level ML probability from JSON snapshot (when not present in XLSX).
    # This keeps Killer Legs avg_ml_prob meaningful even if the workbook layout changes.
    if "ml_prob" not in legs_df.columns:
        legs_df["ml_prob"] = np.nan
    need_fill = pd.to_numeric(legs_df["ml_prob"], errors="coerce").isna()
    if need_fill.any():
        prob_map = _load_ticket_json_leg_probs(ROOT, tickets_xlsx)
        if prob_map:
            def _lookup_prob(r):
                return prob_map.get((str(r.get("sheet") or ""), int(r.get("ticket_no")), int(r.get("leg_no"))), np.nan)
            legs_df.loc[need_fill, "ml_prob"] = legs_df.loc[need_fill].apply(_lookup_prob, axis=1)

    _slate_payload = _load_combined_slate_payload(ROOT, tickets_path)
    _sl_lookup = _standard_line_lookup_from_payload(_slate_payload)
    if _sl_lookup:
        legs_df = merge_standard_lines_into_legs_df(legs_df, _sl_lookup)
    elif "standard_line" not in legs_df.columns:
        legs_df["standard_line"] = np.nan

    _gd_params = gd_load_params()

    def _row_delta_pct(r: pd.Series) -> float:
        v = gd_leg_delta_pct(r.get("line"), r.get("standard_line"))
        return float(v) if v is not None else np.nan

    legs_df["delta_pct"] = legs_df.apply(_row_delta_pct, axis=1)

    def _row_gd_factor(r: pd.Series) -> float:
        v = r.get("delta_pct")
        if v is None or (isinstance(v, float) and np.isnan(v)):
            dp = None
        else:
            dp = float(v)
        return float(gd_leg_factor(dp, str(r.get("pick_type", "")), _gd_params))

    legs_df["gd_leg_factor"] = legs_df.apply(_row_gd_factor, axis=1)

    legs_df["payout_method"] = legs_df.apply(
        lambda row: gd_leg_payout_method(row.get("delta_pct"), str(row.get("pick_type", ""))),
        axis=1,
    )

    # per-ticket modifiers
    mods_df = (legs_df.groupby("ticket_id", as_index=False)
               .agg(
                    leg_types=("leg_type", lambda s: list(s)),
                    sports=("sport", lambda s: ",".join(sorted(set([x for x in s if str(x).strip()])))),
                    pick_types=("pick_type", lambda s: ",".join(sorted(set([x for x in s if str(x).strip()])))),
                    tiers=("tier", lambda s: ",".join(sorted(set([x for x in s if str(x).strip()])))),
                ))
    mods_df[["power_mod", "flex_mod"]] = mods_df["leg_types"].apply(lambda L: pd.Series(leg_modifiers(L)))

    # ticket base stats
    ticket_base = (legs_df
        .groupby(["sheet", "ticket_no", "ticket_id"], as_index=False)
        .agg(
            legs=("leg_no", "max"),
            hits=("leg_result", lambda s: int((s == "HIT").sum())),
            misses=("leg_result", lambda s: int((s == "MISS").sum())),
            pushes=("leg_result", lambda s: int((s == "PUSH").sum())),
            voids=("leg_result", lambda s: int((s == "VOID").sum())),
            no_actual=("leg_result", lambda s: int((s == "NO_ACTUAL").sum())),
        ))
    ticket_base["effective_legs"] = ticket_base["legs"] - ticket_base["pushes"] - ticket_base["voids"]
    ticket_base["stake"] = float(args.stake)
    ticket_base = ticket_base.merge(mods_df[["ticket_id", "sports", "pick_types", "tiers", "power_mod", "flex_mod"]], on="ticket_id", how="left")

    legs_by_ticket_pre = {k: g.sort_values("leg_no") for k, g in legs_df.groupby("ticket_id")}

    modes = ["power", "flex"] if args.mode == "both" else [args.mode]
    ticket_rows = []
    for mode in modes:
        t = ticket_base.copy()
        payouts_out = []
        statuses = []
        mults = []
        est_curve_mults: List[float] = []
        flat_standard_mults: List[float] = []
        mult_delta_std: List[float] = []
        payouts_est: List[float] = []
        ticket_payout_methods: List[str] = []
        for _, r in t.iterrows():
            tid = str(r["ticket_id"])
            g_legs = legs_by_ticket_pre.get(tid)
            n_eff = int(r["effective_legs"])
            stake = float(r["stake"])
            p_prm = gd_load_params()
            factors: List[float] = []
            leg_methods: List[str] = []
            if g_legs is not None and n_eff >= 2:
                for _, lr in g_legs.iterrows():
                    if str(lr.get("leg_result", "")).upper() == "VOID":
                        continue
                    leg_methods.append(str(lr.get("payout_method", "curve")))
                    dp = gd_leg_delta_pct(lr["line"], lr.get("standard_line"))
                    factors.append(
                        gd_leg_factor(dp, pick_category_from_cell(str(lr.get("pick_type", ""))), p_prm)
                    )
                while len(factors) > n_eff:
                    factors.pop()
                while len(factors) < n_eff:
                    factors.append(1.0)
            else:
                factors = [1.0] * max(0, n_eff)

            ticket_payout_methods.append(
                "flat_fallback" if leg_methods and "flat_fallback" in leg_methods else "curve"
            )

            if n_eff >= 2:
                if mode == "power":
                    est_m = float(gd_ticket_multiplier(n_eff, factors[:n_eff], "power"))
                    flat_std = float(POWER_BASE.get(n_eff, 0.0))
                else:
                    hi = int(r["hits"])
                    est_m = float(gd_ticket_multiplier(n_eff, factors[:n_eff], "flex", hits=hi))
                    flat_std = float(FLEX_BASE.get(n_eff, {}).get(hi, 0.0))
            else:
                est_m = 1.0 if mode == "power" else 0.0
                flat_std = 1.0 if mode == "power" else 0.0

            payout_amt, status, mult = compute_ticket_payout(
                stake=stake,
                mode=mode,
                legs=int(r["legs"]),
                hits=int(r["hits"]),
                misses=int(r["misses"]),
                pushes=int(r["pushes"]),
                no_actual=int(r["no_actual"]),
                voids=int(r["voids"]),
                power_mod=float(r["power_mod"]),
                flex_mod=float(r["flex_mod"]),
            )
            payouts_out.append(payout_amt)
            statuses.append(status)
            mults.append(mult)

            payout_e = float(payout_amt) if pd.notna(payout_amt) else np.nan
            if mode == "power" and status in ("WIN", "WIN_NO_MULT") and np.isfinite(est_m):
                payout_e = stake * est_m
            elif mode == "flex" and status == "CASH" and np.isfinite(est_m):
                payout_e = stake * est_m

            if (
                status in ("WIN", "WIN_NO_MULT", "CASH")
                and np.isfinite(mult)
                and np.isfinite(est_m)
                and float(mult) > 0
                and abs(float(mult) - est_m) / float(mult) > 0.15
            ):
                _log.warning(
                    "Payout mult mismatch ticket=%s mode=%s legacy_applied=%.4f est_curve=%.4f (>15%%)",
                    tid,
                    mode,
                    float(mult),
                    est_m,
                )

            est_curve_mults.append(est_m)
            flat_standard_mults.append(flat_std)
            mult_delta_std.append(round(est_m - flat_std, 4))
            payouts_est.append(payout_e)

        t["mode"] = mode
        t["payout_status"] = statuses
        t["applied_mult"] = mults
        t["payout"] = payouts_out
        t["profit"] = t["payout"] - t["stake"]
        t["est_curve_mult"] = est_curve_mults
        t["flat_standard_mult"] = flat_standard_mults
        t["mult_delta_vs_standard"] = mult_delta_std
        t["payout_est_curve"] = payouts_est
        t["profit_est_curve"] = t["payout_est_curve"] - t["stake"]
        t["payout_method"] = ticket_payout_methods
        t["is_win"] = ((t["payout_status"] == "WIN") | (t["payout_status"] == "WIN_NO_MULT")).astype(int)
        t["is_cash"] = ((t["payout_status"] == "WIN") | (t["payout_status"] == "WIN_NO_MULT") | (t["payout_status"] == "CASH")).astype(int)
        ticket_rows.append(t)

    ticket_results = pd.concat(ticket_rows, ignore_index=True)

    ticket_bt_df = build_ticket_backtest_dataframe(
        legs_df, tickets_xlsx, str(grade_date), tickets_xlsx.name
    )
    ticket_bt_df = enrich_ticket_backtest_with_payouts(ticket_bt_df, ticket_results)
    ticket_obj_deciles = build_ticket_objective_decile_summary(ticket_bt_df)

    # overall stats
    overall = {}
    for mode in modes:
        sub = ticket_results[ticket_results["mode"] == mode].copy()
        eligible = sub[sub["payout_status"] != "NO_ACTUAL"]
        overall[f"{mode}_tickets"] = int(sub["ticket_id"].nunique())
        overall[f"{mode}_eligible_tickets"] = int(eligible["ticket_id"].nunique())
        overall[f"{mode}_no_actual_tickets"] = int((sub["payout_status"] == "NO_ACTUAL").sum())
        overall[f"{mode}_staked"] = float(eligible["stake"].sum())
        overall[f"{mode}_payout"] = float(eligible["payout"].sum())
        overall[f"{mode}_profit"] = float(eligible["profit"].sum())
        overall[f"{mode}_roi"] = round(float(eligible["profit"].sum() / eligible["stake"].sum()) if eligible["stake"].sum() > 0 else 0.0, 4)
        overall[f"{mode}_win_rate"] = round(float(eligible["is_win"].mean()) if len(eligible) else 0.0, 4)
        overall[f"{mode}_cash_rate"] = round(float(eligible["is_cash"].mean()) if len(eligible) else 0.0, 4)
        pec = pd.to_numeric(eligible["profit_est_curve"], errors="coerce").fillna(0.0)
        stk = float(eligible["stake"].sum())
        overall[f"{mode}_profit_est_curve"] = float(pec.sum())
        overall[f"{mode}_roi_est_curve"] = round(float(pec.sum() / stk) if stk > 0 else 0.0, 4)

    tables = {}
    for mode in modes:
        sub = ticket_results[ticket_results["mode"] == mode].copy()
        eligible = sub[sub["payout_status"] != "NO_ACTUAL"].copy()
        tables[f"{mode}_BY_SHEET"] = pivot_roi(eligible, ["sheet"], f"{mode}_BY_SHEET")
        tables[f"{mode}_BY_LEGS"] = pivot_roi(eligible, ["effective_legs"], f"{mode}_BY_LEGS")
        tables[f"{mode}_BY_SPORTS"] = pivot_roi(eligible, ["sports"], f"{mode}_BY_SPORTS")
        tables[f"{mode}_BY_PICK_TYPES"] = pivot_roi(eligible, ["pick_types"], f"{mode}_BY_PICK_TYPES")

    summary_kv = build_summary_kv(overall)

    leg_breakdowns = build_leg_breakdown_tables(legs_df)
    insights_df = build_analysis_insights(overall, ticket_results, leg_breakdowns)
    deep_df = build_ticket_deep_dive(ticket_results, legs_df)

    # Killer legs report (loss drivers)
    graded_tickets: list[dict] = []
    legs_by_ticket = {k: g for k, g in legs_df.groupby("ticket_id")}
    for _, tr in ticket_results.iterrows():
        tid = tr.get("ticket_id")
        legs = []
        g = legs_by_ticket.get(tid)
        if g is not None:
            for _, lr in g.iterrows():
                legs.append(
                    {
                        "sport": lr.get("sport"),
                        "prop_type": lr.get("prop_norm") or lr.get("prop"),
                        "pick_type": lr.get("pick_type"),
                        "ml_prob": lr.get("ml_prob"),
                        "leg_result": lr.get("leg_result"),
                    }
                )
        graded_tickets.append({"result": tr.get("payout_status"), "legs": legs})
    killer_df = build_leg_contribution_report(graded_tickets)
    ml_calibration_df = _ml_prob_calibration_bins(legs_df)

    ml_pack: Dict[str, Any] = {}
    if not args.no_ml:
        ml_pack = run_ml_profit_layers(
            legs_df,
            ticket_results,
            min_graded_legs=int(args.ml_min_legs),
        )

    pay_acc_a = legs_df.copy()
    pay_acc_a["pick_cat"] = pay_acc_a["pick_type"].astype(str).map(pick_category_from_cell)
    _plc = [
        "ticket_id",
        "leg_no",
        "player",
        "sport",
        "prop_norm",
        "pick_cat",
        "pick_type",
        "standard_line",
        "line",
        "delta_pct",
        "gd_leg_factor",
        "payout_method",
        "leg_result",
    ]
    _plc = [c for c in _plc if c in pay_acc_a.columns]
    pay_acc_legs = pay_acc_a[_plc].copy()

    _ptc = [
        "ticket_id",
        "mode",
        "sheet",
        "ticket_no",
        "applied_mult",
        "est_curve_mult",
        "flat_standard_mult",
        "mult_delta_vs_standard",
        "payout",
        "payout_est_curve",
        "profit",
        "profit_est_curve",
        "stake",
        "payout_status",
        "payout_method",
    ]
    _ptc = [c for c in _ptc if c in ticket_results.columns]
    pay_acc_tickets = ticket_results[_ptc].copy() if _ptc else pd.DataFrame()

    n_delta_known = int(legs_df["delta_pct"].notna().sum()) if "delta_pct" in legs_df.columns else 0
    n_flat_fb = (
        int((legs_df["payout_method"] == "flat_fallback").sum())
        if "payout_method" in legs_df.columns
        else 0
    )
    avg_md = float(pd.to_numeric(ticket_results.get("mult_delta_vs_standard"), errors="coerce").mean())
    summ_rows: List[Dict[str, Any]] = [
        {"metric": "legs_with_delta_pct", "value": n_delta_known},
        {"metric": "legs_payout_flat_fallback", "value": n_flat_fb},
        {"metric": "avg_mult_delta_vs_standard", "value": round(avg_md, 4)},
    ]
    for _m in modes:
        summ_rows.append({"metric": f"{_m}_flat_roi", "value": overall.get(f"{_m}_roi")})
        summ_rows.append({"metric": f"{_m}_est_curve_roi", "value": overall.get(f"{_m}_roi_est_curve")})
    pay_acc_summary = pd.DataFrame(summ_rows)

    if not pay_acc_tickets.empty and "applied_mult" in pay_acc_tickets.columns and "est_curve_mult" in pay_acc_tickets.columns:
        _dff = (
            pd.to_numeric(pay_acc_tickets["applied_mult"], errors="coerce")
            - pd.to_numeric(pay_acc_tickets["est_curve_mult"], errors="coerce")
        ).abs() > 1.0
        pay_acc_warnings = pay_acc_tickets[_dff.fillna(False)].copy()
    else:
        pay_acc_warnings = pd.DataFrame()

    with pd.ExcelWriter(out_xlsx, engine="openpyxl") as xw:
        summary_kv.to_excel(xw, index=False, sheet_name="SUMMARY")
        insights_df.to_excel(xw, index=False, sheet_name="ANALYSIS_INSIGHTS")
        ticket_results.to_excel(xw, index=False, sheet_name="TICKET_RESULTS")
        legs_df.to_excel(xw, index=False, sheet_name="LEG_RESULTS")
        ml_calibration_df.to_excel(xw, index=False, sheet_name="ML_CALIBRATION")
        # Always write the sheet (even empty) so it is visible/stable for downstream tooling.
        killer_df.to_excel(xw, index=False, sheet_name="Killer Legs")
        if deep_df is not None and not deep_df.empty:
            deep_df.to_excel(xw, index=False, sheet_name="TICKET_DEEP_DIVE")
        for name, tab in leg_breakdowns.items():
            if tab is not None and not tab.empty:
                tab.to_excel(xw, index=False, sheet_name=str(name)[:31])
        if not args.no_ml:
            for name, tab in ml_pack.get("sheets", {}).items():
                if tab is not None and not tab.empty:
                    tab.to_excel(xw, index=False, sheet_name=str(name)[:31])
        for name, tab in tables.items():
            tab.to_excel(xw, index=False, sheet_name=str(name)[:31])
        ticket_bt_df.to_excel(xw, index=False, sheet_name="TICKET_BACKTEST")
        ticket_obj_deciles.to_excel(xw, index=False, sheet_name="TICKET_OBJ_DECILES")
        pay_acc_legs.to_excel(xw, index=False, sheet_name="PAYOUT_LEG_DETAIL")
        pay_acc_tickets.to_excel(xw, index=False, sheet_name="PAYOUT_TICKET_DETAIL")
        pay_acc_summary.to_excel(xw, index=False, sheet_name="PAYOUT_ACCURACY")
        if not pay_acc_warnings.empty:
            pay_acc_warnings.to_excel(xw, index=False, sheet_name="PAYOUT_WARNINGS")
        apply_graded_workbook_styles(xw.book)

    # Keep this ASCII-only so the script runs in non-UTF8 consoles.
    print(f"Wrote graded workbook -> {out_xlsx}")
    for _m in modes:
        print(
            f"  {_m.upper()} ROI: flat(legacy)={overall.get(_m + '_roi')} "
            f"est_curve={overall.get(_m + '_roi_est_curve')}"
        )

    ex = str(args.export_graded_legs_csv or "").strip()
    if ex:
        export_graded_legs_for_ml_training(
            legs_df,
            Path(ex),
            grade_date=str(grade_date),
            source_workbook=str(tickets_xlsx.name),
            append=bool(args.append_graded_legs_csv),
            ticket_bt=ticket_bt_df,
        )
        print(f"Wrote graded legs training export -> {ex} (append={bool(args.append_graded_legs_csv)})")

    tx = str(args.export_graded_tickets_csv or "").strip()
    if tx:
        export_graded_tickets_for_ml_training(
            ticket_bt_df,
            Path(tx),
            append=bool(args.append_graded_tickets_csv),
        )
        print(f"Wrote graded tickets export -> {tx} (append={bool(args.append_graded_tickets_csv)})")


if __name__ == "__main__":
    main()
