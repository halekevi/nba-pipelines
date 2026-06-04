#!/usr/bin/env python3
"""
combined_slate_tickets.py

Combined NBA + CBB + NHL + Soccer + Tennis + WNBA Slate & Ticket Generator
Merges NBA (step8_all_direction_clean.xlsx) and CBB (step6_ranked_cbb.xlsx ELIGIBLE)
Outputs:
  - combined_slate_tickets_YYYY-MM-DD.xlsx
  - tickets_latest.json (web; /tickets renders from this). Graded HTML: build_ticket_eval.py → ticket_eval_<date>.html

Cross-book lines (optional):
  Place CSVs next to the combined output, then pass --underdog-csv / --draftkings-csv (or use run_pipeline
  when files exist under outputs/<date>/):
    outputs/<date>/underdog_props.csv   ← fetch_underdog_pickem.py --output ...
    outputs/<date>/draftkings_props_all.csv ← merge_draftkings_pickem_csvs.py (NBA+NHL+MLB+CBB), or
    outputs/<date>/draftkings_props_nba.csv ← fetch_draftkings_player_props.py --league nba -o ...
  Join is on sport + team + normalized player + normalized prop label, and the numeric line must match
  PrizePicks ``line`` within ~0.05. Matched rows populate ``line_underdog`` / ``line_draftkings``; ladder rows
  that do not match any PP line are appended as extra slate rows with ``pick_platform`` underdog/draftkings.
  After merge, each row gets cross-book comparison: best_cross_line / best_cross_book / cross_edge_vs_pp
  (OVER favors lowest line; UNDER favors highest; edge_vs_pp is points vs PrizePicks).

Sheets: SUMMARY, Full Slate (reordered + STRONG/LEAN/RISK + pace beside Def Tier), NBA Slate, CBB Slate,
        2–6-Leg tickets per sport (Goblin / Standard / Std+Gob mix),
        cross-sport Standard / Std+Gob / Goblin mixes,
        plus up to three cross-pipeline slips (max 6 legs each): Standard-only, Goblin-only, Std+Gob mix

NEW (Web):
- Adds player headshot thumbnails when an ID is available:
    NBA: uses nba_player_id (if present) -> cdn.nba.com headshot
    CBB: uses espn_player_id (if present) -> espncdn headshot
  If no ID exists, it falls back to a simple initials avatar.
- JSON includes image_url per leg.
- More helpful file-path resolution (tries script dir + recursive search if file not found)

Ticket modes (defaults favor volume; optional strict mode):
- --high-conviction is OFF by default (--high-conviction for stricter pools).
- Default pool: tiers A,B,C, min hit rate 0.65, per-leg floors LEG_MIN_HIT_RATE; --high-conviction raises floors via max().
- With --high-conviction: pool min hit rate >= 0.65, max 4 legs on FINAL slips; structured 2–3 leg tickets use 0.65 leg floor unless --min-leg-hit-rate set.
- --min-leg-hit-rate / --max-ticket-legs: optional overrides (see argparse help).
- --prioritize-ticket-hit: optional; raises per-leg floors and drops slips below modeled P(payout).
  This maximizes *expected* ticket success — no generator can guarantee a literal 100% hit rate.
- --ticket-candidate-sort: how to rank props when *choosing* legs (default blend = ML prob + rank composite).
  L5-backed hit_rate (when sample ≥3) drives est_win_prob via _resolve_leg_prob; else ML (capped), rank, edge.
- Improve ml_prob over time: run combined_ticket_grader.py with --export-graded-legs-csv (stack slates) and read ML_CALIBRATION in the graded workbook.
- --ticket-gen-starts (default 10): structured slips try K alternative first legs and keep the best modeled ticket payout (flex cash or all-hit prob).

HOTFIX:
- Fixes crash when CBB "direction" becomes a DataFrame due to duplicate columns.
  We de-duplicate columns BEFORE touching df["direction"].str.upper().
"""

from __future__ import annotations

import argparse
import hashlib
import itertools
import json
try:
    import joblib
except Exception:  # optional on web runtime; rerank gracefully disables
    joblib = None
import logging
import math
import os
import re
import sys
import unicodedata
from collections import Counter, defaultdict
from datetime import datetime, timedelta, timezone
from pathlib import Path
from typing import Any, Iterable, List, Optional
from zoneinfo import ZoneInfo

_SLATE_TZ = ZoneInfo("America/New_York")


def slate_calendar_date_ymd() -> str:
    """US Eastern calendar date for outputs/ and tickets JSON (avoids UTC-midnight skew on Railway)."""
    return datetime.now(_SLATE_TZ).date().strftime("%Y-%m-%d")

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from usage_redistribution import apply_usage_redistribution

# Repo root = parent of scripts/ (this file lives in scripts/)
REPO_ROOT = os.path.abspath(os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
if REPO_ROOT not in sys.path:
    sys.path.insert(0, REPO_ROOT)
from utils.defense_tiers import normalize_def_tier_label
from utils.kelly_staking import fractional_kelly, leg_edge_pct_for_kelly
from utils.cbb_tourney_metadata import CBB_AP_TOP25_2026, CBB_TOURNEY_2026
from utils.goblin_demon_multiplier import (
    compute_ticket_ev as gd_compute_ticket_ev,
    leg_delta_pct as gd_leg_delta_pct,
    multiplier_summary as gd_multiplier_summary,
)
from utils.ticket_diversity import apply_diversity_filter
from utils.pipeline_read_enrichment import (
    READ_SLATE_EXPORT_KEYS,
    enrich_read_fields_dataframe,
)
from utils.ticket_ev_tiers import (
    apply_slate_ev_tier_recommendations,
    recommendation_from_ev,
)

_log_slate = logging.getLogger("combined_slate_tickets")


def _norm_def_tier_cell_upper(raw: object) -> str:
    """Canonical def tier label, uppercased for legacy comparisons (maps SOLID → ABOVE AVG, etc.)."""
    base = normalize_def_tier_label(raw)
    if base:
        return base.upper()
    s = str(raw or "").strip()
    return s.upper()
TICKET_MODEL_PATH = os.path.join(REPO_ROOT, "models", "ticket_model.pkl")
TICKET_MODEL_2LEG_PATH = os.path.join(REPO_ROOT, "models", "ticket_model_2leg.pkl")
TICKET_MODEL_3LEG_PATH = os.path.join(REPO_ROOT, "models", "ticket_model_3leg.pkl")
TICKET_MODEL_4PLUS_PATH = os.path.join(REPO_ROOT, "models", "ticket_model_4plus.pkl")
TICKET_MODEL_FEATURES_PATH = os.path.join(REPO_ROOT, "models", "ticket_model_features.json")

_TICKET_MODEL_RERANK_ENABLED = False
_TICKET_MODEL_RERANK_WEIGHT = 0.15
_TICKET_MODEL_RERANK_TOP_N = 5
_TICKET_MODEL_USE_BUCKETS = True
_TICKET_MODEL = None
_TICKET_MODEL_BUCKETS: dict[str, Any] = {}
_TICKET_MODEL_FEATURES: list[str] = []

# Primary layout: outputs/<slate-date>/<sport>/... (canonical runtime),
# plus outputs/<slate-date>/step8_*_<date>.xlsx (dated exports).
# Defaults are applied in apply_default_sport_inputs() after --date is resolved.
DEFAULT_NBA_PATH = os.path.join(REPO_ROOT, "Sports", "NBA", "data", "outputs", "step8_all_direction_clean.xlsx")
DEFAULT_CBB_PATH = os.path.join(REPO_ROOT, "Sports", "CBB", "step6_ranked_cbb.xlsx")
DEFAULT_CFB_PATH = os.path.join(REPO_ROOT, "Sports", "CFB", "step6_ranked_cfb.xlsx")
DEFAULT_NBA1H_PATH = os.path.join(REPO_ROOT, "Sports", "NBA", "step8_nba1h_direction_clean.xlsx")
DEFAULT_NBA1Q_PATH = os.path.join(REPO_ROOT, "Sports", "NBA", "step8_nba1q_direction_clean.xlsx")
DEFAULT_WCBB_PATH = os.path.join(REPO_ROOT, "Sports", "CBB", "step6_ranked_wcbb.xlsx")
DEFAULT_MLB_PATH = os.path.join(REPO_ROOT, "Sports", "MLB", "step8_mlb_direction_clean.xlsx")
DEFAULT_NFL_PATH = os.path.join(REPO_ROOT, "Sports", "NFL", "outputs", "step8_nfl_direction_clean.xlsx")
DISABLED_SPORTS: set[str] = set()
DEFAULT_SOCCER_PATH = os.path.join(REPO_ROOT, "Sports", "Soccer", "outputs", "step8_soccer_direction_clean.xlsx")
DEFAULT_TENNIS_PATH = os.path.join(REPO_ROOT, "Tennis", "outputs", "step8_tennis_direction_clean.xlsx")
DEFAULT_WNBA_PATH = os.path.join(REPO_ROOT, "Sports", "WNBA", "outputs", "step8_wnba_direction_clean.xlsx")
DEFAULT_NHL_PATH = os.path.join(REPO_ROOT, "Sports", "NHL", "outputs", "step8_nhl_direction_clean.xlsx")
DEFAULT_WEB_OUTDIR = os.path.join(REPO_ROOT, "ui_runner", "templates")


def _outputs_dir_for_date(date_str: str) -> str:
    d = str(date_str).strip()[:10]
    return os.path.join(REPO_ROOT, "outputs", d)


def _first_existing_path(*candidates: str) -> str:
    for p in candidates:
        if p and os.path.isfile(p):
            return str(p)
    return ""


def _required_placeholder(*candidates: str) -> str:
    """First existing path, else first non-empty candidate (for clear load errors / validation)."""
    hit = _first_existing_path(*candidates)
    if hit:
        return hit
    for p in candidates:
        if p:
            return str(p)
    return ""


def apply_default_sport_inputs(args: argparse.Namespace) -> None:
    """
    Fill empty --sport paths. Order: outputs/<date>/step8_*_<date>.xlsx, then Sports/*, then legacy repo-root paths.
    Optional sports stay empty when nothing exists on disk.
    """
    d = str(args.date).strip()[:10]
    out = _outputs_dir_for_date(d)

    if not str(args.nba).strip():
        args.nba = _required_placeholder(
            os.path.join(out, "nba", "step8_all_direction_clean.xlsx"),
            os.path.join(out, f"step8_nba_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "NBA", "data", "outputs", "step8_all_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "NBA", "data", "outputs", "step8_all_direction_clean.xlsx"),
        )

    if not str(args.cbb).strip():
        args.cbb = _required_placeholder(
            os.path.join(out, "cbb", "step6_ranked_cbb.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "CBB", "step6_ranked_cbb.xlsx"),
            os.path.join(REPO_ROOT, "CBB", "step6_ranked_cbb.xlsx"),
        )

    if not str(args.wcbb).strip():
        args.wcbb = _first_existing_path(
            os.path.join(REPO_ROOT, "Sports", "CBB", "step6_ranked_wcbb.xlsx"),
            os.path.join(REPO_ROOT, "CBB", "step6_ranked_wcbb.xlsx"),
        )

    if not str(args.nhl).strip():
        args.nhl = _first_existing_path(
            os.path.join(out, "nhl", "step8_nhl_direction_clean.xlsx"),
            os.path.join(out, f"step8_nhl_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "NHL", "outputs", "step8_nhl_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "NHL", "outputs", "step8_nhl_direction_clean.xlsx"),
        )

    if not str(args.soccer).strip():
        args.soccer = _first_existing_path(
            os.path.join(out, "soccer", "step8_soccer_direction_clean.xlsx"),
            os.path.join(out, f"step8_soccer_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "Soccer", "outputs", "step8_soccer_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Soccer", "outputs", "step8_soccer_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Soccer", "step8_soccer_direction_clean.xlsx"),
        )

    if not str(args.tennis).strip():
        tennis_d = str(getattr(args, "tennis_date", None) or "").strip()[:10] or d
        args.tennis = _first_existing_path(
            os.path.join(out, "tennis", f"step8_tennis_direction_clean_{tennis_d}.xlsx"),
            os.path.join(out, f"step8_tennis_direction_clean_{tennis_d}.xlsx"),
            os.path.join(out, "tennis", "step8_tennis_direction_clean.xlsx"),
            os.path.join(out, "step8_tennis_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "Tennis", "outputs", "step8_tennis_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Tennis", "outputs", "step8_tennis_direction_clean.xlsx"),
        )

    if not str(args.wnba).strip():
        args.wnba = _first_existing_path(
            os.path.join(out, "wnba", "step8_wnba_direction_clean.xlsx"),
            os.path.join(out, "wnba", "step8_wnba_direction.xlsx"),
            os.path.join(out, f"step8_wnba_direction_clean_{d}.xlsx"),
            os.path.join(out, f"step8_wnba_direction_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "WNBA", "outputs", "step8_wnba_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "WNBA", "outputs", "step8_wnba_direction.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "WNBA", "step8_wnba_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "WNBA", "step8_wnba_direction.xlsx"),
            os.path.join(REPO_ROOT, "WNBA", "step8_wnba_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "WNBA", "step8_wnba_direction.xlsx"),
        )

    if not str(args.mlb).strip():
        args.mlb = _first_existing_path(
            os.path.join(out, "mlb", "step8_mlb_direction_clean.xlsx"),
            os.path.join(out, f"step8_mlb_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "MLB", "step8_mlb_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "MLB", "step8_mlb_direction_clean.xlsx"),
        )

    if not str(args.nba1q).strip():
        args.nba1q = _first_existing_path(
            os.path.join(out, "nba1q", "step8_nba1q_direction_clean.xlsx"),
            os.path.join(out, f"step8_nba1q_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "NBA", "step8_nba1q_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "NBA", "step8_nba1q_direction_clean.xlsx"),
        )

    if not str(args.nba1h).strip():
        args.nba1h = _first_existing_path(
            os.path.join(out, "nba1h", "step8_nba1h_direction_clean.xlsx"),
            os.path.join(out, f"step8_nba1h_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "NBA", "step8_nba1h_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "NBA", "step8_nba1h_direction_clean.xlsx"),
        )

    if not str(args.nfl).strip():
        args.nfl = _first_existing_path(
            os.path.join(out, "nfl", "step8_nfl_direction_clean.xlsx"),
            os.path.join(out, f"step8_nfl_direction_clean_{d}.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "NFL", "outputs", "step8_nfl_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "NFL", "data", "outputs", "step8_nfl_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "NFL", "outputs", "step8_nfl_direction_clean.xlsx"),
            os.path.join(REPO_ROOT, "NFL", "data", "outputs", "step8_nfl_direction_clean.xlsx"),
        )

    if not str(getattr(args, "cfb", "") or "").strip():
        args.cfb = _first_existing_path(
            os.path.join(out, "cfb", "step6_ranked_cfb.xlsx"),
            os.path.join(REPO_ROOT, "Sports", "CFB", "step6_ranked_cfb.xlsx"),
            os.path.join(REPO_ROOT, "CFB", "step6_ranked_cfb.xlsx"),
        )


def print_combined_slate_input_paths(args: argparse.Namespace) -> None:
    """Echo resolved inputs so missing step8 files are obvious before loading."""

    def _rel(p: str) -> str:
        p = str(p or "").strip()
        if not p:
            return ""
        try:
            return os.path.relpath(p, REPO_ROOT)
        except ValueError:
            return p

    def _line(label: str, path: str, *, optional: bool) -> None:
        p = str(path or "").strip()
        if optional and not p:
            print(f"  {label:<7} (none) [optional — skipped]")
            return
        if not p:
            print(f"  {label:<7} (none) [MISSING]")
            return
        exists = os.path.isfile(p)
        rel = _rel(p)
        if optional:
            tag = "[EXISTS]" if exists else "[MISSING — skipped]"
        else:
            tag = "[EXISTS]" if exists else "[MISSING — expect load warning]"
        print(f"  {label:<7} {rel} {tag}")

    print("[combined_slate] Input paths:")
    _line("NBA", args.nba, optional=False)
    if "CBB" in DISABLED_SPORTS:
        print("  CBB     (season deactivated — skipped)")
    else:
        _line("CBB", args.cbb, optional=False)
    _line("WCBB", args.wcbb, optional=True)
    _line("NBA1Q", args.nba1q, optional=True)
    _line("NBA1H", args.nba1h, optional=True)
    _line("MLB", args.mlb, optional=True)
    _line("NHL", args.nhl, optional=True)
    _line("Soccer", args.soccer, optional=True)
    _line("Tennis", args.tennis, optional=True)
    _line("WNBA", args.wnba, optional=True)
    _line("NFL", args.nfl, optional=True)
    _line("CFB", getattr(args, "cfb", ""), optional=True)


DIVERSITY_CONFIG_PATH = os.path.join(REPO_ROOT, "config", "diversity_config.json")
PROP_RELIABILITY_LATEST_PATH = os.path.join(REPO_ROOT, "data", "reports", "prop_reliability_latest.json")
PROP_STRAT_BOARD_LATEST_PATH = os.path.join(REPO_ROOT, "ui_runner", "data", "prop_stratification_board_latest.json")


def _sanitize_for_json(obj: Any) -> Any:
    """Replace float nan/inf with None so json.dump emits strict JSON null (JavaScript-safe)."""
    if isinstance(obj, dict):
        return {k: _sanitize_for_json(v) for k, v in obj.items()}
    if isinstance(obj, list):
        return [_sanitize_for_json(v) for v in obj]
    if isinstance(obj, float):
        return obj if math.isfinite(obj) else None
    if obj is None or isinstance(obj, (bool, str)):
        return obj
    if isinstance(obj, int):
        return obj
    try:
        xf = float(obj)
    except (TypeError, ValueError):
        return obj
    return xf if math.isfinite(xf) else None


def _write_json_file(path: str, payload: Any) -> None:
    """Atomic JSON write (temp + replace) with short retries for Windows file locks."""
    import time

    data = json.dumps(_sanitize_for_json(payload), ensure_ascii=False, default=str, allow_nan=False)
    directory = os.path.dirname(os.path.abspath(path)) or "."
    os.makedirs(directory, exist_ok=True)
    last_err: Exception | None = None
    for attempt in range(4):
        tmp = os.path.join(directory, f".{os.path.basename(path)}.{os.getpid()}.{attempt}.tmp")
        try:
            with open(tmp, "w", encoding="utf-8", newline="\n") as wf:
                wf.write(data)
            os.replace(tmp, path)
            return
        except OSError as exc:
            last_err = exc
            try:
                if os.path.isfile(tmp):
                    os.remove(tmp)
            except OSError:
                pass
            time.sleep(0.2 * (attempt + 1))
    if last_err is not None:
        raise last_err
    raise OSError(f"Failed to write JSON: {path}")


def _js_literal_float(x: Any) -> str:
    """Emit a JS literal number or null — never NaN/Infinity (breaks JSON.parse and inline scripts)."""
    if x is None:
        return "null"
    try:
        xf = float(x)
    except (TypeError, ValueError):
        return "null"
    if not math.isfinite(xf):
        return "null"
    return json.dumps(xf)


# ── Color palette ─────────────────────────────────────────────────────────────
C = {
    "hdr": "1C1C1C",
    "hdr_nba": "1A5276",
    "hdr_cbb": "1E8449",
    "hdr_mix": "6C3483",
    "hdr_sum": "117A65",
    "hit": "27AE60",
    "miss": "E74C3C",
    "push": "F39C12",
    "tier_a": "D5F5E3",
    "tier_b": "D6EAF8",
    "tier_c": "FEF9E7",
    "tier_d": "FDEDEC",
    "goblin": "E8D5F5",
    "demon": "FDEDEC",
    "standard": "F2F3F4",
    "over": "D6EAF8",
    "under": "FDEBD0",
    "alt": "F2F3F4",
    "white": "FFFFFF",
    "nba": "EBF5FB",
    "cbb": "EAFAF1",
    "nhl": "EBF4FD",
    "hdr_nhl": "1A3A5C",
    "hdr_soccer": "1A5C2E",
    "soccer": "EAFBF1",
    "hdr_tennis": "4A6741",
    "tennis": "E8F5E9",
    "hdr_wcbb": "4A235A",
    "hdr_mlb": "922B21",
    "hdr_nba1q": "1F618D",
    "hdr_nba1h": "117A65",
    "wcbb": "F5EEF8",
    "mlb": "FDEDEC",
    "nba1q": "D6EAF8",
    "nba1h": "D5F5E3",
    "hdr_nfl": "5D4037",
    "nfl": "FFF3E0",
    "mix": "F5EEF8",
    "gold": "F9E79F",
}

PAYOUT = {
    2: {"power": 3.0,  "flex": 3.0},
    3: {"power": 6.0,  "flex": 3.0},
    4: {"power": 10.0, "flex": 6.0},
    5: {"power": 20.0, "flex": 10.0},
    6: {"power": 40.0, "flex": 25.0},
}


# Cross-pipeline showcase slips: PrizePicks caps at 6 legs.
CROSS_PIPELINE_MAX_LEGS = 6


def power_flex_payout_for_n(n_legs: int) -> tuple[float, float]:
    """PrizePicks-style multipliers; extrapolate beyond 6 legs conservatively."""
    n = int(n_legs)
    if n < 2:
        return 0.0, 0.0
    if n in PAYOUT:
        p = PAYOUT[n]
        return float(p["power"]), float(p["flex"])
    base_n = max(PAYOUT.keys())
    base = PAYOUT[base_n]
    extra = max(0, n - base_n)
    power = float(base["power"]) * (1.10**extra)
    flex = float(base["flex"]) * (1.08**extra)
    return round(power, 2), round(flex, 2)


# ── Empirical Goblin/Demon payout (manual PrizePicks observations, April 2026) ─
# Multiplicative per leg; line_distance = |standard_line - played_line|.
# Goblin: linear distance fit from obs_A–obs_D (see data/payout_formula_coefficients.json).
GOBLIN_FACTOR_INTERCEPT = 0.838
GOBLIN_FACTOR_SLOPE = 0.031
GOBLIN_FACTOR_MIN = 0.40
DEMON_POWER_COEFF = 0.1782
DEMON_POWER_EXP = 1.287
SWEEP_PAYOUT = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 40.0}
# Power-style min guarantees (all-standard baseline) and goblin distance factors.
POWER_MIN_GUARANTEE_STANDARD: dict[int, float] = {
    2: 3.0,
    3: 1.6,
    4: 2.5,
    5: 8.5,
    6: 40.0,
}
GOBLIN_MIN_FACTOR_INTERCEPT = 1.0
GOBLIN_MIN_FACTOR_SLOPE = 0.074
GOBLIN_MIN_FACTOR_FLOOR = 0.30
# Runtime toggle wired from CLI (--debug-payout).
PAYOUT_DEBUG: bool = False
PAYOUT_LADDER_STANDARD_PATH = os.path.join(REPO_ROOT, "data", "payout_ladder.json")
PAYOUT_LADDER_REVERTED_PATH = os.path.join(REPO_ROOT, "data", "payout_ladder_reverted.json")
PAYOUT_LADDER_PATH = PAYOUT_LADDER_STANDARD_PATH
_PAYOUT_LADDER_CACHE: list[dict[str, Any]] | None = None
# Flex: published ladder (all-standard) aligned with data/payout_ladder.json.
# Keys are legs correct (n = all hit, n-1 = one miss). Goblin/demon scaling is applied on top.
FLEX_GUARANTEE: dict[int, dict[int, float]] = {
    2: {2: 3.0},
    3: {3: 2.25, 2: 1.25},
    4: {4: 5.0, 3: 1.5, 2: 0.4},
    5: {5: 10.0, 4: 2.0, 3: 0.4},
    6: {6: 25.0, 5: 2.0, 4: 0.4},
}
# Fallback when n not in FLEX_GUARANTEE (extrapolation)
BASE_FLEX_FIRST = {2: 3.0, 3: 2.25, 4: 5.0, 5: 10.0, 6: 25.0}
BASE_FLEX_MIN = {2: 0.0, 3: 1.25, 4: 1.5, 5: 2.0, 6: 2.0}
KNOWN_SWEEP_BOUNDS: dict[tuple[int, str], tuple[float, float]] = {
    (2, "power"): (3.0, 3.0),
    (3, "power"): (3.5, 6.0),
    (4, "power"): (7.5, 10.0),
    (5, "power"): (15.0, 20.0),
    (6, "power"): (28.0, 40.0),
    (3, "flex"): (1.5, 3.0),
    (4, "flex"): (2.5, 5.0),
    (5, "flex"): (4.0, 10.0),
    (6, "flex"): (8.0, 25.0),
}


def _load_payout_ladder_entries() -> list[dict[str, Any]]:
    global _PAYOUT_LADDER_CACHE
    if _PAYOUT_LADDER_CACHE is not None:
        return _PAYOUT_LADDER_CACHE
    try:
        with open(PAYOUT_LADDER_PATH, "r", encoding="utf-8") as f:
            data = json.load(f)
        entries = data.get("entries") if isinstance(data, dict) else []
        _PAYOUT_LADDER_CACHE = [e for e in (entries or []) if isinstance(e, dict)]
    except Exception:
        _PAYOUT_LADDER_CACHE = []
    return _PAYOUT_LADDER_CACHE


def configure_payout_ladder(*, use_reverted_ladder: bool) -> None:
    """Switch exact ladder source between standard and reverted payout contexts."""
    global PAYOUT_LADDER_PATH, _PAYOUT_LADDER_CACHE
    PAYOUT_LADDER_PATH = (
        PAYOUT_LADDER_REVERTED_PATH if bool(use_reverted_ladder) else PAYOUT_LADDER_STANDARD_PATH
    )
    _PAYOUT_LADDER_CACHE = None


def _mix_signature_from_legs(legs: list[dict[str, Any]]) -> dict[str, int]:
    sig = {"goblin": 0, "standard": 0, "demon": 0}
    for leg in legs or []:
        pt = str(leg.get("pick_type", "standard")).strip().lower()
        if pt not in sig:
            pt = "standard"
        sig[pt] += 1
    return sig


def _goblin_line_distance_half_steps(line_distance: float) -> int:
    """
    Ladder keys use half-point resolution (0, 0.5, 1, 1.5, …) so 1.0 and 1.5 do not collapse.
    Internal representation: integer half-steps n where distance in points = n / 2.
    """
    try:
        d = abs(float(line_distance or 0.0))
        if not math.isfinite(d):
            d = 0.0
    except (TypeError, ValueError):
        d = 0.0
    return int(round(d * 2.0))


def _goblin_distance_signature(legs: list[dict[str, Any]]) -> list[int]:
    """
    Sorted half-step ints per Goblin leg; see _goblin_line_distance_half_steps.

    Goblin lines are eased vs standard (delta > 0). Half-steps below 1
    (i.e. < 0.5 points) clamp to 1 so export noise at 0 still matches the
    shallowest ladder bucket (0.5 points).
    """
    out: list[int] = []
    for leg in legs or []:
        pt = str(leg.get("pick_type", "standard")).strip().lower()
        if pt != "goblin":
            continue
        try:
            d = abs(float(leg.get("line_distance", 0.0) or 0.0))
            if not math.isfinite(d):
                d = 0.0
        except (TypeError, ValueError):
            d = 0.0
        hs = _goblin_line_distance_half_steps(d)
        if hs < 1:
            hs = 1
        out.append(hs)
    out.sort()
    return out


def _ladder_goblin_distance_to_half_steps(raw: Any) -> int:
    """
    Parse payout_ladder.json goblin_distances element (e.g. 2, 1.5, 2.0) to half-steps.

    JSON 0 is treated as the shallowest goblin bucket (same half-step as 0.5 points).
    """
    h = int(round(float(raw) * 2.0))
    return max(1, h)


def _ladder_entry_goblin_sig(ent: dict[str, Any]) -> list[int] | None:
    """Return sorted goblin distance signature in half-steps, or None if entry matches any."""
    if "goblin_distances" not in ent:
        return None
    e_g = ent.get("goblin_distances")
    if not isinstance(e_g, list) or len(e_g) == 0:
        return None
    return sorted(_ladder_goblin_distance_to_half_steps(x) for x in e_g)


def _lookup_exact_payout_ladder(
    ticket_type: str,
    n_legs: int,
    legs: list[dict[str, Any]],
) -> tuple[float, float, str] | None:
    """
    Prefer mix + goblin_distances exact rows over mix-only wildcards so
    distance-specific ladder entries beat generic same-mix rows.

    goblin_distances in JSON are in **points** (ints or half-points like 1.5);
    matching uses half-step ints so 1.0 and 1.5 do not round into the same bucket.
    """
    mix_sig = _mix_signature_from_legs(legs)
    gob_sig = _goblin_distance_signature(legs)
    is_reverted_ladder = os.path.abspath(PAYOUT_LADDER_PATH) == os.path.abspath(PAYOUT_LADDER_REVERTED_PATH)
    specific: list[dict[str, Any]] = []
    generic: list[dict[str, Any]] = []
    for ent in _load_payout_ladder_entries():
        try:
            if str(ent.get("entry_type", "")).strip().lower() != str(ticket_type).strip().lower():
                continue
            if int(ent.get("n_legs", -1)) != int(n_legs):
                continue
            emix = ent.get("mix") if isinstance(ent.get("mix"), dict) else {}
            if {
                "goblin": int(emix.get("goblin", 0)),
                "standard": int(emix.get("standard", 0)),
                "demon": int(emix.get("demon", 0)),
            } != mix_sig:
                continue
            req = _ladder_entry_goblin_sig(ent)
            if req is None:
                # Reverted/Leaderboard payouts are highly delta-sensitive for Goblin legs.
                # Do not let mix-only wildcard rows over-apply in reverted mode.
                if is_reverted_ladder and int(mix_sig.get("goblin", 0)) > 0:
                    continue
                generic.append(ent)
            elif req == gob_sig:
                specific.append(ent)
        except Exception:
            continue
    chosen = specific[0] if specific else (generic[0] if generic else None)
    if chosen is None:
        return None
    try:
        sweep = float(chosen["sweep_payout_x"])
        min_p = float(chosen["min_payout_x"])
        return sweep, min_p, "exact"
    except Exception:
        return None


def goblin_per_leg_factor(line_distance: float) -> float:
    """Empirical goblin multiplier vs line distance (linear fit, floored)."""
    try:
        dist = abs(float(line_distance or 0.0))
        if not math.isfinite(dist):
            dist = 0.0
    except (TypeError, ValueError):
        dist = 0.0
    return max(
        GOBLIN_FACTOR_MIN,
        GOBLIN_FACTOR_INTERCEPT - GOBLIN_FACTOR_SLOPE * dist,
    )


def compute_leg_adjustment(pick_type: str, line_distance: float) -> float:
    """
    Per-leg payout adjustment factor.

    pick_type: 'goblin' | 'standard' | 'demon'
    line_distance: abs(standard_line - played_line)

    Standard = 1.0. Goblin < 1.0. Demon >= 1.0 (capped at 3).
    """
    try:
        dist = abs(float(line_distance or 0.0))
        if not math.isfinite(dist):
            dist = 0.0
    except (TypeError, ValueError):
        dist = 0.0
    pt = str(pick_type or "standard").strip().lower()

    if pt == "goblin":
        return goblin_per_leg_factor(dist)

    if pt == "demon":
        if dist <= 0:
            return 1.0
        raw = DEMON_POWER_COEFF * (dist ** DEMON_POWER_EXP)
        return min(3.0, max(1.0, raw))

    return 1.0


def compute_min_guarantee_adjustment(legs: list) -> float:
    """
    Combined min_guarantee / power adjustment: product of per-leg factors.

    legs: list of dicts with 'pick_type', 'line_distance'.
    """
    adjustment = 1.0
    for leg in legs:
        adjustment *= compute_leg_adjustment(
            str(leg.get("pick_type", "standard")),
            float(leg.get("line_distance", 0.0) or 0.0),
        )
    return round(adjustment, 4)


def compute_flex_normalized_adjustment(legs: list) -> float:
    """
    Flex ladder entries are defined for all-standard legs (factor 1). For goblins, the raw
    per-leg factor at distance 0 is GOBLIN_FACTOR_INTERCEPT; multiplying n legs would
    compound that baseline and crush partial payouts. Normalize each goblin leg by the
    intercept so distance-0 goblins do not add extra discount beyond the ladder; deeper
    lines still reduce payouts via ratio f/intercept < 1.
    """
    adjustment = 1.0
    intercept = float(GOBLIN_FACTOR_INTERCEPT) or 1.0
    for leg in legs or []:
        pt = str(leg.get("pick_type", "standard")).strip().lower()
        dist = float(leg.get("line_distance", 0.0) or 0.0)
        if pt == "goblin":
            adjustment *= goblin_per_leg_factor(dist) / intercept
        elif pt == "demon":
            adjustment *= compute_leg_adjustment(pt, dist)
    return round(float(adjustment), 4)


def goblin_min_factor(dist: float) -> float:
    try:
        d = abs(float(dist or 0.0))
        if not math.isfinite(d):
            d = 0.0
    except (TypeError, ValueError):
        d = 0.0
    return max(
        float(GOBLIN_MIN_FACTOR_FLOOR),
        float(GOBLIN_MIN_FACTOR_INTERCEPT) - float(GOBLIN_MIN_FACTOR_SLOPE) * d,
    )


def _compute_power_min_guarantee(legs: list, n_legs: int) -> tuple[float, int]:
    """
    Power-play min guarantee by all-standard base × per-leg Goblin/Demon distance factors.
    (Sweep for mixed tickets is handled separately in compute_ticket_ev; all-Goblin Power uses
    a dedicated scaled path so we do not show Standard-tier jackpots on Goblin-only slips.)
    """
    n = int(n_legs)
    base = float(POWER_MIN_GUARANTEE_STANDARD.get(n, 1.6))
    g_count = sum(
        1
        for leg in (legs or [])
        if str(leg.get("pick_type", "standard")).strip().lower() == "goblin"
    )
    adj = 1.0
    for leg in legs or []:
        pt = str(leg.get("pick_type", "") or "").strip().lower()
        try:
            dist = abs(float(leg.get("line_distance", 0.0) or 0.0))
            if not math.isfinite(dist):
                dist = 0.0
        except (TypeError, ValueError):
            dist = 0.0
        if pt == "goblin":
            f = float(goblin_min_factor(dist))
            if PAYOUT_DEBUG:
                print(f"[PAYOUT] goblin dist={dist:.1f} factor={f:.3f}")
            adj *= f
        elif pt == "demon":
            raw = DEMON_POWER_COEFF * (dist ** DEMON_POWER_EXP)
            f = min(3.0, max(1.0, raw))
            if PAYOUT_DEBUG:
                print(f"[PAYOUT] demon dist={dist:.1f} factor={f:.3f}")
            adj *= f
    mg = round(float(base * adj), 2)
    if PAYOUT_DEBUG:
        print(f"[PAYOUT] total_adj={adj:.3f}")
        print(f"[PAYOUT] min_guarantee={mg:.2f}")
    return mg, g_count


def _all_goblin_power_legs(legs: list, n: int) -> bool:
    """True when this is an n-leg Power slip with every leg marked Goblin (no Std/Demon)."""
    if n < 1:
        return False
    mx = _mix_signature_from_legs(legs)
    return (
        int(mx.get("goblin", 0)) == int(n)
        and int(mx.get("standard", 0)) == 0
        and int(mx.get("demon", 0)) == 0
    )


def _all_goblin_power_sweep_min_from_standard_tier(legs: list, n: int) -> tuple[float, float]:
    """
    PrizePicks all-Goblin Power boards pay far below all-Standard; there is often no ladder row
    beyond 3 legs. Scale the published Standard sweep + min tier by the empirical Goblin
    per-leg factor product (same family as Flex EV adjustments).
    """
    prod = 1.0
    for leg in legs or []:
        try:
            d = abs(float(leg.get("line_distance", 0.0) or 0.0))
            if not math.isfinite(d):
                d = 0.0
        except (TypeError, ValueError):
            d = 0.0
        prod *= float(goblin_per_leg_factor(d))
    std_sw = float(SWEEP_PAYOUT.get(int(n), 6.0))
    std_mn = float(POWER_MIN_GUARANTEE_STANDARD.get(int(n), std_sw))
    sweep = round(max(1.05, std_sw * prod), 4)
    mn = round(max(1.0, std_mn * prod), 4)
    if mn > sweep:
        mn = sweep
    return sweep, mn


def compute_ticket_ev(
    legs: list,
    ticket_type: str,
    n_legs: int,
) -> dict[str, Any]:
    """
    EV (per $1 stake style) using empirical payout formula.

    Power: all-Standard uses SWEEP_PAYOUT / POWER_MIN_GUARANTEE_STANDARD (see ladder + 6-leg fix).
           Mixed or all-Goblin uses ladder when present; otherwise all-Goblin Power is scaled down
           from the Standard tier by the Goblin per-leg factor product (sweep is not 40× on 6 Goblin).
    Flex: FLEX_GUARANTEE payouts × per-leg goblin/demon adjustment product.
    """
    n = int(n_legs)
    tt = str(ticket_type or "power").strip().lower()
    adj = float(compute_min_guarantee_adjustment(legs))
    flex_adj = float(compute_flex_normalized_adjustment(legs)) if tt == "flex" else adj
    payout_source = "calibrated"

    exact = _lookup_exact_payout_ladder(tt, n, legs)
    if exact is not None:
        adjusted_first, adjusted_min_g, payout_source = exact
    else:
        if tt == "flex":
            flex_tbl = FLEX_GUARANTEE.get(n, {})
            base_first = float(flex_tbl.get(n, BASE_FLEX_FIRST.get(n, 3.0)))
            base_partial = float(flex_tbl.get(n - 1, BASE_FLEX_MIN.get(n, 0.0))) if n >= 2 else 0.0
            adjusted_first = round(base_first * flex_adj, 4)
            adjusted_min_g = round(base_partial * flex_adj, 4)
        else:
            adjusted_first = round(float(SWEEP_PAYOUT.get(n, 6.0)), 4)
            adjusted_min_g, g_count = _compute_power_min_guarantee(legs, n)
            if PAYOUT_DEBUG:
                print(f"[PAYOUT DEBUG] n_legs={n} base={adjusted_first}")
                print(f"[PAYOUT DEBUG] goblin_legs={g_count}")
                for leg in legs:
                    if str(leg.get("pick_type", "standard")).strip().lower() != "goblin":
                        continue
                    try:
                        dist = abs(float(leg.get("line_distance", 0.0) or 0.0))
                        if not math.isfinite(dist):
                            dist = 0.0
                    except (TypeError, ValueError):
                        dist = 0.0
                    factor = goblin_per_leg_factor(dist)
                    print(f"[PAYOUT DEBUG] dist={dist} factor={factor}")
                print(f"[PAYOUT DEBUG] total_adj={adj} sweep={adjusted_first} min_guarantee={adjusted_min_g}")
                print(f"[PAYOUT] min_guarantee={adjusted_min_g}")

    # 6-leg all-Standard Power: published board is 40× min guarantee and 40× sweep.
    # `payout_ladder.json` may still carry an older 37.5× row; ladder "exact" must not win here.
    if tt == "power" and n == 6:
        mix6 = _mix_signature_from_legs(legs)
        if int(mix6.get("goblin", 0)) == 0 and int(mix6.get("demon", 0)) == 0:
            if payout_source == "exact":
                payout_source = "calibrated"
            adjusted_first = float(SWEEP_PAYOUT.get(6, 40.0))
            adjusted_min_g = float(POWER_MIN_GUARANTEE_STANDARD.get(6, 40.0))

    # All-Goblin Power: never inherit Standard 40× sweep + PP "all-standard" sweep clamps.
    if tt == "power" and _all_goblin_power_legs(legs, n) and payout_source != "exact":
        adjusted_first, adjusted_min_g = _all_goblin_power_sweep_min_from_standard_tier(legs, n)
        payout_source = "calibrated"

    bounds = KNOWN_SWEEP_BOUNDS.get((n, tt))
    skip_sweep_bounds = tt == "power" and _all_goblin_power_legs(legs, n) and payout_source != "exact"
    if bounds and not skip_sweep_bounds:
        lo, hi = float(bounds[0]), float(bounds[1])
        adjusted_first = max(lo, min(hi, float(adjusted_first)))
        if tt == "flex":
            # Partial tier can be below sweep bounds' lo; do not force a bogus floor from sweep.
            adjusted_min_g = max(0.0, min(float(hi), float(adjusted_min_g)))
        else:
            min_floor = lo * 0.3
            adjusted_min_g = max(min_floor, min(hi, float(adjusted_min_g)))

    mg_adjustment = round(adjusted_min_g / adjusted_first, 4) if adjusted_first > 0 else 0.0

    probs = [float(leg.get("hit_prob", 0.65)) for leg in legs]
    p_all = 1.0
    for p in probs:
        p_all *= p

    p_miss_1 = 0.0
    if n >= 2:
        for i in range(len(probs)):
            term = 1.0 - probs[i]
            for j, p in enumerate(probs):
                if j != i:
                    term *= p
            p_miss_1 += term

    p_lose = max(0.0, 1.0 - p_all - p_miss_1)
    if payout_source != "exact":
        hsrcs = {
            str(leg.get("hit_prob_source", "")).strip().lower()
            for leg in (legs or [])
            if isinstance(leg, dict)
        }
        if any(s in {"ml_prob", "fallback_const"} for s in hsrcs):
            payout_source = "fallback"
    if tt == "power":
        ev = (p_all * adjusted_min_g) - 1.0
    else:
        ev = (p_all * adjusted_first) + (p_miss_1 * adjusted_min_g) - p_lose
    min_payout_x = float(adjusted_min_g)
    sweep_payout_x = float(adjusted_first)
    ev_formula = (
        f"EV = P(all)*{sweep_payout_x:.2f} + P(miss-1)*{min_payout_x:.2f} - 1.0"
        if tt == "flex"
        else f"EV = P(all)*{min_payout_x:.2f} - 1.0"
    )

    return {
        "ev": round(ev, 4),
        "p_all_win": round(p_all, 4),
        "p_miss_1": round(p_miss_1, 4),
        "p_lose": round(p_lose, 4),
        "first_place_payout": adjusted_first,
        "min_guarantee": adjusted_min_g,
        "min_payout_x": round(min_payout_x, 4),
        "sweep_payout_x": round(sweep_payout_x, 4),
        "min_guarantee_adjustment": mg_adjustment,
        "payout_adjustment": (flex_adj if tt == "flex" else 1.0),
        "payout_source": payout_source,
        "ev_formula": ev_formula,
        "ticket_type": tt,
        "n_legs": n,
        "recommendation": recommendation_from_ev(ev),
    }


def _ticket_row_get(row: Any, field: str) -> Any:
    """Read one field from a leg row (dict or pandas Series)."""
    if isinstance(row, dict):
        return row.get(field)
    try:
        if hasattr(row, "index") and field in row.index:
            return row[field]
    except Exception:
        pass
    try:
        return getattr(row, field, None)
    except Exception:
        return None


def _ticket_payout_line_distance(row: Any) -> float:
    ld = _ticket_row_get(row, "line_distance")
    if ld is not None and str(ld).strip() != "":
        try:
            v = float(ld)
            if not math.isfinite(v):
                return 0.0
            return abs(v)
        except (TypeError, ValueError):
            pass
    try:
        sl = _ticket_row_get(row, "standard_line")
        ln = _ticket_row_get(row, "line")
        if sl is not None and ln is not None and str(sl).strip() != "" and str(ln).strip() != "":
            sv = float(sl)
            lv = float(ln)
            if not (math.isfinite(sv) and math.isfinite(lv)):
                return 0.0
            return abs(sv - lv)
    except (TypeError, ValueError):
        pass
    return 0.0


def _normalize_historical_hit_rate_to_prob(hr_raw: Any, default: float = 0.65) -> float:
    """Convert spreadsheet/UI hit rate to Bernoulli p in (0,1). Supports 0.8 or 80 or '80%'."""
    if hr_raw is None or (isinstance(hr_raw, str) and not str(hr_raw).strip()):
        return default
    try:
        s = str(hr_raw).strip().replace("%", "")
        v = float(s)
    except (TypeError, ValueError):
        return default
    if isinstance(v, float) and math.isnan(v):
        return default
    if v <= 0:
        return default
    if v > 1.0:
        v = v / 100.0 if v <= 100.0 else 1.0
    return max(0.05, min(0.99, float(v)))


def _ticket_payout_hit_prob_with_source(row: Any) -> tuple[float, str]:
    """
    Per-leg hit probability for empirical EV: direction-aware historical hit rate
    (hit_rate / over-under splits), not blended_score or ml_prob.
    """
    direction_raw = (
        _ticket_row_get(row, "bet_direction")
        or _ticket_row_get(row, "direction_used")
        or _ticket_row_get(row, "direction")
        or "OVER"
    )
    direction = str(direction_raw).strip().upper()
    if "UNDER" in direction:
        direction = "UNDER"
    elif "OVER" in direction:
        direction = "OVER"
    else:
        direction = "OVER"

    hr_raw = None
    if direction == "UNDER":
        for key in ("under_hit_rate", "hit_rate_under_L5", "hit_rate_under_L10", "hit_rate"):
            hr_raw = _ticket_row_get(row, key)
            if hr_raw is not None and str(hr_raw).strip() != "":
                break
        if hr_raw is None or str(hr_raw).strip() == "":
            hr_raw = _ticket_row_get(row, "hr")
    else:
        for key in ("over_hit_rate", "hit_rate_over_L5", "hit_rate_over_L10", "hit_rate"):
            hr_raw = _ticket_row_get(row, key)
            if hr_raw is not None and str(hr_raw).strip() != "":
                break
        if hr_raw is None or str(hr_raw).strip() == "":
            hr_raw = _ticket_row_get(row, "hr")

    if hr_raw is not None and str(hr_raw).strip() != "":
        return float(_normalize_historical_hit_rate_to_prob(hr_raw, default=0.65)), "hit_rate"

    mlp = pd.to_numeric(_ticket_row_get(row, "ml_prob"), errors="coerce")
    if pd.notna(mlp):
        try:
            mv = float(mlp)
            if 0.0 < mv < 1.0 and math.isfinite(mv):
                return mv, "ml_prob"
        except (TypeError, ValueError):
            pass
    return 0.65, "fallback_const"


def _ticket_payout_pick_type_token(row: Any) -> str:
    pt = str(
        _ticket_row_get(row, "pick_type")
        or _ticket_row_get(row, "Pick Type")
        or "standard"
    ).strip().lower()
    if "goblin" in pt:
        return "goblin"
    if "demon" in pt:
        return "demon"
    return "standard"


def build_ticket_payout_json(group_name: str, ticket_rows: list) -> dict[str, Any] | None:
    """
    Empirical payout + EV block for tickets_latest.json / UI.

    ``payout`` / ``min_guarantee`` are the primary min-guarantee multipliers.
    ``sweep_payout`` is the all-legs-hit upside multiplier ("jackpot").
    On any error, returns None (caller stores null in JSON).
    """
    if not ticket_rows:
        return None
    legs: list[dict[str, Any]] = []
    for row in ticket_rows:
        hp, hp_src = _ticket_payout_hit_prob_with_source(row)
        legs.append(
            {
                "pick_type": _ticket_payout_pick_type_token(row),
                "line_distance": float(_ticket_payout_line_distance(row) or 0.0),
                "hit_prob": float(hp),
                "hit_prob_source": hp_src,
            }
        )
    n = len(legs)
    gname = str(group_name or "")
    # PrizePicks: Flex cash slips are named "… Flex …"; Power Play / Standard / Goblin sheets are power path.
    tt = "flex" if "Flex" in gname else "power"
    try:
        ev_result = compute_ticket_ev(legs=legs, ticket_type=tt, n_legs=n)
    except Exception:
        return None
    try:
        mg = float(ev_result["min_guarantee"])
        sweep = float(ev_result["first_place_payout"])
        paw = float(ev_result["p_all_win"])
        return {
            "ticket_type": tt,
            "payout": ev_result["min_guarantee"],
            "min_guarantee": ev_result["min_guarantee"],
            "min_payout_x": ev_result.get("min_payout_x", ev_result["min_guarantee"]),
            "min_guarantee_adjustment": ev_result["min_guarantee_adjustment"],
            "payout_adjustment": ev_result.get("payout_adjustment", 1.0),
            "p_all_win": ev_result["p_all_win"],
            "p_miss_1": ev_result["p_miss_1"],
            "ev": ev_result["ev"],
            "recommendation": ev_result["recommendation"],
            "entry_10_to_win_guarantee": round(10 * mg, 2),
            "entry_20_to_win_guarantee": round(20 * mg, 2),
            "sweep_payout": sweep,
            "sweep_payout_x": ev_result.get("sweep_payout_x", sweep),
            "entry_10_to_win_sweep": round(10 * sweep, 2),
            "payout_confidence_score": round(sweep * paw, 4),
            "payout_source": ev_result.get("payout_source", "calibrated"),
            "ev_formula": ev_result.get("ev_formula", ""),
        }
    except Exception:
        return None


def _ticket_passes_positive_ev_gate(ticket: dict) -> bool:
    """
    True if slip should appear on /tickets JSON and page.

    Hygiene (all sports): never show SKIP, negative empirical payout.ev, negative est_ev,
    or modeled win prob below MIN_WEB_DISPLAY_EST_WIN_PROB.

    Then: pure Tennis / Soccer bypass only the *min-EV threshold* (sparse boards).

    Other sports: payout.ev >= bar OR STRONG/OK, else est_ev >= bar.
    """
    n_legs = _ticket_n_legs(ticket)
    min_ev = float(MIN_TICKET_EV_BY_LEGS.get(int(n_legs), MIN_TICKET_EV_DEFAULT))
    legs = list(ticket.get("legs") or [])
    leg_sports = {
        str(leg.get("sport") or "").strip().upper()
        for leg in legs
        if isinstance(leg, dict)
    }

    wp = ticket.get("est_win_prob")
    if wp is not None:
        try:
            wpf = float(wp)
            if math.isfinite(wpf) and wpf < float(MIN_WEB_DISPLAY_EST_WIN_PROB):
                return False
        except (TypeError, ValueError):
            pass

    pay = ticket.get("payout")
    if isinstance(pay, dict):
        rec_u = str(pay.get("recommendation") or "").strip().upper()
        if rec_u == "SKIP":
            return False
        ev_raw = pay.get("ev")
        if ev_raw is not None:
            try:
                v = float(ev_raw)
                if math.isfinite(v) and v < 0:
                    return False
            except (TypeError, ValueError):
                pass

    est_neg = ticket.get("est_ev")
    if est_neg is not None:
        try:
            ve = float(est_neg)
            if math.isfinite(ve) and ve < 0:
                return False
        except (TypeError, ValueError):
            pass

    # Pure Tennis / Soccer: after hygiene, skip the global min-EV bar only.
    if leg_sports and (
        leg_sports.issubset({"TENNIS"})
        or leg_sports.issubset({"SOCCER", "SOC"})
    ):
        return True

    if isinstance(pay, dict):
        rec_ok = str(pay.get("recommendation") or "").strip().upper() in (
            "STRONG",
            "OK",
        )
        ev_ok = False
        ev_raw = pay.get("ev")
        if ev_raw is not None:
            try:
                v = float(ev_raw)
                if isinstance(v, float) and math.isnan(v):
                    ev_ok = False
                elif math.isfinite(v):
                    ev_ok = v >= min_ev
            except (TypeError, ValueError):
                pass
        if ev_ok or rec_ok:
            return True
    est = ticket.get("est_ev")
    if est is not None:
        try:
            v = float(est)
            return math.isfinite(v) and v >= min_ev
        except (TypeError, ValueError):
            return False
    return False


def _ticket_meets_min_web_payout(ticket: dict, *, group_name: str = "") -> bool:
    """Require selected ticket-mode payout multipliers to be >= MIN_WEB_PAYOUT_X."""
    payout_vals: list[float] = []
    pay = ticket.get("payout")
    mode = ""
    if isinstance(pay, dict):
        mode = str(pay.get("ticket_type") or pay.get("entry_type") or "").strip().lower()
    if mode not in {"power", "flex"}:
        gn = str(group_name or "").strip().lower()
        mode = "flex" if "flex" in gn else "power"

    if isinstance(pay, dict):
        pay_keys = (
            ("min_payout_x", "payout")
            if mode == "flex"
            else ("sweep_payout_x", "sweep_payout", "min_payout_x", "payout")
        )
        for k in pay_keys:
            v = pay.get(k)
            if v is None:
                continue
            try:
                vf = float(v)
                if math.isfinite(vf):
                    payout_vals.append(vf)
            except (TypeError, ValueError):
                pass
    ticket_keys = (
        ("flex_payout",)
        if mode == "flex"
        else ("power_payout", "base_power_payout")
    )
    for k in ticket_keys:
        v = ticket.get(k)
        if v is None:
            continue
        try:
            vf = float(v)
            if math.isfinite(vf):
                payout_vals.append(vf)
        except (TypeError, ValueError):
            pass
    if not payout_vals:
        return False
    min_required = float(MIN_WEB_PAYOUT_X)
    gn_u = str(group_name or "").upper()
    if "GOBLIN" in gn_u:
        m = re.search(r"(\d+)\s*-\s*LEG", gn_u)
        if m:
            try:
                n_legs = int(m.group(1))
                if n_legs in (3, 4):
                    # Allow short goblin slips on /tickets; these usually do not clear the 3x floor.
                    min_required = float(MIN_WEB_PAYOUT_X_GOBLIN_SHORT)
            except (TypeError, ValueError):
                pass
    return min(payout_vals) >= min_required


def _ticket_primary_sport(ticket: dict) -> str:
    legs = list(ticket.get("legs") or [])
    sports = {
        str(leg.get("sport") or "").strip().upper()
        for leg in legs
        if isinstance(leg, dict) and str(leg.get("sport") or "").strip()
    }
    if len(sports) != 1:
        return ""
    return next(iter(sports))


def _ticket_leg_sports(ticket: dict) -> set[str]:
    legs = list(ticket.get("legs") or [])
    return {
        str(leg.get("sport") or "").strip().upper()
        for leg in legs
        if isinstance(leg, dict) and str(leg.get("sport") or "").strip()
    }


def _ticket_rank_tuple(ticket: dict) -> tuple[float, int, float]:
    pay = ticket.get("payout")
    ev = None
    rec = ""
    payout_conf = 0.0
    if isinstance(pay, dict):
        ev = pay.get("ev")
        rec = str(pay.get("recommendation") or "").strip().upper()
        try:
            pc = float(pay.get("payout_confidence_score"))
            if math.isfinite(pc):
                payout_conf = pc
        except (TypeError, ValueError):
            payout_conf = 0.0
    if ev is None:
        ev = ticket.get("est_ev")
    try:
        evf = float(ev)
        if not math.isfinite(evf):
            evf = -1e9
    except (TypeError, ValueError):
        evf = -1e9
    rec_rank = {"STRONG": 3, "OK": 2, "MARGINAL": 1, "SKIP": 0}.get(rec, -1)
    try:
        winf = float(ticket.get("est_win_prob"))
        if not math.isfinite(winf):
            winf = -1e9
    except (TypeError, ValueError):
        winf = -1e9
    return (payout_conf, evf, rec_rank, winf)


def _apply_web_ticket_template(groups: list[dict]) -> list[dict]:
    if not groups:
        return []

    candidates_by_key: dict[tuple[str, int], list[tuple[tuple[int, int], dict]]] = defaultdict(list)

    for gi, g in enumerate(groups):
        tickets = list(g.get("tickets") or [])
        for ti, t in enumerate(tickets):
            if not isinstance(t, dict):
                continue
            n_legs = _ticket_n_legs(t)
            if n_legs not in WEB_TICKET_TEMPLATE_BY_LEGS:
                continue
            sport = _ticket_primary_sport(t)
            if not sport:
                continue
            candidates_by_key[(sport, n_legs)].append(((gi, ti), t))

    selected_ids: set[tuple[int, int]] = set()
    sports = sorted({sport for sport, _ in candidates_by_key.keys()})
    for sport in sports:
        for n_legs, quota in WEB_TICKET_TEMPLATE_BY_LEGS.items():
            cand = list(candidates_by_key.get((sport, n_legs), []))
            if not cand:
                continue
            cand.sort(key=lambda x: _ticket_rank_tuple(x[1]), reverse=True)
            for tid, _t in cand[: int(quota)]:
                selected_ids.add(tid)

    if not selected_ids:
        return groups

    out_groups: list[dict] = []
    for gi, g in enumerate(groups):
        tickets = list(g.get("tickets") or [])
        kept = [t for ti, t in enumerate(tickets) if (gi, ti) in selected_ids]
        # Keep template-sized slips that never entered the per-sport quota pool (e.g. cross-sport,
        # or legs missing sport before payload fixes). Otherwise they vanish whenever any other slip
        # activates the template filter.
        for ti, t in enumerate(tickets):
            if (gi, ti) in selected_ids or not isinstance(t, dict):
                continue
            nl = _ticket_n_legs(t)
            if nl not in WEB_TICKET_TEMPLATE_BY_LEGS:
                continue
            if _ticket_primary_sport(t):
                continue
            kept.append(t)
        if not kept:
            continue
        ng = dict(g)
        ng["tickets"] = kept
        out_groups.append(ng)
    return out_groups


def filter_web_tickets_for_ui(
    payload: dict,
    *,
    require_positive_ev: bool,
    apply_template_cap: bool = False,
    discard_tracker: "DiscardTracker | None" = None,
) -> dict:
    """Build /tickets JSON groups: optional positive-EV gate, then optional WEB_TICKET_TEMPLATE quotas."""
    groups_in = list(payload.get("groups") or [])
    out_groups: list[dict] = []
    sport_candidates: dict[str, tuple[tuple[float, int, float], dict, str]] = {}
    for g in groups_in:
        if not isinstance(g, dict):
            continue
        group_name = str(g.get("group_name") or "")
        tickets_in = list(g.get("tickets") or [])
        sport_key = _group_sport(str(g.get("group_name") or "")) or "ALL"
        for t in tickets_in:
            if not isinstance(t, dict):
                continue
            rnk = _ticket_rank_tuple(t)
            leg_sports = _ticket_leg_sports(t)
            for sp in leg_sports:
                cur = sport_candidates.get(sp)
                if cur is None or rnk > cur[0]:
                    sport_candidates[sp] = (rnk, t, group_name)
        if require_positive_ev:
            kept = []
            for t in tickets_in:
                if not isinstance(t, dict):
                    continue
                if not _ticket_passes_positive_ev_gate(t):
                    if discard_tracker is not None:
                        discard_tracker.log_count(sport_key, "web_ev_gate_fail", 1)
                    continue
                if not _ticket_meets_min_web_payout(t, group_name=str(g.get("group_name") or "")):
                    if discard_tracker is not None:
                        discard_tracker.log_count(sport_key, "payout_below_3x", 1)
                    continue
                kept.append(t)
        else:
            kept = []
            for t in tickets_in:
                if not isinstance(t, dict):
                    continue
                if not _ticket_meets_min_web_payout(t, group_name=str(g.get("group_name") or "")):
                    if discard_tracker is not None:
                        discard_tracker.log_count(sport_key, "payout_below_3x", 1)
                    continue
                kept.append(t)
        if not kept:
            continue
        ng = dict(g)
        ng["tickets"] = kept
        out_groups.append(ng)
    if apply_template_cap:
        out_groups = _apply_web_ticket_template(out_groups)

    # Keep at least one single-sport ticket visible per sport in /tickets UI.
    # This prevents strict web gates from hiding entire sports (e.g. NBA1H/SOCCER/WNBA)
    # when they still have generated candidates in the payload.
    ensure_cov_raw = os.getenv("PROPORACLE_WEB_ENSURE_SPORT_COVERAGE", "1").strip().lower()
    ensure_sport_coverage = ensure_cov_raw not in {"0", "false", "no", "off"}
    if ensure_sport_coverage and sport_candidates:
        present_sports: set[str] = set()
        existing_group_names: set[str] = set()
        for g in out_groups:
            if not isinstance(g, dict):
                continue
            gn = str(g.get("group_name") or "")
            if gn:
                existing_group_names.add(gn)
                present_sports.add(_group_sport(gn))

        missing = [sp for sp in sport_candidates.keys() if sp not in present_sports]
        for sp in sorted(missing):
            _rank, best_ticket, src_group_name = sport_candidates[sp]
            inserted = False
            for g in out_groups:
                if not isinstance(g, dict):
                    continue
                gn = str(g.get("group_name") or "")
                if gn == src_group_name and _group_sport(gn) == sp:
                    tickets_cur = [t for t in list(g.get("tickets") or []) if isinstance(t, dict)]
                    if not any(sp in _ticket_leg_sports(t) for t in tickets_cur):
                        tickets_cur.append(best_ticket)
                        g["tickets"] = tickets_cur
                    inserted = True
                    break
            if inserted:
                continue

            # Source group was filtered out; create a minimal sport group with its best ticket.
            out_groups.append(
                {
                    "group_name": f"{sp} Coverage",
                    "n_legs": _ticket_n_legs(best_ticket),
                    "power_payout": best_ticket.get("power_payout") or best_ticket.get("base_power_payout"),
                    "flex_payout": best_ticket.get("flex_payout"),
                    "tickets": [best_ticket],
                }
            )

    out = dict(payload)
    out["groups"] = out_groups
    return out


def filter_positive_ev_tickets_payload(payload: dict, *, apply_template_cap: bool = False) -> dict:
    """Only persist / show slips that pass _ticket_passes_positive_ev_gate; drop empty groups."""
    return filter_web_tickets_for_ui(
        payload,
        require_positive_ev=True,
        apply_template_cap=apply_template_cap,
    )


def print_positive_ev_gate_report(gated_preview: dict) -> None:
    """Console verification: leg-count histogram, best EV slip, sport mix, fingerprint dedupe."""
    n_g_g = len(gated_preview["groups"])
    n_s_g = sum(len(g["tickets"]) for g in gated_preview["groups"])
    print(f"  [gate positive-EV by leg] groups: {n_g_g}  slips: {n_s_g}")
    lc_grp: Counter[int] = Counter()
    for _g in gated_preview["groups"]:
        _nl0 = int(_g.get("n_legs") or 0)
        if _nl0 > 0:
            lc_grp[_nl0] += 1
    print(f"  [verify] gated groups by leg count: {dict(sorted(lc_grp.items()))}")
    lc_g: Counter[int] = Counter()
    best_ev_o = -1e9
    best_n_o = 0
    best_gn_o = ""
    for _g in gated_preview["groups"]:
        _gn = str(_g.get("group_name") or "")
        for _t in _g.get("tickets") or []:
            if not isinstance(_t, dict):
                continue
            _nl = _ticket_n_legs(_t)
            lc_g[_nl] += 1
            _evv = None
            _pay = _t.get("payout")
            if isinstance(_pay, dict) and _pay.get("ev") is not None:
                try:
                    _evv = float(_pay["ev"])
                except (TypeError, ValueError):
                    pass
            if _evv is None and _t.get("est_ev") is not None:
                try:
                    _evv = float(_t["est_ev"])
                except (TypeError, ValueError):
                    pass
            if _evv is not None and math.isfinite(_evv) and _evv > best_ev_o:
                best_ev_o = _evv
                best_n_o = _nl
                best_gn_o = _gn
    print(f"  [verify] gated slips by leg count: {dict(sorted(lc_g.items()))}")
    print(
        f"  [verify] highest EV slip (gated): ev={best_ev_o:.4f} n_legs={best_n_o} "
        f"group={best_gn_o[:72]}"
    )
    sport_slip_ctr: Counter[str] = Counter()
    for _g in gated_preview["groups"]:
        _gn = str(_g.get("group_name") or "")
        sport_slip_ctr[_group_sport(_gn)] += len(_g.get("tickets") or [])
    print(f"  [gate positive-EV by leg] slips by sport: {dict(sport_slip_ctr)}")
    _fps: list[frozenset] = []
    for _g in gated_preview["groups"]:
        _acc: set[tuple[str, str, str, str]] = set()
        for _t in _g.get("tickets") or []:
            for _L in _t.get("legs") or []:
                if isinstance(_L, dict):
                    _acc.add(_leg_fp_tuple(_L))
        _fps.append(frozenset(_acc))
    _dup = len(_fps) != len(set(_fps))
    print(f"  [gate positive-EV by leg] duplicate fingerprints among groups: {'YES' if _dup else 'none'}")


def _norm_line_for_leg_fp(val: Any) -> str:
    """Normalize line for stable dedupe keys."""
    if val is None or val == "":
        return ""
    try:
        x = float(val)
        if isinstance(x, float) and math.isnan(x):
            return ""
        return f"{x:.6g}"
    except (TypeError, ValueError):
        return str(val).strip().lower()


def _leg_fp_tuple(r: Any) -> tuple[str, str, str, str]:
    """One leg key: (player, prop, line, direction). Works on ticket row dict/Series or JSON leg dict."""

    def gv(field: str) -> Any:
        if isinstance(r, dict):
            return r.get(field)
        try:
            if hasattr(r, "index") and field in r.index:
                return r[field]
        except Exception:
            pass
        try:
            return getattr(r, field, None)
        except Exception:
            return None

    p = str(gv("player_name") or gv("player") or "").strip().lower()
    pt = str(gv("prop_type") or gv("prop") or "").strip().lower()
    ln_raw = gv("line_score") if gv("line_score") is not None else gv("line")
    line_s = _norm_line_for_leg_fp(ln_raw)
    d_raw = gv("bet_direction") or gv("direction") or gv("direction_used")
    d = str(d_raw or "").strip().upper()
    if "UNDER" in d:
        d = "UNDER"
    elif "OVER" in d:
        d = "OVER"
    return (p, pt, line_s, d)


def _ticket_group_leg_fingerprint(tickets: list) -> frozenset:
    """All legs across all slips in a workbook group (player+prop+line+direction)."""
    acc: set[tuple[str, str, str, str]] = set()
    for t in tickets or []:
        if not isinstance(t, dict):
            continue
        for row in t.get("rows") or []:
            acc.add(_leg_fp_tuple(row))
    return frozenset(acc)


def _ticket_group_dedupe_key(tickets: list) -> tuple[frozenset, tuple[str, ...]]:
    """Leg set + ticket_type per slip so Flex vs Power with the same legs are not merged."""
    leg_fp = _ticket_group_leg_fingerprint(tickets)
    types = tuple(
        str(t.get("ticket_type", "")).strip().lower()
        for t in (tickets or [])
        if isinstance(t, dict)
    )
    return (leg_fp, types)


def dedupe_ticket_groups_by_leg_set(all_ticket_groups: list) -> tuple[list, int, int]:
    # Deduplicate: drop groups with identical legs AND ticket product (Flex vs Power Std 3, etc.)
    n_before = len(all_ticket_groups)
    seen: set[tuple[frozenset, tuple[str, ...]]] = set()
    out: list = []
    for item in all_ticket_groups:
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            out.append(item)
            continue
        group_name, tickets = item[0], item[1]
        tail = item[2:] if len(item) > 2 else ()
        tickets = tickets or []
        fp = _ticket_group_leg_fingerprint(tickets)
        if len(fp) == 0:
            out.append((group_name, tickets, *tail))
            continue
        key = _ticket_group_dedupe_key(tickets)
        if key in seen:
            continue
        seen.add(key)
        out.append((group_name, tickets, *tail))
    return out, n_before, len(out)


def enforce_group_jaccard_diversity(
    all_ticket_groups: list,
    max_jaccard_overlap: float = 0.55,
) -> tuple[list, int, int]:
    """
    Hard diversity guard at group level:
    after exact dedupe, keep only groups whose combined leg-set overlap is <= threshold.
    """
    n_before = len(all_ticket_groups or [])
    if n_before <= 1:
        return list(all_ticket_groups or []), n_before, n_before
    thr = float(max(0.0, min(1.0, max_jaccard_overlap)))
    kept: list = []
    kept_leg_sets: list[frozenset] = []
    for item in (all_ticket_groups or []):
        if not isinstance(item, (list, tuple)) or len(item) < 2:
            kept.append(item)
            continue
        tickets = item[1] or []
        leg_set = _ticket_group_leg_fingerprint(tickets)
        if not leg_set:
            kept.append(item)
            continue
        too_similar = False
        for prev in kept_leg_sets:
            union = len(leg_set | prev)
            overlap = (len(leg_set & prev) / union) if union > 0 else 0.0
            if overlap > thr:
                too_similar = True
                break
        if too_similar:
            continue
        kept.append(item)
        kept_leg_sets.append(leg_set)
    return kept, n_before, len(kept)


def _load_diversity_config(path: str = DIVERSITY_CONFIG_PATH) -> dict[str, Any]:
    defaults: dict[str, Any] = {
        "max_leg_exposure": 4,
        "max_player_exposure": 8,
        "void_risk_min_sample": 10,
        "max_jaccard_overlap": 0.8,
        "exposure_penalty_weight": 0.1,
        "overlap_penalty_weight": 0.2,
        "void_penalty_weight": 0.5,
        "enabled": False,
    }
    try:
        if not os.path.isfile(path):
            _log_slate.info("[diversity] config missing at %s; using defaults", path)
            return defaults
        raw = json.loads(Path(path).read_text(encoding="utf-8"))
        if not isinstance(raw, dict):
            _log_slate.info("[diversity] invalid config shape at %s; using defaults", path)
            return defaults
        merged = dict(defaults)
        merged.update(raw)
        return merged
    except Exception as e:
        _log_slate.info("[diversity] failed to read config at %s (%s); using defaults", path, e)
        return defaults


def _apply_diversity_filter_to_ticket_groups(
    all_ticket_groups: list[tuple[str, list[dict[str, Any]], Any]],
    config: dict[str, Any],
) -> list[tuple[str, list[dict[str, Any]], Any]]:
    flat: list[dict[str, Any]] = []
    for _gname, tickets, _bg in all_ticket_groups:
        for t in (tickets or []):
            if isinstance(t, dict):
                flat.append(t)
    n_before_groups = len(all_ticket_groups)
    n_before_slips = len(flat)
    if n_before_slips == 0:
        _log_slate.info("[diversity] no candidate slips to filter")
        return all_ticket_groups

    kept = apply_diversity_filter(flat, config)
    kept_ids = {id(t) for t in kept}

    out: list[tuple[str, list[dict[str, Any]], Any]] = []
    dropped_groups = 0
    for group_name, tickets, bg in all_ticket_groups:
        filtered = [t for t in (tickets or []) if id(t) in kept_ids]
        if filtered:
            out.append((group_name, filtered, bg))
        else:
            dropped_groups += 1

    n_after_slips = sum(len(g[1]) for g in out)
    _log_slate.info(
        "[diversity] groups %d -> %d (dropped %d), slips %d -> %d",
        n_before_groups,
        len(out),
        dropped_groups,
        n_before_slips,
        n_after_slips,
    )
    return out


# Props excluded from ticket pools based on empirical hit rates below break-even
# Blocked Shots NBA: 41.9% overall, too low for any ticket
# Combo props: small sample, unreliable
# NHL OVER props: 21.5% hit rate — never use OVER direction in NHL tickets
# Fantasy Score is excluded from ticket generation pending data integrity
# validation. It remains in all grade/ranking outputs so hit rates can be
# monitored. Remove from this set once validated.
TICKET_EXCLUDED_PROPS = {
    "fantasy score", "fantasy_score", "fantasy",
    "fg made",
    "personal fouls",
    "blks+stls",
}

ATTEMPT_PROPS = {
    "fg attempted",
    "field goals attempted",
    "3-pt attempted",
    "three pointers attempted",
    "two pointers attempted",
}

# Keep priority tiers focused on single-stat regular props.
# Combo props are still allowed (unless otherwise excluded) but should not get
# an extra priority bonus versus regular markets.
TIER1_PROPS = {"points", "rebounds"}
TIER2_PROPS = {"assists", "3-pt made"}
TIER3_PROPS = {"steals", "blocked shots", "turnovers", "free throws made"}

UNDER_ALLOWED_PROPS = {"free throws attempted", "turnovers"}

NBA_EXCLUDED_PROPS = {
    "blocked shots",
    "free throws attempted",
    "defensive rebounds", "offensive rebounds",
    "dunks", "quarters with 5+ points", "quarters with 3+ points",
    "points - 1st 3 minutes", "assists - 1st 3 minutes", "rebounds - 1st 3 minutes",
    "assists (combo)", "points (combo)", "rebounds (combo)", "3-pt made (combo)",
}

CBB_EXCLUDED_PROPS = {
    "fantasy",
    "points (combo)",
}

# MLB: graded slate (Apr 13) — very low signal / high variance for ticket pool.
MLB_EXCLUDED_PROPS = frozenset({"home_runs", "stolen_bases"})

# NHL: Apr 13 graded slate — SOG OVER only (UNDER retained); faceoffs_won both directions.
# Legacy exclusions kept from prior calibration (goals / assists / plus-minus).
NHL_EXCLUDED_PROPS_SOG_OVER_ONLY = frozenset({"shots_on_goal"})
NHL_EXCLUDED_PROPS_ALL_DIRS = frozenset({"faceoffs_won"})
NHL_POOL_EXCLUDE_LEGACY = frozenset({"goals", "assists", "plus/minus"})

SOCCER_EXCLUDED_PROPS = {
    "passes attempted",
    "tackles",
    "fouls",
    "clearances",
}

# Soccer tickets now allow UNDER legs in standard flow; OVER legs are additionally edge-gated.
# Raise per-leg hit floors vs global defaults (still below NBA; soccer hit_rate is often a proxy).
SOCCER_LEG_MIN_HIT_RATE = {
    2: 0.56,
    3: 0.58,
    4: 0.60,
}


def _norm_prop_label(v: object) -> str:
    s = str(v or "").strip().lower()
    s = s.replace("_", " ")
    s = re.sub(r"\s+", " ", s)
    return s


def _fantasy_prop_mask(df: pd.DataFrame) -> pd.Series:
    """True for rows that should be excluded as fantasy props."""
    mask = pd.Series([False] * len(df), index=df.index)
    for col in ("prop_type", "prop", "prop_name"):
        if col in df.columns:
            txt = df[col].astype(str).str.lower()
            mask |= txt.str.contains("fantasy", na=False)
    return mask


def _line_bucket_label(v: object) -> str:
    try:
        x = abs(float(v))
    except Exception:
        return "(missing)"
    if x < 1.5:
        return "micro"
    if x < 5:
        return "low"
    if x < 15:
        return "mid"
    if x < 30:
        return "high"
    return "xl"


def _load_prop_reliability_index(path: str = PROP_RELIABILITY_LATEST_PATH) -> dict[tuple[str, str, str, str], dict]:
    if not path or not os.path.exists(path):
        return {}
    try:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return {}
    rows = payload.get("rows", []) if isinstance(payload, dict) else []
    out: dict[tuple[str, str, str, str], dict] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        k = (
            str(r.get("sport", "")).strip().upper(),
            _norm_prop_label(r.get("prop_type", "")),
            str(r.get("direction", "")).strip().upper(),
            str(r.get("line_bucket", "")).strip().lower(),
        )
        if not k[0] or not k[1] or not k[2] or not k[3]:
            continue
        out[k] = r
    return out


def _apply_reliability_pool_filter(df: pd.DataFrame, reliability_index: dict[tuple[str, str, str, str], dict]) -> tuple[pd.DataFrame, int]:
    if df is None or df.empty:
        return df, 0
    if not reliability_index:
        return df, 0
    if not {"sport", "prop_type", "direction"}.issubset(df.columns):
        return df, 0
    sp = df["sport"].astype(str).str.upper().str.strip()
    prop = df["prop_type"].apply(_norm_prop_label)
    direc = df["direction"].astype(str).str.upper().str.strip()
    bucket = df["line"].apply(_line_bucket_label) if "line" in df.columns else pd.Series("(missing)", index=df.index)
    drop_mask = pd.Series(False, index=df.index)
    for i in df.index:
        row = reliability_index.get((sp.loc[i], prop.loc[i], direc.loc[i], bucket.loc[i]))
        if not row:
            continue
        n = int(float(row.get("decided_n", 0) or 0))
        status = str(row.get("status", "")).strip().upper()
        reliability = float(row.get("reliability_score", 1.0) or 1.0)
        if n >= 40 and (status == "UNRELIABLE" or reliability < 0.35):
            drop_mask.loc[i] = True
    return df.loc[~drop_mask].copy(), int(drop_mask.sum())


def _attach_reliability_columns(df: pd.DataFrame, reliability_index: dict[tuple[str, str, str, str], dict]) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    if not reliability_index:
        out = df.copy()
        out["prop_reliability_score"] = np.nan
        out["prop_reliability_status"] = ""
        return out
    if not {"sport", "prop_type", "direction"}.issubset(df.columns):
        return df
    out = df.copy()
    sp = out["sport"].astype(str).str.upper().str.strip()
    prop = out["prop_type"].apply(_norm_prop_label)
    direc = out["direction"].astype(str).str.upper().str.strip()
    bucket = out["line"].apply(_line_bucket_label) if "line" in out.columns else pd.Series("(missing)", index=out.index)
    scores = []
    statuses = []
    for i in out.index:
        row = reliability_index.get((sp.loc[i], prop.loc[i], direc.loc[i], bucket.loc[i]))
        if row:
            scores.append(float(row.get("reliability_score", np.nan)))
            statuses.append(str(row.get("status", "")).strip().upper())
        else:
            scores.append(np.nan)
            statuses.append("")
    out["prop_reliability_score"] = scores
    out["prop_reliability_status"] = statuses
    return out


def _norm_optional_bucket(v: object) -> str:
    s = str(v or "").strip()
    if not s:
        return "(unknown)"
    return s


def _load_prop_strat_index(path: str = PROP_STRAT_BOARD_LATEST_PATH) -> dict[tuple[str, ...], dict]:
    if not path or not os.path.exists(path):
        return {}
    try:
        payload = json.loads(Path(path).read_text(encoding="utf-8"))
    except Exception:
        return {}
    rows = payload.get("top_trusted_segments", []) if isinstance(payload, dict) else []
    out: dict[tuple[str, ...], dict] = {}
    for r in rows:
        if not isinstance(r, dict):
            continue
        key = (
            str(r.get("sport", "")).strip().upper(),
            _norm_prop_label(r.get("prop_type", "")),
            str(r.get("direction", "")).strip().upper(),
            str(r.get("pick_type", "")).strip().upper(),
            str(r.get("tier", "")).strip().upper(),
            str(r.get("line_bucket", "")).strip().lower(),
            _norm_optional_bucket(r.get("def_tier")),
            _norm_optional_bucket(r.get("h2h_bucket")),
            _norm_optional_bucket(r.get("minutes_tier")),
            _norm_optional_bucket(r.get("role_tier")),
            _norm_optional_bucket(r.get("game_total_bucket")),
        )
        if not key[0] or not key[1] or not key[2]:
            continue
        out[key] = r
    return out


def _attach_strat_columns(df: pd.DataFrame, strat_index: dict[tuple[str, ...], dict]) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    if not strat_index:
        out["strat_hit_rate"] = np.nan
        out["strat_last5_hit_rate"] = np.nan
        out["strat_n"] = np.nan
        return out
    sp = out["sport"].astype(str).str.upper().str.strip() if "sport" in out.columns else pd.Series("", index=out.index)
    prop = out["prop_type"].apply(_norm_prop_label) if "prop_type" in out.columns else pd.Series("", index=out.index)
    direc = out["direction"].astype(str).str.upper().str.strip() if "direction" in out.columns else pd.Series("", index=out.index)
    pick = out["pick_type"].astype(str).str.upper().str.strip() if "pick_type" in out.columns else pd.Series("UNKNOWN", index=out.index)
    tier = out["tier"].astype(str).str.upper().str.strip() if "tier" in out.columns else pd.Series("UNKNOWN", index=out.index)
    lb = out["line"].apply(_line_bucket_label) if "line" in out.columns else pd.Series("(missing)", index=out.index)
    def_tier = out["def_tier"].apply(_norm_optional_bucket) if "def_tier" in out.columns else pd.Series("(unknown)", index=out.index)
    h2h = out["h2h_bucket"].apply(_norm_optional_bucket) if "h2h_bucket" in out.columns else pd.Series("(unknown)", index=out.index)
    minutes = out["minutes_tier"].apply(_norm_optional_bucket) if "minutes_tier" in out.columns else pd.Series("(unknown)", index=out.index)
    role = out["role_tier"].apply(_norm_optional_bucket) if "role_tier" in out.columns else pd.Series("(unknown)", index=out.index)
    ou = out["game_total_bucket"].apply(_norm_optional_bucket) if "game_total_bucket" in out.columns else pd.Series("(unknown)", index=out.index)

    s_hr, s_l5, s_n = [], [], []
    for i in out.index:
        key = (sp.loc[i], prop.loc[i], direc.loc[i], pick.loc[i], tier.loc[i], lb.loc[i], def_tier.loc[i], h2h.loc[i], minutes.loc[i], role.loc[i], ou.loc[i])
        row = strat_index.get(key)
        if not row:
            s_hr.append(np.nan); s_l5.append(np.nan); s_n.append(np.nan); continue
        s_hr.append(float(row.get("hit_rate", np.nan)))
        s_l5.append(float(row.get("last5_hit_rate", np.nan)))
        s_n.append(float(row.get("n", np.nan)))
    out["strat_hit_rate"] = s_hr
    out["strat_last5_hit_rate"] = s_l5
    out["strat_n"] = s_n
    return out


def _pool_prop_snake(pt: object) -> str:
    """Normalize prop_type for pool exclusion keys (e.g. 'Home Runs' -> 'home_runs')."""
    return _norm_prop_label(pt).replace(" ", "_")


def _mlb_ticket_pool_exclusion_mask(df: pd.DataFrame) -> tuple[pd.Series, int]:
    if not {"sport", "prop_type"}.issubset(df.columns):
        return pd.Series(False, index=df.index), 0
    sp = df["sport"].astype(str).str.upper().str.strip()
    ps = df["prop_type"].apply(_pool_prop_snake)
    m = sp.eq("MLB") & ps.isin(MLB_EXCLUDED_PROPS)
    return m, int(m.sum())


def _nhl_ticket_pool_exclusion_mask(df: pd.DataFrame) -> tuple[pd.Series, int, int, int]:
    """Mask rows to drop from NHL (or combined) ticket pools; counts by rule bucket."""
    if not {"sport", "prop_type"}.issubset(df.columns):
        z = pd.Series(False, index=df.index)
        return z, 0, 0, 0
    sp = df["sport"].astype(str).str.upper().str.strip()
    ps = df["prop_type"].apply(_pool_prop_snake)
    nhl = sp.eq("NHL")
    if "direction" in df.columns:
        dir_u = df["direction"].astype(str).str.upper().str.strip()
    else:
        dir_u = pd.Series("", index=df.index)
    m_sog = nhl & ps.isin(NHL_EXCLUDED_PROPS_SOG_OVER_ONLY) & dir_u.eq("OVER")
    m_fc = nhl & ps.isin(NHL_EXCLUDED_PROPS_ALL_DIRS)
    m_leg = nhl & ps.isin(NHL_POOL_EXCLUDE_LEGACY)
    m_all = m_sog | m_fc | m_leg
    return m_all, int(m_sog.sum()), int(m_fc.sum()), int(m_leg.sum())


def _is_attempt_prop_series(prop_s: pd.Series) -> pd.Series:
    return prop_s.astype(str).apply(_norm_prop_label).isin(ATTEMPT_PROPS)


# NBA abbreviations sometimes differ vs other books (align to slate/step8 style).
_CROSSBOOK_NBA_TEAM_ALIASES: dict[str, str] = {
    "BKN": "BRK",
    "CHA": "CHA",
    "PHX": "PHO",
    "GSW": "GS",
    "NOP": "NO",
    "NYK": "NY",
    "SAS": "SA",
    "UTH": "UTA",
    "UTAH": "UTA",
}


def _join_sport_key(sport: object) -> str:
    s = str(sport or "").strip().upper()
    if s in ("NBA1Q", "NBA1H"):
        return "NBA"
    if s == "SOCCER":
        return "SOCCER"
    return s


def _norm_player_join(v: object) -> str:
    s = str(v or "").strip().lower()
    s = unicodedata.normalize("NFKD", s)
    s = "".join(c for c in s if not unicodedata.combining(c))
    s = re.sub(r"[^a-z0-9\s]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = re.sub(r"\b(jr|sr|ii|iii|iv|v)\b", "", s).strip()
    return s


def _norm_team_join(team: object, sport: object) -> str:
    t = str(team or "").strip().upper()
    sp = str(sport or "").strip().upper()
    if sp in ("NBA", "NBA1Q", "NBA1H"):
        return _CROSSBOOK_NBA_TEAM_ALIASES.get(t, t)
    return t


def _ud_join_sport(ud_sid: object) -> str:
    u = str(ud_sid or "").strip().upper()
    m = {"FIFA": "SOCCER", "MASL": "SOCCER", "UFL": "SOCCER", "EPL": "SOCCER"}
    return m.get(u, u)


# Join UD/DK ladder rows to PrizePicks only when this numeric line matches PP ``line``.
_ALT_LINE_MATCH_ATOL = 0.051


def _sport_display_from_join_key(js: object) -> str:
    j = str(js or "").strip().upper()
    if j == "SOCCER":
        return "Soccer"
    return j


def _load_underdog_alt_lines_detail(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    if df.empty:
        return pd.DataFrame()
    df["line"] = pd.to_numeric(df.get("line", np.nan), errors="coerce")
    df = df[df["line"].notna()].copy()
    df["_js"] = df["ud_sport_id"].map(_ud_join_sport)
    df["_jt"] = df["team"].map(lambda x: str(x).strip().upper())
    df["_jp"] = df["player"].map(_norm_player_join)
    df["_jpr"] = df["prop_type"].map(_norm_prop_label)
    df = df[(df["_jp"] != "") & (df["_jpr"] != "")].copy()
    return df


def _load_draftkings_alt_lines_detail(path: str) -> pd.DataFrame:
    df = pd.read_csv(path, dtype=str, encoding="utf-8-sig").fillna("")
    if df.empty:
        return pd.DataFrame()
    df["line"] = pd.to_numeric(df.get("line", np.nan), errors="coerce")
    df = df[df["line"].notna()].copy()
    if "board_sport" in df.columns:
        df["_js"] = df["board_sport"].map(lambda x: _join_sport_key(str(x).strip()))
    else:
        df["_js"] = ""
    lbl = df["dk_market_label"] if "dk_market_label" in df.columns else df.get("prop_type", "")
    df["_jpr"] = lbl.map(_norm_prop_label)
    df["_jt"] = df["team"].map(lambda x: str(x).strip().upper())
    df["_jp"] = df["player"].map(_norm_player_join)
    df = df[(df["_js"] != "") & (df["_jp"] != "") & (df["_jpr"] != "")].copy()
    return df


def _empty_row_from_template(cols: pd.Index) -> dict:
    return {c: np.nan for c in cols}


def _append_alt_book_orphan_rows(
    out: pd.DataFrame,
    detail: pd.DataFrame,
    *,
    book: str,
) -> pd.DataFrame:
    """Append alt-book-only ladder rows that did not line-match any PrizePicks row."""
    if detail is None or detail.empty or "_matched" not in detail.columns:
        return out
    miss = detail.loc[~detail["_matched"]].copy()
    if miss.empty:
        return out
    miss = miss.drop_duplicates(subset=["_js", "_jt", "_jp", "_jpr", "line"], keep="first")
    before = len(out)
    new_rows: list[dict] = []
    cols = out.columns
    for _, r in miss.iterrows():
        base = _empty_row_from_template(cols)
        js = str(r.get("_js") or "").strip().upper()
        sport_disp = _sport_display_from_join_key(js)
        ln = float(r["line"])
        pt = str(r.get("pick_type") or "Standard").strip() or "Standard"
        base.update(
            {
                "sport": sport_disp,
                "tier": "",
                "rank_score": np.nan,
                "player": str(r.get("player") or ""),
                "team": str(r.get("team") or ""),
                "opp": str(r.get("opp_team") or ""),
                "prop_type": str(r.get("prop_type") or ""),
                "pick_type": pt,
                "line": ln,
                "direction": "OVER",
                "edge": np.nan,
                "projection": np.nan,
                "hit_rate": np.nan,
                "game_time": str(r.get("start_time") or ""),
                "pick_platform": book,
                "line_underdog": ln if book == "underdog" else np.nan,
                "line_draftkings": ln if book == "draftkings" else np.nan,
            }
        )
        new_rows.append(base)
    if not new_rows:
        return out
    add = pd.DataFrame(new_rows)
    add = add.reindex(columns=out.columns, fill_value=np.nan)
    out = pd.concat([out, add], ignore_index=True)
    print(f"  [alt-books] {book}: appended {len(out) - before} orphan line row(s) (no PP line match)")
    return out


def attach_alt_book_lines(
    combined: pd.DataFrame,
    *,
    underdog_csv: str = "",
    draftkings_csv: str = "",
) -> pd.DataFrame:
    """
    When UD/DK numeric lines match a PrizePicks row (same player/prop/team ± line), fill
    ``line_underdog`` / ``line_draftkings``. Unmatched alt ladder entries become extra combined rows
    tagged with ``pick_platform`` = ``underdog`` / ``draftkings``.
    """
    out = combined.copy()
    out["_js"] = out["sport"].map(_join_sport_key)
    out["_jt"] = [_norm_team_join(t, s) for t, s in zip(out["team"], out["sport"])]
    out["_jp"] = out["player"].map(_norm_player_join)
    out["_jpr"] = out["prop_type"].map(_norm_prop_label)
    join_on = ["_js", "_jt", "_jp", "_jpr"]

    if "pick_platform" not in out.columns:
        out["pick_platform"] = "prizepicks"
    else:
        out["pick_platform"] = (
            out["pick_platform"].fillna("prizepicks").replace("", "prizepicks").astype(str).str.lower()
        )
        out.loc[out["pick_platform"].str.strip() == "", "pick_platform"] = "prizepicks"

    out["line_underdog"] = np.nan
    out["line_draftkings"] = np.nan

    atol = _ALT_LINE_MATCH_ATOL

    u_path = (underdog_csv or "").strip()
    if u_path and os.path.isfile(u_path):
        try:
            ud_detail = _load_underdog_alt_lines_detail(u_path)
            if not ud_detail.empty:
                ud_detail = ud_detail.copy()
                ud_detail["_matched"] = False
                for i in out.index:
                    if str(out.at[i, "pick_platform"] or "prizepicks").lower() != "prizepicks":
                        continue
                    pp_line = pd.to_numeric(out.at[i, "line"], errors="coerce")
                    if pd.isna(pp_line):
                        continue
                    key = (
                        str(out.at[i, "_js"]),
                        str(out.at[i, "_jt"]),
                        str(out.at[i, "_jp"]),
                        str(out.at[i, "_jpr"]),
                    )
                    sub = ud_detail[
                        (ud_detail["_js"].astype(str) == key[0])
                        & (ud_detail["_jt"].astype(str) == key[1])
                        & (ud_detail["_jp"].astype(str) == key[2])
                        & (ud_detail["_jpr"].astype(str) == key[3])
                        & np.isclose(ud_detail["line"].astype(float), float(pp_line), rtol=0.0, atol=atol)
                    ]
                    if sub.empty:
                        continue
                    j = (sub["line"].astype(float) - float(pp_line)).abs().idxmin()
                    ud_ln = float(sub.loc[j, "line"])
                    out.at[i, "line_underdog"] = ud_ln
                    ud_detail.loc[j, "_matched"] = True
                u_n = int(pd.to_numeric(out["line_underdog"], errors="coerce").notna().sum())
                print(f"  [alt-books] Underdog lines joined (line-matched): {u_n} / {len(out)} rows ({u_path})")
                out = _append_alt_book_orphan_rows(out, ud_detail, book="underdog")
        except Exception as e:
            print(f"  [alt-books] WARN Underdog merge skipped: {e}")

    d_path = (draftkings_csv or "").strip()
    if d_path and os.path.isfile(d_path):
        try:
            dk_detail = _load_draftkings_alt_lines_detail(d_path)
            if not dk_detail.empty:
                dk_detail = dk_detail.copy()
                dk_detail["_matched"] = False
                for i in out.index:
                    if str(out.at[i, "pick_platform"] or "prizepicks").lower() != "prizepicks":
                        continue
                    pp_line = pd.to_numeric(out.at[i, "line"], errors="coerce")
                    if pd.isna(pp_line):
                        continue
                    key = (
                        str(out.at[i, "_js"]),
                        str(out.at[i, "_jt"]),
                        str(out.at[i, "_jp"]),
                        str(out.at[i, "_jpr"]),
                    )
                    sub = dk_detail[
                        (dk_detail["_js"].astype(str) == key[0])
                        & (dk_detail["_jt"].astype(str) == key[1])
                        & (dk_detail["_jp"].astype(str) == key[2])
                        & (dk_detail["_jpr"].astype(str) == key[3])
                        & np.isclose(dk_detail["line"].astype(float), float(pp_line), rtol=0.0, atol=atol)
                    ]
                    if sub.empty:
                        continue
                    j = (sub["line"].astype(float) - float(pp_line)).abs().idxmin()
                    dk_ln = float(sub.loc[j, "line"])
                    out.at[i, "line_draftkings"] = dk_ln
                    dk_detail.loc[j, "_matched"] = True
                d_n = int(pd.to_numeric(out["line_draftkings"], errors="coerce").notna().sum())
                print(f"  [alt-books] DraftKings lines joined (line-matched): {d_n} / {len(out)} rows ({d_path})")
                out = _append_alt_book_orphan_rows(out, dk_detail, book="draftkings")
        except Exception as e:
            print(f"  [alt-books] WARN DraftKings merge skipped: {e}")

    if "line_underdog" not in out.columns:
        out["line_underdog"] = np.nan
    if "line_draftkings" not in out.columns:
        out["line_draftkings"] = np.nan

    out = out.drop(columns=join_on, errors="ignore")
    return out


# Columns produced by add_cross_platform_best_lines (propagated to per-sport slate dataframes).
CROSS_LINE_COLS: tuple[str, ...] = (
    "best_cross_line",
    "best_cross_book",
    "cross_edge_vs_pp",
    "cross_n_books",
)


def add_cross_platform_best_lines(df: pd.DataFrame) -> pd.DataFrame:
    """
    For each prop, compare PrizePicks ``line`` with ``line_underdog`` and ``line_draftkings``.
    For OVER, the best line is the lowest; for UNDER, the highest (among books with data).

    Adds:
      best_cross_line — optimal line for the row's direction
      best_cross_book — PP / UD / DK, or ties like PP+UD (short codes)
      cross_edge_vs_pp — points of line value vs PrizePicks (0 when PP is best among available;
                         NaN when PP line is missing)
      cross_n_books — count of books with a finite line
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    if "line_underdog" not in out.columns:
        out["line_underdog"] = np.nan
    if "line_draftkings" not in out.columns:
        out["line_draftkings"] = np.nan

    pp = pd.to_numeric(out["line"], errors="coerce")
    ud = pd.to_numeric(out["line_underdog"], errors="coerce")
    dk = pd.to_numeric(out["line_draftkings"], errors="coerce")
    mat = pd.DataFrame({"PrizePicks": pp, "Underdog": ud, "DraftKings": dk})

    dir_u = (
        out["direction"].astype(str).str.upper().str.strip()
        if "direction" in out.columns
        else pd.Series("", index=out.index)
    )
    is_under = dir_u == "UNDER"
    order = ("PrizePicks", "Underdog", "DraftKings")
    short = {"PrizePicks": "PP", "Underdog": "UD", "DraftKings": "DK"}

    pick_plat = (
        out["pick_platform"].astype(str).str.lower().str.strip()
        if "pick_platform" in out.columns
        else pd.Series("prizepicks", index=out.index)
    )

    n = len(out)
    best_line = np.full(n, np.nan, dtype=float)
    best_book = np.array([""] * n, dtype=object)
    edge_pp = np.full(n, np.nan, dtype=float)
    n_books = np.zeros(n, dtype=np.int16)

    for i in range(n):
        plat = str(pick_plat.iloc[i] or "prizepicks").strip().lower()
        vals: dict[str, float] = {}
        if plat == "prizepicks":
            for j, name in enumerate(order):
                v = mat.iat[i, j]
                if pd.notna(v) and np.isfinite(float(v)):
                    vals[name] = float(v)
        elif plat == "underdog":
            v = ud.iloc[i] if i < len(ud) else float("nan")
            if pd.notna(v) and np.isfinite(float(v)):
                vals["Underdog"] = float(v)
        elif plat == "draftkings":
            v = dk.iloc[i] if i < len(dk) else float("nan")
            if pd.notna(v) and np.isfinite(float(v)):
                vals["DraftKings"] = float(v)
        else:
            for j, name in enumerate(order):
                v = mat.iat[i, j]
                if pd.notna(v) and np.isfinite(float(v)):
                    vals[name] = float(v)
        n_books[i] = len(vals)
        if not vals:
            continue
        iu = bool(is_under.iloc[i])
        if iu:
            bv = max(vals.values())
            winners = [k for k, v in vals.items() if v == bv]
        else:
            bv = min(vals.values())
            winners = [k for k, v in vals.items() if v == bv]
        win_codes = "+".join(short[w] for w in sorted(winners, key=lambda x: order.index(x)))
        best_line[i] = bv
        best_book[i] = win_codes
        ppv = vals.get("PrizePicks")
        if ppv is not None and np.isfinite(ppv):
            edge_pp[i] = (bv - ppv) if iu else (ppv - bv)

    out["best_cross_line"] = best_line
    out["best_cross_book"] = best_book
    out["cross_edge_vs_pp"] = edge_pp
    out["cross_n_books"] = n_books

    finite_edge = np.isfinite(edge_pp) & (edge_pp > 1e-9)
    n_edge = int(finite_edge.sum())
    _nb = n_books[n_books > 0]
    _mean_b = float(_nb.mean()) if len(_nb) else 0.0
    print(
        f"  [cross-book] rows with cross_edge_vs_pp > 0 vs PrizePicks: {n_edge} / {len(out)} "
        f"(mean books/row with lines: {_mean_b:.2f})"
    )
    return out


def propagate_alt_book_lines_to_sport_frame(
    sport_df: pd.DataFrame | None,
    combined: pd.DataFrame,
    sport_labels: tuple[str, ...],
) -> pd.DataFrame | None:
    """
    Copy line_underdog / line_draftkings / cross-book best-line columns from combined onto each sport slate.
    Join keys match attach_alt_book_lines (team + player + prop norms).
    """
    if sport_df is None or len(sport_df) == 0:
        return sport_df
    out = sport_df.copy()
    if "line_underdog" not in combined.columns:
        out["line_underdog"] = np.nan
        out["line_draftkings"] = np.nan
        for c in CROSS_LINE_COLS:
            if c == "best_cross_book":
                out[c] = ""
            elif c == "cross_n_books":
                out[c] = 0
            else:
                out[c] = np.nan
        return out
    labels = {s.upper() for s in sport_labels}
    sub = combined[combined["sport"].astype(str).str.upper().isin(labels)].copy()
    if sub.empty:
        out["line_underdog"] = np.nan
        out["line_draftkings"] = np.nan
        for c in CROSS_LINE_COLS:
            if c == "best_cross_book":
                out[c] = ""
            elif c == "cross_n_books":
                out[c] = 0
            else:
                out[c] = np.nan
        return out
    sub = sub.copy()
    sub["_jt"] = [_norm_team_join(t, s) for t, s in zip(sub["team"], sub["sport"])]
    sub["_jp"] = sub["player"].map(_norm_player_join)
    sub["_jpr"] = sub["prop_type"].map(_norm_prop_label)
    sub["_jln"] = pd.to_numeric(sub["line"], errors="coerce").round(4)
    agg_map: dict = {"line_underdog": "first", "line_draftkings": "first"}
    for c in CROSS_LINE_COLS:
        if c in sub.columns:
            agg_map[c] = "first"
    if "pick_platform" in sub.columns:
        agg_map["pick_platform"] = "first"
    agg = sub.groupby(["_jt", "_jp", "_jpr", "_jln"], as_index=False).agg(agg_map)
    out["_jt"] = [_norm_team_join(t, s) for t, s in zip(out["team"], out["sport"])]
    out["_jp"] = out["player"].map(_norm_player_join)
    out["_jpr"] = out["prop_type"].map(_norm_prop_label)
    out["_jln"] = pd.to_numeric(out["line"], errors="coerce").round(4)
    out = out.merge(agg, on=["_jt", "_jp", "_jpr", "_jln"], how="left")
    return out.drop(columns=["_jt", "_jp", "_jpr", "_jln"])


def _prop_priority_bonus(v: object) -> float:
    p = _norm_prop_label(v)
    if p in TIER1_PROPS:
        return 0.10
    if p in TIER3_PROPS:
        return -0.10
    return 0.0


# ── Goblin / Demon payout adjustment ─────────────────────────────────────────
# Goblin: line moved in your favor → REDUCES payout LINEARLY per unit of
#   line_discount_vs_standard.  Each 0.5pt of line movement cuts payout by
#   GOBLIN_REDUCTION_PER_UNIT * 0.5.  Tune this constant against real
#   PrizePicks payout tables once you have more graded data.
#   Current value: 0.18 per unit → 0.5pt goblin ≈ 9% payout reduction per leg.
GOBLIN_REDUCTION_PER_UNIT: float = 0.18   # ← tune me
GOBLIN_MAX_REDUCTION:      float = 0.60   # single-leg reduction cap (60%)

# Demon: line moved against you → BOOSTS payout MULTIPLICATIVELY per unit.
#   DEMON_BOOST_PER_UNIT * discount applied as a multiplier per leg.
#   Current value: 0.28 per unit → 0.5pt demon ≈ 14% payout boost per leg.
DEMON_BOOST_PER_UNIT: float = 0.28        # ← tune me
DEMON_MAX_BOOST:      float = 2.50        # single-leg multiplier cap (2.5×)

# Min EV for generation (ev_power = est_win_prob × adj_power) and for positive-EV web gate
# (see _ticket_passes_positive_ev_gate). Longer slips need higher EV vs variance.
MIN_TICKET_EV_DEFAULT: float = 0.80
MIN_TICKET_EV_BY_LEGS: dict[int, float] = {
    2: 1.00,
    3: 1.00,
    4: 1.10,
    5: 1.25,
    6: 1.50,
}

# /tickets: hide slips at or below this modeled all-hit win prob (clutter / lottery tickets).
MIN_WEB_DISPLAY_EST_WIN_PROB: float = 0.06

# /tickets: hard floor for displayed/generated slip payout multipliers.
MIN_WEB_PAYOUT_X: float = 3.0
MIN_WEB_PAYOUT_X_GOBLIN_SHORT: float = 2.0
DEBUG_PAYOUT_DIAGNOSTIC: bool = os.getenv("PROPORACLE_DEBUG_PAYOUT", "false").lower() == "true"

# /tickets page target volumes per sport after EV gate.
WEB_TICKET_TEMPLATE_BY_LEGS: dict[int, int] = {
    6: 6,
    5: 8,
    4: 12,
    3: 12,
    2: 12,
}

# Cap sorted candidate pool size per leg count (top rows by ticket sort) to bound greedy work.
MAX_TICKET_POOL_ROWS_BY_LEG_COUNT: dict[int, int] = {
    2: 50_000,
    3: 50_000,
    4: 30_000,
    5: 20_000,
    6: 10_000,
}

# Max legs for FINAL / cross-sport builders (argparse default matches).
MAX_TICKET_LEGS: int = 6

LEG_PROB_FLOOR = 0.35
# Source-aware caps (ML > hit-rate > rank-score fallback)
LEG_PROB_CAPS = {
    "ml_prob": 0.92,
    "ml_prob_demon": 0.72,
    "rank_score": 0.72,
    "edge": 0.70,
    "hit_rate": 0.72,
    "hit_rate_demon": 0.75,
    "fallback_const": 0.65,
}
TICKET_PROB_FLOOR = 1e-6
TICKET_PROB_CAP = 0.999
RANK_SCORE_SIGMOID_SCALE = 0.4
DEFAULT_LEG_PROB_FALLBACK = 0.50

# Limit how many distinct generated slips may include the same player (reduces single-leg cascade risk).
MAX_SLIPS_PER_PLAYER = 4


def _ticket_cap_players_from_rows(rows: list) -> set[str]:
    out: set[str] = set()
    for r in rows:
        p = _norm_player_join(r.get("player"))
        if p:
            out.add(p)
    return out


def _ticket_cap_can_add(rows: list, counts: dict[str, int] | None, cap: int = MAX_SLIPS_PER_PLAYER) -> bool:
    if counts is None or cap <= 0:
        return True
    for p in _ticket_cap_players_from_rows(rows):
        if int(counts.get(p, 0)) >= cap:
            return False
    return True


def _ticket_cap_register(rows: list, counts: dict[str, int] | None) -> None:
    if not counts:
        return
    for p in _ticket_cap_players_from_rows(rows):
        counts[p] = int(counts.get(p, 0)) + 1


def min_ev_for_ticket(n_legs: int) -> float:
    return float(MIN_TICKET_EV_BY_LEGS.get(int(n_legs), MIN_TICKET_EV_DEFAULT))


def _trim_pool_by_leg_count(df: pd.DataFrame, n_legs: int) -> pd.DataFrame:
    if df is None or len(df) == 0:
        return df
    cap = int(MAX_TICKET_POOL_ROWS_BY_LEG_COUNT.get(int(n_legs), 50_000))
    if len(df) <= cap:
        return df
    return df.iloc[:cap].reset_index(drop=True)


def _ticket_n_legs(ticket: dict) -> int:
    """Leg count for EV gates / stats (payload uses 'legs'; builders use 'rows')."""
    n = ticket.get("n_legs")
    if n is not None:
        try:
            return max(1, int(n))
        except (TypeError, ValueError):
            pass
    legs = ticket.get("legs") if isinstance(ticket.get("legs"), list) else None
    rows = ticket.get("rows") if isinstance(ticket.get("rows"), list) else None
    seq = legs if legs else rows
    if seq:
        return len(seq)
    return 2

def _norm_team_abbr(v: object) -> str:
    return str(v or "").strip().upper()


def calc_adjusted_payout(base_payout: float, legs: list) -> float:
    """
    Adjusts base_payout for Goblin (linear reduction) and Demon
    (multiplicative boost) legs based on how far each line was moved
    from the Standard reference (line_discount_vs_standard).

    Goblin legs reduce payout linearly:
        reduction = min(discount * GOBLIN_REDUCTION_PER_UNIT, GOBLIN_MAX_REDUCTION)
        multiplier *= (1 - reduction)

    Demon legs boost payout multiplicatively:
        boost = min(1 + discount * DEMON_BOOST_PER_UNIT, DEMON_MAX_BOOST)
        multiplier *= boost

    Standard legs: no adjustment (multiplier unchanged).

    Returns adjusted payout rounded to 2 decimal places.
    """
    multiplier = 1.0
    for leg in legs:
        pt = str(leg.get("pick_type", "")).strip()
        raw_discount = leg.get("line_discount_vs_standard")
        try:
            discount = abs(float(raw_discount))
        except (TypeError, ValueError):
            discount = 0.0

        if pt == "Goblin":
            reduction = min(discount * GOBLIN_REDUCTION_PER_UNIT, GOBLIN_MAX_REDUCTION)
            multiplier *= max(0.0, 1.0 - reduction)
        elif pt == "Demon":
            boost = min(1.0 + discount * DEMON_BOOST_PER_UNIT, DEMON_MAX_BOOST)
            multiplier *= boost
        # Standard: no change

    return round(base_payout * multiplier, 2)


def enrich_ticket_curve_payouts(ticket: dict, stake_unit: float = 1.0) -> None:
    """
    Adds flat_multiplier, est_multiplier, mult_error, EV columns using
    utils.goblin_demon_multiplier (delta_pct from line / standard_line per leg).
    Mutates ticket dict in place.
    """
    rows = ticket.get("rows") or []
    n = int(ticket.get("n_legs", len(rows)) or 0) or len(rows)
    legs_payload: list[dict] = []
    using_fb = False
    for r in rows:
        rd = r if isinstance(r, dict) else dict(r)
        sl = rd.get("standard_line")
        ln = rd.get("line")
        dp = gd_leg_delta_pct(ln, sl)
        pt = str(rd.get("pick_type") or "Standard")
        pl = pt.lower()
        if ("goblin" in pl or "demon" in pl) and dp is None:
            using_fb = True
        pr = rd.get("leg_prob_used")
        if pr is None:
            pr = rd.get("ml_prob")
        legs_payload.append(
            {
                "pick_type": pt,
                "delta_pct": dp,
                "line": ln,
                "standard_line": sl,
                "prob": pr,
            }
        )
    summ = gd_multiplier_summary(legs_payload, mode="power", stake=float(stake_unit))
    flat_p = float(summ.get("flat_mult") or 0.0)
    est_p = float(summ.get("est_mult") or flat_p)
    mult_err = float(summ.get("mult_delta") or 0.0)
    ticket["flat_multiplier"] = round(flat_p, 4)
    ticket["est_multiplier"] = round(est_p, 4)
    ticket["mult_error"] = round(mult_err, 4)
    ticket["est_payout"] = round(est_p * float(stake_unit), 4)
    ticket["flat_payout"] = round(flat_p * float(stake_unit), 4)
    ticket["payout_delta"] = round(ticket["est_payout"] - ticket["flat_payout"], 4)
    cp = summ.get("combined_prob")
    ticket["combined_hit_prob_curve"] = cp
    if cp is not None:
        ticket["est_ev"] = round(gd_compute_ticket_ev(est_p, float(cp), float(stake_unit)), 4)
        ticket["flat_ev"] = round(gd_compute_ticket_ev(flat_p, float(cp), float(stake_unit)), 4)
    if n >= 3:
        summ_f = gd_multiplier_summary(legs_payload, mode="flex", hits=n, stake=float(stake_unit))
        ticket["est_multiplier_flex_nn"] = summ_f.get("est_mult")
        ticket["flat_multiplier_flex_nn"] = summ_f.get("flat_mult")
    ticket["using_flat_fallback"] = bool(using_fb)
    try:
        emp_legs: list[dict] = []
        for r in rows:
            rd = r if isinstance(r, dict) else dict(r)
            pt_raw = str(rd.get("pick_type") or "Standard")
            pll = pt_raw.lower()
            if "goblin" in pll:
                pt_e = "goblin"
            elif "demon" in pll:
                pt_e = "demon"
            else:
                pt_e = "standard"
            ld = 0.0
            try:
                sl = rd.get("standard_line")
                ln = rd.get("line")
                if sl is not None and ln is not None and str(sl).strip() != "" and str(ln).strip() != "":
                    ld = abs(float(sl) - float(ln))
            except (TypeError, ValueError):
                ld = 0.0
            pr = rd.get("leg_prob_used")
            if pr is None:
                pr = rd.get("ml_prob")
            try:
                prf = float(pr)
            except (TypeError, ValueError):
                prf = 0.52
            if not (0.0 < prf <= 1.0):
                prf = 0.52
            emp_legs.append({"pick_type": pt_e, "line_distance": ld, "hit_prob": prf})
        flow = str(ticket.get("flow") or "power").strip().lower()
        tt = "flex" if flow == "flex" else "power"
        emp = compute_ticket_ev(emp_legs, tt, n)
        ticket["empirical_ev"] = emp["ev"]
        ticket["empirical_first_place"] = emp["first_place_payout"]
        ticket["empirical_min_guarantee"] = emp["min_guarantee"]
        ticket["empirical_min_guarantee_adjustment"] = emp["min_guarantee_adjustment"]
        ticket["empirical_recommendation"] = emp["recommendation"]
    except Exception:
        pass
    if abs(mult_err) > 1.0:
        _log_slate.warning(
            "Ticket curve mult_error %.4f (>1.0 vs flat base): %s-leg slip",
            mult_err,
            n,
        )


# ── Per-leg count quality thresholds (used by smart ticket builder) ───────────
# Min hit rate required per leg depending on ticket length
# Longer tickets need higher floor because win prob = product of all hit rates
LEG_MIN_HIT_RATE = {
    2: 0.65,
    3: 0.67,
    4: 0.69,
    5: 0.71,
    6: 0.73,
}

MLB_LEG_MIN_HIT_RATE = {
    2: 0.58,
    3: 0.60,
    4: 0.62,
}

# NHL: smaller quality pool vs NBA — cap per-leg floors so structured/FINAL builders can fill slips.
NHL_LEG_MIN_HIT_RATE = {
    2: 0.55,
    3: 0.57,
    4: 0.60,
}
# Soccer OVER legs have materially weaker realized performance; require stronger edge or drop.
# TODO: confirm 0.60 vs 0.65 once post-fix Soccer graded sample grows.
SOCCER_OVER_MIN_EDGE = 0.0  # Edge is leaky for Soccer; gate disabled until replacement signal exists

# ── Model-performance ticket gates (Track A + auto-gate from tracker) ─────────
# REVERT NBA1H WHEN: model_gate_recommendations.json NBA1H block has
# consecutive_days_above_052 >= 3 (written by track_model_performance.py --nba1h-monitor).
# HOW TO REVERT: Set gate false on NBA1H in model_gate_recommendations.json, or wait for streak.
# NBA1H props still flow through pipeline, slate explorer, and graded archive.
ALWAYS_ALLOW_SPORTS = frozenset({"NBA", "MLB"})
NBA1H_TICKET_GATE = True  # kept for reference; ticket block driven by _nba1h_gated() + JSON
NBA1H_TICKET_GATE_MIN_AUC = 0.52  # revert threshold; streak tracked in consecutive_days_above_052
_MODEL_GATE_RECOMMENDATIONS_PATH = os.path.join(REPO_ROOT, "data", "model_gate_recommendations.json")
_MODEL_GATE_CACHE: dict[str, dict] | None = None
_MODEL_GATE_LOGGED: set[str] = set()


def _load_model_gate_recommendations() -> dict[str, dict]:
    path = Path(_MODEL_GATE_RECOMMENDATIONS_PATH)
    if not path.is_file():
        return {}
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
        return data if isinstance(data, dict) else {}
    except Exception:
        return {}


def _nba1h_ticket_gate_reason() -> str:
    """Reason string for NBA1H ticket gate (reads live AUC from gate JSON)."""
    global _MODEL_GATE_CACHE
    if _MODEL_GATE_CACHE is None:
        _MODEL_GATE_CACHE = _load_model_gate_recommendations()
    rec = (_MODEL_GATE_CACHE or {}).get("NBA1H") or {}
    auc = rec.get("auc")
    if isinstance(auc, (int, float)):
        return (
            f"AUC {float(auc):.4f} — below {NBA1H_TICKET_GATE_MIN_AUC:.2f} revert threshold"
        )
    return f"below {NBA1H_TICKET_GATE_MIN_AUC:.2f} revert threshold (see model_performance_log)"


def _nba1h_gated(gate_recs: dict) -> bool:
    """True when NBA1H legs must not enter ticket pools (JSON + streak/AUC unblock)."""
    block = (gate_recs or {}).get("NBA1H") or {}
    streak = int(block.get("consecutive_days_above_052") or 0)
    auc = block.get("rolling_30d_auc")
    if streak >= 3 and auc is not None and float(auc) >= NBA1H_TICKET_GATE_MIN_AUC:
        return False
    return bool(block.get("gate", True))


def _sport_ticket_gated(sport: str) -> tuple[bool, str]:
    """True if sport must not enter EV / win-rate ticket pools (slate unchanged)."""
    su = str(sport or "").strip().upper()
    if not su or su in ALWAYS_ALLOW_SPORTS:
        return False, ""
    if su == "NFL" and NFL_TICKET_GATE:
        return True, NFL_TICKET_GATE_REASON
    global _MODEL_GATE_CACHE
    if _MODEL_GATE_CACHE is None:
        _MODEL_GATE_CACHE = _load_model_gate_recommendations()
    if _nba1h_gated(_MODEL_GATE_CACHE or {}) and su == "NBA1H":
        return True, _nba1h_ticket_gate_reason()
    rec = (_MODEL_GATE_CACHE or {}).get(su)
    if isinstance(rec, dict) and rec.get("gate"):
        return True, str(rec.get("reason") or "model performance gate")
    return False, ""


def _log_auto_gate_once(sport: str, reason: str) -> None:
    global _MODEL_GATE_LOGGED
    su = str(sport or "").strip().upper()
    if su in _MODEL_GATE_LOGGED:
        return
    _MODEL_GATE_LOGGED.add(su)
    print(f"  [AUTO-GATE] {su} gated by model performance tracker — {reason}")


DIRECTIONAL_HR_THRESHOLDS: dict[str, dict[str, float]] = {
    "NBA": {"over": 0.70, "under": 0.30, "standard_over_min_edge": 2.45, "standard_under_min_edge": 1.33},
    "NBA1Q": {"over": 0.65, "under": 0.35, "standard_over_min_edge": 2.45, "standard_under_min_edge": 1.33},
    "NBA1H": {"over": 0.65, "under": 0.35, "standard_over_min_edge": 2.45, "standard_under_min_edge": 1.33},
    "NHL": {"over": 0.65, "under": 0.35, "standard_over_min_edge": 0.5, "standard_under_min_edge": 0.5},
    "MLB": {"over": 0.60, "under": 0.40, "standard_over_min_edge": 0.5, "standard_under_min_edge": 0.5},
    "TENNIS": {"over": 0.60, "under": 0.40},
    "SOCCER": {"over": 0.60, "under": 0.40, "standard_over_min_rank_score": 0.25},
    "CBB": {"over": 0.65, "under": 0.35},
    "WCBB": {"over": 0.65, "under": 0.35},
}
DEFAULT_DIRECTIONAL_THRESHOLD: dict[str, float] = {"over": 0.65, "under": 0.35}
MLB_MAX_LEGS = 4
MLB_PITCHING_OVER_ONLY_PROPS = {"strikeouts", "hits allowed"}

# Tennis: short slips only (max 3 legs). Relaxed per-leg floors vs NBA — no graded PP history yet.
MAX_LEGS_TENNIS = 3
TENNIS_LEG_MIN_HIT_RATE = {2: 0.55, 3: 0.58, 4: 0.62}

# Pipelines that emit step8 boards into combined slate (reference for docs / tooling).
ACTIVE_SPORTS = ("NBA", "NHL", "SOCCER", "TENNIS", "WNBA", "MLB", "NBA1H", "NBA1Q", "WCBB", "NFL", "CFB")
# NFL — Phase 1 scaffold only; keep off slate until step8 + historical hit rates exist (Sept 2026).
# Reference: {"NFL": False}  # activate September 2026 — do not add "NFL" to ACTIVE_SPORTS yet.
NFL_TICKET_GATE = True
NFL_TICKET_GATE_REASON = "Off-season scaffold — activate Week 1 2026"

# When --high-conviction: per-leg hit_rate floors (merged with LEG_MIN_HIT_RATE via max())
HIGH_CONVICTION_LEG_MIN_HIT_RATE = {
    2: 0.65,
    3: 0.67,
    4: 0.69,
    5: 0.71,
    6: 0.73,
}

# With --prioritize-ticket-hit: extra leg floors (merged via max() on top of strict pools).
PRIORITIZE_TICKET_HIT_LEG_MIN = {
    2: 0.75,
    3: 0.77,
    4: 0.79,
    5: 0.81,
    6: 0.83,
}

# Modeled probability floors (after leg-prob caps + correlation penalty). Tune if too few tickets emit.
MIN_PRIORITIZE_MODELED_POWER_WIN_PROB = 0.38
MIN_PRIORITIZE_MODELED_FLEX_CASH_PROB = 0.52


def effective_leg_min_hit_rates(
    high_conviction: bool,
    override: dict[int, float] | None = None,
    prioritize_ticket_hit: bool = False,
) -> dict[int, float]:
    """Per ticket length: minimum hit_rate (0–1) for each leg in build_tickets / FINAL groups."""
    out: dict[int, float] = {}
    for n in TICKET_LEG_SIZES:
        base = float(LEG_MIN_HIT_RATE.get(n, 0.55))
        if high_conviction:
            base = max(base, float(HIGH_CONVICTION_LEG_MIN_HIT_RATE.get(n, 0.70)))
        if prioritize_ticket_hit:
            base = max(base, float(PRIORITIZE_TICKET_HIT_LEG_MIN.get(n, 0.75)))
        if override and n in override:
            base = max(base, float(override[n]))
        out[n] = base
    return out


def ticket_leg_sizes_for_max(max_legs: int) -> list:
    m = int(max_legs)
    if m < MIN_TICKET_POOL:
        m = MIN_TICKET_POOL
    return [n for n in TICKET_LEG_SIZES if n <= m]

# Min tier per leg count for Power mode tickets
POWER_MIN_TIER = {
    2: ["A", "B", "C"],
    3: ["A", "B", "C"],   # 3-leg power: Tier A/B/C ok
    4: ["A", "B"],
    5: ["A", "B"],         # 5-leg power: Tier A/B only
    6: ["A", "B"],         # 6-leg power: Tier A/B only
}

# Pick-type tier policy for NBA-family tickets.
# Keep broad tier eligibility; only hard-exclude Demon OVER.
# Key: (pick_type, direction) in uppercase.
ELIGIBLE_TIERS_BY_PICK_TYPE_DIRECTION: dict[tuple[str, str], set[str]] = {
    ("DEMON", "OVER"): set(),
}
TIER_POLICY_SPORTS = {"NBA", "NBA1H", "NBA1Q"}

# Cap fantasy-score concentration per ticket so slips are more diversified.
MAX_FANTASY_LEGS = {
    2: 1,
    3: 1,
    4: 2,
    5: 2,
    6: 2,
}
MAX_SAME_PROP_TYPE_PER_TICKET = 2

# Ticket leg counts written to workbook + FINAL web payload
TICKET_LEG_SIZES = [2, 3, 4, 5, 6]
MIN_TICKET_POOL = min(TICKET_LEG_SIZES)


def _is_fantasy_prop(row: pd.Series) -> bool:
    return "fantasy" in str(row.get("prop_type", "")).strip().lower()


def _ticket_prop_token(row: pd.Series | dict) -> str:
    raw = ""
    if isinstance(row, pd.Series):
        raw = str(row.get("prop_type", "") or "")
    elif isinstance(row, dict):
        raw = str(row.get("prop_type", "") or "")
    return _norm_prop_label(raw) or raw.strip().lower()


def _can_add_row_with_prop_cap(
    row: pd.Series | dict,
    prop_type_counts: Counter[str],
    max_same_prop: int = MAX_SAME_PROP_TYPE_PER_TICKET,
) -> bool:
    tok = _ticket_prop_token(row)
    if not tok:
        return True
    return int(prop_type_counts.get(tok, 0)) < int(max_same_prop)

# Demon legs are only allowed in Flex-mode analysis (too low hit rate for Power)
# This is enforced in build_tickets_smart() below


# ── Excel style helpers ───────────────────────────────────────────────────────
def side(color: str = "CCCCCC") -> Border:
    s = Side(style="thin", color=color)
    return Border(left=s, right=s, top=s, bottom=s)


# ── Style object caches (avoid recreating identical objects per cell) ──────────
_font_cache: dict = {}
_fill_cache: dict = {}
_align_cache: dict = {}
_border_obj = None

def _side_obj():
    global _border_obj
    if _border_obj is None:
        _border_obj = side()
    return _border_obj

def _font(bold=False, color="000000", sz=9):
    key = (bold, color, sz)
    if key not in _font_cache:
        _font_cache[key] = Font(bold=bold, name="Arial", size=sz, color=color)
    return _font_cache[key]

def _fill(bg):
    if bg not in _fill_cache:
        _fill_cache[bg] = PatternFill("solid", start_color=bg)
    return _fill_cache[bg]

def _align(horizontal="center", wrap=False):
    key = (horizontal, wrap)
    if key not in _align_cache:
        _align_cache[key] = Alignment(horizontal=horizontal, vertical="center", wrap_text=wrap)
    return _align_cache[key]


def hc(ws, r, c, v, bg=None, fc="FFFFFF", bold=True, sz=9, align="center"):
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _font(bold=bold, color=fc, sz=sz)
    if bg:
        cell.fill = _fill(bg)
    cell.alignment = _align(horizontal=align, wrap=True)
    cell.border = _side_obj()
    return cell


def dc(ws, r, c, v, bg=None, bold=False, sz=9, align="center", fc="000000", fmt=None):
    if v is pd.NA or (isinstance(v, float) and np.isnan(v)) or v is None:
        v = ""
    cell = ws.cell(row=r, column=c, value=v)
    cell.font = _font(bold=bold, color=fc, sz=sz)
    cell.fill = _fill(bg or C["white"])
    cell.alignment = _align(horizontal=align)
    cell.border = _side_obj()
    if fmt:
        cell.number_format = fmt
    return cell


def sw(ws, widths: List[int]):
    for ci, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w


def tier_bg(t) -> str:
    return {"A": C["tier_a"], "B": C["tier_b"], "C": C["tier_c"], "D": C["tier_d"]}.get(
        str(t).upper(), C["white"]
    )


def pt_bg(pt) -> str:
    return {"Goblin": C["goblin"], "Demon": C["demon"], "Standard": C["standard"]}.get(pt, C["white"])


def hr_bg(v) -> str:
    if v is None or (isinstance(v, float) and np.isnan(v)):
        return "DDDDDD"
    if v >= 0.65:
        return C["hit"]
    if v >= 0.50:
        return C["push"]
    return C["miss"]


def pct_cell(ws, r, c, val):
    nan = val is None or (isinstance(val, float) and np.isnan(val))
    bg = hr_bg(val) if not nan else "DDDDDD"
    cell = dc(ws, r, c, val if not nan else "", bg=bg, bold=True)
    if not nan:
        cell.number_format = "0%"
        cell.font = _font(bold=True, color="FFFFFF", sz=9)
    return cell


def _signal_float(v):
    """Parse numeric for bet-signal / context scoring (shared HTML + Excel)."""
    try:
        if v is None or v == "":
            return None
        if isinstance(v, float) and np.isnan(v):
            return None
        return float(v)
    except Exception:
        return None


def _clip_prob(p: float, source: str = "hit_rate") -> float:
    """
    Clip probability to floor/cap bounds.
    Caps are source-aware so ML probabilities are not suppressed as aggressively as
    hit-rate/rank-score derived fallbacks.
    """
    if p is None or not np.isfinite(p):
        return DEFAULT_LEG_PROB_FALLBACK
    src = str(source or "").strip()
    cap = float(LEG_PROB_CAPS.get(src, LEG_PROB_CAPS["hit_rate"]))
    return float(np.clip(float(p), LEG_PROB_FLOOR, cap))


def _rank_score_to_prob(rank_score: float) -> float:
    if rank_score is None or not np.isfinite(rank_score):
        return DEFAULT_LEG_PROB_FALLBACK
    prob = 1.0 / (1.0 + math.exp(-float(rank_score) * RANK_SCORE_SIGMOID_SCALE))
    return float(prob)


def _scalar_rank_to_prob_for_sort(x: object) -> float:
    try:
        if x is None or pd.isna(x):
            return DEFAULT_LEG_PROB_FALLBACK
        xf = float(x)
        if not np.isfinite(xf):
            return DEFAULT_LEG_PROB_FALLBACK
        return _rank_score_to_prob(xf)
    except (TypeError, ValueError):
        return DEFAULT_LEG_PROB_FALLBACK


def _attach_ticket_pick_order(df: pd.DataFrame, mode: str) -> pd.DataFrame:
    """
    Add __ts_pri / __ts_sec for descending sort when assembling tickets.
    rank: primary = rank_score; ml: primary = ml_prob (missing last);
    blend: 0.5*ml + 0.5*sigmoid(rank), with missing ml falling back to sigmoid(rank) only.
    rule: user ruleset (L5, L10/season consistency, defense, minutes, role, H2H, edge),
    with Demon legs deprioritized.
    """
    out = df.copy()
    if out.empty:
        out["__ts_pri"] = pd.Series(dtype=float)
        out["__ts_sec"] = pd.Series(dtype=float)
        return out
    rs = pd.to_numeric(out["rank_score"], errors="coerce") if "rank_score" in out.columns else pd.Series(np.nan, index=out.index)
    ml = pd.to_numeric(out["ml_prob"], errors="coerce") if "ml_prob" in out.columns else pd.Series(np.nan, index=out.index)
    rs_p = rs.map(_scalar_rank_to_prob_for_sort)
    m = (mode or "rank").strip().lower()
    if m not in ("rank", "ml", "blend", "rule"):
        m = "rank"
    if m == "ml":
        out["__ts_pri"] = ml.fillna(-1.0)
        out["__ts_sec"] = rs.fillna(-1e9)
    elif m == "blend":
        ml_b = ml.where(ml.notna(), rs_p)
        out["__ts_pri"] = 0.5 * ml_b + 0.5 * rs_p
        out["__ts_sec"] = rs.fillna(-1e9)
    elif m == "rule":
        # User methodology:
        # 1) Remove/deprioritize Demon
        # 2) Prefer L5 side hit count >= 4
        # 3) Consistency from L10 + L5 vs season average
        # 4) Defense/mintier directional context
        # 5) Role context
        # 6) H2H consistency
        # 7) Edge as a smaller tie-breaker
        direction = out.get("direction", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()
        pick_type = out.get("pick_type", pd.Series("Standard", index=out.index)).astype(str).str.upper().str.strip()
        def_tier_raw = out.get("def_tier", pd.Series("", index=out.index))
        def_tier = def_tier_raw.map(_norm_def_tier_cell_upper)
        min_tier = out.get("min_tier", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()
        role = (
            out.get("usage_role", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()
            + " "
            + out.get("shot_role", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()
        ).str.strip()

        nan_s = pd.Series(np.nan, index=out.index)
        l5_over = pd.to_numeric(out.get("l5_over", nan_s), errors="coerce")
        l5_under = pd.to_numeric(out.get("l5_under", nan_s), errors="coerce")
        l10_over = pd.to_numeric(out.get("l10_over", nan_s), errors="coerce")
        l10_under = pd.to_numeric(out.get("l10_under", nan_s), errors="coerce")
        l5_avg = pd.to_numeric(out.get("l5_avg", nan_s), errors="coerce")
        szn_avg = pd.to_numeric(out.get("szn_avg", nan_s), errors="coerce")
        h2h_over = pd.to_numeric(out.get("h2h_over_pct", nan_s), errors="coerce")
        h2h_gp = pd.to_numeric(out.get("h2h_gp", nan_s), errors="coerce")
        edge = pd.to_numeric(out.get("edge", nan_s), errors="coerce")
        mlp = pd.to_numeric(out.get("ml_prob", nan_s), errors="coerce")

        # Do not anchor ticket ordering on hit_rate.
        pri = mlp.where(mlp.notna(), rs_p).fillna(DEFAULT_LEG_PROB_FALLBACK)

        side_l5 = np.where(direction.eq("UNDER"), l5_under, l5_over)
        side_l10 = np.where(direction.eq("UNDER"), l10_under, l10_over)

        # L5 threshold behavior (>=4 strong, <=2 weak)
        pri = pri + np.where(pd.isna(side_l5), 0.0, np.where(side_l5 >= 4, 0.10, np.where(side_l5 <= 2, -0.07, 0.0)))
        # L10 consistency
        pri = pri + np.where(pd.isna(side_l10), 0.0, np.where(side_l10 >= 7, 0.05, np.where(side_l10 >= 6, 0.02, np.where(side_l10 <= 4, -0.03, 0.0))))
        # L5 vs season average stability
        avg_diff = (l5_avg - szn_avg).abs()
        pri = pri + np.where(pd.isna(avg_diff), 0.0, np.where(avg_diff <= 0.5, 0.03, np.where(avg_diff <= 1.0, 0.015, np.where(avg_diff >= 2.0, -0.02, 0.0))))

        over_mask = direction.eq("OVER")
        under_mask = direction.eq("UNDER")

        # Defense directional preference
        pri = pri + np.where(over_mask & def_tier.eq("WEAK"), 0.05, 0.0)
        pri = pri + np.where(over_mask & def_tier.isin(["ABOVE AVG", "ELITE"]), -0.03, 0.0)
        pri = pri + np.where(under_mask & def_tier.isin(["ELITE", "ABOVE AVG"]), 0.04, 0.0)
        pri = pri + np.where(under_mask & def_tier.eq("WEAK"), -0.04, 0.0)

        # Minutes directional preference
        pri = pri + np.where(over_mask & min_tier.isin(["HIGH"]), 0.04, 0.0)
        pri = pri + np.where(over_mask & min_tier.isin(["LOW"]), -0.04, 0.0)
        pri = pri + np.where(under_mask & min_tier.isin(["LOW"]), 0.04, 0.0)
        pri = pri + np.where(under_mask & min_tier.isin(["HIGH"]), -0.04, 0.0)

        # Role preference
        pri = pri + np.where(role.str.contains("PRIMARY", regex=False, na=False) & over_mask, 0.03, 0.0)
        pri = pri + np.where(role.str.contains("PRIMARY", regex=False, na=False) & under_mask, -0.01, 0.0)
        pri = pri + np.where(role.str.contains("SECONDARY", regex=False, na=False), 0.015, 0.0)
        pri = pri + np.where(role.str.contains("SUPPORT", regex=False, na=False) & under_mask, 0.015, 0.0)
        pri = pri + np.where(role.str.contains("SUPPORT", regex=False, na=False) & over_mask, -0.015, 0.0)

        # H2H (requires small but non-trivial sample)
        valid_h2h = h2h_gp.fillna(0) >= 2
        pri = pri + np.where(valid_h2h & over_mask & (h2h_over >= 0.60), 0.03, 0.0)
        pri = pri + np.where(valid_h2h & over_mask & (h2h_over <= 0.40), -0.03, 0.0)
        pri = pri + np.where(valid_h2h & under_mask & (h2h_over <= 0.40), 0.03, 0.0)
        pri = pri + np.where(valid_h2h & under_mask & (h2h_over >= 0.60), -0.03, 0.0)

        # Edge last, with directional sign handling:
        # OVER: higher edge is better; UNDER: lower (more negative) edge is better.
        edge_adj = pd.Series(np.where(over_mask, edge, -edge), index=out.index)
        edge_adj = pd.to_numeric(edge_adj, errors="coerce").fillna(0.0).clip(-12.0, 12.0) / 40.0
        pri = pri + edge_adj

        # Explicitly deprioritize Demon as requested.
        pri = np.where(pick_type.eq("DEMON"), -9.0, pri)

        pri_s = pd.Series(pri, index=out.index)
        out["__ts_pri"] = pd.to_numeric(pri_s, errors="coerce").fillna(-9.0)
        out["__ts_sec"] = rs.fillna(-1e9)
    else:
        out["__ts_pri"] = rs.fillna(-1e9)
        out["__ts_sec"] = ml.fillna(-1.0)
    return out


def get_edge_threshold(sport: str, prop_type: str, pick_type: str) -> float:
    """
    Edge threshold used to center raw edge -> probability conversion.
    Kept simple for now; can be expanded to sport/prop/pick specific thresholds.
    """
    _ = (sport, prop_type, pick_type)
    return 0.0


def _to_prob_0_1(v: Any) -> float | None:
    p = pd.to_numeric(v, errors="coerce")
    if pd.isna(p):
        return None
    pf = float(p)
    if pf > 1.0 and pf <= 100.0:
        pf = pf / 100.0
    if 0.0 <= pf <= 1.0:
        return pf
    return None


def _demon_passes_quality_gate(row: pd.Series | dict) -> bool:
    getv = row.get if hasattr(row, "get") else (lambda _k, _d=None: _d)
    hr = _to_prob_0_1(getv("hit_rate"))
    if hr is None:
        hr = _to_prob_0_1(getv("Hit Rate (5g)"))
    ml = _to_prob_0_1(getv("ml_prob"))
    if ml is None:
        ml = _to_prob_0_1(getv("ML Prob"))
    l5_over = pd.to_numeric(getv("l5_over", getv("L5 Over", 0)), errors="coerce")
    l5_under = pd.to_numeric(getv("l5_under", getv("L5 Under", 0)), errors="coerce")
    l5_sample = max(
        0.0,
        float(0.0 if pd.isna(l5_over) else l5_over),
        float(0.0 if pd.isna(l5_under) else l5_under),
    )
    tier = str(getv("tier", getv("Tier", "D")) or "D").strip().upper()

    if tier not in ("A", "B", "C"):
        return False
    has_sample = l5_sample >= 3.0
    hr_val = float(hr) if hr is not None else 0.0
    ml_val = float(ml) if ml is not None else 0.0
    hr_strong = hr_val >= 0.70 and has_sample
    ml_strong = ml_val >= 0.72
    if not hr_strong and not ml_strong:
        return False
    if hr_val < 0.55 and ml_val < 0.65:
        return False
    return True


def format_hit_window_fraction(n_games: int, raw) -> str:
    """
    Format L5/L10 over/under for UI pills.

    step4/8 store integer hit *counts* (0..n). Older formatters treated any
    value <= 1 as a rate, so l5_under=1 displayed as 5/5 instead of 1/5.
    """
    try:
        x = float(raw)
    except (TypeError, ValueError):
        return str(raw)
    if not math.isfinite(x):
        return str(raw)
    xi = int(round(x))
    if abs(x - xi) < 1e-6 and 0 <= xi <= int(n_games):
        k = xi
    elif 0.0 < x <= 1.0:
        k = int(round(x * n_games))
    elif float(n_games) < x <= 100.0:
        k = int(round((x / 100.0) * n_games))
    else:
        k = int(round(x))
    k = max(0, min(int(n_games), k))
    return f"{k}/{n_games}"


def _resolve_l5_cols(row: pd.Series, direction: str) -> tuple[float, float]:
    """
    Return (l5_hits, l5_games_played) for the play direction.

    Column aliases — step8 clean xlsx uses last5_over / last5_under;
    raw step8 CSV uses l5_over / l5_under. Both are hit counts (0–5),
    NOT games played. Games played comes from l5_games_played,
    n_legs_sample, or line_games_played_5 (whichever is present).

    Returns (0.0, 0.0) when nothing is resolvable.
    """
    if direction == "UNDER":
        hits_keys = ("l5_under", "last5_under")
    else:
        hits_keys = ("l5_over", "last5_over")

    hits_raw = None
    for k in hits_keys:
        v = row.get(k)
        if v is not None and str(v).strip() not in ("", "nan", "None"):
            hits_raw = v
            break

    gp_raw = (
        row.get("l5_games_played")
        or row.get("n_legs_sample")
        or row.get("line_games_played_5")
    )
    gp_num = pd.to_numeric(gp_raw, errors="coerce")
    has_gp = pd.notna(gp_num) and float(gp_num) > 0
    has_hits = hits_raw is not None

    if not has_hits and not has_gp:
        return 0.0, 0.0

    hits = pd.to_numeric(hits_raw, errors="coerce") if has_hits else np.nan
    hits = 0.0 if pd.isna(hits) else float(hits)
    gp = float(gp_num) if has_gp else 5.0
    return hits, gp


def _resolve_leg_prob(row: pd.Series) -> tuple[float, str]:
    """
    Selection / est_win_prob leg probability.
    Prefer pipeline_read enrichment (hit_prob_actionable); else empirical/ML/rank/edge chain.
    """
    for key, src in (
        ("hit_prob_actionable", "hit_prob_actionable"),
        ("hit_prob_selected", "hit_prob_selected"),
    ):
        p = _to_prob_0_1(row.get(key))
        if p is not None:
            return _clip_prob(p, src), src

    direction = str(
        row.get("bet_direction") or row.get("direction_used") or row.get("direction") or "OVER"
    ).strip().upper()

    def _directional_hr_raw() -> tuple[Any, str]:
        """Resolve direction-aware hit rate with safe fallbacks."""
        if direction == "UNDER":
            u = row.get("under_hit_rate")
            if u is not None and str(u).strip() != "":
                return u, "under_hit_rate"
            o = row.get("over_hit_rate")
            if o is not None and str(o).strip() != "":
                try:
                    ov = float(o)
                    if ov > 1.0:
                        ov = ov / 100.0
                    if math.isfinite(ov):
                        return 1.0 - ov, "over_hit_rate_inverted"
                except (TypeError, ValueError):
                    pass
            _l5u_hits, _l5u_gp = _resolve_l5_cols(row, "UNDER")
            if _l5u_gp > 0:
                return _l5u_hits / _l5u_gp, "l5_under_proxy"
            return row.get("hit_rate"), "hit_rate_fallback"
        o = row.get("over_hit_rate")
        if o is not None and str(o).strip() != "":
            return o, "over_hit_rate"
        _l5o_hits, _l5o_gp = _resolve_l5_cols(row, "OVER")
        if _l5o_gp > 0:
            return _l5o_hits / _l5o_gp, "l5_over_proxy"
        return row.get("hit_rate"), "hit_rate_fallback"

    hr_raw, hr_source = _directional_hr_raw()

    try:
        if hr_raw is not None and isinstance(hr_raw, float) and (math.isnan(hr_raw) or not math.isfinite(hr_raw)):
            hr_raw = None
    except TypeError:
        pass
    if hr_raw is not None and pd.isna(hr_raw):
        hr_raw = None

    l5_hits, l5_gp = _resolve_l5_cols(row, direction)
    l5_n = l5_gp
    pick_type = str(row.get("pick_type", "") or "").strip().lower()
    sport = str(row.get("sport", "") or "").strip().upper()

    if "demon" in pick_type and sport in ("NHL", "SOCCER", "SOC"):
        hr = _to_prob_0_1(hr_raw)
        ml = _to_prob_0_1(row.get("ml_prob"))
        has_sample = l5_gp >= 3.0
        hr_val = float(hr) if hr is not None else 0.0
        ml_val = float(ml) if ml is not None else 0.0
        hr_prob = min(0.75, hr_val)
        ml_prob = min(0.75, ml_val)
        if has_sample and hr_val >= 0.60:
            demon_prob = min(hr_prob, max(ml_prob, 0.50))
            return _clip_prob(demon_prob, "hit_rate_demon"), "hit_rate_demon"
        demon_prob = min(ml_prob, 0.72)
        return _clip_prob(demon_prob, "ml_prob_demon"), "ml_prob_demon"

    if hr_raw is not None and str(hr_raw).strip() != "" and l5_n >= 3.0:
        try:
            hit_prob = float(hr_raw)
            if hit_prob > 1.0:
                hit_prob = hit_prob / 100.0
            hit_prob = max(0.50, min(_leg_l5_hit_prob_cap(row), hit_prob))
            return hit_prob, hr_source
        except (TypeError, ValueError):
            pass

    # Priority 1: calibrated ML probability (cap via LEG_PROB_CAPS["ml_prob"]).
    mlp = pd.to_numeric(row.get("ml_prob"), errors="coerce")
    if pd.notna(mlp) and 0.0 < float(mlp) < 1.0:
        return _clip_prob(float(mlp), "ml_prob"), "ml_prob"

    # Priority 2: rank_score sigmoid (composite signal).
    rs = pd.to_numeric(row.get("rank_score"), errors="coerce")
    if pd.notna(rs):
        return _clip_prob(_rank_score_to_prob(float(rs)), "rank_score"), "rank_score"

    # Priority 3: edge-to-probability (magnitude — signed raw edge punishes UNDERs).
    ae = pd.to_numeric(row.get("abs_edge"), errors="coerce")
    edge_raw = pd.to_numeric(row.get("edge"), errors="coerce")
    edge_mag = ae if pd.notna(ae) else (abs(float(edge_raw)) if pd.notna(edge_raw) else float("nan"))
    thresh = get_edge_threshold(
        row.get("sport", ""), row.get("prop_type", ""), row.get("pick_type", "")
    )
    if pd.notna(edge_mag):
        shifted = float(edge_mag) - float(thresh)
        prob = 1.0 / (1.0 + math.exp(-shifted * 0.6))
        return _clip_prob(prob, "edge"), "edge"

    # Priority 4: shrunk hit rate.
    hr = pd.to_numeric(row.get("hit_rate"), errors="coerce")
    if pd.notna(hr):
        hr_val = float(hr)
        if 1.0 < hr_val <= 100.0:
            hr_val = hr_val / 100.0
        if 0.0 < hr_val < 1.0:
            n = pd.to_numeric(row.get("l5_games", row.get("sample_n", 5)), errors="coerce")
            if pd.isna(n) or float(n) <= 0:
                n = 5.0
            hit_rate_shrunk = (hr_val * float(n) + 0.55 * 5.0) / (float(n) + 5.0)
            return _clip_prob(float(hit_rate_shrunk), "hit_rate"), "hit_rate"

    return DEFAULT_LEG_PROB_FALLBACK, "fallback_const"


def win_prob(leg_probs_with_source, _n_legs: int) -> float:
    vals = []
    for p, src in leg_probs_with_source:
        try:
            if p is None:
                continue
            if isinstance(p, float) and np.isnan(p):
                continue
            vals.append(_clip_prob(float(p), src))
        except Exception:
            continue
    if not vals:
        return TICKET_PROB_FLOOR
    return float(np.clip(np.prod(vals), TICKET_PROB_FLOOR, TICKET_PROB_CAP))


_WIN_RATE_PRIMARY_SPORTS = frozenset({"NBA", "WNBA", "NBA1Q"})
_WIN_RATE_EXTRA_SPORTS = frozenset({"MLB", "NHL", "TENNIS"})
# Cap per-leg factor in p_win product — l5_over_proxy can return 0.99 on hot streaks (artifact).
MAX_LEG_PROB_FOR_P_WIN = 0.72
# Deep bench legs: L5 "5/5 over" on 0.5 Goblin lines is not a trustworthy parlay factor.
MAX_LEG_PROB_BENCH_SUPPORT = 0.62
MAX_L5_HIT_PROB_BENCH = 0.68
MAX_L5_HIT_PROB_DEFAULT = 0.85


def _winrate_leg_bench_risk(leg: dict) -> bool:
    """True for low-minute support bench (e.g. 0-pt DNP-risk legs in playoff rotations)."""
    su = str(leg.get("sport") or "").strip().upper()
    if su not in ("NBA", "WNBA", "NBA1H"):
        return False
    mt = str(leg.get("min_tier") or leg.get("minutes_tier") or "").strip().upper()
    ur = str(leg.get("usage_role") or "").strip().upper()
    sr = str(leg.get("shot_role") or "").strip().upper()
    return mt == "LOW" and ur == "SUPPORT" and sr in ("LOW_VOL", "", "LOW")


def _leg_l5_hit_prob_cap(row: pd.Series | dict) -> float:
    if isinstance(row, pd.Series):
        row = row.to_dict()
    if _winrate_leg_bench_risk(row):
        return MAX_L5_HIT_PROB_BENCH
    return MAX_L5_HIT_PROB_DEFAULT


def _leg_prob_for_p_win_from_mapping(leg: dict | pd.Series) -> float:
    """P(win) leg factor: leg_prob_used → composite_hit_rate/hit_rate → ml_prob → 0.55 (capped)."""
    if isinstance(leg, pd.Series):
        leg = leg.to_dict()
    cap = MAX_LEG_PROB_BENCH_SUPPORT if _winrate_leg_bench_risk(leg) else MAX_LEG_PROB_FOR_P_WIN
    for key in ("leg_prob_used", "composite_hit_rate", "hit_rate", "ml_prob"):
        raw = leg.get(key)
        if raw is None or raw == "":
            continue
        try:
            v = float(raw)
            if math.isfinite(v):
                return float(np.clip(v, 0.0, min(1.0, cap)))
        except (TypeError, ValueError):
            continue
    return min(0.55, cap)


def _compute_p_win_from_rows(rows: list) -> float:
    p = 1.0
    for row in rows or []:
        p *= _leg_prob_for_p_win_from_mapping(row)
    return float(np.clip(p, 0.0, 1.0))


def _enrich_slip_p_win_fields(slip: dict, *, mode: str = "ev") -> None:
    legs = slip.get("legs") or []
    if legs:
        p_win = _compute_p_win_from_rows(legs)
    else:
        rows = slip.get("rows") or []
        p_win = _compute_p_win_from_rows(rows)
    slip["p_win"] = round(p_win, 6)
    slip["expected_wins_per_100"] = round(p_win * 100, 1)
    slip["mode"] = mode
    if mode == "win_rate":
        pay_mult = float(slip.get("payout_multiplier") or 1.0)
        slip["win_rate_score"] = round(p_win * math.log(1.0 + max(pay_mult, 0.0)), 6)


def _group_max_p_win(group: dict) -> float:
    best = 0.0
    for t in group.get("tickets") or []:
        if not isinstance(t, dict):
            continue
        try:
            pw = float(t.get("p_win") or 0.0)
            if math.isfinite(pw):
                best = max(best, pw)
        except (TypeError, ValueError):
            pass
    return best


_GRADED_ANALYSIS_JSON = os.path.join(REPO_ROOT, "data", "graded_analysis_latest.json")
_GRADED_ANALYSIS_AVOID_SLICES: frozenset[tuple[str, str, str, str]] = frozenset(
    {
        ("NBA1Q", "DEMON", "OVER", "D"),
        ("NBA1Q", "STANDARD", "UNDER", "D"),
    }
)


def _load_graded_analysis(path: str | None = None) -> dict | None:
    p = path or _GRADED_ANALYSIS_JSON
    if not os.path.isfile(p):
        return None
    try:
        with open(p, encoding="utf-8") as f:
            return json.load(f)
    except Exception:
        return None


def _graded_analysis_context(analysis: dict | None) -> dict[str, Any]:
    ctx: dict[str, Any] = {
        "slice_priority": {},
        "avoid_slices": set(_GRADED_ANALYSIS_AVOID_SLICES),
        "top_players": set(),
        "bottom_players": set(),
    }
    if not analysis:
        return ctx
    for s in analysis.get("top_slices") or []:
        if not isinstance(s, dict):
            continue
        key = (
            str(s.get("sport", "")).strip().upper(),
            str(s.get("pick_type", "")).strip().lower(),
            str(s.get("direction", "")).strip().upper(),
            str(s.get("tier", "")).strip().upper(),
        )
        if key[0] and key[1]:
            ctx["slice_priority"][key] = int(s.get("priority", 99))
    for s in analysis.get("avoid_slices") or []:
        if not isinstance(s, dict):
            continue
        try:
            hr = float(s.get("hit_rate", 1.0))
        except (TypeError, ValueError):
            hr = 1.0
        if hr >= 0.50:
            continue
        key = (
            str(s.get("sport", "")).strip().upper(),
            str(s.get("pick_type", "")).strip().lower(),
            str(s.get("direction", "")).strip().upper(),
            str(s.get("tier", "")).strip().upper(),
        )
        if key[0]:
            ctx["avoid_slices"].add(key)
    pr = analysis.get("player_rankings") or {}
    for p in pr.get("top_30") or []:
        if isinstance(p, dict) and p.get("player"):
            ctx["top_players"].add(
                (str(p.get("player", "")).strip().casefold(), str(p.get("sport", "")).strip().upper())
            )
    for p in pr.get("bottom_20") or []:
        if isinstance(p, dict) and p.get("player"):
            ctx["bottom_players"].add(
                (str(p.get("player", "")).strip().casefold(), str(p.get("sport", "")).strip().upper())
            )
    return ctx


def _row_slice_key(row_d: dict) -> tuple[str, str, str, str]:
    return (
        str(row_d.get("sport") or "").strip().upper(),
        str(row_d.get("pick_type") or "").strip().lower(),
        str(row_d.get("direction") or row_d.get("bet_direction") or "").strip().upper(),
        str(row_d.get("tier") or "").strip().upper(),
    )


def _row_hot_l10_streak(row_d: dict) -> bool:
    direction = str(row_d.get("direction") or row_d.get("bet_direction") or "OVER").strip().upper()
    if direction == "UNDER":
        pct = pd.to_numeric(row_d.get("l10_under_pct"), errors="coerce")
        raw = pd.to_numeric(row_d.get("l10_under"), errors="coerce")
    else:
        pct = pd.to_numeric(row_d.get("l10_over_pct"), errors="coerce")
        raw = pd.to_numeric(row_d.get("l10_over"), errors="coerce")
    if pd.notna(pct) and float(pct) >= 0.70:
        return True
    if pd.notna(raw) and float(raw) >= 7.0:
        return True
    return False


def _graded_analysis_row_boost(row_d: dict, ctx: dict[str, Any]) -> float:
    """Higher = prefer leg in win-rate pool / anchor picks."""
    boost = 0.0
    sk = _row_slice_key(row_d)
    pri = ctx.get("slice_priority", {}).get(sk)
    if pri is not None:
        boost += max(0.0, 0.15 - 0.01 * float(pri))
    player_key = (str(row_d.get("player", "")).strip().casefold(), sk[0])
    if player_key in ctx.get("top_players", set()):
        boost += 0.04
    if player_key in ctx.get("bottom_players", set()):
        boost -= 0.06
    if _row_hot_l10_streak(row_d):
        boost += 0.05
    try:
        ln = float(row_d.get("line") or row_d.get("line_score") or 0)
    except (TypeError, ValueError):
        ln = 0.0
    if sk[0] in ("NBA", "NBA1H") and sk[1] == "goblin" and ln >= 3.0:
        boost += 0.02
    return boost


def _row_in_avoid_slice(row_d: dict, ctx: dict[str, Any]) -> bool:
    sk = _row_slice_key(row_d)
    avoid = ctx.get("avoid_slices") or set()
    return sk in avoid


def _pick_win_rate_leg(
    df: pd.DataFrame | None,
    *,
    sport: str,
    pick_type: str,
    tier: str,
    direction: str = "OVER",
    require_hot: bool,
    min_leg_prob: float,
    min_composite_hr: float,
    graded_ctx: dict[str, Any],
    prefer_high_line: bool = False,
) -> dict | None:
    if df is None or df.empty:
        return None
    best: dict | None = None
    best_score = -1e9
    for _, r in df.iterrows():
        row_d = r.to_dict()
        if str(row_d.get("sport", "")).strip().upper() != sport.upper():
            continue
        if str(row_d.get("pick_type", "")).strip().lower() != pick_type.lower():
            continue
        if str(row_d.get("tier", "")).strip().upper() != tier.upper():
            continue
        if str(row_d.get("direction") or row_d.get("bet_direction") or "").strip().upper() != direction.upper():
            continue
        if require_hot and not _row_hot_l10_streak(row_d):
            continue
        if _row_in_avoid_slice(row_d, graded_ctx):
            continue
        if not _row_win_rate_eligible(
            row_d, min_leg_prob=min_leg_prob, min_composite_hr=min_composite_hr
        ):
            continue
        leg_p = _leg_prob_for_p_win_from_mapping(row_d)
        score = leg_p + _graded_analysis_row_boost(row_d, graded_ctx)
        if prefer_high_line:
            try:
                score += 0.002 * float(row_d.get("line") or 0)
            except (TypeError, ValueError):
                pass
        if score > best_score:
            best_score = score
            best = dict(row_d)
    return best


def build_win_rate_anchor_ticket(
    frames_by_sport: dict[str, pd.DataFrame],
    *,
    min_leg_prob: float,
    min_composite_hr: float,
    graded_ctx: dict[str, Any],
) -> dict | None:
    """
    Preferred 3-leg win-rate structure from clean graded analysis:
      NBA1Q Goblin OVER A (HOT) anchor + NBA Goblin A/B (HOT) for payout legs.
    """
    nba1q = frames_by_sport.get("NBA1Q")
    nba = frames_by_sport.get("NBA")
    leg1 = _pick_win_rate_leg(
        nba1q,
        sport="NBA1Q",
        pick_type="goblin",
        tier="A",
        direction="OVER",
        require_hot=True,
        min_leg_prob=min_leg_prob,
        min_composite_hr=min_composite_hr,
        graded_ctx=graded_ctx,
    )
    leg2 = _pick_win_rate_leg(
        nba,
        sport="NBA",
        pick_type="goblin",
        tier="A",
        direction="OVER",
        require_hot=True,
        min_leg_prob=min_leg_prob,
        min_composite_hr=min_composite_hr,
        graded_ctx=graded_ctx,
        prefer_high_line=True,
    )
    leg3 = _pick_win_rate_leg(
        nba,
        sport="NBA",
        pick_type="goblin",
        tier="B",
        direction="OVER",
        require_hot=True,
        min_leg_prob=min_leg_prob,
        min_composite_hr=min_composite_hr,
        graded_ctx=graded_ctx,
        prefer_high_line=True,
    )
    if not leg1 or not leg2 or not leg3:
        return None
    rows = [leg1, leg2, leg3]
    if len({_ticket_row_dedup_key([r]) for r in rows}) < 3:
        return None
    fin = _finalize_structure_ticket_dict(
        rows, "power", "Win-Rate Anchor", "power", 3, None, False
    )
    if fin is None:
        return None
    fin["mode"] = "win_rate"
    fin["anchor_template"] = "NBA1Q_GobA_HOT + NBA_GobA_HOT + NBA_GobB_HOT"
    fin["_sport_label"] = "Win-Rate Anchor"
    return fin


def _win_rate_sport_allowed(sport_key: str, leg_prob: float) -> bool:
    su = str(sport_key or "").strip().upper()
    gated, reason = _sport_ticket_gated(su)
    if gated:
        return False
    if su in _WIN_RATE_PRIMARY_SPORTS:
        return True
    if su in _WIN_RATE_EXTRA_SPORTS:
        return float(leg_prob) >= 0.60
    return False


def _row_win_rate_eligible(
    row: pd.Series | dict,
    *,
    min_leg_prob: float,
    min_composite_hr: float,
    graded_ctx: dict[str, Any] | None = None,
) -> bool:
    if isinstance(row, pd.Series):
        row_d = row.to_dict()
    else:
        row_d = dict(row)
    if graded_ctx and _row_in_avoid_slice(row_d, graded_ctx):
        return False
    pt = str(row_d.get("pick_type") or "").strip().lower()
    tier = str(row_d.get("tier") or "").strip().upper()
    if pt == "goblin":
        pass
    elif pt == "standard" and tier == "A":
        pass
    else:
        return False
    leg_prob = _leg_prob_for_p_win_from_mapping(row_d)
    if leg_prob < float(min_leg_prob):
        return False
    comp = row_d.get("composite_hit_rate")
    if comp is None or comp == "":
        comp = row_d.get("hit_rate")
    try:
        comp_f = float(comp) if comp is not None and comp != "" else 0.0
    except (TypeError, ValueError):
        comp_f = 0.0
    if comp_f < float(min_composite_hr):
        return False
    sport = str(row_d.get("sport") or "").strip().upper()
    if not _win_rate_sport_allowed(sport, leg_prob):
        return False
    if _winrate_leg_bench_risk(row_d):
        return False
    return True


def _winrate_ticket_same_game_bench_stack(ticket: dict) -> bool:
    """Two+ legs from the same game where every leg is deep-bench SUPPORT (high DNP risk)."""
    legs = [leg for leg in (ticket.get("legs") or []) if isinstance(leg, dict)]
    if len(legs) < 2:
        return False
    by_game: dict[tuple[str, str, str], list[dict]] = {}
    for leg in legs:
        key = (
            str(leg.get("sport") or "").strip().upper(),
            str(leg.get("team") or "").strip().upper(),
            str(leg.get("opp") or "").strip().upper(),
        )
        by_game.setdefault(key, []).append(leg)
    for grp in by_game.values():
        if len(grp) >= 2 and all(_winrate_leg_bench_risk(leg) for leg in grp):
            return True
    return False


def _winrate_ticket_win_prob(ticket: dict) -> float:
    """Modeled probability ticket wins (power = all legs hit), not payout/cash EV."""
    for key in ("est_win_prob", "p_win", "combined_hit_prob_curve"):
        raw = ticket.get(key)
        if raw is None or raw == "":
            continue
        try:
            v = float(raw)
            if math.isfinite(v) and v > 0:
                return float(np.clip(v, 0.0, 1.0))
        except (TypeError, ValueError):
            continue
    return 0.0


def _winrate_ticket_rank_score(ticket: dict) -> float:
    """Panel/build sort: highest modeled win probability (not ticket_model_p_cash)."""
    v = _winrate_ticket_win_prob(ticket)
    if _winrate_ticket_same_game_bench_stack(ticket):
        v *= 0.85
    return v


def _winrate_ticket_panel_pcash_optional(ticket: dict) -> float | None:
    """P(cash) only shown when it differs from win prob (secondary line)."""
    pwin = _winrate_ticket_win_prob(ticket)
    raw = ticket.get("ticket_model_p_cash")
    if raw is None or raw == "":
        return None
    try:
        pcash = float(raw)
        if not math.isfinite(pcash) or pcash <= 0:
            return None
        pcash = float(np.clip(pcash, 0.0, 1.0))
        if abs(pcash - pwin) < 0.05:
            return None
        return pcash
    except (TypeError, ValueError):
        return None


def _filter_win_rate_pool(
    df: pd.DataFrame | None,
    *,
    min_leg_prob: float,
    min_composite_hr: float,
    graded_ctx: dict[str, Any] | None = None,
) -> pd.DataFrame:
    if df is None or df.empty:
        return pd.DataFrame()
    out_rows: list[dict] = []
    for _, r in df.iterrows():
        if _row_win_rate_eligible(
            r,
            min_leg_prob=min_leg_prob,
            min_composite_hr=min_composite_hr,
            graded_ctx=graded_ctx,
        ):
            out_rows.append(r.to_dict())
    if not out_rows:
        return pd.DataFrame()
    return pd.DataFrame(out_rows)


def build_win_rate_ticket_groups(
    sport_frames: list[tuple[str, pd.DataFrame]],
    *,
    min_leg_prob: float,
    min_composite_hr: float,
    max_legs: int,
    max_tickets: int,
    graded_analysis: dict | None = None,
) -> list[tuple[str, list, None]]:
    """Build up to max_tickets win-rate slips (2–max_legs legs), sorted by p_win."""
    max_legs = max(2, min(3, int(max_legs)))
    graded_ctx = _graded_analysis_context(graded_analysis)
    frames_by_sport: dict[str, pd.DataFrame] = {}
    for label, raw_df in sport_frames:
        if raw_df is not None and not raw_df.empty:
            su = str(label).strip().upper()
            frames_by_sport[su] = raw_df
            if "sport" in raw_df.columns:
                for sp, g in raw_df.groupby(raw_df["sport"].astype(str).str.upper()):
                    frames_by_sport[str(sp).strip().upper()] = g

    candidates: list[dict] = []
    anchor = build_win_rate_anchor_ticket(
        frames_by_sport,
        min_leg_prob=min_leg_prob,
        min_composite_hr=min_composite_hr,
        graded_ctx=graded_ctx,
    )
    if anchor is not None:
        p_win = _compute_p_win_from_rows(anchor.get("rows") or [])
        anchor["p_win"] = p_win
        anchor["win_rate_score"] = p_win * math.log(1.0 + max(float(anchor.get("payout_multiplier") or 1.0), 0.0))
        candidates.append(anchor)

    for label, raw_df in sport_frames:
        wr_df = _filter_win_rate_pool(
            raw_df,
            min_leg_prob=min_leg_prob,
            min_composite_hr=min_composite_hr,
            graded_ctx=graded_ctx,
        )
        if wr_df is None or len(wr_df) < 2:
            continue
        for n in range(2, max_legs + 1):
            if len(wr_df) < n:
                continue
            built = build_tickets(
                wr_df,
                n,
                max_tickets=max(5, int(max_tickets) * 2),
                ticket_sort_mode="rank",
                player_ticket_counts=defaultdict(int),
            )
            for t in built:
                rows = list(t.get("rows") or [])
                p_win = _compute_p_win_from_rows(rows)
                pay_mult = float(t.get("payout_multiplier") or 1.0)
                t = dict(t)
                t["p_win"] = p_win
                t["win_rate_score"] = p_win * math.log(1.0 + max(pay_mult, 0.0))
                t["mode"] = "win_rate"
                t["_sport_label"] = label
                candidates.append(t)

    candidates.sort(
        key=lambda x: (
            -_winrate_ticket_win_prob(x),
            -_winrate_ticket_rank_score(x),
            -float(x.get("win_rate_score") or 0.0),
        )
    )
    seen: set[frozenset] = set()
    picked: list[dict] = []
    for t in candidates:
        if _winrate_ticket_same_game_bench_stack(t):
            continue
        rows = [dict(r) for r in (t.get("rows") or [])]
        key = _ticket_row_dedup_key(rows)
        if key in seen:
            continue
        seen.add(key)
        picked.append(t)
        if len(picked) >= int(max_tickets):
            break

    groups: list[tuple[str, list, None]] = []
    for t in picked:
        rows = list(t.get("rows") or [])
        n = len(rows)
        sport_label = str(t.get("_sport_label") or "Mixed")
        pts = {str(r.get("pick_type") or "").strip().lower() for r in rows}
        if pts == {"goblin"}:
            pool_tag = "Goblin"
        elif pts == {"standard"}:
            pool_tag = "Standard"
        else:
            pool_tag = "Mixed"
        gn = f"{sport_label} {n}-Leg {pool_tag}"
        groups.append((gn, [t], None))
    return groups


def flex_cash_prob(leg_probs_with_source: list) -> float:
    """
    Modeled P(flex pays) for independent legs: all hit OR exactly one miss (PrizePicks-style flex, n≥3).
    For n<3, same as power (product of leg probs).
    """
    vals: list[float] = []
    for p, src in leg_probs_with_source:
        try:
            if p is None or (isinstance(p, float) and np.isnan(p)):
                continue
            vals.append(_clip_prob(float(p), str(src)))
        except Exception:
            continue
    n = len(vals)
    if n < 3:
        return float(np.clip(np.prod(vals), TICKET_PROB_FLOOR, TICKET_PROB_CAP)) if vals else TICKET_PROB_FLOOR
    prod_all = float(np.prod(vals))
    one_miss = 0.0
    for i in range(n):
        miss = 1.0 - vals[i]
        rest = 1.0
        for j in range(n):
            if j != i:
                rest *= vals[j]
        one_miss += miss * rest
    return float(np.clip(prod_all + one_miss, TICKET_PROB_FLOOR, TICKET_PROB_CAP))


def _best_flex3_rows_from_long_ticket(rows: list[dict]) -> list[dict] | None:
    """
    If a 4+ leg slip was built but greedy Flex 3 failed, pick the 3-leg subset with
    highest modeled flex cash (independent-leg approximation used in finalize).
    """
    if not rows or len(rows) < 3:
        return None
    if len(rows) == 3:
        return [dict(r) for r in rows]
    best: list[dict] | None = None
    best_sc = -1.0
    for combo in itertools.combinations(range(len(rows)), 3):
        r3 = [rows[i] for i in combo]
        leg_probs = [_resolve_leg_prob(pd.Series(r)) for r in r3]
        sc = flex_cash_prob(leg_probs)
        if sc > best_sc:
            best_sc = sc
            best = [dict(r) for r in r3]
    return best


_GUARANTEE_FLEX3_SPORT_KEYS: frozenset[str] = frozenset({"SOCCER", "TENNIS", "MLB", "NHL"})


def _normalize_sport_key_for_flex3_guarantee(raw: object) -> str:
    u = str(raw or "").strip().upper()
    if u in ("SOC", "SOCCER"):
        return "SOCCER"
    return u


def _ticket_normalized_single_sport(ticket: dict) -> str | None:
    rows = list(ticket.get("rows") or [])
    sports = {
        _normalize_sport_key_for_flex3_guarantee(r.get("sport"))
        for r in rows
        if isinstance(r, dict) and str(r.get("sport") or "").strip()
    }
    if len(sports) == 1:
        return next(iter(sports))
    if not sports:
        ts = _normalize_sport_key_for_flex3_guarantee(ticket.get("sport"))
        if ts in _GUARANTEE_FLEX3_SPORT_KEYS:
            return ts
    return None


def _ticket_is_cross_sport(ticket: dict) -> bool:
    rows = list(ticket.get("rows") or [])
    sports = {
        _normalize_sport_key_for_flex3_guarantee(r.get("sport"))
        for r in rows
        if isinstance(r, dict) and str(r.get("sport") or "").strip()
    }
    return len(sports) >= 2


def _builder_ticket_n_legs(ticket: dict) -> int:
    try:
        n = int(ticket.get("n_legs") or 0)
        if n > 0:
            return n
    except (TypeError, ValueError):
        pass
    return len(list(ticket.get("rows") or []))


def _top3_rows_by_hit_rate(rows: list[dict]) -> list[dict] | None:
    """Pick three legs with highest numeric hit_rate (for guaranteed Flex 3 from 4+)."""
    if len(rows) < 3:
        return None
    scored: list[tuple[float, dict]] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        hr = float(pd.to_numeric(r.get("hit_rate"), errors="coerce") or 0.0)
        scored.append((hr, dict(r)))
    if len(scored) < 3:
        return None
    scored.sort(key=lambda x: (-x[0], str(x[1].get("player", ""))))
    return [dict(x[1]) for x in scored[:3]]


def _flex3_guarantee_prefix_and_hdr(skey: str) -> tuple[str, str]:
    if skey == "SOCCER":
        return "Soccer", C["hdr_soccer"]
    if skey == "TENNIS":
        return "Tennis", C["hdr_tennis"]
    if skey == "MLB":
        return "MLB", C["hdr_mlb"]
    if skey == "NHL":
        return "NHL", C["hdr_nhl"]
    return "MIX", C["hdr_mix"]


def apply_guaranteed_flex3_backfill_from_four_plus(
    all_ticket_groups: list,
    counters: dict,
    wb: Any,
) -> None:
    """
    After all ticket groups are built: for Soccer/Tennis/MLB/NHL, if any slip has n>=4 and none
    has n==3, add Flex 3 from the best 4+ slip's top 3 legs by hit_rate. Same for cross-sport slips.
    """
    for skey in sorted(_GUARANTEE_FLEX3_SPORT_KEYS):
        has3 = False
        cands: list[tuple[int, float, dict, str]] = []
        for gname, tickets, *_ in all_ticket_groups:
            for t in tickets or []:
                if not isinstance(t, dict):
                    continue
                if _ticket_normalized_single_sport(t) != skey:
                    continue
                n = _builder_ticket_n_legs(t)
                if n == 3:
                    has3 = True
                if n >= 4:
                    avg_hr = float(t.get("avg_hit_rate") or 0.0)
                    cands.append((n, avg_hr, t, str(gname)))
        if has3 or not cands:
            continue
        cands.sort(key=lambda x: (-x[0], -x[1]))
        best_t = cands[0][2]
        rows = list(best_t.get("rows") or [])
        rows3 = _top3_rows_by_hit_rate(rows)
        if not rows3:
            continue
        prefix, bg_hdr = _flex3_guarantee_prefix_and_hdr(skey)
        fin = _finalize_structure_ticket_dict(
            rows3, "flex", prefix, "flex", 3, counters, False
        )
        if fin is None:
            print(f"  [flex3-guarantee] WARN: could not finalize {prefix} Flex 3 from 4+ slip")
            continue
        fin.setdefault("ticket_type", "Flex")
        display = f"{prefix} Flex 3-Leg · from 4+"
        sname = _excel_ticket_sheet_title_unique(display, wb.sheetnames)
        write_ticket_sheet(wb, [fin], sname, bg_hdr, label=f"{prefix} Flex")
        all_ticket_groups.append((display, [fin], None))
        print(
            f"  [flex3-guarantee] {prefix}: Flex 3 from top HR legs of {cands[0][0]}-leg slip "
            f"(no natural 3-leg found)"
        )

    has3x = False
    cands_x: list[tuple[int, float, dict, str]] = []
    for gname, tickets, *_ in all_ticket_groups:
        for t in tickets or []:
            if not isinstance(t, dict):
                continue
            if not _ticket_is_cross_sport(t):
                continue
            n = _builder_ticket_n_legs(t)
            if n == 3:
                has3x = True
            if n >= 4:
                avg_hr = float(t.get("avg_hit_rate") or 0.0)
                cands_x.append((n, avg_hr, t, str(gname)))
    if has3x or not cands_x:
        return
    cands_x.sort(key=lambda x: (-x[0], -x[1]))
    rows = list(cands_x[0][2].get("rows") or [])
    rows3 = _top3_rows_by_hit_rate(rows)
    if not rows3:
        return
    fin = _finalize_structure_ticket_dict(rows3, "flex", "MIX", "flex", 3, counters, False)
    if fin is None:
        print("  [flex3-guarantee] WARN: could not finalize cross-sport Flex 3 from 4+ slip")
        return
    fin.setdefault("ticket_type", "Flex")
    display = "Cross-sport Flex 3-Leg · from 4+"
    sname = _excel_ticket_sheet_title_unique(display, wb.sheetnames)
    write_ticket_sheet(wb, [fin], sname, C["hdr_mix"], label="Cross-sport Flex")
    all_ticket_groups.append((display, [fin], None))
    print(
        f"  [flex3-guarantee] cross-sport: Flex 3 from top HR legs of {cands_x[0][0]}-leg slip "
        "(no natural 3-leg found)"
    )


def _greedy_ticket_with_first_leg(cand: pd.DataFrame, n_legs: int, first_idx: int) -> list[pd.Series] | None:
    """Greedy unique-player fill in cand row order; first leg locked to cand.iloc[first_idx]."""
    cand = cand.reset_index(drop=True)
    if first_idx < 0 or first_idx >= len(cand):
        return None
    first = cand.iloc[first_idx]
    p0 = str(first.get("player", "") or "").strip().lower()
    if not p0:
        return None
    chosen: list[pd.Series] = [first]
    used = {p0}
    prop_counts: Counter[str] = Counter()
    prop_counts[_ticket_prop_token(first)] += 1
    for i in range(len(cand)):
        if len(chosen) >= n_legs:
            break
        if i == first_idx:
            continue
        r = cand.iloc[i]
        p = str(r.get("player", "") or "").strip().lower()
        if not p or p in used:
            continue
        if not _can_add_row_with_prop_cap(r, prop_counts):
            continue
        chosen.append(r)
        used.add(p)
        prop_counts[_ticket_prop_token(r)] += 1
    if len(chosen) < n_legs:
        return None
    return chosen


def _modeled_ticket_paid_score(rows: list[dict], flow: str, n_legs: int) -> tuple[float, float, float]:
    """(objective_score, ep_with_penalty, flex_cash_with_penalty). Objective = flex cash for flex n≥3 else power."""
    leg_probs = [_resolve_leg_prob(pd.Series(r)) for r in rows]
    cmult, _ = _correlation_multiplier_and_audit(rows)
    ep = win_prob(leg_probs, n_legs) * cmult
    flex_cash = flex_cash_prob(leg_probs) * cmult if n_legs >= 3 else ep
    if flow == "flex" and n_legs >= 3:
        score = flex_cash
    else:
        score = ep
    return score, ep, flex_cash


def _greedy_ticket_ev_rank_metric(rows: list[dict], flow: str, n_legs: int) -> float:
    """
    Payout-adjusted EV proxy for ranking greedy candidates: flex n≥3 uses flex_cash×adj_flex;
    otherwise est win prob × adj power payout (matches ticket ev_power spirit).
    """
    leg_probs = [_resolve_leg_prob(pd.Series(r)) for r in rows]
    cmult, _ = _correlation_multiplier_and_audit(rows)
    ep = win_prob(leg_probs, n_legs) * cmult
    flex_cash = flex_cash_prob(leg_probs) * cmult if n_legs >= 3 else ep
    po = PAYOUT.get(int(n_legs), {"power": 0.0, "flex": 0.0})
    adj_p = calc_adjusted_payout(float(po["power"]), rows)
    adj_f = calc_adjusted_payout(float(po["flex"]), rows)
    if str(flow) == "flex" and int(n_legs) >= 3:
        return float(flex_cash * adj_f)
    return float(ep * adj_p)


STRUCTURE_VARIANT_MAX_JACCARD = 0.55


def _ticket_row_dedup_key(rows: list[dict]) -> frozenset[str]:
    """
    Canonical leg-set key for slip dedupe.
    Include sport/player/prop/line/direction so near-identical variants collapse.
    """
    out: set[str] = set()
    for r in rows or []:
        if not isinstance(r, dict):
            continue
        line_s = _norm_line_for_leg_fp(r.get("line"))
        d = str(r.get("direction", "") or "").strip().upper()
        if "UNDER" in d:
            d = "UNDER"
        elif "OVER" in d:
            d = "OVER"
        out.add(
            "|".join(
                [
                    str(r.get("sport", "") or "").strip().upper(),
                    str(r.get("player", "") or "").strip().lower(),
                    str(r.get("prop_type", "") or "").strip().lower(),
                    line_s,
                    d,
                ]
            )
        )
    return frozenset(x for x in out if x.strip())


def _jaccard_overlap(a: frozenset[str], b: frozenset[str]) -> float:
    if not a and not b:
        return 1.0
    if not a or not b:
        return 0.0
    inter = len(a & b)
    union = len(a | b)
    return float(inter) / float(union) if union else 0.0


def _pctl(vals: list[float], q: float) -> float:
    if not vals:
        return 0.0
    arr = sorted(float(v) for v in vals)
    qq = max(0.0, min(1.0, float(q)))
    idx = int(round((len(arr) - 1) * qq))
    idx = max(0, min(len(arr) - 1, idx))
    return float(arr[idx])


def compute_ticket_diversity_audit(
    all_ticket_groups: list[tuple[str, list[dict[str, Any]], Any]],
) -> dict[str, Any]:
    slips: list[tuple[int, frozenset[str]]] = []
    for _group_name, tickets, _bg in all_ticket_groups or []:
        for t in (tickets or []):
            if not isinstance(t, dict):
                continue
            rows = t.get("rows") or t.get("legs") or []
            key = _ticket_row_dedup_key(rows)
            if not key:
                continue
            n_legs = int(t.get("n_legs") or len(rows) or 0)
            slips.append((n_legs, key))

    keys_all = [k for _, k in slips]
    total = len(keys_all)
    unique = len(set(keys_all))
    dup_rate = (1.0 - (float(unique) / float(total))) if total > 0 else 0.0

    pair_j: list[float] = []
    for i in range(len(keys_all)):
        a = keys_all[i]
        for j in range(i + 1, len(keys_all)):
            pair_j.append(_jaccard_overlap(a, keys_all[j]))

    by_legs: dict[str, Any] = {}
    for n in sorted({n for n, _ in slips if n > 0}):
        nk = [k for n0, k in slips if n0 == n]
        n_total = len(nk)
        n_unique = len(set(nk))
        n_dup = (1.0 - (float(n_unique) / float(n_total))) if n_total > 0 else 0.0
        n_pair_j: list[float] = []
        for i in range(len(nk)):
            a = nk[i]
            for j in range(i + 1, len(nk)):
                n_pair_j.append(_jaccard_overlap(a, nk[j]))
        by_legs[str(n)] = {
            "tickets": int(n_total),
            "unique_leg_sets": int(n_unique),
            "duplicate_rate": float(n_dup),
            "jaccard_mean": float(sum(n_pair_j) / len(n_pair_j)) if n_pair_j else 0.0,
            "jaccard_p90": float(_pctl(n_pair_j, 0.90)) if n_pair_j else 0.0,
            "jaccard_max": float(max(n_pair_j)) if n_pair_j else 0.0,
        }

    return {
        "tickets": int(total),
        "unique_leg_sets": int(unique),
        "duplicate_rate": float(dup_rate),
        "pair_count": int(len(pair_j)),
        "jaccard_mean": float(sum(pair_j) / len(pair_j)) if pair_j else 0.0,
        "jaccard_p90": float(_pctl(pair_j, 0.90)) if pair_j else 0.0,
        "jaccard_max": float(max(pair_j)) if pair_j else 0.0,
        "by_legs": by_legs,
    }


def _rank_greedy_tickets_by_paid_metric(
    cand: pd.DataFrame,
    n_legs: int,
    flow: str,
    ticket_gen_starts: int,
    max_tickets: int,
) -> list[tuple[list[dict], float, float, float]]:
    """
    Rank up to max_tickets distinct greedy slips from the first K first-leg seeds (sorted-candidate order).
    Sorts by payout-adjusted EV proxy first, then modeled hit/flex score (see _greedy_ticket_ev_rank_metric).
    """
    cand = cand.reset_index(drop=True)
    eligible_idx = [i for i in range(len(cand)) if str(cand.iloc[i].get("player", "") or "").strip()]
    if not eligible_idx:
        return []
    n_starts = max(1, min(int(ticket_gen_starts), len(eligible_idx)))
    ranked_work: list[tuple[list[dict], float, float, float, float]] = []
    seen: set[frozenset[str]] = set()
    for s in range(n_starts):
        first_idx = eligible_idx[s]
        chosen = _greedy_ticket_with_first_leg(cand, n_legs, first_idx)
        if not chosen:
            continue
        rows = [x.to_dict() for x in chosen]
        key = _ticket_row_dedup_key(rows)
        if key in seen:
            continue
        seen.add(key)
        score, ep, fc = _modeled_ticket_paid_score(rows, flow, n_legs)
        ev_m = _greedy_ticket_ev_rank_metric(rows, flow, n_legs)
        ranked_work.append((rows, score, ep, fc, ev_m))
    ranked_work.sort(key=lambda x: (-x[4], -x[1]))
    max_keep = max(1, int(max_tickets))
    return [(t[0], t[1], t[2], t[3]) for t in ranked_work[:max_keep]]


def _pick_best_greedy_ticket_by_paid_metric(
    cand: pd.DataFrame,
    n_legs: int,
    flow: str,
    ticket_gen_starts: int,
) -> tuple[list[dict] | None, float, float, float]:
    """
    Use the first K eligible rows (in sorted-candidate order) as alternative first legs; keep the
    combination with highest modeled ticket payout (P(all hit) for power-style, P(flex cash) for flex 3+).
    """
    ranked = _rank_greedy_tickets_by_paid_metric(cand, n_legs, flow, ticket_gen_starts, max_tickets=1)
    if not ranked:
        return None, 0.0, 0.0, 0.0
    rows, best_score, best_ep, best_flex = ranked[0]
    return rows, best_score, best_ep, best_flex


def _collect_row_candidates_for_structure(
    cand: pd.DataFrame,
    n_legs: int,
    flow: str,
    ticket_gen_starts: int,
    max_variants: int,
) -> list[list[dict]]:
    max_variants = max(1, int(max_variants))
    tg_starts = max(1, int(ticket_gen_starts))
    if max_variants == 1 and tg_starts <= 1:
        chosen: list[pd.Series] = []
        used_players: set[str] = set()
        prop_counts: Counter[str] = Counter()
        for _, r in cand.iterrows():
            p = str(r.get("player", "")).strip().lower()
            if not p or p in used_players:
                continue
            if not _can_add_row_with_prop_cap(r, prop_counts):
                continue
            chosen.append(r)
            used_players.add(p)
            prop_counts[_ticket_prop_token(r)] += 1
            if len(chosen) == n_legs:
                break
        if len(chosen) < n_legs:
            return []
        return [[x.to_dict() for x in chosen]]
    eff_starts = max(tg_starts, max_variants * 3) if max_variants > 1 else tg_starts
    # Pull a larger candidate set, then keep only diverse slips by leg-set overlap.
    ranked = _rank_greedy_tickets_by_paid_metric(
        cand,
        n_legs,
        flow,
        eff_starts,
        max_tickets=max(3, max_variants * 4),
    )
    selected: list[list[dict]] = []
    selected_keys: list[frozenset[str]] = []
    for t in ranked:
        rows = t[0]
        key = _ticket_row_dedup_key(rows)
        if not key:
            continue
        if any(_jaccard_overlap(key, k) > float(STRUCTURE_VARIANT_MAX_JACCARD) for k in selected_keys):
            continue
        selected.append(rows)
        selected_keys.append(key)
        if len(selected) >= max_variants:
            break
    return selected


def _finalize_structure_ticket_dict(
    rows: list[dict],
    structure: str,
    sport_label: str,
    flow: str,
    n_legs: int,
    counters: dict | None,
    prioritize_ticket_hit: bool,
) -> dict | None:
    # Many step8 rows omit sport; web template / EV grouping need per-leg sport for _ticket_primary_sport.
    sl = str(sport_label or "").strip()
    rows_use: list[dict] = []
    for r in rows:
        if not isinstance(r, dict):
            continue
        rd = dict(r)
        if sl and not str(rd.get("sport") or "").strip():
            rd["sport"] = sl
        rows_use.append(rd)
    if len(rows_use) < int(n_legs):
        return None
    rows = rows_use

    leg_probs = [_resolve_leg_prob(pd.Series(r)) for r in rows]
    cmult, caudit = _correlation_multiplier_and_audit(rows)
    ep = win_prob(leg_probs, n_legs) * cmult
    flex_cash = flex_cash_prob(leg_probs) * cmult if n_legs >= 3 else ep
    obj_score = flex_cash if flow == "flex" and n_legs >= 3 else ep

    hrs = [float(r.get("hit_rate", 0.5) or 0.5) for r in rows]
    rss = [float(r.get("rank_score", 0.0) or 0.0) for r in rows]

    if prioritize_ticket_hit:
        if flow == "flex" and n_legs >= 3:
            if flex_cash < float(MIN_PRIORITIZE_MODELED_FLEX_CASH_PROB):
                return None
        elif ep < float(MIN_PRIORITIZE_MODELED_POWER_WIN_PROB):
            return None

    payout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})
    pwr = payout["power"]
    flx = payout["flex"]
    adj_power = calc_adjusted_payout(pwr, rows)
    adj_flex = calc_adjusted_payout(flx, rows)

    tiers = [str(r.get("tier", "")).upper() for r in rows]
    defs = [_norm_prop_label(r.get("def_tier", r.get("opp_def_tier", ""))) for r in rows]
    if all(t in {"A", "B"} for t in tiers) and all(d in {"avg", "weak"} for d in defs):
        expected_win_rate = 0.78
    else:
        expected_win_rate = 0.68

    pct_out = counters.get("player_ticket_counts") if counters else None
    if pct_out is not None and not _ticket_cap_can_add(rows, pct_out):
        return None
    _ticket_cap_register(rows, pct_out)

    return {
        "key": _ticket_row_dedup_key(rows),
        "rows": rows,
        "avg_hit_rate": float(np.mean(hrs)) if hrs else 0.0,
        "avg_rank_score": float(np.mean(rss)) if rss else 0.0,
        "est_win_prob": ep,
        "ticket_objective_score": round(float(obj_score), 4),
        "est_flex_cash_prob": round(float(flex_cash), 4) if n_legs >= 3 else None,
        "power_payout": adj_power,
        "flex_payout": adj_flex,
        "base_power_payout": pwr,
        "payout_multiplier": round(adj_power / pwr, 4) if pwr else 1.0,
        "ev_power": round(ep * adj_power, 4),
        "kelly_units": round(kelly_fraction(ep, adj_power, fraction=0.25), 2),
        "n_legs": n_legs,
        "expected_win_rate": expected_win_rate,
        "correlation_multiplier": cmult,
        "correlation_audit": caudit,
        "ticket_type": (
            "Standard"
            if structure == "standard"
            else (
                "Flex"
                if structure == "flex"
                else (
                    "Power Standard 3"
                    if structure == "power_std3"
                    else "Goblin 3" if structure == "goblin3" else "Power Play"
                )
            )
        ),
        "sport": sport_label,
    }


def _same_game_density_multiplier(ticket_rows: list) -> float:
    """
    Same-game correlation discount (legacy).
    Any game with 2+ legs gets multiplied by 0.94 ** (n_same_game_legs - 1).
    """
    from collections import Counter

    keys = []
    for r in ticket_rows:
        team = str(r.get("team", "")).strip().upper()
        opp = str(r.get("opp", r.get("opp_team", ""))).strip().upper()
        if not team or not opp:
            continue
        keys.append("|".join(sorted([team, opp])))

    counts = Counter(keys)
    mult = 1.0
    for _, n in counts.items():
        if n >= 2:
            mult *= 0.94 ** (n - 1)
    return float(mult)


def _correlation_multiplier_and_audit(ticket_rows: list) -> tuple[float, list[str]]:
    """
    Ticket score multiplier + human-readable audit trail.
    - Same-game density (0.94^(n-1) per congested game), logged as same_game_density.
    - Same team, same game (2+ legs on one side): ×0.85 (−15%).
    - Same player correlated stack (e.g. PTS+AST or combo props): ×1.05 (+5%), at most once per ticket.
    """
    audit: list[str] = []
    mult = _same_game_density_multiplier(ticket_rows)
    audit.append(f"same_game_density×{mult:.4f}")

    from collections import defaultdict

    game_team_counts: dict[str, dict[str, int]] = {}
    for r in ticket_rows:
        team = str(r.get("team", "")).strip().upper()
        opp = str(r.get("opp", r.get("opp_team", ""))).strip().upper()
        if not team or not opp:
            continue
        gk = "|".join(sorted([team, opp]))
        if gk not in game_team_counts:
            game_team_counts[gk] = defaultdict(int)
        game_team_counts[gk][team] += 1
    if any(any(n >= 2 for n in tc.values()) for tc in game_team_counts.values()):
        mult *= 0.85
        audit.append("same_team_same_game:-15%")

    by_player: dict[str, set[str]] = {}
    for r in ticket_rows:
        pl = str(r.get("player", "")).strip().lower()
        if not pl:
            continue
        tok = _norm_prop_label(r.get("prop_type", ""))
        by_player.setdefault(pl, set()).add(tok)

    for pl, pset in by_player.items():
        if len(pset) < 2:
            continue
        has_pts = "points" in pset
        has_ast = "assists" in pset
        has_reb = "rebounds" in pset
        combo = bool(
            pset
            & {
                "pts+asts",
                "pts+rebs",
                "rebs+asts",
                "pts+rebs+asts",
            }
        )
        if combo or (has_pts and has_ast) or (has_pts and has_ast and has_reb):
            mult *= 1.05
            audit.append(f"player_correlated_stack:+5%:{pl[:28]}")
            break

    return float(mult), audit


def _correlation_penalty(ticket_rows: list) -> float:
    """Backward-compatible multiplier only (no audit). Used by combined_ticket_grader."""
    return _correlation_multiplier_and_audit(ticket_rows)[0]


def kelly_fraction(win_prob: float, payout_mult: float, fraction: float = 0.25) -> float:
    """
    Fractional Kelly sizing.

    payout_mult is the total return multiplier (e.g., 3.0 means win returns 3x stake).
    Returns recommended fraction of bankroll to wager (0..1).
    """
    try:
        p = float(win_prob)
        b = float(payout_mult) - 1.0
    except Exception:
        return 0.0
    if b <= 0.0 or p <= 0.0 or p >= 1.0:
        return 0.0
    k = (p * b - (1.0 - p)) / b
    k = max(0.0, float(k))
    return float(k * float(fraction))


# ──────────────────────────────────────────────────────────────────────────────
# Path resolution helpers (fixes the “file not found” headaches)
# ──────────────────────────────────────────────────────────────────────────────
def _norm_path(p: str) -> str:
    p = (p or "").strip().strip('"').strip("'")
    p = os.path.expanduser(p)
    return os.path.abspath(p)


def _find_most_recent_by_filename(root_dir: str, filename: str) -> Optional[str]:
    """
    Recursively search root_dir for files matching filename (case-insensitive).
    Returns the most recently modified match so stale archive copies never win.
    Skips archive and old_* directories entirely.
    """
    SKIP_DIRS = {"archive", "old_outputs", "old_csv", "old_runs", "old_scripts", "__pycache__"}
    matches = []
    try:
        for base, dirs, files in os.walk(root_dir):
            dirs[:] = [d for d in dirs if d.lower() not in SKIP_DIRS]
            for f in files:
                if f.lower() == filename.lower():
                    full = os.path.join(base, f)
                    try:
                        matches.append((os.path.getmtime(full), full))
                    except OSError:
                        pass
    except Exception:
        return None
    if not matches:
        return None
    matches.sort(key=lambda x: x[0], reverse=True)
    return matches[0][1]


def _is_plain_filename(raw: str) -> bool:
    """True if `raw` is a single filename with no directory or drive (safe for repo-wide search)."""
    r = os.path.expanduser(raw.strip().strip('"').strip("'"))
    return os.path.dirname(r) == ""


def resolve_input_path(path: str, fallback_filename: Optional[str] = None) -> str:
    """
    Tries:
    1) exact path as provided
    2) relative to repo root (parent of scripts/)
    3) relative to script directory
    4) recursive search — only if `path` is a plain filename (no folders); picks most recent
       under repo root then scripts/, skips archive/old_* dirs. Never substitutes a different
       basename when the user gave an explicit relative/absolute path that is missing.
    """
    if not path:
        raise FileNotFoundError("Empty input path.")

    raw = path.strip().strip('"').strip("'")
    p = _norm_path(raw)
    if os.path.exists(p):
        return p

    script_dir = os.path.dirname(os.path.abspath(__file__))
    repo_root = os.path.abspath(os.path.join(script_dir, ".."))

    p_repo = os.path.abspath(os.path.join(repo_root, raw))
    if os.path.exists(p_repo):
        return p_repo

    p2 = os.path.abspath(os.path.join(script_dir, raw))
    if os.path.exists(p2):
        return p2

    if not _is_plain_filename(raw):
        filename = os.path.basename(raw)
        raise FileNotFoundError(
            f"Could not find file: {path}\nTried:\n- {p}\n- {p_repo}\n- {p2}\n"
            f"(no recursive search — path includes a directory and file is missing)"
        )

    filename = fallback_filename or os.path.basename(raw)
    for root in (repo_root, script_dir):
        found = _find_most_recent_by_filename(root, filename)
        if found and os.path.exists(found):
            print(f"  [resolve] Fallback found (most recent under {root}): {found}")
            return os.path.abspath(found)

    raise FileNotFoundError(
        f"Could not find file: {path}\nTried:\n- {p}\n- {p_repo}\n- {p2}\n- recursive search for: {filename}"
    )


# Columns expected after each loader normalizes (shared slate / attach_standard_refs)
_SLATE_CORE_COLS = (
    "tier", "rank_score", "player", "team", "opp", "prop_type", "pick_type",
    "line", "direction", "hit_rate", "edge", "game_time",
)


def _load_audit_row(sport: str, path: str, df: pd.DataFrame) -> None:
    pe = os.path.isfile(_norm_path(path.strip())) if (path or "").strip() else False
    miss = [c for c in _SLATE_CORE_COLS if c not in df.columns]
    miss_s = ",".join(miss) if miss else "(none)"
    print(f"  [audit {sport}] file_exists={'Y' if pe else 'N'} rows={len(df)} missing_core_cols={miss_s}")


def _extract_game_dates(game_time: pd.Series, target_year: int) -> pd.Series:
    """
    Build canonical YYYY-MM-DD game dates from mixed game_time formats:
    - ISO timestamps (e.g. 2026-03-23T19:00:00-04:00)
    - MM/DD HH:MMAM/PM (e.g. 03/23 07:00PM)
    """
    s = game_time.astype(str).fillna("").str.strip()
    out = pd.Series(pd.NA, index=s.index, dtype="object")

    # ISO-like or full datetime strings that already include a year.
    iso_like = s.str.match(r"^\d{4}[-/]\d{1,2}[-/]\d{1,2}", na=False)
    dt_iso = pd.to_datetime(s.where(iso_like, None), errors="coerce")
    iso_mask = iso_like & dt_iso.notna()
    if iso_mask.any():
        out.loc[iso_mask] = dt_iso.loc[iso_mask].dt.strftime("%Y-%m-%d")

    # MM/DD fallback strings used in slate sheets.
    mmdd = s.str.extract(r"^\s*(\d{1,2})/(\d{1,2})(?:\b|[\sT])")
    mm_mask = mmdd[0].notna() & out.isna()
    if mm_mask.any():
        m = pd.to_numeric(mmdd.loc[mm_mask, 0], errors="coerce")
        d = pd.to_numeric(mmdd.loc[mm_mask, 1], errors="coerce")
        built = pd.to_datetime(
            {"year": target_year, "month": m, "day": d},
            errors="coerce",
        )
        ok = built.notna()
        if ok.any():
            out.loc[mm_mask[mm_mask].index[ok]] = built.loc[ok].dt.strftime("%Y-%m-%d")

    return out


def enforce_target_date(
    df: pd.DataFrame,
    sport: str,
    target_date: str,
    allow_cross_date_fallback: bool = False,
) -> pd.DataFrame:
    """
    Strict date behavior by default:
    - keep only rows matching target_date
    - if none, return empty (sport skipped)
    Optional fallback mode:
    - choose largest upcoming date (>= target) and tie-break by nearest date.
    Soccer: always applies that fallback when the target date has zero rows (PP slate often
    rolls to the next ET calendar day before US sports).
    """
    if df is None or df.empty:
        print(f"  [{sport} date] no rows to filter")
        return df
    if "game_time" not in df.columns:
        if allow_cross_date_fallback:
            print(f"  [{sport} date] missing game_time column; fallback enabled -> keeping {len(df)} rows")
            return df
        print(f"  [{sport} date] missing game_time column; strict mode -> skipping {sport}")
        return df.iloc[0:0].copy()

    out = df.copy()
    target_year = int(str(target_date)[:4])
    parsed = _extract_game_dates(out["game_time"], target_year)
    if "game_date" in out.columns:
        ts = pd.to_datetime(out["game_date"], errors="coerce")
        existing = ts.dt.strftime("%Y-%m-%d").where(ts.notna(), other=pd.NA)
        out["game_date"] = parsed.combine_first(existing)
    else:
        out["game_date"] = parsed

    counts = out["game_date"].value_counts(dropna=True)
    if not counts.empty:
        top = ", ".join([f"{str(k)}={int(v)}" for k, v in counts.head(5).items()])
        print(f"  [{sport} date] available: {top}")
    else:
        if allow_cross_date_fallback:
            print(f"  [{sport} date] no parseable game_date values; fallback enabled -> keeping {len(out)} rows")
            return out
        print(f"  [{sport} date] no parseable game_date values; strict mode -> skipping {sport}")
        return out.iloc[0:0].copy()

    keep_mask = out["game_date"].eq(target_date)
    kept = int(keep_mask.sum())
    total = len(out)

    # Keep cross-date fallback limited to sparse/overnight boards.
    # NBA/MLB period boards and Soccer should not silently roll dates.
    sport_u = str(sport).upper()
    fallback_sports = {"TENNIS"}
    if sport_u in {"NBA", "NBA1Q", "NBA1H", "MLB", "SOCCER", "SOC", "WNBA", "NFL", "NHL"}:
        use_date_fallback = False
    else:
        use_date_fallback = allow_cross_date_fallback or (sport_u in fallback_sports)
    if kept == 0 and use_date_fallback:
        avail = [str(d) for d in counts.index.tolist() if str(d)]
        if avail:
            future_dates = sorted([d for d in avail if d >= target_date])
            chosen = future_dates[0] if future_dates else max(avail)
            keep_mask = out["game_date"].eq(chosen)
            kept = int(keep_mask.sum())
            print(
                f"  [{sport} date] no rows for {target_date}; "
                f"date fallback -> using nearest date {chosen} ({kept} rows)"
            )

    filtered = out.loc[keep_mask].copy()
    dropped = total - len(filtered)
    print(f"  [{sport} date] kept {len(filtered)}/{total}, dropped {dropped}")

    if filtered.empty:
        print(f"  [{sport} date] WARNING: sport skipped for target date {target_date}")

    return filtered


# ──────────────────────────────────────────────────────────────────────────────
# Web outputs (static HTML + JSON) + player images
# ──────────────────────────────────────────────────────────────────────────────
def _safe_float(x, default=None):
    try:
        if x is None:
            return default
        if isinstance(x, float) and np.isnan(x):
            return default
        f = float(x)
        return f if np.isfinite(f) else default
    except Exception:
        return default


def _clean_id(x) -> str:
    """Return a clean integer-like string for IDs, or ''."""
    if x is None:
        return ""
    s = str(x).strip()
    if s == "" or s.lower() == "nan":
        return ""
    # handle 1628368.0
    if re.fullmatch(r"\d+\.0", s):
        s = s[:-2]
    if re.fullmatch(r"\d+", s):
        return s
    return ""

def attach_standard_refs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Adds Standard sibling references to every row (Standard/Goblin/Demon):
      - standard_line
      - standard_edge
      - standard_projection
      - line_discount_vs_standard (direction-aware)

    Matching key uses: sport, player, team, opp, prop_type, game_time
    Bulletproof: supports 'Projection' vs 'projection' and missing cols.
    """
    if df is None or df.empty:
        return df

    out = df.copy()

    # --- unify projection column name (Projection -> projection) ---
    if "projection" not in out.columns and "Projection" in out.columns:
        out["projection"] = out["Projection"]

    # Ensure required columns exist
    for c in [
        "sport", "player", "team", "opp", "prop_type", "pick_type",
        "direction", "line", "edge", "projection", "game_time"
    ]:
        if c not in out.columns:
            out[c] = pd.NA

    key_cols = ["sport", "player", "team", "opp", "prop_type", "game_time"]

    # Build Standard reference table
    std = out[out["pick_type"].astype(str).str.lower() == "standard"].copy()
    if std.empty:
        out["standard_line"] = pd.NA
        out["standard_edge"] = pd.NA
        out["standard_projection"] = pd.NA
        out["line_discount_vs_standard"] = pd.NA
        return out

    std_ref = (
        std[key_cols + ["line", "edge", "projection"]]
        .rename(columns={
            "line": "standard_line",
            "edge": "standard_edge",
            "projection": "standard_projection",
        })
        .drop_duplicates(subset=key_cols, keep="first")
    )

    prior_std = (
        pd.to_numeric(out["standard_line"], errors="coerce")
        if "standard_line" in out.columns
        else pd.Series(pd.NA, index=out.index, dtype="float64")
    )
    out = out.drop(columns=[c for c in ("standard_line", "standard_edge", "standard_projection") if c in out.columns])
    out = out.merge(std_ref, on=key_cols, how="left")
    merged_std = pd.to_numeric(out["standard_line"], errors="coerce")
    out["standard_line"] = prior_std.combine_first(merged_std)
    is_standard = out["pick_type"].astype(str).str.lower().eq("standard")
    out.loc[is_standard, "standard_line"] = out.loc[is_standard, "standard_line"].fillna(
        pd.to_numeric(out.loc[is_standard, "line"], errors="coerce")
    )

    # Direction-aware "discount vs standard"
    def _discount(row):
        try:
            s = row.get("standard_line", pd.NA)
            l = row.get("line", pd.NA)
            if pd.isna(s) or pd.isna(l):
                return pd.NA
            d = str(row.get("direction", "")).upper().strip()
            s = float(s)
            l = float(l)
            if d == "OVER":
                return s - l
            if d == "UNDER":
                return l - s
            return pd.NA
        except Exception:
            return pd.NA

    out["line_discount_vs_standard"] = out.apply(_discount, axis=1)
    return out

def _safe_int_cross_books(v) -> Optional[int]:
    if v is None or v == "":
        return None
    try:
        if isinstance(v, float) and math.isnan(v):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return int(round(float(v)))
    except (TypeError, ValueError):
        return None


def player_initials(name: str) -> str:
    s = (name or "").strip()
    if not s:
        return "?"
    parts = [p for p in re.split(r"\s+", s) if p]
    if len(parts) == 1:
        return parts[0][:1].upper()
    return (parts[0][:1] + parts[-1][:1]).upper()


def compute_image_url(leg: dict) -> Optional[str]:
    """
    NBA:
      needs nba_player_id -> https://cdn.nba.com/headshots/nba/latest/1040x760/<id>.png
    CBB:
      needs espn_player_id -> https://a.espncdn.com/i/headshots/mens-college-basketball/players/full/<id>.png
    WNBA:
      ESPN athlete id when available
    """
    raw = str(leg.get("image_url") or "").strip()
    if raw and raw.lower() not in ("nan", "none", ""):
        return raw
    sport = (leg.get("sport") or "").upper()
    if sport == "NBA":
        pid = _clean_id(leg.get("nba_player_id") or leg.get("player_id"))
        if pid:
            return f"https://cdn.nba.com/headshots/nba/latest/1040x760/{pid}.png"
        return None
    if sport == "CBB":
        eid = _clean_id(leg.get("espn_player_id"))
        if eid:
            return f"https://a.espncdn.com/i/headshots/mens-college-basketball/players/full/{eid}.png"
        return None
    if sport == "WNBA":
        eid = _clean_id(leg.get("espn_player_id") or leg.get("espn_athlete_id"))
        if eid:
            return f"https://a.espncdn.com/i/headshots/wnba/players/full/{eid}.png"
        return None
    return None


_READ_EXPORT_BOOL_KEYS = frozenset({"pick_type_eligible"})
_READ_EXPORT_STR_KEYS = frozenset({"prob_over_source", "prob_under_source"})


def _merge_read_export_fields_into_leg(leg: dict, gv) -> None:
    """Pass pipeline read enrichment columns from source row onto ticket leg dict."""
    for rk in READ_SLATE_EXPORT_KEYS:
        val = gv(rk)
        if val is None or val == "":
            continue
        try:
            if pd.isna(val):
                continue
        except (TypeError, ValueError):
            pass
        if rk in _READ_EXPORT_BOOL_KEYS:
            leg[rk] = bool(val)
        elif rk in _READ_EXPORT_STR_KEYS:
            leg[rk] = str(val).strip()
        else:
            leg[rk] = _safe_float(val)
    gd = gv("game_date")
    if gd is not None and str(gd).strip() not in ("", "nan", "None"):
        leg["game_date"] = str(gd)[:10]
    miss_raw = gv("read_fields_missing")
    if miss_raw:
        try:
            leg["read_fields_missing"] = (
                json.loads(str(miss_raw)) if isinstance(miss_raw, str) else miss_raw
            )
        except json.JSONDecodeError:
            pass
    proj = gv("projection")
    if proj is not None and str(proj).strip() not in ("", "nan", "None"):
        leg["projection"] = _safe_float(proj)


def ticket_groups_to_payload(
    all_ticket_groups, date_str, thresholds, bankroll: float = 0.0, curve_stake_usd: float = 1.0
):
    payload = {
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "date": date_str,
        "filters": thresholds,
        "bankroll": float(bankroll) if bankroll and bankroll > 0 else None,
        "groups": [],
    }

    _ordered_groups = sorted(
        enumerate(all_ticket_groups),
        key=lambda ix_t: (
            _ticket_group_sort_rank(str(ix_t[1][0])),
            _ticket_group_picktype_rank(str(ix_t[1][0])),
            _ticket_group_leg_count(str(ix_t[1][0])),
            _ticket_group_serial(str(ix_t[1][0])),
            ix_t[0],
        ),
    )
    for _, (group_name, tickets, _bg) in _ordered_groups:
        if not tickets:
            continue

        group = {
            "group_name": str(group_name),
            "n_legs": int(tickets[0].get("n_legs", 0) or 0),
            "power_payout": _safe_float(tickets[0].get("power_payout")),
            "flex_payout": _safe_float(tickets[0].get("flex_payout")),
            "tickets": [],
        }

        for ti, t in enumerate(tickets, start=1):
            enrich_ticket_curve_payouts(t, stake_unit=float(curve_stake_usd))
            rows = t.get("rows", [])
            _gn_safe = re.sub(r"[|]+", "_", str(group_name).strip())[:80]
            ticket_id = f"{date_str}|{_gn_safe}|{ti}"
            slip = {
                "web_group_name": str(group_name),
                "ticket_id": ticket_id,
                "ticket_no": ti,
                "avg_hit_rate": _safe_float(t.get("avg_hit_rate")),
                "avg_rank_score": _safe_float(t.get("avg_rank_score")),
                "est_win_prob": _safe_float(t.get("est_win_prob")),
                "ticket_objective_score": _safe_float(t.get("ticket_objective_score")),
                "ticket_live_score": _safe_float(t.get("ticket_live_score")),
                "ticket_model_p_cash": _safe_float(t.get("ticket_model_p_cash")),
                "est_flex_cash_prob": _safe_float(t.get("est_flex_cash_prob")),
                "power_payout": _safe_float(t.get("power_payout")),
                "flex_payout": _safe_float(t.get("flex_payout")),
                "base_power_payout": _safe_float(t.get("base_power_payout")),
                "payout_multiplier": _safe_float(t.get("payout_multiplier")),
                "ev_power": _safe_float(t.get("ev_power")),
                "kelly_units": _safe_float(t.get("kelly_units")),
                "correlation_multiplier": _safe_float(t.get("correlation_multiplier")),
                "correlation_audit": list(t.get("correlation_audit") or []),
                "flat_multiplier": _safe_float(t.get("flat_multiplier")),
                "est_multiplier": _safe_float(t.get("est_multiplier")),
                "mult_error": _safe_float(t.get("mult_error")),
                "est_payout": _safe_float(t.get("est_payout")),
                "flat_payout": _safe_float(t.get("flat_payout")),
                "payout_delta": _safe_float(t.get("payout_delta")),
                "est_ev": _safe_float(t.get("est_ev")),
                "flat_ev": _safe_float(t.get("flat_ev")),
                "combined_hit_prob_curve": _safe_float(t.get("combined_hit_prob_curve")),
                "est_multiplier_flex_nn": _safe_float(t.get("est_multiplier_flex_nn")),
                "flat_multiplier_flex_nn": _safe_float(t.get("flat_multiplier_flex_nn")),
                "using_flat_fallback": bool(t.get("using_flat_fallback")),
                "has_data_warning": False,
                "legs": [],
            }

            for row in rows:

                def gv(field):
                    return row.get(field, "") if isinstance(row, dict) else getattr(row, field, "")

                _dpv = gd_leg_delta_pct(gv("line"), gv("standard_line"))
                sport_s = str(gv("sport") or t.get("sport") or "").strip()
                player_s = str(gv("player") or "")
                team_s = str(gv("team") or "")
                opp_s = str(gv("opp") or gv("opp_team") or "").strip()
                prop_s = str(gv("prop_type") or "")
                dir_s = str(gv("direction") or "")
                game_time_s = str(gv("game_time") or "")
                best_book_s = str(gv("best_cross_book") or "")
                line_f = _safe_float(gv("line"))
                line_key = f"{float(line_f):.3f}" if line_f is not None else ""
                id_material = "|".join(
                    [
                        sport_s.strip().lower(),
                        player_s.strip().lower(),
                        team_s.strip().lower(),
                        opp_s.strip().lower(),
                        prop_s.strip().lower(),
                        line_key,
                        dir_s.strip().lower(),
                        game_time_s.strip().lower(),
                        best_book_s.strip().lower(),
                        str(date_str).strip().lower(),
                    ]
                )
                canonical_leg_id = "leg_" + hashlib.sha1(id_material.encode("utf-8")).hexdigest()[:20]
                leg = {
                    "ticket_id": ticket_id,
                    "sport": sport_s,
                    "player": player_s,
                    "team": team_s,
                    "opp": opp_s,
                    "prop_type": prop_s,
                    "pick_type": str(gv("pick_type") or ""),
                    "direction": dir_s,
                    "line": line_f,
                    "edge": _safe_float(gv("edge")),
                    "abs_edge": _safe_float(gv("abs_edge")),
                    "standard_line": _safe_float(gv("standard_line")),
                    "standard_edge": _safe_float(gv("standard_edge")),
                    "standard_projection": _safe_float(gv("standard_projection")),
                    "line_discount_vs_standard": _safe_float(gv("line_discount_vs_standard")),
                    "delta_pct": round(float(_dpv), 4) if _dpv is not None else None,
                    "line_underdog": _safe_float(gv("line_underdog")),
                    "line_draftkings": _safe_float(gv("line_draftkings")),
                    "pick_platform": str(gv("pick_platform") or "prizepicks"),
                    "best_cross_line": _safe_float(gv("best_cross_line")),
                    "best_cross_book": best_book_s,
                    "cross_edge_vs_pp": _safe_float(gv("cross_edge_vs_pp")),
                    "cross_n_books": _safe_int_cross_books(gv("cross_n_books")),
                    "hit_rate": _safe_float(gv("hit_rate")),
                    "over_hit_rate": _safe_float(gv("over_hit_rate") or gv("hit_rate_over_L5")),
                    "under_hit_rate": _safe_float(gv("under_hit_rate") or gv("hit_rate_under_L5")),
            "ml_prob": _safe_float(gv("ml_prob")),
            "rank_score": _safe_float(gv("rank_score")),
            "tier": str(gv("tier") or gv("Tier") or ""),
            "opponent_def_rank": _safe_float(
                gv("opponent_def_rank")
                or gv("opp_def_rank")
                or gv("OVERALL_DEF_RANK")
                or gv("def_rank")
            ),
                    "game_time": game_time_s,
                    "event_start_time": game_time_s or None,
                    "posted_at": str(gv("posted_at") or "") or None,
                    "lock_at": str(gv("lock_at") or gv("game_time") or "") or None,
                    "source_priority": "prizepicks_primary",
                    "lineage_version": "v1",
                    "canonical_leg_id": canonical_leg_id,
                    "nba_player_id": gv("nba_player_id"),
                    "espn_player_id": gv("espn_player_id"),
                    "min_tier": str(gv("min_tier") or gv("minutes_tier") or gv("Min Tier") or ""),
                    "shot_role": str(gv("shot_role") or gv("Shot Role") or ""),
                    "usage_role": str(gv("usage_role") or gv("Usage Role") or ""),
                    "l5_avg": _safe_float(gv("l5_avg") or gv("Last 5 Avg") or gv("last_5_avg") or gv("intel_l5_avg")),
                    "season_avg": _safe_float(gv("season_avg") or gv("Season Avg") or gv("avg_season") or gv("intel_season_avg")),
                    "intel_season_hit_rate":   _safe_float(gv("intel_season_hit_rate")),
                    "intel_cushion":           _safe_float(gv("intel_cushion")),
                    "intel_cv_pct":            _safe_float(gv("intel_cv_pct")),
                    "intel_opp_vs_league_pct": _safe_float(gv("intel_opp_vs_league_pct")),
                    "intel_l5_vs_season":      _safe_float(gv("intel_l5_vs_season")),
                    "l5_over": _safe_float(gv("l5_over") or gv("L5 Over") or gv("line_hits_over_5")),
                    "l5_under": _safe_float(gv("l5_under") or gv("L5 Under") or gv("line_hits_under_5")),
                    "l5_games_played": _safe_float(
                        gv("l5_games_played") or gv("line_games_played_5") or gv("Games (5g)")
                    ),
                    "l10_games_played": _safe_float(
                        gv("l10_games_played") or gv("line_games_played_10") or gv("Games (10g)")
                    ),
                    "l5_side_hits": _safe_float(gv("l5_side_hits")),
                    "l5_consistency": _safe_float(gv("l5_consistency")),
                    "l10_over": _safe_float(gv("l10_over") or gv("L10 Over") or gv("hit_rate_over_L10") or gv("over_L10")),
                    "l10_under": _safe_float(gv("l10_under") or gv("L10 Under") or gv("hit_rate_under_L10") or gv("under_L10")),
                    "def_tier": str(gv("def_tier") or gv("Def Tier") or ""),
                    "pace_tier": str(gv("pace_tier") or gv("Pace Tier") or ""),
                    "context_score": _safe_float(gv("context_score")),
                    "usage_boost": _safe_float(gv("usage_boost")),
                    "usage_boost_reason": str(gv("usage_boost_reason") or ""),
                }
                # Preserve per-game history so UI can render true actual-vs-line charts.
                for _i in range(1, 11):
                    _gv = gv(f"G{_i}")
                    if _gv is None or _gv == "":
                        _gv = gv(f"g{_i}")
                    if _gv is None or _gv == "":
                        _gv = gv(f"stat_g{_i}")
                    _hist_v = _safe_float(_gv)
                    if _hist_v is not None:
                        leg[f"g{_i}"] = _hist_v
                        leg[f"stat_g{_i}"] = _hist_v
                    _lv = gv(f"line_g{_i}")
                    if _lv is None or _lv == "":
                        _lv = gv(f"prop_line_g{_i}")
                    _line_hist_v = _safe_float(_lv)
                    if _line_hist_v is not None:
                        leg[f"line_g{_i}"] = _line_hist_v
                _merge_read_export_fields_into_leg(leg, gv)
                leg["data_warning"] = "LIMITED_Q1_HISTORY" if str(leg.get("sport", "")).upper() == "NBA1Q" else None
                leg_prob_used, leg_prob_source = _resolve_leg_prob(pd.Series(leg))
                leg["leg_prob_used"] = _safe_float(leg_prob_used)
                leg["leg_prob_source"] = leg_prob_source
                leg["image_url"] = compute_image_url(leg)
                leg["initials"] = player_initials(leg.get("player", ""))
                br = float(bankroll) if bankroll and float(bankroll) > 0 else 0.0
                if br > 0:
                    p_raw = leg_prob_used if leg_prob_used is not None else _safe_float(gv("ml_prob"))
                    try:
                        p_f = float(p_raw) if p_raw is not None and not (isinstance(p_raw, float) and math.isnan(p_raw)) else 0.5
                    except (TypeError, ValueError):
                        p_f = 0.5
                    e_pct = leg_edge_pct_for_kelly(_safe_float(gv("ml_prob")), _safe_float(gv("edge")))
                    leg["recommended_stake_usd"] = fractional_kelly(e_pct, p_f, br)
                else:
                    leg["recommended_stake_usd"] = None

                slip["legs"].append(leg)
            slip["has_data_warning"] = any(bool(x.get("data_warning")) for x in slip["legs"])

            try:
                slip["payout"] = build_ticket_payout_json(str(group_name), rows)
            except Exception:
                slip["payout"] = None

            _enrich_slip_p_win_fields(slip, mode=str(t.get("mode") or "ev"))

            group["tickets"].append(slip)

        payload["groups"].append(group)

    apply_slate_ev_tier_recommendations(payload)
    return payload


def dataframe_to_slate_sport_rows(df: Optional[pd.DataFrame]) -> List[dict]:
    """Convert a step7/step8 direction dataframe to Slate Explorer row dicts (see _SLATE_SPORT_UI_KEYS)."""
    import math

    if df is None or len(df) == 0:
        return []
    df = enrich_read_fields_dataframe(df)

    def safe(v):
        if v is None:
            return None
        try:
            if pd.isna(v):
                return None
        except (TypeError, ValueError):
            pass
        try:
            if isinstance(v, float) and math.isnan(v):
                return None
        except Exception:
            pass
        if hasattr(v, "item"):
            return v.item()
        return v

    rows: List[dict] = []
    for _, r in df.iterrows():
        def g(c):
            return safe(r[c]) if c in df.columns else None

        row = {
            "tier":       g("tier"),
            "rank_score": g("rank_score"),
            "player":     g("player") or "",
            "team":       g("team") or "",
            "pos":        g("pos") or g("Pos"),
            "opp":        str(g("opp") or g("opp_team") or "").strip(),
            "prop":       g("prop_type") or g("prop") or "",
            "pick_type":  g("pick_type") or "",
            "line":       g("line"),
            "dir":        g("direction") or g("dir") or "",
            "edge":       g("edge"),
            "abs_edge":   g("abs_edge"),
            "hit_rate":   g("hit_rate"),
            "l5_over":    g("l5_over"),
            "l5_under":   g("l5_under"),
            "l5_games_played": g("l5_games_played"),
            "l10_over":   g("l10_over"),
            "l10_under":  g("l10_under"),
            "l10_games_played": g("l10_games_played"),
            "season_avg": g("season_avg") or g("szn_avg"),
            "ml_prob":    g("ml_prob"),
            "def_tier":   g("def_tier") if g("def_tier") else g("Def Tier"),
            "opponent_def_rank": g("opponent_def_rank")
            or g("opp_def_rank")
            or g("OVERALL_DEF_RANK")
            or g("def_rank"),
            "standard_line": g("standard_line"),
            "standard_projection": g("standard_projection"),
            "projection": g("projection") or g("intel_projection"),
            "game_time":  str(g("game_time") or "") or None,
        }
        if "pick_platform" in df.columns and g("pick_platform") is not None:
            row["pick_platform"] = g("pick_platform")
        if "line_underdog" in df.columns and g("line_underdog") is not None:
            row["line_underdog"] = g("line_underdog")
        if "line_draftkings" in df.columns and g("line_draftkings") is not None:
            row["line_draftkings"] = g("line_draftkings")
        sport_val = str(g("sport") or "").strip().upper()
        if sport_val:
            row["sport"] = sport_val

        leg_s = pd.Series(
            {
                **row,
                "direction": row.get("dir"),
                "bet_direction": row.get("dir"),
                "hit_rate": row.get("hit_rate"),
                "l5_over": row.get("l5_over"),
                "l5_under": row.get("l5_under"),
                "ml_prob": row.get("ml_prob"),
            }
        )
        try:
            prob_used, prob_src = _resolve_leg_prob(leg_s)
            row["leg_prob_used"] = round(float(prob_used), 4)
            row["leg_prob_source"] = prob_src
        except Exception:
            pass

        for rk in READ_SLATE_EXPORT_KEYS:
            if rk not in df.columns:
                continue
            val = g(rk)
            if val is None:
                continue
            if rk == "pick_type_eligible":
                row[rk] = bool(val)
            elif rk == "read_fields_missing":
                continue
            else:
                row[rk] = val
        miss_raw = g("read_fields_missing")
        if miss_raw:
            try:
                row["read_fields_missing"] = (
                    json.loads(str(miss_raw)) if isinstance(miss_raw, str) else miss_raw
                )
            except json.JSONDecodeError:
                row["read_fields_missing"] = []
        mt = g("minutes_tier") or g("min_tier")
        if mt:
            row["minutes_tier"] = mt

        img = g("image_url")
        if img and str(img).strip().lower() not in ("nan", "none"):
            row["image_url"] = str(img).strip()
        else:
            computed = compute_image_url({**row, "sport": sport_val})
            if computed:
                row["image_url"] = computed
        nba_pid = g("nba_player_id") or g("player_id")
        if nba_pid:
            row["nba_player_id"] = _clean_id(nba_pid)
        eid = g("espn_player_id")
        if eid:
            row["espn_player_id"] = _clean_id(eid)

        base_line = safe(g("standard_line")) or safe(g("line"))
        actuals: List[float] = []
        line_hist: List[float] = []
        for _i in range(1, 11):
            av = safe(g(f"stat_g{_i}") or g(f"g{_i}") or g(f"G{_i}"))
            lv = safe(g(f"line_g{_i}"))
            if av is not None:
                try:
                    av_num = float(av)
                except (TypeError, ValueError):
                    continue
                actuals.append(av_num)
                if lv is not None:
                    try:
                        line_hist.append(float(lv))
                    except (TypeError, ValueError):
                        line_hist.append(float(base_line) if base_line is not None else av_num)
                else:
                    line_hist.append(float(base_line) if base_line is not None else av_num)
                row[f"stat_g{_i}"] = av
                row[f"g{_i}"] = av
                if lv is not None:
                    row[f"line_g{_i}"] = lv
        if actuals:
            row["actual_series"] = actuals
            if base_line is not None:
                row["line_series"] = line_hist if line_hist else [float(base_line)] * len(actuals)
            elif line_hist:
                row["line_series"] = line_hist

        gd = g("game_date")
        if gd:
            row["game_date"] = str(gd)[:10]
        dr = g("days_rest")
        if dr is not None:
            row["rest_days"] = dr
        b2b = g("b2b_flag")
        if b2b is not None:
            row["back_to_back"] = str(b2b).strip().lower() in ("1", "true", "yes", "y")
        if row.get("projection") is None:
            l5a = g("l5_avg")
            if l5a is not None:
                row["projection"] = l5a

        rows.append({k: v for k, v in row.items() if v is not None})
    return rows


def resolve_default_wnba_step8_path(date_str: str) -> str:
    """Same resolution order as apply_default_sport_inputs --wnba (standalone canonical tree)."""
    d = str(date_str).strip()[:10]
    if len(d) != 10:
        return ""
    out = _outputs_dir_for_date(d)
    return _first_existing_path(
        os.path.join(out, "wnba", "step8_wnba_direction_clean.xlsx"),
        os.path.join(out, "wnba", "step8_wnba_direction.xlsx"),
        os.path.join(out, f"step8_wnba_direction_clean_{d}.xlsx"),
        os.path.join(out, f"step8_wnba_direction_{d}.xlsx"),
        os.path.join(REPO_ROOT, "Sports", "WNBA", "outputs", "step8_wnba_direction_clean.xlsx"),
        os.path.join(REPO_ROOT, "Sports", "WNBA", "outputs", "step8_wnba_direction.xlsx"),
        os.path.join(REPO_ROOT, "Sports", "WNBA", "step8_wnba_direction_clean.xlsx"),
        os.path.join(REPO_ROOT, "Sports", "WNBA", "step8_wnba_direction.xlsx"),
        os.path.join(REPO_ROOT, "WNBA", "step8_wnba_direction_clean.xlsx"),
        os.path.join(REPO_ROOT, "WNBA", "step8_wnba_direction.xlsx"),
    )


def _overlay_wnba_defense_ranks(df: pd.DataFrame) -> pd.DataFrame:
    """Refresh opp def rank/tier from wnba_defense_summary.csv (slate + matchup panel parity)."""
    if df is None or len(df) == 0:
        return df
    def_path = os.path.join(REPO_ROOT, "Sports", "WNBA", "wnba_defense_summary.csv")
    if not os.path.isfile(def_path):
        return df
    try:
        from utils.wnba_team_keys import defense_team_key
    except ImportError:
        return df
    try:
        ddef = pd.read_csv(def_path, encoding="utf-8-sig")
    except Exception:
        return df
    if "TEAM_ABBREVIATION" not in ddef.columns:
        return df
    rank_by_key: dict[str, float] = {}
    tier_by_key: dict[str, str] = {}
    for r in ddef.itertuples(index=False):
        key = defense_team_key(getattr(r, "TEAM_ABBREVIATION", ""))
        if not key:
            continue
        rk = pd.to_numeric(getattr(r, "OVERALL_DEF_RANK", np.nan), errors="coerce")
        if pd.notna(rk):
            rank_by_key[key] = float(rk)
        tier = str(getattr(r, "DEF_TIER", "") or "").strip()
        if tier:
            tier_by_key[key] = tier
    if not rank_by_key:
        return df
    out = df.copy()
    opp_col = "opp" if "opp" in out.columns else ("opp_team" if "opp_team" in out.columns else None)
    if not opp_col:
        return out
    opp_keys = out[opp_col].astype(str).str.strip().str.upper().map(defense_team_key)
    out["opponent_def_rank"] = opp_keys.map(rank_by_key)
    out["OVERALL_DEF_RANK"] = out["opponent_def_rank"]
    out["def_rank"] = out["opponent_def_rank"]
    if tier_by_key:
        out["def_tier"] = opp_keys.map(tier_by_key)
        out["DEF_TIER"] = out["def_tier"]
    return out


def publish_wnba_slate_merge_into_web(
    date_str: str,
    web_outdirs: Optional[str | List[str]] = None,
) -> bool:
    """
    Merge WNBA step8 rows into existing slate_latest.json (standalone WNBA pipeline → UI).

    Preserves other sports already in the file; replaces only sports[\"wnba\"].
    Writes slate_sport_wnba.json alongside (static /api/slate-sport/wnba shape).
    """
    d = str(date_str).strip()[:10]
    if len(d) != 10:
        print("[wnba-slate-web] invalid date")
        return False

    wnba_path = resolve_default_wnba_step8_path(d)
    if not wnba_path:
        print("[wnba-slate-web] no WNBA step8 workbook found; skip web slate merge")
        return False
    try:
        wnba_df = load_wnba(wnba_path)
    except Exception as exc:
        print(f"[wnba-slate-web] load_wnba failed: {exc}")
        return False
    if wnba_df is None or len(wnba_df) == 0:
        print("[wnba-slate-web] WNBA board empty; skip web slate merge")
        return False

    wnba_df = _overlay_wnba_defense_ranks(wnba_df)
    rows = dataframe_to_slate_sport_rows(wnba_df)
    if web_outdirs is None:
        web_outdirs = [os.path.join(REPO_ROOT, "ui_runner", "templates")]
        _mob = os.path.join(REPO_ROOT, "mobile", "www")
        if os.path.isdir(_mob):
            web_outdirs.append(_mob)
    elif isinstance(web_outdirs, str):
        web_outdirs = [web_outdirs]

    gen_at = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    for outdir in web_outdirs:
        if not str(outdir).strip():
            continue
        os.makedirs(outdir, exist_ok=True)
        slate_path = os.path.join(outdir, "slate_latest.json")
        payload = None
        if os.path.isfile(slate_path):
            try:
                with open(slate_path, encoding="utf-8") as sf:
                    payload = json.load(sf)
            except Exception as exc:
                print(f"[wnba-slate-web] WARN resetting {slate_path}: {exc}")
                payload = None
        if not isinstance(payload, dict):
            payload = {"date": d, "generated_at": gen_at, "sports": {}}
        sports = payload.get("sports")
        if not isinstance(sports, dict):
            sports = {}
        else:
            sports = dict(sports)
        sports["wnba"] = rows
        payload["sports"] = sports
        if not str(payload.get("date") or "").strip():
            payload["date"] = d
        payload["generated_at"] = gen_at
        _write_json_file(slate_path, payload)
        sport_path = os.path.join(outdir, "slate_sport_wnba.json")
        _write_json_file(sport_path, {"ok": True, "sport": "wnba", "rows": rows})
        n_tot = sum(len(v or []) for v in (payload.get("sports") or {}).values() if isinstance(v, list))
        print(f"  [wnba-slate-web] {slate_path}  (wnba={len(rows)} rows, all_sports={n_tot} props)")
        print(f"  [wnba-slate-web] {sport_path}")
    return True


def write_slate_json(nba, cbb, nhl, soccer, date_str, outdir,
                     wcbb=None, mlb=None, nba1q=None, nba1h=None, tennis=None, nfl=None, wnba=None, cfb=None,
                     tennis_date=None):
    """Write full per-sport ranked slate to slate_latest.json for the web UI.

    Sport keys in ``sports`` are lowercase (nba, nfl, …) so /api/slate-sport and the
    home Slate Explorer stay aligned with templates/index.html (SPORTS + _slate_counts).
    """
    payload = {
        "date": date_str,
        "generated_at": datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC"),
        "sports": {
            "nba":    dataframe_to_slate_sport_rows(nba),
            "cbb":    dataframe_to_slate_sport_rows(cbb),
            "cfb":    dataframe_to_slate_sport_rows(cfb),
            "nhl":    dataframe_to_slate_sport_rows(nhl),
            "soccer": dataframe_to_slate_sport_rows(soccer),
            "tennis": dataframe_to_slate_sport_rows(tennis),
            "wcbb":   dataframe_to_slate_sport_rows(wcbb),
            "mlb":    dataframe_to_slate_sport_rows(mlb),
            "nba1q":  dataframe_to_slate_sport_rows(nba1q),
            "nba1h":  dataframe_to_slate_sport_rows(nba1h),
            "nfl":    dataframe_to_slate_sport_rows(nfl),
            "wnba":   dataframe_to_slate_sport_rows(wnba),
        }
    }
    tennis_rows = payload["sports"].get("tennis") or []
    if tennis_date and isinstance(tennis_rows, list) and len(tennis_rows) > 0:
        payload["tennis_date"] = str(tennis_date).strip()[:10]

    os.makedirs(outdir, exist_ok=True)
    out_path = os.path.join(outdir, "slate_latest.json")
    payload = _sanitize_for_json(payload)
    with open(out_path, "w", encoding="utf-8") as f:
        import json as _json

        _json.dump(payload, f, ensure_ascii=False, default=str, allow_nan=False)
    print(f"  slate_latest.json -> {out_path}  ({sum(len(v) for v in payload['sports'].values())} props)")

    # Static replacements for /api/slate-sport/<sport> and /api/slate-sport/combined.
    sports_payload = payload.get("sports") if isinstance(payload, dict) else {}
    if not isinstance(sports_payload, dict):
        sports_payload = {}

    combined_rows: list[dict] = []
    for sport_key, rows in sports_payload.items():
        safe_rows = rows if isinstance(rows, list) else []
        sport_path = os.path.join(outdir, f"slate_sport_{sport_key}.json")
        with open(sport_path, "w", encoding="utf-8") as sf:
            _json.dump(
                {"ok": True, "sport": sport_key, "rows": safe_rows},
                sf,
                ensure_ascii=False,
                default=str,
                allow_nan=False,
            )
        for r in safe_rows:
            rr = dict(r) if isinstance(r, dict) else {"value": r}
            if isinstance(rr, dict) and not str(rr.get("sport") or "").strip():
                rr["sport"] = str(sport_key).upper()
            combined_rows.append(rr)

    combined_path = os.path.join(outdir, "slate_sport_combined.json")
    with open(combined_path, "w", encoding="utf-8") as cf:
        _json.dump(
            {"ok": True, "sport": "combined", "rows": combined_rows},
            cf,
            ensure_ascii=False,
            default=str,
            allow_nan=False,
        )
    print(f"  slate_sport_combined.json -> {combined_path}  ({len(combined_rows)} rows)")


def render_tickets_html(payload: dict) -> str:
    """Build full tickets page HTML from the same structure as tickets_latest.json."""

    def fmt_pct(x) -> str:
        try:
            if x is None:
                return ""
            return f"{float(x) * 100:.2f}%"
        except Exception:
            return ""

    def fmt_2(x) -> str:
        try:
            if x is None:
                return ""
            return f"{float(x):.2f}"
        except Exception:
            return ""

    def fmt_line(x) -> str:
        # keep lines readable (avoid 5.5000000003)
        try:
            if x is None:
                return ""
            xf = float(x)
            if abs(xf - round(xf)) < 1e-9:
                return str(int(round(xf)))
            return f"{xf:.2f}".rstrip("0").rstrip(".")
        except Exception:
            return str(x) if x is not None else ""

    # ── helpers ────────────────────────────────────────────────────────────────
    def hit_color(x) -> str:
        try:
            v = float(x)
            if v >= 0.65:
                return "#00F2FF"
            if v >= 0.50:
                return "#f0a500"
            return "#c96a74"
        except Exception:
            return "#c96a74"

    def sport_badge(sport: str) -> str:
        s = (sport or "").upper()
        if "NBA" in s:
            return "<span style='background:#c8ff00;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>NBA</span>"
        if "CFB" in s or s == "NCAAF":
            return "<span style='background:#8B4513;color:#fff;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>CFB</span>"
        if "CBB" in s or ("NCAA" in s and "CFB" not in s):
            return "<span style='background:#00e5ff;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>CBB</span>"
        if "NHL" in s:
            return "<span style='background:#5bc4f5;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>NHL</span>"
        if "SOCCER" in s:
            return "<span style='background:#57e87d;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>SOC</span>"
        if "TENNIS" in s:
            return "<span style='background:#aed581;color:#000;font-size:11px;font-weight:700;padding:2px 7px;border-radius:4px;letter-spacing:.04em;'>TEN</span>"
        return f"<span style='background:#333;color:#ccc;font-size:11px;padding:2px 7px;border-radius:4px;'>{sport or ''}</span>"

    def badge(val, color="#00F2FF") -> str:
        if not val:
            return "<span style='color:#555;font-size:12px;'>—</span>"
        return f"<span style='background:rgba(0,0,0,.35);color:{color};font-size:12px;padding:2px 8px;border-radius:4px;border:1px solid {color}33;'>{val}</span>"

    def wp_bar(wp) -> str:
        try:
            pct = float(wp) * 100
            w = max(2, min(100, pct))
            col = "#00F2FF" if pct >= 50 else "#f0a500"
            return (
                f"<div style='display:flex;align-items:center;gap:8px;'>"
                f"<div style='flex:1;height:6px;background:#1a1a2e;border-radius:3px;overflow:hidden;'>"
                f"<div style='width:{w:.1f}%;height:100%;background:{col};border-radius:3px;'></div></div>"
                f"<span style='color:{col};font-family:\"Bebas Neue\",sans-serif;font-size:15px;letter-spacing:.05em;'>{pct:.1f}%</span>"
                f"</div>"
            )
        except Exception:
            return ""

    def direction_signal(leg: dict):
        """
        User-facing decision helper.
        Returns (signal_html, reason_text) based on available context.
        """
        score, reasons = compute_bet_signal_core(leg)
        joined = " + ".join(reasons) if reasons else ""
        if score >= 3:
            return "<span class='sig-strong'>STRONG</span>", joined or "aligned context"
        if score >= 2:
            return "<span class='sig-lean'>LEAN</span>", joined or "partial context"
        return "<span class='sig-risk'>RISKY</span>", joined or "limited context support"

    # ── HTML ───────────────────────────────────────────────────────────────────
    filters = payload.get("filters", {})
    gen_at  = payload.get("generated_at", "")
    date_declared_raw = (payload.get("date") or "").strip()
    date_declared = date_declared_raw[:10] if len(date_declared_raw) >= 10 else date_declared_raw

    def fmt_slate_date_pretty(iso: str) -> str:
        """M/D/YYYY from YYYY-MM-DD (no ambiguous 04-04 style)."""
        s = (iso or "").strip()[:10]
        if len(s) != 10 or s[4] != "-" or s[7] != "-":
            return (iso or "").strip() or "—"
        try:
            y, m, d = int(s[0:4]), int(s[5:7]), int(s[8:10])
            if not (1 <= m <= 12 and 1 <= d <= 31):
                return iso
            return f"{m}/{d}/{y}"
        except (TypeError, ValueError):
            return iso or "—"

    def _calendar_date_from_game_time(gs: str) -> str | None:
        """Calendar YYYY-MM-DD in the prop's local offset (or parsed instant)."""
        s = (gs or "").strip()
        if not s:
            return None
        candidates = [s]
        if " " in s and "T" not in s.split(" ", 1)[0]:
            candidates.append(s.replace(" ", "T", 1))
        for cand in candidates:
            try:
                c2 = cand.replace("Z", "+00:00") if cand.endswith("Z") else cand
                dt = datetime.fromisoformat(c2)
                return dt.date().isoformat()
            except ValueError:
                continue
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            head = s[:10]
            if head[0:4].isdigit() and head[5:7].isdigit() and head[8:10].isdigit():
                return head
        return None

    def _modal_slate_date_from_legs(p: dict) -> str | None:
        counts: dict[str, int] = {}
        for g in p.get("groups") or []:
            for t in g.get("tickets") or []:
                for leg in t.get("legs") or []:
                    cd = _calendar_date_from_game_time(str(leg.get("game_time") or ""))
                    if cd:
                        counts[cd] = counts.get(cd, 0) + 1
        if not counts:
            return None
        return max(counts.items(), key=lambda kv: (kv[1], kv[0]))[0]

    date_from_legs = _modal_slate_date_from_legs(payload)
    date_eff = date_from_legs or date_declared or ""
    if len(date_eff) > 10:
        date_eff = date_eff[:10]

    date_mismatch_html = ""
    if date_from_legs and date_declared and date_from_legs != date_declared:
        date_mismatch_html = (
            f' <span style="opacity:.65;font-size:11px;">file date {fmt_slate_date_pretty(date_declared)}</span>'
        )

    date_pretty = fmt_slate_date_pretty(date_eff)

    CSS = """
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&family=Inter:wght@400;500;600;700;800&display=swap');
*{box-sizing:border-box;margin:0;padding:0;}
:root{
  --bg:#050505;--surface:rgba(20,20,20,0.60);--card:rgba(20,20,20,0.60);--border:rgba(212,175,55,0.15);
  --accent:#d4af37;--cyan:#00F2FF;--muted:#ffffff;--muted2:#f0f0f0;--text:#e8e8f0;
}
body{background:var(--bg);color:var(--text);font-family:'Inter',sans-serif;min-height:100vh;overflow-x:hidden;}

body::before{
  content:'';position:fixed;inset:0;
  background:
    radial-gradient(1200px 760px at -8% -18%, rgba(0,242,255,.12) 0%, transparent 56%),
    radial-gradient(980px 620px at 108% -8%, rgba(212,175,55,.18) 0%, transparent 54%),
    linear-gradient(180deg,#050505 0%,#080808 52%,#0f0f0f 100%);
  pointer-events:none;z-index:0;
}

/* scanlines */
body::after{content:'';position:fixed;inset:0;pointer-events:none;z-index:0;}

#app{position:relative;z-index:1;width:100%;max-width:min(1920px,98vw);margin:0 auto;padding:0 clamp(10px,2.5vw,28px) 24px;box-sizing:border-box;}

/* nav */
nav{display:flex;align-items:center;gap:16px;padding:10px 0 12px;border-bottom:1px solid rgba(196,166,107,.22);flex-wrap:wrap;position:sticky;top:0;z-index:220;background:rgba(7,10,19,0.90);backdrop-filter:blur(22px) saturate(180%);}
.nav-logo{display:flex;align-items:center;gap:12px;text-decoration:none;}
.brain-wrap{display:none;}
.nav-logo::before{
  content:"";width:64px;height:40px;flex:0 0 64px;border-radius:10px;
  background:url('/static/hybrid-logo.png?v=20260320a') center/contain no-repeat;
  filter:drop-shadow(0 6px 14px rgba(0,0,0,.35));
}
.brain-slate{position:absolute;inset:0;border-radius:7px;background:linear-gradient(145deg,#12122a 0%,#080818 100%);border:1px solid #252545;animation:slateBreak 3.5s ease-in-out infinite;}
.brain-slate::before{content:'';position:absolute;inset:0;background:linear-gradient(to bottom right,transparent 47%,#c8ff0044 49%,transparent 51%),linear-gradient(to bottom left,transparent 44%,#c8ff0022 46%,transparent 48%),linear-gradient(to right,transparent 30%,#00e5ff22 31%,transparent 33%);border-radius:7px;animation:crackGlow 3.5s ease-in-out infinite;}
@keyframes slateBreak{0%,100%{transform:scale(1);box-shadow:0 0 0px #c8ff0000;}48%{transform:scale(1.06) rotate(-0.5deg);box-shadow:0 0 24px #c8ff0055;}50%{transform:scale(1.10) rotate(0.5deg);box-shadow:0 0 40px #c8ff0088;}52%{transform:scale(1.06) rotate(-0.3deg);box-shadow:0 0 24px #c8ff0055;}}
@keyframes crackGlow{0%,100%{opacity:0.2;}50%{opacity:1;}}
.brain-svg{position:absolute;inset:3px;animation:brainBreakthrough 3.5s ease-in-out infinite;transform-origin:center bottom;}
@keyframes brainBreakthrough{0%,100%{transform:scale(1) translateY(0px);filter:drop-shadow(0 0 5px #c8ff0099) drop-shadow(0 0 2px #00e5ff66);}48%{transform:scale(1.07) translateY(-1px);filter:drop-shadow(0 0 12px #c8ff00cc) drop-shadow(0 0 8px #00e5ffaa);}50%{transform:scale(1.18) translateY(-3px);filter:drop-shadow(0 0 20px #c8ff00ff) drop-shadow(0 0 14px #00e5ffcc) drop-shadow(0 0 40px #c8ff0044);}52%{transform:scale(1.07) translateY(-1px);filter:drop-shadow(0 0 12px #c8ff00cc) drop-shadow(0 0 8px #00e5ffaa);}}
.brain-pulse-ring{position:absolute;border-radius:9px;border:1.5px solid #c8ff00;opacity:0;animation:brainRingExpand 3.5s ease-out infinite;inset:-3px;}
.brain-pulse-ring:nth-child(2){border-color:#00e5ff;animation-delay:0.15s;}
.brain-pulse-ring:nth-child(3){border-color:#c8ff0088;animation-delay:0.3s;}
.brain-pulse-ring:nth-child(4){border-color:#00e5ff66;animation-delay:0.45s;}
@keyframes brainRingExpand{0%,48%{transform:scale(1);opacity:0;}50%{transform:scale(1);opacity:0.9;}85%{transform:scale(2.4);opacity:0;}100%{transform:scale(2.4);opacity:0;}}
.bspark{position:absolute;border-radius:50%;opacity:0;animation:bsparkFly 3.5s ease-out infinite;}
.bspark.lg{width:4px;height:4px;background:#c8ff00;box-shadow:0 0 6px #c8ff00;}
.bspark.md{width:3px;height:3px;background:#00e5ff;box-shadow:0 0 5px #00e5ff;}
.bspark.sm{width:2px;height:2px;background:#c8ff00cc;}
.bspark.cy{width:2px;height:2px;background:#00e5ffcc;}
.bspark.wh{width:2px;height:2px;background:#ffffffaa;}
.bspark:nth-child(5) {top:10%;left:5%; --tx:-18px;--ty:-16px;animation-delay:0.50s;}
.bspark:nth-child(6) {top:5%; left:40%;--tx:2px;  --ty:-22px;animation-delay:0.52s;}
.bspark:nth-child(7) {top:8%; left:75%;--tx:16px; --ty:-18px;animation-delay:0.54s;}
.bspark:nth-child(8) {top:30%;left:96%;--tx:22px; --ty:-8px; animation-delay:0.51s;}
.bspark:nth-child(9) {top:55%;left:96%;--tx:20px; --ty:8px;  animation-delay:0.53s;}
.bspark:nth-child(10){top:80%;left:86%;--tx:14px; --ty:16px; animation-delay:0.55s;}
.bspark:nth-child(11){top:92%;left:55%;--tx:4px;  --ty:22px; animation-delay:0.50s;}
.bspark:nth-child(12){top:90%;left:25%;--tx:-10px;--ty:20px; animation-delay:0.52s;}
.bspark:nth-child(13){top:72%;left:2%; --tx:-20px;--ty:12px; animation-delay:0.54s;}
.bspark:nth-child(14){top:45%;left:0%; --tx:-22px;--ty:0px;  animation-delay:0.51s;}
.bspark:nth-child(15){top:20%;left:2%; --tx:-18px;--ty:-12px;animation-delay:0.56s;}
.bspark:nth-child(16){top:15%;left:60%;--tx:10px; --ty:-20px;animation-delay:0.53s;}
.bspark:nth-child(17){top:18%;left:20%;--tx:-14px;--ty:-18px;animation-delay:0.65s;}
.bspark:nth-child(18){top:12%;left:55%;--tx:6px;  --ty:-20px;animation-delay:0.67s;}
.bspark:nth-child(19){top:25%;left:88%;--tx:18px; --ty:-14px;animation-delay:0.66s;}
.bspark:nth-child(20){top:60%;left:93%;--tx:18px; --ty:10px; animation-delay:0.68s;}
.bspark:nth-child(21){top:82%;left:70%;--tx:10px; --ty:18px; animation-delay:0.65s;}
.bspark:nth-child(22){top:80%;left:10%;--tx:-16px;--ty:14px; animation-delay:0.67s;}
.bspark:nth-child(23){top:40%;left:2%; --tx:-20px;--ty:4px;  animation-delay:0.69s;}
.bspark:nth-child(24){top:35%;left:93%;--tx:20px; --ty:-4px; animation-delay:0.66s;}
.bspark:nth-child(25){top:3%; left:30%;--tx:-6px; --ty:-24px;animation-delay:0.72s;}
.bspark:nth-child(26){top:3%; left:65%;--tx:8px;  --ty:-24px;animation-delay:0.70s;}
.bspark:nth-child(27){top:50%;left:98%;--tx:24px; --ty:2px;  animation-delay:0.73s;}
.bspark:nth-child(28){top:50%;left:0%; --tx:-24px;--ty:2px;  animation-delay:0.71s;}
@keyframes bsparkFly{0%,47%{opacity:0;transform:translate(0,0) scale(0);}50%{opacity:1;transform:translate(0,0) scale(1);}75%{opacity:0.5;}95%{opacity:0;transform:translate(var(--tx),var(--ty)) scale(0.2);}100%{opacity:0;transform:translate(var(--tx),var(--ty)) scale(0);}}
.brand{font-family:'Inter',sans-serif;font-size:34px;font-weight:700;letter-spacing:-0.5px;color:#ffffff;line-height:1;text-shadow:0 1px 10px rgba(0,0,0,.35);}
.brand span{color:var(--accent);font-weight:800;}
.nav-links{display:flex;gap:8px;margin-left:auto;flex-wrap:wrap;}
.nav-links a{color:rgba(255,255,255,0.95);text-decoration:none;font-size:13px;padding:6px 14px;border-radius:6px;border:1px solid transparent;transition:all .2s;}
.nav-links a:hover{color:var(--text);border-color:var(--border);}
.nav-links a.active{color:var(--accent);border-color:var(--accent);background:rgba(225,188,101,.10);}
/* player graph expand */
.leg-row{cursor:pointer;transition:background .15s;}
.leg-row:hover{background:rgba(200,255,0,.04);}
.leg-graph-row{display:none;}
.leg-graph-row.open{display:table-row;}
.leg-graph-cell{padding:12px 16px 16px;background:#0d1117;border-bottom:1px solid var(--border);}
.graph-wrap{display:flex;gap:16px;align-items:flex-start;flex-wrap:wrap;}
.graph-stats{display:flex;gap:12px;flex-wrap:wrap;margin-bottom:10px;}
.gstat{background:#1a1f2e;border:1px solid var(--border);border-radius:6px;padding:6px 12px;min-width:80px;text-align:center;}
.gstat-label{font-size:10px;color:rgba(255,255,255,0.92);text-transform:uppercase;letter-spacing:.5px;}
.gstat-val{font-size:15px;font-weight:700;color:var(--accent);margin-top:2px;}
.graph-canvas-wrap,.leg-game-log-wrap{flex:1;min-width:260px;max-width:520px;}
table.leg-game-log{width:100%;font-size:13px;border-collapse:collapse;}
table.leg-game-log th{text-align:left;padding:6px 8px;border-bottom:1px solid rgba(255,255,255,0.14);color:rgba(255,255,255,0.65);font-size:11px;text-transform:uppercase;letter-spacing:0.06em;}
table.leg-game-log td{padding:6px 8px;border-bottom:1px solid rgba(255,255,255,0.06);}
table.leg-game-log tr:last-child td{border-bottom:none;}
.leg-game-hit{color:#00ff88;font-weight:600;}
.leg-game-miss{color:#c96a74;font-weight:600;}
.leg-game-log-empty{margin:0;font-size:12px;color:rgba(255,255,255,0.55);}
canvas.leg-chart{width:100%!important;height:140px!important;}

/* hero */
.hero{margin:28px 0 20px;display:flex;flex-wrap:wrap;gap:12px;align-items:flex-end;}
.hero h1{font-family:'Bebas Neue',sans-serif;font-size:clamp(32px,5vw,52px);letter-spacing:.08em;line-height:1;color:var(--accent);}
.hero h1 span{color:var(--cyan);}
.meta{color:var(--muted);font-size:12px;margin-top:4px;}

/* filter pill */
.filter-pill{background:rgba(14,18,34,.72);border:1px solid rgba(196,166,107,.20);border-radius:12px;padding:10px 16px;font-size:12px;color:rgba(255,255,255,0.92);margin-bottom:24px;backdrop-filter:blur(10px);}
.filter-pill strong{color:var(--cyan);}

/* slip card (group title + slip body unified) */
.ticket-group-band{display:flex;align-items:center;gap:12px;flex-wrap:wrap;padding-bottom:12px;margin-bottom:12px;border-bottom:1px solid rgba(255,255,255,.08);}
.group-title{font-family:'Bebas Neue',sans-serif;font-size:22px;letter-spacing:.08em;color:var(--accent);line-height:1.15;max-width:100%;overflow-wrap:anywhere;word-break:break-word;}
.group-meta{color:var(--muted);font-size:12px;}

/* ticket card */
.ticket{background:linear-gradient(160deg,rgba(24,30,52,.72) 0%,rgba(13,18,35,.68) 100%);border:1px solid rgba(196,166,107,.22);border-radius:14px;margin-bottom:24px;overflow:hidden;transition:transform .2s,box-shadow .2s;backdrop-filter:blur(10px);}
.ticket:hover{transform:translateY(-2px);box-shadow:0 10px 28px rgba(0,0,0,.35),0 0 0 1px rgba(196,166,107,.20) inset;}
.ticket-body{padding:16px 18px;}
.ticket-hdr{display:flex;align-items:center;gap:10px;margin-bottom:10px;flex-wrap:wrap;}
.ticket-no{font-family:'Bebas Neue',sans-serif;font-size:18px;letter-spacing:.08em;color:var(--text);}
.kpi-row{display:flex;gap:clamp(20px,4.5vw,44px);row-gap:14px;flex-wrap:wrap;margin-bottom:14px;justify-content:flex-start;}
.kpi{display:flex;flex-direction:column;gap:6px;min-width:4.75rem;}
.kpi-label{font-size:12px;color:var(--muted);letter-spacing:.1em;text-transform:uppercase;}
.kpi-val{font-family:'Bebas Neue',sans-serif;font-size:clamp(22px,2.4vw,28px);letter-spacing:.05em;line-height:1.1;}

/* table */
table{width:100%;border-collapse:collapse;}
th{background:rgba(225,188,101,.10);color:var(--accent);font-family:'Bebas Neue',sans-serif;font-size:clamp(14px,1.15vw,16px);letter-spacing:.08em;padding:10px 12px;text-align:left;border-bottom:1px solid rgba(196,166,107,.28);}
td{padding:10px 12px;border-bottom:1px solid rgba(255,255,255,.06);font-size:clamp(13px,1.05vw,15px);vertical-align:middle;}
tr:last-child td{border-bottom:none;}
tr:hover td{background:rgba(225,188,101,.06);}

html[data-theme="light"] body{
  background:
    radial-gradient(1200px 760px at -12% -22%, rgba(213,225,255,.76) 0%, transparent 56%),
    radial-gradient(980px 640px at 108% -8%, rgba(255,227,190,.72) 0%, transparent 54%),
    linear-gradient(180deg,#fcfdff 0%,#f6f8ff 45%,#f8f2e8 100%);
  color:#1f2430;
}
html[data-theme="light"] body::before{
  background:
    radial-gradient(1200px 760px at -12% -22%, rgba(213,225,255,.76) 0%, transparent 56%),
    radial-gradient(980px 640px at 108% -8%, rgba(255,227,190,.72) 0%, transparent 54%);
}
html[data-theme="light"] body::after{display:none;}
html[data-theme="light"] nav{
  background:rgba(255,255,255,.84);
  border-bottom:1px solid rgba(196,166,107,.24);
}
html[data-theme="light"] .filter-pill,
html[data-theme="light"] .ticket{
  background:rgba(255,255,255,.74);
  border:1px solid rgba(196,166,107,.22);
}

/* player cell */
.pwrap{display:flex;gap:8px;align-items:center;}
.avatar{width:34px;height:34px;border-radius:50%;overflow:hidden;border:1px solid var(--border);flex-shrink:0;background:#1a1a2e;display:flex;align-items:center;justify-content:center;font-size:12px;font-weight:700;color:var(--accent);}
.avatar img{width:100%;height:100%;object-fit:cover;}

/* dir badges */
.dir-over{background:rgba(0,242,255,.15);color:#00F2FF;padding:3px 10px;border-radius:4px;font-size:13px;font-weight:700;}
.dir-under{background:rgba(240,165,0,.15);color:#f0a500;padding:3px 10px;border-radius:4px;font-size:13px;font-weight:700;}
.delta-badge{font-family:'Inter',sans-serif;font-size:10px;padding:2px 6px;border-radius:6px;border:1px solid;margin-left:6px;vertical-align:middle;white-space:nowrap;}
.sig-strong{background:rgba(0,242,255,.16);color:#00F2FF;border:1px solid rgba(0,242,255,.35);padding:3px 8px;border-radius:5px;font-size:11px;font-weight:700;display:inline-block;}
.sig-lean{background:rgba(240,165,0,.16);color:#f0a500;border:1px solid rgba(240,165,0,.35);padding:3px 8px;border-radius:5px;font-size:11px;font-weight:700;display:inline-block;}
.sig-risk{background:rgba(201,106,116,.16);color:#c96a74;border:1px solid rgba(201,106,116,.35);padding:3px 8px;border-radius:5px;font-size:11px;font-weight:700;display:inline-block;}
.why-note{color:#bfc5d4;font-size:11px;line-height:1.25;}

.ca-border{position:relative;isolation:isolate;}
.ca-border::before{
  content:"";position:absolute;inset:0;padding:1px;border-radius:inherit;pointer-events:none;
  background:linear-gradient(120deg, rgba(212,175,55,.35), rgba(0,242,255,.20), rgba(212,175,55,.35));
  -webkit-mask:linear-gradient(#fff 0 0) content-box, linear-gradient(#fff 0 0);
  -webkit-mask-composite:xor;mask-composite:exclude;opacity:.45;
}
.mouse-glow{position:relative;overflow:hidden;}
.mouse-glow::after{
  content:"";position:absolute;left:var(--mx,50%);top:var(--my,50%);width:340px;height:340px;pointer-events:none;
  transform:translate(-50%,-50%);background:radial-gradient(circle, rgba(212,175,55,.18) 0%, rgba(0,242,255,.08) 28%, transparent 65%);
  opacity:0;transition:opacity .25s;
}
.mouse-glow:hover::after{opacity:1;}

/* responsive */
@media(max-width:640px){
  .kpi-row{gap:18px 22px;row-gap:12px;}
  .kpi-label{font-size:11px;}
  .kpi-val{font-size:clamp(20px,5.2vw,24px);}
  th{padding:8px 8px;font-size:clamp(12px,3.4vw,14px);}
  td{padding:8px 8px;font-size:clamp(12px,3.2vw,14px);}
  .avatar{width:30px;height:30px;font-size:11px;}
  .dir-over,.dir-under{font-size:12px;padding:2px 8px;}
}
"""

    html_parts = []
    html_parts.append(f"""<!doctype html>
<html lang="en">
<head>
<meta charset="utf-8"/>
<meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>PropOracle — Tickets · {date_pretty}</title>
<style>{CSS}</style>
</head>
<body>
<div id="app">

<nav>
  <a class="nav-logo" href="/">
    <div class="brain-wrap">
      <div class="brain-slate"></div>
      <div class="brain-pulse-ring"></div>
      <div class="brain-pulse-ring"></div>
      <div class="brain-pulse-ring"></div>
      <div class="brain-pulse-ring"></div>
      <div class="bspark lg"></div><div class="bspark md"></div>
      <div class="bspark sm"></div><div class="bspark lg"></div>
      <div class="bspark cy"></div><div class="bspark md"></div>
      <div class="bspark sm"></div><div class="bspark lg"></div>
      <div class="bspark cy"></div><div class="bspark md"></div>
      <div class="bspark lg"></div><div class="bspark sm"></div>
      <div class="bspark cy"></div><div class="bspark sm"></div>
      <div class="bspark md"></div><div class="bspark cy"></div>
      <div class="bspark sm"></div><div class="bspark md"></div>
      <div class="bspark lg"></div><div class="bspark wh"></div>
      <svg class="brain-svg" viewBox="0 0 50 50" fill="none" xmlns="http://www.w3.org/2000/svg">
        <defs>
          <linearGradient id="lgL" x1="6" y1="6" x2="25" y2="44" gradientUnits="userSpaceOnUse">
            <stop offset="0%" stop-color="#c8ff00" stop-opacity="0.35"/>
            <stop offset="60%" stop-color="#c8ff00" stop-opacity="0.12"/>
            <stop offset="100%" stop-color="#c8ff00" stop-opacity="0.06"/>
          </linearGradient>
          <linearGradient id="lgR" x1="44" y1="6" x2="25" y2="44" gradientUnits="userSpaceOnUse">
            <stop offset="0%" stop-color="#00e5ff" stop-opacity="0.35"/>
            <stop offset="60%" stop-color="#00e5ff" stop-opacity="0.12"/>
            <stop offset="100%" stop-color="#00e5ff" stop-opacity="0.06"/>
          </linearGradient>
          <filter id="nglow"><feGaussianBlur stdDeviation="0.8" result="blur"/><feMerge><feMergeNode in="blur"/><feMergeNode in="SourceGraphic"/></feMerge></filter>
        </defs>
        <path d="M25 7 C22 7 18 8 15 10 C12 12 10 15 9 18 C8 21 8.5 24 9 26 C7.5 27.5 7 30 7.5 32.5 C8 35 10 37.5 13 39 C15 40 17 40 19 39.5 C20.5 39 22 38 23 37 L23 9 C23.5 8 24 7.5 25 7Z" fill="url(#lgL)" stroke="#c8ff00" stroke-width="0.9"/>
        <path d="M15 10 C13 11 11 13 11 15 C11 17 12.5 18.5 14 18" stroke="#c8ff00" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.6"/>
        <path d="M9 19 C10.5 18 12 19 13.5 18"    stroke="#c8ff00" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M8.5 23 C10 22 12 23 13.5 22"    stroke="#c8ff00" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M8 27 C9.5 26 11.5 27 13 26.5"   stroke="#c8ff00" stroke-width="0.7"  stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M8.5 31 C10 30.5 12 31 13.5 30.5" stroke="#c8ff00" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M10 35 C11.5 34.5 13.5 35 15 34.5" stroke="#c8ff00" stroke-width="0.65" stroke-linecap="round" fill="none" opacity="0.55"/>
        <path d="M16 14 C17 13.5 18.5 14 19.5 13.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M15.5 20 C17 19.5 18.5 20 20 19.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M15 27 C16.5 26.5 18 27 19.5 26.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M15 33 C16.5 32.5 18.5 33 20 32.5"  stroke="#c8ff00" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M25 7 C28 7 32 8 35 10 C38 12 40 15 41 18 C42 21 41.5 24 41 26 C42.5 27.5 43 30 42.5 32.5 C42 35 40 37.5 37 39 C35 40 33 40 31 39.5 C29.5 39 28 38 27 37 L27 9 C26.5 8 26 7.5 25 7Z" fill="url(#lgR)" stroke="#00e5ff" stroke-width="0.9"/>
        <path d="M35 10 C37 11 39 13 39 15 C39 17 37.5 18.5 36 18" stroke="#00e5ff" stroke-width="0.7" stroke-linecap="round" fill="none" opacity="0.6"/>
        <path d="M41 19 C39.5 18 38 19 36.5 18"      stroke="#00e5ff" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M41.5 23 C40 22 38 23 36.5 22"      stroke="#00e5ff" stroke-width="0.75" stroke-linecap="round" fill="none" opacity="0.8"/>
        <path d="M42 27 C40.5 26 38.5 27 37 26.5"    stroke="#00e5ff" stroke-width="0.7"  stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M41.5 31 C40 30.5 38 31 36.5 30.5"  stroke="#00e5ff" stroke-width="0.7"  stroke-linecap="round" fill="none" opacity="0.7"/>
        <path d="M40 35 C38.5 34.5 36.5 35 35 34.5"  stroke="#00e5ff" stroke-width="0.65" stroke-linecap="round" fill="none" opacity="0.55"/>
        <path d="M34 14 C33 13.5 31.5 14 30.5 13.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M34.5 20 C33 19.5 31.5 20 30 19.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M35 27 C33.5 26.5 32 27 30.5 26.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <path d="M35 33 C33.5 32.5 31.5 33 30 32.5"  stroke="#00e5ff" stroke-width="0.6" stroke-linecap="round" fill="none" opacity="0.5"/>
        <line x1="25" y1="8" x2="25" y2="38" stroke="#ffffff22" stroke-width="0.6" stroke-dasharray="2.5,2"/>
        <circle cx="13" cy="16" r="1.4" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="1;0.15;1" dur="1.7s" repeatCount="indefinite"/><animate attributeName="r" values="1.4;0.9;1.4" dur="1.7s" repeatCount="indefinite"/></circle>
        <circle cx="11" cy="22" r="1.2" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.8;0.1;0.8" dur="2.2s" repeatCount="indefinite" begin="0.3s"/></circle>
        <circle cx="12" cy="28.5" r="1.3" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.9;0.2;0.9" dur="1.9s" repeatCount="indefinite" begin="0.6s"/></circle>
        <circle cx="15" cy="34.5" r="1.1" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.1;0.7" dur="2.4s" repeatCount="indefinite" begin="0.9s"/></circle>
        <circle cx="19" cy="18" r="1.0" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.6;0.1;0.6" dur="2.0s" repeatCount="indefinite" begin="1.1s"/></circle>
        <circle cx="18" cy="30" r="1.0" fill="#c8ff00" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.15;0.7" dur="1.6s" repeatCount="indefinite" begin="0.5s"/></circle>
        <circle cx="37" cy="16" r="1.4" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="1;0.15;1" dur="2.0s" repeatCount="indefinite" begin="0.2s"/><animate attributeName="r" values="1.4;0.9;1.4" dur="2.0s" repeatCount="indefinite" begin="0.2s"/></circle>
        <circle cx="39" cy="22" r="1.2" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.8;0.1;0.8" dur="1.8s" repeatCount="indefinite" begin="0.5s"/></circle>
        <circle cx="38" cy="28.5" r="1.3" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.9;0.2;0.9" dur="2.3s" repeatCount="indefinite" begin="0.8s"/></circle>
        <circle cx="35" cy="34.5" r="1.1" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.1;0.7" dur="1.7s" repeatCount="indefinite" begin="1.0s"/></circle>
        <circle cx="31" cy="18" r="1.0" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.6;0.1;0.6" dur="2.1s" repeatCount="indefinite" begin="1.2s"/></circle>
        <circle cx="32" cy="30" r="1.0" fill="#00e5ff" filter="url(#nglow)"><animate attributeName="opacity" values="0.7;0.15;0.7" dur="1.5s" repeatCount="indefinite" begin="0.4s"/></circle>
        <line x1="13" y1="16" x2="37" y2="16" stroke="#c8ff0030" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.9;0.2" dur="1.7s" repeatCount="indefinite"/></line>
        <line x1="11" y1="22" x2="39" y2="22" stroke="#00e5ff30" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.9;0.2" dur="2.2s" repeatCount="indefinite" begin="0.4s"/></line>
        <line x1="12" y1="28.5" x2="38" y2="28.5" stroke="#c8ff0030" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.8;0.2" dur="1.9s" repeatCount="indefinite" begin="0.7s"/></line>
        <line x1="15" y1="34.5" x2="35" y2="34.5" stroke="#00e5ff30" stroke-width="0.6"><animate attributeName="opacity" values="0.2;0.8;0.2" dur="2.4s" repeatCount="indefinite" begin="1.0s"/></line>
        <line x1="13" y1="16" x2="39" y2="22" stroke="#c8ff0018" stroke-width="0.5"><animate attributeName="opacity" values="0;0.6;0" dur="2.5s" repeatCount="indefinite" begin="0.3s"/></line>
        <line x1="11" y1="22" x2="38" y2="28.5" stroke="#00e5ff18" stroke-width="0.5"><animate attributeName="opacity" values="0;0.6;0" dur="2.1s" repeatCount="indefinite" begin="0.8s"/></line>
        <line x1="19" y1="18" x2="31" y2="18" stroke="#ffffff18" stroke-width="0.5"><animate attributeName="opacity" values="0;0.7;0" dur="1.6s" repeatCount="indefinite" begin="1.1s"/></line>
        <line x1="18" y1="30" x2="32" y2="30" stroke="#ffffff18" stroke-width="0.5"><animate attributeName="opacity" values="0;0.7;0" dur="2.0s" repeatCount="indefinite" begin="0.5s"/></line>
        <path d="M22 38 C22 40.5 23 43 25 43 C27 43 28 40.5 28 38" stroke="#c8ff0066" stroke-width="0.9" fill="none" stroke-linecap="round"/>
        <line x1="25" y1="38" x2="25" y2="43" stroke="#00e5ff55" stroke-width="0.7" stroke-dasharray="1.5,1.5"/>
      </svg>
    </div>
    <div><div class="brand">Prop<span>ORACLE</span></div></div>
  </a>
  <div class="nav-links">
    <a href="/">Home</a>
    <a href="/tickets" class="active">Tickets</a>
    <a href="/grades">Grades</a>
    <a href="/payout">Payouts</a>
  </div>
</nav>

<div class="hero">
  <div>
    <h1>🎟 Latest <span>Tickets</span></h1>
    <div class="meta">Generated: {gen_at} &nbsp;|&nbsp; Slate date: <strong>{date_pretty}</strong> <span style="opacity:.72">({date_eff})</span>{date_mismatch_html}</div>
  </div>
  """)



    if not payload.get("groups"):
        html_parts.append("""
<div class="filter-pill" style="margin-top:-12px;border-color:rgba(201,106,116,.35);color:#d4b5b8;">
  <strong>No tickets in this JSON.</strong> Run the combined slate ticket script with <code>--write-web</code> after building slates, or relax filters. Download JSON (link above) to confirm <code>groups</code> is non-empty.
</div>
""")

    for g in payload.get("groups", []):
        gname = g.get("group_name", "Group")
        accent = _sport_accent(_group_sport(gname))
        group_meta = (
            f"Legs: {g.get('n_legs', '')} &nbsp;|&nbsp; Power: {g.get('power_payout', '')}x "
            f"&nbsp;|&nbsp; Flex: {g.get('flex_payout', '')}x"
        )
        for t in g.get("tickets", []):
            avg_hr = t.get("avg_hit_rate")
            avg_rs = t.get("avg_rank_score")
            wp     = t.get("est_win_prob")

            try:
                hr_disp = f"{float(avg_hr)*100:.1f}%"
                hr_col  = hit_color(avg_hr)
            except Exception:
                hr_disp, hr_col = "—", "#aaa"

            try:
                rs_disp = f"{float(avg_rs):.2f}"
            except Exception:
                rs_disp = "—"

            em = t.get("est_multiplier")
            fm = t.get("flat_multiplier")
            mult_kpi = ""
            if em is not None and fm is not None:
                try:
                    mult_kpi = f"""
          <div class="kpi">
            <div class="kpi-label">Curve est</div>
            <div class="kpi-val" style="color:var(--accent);">{float(em):.2f}x</div>
            <div class="kpi-label" style="margin-top:4px">Flat PP base</div>
            <div style="font-family:'Bebas Neue',sans-serif;font-size:15px;color:var(--muted);">{float(fm):.2f}x</div>
          </div>"""
                except (TypeError, ValueError):
                    mult_kpi = ""

            html_parts.append(f"""
<div class="ticket" style="border-left:4px solid {accent};">
  <div class="ticket-body">
    <div class="ticket-group-band">
      <div class="group-title" style="color:{accent};">{gname}</div>
      <div class="group-meta">{group_meta}</div>
    </div>
        <div class="ticket-hdr">
          <div class="ticket-no">Ticket #{t.get('ticket_no','')}</div>
        </div>
        <div class="kpi-row">
          <div class="kpi">
            <div class="kpi-label">Hit Rate</div>
            <div class="kpi-val" style="color:{hr_col};">{hr_disp}</div>
          </div>
          <div class="kpi">
            <div class="kpi-label">Avg Rank</div>
            <div class="kpi-val" style="color:var(--cyan);">{rs_disp}</div>
          </div>
          <div class="kpi" style="flex:1;min-width:140px;">
            <div class="kpi-label">Win Prob</div>
            {wp_bar(wp)}
          </div>{mult_kpi}
        </div>
        <table>
          <thead><tr>
            <th>#</th><th>Sport</th><th>Player</th><th>Prop</th><th>Line</th>
            <th>Pick</th><th>Min</th><th>Shot</th><th>Usage</th>
            <th>Dir</th><th>Signal</th><th>Why</th><th>Hit%</th><th>Edge</th><th>Rank</th>
          </tr></thead>
          <tbody>
""")
            for i, leg in enumerate(t.get("legs", []), start=1):
                dirv = (leg.get("direction") or "").upper()
                dir_span = (
                    "<span class='dir-over'>OVER</span>"
                    if dirv == "OVER"
                    else f"<span class='dir-under'>{dirv or '—'}</span>"
                )
                img      = leg.get("image_url")
                initials = leg.get("initials") or "?"
                if img:
                    avatar = f"<div class='avatar'><img src='{img}' alt='{initials}' onerror=\"this.style.display='none'\"></div>"
                else:
                    avatar = f"<div class='avatar'>{initials}</div>"

                player_cell = f"<div class='pwrap'>{avatar}<div>{leg.get('player','')}</div></div>"

                hr_val = leg.get("hit_rate")
                hr_fmt = fmt_pct(hr_val) if hr_val is not None else "—"
                hr_c   = hit_color(hr_val) if hr_val is not None else "#aaa"

                min_tier  = badge(leg.get("min_tier") or leg.get("minutes_tier"), "#39ff6e")
                shot_role = badge(leg.get("shot_role"), "#00e5ff")
                usg_role  = badge(leg.get("usage_role"), "#888")

                # build graph data for expand panel
                l5_avg     = leg.get("l5_avg")
                season_avg = leg.get("season_avg")
                l5_over    = leg.get("l5_over")
                l5_under   = leg.get("l5_under")
                l10_over   = leg.get("l10_over")
                l10_under  = leg.get("l10_under")
                line_val   = leg.get("line")
                dir_txt    = str(leg.get("direction") or "").upper()
                row_id     = f"lgr-{id(leg)}-{i}"
                sig_html, sig_reason = direction_signal(leg)

                dp = leg.get("delta_pct")
                ptl = str(leg.get("pick_type") or "").lower()
                delta_badge = ""
                if dp is not None:
                    try:
                        dpf = float(dp)
                        bc = "#888888"
                        if "goblin" in ptl:
                            if dpf >= 0.9:
                                bc = "#7dcf9a"
                            elif dpf >= 0.7:
                                bc = "#c89a4a"
                            else:
                                bc = "#e67e22"
                        elif "demon" in ptl:
                            if dpf <= 1.15:
                                bc = "#e8a0a0"
                            else:
                                bc = "#ff3333"
                        delta_badge = (
                            f" <span class='delta-badge' style='border-color:{bc};color:{bc}' "
                            f"title='played ÷ standard'>{dpf * 100:.1f}%</span>"
                        )
                    except (TypeError, ValueError):
                        pass
                pick_cell = f"{leg.get('pick_type', '')}{delta_badge}"

                # stat pills
                def _pill(label, val, fmt=None):
                    if val is None: return ""
                    v = fmt(val) if fmt else str(val)
                    return f'<div class="gstat"><div class="gstat-label">{label}</div><div class="gstat-val">{v}</div></div>'

                pills = "".join([
                    _pill("L5 Avg",     l5_avg,     lambda x: f"{x:.1f}"),
                    _pill("Season Avg", season_avg, lambda x: f"{x:.1f}"),
                    _pill("L5 Over",    l5_over,    lambda x: format_hit_window_fraction(5, x)),
                    _pill("L5 Under",   l5_under,   lambda x: format_hit_window_fraction(5, x)),
                    _pill("L10 Over",   l10_over,   lambda x: format_hit_window_fraction(10, x)),
                    _pill("L10 Under",  l10_under,  lambda x: format_hit_window_fraction(10, x)),
                    _pill("Hit Rate",   hr_val,     lambda x: f"{x*100:.0f}%"),
                ])

                game_log_html = _tickets_leg_game_log_table_html(leg)
                graph_row = f"""
<tr class="leg-graph-row" id="{row_id}">
  <td class="leg-graph-cell" colspan="15">
    <div class="graph-wrap">
      <div style="flex:1;min-width:200px;">
        <div style="font-size:11px;color:rgba(255,255,255,0.92);margin-bottom:6px;text-transform:uppercase;letter-spacing:.5px;">{leg.get('player','')} · {leg.get('prop_type','')} · Line {fmt_line(line_val)}</div>
        <div class="graph-stats">{pills}</div>
      </div>
      <div class="leg-game-log-wrap">
        {game_log_html}
      </div>
    </div>
  </td>
</tr>"""

                html_parts.append(
                    f"<tr class='leg-row' onclick=\"var r=document.getElementById('{row_id}');r.classList.toggle('open');\">"
                    f"<td>{i}</td>"
                    f"<td>{sport_badge(leg.get('sport',''))}</td>"
                    f"<td>{player_cell}</td>"
                    f"<td>{leg.get('prop_type','')}</td>"
                    f"<td style='color:var(--text);'>{fmt_line(leg.get('line'))}</td>"
                    f"<td>{pick_cell}</td>"
                    f"<td>{min_tier}</td>"
                    f"<td>{shot_role}</td>"
                    f"<td>{usg_role}</td>"
                    f"<td>{dir_span}</td>"
                    f"<td>{sig_html}</td>"
                    f"<td class='why-note'>{sig_reason}</td>"
                    f"<td style='color:{hr_c};font-weight:600;'>{hr_fmt}</td>"
                    f"<td>{fmt_2(leg.get('edge')) if leg.get('edge') is not None else '—'}</td>"
                    f"<td>{fmt_2(leg.get('rank_score')) if leg.get('rank_score') is not None else '—'}</td>"
                    f"</tr>"
                )
                html_parts.append(graph_row)

            html_parts.append("""
          </tbody>
        </table>
    </div>
</div>
""")

    html_parts.append("""
</div><!-- #app -->
<script>
(() => {
  document.querySelectorAll('.ticket,.filter-pill,.kpi').forEach(el => {
    el.classList.add('ca-border');
    el.classList.add('mouse-glow');
    el.addEventListener('mousemove', e => {
      const r = el.getBoundingClientRect();
      el.style.setProperty('--mx', (e.clientX - r.left) + 'px');
      el.style.setProperty('--my', (e.clientY - r.top) + 'px');
    });
  });
})();
</script>
</body>
</html>""")

    return "\n".join(html_parts)


def _safe_group_key(group: dict[str, Any]) -> tuple[str, int, float, float]:
    return (
        str(group.get("group_name") or ""),
        int(group.get("n_legs") or 0),
        float(group.get("power_payout") or 0.0),
        float(group.get("flex_payout") or 0.0),
    )


def _safe_leg_sig(leg: dict[str, Any]) -> tuple:
    def _s(v: Any) -> str:
        return str(v or "").strip().lower()

    try:
        line_v = round(float(leg.get("line")), 4)
    except Exception:
        line_v = _s(leg.get("line"))
    return (
        _s(leg.get("sport")),
        _s(leg.get("player")),
        _s(leg.get("team")),
        _s(leg.get("opp")),
        _s(leg.get("prop_type")),
        _s(leg.get("pick_type")),
        _s(leg.get("direction")),
        line_v,
    )


def _safe_ticket_sig(ticket: dict[str, Any]) -> tuple:
    legs = ticket.get("legs") or []
    if isinstance(legs, list) and legs:
        leg_sigs = sorted(_safe_leg_sig(leg) for leg in legs if isinstance(leg, dict))
        if leg_sigs:
            return tuple(leg_sigs)
    # Fallback when legs are malformed/missing.
    return (json.dumps(ticket, sort_keys=True, ensure_ascii=False, default=str),)


def merge_web_payloads_by_group(new_payload: dict[str, Any], old_payload: dict[str, Any]) -> dict[str, Any]:
    """Merge new + existing tickets_latest payload for the same date, de-duping by leg set."""
    merged = dict(new_payload)
    new_groups = list(new_payload.get("groups") or [])
    old_groups = list(old_payload.get("groups") or [])
    grouped: dict[tuple[str, int, float, float], dict[str, Any]] = {}
    order: list[tuple[str, int, float, float]] = []

    # Seed with new groups (so latest run remains primary ordering).
    for g in new_groups:
        if not isinstance(g, dict):
            continue
        k = _safe_group_key(g)
        grouped[k] = {**g, "tickets": list(g.get("tickets") or [])}
        order.append(k)

    # Merge in old groups/tickets.
    for g in old_groups:
        if not isinstance(g, dict):
            continue
        k = _safe_group_key(g)
        if k not in grouped:
            grouped[k] = {**g, "tickets": []}
            order.append(k)
        grouped[k]["tickets"].extend(list(g.get("tickets") or []))

    out_groups: list[dict[str, Any]] = []
    for k in order:
        g = grouped[k]
        seen: set[tuple] = set()
        deduped_tickets: list[dict[str, Any]] = []
        for t in list(g.get("tickets") or []):
            if not isinstance(t, dict):
                continue
            sig = _safe_ticket_sig(t)
            if sig in seen:
                continue
            seen.add(sig)
            deduped_tickets.append(t)
        for i, t in enumerate(deduped_tickets, start=1):
            t["ticket_no"] = i
        g["tickets"] = deduped_tickets
        out_groups.append(g)

    merged["groups"] = out_groups
    merged["generated_at"] = datetime.now(timezone.utc).strftime("%Y-%m-%d %H:%M:%S UTC")
    return merged


def write_web_outputs(
    payload,
    outdir: str,
    *,
    require_positive_ev: bool = True,
    merge_existing_for_date: bool = False,
    apply_template_cap: bool = False,
    discard_tracker: DiscardTracker | None = None,
    json_filename: str = "tickets_latest.json",
    skip_ui_filters: bool = False,
):
    """Write tickets_latest.json for /tickets; graded HTML is build_ticket_eval.py → ticket_eval_<date>.html."""
    os.makedirs(outdir, exist_ok=True)
    json_path = os.path.join(outdir, json_filename)
    if skip_ui_filters:
        apply_slate_ev_tier_recommendations(payload)
        payload = _sanitize_for_json(payload)
        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False, allow_nan=False)
        print(f"[OK] Web JSON  -> {json_path}")
        return
    apply_slate_ev_tier_recommendations(payload)
    payload = filter_web_tickets_for_ui(
        payload,
        require_positive_ev=require_positive_ev,
        apply_template_cap=apply_template_cap,
        discard_tracker=discard_tracker,
    )
    if not require_positive_ev:
        print("  [web] EV gate OFF — JSON includes top slips per sport/leg-count (workbook pool, template caps)")
    if merge_existing_for_date and os.path.isfile(json_path):
        try:
            with open(json_path, "r", encoding="utf-8") as f:
                existing = json.load(f)
            if isinstance(existing, dict) and str(existing.get("date")) == str(payload.get("date")):
                old_count = sum(len(g.get("tickets") or []) for g in (existing.get("groups") or []))
                new_count = sum(len(g.get("tickets") or []) for g in (payload.get("groups") or []))
                payload = merge_web_payloads_by_group(payload, existing)
                merged_count = sum(len(g.get("tickets") or []) for g in (payload.get("groups") or []))
                print(f"  [web-merge] same-date merge enabled: new={new_count}, existing={old_count}, merged={merged_count}")
            else:
                print("  [web-merge] skipped: existing tickets_latest.json has a different slate date")
        except Exception as exc:
            print(f"  [web-merge] WARN: could not merge existing tickets_latest.json ({exc})")
    payload = _sanitize_for_json(payload)
    with open(json_path, "w", encoding="utf-8") as f:
        json.dump(payload, f, indent=2, ensure_ascii=False, allow_nan=False)
    print(f"[OK] Web JSON  -> {json_path}")
    # Keep docs JSON in sync for static/GitHub Pages views that read ui_runner/docs.
    try:
        outdir_p = Path(outdir).resolve()
        docs_json = outdir_p.parent / "docs" / "tickets_latest.json"
        docs_json.parent.mkdir(parents=True, exist_ok=True)
        with open(docs_json, "w", encoding="utf-8") as f:
            json.dump(payload, f, indent=2, ensure_ascii=False, allow_nan=False)
        print(f"[OK] Docs JSON -> {docs_json}")
    except Exception as exc:
        print(f"[WARN] Docs JSON sync skipped: {exc}")
    print("  (Graded eval HTML) Run: py -3.14 scripts/build_ticket_eval.py --date <YYYY-MM-DD>")


def generate_payout_ladder_examples(payload: dict, out_path: str) -> None:
    """
    Build composition-oriented ticket examples for payout ladder collection.
    """
    all_legs: list[dict] = []
    for grp in (payload or {}).get("groups") or []:
        for t in (grp or {}).get("tickets") or []:
            for leg in (t or {}).get("legs") or []:
                if isinstance(leg, dict):
                    all_legs.append(leg)
    if not all_legs:
        os.makedirs(os.path.dirname(out_path), exist_ok=True)
        with open(out_path, "w", encoding="utf-8") as f:
            json.dump(
                {"generated_at": datetime.now(timezone.utc).isoformat(), "examples": []},
                f,
                ensure_ascii=False,
                indent=2,
                allow_nan=False,
            )
        return

    def _num_or_nan(raw: Any) -> float:
        try:
            return float(raw)
        except (TypeError, ValueError):
            return float("nan")

    std_pool: list[tuple[dict, float]] = []
    gob_pool: list[tuple[dict, float]] = []
    dem_pool: list[tuple[dict, float]] = []
    for leg in all_legs:
        pt = str(leg.get("pick_type") or "Standard").strip().title()
        line = _num_or_nan(leg.get("line"))
        std_line = _num_or_nan(leg.get("standard_line"))
        if not math.isfinite(std_line):
            std_line = line
        delta = abs(std_line - line) if (math.isfinite(std_line) and math.isfinite(line)) else 0.0
        if pt == "Goblin" and 0.5 <= delta <= 3.0:
            gob_pool.append((leg, delta))
        elif pt == "Demon" or delta > 3.0:
            dem_pool.append((leg, delta))
        else:
            std_pool.append((leg, delta))

    seed_std = std_pool[0] if std_pool else (all_legs[0], 0.0)
    seed_gob = gob_pool[0] if gob_pool else seed_std
    seed_dem = dem_pool[0] if dem_pool else seed_std
    delta_targets = [0.5, 1.0, 1.5, 2.0, 2.5, 3.0]
    examples: list[dict] = []

    def _leg(seed: tuple[dict, float], pick_type: str, delta_override: float | None = None) -> dict:
        row, dflt = seed
        delta_val = float(dflt if delta_override is None else delta_override)
        return {
            "player": str(row.get("player") or ""),
            "prop_type": str(row.get("prop_type") or row.get("prop") or ""),
            "line": row.get("line"),
            "standard_line": row.get("standard_line"),
            "pick_type": pick_type,
            "delta": round(delta_val, 4),
            "direction": str(row.get("direction") or "OVER").upper(),
            "hit_rate": row.get("hit_rate"),
        }

    def _pick_unique_leg(
        pool: list[tuple[dict, float]],
        used_players: set[str],
        fallback: tuple[dict, float],
        pick_type: str,
        delta_override: float | None = None,
    ) -> dict | None:
        for cand in pool:
            player = str((cand[0] or {}).get("player") or "").strip().lower()
            if player and player in used_players:
                continue
            out = _leg(cand, pick_type, delta_override)
            p_out = str(out.get("player") or "").strip().lower()
            if p_out:
                used_players.add(p_out)
            return out
        # fallback only if it doesn't violate uniqueness
        fb_player = str((fallback[0] or {}).get("player") or "").strip().lower()
        if fb_player and fb_player in used_players:
            return None
        out = _leg(fallback, pick_type, delta_override)
        p_out = str(out.get("player") or "").strip().lower()
        if p_out:
            used_players.add(p_out)
        return out

    def _emit(n: int, g_count: int, d_count: int, g_delta: float | None):
        s_count = max(0, n - g_count - d_count)
        legs: list[dict] = []
        used_players: set[str] = set()
        for _ in range(s_count):
            leg = _pick_unique_leg(std_pool, used_players, seed_std, "Standard", 0.0)
            if not leg:
                return
            legs.append(leg)
        for _ in range(g_count):
            leg = _pick_unique_leg(gob_pool, used_players, seed_gob, "Goblin", g_delta)
            if not leg:
                return
            legs.append(leg)
        for _ in range(d_count):
            leg = _pick_unique_leg(dem_pool, used_players, seed_dem, "Demon")
            if not leg:
                return
            legs.append(leg)
        examples.append(
            {
                "composition": f"{s_count}S+{g_count}G+{d_count}D",
                "goblin_delta": (round(float(g_delta), 3) if g_delta is not None else None),
                "legs": legs,
            }
        )

    for n in range(2, 7):
        for dlt in delta_targets:                          # 1 goblin + standard
            _emit(n, g_count=1, d_count=0, g_delta=dlt)
        if n >= 3:
            _emit(n, g_count=2, d_count=0, g_delta=2.0)   # 2 goblin + standard
            _emit(n, g_count=1, d_count=1, g_delta=1.5)   # mixed goblin + demon
        _emit(n, g_count=0, d_count=1, g_delta=None)      # 1 demon + standard

    # Safety net: never publish examples with duplicate players in the same ticket.
    deduped_examples: list[dict] = []
    for ex in examples:
        legs = [l for l in (ex.get("legs") or []) if isinstance(l, dict)]
        players = [str((l.get("player") or "")).strip().lower() for l in legs if str((l.get("player") or "")).strip()]
        if len(players) != len(set(players)):
            continue
        deduped_examples.append(ex)
    examples = deduped_examples

    os.makedirs(os.path.dirname(out_path), exist_ok=True)
    payload_out = _sanitize_for_json(
        {"generated_at": datetime.now(timezone.utc).isoformat(), "examples": examples}
    )
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(payload_out, f, ensure_ascii=False, indent=2, allow_nan=False)
    print(f"[OK] Payout ladder examples -> {out_path}")


def _apply_l5_truth_from_stat_games(
    df: pd.DataFrame, sport_label: str, *, min_stat_games: int = 3
) -> pd.DataFrame:
    """
    Source-of-truth guardrail:
    when stat_g1..stat_g5 are present, derive L5 Over/Under, L5 Avg, and HIT%
    directly from those raw values so downstream UI cannot drift from game logs.
    """
    if df is None or df.empty:
        return df

    stat_cols = [c for c in [f"stat_g{i}" for i in range(1, 6)] if c in df.columns]
    if not stat_cols or "line" not in df.columns:
        return df

    vals = df[stat_cols].apply(pd.to_numeric, errors="coerce")
    line = pd.to_numeric(df["line"], errors="coerce")
    valid_n = vals.notna().sum(axis=1)
    use_mask = valid_n >= int(min_stat_games)
    if not bool(use_mask.any()):
        return df

    over = vals.gt(line, axis=0).sum(axis=1).astype(float)
    under = vals.lt(line, axis=0).sum(axis=1).astype(float)
    l5_avg = vals.mean(axis=1)
    total_ou = over + under

    direction = (
        df.get("direction", pd.Series(["OVER"] * len(df), index=df.index))
        .astype(str)
        .str.strip()
        .str.upper()
    )
    hit_over = over.divide(total_ou.where(total_ou > 0))
    hit_under = under.divide(total_ou.where(total_ou > 0))
    hit_dir = hit_over.where(direction.ne("UNDER"), hit_under)

    if "l5_over" not in df.columns:
        df["l5_over"] = np.nan
    if "l5_under" not in df.columns:
        df["l5_under"] = np.nan
    if "l5_avg" not in df.columns:
        df["l5_avg"] = np.nan
    if "hit_rate" not in df.columns:
        df["hit_rate"] = np.nan

    df.loc[use_mask, "l5_over"] = over[use_mask]
    df.loc[use_mask, "l5_under"] = under[use_mask]
    df.loc[use_mask, "l5_avg"] = l5_avg[use_mask]
    df.loc[use_mask, "hit_rate"] = hit_dir[use_mask]
    return df


def _apply_l10_truth_from_stat_games(
    df: pd.DataFrame, sport_label: str, *, min_stat_games: int = 6
) -> pd.DataFrame:
    """Derive L10 Over/Under and last-10 avg from stat_g1..stat_g10 when present."""
    if df is None or df.empty:
        return df

    stat_cols = [c for c in [f"stat_g{i}" for i in range(1, 11)] if c in df.columns]
    if not stat_cols or "line" not in df.columns:
        return df

    vals = df[stat_cols].apply(pd.to_numeric, errors="coerce")
    line = pd.to_numeric(df["line"], errors="coerce")
    valid_n = vals.notna().sum(axis=1)
    use_mask = valid_n >= int(min_stat_games)
    if not bool(use_mask.any()):
        return df

    over = vals.gt(line, axis=0).sum(axis=1).astype(float)
    under = vals.lt(line, axis=0).sum(axis=1).astype(float)
    l10_avg = vals.mean(axis=1)

    if "l10_over" not in df.columns:
        df["l10_over"] = np.nan
    if "l10_under" not in df.columns:
        df["l10_under"] = np.nan
    if "l10_games_played" not in df.columns:
        df["l10_games_played"] = valid_n.astype(float)

    df.loc[use_mask, "l10_over"] = over[use_mask]
    df.loc[use_mask, "l10_under"] = under[use_mask]
    df.loc[use_mask, "l10_games_played"] = valid_n[use_mask].astype(float)
    if "stat_last10_avg" in df.columns:
        df.loc[use_mask, "stat_last10_avg"] = l10_avg[use_mask]
    return df


def _ensure_stat_g_columns(df: pd.DataFrame) -> pd.DataFrame:
    """Mirror G1..G10 <-> stat_g1..stat_g10 from step8 clean exports."""
    if df is None or df.empty:
        return df
    for i in range(1, 11):
        gcol, scol = f"G{i}", f"stat_g{i}"
        if gcol in df.columns and scol not in df.columns:
            df[scol] = df[gcol]
        elif scol in df.columns and gcol not in df.columns:
            df[gcol] = df[scol]
    return df


def _fill_projection_from_avgs(df: pd.DataFrame) -> pd.DataFrame:
    """Use last-5 / season avg when model projection is missing (common on Goblin micro-lines)."""
    if df is None or df.empty or "projection" not in df.columns:
        return df
    proj = pd.to_numeric(df["projection"], errors="coerce")
    miss = proj.isna() | ~np.isfinite(proj)
    if not bool(miss.any()):
        return df
    l5a = pd.to_numeric(df.get("l5_avg", pd.Series(dtype=float)), errors="coerce")
    sa = pd.to_numeric(df.get("season_avg", pd.Series(dtype=float)), errors="coerce")
    fill = l5a.where(l5a.notna(), sa)
    df.loc[miss, "projection"] = fill[miss]
    return df


def _resolve_step1_pp_path(board_path: str, sport: str) -> Optional[Path]:
    """Locate step1 PP CSV beside the board or under outputs/<date>/<sport>/."""
    sport_u = str(sport or "").upper()
    sport_low = sport_u.lower()
    step1_names = {
        "NBA": ("step1_pp_props_today.csv",),
        "WNBA": ("step1_wnba_props.csv",),
    }.get(sport_u, ())
    if not step1_names:
        return None
    try:
        parent = Path(board_path).resolve().parent
    except Exception:
        parent = None
    out_root = Path(REPO_ROOT) / "outputs"
    if out_root.is_dir():
        dated_hits = sorted(
            (p for p in out_root.glob(f"*/{sport_low}/{step1_names[0]}") if p.is_file()),
            reverse=True,
        )
        if dated_hits:
            return dated_hits[0]
    if parent is not None:
        hit = next((parent / n for n in step1_names if (parent / n).is_file()), None)
        if hit is not None:
            return hit
    if parent is not None:
        for part in reversed(parent.parts):
            if len(part) == 10 and part[4] == "-" and part[7] == "-":
                dated_dir = Path(REPO_ROOT) / "outputs" / part / sport_low
                hit = next((dated_dir / n for n in step1_names if (dated_dir / n).is_file()), None)
                if hit is not None:
                    return hit
    legacy_cand = Path(REPO_ROOT) / "Sports" / sport_u / "data" / "outputs"
    if sport_u == "NBA":
        legacy_file = legacy_cand / "step1_pp_props_today.csv"
        if legacy_file.is_file():
            return legacy_file
    return None


def _merge_step1_pp_metadata(df: pd.DataFrame, board_path: str, sport: str) -> pd.DataFrame:
    """Attach PrizePicks image_url / player_id from dated step1 when step8 omitted them."""
    if df is None or df.empty:
        return df
    sport_u = str(sport or "").upper()
    step1_path = _resolve_step1_pp_path(board_path, sport_u)
    if step1_path is None or not step1_path.is_file():
        return df
    try:
        s1 = pd.read_csv(step1_path, dtype=str, low_memory=False).fillna("")
    except Exception:
        return df
    if s1.empty:
        return df
    key_cols = [c for c in ("projection_id", "pp_projection_id", "player", "prop_type", "line") if c in s1.columns]
    if "projection_id" not in s1.columns and "player" not in s1.columns:
        return df
    meta_cols = [c for c in ("image_url", "player_id", "nba_player_id") if c in s1.columns]
    if not meta_cols:
        return df
    s1_sub = s1[key_cols + meta_cols].drop_duplicates(subset=key_cols, keep="first")
    out = df.copy()
    if "image_url" not in out.columns:
        out["image_url"] = ""
    if "nba_player_id" not in out.columns and sport_u == "NBA":
        out["nba_player_id"] = ""
    merge_on = None
    if "projection_id" in out.columns and "projection_id" in s1_sub.columns:
        merge_on = ["projection_id"]
    elif all(c in out.columns for c in ("player", "prop_type", "line")) and all(
        c in s1_sub.columns for c in ("player", "prop_type", "line")
    ):
        merge_on = ["player", "prop_type", "line"]
    elif all(c in out.columns for c in ("player", "prop_type")) and all(
        c in s1_sub.columns for c in ("player", "prop_type")
    ):
        merge_on = ["player", "prop_type"]
        s1_sub = s1_sub.sort_values("line", ascending=False).drop_duplicates(
            subset=["player", "prop_type"], keep="first"
        )
    if not merge_on:
        return df
    for col in merge_on:
        if col == "line":
            out[col] = pd.to_numeric(out[col], errors="coerce").round(2)
            s1_sub[col] = pd.to_numeric(s1_sub[col], errors="coerce").round(2)
        else:
            out[col] = out[col].astype(str).str.strip()
            s1_sub[col] = s1_sub[col].astype(str).str.strip()
    out = out.merge(s1_sub, on=merge_on, how="left", suffixes=("", "_pp"))
    if "image_url_pp" in out.columns:
        cur = out.get("image_url", pd.Series("", index=out.index))
        cur_s = cur.astype(str).str.strip()
        cur_bad = cur.isna() | cur_s.eq("") | cur_s.str.lower().isin(["nan", "none"])
        pp_s = out["image_url_pp"].astype(str).str.strip()
        pp_ok = pp_s.ne("") & ~pp_s.str.lower().isin(["nan", "none"])
        out.loc[cur_bad & pp_ok, "image_url"] = out.loc[cur_bad & pp_ok, "image_url_pp"]
        out.drop(columns=["image_url_pp"], inplace=True)
    if sport_u == "NBA":
        if "nba_player_id" not in out.columns:
            out["nba_player_id"] = ""
        for src in ("player_id_pp", "nba_player_id_pp"):
            if src in out.columns:
                m = out["nba_player_id"].astype(str).str.strip().eq("")
                out.loc[m, "nba_player_id"] = out.loc[m, src].apply(_clean_id)
                out.drop(columns=[src], inplace=True, errors="ignore")
    return out


def _board_history_enrichment(df: pd.DataFrame, sport_label: str) -> pd.DataFrame:
    """Stat-g truth for L5/L10, projection fallback — shared by NBA load and step8 boards."""
    if df is None or df.empty:
        return df
    df = _ensure_stat_g_columns(df)
    df = _apply_l5_truth_from_stat_games(df, sport_label)
    df = _apply_l10_truth_from_stat_games(df, sport_label)
    df = _fill_projection_from_avgs(df)
    return df


# ── Load & normalize NBA ───────────────────────────────────────────────────────
def load_nba(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_all_direction_clean.xlsx")

    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path, low_memory=False)
        df = df.loc[:, ~df.columns.duplicated()].copy()
    else:
        xl = pd.ExcelFile(path, engine="openpyxl")
        sheet = "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0]
        df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(
        columns={
            "Tier": "tier",
            "Rank Score": "rank_score",
            "Player": "player",
            "Pos": "pos",
            "Team": "team",
            "Opp": "opp",
            "Game Time": "game_time",
            "Prop": "prop_type",
            "Pick Type": "pick_type",
            "Line": "line",
            "Direction": "direction",
            "Edge": "edge",
            "Abs Edge": "abs_edge",
            "Projection": "projection",
            "ML Prob": "ml_prob",
            "Hit Rate (5g)": "hit_rate",
            "Last 5 Avg": "l5_avg",
            "Season Avg": "season_avg",
            "L5 Over": "l5_over",
            "L5 Under": "l5_under",
            "L10 Over": "l10_over",
            "L10 Under": "l10_under",
            "Games (5g)": "l5_games_played",
            "Games (10g)": "l10_games_played",
            "Standard Line": "standard_line",
            "Game Date": "game_date",
            "Days Rest": "days_rest",
            "image_url": "image_url",
            "Image URL": "image_url",
            "G1": "stat_g1",
            "G2": "stat_g2",
            "G3": "stat_g3",
            "G4": "stat_g4",
            "G5": "stat_g5",
            "G6": "stat_g6",
            "G7": "stat_g7",
            "G8": "stat_g8",
            "G9": "stat_g9",
            "G10": "stat_g10",
            "Def Rank": "def_rank",
            "Def Tier": "def_tier",
            "Pace Tier": "pace_tier",
            "pace_tier": "pace_tier",
            "Context Score": "context_score",
            "context_score": "context_score",
            "Min Tier": "min_tier",
            "Shot Role": "shot_role",
            "Usage Role": "usage_role",
            "Void Reason": "void_reason",
            # OPTIONAL if your NBA file has it:
            "nba_player_id": "nba_player_id",
            "NBA Player ID": "nba_player_id",
            "player_id": "nba_player_id",
            "Player ID": "nba_player_id",
            # H2H / B2B / CV / Opp vs Avg
            "H2H Avg":      "h2h_avg",
            "H2H Over%":    "h2h_over_rate",
            "H2H Games":    "h2h_games",
            "H2H Last":     "h2h_last",
            "B2B":          "b2b_flag",
            "CV%":          "cv_pct",
            "Opp vs Avg%":  "opp_vs_avg_pct",
            "Game Script Mult": "game_script_mult",
            "Game Script Note": "game_script_note",
            "game_script_mult": "game_script_mult",
            "game_script_note": "game_script_note",
            # snake_case step8 CSV (dated outputs/*/nba/)
            "final_bet_direction": "direction",
            "bet_direction": "direction",
            "opp_team": "opp",
            "prop_norm": "prop_type",
            "stat_last5_avg": "l5_avg",
            "stat_last10_avg": "season_avg",
            "last5_over": "l5_over",
            "last5_under": "l5_under",
            "line_games_played_5": "l5_games_played",
            "line_games_played_10": "l10_games_played",
            "line_hits_over_10": "l10_over",
            "line_hits_under_10": "l10_under",
            "b2b_flag": "b2b_flag",
            "days_rest": "days_rest",
            "ESPN_ATHLETE_ID": "espn_player_id",
        }
    )

    # ✅ IMPORTANT: de-dupe before using any column as Series
    df = df.loc[:, ~df.columns.duplicated()].copy()

    df["sport"] = "NBA"

    if "direction" not in df.columns and "final_bet_direction" in df.columns:
        df["direction"] = df["final_bet_direction"]

    if "direction" in df.columns:
        if isinstance(df["direction"], pd.DataFrame):
            df["direction"] = df["direction"].iloc[:, 0]
        df["direction"] = df["direction"].astype(str).str.upper()

    if "tier" in df.columns:
        if isinstance(df["tier"], pd.DataFrame):
            df["tier"] = df["tier"].iloc[:, 0]
        df["tier"] = df["tier"].astype(str).str.upper()

    # Keep void_reason metadata but do not prune NBA board rows here.
    # The product requirement is to show board parity with available PP lines.
    if "void_reason" in df.columns:
        if isinstance(df["void_reason"], pd.DataFrame):
            df["void_reason"] = df["void_reason"].iloc[:, 0]

    # Drop "1st 3 Minutes" props — no historical data, not bettable on PrizePicks standard
    if "prop_type" in df.columns:
        before = len(df)
        df = df[~df["prop_type"].astype(str).str.contains("1st 3 Min", case=False, na=False)].copy()
        dropped = before - len(df)
        if dropped:
            print(f"  [load_nba] Dropped {dropped} '1st 3 Min' props")

    # Clean ID if present
    if "nba_player_id" in df.columns:
        df["nba_player_id"] = df["nba_player_id"].apply(_clean_id)

    df = _board_history_enrichment(df, "NBA")
    df = _merge_step1_pp_metadata(df, path, "NBA")
    df = add_l5_play_side_columns(df)
    if "abs_edge" not in df.columns and "edge" in df.columns:
        df["abs_edge"] = pd.to_numeric(df["edge"], errors="coerce").abs()
    elif "abs_edge" in df.columns:
        df["abs_edge"] = pd.to_numeric(df["abs_edge"], errors="coerce")
    return df


# ── Load & normalize CBB ───────────────────────────────────────────────────────
def load_cbb(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step6_ranked_cbb.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = (
        "ELIGIBLE"
        if "ELIGIBLE" in xl.sheet_names
        else ("ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    )
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(
        columns={
            "final_bet_direction": "direction",
            "bet_direction": "direction",
            "opp_team_abbr": "opp",
            "start_time": "game_time",
            "line_hit_rate": "hit_rate",
            "stat_last5_avg": "l5_avg",
            "stat_season_avg": "season_avg",
            "line_hits_over_5": "l5_over",
            # Intel layer columns (step6e)
            "intel_season_avg":        "intel_season_avg",
            "intel_l5_avg":            "intel_l5_avg",
            "intel_l10_avg":           "intel_l10_avg",
            "intel_season_hit_rate":   "intel_season_hit_rate",
            "intel_cushion":           "intel_cushion",
            "intel_cv_pct":            "intel_cv_pct",
            "intel_opp_vs_league_pct": "intel_opp_vs_league_pct",
            "intel_l5_vs_season":      "intel_l5_vs_season",
            "line_hits_under_5": "l5_under",
            "Def Tier": "def_tier",
            "DEF_TIER": "def_tier",
            "Defense Tier": "def_tier",
            "minutes_tier": "min_tier",
            "Min Tier": "min_tier",
            "shot_role": "shot_role",
            "Shot Role": "shot_role",
            "usage_role": "usage_role",
            "Usage Role": "usage_role",
            # OPTIONAL IDs
            "espn_player_id": "espn_player_id",
            "ESPN Player ID": "espn_player_id",
            "player_id": "espn_player_id",
            # Optional NCAA ranking fields (when present in CBB pipeline output)
            "NCAA Rank": "ncaa_rank",
            "ncaa_rank": "ncaa_rank",
            "OVERALL_DEF_RANK": "ncaa_rank",
            "opp_def_rank": "ncaa_rank",
        }
    )

    # ✅ CRITICAL HOTFIX: de-duplicate columns BEFORE df["direction"].str.upper()
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # ✅ If direction is still a DataFrame for any reason, take the first column.
    if "direction" in df.columns and isinstance(df["direction"], pd.DataFrame):
        df["direction"] = df["direction"].iloc[:, 0]

    df["sport"] = "CBB"

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()

    if "tier" in df.columns:
        if isinstance(df["tier"], pd.DataFrame):
            df["tier"] = df["tier"].iloc[:, 0]
        df["tier"] = df["tier"].astype(str).str.upper()

    if "void_reason" in df.columns:
        if isinstance(df["void_reason"], pd.DataFrame):
            df["void_reason"] = df["void_reason"].iloc[:, 0]
        df = df[df["void_reason"].isna() | (df["void_reason"].astype(str).str.strip() == "")]

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    # Enrich CBB rows with tournament + AP metadata for team and opponent.
    team_src = "team" if "team" in df.columns else ("pp_team" if "pp_team" in df.columns else "")
    opp_src = "opp" if "opp" in df.columns else ("opp_team_abbr" if "opp_team_abbr" in df.columns else "")
    if team_src:
        t_abbr = df[team_src].map(_norm_team_abbr)
        df["team_seed"] = t_abbr.map(lambda a: CBB_TOURNEY_2026.get(a, ("", ""))[0])
        df["team_region"] = t_abbr.map(lambda a: CBB_TOURNEY_2026.get(a, ("", ""))[1])
        df["team_ap_rank"] = t_abbr.map(lambda a: CBB_AP_TOP25_2026.get(a, ""))
    if opp_src:
        o_abbr = df[opp_src].map(_norm_team_abbr)
        df["opp_seed"] = o_abbr.map(lambda a: CBB_TOURNEY_2026.get(a, ("", ""))[0])
        df["opp_region"] = o_abbr.map(lambda a: CBB_TOURNEY_2026.get(a, ("", ""))[1])
        df["opp_ap_rank"] = o_abbr.map(lambda a: CBB_AP_TOP25_2026.get(a, ""))

    # Ensure NCAA rank is numeric when available.
    if "ncaa_rank" in df.columns:
        df["ncaa_rank"] = pd.to_numeric(df["ncaa_rank"], errors="coerce")

    if "team_seed" in df.columns and "opp_seed" in df.columns:
        _ts = pd.to_numeric(df["team_seed"], errors="coerce")
        _os = pd.to_numeric(df["opp_seed"], errors="coerce")
        df["is_tournament_game"] = ((_ts.notna()) & (_ts > 0)) | ((_os.notna()) & (_os > 0))
    else:
        df["is_tournament_game"] = False

    return df


def load_cfb(path: str) -> pd.DataFrame:
    """College Football step6 ranked workbook (same layout as CBB step6)."""
    from utils.cfb_playoff_metadata import (
        CFB_AP_TOP25_2026,
        cfb_playoff_info,
        cfb_row_in_playoff,
        norm_cfb_team_abbr,
    )

    path = resolve_input_path(path, fallback_filename="step6_ranked_cfb.xlsx")
    df = load_cbb(path)
    if df is None or len(df) == 0:
        return df
    df = df.copy()
    df["sport"] = "CFB"

    team_src = "team" if "team" in df.columns else ("pp_team" if "pp_team" in df.columns else "")
    opp_src = "opp" if "opp" in df.columns else ("opp_team_abbr" if "opp_team_abbr" in df.columns else "")
    if team_src:
        t_abbr = df[team_src].map(norm_cfb_team_abbr)
        if "team_seed" not in df.columns or df["team_seed"].isna().all():
            df["team_seed"] = t_abbr.map(lambda a: cfb_playoff_info(a)[0])
        if "playoff_round" not in df.columns:
            df["playoff_round"] = t_abbr.map(lambda a: cfb_playoff_info(a)[1])
        df["team_ap_rank"] = t_abbr.map(lambda a: CFB_AP_TOP25_2026.get(a, ""))
    if opp_src:
        o_abbr = df[opp_src].map(norm_cfb_team_abbr)
        df["opp_seed"] = o_abbr.map(lambda a: cfb_playoff_info(a)[0])
        df["opp_playoff_round"] = o_abbr.map(lambda a: cfb_playoff_info(a)[1])
        df["opp_ap_rank"] = o_abbr.map(lambda a: CFB_AP_TOP25_2026.get(a, ""))

    if "is_playoff_game" not in df.columns:
        if team_src and opp_src:
            df["is_playoff_game"] = df.apply(
                lambda r: int(cfb_row_in_playoff(r[team_src], r[opp_src])),
                axis=1,
            )
        else:
            df["is_playoff_game"] = 0
    df["is_tournament_game"] = df["is_playoff_game"]

    if "team_playoff_seed" in df.columns and "team_seed" in df.columns and df["team_seed"].isna().all():
        df["team_seed"] = df["team_playoff_seed"]

    return df


def _fill_nhl_l5_season_avgs(df: pd.DataFrame) -> pd.DataFrame:
    """
    Step8 NHL often writes blank avg_L5 / avg_season while projection and last-N raw
    game cells are populated. Backfill L5 Avg and Szn Avg for combined / Excel / web.
    """
    if df is None or len(df) == 0:
        return df
    out = df.copy()
    if "l5_avg" not in out.columns:
        out["l5_avg"] = np.nan
    if "season_avg" not in out.columns:
        out["season_avg"] = np.nan

    l5 = pd.to_numeric(out["l5_avg"], errors="coerce")
    szn = pd.to_numeric(out["season_avg"], errors="coerce")
    proj = (
        pd.to_numeric(out["projection"], errors="coerce")
        if "projection" in out.columns
        else pd.Series(np.nan, index=out.index)
    )
    a10 = (
        pd.to_numeric(out["avg_L10"], errors="coerce")
        if "avg_L10" in out.columns
        else pd.Series(np.nan, index=out.index)
    )
    a20 = (
        pd.to_numeric(out["avg_L20"], errors="coerce")
        if "avg_L20" in out.columns
        else pd.Series(np.nan, index=out.index)
    )

    raw_cols = [c for c in ("last1_raw", "last2_raw", "last3_raw") if c in out.columns]
    l3_mean = pd.Series(np.nan, index=out.index)
    if raw_cols:
        mats = [pd.to_numeric(out[c], errors="coerce") for c in raw_cols]
        l3_mean = pd.concat(mats, axis=1).mean(axis=1, skipna=True)

    l5_filled = l5.combine_first(proj).combine_first(l3_mean)
    n_l5 = int((l5.isna() & l5_filled.notna()).sum())
    out["l5_avg"] = l5_filled

    szn_filled = szn.combine_first(a10).combine_first(a20).combine_first(l5_filled)
    n_szn = int((szn.isna() & szn_filled.notna()).sum())
    out["season_avg"] = szn_filled

    if n_l5 or n_szn:
        print(
            f"  [load_nhl] backfilled l5_avg on {n_l5} row(s), season_avg on {n_szn} row(s) "
            "(projection / avg_L10 / avg_L20 / last1–3 raw fallbacks)"
        )
    return out


# ── Load & normalize NHL ──────────────────────────────────────────────────────
def load_nhl(path: str) -> pd.DataFrame:
    raw = (path or "").strip()
    if not raw:
        return pd.DataFrame()

    try:
        path = resolve_input_path(raw, fallback_filename="step8_nhl_direction_clean.xlsx")
    except FileNotFoundError:
        print("  [load_nhl] NHL file not found — skipping NHL")
        return pd.DataFrame()

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "NHL" if "NHL" in xl.sheet_names else ("ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        "Game Script Mult": "game_script_mult",
        "Game Script Note": "game_script_note",
        "player_name":        "player",
        "position":           "pos",
        "stat_type":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "Composite Hit Rate": "hit_rate",
        "composite_hr":       "hit_rate",
        "hr_L10":             "hit_rate_over_L10",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "Last 5 Avg":         "l5_avg",
        "Season Avg":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "Rank Score":         "rank_score",
        "Line Combo":         "line_combo",
        "line_combo":         "line_combo",
        "game_start":         "game_time",
        "Fetched At":         "fetched_at",
        "fetched_at":         "fetched_at",
    })
    # Deduplicate columns immediately after rename — multiple source cols may map to same target
    df = df.loc[:, ~df.columns.duplicated()].copy()

    # Fallback: derive hit_rate from hit_rate_over_L10 if still missing
    if "hit_rate" not in df.columns or df["hit_rate"].isna().all():
        for fallback_col in ("hit_rate_over_L10", "hit_rate_over_L5", "hit_rate_over_L20",
                             "over_L10", "over_L5"):
            if fallback_col in df.columns:
                df["hit_rate"] = pd.to_numeric(df[fallback_col], errors="coerce")
                print(f"  [load_nhl] hit_rate sourced from '{fallback_col}'")
                break

    # Normalize hit_rate to 0-1 — handle "94.0%", "0.94", or 94.0
    if "hit_rate" in df.columns:
        hr = df["hit_rate"].astype(str).str.replace("%", "", regex=False).str.strip()
        hr = pd.to_numeric(hr, errors="coerce")
        if hr.dropna().max() > 1.5:   # clearly a percentage value (e.g. 94.0)
            hr = hr / 100.0
        df["hit_rate"] = hr

    # NHL proxy hotfix:
    # If hit_rate is present but effectively zeros, source a proxy from directional windows
    # so strict min-leg gates do not collapse a healthy NHL pool.
    if "hit_rate" in df.columns:
        hr_now = pd.to_numeric(df["hit_rate"], errors="coerce").fillna(0.0)
        zero_like = bool((hr_now <= 0.001).mean() >= 0.80)
        if zero_like:
            proxy_col = None
            for c in ("hit_rate_over_L10", "hit_rate_over_L5", "hit_rate_over_L20", "over_L10", "over_L5", "composite_hr"):
                if c in df.columns:
                    proxy_col = c
                    break
            if proxy_col is not None:
                proxy = pd.to_numeric(df[proxy_col], errors="coerce")
                # Convert percentages to 0-1 when needed.
                if proxy.dropna().max() > 1.5:
                    proxy = proxy / 100.0
                # Clamp to realistic range for NHL leg-level selection.
                proxy = proxy.clip(lower=0.52, upper=0.90)
                df["hit_rate"] = proxy.where(proxy.notna(), hr_now)
                print(f"  [load_nhl] hit_rate proxy applied from '{proxy_col}' (>=80% zero-like source)")

    # opponent is stored in 'description' column
    if "opp" not in df.columns:
        if "description" in df.columns:
            df["opp"] = df["description"]
        else:
            df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "NHL"

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)
    forced = df["pick_type"].isin(["Goblin", "Demon"])

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
        df.loc[forced, "direction"] = "OVER"
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    # Extra fallback: NHL step8 may still have line_score if rename didn't catch it
    if "line" not in df.columns:
        for alt in ("line_score", "Line", "line_value", "prop_line"):
            if alt in df.columns:
                df["line"] = df[alt]
                break
    if "line" not in df.columns:
        df["line"] = np.nan

    for col in ["rank_score", "hit_rate", "line"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")
    if "edge" not in df.columns:
        df["edge"] = np.nan
    df["edge"] = pd.to_numeric(df["edge"], errors="coerce")
    if "abs_edge" not in df.columns:
        df["abs_edge"] = np.nan
    df["abs_edge"] = pd.to_numeric(df["abs_edge"], errors="coerce")
    df["abs_edge"] = df["abs_edge"].where(df["abs_edge"].notna(), df["edge"].abs())

    # Faceoff wins are currently not sourced with reliable per-player counts
    # in our actuals pipeline; exclude until a stable source is added.
    if "prop_type" in df.columns:
        before = len(df)
        p = df["prop_type"].astype(str).str.lower()
        df = df[~p.str.contains(r"faceoff", regex=True, na=False)].copy()
        dropped = before - len(df)
        if dropped > 0:
            print(f"  [load_nhl] Dropped {dropped} unreliable faceoff props")

    # Last-5 game counts vs line (step8 raw cols) → slate L5 Over / L5 Under
    if "over_L5_raw" in df.columns:
        df["l5_over"] = pd.to_numeric(df["over_L5_raw"], errors="coerce")
    elif "l5_over" not in df.columns:
        df["l5_over"] = np.nan
    if "under_L5_raw" in df.columns:
        df["l5_under"] = pd.to_numeric(df["under_L5_raw"], errors="coerce")
    elif "l5_under" not in df.columns:
        df["l5_under"] = np.nan

    # Keep NHL on the same source-of-truth L5 logic as other sports whenever
    # stat_g* windows are present in the board payload.
    df = _ensure_stat_g_columns(df)
    df = _apply_l5_truth_from_stat_games(df, "NHL")
    df = add_l5_play_side_columns(df)
    df = _fill_nhl_l5_season_avgs(df)

    df = df[df["line"].notna() & (df["line"] > 0)]
    # Convert all pandas NA/NaT to None so openpyxl can handle them
    df = df.astype(object).where(df.notna(), other=None)
    return df



def _tennis_hit_rate_zero_like_proxy(df: pd.DataFrame, log_prefix: str) -> None:
    """
    Tennis step8 often carries Hit Rate (5g) as zeros while L10/L5 windows are populated.
    Mirror the NHL hotfix: if >=80% of rows look like zero hit_rate, backfill from directional windows.
    Mutates df['hit_rate'] in place (expects numeric 0–1 hit_rate column after normalization).
    """
    if "hit_rate" not in df.columns or len(df) == 0:
        return
    hr_now = pd.to_numeric(df["hit_rate"], errors="coerce").fillna(0.0)
    if bool((hr_now <= 0.001).mean() < 0.80):
        return
    proxy_col = None
    for c in ("l10_over", "l10_under", "l5_over", "l5_under"):
        if c in df.columns:
            proxy_col = c
            break
    if proxy_col is None:
        return
    proxy = pd.to_numeric(df[proxy_col], errors="coerce")
    if proxy.notna().any() and float(proxy.dropna().max()) > 1.5:
        proxy = proxy / 100.0
    proxy = proxy.clip(lower=0.52, upper=0.90)
    df["hit_rate"] = proxy.where(proxy.notna(), hr_now)
    print(
        f"  [{log_prefix}] Tennis hit_rate proxy applied from '{proxy_col}' "
        "(>=80% zero-like Hit Rate (5g))"
    )


# ── Load & normalize step8 "direction clean" boards (Soccer, Tennis, …) ───────
def _load_step8_board_like(
    path: str,
    *,
    fallback_filename: str,
    sheet_order: tuple[str, ...],
    sport: str,
    log_prefix: str,
) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename=fallback_filename)
    df: pd.DataFrame
    if str(path).lower().endswith(".csv"):
        df = pd.read_csv(path, low_memory=False)
        df = df.loc[:, ~df.columns.duplicated()].copy()
    else:
        try:
            xl = pd.ExcelFile(path, engine="openpyxl")
            sheet = next((s for s in sheet_order if s in xl.sheet_names), xl.sheet_names[0])
            df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")
        except PermissionError:
            base, _ext = os.path.splitext(path)
            csv_candidates = [
                f"{base}.csv",
                f"{base.replace('_clean', '')}.csv",
            ]
            csv_path = next((p for p in csv_candidates if os.path.exists(p)), "")
            if not csv_path:
                raise
            print(f"  [{log_prefix}] XLSX locked; using CSV fallback: {csv_path}")
            df = pd.read_csv(csv_path)

    df = df.rename(columns={
        # title-case (from step8 clean xlsx)
        "Player":           "player",
        "Tier":             "tier",
        "Rank Score":       "rank_score",
        "Pos":              "pos",
        "Team":             "team",
        "Opp":              "opp",
        "Game Time":        "game_time",
        "Game Date":        "game_date",
        "Prop":             "prop_type",
        "Pick Type":        "pick_type",
        "Line":             "line",
        "Direction":        "direction",
        "Edge":             "edge",
        "Abs Edge":         "abs_edge",
        "Projection":       "projection",
        "ESPN ID":          "espn_player_id",
        "Hit Rate (5g)":    "hit_rate",
        "Hit Rate Status":  "hit_rate_status",
        "Reliability Note": "reliability_note",
        # Kept separate so we can coalesce into hit_rate when 5g is blank (common when
        # line-hit columns aren't populated yet).
        "Hit Rate (10g)":   "_board_hit10",
        "Last 5 Avg":       "l5_avg",
        "Season Avg":       "season_avg",
        "L5 Over":          "l5_over",
        "L5 Under":         "l5_under",
        "Games (5g)":       "l5_games_played",
        "L10 Over":         "l10_over",
        "L10 Under":        "l10_under",
        "Games (10g)":      "l10_games_played",
        "line_games_played_5":  "l5_games_played",
        "line_games_played_10": "l10_games_played",
        "Def Rank":         "def_rank",
        "Def Tier":         "def_tier",
        "Min Tier":         "min_tier",
        "Shot Role":        "shot_role",
        "Usage Role":       "usage_role",
        "League":           "league",
        "Pos Group":        "position_group",
        "Void Reason":      "void_reason",
        "Rank Score Penalized": "rank_score_penalized",
        "Surface":          "surface",
        "Line Combo":       "line_combo",
        "distribution_std": "distribution_std",
        "Distribution Std": "distribution_std",
        "distribution_n":   "distribution_n",
        "Distribution N":   "distribution_n",
        "G1": "stat_g1", "G2": "stat_g2", "G3": "stat_g3", "G4": "stat_g4", "G5": "stat_g5",
        "G6": "stat_g6", "G7": "stat_g7", "G8": "stat_g8", "G9": "stat_g9", "G10": "stat_g10",
        # snake_case fallbacks
        "rank_score_penalized": "rank_score_penalized",
        "surface":          "surface",
        "line_combo":       "line_combo",
        "player_name":        "player",
        "player_id":          "nba_player_id",
        "wnba_player_id":     "nba_player_id",
        "espn_athlete_id":    "espn_player_id",
        "image_url":          "image_url",
        "stat_type":          "prop_type",
        "stat_norm":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
        "start_time":         "game_time",
        "Fetched At":         "fetched_at",
        "fetched_at":         "fetched_at",
        "opponent":           "opp",
        "opp_team":           "opp",
        "pos":                "pos",
        "line_hit_rate_over_ou_5":  "hit_rate",
        "line_hit_rate_over_ou_10": "_board_hit10",
        "hit_rate_over_L10": "l10_over",
        "hit_rate_under_L10": "l10_under",
        "over_L10": "l10_over",
        "under_L10": "l10_under",
        "Last 10 Avg": "season_avg",
        "Game Script Mult": "game_script_mult",
        "Game Script Note": "game_script_note",
        "game_script_mult": "game_script_mult",
        "game_script_note": "game_script_note",
        "Blended Score": "blended_score",
        "blended_score": "blended_score",
        "Open Line": "open_line",
        "Line Movement": "line_movement",
        "Line Shift": "line_direction_shift",
        "open_line": "open_line",
        "line_movement": "line_movement",
        "line_direction_shift": "line_direction_shift",
    })

    if "opp" not in df.columns:
        df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = sport

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    if "hit_rate" not in df.columns:
        df["hit_rate"] = np.nan

    for col in [
        "rank_score",
        "hit_rate",
        "line",
        "_board_hit10",
        "blended_score",
        "l5_avg",
        "season_avg",
        "l5_over",
        "l5_under",
        "l10_over",
        "l10_under",
        "projection",
        "edge",
        "abs_edge",
    ]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Prefer L5 hit rate, then L10, when either is present.
    if "_board_hit10" in df.columns:
        df["hit_rate"] = df["hit_rate"].combine_first(df["_board_hit10"])
        df.drop(columns=["_board_hit10"], inplace=True)

    # Normalize hit_rate to 0–1 (handles "62%" or 62.0 from spreadsheets)
    if "hit_rate" in df.columns and df["hit_rate"].notna().any():
        hr = df["hit_rate"]
        if hr.dtype == object:
            hr = hr.astype(str).str.replace("%", "", regex=False).str.strip()
            hr = pd.to_numeric(hr, errors="coerce")
        else:
            hr = pd.to_numeric(hr, errors="coerce")
        if hr.dropna().max() is not None and hr.dropna().max() > 1.5:
            hr = hr / 100.0
        df["hit_rate"] = hr

    # NBA1Q can overstate hit_rate on tiny windows (e.g., 5/5 => 100%).
    def _num_col(name: str) -> pd.Series:
        if name not in df.columns:
            return pd.Series(np.nan, index=df.index, dtype="float64")
        col = df[name]
        if isinstance(col, pd.DataFrame):
            col = col.iloc[:, 0]
        return pd.to_numeric(col, errors="coerce")

    l5o = _num_col("l5_over")
    l5u = _num_col("l5_under")
    l5n = l5o.add(l5u, fill_value=0)
    hits = np.where(
        df["direction"].astype(str).str.upper().eq("UNDER").to_numpy(),
        l5u.to_numpy(dtype=float, copy=True),
        l5o.to_numpy(dtype=float, copy=True),
    )
    prior_p = 0.55
    prior_n = 15.0
    shrunk = (hits + prior_p * prior_n) / (l5n.to_numpy(dtype=float, copy=True) + prior_n)
    use_shrink = (l5n > 0) & (l5n <= 10) & np.isfinite(shrunk)
    df.loc[use_shrink, "hit_rate"] = shrunk[use_shrink.to_numpy()]
    df["hit_rate"] = pd.to_numeric(df["hit_rate"], errors="coerce").clip(lower=0.35, upper=0.92)

    # Still no usable hit rate (common when line-hit columns aren't wired yet).
    # Use a mild rank_score-based proxy so tier/rank ticket gates still run.
    hr_series = pd.to_numeric(df["hit_rate"], errors="coerce")
    if hr_series.notna().sum() == 0:
        rs = pd.to_numeric(df.get("rank_score", 0), errors="coerce").fillna(0.0)
        q25, q75 = float(rs.quantile(0.25)), float(rs.quantile(0.75))
        span = (q75 - q25) + 1e-6
        proxy = 0.54 + ((rs - q25) / span).clip(lower=0.0, upper=1.0) * 0.12
        df["hit_rate"] = proxy.clip(0.50, 0.68)
        print(
            f"  [{log_prefix}] NOTE: Hit Rate (5g)/(10g) empty - using rank_score proxy for ticket eligibility."
        )

    # Backfill sparse board stats so slate columns are populated.
    hr_for_counts = pd.to_numeric(df.get("hit_rate", np.nan), errors="coerce").clip(lower=0.0, upper=1.0)
    if "l5_over" not in df.columns:
        df["l5_over"] = np.nan
    if "l5_under" not in df.columns:
        df["l5_under"] = np.nan
    if "l10_over" not in df.columns:
        df["l10_over"] = np.nan
    if "l10_under" not in df.columns:
        df["l10_under"] = np.nan
    if "l5_avg" not in df.columns:
        df["l5_avg"] = np.nan
    if "season_avg" not in df.columns:
        df["season_avg"] = np.nan

    df = _ensure_stat_g_columns(df)
    df = _apply_l5_truth_from_stat_games(df, sport, min_stat_games=5)
    df = _apply_l10_truth_from_stat_games(df, sport, min_stat_games=6)
    df = _fill_projection_from_avgs(df)

    # IMPORTANT:
    # For these boards, upstream "Hit Rate (5g)/(10g)" is direction-aware in many cases
    # (e.g. when stat_g* are missing, step7 fills line_hit_rate into the over_* columns).
    # If hit_rate represents the chosen bet side, we must derive L5/L10 counts
    # directionally so UNDER rows don't get reversed.
    dirv = df.get("direction", pd.Series(["OVER"] * len(df), index=df.index)).astype(str).str.upper().fillna("OVER")

    l5_hit_as_over = (hr_for_counts * 5.0).round()
    l5_over_fill = l5_hit_as_over.where(dirv.ne("UNDER"), 5.0 - l5_hit_as_over)
    l5_under_fill = 5.0 - l5_over_fill

    l10_hit_as_over = (hr_for_counts * 10.0).round()
    l10_over_fill = l10_hit_as_over.where(dirv.ne("UNDER"), 10.0 - l10_hit_as_over)
    l10_under_fill = 10.0 - l10_over_fill

    df["l5_over"] = pd.to_numeric(df["l5_over"], errors="coerce").combine_first(l5_over_fill)
    df["l5_under"] = pd.to_numeric(df["l5_under"], errors="coerce").combine_first(l5_under_fill)
    df["l10_over"] = pd.to_numeric(df["l10_over"], errors="coerce").combine_first(l10_over_fill)
    df["l10_under"] = pd.to_numeric(df["l10_under"], errors="coerce").combine_first(l10_under_fill)

    proj = pd.to_numeric(df.get("projection", np.nan), errors="coerce")
    if not isinstance(proj, pd.Series):
        proj = pd.Series(np.nan, index=df.index, dtype="float64")
    else:
        proj = proj.reindex(df.index).astype("float64", copy=False)

    df["l5_avg"] = pd.to_numeric(df["l5_avg"], errors="coerce").combine_first(proj)
    df["season_avg"] = pd.to_numeric(df["season_avg"], errors="coerce").combine_first(df["l5_avg"]).combine_first(proj)

    # Edge is only meaningful when projection exists (e.g. WNBA step8 often omits it; sheet edge
    # can be 0 - line = -line with abs_edge = line, tripping distance/Demon heuristics). When
    # projection is missing or non-finite, leave edge and abs_edge unset.
    if "line" in df.columns:
        line_sr = pd.to_numeric(df["line"], errors="coerce").reindex(df.index).astype("float64", copy=False)
    else:
        line_sr = pd.Series(np.nan, index=df.index, dtype="float64")
    proj_arr = proj.to_numpy(dtype=np.float64, copy=False)
    line_arr = line_sr.to_numpy(dtype=np.float64, copy=False)
    proj_ok = pd.Series(np.isfinite(proj_arr) & np.isfinite(line_arr), index=df.index)
    computed_edge = proj - line_sr
    df["edge"] = computed_edge.where(proj_ok, np.nan)
    df["abs_edge"] = computed_edge.abs().where(proj_ok, np.nan)

    # Demon is distance-sensitive in upstream step8; without a direction/edge gate, UNDER legs
    # (or rows with missing projection → edge <= 0) can be mislabeled. PP-style Demon here
    # is only retained on OVER with strictly positive model edge; others → Goblin if
    # abs_edge suggests a softened line, else Standard.
    _pt_low = df["pick_type"].astype(str).str.strip().str.lower()
    _dmask = _pt_low.eq("demon")
    if bool(_dmask.any()):
        _dir_u = df["direction"].astype(str).str.strip().str.upper()
        _edge = pd.to_numeric(df["edge"], errors="coerce")
        _ae = pd.to_numeric(df["abs_edge"], errors="coerce")
        _ae = _ae.where(_ae.notna(), _edge.abs())
        _bad = _dmask & (~_dir_u.eq("OVER") | ~_edge.gt(0))
        if bool(_bad.any()):
            n_bad = int(_bad.sum())
            _use_gob = _bad & _ae.ge(0.5)
            _use_std = _bad & ~_use_gob
            df.loc[_use_gob, "pick_type"] = "Goblin"
            df.loc[_use_std, "pick_type"] = "Standard"
            print(
                f"  [{log_prefix}] demoted {n_bad} invalid Demon row(s) "
                "(require OVER + positive edge; Goblin if abs_edge>=0.5 else Standard)."
            )

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    df = df[df["line"].notna() & (df["line"] >= 0)]
    df = _board_history_enrichment(df, sport)
    df = _merge_step1_pp_metadata(df, path, sport)
    df = df.astype(object).where(df.notna(), other=None)
    return df


def load_soccer(path: str) -> pd.DataFrame:
    df = _load_step8_board_like(
        path,
        fallback_filename="step8_soccer_direction_clean.xlsx",
        sheet_order=("Soccer", "ALL"),
        sport="Soccer",
        log_prefix="load_soccer",
    )
    if df is None or df.empty:
        return df

    # Guardrail: keep only props that still exist in the current Soccer PP fetch snapshot.
    # This prevents stale names from older step8 rows appearing on /tickets after board churn.
    try:
        step8_path = resolve_input_path(path, fallback_filename="step8_soccer_direction_clean.xlsx")
        step8_dir = os.path.dirname(step8_path)
        step1_csv = os.path.join(step8_dir, "step1_soccer_props.csv")
        if os.path.exists(step1_csv):
            s1 = pd.read_csv(step1_csv)
            live_ids: set[str] = set()
            for c in ("pp_projection_id", "projection_id"):
                if c in s1.columns:
                    vals = (
                        pd.to_numeric(s1[c], errors="coerce")
                        .dropna()
                        .astype(int)
                        .astype(str)
                    )
                    live_ids.update(vals.tolist())
            if live_ids:
                id_series = pd.Series([""] * len(df), index=df.index)
                for c in ("pp_projection_id", "projection_id"):
                    if c in df.columns:
                        vals = (
                            pd.to_numeric(df[c], errors="coerce")
                            .fillna(-1)
                            .astype(int)
                            .astype(str)
                        )
                        id_series = id_series.where(id_series.ne(""), vals)
                keep_id = id_series.isin(live_ids)

                # Fallback matcher: some Soccer step8 runs remap projection IDs.
                # In that case, keep rows that match current PP board keys.
                def _norm_txt(v: object) -> str:
                    return re.sub(r"\s+", " ", str(v or "").strip().lower())

                live_keys: set[tuple[str, str, str, str]] = set()
                for _, r in s1.iterrows():
                    live_keys.add(
                        (
                            _norm_txt(r.get("player")),
                            _norm_txt(r.get("prop_type")),
                            str(pd.to_numeric(r.get("line"), errors="coerce")),
                            _norm_txt(r.get("pick_type")),
                        )
                    )

                row_keys = df.apply(
                    lambda r: (
                        _norm_txt(r.get("player")),
                        _norm_txt(r.get("prop_type")),
                        str(pd.to_numeric(r.get("line"), errors="coerce")),
                        _norm_txt(r.get("pick_type")),
                    ),
                    axis=1,
                )
                keep_key = row_keys.isin(live_keys)
                keep = keep_id | keep_key

                before = len(df)
                kept_id = int(keep_id.sum())
                kept_key = int((~keep_id & keep_key).sum())
                df = df.loc[keep].copy()
                dropped = before - len(df)
                if dropped > 0:
                    print(
                        f"  [load_soccer] reconciled against step1_soccer_props.csv: "
                        f"dropped {dropped} stale rows; kept {len(df)} "
                        f"(id={kept_id}, key={kept_key})"
                    )
            else:
                print("  [load_soccer] step1 snapshot has no projection IDs; skipping stale-row reconciliation")
        else:
            print("  [load_soccer] step1_soccer_props.csv not found; skipping stale-row reconciliation")
    except Exception as e:
        print(f"  [load_soccer] WARN: stale-row reconciliation skipped ({e})")

    if "opp" in df.columns:
        opp_norm = (
            df["opp"]
            .astype(str)
            .str.strip()
            .str.upper()
        )
        bad_opp = opp_norm.isin({"", "UNKNOWN_OPP", "UNKNOWN"})
        # Soft-degrade unknown opponent rows instead of hard dropping the entire sport pool.
        # Keep the prop leg but strip opponent-dependent context fields.
        df["opp_known"] = (~bad_opp).astype(bool)
        if bad_opp.any():
            for dep_col in ("def_tier", "vs_def", "opp_def_tier"):
                if dep_col in df.columns:
                    df.loc[bad_opp, dep_col] = None
            if "load_warn" not in df.columns:
                df["load_warn"] = None
            df.loc[bad_opp, "load_warn"] = "UNKNOWN_OPP"
            print(
                f"  [load_soccer] kept {int(bad_opp.sum())} rows with unknown opponent metadata "
                f"(opp_known=False; opponent-dependent fields nulled)"
            )
    return df


def _tennis_board_hit_rate_proxy(df: pd.DataFrame) -> pd.DataFrame:
    """
    PrizePicks Tennis has no graded history — step8 often leaves hit_rate at 0.
    1) NHL-style L10/L5 window proxy when >=80% of rows are zero-like (realize counts /5).
    2) Remaining zeros: derive from blended_score (zero-history board) when still mostly flat.
    """
    if df is None or len(df) == 0 or "hit_rate" not in df.columns:
        return df
    out = df.copy()
    _tennis_hit_rate_zero_like_proxy(out, "load_tennis")
    hr0 = pd.to_numeric(out["hit_rate"], errors="coerce").fillna(0.0)
    if float((hr0 <= 0.001).mean()) < 0.60:
        return out
    bs = pd.to_numeric(out.get("blended_score", np.nan), errors="coerce")
    if bs.notna().sum() == 0:
        return out
    high = bs >= 0.70
    proxy = (bs * 0.65).where(high, (bs * 0.58))
    proxy = proxy.clip(0.52, 0.90)
    m0 = hr0 <= 0.001
    use = m0 & proxy.notna()
    out.loc[use, "hit_rate"] = proxy.loc[use]
    print("  [load_tennis] hit_rate proxy from blended_score (zero-history board)")
    return out


def load_tennis(path: str) -> pd.DataFrame:
    base = _load_step8_board_like(
        path,
        fallback_filename="step8_tennis_direction_clean.xlsx",
        sheet_order=("Tennis", "ALL"),
        sport="Tennis",
        log_prefix="load_tennis",
    )
    return _tennis_board_hit_rate_proxy(base)


def load_wnba(path: str) -> pd.DataFrame:
    """WNBA step8 direction workbook (same column contract as other step8 boards)."""
    return _load_step8_board_like(
        path,
        fallback_filename="step8_wnba_direction.xlsx",
        sheet_order=("WNBA", "ALL"),
        sport="WNBA",
        log_prefix="load_wnba",
    )


def load_nfl(path: str) -> pd.DataFrame:
    """NFL step8 direction clean workbook."""
    return _load_step8_board_like(
        path,
        fallback_filename="step8_nfl_direction_clean.xlsx",
        sheet_order=("NFL", "ALL"),
        sport="NFL",
        log_prefix="load_nfl",
    )


def load_wcbb(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_wcbb_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "Soccer" if "Soccer" in xl.sheet_names else (
        "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        # title-case (from step8 clean xlsx)
        "Player":           "player",
        "Tier":             "tier",
        "Rank Score":       "rank_score",
        "Pos":              "pos",
        "Team":             "team",
        "Opp":              "opp",
        "Game Time":        "game_time",
        "Prop":             "prop_type",
        "Pick Type":        "pick_type",
        "Line":             "line",
        "Direction":        "direction",
        "Edge":             "edge",
        "Abs Edge":         "abs_edge",
        "Projection":       "projection",
        "ESPN ID":          "espn_player_id",
        "Hit Rate (5g)":    "hit_rate",
        "Hit Rate Status":  "hit_rate_status",
        "Reliability Note": "reliability_note",
        # Kept separate so we can coalesce into hit_rate when 5g is blank (common when
        # Soccer step5/7 line-hit columns aren't populated yet).
        "Hit Rate (10g)":   "_soccer_hit10",
        "Last 5 Avg":       "l5_avg",
        "Season Avg":       "season_avg",
        "L5 Over":          "l5_over",
        "L5 Under":         "l5_under",
        "Def Rank":         "def_rank",
        "Def Tier":         "def_tier",
        "Min Tier":         "min_tier",
        "Shot Role":        "shot_role",
        "Usage Role":       "usage_role",
        "League":           "league",
        "Pos Group":        "position_group",
        "Void Reason":      "void_reason",
        # snake_case fallbacks
        "player_name":        "player",
        "stat_type":          "prop_type",
        "stat_norm":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
        "opponent":           "opp",
        "line_hit_rate_over_ou_5":  "hit_rate",
        "line_hit_rate_over_ou_10": "_soccer_hit10",
        "Game Script Mult": "game_script_mult",
        "Game Script Note": "game_script_note",
        "game_script_mult": "game_script_mult",
        "game_script_note": "game_script_note",
    })

    if "opp" not in df.columns:
        df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "WCBB"

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    if "hit_rate" not in df.columns:
        df["hit_rate"] = np.nan

    for col in ["rank_score", "hit_rate", "line", "_soccer_hit10"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Prefer L5 hit rate, then L10, when either is present.
    if "_soccer_hit10" in df.columns:
        df["hit_rate"] = df["hit_rate"].combine_first(df["_soccer_hit10"])
        df.drop(columns=["_soccer_hit10"], inplace=True)

    # Normalize hit_rate to 0–1 (handles "62%" or 62.0 from spreadsheets)
    if "hit_rate" in df.columns and df["hit_rate"].notna().any():
        hr = df["hit_rate"]
        if hr.dtype == object:
            hr = hr.astype(str).str.replace("%", "", regex=False).str.strip()
            hr = pd.to_numeric(hr, errors="coerce")
        else:
            hr = pd.to_numeric(hr, errors="coerce")
        if hr.dropna().max() is not None and hr.dropna().max() > 1.5:
            hr = hr / 100.0
        df["hit_rate"] = hr

    # NBA1H can overstate hit_rate on tiny windows (e.g., 5/5 => 100%).
    l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
    l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
    l5n = l5o.add(l5u, fill_value=0)
    hits = np.where(
        df["direction"].astype(str).str.upper().eq("UNDER").to_numpy(),
        l5u.to_numpy(dtype=float, copy=True),
        l5o.to_numpy(dtype=float, copy=True),
    )
    prior_p = 0.55
    prior_n = 15.0
    shrunk = (hits + prior_p * prior_n) / (l5n.to_numpy(dtype=float, copy=True) + prior_n)
    use_shrink = (l5n > 0) & (l5n <= 10) & np.isfinite(shrunk)
    df.loc[use_shrink, "hit_rate"] = shrunk[use_shrink.to_numpy()]
    df["hit_rate"] = pd.to_numeric(df["hit_rate"], errors="coerce").clip(lower=0.35, upper=0.92)

    # NBA split boards can show extreme 5/5 streaks; shrink tiny-window rates so UI/tickets
    # do not overstate confidence (e.g., 100% from only 5 samples).
    l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
    l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
    l5n = l5o.add(l5u, fill_value=0)
    hits = np.where(
        df["direction"].astype(str).str.upper().eq("UNDER").to_numpy(),
        l5u.to_numpy(dtype=float, copy=True),
        l5o.to_numpy(dtype=float, copy=True),
    )
    # Strong prior toward 55% over a 15-game pseudo sample.
    prior_p = 0.55
    prior_n = 15.0
    shrunk = (hits + prior_p * prior_n) / (l5n.to_numpy(dtype=float, copy=True) + prior_n)
    hr_now = pd.to_numeric(df["hit_rate"], errors="coerce")
    use_shrink = (l5n > 0) & (l5n <= 10) & np.isfinite(shrunk)
    df.loc[use_shrink, "hit_rate"] = shrunk[use_shrink.to_numpy()]
    df["hit_rate"] = pd.to_numeric(df["hit_rate"], errors="coerce").clip(lower=0.35, upper=0.92)

    # Still no usable hit rate (common on current Soccer pipeline when game logs are empty).
    # Use a mild rank_score-based proxy so tier/rank ticket gates still run; re-run step5 when HRs exist.
    hr_series = pd.to_numeric(df["hit_rate"], errors="coerce")
    if hr_series.notna().sum() == 0:
        rs = pd.to_numeric(df.get("rank_score", 0), errors="coerce").fillna(0.0)
        q25, q75 = float(rs.quantile(0.25)), float(rs.quantile(0.75))
        span = (q75 - q25) + 1e-6
        proxy = 0.54 + ((rs - q25) / span).clip(lower=0.0, upper=1.0) * 0.12
        df["hit_rate"] = proxy.clip(0.50, 0.68)
        print(
            "  [load_soccer] NOTE: Hit Rate (5g)/(10g) empty - using rank_score proxy for ticket eligibility. "
            "Fix Soccer step5 line-hit output when possible."
        )

    if "edge" not in df.columns:
        df["edge"] = np.nan
    df["edge"] = pd.to_numeric(df["edge"], errors="coerce")
    if "abs_edge" not in df.columns:
        df["abs_edge"] = np.nan
    df["abs_edge"] = pd.to_numeric(df["abs_edge"], errors="coerce")
    df["abs_edge"] = df["abs_edge"].where(df["abs_edge"].notna(), df["edge"].abs())

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    df = df[df["line"].notna() & (df["line"] >= 0)]
    df = _apply_l5_truth_from_stat_games(df, "NBA1H")
    df = df.astype(object).where(df.notna(), other=None)
    return df


def load_mlb(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_mlb_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "MLB" if "MLB" in xl.sheet_names else (
        "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        # title-case (from step8 clean xlsx)
        "Player":           "player",
        "Tier":             "tier",
        "Rank Score":       "rank_score",
        "Pos":              "pos",
        "Team":             "team",
        "Opp":              "opp",
        "Game Time":        "game_time",
        "Prop":             "prop_type",
        "Pick Type":        "pick_type",
        "Line":             "line",
        "Direction":        "direction",
        "Edge":             "edge",
        "Abs Edge":         "abs_edge",
        "Projection":       "projection",
        "ESPN ID":          "espn_player_id",
        "Hit Rate (5g)":    "hit_rate",
        "Hit Rate Status":  "hit_rate_status",
        "Reliability Note": "reliability_note",
        # Kept separate so we can coalesce into hit_rate when 5g is blank (common when
        # Soccer step5/7 line-hit columns aren't populated yet).
        "Hit Rate (10g)":   "_soccer_hit10",
        "Last 5 Avg":       "l5_avg",
        "Season Avg":       "season_avg",
        "L5 Over":          "l5_over",
        "L5 Under":         "l5_under",
        "Def Rank":         "def_rank",
        "Def Tier":         "def_tier",
        "Min Tier":         "min_tier",
        "Shot Role":        "shot_role",
        "Usage Role":       "usage_role",
        "League":           "league",
        "Pos Group":        "position_group",
        "Void Reason":      "void_reason",
        # snake_case fallbacks
        "player_name":        "player",
        "stat_type":          "prop_type",
        "stat_norm":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "bet_direction":      "direction",
        "final_bet_direction": "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
        "opponent":           "opp",
        "line_hit_rate_over_ou_5":  "hit_rate",
        "line_hit_rate_over_ou_10": "_soccer_hit10",
        "Game Script Mult": "game_script_mult",
        "Game Script Note": "game_script_note",
        "game_script_mult": "game_script_mult",
        "game_script_note": "game_script_note",
        "Open Line": "open_line",
        "Line Movement": "line_movement",
        "Line Shift": "line_direction_shift",
        "open_line": "open_line",
        "line_movement": "line_movement",
        "line_direction_shift": "line_direction_shift",
    })

    if "opp" not in df.columns:
        df["opp"] = ""

    # Step8 CSV exports often populate opp_team while "Opp" is blank.
    if "opp_team" in df.columns:
        om = df["opp"].astype(str).str.strip()
        om = om.mask(om.str.lower().isin(["nan", "none"]), "")
        ot = df["opp_team"].astype(str).str.strip()
        ot = ot.mask(ot.str.lower().isin(["nan", "none"]), "")
        df["opp"] = np.where(om.ne(""), om, ot)

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "MLB"

    # Ensure ml_prob is numeric and always present for downstream leg scoring.
    if "ml_prob" not in df.columns:
        df["ml_prob"] = np.nan
    df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    if "hit_rate" not in df.columns:
        df["hit_rate"] = np.nan

    for col in ["rank_score", "hit_rate", "line", "_soccer_hit10"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Prefer L5 hit rate, then L10, when either is present.
    if "_soccer_hit10" in df.columns:
        df["hit_rate"] = df["hit_rate"].combine_first(df["_soccer_hit10"])
        df.drop(columns=["_soccer_hit10"], inplace=True)

    # Normalize hit_rate to 0–1 (handles "62%" or 62.0 from spreadsheets)
    if "hit_rate" in df.columns and df["hit_rate"].notna().any():
        hr = df["hit_rate"]
        if hr.dtype == object:
            hr = hr.astype(str).str.replace("%", "", regex=False).str.strip()
            hr = pd.to_numeric(hr, errors="coerce")
        else:
            hr = pd.to_numeric(hr, errors="coerce")
        if hr.dropna().max() is not None and hr.dropna().max() > 1.5:
            hr = hr / 100.0
        df["hit_rate"] = hr

    # Reconcile L5 Over/Under (and directional hit%) with stat_g* vs line when step8 carries
    # rolling games — avoids published slates disagreeing with game-log charts.
    df = _apply_l5_truth_from_stat_games(df, "MLB", min_stat_games=3)

    # NBA split boards can show extreme 5/5 streaks; shrink tiny-window rates so UI/tickets
    # do not overstate confidence (e.g., 100% from only 5 samples).
    l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
    l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
    l5n = l5o.add(l5u, fill_value=0)
    hits = np.where(
        df["direction"].astype(str).str.upper().eq("UNDER").to_numpy(),
        l5u.to_numpy(dtype=float, copy=True),
        l5o.to_numpy(dtype=float, copy=True),
    )
    # Strong prior toward 55% over a 15-game pseudo sample.
    prior_p = 0.55
    prior_n = 15.0
    shrunk = (hits + prior_p * prior_n) / (l5n.to_numpy(dtype=float, copy=True) + prior_n)
    hr_now = pd.to_numeric(df["hit_rate"], errors="coerce")
    use_shrink = (l5n > 0) & (l5n <= 10) & np.isfinite(shrunk)
    df.loc[use_shrink, "hit_rate"] = shrunk[use_shrink.to_numpy()]
    df["hit_rate"] = pd.to_numeric(df["hit_rate"], errors="coerce").clip(lower=0.35, upper=0.92)

    # Still no usable hit rate (common on current Soccer pipeline when game logs are empty).
    # Use a mild rank_score-based proxy so tier/rank ticket gates still run; re-run step5 when HRs exist.
    hr_series = pd.to_numeric(df["hit_rate"], errors="coerce")
    if hr_series.notna().sum() == 0:
        rs = pd.to_numeric(df.get("rank_score", 0), errors="coerce").fillna(0.0)
        q25, q75 = float(rs.quantile(0.25)), float(rs.quantile(0.75))
        span = (q75 - q25) + 1e-6
        proxy = 0.54 + ((rs - q25) / span).clip(lower=0.0, upper=1.0) * 0.12
        df["hit_rate"] = proxy.clip(0.50, 0.68)
        print(
            "  [load_mlb] NOTE: Hit Rate (5g)/(10g) empty - using rank_score proxy for ticket eligibility. "
            "Fix Soccer step5 line-hit output when possible."
        )

    # Sample size for L5 (after stat_g reconciliation and shrink).
    l5o_g = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
    l5u_g = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
    _l5sum = l5o_g.add(l5u_g, fill_value=0)
    df["l5_games"] = _l5sum.where(_l5sum > 0, np.nan)

    if "edge" not in df.columns:
        df["edge"] = np.nan
    df["edge"] = pd.to_numeric(df["edge"], errors="coerce")
    if "abs_edge" not in df.columns:
        df["abs_edge"] = np.nan
    df["abs_edge"] = pd.to_numeric(df["abs_edge"], errors="coerce")
    df["abs_edge"] = df["abs_edge"].where(df["abs_edge"].notna(), df["edge"].abs())

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    df = df[df["line"].notna() & (df["line"] >= 0)]
    df = df.astype(object).where(df.notna(), other=None)
    return df


def load_nba1q(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_nba1q_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "NBA1Q" if "NBA1Q" in xl.sheet_names else (
        "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        # title-case (from step8 clean xlsx)
        "Player":           "player",
        "Tier":             "tier",
        "Rank Score":       "rank_score",
        "Pos":              "pos",
        "Team":             "team",
        "Opp":              "opp",
        "Game Time":        "game_time",
        "Prop":             "prop_type",
        "Pick Type":        "pick_type",
        "Line":             "line",
        "Direction":        "direction",
        "Edge":             "edge",
        "Abs Edge":         "abs_edge",
        "Projection":       "projection",
        "ESPN ID":          "espn_player_id",
        "ML Prob":          "ml_prob",
        "Hit Rate (5g)":    "hit_rate",
        "Hit Rate Status":  "hit_rate_status",
        "Reliability Note": "reliability_note",
        # Kept separate so we can coalesce into hit_rate when 5g is blank (common when
        # Soccer step5/7 line-hit columns aren't populated yet).
        "Hit Rate (10g)":   "_soccer_hit10",
        "Last 5 Avg":       "l5_avg",
        "Season Avg":       "season_avg",
        "L5 Over":          "l5_over",
        "L5 Under":         "l5_under",
        "Def Rank":         "def_rank",
        "Def Tier":         "def_tier",
        "Min Tier":         "min_tier",
        "Shot Role":        "shot_role",
        "Usage Role":       "usage_role",
        "League":           "league",
        "Pos Group":        "position_group",
        "Void Reason":      "void_reason",
        # snake_case fallbacks
        "player_name":        "player",
        "stat_type":          "prop_type",
        "stat_norm":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
        "opponent":           "opp",
        "line_hit_rate_over_ou_5":  "hit_rate",
        "line_hit_rate_over_ou_10": "_soccer_hit10",
        "Game Script Mult": "game_script_mult",
        "Game Script Note": "game_script_note",
        "game_script_mult": "game_script_mult",
        "game_script_note": "game_script_note",
    })

    if "opp" not in df.columns:
        df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "NBA1Q"

    # Ensure ml_prob is numeric and always present for downstream leg scoring.
    if "ml_prob" not in df.columns:
        df["ml_prob"] = np.nan
    df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")

    # Derive L5 game sample size from over+under counts when available.
    if "l5_games" not in df.columns:
        l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
        l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
        derived = l5o.add(l5u, fill_value=0)
        df["l5_games"] = derived.where(derived > 0, np.nan)

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    if "hit_rate" not in df.columns:
        df["hit_rate"] = np.nan

    for col in ["rank_score", "hit_rate", "line", "_soccer_hit10"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Prefer L5 hit rate, then L10, when either is present.
    if "_soccer_hit10" in df.columns:
        df["hit_rate"] = df["hit_rate"].combine_first(df["_soccer_hit10"])
        df.drop(columns=["_soccer_hit10"], inplace=True)

    # Normalize hit_rate to 0–1 (handles "62%" or 62.0 from spreadsheets)
    if "hit_rate" in df.columns and df["hit_rate"].notna().any():
        hr = df["hit_rate"]
        if hr.dtype == object:
            hr = hr.astype(str).str.replace("%", "", regex=False).str.strip()
            hr = pd.to_numeric(hr, errors="coerce")
        else:
            hr = pd.to_numeric(hr, errors="coerce")
        if hr.dropna().max() is not None and hr.dropna().max() > 1.5:
            hr = hr / 100.0
        df["hit_rate"] = hr

    # NBA1Q: shrink tiny-window streak rates (e.g., 5/5) toward a prior.
    l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
    l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
    l5n = l5o.add(l5u, fill_value=0)
    hits = np.where(
        df["direction"].astype(str).str.upper().eq("UNDER").to_numpy(),
        l5u.to_numpy(dtype=float, copy=True),
        l5o.to_numpy(dtype=float, copy=True),
    )
    prior_p = 0.55
    prior_n = 15.0
    shrunk = (hits + prior_p * prior_n) / (l5n.to_numpy(dtype=float, copy=True) + prior_n)
    use_shrink = (l5n > 0) & (l5n <= 10) & np.isfinite(shrunk)
    df.loc[use_shrink, "hit_rate"] = shrunk[use_shrink.to_numpy()]
    df["hit_rate"] = pd.to_numeric(df["hit_rate"], errors="coerce").clip(lower=0.35, upper=0.92)

    # Still no usable hit rate (common on current Soccer pipeline when game logs are empty).
    # Use a mild rank_score-based proxy so tier/rank ticket gates still run; re-run step5 when HRs exist.
    hr_series = pd.to_numeric(df["hit_rate"], errors="coerce")
    if hr_series.notna().sum() == 0:
        rs = pd.to_numeric(df.get("rank_score", 0), errors="coerce").fillna(0.0)
        q25, q75 = float(rs.quantile(0.25)), float(rs.quantile(0.75))
        span = (q75 - q25) + 1e-6
        proxy = 0.54 + ((rs - q25) / span).clip(lower=0.0, upper=1.0) * 0.12
        df["hit_rate"] = proxy.clip(0.50, 0.68)
        print(
            "  [load_nba1q] NOTE: Hit Rate (5g)/(10g) empty - using rank_score proxy for ticket eligibility. "
            "Fix Soccer step5 line-hit output when possible."
        )

    if "edge" not in df.columns:
        df["edge"] = np.nan
    df["edge"] = pd.to_numeric(df["edge"], errors="coerce")
    if "abs_edge" not in df.columns:
        df["abs_edge"] = np.nan
    df["abs_edge"] = pd.to_numeric(df["abs_edge"], errors="coerce")
    df["abs_edge"] = df["abs_edge"].where(df["abs_edge"].notna(), df["edge"].abs())

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    df = df[df["line"].notna() & (df["line"] >= 0)]
    df = df.astype(object).where(df.notna(), other=None)
    return df


def load_nba1h(path: str) -> pd.DataFrame:
    path = resolve_input_path(path, fallback_filename="step8_nba1h_direction_clean.xlsx")

    xl = pd.ExcelFile(path, engine="openpyxl")
    sheet = "NBA1H" if "NBA1H" in xl.sheet_names else (
        "ALL" if "ALL" in xl.sheet_names else xl.sheet_names[0])
    df = pd.read_excel(path, sheet_name=sheet, engine="openpyxl")

    df = df.rename(columns={
        # title-case (from step8 clean xlsx)
        "Player":           "player",
        "Tier":             "tier",
        "Rank Score":       "rank_score",
        "Pos":              "pos",
        "Team":             "team",
        "Opp":              "opp",
        "Game Time":        "game_time",
        "Prop":             "prop_type",
        "Pick Type":        "pick_type",
        "Line":             "line",
        "Direction":        "direction",
        "Edge":             "edge",
        "Abs Edge":         "abs_edge",
        "Projection":       "projection",
        "ESPN ID":          "espn_player_id",
        "ML Prob":          "ml_prob",
        "Hit Rate (5g)":    "hit_rate",
        # Kept separate so we can coalesce into hit_rate when 5g is blank (common when
        # Soccer step5/7 line-hit columns aren't populated yet).
        "Hit Rate (10g)":   "_soccer_hit10",
        "Last 5 Avg":       "l5_avg",
        "Season Avg":       "season_avg",
        "L5 Over":          "l5_over",
        "L5 Under":         "l5_under",
        "Def Rank":         "def_rank",
        "Def Tier":         "def_tier",
        "Min Tier":         "min_tier",
        "Shot Role":        "shot_role",
        "Usage Role":       "usage_role",
        "League":           "league",
        "Pos Group":        "position_group",
        "Void Reason":      "void_reason",
        # snake_case fallbacks
        "player_name":        "player",
        "stat_type":          "prop_type",
        "stat_norm":          "prop_type",
        "line_score":         "line",
        "recommended_side":   "direction",
        "composite_hit_rate": "hit_rate",
        "avg_L5":             "l5_avg",
        "avg_season":         "season_avg",
        "def_tier":           "def_tier",
        "def_rank":           "def_rank",
        "prop_score":         "rank_score",
        "game_start":         "game_time",
        "opponent":           "opp",
        "line_hit_rate_over_ou_5":  "hit_rate",
        "line_hit_rate_over_ou_10": "_soccer_hit10",
        "Game Script Mult": "game_script_mult",
        "Game Script Note": "game_script_note",
        "game_script_mult": "game_script_mult",
        "game_script_note": "game_script_note",
    })

    if "opp" not in df.columns:
        df["opp"] = ""

    df = df.loc[:, ~df.columns.duplicated()].copy()
    df["sport"] = "NBA1H"

    # Ensure ml_prob is numeric and always present for downstream leg scoring.
    if "ml_prob" not in df.columns:
        df["ml_prob"] = np.nan
    df["ml_prob"] = pd.to_numeric(df["ml_prob"], errors="coerce")

    # Derive L5 game sample size from over+under counts when available.
    if "l5_games" not in df.columns:
        l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
        l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
        derived = l5o.add(l5u, fill_value=0)
        df["l5_games"] = derived.where(derived > 0, np.nan)

    def _norm_pick(x):
        t = str(x).strip().lower() if x else ""
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    if "pick_type" not in df.columns:
        df["pick_type"] = "Standard"
    df["pick_type"] = df["pick_type"].apply(_norm_pick)

    if "direction" in df.columns:
        df["direction"] = df["direction"].astype(str).str.upper()
    else:
        df["direction"] = "OVER"

    if "tier" in df.columns:
        df["tier"] = df["tier"].astype(str).str.upper()
    else:
        df["tier"] = "C"

    if "hit_rate" not in df.columns:
        df["hit_rate"] = np.nan

    for col in ["rank_score", "hit_rate", "line", "_soccer_hit10"]:
        if col in df.columns:
            df[col] = pd.to_numeric(df[col], errors="coerce")

    # Prefer L5 hit rate, then L10, when either is present.
    if "_soccer_hit10" in df.columns:
        df["hit_rate"] = df["hit_rate"].combine_first(df["_soccer_hit10"])
        df.drop(columns=["_soccer_hit10"], inplace=True)

    # Normalize hit_rate to 0–1 (handles "62%" or 62.0 from spreadsheets)
    if "hit_rate" in df.columns and df["hit_rate"].notna().any():
        hr = df["hit_rate"]
        if hr.dtype == object:
            hr = hr.astype(str).str.replace("%", "", regex=False).str.strip()
            hr = pd.to_numeric(hr, errors="coerce")
        else:
            hr = pd.to_numeric(hr, errors="coerce")
        if hr.dropna().max() is not None and hr.dropna().max() > 1.5:
            hr = hr / 100.0
        df["hit_rate"] = hr

    # NBA1H: shrink tiny-window streak rates (e.g., 5/5) toward a prior.
    l5o = pd.to_numeric(df.get("l5_over", np.nan), errors="coerce")
    l5u = pd.to_numeric(df.get("l5_under", np.nan), errors="coerce")
    l5n = l5o.add(l5u, fill_value=0)
    hits = np.where(
        df["direction"].astype(str).str.upper().eq("UNDER").to_numpy(),
        l5u.to_numpy(dtype=float, copy=True),
        l5o.to_numpy(dtype=float, copy=True),
    )
    prior_p = 0.55
    prior_n = 15.0
    shrunk = (hits + prior_p * prior_n) / (l5n.to_numpy(dtype=float, copy=True) + prior_n)
    use_shrink = (l5n > 0) & (l5n <= 10) & np.isfinite(shrunk)
    df.loc[use_shrink, "hit_rate"] = shrunk[use_shrink.to_numpy()]
    df["hit_rate"] = pd.to_numeric(df["hit_rate"], errors="coerce").clip(lower=0.35, upper=0.92)

    # Still no usable hit rate (common on current Soccer pipeline when game logs are empty).
    # Use a mild rank_score-based proxy so tier/rank ticket gates still run; re-run step5 when HRs exist.
    hr_series = pd.to_numeric(df["hit_rate"], errors="coerce")
    if hr_series.notna().sum() == 0:
        rs = pd.to_numeric(df.get("rank_score", 0), errors="coerce").fillna(0.0)
        q25, q75 = float(rs.quantile(0.25)), float(rs.quantile(0.75))
        span = (q75 - q25) + 1e-6
        proxy = 0.54 + ((rs - q25) / span).clip(lower=0.0, upper=1.0) * 0.12
        df["hit_rate"] = proxy.clip(0.50, 0.68)
        print(
            "  [load_nba1h] NOTE: Hit Rate (5g)/(10g) empty - using rank_score proxy for ticket eligibility. "
            "Fix Soccer step5 line-hit output when possible."
        )

    if "edge" not in df.columns:
        df["edge"] = np.nan
    df["edge"] = pd.to_numeric(df["edge"], errors="coerce")
    if "abs_edge" not in df.columns:
        df["abs_edge"] = np.nan
    df["abs_edge"] = pd.to_numeric(df["abs_edge"], errors="coerce")
    df["abs_edge"] = df["abs_edge"].where(df["abs_edge"].notna(), df["edge"].abs())

    if "espn_player_id" in df.columns:
        df["espn_player_id"] = df["espn_player_id"].apply(_clean_id)

    df = df[df["line"].notna() & (df["line"] >= 0)]
    df = df.astype(object).where(df.notna(), other=None)
    return df


def add_l5_play_side_columns(df: pd.DataFrame) -> pd.DataFrame:
    """
    For each row: L5 hits on the recommended side (over vs under vs line) and
    hits / (l5_over + l5_under) when that sample size is known (>0).
    """
    if df is None or df.empty:
        return df
    out = df.copy()
    lo = pd.to_numeric(out.get("l5_over"), errors="coerce")
    lu = pd.to_numeric(out.get("l5_under"), errors="coerce")
    if "direction" not in out.columns:
        out["l5_side_hits"] = np.nan
        out["l5_consistency"] = np.nan
        return out
    d = out["direction"].astype(str).str.strip().str.upper()
    lo_a = lo.to_numpy(dtype=float, copy=True)
    lu_a = lu.to_numpy(dtype=float, copy=True)
    hits = np.select(
        [d.eq("OVER").to_numpy(), d.eq("UNDER").to_numpy()],
        [lo_a, lu_a],
        default=np.nan,
    )
    out["l5_side_hits"] = hits
    denom = lo_a + lu_a
    denom = np.where(denom > 0, denom, np.nan)
    out["l5_consistency"] = hits / denom
    return out


def drop_demon_over_rows(df: Optional[pd.DataFrame], sport_label: str) -> Optional[pd.DataFrame]:
    """
    PrizePicks Demon is a line-hardening pick type and is not offered on the OVER side for
    these boards — upstream step8 can still emit Demon+OVER rows. Remove them so Excel
    slates, Full Slate, ticket pools, and slate_latest.json never surface unbookable legs.
    """
    if df is None or len(df) == 0:
        return df
    if "pick_type" not in df.columns or "direction" not in df.columns:
        return df
    pt = df["pick_type"].astype(str).str.strip().str.upper()
    dr = df["direction"].astype(str).str.strip().str.upper()
    bad = pt.eq("DEMON") & dr.eq("OVER")
    n_drop = int(bad.sum())
    if n_drop:
        print(f"  [bookability] {sport_label}: dropping {n_drop} Demon+OVER row(s) (not a valid PP offering).")
        return df.loc[~bad].copy()
    return df


# ── Merge to full slate ────────────────────────────────────────────────────────
def build_combined_slate(
    nba: pd.DataFrame,
    cbb: pd.DataFrame,
    nhl: pd.DataFrame = None,
    soccer: pd.DataFrame = None,
    tennis: pd.DataFrame = None,
    wnba: pd.DataFrame = None,
    wcbb: pd.DataFrame = None,
    mlb: pd.DataFrame = None,
    nba1q: pd.DataFrame = None,
    nba1h: pd.DataFrame = None,
    nfl: pd.DataFrame = None,
    cfb: pd.DataFrame = None,
) -> pd.DataFrame:
    keep = [
        "sport",
        "tier",
        "rank_score",
        "rank_score_penalized",
        "surface",
        "line_combo",
        "player",
        "team",
        "opp",
        "team_seed",
        "team_region",
        "team_ap_rank",
        "opp_seed",
        "opp_region",
        "opp_ap_rank",
        "ncaa_rank",
        "game_time",
        "fetched_at",
        "game_date",
        "prop_type",
        "pick_type",
        "line",
        "standard_line",
        "direction",
        "edge",
        "abs_edge",
        "projection",
        "hit_rate",
        "ml_prob",
        "l5_avg",
        "season_avg",
        "l5_over",
        "l5_under",
        "l10_over",
        "l10_under",
        "hit_rate_status",
        "reliability_note",
        "def_tier",
        "opponent_def_rank",
        "def_rank",
        "pace_tier",
        "context_score",
        "prop_quality_score",
        "min_tier",
        "shot_role",
        "usage_role",
        "nba_player_id",
        "espn_player_id",
        "league",
        "position_group",
        "distribution_std",
        "distribution_n",
        "G1",
        "G2",
        "G3",
        "G4",
        "G5",
        "G6",
        "G7",
        "G8",
        "G9",
        "G10",
        "stat_g1",
        "stat_g2",
        "stat_g3",
        "stat_g4",
        "stat_g5",
        "stat_g6",
        "stat_g7",
        "stat_g8",
        "stat_g9",
        "stat_g10",
    ]

    def safe_keep(df, cols):
        df = df.loc[:, ~df.columns.duplicated()].copy()
        return df.reindex(columns=cols).copy()

    def _prep_for_combined(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or len(df) == 0:
            return df
        return safe_keep(_ensure_stat_g_columns(df), keep)

    frames = [_prep_for_combined(nba), _prep_for_combined(cbb)]
    if nhl is not None and len(nhl) > 0:
        frames.append(_prep_for_combined(nhl))
    if soccer is not None and len(soccer) > 0:
        frames.append(_prep_for_combined(soccer))
    if tennis is not None and len(tennis) > 0:
        frames.append(_prep_for_combined(tennis))
    if wnba is not None and len(wnba) > 0:
        frames.append(_prep_for_combined(wnba))
    if wcbb is not None and len(wcbb) > 0:
        frames.append(_prep_for_combined(wcbb))
    if mlb is not None and len(mlb) > 0:
        frames.append(_prep_for_combined(mlb))
    if nba1q is not None and len(nba1q) > 0:
        frames.append(_prep_for_combined(nba1q))
    if nba1h is not None and len(nba1h) > 0:
        frames.append(_prep_for_combined(nba1h))
    if nfl is not None and len(nfl) > 0:
        frames.append(_prep_for_combined(nfl))
    if cfb is not None and len(cfb) > 0:
        frames.append(_prep_for_combined(cfb))
    combined = pd.concat(frames, ignore_index=True)

    if "rank_score" in combined.columns:
        combined["rank_score"] = pd.to_numeric(combined["rank_score"], errors="coerce")
    if "hit_rate" in combined.columns:
        combined["hit_rate"] = pd.to_numeric(combined["hit_rate"], errors="coerce")
    if "ml_prob" in combined.columns:
        combined["ml_prob"] = pd.to_numeric(combined["ml_prob"], errors="coerce")
    if "edge" in combined.columns:
        combined["edge"] = pd.to_numeric(combined["edge"], errors="coerce")
    if "abs_edge" in combined.columns:
        combined["abs_edge"] = pd.to_numeric(combined["abs_edge"], errors="coerce")

    combined = add_l5_play_side_columns(combined)
    combined = add_prop_quality_score(combined)

    combined = combined.sort_values("rank_score", ascending=False, na_position="last").reset_index(drop=True)
    return combined


def _edge_magnitude_series(df: pd.DataFrame) -> pd.Series:
    """Use abs_edge when present so UNDER legs are not dropped by min_edge filters."""
    if df is None or len(df) == 0:
        return pd.Series(dtype=float)
    if "abs_edge" in df.columns:
        ae = pd.to_numeric(df["abs_edge"], errors="coerce")
        if ae.notna().any():
            return ae
    return pd.to_numeric(df.get("edge", np.nan), errors="coerce").abs()


def _directional_edge_series(df: pd.DataFrame) -> pd.Series:
    """
    Direction-aware edge score where larger is better for both sides:
    - OVER:  uses +edge
    - UNDER: uses -edge
    """
    if df is None or len(df) == 0:
        return pd.Series(dtype=float)
    edge = pd.to_numeric(df.get("edge", np.nan), errors="coerce")
    direction = df.get("direction", pd.Series("", index=df.index)).astype(str).str.upper().str.strip()
    return pd.Series(np.where(direction.eq("UNDER"), -edge, edge), index=df.index)


PROP_QUALITY_MIN_BY_SPORT: dict[str, dict[str, float] | float] = {
    "NBA": {"over": 0.60, "under": 0.35},
    "NBA1Q": {"over": 0.60, "under": 0.35},
    "NBA1H": {"over": 0.60, "under": 0.35},
    "CBB": {"over": 0.60, "under": 0.35},
    "WCBB": {"over": 0.60, "under": 0.35},
    "MLB": {"over": 0.35, "under": 0.35},
    "NHL": {"over": 0.35, "under": 0.35},
    "SOCCER": {"over": 0.35, "under": 0.35},
    "SOC": {"over": 0.35, "under": 0.35},
    "TENNIS": {"over": 0.35, "under": 0.35},
}


def _prop_quality_floor_over_under(sport: str) -> tuple[float, float]:
    spec = PROP_QUALITY_MIN_BY_SPORT.get(str(sport).strip().upper(), 0.58)
    if isinstance(spec, dict):
        o = float(spec.get("over", 0.58))
        u = float(spec.get("under", spec.get("over", 0.58)))
        return o, u
    fv = float(spec)
    return fv, fv


def _prop_quality_min_threshold_series(df: pd.DataFrame, min_prop_quality: float) -> pd.Series:
    if df is None or len(df) == 0:
        return pd.Series(dtype=float)
    if min_prop_quality >= 0:
        return pd.Series(float(min_prop_quality), index=df.index)
    sp = df.get("sport", pd.Series("", index=df.index)).astype(str).str.upper().str.strip()
    d = df.get("direction", pd.Series("", index=df.index)).astype(str).str.upper().str.strip()
    is_under = d.isin({"UNDER", "LOWER"})
    over_map = {k: _prop_quality_floor_over_under(k)[0] for k in sp.unique()}
    under_map = {k: _prop_quality_floor_over_under(k)[1] for k in sp.unique()}
    ov_th = sp.map(over_map).fillna(0.58)
    un_th = sp.map(under_map).fillna(0.58)
    return pd.Series(np.where(is_under, un_th, ov_th), index=df.index, dtype=float)


def add_prop_quality_score(df: pd.DataFrame) -> pd.DataFrame:
    """Attach prop_quality_score in [0,1] from core leg quality signals."""
    if df is None or len(df) == 0:
        return df
    out = df.copy()

    direction_u = (
        out.get("direction", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()
    )
    is_under_side = direction_u.isin({"UNDER", "LOWER"})

    hr = pd.to_numeric(out.get("hit_rate"), errors="coerce")
    hr = hr.apply(_to_prob_0_1)

    l5o = pd.to_numeric(out.get("l5_over"), errors="coerce").fillna(0.0)
    l5u = pd.to_numeric(out.get("l5_under"), errors="coerce").fillna(0.0)
    l5_side = np.maximum(l5o, l5u)
    l5_side_rate = np.clip(l5_side / 5.0, 0.0, 1.0)
    hr = hr.fillna(pd.Series(l5_side_rate, index=out.index))
    # Direction-aware: sheet hit_rate is OVER-side rate; UNDER legs use implied under-side rate.
    hr_eff = pd.Series(np.where(is_under_side, 1.0 - hr, hr), index=out.index).clip(0.0, 1.0)
    out["effective_hit_rate"] = hr_eff

    edge_dir = _directional_edge_series(out).fillna(0.0)
    edge_norm = np.clip(np.abs(edge_dir) / 15.0, 0.0, 1.0)

    rs = pd.to_numeric(out.get("rank_score"), errors="coerce")
    rank_prob = rs.apply(lambda x: _rank_score_to_prob(float(x)) if pd.notna(x) else np.nan).fillna(0.5)

    sample_strength = pd.Series(np.clip(l5_side / 5.0, 0.0, 1.0), index=out.index)

    tier_raw = out.get("tier", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()
    tier_norm = tier_raw.map({"A": 1.00, "B": 0.86, "C": 0.70, "D": 0.45}).fillna(0.55)

    score = (
        0.33 * hr_eff
        + 0.22 * edge_norm
        + 0.20 * rank_prob
        + 0.15 * sample_strength
        + 0.10 * tier_norm
    )

    rel = out.get("reliability_note", pd.Series("", index=out.index)).astype(str).str.upper()
    hs = out.get("hit_rate_status", pd.Series("", index=out.index)).astype(str).str.upper()
    score = score - np.where(rel.str.contains("THIN_SAMPLE_", na=False), 0.08, 0.0)
    score = score - np.where(hs.str.startswith("BLENDED_N"), 0.05, 0.0)

    out["prop_quality_score"] = np.clip(score, 0.0, 1.0)
    return out


class DiscardTracker:
    """Track discard counts by sport and reason."""

    def __init__(self) -> None:
        self.records: defaultdict[str, defaultdict[str, int]] = defaultdict(lambda: defaultdict(int))

    def log_df(self, df_slice: pd.DataFrame | None, reason: str, *, sport_col: str = "sport", default_sport: str = "ALL") -> None:
        if df_slice is None or len(df_slice) == 0:
            return
        if sport_col in df_slice.columns:
            for sport, grp in df_slice.groupby(sport_col, dropna=False):
                key = str(sport).strip().upper() if str(sport).strip() else default_sport
                self.records[key][reason] += int(len(grp))
        else:
            self.records[default_sport][reason] += int(len(df_slice))

    def log_count(self, sport: str, reason: str, count: int) -> None:
        n = int(count)
        if n <= 0:
            return
        key = str(sport).strip().upper() if str(sport).strip() else "ALL"
        self.records[key][reason] += n

    def report(self) -> str:
        lines: list[str] = []
        for sport in sorted(self.records.keys()):
            lines.append(f"\n  [{sport}]")
            sport_total = int(sum(self.records[sport].values()))
            rows = sorted(self.records[sport].items(), key=lambda kv: (-kv[1], kv[0]))
            for reason, count in rows:
                pct = (100.0 * float(count) / float(sport_total)) if sport_total > 0 else 0.0
                lines.append(f"    {reason:<36} {count:>5} ({pct:>4.0f}%)")
        return "\n".join(lines)

    def to_dict(self) -> dict[str, dict[str, int]]:
        return {sport: dict(reasons) for sport, reasons in self.records.items()}


class FunnelTracker:
    """Track survivor counts by stage and sport."""

    def __init__(self) -> None:
        self.stage_order: list[str] = [
            "input",
            "after_prop_ban",
            "after_fantasy_score",
            "after_directional_l5_hr",
            "after_pick_type",
            "after_pick_type_eligible",
            "after_void_reason",
            "after_prop_quality",
        ]
        self.stages: list[str] = []
        self.snapshots: dict[str, defaultdict[str, int]] = {}

    def checkpoint_df(
        self,
        label: str,
        df: pd.DataFrame | None,
        *,
        sport_col: str = "sport",
        default_sport: str = "ALL",
    ) -> None:
        if label not in self.snapshots:
            self.snapshots[label] = defaultdict(int)
            self.stages.append(label)
        if df is None:
            return
        n_all = int(len(df))
        self.snapshots[label]["ALL"] += n_all
        if n_all == 0:
            key = str(default_sport).strip().upper() if str(default_sport).strip() else "ALL"
            # Preserve explicit 0 checkpoint so funnel reports true stage drop to zero
            # instead of carrying forward prior-stage counts for this sport.
            self.snapshots[label][key] = int(self.snapshots[label].get(key, 0))
            return
        if sport_col in df.columns:
            for sport, grp in df.groupby(sport_col, dropna=False):
                key = str(sport).strip().upper() if str(sport).strip() else default_sport
                self.snapshots[label][key] += int(len(grp))
        else:
            key = str(default_sport).strip().upper() if str(default_sport).strip() else "ALL"
            self.snapshots[label][key] += n_all

    def report(self) -> str:
        if not self.stages:
            return "  (no funnel checkpoints recorded)"
        sports = sorted({s for snap in self.snapshots.values() for s in snap.keys() if s != "ALL"})
        rows: list[str] = []
        for sport in ["ALL", *sports]:
            rows.append(f"\n  [{sport}]")
            first_stage = "input" if "input" in self.snapshots else self.stages[0]
            base = int(self.snapshots[first_stage].get(sport, 0))
            prev = None
            for i, stage in enumerate(self.stage_order):
                if stage not in self.snapshots:
                    continue
                snap = self.snapshots.get(stage, {})
                if sport in snap:
                    cur_raw = int(snap.get(sport, 0))
                    cur = cur_raw if prev is None else min(cur_raw, int(prev))
                else:
                    # If a sport has no rows at a stage, carry forward prior value.
                    cur = int(prev) if prev is not None else int(base)
                if i == 0:
                    rows.append(f"    {'INPUT':<34} {cur:>5}")
                else:
                    drop = max(0, int(prev or 0) - int(cur))
                    pct = (100.0 * float(cur) / float(base)) if base > 0 else 0.0
                    drop_txt = f"(-{drop})" if drop > 0 else "(+0)"
                    rows.append(f"    {stage:<34} {cur:>5}  {drop_txt:>6}  [{pct:>3.0f}% of input]")
                prev = cur
        return "\n".join(rows)

    def to_dict(self) -> dict[str, Any]:
        return {
            "stages": list(self.stages),
            "snapshots": {k: dict(v) for k, v in self.snapshots.items()},
        }


# ── Filter eligible props for tickets ─────────────────────────────────────────
def filter_eligible(
    df: pd.DataFrame,
    min_hit_rate=0.55,
    min_edge=0.0,
    min_rank=None,
    tiers=None,
    pick_types=None,
    *,
    allow_strong_l5_bypass: bool = True,
    min_prop_quality: float = -1.0,
    discard_tracker: DiscardTracker | None = None,
    funnel_tracker: FunnelTracker | None = None,
    discard_sport: str = "",
):
    def _norm_pick_type(v: Any) -> str:
        s = str(v or "").strip().upper()
        if not s:
            return "STANDARD"
        if "GOBLIN" in s:
            return "GOBLIN"
        if "DEMON" in s:
            return "DEMON"
        return "STANDARD"

    def _norm_direction(v: Any) -> str:
        s = str(v or "").strip().upper()
        if s in {"UNDER", "LOWER"}:
            return "UNDER"
        return "OVER"

    def _passes_picktype_direction_tier_policy(row: pd.Series) -> bool:
        sport_u = str(row.get("sport", "") or "").strip().upper()
        if sport_u not in TIER_POLICY_SPORTS:
            return True
        pick_u = _norm_pick_type(row.get("pick_type", row.get("Pick Type", "")))
        dir_u = _norm_direction(row.get("direction", row.get("bet_direction", row.get("direction_used", ""))))
        tier_u = str(row.get("tier", row.get("Tier", "")) or "").strip().upper()
        allowed = ELIGIBLE_TIERS_BY_PICK_TYPE_DIRECTION.get((pick_u, dir_u))
        if allowed is None:
            return True
        return tier_u in allowed

    def apply_directional_l5_hr(
        in_df: pd.DataFrame,
        thresholds: dict[str, dict[str, float]],
        default: dict[str, float],
    ) -> pd.DataFrame:
        if in_df is None or len(in_df) == 0:
            return pd.DataFrame(columns=getattr(in_df, "columns", []))
        result: list[pd.DataFrame] = []
        sport_series = in_df.get("sport", pd.Series([""] * len(in_df), index=in_df.index)).astype(str).str.upper().str.strip()
        for sport, grp in in_df.groupby(sport_series):
            t = thresholds.get(str(sport).upper(), default)
            # Tennis/WNBA boards often have sparse/placeholder L5 directional windows.
            # Keep them in the candidate pool and rely on downstream ranking/EV gates.
            if str(sport).upper() in {"TENNIS", "WNBA"}:
                result.append(grp)
                continue
            l5_over = pd.to_numeric(grp.get("l5_over"), errors="coerce").fillna(0)
            l5_under = pd.to_numeric(grp.get("l5_under"), errors="coerce").fillna(0)
            hit_rate = pd.to_numeric(grp.get("hit_rate"), errors="coerce")
            dir_s = grp.get("direction", pd.Series([""] * len(grp), index=grp.index)).astype(str).str.upper().str.strip()
            over_dir = dir_s.isin({"OVER", "HIGHER"})
            under_dir = dir_s.isin({"UNDER", "LOWER"})
            over_pass = over_dir & (l5_over >= 4) & (hit_rate >= float(t.get("over", default["over"])))
            under_pass = under_dir & (l5_under >= 4) & (hit_rate <= float(t.get("under", default["under"])))
            passed = grp[over_pass | under_pass]
            if len(passed) > 0:
                result.append(passed)
        if not result:
            return in_df.iloc[0:0].copy()
        return pd.concat(result, axis=0)

    df = add_prop_quality_score(df)
    if "pick_type_eligible" not in df.columns:
        df = enrich_read_fields_dataframe(df)
    mask = pd.Series([True] * len(df), index=df.index)
    sport_hint = str(discard_sport or "").strip().upper() or "ALL"
    if funnel_tracker is not None:
        funnel_tracker.checkpoint_df("input", df, default_sport=sport_hint)

    def _apply_gate(cond: pd.Series, reason: str, stage: str) -> None:
        nonlocal mask
        cond = cond.fillna(False)
        dropped = df[mask & ~cond]
        if discard_tracker is not None and len(dropped) > 0:
            if "sport" in dropped.columns:
                discard_tracker.log_df(dropped, reason)
            else:
                discard_tracker.log_count(sport_hint, reason, int(len(dropped)))
        mask &= cond
        if funnel_tracker is not None:
            funnel_tracker.checkpoint_df(stage, df[mask], default_sport=sport_hint)

    if "prop_type" in df.columns:
        prop_norm = df["prop_type"].apply(_norm_prop_label)
        _apply_gate(~prop_norm.isin(TICKET_EXCLUDED_PROPS), "prop_banned", "after_prop_ban")
    if "stat_coverage" in df.columns:
        cov = df["stat_coverage"].astype(str).str.strip().str.lower()
        _apply_gate(~cov.eq("unsupported"), "stat_coverage_unsupported", "after_stat_coverage")
    _apply_gate(~_fantasy_prop_mask(df), "fantasy_score_excluded", "after_fantasy_score")
    before_dir = df[mask].copy()
    after_dir = apply_directional_l5_hr(before_dir, DIRECTIONAL_HR_THRESHOLDS, DEFAULT_DIRECTIONAL_THRESHOLD)
    dir_cond = pd.Series(df.index.isin(after_dir.index), index=df.index)
    _apply_gate(dir_cond, "directional_l5_hr_fail", "after_directional_l5_hr")
    if {"sport", "pick_type", "direction", "edge"}.issubset(df.columns):
        sport_s = df["sport"].astype(str).str.upper().str.strip()
        pick_s = df["pick_type"].astype(str).str.upper().str.strip()
        dir_s = df["direction"].astype(str).str.upper().str.strip()
        edge_s = pd.to_numeric(df["edge"], errors="coerce")
        effective_edge = pd.Series(
            np.where(dir_s.isin({"UNDER", "LOWER"}), -edge_s, edge_s),
            index=df.index,
        )

        over_floor = sport_s.map(
            lambda sp: DIRECTIONAL_HR_THRESHOLDS.get(str(sp).upper(), {}).get("standard_over_min_edge")
        )
        cond_std_over = pd.Series([True] * len(df), index=df.index)
        std_over_mask = pick_s.eq("STANDARD") & dir_s.isin({"OVER", "HIGHER"}) & over_floor.notna()
        cond_std_over.loc[std_over_mask] = effective_edge.loc[std_over_mask] >= pd.to_numeric(
            over_floor.loc[std_over_mask], errors="coerce"
        )
        _apply_gate(cond_std_over, "standard_over_edge_below_floor", "after_standard_over_edge_floor")

        # Soccer rank_score gate — applied to STANDARD OVER when rank_score is present
        if "rank_score" in df.columns:
            rs_floor = sport_s.map(
                lambda sp: DIRECTIONAL_HR_THRESHOLDS.get(str(sp).upper(), {}).get("standard_over_min_rank_score")
            )
            rs_val = pd.to_numeric(df["rank_score"], errors="coerce")
            cond_rs_over = pd.Series([True] * len(df), index=df.index)
            rs_over_mask = pick_s.eq("STANDARD") & dir_s.isin({"OVER", "HIGHER"}) & rs_floor.notna() & rs_val.notna()
            cond_rs_over.loc[rs_over_mask] = rs_val.loc[rs_over_mask] >= pd.to_numeric(
                rs_floor.loc[rs_over_mask], errors="coerce"
            )
            _apply_gate(cond_rs_over, "standard_over_rank_score_below_floor", "after_standard_over_rank_score_floor")

        under_floor = sport_s.map(
            lambda sp: DIRECTIONAL_HR_THRESHOLDS.get(str(sp).upper(), {}).get("standard_under_min_edge")
        )
        cond_std_under = pd.Series([True] * len(df), index=df.index)
        std_under_mask = pick_s.eq("STANDARD") & dir_s.isin({"UNDER", "LOWER"}) & under_floor.notna()
        cond_std_under.loc[std_under_mask] = effective_edge.loc[std_under_mask] >= pd.to_numeric(
            under_floor.loc[std_under_mask], errors="coerce"
        )
        _apply_gate(cond_std_under, "standard_under_edge_below_floor", "after_standard_under_edge_floor")
    policy_cond = df.apply(_passes_picktype_direction_tier_policy, axis=1)
    _apply_gate(policy_cond, "picktype_direction_tier_policy_fail", "after_picktype_direction_tier_policy")
    if pick_types and "pick_type" in df.columns:
        _apply_gate(df["pick_type"].isin(pick_types), "pick_type_not_allowed", "after_pick_type")
    if "pick_type_eligible" in df.columns:
        elig = df["pick_type_eligible"].fillna(True).astype(bool)
        _apply_gate(elig, "pick_type_not_eligible", "after_pick_type_eligible")
    if "void_reason" in df.columns:
        vs = df["void_reason"]
        void_str = vs.astype(str).str.strip()
        _apply_gate(~void_str.eq("NO_PROJECTION_OR_LINE"), "no_projection_or_line", "after_void_reason")
    if "prop_quality_score" in df.columns:
        pq = pd.to_numeric(df["prop_quality_score"], errors="coerce").fillna(0.0)
        pq_min = _prop_quality_min_threshold_series(df, float(min_prop_quality))
        _apply_gate(pq >= pq_min, "prop_quality_below_floor", "after_prop_quality")
    out = df[mask].copy()
    if len(out) > 0:
        out["effective_edge"] = _directional_edge_series(out)
        edge_s = pd.to_numeric(out["effective_edge"], errors="coerce").fillna(-1e9)
        def _pool_sort_series(col: str) -> pd.Series:
            if col not in out.columns:
                return pd.Series(np.nan, index=out.index, dtype=float)
            val = out[col]
            if isinstance(val, pd.DataFrame):
                val = val.iloc[:, 0]
            return pd.to_numeric(val, errors="coerce")

        win_s = _pool_sort_series("win_probability")
        if not win_s.notna().any():
            win_s = _pool_sort_series("ml_prob")
        if not win_s.notna().any():
            win_s = _pool_sort_series("hit_rate")
        win_s = win_s.fillna(0.0)
        out = out.assign(_edge_sort=edge_s, _win_sort=win_s).sort_values(
            ["_edge_sort", "_win_sort"], ascending=[False, False]
        )
        out = out.drop(columns=["_edge_sort", "_win_sort"], errors="ignore")
    return out


# Per-sport ticket structures: n_legs, pick pool (goblin vs standard), and
# direction/sort flow (power vs flex vs standard 2-leg).
_STRUCTURE_SPECS: dict[str, dict[str, object]] = {
    "power": {"n_legs": 2, "pool": "goblin", "flow": "power"},
    "power4": {"n_legs": 4, "pool": "goblin", "flow": "power"},
    "power5": {"n_legs": 5, "pool": "goblin", "flow": "power"},
    "power6": {"n_legs": 6, "pool": "goblin", "flow": "power"},
    "flex": {"n_legs": 3, "pool": "goblin", "flow": "flex"},
    "flex4": {"n_legs": 4, "pool": "goblin", "flow": "flex"},
    "flex5": {"n_legs": 5, "pool": "goblin", "flow": "flex"},
    "flex6": {"n_legs": 6, "pool": "goblin", "flow": "flex"},
    "standard": {"n_legs": 2, "pool": "standard", "flow": "standard"},
    "power_std3": {"n_legs": 3, "pool": "standard", "flow": "power"},
    "goblin3": {"n_legs": 3, "pool": "goblin", "flow": "power"},
}


def build_single_structure_ticket(
    pool_df: pd.DataFrame,
    sport_label: str,
    structure: str,
    counters: dict | None = None,
    relaxed: bool = False,
    min_leg_hit_rate: float | None = None,
    prioritize_ticket_hit: bool = False,
    ticket_sort_mode: str = "rank",
    ticket_gen_starts: int = 10,
) -> dict | None:
    """
    Build exactly one best ticket for a sport+structure.
    With ticket_gen_starts>1, tries several first-leg seeds and keeps the combo that maximizes
    modeled ticket payout (flex cash for flex 3+, else all-hit prob).
    """
    if pool_df is None or pool_df.empty:
        return None

    spec = _STRUCTURE_SPECS.get(structure)
    if not spec:
        return None

    n_legs = int(spec["n_legs"])
    pool_kind = str(spec["pool"])
    flow = str(spec["flow"])
    allowed_tiers = {"A", "B", "C", "D"}
    q = 0.70 if flow in ("power", "standard") else 0.50  # top 30% / top 50%

    # NHL, Soccer, and Tennis don't use Goblin/Standard split for tickets — all props behave as Standard.
    # Skip pick_type filtering for these sports so Power/Flex can use Standard props.
    sport_up = sport_label.upper()
    skip_picktype_filter = sport_up in ("NHL", "SOCCER", "SOC", "TENNIS")

    df = pool_df.copy()
    if "pick_type" in df.columns and not skip_picktype_filter:
        pt = df["pick_type"].astype(str).str.strip().str.lower()
        if pool_kind == "standard":
            df = df[pt == "standard"]
        else:
            df = df[pt == "goblin"]
    if "tier" in df.columns and not (relaxed and structure == "standard"):
        df = df[df["tier"].astype(str).str.upper().isin(allowed_tiers)]

    prop_norm = df["prop_type"].apply(_norm_prop_label) if "prop_type" in df.columns else pd.Series([""] * len(df))
    excl_mask = prop_norm.isin(TICKET_EXCLUDED_PROPS) | _fantasy_prop_mask(df)
    if counters is not None:
        fantasy_mask = prop_norm.str.contains("fantasy", na=False)
        counters["fantasy_excluded_count"] += int((excl_mask & fantasy_mask).sum())
        counters["ban_list_filtered_count"] += int((excl_mask & ~fantasy_mask).sum())
    df = df[~excl_mask].copy()

    if flow in ("power", "standard"):
        df_prop_norm = df["prop_type"].apply(_norm_prop_label) if "prop_type" in df.columns else pd.Series([""] * len(df))
        df = df[~df_prop_norm.isin(TIER3_PROPS)].copy()
        # Explicitly block steals from Power Play and Standard 2-leg tickets.
        df_prop_norm = df["prop_type"].apply(_norm_prop_label) if "prop_type" in df.columns else pd.Series([""] * len(df))
        df = df[~df_prop_norm.eq("steals")].copy()

    # Keep generator aligned with ticket_eval render filters so tickets do not
    # collapse into partials during HTML rendering.
    if "direction" in df.columns and "line" in df.columns and "prop_type" in df.columns:
        ddir = df["direction"].astype(str).str.strip().str.upper()
        dprop = df["prop_type"].apply(_norm_prop_label)
        dline = pd.to_numeric(df["line"], errors="coerce")
        line_ok = pd.Series([True] * len(df), index=df.index)
        line_ok &= ~((dprop == "points") & (ddir == "OVER") & (dline < 8.0))
        line_ok &= ~((dprop == "rebounds") & (ddir == "OVER") & (dline < 2.5))
        df = df[line_ok].copy()

    if "rank_score" in df.columns and len(df) > 0 and not (relaxed and structure == "standard"):
        rs = pd.to_numeric(df["rank_score"], errors="coerce")
        cutoff = float(rs.quantile(q))
        df_q = df[rs >= cutoff].copy()
        # Top-quantile cuts collapse tiny pools (NHL/MLB especially) to <2 legs after other gates.
        if sport_up in ("NHL", "MLB", "TENNIS", "SOCCER", "SOC") and len(df_q) < max(n_legs * 2, 6):
            df = df  # keep pre-quantile pool for thin sports
        else:
            df = df_q
    if df.empty:
        return None

    # Direction rules
    dirs = df.get("direction", pd.Series([""] * len(df), index=df.index)).astype(str).str.upper().str.strip()
    over_df = df[dirs == "OVER"].copy()
    under_df = df[dirs == "UNDER"].copy()

    if flow == "standard":
        # Standard: allow both directions; directional edge ranking picks best side.
        cand = pd.concat([over_df, under_df], ignore_index=True)
    elif flow == "power":
        # Power: allow both directions; directional edge ranking picks best side.
        cand = pd.concat([over_df, under_df], ignore_index=True)
    else:
        # Flex: allow UNDER when directional L5 supports it.
        if not under_df.empty:
            l5_u = pd.to_numeric(under_df.get("l5_under", 0), errors="coerce").fillna(0.0)
            under_df = under_df[l5_u >= 4].copy()
        cand = pd.concat([over_df, under_df], ignore_index=True)

    if cand.empty:
        return None

    # Tennis: OVER only for Aces + Games Won; Double Faults (and other props) allow UNDER.
    if sport_up == "TENNIS" and "prop_type" in cand.columns and "direction" in cand.columns:
        pn = cand["prop_type"].apply(_norm_prop_label)
        ddir = cand["direction"].astype(str).str.upper().str.strip()
        ace_games_won = pn.str.contains("ace", na=False) | (
            pn.str.contains("game", na=False) & pn.str.contains("won", na=False) & ~pn.str.contains("set", na=False)
        )
        cand = cand[~(ace_games_won & (ddir == "UNDER"))].copy()

    if cand.empty:
        return None

    # NHL / Tennis hit-rate proxy in ticket builder:
    # pool() may pass strong-L5 candidates even when raw hit_rate is near zero.
    # For structured-ticket leg floors, use directional L10/L5 proxy when hit_rate is mostly zero.
    # Tennis: no PP graded history — usually use blended_score (see load_tennis board proxy too).
    if sport_up == "NHL" and "hit_rate" in cand.columns:
        hr0 = pd.to_numeric(cand["hit_rate"], errors="coerce").fillna(0.0)
        if bool((hr0 <= 0.001).mean() >= 0.60):
            proxy_col = None
            for c in ("hit_rate_over_L10", "hit_rate_over_L5", "hit_rate_over_L20", "over_L10", "over_L5"):
                if c in cand.columns:
                    proxy_col = c
                    break
            if proxy_col is not None:
                proxy = pd.to_numeric(cand[proxy_col], errors="coerce")
                if proxy.dropna().max() > 1.5:
                    proxy = proxy / 100.0
                proxy = proxy.clip(lower=0.52, upper=0.90)
                cand["hit_rate"] = proxy.where(proxy.notna(), hr0)
                print(f"  [NHL GATE TRACE] build_single_structure_ticket hit_rate proxy='{proxy_col}' applied")

    if sport_up == "TENNIS" and "hit_rate" in cand.columns:
        hr0 = pd.to_numeric(cand["hit_rate"], errors="coerce").fillna(0.0)
        if bool((hr0 <= 0.001).mean() >= 0.60):
            proxy_col = None
            for c in ("hit_rate_over_L10", "hit_rate_over_L5", "hit_rate_over_L20", "over_L10", "over_L5"):
                if c in cand.columns:
                    proxy_col = c
                    break
            if proxy_col is not None:
                proxy = pd.to_numeric(cand[proxy_col], errors="coerce")
                if proxy.dropna().max() > 1.5:
                    proxy = proxy / 100.0
                proxy = proxy.clip(lower=0.52, upper=0.90)
                cand["hit_rate"] = proxy.where(proxy.notna(), hr0)
                print(f"  [TENNIS GATE TRACE] build_single_structure_ticket hit_rate proxy='{proxy_col}' applied")
            elif "blended_score" in cand.columns:
                bs = pd.to_numeric(cand["blended_score"], errors="coerce")
                high = bs >= 0.70
                proxy = (bs * 0.65).where(high, (bs * 0.58)).clip(0.52, 0.90)
                m0 = hr0 <= 0.001
                use = m0 & proxy.notna()
                cand.loc[use, "hit_rate"] = proxy.loc[use]
                print("  [TENNIS GATE TRACE] build_single_structure_ticket hit_rate proxy=blended_score applied")

    # No per-leg hit_rate floor in ticket construction.
    if cand.empty:
        return None

    if counters is not None:
        pct_cap = counters.get("player_ticket_counts")
        if pct_cap is not None and len(cand) > 0 and "player" in cand.columns:
            pn = cand["player"].map(_norm_player_join)
            cap_ok = pn.eq("") | pn.map(lambda p: int(pct_cap.get(p, 0)) < MAX_SLIPS_PER_PLAYER)
            cand = cand[cap_ok].copy()
    if cand.empty:
        return None

    if counters is not None:
        counters["total_eligible_count"] += int(len(cand))

    cand = _attach_ticket_pick_order(cand, ticket_sort_mode)
    cand["__dir_edge"] = _directional_edge_series(cand).fillna(0.0)
    bonus = cand["prop_type"].apply(_prop_priority_bonus) if "prop_type" in cand.columns else 0.0
    cand["__score_adj"] = cand["__ts_pri"] + bonus
    if flow == "standard":
        cand = cand.sort_values(
            ["__dir_edge", "__ts_pri", "__ts_sec"], ascending=[False, False, False], na_position="last"
        )
    else:
        cand = cand.sort_values(
            ["__dir_edge", "__score_adj", "__ts_sec"], ascending=[False, False, False], na_position="last"
        )

    tg_starts = max(1, int(ticket_gen_starts))
    row_sets = _collect_row_candidates_for_structure(cand, n_legs, flow, tg_starts, 1)
    if not row_sets:
        return None
    return _finalize_structure_ticket_dict(
        row_sets[0], structure, sport_label, flow, n_legs, counters, prioritize_ticket_hit
    )


def build_structure_ticket_variants(
    pool_df: pd.DataFrame,
    sport_label: str,
    structure: str,
    counters: dict | None = None,
    relaxed: bool = False,
    min_leg_hit_rate: float | None = None,
    prioritize_ticket_hit: bool = False,
    ticket_sort_mode: str = "rank",
    ticket_gen_starts: int = 10,
    max_variants: int = 3,
) -> list[dict]:
    """
    Same filters as build_single_structure_ticket, but return up to max_variants distinct greedy slips
    (by player+prop set), ranked by modeled ticket payout.
    """
    max_variants = max(1, int(max_variants))
    if max_variants <= 1:
        one = build_single_structure_ticket(
            pool_df,
            sport_label,
            structure,
            counters=counters,
            relaxed=relaxed,
            min_leg_hit_rate=min_leg_hit_rate,
            prioritize_ticket_hit=prioritize_ticket_hit,
            ticket_sort_mode=ticket_sort_mode,
            ticket_gen_starts=ticket_gen_starts,
        )
        return [one] if one is not None else []

    if pool_df is None or pool_df.empty:
        return []

    spec = _STRUCTURE_SPECS.get(structure)
    if not spec:
        return []

    n_legs = int(spec["n_legs"])
    pool_kind = str(spec["pool"])
    flow = str(spec["flow"])
    allowed_tiers = {"A", "B", "C", "D"}
    q = 0.70 if flow in ("power", "standard") else 0.50

    sport_up = sport_label.upper()
    skip_picktype_filter = sport_up in ("NHL", "SOCCER", "SOC", "TENNIS")

    df = pool_df.copy()
    if "pick_type" in df.columns and not skip_picktype_filter:
        pt = df["pick_type"].astype(str).str.strip().str.lower()
        if pool_kind == "standard":
            df = df[pt == "standard"]
        else:
            df = df[pt == "goblin"]
    if "tier" in df.columns and not (relaxed and structure == "standard"):
        df = df[df["tier"].astype(str).str.upper().isin(allowed_tiers)]

    prop_norm = df["prop_type"].apply(_norm_prop_label) if "prop_type" in df.columns else pd.Series([""] * len(df))
    excl_mask = prop_norm.isin(TICKET_EXCLUDED_PROPS) | _fantasy_prop_mask(df)
    df = df[~excl_mask].copy()

    if flow in ("power", "standard"):
        df_prop_norm = df["prop_type"].apply(_norm_prop_label) if "prop_type" in df.columns else pd.Series([""] * len(df))
        df = df[~df_prop_norm.isin(TIER3_PROPS)].copy()
        df_prop_norm = df["prop_type"].apply(_norm_prop_label) if "prop_type" in df.columns else pd.Series([""] * len(df))
        df = df[~df_prop_norm.eq("steals")].copy()

    if "direction" in df.columns and "line" in df.columns and "prop_type" in df.columns:
        ddir = df["direction"].astype(str).str.strip().str.upper()
        dprop = df["prop_type"].apply(_norm_prop_label)
        dline = pd.to_numeric(df["line"], errors="coerce")
        line_ok = pd.Series([True] * len(df), index=df.index)
        line_ok &= ~((dprop == "points") & (ddir == "OVER") & (dline < 8.0))
        line_ok &= ~((dprop == "rebounds") & (ddir == "OVER") & (dline < 2.5))
        df = df[line_ok].copy()

    if "rank_score" in df.columns and len(df) > 0 and not (relaxed and structure == "standard"):
        rs = pd.to_numeric(df["rank_score"], errors="coerce")
        cutoff = float(rs.quantile(q))
        df_q = df[rs >= cutoff].copy()
        if sport_up in ("NHL", "MLB", "TENNIS", "SOCCER", "SOC") and len(df_q) < max(n_legs * 2, 6):
            df = df
        else:
            df = df_q
    if df.empty:
        return []

    dirs = df.get("direction", pd.Series([""] * len(df), index=df.index)).astype(str).str.upper().str.strip()
    over_df = df[dirs == "OVER"].copy()
    under_df = df[dirs == "UNDER"].copy()

    if flow == "standard":
        cand = pd.concat([over_df, under_df], ignore_index=True)
    elif flow == "power":
        cand = pd.concat([over_df, under_df], ignore_index=True)
    else:
        if not under_df.empty:
            l5_u = pd.to_numeric(under_df.get("l5_under", 0), errors="coerce").fillna(0.0)
            under_df = under_df[l5_u >= 4].copy()
        cand = pd.concat([over_df, under_df], ignore_index=True)

    if cand.empty:
        return []

    if sport_up == "TENNIS" and "prop_type" in cand.columns and "direction" in cand.columns:
        pn = cand["prop_type"].apply(_norm_prop_label)
        ddir = cand["direction"].astype(str).str.upper().str.strip()
        ace_games_won = pn.str.contains("ace", na=False) | (
            pn.str.contains("game", na=False) & pn.str.contains("won", na=False) & ~pn.str.contains("set", na=False)
        )
        cand = cand[~(ace_games_won & (ddir == "UNDER"))].copy()

    if cand.empty:
        return []

    if sport_up == "NHL" and "hit_rate" in cand.columns:
        hr0 = pd.to_numeric(cand["hit_rate"], errors="coerce").fillna(0.0)
        if bool((hr0 <= 0.001).mean() >= 0.60):
            proxy_col = None
            for c in ("hit_rate_over_L10", "hit_rate_over_L5", "hit_rate_over_L20", "over_L10", "over_L5"):
                if c in cand.columns:
                    proxy_col = c
                    break
            if proxy_col is not None:
                proxy = pd.to_numeric(cand[proxy_col], errors="coerce")
                if proxy.dropna().max() > 1.5:
                    proxy = proxy / 100.0
                proxy = proxy.clip(lower=0.52, upper=0.90)
                cand["hit_rate"] = proxy.where(proxy.notna(), hr0)

    if sport_up == "TENNIS" and "hit_rate" in cand.columns:
        hr0 = pd.to_numeric(cand["hit_rate"], errors="coerce").fillna(0.0)
        if bool((hr0 <= 0.001).mean() >= 0.60):
            proxy_col = None
            for c in ("hit_rate_over_L10", "hit_rate_over_L5", "hit_rate_over_L20", "over_L10", "over_L5"):
                if c in cand.columns:
                    proxy_col = c
                    break
            if proxy_col is not None:
                proxy = pd.to_numeric(cand[proxy_col], errors="coerce")
                if proxy.dropna().max() > 1.5:
                    proxy = proxy / 100.0
                proxy = proxy.clip(lower=0.52, upper=0.90)
                cand["hit_rate"] = proxy.where(proxy.notna(), hr0)
            elif "blended_score" in cand.columns:
                bs = pd.to_numeric(cand["blended_score"], errors="coerce")
                high = bs >= 0.70
                proxy = (bs * 0.65).where(high, (bs * 0.58)).clip(0.52, 0.90)
                m0 = hr0 <= 0.001
                use = m0 & proxy.notna()
                cand.loc[use, "hit_rate"] = proxy.loc[use]

    # No per-leg hit_rate floor in ticket construction.
    if cand.empty:
        return []

    if counters is not None:
        pct_cap = counters.get("player_ticket_counts")
        if pct_cap is not None and len(cand) > 0 and "player" in cand.columns:
            pn = cand["player"].map(_norm_player_join)
            cap_ok = pn.eq("") | pn.map(lambda p: int(pct_cap.get(p, 0)) < MAX_SLIPS_PER_PLAYER)
            cand = cand[cap_ok].copy()
    if cand.empty:
        return []

    if counters is not None:
        counters["total_eligible_count"] += int(len(cand))

    cand = _attach_ticket_pick_order(cand, ticket_sort_mode)
    cand["__dir_edge"] = _directional_edge_series(cand).fillna(0.0)
    bonus = cand["prop_type"].apply(_prop_priority_bonus) if "prop_type" in cand.columns else 0.0
    cand["__score_adj"] = cand["__ts_pri"] + bonus
    if flow == "standard":
        cand = cand.sort_values(
            ["__dir_edge", "__ts_pri", "__ts_sec"], ascending=[False, False, False], na_position="last"
        )
    else:
        cand = cand.sort_values(
            ["__dir_edge", "__score_adj", "__ts_sec"], ascending=[False, False, False], na_position="last"
        )

    tg_starts = max(1, int(ticket_gen_starts))
    row_sets = _collect_row_candidates_for_structure(cand, n_legs, flow, tg_starts, max_variants)
    out: list[dict] = []
    seen_keys: set[frozenset[str]] = set()
    for rows in row_sets:
        key = _ticket_row_dedup_key(rows)
        if key in seen_keys:
            continue
        seen_keys.add(key)
        fin = _finalize_structure_ticket_dict(
            rows, structure, sport_label, flow, n_legs, counters, prioritize_ticket_hit
        )
        if fin is not None:
            out.append(fin)
    return out


def apply_nba_context_confidence_filter(
    df: pd.DataFrame,
    enabled: bool = True,
    min_context_score: int = 2,
    min_l5_sample: int = 5,
) -> pd.DataFrame:
    """
    Context-aware filter calibrated from NBA backtest tendencies:
    - OVER performs better vs weaker defenses and faster pace
    - UNDER performs better vs stronger defenses and normal/slower pace
    - Require minimal recent sample size (L5 over+under count)

    Applies only to NBA Standard picks. Other sports/pick-types pass through unchanged.
    """
    if not enabled or df.empty:
        return df

    out = df.copy()
    if "sport" not in out.columns:
        return out

    sport = out["sport"].astype(str).str.upper()
    if "pick_type" in out.columns:
        pick = out["pick_type"].astype(str).str.title()
    else:
        pick = pd.Series(["Standard"] * len(out), index=out.index)

    is_nba_standard = (sport == "NBA") & (pick == "Standard")
    if not is_nba_standard.any():
        return out

    direction = out.get("direction", pd.Series("", index=out.index)).astype(str).str.upper()
    def_tier = out.get("def_tier", pd.Series("", index=out.index)).map(_norm_def_tier_cell_upper)
    pace_tier = out.get("pace_tier", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()

    l5_over = pd.to_numeric(out.get("l5_over", 0), errors="coerce").fillna(0)
    l5_under = pd.to_numeric(out.get("l5_under", 0), errors="coerce").fillna(0)
    l5_sample = l5_over + l5_under

    def_over_good = def_tier.isin(["WEAK", "AVG", "ABOVE AVG", "AVERAGE", "BELOW AVG"])
    def_under_good = def_tier.isin(["ELITE", "ABOVE AVG"])
    pace_over_good = pace_tier.eq("FAST")
    pace_under_good = pace_tier.isin(["NORMAL", "SLOW"])

    score = pd.Series(0, index=out.index, dtype="int64")
    score += (l5_sample >= min_l5_sample).astype(int)
    score += (((direction == "OVER") & def_over_good) | ((direction == "UNDER") & def_under_good)).astype(int)
    score += (((direction == "OVER") & pace_over_good) | ((direction == "UNDER") & pace_under_good)).astype(int)

    keep = (~is_nba_standard) | (score >= int(min_context_score))

    kept_before = len(out)
    out = out[keep].copy()
    dropped = kept_before - len(out)
    if dropped > 0:
        print(
            f"  [context_filter] Dropped {dropped} NBA standard rows "
            f"(score < {min_context_score}, min_l5_sample={min_l5_sample})"
        )
    return out


def compute_bet_signal_core(leg: dict) -> tuple[int, list[str]]:
    """
    Context score 0–3+ and human reasons; same rules as web direction_signal().
    """
    sport = str(leg.get("sport", "") or "").upper()
    direction = str(leg.get("direction", "") or "").upper()
    def_tier = _norm_def_tier_cell_upper(leg.get("def_tier", ""))
    pace_tier = str(leg.get("pace_tier", "") or "").upper().strip()
    l5o = _signal_float(leg.get("l5_over"))
    l5u = _signal_float(leg.get("l5_under"))
    l5_sample = int(round((l5o or 0) + (l5u or 0)))

    explicit_score = _signal_float(leg.get("context_score"))
    score = int(round(explicit_score)) if explicit_score is not None else 0
    reasons: list[str] = []

    if explicit_score is None:
        if l5_sample >= 5:
            score += 1
            reasons.append("enough recent sample")

        over_def_good = def_tier in {"WEAK", "AVG", "ABOVE AVG", "AVERAGE", "BELOW AVG"}
        under_def_good = def_tier in {"ELITE", "ABOVE AVG"}
        if (direction == "OVER" and over_def_good) or (direction == "UNDER" and under_def_good):
            score += 1
            reasons.append(f"defense supports {direction}")

        over_pace_good = pace_tier == "FAST"
        under_pace_good = pace_tier in {"NORMAL", "SLOW"}
        if (direction == "OVER" and over_pace_good) or (direction == "UNDER" and under_pace_good):
            score += 1
            reasons.append(f"pace supports {direction}")
    else:
        reasons.append(f"context score {int(round(explicit_score))}")
        if l5_sample > 0:
            reasons.append(f"L5 sample {l5_sample}")

    if sport != "NBA" and not reasons:
        hr = _signal_float(leg.get("hit_rate")) or 0.0
        ae = _signal_float(leg.get("abs_edge"))
        edge = ae if ae is not None else abs(_signal_float(leg.get("edge")) or 0.0)
        if hr >= 0.62 and edge > 0:
            score = 2
            reasons.append("strong model profile")
        elif hr >= 0.55:
            score = 1
            reasons.append("model lean")
        else:
            reasons.append("model-only read")

    return score, reasons


def excel_signal_columns_from_leg(leg: dict) -> dict[str, str]:
    score, _ = compute_bet_signal_core(leg)
    return {
        "bet_strong": "Y" if score >= 3 else "",
        "bet_lean": "Y" if score == 2 else "",
        "bet_risk": "LOW" if score >= 3 else ("MED" if score == 2 else "HIGH"),
    }


def apply_full_slate_signal_columns(df: pd.DataFrame) -> pd.DataFrame:
    if df is None or df.empty:
        return df
    out = df.copy()
    # Avoid DataFrame.apply(axis=1): it materializes a Series per row and is very slow here (~3k+ legs).
    sig = pd.DataFrame(
        (excel_signal_columns_from_leg(r) for r in out.to_dict(orient="records")),
        index=out.index,
    )
    return pd.concat([out, sig], axis=1)


def _ticket_bucket_name(n_legs: int) -> str:
    n = int(n_legs or 0)
    if n <= 2:
        return "2leg"
    if n == 3:
        return "3leg"
    return "4plus"


def _load_ticket_rerank_models() -> None:
    global _TICKET_MODEL, _TICKET_MODEL_BUCKETS, _TICKET_MODEL_FEATURES
    if _TICKET_MODEL is not None:
        return
    _TICKET_MODEL = False
    _TICKET_MODEL_BUCKETS = {}
    _TICKET_MODEL_FEATURES = []
    try:
        if joblib is None:
            _log_slate.warning("[ticket-rerank] joblib unavailable; rerank disabled")
            return
        if os.path.exists(TICKET_MODEL_FEATURES_PATH):
            with open(TICKET_MODEL_FEATURES_PATH, "r", encoding="utf-8") as f:
                _TICKET_MODEL_FEATURES = list(json.load(f) or [])
        if os.path.exists(TICKET_MODEL_PATH):
            _TICKET_MODEL = joblib.load(TICKET_MODEL_PATH)
        if bool(_TICKET_MODEL_USE_BUCKETS):
            for bname, p in {
                "2leg": TICKET_MODEL_2LEG_PATH,
                "3leg": TICKET_MODEL_3LEG_PATH,
                "4plus": TICKET_MODEL_4PLUS_PATH,
            }.items():
                if os.path.exists(p):
                    _TICKET_MODEL_BUCKETS[bname] = joblib.load(p)
    except Exception as e:
        _log_slate.warning("[ticket-rerank] model load failed: %s", e)
        _TICKET_MODEL = False
        _TICKET_MODEL_BUCKETS = {}
        _TICKET_MODEL_FEATURES = []


def _ticket_ev_signal(ticket: dict) -> float:
    for k in ("est_ev", "predicted_ev", "ev_power", "ticket_objective_score"):
        v = ticket.get(k)
        try:
            if v is not None and np.isfinite(float(v)):
                return float(v)
        except Exception:
            continue
    return 0.0


def _ticket_feature_vector(ticket: dict) -> tuple[pd.DataFrame, list[str]]:
    rows = list(ticket.get("rows") or [])
    sport_counts: Counter[str] = Counter()
    pick_counts: Counter[str] = Counter()
    ml_probs: list[float] = []
    hit_rates: list[float] = []
    edges: list[float] = []
    abs_edges: list[float] = []
    rank_scores: list[float] = []
    context_scores: list[float] = []
    intel_rates: list[float] = []
    leg_probs_used: list[float] = []
    for r in rows:
        sp = str(r.get("sport") or "").strip().upper()
        if sp:
            sport_counts[sp] += 1
        pt = str(r.get("pick_type") or "").strip().upper()
        if "DEMON" in pt:
            pick_counts["DEMON"] += 1
        elif "GOBLIN" in pt:
            pick_counts["GOBLIN"] += 1
        else:
            pick_counts["STANDARD"] += 1
        for dst, key in (
            (ml_probs, "ml_prob"),
            (hit_rates, "hit_rate"),
            (edges, "edge"),
            (abs_edges, "abs_edge"),
            (rank_scores, "rank_score"),
            (context_scores, "context_score"),
            (intel_rates, "intel_season_hit_rate"),
            (leg_probs_used, "leg_prob_used"),
        ):
            try:
                v = r.get(key)
                if v is not None and np.isfinite(float(v)):
                    dst.append(float(v))
            except Exception:
                pass

    n_legs = int(ticket.get("n_legs") or len(rows) or 0)
    dominant = sport_counts.most_common(1)[0][0] if sport_counts else ""
    group_type = str(ticket.get("group_type") or "").upper()
    if not group_type:
        group_type = "FLEX" if "flex" in str(ticket.get("sheet", "")).lower() else "POWER"

    base_num = {
        "n_legs": float(n_legs),
        "is_flex_structure": float(1 if group_type == "FLEX" else 0),
        "sports_in_ticket": float(len(sport_counts)),
        "legs_nba": float(sport_counts.get("NBA", 0) + sport_counts.get("NBA1H", 0) + sport_counts.get("NBA1Q", 0)),
        "legs_cbb": float(sport_counts.get("CBB", 0) + sport_counts.get("WCBB", 0)),
        "legs_nhl": float(sport_counts.get("NHL", 0)),
        "legs_soccer": float(sport_counts.get("SOCCER", 0)),
        "legs_mlb": float(sport_counts.get("MLB", 0)),
        "pick_standard_count": float(pick_counts.get("STANDARD", 0)),
        "pick_goblin_count": float(pick_counts.get("GOBLIN", 0)),
        "pick_demon_count": float(pick_counts.get("DEMON", 0)),
        "ticket_objective_score": float(ticket.get("ticket_objective_score") or 0.0),
        "ev_power": float(ticket.get("ev_power") or 0.0),
        "est_ev": float(ticket.get("est_ev") or 0.0),
        "flat_ev": float(ticket.get("flat_ev") or 0.0),
        "payout_multiplier": float(ticket.get("payout_multiplier") or 0.0),
        "power_payout": float(ticket.get("power_payout") or 0.0),
        "flex_payout": float(ticket.get("flex_payout") or 0.0),
        "est_win_prob": float(ticket.get("est_win_prob") or 0.0),
        "predicted_payout_mult": float(ticket.get("predicted_payout_mult") or 0.0),
        "predicted_p_win": float(ticket.get("predicted_p_win") or 0.0),
        "predicted_ev": float(ticket.get("predicted_ev") or 0.0),
        "avg_hit_rate_leg": float(np.mean(hit_rates)) if hit_rates else 0.0,
        "avg_ml_prob_leg": float(np.mean(ml_probs)) if ml_probs else 0.0,
        "min_ml_prob_leg": float(np.min(ml_probs)) if ml_probs else 0.0,
        "max_ml_prob_leg": float(np.max(ml_probs)) if ml_probs else 0.0,
        "std_ml_prob_leg": float(np.std(ml_probs)) if ml_probs else 0.0,
        "avg_leg_prob_used": float(np.mean(leg_probs_used)) if leg_probs_used else 0.0,
        "min_leg_prob_used": float(np.min(leg_probs_used)) if leg_probs_used else 0.0,
        "avg_edge_leg": float(np.mean(edges)) if edges else 0.0,
        "min_edge_leg": float(np.min(edges)) if edges else 0.0,
        "max_edge_leg": float(np.max(edges)) if edges else 0.0,
        "avg_abs_edge_leg": float(np.mean(abs_edges)) if abs_edges else 0.0,
        "avg_rank_score_leg": float(np.mean(rank_scores)) if rank_scores else 0.0,
        "min_rank_score_leg": float(np.min(rank_scores)) if rank_scores else 0.0,
        "avg_context_score_leg": float(np.mean(context_scores)) if context_scores else 0.0,
        "avg_intel_hit_rate_leg": float(np.mean(intel_rates)) if intel_rates else 0.0,
    }
    cat_vals = {
        f"group_type_{group_type}": 1.0 if group_type else 0.0,
        f"dominant_sport_{dominant}": 1.0 if dominant else 0.0,
    }
    feat_names = _TICKET_MODEL_FEATURES or (list(base_num.keys()) + list(cat_vals.keys()))
    vec = np.zeros((1, len(feat_names)), dtype=float)
    for i, name in enumerate(feat_names):
        if name in base_num:
            vec[0, i] = float(base_num[name])
        elif name in cat_vals:
            vec[0, i] = float(cat_vals[name])
    return pd.DataFrame(vec, columns=feat_names), feat_names


def _ticket_model_prob(ticket: dict) -> float | None:
    _load_ticket_rerank_models()
    if _TICKET_MODEL is False:
        return None
    try:
        model = _TICKET_MODEL_BUCKETS.get(_ticket_bucket_name(int(ticket.get("n_legs") or 0))) if _TICKET_MODEL_BUCKETS else _TICKET_MODEL
        if model is None:
            model = _TICKET_MODEL
        if model is None or model is False:
            return None
        X, _ = _ticket_feature_vector(ticket)
        p = float(model.predict_proba(X)[0, 1])
        return p
    except Exception:
        return None


def _rerank_tickets_live(tickets: list[dict], max_tickets: int) -> list[dict]:
    if not tickets:
        return []
    if not bool(_TICKET_MODEL_RERANK_ENABLED):
        return tickets[:max_tickets]
    scored: list[dict] = []
    evs = np.array([_ticket_ev_signal(t) for t in tickets], dtype=float)
    if len(evs) and float(np.max(evs) - np.min(evs)) > 1e-12:
        evn = (evs - np.min(evs)) / (np.max(evs) - np.min(evs))
    else:
        evn = np.full(len(tickets), 0.5, dtype=float)
    w = max(0.0, min(1.0, float(_TICKET_MODEL_RERANK_WEIGHT)))
    for i, t in enumerate(tickets):
        mp = _ticket_model_prob(t)
        if mp is None:
            mp = float(t.get("est_win_prob") or 0.0)
        score = (1.0 - w) * float(evn[i]) + w * float(mp)
        x = dict(t)
        x["ticket_model_p_cash"] = round(float(mp), 6)
        x["ticket_live_score"] = round(float(score), 6)
        scored.append(x)
    scored.sort(
        key=lambda x: (
            -float(x.get("ticket_live_score", 0.0)),
            -float(x.get("ticket_model_p_cash", 0.0)),
            -float(x.get("est_win_prob", 0.0)),
        )
    )
    keep_n = min(max_tickets, max(1, int(_TICKET_MODEL_RERANK_TOP_N)))
    return scored[:keep_n]


# ── Build tickets ──────────────────────────────────────────────────────────────
def build_tickets(
    pool: pd.DataFrame,
    n_legs: int,
    max_tickets=20,
    require_mix=False,
    leg_min_hit_by_n: dict[int, float] | None = None,
    prioritize_ticket_hit: bool = False,
    ticket_sort_mode: str = "rank",
    player_ticket_counts: dict[str, int] | None = None,
) -> list:
    """
    Exhaustive combination builder (dedupe-first, no quality gates).

    This intentionally avoids min hit-rate / tier / EV filters so we can generate
    large candidate sets for downstream ML and hit-rate analysis.
    """
    pool = pool.copy().reset_index(drop=True)
    if len(pool) < int(n_legs):
        return []

    has_sport_col = "sport" in pool.columns
    can_mix = require_mix and has_sport_col and pool["sport"].nunique(dropna=True) >= 2

    eligible = (
        _attach_ticket_pick_order(pool, ticket_sort_mode)
        .sort_values(["__ts_pri", "__ts_sec"], ascending=[False, False], na_position="last")
        .reset_index(drop=True)
    )
    eligible = _trim_pool_by_leg_count(eligible, n_legs)
    if len(eligible) < int(n_legs):
        return []

    # Pick a bounded top-K subset that still yields many combinations.
    target_combo_count = max(int(max_tickets) * 80, int(max_tickets))
    max_k = min(len(eligible), 120)
    k = max(int(n_legs), min(20, max_k))
    while k < max_k and math.comb(k, int(n_legs)) < target_combo_count:
        k += 1
    candidates = [r for _, r in eligible.head(k).iterrows()]

    seen_ticket_keys: set[frozenset] = set()
    tickets: list[dict] = []
    # Hard scan cap to avoid pathological runtimes.
    scan_cap = max(20_000, int(max_tickets) * 8_000)
    scanned = 0

    for combo in itertools.combinations(candidates, int(n_legs)):
        scanned += 1
        if scanned > scan_cap:
            break
        rows = list(combo)
        if can_mix:
            sports = {str(r.get("sport", "")).strip().upper() for r in rows if str(r.get("sport", "")).strip()}
            if len(sports) < 2:
                continue

        key = frozenset(_leg_fp_tuple(r) for r in rows)
        if key in seen_ticket_keys:
            continue
        if not _ticket_cap_can_add(rows, player_ticket_counts):
            continue

        hrs = [float(r.get("hit_rate", 0.5) or 0.5) for r in rows]
        rss = [float(r.get("rank_score", 0) or 0) for r in rows]
        leg_probs = [_resolve_leg_prob(r) for r in rows]
        prob_srcs = [src for _, src in leg_probs]
        avg_hr = float(np.mean(hrs)) if hrs else 0.0
        avg_rs = float(np.mean(rss)) if rss else 0.0
        cmult, caudit = _correlation_multiplier_and_audit(rows)
        ep = win_prob(leg_probs, n_legs) * cmult
        pout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})
        adj_power = calc_adjusted_payout(pout["power"], rows)
        adj_flex = calc_adjusted_payout(pout["flex"], rows)
        ev_power = ep * adj_power

        tickets.append(
            {
                "key": key,
                "rows": rows,
                "avg_hit_rate": avg_hr,
                "avg_rank_score": avg_rs,
                "est_win_prob": ep,
                "power_payout": adj_power,
                "flex_payout": adj_flex,
                "base_power_payout": pout["power"],
                "payout_multiplier": round(adj_power / pout["power"], 4) if pout["power"] else 1.0,
                "ev_power": round(ev_power, 4),
                "kelly_units": round(kelly_fraction(ep, adj_power, fraction=0.25), 2),
                "n_legs": n_legs,
                "leg_prob_sources": ",".join(sorted(set(prob_srcs))),
                "correlation_multiplier": cmult,
                "correlation_audit": caudit,
            }
        )
        seen_ticket_keys.add(key)
        _ticket_cap_register(rows, player_ticket_counts)

    tickets.sort(key=lambda x: (-x["est_win_prob"], -x["avg_rank_score"]))
    return _rerank_tickets_live(tickets, max_tickets=max_tickets)


# ──────────────────────────────────────────────────────────────────────────────
# Long-leg web ticket groups (per-sport + cross-sport) with enforced Std/Gob mix where applicable
# ──────────────────────────────────────────────────────────────────────────────
def build_mixed_picktype_tickets(
    pool_df: pd.DataFrame,
    n_legs: int,
    max_tickets: int,
    min_standard: int,
    min_leg_hit_rate: float | None = None,
    prioritize_ticket_hit: bool = False,
    ticket_sort_mode: str = "rank",
    player_ticket_counts: dict[str, int] | None = None,
) -> list:
    """Exhaustive combination builder with a minimum Standard-leg constraint."""
    pool_df = pool_df.copy()
    if "rank_score" not in pool_df.columns or "pick_type" not in pool_df.columns or len(pool_df) < int(n_legs):
        return []

    std_available = int((pool_df["pick_type"] == "Standard").sum())
    if std_available < int(min_standard):
        return []

    eligible = (
        _attach_ticket_pick_order(pool_df, ticket_sort_mode)
        .sort_values(["__ts_pri", "__ts_sec"], ascending=[False, False], na_position="last")
        .reset_index(drop=True)
    )
    eligible = _trim_pool_by_leg_count(eligible, n_legs)
    if len(eligible) < int(n_legs):
        return []

    target_combo_count = max(int(max_tickets) * 100, int(max_tickets))
    max_k = min(len(eligible), 120)
    k = max(int(n_legs), min(24, max_k))
    while k < max_k and math.comb(k, int(n_legs)) < target_combo_count:
        k += 1
    candidates = [r for _, r in eligible.head(k).iterrows()]

    tickets: list[dict] = []
    seen_ticket_keys: set[frozenset] = set()
    scan_cap = max(30_000, int(max_tickets) * 10_000)
    scanned = 0

    for combo in itertools.combinations(candidates, int(n_legs)):
        scanned += 1
        if scanned > scan_cap:
            break
        rows = list(combo)
        std_count = sum(1 for r in rows if str(r.get("pick_type", "")) == "Standard")
        if std_count < int(min_standard):
            continue

        key = frozenset(_leg_fp_tuple(r) for r in rows)
        if key in seen_ticket_keys:
            continue
        if not _ticket_cap_can_add(rows, player_ticket_counts):
            continue

        hrs = [float(r.get("hit_rate", 0.5) or 0.5) for r in rows]
        rss = [float(r.get("rank_score", 0) or 0) for r in rows]
        leg_probs = [_resolve_leg_prob(r) for r in rows]
        prob_srcs = [src for _, src in leg_probs]
        avg_hr = float(np.mean(hrs)) if hrs else 0.0
        avg_rs = float(np.mean(rss)) if rss else 0.0
        cmult, caudit = _correlation_multiplier_and_audit(rows)
        ep = win_prob(leg_probs, n_legs) * cmult
        pout = PAYOUT.get(n_legs, {"power": 0, "flex": 0})
        adj_power = calc_adjusted_payout(pout["power"], rows)
        adj_flex = calc_adjusted_payout(pout["flex"], rows)
        ev_power = ep * adj_power

        tickets.append(
            {
                "key": key,
                "rows": rows,
                "avg_hit_rate": avg_hr,
                "avg_rank_score": avg_rs,
                "est_win_prob": ep,
                "power_payout": adj_power,
                "flex_payout": adj_flex,
                "base_power_payout": pout["power"],
                "payout_multiplier": round(adj_power / pout["power"], 4) if pout["power"] else 1.0,
                "ev_power": round(ev_power, 4),
                "kelly_units": round(kelly_fraction(ep, adj_power, fraction=0.25), 2),
                "n_legs": n_legs,
                "leg_prob_sources": ",".join(sorted(set(prob_srcs))),
                "correlation_multiplier": cmult,
                "correlation_audit": caudit,
            }
        )
        seen_ticket_keys.add(key)
        _ticket_cap_register(rows, player_ticket_counts)

    tickets.sort(key=lambda x: (-x["est_win_prob"], -x["avg_rank_score"]))
    return _rerank_tickets_live(tickets, max_tickets=max_tickets)


def _sport_display_label(label: str) -> str:
    """Readable sport token for ticket group titles (web + SUMMARY)."""
    s = str(label or "").strip()
    if s.upper() in ("NBA+CBB", "NBA + CBB"):
        return "NBA/CBB"
    return s


def _sanitize_excel_sheet_title(raw: str) -> str:
    """Excel disallows \\ / * ? : [ ] in sheet names."""
    s = str(raw or "").strip()
    for ch in ("\\", "/", "*", "?", ":", "[", "]"):
        s = s.replace(ch, "-")
    return " ".join(s.split())


def _excel_ticket_sheet_title(display_name: str, max_len: int = 31) -> str:
    """Fit a workbook tab label; may abbreviate (display_name stays full in JSON)."""
    s = _sanitize_excel_sheet_title(display_name)
    if len(s) <= max_len:
        return s
    compact = (
        s.replace("Cross-sport", "X-Sport")
        .replace("(all pipes)", "(pipes)")
        .replace("Standard only", "Std only")
        .replace("Std+Gob", "S+G")
    )
    compact = " ".join(compact.split())
    if len(compact) <= max_len:
        return compact
    return compact[:max_len]


def _excel_ticket_sheet_title_unique(display_name: str, existing: Iterable[str]) -> str:
    base = _excel_ticket_sheet_title(display_name, 31)
    used = {str(x) for x in existing}
    if base not in used:
        return base
    for i in range(2, 30):
        suff = f" ({i})"
        head = base[: max(0, 31 - len(suff))].rstrip()
        cand = (head + suff).strip()[:31]
        if cand not in used:
            return cand
    return base[:28] + " +"


def build_final_web_ticket_groups(
    nba_pool: pd.DataFrame,
    cbb_pool: pd.DataFrame,
    nhl_pool: pd.DataFrame = None,
    soccer_pool: pd.DataFrame = None,
    tennis_pool: pd.DataFrame = None,
    mlb_pool: pd.DataFrame = None,
    min_hit_rate=0.65,
    min_edge=2.0,
    min_rank=5.0,
    ticket_leg_sizes: list | None = None,
    leg_min_hit_by_n: dict[int, float] | None = None,
    prioritize_ticket_hit: bool = False,
    ticket_sort_mode: str = "rank",
    player_ticket_counts: dict[str, int] | None = None,
    max_tickets_per_group: int = 3,
):
    def apply_filters(df):
        mask = pd.Series(True, index=df.index)
        if "sport" in df.columns and "reliability_note" in df.columns:
            sp = df["sport"].astype(str).str.upper().str.strip()
            rel = df["reliability_note"].astype(str).str.upper()
            mask &= ~(sp.eq("MLB") & rel.str.contains("THIN_SAMPLE_", na=False))
        if {"sport", "hit_rate_status", "prop_type"}.issubset(df.columns):
            sp = df["sport"].astype(str).str.upper().str.strip()
            hs = df["hit_rate_status"].astype(str).str.upper()
            pp = df["prop_type"].astype(str).str.lower()
            pitch_kw = pp.str.contains(
                "strikeout|pitching out|earned run|walks allowed|hits allowed|pitches thrown|innings",
                regex=True,
                na=False,
            )
            mask &= ~(sp.eq("MLB") & hs.str.startswith("BLENDED_N") & pitch_kw)
        l5_o = pd.to_numeric(df.get("l5_over"), errors="coerce").fillna(0)
        l5_u = pd.to_numeric(df.get("l5_under"), errors="coerce").fillna(0)
        mask &= (l5_o >= 4) | (l5_u >= 4)
        if min_edge > 0:
            mask &= _directional_edge_series(df).fillna(0) >= min_edge
        if min_rank is not None and "rank_score" in df.columns:
            mask &= df["rank_score"].fillna(-99) >= min_rank
        return df[mask].copy()

    def _split_sg(df_f: pd.DataFrame):
        if df_f is None or len(df_f) == 0:
            empty = pd.DataFrame()
            return empty, empty, empty
        if "pick_type" not in df_f.columns:
            return df_f.copy(), df_f.copy(), df_f.iloc[0:0].copy()
        mix = df_f[df_f["pick_type"].isin(["Standard", "Goblin"])].copy()
        std = df_f[df_f["pick_type"] == "Standard"].copy()
        gob = df_f[df_f["pick_type"] == "Goblin"].copy()
        return mix, std, gob

    def _min_std_mixed(n: int) -> int:
        return 2 if n >= 4 else 1

    def _sort_rank(df: pd.DataFrame) -> pd.DataFrame:
        if df is None or len(df) == 0 or "rank_score" not in df.columns:
            return df
        return df.sort_values("rank_score", ascending=False, na_position="last")

    _pct: dict[str, int] = player_ticket_counts if player_ticket_counts is not None else defaultdict(int)

    nba_filtered = apply_filters(nba_pool)
    nba_mix, nba_std, nba_gob = _split_sg(nba_filtered)

    groups = []
    leg_sizes = ticket_leg_sizes if ticket_leg_sizes is not None else TICKET_LEG_SIZES

    def _min_hr_for_n(n: int, label: str = "") -> float | None:
        base: float | None
        if not leg_min_hit_by_n:
            base = None
        else:
            base = float(leg_min_hit_by_n.get(int(n), LEG_MIN_HIT_RATE.get(int(n), 0.55)))
        if str(label).strip().upper() == "NHL":
            cap = NHL_LEG_MIN_HIT_RATE.get(int(n))
            if cap is not None:
                if base is None:
                    return float(cap)
                return min(base, float(cap))
        if str(label).strip().upper() == "TENNIS":
            cap = TENNIS_LEG_MIN_HIT_RATE.get(int(n))
            if cap is not None:
                if base is None:
                    return float(cap)
                return min(base, float(cap))
        if str(label).strip().upper() in ("SOCCER", "SOC"):
            cap = SOCCER_LEG_MIN_HIT_RATE.get(int(n))
            if cap is not None:
                if base is None:
                    return float(cap)
                return max(base, float(cap))
        return base

    max_tix = max(1, int(max_tickets_per_group))
    max_tix_3 = max(2, min(max_tix, max_tix))
    max_tix_other = max(1, max_tix)

    def _add_mixed_std_gob(sub: pd.DataFrame, label: str, leg_sizes_override: list | None = None):
        _ls = leg_sizes_override if leg_sizes_override is not None else leg_sizes
        for n in _ls:
            if len(sub) < n:
                continue
            mt = max_tix_3 if n == 3 else max_tix_other
            tix = build_mixed_picktype_tickets(
                sub,
                n,
                max_tickets=mt,
                min_standard=_min_std_mixed(n),
                min_leg_hit_rate=_min_hr_for_n(n, label),
                prioritize_ticket_hit=prioritize_ticket_hit,
                ticket_sort_mode=ticket_sort_mode,
                player_ticket_counts=_pct,
            )
            if tix:
                dlab = _sport_display_label(label)
                groups.append((f"{dlab} {n}-Leg · Std+Gob", tix, None))

    def _add_std_only(sub: pd.DataFrame, label: str, leg_sizes_override: list | None = None):
        _ls = leg_sizes_override if leg_sizes_override is not None else leg_sizes
        for n in _ls:
            if len(sub) < n:
                continue
            tix = build_tickets(
                sub,
                n,
                max_tickets=max_tix_other,
                leg_min_hit_by_n=leg_min_hit_by_n,
                prioritize_ticket_hit=prioritize_ticket_hit,
                ticket_sort_mode=ticket_sort_mode,
                player_ticket_counts=_pct,
            )
            if tix:
                dlab = _sport_display_label(label)
                groups.append((f"{dlab} {n}-Leg · Standard only", tix, None))

    _add_mixed_std_gob(nba_mix, "NBA")
    _add_std_only(nba_std, "NBA")

    cbb_mix = cbb_std = cbb_gob = pd.DataFrame()
    if cbb_pool is not None and len(cbb_pool):
        cbb_f = apply_filters(cbb_pool)
        cbb_mix, cbb_std, cbb_gob = _split_sg(cbb_f)
        _add_mixed_std_gob(cbb_mix, "CBB")
        _add_std_only(cbb_std, "CBB")
        combo_ncaa = pd.concat([nba_mix, cbb_mix], ignore_index=True)
        _add_mixed_std_gob(combo_ncaa, "NBA/CBB")
        _add_std_only(pd.concat([nba_std, cbb_std], ignore_index=True), "NBA/CBB")

    nhl_mix = nhl_std = nhl_gob = pd.DataFrame()
    if nhl_pool is not None and len(nhl_pool):
        nhl_f = apply_filters(nhl_pool)
        nhl_mix, nhl_std, nhl_gob = _split_sg(nhl_f)
        _add_mixed_std_gob(nhl_mix, "NHL")
        _add_std_only(nhl_std, "NHL")

    soc_mix = soc_std = soc_gob = pd.DataFrame()
    if soccer_pool is not None and len(soccer_pool):
        soc_f = apply_filters(soccer_pool)
        soc_mix, soc_std, soc_gob = _split_sg(soc_f)
        _add_mixed_std_gob(soc_mix, "Soccer")
        _add_std_only(soc_std, "Soccer")

    ten_mix = ten_std = ten_gob = pd.DataFrame()
    if tennis_pool is not None and len(tennis_pool):
        ten_f = apply_filters(tennis_pool)
        ten_mix, ten_std, ten_gob = _split_sg(ten_f)
        ten_ls = [n for n in leg_sizes if n <= MAX_LEGS_TENNIS]
        _add_mixed_std_gob(ten_mix, "Tennis", ten_ls)
        _add_std_only(ten_std, "Tennis", ten_ls)

    mlb_mix = mlb_std = mlb_gob = pd.DataFrame()
    if mlb_pool is not None and len(mlb_pool):
        mlb_f = apply_filters(mlb_pool)
        mlb_mix, mlb_std, mlb_gob = _split_sg(mlb_f)
        _add_mixed_std_gob(mlb_mix, "MLB")
        _add_std_only(mlb_std, "MLB")

    mix_frames = [f for f in (nba_mix, cbb_mix, nhl_mix, soc_mix, ten_mix, mlb_mix) if len(f) > 0]
    if mix_frames:
        all_sg = _sort_rank(pd.concat(mix_frames, ignore_index=True))
        if "sport" in all_sg.columns and all_sg["sport"].nunique() >= 2:
            for n in leg_sizes:
                if len(all_sg) < n:
                    continue
                tix = build_tickets(
                    all_sg,
                    n,
                    max_tickets=max_tix_3 if n == 3 else max_tix_other,
                    require_mix=True,
                    leg_min_hit_by_n=leg_min_hit_by_n,
                    prioritize_ticket_hit=prioritize_ticket_hit,
                    ticket_sort_mode=ticket_sort_mode,
                    player_ticket_counts=_pct,
                )
                if tix:
                    groups.append((f"Cross-sport {n}-Leg · Std+Gob", tix, None))

    std_frames = [f for f in (nba_std, cbb_std, nhl_std, soc_std, ten_std, mlb_std) if len(f) > 0]
    if std_frames:
        all_std = _sort_rank(pd.concat(std_frames, ignore_index=True))
        if "sport" in all_std.columns and all_std["sport"].nunique() >= 2:
            for n in leg_sizes:
                if len(all_std) < n:
                    continue
                tix = build_tickets(
                    all_std,
                    n,
                    max_tickets=max_tix_3 if n == 3 else max_tix_other,
                    require_mix=True,
                    leg_min_hit_by_n=leg_min_hit_by_n,
                    prioritize_ticket_hit=prioritize_ticket_hit,
                    ticket_sort_mode=ticket_sort_mode,
                    player_ticket_counts=_pct,
                )
                if tix:
                    groups.append((f"Cross-sport {n}-Leg · Standard", tix, None))

    return groups


def _filter_pool_cross_pick_mode(pool_df: pd.DataFrame | None, sport_label: str, mode: str) -> pd.DataFrame:
    """
    mode: standard | goblin | mix (Standard+Goblin only; Demon excluded when pick_type present).
    NHL/Soccer pools have no Goblin split — all rows count as Standard for standard/mix; goblin-only skips them.
    """
    if pool_df is None or len(pool_df) == 0:
        return pd.DataFrame()
    df = pool_df.copy()
    su = str(sport_label).upper()
    skip_pt = su in ("NHL", "SOCCER", "SOC", "TENNIS")
    if "pick_type" in df.columns and not skip_pt:
        pt = df["pick_type"].astype(str).str.strip().str.lower()
        if mode == "standard":
            df = df[pt == "standard"].copy()
        elif mode == "goblin":
            df = df[pt == "goblin"].copy()
        elif mode == "mix":
            df = df[pt.isin(["standard", "goblin"])].copy()
    elif mode == "goblin" and skip_pt:
        return pd.DataFrame()
    return df


def _pick_top_row_from_eligible_pool(pool_df: pd.DataFrame, sort_mode: str = "rank") -> dict | None:
    """
    Best single prop in a sport's eligible pool for cross-pipeline ticket.
    With sort_mode rank/ml/blend, uses _attach_ticket_pick_order (ML-aware when not rank).
    Fallback: rank_score, blended_score, ml distance, confidence_score.
    Applies global TICKET_EXCLUDED_PROPS + fantasy ban (same family as structured tickets).
    """
    if pool_df is None or pool_df.empty:
        return None
    df = pool_df.copy()
    if "prop_type" in df.columns:
        pn = df["prop_type"].apply(_norm_prop_label)
        df = df[~pn.isin(TICKET_EXCLUDED_PROPS) & ~pn.str.contains("fantasy", na=False)].copy()
    if df.empty:
        return None
    work = df.copy()
    sm = (sort_mode or "rank").strip().lower()
    if sm in ("rank", "ml", "blend"):
        work = _attach_ticket_pick_order(work, sm)
        work["_hr"] = pd.to_numeric(work.get("hit_rate"), errors="coerce").fillna(0)
        work = work.sort_values(["__ts_pri", "__ts_sec", "_hr"], ascending=[False, False, False], na_position="last")
        return work.iloc[0].to_dict()
    if "rank_score" in work.columns and pd.to_numeric(work["rank_score"], errors="coerce").notna().any():
        work["_k"] = pd.to_numeric(work["rank_score"], errors="coerce")
    elif "blended_score" in work.columns and pd.to_numeric(work["blended_score"], errors="coerce").notna().any():
        work["_k"] = pd.to_numeric(work["blended_score"], errors="coerce")
    elif "ml_prob" in work.columns and pd.to_numeric(work["ml_prob"], errors="coerce").notna().any():
        work["_k"] = (pd.to_numeric(work["ml_prob"], errors="coerce") - 0.5).abs()
    elif "confidence_score" in work.columns and pd.to_numeric(
        work["confidence_score"], errors="coerce"
    ).notna().any():
        work["_k"] = pd.to_numeric(work["confidence_score"], errors="coerce")
    else:
        return work.iloc[0].to_dict()
    work["_hr"] = pd.to_numeric(work.get("hit_rate"), errors="coerce").fillna(0)
    work = work.sort_values(["_k", "_hr"], ascending=[False, False], na_position="last")
    return work.iloc[0].to_dict()


def _collect_cross_pipeline_rows(
    sport_pools: list[tuple[str, pd.DataFrame | None]],
    mode: str,
    max_legs: int,
    ticket_sort_mode: str = "rank",
) -> list[dict]:
    """Up to max_legs legs, one per pipeline in order, after pick-mode filter."""
    rows: list[dict] = []
    for sport_label, pool_df in sport_pools:
        if len(rows) >= max_legs:
            break
        sub = _filter_pool_cross_pick_mode(pool_df, sport_label, mode)
        picked = _pick_top_row_from_eligible_pool(sub, sort_mode=ticket_sort_mode)
        if not picked:
            continue
        d = dict(picked)
        d["sport"] = str(sport_label).upper()
        rows.append(d)
    return rows


def _finalize_cross_pipeline_ticket(rows: list[dict], ticket_type: str) -> dict | None:
    """Build ticket dict from leg rows (min 2 legs). No extra EV gate."""
    if len(rows) < 2:
        return None
    n_legs = len(rows)
    leg_probs: list = []
    prob_srcs: list[str] = []
    hrs: list[float] = []
    rss: list[float] = []
    for r in rows:
        _p, _src = _resolve_leg_prob(pd.Series(r))
        leg_probs.append((_p, _src))
        prob_srcs.append(_src)
        hrs.append(float(r.get("hit_rate", 0.5) or 0.5))
        rss.append(float(r.get("rank_score", 0.0) or 0.0))
    cmult, caudit = _correlation_multiplier_and_audit(rows)
    ep = win_prob(leg_probs, n_legs) * cmult
    pwr, flx = power_flex_payout_for_n(n_legs)
    adj_power = calc_adjusted_payout(pwr, rows)
    adj_flex = calc_adjusted_payout(flx, rows)

    tiers = [str(r.get("tier", "")).upper() for r in rows]
    defs = [_norm_prop_label(r.get("def_tier", r.get("opp_def_tier", ""))) for r in rows]
    if all(t in {"A", "B"} for t in tiers) and all(d in {"avg", "weak"} for d in defs):
        expected_win_rate = 0.78
    else:
        expected_win_rate = 0.68

    return {
        "key": frozenset(
            (str(r.get("player", "")) + "|" + str(r.get("prop_type", ""))).strip() for r in rows
        ),
        "rows": rows,
        "avg_hit_rate": float(np.mean(hrs)) if hrs else 0.0,
        "avg_rank_score": float(np.mean(rss)) if rss else 0.0,
        "est_win_prob": ep,
        "power_payout": adj_power,
        "flex_payout": adj_flex,
        "base_power_payout": pwr,
        "payout_multiplier": round(adj_power / pwr, 4) if pwr else 1.0,
        "ev_power": round(ep * adj_power, 4),
        "kelly_units": round(kelly_fraction(ep, adj_power, fraction=0.25), 2),
        "n_legs": n_legs,
        "expected_win_rate": expected_win_rate,
        "ticket_type": ticket_type,
        "sport": "MIX",
        "leg_prob_sources": ",".join(sorted(set(prob_srcs))),
        "correlation_multiplier": cmult,
        "correlation_audit": caudit,
    }


def build_cross_pipeline_ticket_bundle(
    sport_pools: list[tuple[str, pd.DataFrame | None]],
    max_legs: int = CROSS_PIPELINE_MAX_LEGS,
    ticket_sort_mode: str = "rank",
    player_ticket_counts: dict[str, int] | None = None,
) -> list[tuple[str, dict]]:
    """
    Up to three tickets (each ≤ max_legs, default 6):
      Standard-only, Goblin-only, Std+Gob mix — best eligible prop per pipeline.
    Returns [(display_group_name, ticket_dict), ...] (Excel tab is derived separately).
    """
    specs = [
        ("standard", "Cross-Pipeline Standard", "Cross-sport · Standard (all pipes)"),
        ("goblin", "Cross-Pipeline Goblin", "Cross-sport · Goblin (all pipes)"),
        ("mix", "Cross-Pipeline Mix", "Cross-sport · Std+Gob (all pipes)"),
    ]
    out: list[tuple[str, dict]] = []
    for mode, ttype, display in specs:
        legs = _collect_cross_pipeline_rows(sport_pools, mode, max_legs, ticket_sort_mode=ticket_sort_mode)
        tix = _finalize_cross_pipeline_ticket(legs, ttype)
        if tix is not None:
            if not _ticket_cap_can_add(tix["rows"], player_ticket_counts):
                continue
            _ticket_cap_register(tix["rows"], player_ticket_counts)
            out.append((display, tix))
    return out


# ── Write slate sheet ──────────────────────────────────────────────────────────
SLATE_COLS = [
    "sport",
    "tier",
    "rank_score",
    "player",
    "team",
    "opp",
    "team_seed",
    "team_region",
    "team_ap_rank",
    "opp_seed",
    "opp_region",
    "opp_ap_rank",
    "ncaa_rank",
    "prop_type",
    "pick_type",
    "pick_platform",
    "line",
    "line_underdog",
    "line_draftkings",
    "best_cross_line",
    "best_cross_book",
    "cross_edge_vs_pp",
    "cross_n_books",
    "direction",
    "edge",
    "projection",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "l5_side_hits",
    "l5_consistency",
    "l10_over",
    "l10_under",
    "def_tier",
    "min_tier",
    "shot_role",
    "usage_role",
    "h2h_avg",
    "h2h_over_rate",
    "h2h_games",
    "h2h_last",
    "b2b_flag",
    "cv_pct",
    "opp_vs_avg_pct",
    "game_time",
]
SLATE_WIDTHS = [6, 5, 10, 20, 6, 6, 7, 10, 8, 7, 10, 8, 10, 18, 9, 10, 6, 11, 11, 9, 10, 9, 6, 8, 7, 10, 10, 8, 10, 7, 7, 9, 10, 8, 8, 10, 9, 10, 10, 8, 9, 8, 10, 7, 8, 10, 16]
SLATE_HDRS = [
    "Sport",
    "Tier",
    "Rank Score",
    "Player",
    "Team",
    "Opp",
    "Team Seed",
    "Team Region",
    "Team AP",
    "Opp Seed",
    "Opp Region",
    "Opp AP",
    "NCAA Rank",
    "Prop",
    "Pick Type",
    "Platform",
    "Line",
    "Line UD",
    "Line DK",
    "Best Line",
    "Best Book",
    "Edge vs PP",
    "#Books",
    "Dir",
    "Edge",
    "Proj",
    "Hit Rate",
    "L5 Avg",
    "Szn Avg",
    "L5 Over",
    "L5 Under",
    "L5 Side Hits",
    "L5 Match %",
    "L10 Over",
    "L10 Under",
    "Def Tier",
    "Min Tier",
    "Shot Role",
    "Usage Role",
    "H2H Avg",
    "H2H Over%",
    "H2H GP",
    "H2H Last",
    "B2B",
    "CV%",
    "Opp vs Avg%",
    "Game Time",
]

_SLATE_HDR_BY_COL = dict(zip(SLATE_COLS, SLATE_HDRS))
_SLATE_WIDTH_BY_COL = dict(zip(SLATE_COLS, SLATE_WIDTHS))

# Full Slate only: scan order, pace + STRONG/LEAN/RISK beside Def Tier (per-sport sheets keep SLATE_COLS).
FULL_SLATE_EXTRA_HDRS = {
    "hit_prob_over": "Hit Prob Over",
    "hit_prob_under": "Hit Prob Under",
    "hit_prob_selected": "Hit Prob Selected",
    "hit_prob_actionable": "Hit Prob Actionable",
    "rank_read_score": "Rank Read Score",
    "prop_quality_score": "Prop Quality Score",
    "data_completeness_score": "Data Completeness Score",
    "pick_type_eligible": "Pick Type Eligible",
    "pace_tier": "Pace Tier",
    "bet_strong": "STRONG",
    "bet_lean": "LEAN",
    "bet_risk": "RISK",
    "game_script_mult": "Game Script",
    "game_script_note": "Script Note",
    "l5_side_hits": "L5 Side Hits",
    "l5_consistency": "L5 Match %",
    "line_underdog": "Line (UD)",
    "pick_platform": "Platform",
    "line_draftkings": "Line (DK)",
    "best_cross_line": "Best Line",
    "best_cross_book": "Best Book",
    "cross_edge_vs_pp": "Edge vs PP",
    "cross_n_books": "#Books",
    "fetched_at": "fetched_at",
    "game_date": "Game Date",
    "standard_line": "Standard Line",
}
FULL_SLATE_EXTRA_WIDTHS = {
    "pace_tier": 10,
    "bet_strong": 9,
    "bet_lean": 7,
    "bet_risk": 9,
    "game_script_mult": 12,
    "game_script_note": 42,
    "l5_side_hits": 9,
    "l5_consistency": 10,
    "line_underdog": 11,
    "pick_platform": 10,
    "line_draftkings": 11,
    "best_cross_line": 9,
    "best_cross_book": 10,
    "cross_edge_vs_pp": 9,
    "cross_n_books": 6,
    "distribution_std": 10,
    "distribution_n": 8,
    "G1": 6,
    "G2": 6,
    "G3": 6,
    "G4": 6,
    "G5": 6,
    "G6": 6,
    "G7": 6,
    "G8": 6,
    "G9": 6,
    "G10": 6,
    "fetched_at": 20,
}

FULL_SLATE_COLS = [
    "sport",
    "tier",
    "rank_score",
    "rank_score_penalized",
    "surface",
    "line_combo",
    "distribution_std",
    "distribution_n",
    "G1",
    "G2",
    "G3",
    "G4",
    "G5",
    "G6",
    "G7",
    "G8",
    "G9",
    "G10",
    "player",
    "team",
    "opp",
    "game_time",
    "game_date",
    "fetched_at",
    "team_seed",
    "team_region",
    "team_ap_rank",
    "opp_seed",
    "opp_region",
    "opp_ap_rank",
    "ncaa_rank",
    "prop_type",
    "pick_type",
    "pick_platform",
    "line",
    "standard_line",
    "line_underdog",
    "line_draftkings",
    "best_cross_line",
    "best_cross_book",
    "cross_edge_vs_pp",
    "cross_n_books",
    "direction",
    "edge",
    "projection",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "l5_side_hits",
    "l5_consistency",
    "l10_over",
    "l10_under",
    "hit_prob_over",
    "hit_prob_under",
    "hit_prob_selected",
    "hit_prob_actionable",
    "leg_prob_used",
    "rank_read_score",
    "prop_quality_score",
    "data_completeness_score",
    "pick_type_eligible",
    "def_tier",
    "opponent_def_rank",
    "pace_tier",
    "bet_strong",
    "bet_lean",
    "bet_risk",
    "min_tier",
    "shot_role",
    "usage_role",
    "h2h_avg",
    "h2h_over_rate",
    "h2h_games",
    "h2h_last",
    "b2b_flag",
    "cv_pct",
    "opp_vs_avg_pct",
    "game_script_mult",
    "game_script_note",
]


def _slate_hdr_for(col: str) -> str:
    if col in FULL_SLATE_EXTRA_HDRS:
        return FULL_SLATE_EXTRA_HDRS[col]
    return _SLATE_HDR_BY_COL.get(col, col.replace("_", " ").title())


def _slate_width_for(col: str) -> int:
    if col in FULL_SLATE_EXTRA_WIDTHS:
        return FULL_SLATE_EXTRA_WIDTHS[col]
    return _SLATE_WIDTH_BY_COL.get(col, 11)


def full_slate_column_order(df: pd.DataFrame) -> List[str]:
    """Preferred Full Slate order first, then legacy slate columns, then any extras."""
    seen: set[str] = set()
    out: List[str] = []
    for c in FULL_SLATE_COLS:
        if c in df.columns and c not in seen:
            out.append(c)
            seen.add(c)
    for c in SLATE_COLS:
        if c in df.columns and c not in seen:
            out.append(c)
            seen.add(c)
    for c in df.columns:
        if c not in seen:
            out.append(c)
            seen.add(c)
    return out


def write_slate_sheet(
    wb,
    df,
    sheet_name,
    bg_hdr,
    sport_label="",
    *,
    column_order: Optional[List[str]] = None,
    full_slate_visual: bool = False,
):
    ws = wb.create_sheet(sheet_name)
    if column_order is not None:
        cols = [c for c in column_order if c in df.columns]
    else:
        cols = [c for c in SLATE_COLS if c in df.columns]
    hdrs = [_slate_hdr_for(c) for c in cols]
    widths = [_slate_width_for(c) for c in cols]
    sw(ws, widths)
    hdr_h = 28 if full_slate_visual else 22
    hdr_sz = 10 if full_slate_visual else 9
    ws.row_dimensions[1].height = hdr_h
    for ci, h in enumerate(hdrs, 1):
        hc(ws, 1, ci, h, bg=bg_hdr, sz=hdr_sz)
    ws.freeze_panes = "A2"

    for ri, row in enumerate(df[cols].itertuples(index=False), 2):
        if full_slate_visual:
            ws.row_dimensions[ri].height = 19
        bg = C["alt"] if ri % 2 == 0 else C["white"]
        sp = getattr(row, "sport", "")
        spu = str(sp).upper() if sp else ""
        if spu == "NBA":
            bg_row = C["nba"] if ri % 2 == 0 else C["white"]
        elif spu == "CBB":
            bg_row = C["cbb"] if ri % 2 == 0 else C["white"]
        elif spu == "NHL":
            bg_row = C["nhl"] if ri % 2 == 0 else C["white"]
        elif spu == "SOCCER":
            bg_row = C["soccer"] if ri % 2 == 0 else C["white"]
        elif spu == "TENNIS":
            bg_row = C["tennis"] if ri % 2 == 0 else C["white"]
        else:
            bg_row = bg

        for ci, col in enumerate(cols, 1):
            val = getattr(row, col, "")
            if val is None or (isinstance(val, float) and np.isnan(val)):
                val = ""
            if col == "tier":
                dc(ws, ri, ci, val, bg=tier_bg(val), bold=True, align="center")
            elif col == "pick_type":
                dc(ws, ri, ci, val, bg=pt_bg(val), align="center")
            elif col == "hit_rate":
                pct_cell(ws, ri, ci, val if val != "" else np.nan)
                continue
            elif col == "rank_score":
                dc(ws, ri, ci, round(val, 2) if val != "" else "", bg=bg_row, bold=True, fmt="0.00")
            elif col == "direction":
                dbg = C["over"] if str(val).upper() == "OVER" else C["under"]
                dc(ws, ri, ci, val, bg=dbg, bold=True)
            elif col == "sport":
                vu = str(val).upper() if val else ""
                if vu == "NBA":
                    sbg = C["hdr_nba"]
                elif vu == "CBB":
                    sbg = C["hdr_cbb"]
                elif vu == "NHL":
                    sbg = C["hdr_nhl"]
                elif vu == "SOCCER":
                    sbg = C["hdr_soccer"]
                elif vu == "TENNIS":
                    sbg = C["hdr_tennis"]
                else:
                    sbg = C["hdr"]
                dc(ws, ri, ci, val, bg=sbg, bold=True, fc="FFFFFF")
            elif col == "player":
                dc(ws, ri, ci, val, bg=bg_row, align="left", bold=True)
            elif col == "def_tier" and full_slate_visual:
                dc(ws, ri, ci, val, bg=bg_row, align="center", bold=True)
            elif col == "pace_tier" and full_slate_visual:
                dc(ws, ri, ci, val, bg=bg_row, align="center", bold=(val != ""))
            elif col == "bet_strong" and str(val).upper() == "Y":
                dc(ws, ri, ci, "Y", bg=C["hit"], bold=True, fc="FFFFFF", align="center")
            elif col == "bet_lean" and str(val).upper() == "Y":
                dc(ws, ri, ci, "Y", bg=C["gold"], bold=True, fc="000000", align="center")
            elif col == "bet_risk":
                vs = str(val).upper()
                if vs == "LOW":
                    dc(ws, ri, ci, val, bg=C["hit"], bold=True, fc="FFFFFF", align="center")
                elif vs == "MED":
                    dc(ws, ri, ci, val, bg=C["push"], bold=True, fc="000000", align="center")
                elif vs == "HIGH":
                    dc(ws, ri, ci, val, bg=C["miss"], bold=True, fc="FFFFFF", align="center")
                else:
                    dc(ws, ri, ci, val, bg=bg_row, align="center")
            elif col == "l5_side_hits" and val != "":
                try:
                    dc(ws, ri, ci, int(round(float(val))), bg=bg_row, align="center", fmt="0")
                except (TypeError, ValueError):
                    dc(ws, ri, ci, val, bg=bg_row, align="center")
            elif col == "l5_consistency":
                pct_cell(ws, ri, ci, val if val != "" else np.nan)
                continue
            elif col in ("h2h_over_rate", "opp_vs_avg_pct"):
                cell = dc(ws, ri, ci, val if val != "" else "", bg=bg_row, align="center")
                if val != "":
                    try:
                        cell.number_format = "0.0%"
                    except Exception:
                        pass
            elif col == "cv_pct":
                cell = dc(ws, ri, ci, val if val != "" else "", bg=bg_row, align="center")
                if val != "":
                    try:
                        cell.number_format = "0.0"
                    except Exception:
                        pass
            elif col == "b2b_flag":
                b2b_str = "YES" if str(val).lower() in ("true", "1", "yes") else ("NO" if val != "" else "")
                b2b_bg = C["miss"] if b2b_str == "YES" else bg_row
                dc(ws, ri, ci, b2b_str, bg=b2b_bg, bold=(b2b_str == "YES"), align="center",
                   fc="FFFFFF" if b2b_str == "YES" else "000000")
                continue
            elif col == "game_time":
                try:
                    if val and val != "":
                        dt = pd.to_datetime(val)
                        dc(ws, ri, ci, dt.strftime("%m/%d %I:%M%p"), bg=bg_row, align="center")
                    else:
                        dc(ws, ri, ci, "", bg=bg_row)
                except Exception:
                    dc(ws, ri, ci, str(val)[:16], bg=bg_row)
                continue
            elif col == "edge" and full_slate_visual and val != "":
                try:
                    ev = float(val)
                    dc(ws, ri, ci, round(ev, 2), bg=bg_row, align="center", bold=True, fmt="0.00")
                except Exception:
                    dc(ws, ri, ci, val, bg=bg_row, align="center")
            elif col == "game_script_mult" and full_slate_visual:
                fv = None
                try:
                    if val != "" and val is not None:
                        fv = float(val)
                except (TypeError, ValueError):
                    fv = None
                fill_gs = bg_row
                fc_gs = "000000"
                if fv is not None:
                    if fv >= 1.03:
                        fill_gs = C["hit"]
                        fc_gs = "FFFFFF"
                    elif fv < 0.90:
                        fill_gs = C["miss"]
                        fc_gs = "FFFFFF"
                    elif 0.90 <= fv <= 0.96:
                        fill_gs = C["push"]
                        fc_gs = "000000"
                dc(
                    ws,
                    ri,
                    ci,
                    round(fv, 3) if fv is not None else "",
                    bg=fill_gs,
                    align="center",
                    bold=(fv is not None),
                    fc=fc_gs,
                    fmt="0.000" if fv is not None else None,
                )
            elif col == "game_script_note" and full_slate_visual:
                dc(ws, ri, ci, val, bg=bg_row, align="left")
            elif col == "best_cross_line" and val != "":
                try:
                    dc(ws, ri, ci, round(float(val), 2), bg=bg_row, align="center", fmt="0.00")
                except (TypeError, ValueError):
                    dc(ws, ri, ci, val, bg=bg_row, align="center")
            elif col == "cross_edge_vs_pp" and val != "":
                try:
                    fv = float(val)
                    cbg = PatternFill("solid", start_color="C8F7C5") if fv > 0.01 else bg_row
                    dc(ws, ri, ci, round(fv, 2), bg=cbg, align="center", fmt="0.00")
                except (TypeError, ValueError):
                    dc(ws, ri, ci, val, bg=bg_row, align="center")
            elif col == "cross_n_books" and val != "":
                try:
                    dc(ws, ri, ci, int(round(float(val))), bg=bg_row, align="center", fmt="0")
                except (TypeError, ValueError):
                    dc(ws, ri, ci, val, bg=bg_row, align="center")
            else:
                dc(ws, ri, ci, val, bg=bg_row, align="center")

    if cols:
        ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"


# ── Write ticket sheet ─────────────────────────────────────────────────────────
TICKET_COLS = [
    "#",
    "player",
    "team",
    "opp",
    "prop_type",
    "pick_type",
    "line",
    "direction",
    "edge",
    "hit_rate",
    "l5_avg",
    "season_avg",
    "l5_over",
    "l5_under",
    "l10_over",
    "l10_under",
    "rank_score",
    "ml_prob",
    "def_tier",
    "h2h_avg",
    "h2h_over_rate",
    "h2h_games",
    "b2b_flag",
    "cv_pct",
    "opp_vs_avg_pct",
    "standard_line",
    "delta_pct",
    "best_cross_line",
    "best_cross_book",
    "cross_edge_vs_pp",
    "sport",
]
TICKET_HDRS = [
    "#",
    "Player",
    "Team",
    "Opp",
    "Prop",
    "Pick Type",
    "Line",
    "Dir",
    "Edge",
    "Hit Rate",
    "L5 Avg",
    "Szn Avg",
    "L5 Over",
    "L5 Under",
    "L10 Over",
    "L10 Under",
    "Rank Score",
    "ML Prob",
    "Def Tier",
    "H2H Avg",
    "H2H Over%",
    "H2H GP",
    "B2B",
    "CV%",
    "Opp vs Avg%",
    "Std Line",
    "Delta %",
    "Best Line",
    "Best Book",
    "Edge vs PP",
    "Sport",
]
TICKET_W = [4, 20, 6, 6, 18, 10, 6, 6, 7, 9, 8, 9, 7, 8, 8, 9, 11, 8, 10, 8, 9, 7, 7, 8, 10, 8, 8, 9, 10, 9, 6]


def write_ticket_sheet(wb, tickets, sheet_name, bg_hdr, label=""):
    if not tickets:
        return
    ws = wb.create_sheet(sheet_name)
    sw(ws, TICKET_W)
    ws.freeze_panes = "A2"

    ri = 1
    for ti, ticket in enumerate(tickets, 1):
        n = ticket["n_legs"]
        pout = ticket["power_payout"]
        fout = ticket["flex_payout"]
        cost = round(100 / pout, 0) if pout else 0
        avg_hr = ticket["avg_hit_rate"]
        ep = ticket["est_win_prob"]
        avg_rs = ticket["avg_rank_score"]

        base_pout  = ticket.get("base_power_payout", pout)
        pay_mult   = ticket.get("payout_multiplier", 1.0)
        ev_pow     = ticket.get("ev_power", round(ep * pout, 4))
        exp_wr     = ticket.get("expected_win_rate", None)
        ttype      = ticket.get("ticket_type", "")
        em_curve = ticket.get("est_multiplier")
        fm_curve = ticket.get("flat_multiplier")
        curve_lbl = ""
        if em_curve is not None and fm_curve is not None:
            try:
                curve_lbl = f"  ·  Curve est: {float(em_curve):.2f}x vs flat PP base {float(fm_curve):.2f}x"
            except (TypeError, ValueError):
                curve_lbl = ""
        _pl_parts: list[str] = []
        if abs(pay_mult - 1.0) > 0.001:
            _pl_parts.append(f"Payout Mult: {pay_mult:.2f}x (base {base_pout}x → adj {pout}x)")
        if curve_lbl:
            _pl_parts.append(curve_lbl.strip())
        mult_label = ("  ·  " + "  ·  ".join(_pl_parts)) if _pl_parts else ""
        wr_label = f"  ·  Expected Win Rate: {float(exp_wr):.0%}" if exp_wr is not None else ""
        banner = (
            f"  Ticket #{ti}  ·  {n}-Leg {label} {ttype}  ·  "
            f"Power: {pout}x (${cost:.0f} to win $100)  ·  Flex: {fout}x  ·  "
            f"Avg Hit Rate: {avg_hr:.0%}  ·  Est Win Prob: {ep:.0%}  ·  EV: {ev_pow:.2f}  ·  "
            f"Avg Rank Score: {avg_rs:.2f}{wr_label}{mult_label}"
        )
        ws.merge_cells(start_row=ri, start_column=1, end_row=ri, end_column=len(TICKET_COLS))
        hc(ws, ri, 1, banner, bg=bg_hdr, sz=9, align="left")
        ws.row_dimensions[ri].height = 16
        ri += 1

        for ci, h in enumerate(TICKET_HDRS, 1):
            hc(ws, ri, ci, h, bg=C["hdr"], sz=8)
        ws.row_dimensions[ri].height = 14
        ri += 1

        for leg_i, row in enumerate(ticket["rows"], 1):
            bg = C["alt"] if leg_i % 2 == 0 else C["white"]
            sp = row.get("sport", "")
            if sp == "NBA":
                bg = C["nba"]
            elif sp == "CBB":
                bg = C["cbb"]
            elif sp == "MLB":
                bg = C["mlb"]
            elif sp == "NHL":
                bg = C["nhl"]
            elif sp in ("SOCCER", "SOC"):
                bg = C["soccer"]
            elif sp == "TENNIS":
                bg = C["tennis"]
            elif sp == "WCBB":
                bg = C["wcbb"]
            elif sp == "NBA1Q":
                bg = C["nba1q"]
            elif sp == "NBA1H":
                bg = C["nba1h"]

            def gv(field):
                return row.get(field, "")

            def _fmt_team_with_meta(team_val, seed_val, region_val, ap_val):
                t = str(team_val or "").strip()
                if not t:
                    return ""
                tags = []
                ap_missing = ap_val in ("", None) or (isinstance(ap_val, float) and np.isnan(ap_val))
                seed_missing = seed_val in ("", None) or (isinstance(seed_val, float) and np.isnan(seed_val))
                region_missing = region_val in ("", None) or (isinstance(region_val, float) and np.isnan(region_val))
                if not ap_missing:
                    tags.append(f"AP#{int(ap_val)}")
                if not seed_missing:
                    try:
                        s = int(seed_val)
                    except Exception:
                        s = seed_val
                    tags.append(f"S{s}")
                if not region_missing:
                    tags.append(str(region_val))
                return f"{t} ({' | '.join(tags)})" if tags else t

            dc(ws, ri, 1, leg_i, bg=bg, bold=True, align="center")
            dc(ws, ri, 2, gv("player"), bg=bg, align="left", bold=True)
            dc(ws, ri, 3, _fmt_team_with_meta(gv("team"), gv("team_seed"), gv("team_region"), gv("team_ap_rank")), bg=bg)
            dc(ws, ri, 4, _fmt_team_with_meta(gv("opp"), gv("opp_seed"), gv("opp_region"), gv("opp_ap_rank")), bg=bg)
            dc(ws, ri, 5, gv("prop_type"), bg=bg, align="left")
            ptv = gv("pick_type")
            dc(ws, ri, 6, ptv, bg=pt_bg(str(ptv)), align="center")
            dc(ws, ri, 7, gv("line"), bg=bg)
            dirv = str(gv("direction")).upper()
            dc(ws, ri, 8, dirv, bg=C["over"] if dirv == "OVER" else C["under"], bold=True)
            dc(ws, ri, 9, gv("edge"), bg=bg)
            pct_cell(ws, ri, 10, gv("hit_rate") if gv("hit_rate") != "" else np.nan)
            dc(ws, ri, 11, gv("l5_avg"), bg=bg)
            dc(ws, ri, 12, gv("season_avg"), bg=bg)
            dc(ws, ri, 13, gv("l5_over"), bg=bg)
            dc(ws, ri, 14, gv("l5_under"), bg=bg)
            dc(ws, ri, 15, gv("l10_over"), bg=bg)
            dc(ws, ri, 16, gv("l10_under"), bg=bg)
            rs = gv("rank_score")
            try:
                rs_out = round(float(rs), 2) if rs != "" and rs is not None else ""
            except Exception:
                rs_out = ""
            dc(ws, ri, 17, rs_out, bg=bg, bold=True)
            # ML prob (if present); keep formatting consistent with other probability fields
            mp = gv("ml_prob")
            try:
                mp_out = round(float(mp), 4) if mp != "" and mp is not None else ""
            except Exception:
                mp_out = ""
            dc(ws, ri, 18, mp_out, bg=bg, align="center", bold=(mp_out != ""), fmt="0.0000" if mp_out != "" else None)
            dc(ws, ri, 19, gv("def_tier"), bg=bg)
            # H2H Avg
            dc(ws, ri, 20, gv("h2h_avg"), bg=bg, align="center")
            # H2H Over%
            h2h_or = gv("h2h_over_rate")
            cell_h2h = dc(ws, ri, 21, h2h_or if h2h_or != "" else "", bg=bg, align="center")
            if h2h_or != "":
                try:
                    cell_h2h.number_format = "0.0%"
                except Exception:
                    pass
            # H2H GP
            dc(ws, ri, 22, gv("h2h_games"), bg=bg, align="center")
            # B2B
            b2b_raw = gv("b2b_flag")
            b2b_str = "YES" if str(b2b_raw).lower() in ("true", "1", "yes") else ("NO" if b2b_raw != "" else "")
            b2b_bg = C["miss"] if b2b_str == "YES" else bg
            dc(ws, ri, 23, b2b_str, bg=b2b_bg, bold=(b2b_str == "YES"), align="center",
               fc="FFFFFF" if b2b_str == "YES" else "000000")
            # CV%
            cv_val = gv("cv_pct")
            cell_cv = dc(ws, ri, 24, cv_val if cv_val != "" else "", bg=bg, align="center")
            if cv_val != "":
                try:
                    cell_cv.number_format = "0.0"
                except Exception:
                    pass
            # Opp vs Avg%
            opp_val = gv("opp_vs_avg_pct")
            cell_opp = dc(ws, ri, 25, opp_val if opp_val != "" else "", bg=bg, align="center")
            if opp_val != "":
                try:
                    cell_opp.number_format = "0.0%"
                except Exception:
                    pass
            std_v = gv("standard_line")
            try:
                std_out = round(float(std_v), 2) if std_v != "" and std_v is not None else ""
            except (TypeError, ValueError):
                std_out = ""
            dc(ws, ri, 26, std_out, bg=bg, align="center")
            _dpx = gd_leg_delta_pct(gv("line"), gv("standard_line"))
            dp_out = round(float(_dpx), 4) if _dpx is not None else ""
            dc(ws, ri, 27, dp_out, bg=bg, align="center", fmt="0.0000" if dp_out != "" else None)
            bcl = gv("best_cross_line")
            try:
                bcl_out = round(float(bcl), 2) if bcl != "" and bcl is not None else ""
            except (TypeError, ValueError):
                bcl_out = ""
            dc(ws, ri, 28, bcl_out, bg=bg, align="center", fmt="0.00" if bcl_out != "" else None)
            dc(ws, ri, 29, gv("best_cross_book"), bg=bg, align="center")
            cep = gv("cross_edge_vs_pp")
            try:
                cep_out = round(float(cep), 2) if cep != "" and cep is not None else ""
            except (TypeError, ValueError):
                cep_out = ""
            cell_cep = dc(ws, ri, 30, cep_out, bg=bg, align="center", fmt="0.00" if cep_out != "" else None)
            if cep_out != "":
                try:
                    if float(cep) > 0.01:
                        cell_cep.fill = PatternFill("solid", start_color="C8F7C5")
                except (TypeError, ValueError):
                    pass
            # Sport
            sv = gv("sport")
            sbg = C["hdr_nba"] if sv == "NBA" else (C["hdr_cbb"] if sv == "CBB" else C["hdr"])
            dc(ws, ri, 31, sv, bg=sbg, bold=True, fc="FFFFFF")
            ws.row_dimensions[ri].height = 14
            ri += 1

        ws.row_dimensions[ri].height = 6
        ri += 1


# ── Write SUMMARY sheet ───────────────────────────────────────────────────────
def write_summary(wb, nba, cbb, combined, all_ticket_groups, date_str, thresholds,
                  nhl=None, soccer=None, tennis=None, wcbb=None, mlb=None, nba1q=None, nba1h=None,
                  nfl=None):
    ws = wb.create_sheet("SUMMARY", 0)
    sw(ws, [28, 14, 10, 10, 10, 10, 10, 12, 18])

    ws.merge_cells("A1:I1")
    c = ws["A1"]
    c.value = (
        f"COMBINED NBA + CBB SLATE  |  {date_str}  |  Generated "
        f"{datetime.now(_SLATE_TZ).strftime('%Y-%m-%d %I:%M %p %Z')}"
    )
    c.font = Font(bold=True, name="Arial", size=13, color="FFFFFF")
    c.fill = PatternFill("solid", start_color=C["hdr"])
    c.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 30

    ws.merge_cells("A2:I2")
    c2 = ws["A2"]
    c2.value = (
        f"Filters: Tier {thresholds.get('tiers','ALL')} | "
        f"Min Hit Rate: {thresholds.get('min_hit_rate',0):.0%} | "
        f"Min Edge: {thresholds.get('min_edge',0)} | "
        f"Min Rank Score: {thresholds.get('min_rank','None')} | "
        f"Pick Types: {thresholds.get('pick_types','ALL')} | "
        f"Context Filter: {thresholds.get('context_filter', False)} "
        f"(score>={thresholds.get('context_min_score', 2)}, "
        f"L5 sample>={thresholds.get('context_min_l5_sample', 5)})"
    )
    c2.font = Font(bold=False, name="Arial", size=9, color="000000")
    c2.fill = PatternFill("solid", start_color=C["gold"])
    c2.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[2].height = 16

    row = 4

    def sec(r, label, bg):
        ws.merge_cells(start_row=r, start_column=1, end_row=r, end_column=9)
        hc(ws, r, 1, label, bg=bg, sz=10, align="left")
        ws.row_dimensions[r].height = 20
        return r + 1

    def stat_row(r, label, total, elig, bg=None):
        bg = bg or (C["alt"] if r % 2 == 0 else C["white"])
        dc(ws, r, 1, label, bg=bg, align="left", bold=True)
        dc(ws, r, 2, total, bg=bg)
        dc(ws, r, 3, elig, bg=bg)
        for ci in range(4, 10):
            dc(ws, r, ci, "", bg=bg)
        return r + 1

    row = sec(row, "📊 SLATE OVERVIEW", C["hdr_sum"])
    for ci, h in enumerate(["Category", "Total Props", "Eligible", "", "", "", "", "", ""], 1):
        hc(ws, row, ci, h, bg=C["hdr"], sz=8)
    ws.row_dimensions[row].height = 14
    row += 1

    elig_nba = len(nba[nba.get("tier", "").isin(["A", "B"])]) if "tier" in nba.columns else 0
    elig_cbb = len(cbb[cbb.get("tier", "").isin(["A", "B"])]) if "tier" in cbb.columns else 0
    elig_all = len(combined[combined.get("tier", "").isin(["A", "B"])]) if "tier" in combined.columns else 0
    row = stat_row(row, "NBA Props", len(nba), elig_nba, C["nba"])
    row = stat_row(row, "CBB Props", len(cbb), elig_cbb, C["cbb"])
    if nhl is not None and len(nhl) > 0:
        elig_nhl = len(nhl[nhl.get("tier", "").isin(["A", "B"])]) if "tier" in nhl.columns else 0
        row = stat_row(row, "NHL Props", len(nhl), elig_nhl, C["nhl"])
    if soccer is not None and len(soccer) > 0:
        elig_soc = len(soccer[soccer.get("tier", "").isin(["A", "B"])]) if "tier" in soccer.columns else 0
        row = stat_row(row, "Soccer Props", len(soccer), elig_soc, C["soccer"])
    if tennis is not None and len(tennis) > 0:
        elig_tn = len(tennis[tennis["tier"].isin(["A", "B"])]) if "tier" in tennis.columns else 0
        row = stat_row(row, "Tennis Props", len(tennis), elig_tn, C["tennis"])
    if wcbb is not None and len(wcbb) > 0:
        elig_wcbb = len(wcbb[wcbb["tier"].isin(["A", "B"])]) if "tier" in wcbb.columns else 0
        row = stat_row(row, "WCBB Props", len(wcbb), elig_wcbb, C["wcbb"])
    if mlb is not None and len(mlb) > 0:
        elig_mlb = len(mlb[mlb["tier"].isin(["A", "B"])]) if "tier" in mlb.columns else 0
        row = stat_row(row, "MLB Props", len(mlb), elig_mlb, C["mlb"])
    if nba1q is not None and len(nba1q) > 0:
        elig_nba1q = len(nba1q[nba1q["tier"].isin(["A", "B"])]) if "tier" in nba1q.columns else 0
        row = stat_row(row, "NBA1Q Props", len(nba1q), elig_nba1q, C["nba1q"])
    if nba1h is not None and len(nba1h) > 0:
        elig_nba1h = len(nba1h[nba1h["tier"].isin(["A", "B"])]) if "tier" in nba1h.columns else 0
        row = stat_row(row, "NBA1H Props", len(nba1h), elig_nba1h, C["nba1h"])
    if nfl is not None and len(nfl) > 0:
        elig_nfl = len(nfl[nfl["tier"].isin(["A", "B"])]) if "tier" in nfl.columns else 0
        row = stat_row(row, "NFL Props", len(nfl), elig_nfl, C["nfl"])
    row = stat_row(row, "Combined Slate", len(combined), elig_all)
    row += 1

    row = sec(row, "🎟️ TICKET SUMMARY", C["hdr_mix"])
    for ci, h in enumerate(
        ["Sheet", "Legs", "Type", "# Tickets", "Avg Hit Rate", "Avg Win Prob", "Avg Rank Score", "Adj Power Payout", "Avg EV", "Payout Mult", "Players"],
        1,
    ):
        hc(ws, row, ci, h, bg=C["hdr"], sz=8)
    ws.row_dimensions[row].height = 14
    row += 1

    sw(ws, [28, 14, 10, 10, 10, 10, 10, 12, 10, 11, 18])

    for group_name, tickets, bg_row in all_ticket_groups:
        if not tickets:
            continue
        avg_hr  = np.mean([t["avg_hit_rate"] for t in tickets])
        avg_wp  = np.mean([t["est_win_prob"] for t in tickets])
        avg_rs  = np.mean([t["avg_rank_score"] for t in tickets])
        avg_ev  = np.mean([t.get("ev_power", t["est_win_prob"] * t["power_payout"]) for t in tickets])
        avg_pm  = np.mean([t.get("payout_multiplier", 1.0) for t in tickets])
        n    = tickets[0]["n_legs"]
        pout = tickets[0]["power_payout"]
        bg   = bg_row if bg_row else (C["alt"] if row % 2 == 0 else C["white"])
        # colour the EV cell: green ≥ 1.2, amber 1.0–1.2, red < 1.0
        ev_bg = C["hit"] if avg_ev >= 1.2 else (C["push"] if avg_ev >= 1.0 else C["miss"])
        dc(ws, row, 1, group_name, bg=bg, align="left", bold=True)
        dc(ws, row, 2, n, bg=bg)
        lbl = group_name.split(" ")[0] if group_name else ""
        dc(ws, row, 3, lbl, bg=bg)
        dc(ws, row, 4, len(tickets), bg=bg)
        pct_cell(ws, row, 5, avg_hr)
        pct_cell(ws, row, 6, avg_wp)
        dc(ws, row, 7, round(avg_rs, 2), bg=bg)
        dc(ws, row, 8, f"{pout}x", bg=bg)
        dc(ws, row, 9, round(avg_ev, 2), bg=ev_bg, bold=True, fc="FFFFFF")
        dc(ws, row, 10, f"{avg_pm:.2f}x", bg=bg)
        sample = " | ".join(f"{r.get('player','')}" for r in tickets[0]["rows"][:3]) + ("..." if n > 3 else "")
        dc(ws, row, 11, sample, bg=bg, align="left", sz=8)
        row += 1


# ── Main ──────────────────────────────────────────────────────────────────────
def main():
    ap = argparse.ArgumentParser()
    ap.add_argument(
        "--nba",
        default="",
        help=(
            "NBA step8 xlsx. When omitted: outputs/<date>/step8_nba_direction_clean_<date>.xlsx if present, "
            f"else {DEFAULT_NBA_PATH}"
        ),
    )
    ap.add_argument(
        "--cbb",
        default="",
        help=f"CBB step6_ranked_cbb.xlsx. When omitted: {DEFAULT_CBB_PATH} (legacy CBB\\ fallback).",
    )
    ap.add_argument(
        "--nhl",
        default="",
        help=(
            "NHL step8. When omitted: outputs/<date>/step8_nhl_direction_clean_<date>.xlsx, "
            f"then {DEFAULT_NHL_PATH}"
        ),
    )
    ap.add_argument(
        "--soccer",
        default="",
        help=(
            "Soccer step8. When omitted: outputs/<date>/step8_soccer_direction_clean_<date>.xlsx, "
            f"then {DEFAULT_SOCCER_PATH}"
        ),
    )
    ap.add_argument(
        "--tennis",
        default="",
        help=(
            "Tennis step8. When omitted: outputs/<date>/step8_tennis_direction_clean_<date>.xlsx, "
            f"then {DEFAULT_TENNIS_PATH}"
        ),
    )
    ap.add_argument(
        "--wnba",
        default="",
        help=(
            "WNBA step8. When omitted: outputs/<date>/wnba/step8_wnba_direction_clean.xlsx, "
            "then outputs/<date>/step8_wnba_direction_clean_<date>.xlsx, "
            f"then Sports/WNBA/outputs/...; legacy paths still tried. Default constant: {DEFAULT_WNBA_PATH}"
        ),
    )
    ap.add_argument("--wcbb", default="", help="WCBB step8 direction clean xlsx (optional)")
    ap.add_argument("--mlb", default="", help="MLB step8 direction clean xlsx (optional)")
    ap.add_argument("--nba1q", default="", help="NBA 1st Quarter step8 direction clean xlsx (optional)")
    ap.add_argument("--nba1h", default="", help="NBA 1st Half step8 direction clean xlsx (optional)")
    ap.add_argument(
        "--nfl",
        default="",
        help=(
            "NFL step8. When omitted: outputs/<date>/step8_nfl_direction_clean_<date>.xlsx, "
            f"then {DEFAULT_NFL_PATH}"
        ),
    )
    ap.add_argument(
        "--cfb",
        default="",
        help=(
            "CFB step6. When omitted: outputs/<date>/cfb/step6_ranked_cfb.xlsx, "
            f"then {DEFAULT_CFB_PATH}"
        ),
    )
    ap.add_argument("--output", default="")
    ap.add_argument(
        "--date",
        default="",
        help="Slate date YYYY-MM-DD, or 'today' / 'now' (default: US Eastern calendar date)",
    )
    ap.add_argument(
        "--tennis-date",
        dest="tennis_date",
        default=None,
        help="Override date for Tennis step8 path + ET match-day filter (default: same as --date).",
    )
    ap.add_argument("--tiers", default="A,B,C", help="Comma-separated tiers e.g. A,B")
    ap.add_argument(
        "--high-conviction",
        action=argparse.BooleanOptionalAction,
        default=False,
        help=(
            "Strict ticket pool (optional): min pool hit rate >= 0.65; cap FINAL slips at 4 legs; "
            "merges HIGH_CONVICTION_LEG_MIN_HIT_RATE into per-leg floors. Default off for wider pools."
        ),
    )
    ap.add_argument(
        "--prioritize-ticket-hit",
        action=argparse.BooleanOptionalAction,
        default=False,
        help=(
            "Bias generation toward modeled payout probability (not a real-world 100%% guarantee): "
            "higher per-leg hit_rate floors, min modeled P(all legs hit) for power-style slips, "
            "min modeled P(flex cash) for Flex 3+ sheets; FINAL builders use the power threshold. "
            "May yield fewer or empty slips — loosen with --no-prioritize-ticket-hit (default)."
        ),
    )
    ap.add_argument(
        "--ticket-candidate-sort",
        choices=("rank", "ml", "blend", "rule"),
        default="rule",
        dest="ticket_candidate_sort",
        help=(
            "Order slate rows when choosing ticket legs. rank=rank_score only; ml=ml_prob first (NaN last); "
            "blend=avg(ml_prob, sigmoid(rank_score)) with missing ml using sigmoid(rank) only; "
            "rule=full context rules (remove Demon, L5/L10, defense, minutes, role, H2H, edge). "
            "Default blend uses your step8 ML Prob column when present (same signal as _resolve_leg_prob priority)."
        ),
    )
    ap.add_argument(
        "--ticket-gen-starts",
        type=int,
        default=48,
        dest="ticket_gen_starts",
        help=(
            "Structured tickets only: try the first K eligible rows as the first leg (after sort) and keep the slip "
            "with highest modeled ticket payout (flex cash for Flex 3+, else P(all hit)). Use 1 for legacy single-pass."
        ),
    )
    ap.add_argument(
        "--nba-structured-variants",
        type=int,
        default=8,
        dest="nba_structured_variants",
        help=(
            "NBA only: up to N distinct structured slips per sheet type (Power 2, Flex 3, Standard 2, Pwr Std 3, "
            "Goblin 3) from different first-leg seeds. Other sports still emit one slip per type. Clamped to 1-8."
        ),
    )
    ap.add_argument(
        "--min-leg-hit-rate",
        type=float,
        default=None,
        dest="min_leg_hit_rate",
        help="Every ticket leg must have hit_rate >= this (0-1). When strict mode is on, defaults to 0.70 if omitted.",
    )
    ap.add_argument(
        "--max-ticket-legs",
        type=int,
        default=6,
        dest="max_ticket_legs",
        help="FINAL / long-slip builders: max leg count (2-6). In strict mode (default), capped at 4 unless already lower.",
    )
    ap.add_argument("--min-hit-rate", type=float, default=0.65, dest="min_hit_rate")
    ap.add_argument("--min-edge", type=float, default=0.0, dest="min_edge")
    ap.add_argument(
        "--min-prop-quality",
        type=float,
        default=-1.0,
        dest="min_prop_quality",
        help=(
            "Minimum prop quality score in [0,1]. "
            "Use -1 (default) for sport-specific conservative defaults."
        ),
    )
    ap.add_argument("--min-rank", type=float, default=None, dest="min_rank")
    ap.add_argument(
        "--pick-types",
        default="Goblin,Standard",
        dest="pick_types",
        help="Comma-separated pick types for ticket eligibility (Demon excluded from tickets).",
    )
    ap.add_argument("--max-tickets", type=int, default=10, dest="max_tickets")
    ap.add_argument(
        "--ticket-model-rerank",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Use trained ticket ML models to rerank built tickets within each group.",
    )
    ap.add_argument(
        "--ticket-model-weight",
        type=float,
        default=0.15,
        help="Blend weight for model p_cash vs EV signal when reranking tickets (0..1).",
    )
    ap.add_argument(
        "--ticket-model-top-n",
        type=int,
        default=5,
        help="Keep top N tickets per generated group after model reranking.",
    )
    ap.add_argument(
        "--ticket-model-use-buckets",
        action=argparse.BooleanOptionalAction,
        default=True,
        help="Use bucketed ticket models (2-leg, 3-leg, 4+) when available.",
    )
    ap.add_argument("--use-context-filter", action="store_true", dest="use_context_filter", default=True,
                    help="Apply NBA direction+defense+pace context confidence filter")
    ap.add_argument("--no-context-filter", action="store_false", dest="use_context_filter",
                    help="Disable NBA direction+defense+pace context confidence filter")
    ap.add_argument("--context-min-score", type=int, default=2, dest="context_min_score",
                    help="Minimum NBA context score for Standard picks")
    ap.add_argument("--context-min-l5-sample", type=int, default=5, dest="context_min_l5_sample",
                    help="Minimum (L5 over+under) sample size for NBA context filter")
    ap.add_argument(
        "--allow-cross-date-fallback",
        action="store_true",
        help="Allow non-target game dates when target date has zero rows (default: strict target-date only).",
    )
    ap.add_argument(
        "--underdog-csv",
        default="",
        help="Optional Underdog fetch CSV (PP-shaped). If omitted and outputs/<date>/underdog_props.csv exists, it is used.",
    )
    ap.add_argument(
        "--draftkings-csv",
        default="",
        help="Optional DraftKings sportsbook CSV with board_sport column. "
        "If omitted, uses outputs/<date>/draftkings_props_all.csv if present, else draftkings_props_nba.csv.",
    )

    # Web outputs
    ap.add_argument(
        "--write-web",
        action="store_true",
        help="Write tickets_latest.json for web/Railway (graded HTML via build_ticket_eval.py)",
    )
    ap.add_argument(
        "--long-leg-supplement",
        action="store_true",
        help=(
            "After slate sheets, run build_final_web_ticket_groups for leg sizes 4–6 (adds extra Excel sheets). "
            "Very slow on large pools; main ticket emission already includes 4–6 leg groups for JSON. "
            "Default: skipped when --write-web is set; always runs when building workbook without --write-web."
        ),
    )
    ap.add_argument(
        "--no-web-ev-gate",
        action="store_true",
        dest="no_web_ev_gate",
        help="Put more sports on /tickets: skip positive-EV filter for JSON (template per sport/leg-count still applies). "
        "Default web JSON drops SKIP / negative-EV / very-low win-prob slips, then keeps "
        "empirical EV or STRONG/OK (pure Tennis / Soccer bypass only the min-EV threshold).",
    )
    ap.add_argument(
        "--web-outdir",
        default=DEFAULT_WEB_OUTDIR,
        help="Folder to write tickets_latest.json (+ slate_latest.json)",
    )
    ap.add_argument(
        "--also-root",
        action="store_true",
        help="Also write tickets_latest.json in repo root (HTML only from build_ticket_eval.py)",
    )
    ap.add_argument(
        "--merge-web-latest",
        action="store_true",
        help="Merge same-date slips into existing tickets_latest.json instead of replacing it (dedupe by leg set).",
    )
    ap.add_argument(
        "--bankroll",
        type=float,
        default=0.0,
        help="Optional bankroll (USD). When > 0, tickets_latest.json legs include recommended_stake_usd (fractional Kelly, utils/kelly_staking).",
    )
    ap.add_argument(
        "--curve-stake-usd",
        type=float,
        default=1.0,
        dest="curve_stake_usd",
        help="Stake (USD) for est_payout / est_ev / flat_ev columns (Goblin-Demon curve); does not change Kelly stakes.",
    )
    ap.add_argument(
        "--debug-payout",
        action="store_true",
        help="Print empirical payout debug for power tickets (base, goblin dists/factors, total adj, sweep).",
    )
    ap.add_argument(
        "--use-reverted-ladder",
        action="store_true",
        default=False,
        help=(
            "Use data/payout_ladder_reverted.json for exact payouts (Leaderboard/reverted context). "
            "Default uses data/payout_ladder.json."
        ),
    )
    ap.add_argument(
        "--no-diversity-prune",
        action="store_true",
        default=True,
        dest="no_diversity_prune",
        help="Keep all generated ticket groups after dedupe (skip jaccard + diversity pruning).",
    )
    ap.add_argument(
        "--web-template-cap",
        action=argparse.BooleanOptionalAction,
        default=False,
        dest="web_template_cap",
        help="Apply per-sport leg-count template quotas to /tickets JSON. Default off for full deduped coverage.",
    )
    ap.add_argument(
        "--win-rate-mode",
        action="store_true",
        dest="win_rate_mode",
        help="Win-rate ticket pass only: 2-3 leg goblin / Tier-A standard, sort by p_win, separate JSON output.",
    )
    ap.add_argument(
        "--max-legs",
        type=int,
        default=None,
        dest="max_legs",
        help="Max legs per ticket (win-rate mode: capped at 3).",
    )
    ap.add_argument(
        "--min-leg-prob",
        type=float,
        default=0.55,
        dest="min_leg_prob",
        help="Minimum leg_prob_used per leg (win-rate mode).",
    )
    ap.add_argument(
        "--web-filename",
        default="",
        dest="web_filename",
        help="Override web JSON filename (e.g. tickets_winrate_latest.json).",
    )

    args = ap.parse_args()
    global PAYOUT_DEBUG
    PAYOUT_DEBUG = bool(args.debug_payout)
    configure_payout_ladder(use_reverted_ladder=bool(args.use_reverted_ladder))
    print(
        f"[payout-ladder] mode={'reverted' if bool(args.use_reverted_ladder) else 'standard'} "
        f"path={PAYOUT_LADDER_PATH}"
    )

    ds = str(args.date).strip().lower()
    if not ds or ds in ("today", "now"):
        args.date = slate_calendar_date_ymd()

    tennis_ds = str(getattr(args, "tennis_date", None) or "").strip()[:10]
    if tennis_ds:
        args.tennis_date = tennis_ds
    else:
        # Match step8 ET date filter and outputs/<slate>/step8_tennis_direction_clean_<slate>.xlsx
        args.tennis_date = str(args.date).strip()[:10]

    args.max_ticket_legs = max(2, min(6, int(args.max_ticket_legs)))
    if getattr(args, "win_rate_mode", False):
        cap = int(args.max_legs) if args.max_legs is not None else 3
        args.max_ticket_legs = max(2, min(3, cap))
    args.ticket_gen_starts = max(1, min(64, int(args.ticket_gen_starts)))
    args.nba_structured_variants = max(1, min(8, int(args.nba_structured_variants)))
    args.ticket_model_weight = max(0.0, min(1.0, float(args.ticket_model_weight)))
    args.ticket_model_top_n = max(1, int(args.ticket_model_top_n))
    if args.high_conviction:
        args.min_hit_rate = max(float(args.min_hit_rate), 0.65)
        args.max_ticket_legs = min(args.max_ticket_legs, 4)
        print(
            "[tickets] strict pool: min hit rate >= 0.65, "
            f"max FINAL legs={args.max_ticket_legs} (use --no-high-conviction for wider pools)"
        )
    if args.prioritize_ticket_hit:
        args.min_hit_rate = max(float(args.min_hit_rate), 0.72)
        print(
            "[tickets] prioritize-ticket-hit: pool min hit rate >= 0.72, raised per-leg floors, "
            "modeled payout probability gates on structured + FINAL tickets"
        )

    global _TICKET_MODEL_RERANK_ENABLED, _TICKET_MODEL_RERANK_WEIGHT, _TICKET_MODEL_RERANK_TOP_N, _TICKET_MODEL_USE_BUCKETS
    _TICKET_MODEL_RERANK_ENABLED = bool(args.ticket_model_rerank)
    _TICKET_MODEL_RERANK_WEIGHT = float(args.ticket_model_weight)
    _TICKET_MODEL_RERANK_TOP_N = int(args.ticket_model_top_n)
    _TICKET_MODEL_USE_BUCKETS = bool(args.ticket_model_use_buckets)
    if _TICKET_MODEL_RERANK_ENABLED:
        print(
            f"[ticket-rerank] ON: weight={_TICKET_MODEL_RERANK_WEIGHT:.2f} "
            f"top_n={_TICKET_MODEL_RERANK_TOP_N} buckets={_TICKET_MODEL_USE_BUCKETS}"
        )

    effective_max_legs = int(args.max_ticket_legs)
    leg_sizes_runtime = ticket_leg_sizes_for_max(effective_max_legs)
    leg_min_override = None
    if args.min_leg_hit_rate is not None:
        leg_min_override = {n: float(args.min_leg_hit_rate) for n in TICKET_LEG_SIZES}
    leg_min_hit_by_n = effective_leg_min_hit_rates(
        bool(args.high_conviction),
        leg_min_override,
        prioritize_ticket_hit=bool(args.prioritize_ticket_hit),
    )

    structured_min_leg_hr = args.min_leg_hit_rate
    if args.high_conviction and structured_min_leg_hr is None:
        structured_min_leg_hr = 0.65
    nhl_structured_min_leg_hr = structured_min_leg_hr
    if nhl_structured_min_leg_hr is not None and float(nhl_structured_min_leg_hr) > 0.55:
        nhl_structured_min_leg_hr = 0.52
    tennis_structured_min_leg_hr = structured_min_leg_hr
    if tennis_structured_min_leg_hr is not None and float(tennis_structured_min_leg_hr) > 0.55:
        tennis_structured_min_leg_hr = 0.52
    print(f"[NHL TRACE] global structured_min_leg_hr={structured_min_leg_hr}")
    print(f"[NHL TRACE] NHL structured_min_leg_hr_override={nhl_structured_min_leg_hr}")
    print(f"[TENNIS TRACE] Tennis structured_min_leg_hr_override={tennis_structured_min_leg_hr}")

    apply_default_sport_inputs(args)

    if not args.output:
        args.output = f"combined_slate_tickets_{args.date}.xlsx"

    _repo_root = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
    _auto_ud = os.path.join(_repo_root, "outputs", args.date, "underdog_props.csv")
    _auto_dk_all = os.path.join(_repo_root, "outputs", args.date, "draftkings_props_all.csv")
    _auto_dk_nba = os.path.join(_repo_root, "outputs", args.date, "draftkings_props_nba.csv")
    if not str(args.underdog_csv).strip() and os.path.isfile(_auto_ud):
        args.underdog_csv = _auto_ud
        print(f"  [alt-books] Using Underdog CSV: {_auto_ud}")
    if not str(args.draftkings_csv).strip():
        if os.path.isfile(_auto_dk_all):
            args.draftkings_csv = _auto_dk_all
            print(f"  [alt-books] Using DraftKings CSV: {_auto_dk_all}")
        elif os.path.isfile(_auto_dk_nba):
            args.draftkings_csv = _auto_dk_nba
            print(f"  [alt-books] Using DraftKings CSV: {_auto_dk_nba}")

    print_combined_slate_input_paths(args)

    tiers = [t.strip() for t in args.tiers.split(",") if t.strip()]
    pick_types = [p.strip() for p in args.pick_types.split(",") if p.strip()]
    ticket_pick_types = [p for p in pick_types if p != "Demon"]
    thresholds = {
        "tiers": args.tiers,
        "min_hit_rate": args.min_hit_rate,
        "min_edge": args.min_edge,
        "min_rank": args.min_rank,
        # What actually feeds ticket builders (Demon never on tickets)
        "pick_types": ",".join(ticket_pick_types) if ticket_pick_types else "Goblin,Standard",
        "context_filter": args.use_context_filter,
        "context_min_score": args.context_min_score,
        "context_min_l5_sample": args.context_min_l5_sample,
        "high_conviction": bool(args.high_conviction),
        "prioritize_ticket_hit": bool(args.prioritize_ticket_hit),
        "ticket_candidate_sort": str(args.ticket_candidate_sort),
        "ticket_gen_starts": int(args.ticket_gen_starts),
        "nba_structured_variants": int(args.nba_structured_variants),
        "min_leg_hit_rate": args.min_leg_hit_rate,
        "structured_min_leg_hit_rate": structured_min_leg_hr,
        "tennis_structured_min_leg_hit_rate": tennis_structured_min_leg_hr,
        "max_ticket_legs": effective_max_legs,
        "leg_min_hit_by_n": {str(k): round(v, 4) for k, v in leg_min_hit_by_n.items()},
    }

    print(f"Loading NBA slate from {args.nba}...")
    try:
        nba = load_nba(args.nba)
        nba = enforce_target_date(
            nba, "NBA", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
        )
        print(f"  {len(nba)} NBA props loaded")
    except (FileNotFoundError, OSError) as e:
        print(f"  WARNING: NBA slate unavailable ({type(e).__name__}: {e}); continuing with 0 NBA props.")
        nba = pd.DataFrame()
    _load_audit_row("NBA", args.nba, nba)

    if "CBB" in DISABLED_SPORTS:
        print("Loading CBB slate skipped (deactivated season).")
        cbb = pd.DataFrame()
    else:
        print(f"Loading CBB slate from {args.cbb}...")
        cbb = load_cbb(args.cbb)
        cbb = enforce_target_date(
            cbb, "CBB", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
        )
        print(f"  {len(cbb)} CBB props loaded")
        _load_audit_row("CBB", args.cbb, cbb)

    nhl = None
    if str(args.nhl).strip():
        try:
            nhl = load_nhl(args.nhl)
            if nhl is not None and not nhl.empty:
                nhl = enforce_target_date(
                    nhl, "NHL", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
                )
                nhl = attach_standard_refs(nhl)
                print(f"  {len(nhl)} NHL props loaded")
                _load_audit_row("NHL", args.nhl, nhl)
        except Exception as e:
            print(f"  WARNING: Could not load NHL file: {e}")
            nhl = None

    soccer = None
    if str(args.soccer).strip():
        try:
            soccer = load_soccer(args.soccer)
            soccer = enforce_target_date(
                soccer, "Soccer", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
            )
            soccer = attach_standard_refs(soccer)
            print(f"  {len(soccer)} Soccer props loaded")
            _load_audit_row("Soccer", args.soccer, soccer)
        except Exception as e:
            print(f"  WARNING: Could not load Soccer file: {e}")
            soccer = None
    else:
        print("  [Soccer] skipped (empty --soccer)")

    tennis = None
    if str(args.tennis).strip():
        try:
            tennis = load_tennis(args.tennis)
            tennis_match_day = str(getattr(args, "tennis_date", None) or args.date).strip()[:10]
            tennis = enforce_target_date(
                tennis, "Tennis", tennis_match_day, allow_cross_date_fallback=args.allow_cross_date_fallback
            )
            tennis = attach_standard_refs(tennis)
            print(f"  {len(tennis)} Tennis props loaded")
            _load_audit_row("Tennis", args.tennis, tennis)
        except Exception as e:
            print(f"  WARNING: Could not load Tennis file: {e}")
            tennis = None
    else:
        print("  [Tennis] skipped (empty --tennis)")

    wnba = None
    if str(args.wnba).strip():
        try:
            wnba = load_wnba(args.wnba)
            wnba = enforce_target_date(
                wnba, "WNBA", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
            )
            wnba = attach_standard_refs(wnba)
            print(f"  {len(wnba)} WNBA props loaded")
            _load_audit_row("WNBA", args.wnba, wnba)
        except Exception as e:
            print(f"  WARNING: Could not load WNBA file: {e}")
            wnba = None
    else:
        print("  [WNBA] skipped (empty --wnba)")

    wcbb = None
    if args.wcbb:
        try:
            wcbb = load_wcbb(args.wcbb)
            wcbb = attach_standard_refs(wcbb)
            print(f"  {len(wcbb)} WCBB props loaded")
        except Exception as e:
            print(f"  WARNING: Could not load WCBB file: {e}")
            wcbb = None

    mlb = None
    mlb_path = str(args.mlb or "").strip()
    if mlb_path:
        try:
            mlb = load_mlb(mlb_path)
            mlb = enforce_target_date(
                mlb, "MLB", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
            )
            mlb = attach_standard_refs(mlb)
            print(f"  {len(mlb)} MLB props loaded")
            _load_audit_row("MLB", mlb_path, mlb)
        except Exception as e:
            print(f"  WARNING: Could not load MLB file: {e}")
            mlb = None

    nba1q = None
    nba1q_path = str(args.nba1q or "").strip() or (DEFAULT_NBA1Q_PATH if os.path.exists(DEFAULT_NBA1Q_PATH) else "")
    if nba1q_path:
        try:
            nba1q = load_nba1q(nba1q_path)
            nba1q = attach_standard_refs(nba1q)
            print(f"  {len(nba1q)} NBA1Q props loaded")
        except Exception as e:
            print(f"  WARNING: Could not load NBA1Q file: {e}")
            nba1q = None

    nba1h = None
    nba1h_path = str(args.nba1h or "").strip() or (DEFAULT_NBA1H_PATH if os.path.exists(DEFAULT_NBA1H_PATH) else "")
    if nba1h_path:
        try:
            nba1h = load_nba1h(nba1h_path)
            nba1h = attach_standard_refs(nba1h)
            print(f"  {len(nba1h)} NBA1H props loaded")
        except Exception as e:
            print(f"  WARNING: Could not load NBA1H file: {e}")
            nba1h = None

    nfl = None
    nfl_path = str(args.nfl or "").strip()
    if nfl_path:
        try:
            nfl = load_nfl(nfl_path)
            nfl = enforce_target_date(
                nfl, "NFL", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
            )
            nfl = attach_standard_refs(nfl)
            print(f"  {len(nfl)} NFL props loaded")
            _load_audit_row("NFL", nfl_path, nfl)
        except Exception as e:
            print(f"  WARNING: Could not load NFL file: {e}")
            nfl = None
    else:
        print("  [NFL] skipped (empty --nfl / no default file)")

    cfb = None
    cfb_path = str(getattr(args, "cfb", "") or "").strip()
    if cfb_path:
        try:
            cfb = load_cfb(cfb_path)
            cfb = enforce_target_date(
                cfb, "CFB", args.date, allow_cross_date_fallback=args.allow_cross_date_fallback
            )
            cfb = attach_standard_refs(cfb)
            print(f"  {len(cfb)} CFB props loaded")
            _load_audit_row("CFB", cfb_path, cfb)
        except Exception as e:
            print(f"  WARNING: Could not load CFB file: {e}")
            cfb = None
    else:
        print("  [CFB] skipped (empty --cfb / no default file)")

    # ✅ Attach Standard sibling refs AFTER normalized columns exist
    nba = attach_standard_refs(nba)
    cbb = attach_standard_refs(cbb)

    def drop_stale_rows(df, target_date, sport_label):
        if df is None or df.empty:
            return df
        td = str(target_date).strip()[:10]
        if "game_date" not in df.columns:
            # Only synthesize for NBA period sheets; other sports intentionally rely on explicit game_date.
            if sport_label not in ("NBA1Q", "NBA1H") or "game_time" not in df.columns:
                return df
            # Build game_date on the fly for sheets that only carry Game Time.
            target_year = int(td[:4]) if len(td) >= 4 and td[:4].isdigit() else datetime.now().year
            tmp = df.copy()
            tmp["game_date"] = _extract_game_dates(tmp["game_time"], target_year)
            df = tmp
        dated = df["game_date"].notna()
        gd_str = df["game_date"].astype(str).str[:10]
        # NBA boards (full + period) can be posted ahead of the run date.
        # Keep only the nearest future slate date (or latest available if all are past).
        if sport_label in ("NBA", "NFL", "MLB"):
            avail = sorted(gd_str[dated].dropna().unique().tolist())
            if not avail:
                return df
            future = [d for d in avail if d >= td]
            chosen = min(future) if future else max(avail)
            stale = dated & (gd_str != chosen)
            if chosen != td:
                print(
                    f"  [{sport_label}] date fallback: no props on {td}, "
                    f"using nearest date {chosen} ({int((~stale).sum())} rows)"
                )
        # Tennis: boards may not carry the pipeline calendar day (late refresh / ET drift).
        # If a strict "drop everything before td" would empty the slate, keep the nearest available
        # slate date (same idea as NBA), otherwise PP-only runs lose Tennis entirely.
        elif sport_label == "Tennis":
            stale_strict = dated & (gd_str < td)
            # If there is no dated row on/after the pipeline day, the board is from a nearby slate only.
            if dated.any() and not (dated & (gd_str >= td)).any():
                avail = sorted(gd_str[dated].dropna().unique().tolist())
                if avail:
                    future = [d for d in avail if d >= td]
                    chosen = min(future) if future else max(avail)
                    stale = dated & (gd_str != chosen)
                    if chosen != td:
                        print(
                            f"  [{sport_label}] date fallback: no rows on {td}, "
                            f"using nearest slate date {chosen} ({int((~stale).sum())} rows)"
                        )
                else:
                    stale = stale_strict
            else:
                stale = stale_strict
        elif sport_label == "Combined" and "sport" in df.columns:
            # Tennis allows future ET days; other sports (incl. Soccer) must match target.
            su = df["sport"].astype(str).str.upper()
            is_roll = su.isin(["TENNIS", "NBA", "NFL"])
            stale = dated & ((gd_str < td) | (~is_roll & (gd_str != td)))
        else:
            stale = dated & (gd_str != td)
        n_stale = int(stale.sum())
        if n_stale > 0:
            print(f"  [date-filter] {sport_label}: dropped {n_stale} stale-dated rows (target slate {td})")
        return df[~stale].copy()

    nba = drop_stale_rows(nba, args.date, "NBA")
    cbb = drop_stale_rows(cbb, args.date, "CBB")
    nhl = drop_stale_rows(nhl, args.date, "NHL")
    soccer = drop_stale_rows(soccer, args.date, "Soccer")
    tennis = drop_stale_rows(tennis, args.date, "Tennis")
    wnba = drop_stale_rows(wnba, args.date, "WNBA")
    wcbb = drop_stale_rows(wcbb, args.date, "WCBB")
    mlb = drop_stale_rows(mlb, args.date, "MLB")
    nba1q = drop_stale_rows(nba1q, args.date, "NBA1Q")
    nba1h = drop_stale_rows(nba1h, args.date, "NBA1H")
    nfl = drop_stale_rows(nfl, args.date, "NFL")
    cfb = drop_stale_rows(cfb, args.date, "CFB")

    # Apply teammate-absence usage redistribution before ticket eligibility filtering.
    nba = apply_usage_redistribution(nba, "NBA", args.date, REPO_ROOT)
    cbb = apply_usage_redistribution(cbb, "CBB", args.date, REPO_ROOT)
    wcbb = apply_usage_redistribution(wcbb, "WCBB", args.date, REPO_ROOT) if wcbb is not None else wcbb
    nhl = apply_usage_redistribution(nhl, "NHL", args.date, REPO_ROOT) if nhl is not None else nhl
    soccer = apply_usage_redistribution(soccer, "Soccer", args.date, REPO_ROOT) if soccer is not None else soccer
    tennis = apply_usage_redistribution(tennis, "Tennis", args.date, REPO_ROOT) if tennis is not None else tennis
    wnba = apply_usage_redistribution(wnba, "WNBA", args.date, REPO_ROOT) if wnba is not None else wnba
    mlb = apply_usage_redistribution(mlb, "MLB", args.date, REPO_ROOT) if mlb is not None else mlb
    nba1q = apply_usage_redistribution(nba1q, "NBA1Q", args.date, REPO_ROOT) if nba1q is not None else nba1q
    nba1h = apply_usage_redistribution(nba1h, "NBA1H", args.date, REPO_ROOT) if nba1h is not None else nba1h
    nfl = apply_usage_redistribution(nfl, "NFL", args.date, REPO_ROOT) if nfl is not None else nfl
    cfb = apply_usage_redistribution(cfb, "CFB", args.date, REPO_ROOT) if cfb is not None else cfb

    nba = drop_demon_over_rows(nba, "NBA")
    cbb = drop_demon_over_rows(cbb, "CBB")
    nhl = drop_demon_over_rows(nhl, "NHL")
    soccer = drop_demon_over_rows(soccer, "Soccer")
    tennis = drop_demon_over_rows(tennis, "Tennis")
    wnba = drop_demon_over_rows(wnba, "WNBA")
    wcbb = drop_demon_over_rows(wcbb, "WCBB")
    mlb = drop_demon_over_rows(mlb, "MLB")
    nba1q = drop_demon_over_rows(nba1q, "NBA1Q")
    nba1h = drop_demon_over_rows(nba1h, "NBA1H")
    nfl = drop_demon_over_rows(nfl, "NFL")
    cfb = drop_demon_over_rows(cfb, "CFB")

    print("Building combined slate...")
    combined = build_combined_slate(nba, cbb, nhl, soccer,
                                    tennis=tennis,
                                    wnba=wnba,
                                    wcbb=wcbb, mlb=mlb, nba1q=nba1q, nba1h=nba1h,
                                    nfl=nfl, cfb=cfb)
    reliability_index = _load_prop_reliability_index()
    if reliability_index:
        print(f"  [reliability] loaded {len(reliability_index)} prop-direction buckets from {PROP_RELIABILITY_LATEST_PATH}")
    else:
        print("  [reliability] no reliability index found; continuing without reliability gate")
    strat_index = _load_prop_strat_index()
    if strat_index:
        print(f"  [strat] loaded {len(strat_index)} trusted segment rows from {PROP_STRAT_BOARD_LATEST_PATH}")
    else:
        print("  [strat] no trusted stratification board found; continuing without strat weighting")

    # ✅ Attach Standard refs for combined too
    combined = attach_standard_refs(combined)

    combined = attach_alt_book_lines(
        combined,
        underdog_csv=str(args.underdog_csv or ""),
        draftkings_csv=str(args.draftkings_csv or ""),
    )
    combined = add_cross_platform_best_lines(combined)

    combined = drop_stale_rows(combined, args.date, "Combined")
    combined = enrich_read_fields_dataframe(combined)

    # Per-sport Excel sheets use SLATE_COLS — propagate UD/DK lines from combined onto each.
    nba = propagate_alt_book_lines_to_sport_frame(nba, combined, ("NBA",))
    cbb = propagate_alt_book_lines_to_sport_frame(cbb, combined, ("CBB",))
    nhl = propagate_alt_book_lines_to_sport_frame(nhl, combined, ("NHL",))
    soccer = propagate_alt_book_lines_to_sport_frame(soccer, combined, ("Soccer",))
    tennis = propagate_alt_book_lines_to_sport_frame(tennis, combined, ("Tennis",))
    wnba = propagate_alt_book_lines_to_sport_frame(wnba, combined, ("WNBA",))
    wcbb = propagate_alt_book_lines_to_sport_frame(wcbb, combined, ("WCBB",))
    mlb = propagate_alt_book_lines_to_sport_frame(mlb, combined, ("MLB",))
    nba1q = propagate_alt_book_lines_to_sport_frame(nba1q, combined, ("NBA1Q",))
    nba1h = propagate_alt_book_lines_to_sport_frame(nba1h, combined, ("NBA1H",))
    nfl = propagate_alt_book_lines_to_sport_frame(nfl, combined, ("NFL",))

    _n_ud = int(combined["line_underdog"].notna().sum()) if "line_underdog" in combined.columns else 0
    _n_dk = int(combined["line_draftkings"].notna().sum()) if "line_draftkings" in combined.columns else 0
    if _n_ud == 0 and _n_dk == 0:
        print(
            "  [alt-books] No Underdog/DraftKings lines merged (all blank). "
            f"Run run_pipeline.ps1 (alt-book fetch before combined) or write "
            f"outputs/{args.date}/underdog_props.csv and "
            f"outputs/{args.date}/draftkings_props_all.csv (or draftkings_props_nba.csv)."
        )

    print(f"  {len(combined)} total props")
    for s in ("NBA", "CBB", "NHL", "Soccer", "Tennis", "MLB", "NBA1H", "NBA1Q", "WCBB", "WNBA", "NFL"):
        n_s = int((combined["sport"] == s).sum()) if "sport" in combined.columns else 0
        if n_s > 0:
            print(f"  Full Slate rows — {s}: {n_s}")
    if len(combined) > 0 and {"sport", "prop_quality_score"}.issubset(combined.columns):
        print("\n[PROP QUALITY] percentiles by sport:")
        pq_summary = (
            combined.groupby("sport")["prop_quality_score"]
            .describe(percentiles=[0.25, 0.50, 0.75, 0.90])
            .round(3)
        )
        print(pq_summary.to_string())

    # ── CBB Goblin rank floor ─────────────────────────────────────────────────
    # CBB Goblin hits at ~55-58% vs NBA Goblin at ~67%.
    # We raise the minimum rank score for CBB Goblin-only pools so only
    # the model's highest-confidence CBB Goblin props enter tickets.
    CBB_GOBLIN_MIN_RANK = 5.0   # tune this up/down based on graded results

    discard_tracker = DiscardTracker()
    funnel_tracker = FunnelTracker()
    funnel_seen_sports: set[str] = set()

    global _MODEL_GATE_CACHE
    _MODEL_GATE_CACHE = _load_model_gate_recommendations()
    if _MODEL_GATE_CACHE:
        for sp, rec in _MODEL_GATE_CACHE.items():
            if isinstance(rec, dict) and rec.get("gate") and sp not in ALWAYS_ALLOW_SPORTS:
                if not (_nba1h_gated(_MODEL_GATE_CACHE or {}) and sp == "NBA1H"):
                    _log_auto_gate_once(sp, str(rec.get("reason") or "AUC gate"))
    if _nba1h_gated(_MODEL_GATE_CACHE or {}):
        print(f"  [ticket-gate] NBA1H excluded from tickets — {_nba1h_ticket_gate_reason()}")

    def pool(df, pt=None):
        if df is None or len(df) == 0:
            return df

        sport = str(df["sport"].iloc[0]).upper() if "sport" in df.columns and len(df) > 0 else ""
        total_loaded = int(len(df))

        gated, gate_reason = _sport_ticket_gated(sport)
        if gated:
            if sport and sport not in ALWAYS_ALLOW_SPORTS and not (
                _nba1h_gated(_MODEL_GATE_CACHE or {}) and sport == "NBA1H"
            ):
                _log_auto_gate_once(sport, gate_reason)
            gate_reason_key = "NBA1H_AUC_GATE" if sport == "NBA1H" else "MODEL_AUC_GATE"
            discard_tracker.log_count(sport or "ALL", gate_reason_key, total_loaded)
            funnel_tracker.checkpoint_df("after_model_auc_gate", pd.DataFrame(), default_sport=sport or "ALL")
            return pd.DataFrame()

        # Sport-specific prop exclusions
        excluded = set()
        if sport == "NBA":
            excluded = NBA_EXCLUDED_PROPS
        elif sport == "CBB":
            excluded = CBB_EXCLUDED_PROPS
        elif sport == "SOCCER":
            excluded = SOCCER_EXCLUDED_PROPS
        elif sport == "TENNIS":
            excluded = set()
        # NHL + MLB: row-wise low-signal exclusions.

        filtered_df = df.copy()
        filtered_df = _attach_reliability_columns(filtered_df, reliability_index)
        filtered_df = _attach_strat_columns(filtered_df, strat_index)
        voided_excluded = 0
        demon_candidates = None
        demon_passed = None
        demon_example = None
        if "void_reason" in filtered_df.columns:
            void_mask = filtered_df["void_reason"].apply(
                lambda x: bool(x) and str(x).strip().lower() not in ("", "nan", "none", "null")
            )
            voided_excluded = int(void_mask.sum())
        if excluded and "prop_type" in filtered_df.columns:
            filtered_df = filtered_df[
                ~filtered_df["prop_type"].astype(str).str.lower().isin(excluded)
            ]

        mlb_ex_n = 0
        rel_ex_n = 0
        nhl_sog_ex_n = nhl_fc_ex_n = nhl_leg_ex_n = 0
        if {"sport", "prop_type"}.issubset(filtered_df.columns):
            m_mlb, mlb_ex_n = _mlb_ticket_pool_exclusion_mask(filtered_df)
            m_nhl, nhl_sog_ex_n, nhl_fc_ex_n, nhl_leg_ex_n = _nhl_ticket_pool_exclusion_mask(filtered_df)
            filtered_df = filtered_df[~(m_mlb | m_nhl)].copy()
            filtered_df, rel_ex_n = _apply_reliability_pool_filter(filtered_df, reliability_index)
        # Normalize pick types for assembly compatibility.
        # Some sheets provide blank/variant labels; treat unknowns as Standard so
        # post-filter survivors can still participate in ticket building.
        if "pick_type" in filtered_df.columns:
            _pt_u = filtered_df["pick_type"].astype(str).str.strip().str.upper()
            filtered_df["pick_type"] = np.where(
                _pt_u.eq("GOBLIN"),
                "Goblin",
                np.where(_pt_u.eq("DEMON"), "Demon", "Standard"),
            )
        else:
            filtered_df["pick_type"] = "Standard"
        if mlb_ex_n:
            print(f"  [pool] Excluded {mlb_ex_n} MLB legs (home_runs/stolen_bases)")
        if rel_ex_n:
            print(f"  [pool] Excluded {rel_ex_n} legs via prop reliability gate")
        nhl_apr13_n = int(nhl_sog_ex_n + nhl_fc_ex_n)
        if nhl_apr13_n:
            print(
                f"  [pool] Excluded {nhl_apr13_n} NHL legs "
                f"(shots_on_goal OVER: {nhl_sog_ex_n}, faceoffs_won: {nhl_fc_ex_n})"
            )
        if nhl_leg_ex_n:
            print(
                f"  [pool] Excluded {nhl_leg_ex_n} NHL legs (goals/assists/plus-minus)"
            )
        # Direction-aware defense tier bonus/penalty before threshold checks.
        # Research-calibrated behavior:
        # - OVER + Above Avg defense tier: +0.05
        # - UNDER + Elite defense tier: +0.05
        # - OVER + Elite defense tier: -0.03
        if {"direction", "def_tier"}.issubset(filtered_df.columns):
            ddir = filtered_df["direction"].astype(str).str.upper().str.strip()
            dtier = filtered_df["def_tier"].map(_norm_def_tier_cell_upper)
            def_bonus = pd.Series(0.0, index=filtered_df.index)
            def_bonus = def_bonus + (((ddir == "OVER") & dtier.eq("ABOVE AVG")).astype(float) * 0.05)
            def_bonus = def_bonus + (((ddir == "UNDER") & dtier.eq("ELITE")).astype(float) * 0.05)
            def_bonus = def_bonus - (((ddir == "OVER") & dtier.eq("ELITE")).astype(float) * 0.03)
            filtered_df["_def_tier_bonus"] = def_bonus

            if "blended_score" in filtered_df.columns:
                filtered_df["blended_score"] = (
                    pd.to_numeric(filtered_df["blended_score"], errors="coerce").fillna(0.0) + def_bonus
                )
            if "rank_score" in filtered_df.columns:
                filtered_df["rank_score"] = (
                    pd.to_numeric(filtered_df["rank_score"], errors="coerce").fillna(0.0) + def_bonus
                )

        # Reliability-aware score shaping: keep rows for exploration, but softly downweight watchlist
        # buckets before threshold/ranking and hard filter later via _apply_reliability_pool_filter.
        if {"prop_reliability_score", "prop_reliability_status"}.issubset(filtered_df.columns):
            rel_score = pd.to_numeric(filtered_df["prop_reliability_score"], errors="coerce")
            rel_status = filtered_df["prop_reliability_status"].astype(str).str.upper().str.strip()
            soft_mult = pd.Series(1.0, index=filtered_df.index)
            soft_mult = soft_mult * np.where(rel_status.eq("WATCHLIST"), 0.85, 1.0)
            soft_mult = soft_mult * np.where(rel_score.notna(), np.clip(rel_score.fillna(1.0), 0.35, 1.0), 1.0)
            if "blended_score" in filtered_df.columns:
                filtered_df["blended_score"] = pd.to_numeric(filtered_df["blended_score"], errors="coerce").fillna(0.0) * soft_mult
            if "rank_score" in filtered_df.columns:
                filtered_df["rank_score"] = pd.to_numeric(filtered_df["rank_score"], errors="coerce").fillna(0.0) * soft_mult

        strat_boost_n = 0
        if {"strat_hit_rate", "strat_last5_hit_rate", "strat_n"}.issubset(filtered_df.columns):
            hr = pd.to_numeric(filtered_df["strat_hit_rate"], errors="coerce")
            l5 = pd.to_numeric(filtered_df["strat_last5_hit_rate"], errors="coerce")
            n = pd.to_numeric(filtered_df["strat_n"], errors="coerce")
            hr_component = ((hr - 0.56) * 1.2).clip(lower=-0.18, upper=0.22)
            l5_component = ((l5 - 0.56) * 0.5).clip(lower=-0.10, upper=0.10)
            n_conf = np.clip((n.fillna(0.0) / 80.0), 0.0, 1.0)
            strat_mult = (1.0 + (hr_component.fillna(0.0) + l5_component.fillna(0.0)) * (0.35 + 0.65 * n_conf)).clip(lower=0.78, upper=1.25)
            strat_boost_n = int(hr.notna().sum())
            if "blended_score" in filtered_df.columns:
                filtered_df["blended_score"] = pd.to_numeric(filtered_df["blended_score"], errors="coerce").fillna(0.0) * strat_mult
            if "rank_score" in filtered_df.columns:
                filtered_df["rank_score"] = pd.to_numeric(filtered_df["rank_score"], errors="coerce").fillna(0.0) * strat_mult

        # Sport-specific hit rate floors based on empirical data
        effective_min_hit = args.min_hit_rate

        if pt == ["Goblin"]:
            if sport == "NBA":
                effective_min_hit = max(args.min_hit_rate, 0.62)   # NBA Goblin: 64.3% overall
            elif sport == "CBB":
                effective_min_hit = max(args.min_hit_rate, 0.58)   # CBB Goblin: 61.9%
            elif sport == "NHL":
                effective_min_hit = max(args.min_hit_rate, 0.38)   # NHL Goblin is weak (40%)
            elif sport == "SOCCER":
                # Soccer OVER legs grade poorly vs model hit_rate — tighten pool vs generic 0.55.
                effective_min_hit = max(args.min_hit_rate, 0.58)
            else:
                effective_min_hit = max(args.min_hit_rate, 0.55)

        elif pt == ["Standard"]:
            if sport == "NBA":
                effective_min_hit = max(args.min_hit_rate, 0.50)   # NBA Standard: 50.7% — only Tier A viable
            elif sport == "CBB":
                effective_min_hit = max(args.min_hit_rate, 0.50)   # CBB Standard: 51.6%
            elif sport == "NHL":
                effective_min_hit = max(args.min_hit_rate, 0.65)   # NHL Standard: 67.9% — very strong
            elif sport == "SOCCER":
                effective_min_hit = max(args.min_hit_rate, 0.55)
            else:
                effective_min_hit = max(args.min_hit_rate, 0.50)

        # Tennis: pool() is often called with pt=None (structured + cross). Without a branch, the global
        # min_hit_rate (0.65+) applies while hit_rate is rank/blended proxy as low as ~0.50 — pool collapses.
        elif pt is None and sport == "TENNIS":
            effective_min_hit = min(float(args.min_hit_rate), 0.50)

        # Soccer OVER legs require stronger edge support; keep UNDER legs unchanged.
        if sport == "SOCCER" and "direction" in filtered_df.columns:
            _dir = filtered_df["direction"].astype(str).str.upper().str.strip()
            _edge = _edge_magnitude_series(filtered_df).fillna(0.0)
            _over_mask = _dir.eq("OVER")
            _under_mask = _dir.eq("UNDER")
            _over_total = int(_over_mask.sum())
            _under_total = int(_under_mask.sum())
            _keep_mask = (~_over_mask) | (_edge >= float(SOCCER_OVER_MIN_EDGE))
            filtered_df = filtered_df[_keep_mask].copy()
            _over_kept = int((_over_mask & _keep_mask).sum())
            _under_kept = int((_under_mask & _keep_mask).sum())
            print(
                "  [SOCCER GATE TRACE] "
                f"OVER kept={_over_kept}/{_over_total} removed={_over_total - _over_kept} "
                f"(edge>={SOCCER_OVER_MIN_EDGE:.2f}); "
                f"UNDER kept={_under_kept}/{_under_total} removed={_under_total - _under_kept}"
            )

        # MLB: allow both OVER and UNDER; directional edge + L5 consistency now controls selection.

        # Tennis: OVER only for Aces + Games Won; other props keep both directions.
        if sport == "TENNIS" and {"direction", "prop_type"}.issubset(filtered_df.columns):
            _pn = filtered_df["prop_type"].apply(_norm_prop_label)
            _dd = filtered_df["direction"].astype(str).str.upper().str.strip()
            _og = _pn.str.contains("ace", na=False) | (
                _pn.str.contains("game", na=False) & _pn.str.contains("won", na=False)
                & ~_pn.str.contains("set", na=False)
            )
            filtered_df = filtered_df[~(_og & (_dd == "UNDER"))].copy()

        # NHL/Soccer demon pool inclusion requires quality gate.
        if sport in ("NHL", "SOCCER") and "pick_type" in filtered_df.columns:
            _pt = filtered_df["pick_type"].astype(str).str.strip().str.upper()
            _is_demon = _pt.eq("DEMON")
            demon_candidates = int(_is_demon.sum())
            if demon_candidates > 0:
                _demon_pass = filtered_df.apply(_demon_passes_quality_gate, axis=1)
                _keep = (~_is_demon) | (_is_demon & _demon_pass)
                passed_rows = filtered_df[_is_demon & _demon_pass]
                demon_passed = int(len(passed_rows))
                if demon_passed > 0:
                    _ex = passed_rows.iloc[0]
                    demon_example = (
                        f"player={str(_ex.get('player', '')).strip()} | "
                        f"prop={str(_ex.get('prop_type', '')).strip()} | "
                        f"line={_fmt(_ex.get('line'), 2)} | "
                        f"hr={_fmt(_to_prob_0_1(_ex.get('hit_rate')), 2)} | "
                        f"ml={_fmt(_to_prob_0_1(_ex.get('ml_prob')), 2)}"
                    )
                filtered_df = filtered_df[_keep].copy()
            else:
                demon_passed = 0

        # Tier floor: exclude Tier D from all pools
        effective_tiers = [t for t in (tiers if tiers else ["A", "B", "C", "D"]) if t != "D"]
        # NHL / Tennis / NFL: strict high-conviction often collapses default tiers to A,B — pool is too small; allow Tier C.
        if sport in ("NHL", "TENNIS", "NFL") and bool(args.high_conviction):
            tier_u = {str(x).strip().upper() for x in effective_tiers}
            if "C" not in tier_u:
                effective_tiers = list(dict.fromkeys([*effective_tiers, "C"]))

        # CBB Goblin rank floor
        effective_min_rank = args.min_rank
        if sport == "CBB" and pt is not None and pt == ["Goblin"]:
            effective_min_rank = max(args.min_rank or 0, CBB_GOBLIN_MIN_RANK)

        if pt is not None:
            effective_pick_types = pt
        else:
            effective_pick_types = list(pick_types if pick_types else ["Goblin", "Standard"])
        effective_pick_types = [p for p in effective_pick_types if str(p).strip().lower() in {"goblin", "standard"}]

        use_funnel = (pt is None) and (sport not in funnel_seen_sports)
        pooled = filter_eligible(
            filtered_df,
            effective_min_hit,
            args.min_edge,
            effective_min_rank,
            effective_tiers,
            effective_pick_types,
            allow_strong_l5_bypass=(sport != "SOCCER"),
            min_prop_quality=float(args.min_prop_quality),
            discard_tracker=discard_tracker,
            funnel_tracker=(funnel_tracker if use_funnel else None),
            discard_sport=sport,
        )
        if use_funnel:
            funnel_seen_sports.add(sport)
        discard_tracker.log_count(str(sport), "final_pool_kept", int(len(pooled)))
        gob_n = std_n = dem_n = 0
        if pooled is not None and "pick_type" in pooled.columns:
            pt_u = pooled["pick_type"].astype(str).str.upper().str.strip()
            gob_n = int(pt_u.eq("GOBLIN").sum())
            std_n = int(pt_u.eq("STANDARD").sum())
            dem_n = int(pt_u.eq("DEMON").sum())
        # Telemetry reconciliation: loaded = voided + in_pool.
        passing = max(0, int(total_loaded - voided_excluded))
        print(
            f"  [pool] {sport}: {total_loaded} legs loaded | {voided_excluded} voided | {passing} in pool"
        )
        if strat_boost_n:
            print(f"         strat: weighted {strat_boost_n} rows using trusted segment history")
        print(f"         pick_types: goblin={gob_n} std={std_n} demon={dem_n}")
        if pooled is not None and len(pooled) > 0 and "prop_quality_score" in pooled.columns:
            pq = pd.to_numeric(pooled["prop_quality_score"], errors="coerce")
            avg_pq = float(pq.mean()) if pq.notna().any() else float("nan")
            sport_pq_floor, _ = _prop_quality_floor_over_under(str(sport))
            active_floor = float(args.min_prop_quality) if float(args.min_prop_quality) >= 0 else sport_pq_floor
            print(f"         quality: avg_pq={avg_pq:.3f} floor={active_floor:.2f}")
        if sport in ("NHL", "SOCCER") and demon_candidates is not None and demon_passed is not None:
            print(f"         demon gate: {demon_candidates} candidates -> {demon_passed} passed quality gate")
            if demon_example:
                print(f"         demon sample pass: {demon_example}")
        return pooled

    def print_nhl_trace(nhl_df: pd.DataFrame | None):
        if nhl_df is None or nhl_df.empty:
            return
        t0 = nhl_df.copy()
        print(f"  [NHL GATE TRACE] After date filter:         {len(t0)}")

        # Tier filter (mirrors pool's NHL behavior with strict-mode Tier C allowance)
        effective_tiers = [t for t in (tiers if tiers else ["A", "B", "C", "D"]) if t != "D"]
        if bool(args.high_conviction):
            tier_u = {str(x).strip().upper() for x in effective_tiers}
            if "C" not in tier_u:
                effective_tiers = list(dict.fromkeys([*effective_tiers, "C"]))
        t_tier = t0.copy()
        if "tier" in t_tier.columns:
            tier_set = {str(t).upper() for t in effective_tiers}
            tier_s = t_tier["tier"].astype(str).str.upper().str.strip()
            tier_ok = tier_s.isin(tier_set)
            if "D" not in tier_set and "prop_type" in t_tier.columns:
                attempt_ok = _is_attempt_prop_series(t_tier["prop_type"])
                tier_ok = tier_ok | ((tier_s == "D") & attempt_ok)
            t_tier = t_tier[tier_ok].copy()
        print(f"  [NHL GATE TRACE] After tier filter:         {len(t_tier)}")

        # Low-signal prop exclusions (mirror pool(): SOG OVER, faceoffs all dirs, legacy props)
        t_prop = t_tier.copy()
        if {"sport", "prop_type"}.issubset(t_prop.columns):
            m_nhl, n_sog, n_fc, n_leg = _nhl_ticket_pool_exclusion_mask(t_prop)
            t_prop = t_prop[~m_nhl].copy()
            print(
                f"  [NHL GATE TRACE] After low-signal prop filter: {len(t_prop)} "
                f"(shots_on_goal OVER={n_sog}, faceoffs_won={n_fc}, legacy={n_leg})"
            )
        else:
            print(f"  [NHL GATE TRACE] After low-signal prop filter: {len(t_prop)}")

        # Global pool min_hit_rate (before NHL cap override)
        t_global_hr = t_prop.copy()
        if "hit_rate" in t_global_hr.columns:
            global_min = float(args.min_hit_rate)
            t_global_hr = t_global_hr[pd.to_numeric(t_global_hr["hit_rate"], errors="coerce").fillna(0) >= global_min].copy()
        print(f"  [NHL GATE TRACE] After global min_hit_rate: {len(t_global_hr)}")

        # NHL hit-rate cap for 2-leg/structured entry (pool-level reference)
        t_nhl_cap = t_global_hr.copy()
        if "hit_rate" in t_nhl_cap.columns:
            nhl_cap = float(NHL_LEG_MIN_HIT_RATE.get(2, 0.55))
            t_nhl_cap = t_nhl_cap[pd.to_numeric(t_nhl_cap["hit_rate"], errors="coerce").fillna(0) >= max(0.52, nhl_cap)].copy()
        print(f"  [NHL GATE TRACE] After NHL hit_rate caps:   {len(t_nhl_cap)}")

        # EV/edge gate
        t_ev = t_nhl_cap.copy()
        if float(args.min_edge) > 0:
            t_ev = t_ev[_edge_magnitude_series(t_ev).fillna(0) >= float(args.min_edge)].copy()
        print(f"  [NHL GATE TRACE] After EV filter:           {len(t_ev)}")

        # Final pool (actual runtime pool() result)
        t_final = pool(nhl_df)
        print(f"  [NHL GATE TRACE] Final pool:                {len(t_final) if t_final is not None else 0}")

    if nhl is not None and len(nhl) > 0:
        print_nhl_trace(nhl)

    nba_pool = pool(nba)
    cbb_pool = pool(cbb)
    mlb_pool = pool(mlb)
    nfl_pool = pool(nfl) if nfl is not None and len(nfl) > 0 else None
    cfb_pool = pool(cfb) if cfb is not None and len(cfb) > 0 else None
    combo_pool = pool(combined)
    mlb_elig = len(mlb_pool) if mlb_pool is not None else 0
    nfl_elig = len(nfl_pool) if nfl_pool is not None else 0
    cfb_elig = len(cfb_pool) if cfb_pool is not None else 0
    _nhl_ticket_pool_n = (
        len(pool(nhl)) if nhl is not None and len(nhl) > 0 else 0
    )
    print(
        f"  NBA eligible: {len(nba_pool)} | CBB eligible: {len(cbb_pool)} | CFB eligible: {cfb_elig} | "
        f"MLB eligible: {mlb_elig} | NFL eligible: {nfl_elig} | Combined: {len(combo_pool)}"
    )
    print(
        f"  NHL ticket-pool legs (relaxed NHL hit-rate caps + Tier C in strict mode): {_nhl_ticket_pool_n}"
    )
    print(f"  CBB Goblin rank floor: {CBB_GOBLIN_MIN_RANK} (NBA uses global floor: {args.min_rank})")

    if getattr(args, "win_rate_mode", False):
        print("\n[win-rate] Generating win-rate optimized tickets (separate from EV pool)...")
        wr_max_legs = max(2, min(3, int(args.max_legs) if args.max_legs is not None else 3))
        wr_min_prob = float(getattr(args, "min_leg_prob", 0.55) or 0.55)
        graded_analysis = _load_graded_analysis()
        if graded_analysis:
            dr = graded_analysis.get("date_range") or {}
            print(
                f"  [win-rate] graded_analysis: {dr.get('min', '?')} → {dr.get('max', '?')} "
                f"({graded_analysis.get('total_props', 0):,} props)"
            )
        else:
            print(f"  [win-rate] graded_analysis not found ({_GRADED_ANALYSIS_JSON})")
        wr_sport_frames: list[tuple[str, pd.DataFrame]] = []
        for label, frame in (
            ("NBA1Q", nba1q),
            ("NBA", nba),
            ("NBA1H", nba1h),
            ("WNBA", wnba),
            ("MLB", mlb),
            ("NHL", nhl),
            ("Tennis", tennis),
        ):
            if _sport_ticket_gated(label)[0]:
                print(f"  [win-rate] {label} excluded (model AUC gate)")
                continue
            if frame is not None and len(frame) > 0:
                wr_sport_frames.append((label, pool(frame)))
        wr_groups = build_win_rate_ticket_groups(
            wr_sport_frames,
            min_leg_prob=wr_min_prob,
            min_composite_hr=0.52,
            max_legs=wr_max_legs,
            max_tickets=int(args.max_tickets),
            graded_analysis=graded_analysis,
        )
        print(f"  [win-rate] Built {len(wr_groups)} ticket groups ({sum(len(g[1]) for g in wr_groups)} slips)")
        wr_payload = ticket_groups_to_payload(
            wr_groups,
            args.date,
            thresholds,
            bankroll=max(0.0, float(args.bankroll)),
            curve_stake_usd=float(args.curve_stake_usd),
        )
        wr_payload["mode"] = "win_rate"
        wr_payload["max_legs"] = wr_max_legs
        wr_payload["sort"] = "p_win"
        for g in wr_payload.get("groups") or []:
            for slip in g.get("tickets") or []:
                _enrich_slip_p_win_fields(slip, mode="win_rate")
        if args.write_web:
            web_name = str(args.web_filename or "").strip() or "tickets_winrate_latest.json"
            write_web_outputs(
                wr_payload,
                args.web_outdir,
                require_positive_ev=False,
                merge_existing_for_date=False,
                apply_template_cap=False,
                json_filename=web_name,
                skip_ui_filters=True,
            )
        if args.output:
            wb_wr = Workbook()
            wb_wr.remove(wb_wr.active)
            for gn, tix, _bg in wr_groups:
                write_ticket_sheet(wb_wr, tix, _excel_ticket_sheet_title(gn), "FFD54F", label="Win-Rate")
            wb_wr.save(args.output)
            print(f"[OK] Win-rate workbook -> {args.output}")
        print("[win-rate] Done (EV ticket generation skipped).")
        return

    print("Generating tickets + workbook...")
    wb = Workbook()
    wb.remove(wb.active)

    all_ticket_groups = []
    fantasy_excluded_count = 0
    def_tier_filtered_count = 0
    ban_list_filtered_count = 0
    total_eligible_count = 0
    generated_tickets: dict = {}
    counters = {
        "fantasy_excluded_count": fantasy_excluded_count,
        "def_tier_filtered_count": def_tier_filtered_count,
        "ban_list_filtered_count": ban_list_filtered_count,
        "total_eligible_count": total_eligible_count,
        "player_ticket_counts": defaultdict(int),
    }
    _diversity_cfg = _load_diversity_config()

    def add_structured_sport_tickets(
        sport_df: pd.DataFrame,
        sport_label: str,
        bg_hdr: str,
        prefix: str,
        min_leg_hit_rate: float | None = None,
        prioritize_ticket_hit: bool = False,
        ticket_sort_mode: str = "rank",
        ticket_gen_starts: int = 10,
        ticket_variant_count: int = 1,
    ):
        if sport_df is None or sport_df.empty:
            print(f"  WARNING: {sport_label} skipped (empty pool).")
            return

        default_variant_count = max(1, min(8, int(args.max_tickets)))
        requested_variants = int(ticket_variant_count) if ticket_variant_count is not None else default_variant_count
        tvc = max(1, min(8, requested_variants))

        def _structured_list(structure: str, relaxed: bool = False) -> list[dict]:
            spec = _STRUCTURE_SPECS.get(structure) or {}
            nlegs = int(spec.get("n_legs", 0) or 0)
            if nlegs and nlegs > int(args.max_ticket_legs):
                return []
            if tvc > 1:
                return build_structure_ticket_variants(
                    sport_df,
                    sport_label,
                    structure,
                    counters=counters,
                    relaxed=relaxed,
                    min_leg_hit_rate=min_leg_hit_rate,
                    prioritize_ticket_hit=prioritize_ticket_hit,
                    ticket_sort_mode=ticket_sort_mode,
                    ticket_gen_starts=ticket_gen_starts,
                    max_variants=tvc,
                )
            t = build_single_structure_ticket(
                sport_df,
                sport_label,
                structure,
                counters=counters,
                relaxed=relaxed,
                min_leg_hit_rate=min_leg_hit_rate,
                prioritize_ticket_hit=prioritize_ticket_hit,
                ticket_sort_mode=ticket_sort_mode,
                ticket_gen_starts=ticket_gen_starts,
            )
            return [t] if t is not None else []

        def _gen_entry(tickets: list[dict]) -> dict | None:
            if not tickets:
                return None
            entry: dict = {
                "legs": [str(x.get("prop_type", "")) for x in tickets[0].get("rows", [])],
            }
            if len(tickets) > 1:
                entry["variant_legs"] = [
                    [str(x.get("prop_type", "")) for x in t.get("rows", [])] for t in tickets[1:]
                ]
            return entry

        p_list = _structured_list("power")
        p4_list = _structured_list("power4")
        p5_list = _structured_list("power5")
        p6_list = _structured_list("power6")
        f_list = _structured_list("flex")
        f4_list = _structured_list("flex4")
        f5_list = _structured_list("flex5")
        f6_list = _structured_list("flex6")
        s_list = _structured_list("standard")
        if not s_list:
            s_list = _structured_list("standard", relaxed=True)
        if not s_list and p_list:
            s_ticket = dict(p_list[0])
            s_ticket["ticket_type"] = "Standard"
            s_ticket["sport"] = sport_label
            s_list = [s_ticket]

        ps3_list = _structured_list("power_std3")
        g3_list = _structured_list("goblin3")

        # Greedy Flex 3 can fail on thin goblin pools while Power/Flex 4+ still fills.
        # If we have any 4–6 leg slip, derive Flex 3 from the best 3-leg subset of that slip.
        if not f_list:
            src_long = None
            for cand in (f4_list, f5_list, f6_list, p4_list, p5_list, p6_list):
                if cand:
                    src_long = cand[0]
                    break
            if src_long is not None:
                rows_long = list(src_long.get("rows") or [])
                rows3 = _best_flex3_rows_from_long_ticket(rows_long)
                if rows3:
                    # Backfill must not re-apply flex-cash floor if greedy Flex failed for pool-order reasons.
                    fin3 = _finalize_structure_ticket_dict(
                        rows3,
                        "flex",
                        sport_label,
                        "flex",
                        3,
                        counters,
                        False,
                    )
                    if fin3 is not None:
                        f_list = [fin3]
            # Tennis / Soccer often cap at 3 legs (no 4+ source): reuse Power Std 3 or Goblin 3 legs as Flex 3.
            if not f_list:
                for src3 in (ps3_list, g3_list):
                    if not src3:
                        continue
                    rows3 = list(src3[0].get("rows") or [])
                    if len(rows3) < 3:
                        continue
                    fin3 = _finalize_structure_ticket_dict(
                        rows3,
                        "flex",
                        sport_label,
                        "flex",
                        3,
                        counters,
                        False,
                    )
                    if fin3 is not None:
                        f_list = [fin3]
                        break

        if (
            not p_list and not p4_list and not p5_list and not p6_list
            and not f_list and not f4_list and not f5_list and not f6_list
            and not s_list and not ps3_list and not g3_list
        ):
            print(f"  WARNING: {sport_label} skipped (<2 eligible legs after strict filters).")
            return

        if p_list:
            sname = f"{prefix} Power Play 2-Leg"[:31]
            write_ticket_sheet(wb, p_list, sname, bg_hdr, label=f"{sport_label} Power Play")
            all_ticket_groups.append((sname, p_list, None))
            print(f"  {sname}: {len(p_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["power_play"] = _gen_entry(p_list)
        else:
            print(f"  WARNING: {sport_label} Power Play 2-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["power_play"] = None

        if p4_list:
            sname = f"{prefix} Power Play 4-Leg"[:31]
            write_ticket_sheet(wb, p4_list, sname, bg_hdr, label=f"{sport_label} Power Play")
            all_ticket_groups.append((sname, p4_list, None))
            print(f"  {sname}: {len(p4_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["power_play_4"] = _gen_entry(p4_list)
        else:
            print(f"  WARNING: {sport_label} Power Play 4-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["power_play_4"] = None

        if p5_list:
            sname = f"{prefix} Power Play 5-Leg"[:31]
            write_ticket_sheet(wb, p5_list, sname, bg_hdr, label=f"{sport_label} Power Play")
            all_ticket_groups.append((sname, p5_list, None))
            print(f"  {sname}: {len(p5_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["power_play_5"] = _gen_entry(p5_list)
        else:
            print(f"  WARNING: {sport_label} Power Play 5-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["power_play_5"] = None

        if p6_list:
            sname = f"{prefix} Power Play 6-Leg"[:31]
            write_ticket_sheet(wb, p6_list, sname, bg_hdr, label=f"{sport_label} Power Play")
            all_ticket_groups.append((sname, p6_list, None))
            print(f"  {sname}: {len(p6_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["power_play_6"] = _gen_entry(p6_list)
        else:
            print(f"  WARNING: {sport_label} Power Play 6-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["power_play_6"] = None

        if f_list:
            sname = f"{prefix} Flex 3-Leg"[:31]
            write_ticket_sheet(wb, f_list, sname, bg_hdr, label=f"{sport_label} Flex")
            all_ticket_groups.append((sname, f_list, None))
            print(f"  {sname}: {len(f_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["flex"] = _gen_entry(f_list)
        else:
            print(f"  WARNING: {sport_label} Flex 3-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["flex"] = None

        if f4_list:
            sname = f"{prefix} Flex 4-Leg"[:31]
            write_ticket_sheet(wb, f4_list, sname, bg_hdr, label=f"{sport_label} Flex")
            all_ticket_groups.append((sname, f4_list, None))
            print(f"  {sname}: {len(f4_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["flex_4"] = _gen_entry(f4_list)
        else:
            print(f"  WARNING: {sport_label} Flex 4-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["flex_4"] = None

        if f5_list:
            sname = f"{prefix} Flex 5-Leg"[:31]
            write_ticket_sheet(wb, f5_list, sname, bg_hdr, label=f"{sport_label} Flex")
            all_ticket_groups.append((sname, f5_list, None))
            print(f"  {sname}: {len(f5_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["flex_5"] = _gen_entry(f5_list)
        else:
            print(f"  WARNING: {sport_label} Flex 5-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["flex_5"] = None

        if f6_list:
            sname = f"{prefix} Flex 6-Leg"[:31]
            write_ticket_sheet(wb, f6_list, sname, bg_hdr, label=f"{sport_label} Flex")
            all_ticket_groups.append((sname, f6_list, None))
            print(f"  {sname}: {len(f6_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["flex_6"] = _gen_entry(f6_list)
        else:
            print(f"  WARNING: {sport_label} Flex 6-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["flex_6"] = None

        if s_list:
            sname = f"{prefix} Standard 2-Leg"[:31]
            write_ticket_sheet(wb, s_list, sname, bg_hdr, label=f"{sport_label} Standard")
            all_ticket_groups.append((sname, s_list, None))
            print(f"  {sname}: {len(s_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["standard"] = _gen_entry(s_list)
        else:
            print(f"  WARNING: {sport_label} Standard 2-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["standard"] = None

        if ps3_list:
            sname = f"{prefix} Pwr Std 3-Leg"[:31]
            write_ticket_sheet(wb, ps3_list, sname, bg_hdr, label=f"{sport_label} Power Std 3")
            all_ticket_groups.append((sname, ps3_list, None))
            print(f"  {sname}: {len(ps3_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["power_std3"] = _gen_entry(ps3_list)
        else:
            print(f"  WARNING: {sport_label} Power Standard 3-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["power_std3"] = None

        if g3_list:
            sname = f"{prefix} Goblin 3-Leg"[:31]
            write_ticket_sheet(wb, g3_list, sname, bg_hdr, label=f"{sport_label} Goblin 3")
            all_ticket_groups.append((sname, g3_list, None))
            print(f"  {sname}: {len(g3_list)} ticket(s)")
            generated_tickets.setdefault(sport_label, {})["goblin3"] = _gen_entry(g3_list)
        else:
            print(f"  WARNING: {sport_label} Goblin 3-Leg unavailable (strict filters).")
            generated_tickets.setdefault(sport_label, {})["goblin3"] = None

    _group_counts_by_size: dict[str, Counter[int]] = defaultdict(Counter)
    _zero_ticket_reasons: dict[str, str] = {}
    _diag_24: list[dict[str, Any]] = []

    def _pick_norm(v: Any) -> str:
        return str(v or "").strip().upper()

    def _header_for_sport(sport_label: str) -> str:
        sm = {
            "NBA": C["hdr_nba"],
            "CBB": C["hdr_cbb"],
            "NHL": C["hdr_nhl"],
            "SOCCER": C["hdr_soccer"],
            "TENNIS": C["hdr_tennis"],
            "MLB": C["hdr_mlb"],
            "NBA1Q": C["hdr_nba1q"],
            "NBA1H": C["hdr_nba1h"],
            "WCBB": C["hdr_wcbb"],
            "WNBA": C.get("hdr_nba", C["hdr"]),
            "NFL": C["hdr_nfl"],
        }
        return sm.get(str(sport_label).upper(), C["hdr_sum"])

    def _select_top_unique_rows(pool_df: pd.DataFrame, n_legs: int) -> list[dict]:
        chosen: list[dict] = []
        players: set[str] = set()
        player_props: set[str] = set()
        for _, row in pool_df.iterrows():
            rd = row.to_dict()
            p = str(rd.get("player", "")).strip().lower()
            if not p or p in players:
                continue
            pp = f"{p}::{_norm_prop_label(rd.get('prop_type', ''))}"
            if pp in player_props:
                continue
            chosen.append(rd)
            players.add(p)
            player_props.add(pp)
            if len(chosen) >= int(n_legs):
                break
        return chosen if len(chosen) == int(n_legs) else []

    def _select_cross_sport_unique_rows(pool_df: pd.DataFrame, n_legs: int) -> list[dict]:
        rows = _select_top_unique_rows(pool_df, n_legs)
        if len(rows) < int(n_legs):
            return []
        sports = {str(r.get("sport", "")).strip().upper() for r in rows if str(r.get("sport", "")).strip()}
        if len(sports) >= 2:
            return rows
        # Force at least one leg from a second sport while preserving uniqueness rules.
        base_sport = str(rows[0].get("sport", "")).strip().upper() if rows else ""
        players = {str(r.get("player", "")).strip().lower() for r in rows}
        player_props = {
            f"{str(r.get('player', '')).strip().lower()}::{_norm_prop_label(r.get('prop_type', ''))}"
            for r in rows
        }
        replacement: dict | None = None
        for _, row in pool_df.iterrows():
            rd = row.to_dict()
            sp = str(rd.get("sport", "")).strip().upper()
            if not sp or sp == base_sport:
                continue
            p = str(rd.get("player", "")).strip().lower()
            if not p or p in players:
                continue
            pp = f"{p}::{_norm_prop_label(rd.get('prop_type', ''))}"
            if pp in player_props:
                continue
            replacement = rd
            break
        if replacement is None:
            return []
        out = list(rows[:-1]) + [replacement]
        out_sports = {str(r.get("sport", "")).strip().upper() for r in out if str(r.get("sport", "")).strip()}
        return out if len(out) == int(n_legs) and len(out_sports) >= 2 else []

    def _ev_tag(ticket: dict) -> str:
        ev_mult = float(ticket.get("ev_power") or 0.0)
        if ev_mult >= 1.50:
            return "STRONG"
        if ev_mult >= 1.15:
            return "OK"
        if ev_mult >= 0.80:
            return "MARGINAL"
        return "SKIP"

    def _build_mode_ticket(
        rows: list[dict],
        sport_label: str,
        mode: str,
        *,
        min_payout_x: float | None = None,
    ) -> dict | None:
        flow = "flex" if mode == "flex" else "power"
        structure = "flex" if mode == "flex" else "power"
        t = _finalize_structure_ticket_dict(
            rows,
            structure,
            sport_label,
            flow,
            len(rows),
            None,
            False,
        )
        if t is None:
            return None
        min_payout_req = float(min_payout_x) if min_payout_x is not None else float(MIN_WEB_PAYOUT_X)
        if mode == "power" and float(t.get("power_payout") or 0.0) < min_payout_req:
            return None
        if mode == "flex" and float(t.get("flex_payout") or 0.0) < min_payout_req:
            return None
        t["ticket_type"] = "Flex" if mode == "flex" else "Power Play"
        t["ev_tag"] = _ev_tag(t)
        for r in t.get("rows", []):
            wp, _ = _resolve_leg_prob(pd.Series(r))
            r["win_probability"] = round(float(wp), 4)
            r["hit_rate_raw"] = r.get("hit_rate")
            r["l5_over_raw"] = r.get("l5_over")
            r["l5_under_raw"] = r.get("l5_under")
        return t

    def _diagnose_2_4(group_prefix: str, sport_label: str, pool_df: pd.DataFrame, require_multi_sport: bool = False) -> None:
        if not DEBUG_PAYOUT_DIAGNOSTIC:
            return
        for n_legs in (2, 3, 4):
            entry: dict[str, Any] = {
                "group_prefix": str(group_prefix),
                "sport_label": str(sport_label),
                "size": int(n_legs),
                "attempted": False,
                "reason": "",
                "best_mode": "",
                "best_payout": 0.0,
                "power_payout": 0.0,
                "flex_payout": 0.0,
            }
            if pool_df is None or len(pool_df) < n_legs:
                entry["reason"] = f"insufficient_pool (pool={0 if pool_df is None else len(pool_df)})"
                _diag_24.append(entry)
                continue
            rows = (
                _select_cross_sport_unique_rows(pool_df, n_legs)
                if require_multi_sport
                else _select_top_unique_rows(pool_df, n_legs)
            )
            if len(rows) < n_legs:
                entry["reason"] = "dedupe_collision"
                _diag_24.append(entry)
                continue
            entry["attempted"] = True
            t_power = _finalize_structure_ticket_dict(
                rows,
                "power",
                sport_label,
                "power",
                len(rows),
                None,
                False,
            )
            t_flex = _finalize_structure_ticket_dict(
                rows,
                "flex",
                sport_label,
                "flex",
                len(rows),
                None,
                False,
            )
            p_pow = float((t_power or {}).get("power_payout") or 0.0)
            p_flex = float((t_flex or {}).get("flex_payout") or 0.0)
            entry["power_payout"] = p_pow
            entry["flex_payout"] = p_flex
            if p_pow >= p_flex:
                entry["best_mode"] = "power"
                entry["best_payout"] = p_pow
            else:
                entry["best_mode"] = "flex"
                entry["best_payout"] = p_flex
            if (p_pow >= float(MIN_WEB_PAYOUT_X)) or (p_flex >= float(MIN_WEB_PAYOUT_X)):
                entry["reason"] = "passes"
            else:
                entry["reason"] = "payout_below_3x"
            _diag_24.append(entry)

    def _multi_sport_fix_indices(work: pd.DataFrame, sel_idx: list[int], used: set[int]) -> list[int] | None:
        """If selection is single-sport, try swapping the last leg for an unused other-sport row."""
        if len(sel_idx) < 2:
            return sel_idx
        rows = [work.iloc[i] for i in sel_idx]
        sports = {str(r.get("sport", "")).strip().upper() for r in rows if str(r.get("sport", "")).strip()}
        if len(sports) >= 2:
            return sel_idx
        base_sport = str(rows[0].get("sport", "")).strip().upper()
        # Replacement must use a player not already on the ticket (same as _select_cross_sport_unique_rows).
        players = {str(r.get("player", "")).strip().lower() for r in rows}
        player_props = {
            f"{str(r.get('player', '')).strip().lower()}::{_norm_prop_label(r.get('prop_type', ''))}"
            for r in rows
        }
        for j in range(len(work)):
            if j in used or j in sel_idx:
                continue
            row = work.iloc[j]
            sp = str(row.get("sport", "")).strip().upper()
            if not sp or sp == base_sport:
                continue
            p = str(row.get("player", "")).strip().lower()
            if not p or p in players:
                continue
            pp = f"{p}::{_norm_prop_label(row.get('prop_type', ''))}"
            if pp in player_props:
                continue
            return sel_idx[:-1] + [j]
        return None

    def _mixed_picktype_fix_indices(work: pd.DataFrame, sel_idx: list[int], used: set[int]) -> list[int] | None:
        """
        Ensure a "Mixed" ticket has at least one Standard and one Goblin leg.
        If missing one side, attempt swapping the last leg for an unused row
        of the missing pick type while preserving unique-player/prop constraints.
        """
        if len(sel_idx) < 2:
            return sel_idx
        rows = [work.iloc[i] for i in sel_idx]

        def _pt_family(v: Any) -> str:
            s = str(v or "").strip().upper()
            if s == "STANDARD":
                return "STANDARD"
            if s == "GOBLIN":
                return "GOBLIN"
            return s

        pts = {_pt_family(r.get("pick_type")) for r in rows}
        has_std = "STANDARD" in pts
        has_gob = "GOBLIN" in pts
        if has_std and has_gob:
            return sel_idx

        missing = "GOBLIN" if not has_gob else "STANDARD"
        players = {str(r.get("player", "")).strip().lower() for r in rows}
        player_props = {
            f"{str(r.get('player', '')).strip().lower()}::{_norm_prop_label(r.get('prop_type', ''))}"
            for r in rows
        }
        for j in range(len(work)):
            if j in used or j in sel_idx:
                continue
            row = work.iloc[j]
            if _pt_family(row.get("pick_type")) != missing:
                continue
            p = str(row.get("player", "")).strip().lower()
            if not p or p in players:
                continue
            pp = f"{p}::{_norm_prop_label(row.get('prop_type', ''))}"
            if pp in player_props:
                continue
            return sel_idx[:-1] + [j]
        return None

    def _emit_groups_for_pool(
        group_prefix: str,
        sport_label: str,
        pool_df: pd.DataFrame,
        bg_hdr: str,
        require_multi_sport: bool = False,
        require_picktype_mix: bool = False,
    ) -> None:
        """Exhaust the pool: repeated greedy N-leg tickets (unique player + unique player+prop), edge order."""
        if pool_df is None or len(pool_df) < 2:
            return
        work = pool_df.reset_index(drop=True)
        max_n = min(int(len(work)), 6)
        for n_legs in range(2, max_n + 1):
            used: set[int] = set()
            ticket_num = 0
            while True:
                sel_idx: list[int] = []
                players: set[str] = set()
                player_props: set[str] = set()
                for i in range(len(work)):
                    if i in used:
                        continue
                    row = work.iloc[i]
                    p = str(row.get("player", "")).strip().lower()
                    if not p or p in players:
                        continue
                    pp = f"{p}::{_norm_prop_label(row.get('prop_type', ''))}"
                    if pp in player_props:
                        continue
                    sel_idx.append(i)
                    players.add(p)
                    player_props.add(pp)
                    if len(sel_idx) >= n_legs:
                        break
                if len(sel_idx) < n_legs:
                    break
                if require_multi_sport:
                    fixed = _multi_sport_fix_indices(work, sel_idx, used)
                    if fixed is None:
                        for i in sel_idx:
                            used.add(i)
                        continue
                    sel_idx = fixed
                if require_picktype_mix:
                    fixed = _mixed_picktype_fix_indices(work, sel_idx, used)
                    if fixed is None:
                        for i in sel_idx:
                            used.add(i)
                        continue
                    sel_idx = fixed
                rows = [work.iloc[i].to_dict() for i in sel_idx]
                tickets: list[dict] = []
                min_px = None
                gp_u = str(group_prefix or "").upper()
                if "GOBLIN" in gp_u and int(n_legs) in (3, 4):
                    min_px = float(MIN_WEB_PAYOUT_X_GOBLIN_SHORT)
                p_ticket = _build_mode_ticket(rows, sport_label, "power", min_payout_x=min_px)
                if p_ticket is not None:
                    tickets.append(p_ticket)
                f_ticket = _build_mode_ticket(rows, sport_label, "flex", min_payout_x=min_px)
                if f_ticket is not None:
                    tickets.append(f_ticket)
                if not tickets:
                    for i in sel_idx:
                        used.add(i)
                    continue
                ticket_num += 1
                display = f"{group_prefix} {n_legs}-Leg #{ticket_num}"
                sname = _excel_ticket_sheet_title_unique(display, wb.sheetnames)
                write_ticket_sheet(wb, tickets, sname, bg_hdr, label=display)
                all_ticket_groups.append((display, tickets, None))
                _group_counts_by_size[group_prefix][n_legs] += 1
                print(
                    f"  {display}: {len(tickets)} slip(s) | "
                    f"Power {float(tickets[0].get('power_payout') or 0.0):.2f}x / "
                    f"Flex {float(tickets[0].get('flex_payout') or 0.0):.2f}x | "
                    f"EV {str(tickets[0].get('ev_tag', ''))}"
                )
                for i in sel_idx:
                    used.add(i)

    sport_pool_map: list[tuple[str, pd.DataFrame | None]] = [
        ("NBA", nba_pool),
        ("CBB", cbb_pool),
        ("CFB", cfb_pool),
        ("NHL", pool(nhl) if nhl is not None and len(nhl) > 0 else None),
        ("SOCCER", pool(soccer) if soccer is not None and len(soccer) > 0 else None),
        ("TENNIS", pool(tennis) if tennis is not None and len(tennis) > 0 else None),
        ("WNBA", pool(wnba) if wnba is not None and len(wnba) > 0 else None),
        ("MLB", mlb_pool),
        ("NFL", nfl_pool),
        ("NBA1Q", pool(nba1q) if nba1q is not None and len(nba1q) > 0 else None),
        ("NBA1H", pool(nba1h) if nba1h is not None and len(nba1h) > 0 else None),
        ("WCBB", pool(wcbb) if wcbb is not None and len(wcbb) > 0 else None),
    ]
    sports_to_build = [s for s, _ in sport_pool_map]
    eligible_sports = [s for s, df_ in sport_pool_map if df_ is not None and len(df_) > 0]
    print(f"  [handoff] sports in eligible pool: {eligible_sports}")
    print(f"  [handoff] sports in assembly loop: {sports_to_build}")

    for sport_label, sdf in sport_pool_map:
        if sdf is None or len(sdf) == 0:
            _zero_ticket_reasons[sport_label] = "empty eligible pool"
            continue
        s_pool = sdf.copy()
        if "pick_type" in s_pool.columns:
            pt = s_pool["pick_type"].astype(str).str.upper().str.strip()
            std_pool = s_pool[pt.eq("STANDARD")].copy()
            gob_pool = s_pool[pt.eq("GOBLIN")].copy()
        else:
            std_pool = pd.DataFrame(columns=s_pool.columns)
            gob_pool = pd.DataFrame(columns=s_pool.columns)
        mix_pool = s_pool.copy()
        print(f"  [handoff] {sport_label} std={len(std_pool)} gob={len(gob_pool)} mix={len(mix_pool)}")
        _diagnose_2_4(f"{sport_label} Standard", sport_label, std_pool)
        _diagnose_2_4(f"{sport_label} Goblin", sport_label, gob_pool)
        _diagnose_2_4(f"{sport_label} Mixed", sport_label, mix_pool)
        bg = _header_for_sport(sport_label)
        _emit_groups_for_pool(f"{sport_label} Standard", sport_label, std_pool, bg)
        _emit_groups_for_pool(f"{sport_label} Goblin", sport_label, gob_pool, bg)
        _emit_groups_for_pool(f"{sport_label} Mixed", sport_label, mix_pool, bg, require_picktype_mix=True)
        made = sum(_group_counts_by_size.get(f"{sport_label} Standard", Counter()).values()) + sum(
            _group_counts_by_size.get(f"{sport_label} Goblin", Counter()).values()
        ) + sum(_group_counts_by_size.get(f"{sport_label} Mixed", Counter()).values())
        if made == 0:
            _zero_ticket_reasons[sport_label] = "no 2-6 leg ticket passed uniqueness/payout constraints"

    x_frames = [df for _, df in sport_pool_map if df is not None and len(df) > 0]
    if x_frames:
        x_all = pd.concat(x_frames, ignore_index=True)
        x_pt = x_all["pick_type"].astype(str).str.upper().str.strip() if "pick_type" in x_all.columns else pd.Series([""] * len(x_all))
        x_std = x_all[x_pt.eq("STANDARD")].copy()
        x_gob = x_all[x_pt.eq("GOBLIN")].copy()
        x_mix = x_all.copy()
        _diagnose_2_4("X-Sport Standard", "MIX", x_std, require_multi_sport=True)
        _diagnose_2_4("X-Sport Goblin", "MIX", x_gob, require_multi_sport=True)
        _diagnose_2_4("X-Sport Mixed", "MIX", x_mix, require_multi_sport=True)
        _emit_groups_for_pool("X-Sport Standard", "MIX", x_std, C["hdr_mix"], require_multi_sport=True)
        _emit_groups_for_pool("X-Sport Goblin", "MIX", x_gob, C["hdr_mix"], require_multi_sport=True)
        _emit_groups_for_pool("X-Sport Mixed", "MIX", x_mix, C["hdr_mix"], require_multi_sport=True, require_picktype_mix=True)
    else:
        _zero_ticket_reasons["X-SPORT"] = "no eligible rows for cross-sport pool"

    print("\n[Ticket Group Counts]")
    if _group_counts_by_size:
        for gname in sorted(_group_counts_by_size.keys()):
            print(f"  {gname}: {dict(sorted(_group_counts_by_size[gname].items()))}")
    else:
        print("  (no ticket groups generated)")
    print("\n[Exhaustive: prop-type mix across emitted ticket legs]")
    _leg_props_by_sport: dict[str, Counter[str]] = defaultdict(Counter)
    _leg_props_all: Counter[str] = Counter()
    for _gname, _tickets, _ in all_ticket_groups:
        _sk = _group_sport(str(_gname)) or "ALL"
        for _t in _tickets:
            _legs = list(_t.get("legs") or _t.get("rows") or [])
            for _leg in _legs:
                if not isinstance(_leg, dict):
                    continue
                _pt = str(_leg.get("prop_type") or "").strip()
                if _pt:
                    _leg_props_by_sport[_sk][_pt] += 1
                    _leg_props_all[_pt] += 1
    if _leg_props_all:
        for _sk in sorted(_leg_props_by_sport.keys()):
            _top = _leg_props_by_sport[_sk].most_common(12)
            _n_dist = len(_leg_props_by_sport[_sk])
            print(f"  {_sk}: {_n_dist} prop types | top {dict(_top)}")
        print(f"  ALL sports: {len(_leg_props_all)} distinct | top {dict(_leg_props_all.most_common(15))}")
    else:
        print("  (no legs to summarize)")
    if _zero_ticket_reasons:
        print("\n[Zero-Ticket Sports]")
        for sname in sorted(_zero_ticket_reasons.keys()):
            print(f"  {sname}: {_zero_ticket_reasons[sname]}")
    if DEBUG_PAYOUT_DIAGNOSTIC and _diag_24:
        print("\n[2-4 LEG REJECT DIAGNOSTIC]")
        for d in sorted(_diag_24, key=lambda x: (str(x.get("group_prefix", "")), int(x.get("size", 0)))):
            gp = str(d.get("group_prefix", ""))
            sz = int(d.get("size", 0))
            attempted = bool(d.get("attempted", False))
            reason = str(d.get("reason", ""))
            mode = str(d.get("best_mode", ""))
            best = float(d.get("best_payout", 0.0))
            pp = float(d.get("power_payout", 0.0))
            fp = float(d.get("flex_payout", 0.0))
            if reason == "passes":
                print(f"  {gp} {sz}-Leg: PASSES (attempted={attempted}, power={pp:.2f}x, flex={fp:.2f}x)")
            elif reason == "payout_below_3x":
                print(
                    f"  {gp} {sz}-Leg: payout_below_3x "
                    f"(attempted={attempted}, best={best:.2f}x {mode}, power={pp:.2f}x, flex={fp:.2f}x)"
                )
            else:
                print(f"  {gp} {sz}-Leg: {reason} (attempted={attempted})")

    print("Writing slate sheets...")
    # Strict-mode guardrail: fail if mixed dates survived filtering.
    if not args.allow_cross_date_fallback:
        td = str(args.date).strip()[:10]
        to_check = [
            ("NBA", nba),
            ("CBB", cbb),
            ("NHL", nhl),
            ("Soccer", soccer),
            ("Tennis", tennis),
            ("MLB", mlb),
            ("Combined", combined),
        ]
        mixed = []
        for label, sdf in to_check:
            if sdf is None or len(sdf) == 0 or "game_date" not in sdf.columns:
                continue
            dated = sdf["game_date"].notna()
            gd = sdf["game_date"].astype(str).str[:10]
            if label == "Tennis":
                bad = sdf[dated & (gd < td)]
            elif label in ("NBA", "NBA1Q", "NBA1H"):
                bad = sdf[dated & (gd < td)]
            elif label == "Combined" and "sport" in sdf.columns:
                su = sdf["sport"].astype(str).str.upper()
                is_roll = su.isin(["TENNIS", "NBA", "NFL"])
                bad = sdf[dated & ((gd < td) | (~is_roll & (gd != td)))]
            else:
                bad = sdf[dated & (gd != td)]
            if len(bad) > 0:
                cts = bad["game_date"].astype(str).str[:10].value_counts().to_dict()
                mixed.append((label, cts))
        if mixed:
            msg = "; ".join([f"{lbl}: {cts}" for lbl, cts in mixed])
            raise ValueError(f"Strict date mode violation for target {args.date}: {msg}")

    full_slate_df = apply_full_slate_signal_columns(combined.copy())
    write_slate_sheet(
        wb,
        full_slate_df,
        "Full Slate",
        C["hdr"],
        "ALL",
        column_order=full_slate_column_order(full_slate_df),
        full_slate_visual=True,
    )
    write_slate_sheet(wb, nba, "NBA Slate", C["hdr_nba"], "NBA")
    write_slate_sheet(wb, cbb, "CBB Slate", C["hdr_cbb"], "CBB")
    if nhl is not None and len(nhl) > 0:
        write_slate_sheet(wb, nhl, "NHL Slate", C["hdr_nhl"], "NHL")
    if soccer is not None and len(soccer) > 0:
        write_slate_sheet(wb, soccer, "Soccer Slate", C["hdr_soccer"], "Soccer")
    if tennis is not None and len(tennis) > 0:
        write_slate_sheet(wb, tennis, "Tennis Slate", C["hdr_tennis"], "Tennis")
    if wcbb is not None and len(wcbb) > 0:
        write_slate_sheet(wb, wcbb, "WCBB Slate", C["hdr_wcbb"], "WCBB")
    if mlb is not None and len(mlb) > 0:
        write_slate_sheet(wb, mlb, "MLB Slate", C["hdr_mlb"], "MLB")
    if nba1q is not None and len(nba1q) > 0:
        write_slate_sheet(wb, nba1q, "NBA1Q Slate", C["hdr_nba1q"], "NBA1Q")
    if nba1h is not None and len(nba1h) > 0:
        write_slate_sheet(wb, nba1h, "NBA1H Slate", C["hdr_nba1h"], "NBA1H")
    if nfl is not None and len(nfl) > 0:
        write_slate_sheet(wb, nfl, "NFL Slate", C["hdr_nfl"], "NFL")

    _prio_hit = False
    _ticket_sort = str(args.ticket_candidate_sort)
    long_leg_sizes = [n for n in leg_sizes_runtime if n >= 4]
    _run_long_leg_supplement = bool(long_leg_sizes) and (
        bool(getattr(args, "long_leg_supplement", False)) or not bool(args.write_web)
    )
    if _run_long_leg_supplement:
        nhl_lg = pool(nhl) if nhl is not None and len(nhl) > 0 else None
        soc_lg = pool(soccer) if soccer is not None and len(soccer) > 0 else None
        ten_lg = pool(tennis) if tennis is not None and len(tennis) > 0 else None
        mlb_lg = mlb_pool if mlb_pool is not None and len(mlb_pool) > 0 else None
        final_long = build_final_web_ticket_groups(
            nba_pool,
            cbb_pool,
            nhl_pool=nhl_lg,
            soccer_pool=soc_lg,
            tennis_pool=ten_lg,
            mlb_pool=mlb_lg,
            min_hit_rate=float(thresholds.get("min_hit_rate", 0.65)),
            min_edge=float(thresholds.get("min_edge") or 0.0),
            min_rank=thresholds.get("min_rank"),
            ticket_leg_sizes=long_leg_sizes,
            leg_min_hit_by_n=leg_min_hit_by_n,
            prioritize_ticket_hit=_prio_hit,
            ticket_sort_mode=_ticket_sort,
            player_ticket_counts=counters["player_ticket_counts"],
            max_tickets_per_group=int(args.max_tickets),
        )
        for gname, tix, _bg in final_long:
            display = str(gname)
            sname = _excel_ticket_sheet_title_unique(display, wb.sheetnames)
            write_ticket_sheet(wb, tix, sname, C["hdr_sum"], label=display)
            all_ticket_groups.append((display, tix, _bg))
        print(f"  [long-legs] added {len(final_long)} long-leg sheet(s) for leg sizes {long_leg_sizes}")
    elif long_leg_sizes and bool(args.write_web):
        print(
            f"  [long-legs] skipped (--write-web default; main emission already has 4–6 leg tickets). "
            f"Pass --long-leg-supplement to add the slow extra workbook pass for sizes {long_leg_sizes}."
        )

    _pre_slips = sum(len(t[1]) for t in all_ticket_groups)
    _lc_groups_pre: Counter[int] = Counter()
    for _sn, _tickets, _ in all_ticket_groups:
        if not _tickets:
            continue
        _t0 = _tickets[0]
        _nl_g = int(_t0.get("n_legs") or 0) or len(_t0.get("rows") or [])
        if _nl_g > 0:
            _lc_groups_pre[_nl_g] += 1
    print(
        f"  [verify] groups by leg count (pre-dedupe): {dict(sorted(_lc_groups_pre.items()))}"
    )
    print(
        f"  [verify] pre-dedupe: {len(all_ticket_groups)} groups, {_pre_slips} slips | "
        f"PAYOUT power n=4,5,6: {PAYOUT[4]['power']}, {PAYOUT[5]['power']}, {PAYOUT[6]['power']} "
        f"(compute_ticket_ev SWEEP_PAYOUT[4,5,6] match: 10.0, 20.0, 40.0)"
    )

    _pre_dedupe_n = len(all_ticket_groups)
    _groups_pre_dedupe_snapshot = list(all_ticket_groups)
    all_ticket_groups, _n_groups_before_dedupe, _n_groups_after_dedupe = dedupe_ticket_groups_by_leg_set(
        all_ticket_groups
    )
    print(
        f"  [dedupe] ticket groups: {_n_groups_before_dedupe} -> {_n_groups_after_dedupe} "
        f"({_n_groups_before_dedupe - _n_groups_after_dedupe} duplicate leg sets removed)"
    )
    if bool(args.no_diversity_prune):
        print("  [diversity] skipped (--no-diversity-prune)")
    else:
        _jg_max = float(_diversity_cfg.get("max_jaccard_overlap", 0.55))
        all_ticket_groups, _jg_b, _jg_a = enforce_group_jaccard_diversity(
            all_ticket_groups,
            max_jaccard_overlap=_jg_max,
        )
        if _jg_b != _jg_a:
            print(
                f"  [dedupe-jaccard] ticket groups: {_jg_b} -> {_jg_a} "
                f"({_jg_b - _jg_a} high-overlap groups removed; max_jaccard={_jg_max:.2f})"
            )
        if bool(_diversity_cfg.get("enabled", True)):
            all_ticket_groups = _apply_diversity_filter_to_ticket_groups(all_ticket_groups, _diversity_cfg)
            print(f"  [diversity] groups after filter: {len(all_ticket_groups)}")
    # Diversity can drop every 3-leg group while leaving 4+ — re-check so /tickets always has Flex 3 when 4+ exists.
    apply_guaranteed_flex3_backfill_from_four_plus(all_ticket_groups, counters, wb)
    _post_slips = sum(len(t[1]) for t in all_ticket_groups)
    _lc_groups_post: Counter[int] = Counter()
    for _sn, _tickets, _ in all_ticket_groups:
        if not _tickets:
            continue
        _t0 = _tickets[0]
        _nl_g = int(_t0.get("n_legs") or 0) or len(_t0.get("rows") or [])
        if _nl_g > 0:
            _lc_groups_post[_nl_g] += 1
    print(
        f"  [verify] groups by leg count (post-dedupe): {dict(sorted(_lc_groups_post.items()))}"
    )
    print(f"  [verify] post-dedupe: {len(all_ticket_groups)} groups, {_post_slips} slips")
    _div_audit = compute_ticket_diversity_audit(all_ticket_groups)
    print(
        "  [audit] ticket diversity: "
        f"tickets={_div_audit.get('tickets', 0)} "
        f"unique={_div_audit.get('unique_leg_sets', 0)} "
        f"dup_rate={float(_div_audit.get('duplicate_rate', 0.0)):.4f} "
        f"jaccard_mean={float(_div_audit.get('jaccard_mean', 0.0)):.4f} "
        f"jaccard_p90={float(_div_audit.get('jaccard_p90', 0.0)):.4f}"
    )
    try:
        _audit_dir = Path(REPO_ROOT) / "data" / "reports" / "ticket_diversity"
        _audit_dir.mkdir(parents=True, exist_ok=True)
        _audit_out = _audit_dir / f"ticket_diversity_audit_{args.date}.json"
        _audit_payload = {
            "generated_at_utc": datetime.now(timezone.utc).isoformat(),
            "date": str(args.date),
            "source_output": str(args.output),
            "audit": _div_audit,
        }
        _audit_out.write_text(json.dumps(_audit_payload, indent=2), encoding="utf-8")
        print(f"  [audit] wrote diversity audit -> {_audit_out}")
    except Exception as _audit_exc:
        print(f"  [WARN] failed writing diversity audit: {_audit_exc}")
    _kept_ticket_sheet_names = {str(g[0]) for g in all_ticket_groups}
    for _ent in _groups_pre_dedupe_snapshot:
        _sn = str(_ent[0])
        if _sn not in _kept_ticket_sheet_names and _sn in wb.sheetnames:
            try:
                wb.remove(wb[_sn])
            except Exception:
                pass

    for _gn, _tickets, _bg in all_ticket_groups:
        for _ti in _tickets:
            enrich_ticket_curve_payouts(_ti, stake_unit=float(args.curve_stake_usd))

    write_summary(wb, nba, cbb, combined, all_ticket_groups, args.date, thresholds,
                  nhl=nhl, soccer=soccer, tennis=tennis, wcbb=wcbb, mlb=mlb, nba1q=nba1q, nba1h=nba1h,
                  nfl=nfl)

    # Reorder: put SUMMARY + slate sheets at the front
    desired_first = [
        "SUMMARY", "Full Slate", "NBA Slate", "CBB Slate", "NHL Slate", "Soccer Slate", "Tennis Slate",
        "WCBB Slate", "MLB Slate", "NFL Slate", "NBA1Q Slate", "NBA1H Slate",
    ]
    for sname in reversed(desired_first):
        if sname in wb.sheetnames:
            wb.move_sheet(wb[sname], offset=-(len(wb.sheetnames) - 1))

    wb.save(args.output)
    print(f"\n[OK] Saved -> {args.output}")
    print(f"   Sheets ({len(wb.sheetnames)}): {wb.sheetnames}")

    if args.write_web:
        print("\nWriting web outputs...")
        if all_ticket_groups:
            payload = ticket_groups_to_payload(
                all_ticket_groups,
                args.date,
                thresholds,
                bankroll=max(0.0, float(args.bankroll)),
                curve_stake_usd=float(args.curve_stake_usd),
            )
            n_groups = len(payload["groups"])
            n_slips = sum(len(g["tickets"]) for g in payload["groups"])
            print(f"  Web payload: {n_groups} groups, {n_slips} slips (workbook — all sports).")
            gated_preview = filter_positive_ev_tickets_payload(
                payload,
                apply_template_cap=bool(args.web_template_cap),
            )
            print_positive_ev_gate_report(gated_preview)
        else:
            print("  WARNING: workbook produced 0 groups — falling back to FINAL builder.")
            nhl_pool_web = pool(nhl) if nhl is not None and len(nhl) > 0 else None
            soccer_pool_web = pool(soccer) if soccer is not None and len(soccer) > 0 else None
            tennis_pool_web = pool(tennis) if tennis is not None and len(tennis) > 0 else None
            mlb_pool_web = mlb_pool if mlb_pool is not None and len(mlb_pool) > 0 else None
            final_groups = build_final_web_ticket_groups(
                nba_pool,
                cbb_pool,
                nhl_pool=nhl_pool_web,
                soccer_pool=soccer_pool_web,
                tennis_pool=tennis_pool_web,
                mlb_pool=mlb_pool_web,
                min_hit_rate=thresholds.get("min_hit_rate", 0.65),
                min_edge=thresholds.get("min_edge", 0.5),
                min_rank=thresholds.get("min_rank", 5.0),
                ticket_leg_sizes=leg_sizes_runtime,
                leg_min_hit_by_n=leg_min_hit_by_n,
                prioritize_ticket_hit=bool(args.prioritize_ticket_hit),
                ticket_sort_mode=str(args.ticket_candidate_sort),
                max_tickets_per_group=int(args.max_tickets),
            )
            final_groups, _fg_b, _fg_a = dedupe_ticket_groups_by_leg_set(final_groups)
            if _fg_b != _fg_a:
                print(f"  [dedupe] FINAL fallback groups: {_fg_b} -> {_fg_a}")
            if bool(args.no_diversity_prune):
                print("  [diversity] FINAL fallback skipped (--no-diversity-prune)")
            else:
                final_groups, _fj_b, _fj_a = enforce_group_jaccard_diversity(
                    final_groups,
                    max_jaccard_overlap=float(_diversity_cfg.get("max_jaccard_overlap", 0.55)),
                )
                if _fj_b != _fj_a:
                    print(
                        f"  [dedupe-jaccard] FINAL fallback groups: {_fj_b} -> {_fj_a} "
                        f"(max_jaccard={float(_diversity_cfg.get('max_jaccard_overlap', 0.55)):.2f})"
                    )
                if bool(_diversity_cfg.get("enabled", True)):
                    final_groups = _apply_diversity_filter_to_ticket_groups(final_groups, _diversity_cfg)
                    print(f"  [diversity] FINAL fallback groups: {len(final_groups)}")
            payload = ticket_groups_to_payload(
                final_groups,
                args.date,
                thresholds,
                bankroll=max(0.0, float(args.bankroll)),
                curve_stake_usd=float(args.curve_stake_usd),
            )
            n_groups = len(payload["groups"])
            n_slips = sum(len(g["tickets"]) for g in payload["groups"])
            print(f"  Web payload: {n_groups} groups, {n_slips} slips (FINAL fallback).")
            gated_preview = filter_positive_ev_tickets_payload(
                payload,
                apply_template_cap=bool(args.web_template_cap),
            )
            print_positive_ev_gate_report(gated_preview)
        _web_ev = not bool(args.no_web_ev_gate)
        write_web_outputs(
            payload,
            args.web_outdir,
            require_positive_ev=_web_ev,
            merge_existing_for_date=bool(args.merge_web_latest),
            apply_template_cap=bool(args.web_template_cap),
            discard_tracker=discard_tracker,
        )
        write_slate_json(nba, cbb, nhl, soccer, args.date, args.web_outdir,
                         wcbb=wcbb, mlb=mlb, nba1q=nba1q, nba1h=nba1h, tennis=tennis, nfl=nfl, wnba=wnba, cfb=cfb,
                         tennis_date=getattr(args, "tennis_date", None))
        try:
            ex_out = os.path.join(REPO_ROOT, "ui_runner", "data", "payout_ladder_examples.json")
            generate_payout_ladder_examples(payload, ex_out)
        except Exception as _pex:
            print(f"[WARN] Could not write payout ladder examples: {_pex}")
        if args.also_root:
            write_web_outputs(
                payload,
                outdir=".",
                require_positive_ev=_web_ev,
                merge_existing_for_date=bool(args.merge_web_latest),
                apply_template_cap=bool(args.web_template_cap),
                discard_tracker=discard_tracker,
            )
        # Avoid Windows console codepage issues with unicode checkmarks.
        print("[OK] Web outputs complete.")

    print("\n[TICKETS] -- SUMMARY -----------------------------------------")
    for sport, tickets in generated_tickets.items():
        pp = tickets.get("power_play")
        fl = tickets.get("flex")
        st = tickets.get("standard")
        ps3 = tickets.get("power_std3")
        g3 = tickets.get("goblin3")
        pp_legs = " + ".join(pp["legs"]) if pp else "SKIPPED"
        fl_legs = " + ".join(fl["legs"]) if fl else "SKIPPED"
        st_legs = " + ".join(st["legs"]) if st else "SKIPPED"
        ps3_legs = " + ".join(ps3["legs"]) if ps3 else "SKIPPED"
        g3_legs = " + ".join(g3["legs"]) if g3 else "SKIPPED"
        print(
            f"[TICKETS] {sport}: Power Play ({pp_legs}) | Flex ({fl_legs}) | Standard ({st_legs}) | "
            f"Pwr Std 3 ({ps3_legs}) | Goblin 3 ({g3_legs})"
        )

    print(f"[TICKETS] Fantasy Score excluded : {int(counters['fantasy_excluded_count'])} props removed")
    print(f"[TICKETS] Def tier filtered      : {int(counters['def_tier_filtered_count'])} props removed")
    print(f"[TICKETS] Prop ban list filtered : {int(counters['ban_list_filtered_count'])} props removed")
    print(f"[TICKETS] Total eligible props   : {int(counters['total_eligible_count'])} props used")
    print("\n[TICKETS] -- DISCARD REASON REPORT ---------------------------")
    rep = discard_tracker.report().strip()
    print(rep if rep else "  (no discard data captured)")
    print("\n[TICKETS] -- FILTER FUNNEL (SURVIVAL BY STAGE) --------------")
    frep = funnel_tracker.report().strip()
    print(frep if frep else "  (no funnel checkpoints recorded)")
    try:
        discard_out = os.path.join(REPO_ROOT, "ui_runner", "data", "discard_report_latest.json")
        os.makedirs(os.path.dirname(discard_out), exist_ok=True)
        with open(discard_out, "w", encoding="utf-8") as _f:
            json.dump(
                {
                    "build_date": str(args.date),
                    "generated_at": datetime.now(timezone.utc).isoformat(),
                    "discards": discard_tracker.to_dict(),
                    "funnel": funnel_tracker.to_dict(),
                    "final_pool_size": int(counters.get("total_eligible_count", 0)),
                },
                _f,
                indent=2,
                ensure_ascii=False,
            )
        print(f"[TICKETS] Discard sidecar        : {discard_out}")
    except Exception as _dex:
        print(f"[TICKETS] Discard sidecar WARN   : {_dex}")
    print("[TICKETS] ----------------------------------------------------")


# ── Web render helper ─────────────────────────────────────────────────────────

_SPORT_ACCENT: dict[str, str] = {
    "NBA":    "#36A2FF",
    "WNBA":   "#FF8AC6",
    "CBB":    "#2ECC71",
    "NHL":    "#9B59FF",
    "SOCCER": "#7DFF6B",
    "TENNIS": "#F39C12",
    "MLB":    "#FF5A5F",
    "WCBB":   "#FF66CC",
    "NBA1Q":  "#00E5FF",
    "NBA1H":  "#1ABC9C",
    "CROSS":  "#C77DFF",
    "MIX":    "#C77DFF",
}

_PICK_COLOR: dict[str, str] = {
    "goblin":   "#39ff6e",
    "demon":    "#ff4d4d",
    "standard": "#00e5ff",
}

_TICKETS_BUILT_PAYOUT_CSS = """<style>
.tickets-built .ticket-hdr-bracket {
  font-family: "Bebas Neue", sans-serif;
  font-size: clamp(14px, 1.5vw, 17px);
  letter-spacing: 0.06em;
  color: var(--text);
  border: 1px solid rgba(255,255,255,0.14);
  border-radius: 6px;
  padding: 2px 8px;
  background: rgba(0,0,0,0.2);
}
.tickets-built .payout-rec-badge {
  font-family: "Inter", sans-serif;
  font-size: clamp(11px, 1.1vw, 13px);
  border: 1px solid rgba(255,255,255,0.16);
  border-radius: 6px;
  padding: 3px 10px;
  background: rgba(0,0,0,0.22);
}
.tickets-built .payout-x-badge {
  font-family: "Inter", sans-serif;
  font-size: clamp(11px, 1.1vw, 13px);
  color: var(--cyan);
  border: 1px solid rgba(0,229,255,0.28);
  border-radius: 6px;
  padding: 3px 10px;
  background: rgba(0,229,255,0.06);
}
.tickets-built .ev-strong { color: #00ff88; font-weight: bold; }
.tickets-built .ev-ok { color: #88ccff; }
.tickets-built .ev-marginal { color: #ffaa00; }
.tickets-built .ev-low { color: #ff8844; }
.tickets-built .ev-skip { color: #ff4444; }
.tickets-built .ticket-filter-pill[data-filter="top-payout"].active {
  border-color: rgba(255, 215, 0, 0.42);
  color: #ffd54f;
}
.tickets-built .payout-source-badge {
  font-family: "Inter", sans-serif;
  font-size: 11px;
  margin-left: 6px;
  white-space: nowrap;
}
.tickets-built .payout-source-exact { color: #4caf50; }
.tickets-built .payout-source-calibrated { color: #ffc107; }
.tickets-built .payout-source-fallback { color: #9e9e9e; }
.tickets-built .leg-game-log-wrap { flex: 1; min-width: 280px; max-width: 560px; }
.tickets-built table.leg-game-log { width: 100%; font-size: 13px; border-collapse: collapse; }
.tickets-built table.leg-game-log th {
  text-align: left; padding: 6px 8px; border-bottom: 1px solid rgba(255,255,255,0.14);
  color: var(--muted); font-size: 11px; text-transform: uppercase; letter-spacing: 0.06em;
  font-family: "Bebas Neue", sans-serif;
}
.tickets-built table.leg-game-log td { padding: 6px 8px; border-bottom: 1px solid rgba(255,255,255,0.06); font-family: "Inter", sans-serif; }
.tickets-built table.leg-game-log tr:last-child td { border-bottom: none; }
.tickets-built .leg-game-log-empty { margin: 0; font-size: 12px; color: var(--muted); }
.tickets-built .leg-game-hit { color: #00ff88; font-weight: 600; }
.tickets-built .leg-game-miss { color: #c96a74; font-weight: 600; }
.tickets-built .ticket-filter-sort-wrap {
  display: inline-flex;
  align-items: center;
  gap: 8px;
  margin-right: 6px;
}
.tickets-built .ticket-filter-sort-label {
  font-size: 11px;
  letter-spacing: 0.06em;
  color: var(--muted);
  text-transform: uppercase;
}
.tickets-built .ticket-filter-sort {
  border-radius: 999px;
  border: 1px solid rgba(255,255,255,0.2);
  background: rgba(12,16,26,0.9);
  color: var(--text);
  font-size: 12px;
  padding: 5px 12px;
}
.tickets-built .ticket-filter-bar-action.active {
  border-color: rgba(255, 86, 86, 0.45);
  color: #ff8a8a;
}
.tickets-built .ticket-group-section.group-rec-strong .ticket-group-header { border-left: 4px solid #00ff88; }
.tickets-built .ticket-group-section.group-rec-ok .ticket-group-header { border-left: 4px solid #f0a500; }
.tickets-built .ticket-group-section.group-rec-marginal .ticket-group-header { border-left: 4px solid #ff9f43; }
.tickets-built .ticket-group-section.group-rec-skip .ticket-group-header { border-left: 4px solid #ff5c5c; opacity: 0.78; }
.tickets-built .best-ticket-row {
  display: flex;
  align-items: center;
  justify-content: space-between;
  gap: 12px;
  padding: 6px 0;
  border-bottom: 1px solid rgba(255,255,255,0.08);
  font-size: 13px;
}
.tickets-built .best-ticket-row:last-child { border-bottom: 0; }
.tickets-built .best-ticket-name { color: var(--text); font-weight: 600; }
.tickets-built .best-ticket-meta { font-size: 12px; }
.tickets-built .winrate-best-panel {
  margin: 0 0 20px 0;
  padding: 16px 18px;
  border-radius: 12px;
  border: 1px solid rgba(255, 215, 0, 0.35);
  background: linear-gradient(145deg, rgba(18, 22, 32, 0.98), rgba(8, 12, 20, 0.98));
  box-shadow: 0 8px 28px rgba(0, 0, 0, 0.35);
}
.tickets-built .winrate-best-panel .winrate-best-title {
  font-size: 11px;
  letter-spacing: 2px;
  text-transform: uppercase;
  color: #ffd54f;
  margin-bottom: 4px;
}
.tickets-built .winrate-best-panel .winrate-best-sub {
  font-size: 12px;
  color: var(--muted);
  margin-bottom: 12px;
}
.tickets-built .winrate-best-row {
  display: flex;
  align-items: flex-start;
  justify-content: space-between;
  gap: 12px;
  padding: 10px 0;
  border-bottom: 1px solid rgba(255, 255, 255, 0.08);
  cursor: pointer;
}
.tickets-built .winrate-best-row:last-child { border-bottom: 0; }
.tickets-built .winrate-best-row:hover { background: rgba(255, 215, 0, 0.04); }
.tickets-built .winrate-best-rank { color: #ffd54f; font-weight: 700; min-width: 28px; }
.tickets-built .winrate-best-name { color: var(--text); font-weight: 600; flex: 1; }
.tickets-built .winrate-best-legs { font-size: 12px; color: var(--muted); margin-top: 4px; }
.tickets-built .winrate-best-leg { line-height: 1.35; }
.tickets-built .winrate-best-leg + .winrate-best-leg { margin-top: 2px; }
.tickets-built .winrate-best-stats { text-align: right; font-size: 12px; white-space: nowrap; }
.tickets-built .winrate-best-pwin { color: #00ff88; font-weight: 700; font-size: 14px; }
.tickets-built .winrate-best-pwin-sub { font-size: 10px; color: var(--muted); margin-top: 2px; }
.tickets-built .winrate-best-warn { font-size: 10px; color: #f0a500; margin-top: 4px; }
.tickets-built .ticket-pwin-ev-badge {
  font-size: 12px;
  color: #00ff88;
  margin-left: 8px;
  font-weight: 600;
}
</style>"""


def _payout_ev_class(rec: str) -> str:
    u = (rec or "").strip().upper()
    if u == "STRONG":
        return "ev-strong"
    if u == "OK":
        return "ev-ok"
    if u == "MARGINAL":
        return "ev-marginal"
    if u == "LOW":
        return "ev-low"
    if u == "SKIP":
        return "ev-skip"
    return "ev-skip"


def _payout_rec_prefix(rec: str) -> str:
    u = (rec or "").strip().upper()
    if u == "STRONG":
        return "⚡"
    if u == "OK":
        return "✅"
    if u == "MARGINAL":
        return "⚠"
    if u == "LOW":
        return "▼"
    if u == "SKIP":
        return "⏭"
    return "•"


def _payout_source_badge_html(source: str) -> str:
    src = str(source or "calibrated").strip().lower()
    if src == "exact":
        dot, label = "●", "Exact"
    elif src == "fallback":
        dot, label = "●", "~"
    else:
        src = "calibrated"
        dot, label = "●", "Est"
    return (
        f'<span class="payout-source-badge payout-source-{_h(src)}" title="Payout source: {_h(src)}">'
        f"{dot} {_h(label)}</span>"
    )


def _h(v) -> str:
    """HTML-escape a value."""
    import html as _html
    return _html.escape(str(v)) if v is not None else ""


def _pct(v, decimals: int = 0) -> str:
    try:
        return f"{float(v) * 100:.{decimals}f}%"
    except (TypeError, ValueError):
        return "—"


def _fmt(v, decimals: int = 2, suffix: str = "") -> str:
    try:
        return f"{float(v):.{decimals}f}{suffix}"
    except (TypeError, ValueError):
        return "—"


def _sport_accent(sport: str) -> str:
    key = (sport or "").upper().split()[0]
    return _SPORT_ACCENT.get(key, "#6C3483")


def _group_sport(group_name: str) -> str:
    """Infer sport from group name for accent colouring."""
    name = (group_name or "").upper().replace("\u00a0", " ")
    if "NBA/CBB" in name or "NBA+CBB" in name or "NBA-CBB" in name:
        return "CROSS"
    if name.startswith("CROSS") or name.startswith("MIX"):
        return "CROSS"
    if name.startswith("X-SPORT") or "X-SPORT" in name:
        return "CROSS"
    for sp in ("NBA1Q", "NBA1H", "WNBA", "WCBB", "TENNIS", "SOCCER", "NHL", "MLB", "CBB", "NBA"):
        if sp in name:
            return sp
    return "NBA"


# Align with Slate Explorer sport order; cross-sport / mix buckets sort last.
_TICKET_GROUP_SPORT_SORT_ORDER: dict[str, int] = {
    "NBA": 0,
    "NBA1Q": 1,
    "NBA1H": 2,
    "CBB": 3,
    "WCBB": 4,
    "CFB": 5,
    "NFL": 6,
    "WNBA": 7,
    "MLB": 8,
    "NHL": 9,
    "SOCCER": 10,
    "TENNIS": 11,
    "CROSS": 10_000,
    "MIX": 10_000,
}


def _ticket_group_sort_rank(group_name: str) -> int:
    sk = _group_sport(group_name)
    return _TICKET_GROUP_SPORT_SORT_ORDER.get(sk, 999)


def _ticket_group_picktype_rank(group_name: str) -> int:
    """Order within a sport: Standard, Goblin, Mixed, then everything else."""
    name = (group_name or "").upper().replace("\u00a0", " ")
    if " STANDARD" in name:
        return 0
    if " GOBLIN" in name:
        return 1
    if " MIXED" in name:
        return 2
    return 9


def _ticket_group_leg_count(group_name: str) -> int:
    """Extract N from labels like '... 4-Leg #12' for stable ordering."""
    m = re.search(r"(\d+)\s*-\s*LEG", str(group_name or ""), flags=re.IGNORECASE)
    if not m:
        return 99
    try:
        return int(m.group(1))
    except (TypeError, ValueError):
        return 99


def _ticket_group_serial(group_name: str) -> int:
    """Extract trailing #number if present."""
    m = re.search(r"#\s*(\d+)\s*$", str(group_name or ""))
    if not m:
        return 999_999
    try:
        return int(m.group(1))
    except (TypeError, ValueError):
        return 999_999


_EV_REC_RANK = {"LOW": 0, "SKIP": 0, "MARGINAL": 1, "OK": 2, "STRONG": 3}


def _group_payout_confidence_score(tickets: list) -> float:
    """Max payout_confidence_score (sweep × p_all_win) across slips in a group."""
    best = 0.0
    for t in tickets:
        if not isinstance(t, dict):
            continue
        p = t.get("payout")
        if not isinstance(p, dict):
            continue
        raw = p.get("payout_confidence_score")
        if raw is None:
            continue
        try:
            v = float(raw)
            if math.isfinite(v) and v > best:
                best = v
        except (TypeError, ValueError):
            continue
    return best


def _slip_display_payout_multiplier(
    payout: dict | None, ticket: dict, group: dict
) -> float | None:
    """
    Headline all-hit multiplier for slip UI (not min-guarantee / goblin discount factor).
    Prefer sweep_payout, then ticket/group power/flex, then payout.payout fallback.
    """
    if isinstance(payout, dict):
        sp = payout.get("sweep_payout")
        if sp is not None:
            try:
                v = float(sp)
                if math.isfinite(v) and v > 0:
                    return v
            except (TypeError, ValueError):
                pass
    for k in ("power_payout", "flex_payout"):
        v = ticket.get(k)
        if v is None:
            v = group.get(k)
        if v is not None:
            try:
                vf = float(v)
                if math.isfinite(vf) and vf > 0:
                    return vf
            except (TypeError, ValueError):
                pass
    if isinstance(payout, dict):
        for k in ("payout", "min_guarantee"):
            v = payout.get(k)
            if v is not None:
                try:
                    vf = float(v)
                    if math.isfinite(vf) and vf > 0:
                        return vf
                except (TypeError, ValueError):
                    pass
    return None


def _ticket_group_filter_slugs(group_name: str) -> tuple[str, str, str]:
    """(data_sport, data_type, data_pick) lowercase slugs for /tickets filter pills."""
    name_u = (group_name or "").upper().replace("\u00a0", " ")
    sport_key = _group_sport(group_name)
    sport_sl = sport_key.lower()

    if " FLEX" in name_u or name_u.startswith("FLEX ") or " FLEX " in name_u:
        type_sl = "flex"
    elif "POWER" in name_u:
        type_sl = "power"
    else:
        type_sl = "power"

    if "GOBLIN" in name_u:
        pick_sl = "goblin"
    elif "DEMON" in name_u:
        pick_sl = "demon"
    else:
        pick_sl = "standard"

    return sport_sl, type_sl, pick_sl


def _group_ev_data_attr(tickets: list) -> str:
    """Strongest empirical payout recommendation across tickets in the group."""
    best_r = -1
    best_sl = ""
    for t in tickets:
        p = t.get("payout")
        if not isinstance(p, dict):
            continue
        rec = str(p.get("recommendation") or "").strip().upper()
        r = _EV_REC_RANK.get(rec, -1)
        if r > best_r:
            best_r = r
            best_sl = rec.lower() if rec in _EV_REC_RANK else ""
    return best_sl


def _group_ev_badge_summary_html(tickets: list) -> str:
    """Header line: best empirical EV among tickets with payout JSON."""
    best: tuple[float, str, str] | None = None
    for t in tickets:
        p = t.get("payout")
        if not isinstance(p, dict) or p.get("ev") is None:
            continue
        try:
            evf = float(p["ev"])
        except (TypeError, ValueError):
            continue
        if not math.isfinite(evf):
            continue
        rec = str(p.get("recommendation") or "")
        ev_cls = _payout_ev_class(rec)
        if best is None or evf > best[0]:
            best = (evf, rec, ev_cls)
    if best is None:
        return '<span class="group-ev-badge group-ev-badge--na">—</span>'
    evf, rec, ev_cls = best
    return f'<span class="group-ev-badge {ev_cls}">EV {_fmt(evf, 2)} — {_h(rec)}</span>'


def _group_hit_rate_score(tickets: list) -> float:
    vals: list[float] = []
    for t in tickets:
        if not isinstance(t, dict):
            continue
        v = t.get("avg_hit_rate")
        if v is None:
            continue
        try:
            vf = float(v)
        except (TypeError, ValueError):
            continue
        if math.isfinite(vf):
            vals.append(vf)
    if not vals:
        return 0.0
    return float(sum(vals) / len(vals))


def _tickets_filter_pills_html(attr_rows: list[dict]) -> str:
    """Dynamic filter bar from group-derived slugs (sport / power / flex / goblin / demon / strong)."""
    sports_seen: list[str] = []
    seen_sp: set[str] = set()
    has_power = has_flex = has_goblin = has_demon = has_strong = False
    for row in attr_rows:
        sp = row.get("sport") or ""
        if sp and sp not in seen_sp:
            seen_sp.add(sp)
            sports_seen.append(sp)
        if row.get("type") == "power":
            has_power = True
        if row.get("type") == "flex":
            has_flex = True
        if row.get("pick") == "goblin":
            has_goblin = True
        if row.get("pick") == "demon":
            has_demon = True
        if row.get("ev") == "strong":
            has_strong = True

    # Keep MLB, WNBA, and Soccer filter pills visible even when no groups were generated for that sport
    # (primary /tickets UI is driven by this bar; users expect the sport control to always exist).
    for _sp in ("mlb", "wnba", "soccer"):
        if _sp not in seen_sp:
            seen_sp.add(_sp)
            sports_seen.append(_sp)

    sport_order = (
        "nba",
        "nba1q",
        "nba1h",
        "wnba",
        "cbb",
        "wcbb",
        "nhl",
        "mlb",
        "soccer",
        "tennis",
        "cross",
        "mix",
    )
    sports_sorted = sorted(
        sports_seen,
        key=lambda s: (sport_order.index(s) if s in sport_order else 99, s),
    )

    def _pill(
        data_filter: str,
        label: str,
        *,
        active: bool = False,
        title_attr: str = "",
    ) -> str:
        cls = "ticket-filter-pill active" if active else "ticket-filter-pill"
        return (
            f'<button type="button" class="{cls}" data-filter="{_h(data_filter)}"'
            f"{title_attr}>{label}</button>"
        )

    chunks: list[str] = [
        '<div class="ticket-filter-bar" role="toolbar" aria-label="Filter ticket groups">',
        _pill("all", "ALL", active=True),
    ]
    for sp in sports_sorted:
        chunks.append(_pill(sp, sp.upper()))
    chunks.append(_pill("pp", "PP", title_attr=' title="Any leg priced from PrizePicks row"'))
    chunks.append(_pill("ud", "UD", title_attr=' title="Any leg from Underdog-only ladder row"'))
    chunks.append(_pill("dk", "DK", title_attr=' title="Any leg from DraftKings-only ladder row"'))
    if has_power:
        chunks.append(_pill("power", "POWER"))
    if has_flex:
        chunks.append(_pill("flex", "FLEX"))
    if has_goblin:
        chunks.append(_pill("goblin", "GOBLIN"))
    if has_demon:
        chunks.append(_pill("demon", "DEMON"))
    if has_strong:
        chunks.append(_pill("strong", "⚡ STRONG"))
    chunks.append(
        _pill(
            "top-payout",
            "🏆 TOP PAYOUT",
            title_attr=' title="Highest payout × win probability (top 3 groups)"',
        )
    )
    chunks.append(
        '<button type="button" class="ticket-filter-bar-action utp-bar-toggle" data-utp="toggle" '
        'id="uniform-buckets-toggle" style="border-radius:999px;" aria-expanded="false" '
        'title="Today&apos;s tickets grouped by realized hit-rate band">🎫 UNIFORM</button>'
    )
    chunks.append('<span class="ticket-filter-bar-spacer" aria-hidden="true"></span>')
    chunks.append(
        '<label class="ticket-filter-sort-wrap" for="ticket-sort-select">'
        '<span class="ticket-filter-sort-label">Sort</span>'
        '<select id="ticket-sort-select" class="ticket-filter-sort">'
        '<option value="ev_desc" selected>EV ↓</option>'
        '<option value="ev_asc">EV ↑</option>'
        '<option value="pwin_desc">P(WIN) ↓</option>'
        '<option value="pwin_asc">P(WIN) ↑</option>'
        '<option value="legs_desc">Legs ↓</option>'
        '<option value="group">Group #</option>'
        '<option value="hit_rate">Hit Rate</option>'
        '</select>'
        '</label>'
    )
    chunks.append(
        '<button type="button" class="ticket-filter-bar-action active" id="toggle-skip" '
        'style="border-radius:999px;" aria-pressed="true">HIDE SKIP</button>'
    )
    chunks.append('<button type="button" class="ticket-filter-bar-action" id="expand-all" style="border-radius:999px;">EXPAND ALL</button>')
    chunks.append('<button type="button" class="ticket-filter-bar-action" id="collapse-all" style="border-radius:999px;">COLLAPSE ALL</button>')
    chunks.append("</div>")
    return "".join(chunks)


def _tickets_fmt_line_plain(x) -> str:
    try:
        if x is None:
            return "—"
        xf = float(x)
        if abs(xf - round(xf)) < 1e-9:
            return str(int(round(xf)))
        return f"{xf:.2f}".rstrip("0").rstrip(".")
    except (TypeError, ValueError):
        return str(x) if x is not None else "—"


def _tickets_leg_parse_float(val: Any) -> float | None:
    if val is None:
        return None
    if isinstance(val, str) and not val.strip():
        return None
    try:
        xf = float(val)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(xf):
        return None
    return xf


def _tickets_leg_game_log_table_html(leg: dict) -> str:
    """
    Per-game posted line (line_g*) + actual (stat_g* / g*) — replaces hit/miss bar chart.
    """
    dir_u = str(leg.get("direction") or "").strip().upper()
    rows_html: list[str] = []
    for gi in range(1, 11):
        raw_stat = leg.get(f"stat_g{gi}")
        if raw_stat is None:
            raw_stat = leg.get(f"g{gi}")
        act = _tickets_leg_parse_float(raw_stat)
        ln_raw = leg.get(f"line_g{gi}")
        if ln_raw is None:
            ln_raw = leg.get(f"prop_line_g{gi}")
        line_at_game = _tickets_leg_parse_float(ln_raw)
        if act is None and line_at_game is None:
            continue
        act_disp = _tickets_fmt_line_plain(act) if act is not None else "—"
        line_disp = _tickets_fmt_line_plain(line_at_game) if line_at_game is not None else "—"
        res_disp = "—"
        res_cls = ""
        if act is not None and line_at_game is not None:
            if dir_u == "UNDER":
                ok = act <= line_at_game
            elif dir_u == "OVER":
                ok = act >= line_at_game
            else:
                ok = act >= line_at_game
            res_disp = "Hit" if ok else "Miss"
            res_cls = "leg-game-hit" if ok else "leg-game-miss"
        rows_html.append(
            f"<tr><td>{_h('G' + str(gi))}</td><td>{_h(line_disp)}</td><td>{_h(act_disp)}</td>"
            f'<td class="{res_cls}">{_h(res_disp)}</td></tr>'
        )
    if not rows_html:
        return (
            '<p class="leg-game-log-empty">No per-game line / actual series saved for this leg '
            "(stat_g1.. / line_g1..).</p>"
        )
    return (
        '<table class="leg-game-log" role="grid" aria-label="Recent games vs posted line">'
        "<thead><tr>"
        "<th>Game</th><th>Posted line</th><th>Actual</th><th>vs pick</th>"
        "</tr></thead><tbody>"
        + "".join(rows_html)
        + "</tbody></table>"
    )


def _tickets_leg_graph_row_html(leg: dict, row_id: str, table_cols: int) -> str:
    """Expandable row: stat pills + per-game line/actual table (tickets_built.html)."""
    l5_avg = leg.get("l5_avg")
    season_avg = leg.get("season_avg")
    l5_over = leg.get("l5_over")
    l5_under = leg.get("l5_under")
    l10_over = leg.get("l10_over")
    l10_under = leg.get("l10_under")
    line_val = leg.get("line")
    dir_txt = str(leg.get("direction") or "").upper()
    hr_val = leg.get("hit_rate")

    def _pill(label: str, val, fmt=None) -> str:
        if val is None:
            return ""
        if fmt:
            try:
                v = fmt(val)
            except Exception:
                v = str(val)
        else:
            v = str(val)
        return f'<div class="gstat"><div class="gstat-label">{_h(label)}</div><div class="gstat-val">{_h(v)}</div></div>'

    pills = "".join(
        [
            _pill("L5 Avg", l5_avg, lambda x: f"{float(x):.1f}"),
            _pill("Season Avg", season_avg, lambda x: f"{float(x):.1f}"),
            _pill("L5 Over", l5_over, lambda x: format_hit_window_fraction(5, x)),
            _pill("L5 Under", l5_under, lambda x: format_hit_window_fraction(5, x)),
            _pill("L10 Over", l10_over, lambda x: format_hit_window_fraction(10, x)),
            _pill("L10 Under", l10_under, lambda x: format_hit_window_fraction(10, x)),
            _pill("Hit Rate", hr_val, lambda x: f"{float(x) * 100:.0f}%"),
        ]
    )

    game_log_html = _tickets_leg_game_log_table_html(leg)
    sub = f"{leg.get('player', '')} · {leg.get('prop_type', '')} · Line {_tickets_fmt_line_plain(line_val)}"
    return f"""
<tr class="leg-graph-row" id="{_h(row_id)}">
  <td class="leg-graph-cell" colspan="{table_cols}">
    <div class="graph-wrap">
      <div style="flex:1;min-width:200px;">
        <div style="font-size:11px;color:var(--muted);margin-bottom:6px;text-transform:uppercase;letter-spacing:.5px;">{_h(sub)}</div>
        <div class="graph-stats">{pills}</div>
      </div>
      <div class="leg-game-log-wrap">
        {game_log_html}
      </div>
    </div>
  </td>
</tr>"""

def _winrate_best_leg_label(leg: dict) -> str:
    """One-line leg summary for Today's Best panel: player, prop, direction, line."""
    player = str(leg.get("player") or "").strip()
    prop = str(leg.get("prop_type") or leg.get("prop") or "").strip()
    direction = str(leg.get("direction") or "").strip().upper()
    if direction == "LOWER":
        direction = "UNDER"
    line_s = _tickets_fmt_line_plain(leg.get("line"))

    detail: list[str] = []
    if prop:
        detail.append(prop)
    if direction and line_s != "—":
        dir_short = "O" if direction in ("OVER", "O") else ("U" if direction in ("UNDER", "U") else direction)
        detail.append(f"{dir_short} {line_s}")
    elif line_s != "—":
        detail.append(line_s)
    elif direction:
        detail.append(direction)

    if player and detail:
        return f"{player} — {' · '.join(detail)}"
    if player:
        return player
    if detail:
        return " · ".join(detail)
    return "—"


def _winrate_best_panel_html(winrate_payload: dict | None = None) -> str:
    """Pinned panel: top 5 win-rate tickets (sorted by est_win_prob, bench legs filtered)."""
    _placeholder = (
        '<motionless class="winrate-best-panel" id="winrate-best-panel" aria-live="polite">'
        '<motionless class="winrate-best-title">⚡ TODAY&apos;S BEST — Highest Win Probability</motionless>'
        '<motionless class="winrate-best-sub">Win-rate tickets generating…</motionless>'
        "</motionless>"
    ).replace("motionless", "div")
    data = winrate_payload
    if data is None:
        path = Path(REPO_ROOT) / "ui_runner" / "templates" / "tickets_latest.json"
        if not path.is_file():
            return _placeholder
        try:
            with open(path, "r", encoding="utf-8") as f:
                data = json.load(f)
        except Exception:
            return _placeholder
    generated_at = str((data or {}).get("generated_at") or "")
    flat: list[tuple[float, dict, str]] = []
    for g in (data or {}).get("groups") or []:
        gn = str(g.get("group_name") or "Ticket")
        for t in g.get("tickets") or []:
            if not isinstance(t, dict):
                continue
            if _winrate_ticket_same_game_bench_stack(t):
                continue
            if any(_winrate_leg_bench_risk(leg) for leg in (t.get("legs") or []) if isinstance(leg, dict)):
                continue
            flat.append((_winrate_ticket_rank_score(t), t, gn))
    flat.sort(key=lambda x: -x[0])
    top = flat[:5]
    if not top:
        return (
            '<div class="winrate-best-panel" id="winrate-best-panel">'
            '<div class="winrate-best-title">⚡ TODAY&apos;S BEST — Highest Win Probability</div>'
            '<div class="winrate-best-sub">No qualifying tickets for this slate '
            '(deep-bench SUPPORT legs and same-game bench stacks are excluded). '
            'Rebuild win-rate JSON after the next ticket run.</div>'
            "</div>"
        )
    rows: list[str] = []
    for i, (rank_score, t, gn) in enumerate(top, start=1):
        legs = t.get("legs") or []
        leg_lines: list[str] = []
        for leg in legs:
            if isinstance(leg, dict):
                lbl = _winrate_best_leg_label(leg)
                if lbl and lbl != "—":
                    leg_lines.append(f'<div class="winrate-best-leg">{_h(lbl)}</div>')
        legs_html = "".join(leg_lines) if leg_lines else '<div class="winrate-best-leg">—</div>'
        n_legs = len(legs) or t.get("n_legs") or 0
        ev_v = t.get("ev_power")
        if ev_v is None and isinstance(t.get("payout"), dict):
            ev_v = (t.get("payout") or {}).get("ev")
        try:
            ev_f = float(ev_v) if ev_v is not None else 0.0
        except (TypeError, ValueError):
            ev_f = 0.0
        pay = t.get("payout_multiplier") or t.get("power_payout")
        try:
            pay_f = float(pay) if pay is not None else 0.0
        except (TypeError, ValueError):
            pay_f = 0.0
        pwin = _winrate_ticket_win_prob(t)
        pcash_opt = _winrate_ticket_panel_pcash_optional(t)
        pwin_sub = ""
        if pcash_opt is not None:
            pwin_sub = (
                f'<div class="winrate-best-pwin-sub">P(cash) {_fmt(pcash_opt * 100, 0)}%</div>'
            )
        rows.append(
            f'<div class="winrate-best-row" data-winrate-rank="{i}" role="button" tabindex="0">'
            f'<span class="winrate-best-rank">#{i}</span>'
            f'<span class="winrate-best-name">{_h(gn)}'
            f'<div class="winrate-best-legs">{legs_html}</div>'
            f'</span>'
            f'<span class="winrate-best-stats">'
            f'<div class="winrate-best-pwin">P(win) {_fmt(pwin * 100, 0)}%</div>'
            f'{pwin_sub}'
            f'<div>EV {_fmt(ev_f, 1)} · Payout {_fmt(pay_f, 1)}x · {int(n_legs)}-leg</div>'
            f"</span></div>"
        )
    sub_parts = ["Sorted by modeled win probability (est_win_prob); deep-bench SUPPORT legs excluded"]
    if generated_at:
        sub_parts.append(f"Updated: {generated_at}")
    sub = _h(" · ".join(sub_parts))
    body = "".join(rows)
    return (
        '<div class="winrate-best-panel" id="winrate-best-panel">'
        '<div class="winrate-best-title">⚡ TODAY&apos;S BEST — Highest Win Probability</div>'
        f'<div class="winrate-best-sub">{sub}</div>'
        f"{body}"
        "</div>"
    )


def _group_max_ev_for_ui_cap(group: dict) -> float:
    best = float("-inf")
    for t in group.get("tickets") or []:
        if not isinstance(t, dict):
            continue
        for key in ("ev_power", "est_ev"):
            v = t.get(key)
            if v is None:
                continue
            try:
                vf = float(v)
                if math.isfinite(vf):
                    best = max(best, vf)
            except (TypeError, ValueError):
                pass
        p = t.get("payout")
        if isinstance(p, dict) and p.get("ev") is not None:
            try:
                vf = float(p["ev"])
                if math.isfinite(vf):
                    best = max(best, vf)
            except (TypeError, ValueError):
                pass
    return float(best) if math.isfinite(best) else 0.0


def _parse_ui_group_bucket(group_name: str) -> tuple[str, str, int] | None:
    """Return (sport_key, Standard|Goblin|Mixed, n_legs) for exhaustive group names."""
    gn = (group_name or "").strip()
    m = re.match(r"^(.+?)\s+(Standard|Goblin|Mixed)\s+(\d+)-Leg", gn, flags=re.I)
    if not m:
        return None
    sport_raw = m.group(1).strip()
    pool = str(m.group(2) or "").strip().title()
    n = int(m.group(3))
    su = sport_raw.upper()
    sport_key = "X-Sport" if su.startswith("X-SPORT") else sport_raw.upper()
    return (sport_key, pool, n)


def _cap_ticket_groups_for_ui(groups: list, max_per_bucket: int) -> tuple[list, int, int]:
    """
    Keep the top ``max_per_bucket`` groups per (sport, pick-type bucket, n_legs) by max slip EV.
    Groups that do not match the name pattern are kept. Full JSON is unchanged; this is HTML-only.
    """
    if max_per_bucket <= 0 or not groups:
        return list(groups), len(groups), len(groups)
    buckets: dict[tuple[str, str, int], list[tuple[float, int, dict]]] = defaultdict(list)
    unbucketed: list[dict] = []
    for i, g in enumerate(groups):
        if not isinstance(g, dict):
            continue
        gn = str(g.get("group_name") or "")
        b = _parse_ui_group_bucket(gn)
        ev = _group_max_ev_for_ui_cap(g)
        if b is None:
            unbucketed.append(g)
            continue
        buckets[b].append((ev, i, g))
    out: list[dict] = []
    for _b, items in buckets.items():
        items.sort(key=lambda x: (-x[0], x[1]))
        out.extend([t[2] for t in items[:max_per_bucket]])
    out.extend(unbucketed)

    def _orig_order(g: dict) -> int:
        try:
            return groups.index(g)
        except ValueError:
            return 0

    out.sort(
        key=lambda g: (
            _ticket_group_sort_rank(str(g.get("group_name") or "")),
            _orig_order(g),
        )
    )
    return out, len(groups), len(out)


def _ticket_group_platforms_attr(group: dict) -> str:
    """Space-separated slugs for filter bar: pp, ud, dk."""
    slugs: set[str] = set()
    for t in group.get("tickets") or []:
        for leg in t.get("legs") or []:
            plat = str(leg.get("pick_platform") or "prizepicks").lower().strip()
            if plat == "underdog":
                slugs.add("ud")
            elif plat == "draftkings":
                slugs.add("dk")
            else:
                slugs.add("pp")
    return " ".join(sorted(slugs))


def render_tickets_body_html(
    payload: dict,
    *,
    _non_ev_slips_removed: int = 0,
    winrate_payload: dict | None = None,
) -> tuple[str, str]:
    """
    Render ticket slips from tickets_latest.json payload.
    Returns (body_html, page_title) for injection into tickets_built.html.
    """
    def safe_str(val, default: str = "") -> str:
        if val is None:
            return default
        s = str(val).strip()
        if s.lower() in ("nan", "none", "nat", "null"):
            return default
        return s

    date_declared_raw = (payload.get("date") or "").strip()
    date_declared = date_declared_raw[:10] if len(date_declared_raw) >= 10 else date_declared_raw
    generated_at = payload.get("generated_at") or ""
    groups_all = list(payload.get("groups") or [])
    _ui_cap_raw = os.getenv("PROPORACLE_TICKETS_UI_MAX_GROUPS_PER_BUCKET", "10").strip()
    try:
        _ui_cap = int(_ui_cap_raw) if _ui_cap_raw else 0
    except ValueError:
        _ui_cap = 10
    _ui_cap_note = ""
    if _ui_cap > 0:
        groups, _n_g_full, _n_g_show = _cap_ticket_groups_for_ui(groups_all, _ui_cap)
        if _n_g_show < _n_g_full:
            _ui_cap_note = (
                f' &nbsp;·&nbsp; <span style="opacity:.85;font-size:12px;">'
                f"Showing {_n_g_show} of {_n_g_full} groups</span>"
            )
    else:
        groups = groups_all
    n_slips = sum(len(g.get("tickets") or []) for g in groups)
    n_groups = len(groups)

    def _calendar_date_from_game_time(gs: str) -> str | None:
        """Calendar YYYY-MM-DD from mixed game_time strings."""
        s = (gs or "").strip()
        if not s:
            return None
        candidates = [s]
        if " " in s and "T" not in s.split(" ", 1)[0]:
            candidates.append(s.replace(" ", "T", 1))
        for cand in candidates:
            try:
                c2 = cand.replace("Z", "+00:00") if cand.endswith("Z") else cand
                dt = datetime.fromisoformat(c2)
                return dt.date().isoformat()
            except ValueError:
                continue
        mmdd = re.match(r"^\s*(\d{1,2})/(\d{1,2})\b", s)
        if mmdd and len(date_declared) >= 4 and date_declared[:4].isdigit():
            y = int(date_declared[:4])
            m = int(mmdd.group(1))
            d = int(mmdd.group(2))
            if 1 <= m <= 12 and 1 <= d <= 31:
                return f"{y:04d}-{m:02d}-{d:02d}"
        if len(s) >= 10 and s[4] == "-" and s[7] == "-":
            head = s[:10]
            if head[0:4].isdigit() and head[5:7].isdigit() and head[8:10].isdigit():
                return head
        return None

    def _modal_slate_date_from_legs(p: dict) -> str | None:
        counts: dict[str, int] = {}
        for g in p.get("groups") or []:
            for t in g.get("tickets") or []:
                for leg in t.get("legs") or []:
                    cd = _calendar_date_from_game_time(str(leg.get("game_time") or ""))
                    if cd:
                        counts[cd] = counts.get(cd, 0) + 1
        if not counts:
            return None
        return max(counts.items(), key=lambda kv: (kv[1], kv[0]))[0]

    date_from_legs = _modal_slate_date_from_legs({**payload, "groups": groups_all})
    # Header date should reflect the pipeline target date (file date),
    # not the surviving leg subset date after sport-specific fallbacks.
    date_str = date_declared or date_from_legs or "Today"
    date_note_html = ""
    if date_from_legs and date_declared and date_from_legs != date_declared:
        date_note_html = (
            f' <span style="opacity:.7;font-size:12px;">(file date {_h(date_declared)})</span>'
        )

    page_title = f"PropOracle Tickets — {date_str}"

    parts: list[str] = []
    parts.append('<div class="tickets-built shell">')
    parts.append(_TICKETS_BUILT_PAYOUT_CSS)

    # ── Hero ──────────────────────────────────────────────────────────────────
    built_html = (
        f'<span class="hero-meta-built">{_h(generated_at)}</span>' if generated_at else ""
    )
    if _non_ev_slips_removed > 0:
        counts_line = (
            f"{n_groups} groups &nbsp;·&nbsp; {n_slips} +EV slips "
            f"&nbsp;·&nbsp; <span style=\"color:var(--muted);\">{_non_ev_slips_removed} non-EV filtered</span>"
        )
    else:
        counts_line = f"{n_groups} groups &nbsp;·&nbsp; {n_slips} slips"
    counts_line += _ui_cap_note
    parts.append(f'''
<div class="hero tickets-hero" style="margin-bottom:24px;">
  <div class="hero-copy">
    <div class="hero-eyebrow" style="font-size:11px;letter-spacing:2px;color:var(--muted);text-transform:uppercase;margin-bottom:8px;">Today&rsquo;s Picks</div>
    <h1 class="hero-title" style="font-family:'Bebas Neue',sans-serif;font-size:clamp(32px,5vw,56px);letter-spacing:0.06em;line-height:1.05;color:var(--text);margin:0;">
      PROP<span class="hero-oracle-em">ORACLE</span>&nbsp;TICKETS
    </h1>
  </div>
  <div class="hero-meta-row" role="group" aria-label="Slate summary">
    <span class="hero-meta-date">{_h(date_str)}{date_note_html}</span>
    <span class="hero-meta-counts">{counts_line}</span>
    {built_html}
  </div>
</div>''')

    if not groups:
        parts.append('<div class="filter-pill">No tickets generated for this date.</div>')
        parts.append('</div>')
        return "".join(parts), page_title

    parts.append(_winrate_best_panel_html(winrate_payload))

    # ── Groups ────────────────────────────────────────────────────────────────
    leg_graph_uid = 0
    table_cols = 13

    prepared: list[dict] = []
    for original_index, group in enumerate(groups):
        tickets = group.get("tickets") or []
        if not tickets:
            continue
        gn = group.get("group_name") or "Tickets"
        ds, dt, dpk = _ticket_group_filter_slugs(gn)
        ev_a = _group_ev_data_attr(tickets)
        pc_max = _group_payout_confidence_score(tickets)
        prepared.append(
            {
                "group": group,
                "sport": ds,
                "type": dt,
                "pick": dpk,
                "ev": ev_a,
                "ev_score": _group_max_ev_for_ui_cap(group),
                "hit_score": _group_hit_rate_score(tickets),
                "p_win_score": _group_max_p_win(group),
                "original_index": original_index,
                "payout_confidence": pc_max,
            }
        )

    prepared.sort(
        key=lambda ent: (
            _ticket_group_sort_rank(str(ent["group"].get("group_name") or "")),
            -float(ent.get("ev_score") or 0.0),
            -float(ent.get("payout_confidence") or 0.0),
            -float(ent.get("hit_score") or 0.0),
            int(ent.get("original_index", 0)),
        )
    )
    # Build filter pills from the full payload (not just UI-capped groups) so
    # sports like NBA1H/WNBA remain selectable even when not in today's top-N.
    prepared_all: list[dict] = []
    for original_index, group in enumerate(groups_all):
        tickets = group.get("tickets") or []
        if not tickets:
            continue
        gn = group.get("group_name") or "Tickets"
        ds, dt, dpk = _ticket_group_filter_slugs(gn)
        ev_a = _group_ev_data_attr(tickets)
        prepared_all.append(
            {
                "sport": ds,
                "type": dt,
                "pick": dpk,
                "ev": ev_a,
                "ev_score": _group_max_ev_for_ui_cap(group),
                "original_index": original_index,
            }
        )
    filter_attr_rows = [
        {"sport": x["sport"], "type": x["type"], "pick": x["pick"], "ev": x["ev"]}
        for x in (prepared_all or prepared)
    ]
    parts.append(_tickets_filter_pills_html(filter_attr_rows))

    top_groups = [
        e for e in prepared
        if str(e.get("ev") or "").lower() not in {"skip", "low"}
    ]
    top_groups.sort(
        key=lambda e: (
            -float(e.get("ev_score") or 0.0),
            -float(e.get("payout_confidence") or 0.0),
        )
    )
    if top_groups:
        top_rows: list[str] = []
        for e in top_groups[:5]:
            g = e["group"]
            gn = str(g.get("group_name") or "Tickets")
            ev_score = float(e.get("ev_score") or 0.0)
            rec = str(e.get("ev") or "").upper() or "OK"
            rec_cls = "ev-strong" if rec == "STRONG" else ("ev-ok" if rec == "OK" else "ev-marginal")
            top_rows.append(
                f'<div class="best-ticket-row"><span class="best-ticket-name">{_h(gn)}</span>'
                f'<span class="best-ticket-meta {_h(rec_cls)}">{_h(rec)} · EV {_fmt(ev_score, 2)}×</span></div>'
            )
        parts.append(
            '<div class="filter-pill" style="margin-top:8px;">'
            '<div style="font-size:10px;letter-spacing:2px;color:var(--muted);text-transform:uppercase;margin-bottom:10px;">Today&apos;s Best</div>'
            + "".join(top_rows)
            + "</div>"
        )

    for ent in prepared:
        group = ent["group"]
        group_name = group.get("group_name") or "Tickets"
        n_legs = group.get("n_legs") or 0
        power_pay = group.get("power_payout")
        flex_pay = group.get("flex_payout")
        tickets = group.get("tickets") or []

        sport_key = _group_sport(group_name)
        accent = _sport_accent(sport_key)

        pay_label = ""
        if power_pay and flex_pay and abs(float(power_pay) - float(flex_pay)) > 0.01:
            pay_label = f"Power {_fmt(power_pay, 1)}× &nbsp;·&nbsp; Flex {_fmt(flex_pay, 1)}×"
        elif power_pay:
            pay_label = f"{_fmt(power_pay, 1)}×"

        group_meta_html = f'{n_legs}-leg{(" &nbsp;·&nbsp; " + pay_label) if pay_label else ""}'
        ev_badge_html = _group_ev_badge_summary_html(tickets)
        d_sport = ent["sport"]
        d_type = ent["type"]
        d_pick = ent["pick"]
        d_ev = ent["ev"]
        d_ev_score = float(ent.get("ev_score") or 0.0)
        d_hit_score = float(ent.get("hit_score") or 0.0)
        d_p_win_score = float(ent.get("p_win_score") or 0.0)
        d_pc = float(ent.get("payout_confidence") or 0.0)
        d_oi = int(ent.get("original_index", 0))
        rec_cls = d_ev if d_ev in ("strong", "ok", "marginal", "low", "skip") else "skip"
        d_plat = _ticket_group_platforms_attr(group)
        d_n_legs = int(n_legs) if n_legs else _ticket_group_leg_count(group_name)

        parts.append(f'''
<div class="ticket-group-section collapsed group-rec-{_h(rec_cls)}" data-sport="{_h(d_sport)}" data-type="{_h(d_type)}" data-pick="{_h(d_pick)}" data-ev="{_h(d_ev)}" data-ev-score="{_fmt(d_ev_score, 4)}" data-p-win="{_fmt(d_p_win_score, 6)}" data-hit-score="{_fmt(d_hit_score, 4)}" data-payout-confidence="{_fmt(d_pc, 2)}" data-n-legs="{d_n_legs}" data-original-index="{d_oi}" data-platforms="{_h(d_plat)}">
  <div class="ticket-group-header collapsible-header" role="button" tabindex="0" aria-expanded="false">
    <span class="group-title" style="color:{accent};">{_h(group_name)}</span>
    <span class="group-meta">{group_meta_html}</span>
    {ev_badge_html}
    <span class="collapse-icon" aria-hidden="true">▼</span>
  </div>
  <div class="ticket-group-body">
''')

        for ticket in tickets:
            ticket_no = ticket.get("ticket_no") or ""
            win_prob = ticket.get("est_win_prob")
            try:
                p_win_val = float(ticket.get("p_win")) if ticket.get("p_win") is not None else None
            except (TypeError, ValueError):
                p_win_val = None
            if p_win_val is None:
                try:
                    p_win_val = float(win_prob) if win_prob is not None else None
                except (TypeError, ValueError):
                    p_win_val = None
            avg_hr = ticket.get("avg_hit_rate")
            ev = ticket.get("ev_power")
            t_power_pay = ticket.get("power_payout") or ticket.get("base_power_payout")
            has_warn = ticket.get("has_data_warning", False)
            legs = ticket.get("legs") or []

            ev_f = None
            if ev is not None:
                try:
                    ev_f = float(ev)
                except (TypeError, ValueError):
                    ev_f = None

            payout = ticket.get("payout")
            hdr_brackets = ""
            payout_ok = False
            ev_emp_f = None
            if isinstance(payout, dict) and payout.get("ev") is not None:
                try:
                    ev_emp_f = float(payout["ev"])
                    payout_ok = bool(math.isfinite(ev_emp_f))
                except (TypeError, ValueError):
                    ev_emp_f = None
                    payout_ok = False
            ev_for_badge = ev_emp_f if payout_ok else ev_f
            if ev_for_badge is not None and math.isfinite(ev_for_badge):
                if ev_for_badge >= 1.50:
                    sig_cls, sig_lbl = "sig-strong", "STRONG"
                elif ev_for_badge >= 1.15:
                    sig_cls, sig_lbl = "sig-lean", "OK"
                elif ev_for_badge >= 0.80:
                    sig_cls, sig_lbl = "sig-risk", "MARGINAL"
                else:
                    sig_cls, sig_lbl = "sig-risk", "LOW"
            else:
                sig_cls, sig_lbl = "sig-lean", "—"
            display_ev = ev_emp_f if payout_ok else ev_f
            if display_ev is None:
                display_ev = 0.0
            if payout_ok:
                rec_s = str(payout.get("recommendation") or "")
                ev_cls = _payout_ev_class(rec_s)
                pre = _payout_rec_prefix(rec_s)
                pay_x = payout.get("min_payout_x")
                sweep_x = payout.get("sweep_payout_x")
                psrc = str(payout.get("payout_source") or "calibrated")
                if pay_x is None:
                    pay_x = payout.get("min_guarantee")
                if sweep_x is None:
                    sweep_x = payout.get("sweep_payout")
                if sweep_x is None:
                    sweep_x = _slip_display_payout_multiplier(payout, ticket, group)
                if pay_x is None:
                    pay_x = sweep_x
                pay_tt = str(payout.get("ticket_type") or "").lower()
                payout_badge_label = f"Min {_fmt(pay_x, 2)}x · Sweep {_fmt(sweep_x, 2)}x"
                payout_badge_title = (
                    f' title="Min guarantee {_fmt(pay_x, 2)}x · Sweep {_fmt(sweep_x, 2)}x"'
                    if pay_tt == "power"
                    else ""
                )
                hdr_brackets = f'''
        <span class="ticket-hdr-bracket">[{_h(group_name)}]</span>
        <span class="payout-rec-badge {ev_cls}">[{_h(pre)} {_h(rec_s)} — EV {_fmt(ev_emp_f, 2)}]</span>
        <span class="payout-x-badge"{payout_badge_title}>[{_h(payout_badge_label)}]</span>
        {_payout_source_badge_html(psrc)}
        <span class="{sig_cls}" title="Empirical EV tier (fallback to modeled EV when payout block is missing)">{sig_lbl}</span>'''
            if not hdr_brackets:
                hdr_brackets = (
                    f'<span class="ticket-hdr-bracket">[{_h(group_name)}]</span>'
                    f'<span class="{sig_cls}">{sig_lbl}</span>'
                )

            kpi_payout = None
            kpi_sweep = None
            kpi_source = "calibrated"
            if payout_ok and isinstance(payout, dict):
                kpi_source = str(payout.get("payout_source") or "calibrated")
                kpi_payout = payout.get("min_payout_x")
                kpi_sweep = payout.get("sweep_payout_x")
                if kpi_payout is None:
                    kpi_payout = payout.get("min_guarantee")
                if kpi_sweep is None:
                    kpi_sweep = payout.get("sweep_payout")
            if kpi_payout is None:
                kpi_payout = t_power_pay
            if kpi_sweep is None:
                kpi_sweep = _slip_display_payout_multiplier(payout, ticket, group)

            warn_html = ('<span style="font-size:10px;color:var(--amber);margin-left:auto;">⚠ data warning</span>'
                         if has_warn else "")

            parts.append(f'''
<div class="ticket" style="border-left:4px solid {accent};">
  <div class="ticket-body">
      <div class="ticket-hdr">
        <span class="ticket-no">#{_h(ticket_no)}</span>
        {hdr_brackets}
        {warn_html}
      </div>
      <div class="kpi-row">
        <div class="kpi">
          <div class="kpi-label">Avg Leg HR</div>
          <div class="kpi-val" style="color:var(--green);">{_pct(avg_hr)}</div>
        </div>
        <div class="kpi">
          <div class="kpi-label">Model Win Prob</div>
          <div class="kpi-val" style="color:var(--cyan);">{_pct(win_prob)}</div>
        </div>
        <div class="kpi">
          <div class="kpi-label">EV</div>
          <div class="kpi-val" style="color:var(--accent);" title="{_h(str((payout or {}).get('ev_formula') or 'EV = P(all)*sweep + P(miss-1)*min - 1.0'))}">{_fmt(display_ev, 2)}×</div>
        </div>
        <div class="kpi">
          <div class="kpi-label">MIN PAYOUT</div>
          <div class="kpi-val">{_fmt(kpi_payout, 2)}×</div>
          <div style="font-size:11px;color:var(--muted);margin-top:2px;">Sweep {_fmt(kpi_sweep, 2)}x {_payout_source_badge_html(kpi_source)}</div>
        </div>
      </div>
      <div class="ticket-legs-table-wrapper">
      <table class="ticket-legs-table">
        <thead>
          <tr>
            <th>Player</th>
            <th>Sport</th>
            <th>Prop</th>
            <th>Line</th>
            <th>Dir</th>
            <th>Pick</th>
            <th>HR</th>
            <th>ML</th>
            <th>Edge</th>
            <th>Vs Def</th>
            <th>Best Book</th>
            <th>Best Line</th>
            <th>Edge vs PP</th>
          </tr>
        </thead>
        <tbody>''')

            for leg in legs:
                player = leg.get("player") or ""
                sport = leg.get("sport") or ""
                prop_type = leg.get("prop_type") or ""
                line = leg.get("line")
                std_line = leg.get("standard_line")
                direction = (leg.get("direction") or "").upper()
                if direction == "LOWER":
                    direction = "UNDER"
                pick_type = (leg.get("pick_type") or "").strip()
                hit_rate = leg.get("hit_rate")
                ml_prob = leg.get("ml_prob")
                edge = leg.get("edge")
                def_tier = safe_str(leg.get("def_tier"), "")
                best_book = str(leg.get("best_cross_book") or "").strip()
                best_line = leg.get("best_cross_line")
                cross_edge_vs_pp = leg.get("cross_edge_vs_pp")
                line_underdog = leg.get("line_underdog")
                line_draftkings = leg.get("line_draftkings")
                team = leg.get("team") or ""
                opp = leg.get("opp") or ""
                initials = leg.get("initials") or player[:2].upper()

                # Direction badge
                dir_cls = "dir-over" if direction == "OVER" else "dir-under"
                dir_axis_cls = "direction-over" if direction == "OVER" else "direction-under"
                dir_html = f'<span class="{dir_cls}">{_h(direction)}</span>'

                # Pick type badge
                pk_lower = pick_type.lower()
                pk_color = _PICK_COLOR.get(pk_lower, "#aaa")
                pick_html = f'<span style="font-size:13px;font-weight:700;color:{pk_color};">{_h(pick_type)}</span>'

                # Line display (show goblin discount if applicable)
                if std_line and line and abs(float(std_line) - float(line)) >= 0.1:
                    line_html = f'{_fmt(line, 1)} <span style="font-size:11px;color:var(--muted);text-decoration:line-through;">{_fmt(std_line, 1)}</span>'
                else:
                    line_html = _fmt(line, 1)

                # Cross-book comparison summary (PP vs UD vs DK)
                books_avail = []
                if line is not None:
                    books_avail.append(f'PP {_fmt(line, 1)}')
                if line_underdog is not None:
                    books_avail.append(f'UD {_fmt(line_underdog, 1)}')
                if line_draftkings is not None:
                    books_avail.append(f'DK {_fmt(line_draftkings, 1)}')
                line_tip = " / ".join(books_avail) if books_avail else "No cross-book lines"
                best_book_html = _h(best_book) if best_book else "—"
                best_line_html = _fmt(best_line, 1) if best_line is not None else "—"
                cross_edge_html = _fmt(cross_edge_vs_pp, 2) if cross_edge_vs_pp is not None else "—"
                cross_edge_style = "color:var(--muted);"
                try:
                    if cross_edge_vs_pp is not None and float(cross_edge_vs_pp) > 0.01:
                        cross_edge_style = "color:var(--green);font-weight:700;"
                except (TypeError, ValueError):
                    pass

                plat_raw = str(leg.get("pick_platform") or "prizepicks").lower().strip()
                if plat_raw == "underdog":
                    leg_plat_slug = "ud"
                elif plat_raw == "draftkings":
                    leg_plat_slug = "dk"
                else:
                    leg_plat_slug = "pp"

                # Sport accent chip
                s_accent = _sport_accent(sport)
                sport_html = f'<span style="font-size:12px;font-weight:700;color:{s_accent};background:{s_accent}22;padding:3px 8px;border-radius:4px;border:1px solid {s_accent}44;">{_h(sport)}</span>'

                # Avatar
                av_html = f'<div class="avatar">{_h(initials)}</div>'

                # Matchup sub-label
                matchup = f"{team} vs {opp}" if team and opp else (team or opp)

                hr_disp = (
                    f"Hit rate {_pct(hit_rate)} · ML {_pct(ml_prob)} · Edge {_fmt(edge, 2)}"
                    + (f" · Def {def_tier}" if def_tier else "")
                )

                parts.append(f'''
          <tr class="leg-row" data-hr-display="{_h(hr_disp)}" data-platform="{_h(leg_plat_slug)}">
            <td class="leg-col leg-col-player">
              <div class="pwrap">
                {av_html}
                <div>
                  <div style="font-weight:600;font-size:14px;">{_h(player)}</div>
                  <div style="font-size:12px;color:var(--muted);">{_h(matchup)}</div>
                </div>
              </div>
            </td>
            <td class="leg-col leg-col-sport hide-mobile">{sport_html}</td>
            <td class="leg-col leg-col-prop" style="color:var(--text);font-weight:500;">{_h(prop_type)}</td>
            <td class="leg-col leg-col-line" style="font-family:'Inter',sans-serif;">{line_html}</td>
            <td class="leg-col leg-col-dir direction-cell {dir_axis_cls}">{dir_html}</td>
            <td class="leg-col leg-col-pick">{pick_html}</td>
            <td class="leg-col leg-col-hr hide-mobile" style="font-family:'Inter',sans-serif;color:var(--green);">{_pct(hit_rate)}</td>
            <td class="leg-col leg-col-ml hide-mobile" style="font-family:'Inter',sans-serif;color:var(--cyan);">{_pct(ml_prob)}</td>
            <td class="leg-col leg-col-edge hide-mobile" style="font-family:'Inter',sans-serif;color:var(--accent);">{_fmt(edge, 2)}</td>
            <td class="leg-col leg-col-def hide-mobile" style="font-size:13px;color:var(--muted);">{_h(def_tier)}</td>
            <td class="leg-col leg-col-book hide-mobile" style="font-family:'Inter',sans-serif;color:var(--cyan);" title="{_h(line_tip)}">{best_book_html}</td>
            <td class="leg-col leg-col-bl hide-mobile" style="font-family:'Inter',sans-serif;" title="{_h(line_tip)}">{best_line_html}</td>
            <td class="leg-col leg-col-ce hide-mobile" style="font-family:'Inter',sans-serif;{cross_edge_style}" title="Positive means better line than PP for this direction">{cross_edge_html}</td>
          </tr>''')
                leg_graph_uid += 1
                parts.append(_tickets_leg_graph_row_html(leg, f"lgr-{leg_graph_uid}", table_cols))

            payout_section = ""
            if payout_ok and isinstance(payout, dict):
                try:
                    p_all = float(payout["p_all_win"])
                except (TypeError, ValueError):
                    p_all = 0.0
                rec_s2 = str(payout.get("recommendation") or "")
                ev_cls_row = _payout_ev_class(rec_s2)
                try:
                    ev_disp = float(payout["ev"])
                except (TypeError, ValueError):
                    ev_disp = 0.0
                pay_mult = payout.get("min_payout_x")
                if pay_mult is None:
                    pay_mult = payout.get("min_guarantee")
                if pay_mult is None:
                    pay_mult = payout.get("payout")
                sweep_mult = payout.get("sweep_payout_x")
                if sweep_mult is None:
                    sweep_mult = payout.get("sweep_payout")
                if sweep_mult is None:
                    sweep_mult = _slip_display_payout_multiplier(payout, ticket, group)
                psrc2 = str(payout.get("payout_source") or "calibrated")
                e10g = payout.get("entry_10_to_win_guarantee")
                if e10g is None and pay_mult is not None:
                    try:
                        e10g = round(10 * float(pay_mult), 2)
                    except (TypeError, ValueError):
                        e10g = None
                e10s = payout.get("entry_10_to_win_sweep")
                if e10s is None and sweep_mult is not None:
                    try:
                        e10s = round(10 * float(sweep_mult), 2)
                    except (TypeError, ValueError):
                        e10s = None
                pre_ev = _payout_rec_prefix(rec_s2)
                tt_pay = str(payout.get("ticket_type") or "").lower()
                if tt_pay == "power":
                    try:
                        power_min_mult = float(payout.get("min_payout_x", 1.0))
                    except (TypeError, ValueError):
                        power_min_mult = 1.0
                    e10g = round(10 * power_min_mult, 2)
                    payout_section = f'''
      <div class="ticket-payout">
        <div class="payout-row">
          <span class="payout-label" title="Sweep payout (all correct): {_fmt(sweep_mult, 2)}x">Payout</span>
          <span class="payout-value" title="Sweep payout (all correct): {_fmt(sweep_mult, 2)}x">{_fmt(power_min_mult, 2)}x</span>
          {_payout_source_badge_html(psrc2)}
        </div>
        <div class="payout-row">
          <span class="payout-label">P(Win)</span>
          <span class="payout-value">{_fmt(p_all * 100, 1)}%</span>
        </div>
        <div class="payout-row">
          <span class="payout-label">EV</span>
          <span class="payout-value {ev_cls_row}">{_fmt(ev_disp, 2)} &mdash; {_h(pre_ev)} {_h(rec_s2)}</span>
        </div>
        <div class="payout-entry-guide">
          <span title="Sweep: $10 &rarr; ${_fmt(e10s, 2)}">$10 &rarr; ${_fmt(e10g, 2)} (min guarantee)</span>
        </div>
      </div>'''
                else:
                    payout_section = f'''
      <div class="ticket-payout">
        <div class="payout-row">
          <span class="payout-label">Sweep (all correct)</span>
          <span class="payout-value">{_fmt(sweep_mult, 2)}x</span>
        </div>
        <div class="payout-row">
          <span class="payout-label">Partial ({int(n_legs) - 1} correct)</span>
          <span class="payout-value">{_fmt(pay_mult, 2)}x</span>
          {_payout_source_badge_html(psrc2)}
        </div>
        <div class="payout-row">
          <span class="payout-label">P(Win)</span>
          <span class="payout-value">{_fmt(p_all * 100, 1)}%</span>
        </div>
        <div class="payout-row">
          <span class="payout-label">EV</span>
          <span class="payout-value {ev_cls_row}">{_fmt(ev_disp, 2)} &mdash; {_h(pre_ev)} {_h(rec_s2)}</span>
        </div>
        <div class="payout-entry-guide">
          $10 &rarr; ${_fmt(e10s, 2)} (sweep) / ${_fmt(e10g, 2)} (n&minus;1)
        </div>
      </div>'''

            parts.append(f'''
        </tbody>
      </table>
      </div>
{payout_section}
  </div>
</div>''')

        parts.append("</div></div>")  # ticket-group-body, ticket-group-section

    parts.append('</div>')  # end .tickets-built.shell

    # Inline JS: leg graphs, filter pills, collapsible groups
    parts.append('''
<script>
(function(){
  document.querySelectorAll('.tickets-built .leg-row').forEach(function(row){
    row.addEventListener('click', function(){
      var next = row.nextElementSibling;
      if(next && next.classList.contains('leg-graph-row')){
        next.classList.toggle('open');
      }
    });
  });

  var activeFilter = 'all';
  var sortMode = 'ev_desc';
  var hideSkip = true;

  function parseNum(el, attr){
    var raw = (el.getAttribute(attr) || '').trim();
    var n = parseFloat(raw);
    return Number.isFinite(n) ? n : 0;
  }

  function sortGroups(groups){
    if(sortMode === 'group'){
      groups.sort(function(a,b){
        return parseNum(a, 'data-original-index') - parseNum(b, 'data-original-index');
      });
      return;
    }
    if(sortMode === 'ev_asc'){
      groups.sort(function(a,b){ return parseNum(a, 'data-ev-score') - parseNum(b, 'data-ev-score'); });
      return;
    }
    if(sortMode === 'hit_rate'){
      groups.sort(function(a,b){ return parseNum(b, 'data-hit-score') - parseNum(a, 'data-hit-score'); });
      return;
    }
    if(sortMode === 'pwin_desc'){
      groups.sort(function(a,b){ return parseNum(b, 'data-p-win') - parseNum(a, 'data-p-win'); });
      return;
    }
    if(sortMode === 'pwin_asc'){
      groups.sort(function(a,b){ return parseNum(a, 'data-p-win') - parseNum(b, 'data-p-win'); });
      return;
    }
    if(sortMode === 'legs_desc'){
      groups.sort(function(a,b){
        var dl = parseNum(b, 'data-n-legs') - parseNum(a, 'data-n-legs');
        if(dl !== 0) return dl;
        return parseNum(b, 'data-ev-score') - parseNum(a, 'data-ev-score');
      });
      return;
    }
    groups.sort(function(a,b){ return parseNum(b, 'data-ev-score') - parseNum(a, 'data-ev-score'); });
  }

  function matchesFilter(group, filter){
    if(filter === 'all') return true;
    if(filter === 'top-payout') return true;
    if(filter === 'pp' || filter === 'ud' || filter === 'dk'){
      var raw = (group.getAttribute('data-platforms') || '').toLowerCase().trim();
      if(!raw) return filter === 'pp';
      var parts = raw.split(/\\s+/).filter(Boolean);
      return parts.indexOf(filter) >= 0;
    }
    var ds = (group.getAttribute('data-sport') || '').toLowerCase();
    var dt = (group.getAttribute('data-type') || '').toLowerCase();
    var dp = (group.getAttribute('data-pick') || '').toLowerCase();
    var de = (group.getAttribute('data-ev') || '').toLowerCase();
    return ds === filter || dt === filter || dp === filter || de === filter;
  }

  function applyGroupView(){
    var shell = document.querySelector('.tickets-built.shell');
    if(!shell) return;
    var bar = shell.querySelector('.ticket-filter-bar');
    var allGroups = Array.from(shell.querySelectorAll('.ticket-group-section'));
    var visible = allGroups.filter(function(g){
      if(!matchesFilter(g, activeFilter)) return false;
      if(hideSkip){
        var rec = (g.getAttribute('data-ev') || '').toLowerCase();
        if(rec === 'skip' || rec === 'low') return false;
      }
      return true;
    });

    if(activeFilter === 'top-payout'){
      visible.sort(function(a,b){
        return parseNum(b, 'data-payout-confidence') - parseNum(a, 'data-payout-confidence');
      });
      visible = visible.slice(0, 3);
    } else {
      sortGroups(visible);
    }

    allGroups.forEach(function(g){ g.style.display = 'none'; });
    var frag = document.createDocumentFragment();
    visible.forEach(function(g){ g.style.display = ''; frag.appendChild(g); });
    if(bar){
      var insertBefore = bar.nextElementSibling;
      if(insertBefore && insertBefore.classList && insertBefore.classList.contains('utp-root')){
        insertBefore = insertBefore.nextElementSibling;
      }
      if(insertBefore){
        shell.insertBefore(frag, insertBefore);
      } else {
        shell.appendChild(frag);
      }
    } else {
      shell.appendChild(frag);
    }
  }

  document.querySelectorAll('.ticket-filter-pill').forEach(function(pill){
    pill.addEventListener('click', function(){
      document.querySelectorAll('.ticket-filter-pill').forEach(function(p){ p.classList.remove('active'); });
      pill.classList.add('active');
      activeFilter = (pill.getAttribute('data-filter') || '').toLowerCase();
      applyGroupView();
    });
  });

  var sortSel = document.getElementById('ticket-sort-select');
  if(sortSel){
    sortSel.addEventListener('change', function(){
      sortMode = (sortSel.value || 'ev_desc').toLowerCase();
      applyGroupView();
    });
  }

  var tSkip = document.getElementById('toggle-skip');
  if(tSkip){
    tSkip.addEventListener('click', function(){
      hideSkip = !hideSkip;
      tSkip.classList.toggle('active', hideSkip);
      tSkip.setAttribute('aria-pressed', hideSkip ? 'true' : 'false');
      tSkip.textContent = hideSkip ? 'HIDE SKIP' : 'SHOW SKIP';
      applyGroupView();
    });
  }

  function toggleSectionCollapsed(section){
    if(!section) return;
    section.classList.toggle('collapsed');
    var hdr = section.querySelector('.collapsible-header');
    if(hdr) hdr.setAttribute('aria-expanded', section.classList.contains('collapsed') ? 'false' : 'true');
  }

  document.querySelectorAll('.tickets-built .collapsible-header').forEach(function(header){
    header.addEventListener('click', function(ev){
      ev.preventDefault();
      toggleSectionCollapsed(header.closest('.ticket-group-section'));
    });
    header.addEventListener('keydown', function(ev){
      if(ev.key === 'Enter' || ev.key === ' '){
        ev.preventDefault();
        toggleSectionCollapsed(header.closest('.ticket-group-section'));
      }
    });
  });

  var ex = document.getElementById('expand-all');
  if(ex) ex.addEventListener('click', function(ev){
    ev.preventDefault();
    document.querySelectorAll('.ticket-group-section').forEach(function(s){
      s.classList.remove('collapsed');
      var h = s.querySelector('.collapsible-header');
      if(h) h.setAttribute('aria-expanded', 'true');
    });
  });
  var col = document.getElementById('collapse-all');
  if(col) col.addEventListener('click', function(ev){
    ev.preventDefault();
    document.querySelectorAll('.ticket-group-section').forEach(function(s){
      s.classList.add('collapsed');
      var h = s.querySelector('.collapsible-header');
      if(h) h.setAttribute('aria-expanded', 'false');
    });
  });

  (function(){
    // Always start collapsed so /tickets opens compact on both mobile and desktop.
    function collapseAllGroups(){
      document.querySelectorAll('.ticket-group-section').forEach(function(s){
        s.classList.add('collapsed');
        var h = s.querySelector('.collapsible-header');
        if(h) h.setAttribute('aria-expanded', 'false');
      });
    }
    collapseAllGroups();
    applyGroupView();
  })();
})();
</script>''')

    return "".join(parts), page_title


if __name__ == "__main__":
    main()
