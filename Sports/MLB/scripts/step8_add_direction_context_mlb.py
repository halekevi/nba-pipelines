#!/usr/bin/env python3
"""
step8_add_direction_context_mlb.py  (MLB Pipeline)

Mirrors NBA step8. Reads step7_mlb_ranked.xlsx, adds direction,
outputs CSV + clean formatted XLSX with pitcher/hitter split tabs.

Run:
  py -3.14 step8_add_direction_context_mlb.py \
    --input step7_mlb_ranked.xlsx \
    --output step8_mlb_direction.csv

Writes step8_mlb_direction.csv (full columns, incl. final_dir_reason / direction_override)
and step8_mlb_direction_clean.xlsx (tickets/UI — Direction only). Fade audit: use the CSV.
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook

_REPO = Path(__file__).resolve().parent
for _ in range(10):
    if (_REPO / "utils" / "step8_edge_direction.py").is_file():
        if str(_REPO) not in sys.path:
            sys.path.insert(0, str(_REPO))
        break
    _REPO = _REPO.parent
else:
    raise RuntimeError("Could not locate repo root with utils/step8_edge_direction.py")

from utils.step8_edge_direction import reconcile_signed_edge_abs_dataframe
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date


def _copy_dated_step8_mlb(output_xlsx_path: str, slate_date: str) -> None:
    """Publish dated clean XLSX to repo outputs/<slate>/ and Sports/MLB/outputs/<slate>/ (matches NBA + WNBA pattern)."""
    src = Path(output_xlsx_path)
    if not src.is_file():
        return
    d = (slate_date or "").strip()
    if not d:
        d = date.today().isoformat()
    repo_root = Path(__file__).resolve().parents[3]
    dated_name = f"step8_mlb_direction_clean_{d}.xlsx"
    for dated_dir in (repo_root / "outputs" / d, repo_root / "Sports" / "MLB" / "outputs" / d):
        try:
            dated_dir.mkdir(parents=True, exist_ok=True)
            dated_path = dated_dir / dated_name
            shutil.copy2(src, dated_path)
            print(f"[MLB step8] Dated copy -> {dated_path}")
        except Exception as e:
            print(f"[MLB step8] WARN: dated copy failed ({dated_dir}): {e}")


def _norm_pick_type(x: str) -> str:
    t = str(x or "").strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
PITCHER_TAB_COLOR = "8B0000"
HITTER_TAB_COLOR  = "003366"
HEADER_COLOR      = "1C1C1C"
DIR_OVER          = "C8F7C5"
DIR_UNDER         = "F7C5C5"

PITCHER_PROPS = {
    "strikeouts", "pitching_outs", "innings_pitched",
    "hits_allowed", "earned_runs", "walks_allowed", "batters_faced",
    "pitches_thrown",
}


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def _line_shift_fill(val) -> str:
    if val is None or val == "" or (isinstance(val, float) and pd.isna(val)):
        return "FFFFFF"
    try:
        num = float(val)
        if num > 0:
            return "C8F7C5"
        if num < 0:
            return "F7C5C5"
        return "FFFFFF"
    except (TypeError, ValueError):
        pass
    s = str(val).strip().upper()
    if "UP" in s or s == "MOVED_UP":
        return "C8F7C5"
    if "DOWN" in s or s == "MOVED_DOWN":
        return "F7C5C5"
    return "FFFFFF"


def _row_game_datetimes(df: pd.DataFrame) -> pd.Series:
    """Datetime per row for MLB slate date filtering (prefer game_date, else start_time)."""
    idx = df.index
    gd = pd.Series(pd.NaT, index=idx)
    if "game_date" in df.columns:
        gd = pd.to_datetime(df["game_date"], errors="coerce")
    if "start_time" in df.columns:
        st = pd.to_datetime(df["start_time"], errors="coerce")
        gd = gd.where(gd.notna(), st)
    return gd


def write_sheet(wb, name: str, data: pd.DataFrame, tab_color: str = HEADER_COLOR) -> None:
    ws = wb.create_sheet(name)
    ws.sheet_properties.tabColor = tab_color
    headers = list(data.columns)

    # Keep workbook valid even when slate is empty.
    if not headers:
        ws.cell(row=1, column=1, value="No columns")
        return

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            col_name    = headers[ci - 1]
            display_val = None if pd.isna(val) else val
            cell        = ws.cell(row=ri, column=ci, value=display_val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

            if col_name == "Tier":
                tier_bg, tier_fg = TIER_COLORS.get(str(display_val), ("333333", "FFFFFF"))
                cell.fill = PatternFill("solid", start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name="Arial", size=9)
            elif col_name == "Direction":
                bg = DIR_OVER if display_val == "OVER" else DIR_UNDER if display_val == "UNDER" else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, name="Arial", size=9)
            elif col_name == "Player Type":
                bg = "FFE8E8" if str(display_val).lower() == "pitcher" else "E8F4FF"
                cell.fill = PatternFill("solid", start_color=bg)
            elif col_name in ("Line Shift", "line_direction_shift"):
                cell.fill = PatternFill("solid", start_color=_line_shift_fill(display_val))
            else:
                cell.fill = PatternFill("solid", start_color="F9F9F9" if ri % 2 == 0 else "FFFFFF")

    col_widths = {
        "Tier": 6, "Rank Score": 10, "Player": 22, "Pos": 6, "Player Type": 10,
        "Team": 10, "Opp": 10, "Days Rest": 9, "B2B": 6, "Opp Rest": 9, "Opp B2B": 8, "Game Total": 10, "Spread": 8, "Game Time": 10, "Game Date": 11,
        "Prop": 20, "Pick Type": 10, "Line": 7,
        "Direction": 9, "Edge": 7, "Projection": 10,
        "ML Prob": 9, "Edge Score": 10, "Blended Score": 12,
        "Hit Rate (5g)": 12, "Hit Rate Status": 14, "Reliability Note": 24,
        "Last 5 Avg": 10, "Season Avg": 10,
        "L5 Over": 8, "L5 Under": 8,
        "Def Rank": 9, "Def Tier": 10,
        "Min Tier": 9, "Bat Order": 10, "Pitcher Role": 12,
        "Series HR": 9,
        "Void Reason": 20,
        "Open Line": 8,
        "Line Movement": 12,
        "Line Shift": 10,
        "open_line": 10,
        "line_movement": 12,
        "line_direction_shift": 16,
        "implied_prob": 12,
        "implied_prob_over": 14,
        "implied_prob_under": 15,
        "distribution_std": 10,
        "distribution_n": 8,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


_MIN_TIER_NUM_MAP = {0: "Low", 1: "Med", 2: "High", 3: "Elite"}

MISSING_SENTINELS = {"—", "-", "", "nan", "none", "null"}


def _parse_g_vals(row, prefix: str = "stat_g", n: int = 10) -> list[float]:
    vals: list[float] = []
    for i in range(1, n + 1):
        col = f"{prefix}{i}"
        if isinstance(row, pd.Series):
            raw_val = row.get(col, "")
        else:
            raw_val = row.get(col, "")
        raw = str(raw_val if raw_val is not None else "").strip()
        if raw.lower() in MISSING_SENTINELS:
            continue
        try:
            vals.append(float(raw))
        except ValueError:
            continue
    return vals


def _attach_distribution_std(df: pd.DataFrame, *, g_prefix: str = "stat_g") -> pd.DataFrame:
    """Sample std (ddof=1) of stat_g1..10 for pipeline_read distribution_std."""
    if df is None or len(df) == 0:
        return df
    out = df.copy()
    dist_n: list[int | None] = []
    dist_std: list[float | None] = []
    for _, row in out.iterrows():
        g_vals = _parse_g_vals(row, prefix=g_prefix)
        n = len(g_vals)
        dist_n.append(n)
        if n >= 2:
            dist_std.append(round(float(pd.Series(g_vals).std(ddof=1)), 4))
        else:
            dist_std.append(None)
    out["distribution_n"] = dist_n
    out["distribution_std"] = dist_std
    return out


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    df2 = df2.where(pd.notna(df2), None)
    # Convert numeric minutes_tier (0-3) back to human labels
    if "minutes_tier" in df2.columns:
        _mt_num = pd.to_numeric(df2["minutes_tier"], errors="coerce")
        _mt_valid = _mt_num.notna()
        if _mt_valid.any():
            df2.loc[_mt_valid, "minutes_tier"] = _mt_num[_mt_valid].round().astype(int).map(_MIN_TIER_NUM_MAP).fillna(df2.loc[_mt_valid, "minutes_tier"])
    df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%-I:%M %p")
    # Calendar date for same-day MLB grading (avoids grading full multi-day slate vs one day's games).
    _gd = _row_game_datetimes(df2)
    df2["slate_game_date"] = _gd.dt.strftime("%Y-%m-%d").where(_gd.notna(), "").fillna("")

    keep = [
        "tier", "rank_score",
        "player", "pos", "player_type_norm", "team", "opp_team", "days_rest", "is_back_to_back", "opp_days_rest", "opp_b2b",
        "h2h_avg", "h2h_over_pct", "h2h_games", "h2h_last",
        "game_total", "spread", "game_time", "slate_game_date",
        "prop_type", "pick_type", "line",
        "final_bet_direction",
        "edge", "projection",
        "ml_prob",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "hit_rate_status", "reliability_note",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "OVERALL_DEF_RANK", "DEF_TIER",
        "minutes_tier", "batting_order_tier", "pitcher_role",
        "same_series_hit_rate",
        "void_reason",
        "open_line",
        "line_movement",
        "line_direction_shift",
        "implied_prob",
        "implied_prob_over",
        "implied_prob_under",
    ]
    keep = [c for c in keep if c in df2.columns]
    # Rolling game values (step4): required so combined slate / UI L5 Over|Under match game logs.
    stat_g_cols = sorted(
        (c for c in df2.columns if c.startswith("stat_g") and c[6:].isdigit()),
        key=lambda c: int(c[6:]),
    )
    for c in stat_g_cols:
        if c not in keep:
            keep.append(c)
    for c in ("distribution_std", "distribution_n"):
        if c in df2.columns and c not in keep:
            keep.append(c)
    clean = df2[keep].copy()

    for col in [
        "rank_score",
        "edge",
        "abs_edge",
        "projection",
        "ml_prob",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "same_series_hit_rate",
        "open_line",
        "line_movement",
        "implied_prob",
        "implied_prob_over",
        "implied_prob_under",
    ]:
        if col in clean.columns:
            if col in ("implied_prob", "implied_prob_over", "implied_prob_under"):
                rnd = 4
            else:
                rnd = 4 if col in ("ml_prob", "edge_score", "blended_score") else (3 if col == "line_movement" else 2)
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(rnd)
    if "line_direction_shift" in clean.columns:
        clean["line_direction_shift"] = (
            clean["line_direction_shift"].astype(str).str.strip().replace({"nan": "", "None": ""})
        )
        clean.loc[clean["line_direction_shift"].eq(""), "line_direction_shift"] = "stable"
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")
    if "distribution_std" in clean.columns:
        clean["distribution_std"] = pd.to_numeric(clean["distribution_std"], errors="coerce").round(4)
    if "distribution_n" in clean.columns:
        clean["distribution_n"] = pd.to_numeric(clean["distribution_n"], errors="coerce").astype("Int64")

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "player": "Player", "pos": "Pos", "player_type_norm": "Player Type",
        "team": "Team", "opp_team": "Opp", "game_time": "Game Time",
        "days_rest": "Days Rest",
        "is_back_to_back": "B2B",
        "opp_days_rest": "Opp Rest",
        "opp_b2b": "Opp B2B",
        "h2h_avg": "H2H Avg",
        "h2h_over_pct": "H2H Over%",
        "h2h_games": "H2H Games",
        "h2h_last": "H2H Last",
        "game_total": "Game Total",
        "spread": "Spread",
        "slate_game_date": "Game Date",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "final_bet_direction": "Direction",
        "edge": "Edge", "abs_edge": "Abs Edge", "projection": "Projection",
        "ml_prob": "ML Prob",
        "edge_score": "Edge Score",
        "blended_score": "Blended Score",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "hit_rate_status": "Hit Rate Status",
        "reliability_note": "Reliability Note",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier",
        "minutes_tier": "Min Tier", "batting_order_tier": "Bat Order",
        "pitcher_role": "Pitcher Role",
        "same_series_hit_rate": "Series HR",
        "void_reason": "Void Reason",
    }
    # Keep snake_case line-movement cols (NHL step8 / combined audit contract).
    _lm_cols = (
        "open_line",
        "line_movement",
        "line_direction_shift",
        "implied_prob",
        "implied_prob_over",
        "implied_prob_under",
    )
    rename = {k: v for k, v in rename.items() if k not in _lm_cols}
    clean = clean.rename(columns=rename)
    clean = clean.where(pd.notna(clean), None)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "ALL", clean, HEADER_COLOR)

    for tier in ["A", "B", "C", "D"]:
        subset = clean[clean["Tier"] == tier].copy()
        if len(subset):
            tier_bg = TIER_COLORS.get(tier, ("333333",))[0]
            write_sheet(wb, f"Tier {tier}", subset, tier_bg)

    # Pitcher / Hitter split tabs
    if "Player Type" in clean.columns:
        pitchers = clean[clean["Player Type"].astype(str).str.lower() == "pitcher"].copy()
        hitters  = clean[clean["Player Type"].astype(str).str.lower() == "hitter"].copy()
        if len(pitchers): write_sheet(wb, "Pitchers", pitchers, PITCHER_TAB_COLOR)
        if len(hitters):  write_sheet(wb, "Hitters",  hitters,  HITTER_TAB_COLOR)

    wb.save(xlsx_path)
    print(f"Clean XLSX saved -> {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="step7_mlb_ranked.xlsx")
    ap.add_argument("--sheet", default="ALL")
    ap.add_argument("--output", default="step8_mlb_direction.csv")
    ap.add_argument(
        "--xlsx",
        default="step8_mlb_direction_clean.xlsx",
        help="Path for styled multi-sheet workbook (default: step8_mlb_direction_clean.xlsx).",
    )
    ap.add_argument(
        "--date",
        default="",
        help="YYYY-MM-DD pipeline slate date for outputs/<date>/ archive (default: today).",
    )
    args = ap.parse_args()

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)

    print(f"Loading: {args.input} (sheet={args.sheet})")
    df  = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")
    if df.empty:
        raise SystemExit("ERROR [PropOracle-MLB-S8] Empty input from step7; aborting.")
    out = df.copy()

    reconcile_signed_edge_abs_dataframe(out)

    edge     = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = pd.to_numeric(out["abs_edge"], errors="coerce")

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    forced    = pick_type.isin(["Goblin", "Demon"])

    model_dir = np.where(edge >= 0, "OVER", "UNDER")
    # Step7 sets bet_direction (bottom×tough Standard→UNDER, Goblin/Demon stay OVER).
    step7_dir = (
        out.get("bet_direction", out.get("recommended_side", pd.Series("", index=out.index)))
        .astype(str)
        .str.upper()
        .str.strip()
    )
    from_step7 = step7_dir.isin(("OVER", "UNDER"))
    final_dir = np.where(from_step7, step7_dir, model_dir)
    _ov = out.get("direction_override", pd.Series("", index=out.index)).astype(str).str.strip()
    reason = np.where(
        from_step7 & _ov.eq("BOTTOM3_TOUGH_UNDER"),
        "BOTTOM3_TOUGH_UNDER",
        np.where(abs_edge < 0.03, "MODEL_TIEBREAK_DIFF<0.03", ""),
    )
    final_dir = np.where(forced, "OVER", final_dir)
    reason = np.where(forced, "FORCED_OVER_ONLY_GOB_DEM", reason)

    out["model_dir"]           = model_dir
    out["final_bet_direction"] = final_dir
    out["final_dir_reason"]    = reason

    if out.empty:
        raise SystemExit("ERROR [PropOracle-MLB-S8] Empty output after direction step; aborting.")

    out = _attach_distribution_std(out)
    filled_std = int(pd.to_numeric(out["distribution_std"], errors="coerce").notna().sum())
    print(f"[MLB step8] distribution_std filled {filled_std}/{len(out)} rows")

    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"Saved -> {args.output}")
    print("final_bet_direction:", pd.Series(final_dir).value_counts().to_dict())
    if "tier" in out.columns:
        print("tier:", out["tier"].value_counts().to_dict())

    xlsx_path = args.xlsx if args.xlsx else args.output.replace(".csv", "_clean.xlsx")
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    build_clean_xlsx(out, xlsx_path)
    _copy_dated_step8_mlb(xlsx_path, (args.date or "").strip())


if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR [PropOracle-MLB-S8] {type(e).__name__}: {e}")
        raise SystemExit(1) from e
