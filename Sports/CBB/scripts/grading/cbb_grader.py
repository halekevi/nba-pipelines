#!/usr/bin/env python3
"""
Builds prop_grader.py — a standalone script that:
1. Takes a tickets xlsx (best_tickets.xlsx or cbb_tickets.xlsx)
2. Takes final scores (manual input or ESPN API)
3. Grades each leg HIT/MISS/PUSH
4. Outputs a detailed grading workbook with breakdowns
"""

import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from openpyxl.formatting.rule import ColorScaleRule, DataBarRule
from datetime import datetime
import argparse
import os

# ── Colors ────────────────────────────────────────────────────────────────────
C = {
    'hit':       '27AE60',
    'miss':      'E74C3C',
    'push':      'F39C12',
    'hdr_dark':  '1C1C1C',
    'hdr_blue':  '1A5276',
    'hdr_green': '1E8449',
    'hdr_red':   '922B21',
    'hdr_gold':  '7D6608',
    'row_alt':   'F2F3F4',
    'row_white': 'FFFFFF',
    'tier_a':    'D5F5E3',
    'tier_b':    'D6EAF8',
    'tier_c':    'FEF9E7',
    'tier_d':    'FDEDEC',
}

def thin(color='CCCCCC'):
    s = Side(style='thin', color=color)
    return Border(left=s, right=s, top=s, bottom=s)

def hdr_cell(ws, row, col, value, bg=None, color='FFFFFF', bold=True, sz=9, wrap=False, align='center'):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, color=color, name='Arial', size=sz)
    if bg:
        c.fill = PatternFill('solid', start_color=bg)
    c.alignment = Alignment(horizontal=align, vertical='center', wrap_text=wrap)
    c.border = thin()
    return c

def data_cell(ws, row, col, value, bg=None, bold=False, sz=9, align='center', num_fmt=None):
    c = ws.cell(row=row, column=col, value=value)
    c.font = Font(bold=bold, name='Arial', size=sz)
    c.fill = PatternFill('solid', start_color=bg or C['row_white'])
    c.alignment = Alignment(horizontal=align, vertical='center')
    c.border = thin()
    if num_fmt:
        c.number_format = num_fmt
    return c

def result_color(result):
    r = str(result).upper()
    if r == 'HIT':   return C['hit']
    if r == 'MISS':  return C['miss']
    if r == 'PUSH':  return C['push']
    return 'DDDDDD'

def pct_color(val):
    if val is None: return 'DDDDDD'
    if val >= 0.70: return C['hit']
    if val >= 0.50: return C['push']
    return C['miss']

# ── Sheet 1: RAW GRADES ───────────────────────────────────────────────────────
def build_raw_sheet(wb, df):
    ws = wb.create_sheet('RAW GRADES')
    cols = ['Date','Sport','Ticket #','Leg','Player','Team','Opp',
            'Prop','Pick Type','Line','Direction','Edge','Hit Rate',
            'L5 Avg','Season Avg','Rank Score','Tier',
            'Actual','Result','Margin']
    widths = [11,6,9,4,22,6,6,16,10,7,10,7,9,8,11,11,5,9,8,8]

    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        hdr_cell(ws, 1, ci, h, bg=C['hdr_dark'])
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for i, row in enumerate(df.itertuples(), 2):
        bg = C['row_alt'] if i % 2 == 0 else C['row_white']
        res = str(getattr(row, 'Result', '')).upper()
        res_bg = result_color(res)

        vals = [
            getattr(row, 'Date', ''),
            getattr(row, 'Sport', ''),
            getattr(row, 'Ticket_num', ''),
            getattr(row, 'Leg', ''),
            getattr(row, 'Player', ''),
            getattr(row, 'Team', ''),
            getattr(row, 'Opp', ''),
            getattr(row, 'Prop', ''),
            getattr(row, 'Pick_Type', ''),
            getattr(row, 'Line', ''),
            getattr(row, 'Direction', ''),
            getattr(row, 'Edge', ''),
            getattr(row, 'Hit_Rate', ''),
            getattr(row, 'L5_Avg', ''),
            getattr(row, 'Season_Avg', ''),
            getattr(row, 'Rank_Score', ''),
            getattr(row, 'Tier', ''),
            getattr(row, 'Actual', ''),
            res,
            getattr(row, 'Margin', ''),
        ]
        for ci, val in enumerate(vals, 1):
            c_bg = res_bg if cols[ci-1] == 'Result' else bg
            c = data_cell(ws, i, ci, val, bg=c_bg)
            if cols[ci-1] == 'Result':
                c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
            if cols[ci-1] in ('Hit Rate',):
                if val != '' and val is not None:
                    try: c.number_format = '0%'
                    except: pass

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return ws

# ── Sheet 2: TICKET SUMMARY ───────────────────────────────────────────────────
def build_ticket_summary(wb, df):
    ws = wb.create_sheet('TICKET SUMMARY')

    tickets = df.groupby(['Date','Sport','Ticket_num','Pick_Type','Legs'])
    rows = []
    for (date, sport, tnum, ptype, legs), grp in tickets:
        leg_results = grp['Result'].str.upper().tolist()
        n = len(leg_results)
        hits = leg_results.count('HIT')
        misses = leg_results.count('MISS')
        pushes = leg_results.count('PUSH')
        ticket_result = 'HIT' if misses == 0 and hits > 0 else ('PUSH' if misses == 0 and pushes > 0 else 'MISS')
        avg_score = grp['Rank_Score'].mean() if 'Rank_Score' in grp.columns else None
        avg_hr = grp['Hit_Rate'].mean() if 'Hit_Rate' in grp.columns else None
        players = ' | '.join(f"{r.Player} {r.Direction} {r.Prop} {r.Line}" for r in grp.itertuples())
        rows.append({
            'Date': date, 'Sport': sport, 'Ticket': tnum,
            'Pick Type': ptype, 'Legs': legs,
            'Result': ticket_result,
            'Legs Hit': f"{hits}/{n}",
            'Avg Score': round(avg_score, 2) if avg_score else '',
            'Avg Hit Rate': avg_hr,
            'Players': players,
        })

    cols = ['Date','Sport','Ticket','Pick Type','Legs','Result','Legs Hit','Avg Score','Avg Hit Rate','Players']
    widths = [11,6,8,11,5,8,9,11,13,80]

    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        hdr_cell(ws, 1, ci, h, bg=C['hdr_blue'])
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for i, row in enumerate(rows, 2):
        bg = C['row_alt'] if i % 2 == 0 else C['row_white']
        for ci, col in enumerate(cols, 1):
            val = row[col]
            c_bg = result_color(row['Result']) if col == 'Result' else bg
            c = data_cell(ws, i, ci, val, bg=c_bg,
                         align='left' if col == 'Players' else 'center')
            if col == 'Result':
                c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
            if col == 'Avg Hit Rate' and val:
                c.number_format = '0%'

    ws.auto_filter.ref = f"A1:{get_column_letter(len(cols))}1"
    return ws

# ── Sheet 3: BREAKDOWN ────────────────────────────────────────────────────────
def build_breakdown_sheet(wb, df, title, group_col, bg_hdr):
    ws = wb.create_sheet(title)
    graded = df[df['Result'].str.upper().isin(['HIT','MISS','PUSH'])].copy()

    groups = graded.groupby(group_col)
    rows = []
    for key, grp in groups:
        total = len(grp)
        hits  = (grp['Result'].str.upper() == 'HIT').sum()
        misses= (grp['Result'].str.upper() == 'MISS').sum()
        pushes= (grp['Result'].str.upper() == 'PUSH').sum()
        hr    = hits / total if total else 0
        avg_margin = grp['Margin'].mean() if 'Margin' in grp.columns and grp['Margin'].notna().any() else None
        avg_edge   = grp['Edge'].mean()   if 'Edge' in grp.columns   and grp['Edge'].notna().any()   else None
        rows.append({
            group_col: key, 'Total': total,
            'Hits': hits, 'Misses': misses, 'Pushes': pushes,
            'Hit Rate': hr,
            'Avg Edge': round(avg_edge, 1) if avg_edge is not None else '',
            'Avg Margin': round(avg_margin, 1) if avg_margin is not None else '',
            'Signal': ('Strong ✅' if hr >= 0.70 else ('Weak ⚠' if hr < 0.50 else 'Average')),
        })

    rows.sort(key=lambda x: -x['Total'])

    cols = [group_col,'Total','Hits','Misses','Pushes','Hit Rate','Avg Edge','Avg Margin','Signal']
    widths = [22, 7, 7, 8, 8, 10, 10, 12, 14]

    for ci, (h, w) in enumerate(zip(cols, widths), 1):
        ws.column_dimensions[get_column_letter(ci)].width = w
        hdr_cell(ws, 1, ci, h, bg=bg_hdr)
    ws.row_dimensions[1].height = 20
    ws.freeze_panes = 'A2'

    for i, row in enumerate(rows, 2):
        bg = C['row_alt'] if i % 2 == 0 else C['row_white']
        hr = row['Hit Rate']
        hr_bg = pct_color(hr)
        for ci, col in enumerate(cols, 1):
            val = row[col]
            c_bg = hr_bg if col == 'Hit Rate' else bg
            c = data_cell(ws, i, ci, val, bg=c_bg,
                         align='left' if col == group_col else 'center')
            if col == 'Hit Rate':
                c.number_format = '0.0%'
                c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')

    # Totals row
    total_row = len(rows) + 2
    total_graded = len(graded)
    total_hits   = (graded['Result'].str.upper() == 'HIT').sum()
    overall_hr   = total_hits / total_graded if total_graded else 0

    hdr_cell(ws, total_row, 1, 'TOTAL', bg=C['hdr_dark'])
    data_cell(ws, total_row, 2, total_graded, bg=C['hdr_dark'], bold=True)
    data_cell(ws, total_row, 3, total_hits, bg=C['hdr_dark'], bold=True)
    data_cell(ws, total_row, 4, (graded['Result'].str.upper()=='MISS').sum(), bg=C['hdr_dark'], bold=True)
    data_cell(ws, total_row, 5, (graded['Result'].str.upper()=='PUSH').sum(), bg=C['hdr_dark'], bold=True)
    c = data_cell(ws, total_row, 6, overall_hr, bg=pct_color(overall_hr), bold=True)
    c.number_format = '0.0%'
    c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
    for ci in range(7, len(cols)+1):
        data_cell(ws, total_row, ci, '', bg=C['hdr_dark'])

    return ws

# ── Sheet 4: OVER/UNDER BREAKDOWN ─────────────────────────────────────────────
def build_direction_sheet(wb, df):
    ws = wb.create_sheet('BY DIRECTION')
    graded = df[df['Result'].str.upper().isin(['HIT','MISS','PUSH'])].copy()

    # Cross-tab: Direction x Pick Type
    directions = ['OVER', 'UNDER']
    pick_types  = sorted(graded['Pick_Type'].dropna().unique())

    hdr_cell(ws, 1, 1, 'Direction / Pick Type', bg=C['hdr_dark'])
    for ci, pt in enumerate(pick_types, 2):
        hdr_cell(ws, 1, ci, pt, bg=C['hdr_dark'])
    hdr_cell(ws, 1, len(pick_types)+2, 'TOTAL', bg=C['hdr_dark'])

    ws.column_dimensions['A'].width = 22
    for ci in range(2, len(pick_types)+3):
        ws.column_dimensions[get_column_letter(ci)].width = 16

    for ri, direction in enumerate(directions, 2):
        sub = graded[graded['Direction'].str.upper() == direction]
        hdr_cell(ws, ri, 1, direction, bg=C['hdr_blue'])
        for ci, pt in enumerate(pick_types, 2):
            sub2 = sub[sub['Pick_Type'] == pt]
            total = len(sub2)
            hits  = (sub2['Result'].str.upper() == 'HIT').sum()
            hr = hits/total if total else None
            val = f"{hits}/{total} ({hr:.0%})" if total else '-'
            c = data_cell(ws, ri, ci, val, bg=pct_color(hr))
            if hr is not None:
                c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        # Total column
        total = len(sub)
        hits  = (sub['Result'].str.upper() == 'HIT').sum()
        hr    = hits/total if total else None
        val   = f"{hits}/{total} ({hr:.0%})" if total else '-'
        c = data_cell(ws, ri, len(pick_types)+2, val, bg=pct_color(hr), bold=True)
        if hr is not None:
            c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')

    # Total row
    ri = len(directions) + 2
    hdr_cell(ws, ri, 1, 'TOTAL', bg=C['hdr_dark'])
    for ci, pt in enumerate(pick_types, 2):
        sub2 = graded[graded['Pick_Type'] == pt]
        total = len(sub2)
        hits  = (sub2['Result'].str.upper() == 'HIT').sum()
        hr    = hits/total if total else None
        val   = f"{hits}/{total} ({hr:.0%})" if total else '-'
        c = data_cell(ws, ri, ci, val, bg=pct_color(hr), bold=True)
        if hr is not None:
            c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
    total = len(graded)
    hits  = (graded['Result'].str.upper() == 'HIT').sum()
    hr    = hits/total if total else None
    val   = f"{hits}/{total} ({hr:.0%})" if total else '-'
    c = data_cell(ws, ri, len(pick_types)+2, val, bg=pct_color(hr), bold=True)
    if hr is not None:
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')

    return ws

# ── Sheet 5: DASHBOARD ────────────────────────────────────────────────────────
def build_dashboard(wb, df):
    ws = wb.create_sheet('DASHBOARD', 0)
    graded = df[df['Result'].str.upper().isin(['HIT','MISS','PUSH'])].copy()

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 16
    ws.column_dimensions['C'].width = 16
    ws.column_dimensions['D'].width = 16
    ws.column_dimensions['E'].width = 16
    ws.column_dimensions['F'].width = 20

    # Title
    ws.merge_cells('A1:F1')
    c = ws['A1']
    c.value = f"PROP GRADER DASHBOARD  |  Generated {datetime.now().strftime('%Y-%m-%d %H:%M')}"
    c.font = Font(bold=True, name='Arial', size=12, color='FFFFFF')
    c.fill = PatternFill('solid', start_color=C['hdr_dark'])
    c.alignment = Alignment(horizontal='center', vertical='center')
    ws.row_dimensions[1].height = 28

    def section_hdr(row, title, bg):
        ws.merge_cells(f'A{row}:F{row}')
        c = ws[f'A{row}']
        c.value = title
        c.font = Font(bold=True, name='Arial', size=10, color='FFFFFF')
        c.fill = PatternFill('solid', start_color=bg)
        c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
        ws.row_dimensions[row].height = 22

    def stat_row(row, label, *vals):
        data_cell(ws, row, 1, label, bold=True, align='left')
        for ci, v in enumerate(vals, 2):
            data_cell(ws, row, ci, v, align='center')

    # Overall stats
    section_hdr(2, 'OVERALL LEG PERFORMANCE', C['hdr_blue'])
    hdr_cell(ws, 3, 1, 'Metric', bg=C['hdr_dark'])
    hdr_cell(ws, 3, 2, 'Total', bg=C['hdr_dark'])
    hdr_cell(ws, 3, 3, 'Hits', bg=C['hdr_dark'])
    hdr_cell(ws, 3, 4, 'Misses', bg=C['hdr_dark'])
    hdr_cell(ws, 3, 5, 'Pushes', bg=C['hdr_dark'])
    hdr_cell(ws, 3, 6, 'Hit Rate', bg=C['hdr_dark'])

    sports = ['ALL'] + sorted(graded['Sport'].dropna().unique().tolist())
    for ri, sport in enumerate(sports, 4):
        sub = graded if sport == 'ALL' else graded[graded['Sport'] == sport]
        total  = len(sub)
        hits   = (sub['Result'].str.upper() == 'HIT').sum()
        misses = (sub['Result'].str.upper() == 'MISS').sum()
        pushes = (sub['Result'].str.upper() == 'PUSH').sum()
        hr     = hits/total if total else 0
        bg = C['row_alt'] if ri % 2 == 0 else C['row_white']
        data_cell(ws, ri, 1, sport, bold=(sport=='ALL'), align='left')
        data_cell(ws, ri, 2, total,  bg=bg)
        data_cell(ws, ri, 3, int(hits),   bg=bg)
        data_cell(ws, ri, 4, int(misses), bg=bg)
        data_cell(ws, ri, 5, int(pushes), bg=bg)
        c = data_cell(ws, ri, 6, hr, bg=pct_color(hr), bold=True)
        c.number_format = '0.0%'
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')

    row = 4 + len(sports) + 1

    # By Pick Type
    section_hdr(row, 'BY PICK TYPE', C['hdr_green'])
    row += 1
    hdr_cell(ws, row, 1, 'Pick Type', bg=C['hdr_dark'])
    for ci, h in enumerate(['Total','Hits','Misses','Hit Rate','Signal'], 2):
        hdr_cell(ws, row, ci, h, bg=C['hdr_dark'])
    row += 1
    for pt in sorted(graded['Pick_Type'].dropna().unique()):
        sub    = graded[graded['Pick_Type'] == pt]
        total  = len(sub)
        hits   = (sub['Result'].str.upper() == 'HIT').sum()
        misses = (sub['Result'].str.upper() == 'MISS').sum()
        hr     = hits/total if total else 0
        signal = 'Strong ✅' if hr >= 0.70 else ('Weak ⚠' if hr < 0.50 else 'Average')
        bg = C['row_alt'] if row % 2 == 0 else C['row_white']
        data_cell(ws, row, 1, pt, align='left')
        data_cell(ws, row, 2, total,  bg=bg)
        data_cell(ws, row, 3, int(hits),   bg=bg)
        data_cell(ws, row, 4, int(misses), bg=bg)
        c = data_cell(ws, row, 5, hr, bg=pct_color(hr), bold=True)
        c.number_format = '0.0%'
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        data_cell(ws, row, 6, signal, bg=bg)
        row += 1

    row += 1

    # By Tier
    section_hdr(row, 'BY TIER', C['hdr_gold'])
    row += 1
    hdr_cell(ws, row, 1, 'Tier', bg=C['hdr_dark'])
    for ci, h in enumerate(['Total','Hits','Misses','Hit Rate','Calibration'], 2):
        hdr_cell(ws, row, ci, h, bg=C['hdr_dark'])
    row += 1
    for tier in ['A','B','C','D']:
        if 'Tier' not in graded.columns: break
        sub    = graded[graded['Tier'].astype(str).str.upper() == tier]
        if len(sub) == 0: continue
        total  = len(sub)
        hits   = (sub['Result'].str.upper() == 'HIT').sum()
        misses = (sub['Result'].str.upper() == 'MISS').sum()
        hr     = hits/total if total else 0
        note   = ('Model accurate ✅' if (tier == 'A' and hr >= 0.70) else
                  ('Over-ranked ⚠' if tier == 'A' and hr < 0.55 else
                   ('Under-ranked' if tier in ('C','D') and hr >= 0.70 else '')))
        tier_bg = {'A':C['tier_a'],'B':C['tier_b'],'C':C['tier_c'],'D':C['tier_d']}.get(tier, C['row_white'])
        data_cell(ws, row, 1, f'Tier {tier}', bg=tier_bg, bold=True, align='left')
        data_cell(ws, row, 2, total,  bg=tier_bg)
        data_cell(ws, row, 3, int(hits),   bg=tier_bg)
        data_cell(ws, row, 4, int(misses), bg=tier_bg)
        c = data_cell(ws, row, 5, hr, bg=pct_color(hr), bold=True)
        c.number_format = '0.0%'
        c.font = Font(bold=True, name='Arial', size=9, color='FFFFFF')
        data_cell(ws, row, 6, note, bg=tier_bg, align='left')
        row += 1

    row += 1

    # Model calibration note
    section_hdr(row, 'MODEL CALIBRATION', C['hdr_red'])
    row += 1
    total_graded = len(graded)
    total_hits   = (graded['Result'].str.upper() == 'HIT').sum()
    overall_hr   = total_hits/total_graded if total_graded else 0

    if overall_hr >= 0.65:
        note = f"Overall leg HR {overall_hr:.0%} — model well calibrated ✅"
    elif overall_hr >= 0.55:
        note = f"Overall leg HR {overall_hr:.0%} — model OK, room to improve"
    else:
        note = f"Overall leg HR {overall_hr:.0%} — model needs tuning ⚠  Consider raising min_hit_rate or adjusting tier thresholds"

    ws.merge_cells(f'A{row}:F{row}')
    c = ws[f'A{row}']
    c.value = note
    c.font = Font(name='Arial', size=9, italic=True)
    c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
    c.border = thin()
    ws.row_dimensions[row].height = 18

    return ws


def load_sample_data():
    """Returns a sample DataFrame for testing when no input file provided."""
    import random
    random.seed(42)
    props = [
        ('Jarrett Allen','CLE','BOS','Points','Goblin',9.5,'OVER',8.7,0.80,9.2,9.1,2.31,'A'),
        ('OG Anunoby','NYK','MIA','Pts+Rebs+Asts','Standard',21.5,'OVER',6.2,0.80,22.1,21.8,2.45,'A'),
        ('Grayson Allen','PHX','LAL','Pts+Asts','Standard',17.5,'OVER',5.8,0.80,18.2,17.9,2.20,'A'),
        ('Jamal Murray','DEN','GSW','Pts+Rebs+Asts','Standard',34.5,'OVER',4.1,0.80,35.1,34.8,2.15,'A'),
        ('Harrison Barnes','SAC','POR','Points','Goblin',4.5,'OVER',9.1,0.80,5.2,5.0,2.38,'A'),
        ('Tyrese Haliburton','IND','CHI','Points','Demon',22.5,'OVER',3.2,0.75,23.1,22.8,1.85,'B'),
        ('Darius Garland','CLE','BOS','Assists','Standard',6.5,'OVER',2.1,0.60,6.8,6.6,1.62,'B'),
        ('Evan Mobley','CLE','BOS','Rebounds','Goblin',7.5,'OVER',5.5,0.60,8.1,7.8,1.70,'B'),
        ('Desmond Bane','MEM','OKC','Points','Standard',18.5,'UNDER',2.8,0.60,17.9,18.2,1.55,'C'),
        ('Kyle Kuzma','WAS','ATL','Pts+Rebs','Goblin',26.5,'OVER',1.2,0.55,27.1,26.8,1.20,'C'),
    ]
    rows = []
    for date in ['2026-02-17','2026-02-18','2026-02-19']:
        for sport in ['CBB']:
            for t_num in range(1, 4):
                for leg, (player, team, opp, prop, ptype, line, direction,
                          edge, hr, l5, ssn, score, tier) in enumerate(
                    random.sample(props, 3), 1):
                    actual = round(line + random.uniform(-5, 5), 1)
                    if direction == 'OVER':
                        result = 'HIT' if actual > line else ('PUSH' if actual == line else 'MISS')
                    else:
                        result = 'HIT' if actual < line else ('PUSH' if actual == line else 'MISS')
                    margin = round(actual - line if direction == 'OVER' else line - actual, 1)
                    rows.append({
                        'Date': date, 'Sport': sport,
                        'Ticket_num': t_num, 'Leg': leg,
                        'Player': player, 'Team': team, 'Opp': opp,
                        'Prop': prop, 'Pick_Type': ptype,
                        'Line': line, 'Direction': direction,
                        'Edge': edge, 'Hit_Rate': hr,
                        'L5_Avg': l5, 'Season_Avg': ssn,
                        'Rank_Score': score, 'Tier': tier,
                        'Actual': actual, 'Result': result,
                        'Margin': margin, 'Legs': 3,
                    })
    return pd.DataFrame(rows)


def main():
    ap = argparse.ArgumentParser(description='Prop Pipeline Grader')
    ap.add_argument('--input',  default='',     help='CSV with graded legs (see --template)')
    ap.add_argument('--output', default='cbb_grades.xlsx')
    ap.add_argument('--template', action='store_true', help='Output a blank input template CSV')
    ap.add_argument('--demo',   action='store_true', help='Run with sample data')
    args = ap.parse_args()

    template_cols = ['Date','Sport','Ticket_num','Leg','Player','Team','Opp',
                     'Prop','Pick_Type','Line','Direction','Edge','Hit_Rate',
                     'L5_Avg','Season_Avg','Rank_Score','Tier','Actual','Result',
                     'Margin','Legs']

    if args.template:
        pd.DataFrame(columns=template_cols).to_csv('grader_template.csv', index=False)
        print('Saved grader_template.csv — fill in and run with --input grader_template.csv')
        return

    if args.demo or not args.input:
        print('Running in DEMO mode with sample data...')
        df = load_sample_data()
    else:
        df = pd.read_csv(args.input)
        for col in ['Edge','Hit_Rate','L5_Avg','Season_Avg','Rank_Score','Line','Actual','Margin']:
            if col in df.columns:
                df[col] = pd.to_numeric(df[col], errors='coerce')

    wb = Workbook()
    wb.remove(wb.active)

    build_dashboard(wb, df)
    build_ticket_summary(wb, df)
    build_raw_sheet(wb, df)
    build_breakdown_sheet(wb, df, 'BY PROP TYPE',  'Prop',      C['hdr_blue'])
    build_breakdown_sheet(wb, df, 'BY PICK TYPE',  'Pick_Type', C['hdr_green'])
    build_breakdown_sheet(wb, df, 'BY PLAYER',     'Player',    C['hdr_gold'])
    build_breakdown_sheet(wb, df, 'BY SPORT',      'Sport',     C['hdr_red'])
    build_direction_sheet(wb, df)

    wb.save(args.output)
    print(f'Saved -> {args.output}')
    print()
    print('Sheets:')
    for s in wb.sheetnames:
        print(f'  {s}')
    print()
    print('Usage:')
    print('  Demo:     py -3.14 cbb_grader.py --demo')
    print('  Template: py -3.14 cbb_grader.py --template')
    print('  Grade:    py -3.14 cbb_grader.py --input my_grades.csv --output cbb_grades.xlsx')

if __name__ == '__main__':
    main()
