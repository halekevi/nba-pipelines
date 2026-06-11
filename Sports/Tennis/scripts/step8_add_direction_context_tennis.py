#!/usr/bin/env python3
"""
step8_add_direction_context_tennis.py  (Tennis Pipeline)

Mirrors NBA step8_add_direction_context.py.
Reads step7_tennis_ranked.xlsx, appends direction context,
outputs full CSV + clean formatted XLSX.

Run:
    py -3.14 step8_add_direction_context_tennis.py \
    --input outputs/step7_tennis_ranked.xlsx \
    --output outputs/step8_tennis_direction.csv
"""

from __future__ import annotations

import argparse
import datetime as _dt
import shutil
import sys
import warnings
import zoneinfo
from pathlib import Path

import numpy as np
import pandas as pd

_REPO = Path(__file__).resolve().parent
for _ in range(10):
    if (_REPO / "utils" / "step8_edge_direction.py").is_file():
        if str(_REPO) not in sys.path:
            sys.path.insert(0, str(_REPO))
        break
    _REPO = _REPO.parent
else:
    raise RuntimeError("Could not locate repo root with utils/step8_edge_direction.py")

from scripts.l10_streak_utils import finalize_l10_ui_columns
from utils.hit_tracking_columns import HIT_TRACKING_RENAME, attach_hit_tracking_columns
from utils.step8_edge_direction import reconcile_signed_edge_abs_dataframe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


def _attach_unified_ml_prob(out: pd.DataFrame, repo_root: Path) -> pd.DataFrame:
    """Apply unified edge model ml_prob (non-fatal; keeps heuristic ml_prob on failure)."""
    scripts_dir = repo_root / "scripts"
    if str(scripts_dir) not in sys.path:
        sys.path.insert(0, str(scripts_dir))
    try:
        from edge_predict_utils import predict_unified_edge_scores  # noqa: WPS433

        work = out.copy()
        if "bet_direction" not in work.columns and "final_bet_direction" in work.columns:
            work["bet_direction"] = work["final_bet_direction"]
        pred = predict_unified_edge_scores(work, sport_for_model="TENNIS")
        if pred is None:
            print("[Tennis step8] WARN unified ml_prob unavailable — keeping step7 ml_prob")
            return out
        ml_p, edge_sc, blended = pred
        filled = int(ml_p.notna().sum())
        out["ml_prob"] = pd.to_numeric(ml_p, errors="coerce").round(4)
        out["edge_score"] = pd.to_numeric(edge_sc, errors="coerce").round(4)
        out["blended_score"] = pd.to_numeric(blended, errors="coerce").round(4)
        if "prob_source" in out.columns:
            out.loc[ml_p.notna(), "prob_source"] = "ml_prob_unified"
        print(f"[Tennis step8] unified ml_prob filled {filled}/{len(out)} rows")
    except Exception as exc:
        print(f"[Tennis step8] WARN unified ml_prob failed: {exc}")
        warnings.warn(f"Tennis unified ml_prob skipped: {exc}", stacklevel=2)
    return out


TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
HEADER_COLOR = "1C1C1C"

MISSING_SENTINELS = {"—", "-", "", "nan", "none", "null"}


def _parse_g_vals(row, prefix: str = "stat_g", n: int = 10) -> list[float]:
    vals: list[float] = []
    for i in range(1, n + 1):
        col = f"{prefix}{i}"
        if isinstance(row, pd.Series):
            raw_val = row.get(col, "")
        else:
            raw_val = row.get(col, "")
        raw = str(raw_val if raw_val is not None else "").strip()
        if raw.lower() in MISSING_SENTINELS:
            continue
        try:
            vals.append(float(raw))
        except ValueError:
            continue
    return vals


def _attach_distribution_std(df: pd.DataFrame) -> pd.DataFrame:
    """Population sample std (ddof=1) of stat_g1..10 for pipeline_read distribution_std."""
    if df is None or len(df) == 0:
        return df
    out = df.copy()
    dist_n: list[int | None] = []
    dist_std: list[float | None] = []
    for _, row in out.iterrows():
        g_vals = _parse_g_vals(row, prefix="stat_g")
        n = len(g_vals)
        dist_n.append(n)
        if n >= 2:
            dist_std.append(round(float(pd.Series(g_vals).std(ddof=1)), 4))
        else:
            dist_std.append(None)
    out["distribution_n"] = dist_n
    out["distribution_std"] = dist_std
    return out


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name: str, data: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    tier_bg, tier_fg = TIER_COLORS.get(name, ("333333", "FFFFFF"))
    headers = list(data.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        direction = row[headers.index("Direction")] if "Direction" in headers else ""
        for ci, val in enumerate(row, 1):
            col_name    = headers[ci - 1]
            display_val = "" if pd.isna(val) else val
            cell        = ws.cell(row=ri, column=ci, value=display_val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

            if col_name == "Tier":
                cell.fill = PatternFill("solid", start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name="Arial", size=9)
            elif col_name == "Direction":
                bg = "C8F7C5" if val == "OVER" else "F7C5C5" if val == "UNDER" else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, name="Arial", size=9)
            elif col_name == "League":
                cell.fill = PatternFill("solid", start_color="EBF5FB")
            else:
                cell.fill = PatternFill("solid", start_color="F9F9F9" if ri % 2 == 0 else "FFFFFF")

    col_widths = {
        "Tier": 6, "Rank Score": 10, "Surface": 10, "Player": 20, "Pos": 6, "Pos Group": 9,
        "Team": 12, "Opp": 12, "League": 12, "Game Time": 16,
        "ESPN ID": 10,
        "Prop": 18, "Pick Type": 10, "Line": 7,
        "Direction": 9, "Edge": 7, "Projection": 10,
        "ML Prob": 9, "Edge Score": 10, "Blended Score": 12,
        "Hit Rate (5g)": 12, "Hit Rate (10g)": 12, "Last 5 Avg": 10, "Season Avg": 10,
        "L5 Over": 8, "L5 Under": 8,
        "Def Rank": 9, "Def Tier": 10,
        "Min Tier": 9, "Shot Role": 10, "Usage Role": 10,
        "Void Reason": 20,
        "G1": 6, "G2": 6, "G3": 6, "G4": 6, "G5": 6,
        "G6": 6, "G7": 6, "G8": 6, "G9": 6, "G10": 6,
        "distribution_std": 10,
        "distribution_n": 8,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


_MIN_TIER_NUM_MAP = {0: "Low", 1: "Med", 2: "High", 3: "Elite"}


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    # Convert numeric minutes_tier (0-3) back to human labels
    if "minutes_tier" in df2.columns:
        _mt_num = pd.to_numeric(df2["minutes_tier"], errors="coerce")
        _mt_valid = _mt_num.notna()
        if _mt_valid.any():
            df2.loc[_mt_valid, "minutes_tier"] = _mt_num[_mt_valid].round().astype(int).map(_MIN_TIER_NUM_MAP).fillna(df2.loc[_mt_valid, "minutes_tier"])
    try:
        import platform
        _time_fmt = "%m/%d %#I:%M %p" if platform.system() == "Windows" else "%m/%d %-I:%M %p"
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime(_time_fmt)
    except Exception:
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%m/%d %H:%M")

    # Fill presentation-facing columns with safe fallbacks so the clean workbook
    # is populated even when historical stat cache coverage is sparse.
    hr5 = pd.to_numeric(df2.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")
    hr10 = pd.to_numeric(df2.get("line_hit_rate_over_ou_10", np.nan), errors="coerce").fillna(hr5).fillna(0.50)
    df2["line_hit_rate_over_ou_10"] = hr10

    proj = pd.to_numeric(df2.get("projection", np.nan), errors="coerce")
    df2["stat_last5_avg"] = pd.to_numeric(df2.get("stat_last5_avg", np.nan), errors="coerce").fillna(proj)
    df2["stat_season_avg"] = pd.to_numeric(df2.get("stat_season_avg", np.nan), errors="coerce").fillna(df2["stat_last5_avg"])
    df2["stat_last10_avg"] = pd.to_numeric(df2.get("stat_last10_avg", np.nan), errors="coerce").fillna(df2["stat_last5_avg"])

    l5_over = pd.to_numeric(df2.get("last5_over", np.nan), errors="coerce")
    l5_under = pd.to_numeric(df2.get("last5_under", np.nan), errors="coerce")
    # Approximate L5 split from hr5 when explicit counts are absent.
    l5_over_fallback = (hr5.fillna(0.5) * 5.0).round().clip(0, 5)
    l5_under_fallback = (5 - l5_over_fallback).clip(0, 5)
    df2["last5_over"] = l5_over.fillna(l5_over_fallback)
    df2["last5_under"] = l5_under.fillna(l5_under_fallback)

    if "line" in df2.columns:
        df2 = finalize_l10_ui_columns(df2, line_col="line")
    df2 = attach_hit_tracking_columns(df2, "TENNIS")

    if "OVERALL_DEF_RANK" not in df2.columns:
        df2["OVERALL_DEF_RANK"] = "N/A"
    else:
        df2["OVERALL_DEF_RANK"] = df2["OVERALL_DEF_RANK"].replace("", np.nan).fillna("N/A")
    if "DEF_TIER" not in df2.columns:
        df2["DEF_TIER"] = "N/A"
    else:
        df2["DEF_TIER"] = df2["DEF_TIER"].replace("", np.nan).fillna("N/A")

    for gcol in [f"stat_g{i}" for i in range(1, 11)]:
        if gcol not in df2.columns:
            df2[gcol] = "—"
        else:
            df2[gcol] = df2[gcol].replace("", np.nan).fillna("—")

    if "avg_minutes" in df2.columns:
        df2["avg_minutes"] = pd.to_numeric(df2["avg_minutes"], errors="coerce").fillna(0).round(1)
    else:
        df2["avg_minutes"] = 0.0
    if "game_script_note" not in df2.columns:
        df2["game_script_note"] = "N/A"
    else:
        df2["game_script_note"] = df2["game_script_note"].replace("", np.nan).fillna("N/A")

    if "surface" not in df2.columns:
        df2["surface"] = ""
    else:
        df2["surface"] = df2["surface"].astype(str).str.strip().replace({"nan": "", "None": ""})

    keep = [
        "tier", "rank_score",
        "surface",
        "player", "pos", "position_group", "team", "opp_team", "league", "game_time",
        "espn_player_id",
        "prop_type", "pick_type", "line",
        "final_bet_direction",
        "edge", "abs_edge", "projection",
        "ml_prob",
        "hit_rate", "hit_rate_l5", "hit_rate_l10",
        "strat_hit_rate", "strat_n",
        "player_hr_historical", "opp_hr_historical",
        "sport_signal_maturity", "confidence_tier", "confidence_score", "confidence_note",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "l10_over", "l10_under", "l10_over_pct", "l10_streak", "l10_games_played",
        "OVERALL_DEF_RANK", "DEF_TIER",
        "minutes_tier", "shot_role", "usage_role",
        "void_reason",
        # ── Game log ─────────────────────────────────────────────────────────
        "stat_g1", "stat_g2", "stat_g3", "stat_g4", "stat_g5",
        "stat_g6", "stat_g7", "stat_g8", "stat_g9", "stat_g10",
        "distribution_std", "distribution_n",
        "stat_last10_avg",
        # ── Schedule / context ───────────────────────────────────────────────
        "avg_minutes",
        "game_script_mult",
        "game_script_note",
    ]
    keep  = [c for c in keep if c in df2.columns]
    clean = df2[keep].copy()

    for col in [
        "rank_score",
        "edge",
        "abs_edge",
        "projection",
        "ml_prob",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
    ]:
        if col in clean.columns:
            rnd = 4 if col in ("ml_prob", "edge_score", "blended_score") else 2
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(rnd)
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")
    if "distribution_n" in clean.columns:
        clean["distribution_n"] = pd.to_numeric(clean["distribution_n"], errors="coerce").astype("Int64")
    if "distribution_std" in clean.columns:
        clean["distribution_std"] = pd.to_numeric(clean["distribution_std"], errors="coerce").round(4)

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}

    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "surface": "Surface",
        "player": "Player", "pos": "Pos", "position_group": "Pos Group",
        "team": "Team", "opp_team": "Opp", "league": "League", "game_time": "Game Time",
        "espn_player_id": "ESPN ID",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "final_bet_direction": "Direction",
        "edge": "Edge", "abs_edge": "Abs Edge", "projection": "Projection",
        "ml_prob": "ML Prob",
        "edge_score": "Edge Score",
        "blended_score": "Blended Score",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "line_hit_rate_over_ou_10": "Hit Rate (10g)",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier",
        "minutes_tier": "Min Tier", "shot_role": "Shot Role", "usage_role": "Usage Role",
        "void_reason": "Void Reason",
        # Game log
        "stat_last10_avg": "Last 10 Avg",
        "stat_g1": "G1", "stat_g2": "G2", "stat_g3": "G3",
        "stat_g4": "G4", "stat_g5": "G5", "stat_g6": "G6",
        "stat_g7": "G7", "stat_g8": "G8", "stat_g9": "G9", "stat_g10": "G10",
        "distribution_std": "distribution_std",
        "distribution_n": "distribution_n",
        # Context
        "avg_minutes": "Avg Min",
        "game_script_mult": "Game Script Mult",
        "game_script_note": "Game Script Note",
        **HIT_TRACKING_RENAME,
    }
    clean = clean.rename(columns=rename)

    # Exclude voided rows from Tier sheets (after rename, void col is "Void Reason")
    void_col = "Void Reason" if "Void Reason" in clean.columns else None
    if void_col:
        clean_eligible = clean[clean[void_col].isna() | (clean[void_col] == "")].copy()
    else:
        clean_eligible = clean.copy()

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Tennis", clean)
    write_sheet(wb, "ALL", clean)
    for tier in ["A", "B", "C", "D"]:
        subset = clean_eligible[clean_eligible["Tier"] == tier].copy()
        # Always write every Tier sheet (even if empty) so step9 never crashes
        # on a missing sheet_name
        write_sheet(wb, f"Tier {tier}", subset if len(subset) else clean_eligible.head(0))

    wb.save(xlsx_path)
    print(f"Clean XLSX saved -> {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="outputs/step7_tennis_ranked.xlsx")
    ap.add_argument("--sheet",  default="ALL")
    ap.add_argument("--output", default="outputs/step8_tennis_direction.csv")
    ap.add_argument("--xlsx",   default="outputs/step8_tennis_direction_clean.xlsx")
    ap.add_argument("--date",   default="", help="YYYY-MM-DD target date (default: today ET)")
    args = ap.parse_args()

    print("[Tennis step8] Starting...")
    print(f"Loading: {args.input} (sheet={args.sheet})")
    df  = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")

    if df.empty:
        print("ERROR [Tennis-S8] Empty input from S7 — aborting.")
        sys.exit(1)

    # ── Date filter: keep only target date's games (America/New_York) ─────────
    eastern = zoneinfo.ZoneInfo("America/New_York")
    target_str = (args.date.strip()[:10] if args.date
                  else _dt.datetime.now(tz=eastern).date().isoformat())
    if "start_time" in df.columns:
        before_filter = len(df)
        et_dates = pd.to_datetime(df["start_time"], utc=True, errors="coerce").dt.tz_convert(eastern)
        df["_et_date"] = et_dates.dt.date.apply(
            lambda d: d.isoformat() if isinstance(d, _dt.date) else ""
        )
        mask = df["_et_date"] == target_str
        if not mask.any():
            valid = df["_et_date"].astype(str).str.match(r"^\d{4}-\d{2}-\d{2}$")
            avail = sorted(df.loc[valid, "_et_date"].unique().tolist())
            # Fallback: nearest ET date <= target (folder date), not future
            past_or_equal = [d for d in avail if d <= target_str]
            if past_or_equal:
                fallback_date = past_or_equal[-1]
                print(f"[DateFilter] No exact ET match for {target_str} — falling back to {fallback_date} ({(df['_et_date']==fallback_date).sum()} rows)")
                mask = df["_et_date"] == fallback_date
            elif avail:
                fallback_date = avail[0]
                print(f"[DateFilter] No past ET match — using earliest available {fallback_date} ({(df['_et_date']==fallback_date).sum()} rows)")
                mask = df["_et_date"] == fallback_date
            else:
                print(f"[DateFilter] No valid ET dates found — keeping all {before_filter} rows")
                mask = pd.Series(True, index=df.index)
        df = df.loc[mask].drop(columns="_et_date")
        dropped = before_filter - len(df)
        print(f"[DateFilter] Kept {len(df)}/{before_filter} rows for {target_str} ET (dropped {dropped} rows)")
    else:
        print("[DateFilter] WARNING: no start_time column — skipping date filter")

    out = df.copy()

    # Backfill opponent labels from game context when opp_team is missing.
    # Many tennis rows share pp_game_id; if each match has exactly two players,
    # infer each row's opponent as the "other" team in that game.
    if "opp_team" in out.columns and "pp_game_id" in out.columns and "team" in out.columns:
        opp_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
        if opp_blank.any():
            game_team_map = (
                out.loc[:, ["pp_game_id", "team"]]
                .astype(str)
                .assign(team=lambda x: x["team"].str.strip(), pp_game_id=lambda x: x["pp_game_id"].str.strip())
                .groupby("pp_game_id")["team"]
                .apply(lambda s: sorted({t for t in s.tolist() if t and t.lower() not in ("nan", "none", "null")}))
                .to_dict()
            )
            inferred = []
            for _, r in out.loc[opp_blank, ["pp_game_id", "team"]].iterrows():
                gid = str(r.get("pp_game_id", "")).strip()
                team = str(r.get("team", "")).strip()
                teams = game_team_map.get(gid, [])
                if len(teams) == 2 and team in teams:
                    inferred.append(teams[0] if teams[1] == team else teams[1])
                else:
                    inferred.append("")
            out.loc[opp_blank, "opp_team"] = inferred
            # Keep non-empty placeholder for unresolved rows so downstream views are explicit.
            still_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
            out.loc[still_blank, "opp_team"] = "UNKNOWN_OPP"

    reconcile_signed_edge_abs_dataframe(out)

    edge     = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = pd.to_numeric(out["abs_edge"], errors="coerce")

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)

    # BUG FIX: edge >= 0 is False for NaN, which wrongly assigns UNDER to rows
    # with no projection. Use explicit NaN check.
    has_edge  = edge.notna()
    model_dir = pd.Series(
        np.where(has_edge & (edge >= 0), "OVER",
                 np.where(has_edge & (edge < 0), "UNDER", "NO_EDGE")),
        index=out.index
    )
    out["model_dir"] = model_dir

    final_dir = model_dir.copy()
    reason    = pd.Series(
        np.where(has_edge & (abs_edge < 0.03), "MODEL_TIEBREAK_DIFF<0.03", ""),
        index=out.index
    )

    # Standard-only tie-break: when edge is effectively zero, use L5 side signal.
    # This preserves Goblin/Demon OVER-only behavior while allowing true Standard UNDERS.
    std_mask = pick_type.eq("Standard")
    tie_mask = std_mask & has_edge & (abs_edge < 0.03)
    if tie_mask.any():
        l5_over = pd.to_numeric(out.get("last5_over", np.nan), errors="coerce")
        l5_under = pd.to_numeric(out.get("last5_under", np.nan), errors="coerce")
        hr5 = pd.to_numeric(out.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")

        under_from_l5 = tie_mask & l5_under.notna() & l5_over.notna() & (l5_under > l5_over)
        over_from_l5 = tie_mask & l5_under.notna() & l5_over.notna() & (l5_over > l5_under)
        unresolved = tie_mask & ~(under_from_l5 | over_from_l5)
        under_from_hr = unresolved & hr5.notna() & (hr5 < 0.50)
        over_from_hr = unresolved & ~(under_from_hr)

        final_dir = final_dir.where(~under_from_l5, "UNDER")
        final_dir = final_dir.where(~over_from_l5, "OVER")
        final_dir = final_dir.where(~under_from_hr, "UNDER")
        final_dir = final_dir.where(~over_from_hr, "OVER")

        reason = reason.where(~under_from_l5, "STANDARD_TIEBREAK_L5_UNDER")
        reason = reason.where(~over_from_l5, "STANDARD_TIEBREAK_L5_OVER")
        reason = reason.where(~under_from_hr, "STANDARD_TIEBREAK_HR5_UNDER")
        reason = reason.where(~over_from_hr, "STANDARD_TIEBREAK_HR5_OVER")

    # Additional Standard-only force from Last 5 Avg vs current line:
    # If we actually have stat_last5_avg and it is clearly UNDER the PP line,
    # flip to UNDER even if the edge model is near-zero.
    # (Helps cases where passes-attempted last5 stats exist but edge still ~0.)
    std_force = std_mask.copy()
    last5_avg = pd.to_numeric(out.get("stat_last5_avg", np.nan), errors="coerce")
    linev = pd.to_numeric(out.get("line", np.nan), errors="coerce")
    last5_known = last5_avg.notna() & linev.notna()
    force_under = std_force & last5_known & (last5_avg < (linev - 1e-9))
    force_over  = std_force & last5_known & (last5_avg > (linev + 1e-9))
    if force_under.any():
        final_dir = final_dir.where(~force_under, "UNDER")
        reason = reason.where(~force_under, "STANDARD_L5AVG_UNDER")
    if force_over.any():
        final_dir = final_dir.where(~force_over, "OVER")
        reason = reason.where(~force_over, "STANDARD_L5AVG_OVER")

    forced    = pick_type.isin(["Goblin", "Demon"])
    final_dir = final_dir.where(~forced, "OVER")
    reason    = reason.where(~forced, "FORCED_OVER_ONLY_GOB_DEM")

    out["final_bet_direction"] = final_dir
    out["final_dir_reason"]    = reason
    out["direction"] = out["final_bet_direction"]
    out["bet_direction"] = out["final_bet_direction"]
    out["DEF_TIER"] = "N/A"

    repo_root = Path(__file__).resolve().parents[3]
    out = _attach_unified_ml_prob(out, repo_root)

    hr5b = pd.to_numeric(out.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")
    hr10b = pd.to_numeric(out.get("line_hit_rate_over_ou_10", np.nan), errors="coerce")
    hr10b = hr10b.fillna(hr5b)
    comp_hr = (0.5 * hr5b.fillna(0.5) + 0.5 * hr10b.fillna(0.5)).clip(0.0, 1.0)
    out["composite_hit_rate"] = comp_hr
    mpb = pd.to_numeric(out.get("ml_prob", np.nan), errors="coerce").fillna(0.5)
    out["blended_score"] = (0.3 * mpb + 0.7 * comp_hr).round(4)

    out = _attach_distribution_std(out)
    filled_std = int(pd.to_numeric(out["distribution_std"], errors="coerce").notna().sum())
    print(f"[Tennis step8] distribution_std filled {filled_std}/{len(out)} rows")

    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"Saved -> {args.output}")

    if out.empty:
        print("ERROR [Tennis-S8] Output is empty — aborting.")
        sys.exit(1)

    print("final_bet_direction counts:")
    print(pd.Series(out["final_bet_direction"]).value_counts().to_string())
    if "tier" in out.columns:
        print("tier counts:")
        print(out["tier"].value_counts().to_string())

    xlsx_path = args.xlsx if args.xlsx else "outputs/step8_tennis_direction_clean.xlsx"
    try:
        build_clean_xlsx(out, xlsx_path)
    except Exception as e:
        print(f"WARN build_clean_xlsx failed: {e}")
        print("   Writing raw fallback xlsx so combined pipeline can proceed...")
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
                out.to_excel(w, sheet_name="Tennis", index=False)
                out.to_excel(w, sheet_name="ALL", index=False)
                for _tier in ["A", "B", "C", "D"]:
                    _mask = out.get("tier", pd.Series(dtype=str)) == _tier
                    _void = out.get("void_reason", pd.Series("", index=out.index)).fillna("")
                    _elig = out[_mask & (_void == "")].copy() if _mask.any() else out.head(0)
                    _elig.to_excel(w, sheet_name=f"Tier {_tier}", index=False)
            print(f"Fallback xlsx saved -> {xlsx_path}")
        except Exception as e2:
            print(f"ERROR Fallback xlsx also failed: {e2}")

    # Dated copy: <repo>/outputs/{date}/step8_tennis_direction_clean_{date}.xlsx
    try:
        eastern = zoneinfo.ZoneInfo("America/New_York")
        slate_date = (
            str(args.date).strip()[:10]
            if args.date
            else _dt.datetime.now(tz=eastern).date().isoformat()
        )
        # Sports/Tennis/scripts -> monorepo root (see tennis_grader.py).
        repo_root = Path(__file__).resolve().parents[3]
        dated_dir = repo_root / "outputs" / slate_date
        dated_dir.mkdir(parents=True, exist_ok=True)
        dated_xlsx = dated_dir / f"step8_tennis_direction_clean_{slate_date}.xlsx"
        xp = Path(xlsx_path)
        if not xp.is_file():
            xp = repo_root / "Sports" / "Tennis" / str(xlsx_path).replace("\\", "/").lstrip("./")
        if xp.is_file():
            shutil.copy2(xp, dated_xlsx)
            print(f"[Tennis step8] Dated clean workbook -> {dated_xlsx}")
    except Exception as e:
        print(f"[Tennis step8] WARN dated copy skipped: {e}")


if __name__ == "__main__":
    main()
