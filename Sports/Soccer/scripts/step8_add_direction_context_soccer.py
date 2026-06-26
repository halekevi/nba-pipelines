#!/usr/bin/env python3
"""
step8_add_direction_context_soccer.py  (Soccer Pipeline)

Mirrors NBA step8_add_direction_context.py.
Reads step7_soccer_ranked.xlsx, appends direction context,
outputs full CSV + clean formatted XLSX.

Run:
  py -3.14 step8_add_direction_context_soccer.py \
    --input step7_soccer_ranked.xlsx \
    --output step8_soccer_direction.csv
"""

from __future__ import annotations

import argparse
import shutil
import sys
from pathlib import Path

import numpy as np
import pandas as pd

_REPO = Path(__file__).resolve().parent
for _ in range(10):
    if (_REPO / "utils" / "step8_edge_direction.py").is_file():
        if str(_REPO) not in sys.path:
            sys.path.insert(0, str(_REPO))
        break
    _REPO = _REPO.parent
else:
    raise RuntimeError("Could not locate repo root with utils/step8_edge_direction.py")

from scripts.l10_streak_utils import finalize_l10_ui_columns
from utils.hit_tracking_columns import HIT_TRACKING_RENAME, attach_hit_tracking_columns
from utils.step8_edge_direction import reconcile_signed_edge_abs_dataframe
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from datetime import date

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


def _copy_dated_step8_soccer(output_xlsx_path: str) -> None:
    src = Path(output_xlsx_path)
    if not src.is_file():
        return
    today = date.today().isoformat()
    repo_root = Path(__file__).resolve().parents[3]
    dated_dir = repo_root / "outputs" / today
    try:
        dated_dir.mkdir(parents=True, exist_ok=True)
        dated_path = dated_dir / f"step8_soccer_direction_clean_{today}.xlsx"
        shutil.copy2(src, dated_path)
        print(f"[Soccer step8] Dated copy -> {dated_path}")
    except Exception as e:
        print(f"[Soccer step8] WARN: dated copy failed: {e}")


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


def _warn_step7_goblin_ml_saturation(step7_df: pd.DataFrame) -> None:
    """Non-fatal guard: catch prop-calibrator saturation regressions on Goblin rows."""
    pt = step7_df.get("pick_type", pd.Series("", index=step7_df.index)).astype(str).str.lower()
    if not pt.str.contains("gob", na=False).any():
        return
    ml = pd.to_numeric(step7_df.get("ml_prob", pd.Series(dtype=float)), errors="coerce")
    gob_ml = ml[pt.str.contains("gob", na=False)]
    if gob_ml.notna().any() and float(gob_ml.mean()) > 0.85:
        print(
            "[WARN] Soccer step7 Goblin ml_prob mean > 0.85 — "
            "calibrator saturation may have recurred, check fix"
        )
    missing_7b = not (
        "edge_score" in step7_df.columns and "blended_score" in step7_df.columns
    )
    if missing_7b:
        print(
            "[WARN] Soccer step7 lacks edge_score/blended_score — "
            "step7b may not have run; ml_prob may be stale prop-model output only"
        )


TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
HEADER_COLOR = "1C1C1C"


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def _line_shift_fill(val) -> str:
    if val is None or val == "" or (isinstance(val, float) and pd.isna(val)):
        return "FFFFFF"
    try:
        num = float(val)
        if num > 0:
            return "C8F7C5"
        if num < 0:
            return "F7C5C5"
        return "FFFFFF"
    except (TypeError, ValueError):
        pass
    s = str(val).strip().upper()
    if "UP" in s or s == "MOVED_UP":
        return "C8F7C5"
    if "DOWN" in s or s == "MOVED_DOWN":
        return "F7C5C5"
    return "FFFFFF"


def write_sheet(wb, name: str, data: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    tier_bg, tier_fg = TIER_COLORS.get(name, ("333333", "FFFFFF"))
    headers = list(data.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        direction = row[headers.index("Direction")] if "Direction" in headers else ""
        for ci, val in enumerate(row, 1):
            col_name    = headers[ci - 1]
            display_val = "" if pd.isna(val) else val
            cell        = ws.cell(row=ri, column=ci, value=display_val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

            if col_name == "Tier":
                cell.fill = PatternFill("solid", start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name="Arial", size=9)
            elif col_name == "Direction":
                bg = "C8F7C5" if val == "OVER" else "F7C5C5" if val == "UNDER" else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, name="Arial", size=9)
            elif col_name == "League":
                cell.fill = PatternFill("solid", start_color="EBF5FB")
            elif col_name == "Line Shift":
                cell.fill = PatternFill("solid", start_color=_line_shift_fill(display_val))
            else:
                cell.fill = PatternFill("solid", start_color="F9F9F9" if ri % 2 == 0 else "FFFFFF")

    col_widths = {
        "Tier": 6, "Rank Score": 10, "Player": 20, "Pos": 6, "Pos Group": 9,
        "Team": 12, "Opp": 12, "Days Rest": 9, "B2B": 6, "Opp Rest": 9, "Opp B2B": 8, "Game Total": 10, "Spread": 8, "League": 12, "Game Time": 16,
        "ESPN ID": 10,
        "Prop": 18, "Pick Type": 10, "Line": 7,
        "Direction": 9, "Edge": 7, "Projection": 10,
        "ML Prob": 9, "Edge Score": 10, "Blended Score": 12,
        "Hit Rate (5g)": 12, "Hit Rate (10g)": 12, "Last 5 Avg": 10, "Season Avg": 10,
        "L5 Over": 8, "L5 Under": 8,
        "Def Rank": 9, "Def Tier": 10,
        "Def Boost Hist": 12, "Team Top3 Rank": 10, "Top3 Weak Over": 18, "Top3 Elite Fade": 14,
        "Min Tier": 9, "Starter Tier": 11, "Shot Role": 10, "Usage Role": 10,
        "Void Reason": 20,
        "Open Line": 8,
        "Line Movement": 12,
        "Line Shift": 10,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


_MIN_TIER_NUM_MAP = {0: "Low", 1: "Med", 2: "High", 3: "Elite"}


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    # Convert numeric minutes_tier (0-3) back to human labels
    if "minutes_tier" in df2.columns:
        _mt_num = pd.to_numeric(df2["minutes_tier"], errors="coerce")
        _mt_valid = _mt_num.notna()
        if _mt_valid.any():
            df2.loc[_mt_valid, "minutes_tier"] = _mt_num[_mt_valid].round().astype(int).map(_MIN_TIER_NUM_MAP).fillna(df2.loc[_mt_valid, "minutes_tier"])
    try:
        import platform
        _time_fmt = "%m/%d %#I:%M %p" if platform.system() == "Windows" else "%m/%d %-I:%M %p"
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime(_time_fmt)
    except Exception:
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%m/%d %H:%M")

    if "line" in df2.columns:
        df2 = finalize_l10_ui_columns(df2, line_col="line")
    df2 = attach_hit_tracking_columns(df2, "SOCCER")

    keep = [
        "tier", "rank_score",
        "player", "pos", "position_group", "team", "opp_team", "days_rest", "is_back_to_back", "opp_days_rest", "opp_b2b",
        "h2h_avg", "h2h_over_pct", "h2h_games", "h2h_last",
        "game_total", "spread", "league", "game_time",
        "espn_player_id",
        "prop_type", "pick_type", "line", "standard_line", "standard_line_source",
        "final_bet_direction",
        "edge", "abs_edge", "projection",
        "ml_prob",
        "hit_rate", "hit_rate_l5", "hit_rate_l10",
        "strat_hit_rate", "strat_n",
        "player_hr_historical", "opp_hr_historical",
        "sport_signal_maturity", "confidence_tier", "confidence_score", "confidence_note",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "l10_over", "l10_under", "l10_over_pct", "l10_streak", "l10_games_played",
        "OVERALL_DEF_RANK", "DEF_TIER", "def_tier",
        "def_boost_hist", "team_top3_rank", "top3_weak_overperformer", "top3_elite_fader",
        "deviation_level", "opp_pace",
        "minutes_tier", "starter_tier", "shot_role", "usage_role",
        "void_reason",
        # ── Game log ─────────────────────────────────────────────────────────
        "stat_g1", "stat_g2", "stat_g3", "stat_g4", "stat_g5",
        "stat_g6", "stat_g7", "stat_g8", "stat_g9", "stat_g10",
        "stat_last10_avg",
        # ── Schedule / context ───────────────────────────────────────────────
        "avg_minutes",
        "game_script_mult",
        "game_script_note",
        "open_line",
        "line_movement",
        "line_direction_shift",
    ]
    keep  = [c for c in keep if c in df2.columns]
    clean = df2[keep].copy()

    for col in [
        "rank_score",
        "edge",
        "abs_edge",
        "projection",
        "ml_prob",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
        "open_line",
        "line_movement",
    ]:
        if col in clean.columns:
            rnd = 4 if col in ("ml_prob", "edge_score", "blended_score") else (3 if col == "line_movement" else 2)
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(rnd)
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    if "def_boost_hist" in clean.columns:
        clean["def_boost_hist"] = pd.to_numeric(clean["def_boost_hist"], errors="coerce").round(3)
    if "team_top3_rank" in clean.columns:
        clean["team_top3_rank"] = pd.to_numeric(clean["team_top3_rank"], errors="coerce").round(0)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")
    if "standard_line" in clean.columns:
        clean["standard_line"] = pd.to_numeric(clean["standard_line"], errors="coerce").round(2)

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}

    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "player": "Player", "pos": "Pos", "position_group": "Pos Group",
        "team": "Team", "opp_team": "Opp", "league": "League", "game_time": "Game Time",
        "days_rest": "Days Rest",
        "is_back_to_back": "B2B",
        "opp_days_rest": "Opp Rest",
        "opp_b2b": "Opp B2B",
        "h2h_avg": "H2H Avg",
        "h2h_over_pct": "H2H Over%",
        "h2h_games": "H2H Games",
        "h2h_last": "H2H Last",
        "game_total": "Game Total",
        "spread": "Spread",
        "espn_player_id": "ESPN ID",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "standard_line": "Standard Line", "standard_line_source": "Standard Line Source",
        "final_bet_direction": "Direction",
        "edge": "Edge", "abs_edge": "Abs Edge", "projection": "Projection",
        "ml_prob": "ML Prob",
        "edge_score": "Edge Score",
        "blended_score": "Blended Score",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "line_hit_rate_over_ou_10": "Hit Rate (10g)",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier", "def_tier": "Def Tier",
        "def_boost_hist": "Def Boost Hist",
        "team_top3_rank": "Team Top3 Rank",
        "top3_weak_overperformer": "Top3 Weak Over",
        "top3_elite_fader": "Top3 Elite Fade",
        "deviation_level": "Deviation Level", "opp_pace": "Opp Pace",
        "minutes_tier": "Min Tier", "starter_tier": "Starter Tier", "shot_role": "Shot Role", "usage_role": "Usage Role",
        "void_reason": "Void Reason",
        # Game log
        "stat_last10_avg": "Last 10 Avg",
        "stat_g1": "G1", "stat_g2": "G2", "stat_g3": "G3",
        "stat_g4": "G4", "stat_g5": "G5", "stat_g6": "G6",
        "stat_g7": "G7", "stat_g8": "G8", "stat_g9": "G9", "stat_g10": "G10",
        # Context
        "avg_minutes": "Avg Min",
        "game_script_mult": "Game Script Mult",
        "game_script_note": "Game Script Note",
        "open_line": "Open Line",
        "line_movement": "Line Movement",
        "line_direction_shift": "Line Shift",
        **HIT_TRACKING_RENAME,
    }
    clean = clean.rename(columns=rename)

    # Exclude voided rows from Tier sheets (after rename, void col is "Void Reason")
    void_col = "Void Reason" if "Void Reason" in clean.columns else None
    if void_col:
        clean_eligible = clean[clean[void_col].isna() | (clean[void_col] == "")].copy()
    else:
        clean_eligible = clean.copy()

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "ALL", clean)
    for tier in ["A", "B", "C", "D"]:
        subset = clean_eligible[clean_eligible["Tier"] == tier].copy()
        # Always write every Tier sheet (even if empty) so step9 never crashes
        # on a missing sheet_name
        write_sheet(wb, f"Tier {tier}", subset if len(subset) else clean_eligible.head(0))

    wb.save(xlsx_path)
    print(f"Clean XLSX saved -> {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="s7_soccer_ranked.xlsx")
    ap.add_argument("--sheet",  default="ALL")
    ap.add_argument("--output", default="step8_soccer_direction.csv")
    ap.add_argument("--xlsx",   default="step8_soccer_direction_clean.xlsx")
    ap.add_argument("--date",   default="", help="YYYY-MM-DD target date (default: today ET)")
    args = ap.parse_args()

    print(f"Loading: {args.input} (sheet={args.sheet})")
    df  = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")

    if df.empty:
        print("ERROR [PropOracle-Soccer-S8] Empty input from S7 — aborting.")
        sys.exit(1)

    _warn_step7_goblin_ml_saturation(df)

    # ── Date filter: keep only target date's games ───────────────────────────
    import datetime, zoneinfo
    eastern = zoneinfo.ZoneInfo("America/New_York")
    target_str = (args.date.strip()[:10] if args.date
                  else datetime.datetime.now(tz=eastern).date().isoformat())
    if "start_time" in df.columns:
        before_filter = len(df)
        def _to_et_date(val):
            try:
                dt = pd.to_datetime(val)          # handles -04:00 offset natively
                if dt.tzinfo is None:
                    dt = dt.tz_localize("UTC")
                return dt.tz_convert(eastern).date().isoformat()
            except Exception:
                return ""
        df["_et_date"] = df["start_time"].apply(_to_et_date)
        mask = df["_et_date"] == target_str
        if not mask.any():
            ed = df["_et_date"].astype(str)
            ok = ed.str.len() >= 10
            future = ed[ok & (ed >= target_str)]
            if future.empty:
                fallback = ed[ok]
                pick = str(fallback.mode().iloc[0]) if len(fallback) else target_str
            else:
                pick = str(future.min())
            if pick != target_str:
                print(
                    f"[DateFilter] Kept 0/{before_filter} for {target_str} ET; "
                    f"using nearest slate date {pick} (>= target)"
                )
            mask = df["_et_date"] == pick
        df = df.loc[mask].drop(columns="_et_date")
        dropped = before_filter - len(df)
        print(f"[DateFilter] Kept {len(df)}/{before_filter} rows for slate ET (dropped {dropped} rows)")
    else:
        print("[DateFilter] WARNING: no start_time column — skipping date filter")

    out = df.copy()

    # Backfill opponent from pp_game_id pairing (atomic teams — not raw combo strings).
    if "opp_team" in out.columns:
        from soccer_opp_utils import fill_opp_team_column, sanitize_unknown_opp

        out = fill_opp_team_column(out)
        out["opp_team"] = sanitize_unknown_opp(out["opp_team"])

    reconcile_signed_edge_abs_dataframe(out)

    edge     = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = pd.to_numeric(out["abs_edge"], errors="coerce")

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    step7_dir = out.get("bet_direction", pd.Series("", index=out.index)).astype(str).str.upper().str.strip()

    # BUG FIX: edge >= 0 is False for NaN, which wrongly assigns UNDER to rows
    # with no projection. Use explicit NaN check.
    has_edge  = edge.notna()
    model_dir = pd.Series(
        np.where(has_edge & (edge >= 0), "OVER",
                 np.where(has_edge & (edge < 0), "UNDER", "NO_EDGE")),
        index=out.index
    )
    out["model_dir"] = model_dir

    final_dir = model_dir.copy()
    reason    = pd.Series(
        np.where(has_edge & (abs_edge < 0.03), "MODEL_TIEBREAK_DIFF<0.03", ""),
        index=out.index
    )
    # Preserve step7 direction for Standard rows when it matches reconciled projection − line.
    # Option B: if bet_direction contradicts signed edge, ignore pass-through (model_dir wins).
    # If projection and line both parse but edge is NaN, reconciliation failed to fill — do not
    # silently fall back to Step 7 (distinct audit reason).
    proj_num = (
        pd.to_numeric(out["projection"], errors="coerce")
        if "projection" in out.columns
        else pd.Series(np.nan, index=out.index)
    )
    line_num = (
        pd.to_numeric(out["line"], errors="coerce")
        if "line" in out.columns
        else pd.Series(np.nan, index=out.index)
    )
    has_parseable_proj_line = proj_num.notna() & line_num.notna()

    step7_pick = pick_type.eq("Standard") & step7_dir.isin(["OVER", "UNDER"])
    under_aligns = step7_dir.eq("UNDER") & (edge < 0)
    over_aligns = step7_dir.eq("OVER") & (edge >= 0)
    step7_matches_reconciled = under_aligns | over_aligns
    edge_missing = edge.isna()

    blocked_step7_conflict = step7_pick & (~edge_missing) & (~step7_matches_reconciled)
    blocked_recon_nan_gap = step7_pick & edge_missing & has_parseable_proj_line

    reason = reason.where(~blocked_step7_conflict, "STANDARD_STEP7_BLOCKED_RECON_MISMATCH")
    reason = reason.where(~blocked_recon_nan_gap, "STANDARD_STEP7_BLOCKED_EDGE_NA_FINITE_PROJ_LINE")

    std_from_step7 = step7_pick & (
        (edge_missing & ~has_parseable_proj_line)
        | ((~edge_missing) & step7_matches_reconciled)
    )
    final_dir = final_dir.where(~std_from_step7, step7_dir)
    reason = reason.where(~std_from_step7, "STANDARD_PASS_THROUGH_STEP7")

    # Standard-only tie-break: when edge is effectively zero, use L5 side signal.
    std_mask = pick_type.eq("Standard") & ~std_from_step7
    tie_mask = std_mask & has_edge & (abs_edge < 0.03)
    if tie_mask.any():
        l5_over = pd.to_numeric(out.get("last5_over", np.nan), errors="coerce")
        l5_under = pd.to_numeric(out.get("last5_under", np.nan), errors="coerce")
        hr5 = pd.to_numeric(out.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")

        under_from_l5 = tie_mask & l5_under.notna() & l5_over.notna() & (l5_under > l5_over)
        over_from_l5 = tie_mask & l5_under.notna() & l5_over.notna() & (l5_over > l5_under)
        unresolved = tie_mask & ~(under_from_l5 | over_from_l5)
        under_from_hr = unresolved & hr5.notna() & (hr5 < 0.50)
        over_from_hr = unresolved & ~(under_from_hr)

        final_dir = final_dir.where(~under_from_l5, "UNDER")
        final_dir = final_dir.where(~over_from_l5, "OVER")
        final_dir = final_dir.where(~under_from_hr, "UNDER")
        final_dir = final_dir.where(~over_from_hr, "OVER")

        reason = reason.where(~under_from_l5, "STANDARD_TIEBREAK_L5_UNDER")
        reason = reason.where(~over_from_l5, "STANDARD_TIEBREAK_L5_OVER")
        reason = reason.where(~under_from_hr, "STANDARD_TIEBREAK_HR5_UNDER")
        reason = reason.where(~over_from_hr, "STANDARD_TIEBREAK_HR5_OVER")

    # Additional Standard-only force from Last 5 Avg vs current line:
    # If we actually have stat_last5_avg and it is clearly UNDER the PP line,
    # flip to UNDER even if the edge model is near-zero.
    # (Helps cases where passes-attempted last5 stats exist but edge still ~0.)
    std_force = std_mask.copy()
    last5_avg = pd.to_numeric(out.get("stat_last5_avg", np.nan), errors="coerce")
    linev = pd.to_numeric(out.get("line", np.nan), errors="coerce")
    last5_known = last5_avg.notna() & linev.notna()
    force_under = std_force & last5_known & (last5_avg < (linev - 1e-9))
    force_over  = std_force & last5_known & (last5_avg > (linev + 1e-9))
    if force_under.any():
        final_dir = final_dir.where(~force_under, "UNDER")
        reason = reason.where(~force_under, "STANDARD_L5AVG_UNDER")
    if force_over.any():
        final_dir = final_dir.where(~force_over, "OVER")
        reason = reason.where(~force_over, "STANDARD_L5AVG_OVER")

    forced    = pick_type.isin(["Goblin", "Demon"])
    final_dir = final_dir.where(~forced, "OVER")
    reason    = reason.where(~forced, "FORCED_OVER_ONLY_GOB_DEM")

    out["final_bet_direction"] = final_dir
    out["final_dir_reason"]    = reason

    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"Saved -> {args.output}")

    if out.empty:
        print("ERROR [PropOracle-Soccer-S8] Output is empty — aborting.")
        sys.exit(1)

    print("final_bet_direction counts:")
    print(pd.Series(out["final_bet_direction"]).value_counts().to_string())
    if "tier" in out.columns:
        print("tier counts:")
        print(out["tier"].value_counts().to_string())

    # Always use explicit --xlsx path (default: step8_soccer_direction_clean.xlsx)
    xlsx_path = args.xlsx if args.xlsx else "step8_soccer_direction_clean.xlsx"
    try:
        build_clean_xlsx(out, xlsx_path)
    except Exception as e:
        print(f"WARN build_clean_xlsx failed: {e}")
        print("   Writing raw fallback xlsx so combined pipeline can proceed...")
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
                out.to_excel(w, sheet_name="ALL", index=False)
                for _tier in ["A", "B", "C", "D"]:
                    _mask = out.get("tier", pd.Series(dtype=str)) == _tier
                    _void = out.get("void_reason", pd.Series("", index=out.index)).fillna("")
                    _elig = out[_mask & (_void == "")].copy() if _mask.any() else out.head(0)
                    _elig.to_excel(w, sheet_name=f"Tier {_tier}", index=False)
            print(f"Fallback xlsx saved -> {xlsx_path}")
        except Exception as e2:
            print(f"ERROR Fallback xlsx also failed: {e2}")

    _copy_dated_step8_soccer(xlsx_path)


if __name__ == "__main__":
    main()
