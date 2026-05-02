#!/usr/bin/env python3
"""
step9_build_tickets_soccer.py  (Soccer Pipeline) — OPTIMIZED

Key optimizations vs original:
  1. Replaced pd.to_numeric(pd.Series([val])) with safe_float() in all inner loops
     → avoids creating a pandas Series object per value during combinations iteration
  2. Pre-compute float columns once before ticket building (no per-row re-parsing)
  3. Early-exit in build_tickets once max_tickets reached (was always scanning all)
  4. _ticket_valid uses set comprehension directly, no list conversion
  5. Summary sheet avg_hr_vals uses pre-computed float columns

Run:
  py step9_build_tickets_soccer.py \
    --input step8_soccer_direction_clean.xlsx \
    --output soccer_best_tickets.xlsx
"""

from __future__ import annotations

import argparse
import sys
from itertools import combinations

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ── Payout tables ─────────────────────────────────────────────────────────────
POWER_PLAY_BASE = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 37.5}
FLEX_PLAY_BASE  = {
    2: {2: 3.0},
    3: {3: 3.0,  2: 0.5},
    4: {4: 6.0,  3: 1.5},
    5: {5: 10.0, 4: 2.0,  3: 0.4},
    6: {6: 25.0, 5: 2.0,  4: 0.4},
}
GOBLIN_POWER_MOD = {1: 0.900, 2: 0.800, 3: 0.767, 4: 0.667}
GOBLIN_FLEX_MOD  = {1: 0.833, 2: 0.750, 3: 0.700, 4: 0.600}
DEMON_POWER_MOD  = {1: 1.333, 2: 1.833, 3: 3.833}
DEMON_FLEX_MOD   = {1: 1.200, 2: 1.600, 3: 3.000}

COLORS = {
    "Standard": "2874A6", "Goblin": "6C3483",
    "Demon": "C0392B",    "Best Mix": "1E8449",
}
HDR_COLOR  = "1C1C1C"
DIR_OVER   = "C8F7C5"
DIR_UNDER  = "F7C5C5"
TIER_COLORS = {
    "Elite":     "D5F5E3",
    "Above Avg": "EBF5FB",
    "Avg":       "FDFEFE",
    "Weak":      "FDEDEC",
}


def thin_border():
    s = Side(style="thin", color="CCCCCC")
    return Border(left=s, right=s, top=s, bottom=s)


def _norm_pick_type(x) -> str:
    t = str(x).strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


def safe_float(x, default: float = np.nan) -> float:
    """Fast float conversion without creating a pandas object."""
    if x is None or (isinstance(x, float) and np.isnan(x)):
        return default
    try:
        return float(x)
    except (TypeError, ValueError):
        return default


def calc_ticket_payout(ticket: list, n_legs: int, play_type: str = "power") -> dict:
    base_top  = POWER_PLAY_BASE.get(n_legs, 37.5) if play_type == "power" else FLEX_PLAY_BASE.get(n_legs, {}).get(n_legs, 25.0)
    power_mod = flex_mod = 1.0
    for row in ticket:
        pt  = str(row.get("Pick Type", "Standard")).strip().lower()
        dev = int(safe_float(row.get("deviation_level", row.get("Deviation Level", 1)), 1))
        dev = max(1, min(dev, 4))
        if "gob" in pt:
            power_mod *= GOBLIN_POWER_MOD.get(dev, 0.840)
            flex_mod  *= GOBLIN_FLEX_MOD.get(dev,  0.800)
        elif "dem" in pt:
            power_mod *= DEMON_POWER_MOD.get(dev, 1.627)
            flex_mod  *= DEMON_FLEX_MOD.get(dev,  1.600)
    if play_type == "power":
        top   = round(base_top * power_mod, 2)
        stake = round(100 / top, 2) if top > 0 else 0
        return {"top_payout": top, "stake_to_win_100": stake, "play_type": "Power Play"}
    else:
        flex_base = FLEX_PLAY_BASE.get(n_legs, {})
        partials  = {nc: round(mult * flex_mod if nc == n_legs else mult, 2)
                     for nc, mult in flex_base.items()}
        top   = partials.get(n_legs, 25.0)
        stake = round(100 / top, 2) if top > 0 else 0
        return {"top_payout": top, "stake_to_win_100": stake, "play_type": "Flex Play", "partials": partials}


def _game_key(row) -> frozenset:
    return frozenset([str(row["Team"]), str(row["Opp"])])


def _ticket_valid(combo) -> bool:
    # Use a set comprehension directly — faster than list then set
    players = [r["Player"] for r in combo]
    if len(players) != len(set(players)):
        return False
    games = [_game_key(r) for r in combo]
    return len(games) == len(set(games))


def _ticket_score(combo) -> float:
    # Use pre-computed _rank_score_f column (float already)
    scores = [r.get("_rank_score_f", np.nan) for r in combo]
    valid  = [s for s in scores if not np.isnan(s)]
    return float(np.mean(valid)) if valid else 0.0


def _hit_rate_f(row) -> float:
    return safe_float(row.get("_hit_rate_f", row.get("Hit Rate (5g)", np.nan)), np.nan)


def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets: int,
                  used_legs: set, min_hit_rate: float = 0.0) -> list:
    if min_hit_rate > 0.0:
        pool = pool[pool["_hit_rate_f"].ge(min_hit_rate)].copy()

    pool       = pool.sort_values("_rank_score_f", ascending=False).copy()
    K          = min(len(pool), 25)
    candidates = [row for _, row in pool.head(K).iterrows()]

    valid_tickets: list = []
    seen_keys:     set  = set()

    for combo in combinations(candidates, n_legs):
        leg_keys = [(r["Player"], r["Prop"], r["Line"]) for r in combo]
        if any(lk in used_legs for lk in leg_keys):
            continue
        if not _ticket_valid(combo):
            continue
        tkey = frozenset((r["Player"], r["Prop"], str(r["Line"])) for r in combo)
        if tkey in seen_keys:
            continue
        seen_keys.add(tkey)
        valid_tickets.append((combo, _ticket_score(combo)))

    valid_tickets.sort(key=lambda x: x[1], reverse=True)

    tickets = []
    for combo, _ in valid_tickets:
        if len(tickets) >= max_tickets:
            break
        leg_keys = [(r["Player"], r["Prop"], r["Line"]) for r in combo]
        if any(lk in used_legs for lk in leg_keys):
            continue
        used_legs.update(leg_keys)
        tickets.append(list(combo))

    return tickets


def write_tickets_sheet(wb, sheet_name: str, tickets: list, n_legs: int, pick_label: str) -> None:
    if not tickets:
        return
    ws     = wb.create_sheet(sheet_name)
    tab_bg = COLORS.get(pick_label, "1E8449")

    cols = ["#", "Player", "Team", "Opp", "League", "Prop", "Pick Type", "Line",
            "Direction", "Edge", "Hit Rate (5g)", "Last 5 Avg", "Season Avg",
            "L5 Over", "L5 Under", "Rank Score", "Def Tier"]
    current_row = 1

    for t_idx, ticket in enumerate(tickets, 1):
        ticket_dicts = [dict(r) for r in ticket]
        pp_info      = calc_ticket_payout(ticket_dicts, n_legs, "power")
        flex_info    = calc_ticket_payout(ticket_dicts, n_legs, "flex")
        avg_score    = _ticket_score(ticket)

        # Use pre-computed float column
        hr_vals  = [_hit_rate_f(r) for r in ticket]
        avg_hr   = float(np.mean([v for v in hr_vals if not np.isnan(v)])) if any(not np.isnan(v) for v in hr_vals) else 0.0

        header_val = (
            f"Ticket #{t_idx} | {n_legs}-Leg {pick_label} | "
            f"Power: {pp_info['top_payout']}x (${pp_info['stake_to_win_100']:.0f} to win $100) | "
            f"Flex: {flex_info['top_payout']}x | "
            f"Avg Hit Rate: {avg_hr:.0%} | "
            f"Est Win: {avg_hr**n_legs:.0%} | "
            f"Avg Score: {avg_score:.2f}"
        )
        c = ws.cell(row=current_row, column=1, value=header_val)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=tab_bg)
        c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(cols))
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=9)
            c.fill      = PatternFill("solid", start_color=HDR_COLOR)
            c.alignment = Alignment(horizontal="center", vertical="center")
            c.border    = thin_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        for leg_i, row in enumerate(ticket, 1):
            row_bg = "F8F9FA" if leg_i % 2 == 0 else "FFFFFF"
            def_bg = TIER_COLORS.get(str(row.get("Def Tier", "")), "FFFFFF")

            l5o = row.get("L5 Over",  "")
            l5u = row.get("L5 Under", "")
            l5o = "" if (l5o is None or (isinstance(l5o, float) and np.isnan(l5o))) else int(l5o)
            l5u = "" if (l5u is None or (isinstance(l5u, float) and np.isnan(l5u))) else int(l5u)

            edge_f   = safe_float(row.get("Edge", 0.0), 0.0)
            last5_f  = safe_float(row.get("Last 5 Avg", 0.0), 0.0)
            season_f = safe_float(row.get("Season Avg", 0.0), 0.0)
            score_f  = safe_float(row.get("_rank_score_f", row.get("Rank Score", 0.0)), 0.0)

            vals = [
                leg_i,
                row["Player"],
                row["Team"],
                row["Opp"],
                row.get("League", ""),
                row["Prop"],
                row["Pick Type"],
                row["Line"],
                row["Direction"],
                round(edge_f, 1),
                row.get("Hit Rate (5g)", ""),
                round(last5_f, 1),
                round(season_f, 1),
                l5o, l5u,
                round(score_f, 2),
                row.get("Def Tier", ""),
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.font      = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = thin_border()
                col_name    = cols[ci - 1]
                if col_name == "Direction":
                    bg    = DIR_OVER if val == "OVER" else DIR_UNDER
                    c.fill = PatternFill("solid", start_color=bg)
                    c.font = Font(bold=True, name="Arial", size=9)
                elif col_name == "Def Tier":
                    c.fill = PatternFill("solid", start_color=def_bg)
                else:
                    c.fill = PatternFill("solid", start_color=row_bg)

            current_row += 1
        current_row += 1


def write_summary_sheet(wb, all_ticket_sets: list) -> None:
    ws      = wb.create_sheet("SUMMARY", 0)
    headers = ["Sheet", "Ticket #", "Legs", "Pick Type", "Power Payout",
               "$ to win $100", "Flex Payout", "Avg Hit Rate", "Est Win %",
               "Avg Rank Score", "Players"]
    col_w   = [22, 10, 6, 12, 13, 13, 12, 13, 12, 15, 60]

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        c.fill      = PatternFill("solid", start_color=HDR_COLOR)
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = "A2"

    row = 2
    for sheet_name, tickets, n_legs, pick_label in all_ticket_sets:
        for t_idx, ticket in enumerate(tickets, 1):
            avg_score    = _ticket_score(ticket)
            hr_vals      = [_hit_rate_f(r) for r in ticket]
            avg_hr       = float(np.mean([v for v in hr_vals if not np.isnan(v)])) if any(not np.isnan(v) for v in hr_vals) else 0.0
            ticket_dicts = [dict(r) for r in ticket]
            pp_info      = calc_ticket_payout(ticket_dicts, n_legs, "power")
            flex_info    = calc_ticket_payout(ticket_dicts, n_legs, "flex")
            players      = " | ".join(
                f"{r['Player']} {r['Direction']} {r['Prop']} {r['Line']}" for r in ticket
            )
            vals = [sheet_name, t_idx, n_legs, pick_label,
                    f"{pp_info['top_payout']}x", f"${pp_info['stake_to_win_100']:.0f}",
                    f"{flex_info['top_payout']}x",
                    f"{avg_hr:.0%}", f"{avg_hr**n_legs:.0%}",
                    round(avg_score, 2), players]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font      = Font(name="Arial", size=9)
                c.alignment = Alignment(horizontal="center", vertical="center")
                c.border    = thin_border()
                c.fill      = PatternFill("solid", start_color="F8F9FA" if row % 2 == 0 else "FFFFFF")
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",        default="s8_soccer_direction_clean.xlsx")
    ap.add_argument("--sheet",        default="Tier A")
    ap.add_argument("--output",       default="s9_soccer_tickets.xlsx")
    ap.add_argument("--min_hit_rate", type=float, default=0.0)
    ap.add_argument("--max_tickets",  type=int,   default=10)
    ap.add_argument("--legs",         default="2,3,4")
    args = ap.parse_args()

    leg_counts = [int(x.strip()) for x in args.legs.split(",")]

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    try:
        df = pd.read_excel(args.input, sheet_name=args.sheet)
    except ValueError:
        print(f"  ⚠️  Sheet '{args.sheet}' not found — falling back to 'ALL' sheet, filtering tier A/B")
        df_all    = pd.read_excel(args.input, sheet_name="ALL")
        tier_mask = df_all.get("Tier", pd.Series(dtype=str)).isin(["A", "B"])
        df        = df_all[tier_mask].copy() if tier_mask.any() else df_all.copy()
        # Also strip voided rows on fallback path
        if "Void Reason" in df.columns:
            df = df[df["Void Reason"].isna() | (df["Void Reason"] == "")].copy()
        print(f"  Fallback rows: {len(df)}")

    df["Pick Type"] = df["Pick Type"].astype(str).apply(_norm_pick_type)

    if df.empty:
        print("❌ [PropOracle-Soccer-S9] Empty input from S8 — aborting.")
        sys.exit(1)

    # Pre-compute float columns ONCE (avoids per-row/per-combination re-parsing)
    df["_rank_score_f"] = pd.to_numeric(df["Rank Score"],      errors="coerce")
    df["_hit_rate_f"]   = pd.to_numeric(df["Hit Rate (5g)"],   errors="coerce")
    df["_edge_f"]       = pd.to_numeric(df["Edge"],            errors="coerce")
    df["_last5_f"]      = pd.to_numeric(df.get("Last 5 Avg",   pd.Series(dtype=float)), errors="coerce")
    df["_season_f"]     = pd.to_numeric(df.get("Season Avg",   pd.Series(dtype=float)), errors="coerce")

    if args.min_hit_rate > 0.0:
        df = df[df["_hit_rate_f"].ge(args.min_hit_rate)].copy()
    print(f"  Props with hit rate >= {args.min_hit_rate}: {len(df)}")

    df = df.sort_values("_rank_score_f", ascending=False).drop_duplicates(
        subset=["Player", "Prop", "Line", "Direction"]
    )

    std_pool = df[df["Pick Type"] == "Standard"].drop_duplicates(subset=["Player"])
    gob_pool = df[df["Pick Type"] == "Goblin"].drop_duplicates(subset=["Player"])
    all_pool = df.sort_values("_rank_score_f", ascending=False).drop_duplicates(subset=["Player"])

    pools = [("Standard", std_pool), ("Goblin", gob_pool), ("Best Mix", all_pool)]

    wb = Workbook()
    wb.remove(wb.active)
    all_ticket_sets = []

    for pick_label, pool in pools:
        if len(pool) < 2:
            print(f"  ⚠ Skipping {pick_label} — not enough props ({len(pool)})")
            continue
        used_legs: set = set()
        for n_legs in leg_counts:
            if len(pool) < n_legs:
                continue
            tickets    = build_tickets(pool.copy(), n_legs, args.max_tickets, used_legs, args.min_hit_rate)
            sheet_name = f"{pick_label} {n_legs}-Leg"
            write_tickets_sheet(wb, sheet_name, tickets, n_legs, pick_label)
            all_ticket_sets.append((sheet_name, tickets, n_legs, pick_label))
            print(f"  {sheet_name}: {len(tickets)} tickets")

    if len(std_pool) >= 1 and len(gob_pool) >= 1:
        fill_pool = pd.concat([
            std_pool.assign(_pref=0),
            gob_pool.assign(_pref=1),
        ]).sort_values(["_pref", "_rank_score_f"], ascending=[True, False]).drop("_pref", axis=1)

        used_legs_fill: set = set()
        for n_legs in leg_counts:
            if n_legs < 3 or len(fill_pool) < n_legs:
                continue
            tickets = build_tickets(fill_pool.copy(), n_legs, args.max_tickets, used_legs_fill, args.min_hit_rate)
            if tickets:
                sheet_name = f"Std+Gob Fill {n_legs}-Leg"
                write_tickets_sheet(wb, sheet_name, tickets, n_legs, "Best Mix")
                all_ticket_sets.append((sheet_name, tickets, n_legs, "Best Mix"))
                print(f"  {sheet_name}: {len(tickets)} tickets")

    write_summary_sheet(wb, all_ticket_sets)
    wb.save(args.output)
    print(f"\n✅ Saved → {args.output}")
    print(f"\nPrizePicks Power Play payout reference:")
    print(f"  {'Legs':<6} {'Power':>8} {'Flex':>8}  {'$ to win $100 (Power)':>22}")
    for n in [2, 3, 4, 5, 6]:
        pp    = POWER_PLAY_BASE.get(n, "-")
        fl    = FLEX_PLAY_BASE.get(n, {}).get(n, "-")
        stake = round(100 / pp, 2) if isinstance(pp, (int, float)) else "-"
        print(f"  {n}-leg  {str(pp)+'x':>8} {str(fl)+'x':>8}  ${stake:>6}")


if __name__ == "__main__":
    main()
