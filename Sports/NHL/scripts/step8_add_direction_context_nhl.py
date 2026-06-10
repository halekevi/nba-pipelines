"""
Step 8 — Direction Context + Final Output (NHL)
Reads Step 7 ranked XLSX and produces a clean CSV + XLSX with human-readable
direction context and correct column mappings from the upstream pipeline.

Column mapping (Step 7 → Step 8 display):
  player_name          → player
  stat_norm            → prop_type  (normalized key)
  stat_type            → prop_display (original PrizePicks label)
  line_score           → line
  recommended_side     → direction
  opponent             → opp
  composite_hit_rate   → composite_hr
  hit_rate_over_L5     → hr_L5
  hit_rate_over_L10    → hr_L10
  hit_rate_over_L20    → hr_L20
  hit_rate_over_season → hr_season
  sample_L5/L10/L20    → sample_L5/L10/L20
  scoring_tier         → scoring_tier
  def_tier             → def_tier
  pp_tier              → pp_tier
  prop_score           → prop_score
  tier                 → tier
  rank                 → rank

Exports line_combo / line_combo_toi_pct / on_pp1_line to Full Slate when present in step7.
CAR skaters need NST line-pair cache (CDP or --import-csv); VGK may join from legacy cache rows.

Usage:
    py step8_add_direction_context_nhl.py --input outputs/step7_nhl_ranked.xlsx
        --output outputs/step8_nhl_direction_clean.xlsx --date YYYY-MM-DD
"""

import argparse
import csv
import shutil
import openpyxl
import pandas as pd
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo
try:
    from tqdm import tqdm as _tqdm
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm", "--break-system-packages", "-q"])
    from tqdm import tqdm as _tqdm
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


def _copy_dated_step8_nhl(output_xlsx_path: str, slate_date: str | None = None) -> None:
    src = Path(output_xlsx_path)
    if not src.is_file():
        return
    ds = str(slate_date or "").strip()[:10]
    try:
        dated_key = datetime.strptime(ds, "%Y-%m-%d").date().isoformat() if ds else date.today().isoformat()
    except Exception:
        dated_key = date.today().isoformat()
    repo_root = Path(__file__).resolve().parents[3]
    dated_dir = repo_root / "outputs" / dated_key
    try:
        dated_dir.mkdir(parents=True, exist_ok=True)
        dated_path = dated_dir / f"step8_nhl_direction_clean_{dated_key}.xlsx"
        shutil.copy2(src, dated_path)
        print(f"[NHL step8] Dated copy -> {dated_path}")
    except Exception as e:
        print(f"[NHL step8] WARN: dated copy failed: {e}")


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_float(x, default=0.0) -> float:
    try:
        if x is None or (isinstance(x, str) and x.strip() == ""):
            return default
        return float(x)
    except Exception:
        return default


def hr_output(val):
    """Write hit rates as raw 0-1 floats (4 decimals) for XLSX/CSV; empty if missing."""
    if val is None:
        return ""
    if isinstance(val, str) and val.strip() == "":
        return ""
    try:
        return round(float(val), 4)
    except (TypeError, ValueError):
        return ""


def fmt_num(val, digits=2) -> str:
    f = safe_float(val)
    if f == 0.0 and (val is None or str(val).strip() == ""):
        return ""
    return f"{f:.{digits}f}"


_ET = ZoneInfo("America/New_York")


def _nhl_game_date_time_et(game_start_val) -> tuple[str, str]:
    """Calendar date (YYYY-MM-DD) and clock string in America/New_York (mirrors NBA step8)."""
    if game_start_val is None or str(game_start_val).strip() == "":
        return "", ""
    ts = pd.to_datetime(game_start_val, utc=True, errors="coerce")
    if pd.isna(ts):
        t2 = pd.to_datetime(str(game_start_val).strip(), errors="coerce")
        if pd.isna(t2):
            return "", ""
        if t2.tzinfo is None:
            t2 = t2.tz_localize("UTC", ambiguous="NaT", nonexistent="shift_forward")
            if pd.isna(t2):
                return "", ""
        ts = t2
    if ts.tzinfo is None:
        ts = ts.tz_localize("UTC")
    et = ts.tz_convert(_ET)
    d = et.strftime("%Y-%m-%d")
    clk = et.strftime("%I:%M %p")
    if clk.startswith("0"):
        clk = clk[1:]
    return d, clk


# ── Column name resolution (handles upstream naming variants) ─────────────────

# Maps the canonical Step-8 field name → list of candidate column names in Step 7 output
# First match wins.
COLUMN_ALIASES = {
    "player":           ["player_name", "player"],
    "prop_display":     ["stat_type", "prop_display"],
    "prop_type":        ["stat_norm", "prop_type"],
    "line":             ["line_score", "line"],
    "pick_type":        ["pick_type"],
    "direction":        ["recommended_side", "bet_dir", "dir", "direction"],
    "opp":              ["opponent", "opp_team", "opp"],
    "composite_hr":     ["composite_hit_rate", "composite_hr"],
    "hr_L5":            ["hit_rate_over_L5",  "hr_last5",  "hr_L5"],
    "hr_L10":           ["hit_rate_over_L10", "hr_last10", "hr_L10"],
    "hr_L20":           ["hit_rate_over_L20", "hr_last20", "hr_L20"],
    "hr_season":        ["hit_rate_over_season", "hr_season"],
    "sample_L5":        ["sample_L5"],
    "over_L5":          ["over_L5"],
    "under_L5":         ["under_L5"],
    "over_L10":         ["over_L10"],
    "under_L10":        ["under_L10"],
    "sample_L10":       ["sample_L10"],
    "sample_L20":       ["sample_L20"],
    "sample_season":    ["sample_season"],
    "scoring_tier":     ["scoring_tier", "role_tier"],
    "def_tier":         ["def_tier"],
    "pp_tier":          ["pp_tier"],
    "edge":             ["edge"],
    "prop_score":       ["prop_score"],
    "ml_prob":          ["ml_prob"],
    "edge_score":       ["edge_score"],
    "blended_score":    ["blended_score"],
    "tier":             ["tier"],
    "rank":             ["rank"],
    "team":             ["team"],
    "is_home":          ["is_home"],
    "player_role":      ["player_role"],
    "position_group":   ["position_group"],
    "opp_gaa":          ["opp_gaa"],
    "implied_team_total": ["implied_team_total"],
    "game_total":       ["game_total"],
    "opp_saa":          ["opp_saa"],
    "opp_shots_allowed_avg": ["opp_saa"],
    "opp_pk_pct":       ["opp_pk_pct"],
    "def_rank":         ["def_rank"],
    "avg_L5":           ["avg_L5", "stat_last5_avg"],
    "avg_L10":          ["avg_L10", "stat_last10_avg"],
    "avg_L20":          ["avg_L20"],
    "avg_season":       ["avg_season", "stat_season_avg"],
    "games_played":     ["games_played"],
    "pts_per_game":     ["pts_per_game"],
    "pp_pts_per_game":  ["pp_pts_per_game"],
    "toi_avg_L10":      ["toi_avg_L10"],
    "toi_per_game_api": ["toi_per_game_api"],
    "game_start":       ["game_start", "start_time", "Game Time"],
    "fetched_at":       ["fetched_at"],
    "game_script_mult": ["game_script_mult"],
    "game_script_note": ["game_script_note"],
    "open_line":            ["open_line"],
    "line_movement":        ["line_movement"],
    "line_direction_shift": ["line_direction_shift"],
    "implied_prob":         ["implied_prob"],
    "implied_prob_over":    ["implied_prob_over"],
    "implied_prob_under":   ["implied_prob_under"],
    "rest_days":            ["rest_days"],
    "back_to_back":         ["back_to_back"],
    "goalie_name":          ["goalie_name"],
    "goalie_sv_pct":        ["goalie_sv_pct"],
    "player_avg_shots_L5":  ["player_avg_shots_L5"],
    "player_avg_shots_L10": ["player_avg_shots_L10"],
    "line_combo":         ["line_combo", "Line Combo"],
    "line_combo_toi_pct": ["line_combo_toi_pct", "Line Combo TOI%"],
    "on_pp1_line":        ["on_pp1_line", "On PP1 Line"],
    "player_injury_status": ["player_injury_status", "Player Injury Status"],
    "player_on_il":       ["player_on_il", "Player On IL"],
    "player_dtd":         ["player_dtd", "Player DTD"],
    "team_key_out":       ["team_key_out", "Team Key Out"],
    "team_dtd_count":     ["team_dtd_count", "Team DTD Count"],
    "consistency_grade":  ["consistency_grade", "Consistency Grade"],
    "team_top3_rank":     ["team_top3_rank", "Top3 Rank"],
    "team_bottom3_rank":  ["team_bottom3_rank", "Bottom3 Rank"],
    "def_boost_hist":     ["def_boost_hist", "Def Boost Hist"],
    "top3_weak_overperformer": ["top3_weak_overperformer", "Top3 Weak Over"],
    "top3_elite_fader":   ["top3_elite_fader", "Top3 Elite Fade"],
}

# Ordered like NBA/MLB ET pipelines: prefer full timestamps; time-only columns may not parse.
_GAME_TS_FALLBACK = ["game_start", "start_time", "Game Time"]


def _first_game_timestamp_raw(raw: dict, available_cols: set):
    """First column value that yields an ET calendar date; else first non-empty candidate."""
    for c in _GAME_TS_FALLBACK:
        if c not in available_cols:
            continue
        v = raw.get(c)
        if v in (None, ""):
            continue
        d_et, _ = _nhl_game_date_time_et(v)
        if d_et:
            return v
    for c in _GAME_TS_FALLBACK:
        if c not in available_cols:
            continue
        v = raw.get(c)
        if v not in (None, ""):
            return v
    return None


def resolve(row: dict, canonical: str, available_cols: set) -> str:
    """Return the value from `row` for the first matching alias column."""
    candidates = COLUMN_ALIASES.get(canonical, [canonical])
    for c in candidates:
        if c in available_cols and row.get(c) not in (None, ""):
            return str(row[c])
    return ""


def _game_log_g_columns(raw: dict, available_cols: set, n_games: int = 10) -> dict[str, str]:
    """
    Passthrough rolling game values for combined slate / Tier 3 distribution_std.
    Step 7 carries stat_g1..stat_g10 from step5; export as G1..G10 (NBA step8 convention).
    """
    out: dict[str, str] = {}
    for i in range(1, n_games + 1):
        gcol = f"G{i}"
        val = ""
        for col in (f"stat_g{i}", gcol):
            if col in available_cols and raw.get(col) not in (None, ""):
                val = fmt_num(raw.get(col), 2)
                break
        out[gcol] = val
    return out


def get_last_n_raw(raw: dict, available_cols: set, n: int) -> str:
    """
    Find the last-N raw game value for the prop's stat type.
    Step 4 writes columns like last1_shots_on_goal, last2_goals, etc.
    We scan for any column matching last{n}_<anything> excluding fantasy_score.
    """
    prefix = f"last{n}_"
    for col in sorted(available_cols):
        if col.startswith(prefix) and "fantasy_score" not in col:
            v = raw.get(col)
            if v not in (None, ""):
                return str(v)
    return ""


MISSING_SENTINELS = {"—", "-", "", "nan", "none", "null"}


def _parse_g_vals(row, prefix: str = "stat_g", n: int = 10) -> list[float]:
    vals: list[float] = []
    for i in range(1, n + 1):
        col = f"{prefix}{i}"
        if isinstance(row, pd.Series):
            raw_val = row.get(col, "")
        else:
            raw_val = row.get(col, "")
        raw = str(raw_val if raw_val is not None else "").strip()
        if raw.lower() in MISSING_SENTINELS:
            continue
        try:
            vals.append(float(raw))
        except ValueError:
            continue
    return vals


def _attach_distribution_std(df: pd.DataFrame, *, g_prefix: str = "stat_g") -> pd.DataFrame:
    """Sample std (ddof=1) of G1..10 / stat_g1..10 for pipeline_read distribution_std."""
    if df is None or len(df) == 0:
        return df
    out = df.copy()
    dist_n: list[int | None] = []
    dist_std: list[float | None] = []
    for _, row in out.iterrows():
        g_vals = _parse_g_vals(row, prefix=g_prefix)
        n = len(g_vals)
        dist_n.append(n)
        if n >= 2:
            dist_std.append(round(float(pd.Series(g_vals).std(ddof=1)), 4))
        else:
            dist_std.append(None)
    out["distribution_n"] = dist_n
    out["distribution_std"] = dist_std
    return out


# ── Build display row ─────────────────────────────────────────────────────────

def build_display_row(raw: dict, available_cols: set) -> dict:
    """Build a clean, fully-populated display row from a raw Step 7 row."""
    def r(key):
        return resolve(raw, key, available_cols)

    gs_raw = _first_game_timestamp_raw(raw, available_cols)
    game_date_et, game_time_et = _nhl_game_date_time_et(gs_raw)

    avg_L5     = safe_float(r("avg_L5"))
    avg_L10    = safe_float(r("avg_L10"))
    avg_L20    = safe_float(r("avg_L20"))
    avg_season = safe_float(r("avg_season"))
    line_val   = safe_float(r("line"))

    pick_pt = str(r("pick_type") or "").strip().upper()
    forced_od = pick_pt in ("GOBLIN", "DEMON")

    proj_num = None
    for av in (avg_L5, avg_L10, avg_L20):
        if av is not None:
            proj_num = av
            break

    upstream_dir = str(r("direction") or "").strip().upper()
    upstream_edge = safe_float(r("edge"))

    signed_e = (proj_num - line_val) if proj_num is not None and line_val is not None else upstream_edge

    if forced_od:
        direction = "OVER"
    elif signed_e is not None:
        direction = "OVER" if signed_e >= 0 else "UNDER"
    elif upstream_dir in ("OVER", "UNDER"):
        direction = upstream_dir
    else:
        direction = "OVER"

    composite_hr = safe_float(r("composite_hr"))
    hr_L5        = safe_float(r("hr_L5"))
    hr_L10       = safe_float(r("hr_L10"))
    hr_L20       = safe_float(r("hr_L20"))
    hr_season    = safe_float(r("hr_season"))

    # Direction-adjust hit rates: if UNDER, flip all rates.
    # Use explicit None checks — 0.0 is a valid hit rate (falsy but meaningful).
    def _has_val(v) -> bool:
        return v is not None and str(v).strip() not in ("", "nan")

    if direction == "UNDER":
        composite_hr_adj = (1 - composite_hr) if _has_val(r("composite_hr")) else ""
        hr_L5_adj        = (1 - hr_L5)        if _has_val(r("hr_L5"))        else ""
        hr_L10_adj       = (1 - hr_L10)       if _has_val(r("hr_L10"))       else ""
        hr_L20_adj       = (1 - hr_L20)       if _has_val(r("hr_L20"))       else ""
        hr_season_adj    = (1 - hr_season)     if _has_val(r("hr_season"))    else ""
    else:
        composite_hr_adj = composite_hr if _has_val(r("composite_hr")) else ""
        hr_L5_adj        = hr_L5        if _has_val(r("hr_L5"))        else ""
        hr_L10_adj       = hr_L10       if _has_val(r("hr_L10"))       else ""
        hr_L20_adj       = hr_L20       if _has_val(r("hr_L20"))       else ""
        hr_season_adj    = hr_season    if _has_val(r("hr_season"))    else ""

    # Gap: rolling avg vs line (how far above/below)
    gap_L5  = round(avg_L5  - line_val, 3) if avg_L5  and line_val else ""
    gap_L10 = round(avg_L10 - line_val, 3) if avg_L10 and line_val else ""

    # Over/under counts — swap when direction is UNDER so columns reflect bet direction
    raw_over_L5  = safe_float(r("over_L5"))
    raw_under_L5 = safe_float(r("under_L5"))
    raw_over_L10  = safe_float(r("over_L10"))
    raw_under_L10 = safe_float(r("under_L10"))
    s5  = safe_float(r("sample_L5"))
    s10 = safe_float(r("sample_L10"))
    if direction == "UNDER":
        # UNDER bets win when value < line, so "wins" = under count
        win_L5  = int(raw_under_L5) if raw_under_L5 is not None else ""
        win_L10 = int(raw_under_L10) if raw_under_L10 is not None else ""
    else:
        win_L5  = int(raw_over_L5)  if raw_over_L5  is not None else ""
        win_L10 = int(raw_over_L10) if raw_over_L10 is not None else ""
    tot_L5  = int(s5)  if s5  is not None else ""
    tot_L10 = int(s10) if s10 is not None else ""
    hit_L5_display  = f"{win_L5}/{tot_L5}"   if win_L5  != "" and tot_L5  != "" else ""
    hit_L10_display = f"{win_L10}/{tot_L10}" if win_L10 != "" and tot_L10 != "" else ""

    return {
        "rank":             r("rank"),
        "tier":             r("tier"),
        "player":           r("player"),
        "team":             r("team"),
        "opp":              r("opp"),
        "is_home":          r("is_home"),
        "player_role":      r("player_role"),
        "position_group":   r("position_group"),
        "prop_display":     r("prop_display"),
        "prop_type":        r("prop_type"),
        "line":             fmt_num(r("line"), 1),
        "direction":        direction,
        # Hit rates (direction-adjusted) — raw 0-1 floats for step9 / Excel
        "composite_hr":     hr_output(composite_hr_adj),
        # Canonical alias used by upstream validators and cross-sport tooling.
        "hit_rate":         hr_output(composite_hr_adj),
        "hr_L5":            hr_output(hr_L5_adj),
        "hr_L10":           hr_output(hr_L10_adj),
        "hr_L20":           hr_output(hr_L20_adj),
        "hr_season":        hr_output(hr_season_adj),
        "sample_L5":        r("sample_L5"),
        "hit_L5":           hit_L5_display,
        "sample_L10":       r("sample_L10"),
        "hit_L10":          hit_L10_display,
        "sample_L20":       r("sample_L20"),
        "sample_season":    r("sample_season"),
        # Rolling averages
        "avg_L5":           fmt_num(r("avg_L5"), 2),
        "avg_L10":          fmt_num(r("avg_L10"), 2),
        "avg_L20":          fmt_num(r("avg_L20"), 2),
        "avg_season":       fmt_num(r("avg_season"), 2),
        "gap_vs_line_L5":   gap_L5,
        "gap_vs_line_L10":  gap_L10,
        # Projection: best available rolling avg (used by combined_slate_tickets as "Proj")
        "projection":       fmt_num(r("avg_L5"), 2) or fmt_num(r("avg_L10"), 2) or fmt_num(r("avg_L20"), 2) or "",
        # Raw over/under counts for combined slate L5 Over / L5 Under columns
        "over_L5_raw":      int(raw_over_L5)  if raw_over_L5  is not None else "",
        "under_L5_raw":     int(raw_under_L5) if raw_under_L5 is not None else "",
        "over_L10_raw":     int(raw_over_L10) if raw_over_L10 is not None else "",
        "under_L10_raw":    int(raw_under_L10) if raw_under_L10 is not None else "",
        # Last 3 raw game values (from Step 4 fix)
        "last1_raw":        get_last_n_raw(raw, available_cols, 1),
        "last2_raw":        get_last_n_raw(raw, available_cols, 2),
        "last3_raw":        get_last_n_raw(raw, available_cols, 3),
        "games_played":     r("games_played"),
        # Context
        "scoring_tier":     r("scoring_tier"),
        "pp_tier":          r("pp_tier"),
        "def_tier":         r("def_tier"),
        "def_rank":         r("def_rank"),
        "consistency_grade": r("consistency_grade"),
        "team_top3_rank":   r("team_top3_rank"),
        "team_bottom3_rank": r("team_bottom3_rank"),
        "def_boost_hist":   r("def_boost_hist"),
        "top3_weak_overperformer": r("top3_weak_overperformer"),
        "top3_elite_fader": r("top3_elite_fader"),
        "opp_gaa":          fmt_num(r("opp_gaa"), 3),
        "implied_team_total": fmt_num(r("implied_team_total"), 2),
        "game_total":       fmt_num(r("game_total"), 1),
        "opp_saa":          fmt_num(r("opp_saa"), 3),
        "opp_shots_allowed_avg": fmt_num(r("opp_saa"), 1),
        "opp_pk_pct":       fmt_num(r("opp_pk_pct"), 3),
        "player_avg_shots_L5": fmt_num(r("player_avg_shots_L5"), 2),
        "player_avg_shots_L10": fmt_num(r("player_avg_shots_L10"), 2),
        "pts_per_game":     fmt_num(r("pts_per_game"), 3),
        "pp_pts_per_game":  fmt_num(r("pp_pts_per_game"), 3),
        "toi_avg_L10":      fmt_num(r("toi_avg_L10"), 2),
        "toi_per_game_api": fmt_num(r("toi_per_game_api"), 2),
        # Model scores
        "prop_score":       fmt_num(r("prop_score"), 5),
        "edge":             fmt_num(signed_e, 4) if signed_e is not None else fmt_num(r("edge"), 4),
        "abs_edge":         fmt_num(abs(signed_e), 4) if signed_e is not None else (
            fmt_num(abs(upstream_edge), 4) if upstream_edge is not None else ""
        ),
        "ml_prob":          fmt_num(r("ml_prob"), 4),
        "edge_score":       fmt_num(r("edge_score"), 4),
        "blended_score":    fmt_num(r("blended_score"), 4),
        "pick_type":        r("pick_type"),
        # Game info — ET calendar + wall clock (grader / MLB-style slate_game_date filter)
        "Game Date":        game_date_et,
        # Lowercase alias for downstream tooling / CSV consumers (mirrors other sports pipelines).
        "game_date":        game_date_et,
        "Game Time":        game_time_et or (str(gs_raw) if gs_raw not in (None, "") else r("game_start")),
        "game_start":       (str(gs_raw) if gs_raw not in (None, "") else r("game_start")),
        "fetched_at":       str(r("fetched_at") or ""),
        "game_script_mult": (
            fmt_num(raw.get("game_script_mult"), 3)
            if raw.get("game_script_mult") not in (None, "")
            else ""
        ),
        "game_script_note": str(raw.get("game_script_note", "") or ""),
        "open_line": (
            fmt_num(r("open_line"), 2)
            if r("open_line") not in (None, "")
            else ""
        ),
        "line_movement": fmt_num(r("line_movement"), 3),
        "line_direction_shift": str(r("line_direction_shift") or "stable"),
        "implied_prob": fmt_num(r("implied_prob"), 4),
        "implied_prob_over": fmt_num(r("implied_prob_over"), 4),
        "implied_prob_under": fmt_num(r("implied_prob_under"), 4),
        "rest_days": str(r("rest_days") or ""),
        "back_to_back": "1" if str(r("back_to_back") or "").strip() in ("1", "1.0", "true", "True") else "0",
        "goalie_name": str(r("goalie_name") or ""),
        "goalie_sv_pct": fmt_num(r("goalie_sv_pct"), 3),
        "line_combo": str(r("line_combo") or ""),
        "line_combo_toi_pct": fmt_num(r("line_combo_toi_pct"), 1),
        "on_pp1_line": str(r("on_pp1_line") or ""),
        "player_injury_status": str(r("player_injury_status") or ""),
        "player_on_il": str(r("player_on_il") or ""),
        "player_dtd": str(r("player_dtd") or ""),
        "team_key_out": str(r("team_key_out") or ""),
        "team_dtd_count": fmt_num(r("team_dtd_count"), 0),
        **_game_log_g_columns(raw, available_cols),
    }


# ── Writers ───────────────────────────────────────────────────────────────────

TIER_FILL = {
    "A": "C6EFCE",
    "B": "FFEB9C",
    "C": "FCE4D6",
    "D": "E0E0E0",
}
HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"
_COL_WIDTH_OVERRIDES = {
    "fetched_at": 20,
    "distribution_std": 10,
    "distribution_n": 8,
    "implied_prob": 12,
    "implied_prob_over": 14,
    "implied_prob_under": 15,
    "player_injury_status": 16,
    "player_on_il": 10,
    "player_dtd": 10,
    "team_key_out": 12,
    "team_dtd_count": 12,
}


def write_csv(rows: list[dict], path: str):
    if not rows:
        open(path, "w").close()
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"CSV saved -> {path}")


def _write_sheet(ws, rows: list[dict], headers: list[str]):
    """Write data + formatting to a worksheet."""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")

    for row_data in rows:
        ws.append([row_data.get(h, "") for h in headers])
        last_row = ws.max_row
        tier = str(row_data.get("tier", "D")).upper()
        fill_color = TIER_FILL.get(tier, "FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            ws.cell(last_row, col_idx).fill = PatternFill("solid", fgColor=fill_color)

    ws.freeze_panes = "A2"
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_data in rows[:500]:
            v = row_data.get(h, "")
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = _COL_WIDTH_OVERRIDES.get(
            h, min(40, max(10, max_len + 2))
        )


def write_xlsx(rows: list[dict], path: str):
    if not rows:
        openpyxl.Workbook().save(path)
        return

    # Human + machine calendar columns; grader maps only "Game Date" -> slate_game_date
    # so lowercase `game_date` can coexist in the XLSX (dated copy included).
    rows_xlsx = list(rows)
    headers = list(rows_xlsx[0].keys())
    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "All Props"
    _write_sheet(ws_all, rows_xlsx, headers)

    skaters = [r for r in rows_xlsx if str(r.get("player_role", "")).upper() == "SKATER"]
    if skaters:
        _write_sheet(wb.create_sheet("Skaters"), skaters, headers)

    goalies = [r for r in rows_xlsx if str(r.get("player_role", "")).upper() == "GOALIE"]
    if goalies:
        _write_sheet(wb.create_sheet("Goalies"), goalies, headers)

    a_rows = [r for r in rows_xlsx if str(r.get("tier", "")).upper() == "A"]
    if a_rows:
        _write_sheet(wb.create_sheet("A-Tier Best"), a_rows, headers)

    wb.save(path)
    print(f"XLSX saved -> {path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  required=True, help="Step 7 ranked XLSX")
    ap.add_argument("--output", default="outputs/step8_nhl_direction_clean.xlsx")
    ap.add_argument(
        "--date",
        default="",
        help="Target slate date YYYY-MM-DD (default: local today; matches run_pipeline --Date)",
    )
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    if "All Props" in wb.sheetnames:
        ws = wb["All Props"]
    else:
        ws = wb.active
        print(
            f"WARNING: 'All Props' sheet not found, reading active sheet: {ws.title}"
        )
    raw_headers = []
    raw_rows = []
    for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_i == 1:
            raw_headers = [str(x).strip() if x is not None else f"col_{i}"
                           for i, x in enumerate(row)]
            continue
        r = {}
        for j, h in enumerate(raw_headers):
            r[h] = row[j] if j < len(row) else ""
        raw_rows.append(r)
    wb.close()

    available_cols = set(raw_headers)
    print(f"Loaded {len(raw_rows)} props from {args.input}")
    print(f"Upstream columns ({len(available_cols)}): {sorted(available_cols)}\n")

    display_rows = [build_display_row(r, available_cols) for r in _tqdm(raw_rows, desc="  Building display rows", unit="prop")]

    # ── Date filter: keep only target slate date games (ET calendar via Game Date) ─
    ds = str(args.date).strip()[:10] if args.date and str(args.date).strip() else ""
    target_local = (
        datetime.strptime(ds, "%Y-%m-%d").date() if ds else date.today()
    )
    target_str = target_local.isoformat()
    before_filter = len(display_rows)
    unfiltered = list(display_rows)

    def _game_start_local_date(gs):
        """Return local calendar date for game_start, or None if unparseable."""
        if gs is None or gs == "":
            return None
        try:
            if isinstance(gs, datetime):
                if gs.tzinfo is not None:
                    return gs.astimezone().date()
                return gs.date()
            gs_str = str(gs).strip()
            if not gs_str:
                return None
            for fmt in (
                "%Y-%m-%dT%H:%M:%S%z",
                "%Y-%m-%d %H:%M:%S%z",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M:%S",
            ):
                try:
                    dt = datetime.strptime(gs_str[:25], fmt)
                    if dt.tzinfo is not None:
                        return dt.astimezone().date()
                    return dt.date()
                except ValueError:
                    continue
            if len(gs_str) >= 10 and gs_str[4] == "-" and gs_str[7] == "-":
                try:
                    return datetime.strptime(gs_str[:10], "%Y-%m-%d").date()
                except ValueError:
                    return None
        except Exception:
            return None
        return None

    def _is_target_date(gs) -> bool:
        """
        Fallback when ``Game Date`` / ``game_date`` is empty: compare ``game_start``
        using the ET calendar date (mirrors NBA step8 behavior; avoids local-midnight skew).
        """
        d_et, _t_et = _nhl_game_date_time_et(gs)
        if d_et:
            return d_et == target_str
        return False

    def _row_matches_target(r: dict) -> bool:
        gd = str(r.get("Game Date", "") or r.get("game_date", "")).strip()
        if gd:
            return gd == target_str
        return _is_target_date(r.get("game_start", ""))

    display_rows = [r for r in display_rows if _row_matches_target(r)]
    dropped = before_filter - len(display_rows)
    if len(display_rows) == 0 and len(unfiltered) > 0:
        dates_seen: list[str] = []
        for r in unfiltered:
            gd = str(r.get("Game Date", "") or r.get("game_date", "")).strip()
            if gd:
                dates_seen.append(gd)
            else:
                d_et, _t_et = _nhl_game_date_time_et(r.get("game_start", ""))
                if d_et:
                    dates_seen.append(d_et)
                else:
                    d = _game_start_local_date(r.get("game_start", ""))
                    if d is not None:
                        dates_seen.append(d.isoformat())
        if dates_seen:
            unique = sorted(set(dates_seen))
            future = [d for d in unique if d >= target_str]
            chosen = future[0] if future else unique[-1]

            def _is_chosen_row(row: dict) -> bool:
                gd = str(row.get("Game Date", "") or row.get("game_date", "")).strip()
                if gd:
                    return gd == chosen
                d_et, _t_et = _nhl_game_date_time_et(row.get("game_start", ""))
                if d_et:
                    return d_et == chosen
                d2 = _game_start_local_date(row.get("game_start", ""))
                return d2 is not None and d2.isoformat() == chosen

            display_rows = [r for r in unfiltered if _is_chosen_row(r)]
            print(
                f"[DateFilter] Kept 0/{before_filter} for {target_str}; "
                f"using nearest slate date {chosen} (>= target when possible) -> {len(display_rows)} rows"
            )
        else:
            print(
                f"[DateFilter] Kept {len(display_rows)}/{before_filter} rows for {target_str} "
                f"(dropped {dropped} future/past rows)"
            )
    else:
        print(
            f"[DateFilter] Kept {len(display_rows)}/{before_filter} rows for {target_str} "
            f"(dropped {dropped} future/past rows)"
        )

    display_rows.sort(
        key=lambda r: int(r.get("rank") or 9999)
        if str(r.get("rank", "")).isdigit() else 9999
    )

    if display_rows:
        display_rows = _attach_distribution_std(
            pd.DataFrame(display_rows), g_prefix="G"
        ).to_dict(orient="records")
        filled_std = int(
            pd.to_numeric(
                pd.Series([r.get("distribution_std") for r in display_rows]),
                errors="coerce",
            ).notna().sum()
        )
        print(f"[NHL step8] distribution_std filled {filled_std}/{len(display_rows)} rows")

    if args.output.lower().endswith(".xlsx"):
        write_xlsx(display_rows, args.output)
        csv_out = args.output.replace(".xlsx", ".csv")
        write_csv(display_rows, csv_out)
        xlsx_final = args.output
    else:
        write_csv(display_rows, args.output)
        xlsx_out = args.output.replace(".csv", "_clean.xlsx")
        write_xlsx(display_rows, xlsx_out)
        xlsx_final = xlsx_out

    _copy_dated_step8_nhl(xlsx_final, slate_date=target_str)

    # ── Console summary ───────────────────────────────────────────────────────
    tier_counts = {}
    for r in display_rows:
        t = str(r.get("tier", "?")).upper()
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"\nTier breakdown: {tier_counts}")

    for tier_label in ("A", "B"):
        tier_rows = [r for r in display_rows if str(r.get("tier", "")).upper() == tier_label]
        if not tier_rows:
            continue
        print(f"\n{'='*65}")
        print(f"{tier_label}-TIER PICKS ({len(tier_rows)} total)")
        print("="*65)
        for r in tier_rows[:25]:
            print(
                f"  [{tier_label}] #{str(r.get('rank','?')):>3}  "
                f"{str(r.get('player','')):25s}  "
                f"{str(r.get('direction','')):5s} {str(r.get('line','')):>5}  "
                f"{str(r.get('prop_display') or r.get('prop_type','')):20s}  "
                f"vs {str(r.get('opp','?')):4s}  |  "
                f"{str(r.get('scoring_tier','')):12s}  {str(r.get('def_tier','')):7s}"
            )
            print(
                f"       Composite HR: {str(r.get('composite_hr','')):>6}  |  "
                f"L5:{str(r.get('hr_L5','')):>6}  "
                f"L10:{str(r.get('hr_L10','')):>6}  "
                f"L20:{str(r.get('hr_L20','')):>6}  "
                f"Szn:{str(r.get('hr_season','')):>6}  |  "
                f"AvgL10:{str(r.get('avg_L10','')):>5}  "
                f"Line:{str(r.get('line','')):>5}  "
                f"Gap:{str(r.get('gap_vs_line_L10','')):>6}  |  "
                f"score={r.get('prop_score','')}"
            )


if __name__ == "__main__":
    main()
