#!/usr/bin/env python3
"""
step8_add_direction_context.py
-------------------------------
Reads step7_ranked_props.xlsx, appends direction context columns,
and outputs both a full CSV and a clean formatted XLSX for decision making.

Adds:
- model_dir
- final_bet_direction
- final_dir_reason

Run:
  py -3.14 step8_add_direction_context.py --input step7_ranked_props.xlsx --sheet ALL --output step8_all_direction.csv
"""

from __future__ import annotations

import argparse
from pathlib import Path
from zoneinfo import ZoneInfo

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
import sys

_REPO = Path(__file__).resolve().parent
for _ in range(10):
    if (_REPO / "utils" / "step8_edge_direction.py").is_file():
        if str(_REPO) not in sys.path:
            sys.path.insert(0, str(_REPO))
        break
    _REPO = _REPO.parent
else:
    raise RuntimeError("Could not locate repo root with utils/step8_edge_direction.py")

from scripts.l10_streak_utils import finalize_l10_ui_columns
from utils.hit_tracking_columns import HIT_TRACKING_RENAME, attach_hit_tracking_columns
from utils.step8_edge_direction import reconcile_signed_edge_abs_dataframe

_ET = ZoneInfo("America/New_York")


def _start_times_et(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, utc=True, errors="coerce").dt.tz_convert(_ET)


def _format_et_clock(et: pd.Series) -> pd.Series:
    clk = et.dt.strftime("%I:%M %p")
    return clk.str.replace(r"^0(\d:)", r"\1", regex=True)


def _copy_dated_step8_wnba(output_xlsx_path: str, slate_date: str) -> None:
    """Dated mirror of clean step8 is handled in scripts/run_wnba_pipeline.ps1 (Publish-WnbaStep8CleanArtifacts).

    Avoid writing step1–7 or extra step8 copies into repo ``outputs/<date>/`` from Python; that script
    publishes ``outputs/<date>/step8_wnba_direction_clean_<date>.xlsx`` and ``data/outputs/`` only.
    """
    _ = (output_xlsx_path, slate_date)
    return


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t:
        return "Goblin"
    if "dem" in t:
        return "Demon"
    return "Standard"

TIER_COLORS = {
    'A': ('1E8449', 'FFFFFF'),
    'B': ('2874A6', 'FFFFFF'),
    'C': ('D4AC0D', '000000'),
    'D': ('717D7E', 'FFFFFF'),
}
HEADER_COLOR = '1C1C1C'

def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

def _line_shift_fill(val) -> str:
    if val is None or val == '' or (isinstance(val, float) and pd.isna(val)):
        return 'FFFFFF'
    try:
        num = float(val)
        if num > 0:
            return 'C8F7C5'
        if num < 0:
            return 'F7C5C5'
        return 'FFFFFF'
    except (TypeError, ValueError):
        pass
    s = str(val).strip().upper()
    if 'UP' in s or s == 'MOVED_UP':
        return 'C8F7C5'
    if 'DOWN' in s or s == 'MOVED_DOWN':
        return 'F7C5C5'
    return 'FFFFFF'

MISSING_SENTINELS = {"—", "-", "", "nan", "none", "null"}


def _parse_g_vals(row, prefix: str = "stat_g", n: int = 10) -> list[float]:
    vals: list[float] = []
    for i in range(1, n + 1):
        col = f"{prefix}{i}"
        if isinstance(row, pd.Series):
            raw_val = row.get(col, "")
        else:
            raw_val = row.get(col, "")
        raw = str(raw_val if raw_val is not None else "").strip()
        if raw.lower() in MISSING_SENTINELS:
            continue
        try:
            vals.append(float(raw))
        except ValueError:
            continue
    return vals


def _attach_distribution_std(df: pd.DataFrame, *, g_prefix: str = "stat_g") -> pd.DataFrame:
    """Sample std (ddof=1) of stat_g1..10 for pipeline_read distribution_std."""
    if df is None or len(df) == 0:
        return df
    out = df.copy()
    dist_n: list[int | None] = []
    dist_std: list[float | None] = []
    for _, row in out.iterrows():
        g_vals = _parse_g_vals(row, prefix=g_prefix)
        n = len(g_vals)
        dist_n.append(n)
        if n >= 2:
            dist_std.append(round(float(pd.Series(g_vals).std(ddof=1)), 4))
        else:
            dist_std.append(None)
    out["distribution_n"] = dist_n
    out["distribution_std"] = dist_std
    return out

def write_sheet(wb, name, data):
    ws = wb.create_sheet(name)
    tier_bg, tier_fg = TIER_COLORS.get(name, ('333333', 'FFFFFF'))
    headers = list(data.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        direction = row[headers.index('Direction')] if 'Direction' in headers else ''
        for ci, val in enumerate(row, 1):
            col_name = headers[ci - 1]
            display_val = '' if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=display_val)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()

            if col_name == 'Tier':
                cell.fill = PatternFill('solid', start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name='Arial', size=9)
            elif col_name == 'Direction':
                bg = 'C8F7C5' if val == 'OVER' else 'F7C5C5' if val == 'UNDER' else 'FFFFFF'
                cell.fill = PatternFill('solid', start_color=bg)
                cell.font = Font(bold=True, name='Arial', size=9)
            elif col_name in ('Line Shift', 'line_direction_shift'):
                cell.fill = PatternFill('solid', start_color=_line_shift_fill(display_val))
            else:
                cell.fill = PatternFill('solid', start_color='F9F9F9' if ri % 2 == 0 else 'FFFFFF')

    col_widths = {
        'Tier': 6, 'Rank Score': 10, 'Rank Score Penalized': 14, 'Player': 18, 'Pos': 6,
        'Team': 6, 'Opp': 6, 'Days Rest': 9, 'B2B': 6, 'Opp Rest': 9, 'Opp B2B': 8, 'Game Total': 10, 'Spread': 8, 'Game Date': 12, 'Game Time': 10,
        'Prop': 16, 'Pick Type': 10, 'Line': 7,
        'Direction': 9, 'Edge': 7, 'Abs Edge': 7, 'Projection': 10,
        'ML Prob': 9, 'Edge Score': 10, 'Blended Score': 12,
        'Hit Rate (5g)': 12, 'Last 5 Avg': 10, 'Season Avg': 10,
        'L5 Over': 8, 'L5 Under': 8,
        'Def Rank': 9, 'Def Tier': 10,
        'Min Tier': 9, 'Shot Role': 10, 'Usage Role': 10,
        'Usage Pct': 9, 'Usage Tier': 9, 'Star Tier': 8, 'Franchise Star': 12,
        'Usage Boost': 10, 'Usage Boost Proj': 12, 'Usage Boost Reason': 18,
        'Usage Boost Source': 14,
        'Team Star Out': 12, 'Key Facilitator Out': 14, 'Injury Boost': 12, 'Usage Vacuum': 11,
        'Last 10 Avg': 10,
        'G1': 6, 'G2': 6, 'G3': 6, 'G4': 6, 'G5': 6,
        'G6': 6, 'G7': 6, 'G8': 6, 'G9': 6, 'G10': 6,
        'Void Reason': 20,
        'Open Line': 8, 'Line Movement': 12, 'Line Shift': 10,
        'open_line': 10, 'line_movement': 12, 'line_direction_shift': 16,
        'implied_prob': 12, 'implied_prob_over': 14, 'implied_prob_under': 15,
        'distribution_std': 10, 'distribution_n': 8,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str):
    df2 = df.copy()
    # Align L5 Over/Under and 5g hit rate with stat_g1..5 vs line (matches NBA step8; fixes sparse early-season rows).
    g5_cols = [c for c in ("stat_g1", "stat_g2", "stat_g3", "stat_g4", "stat_g5") if c in df2.columns]
    if g5_cols and "line" in df2.columns:
        for c in ("last5_over", "last5_under"):
            if c in df2.columns:
                df2[c] = pd.to_numeric(df2[c], errors="coerce")
        g5 = df2[g5_cols].apply(pd.to_numeric, errors="coerce")
        line = pd.to_numeric(df2["line"], errors="coerce")
        valid_n = g5.notna().sum(axis=1)
        over_n = g5.gt(line, axis=0).sum(axis=1)
        under_n = g5.lt(line, axis=0).sum(axis=1)
        has_hist = valid_n > 0
        if "last5_over" not in df2.columns:
            df2["last5_over"] = np.nan
        if "last5_under" not in df2.columns:
            df2["last5_under"] = np.nan
        df2.loc[has_hist, "last5_over"] = over_n[has_hist]
        df2.loc[has_hist, "last5_under"] = under_n[has_hist]
        if "line_hit_rate_over_ou_5" in df2.columns:
            df2["line_hit_rate_over_ou_5"] = pd.to_numeric(df2["line_hit_rate_over_ou_5"], errors="coerce")
            hr = (over_n / valid_n.replace(0, np.nan)).where(valid_n > 0)
            df2.loc[has_hist, "line_hit_rate_over_ou_5"] = hr[has_hist]

    if "start_time" in df2.columns:
        et = _start_times_et(df2["start_time"])
        parsed_gd = et.dt.strftime("%Y-%m-%d").where(et.notna(), "").astype(str).str.strip()
        prev = df2.get("game_date", pd.Series([""] * len(df2))).astype(str).str.strip().str[:10]
        prev_ok = prev.str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)
        df2["game_time"] = _format_et_clock(et)
        # Prefer upstream game_date when it is a valid YYYY-MM-DD (step1 may anchor it to pipeline
        # --date for full boards). Only fall back to start_time ET when game_date is missing/wrong.
        df2["game_date"] = np.where(
            prev_ok,
            prev,
            np.where(parsed_gd.str.len() > 0, parsed_gd, ""),
        )
    else:
        if "game_date" not in df2.columns:
            df2["game_date"] = ""
        if "game_time" not in df2.columns:
            df2["game_time"] = ""

    if "line" in df2.columns:
        df2 = finalize_l10_ui_columns(df2, line_col="line")
    df2 = attach_hit_tracking_columns(df2, "WNBA")

    keep = [
        'tier', 'rank_score', 'rank_score_penalized',
        'player', 'pos', 'team', 'opp_team', 'days_rest', 'is_back_to_back', 'opp_days_rest', 'opp_b2b',
        'h2h_avg', 'h2h_over_pct', 'h2h_games', 'h2h_last',
        'game_total', 'spread', 'game_date', 'game_time',
        'prop_type', 'pick_type', 'line',
        'final_bet_direction',
        'edge', 'projection',
        'ml_prob',
        'hit_rate', 'hit_rate_l5', 'hit_rate_l10',
        'strat_hit_rate', 'strat_n',
        'player_hr_historical', 'opp_hr_historical',
        'sport_signal_maturity', 'confidence_tier', 'confidence_score', 'confidence_note',
        'edge_score',
        'blended_score',
        'line_hit_rate_over_ou_5',
        'line_games_played_5', 'line_games_played_10',
        'stat_last5_avg', 'stat_season_avg',
        'stat_last10_avg',
        'stat_g1', 'stat_g2', 'stat_g3', 'stat_g4', 'stat_g5',
        'stat_g6', 'stat_g7', 'stat_g8', 'stat_g9', 'stat_g10',
        'last5_over', 'last5_under',
        'l10_over', 'l10_under', 'l10_over_pct', 'l10_streak', 'l10_games_played',
        'line_hits_over_10', 'line_hits_under_10',
        'OVERALL_DEF_RANK', 'DEF_TIER',
        'minutes_tier', 'shot_role', 'usage_role',
        'usage_pct', 'usage_tier', 'star_tier', 'is_franchise_star',
        'usage_boost', 'usage_boost_proj', 'usage_boost_reason', 'usage_boost_source',
        'team_star_out', 'key_facilitator_out', 'injury_boost_candidate', 'usage_vacuum',
        'void_reason',
        'open_line', 'line_movement', 'line_direction_shift',
        'implied_prob', 'implied_prob_over', 'implied_prob_under',
        'distribution_std', 'distribution_n',
        'consistency_grade',
        'team_top3_rank', 'team_bottom3_rank', 'def_boost_hist',
        'top3_weak_overperformer', 'top3_elite_fader',
        'top3_def_context', 'top3_under_context',
    ]
    # only keep cols that exist
    keep = [c for c in keep if c in df2.columns]
    clean = df2[keep].copy()

    for col in ['rank_score', 'edge', 'projection', 'ml_prob', 'edge_score', 'blended_score', 'line_hit_rate_over_ou_5']:
        if col in clean.columns:
            rnd = 4 if col in ('ml_prob', 'edge_score', 'blended_score') else 2
            clean[col] = pd.to_numeric(clean[col], errors='coerce').round(rnd)
    for col in ['open_line', 'line_movement', 'implied_prob', 'implied_prob_over', 'implied_prob_under']:
        if col in clean.columns:
            if col in ('implied_prob', 'implied_prob_over', 'implied_prob_under'):
                rnd = 4
            else:
                rnd = 2 if col == 'open_line' else 3
            clean[col] = pd.to_numeric(clean[col], errors='coerce').round(rnd)
    if 'line_direction_shift' in clean.columns:
        clean['line_direction_shift'] = (
            clean['line_direction_shift'].astype(str).str.strip().replace({'nan': '', 'None': ''})
        )
        clean.loc[clean['line_direction_shift'].eq(''), 'line_direction_shift'] = 'stable'
    for col in ['stat_last5_avg', 'stat_season_avg', 'stat_last10_avg']:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors='coerce').round(1)
    for col in [f'stat_g{i}' for i in range(1, 11)]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors='coerce').round(1)
    for col in ['last5_over', 'last5_under']:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors='coerce').astype('Int64')
    if 'distribution_std' in clean.columns:
        clean['distribution_std'] = pd.to_numeric(clean['distribution_std'], errors='coerce').round(4)
    if 'distribution_n' in clean.columns:
        clean['distribution_n'] = pd.to_numeric(clean['distribution_n'], errors='coerce').astype('Int64')

    tier_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
    clean['_tier_sort'] = clean['tier'].map(tier_order)
    clean = clean.sort_values(['_tier_sort', 'rank_score'], ascending=[True, False]).drop(columns='_tier_sort')

    rename = {
        'tier': 'Tier', 'rank_score': 'Rank Score', 'rank_score_penalized': 'Rank Score Penalized',
        'player': 'Player', 'pos': 'Pos', 'team': 'Team', 'opp_team': 'Opp',
        'days_rest': 'Days Rest',
        'is_back_to_back': 'B2B',
        'opp_days_rest': 'Opp Rest',
        'opp_b2b': 'Opp B2B',
        'h2h_avg': 'H2H Avg',
        'h2h_over_pct': 'H2H Over%',
        'h2h_games': 'H2H Games',
        'h2h_last': 'H2H Last',
        'game_total': 'Game Total',
        'spread': 'Spread',
        'game_date': 'Game Date', 'game_time': 'Game Time',
        'prop_type': 'Prop', 'pick_type': 'Pick Type', 'line': 'Line',
        'final_bet_direction': 'Direction',
        'edge': 'Edge', 'abs_edge': 'Abs Edge', 'projection': 'Projection',
        'ml_prob':            'ML Prob',
        'edge_score':         'Edge Score',
        'blended_score':      'Blended Score',
        'line_hit_rate_over_ou_5': 'Hit Rate (5g)',
        'line_games_played_5': 'Games (5g)',
        'line_games_played_10': 'Games (10g)',
        'stat_last5_avg': 'Last 5 Avg', 'stat_season_avg': 'Season Avg',
        'stat_last10_avg': 'Last 10 Avg',
        'stat_g1': 'G1', 'stat_g2': 'G2', 'stat_g3': 'G3',
        'stat_g4': 'G4', 'stat_g5': 'G5', 'stat_g6': 'G6',
        'stat_g7': 'G7', 'stat_g8': 'G8', 'stat_g9': 'G9', 'stat_g10': 'G10',
        'last5_over': 'L5 Over', 'last5_under': 'L5 Under',
        'line_hits_over_10': 'L10 Over', 'line_hits_under_10': 'L10 Under',
        'OVERALL_DEF_RANK': 'Def Rank', 'DEF_TIER': 'Def Tier',
        'minutes_tier': 'Min Tier', 'shot_role': 'Shot Role', 'usage_role': 'Usage Role',
        'usage_pct': 'Usage Pct', 'usage_tier': 'Usage Tier',
        'star_tier': 'Star Tier', 'is_franchise_star': 'Franchise Star',
        'usage_boost': 'Usage Boost', 'usage_boost_proj': 'Usage Boost Proj',
        'usage_boost_reason': 'Usage Boost Reason', 'usage_boost_source': 'Usage Boost Source',
        'team_star_out': 'Team Star Out', 'key_facilitator_out': 'Key Facilitator Out',
        'injury_boost_candidate': 'Injury Boost', 'usage_vacuum': 'Usage Vacuum',
        'void_reason': 'Void Reason',
        'consistency_grade': 'Consistency Grade',
        'team_top3_rank': 'Top3 Rank',
        'team_bottom3_rank': 'Bottom3 Rank',
        'def_boost_hist': 'Def Boost Hist',
        'top3_weak_overperformer': 'Top3 Weak Over',
        'top3_elite_fader': 'Top3 Elite Fade',
        'top3_def_context': 'Top3 Def Context',
        'top3_under_context': 'Top3 Under Context',
        **HIT_TRACKING_RENAME,
    }
    _lm_cols = (
        'open_line', 'line_movement', 'line_direction_shift',
        'implied_prob', 'implied_prob_over', 'implied_prob_under',
    )
    rename = {k: v for k, v in rename.items() if k not in _lm_cols}
    clean = clean.rename(columns=rename)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, 'ALL', clean)
    for tier in ['A', 'B', 'C', 'D']:
        subset = clean[clean['Tier'] == tier].copy()
        if len(subset):
            write_sheet(wb, f'Tier {tier}', subset)

    wb.save(xlsx_path)
    print(f"📊 Clean XLSX saved → {xlsx_path}")

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="step7_ranked_props.xlsx")
    ap.add_argument("--sheet", default="ALL")
    ap.add_argument("--output", default="step8_all_direction.csv")
    ap.add_argument("--xlsx", default="")  # optional override for xlsx path
    ap.add_argument(
        "--date",
        default="",
        help="Slate date YYYY-MM-DD (for dated snapshot copies; same as pipeline -Date)",
    )
    args = ap.parse_args()

    Path(args.output).parent.mkdir(parents=True, exist_ok=True)

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    df = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")

    out = df.copy()

    reconcile_signed_edge_abs_dataframe(out)
    edge = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = pd.to_numeric(out["abs_edge"], errors="coerce")

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    forced = pick_type.isin(["Goblin", "Demon"])

    model_dir = np.where(edge >= 0, "OVER", "UNDER")
    out["model_dir"] = model_dir

    final_dir = model_dir.copy()
    reason = np.where(abs_edge < 0.03, "MODEL_TIEBREAK_DIFF<0.03", "")
    final_dir = np.where(forced, "OVER", final_dir)
    reason = np.where(forced, "FORCED_OVER_ONLY_GOB_DEM", reason)

    out["final_bet_direction"] = final_dir
    out["final_dir_reason"] = reason

    slate_d = str(args.date or "").strip()[:10]
    if len(slate_d) < 10:
        slate_d = ""
    from_start = pd.Series([""] * len(out), index=out.index, dtype=object)
    if "start_time" in out.columns:
        et = _start_times_et(out["start_time"])
        from_start = et.dt.strftime("%Y-%m-%d").where(et.notna(), "").astype(str).str.strip()
    prev_gd = out.get("game_date", pd.Series([""] * len(out))).astype(str).str.strip().str[:10]
    prev_ok = prev_gd.str.match(r"^\d{4}-\d{2}-\d{2}$", na=False)
    # Prefer valid upstream game_date (step1 may anchor full boards to --date). Do not let
    # start_time ET overwrite it — that breaks combined_slate_tickets date filtering.
    merged = prev_gd.where(prev_ok, from_start.where(from_start.str.len() > 0, ""))
    merged = merged.where(merged.str.len() > 0, slate_d)
    out["game_date"] = merged.fillna("")

    out = _attach_distribution_std(out)
    filled_std = int(pd.to_numeric(out["distribution_std"], errors="coerce").notna().sum())
    print(f"[WNBA step8] distribution_std filled {filled_std}/{len(out)} rows")

    # Save full CSV
    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"✅ Saved → {args.output}")
    print("final_bet_direction counts:")
    print(pd.Series(out["final_bet_direction"]).value_counts().to_string())
    if "tier" in out.columns:
        print("tier counts:")
        print(out["tier"].value_counts().to_string())

    # Save clean XLSX
    xlsx_path = args.xlsx if args.xlsx else args.output.replace(".csv", "_clean.xlsx")
    Path(xlsx_path).parent.mkdir(parents=True, exist_ok=True)
    build_clean_xlsx(out, xlsx_path)
    _copy_dated_step8_wnba(xlsx_path, (args.date or "").strip())

if __name__ == "__main__":
    main()
