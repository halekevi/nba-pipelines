#!/usr/bin/env python3
"""
step8_add_direction_context_golf.py — Golf direction clean workbook for combined slate.

Reads step7_golf_ranked.xlsx, applies direction + hit-tracking columns, writes clean XLSX.

Run:
  py -3.14 Sports/Golf/scripts/step8_add_direction_context_golf.py \
      --input outputs/step7_golf_ranked.xlsx --date 2026-06-09
"""

from __future__ import annotations

import argparse
import datetime as _dt
import shutil
import sys
import zoneinfo
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

_REPO = Path(__file__).resolve().parent
for _ in range(10):
    if (_REPO / "utils" / "step8_edge_direction.py").is_file():
        if str(_REPO) not in sys.path:
            sys.path.insert(0, str(_REPO))
        break
    _REPO = _REPO.parent
else:
    raise RuntimeError("Could not locate repo root")

from scripts.l10_streak_utils import finalize_l10_ui_columns
from utils.hit_tracking_columns import HIT_TRACKING_RENAME, attach_hit_tracking_columns
from utils.step8_edge_direction import reconcile_signed_edge_abs_dataframe

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
HEADER_COLOR = "1C1C1C"


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t:
        return "Goblin"
    if "dem" in t:
        return "Demon"
    return "Standard"


def write_sheet(wb: Workbook, name: str, df: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    headers = list(df.columns)
    ws.append(headers)
    hdr_fill = PatternFill("solid", start_color=HEADER_COLOR)
    hdr_font = Font(bold=True, color="FFFFFF", name="Arial", size=10)
    thin = Side(style="thin", color="CCCCCC")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    for ci, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.fill = hdr_fill
        c.font = hdr_font
        c.alignment = Alignment(horizontal="center", vertical="center")
        c.border = border
    tier_col = headers.index("Tier") + 1 if "Tier" in headers else None
    for ri, row in enumerate(df.itertuples(index=False), 2):
        for ci, val in enumerate(row, 1):
            cell = ws.cell(row=ri, column=ci, value=val)
            cell.border = border
            cell.font = Font(name="Arial", size=9)
            if tier_col and ci == tier_col:
                tier = str(val or "").strip().upper()
                bg, fg = TIER_COLORS.get(tier, ("FFFFFF", "000000"))
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, color=fg, name="Arial", size=9)
    col_widths = {
        "Player": 22, "Tier": 6, "Rank Score": 10, "Event": 24, "Tournament": 22,
        "Prop": 18, "Pick Type": 10, "Line": 7, "Direction": 9,
        "Hit Rate": 10, "Hit Rate L5": 10, "Hit Rate L10": 10,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)
    ws.freeze_panes = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    try:
        import platform

        _time_fmt = "%m/%d %#I:%M %p" if platform.system() == "Windows" else "%m/%d %-I:%M %p"
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime(_time_fmt)
    except Exception:
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%m/%d %H:%M")

    hr5 = pd.to_numeric(df2.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")
    hr10 = pd.to_numeric(df2.get("line_hit_rate_over_ou_10", np.nan), errors="coerce").fillna(hr5).fillna(0.50)
    df2["line_hit_rate_over_ou_10"] = hr10
    proj = pd.to_numeric(df2.get("projection", np.nan), errors="coerce")
    df2["stat_last5_avg"] = pd.to_numeric(df2.get("stat_last5_avg", np.nan), errors="coerce").fillna(proj)
    df2["stat_season_avg"] = pd.to_numeric(df2.get("stat_season_avg", np.nan), errors="coerce").fillna(df2["stat_last5_avg"])

    l5_over = pd.to_numeric(df2.get("last5_over", np.nan), errors="coerce")
    l5_under = pd.to_numeric(df2.get("last5_under", np.nan), errors="coerce")
    l5_over_fallback = (hr5.fillna(0.5) * 5.0).round().clip(0, 5)
    l5_under_fallback = (5 - l5_over_fallback).clip(0, 5)
    df2["last5_over"] = l5_over.fillna(l5_over_fallback)
    df2["last5_under"] = l5_under.fillna(l5_under_fallback)

    if "line" in df2.columns:
        df2 = finalize_l10_ui_columns(df2, line_col="line")
    df2 = attach_hit_tracking_columns(df2, "GOLF")

    if "DEF_TIER" not in df2.columns:
        df2["DEF_TIER"] = "LEAGUE AVG"
    if "OVERALL_DEF_RANK" not in df2.columns:
        df2["OVERALL_DEF_RANK"] = "N/A"

    keep = [
        "tier", "rank_score",
        "player", "pos", "team", "event", "tournament", "course", "opp_team", "league", "game_time",
        "prop_type", "pick_type", "line",
        "final_bet_direction",
        "edge", "abs_edge", "projection",
        "ml_prob",
        "hit_rate", "hit_rate_l5", "hit_rate_l10",
        "strat_hit_rate", "strat_n",
        "player_hr_historical", "opp_hr_historical",
        "sport_signal_maturity", "confidence_tier", "confidence_score", "confidence_note",
        "line_hit_rate_over_ou_5", "line_hit_rate_over_ou_10",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "l10_over", "l10_under", "l10_over_pct", "l10_streak", "l10_games_played",
        "DEF_TIER", "OVERALL_DEF_RANK",
        "course_fit_score", "sg_ott", "sg_app", "sg_arg", "weather_signal",
        "void_reason",
    ]
    for c in keep:
        if c not in df2.columns:
            df2[c] = np.nan
    clean = df2[keep].copy()

    for col in ["rank_score", "edge", "abs_edge", "projection", "ml_prob", "line_hit_rate_over_ou_5", "line_hit_rate_over_ou_10"]:
        if col in clean.columns:
            rnd = 4 if col == "ml_prob" else 2
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(rnd)
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}
    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "player": "Player", "pos": "Pos",
        "team": "Team", "event": "Event", "tournament": "Tournament", "course": "Course",
        "opp_team": "Course/Opp", "league": "League", "game_time": "Game Time",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "final_bet_direction": "Direction",
        "edge": "Edge", "abs_edge": "Abs Edge", "projection": "Projection",
        "ml_prob": "ML Prob",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "line_hit_rate_over_ou_10": "Hit Rate (10g)",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier",
        "course_fit_score": "Course Fit",
        "sg_ott": "SG OTT", "sg_app": "SG APP", "sg_arg": "SG ARG",
        "weather_signal": "Weather",
        "void_reason": "Void Reason",
        **HIT_TRACKING_RENAME,
    }
    clean = clean.rename(columns=rename)

    void_col = "Void Reason" if "Void Reason" in clean.columns else None
    if void_col:
        clean_eligible = clean[clean[void_col].isna() | (clean[void_col] == "")].copy()
    else:
        clean_eligible = clean.copy()

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "Golf", clean)
    write_sheet(wb, "ALL", clean)
    for tier in ["A", "B", "C", "D"]:
        subset = clean_eligible[clean_eligible["Tier"] == tier].copy()
        write_sheet(wb, f"Tier {tier}", subset if len(subset) else clean_eligible.head(0))

    wb.save(xlsx_path)
    print(f"[Golf step8] Clean XLSX saved → {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="outputs/step7_golf_ranked.xlsx")
    ap.add_argument("--sheet", default="ALL")
    ap.add_argument("--output", default="outputs/step8_golf_direction.csv")
    ap.add_argument("--xlsx", default="outputs/step8_golf_direction_clean.xlsx")
    ap.add_argument("--date", default="", help="YYYY-MM-DD target date (ET)")
    args = ap.parse_args()

    root = Path(__file__).resolve().parent.parent
    inp = Path(args.input)
    if not inp.is_absolute():
        inp = root / inp

    print("[Golf step8] Starting...")
    df = pd.read_excel(inp, sheet_name=args.sheet, dtype=str).fillna("")
    if df.empty:
        print("ERROR [Golf-S8] Empty input — aborting.")
        sys.exit(1)

    eastern = zoneinfo.ZoneInfo("America/New_York")
    target_str = (
        args.date.strip()[:10]
        if args.date
        else _dt.datetime.now(tz=eastern).date().isoformat()
    )
    if "start_time" in df.columns:
        before_filter = len(df)
        et_dates = pd.to_datetime(df["start_time"], utc=True, errors="coerce").dt.tz_convert(eastern)
        df["_et_date"] = et_dates.dt.date.apply(
            lambda d: d.isoformat() if isinstance(d, _dt.date) else ""
        )
        mask = df["_et_date"] == target_str
        if not mask.any():
            valid = df["_et_date"].astype(str).str.match(r"^\d{4}-\d{2}-\d{2}$")
            avail = sorted(df.loc[valid, "_et_date"].unique().tolist())
            past_or_equal = [d for d in avail if d <= target_str]
            if past_or_equal:
                fallback_date = past_or_equal[-1]
                print(f"[DateFilter] No exact ET match for {target_str} — fallback {fallback_date}")
                mask = df["_et_date"] == fallback_date
            elif avail:
                fallback_date = avail[0]
                print(f"[DateFilter] Using earliest available {fallback_date}")
                mask = df["_et_date"] == fallback_date
            else:
                mask = pd.Series(True, index=df.index)
        df = df.loc[mask].drop(columns="_et_date")
        print(f"[DateFilter] Kept {len(df)}/{before_filter} rows for {target_str} ET")
    else:
        print("[DateFilter] WARNING: no start_time — skipping date filter")

    out = df.copy()
    reconcile_signed_edge_abs_dataframe(out)

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    edge = pd.to_numeric(out.get("edge", np.nan), errors="coerce")
    has_edge = edge.notna()
    final_dir = pd.Series(
        np.where(has_edge & (edge >= 0), "OVER", np.where(has_edge & (edge < 0), "UNDER", "OVER")),
        index=out.index,
    )
    forced = pick_type.isin(["Goblin", "Demon"])
    final_dir = final_dir.where(~forced, "OVER")
    out["final_bet_direction"] = final_dir
    out["direction"] = final_dir
    out["bet_direction"] = final_dir
    out["void_reason"] = ""

    hr5b = pd.to_numeric(out.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")
    hr10b = pd.to_numeric(out.get("line_hit_rate_over_ou_10", np.nan), errors="coerce").fillna(hr5b)
    out["composite_hit_rate"] = (0.5 * hr5b.fillna(0.5) + 0.5 * hr10b.fillna(0.5)).clip(0.0, 1.0)
    mpb = pd.to_numeric(out.get("ml_prob", np.nan), errors="coerce").fillna(0.5)
    out["blended_score"] = (0.3 * mpb + 0.7 * out["composite_hit_rate"]).round(4)

    out = attach_hit_tracking_columns(out, "GOLF")

    csv_path = Path(args.output)
    if not csv_path.is_absolute():
        csv_path = root / csv_path
    csv_path.parent.mkdir(parents=True, exist_ok=True)
    out.to_csv(csv_path, index=False, encoding="utf-8-sig")
    print(f"[Golf step8] Saved → {csv_path}")

    xlsx_path = Path(args.xlsx)
    if not xlsx_path.is_absolute():
        xlsx_path = root / xlsx_path
    build_clean_xlsx(out, str(xlsx_path))

    try:
        repo_root = Path(__file__).resolve().parents[3]
        dated_dir = repo_root / "outputs" / target_str / "golf"
        dated_dir.mkdir(parents=True, exist_ok=True)
        dated_xlsx = dated_dir / f"step8_golf_direction_clean_{target_str}.xlsx"
        if xlsx_path.is_file():
            shutil.copy2(xlsx_path, dated_xlsx)
            print(f"[Golf step8] Dated clean workbook → {dated_xlsx}")
    except Exception as e:
        print(f"[Golf step8] WARN dated copy skipped: {e}")


if __name__ == "__main__":
    main()
