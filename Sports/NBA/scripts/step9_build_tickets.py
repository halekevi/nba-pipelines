#!/usr/bin/env python3
"""
step9_build_tickets.py
-----------------------
Reads step8_all_direction_clean.xlsx (Tier A sheet) and builds
optimized PrizePicks tickets prioritizing 3-leg tickets for best risk/reward.

Ticket building rules:
- Only Tier A props with Hit Rate >= min_hit_rate
- One player per game (no correlated legs)
- One prop per player per ticket
- Deduplicated so same leg doesn't appear in multiple tickets of same type
- Tickets ranked by average rank score

PrizePicks payouts (approximate):
  2-leg: 3x  |  3-leg: 5x  |  4-leg: 10x  |  5-leg: 20x

Run:
  py -3.14 step9_build_tickets.py --input step8_all_direction_clean.xlsx --output best_tickets.xlsx
"""

from __future__ import annotations

import argparse
import pandas as pd
import numpy as np
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# PrizePicks Power Play payouts by leg count and pick type
# Standard = full odds | Goblin/Demon = reduced odds
# ── Payout Tables ─────────────────────────────────────────────────────────────
# Base Power Play payouts (all correct) — Standard lines only
POWER_PLAY_BASE = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 37.5}

# Base Flex Play payouts — Standard lines only
FLEX_PLAY_BASE = {
    2: {2: 3.0},
    3: {3: 3.0,  2: 0.5},  # 2/3 correct = 0.5x (confirmed 2026-02-20)
    4: {4: 6.0,  3: 1.5},
    5: {5: 10.0, 4: 2.0,  3: 0.4},
    6: {6: 25.0, 5: 2.0,  4: 0.4},
}

# ── Goblin Power Play multipliers (CONFIRMED via live testing 2026-02-20) ──────
# 2-leg confirmed: T1=2.7x (mod=0.900), T2=2.4x (mod=0.800), T3=2.3x (mod=0.767), T4+=2.0x (mod=0.667)
# 3-leg confirmed: 1 goblin T1 + 2 std = 3.75x, 2 goblin T1 + 1 std = 3.50x
GOBLIN_POWER_MOD = {1: 0.900, 2: 0.800, 3: 0.767, 4: 0.667}
GOBLIN_FLEX_MOD  = {1: 0.833, 2: 0.750, 3: 0.700, 4: 0.600}

# ── Demon Power Play multipliers (CONFIRMED via live testing 2026-02-20) ───────
# 2-leg confirmed: T1=4.0x (mod=1.333), T2=5.5x (mod=1.833), T3=11.5x (mod=3.833)
DEMON_POWER_MOD  = {1: 1.333, 2: 1.833, 3: 3.833}
DEMON_FLEX_MOD   = {1: 1.200, 2: 1.600, 3: 3.000}

def calc_ticket_payout(ticket: list, n_legs: int, play_type: str = 'power') -> dict:
    """Calculate deviation-aware payout for a mixed ticket."""
    base_top = POWER_PLAY_BASE.get(n_legs, 37.5) if play_type == 'power' else FLEX_PLAY_BASE.get(n_legs, {}).get(n_legs, 25.0)
    power_mod = 1.0
    flex_mod  = 1.0
    for row in ticket:
        pt = str(row.get('Pick Type', 'Standard')).strip().lower()
        dev = int(row.get('deviation_level', row.get('Deviation Level', 1)) or 1)
        dev = max(1, min(dev, 4))  # clamp to 1-4
        if 'gob' in pt:
            power_mod *= GOBLIN_POWER_MOD.get(dev, 0.840)
            flex_mod  *= GOBLIN_FLEX_MOD.get(dev, 0.800)
        elif 'dem' in pt:
            power_mod *= DEMON_POWER_MOD.get(dev, 1.627)
            flex_mod  *= DEMON_FLEX_MOD.get(dev, 1.600)
    if play_type == 'power':
        top = round(base_top * power_mod, 2)
        stake = round(100 / top, 2) if top > 0 else 0
        return {'top_payout': top, 'stake_to_win_100': stake, 'play_type': 'Power Play'}
    else:
        flex_base = FLEX_PLAY_BASE.get(n_legs, {})
        partials = {}
        for n_correct, mult in flex_base.items():
            partials[n_correct] = round(mult * flex_mod if n_correct == n_legs else mult, 2)
        top = partials.get(n_legs, 25.0)
        stake = round(100 / top, 2) if top > 0 else 0
        return {'top_payout': top, 'stake_to_win_100': stake, 'play_type': 'Flex Play', 'partials': partials}

def get_payout(pick_label: str, n_legs: int) -> float:
    return POWER_PLAY_BASE.get(n_legs, 37.5)

def stake_to_win(target: float, pick_label: str, n_legs: int) -> float:
    return round(target / POWER_PLAY_BASE.get(n_legs, 37.5), 2)

COLORS = {
    'Standard': '2874A6',
    'Goblin':   '6C3483',
    'Demon':    'C0392B',
    'Best Mix': '1E8449',
}
HDR_COLOR  = '1C1C1C'
DIR_OVER   = 'C8F7C5'
DIR_UNDER  = 'F7C5C5'
TIER_COLORS = {
    'Elite':     'D5F5E3',
    'Above Avg': 'EBF5FB',
    'Avg':       'FDFEFE',
    'Weak':      'FDEDEC',
}

def thin_border():
    s = Side(style='thin', color='CCCCCC')
    return Border(left=s, right=s, top=s, bottom=s)

def _norm_pick_type(x):
    t = str(x).strip().lower()
    if 'gob' in t: return 'Goblin'
    if 'dem' in t: return 'Demon'
    return 'Standard'

def _game_key(row) -> frozenset:
    """Canonical game identifier (team + opp, order-independent)."""
    return frozenset([str(row['Team']), str(row['Opp'])])


def _ticket_valid(combo) -> bool:
    """A ticket is valid if: no duplicate players, one player per game."""
    players = [r['Player'] for r in combo]
    if len(players) != len(set(players)):
        return False
    games = [_game_key(r) for r in combo]
    if len(games) != len(set(games)):
        return False
    return True


def _ticket_score(combo) -> float:
    """Average rank score — used to rank valid tickets."""
    scores = []
    for r in combo:
        v = pd.to_numeric(pd.Series([r['Rank Score']]), errors='coerce').iloc[0]
        if not pd.isna(v):
            scores.append(float(v))
    return float(np.mean(scores)) if scores else 0.0


def build_tickets(pool: pd.DataFrame, n_legs: int, max_tickets: int,
                  used_legs: set, min_hit_rate: float = 0.0) -> list:
    """
    Combinatorial ticket builder (replaces greedy anchor-fill).

    - Searches all C(N, n_legs) combinations of the top-K candidates
    - Filters: one player per game, no duplicate players, hit rate >= min_hit_rate
    - Ranks by average rank_score → picks top max_tickets
    - Tracks used_legs so legs don't repeat across ticket groups

    K is capped at 25 props so C(25,4) = 12,650 combos — fast enough (~10ms).
    """
    from itertools import combinations

    # Apply real hit-rate floor (this was the bug — was hardcoded >= 0.0)
    if min_hit_rate > 0.0:
        pool = pool[pd.to_numeric(pool['Hit Rate (5g)'], errors='coerce') >= min_hit_rate].copy()

    pool = pool.sort_values('Rank Score', ascending=False).copy()

    # Cap candidate pool so combinations stay fast
    K = min(len(pool), 25)
    candidates = [row for _, row in pool.head(K).iterrows()]

    valid_tickets = []
    seen_keys: set = set()

    for combo in combinations(candidates, n_legs):
        # Skip if any leg already used in another ticket group
        leg_keys = [(r['Player'], r['Prop'], r['Line']) for r in combo]
        if any(lk in used_legs for lk in leg_keys):
            continue
        # Validate game diversity + player uniqueness
        if not _ticket_valid(combo):
            continue
        tkey = frozenset((r['Player'], r['Prop'], str(r['Line'])) for r in combo)
        if tkey in seen_keys:
            continue
        seen_keys.add(tkey)
        valid_tickets.append((combo, _ticket_score(combo)))

    # Best tickets first
    valid_tickets.sort(key=lambda x: x[1], reverse=True)

    # Pick top max_tickets, mark legs used
    tickets = []
    for combo, _ in valid_tickets:
        if len(tickets) >= max_tickets:
            break
        # Re-check used_legs (another ticket in this batch may have claimed a leg)
        leg_keys = [(r['Player'], r['Prop'], r['Line']) for r in combo]
        if any(lk in used_legs for lk in leg_keys):
            continue
        tickets.append(list(combo))
        for lk in leg_keys:
            used_legs.add(lk)

    return tickets


def write_tickets_sheet(wb, sheet_name: str, tickets: list, n_legs: int, pick_label: str):
    if not tickets:
        return

    ws = wb.create_sheet(sheet_name)
    tab_color = COLORS.get(pick_label, '1E8449')
    payout = get_payout(pick_label, n_legs)

    cols    = ['#', 'Player', 'Team', 'Opp', 'Prop', 'Pick Type', 'Line',
               'Direction', 'Edge', 'Hit Rate', 'L5 Avg', 'Season Avg',
               'L5 Over', 'L5 Under', 'Rank Score', 'Def Tier']
    col_w   = [4,   22,      6,      6,     18,     10,          7,
               10,          7,      10,       9,          11,
               8,        8,          11,       11]

    # Set column widths once
    for ci, w in enumerate(col_w, 1):
        ws.column_dimensions[get_column_letter(ci)].width = w

    current_row = 1

    for t_idx, ticket in enumerate(tickets, 1):
        avg_score = sum(float(r['Rank Score']) for r in ticket) / len(ticket)
        avg_hr    = sum(float(r['Hit Rate (5g)']) for r in ticket) / len(ticket)
        ticket_dicts = [dict(r) for r in ticket]
        pp_info   = calc_ticket_payout(ticket_dicts, n_legs, 'power')
        flex_info = calc_ticket_payout(ticket_dicts, n_legs, 'flex')

        # Ticket header banner
        ws.merge_cells(start_row=current_row, start_column=1,
                       end_row=current_row, end_column=len(cols))
        hcell = ws.cell(
            row=current_row, column=1,
            value=(f"  Ticket #{t_idx}  ·  {n_legs}-Leg {pick_label}"
                   f"  ·  Power: {pp_info['top_payout']}x (${pp_info['stake_to_win_100']:.0f} to win $100)"
                   f"  ·  Flex: {flex_info['top_payout']}x"
                   f"  ·  Avg Hit Rate: {avg_hr:.0%}"
                   f"  ·  Est. Win Prob: {avg_hr**n_legs:.0%}"
                   f"  ·  Avg Rank Score: {avg_score:.2f}")
        )
        hcell.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        hcell.fill      = PatternFill('solid', start_color=tab_color)
        hcell.alignment = Alignment(vertical='center', horizontal='left')
        ws.row_dimensions[current_row].height = 22
        current_row += 1

        # Column headers
        for ci, h in enumerate(cols, 1):
            c = ws.cell(row=current_row, column=ci, value=h)
            c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=9)
            c.fill      = PatternFill('solid', start_color=HDR_COLOR)
            c.alignment = Alignment(horizontal='center', vertical='center')
            c.border    = thin_border()
        ws.row_dimensions[current_row].height = 18
        current_row += 1

        # Legs
        for leg_i, row in enumerate(ticket, 1):
            row_bg  = 'F8F9FA' if leg_i % 2 == 0 else 'FFFFFF'
            def_bg  = TIER_COLORS.get(str(row.get('Def Tier', '')), 'FFFFFF')

            l5o = row.get('L5 Over',  '')
            l5u = row.get('L5 Under', '')
            l5o = '' if pd.isna(l5o) else int(l5o)
            l5u = '' if pd.isna(l5u) else int(l5u)

            vals = [
                leg_i,
                row['Player'],
                row['Team'],
                row['Opp'],
                row['Prop'],
                row['Pick Type'],
                row['Line'],
                row['Direction'],
                round(float(row['Edge']), 1),
                row['Hit Rate (5g)'],
                round(float(row['Last 5 Avg']), 1),
                round(float(row['Season Avg']), 1),
                l5o, l5u,
                round(float(row['Rank Score']), 2),
                row['Def Tier'],
            ]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=current_row, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center')
                c.border    = thin_border()

                col_name = cols[ci - 1]
                if col_name == 'Direction':
                    bg = DIR_OVER if val == 'OVER' else DIR_UNDER
                    c.fill = PatternFill('solid', start_color=bg)
                    c.font = Font(bold=True, name='Arial', size=9)
                elif col_name == 'Def Tier':
                    c.fill = PatternFill('solid', start_color=def_bg)
                else:
                    c.fill = PatternFill('solid', start_color=row_bg)

            current_row += 1

        current_row += 1  # spacer


def write_summary_sheet(wb, all_ticket_sets: list):
    """Overview sheet showing all tickets at a glance."""
    ws = wb.create_sheet('SUMMARY', 0)

    headers = ['Sheet', 'Ticket #', 'Legs', 'Pick Type', 'Power Payout',
               '$ to win $100', 'Flex Payout', 'Avg Hit Rate', 'Est Win %', 'Avg Rank Score', 'Players']
    col_w   = [20,      10,       6,      12,   13,
               13,             12,           13,            12,          15,              50]

    for ci, (h, w) in enumerate(zip(headers, col_w), 1):
        c = ws.cell(row=1, column=ci, value=h)
        c.font      = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        c.fill      = PatternFill('solid', start_color=HDR_COLOR)
        c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        c.border    = thin_border()
        ws.column_dimensions[get_column_letter(ci)].width = w
    ws.row_dimensions[1].height = 22
    ws.freeze_panes = 'A2'

    row = 2
    for sheet_name, tickets, n_legs, pick_label in all_ticket_sets:
        tab_color = COLORS.get(pick_label, '1E8449')
        for t_idx, ticket in enumerate(tickets, 1):
            avg_score    = sum(float(r['Rank Score']) for r in ticket) / len(ticket)
            avg_hr       = sum(float(r['Hit Rate (5g)']) for r in ticket) / len(ticket)
            ticket_dicts = [dict(r) for r in ticket]
            pp_info      = calc_ticket_payout(ticket_dicts, n_legs, 'power')
            flex_info    = calc_ticket_payout(ticket_dicts, n_legs, 'flex')
            players      = ' | '.join(f"{r['Player']} {r['Direction']} {r['Prop']} {r['Line']}" for r in ticket)

            vals = [sheet_name, t_idx, n_legs, pick_label,
                    f"{pp_info['top_payout']}x", f"${pp_info['stake_to_win_100']:.0f}",
                    f"{flex_info['top_payout']}x",
                    f"{avg_hr:.0%}", f"{avg_hr**n_legs:.0%}",
                    round(avg_score, 2), players]

            for ci, val in enumerate(vals, 1):
                c = ws.cell(row=row, column=ci, value=val)
                c.font      = Font(name='Arial', size=9)
                c.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                c.border    = thin_border()
                bg = 'F8F9FA' if row % 2 == 0 else 'FFFFFF'
                c.fill = PatternFill('solid', start_color=bg)
            row += 1

    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument('--input',       default='step8_all_direction_clean.xlsx')
    ap.add_argument('--sheet',       default='Tier A')
    ap.add_argument('--output',      default='best_tickets.xlsx')
    ap.add_argument('--min_hit_rate',type=float, default=0.6)
    ap.add_argument('--max_tickets', type=int,   default=10)
    ap.add_argument('--legs',        default='2,3,4',
                    help='Comma-separated leg counts to generate, e.g. 2,3,4,5')
    args = ap.parse_args()

    leg_counts = [int(x.strip()) for x in args.legs.split(',')]

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    df = pd.read_excel(args.input, sheet_name=args.sheet)
    df['Pick Type'] = df['Pick Type'].astype(str).apply(_norm_pick_type)

    # Filter by min hit rate
    df = df[pd.to_numeric(df['Hit Rate (5g)'], errors='coerce') >= args.min_hit_rate].copy()
    print(f"  Props with hit rate >= {args.min_hit_rate}: {len(df)}")

    # Deduplicate - best rank score per player+prop+line
    df = df.sort_values('Rank Score', ascending=False).drop_duplicates(subset=['Player','Prop','Line','Direction'])

    # Separate pools
    std_pool = df[df['Pick Type']=='Standard'].drop_duplicates(subset=['Player'])
    gob_pool = df[df['Pick Type']=='Goblin'].drop_duplicates(subset=['Player'])
    all_pool = df.sort_values('Rank Score', ascending=False).drop_duplicates(subset=['Player'])

    pools = [
        ('Standard', std_pool),
        ('Goblin',   gob_pool),
        ('Best Mix', all_pool),
    ]

    wb = Workbook()
    wb.remove(wb.active)

    all_ticket_sets = []

    for pick_label, pool in pools:
        if len(pool) < 2:
            print(f"  ⚠ Skipping {pick_label} — not enough props in pool ({len(pool)})")
            continue
        # Fresh used_legs per pool type so legs can appear across Standard/Goblin/Mix
        used_legs: set = set()
        for n_legs in leg_counts:
            if len(pool) < n_legs:
                continue
            tickets = build_tickets(pool.copy(), n_legs, args.max_tickets, used_legs, args.min_hit_rate)
            sheet_name = f"{pick_label} {n_legs}-Leg"
            write_tickets_sheet(wb, sheet_name, tickets, n_legs, pick_label)
            all_ticket_sets.append((sheet_name, tickets, n_legs, pick_label))
            print(f"  {sheet_name}: {len(tickets)} tickets")

    # Standard-first Goblin fill: anchor with best Standard, fill remaining with best Goblin
    if len(std_pool) >= 1 and len(gob_pool) >= 1:
        std_fill_pool = pd.concat([
            std_pool.assign(_pref=0),
            gob_pool.assign(_pref=1),
        ]).sort_values(['_pref', 'Rank Score'], ascending=[True, False]).drop('_pref', axis=1)

        used_legs_fill: set = set()
        for n_legs in leg_counts:
            if n_legs < 3: continue  # 2-leg already covered above
            if len(std_fill_pool) < n_legs: continue
            tickets = build_tickets(std_fill_pool.copy(), n_legs, args.max_tickets, used_legs_fill, args.min_hit_rate)
            if tickets:
                sheet_name = f"Std+Gob Fill {n_legs}-Leg"
                write_tickets_sheet(wb, sheet_name, tickets, n_legs, 'Best Mix')
                all_ticket_sets.append((sheet_name, tickets, n_legs, 'Best Mix'))
                print(f"  {sheet_name}: {len(tickets)} tickets (Standard-first, Goblin fill)")

    write_summary_sheet(wb, all_ticket_sets)

    wb.save(args.output)
    print(f"\n✅ Saved → {args.output}")
    print(f"\nPrizePicks Power Play payout reference:")
    print(f"  {'Legs':<6} {'Power':>8} {'Flex':>8}  {'$ to win $100 (Power)':>22}")
    for n in [2, 3, 4, 5, 6]:
        pp    = POWER_PLAY_BASE.get(n, '-')
        fl    = FLEX_PLAY_BASE.get(n, {}).get(n, '-')
        stake = round(100 / pp, 2) if isinstance(pp, float) else '-'
        print(f"  {n}-leg  {str(pp)+chr(120):>8} {str(fl)+chr(120):>8}  ${stake:>6}")

if __name__ == '__main__':
    main()
