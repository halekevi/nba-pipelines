#!/usr/bin/env python3
"""
step8_add_direction_context.py
-------------------------------
Reads step7_ranked_props.xlsx, appends direction context columns,
and outputs both a full CSV and a clean formatted XLSX for decision making.

Adds:
- model_dir
- final_bet_direction
- final_dir_reason

Run:
  py -3.14 step8_add_direction_context.py --input step7_ranked_props.xlsx --sheet ALL --output step8_all_direction.csv
"""

from __future__ import annotations

import argparse
import shutil
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
import os
from datetime import date, datetime
from pathlib import Path
from zoneinfo import ZoneInfo

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass

_repo = Path(__file__).resolve().parent
for _ in range(10):
    if (_repo / "utils" / "step8_edge_direction.py").is_file():
        if str(_repo) not in sys.path:
            sys.path.insert(0, str(_repo))
        break
    _repo = _repo.parent
else:
    raise RuntimeError("Could not locate repo root with utils/step8_edge_direction.py")

from utils.step8_edge_direction import reconcile_signed_edge_abs_dataframe

_REPO_ROOT = Path(__file__).resolve().parents[3]
if str(_REPO_ROOT) not in sys.path:
    sys.path.insert(0, str(_REPO_ROOT))
from scripts.l10_streak_utils import finalize_l10_ui_columns
from nba_enrichment_carry import (
    ENRICHMENT_CARRY_COLS,
    is_nba1h_pipeline,
    reattach_enrichment_carry,
    snapshot_enrichment_carry,
)

_ET = ZoneInfo("America/New_York")


def _start_times_et(s: pd.Series) -> pd.Series:
    return pd.to_datetime(s, utc=True, errors="coerce").dt.tz_convert(_ET)


def _format_et_clock(et: pd.Series) -> pd.Series:
    clk = et.dt.strftime("%I:%M %p")
    return clk.str.replace(r"^0(\d:)", r"\1", regex=True)


def _copy_dated_step8_nba(output_xlsx_path: str, slate_date: str) -> None:
    """Copy clean XLSX to outputs/<slate>/step8_*_direction_clean_<slate>.xlsx.

    slate_date must be YYYY-MM-DD (pipeline -Date). Falls back to today only when empty.
    """
    src = Path(output_xlsx_path)
    if not src.is_file():
        return
    d = (slate_date or "").strip()
    if not d:
        d = date.today().isoformat()
    repo_root = Path(__file__).resolve().parents[3]
    dated_dir = repo_root / "outputs" / d
    src_name = src.name.lower()
    if "nba1q" in src_name:
        dated_name = f"step8_nba1q_direction_clean_{d}.xlsx"
    elif "nba1h" in src_name:
        dated_name = f"step8_nba1h_direction_clean_{d}.xlsx"
    else:
        dated_name = f"step8_nba_direction_clean_{d}.xlsx"
    try:
        dated_dir.mkdir(parents=True, exist_ok=True)
        dated_path = dated_dir / dated_name
        shutil.copy2(src, dated_path)
        print(f"[NBA step8] Dated copy -> {dated_path}")
    except Exception as e:
        print(f"[NBA step8] WARN: dated copy failed: {e}")


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t:
        return "Goblin"
    if "dem" in t:
        return "Demon"
    return "Standard"

TIER_COLORS = {
    'A': ('1E8449', 'FFFFFF'),
    'B': ('2874A6', 'FFFFFF'),
    'C': ('D4AC0D', '000000'),
    'D': ('717D7E', 'FFFFFF'),
}
HEADER_COLOR = '1C1C1C'

def thin_border():
    s = Side(style='thin', color='DDDDDD')
    return Border(left=s, right=s, top=s, bottom=s)

def write_sheet(wb, name, data):
    ws = wb.create_sheet(name)
    tier_bg, tier_fg = TIER_COLORS.get(name, ('333333', 'FFFFFF'))
    headers = list(data.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font = Font(bold=True, color='FFFFFF', name='Arial', size=10)
        cell.fill = PatternFill('solid', start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        cell.border = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        direction = row[headers.index('Direction')] if 'Direction' in headers else ''
        for ci, val in enumerate(row, 1):
            col_name = headers[ci - 1]
            display_val = '' if pd.isna(val) else val
            cell = ws.cell(row=ri, column=ci, value=display_val)
            cell.font = Font(name='Arial', size=9)
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border()

            if col_name == 'Tier':
                cell.fill = PatternFill('solid', start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name='Arial', size=9)
            elif col_name == 'Direction':
                bg = 'C8F7C5' if val == 'OVER' else 'F7C5C5' if val == 'UNDER' else 'FFFFFF'
                cell.fill = PatternFill('solid', start_color=bg)
                cell.font = Font(bold=True, name='Arial', size=9)
            else:
                cell.fill = PatternFill('solid', start_color='F9F9F9' if ri % 2 == 0 else 'FFFFFF')

    col_widths = {
        'Tier': 6, 'Rank Score': 10, 'Player': 18, 'Pos': 6,
        'Team': 6, 'Opp': 6, 'Game Date': 12, 'Game Time': 10,
        'Prop': 16, 'Pick Type': 10, 'Line': 7, 'Standard Line': 12,
        'Direction': 9, 'Edge': 7, 'Abs Edge': 8, 'Projection': 10,
        'Hit Rate (5g)': 12, 'Last 5 Avg': 10, 'Season Avg': 10,
        'L5 Over': 8, 'L5 Under': 8,
        'Def Rank': 9, 'Def Tier': 10,
        'Min Tier': 9, 'Shot Role': 10, 'Usage Role': 10,
        'ML Prob': 9,
        'Void Reason': 20,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes = 'A2'
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"

def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str, source_hint: str = ""):
    df2 = df.copy()
    if "minutes_tier" in df2.columns:
        df2["minutes_tier"] = df2["minutes_tier"].astype(object)
    # step8 main() loads step7 with dtype=str; coerce game-log / L10 numerics before math.
    for col in (
        "line",
        "last5_over",
        "last5_under",
        "l10_over",
        "l10_under",
        "l10_over_pct",
        "line_hits_over_10",
        "line_hits_under_10",
    ):
        if col in df2.columns:
            df2[col] = pd.to_numeric(df2[col], errors="coerce")
    for i in range(1, 11):
        gcol = f"stat_g{i}"
        if gcol in df2.columns:
            df2[gcol] = pd.to_numeric(df2[gcol], errors="coerce")
    # Prefer pipeline minutes labels over ML numeric buckets (0–3).
    if "minutes_tier_label" in df2.columns:
        lbl = df2["minutes_tier_label"].astype(str).str.strip()
        ok = lbl.ne("") & ~lbl.str.lower().isin(["nan", "none", "null"])
        if ok.any():
            df2.loc[ok, "minutes_tier"] = lbl[ok]
    else:
        mt_num = pd.to_numeric(df2.get("minutes_tier", ""), errors="coerce")
        _min_tier_map = {0: "Low", 1: "Med", 2: "High", 3: "Elite"}
        m_ok = mt_num.notna()
        if m_ok.any():
            keys = mt_num.loc[m_ok].round().astype("Int64")
            df2.loc[m_ok, "minutes_tier"] = keys.map(_min_tier_map).astype(str)

    if "abs_edge" not in df2.columns and "edge" in df2.columns:
        df2["abs_edge"] = pd.to_numeric(df2["edge"], errors="coerce").abs()

    if "start_time" in df2.columns:
        et = _start_times_et(df2["start_time"])
        df2["game_date"] = et.dt.strftime("%Y-%m-%d").where(et.notna(), "")
        df2["game_time"] = _format_et_clock(et)
    else:
        if "game_date" not in df2.columns:
            df2["game_date"] = ""
        if "game_time" not in df2.columns:
            df2["game_time"] = ""
    hint = (source_hint or "").lower()
    is_period_slate = ("nba1q" in hint) or ("nba1h" in hint)

    def _num(s: pd.Series) -> pd.Series:
        return pd.to_numeric(s, errors='coerce')

    def _blank_mask(s: pd.Series) -> pd.Series:
        t = s.astype(str).str.strip().str.lower()
        return s.isna() | t.isin(["", "nan", "none", "null", "nat"])

    def _fill_from(col: str, *fallback_cols: str):
        if col not in df2.columns:
            df2[col] = np.nan
        # Avoid dtype setitem issues when fallback source is string-typed.
        df2[col] = df2[col].astype(object)
        base = df2[col]
        miss = _blank_mask(base)
        for fc in fallback_cols:
            if fc not in df2.columns:
                continue
            cand = df2[fc]
            use = miss & (~_blank_mask(cand))
            if use.any():
                df2.loc[use, col] = cand[use]
                miss = _blank_mask(df2[col])

    # Intel/H2H fallback fills (best-effort) so clean output is less sparse.
    _fill_from('intel_season_avg', 'stat_season_avg')
    _fill_from('intel_l5_avg', 'stat_last5_avg')
    _fill_from('intel_l10_avg', 'stat_last10_avg')
    # For period slates (NBA1Q/NBA1H), avoid full-game proxy backfills.
    if not is_period_slate:
        _fill_from('intel_season_games', 'opp_games_played')
        _fill_from('intel_opp_vs_league_pct', 'opp_vs_avg_pct')
        _fill_from('h2h_games_vs_opp', 'h2h_games')
        _fill_from('h2h_avg', 'h2h_last_stat')
        _fill_from('h2h_last_stat', 'h2h_avg')

    if 'h2h_games_vs_opp' not in df2.columns:
        df2['h2h_games_vs_opp'] = np.nan
    m = _blank_mask(df2['h2h_games_vs_opp'])
    if (not is_period_slate) and m.any() and 'h2h_last_stat' in df2.columns:
        last_num = _num(df2['h2h_last_stat'])
        # If we have at least one H2H stat point, mark minimum sample as 1 game.
        df2.loc[m & last_num.notna(), 'h2h_games_vs_opp'] = 1.0

    if 'h2h_over_rate' not in df2.columns:
        df2['h2h_over_rate'] = np.nan
    else:
        df2['h2h_over_rate'] = pd.to_numeric(df2['h2h_over_rate'], errors='coerce')
    m = _blank_mask(df2['h2h_over_rate'])
    if (not is_period_slate) and m.any() and {'h2h_last_stat', 'line'}.issubset(df2.columns):
        hlast = _num(df2['h2h_last_stat'])
        line = _num(df2['line'])
        # One-game fallback proxy: 1.0 if last H2H beat line, else 0.0.
        proxy = np.where(hlast.notna() & line.notna(), (hlast > line).astype(float), np.nan)
        df2.loc[m, 'h2h_over_rate'] = proxy[m]

    # Keep percent-like value in 0-1 range for consistency with existing export.
    if 'h2h_over_rate' in df2.columns:
        h2h_or = _num(df2['h2h_over_rate'])
        df2['h2h_over_rate'] = np.where(h2h_or > 1.0, h2h_or / 100.0, h2h_or)

    if 'intel_l5_vs_season' not in df2.columns:
        df2['intel_l5_vs_season'] = np.nan
    else:
        df2['intel_l5_vs_season'] = pd.to_numeric(df2['intel_l5_vs_season'], errors='coerce')
    m = _blank_mask(df2['intel_l5_vs_season'])
    if m.any() and {'intel_l5_avg', 'intel_season_avg'}.issubset(df2.columns):
        l5 = _num(df2['intel_l5_avg'])
        ss = _num(df2['intel_season_avg'])
        df2.loc[m, 'intel_l5_vs_season'] = (l5 - ss)[m]

    if 'intel_season_hit_rate' not in df2.columns:
        df2['intel_season_hit_rate'] = np.nan
    else:
        df2['intel_season_hit_rate'] = pd.to_numeric(df2['intel_season_hit_rate'], errors='coerce')
    m = _blank_mask(df2['intel_season_hit_rate'])
    if m.any() and (not is_period_slate):
        for fc in ('line_hit_rate', 'line_hit_rate_over_ou_10', 'line_hit_rate_over_ou_5'):
            if fc not in df2.columns:
                continue
            cand = _num(df2[fc])
            # Convert 0-1 style rates to 0-100 for Intel Season Hit%
            cand = np.where(cand <= 1.0, cand * 100.0, cand)
            use = m & pd.notna(cand)
            if np.any(use):
                df2.loc[use, 'intel_season_hit_rate'] = cand[use]
                m = _blank_mask(df2['intel_season_hit_rate'])

    if 'intel_cushion' not in df2.columns:
        df2['intel_cushion'] = np.nan
    else:
        df2['intel_cushion'] = pd.to_numeric(df2['intel_cushion'], errors='coerce')
    m = _blank_mask(df2['intel_cushion'])
    if m.any() and {'projection', 'line'}.issubset(df2.columns) and (not is_period_slate):
        proj = _num(df2['projection'])
        line = _num(df2['line'])
        df2.loc[m, 'intel_cushion'] = (proj - line)[m]

    if 'intel_cv_pct' not in df2.columns:
        df2['intel_cv_pct'] = np.nan
    else:
        df2['intel_cv_pct'] = pd.to_numeric(df2['intel_cv_pct'], errors='coerce')
    m = _blank_mask(df2['intel_cv_pct'])
    gcols = [c for c in [f"stat_g{i}" for i in range(1, 11)] if c in df2.columns]
    if m.any() and gcols:
        g = df2[gcols].apply(pd.to_numeric, errors='coerce')
        mean = g.mean(axis=1)
        std = g.std(axis=1, ddof=0)
        cv = np.where(mean.abs() > 1e-9, (std / mean.abs()) * 100.0, np.nan)
        df2.loc[m, 'intel_cv_pct'] = cv[m]

    # Period-safe L5 over/under + hit-rate rebuild from G1-G5.
    if is_period_slate:
        g5_cols = [c for c in ("stat_g1", "stat_g2", "stat_g3", "stat_g4", "stat_g5") if c in df2.columns]
        if g5_cols and "line" in df2.columns:
            g5 = df2[g5_cols].apply(pd.to_numeric, errors="coerce")
            line = _num(df2["line"])
            valid_n = g5.notna().sum(axis=1)
            over_n = g5.gt(line, axis=0).sum(axis=1)
            under_n = g5.lt(line, axis=0).sum(axis=1)
            has_hist = valid_n > 0

            for c in ("last5_over", "last5_under", "line_hit_rate_over_ou_5", "line_hit_rate_under_ou_5"):
                if c not in df2.columns:
                    df2[c] = np.nan
                else:
                    df2[c] = pd.to_numeric(df2[c], errors="coerce").astype(object)

            df2.loc[has_hist, "last5_over"] = over_n[has_hist]
            df2.loc[has_hist, "last5_under"] = under_n[has_hist]
            df2.loc[has_hist, "line_hit_rate_over_ou_5"] = (over_n[has_hist] / valid_n[has_hist]).round(4)
            df2.loc[has_hist, "line_hit_rate_under_ou_5"] = (under_n[has_hist] / valid_n[has_hist]).round(4)

    # L10 streak columns for UI/tickets (NBA1H/NBA1Q boards use stat_g* from step4/5).
    if "line" in df2.columns:
        df2 = finalize_l10_ui_columns(df2, line_col="line")

    # Sample size proxy used by downstream eligibility void exceptions.
    if "n_legs_sample" not in df2.columns:
        df2["n_legs_sample"] = np.nan
    l5o = pd.to_numeric(df2.get("last5_over", np.nan), errors="coerce")
    l5u = pd.to_numeric(df2.get("last5_under", np.nan), errors="coerce")
    l5_sample = (l5o.fillna(0) + l5u.fillna(0)).where(l5o.notna() | l5u.notna(), np.nan)
    df2["n_legs_sample"] = pd.to_numeric(df2["n_legs_sample"], errors="coerce")
    df2.loc[df2["n_legs_sample"].isna(), "n_legs_sample"] = l5_sample[df2["n_legs_sample"].isna()]

    # Exempt high-confidence low-line OVER rows from strict OVER void tags.
    if "void_reason" in df2.columns:
        vr = df2["void_reason"].astype(str).fillna("")
        hr5 = pd.to_numeric(df2.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")
        hr5 = np.where(hr5 > 1.0, hr5 / 100.0, hr5)
        hr5 = pd.Series(hr5, index=df2.index).fillna(0.0)
        sample_n = pd.to_numeric(df2.get("n_legs_sample", np.nan), errors="coerce").fillna(0)
        line = pd.to_numeric(df2.get("line", np.nan), errors="coerce")
        dir_over = df2.get("final_bet_direction", pd.Series("", index=df2.index)).astype(str).str.upper().eq("OVER")

        exc_std_ml = dir_over & hr5.ge(0.85) & sample_n.ge(4)
        exc_elite_def = dir_over & hr5.ge(0.85) & line.le(2.0)

        def _drop_tag(text: str, tag: str) -> str:
            parts = [p for p in str(text).split(";") if p]
            kept = [p for p in parts if p != tag]
            return ";".join(kept)

        if exc_std_ml.any():
            m = exc_std_ml & vr.str.contains("VOID_STD_OVER_LOW_ML_PROB", na=False)
            df2.loc[m, "void_reason"] = vr[m].map(lambda x: _drop_tag(x, "VOID_STD_OVER_LOW_ML_PROB"))
            vr = df2["void_reason"].astype(str).fillna("")
        if exc_elite_def.any():
            m = exc_elite_def & vr.str.contains("VOID_OVER_ELITE_DEF_NON_Q5", na=False)
            df2.loc[m, "void_reason"] = vr[m].map(lambda x: _drop_tag(x, "VOID_OVER_ELITE_DEF_NON_Q5"))
            vr = df2["void_reason"].astype(str).fillna("")

    keep = [
        'tier', 'rank_score',
        'player', 'pos', 'team', 'opp_team', 'game_date', 'game_time',
        'prop_type', 'pick_type', 'line', 'standard_line',
        'final_bet_direction',
        'edge', 'abs_edge', 'projection',
        'ml_prob',
        'line_hit_rate_over_ou_5',
        'stat_last5_avg', 'stat_season_avg',
        'last5_over', 'last5_under',
        'l10_over', 'l10_under', 'l10_over_pct', 'l10_streak',
        'n_legs_sample',
        'OVERALL_DEF_RANK', 'DEF_TIER',
        'minutes_tier', 'shot_role', 'usage_role',
        'void_reason',
        # ── Intel layer (step6e) ──────────────────────────────────────────
        'intel_season_avg', 'intel_l5_avg', 'intel_l10_avg',
        'intel_season_hit_rate', 'intel_cushion', 'intel_cv_pct',
        'intel_l5_vs_season',
        'intel_season_games',
        'intel_opp_vs_league_pct', 'intel_opp_avg_allowed',
        'intel_opp_avg_allowed_pos', 'intel_opp_vs_league_pct_pos',
        # ── Full game log (g1-g10) and H2H ───────────────────────────────
        'stat_g1', 'stat_g2', 'stat_g3', 'stat_g4', 'stat_g5',
        'stat_g6', 'stat_g7', 'stat_g8', 'stat_g9', 'stat_g10',
        'stat_last10_avg',
        'h2h_avg', 'h2h_over_rate', 'h2h_games_vs_opp', 'h2h_last_stat',
        'h2h_over_rate_l5', 'l5_vs_same_opp_hit_rate',
        'b2b_flag', 'days_rest', 'game_total', 'spread',
        'is_playoff_game',
        'game_script_mult', 'game_script_note',
    ]
    if is_period_slate:
        for c in ENRICHMENT_CARRY_COLS:
            if c not in keep:
                keep.append(c)
    # Force full schema so NBA1Q/NBA1H clean outputs match NBA Step 8 columns.
    for c in keep:
        if c not in df2.columns:
            df2[c] = np.nan
    clean = df2[keep].copy()

    for col in ['rank_score', 'edge', 'abs_edge', 'projection', 'ml_prob', 'line_hit_rate_over_ou_5']:
        if col in clean.columns:
            rnd = 4 if col == 'ml_prob' else 2
            clean[col] = pd.to_numeric(clean[col], errors='coerce').round(rnd)
    if 'standard_line' in clean.columns:
        clean['standard_line'] = pd.to_numeric(clean['standard_line'], errors='coerce').round(2)
    for col in ['stat_last5_avg', 'stat_season_avg']:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors='coerce').round(1)
    for col in ['last5_over', 'last5_under']:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors='coerce').astype('Int64')
    if 'n_legs_sample' in clean.columns:
        clean['n_legs_sample'] = pd.to_numeric(clean['n_legs_sample'], errors='coerce').astype('Int64')

    tier_order = {'A': 0, 'B': 1, 'C': 2, 'D': 3}
    clean['_tier_sort'] = clean['tier'].map(tier_order)
    clean = clean.sort_values(['_tier_sort', 'rank_score'], ascending=[True, False]).drop(columns='_tier_sort')

    rename = {
        'tier': 'Tier', 'rank_score': 'Rank Score',
        'player': 'Player', 'pos': 'Pos', 'team': 'Team', 'opp_team': 'Opp',
        'game_date': 'Game Date', 'game_time': 'Game Time',
        'prop_type': 'Prop', 'pick_type': 'Pick Type', 'line': 'Line',
        'standard_line': 'Standard Line',
        'final_bet_direction': 'Direction',
        'edge': 'Edge', 'abs_edge': 'Abs Edge', 'projection': 'Projection',
        'ml_prob': 'ML Prob',
        'line_hit_rate_over_ou_5': 'Hit Rate (5g)',
        'stat_last5_avg': 'Last 5 Avg', 'stat_season_avg': 'Season Avg',
        'last5_over': 'L5 Over', 'last5_under': 'L5 Under',
        'n_legs_sample': 'N Legs Sample',
        'OVERALL_DEF_RANK': 'Def Rank', 'DEF_TIER': 'Def Tier',
        'minutes_tier': 'Min Tier', 'shot_role': 'Shot Role', 'usage_role': 'Usage Role',
        'void_reason': 'Void Reason',
        # Intel columns
        'intel_season_avg':        'Intel Season Avg',
        'intel_l5_avg':            'Intel L5 Avg',
        'intel_l10_avg':           'Intel L10 Avg',
        'intel_season_hit_rate':   'Season Hit%',
        'intel_cushion':           'Cushion',
        'intel_cv_pct':            'CV%',
        'intel_l5_vs_season':      'L5 vs Season',
        'intel_season_games':      'Season GP',
        'intel_opp_vs_league_pct': 'Opp vs Lg %',
        'intel_opp_avg_allowed':   'Opp Allowed Avg',
        'intel_opp_avg_allowed_pos':   'Opp Allowed Avg (Pos)',
        'intel_opp_vs_league_pct_pos': 'Opp vs Lg % (Pos)',
        # Game log
        'stat_last10_avg':         'Last 10 Avg',
        'stat_g1': 'G1', 'stat_g2': 'G2', 'stat_g3': 'G3',
        'stat_g4': 'G4', 'stat_g5': 'G5', 'stat_g6': 'G6',
        'stat_g7': 'G7', 'stat_g8': 'G8', 'stat_g9': 'G9', 'stat_g10': 'G10',
        # H2H
        'h2h_avg':          'H2H Avg',
        'h2h_over_rate':    'H2H Over%',
        'h2h_games_vs_opp': 'H2H Games',
        'h2h_last_stat':    'H2H Last',
        'h2h_over_rate_l5': 'H2H Over% (L5)',
        'l5_vs_same_opp_hit_rate': 'L5 vs Opp Hit%',
        # Schedule
        'b2b_flag':  'B2B',
        'days_rest': 'Days Rest',
        'game_total':'Game Total',
        'spread':    'Spread',
        'is_playoff_game': 'Playoff Game',
        'game_script_mult': 'Game Script Mult',
        'game_script_note': 'Game Script Note',
    }
    clean = clean.rename(columns=rename)

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, 'ALL', clean)
    for tier in ['A', 'B', 'C', 'D']:
        subset = clean[clean['Tier'] == tier].copy()
        if len(subset):
            write_sheet(wb, f'Tier {tier}', subset)

    wb.save(xlsx_path)
    print(f"📊 Clean XLSX saved → {xlsx_path}")

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input", default="step7_ranked_props.xlsx")
    ap.add_argument("--sheet", default="ALL")
    ap.add_argument("--output", default="step8_all_direction.csv")
    ap.add_argument("--xlsx", default="")  # optional override for xlsx path
    ap.add_argument("--date", default="", help="Filter to YYYY-MM-DD based on start_time")
    args = ap.parse_args()

    print(f"→ Loading: {args.input} (sheet={args.sheet})")
    df = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")

    carry_cols: list[str] = []
    carry_df = None
    if is_nba1h_pipeline(args.input, df):
        carry_cols, carry_df = snapshot_enrichment_carry(df)

    out = df.copy()

    # Backfill opponent labels when opp_team is missing.
    if "opp_team" in out.columns and "team" in out.columns:
        opp_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
        if opp_blank.any():
            if "pp_game_id" in out.columns:
                game_team_map = (
                    out.loc[:, ["pp_game_id", "team"]]
                    .astype(str)
                    .assign(team=lambda x: x["team"].str.strip(), pp_game_id=lambda x: x["pp_game_id"].str.strip())
                    .groupby("pp_game_id")["team"]
                    .apply(lambda s: sorted({t for t in s.tolist() if t and t.lower() not in ("nan", "none", "null")}))
                    .to_dict()
                )
                inferred = []
                for _, r in out.loc[opp_blank, ["pp_game_id", "team"]].iterrows():
                    gid = str(r.get("pp_game_id", "")).strip()
                    team = str(r.get("team", "")).strip()
                    teams = game_team_map.get(gid, [])
                    if len(teams) == 2 and team in teams:
                        inferred.append(teams[0] if teams[1] == team else teams[1])
                    else:
                        inferred.append("")
                out.loc[opp_blank, "opp_team"] = inferred

            still_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
            # Fallback for combo rows where team carries matchup text like "DET/NOP".
            if still_blank.any():
                tm = out.loc[still_blank, "team"].astype(str).str.strip()
                has_pair = tm.str.contains("/", regex=False)
                out.loc[still_blank[still_blank].index[has_pair], "opp_team"] = tm[has_pair]
            still_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
            out.loc[still_blank, "opp_team"] = "UNKNOWN_OPP"

    # Keep only rows for target slate date when start_time is available.
    # This prevents stale historical slates from leaking into today's output.
    if "start_time" in out.columns:
        target_date = (args.date or datetime.now(_ET).strftime("%Y-%m-%d")).strip()
        start_dt_et = pd.to_datetime(out["start_time"], utc=True, errors="coerce").dt.tz_convert(_ET)
        start_dates = start_dt_et.dt.strftime("%Y-%m-%d")
        keep_mask = start_dates.eq(target_date)
        kept = int(keep_mask.sum())
        total = len(out)
        if kept == 0:
            # Fallback to latest available slate date so pipeline does not emit empty NBA outputs
            available_dates = start_dates[start_dates.notna() & (start_dates != "NaT")]
            if len(available_dates):
                fallback_date = available_dates.max()
                keep_mask = start_dates.eq(fallback_date)
                kept = int(keep_mask.sum())
                print(
                    f"[DateFilter] No rows for {target_date}; "
                    f"falling back to latest available date {fallback_date} ({kept} rows)"
                )
            else:
                print(f"[DateFilter] No parseable start_time values; keeping all {total} rows")
                keep_mask = pd.Series(True, index=out.index)
        else:
            print(f"[DateFilter] Kept {kept}/{total} rows for {target_date} (dropped {total - kept} rows)")
        out = out.loc[keep_mask].copy()

    if "start_time" in out.columns:
        et = _start_times_et(out["start_time"])
        out["game_date"] = et.dt.strftime("%Y-%m-%d").where(et.notna(), "")
    else:
        out["game_date"] = ""

    reconcile_signed_edge_abs_dataframe(out)
    edge = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = pd.to_numeric(out["abs_edge"], errors="coerce")

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)
    forced = pick_type.isin(["Goblin", "Demon"])

    model_dir = np.where(edge >= 0, "OVER", "UNDER")
    out["model_dir"] = model_dir

    final_dir = model_dir.copy()
    reason = np.where(abs_edge < 0.03, "MODEL_TIEBREAK_DIFF<0.03", "")
    final_dir = np.where(forced, "OVER", final_dir)
    reason = np.where(forced, "FORCED_OVER_ONLY_GOB_DEM", reason)

    out["final_bet_direction"] = final_dir
    out["final_dir_reason"] = reason

    if carry_df is not None:
        out = reattach_enrichment_carry(out, carry_df, carry_cols, label=args.input)

    # Save full CSV
    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"✅ Saved → {args.output}")
    print("final_bet_direction counts:")
    print(pd.Series(out["final_bet_direction"]).value_counts().to_string())
    if "tier" in out.columns:
        print("tier counts:")
        print(out["tier"].value_counts().to_string())

    # Save clean XLSX
    xlsx_path = args.xlsx if args.xlsx else args.output.replace(".csv", "_clean.xlsx")
    build_clean_xlsx(out, xlsx_path, source_hint=args.input)
    _copy_dated_step8_nba(xlsx_path, (args.date or "").strip())

if __name__ == "__main__":
    try:
        main()
    except Exception as e:
        print(f"ERROR [PropOracle-NBA-S8] {type(e).__name__}: {e}")
        raise SystemExit(1) from e
