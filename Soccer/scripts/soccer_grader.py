#!/usr/bin/env python3
"""
soccer_grader.py
PropOracle Soccer Grader

Grades yesterday's (or any date's) soccer props against actual results
pulled directly from proporacle_ref.db. Outputs a per-prop graded Excel
file into Soccer/outputs/graded/.

Supported props (when the column exists in proporacle_ref.soccer — filled by
``build_boxscore_ref.py`` from ESPN boxscore / roster stat shapes):
  Shots, Shots On Target, Goals, Assists, Goal + Assist,
  Goalie Saves, Fouls, Passes Attempted, Tackles,
  Clearances, Attempted Dribbles (often NULL until ESPN exposes them per player)

Still unsupported / VOID:
  Key Passes, Shots Assisted, Crosses, Yellow Cards,
  Goals Allowed (no per-player goals-conceded column in ref DB)

Run:
  py -3.14 scripts/soccer_grader.py
  py -3.14 scripts/soccer_grader.py --date 2026-03-14
  py -3.14 scripts/soccer_grader.py --date 2026-03-14 --out Soccer/outputs/graded/
"""

from __future__ import annotations

import argparse
import sqlite3
import sys
import unicodedata
from datetime import date, timedelta
from pathlib import Path

import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
    sys.stderr.reconfigure(encoding="utf-8")
except Exception:
    pass


# ── DB path resolution ────────────────────────────────────────────────────────
_here = Path(__file__).resolve().parent
DB_PATH = None
_c = _here
for _ in range(12):
    for rel in (
        ("data", "cache", "proporacle_ref.db"),
        ("NBA", "data", "cache", "proporacle_ref.db"),
    ):
        cand = _c.joinpath(*rel)
        if cand.exists():
            DB_PATH = cand
            break
    if DB_PATH:
        break
    if _c.parent == _c:
        break
    _c = _c.parent

# ── Prop → DB column ──────────────────────────────────────────────────────────
BASE_PROP_TO_DB: dict[str, str] = {
    "shots":                                "sh",
    "shots on target":                      "sog",
    "shots on goal":                        "sog",
    "goals":                                "g",
    "assists":                              "a",
    "goal + assist":                        "g + a",
    "goal+assist":                          "g + a",
    "goalie saves":                         "sv",
    "goalkeeper saves":                     "sv",
    "fouls":                                "fc",
    # Combo
    "shots (combo)":                        "sh",
    "shots on target (combo)":              "sog",
    "goalie saves (combo)":                 "sv",
}

BASE_UNSUPPORTED = {
    "key passes", "shots assisted", "crosses", "yellow cards",
    # proporacle_ref soccer table has no goals-conceded column; do not grade vs saves (sv).
    "goals allowed", "goals allowed (combo)", "goals allowed in first 30 minutes",
}


def resolve_soccer_mappings(con: sqlite3.Connection) -> tuple[dict[str, str], set[str], dict[str, str]]:
    cols = {r[1].lower() for r in con.execute("PRAGMA table_info(soccer)").fetchall()}
    prop_to_db = dict(BASE_PROP_TO_DB)
    unsupported = set(BASE_UNSUPPORTED)
    status: dict[str, str] = {}

    # Task-specific mappings to audit/add.
    checks = [
        ("passes attempted", "pa", "Passes Attempted"),
        ("passes attempted (combo)", "pa", "Passes Attempted (combo)"),
        ("clearances", "clearances", "Clearances"),
        ("tackles", "tk", "Tackles"),
        ("attempted dribbles", "dribble_attempts", "Attempted Dribbles"),
    ]
    for key, col, label in checks:
        if col in cols:
            prop_to_db[key] = col
            status[label] = f"ADDED ({col})"
        else:
            unsupported.add(key)
            status[label] = f"NEEDS_SCRAPER_UPDATE (missing DB col '{col}')"

    return prop_to_db, unsupported, status

# ── Colours ───────────────────────────────────────────────────────────────────
COL = {
    "hit":    "27AE60", "miss":   "E74C3C", "void":   "7F8C8D",
    "push":   "F39C12", "hdr":    "1C1C1C", "tier_a": "D5F5E3",
    "tier_b": "D6EAF8", "tier_c": "FEF9E7", "tier_d": "FDEDEC",
    "over":   "D6EAF8", "under":  "FDEBD0", "white":  "FFFFFF",
}

TIER_HDR = {"A": "1E8449", "B": "2874A6", "C": "D4AC0D", "D": "717D7E"}


def _norm(n: str) -> str:
    s = unicodedata.normalize("NFD", str(n).strip().lower())
    return "".join(c for c in s if unicodedata.category(c) != "Mn")


def _side(c="CCCCCC"):
    s = Side(style="thin", color=c)
    return Border(left=s, right=s, top=s, bottom=s)


def _fill(bg): return PatternFill("solid", start_color=bg)
def _font(bold=False, color="000000", sz=9):
    return Font(bold=bold, color=color, name="Arial", size=sz)
def _align(h="center"): return Alignment(horizontal=h, vertical="center")


def hc(ws, r, c, v, bg="1C1C1C", fc="FFFFFF", bold=True, sz=9):
    cell = ws.cell(row=r, column=c, value=v)
    cell.fill = _fill(bg); cell.font = _font(bold, fc, sz)
    cell.alignment = _align("center"); cell.border = _side()


def dc(ws, r, c, v, bg=None, bold=False, sz=9, align="center", fc="000000"):
    cell = ws.cell(row=r, column=c, value=v)
    if bg: cell.fill = _fill(bg)
    cell.font = _font(bold, fc, sz)
    cell.alignment = _align(align); cell.border = _side()


def sw(ws, widths):
    for i, w in enumerate(widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w


# ── Lookup actual stat from DB ────────────────────────────────────────────────

def get_actual_and_minutes(
    con: sqlite3.Connection, espn_id: str, player_name: str, db_col: str, grade_date: str
) -> tuple[float | None, float | None]:
    """Fetch player's actual stat and minutes_played for grade_date from soccer DB."""
    if not db_col:
        return None, None

    # Handle combo expressions (g + a)
    if "+" in db_col:
        cols = [c.strip() for c in db_col.split("+")]
        vals = []
        mins = []
        for col in cols:
            v, m = get_actual_and_minutes(con, espn_id, player_name, col, grade_date)
            if v is None:
                return None, None
            vals.append(v)
            mins.append(m)
        m_out = None if any(x is None for x in mins) else min(float(x) for x in mins)
        return sum(vals), m_out

    # Primary: ESPN player ID
    if espn_id and str(espn_id).strip() not in ("", "nan", "None"):
        row = con.execute(f"""
            SELECT {db_col}, minutes_played FROM soccer
            WHERE espn_player_id = ? AND game_date = ?
              AND {db_col} IS NOT NULL
        """, [str(espn_id), grade_date]).fetchone()
        if row:
            m = None
            try:
                m = float(row[1]) if row[1] is not None else None
            except (TypeError, ValueError):
                m = None
            return float(row[0]), m

    # Fallback: player name
    if player_name:
        norm = _norm(player_name)
        row = con.execute(f"""
            SELECT {db_col}, minutes_played FROM soccer
            WHERE lower(player) = ? AND game_date = ?
              AND {db_col} IS NOT NULL
        """, [norm, grade_date]).fetchone()
        if row:
            m = None
            try:
                m = float(row[1]) if row[1] is not None else None
            except (TypeError, ValueError):
                m = None
            return float(row[0]), m

    return None, None


# ── Grade a single prop ───────────────────────────────────────────────────────

def grade_prop(actual: float | None, line: float,
               direction: str) -> tuple[str, str]:
    """Return (result, reason)."""
    if actual is None:
        return "VOID", "no_actuals"
    try:
        line_f = float(line)
    except (TypeError, ValueError):
        return "VOID", "bad_line"

    direction = str(direction).strip().upper()

    if abs(actual - line_f) < 0.001:
        return "PUSH", f"actual={actual} == line={line_f}"

    if direction == "OVER":
        result = "HIT" if actual > line_f else "MISS"
    elif direction == "UNDER":
        result = "HIT" if actual < line_f else "MISS"
    else:
        result = "VOID"

    return result, f"actual={actual} {direction} line={line_f}"


# ── Load slate ────────────────────────────────────────────────────────────────

def load_slate(slate_path: Path) -> pd.DataFrame:
    xl = pd.ExcelFile(slate_path)
    sheet = next((s for s in xl.sheet_names if s.upper() == "ALL"), xl.sheet_names[0])
    df = xl.parse(sheet)
    df.columns = [str(c).strip() for c in df.columns]
    return df


# ── Main grading logic ────────────────────────────────────────────────────────

def run_grader(grade_date: str, slate_path: Path, db_path: Path,
               out_dir: Path) -> Path:

    print(f"[Soccer Grader] Date:  {grade_date}")
    print(f"[Soccer Grader] Slate: {slate_path}")
    print(f"[Soccer Grader] DB:    {db_path}")

    df = load_slate(slate_path)
    print(f"[Soccer Grader] Slate rows: {len(df)}")

    con = sqlite3.connect(str(db_path))
    prop_to_db, unsupported, map_status = resolve_soccer_mappings(con)
    print("[Soccer Grader] Mapping audit:")
    for k in ("Passes Attempted", "Clearances", "Tackles", "Attempted Dribbles"):
        print(f"  - {k}: {map_status.get(k, 'N/A')}")
    db_n = con.execute("SELECT COUNT(*) FROM soccer WHERE game_date = ?",
                       [grade_date]).fetchone()[0]
    print(f"[Soccer Grader] DB rows for {grade_date}: {db_n}")

    if db_n == 0:
        print(f"  ⚠️  No soccer data in DB for {grade_date}.")
        print(f"     Run: py NBA/scripts/build_boxscore_ref.py --date {grade_date} --sports soccer")

    # Normalise column names to lowercase keys
    col = {c.lower().replace(" ", "_"): c for c in df.columns}

    def g(key: str):
        return col.get(key, col.get(key.replace("_", " ")))

    player_col    = g("player")
    espnid_col    = g("espn_player_id") or g("espn_id")
    prop_col      = g("prop_type") or g("prop")
    line_col      = g("line")
    dir_col       = g("final_bet_direction") or g("direction")
    tier_col      = g("tier")
    score_col     = g("rank_score")
    team_col      = g("team")
    opp_col       = g("opp_team") or g("opp")
    l5_col        = g("stat_last5_avg") or g("last_5_avg")
    season_col    = g("stat_season_avg") or g("season_avg")
    hr5_col       = g("line_hit_rate_over_ou_5") or g("hit_rate_(5g)")
    deftier_col   = g("def_tier")
    void_col      = g("void_reason")

    results = []
    hit = miss = push = void_n = 0
    legacy_hit = legacy_miss = legacy_push = 0
    no_data_void_count = 0

    for _, row in df.iterrows():
        prop_type = str(row.get(prop_col, "") or "").strip()
        prop_key  = prop_type.lower()
        line      = row.get(line_col)
        direction = str(row.get(dir_col, "") or "").strip().upper()
        espn_id   = str(row.get(espnid_col, "") or "").strip() if espnid_col else ""
        player    = str(row.get(player_col, "") or "").strip() if player_col else ""

        # Check if prop is unsupported
        if prop_key in unsupported:
            result, reason = "VOID", "unsupported_prop"
        elif void_col and pd.notna(row.get(void_col)) and str(row.get(void_col)).strip():
            result, reason = "VOID", str(row.get(void_col))
        else:
            db_col = prop_to_db.get(prop_key)
            actual, minutes_played = get_actual_and_minutes(con, espn_id, player, db_col, grade_date)
            legacy_result, _ = grade_prop(actual, line, direction)
            if legacy_result == "HIT":
                legacy_hit += 1
            elif legacy_result == "MISS":
                legacy_miss += 1
            elif legacy_result == "PUSH":
                legacy_push += 1

            if actual is None:
                result, reason = "VOID", "no_actuals"
                no_data_void_count += 1
            elif float(actual) == 0.0:
                if minutes_played is None:
                    result, reason = "VOID", "no_minutes_data"
                    no_data_void_count += 1
                elif float(minutes_played) <= 0:
                    result, reason = "VOID", "did_not_play"
                    no_data_void_count += 1
                else:
                    result, reason = grade_prop(actual, line, direction)
            else:
                result, reason = grade_prop(actual, line, direction)

        # Tally
        if result == "HIT":   hit  += 1
        elif result == "MISS": miss += 1
        elif result == "PUSH": push += 1
        else:                  void_n += 1

        rec = {
            "Result":       result,
            "Player":       player,
            "Team":         row.get(team_col, "") if team_col else "",
            "Opp":          row.get(opp_col, "")  if opp_col  else "",
            "Prop":         prop_type,
            "Line":         line,
            "Direction":    direction,
            "Tier":         row.get(tier_col, "")  if tier_col  else "",
            "Rank Score":   row.get(score_col, "") if score_col else "",
            "L5 Avg":       row.get(l5_col, "")    if l5_col    else "",
            "Season Avg":   row.get(season_col, "") if season_col else "",
            "Hit Rate 5g":  row.get(hr5_col, "")   if hr5_col   else "",
            "Def Tier":     row.get(deftier_col, "") if deftier_col else "",
            "Reason":       reason,
        }
        results.append(rec)

    con.close()

    graded = pd.DataFrame(results)
    total_gradeable = hit + miss + push
    hit_rate = hit / total_gradeable * 100 if total_gradeable > 0 else 0

    print(f"\n[Soccer Grader] Results for {grade_date}:")
    print(f"  HIT:  {hit}")
    print(f"  MISS: {miss}")
    print(f"  PUSH: {push}")
    print(f"  VOID: {void_n}")
    print(f"  Hit Rate: {hit_rate:.1f}%  ({hit}/{total_gradeable} gradeable props)")
    legacy_total = legacy_hit + legacy_miss + legacy_push
    legacy_hr = (legacy_hit / legacy_total * 100) if legacy_total else 0.0
    print(f"  Est pre-fix hit rate (legacy no-data grading): {legacy_hr:.1f}% ({legacy_hit}/{legacy_total})")
    print(f"  no_data->VOID rows: {no_data_void_count}")

    # ── Write Excel ───────────────────────────────────────────────────────────
    out_dir.mkdir(parents=True, exist_ok=True)
    out_path = out_dir / f"soccer_graded_{grade_date}.xlsx"

    wb = Workbook()
    wb.remove(wb.active)

    # Summary sheet
    ws_sum = wb.create_sheet("SUMMARY")
    hc(ws_sum, 1, 1, f"Soccer Graded — {grade_date}", bg="117A65", sz=12)
    ws_sum.merge_cells("A1:F1")
    summary_rows = [
        ("Date",        grade_date),
        ("HIT",         hit),
        ("MISS",        miss),
        ("PUSH",        push),
        ("VOID",        void_n),
        ("Hit Rate",    f"{hit_rate:.1f}%"),
        ("Gradeable",   total_gradeable),
    ]
    for i, (k, v) in enumerate(summary_rows, 2):
        dc(ws_sum, i, 1, k, bold=True, align="left")
        dc(ws_sum, i, 2, v, align="left")

    # Prop type breakdown
    dc(ws_sum, 10, 1, "Prop Type Breakdown", bold=True, align="left")
    dc(ws_sum, 11, 1, "Prop",    bold=True)
    dc(ws_sum, 11, 2, "HIT",     bold=True)
    dc(ws_sum, 11, 3, "MISS",    bold=True)
    dc(ws_sum, 11, 4, "Hit Rate",bold=True)
    prop_stats = graded[graded["Result"].isin(["HIT","MISS"])].groupby("Prop")["Result"]
    ri = 12
    for prop, grp in prop_stats:
        h = (grp == "HIT").sum()
        m = (grp == "MISS").sum()
        hr = h / (h + m) * 100 if (h + m) > 0 else 0
        dc(ws_sum, ri, 1, prop,          align="left")
        dc(ws_sum, ri, 2, h,             bg="D5F5E3" if hr >= 60 else None)
        dc(ws_sum, ri, 3, m,             bg="FDEDEC" if hr < 40 else None)
        dc(ws_sum, ri, 4, f"{hr:.1f}%",  bg="D5F5E3" if hr >= 60 else "FDEDEC" if hr < 40 else None)
        ri += 1
    sw(ws_sum, [20, 12, 8, 8, 8, 8])

    # Result colour map
    RESULT_BG = {"HIT": COL["hit"], "MISS": COL["miss"],
                 "PUSH": COL["push"], "VOID": COL["void"]}

    HEADERS = ["Result","Player","Team","Opp","Prop","Line","Direction",
               "Tier","Rank Score","L5 Avg","Season Avg","Hit Rate 5g",
               "Def Tier","Reason"]
    WIDTHS   = [8, 24, 12, 12, 22, 7, 10, 6, 10, 9, 10, 12, 10, 30]

    def write_graded_sheet(wb, sheet_name, rows_df, hdr_color="1C1C1C"):
        ws = wb.create_sheet(sheet_name)
        for ci, h in enumerate(HEADERS, 1):
            hc(ws, 1, ci, h, bg=hdr_color)
        ws.row_dimensions[1].height = 22
        ws.freeze_panes = "A2"

        for ri, (_, row) in enumerate(rows_df.iterrows(), 2):
            result = str(row.get("Result", ""))
            bg = RESULT_BG.get(result, COL["white"])
            tier = str(row.get("Tier", ""))
            tier_bg = {"A": COL["tier_a"], "B": COL["tier_b"],
                       "C": COL["tier_c"], "D": COL["tier_d"]}.get(tier, COL["white"])

            for ci, h in enumerate(HEADERS, 1):
                val = row.get(h, "")
                if pd.isna(val): val = ""
                cell_bg = bg if h == "Result" else (
                    tier_bg if h == "Tier" else (
                    COL["over"]  if h == "Direction" and str(val).upper() == "OVER" else (
                    COL["under"] if h == "Direction" and str(val).upper() == "UNDER"
                    else None)))
                dc(ws, ri, ci, val, bg=cell_bg,
                   bold=(h == "Result"),
                   fc="FFFFFF" if h == "Result" else "000000",
                   align="left" if h in ("Player","Prop","Reason") else "center")

        sw(ws, WIDTHS)

    # All results
    write_graded_sheet(wb, "ALL", graded)

    # Split by result
    for result in ["HIT", "MISS", "PUSH", "VOID"]:
        sub = graded[graded["Result"] == result]
        if not sub.empty:
            hdr = {"HIT": COL["hit"], "MISS": COL["miss"],
                   "PUSH": COL["push"], "VOID": COL["void"]}.get(result, "1C1C1C")
            write_graded_sheet(wb, result, sub, hdr_color=hdr)

    # Split by tier (hits/misses only)
    gradeable = graded[graded["Result"].isin(["HIT", "MISS"])]
    for tier in ["A", "B", "C", "D"]:
        sub = gradeable[gradeable["Tier"].astype(str).str.upper() == tier]
        if not sub.empty:
            write_graded_sheet(wb, f"Tier {tier}", sub, hdr_color=TIER_HDR.get(tier, "1C1C1C"))

    wb.save(out_path)
    print(f"\n✅ Saved → {out_path}")
    print(f"   Sheets: {[s.title for s in wb.worksheets]}")
    return out_path


# ── Main ──────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--date",  default="",
                    help="Date to grade YYYY-MM-DD (default: yesterday)")
    ap.add_argument("--slate", default="",
                    help="Path to step8_soccer_direction_clean.xlsx")
    ap.add_argument("--out",   default="",
                    help="Output directory (default: Soccer/outputs/graded/)")
    ap.add_argument("--db",    default="",
                    help="Override DB path")
    args = ap.parse_args()

    grade_date = args.date or str(date.today() - timedelta(days=1))

    # Resolve DB
    db_path = Path(args.db) if args.db else DB_PATH
    if not db_path or not db_path.exists():
        print(f"❌ DB not found. Run: py NBA/scripts/build_boxscore_ref.py --date {grade_date} --sports soccer")
        sys.exit(1)

    # Resolve slate — prefer dated archive, fall back to current
    if args.slate:
        slate_path = Path(args.slate)
    else:
        _p = Path(__file__).resolve().parent
        slate_path = None
        for _ in range(6):
            # 1. Dated archive in outputs/YYYY-MM-DD/ (correct for grading)
            dated = _p / "outputs" / grade_date / f"step8_soccer_direction_clean_{grade_date}.xlsx"
            if dated.exists():
                slate_path = dated
                break
            # 2. Current slate (only valid if grading today)
            current = _p / "Soccer" / "outputs" / "step8_soccer_direction_clean.xlsx"
            if current.exists():
                slate_path = current
                break
            current2 = _p / "outputs" / "step8_soccer_direction_clean.xlsx"
            if current2.exists():
                slate_path = current2
                break
            _p = _p.parent
        if not slate_path:
            print(f"❌ Could not find slate for {grade_date}")
            print(f"   Expected: outputs/{grade_date}/step8_soccer_direction_clean_{grade_date}.xlsx")
            print(f"   Run pipeline for {grade_date} first, or pass --slate explicitly")
            sys.exit(1)

    # Resolve output dir
    if args.out:
        out_dir = Path(args.out)
    else:
        out_dir = slate_path.parent / "graded"

    run_grader(grade_date, slate_path, db_path, out_dir)


if __name__ == "__main__":
    main()
