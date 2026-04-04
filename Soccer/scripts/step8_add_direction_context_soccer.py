#!/usr/bin/env python3
"""
step8_add_direction_context_soccer.py  (Soccer Pipeline)

Mirrors NBA step8_add_direction_context.py.
Reads step7_soccer_ranked.xlsx, appends direction context,
outputs full CSV + clean formatted XLSX.

Run:
  py -3.14 step8_add_direction_context_soccer.py \
    --input step7_soccer_ranked.xlsx \
    --output step8_soccer_direction.csv
"""

from __future__ import annotations

import argparse
import sys
import numpy as np
import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

try:
    sys.stdout.reconfigure(encoding="utf-8")
except Exception:
    pass


def _norm_pick_type(x: str) -> str:
    t = (str(x) if x is not None else "").strip().lower()
    if "gob" in t: return "Goblin"
    if "dem" in t: return "Demon"
    return "Standard"


TIER_COLORS = {
    "A": ("1E8449", "FFFFFF"),
    "B": ("2874A6", "FFFFFF"),
    "C": ("D4AC0D", "000000"),
    "D": ("717D7E", "FFFFFF"),
}
HEADER_COLOR = "1C1C1C"


def thin_border():
    s = Side(style="thin", color="DDDDDD")
    return Border(left=s, right=s, top=s, bottom=s)


def write_sheet(wb, name: str, data: pd.DataFrame) -> None:
    ws = wb.create_sheet(name)
    tier_bg, tier_fg = TIER_COLORS.get(name, ("333333", "FFFFFF"))
    headers = list(data.columns)

    for ci, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci, value=h)
        cell.font      = Font(bold=True, color="FFFFFF", name="Arial", size=10)
        cell.fill      = PatternFill("solid", start_color=HEADER_COLOR)
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border    = thin_border()
    ws.row_dimensions[1].height = 30

    for ri, row in enumerate(data.itertuples(index=False), 2):
        direction = row[headers.index("Direction")] if "Direction" in headers else ""
        for ci, val in enumerate(row, 1):
            col_name    = headers[ci - 1]
            display_val = "" if pd.isna(val) else val
            cell        = ws.cell(row=ri, column=ci, value=display_val)
            cell.font      = Font(name="Arial", size=9)
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border    = thin_border()

            if col_name == "Tier":
                cell.fill = PatternFill("solid", start_color=tier_bg)
                cell.font = Font(bold=True, color=tier_fg, name="Arial", size=9)
            elif col_name == "Direction":
                bg = "C8F7C5" if val == "OVER" else "F7C5C5" if val == "UNDER" else "FFFFFF"
                cell.fill = PatternFill("solid", start_color=bg)
                cell.font = Font(bold=True, name="Arial", size=9)
            elif col_name == "League":
                cell.fill = PatternFill("solid", start_color="EBF5FB")
            else:
                cell.fill = PatternFill("solid", start_color="F9F9F9" if ri % 2 == 0 else "FFFFFF")

    col_widths = {
        "Tier": 6, "Rank Score": 10, "Player": 20, "Pos": 6, "Pos Group": 9,
        "Team": 12, "Opp": 12, "League": 12, "Game Time": 16,
        "ESPN ID": 10,
        "Prop": 18, "Pick Type": 10, "Line": 7,
        "Direction": 9, "Edge": 7, "Projection": 10,
        "ML Prob": 9, "Edge Score": 10, "Blended Score": 12,
        "Hit Rate (5g)": 12, "Hit Rate (10g)": 12, "Last 5 Avg": 10, "Season Avg": 10,
        "L5 Over": 8, "L5 Under": 8,
        "Def Rank": 9, "Def Tier": 10,
        "Min Tier": 9, "Shot Role": 10, "Usage Role": 10,
        "Void Reason": 20,
    }
    for ci, h in enumerate(headers, 1):
        ws.column_dimensions[get_column_letter(ci)].width = col_widths.get(h, 12)

    ws.freeze_panes    = "A2"
    ws.auto_filter.ref = f"A1:{get_column_letter(len(headers))}1"


_MIN_TIER_NUM_MAP = {0: "DNP Risk", 1: "Spot", 2: "Rotation", 3: "Starter"}


def build_clean_xlsx(df: pd.DataFrame, xlsx_path: str) -> None:
    df2 = df.copy()
    # Convert numeric minutes_tier (0-3) back to human labels
    if "minutes_tier" in df2.columns:
        _mt_num = pd.to_numeric(df2["minutes_tier"], errors="coerce")
        _mt_valid = _mt_num.notna()
        if _mt_valid.any():
            df2.loc[_mt_valid, "minutes_tier"] = _mt_num[_mt_valid].round().astype(int).map(_MIN_TIER_NUM_MAP).fillna(df2.loc[_mt_valid, "minutes_tier"])
    try:
        import platform
        _time_fmt = "%m/%d %#I:%M %p" if platform.system() == "Windows" else "%m/%d %-I:%M %p"
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime(_time_fmt)
    except Exception:
        df2["game_time"] = pd.to_datetime(df2.get("start_time", ""), errors="coerce").dt.strftime("%m/%d %H:%M")

    keep = [
        "tier", "rank_score",
        "player", "pos", "position_group", "team", "opp_team", "league", "game_time",
        "espn_player_id",
        "prop_type", "pick_type", "line",
        "final_bet_direction",
        "edge", "projection",
        "ml_prob",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
        "stat_last5_avg", "stat_season_avg",
        "last5_over", "last5_under",
        "OVERALL_DEF_RANK", "DEF_TIER",
        "minutes_tier", "shot_role", "usage_role",
        "void_reason",
        # ── Game log ─────────────────────────────────────────────────────────
        "stat_g1", "stat_g2", "stat_g3", "stat_g4", "stat_g5",
        "stat_g6", "stat_g7", "stat_g8", "stat_g9", "stat_g10",
        "stat_last10_avg",
        # ── Schedule / context ───────────────────────────────────────────────
        "avg_minutes",
        "game_script_mult",
        "game_script_note",
    ]
    keep  = [c for c in keep if c in df2.columns]
    clean = df2[keep].copy()

    for col in [
        "rank_score",
        "edge",
        "projection",
        "ml_prob",
        "edge_score",
        "blended_score",
        "line_hit_rate_over_ou_5",
        "line_hit_rate_over_ou_10",
    ]:
        if col in clean.columns:
            rnd = 4 if col in ("ml_prob", "edge_score", "blended_score") else 2
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(rnd)
    for col in ["stat_last5_avg", "stat_season_avg"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").round(1)
    for col in ["last5_over", "last5_under"]:
        if col in clean.columns:
            clean[col] = pd.to_numeric(clean[col], errors="coerce").astype("Int64")

    tier_order = {"A": 0, "B": 1, "C": 2, "D": 3}

    clean["_tier_sort"] = clean["tier"].map(tier_order)
    clean = clean.sort_values(["_tier_sort", "rank_score"], ascending=[True, False]).drop(columns="_tier_sort")

    rename = {
        "tier": "Tier", "rank_score": "Rank Score",
        "player": "Player", "pos": "Pos", "position_group": "Pos Group",
        "team": "Team", "opp_team": "Opp", "league": "League", "game_time": "Game Time",
        "espn_player_id": "ESPN ID",
        "prop_type": "Prop", "pick_type": "Pick Type", "line": "Line",
        "final_bet_direction": "Direction",
        "edge": "Edge", "projection": "Projection",
        "ml_prob": "ML Prob",
        "edge_score": "Edge Score",
        "blended_score": "Blended Score",
        "line_hit_rate_over_ou_5": "Hit Rate (5g)",
        "line_hit_rate_over_ou_10": "Hit Rate (10g)",
        "stat_last5_avg": "Last 5 Avg", "stat_season_avg": "Season Avg",
        "last5_over": "L5 Over", "last5_under": "L5 Under",
        "OVERALL_DEF_RANK": "Def Rank", "DEF_TIER": "Def Tier",
        "minutes_tier": "Min Tier", "shot_role": "Shot Role", "usage_role": "Usage Role",
        "void_reason": "Void Reason",
        # Game log
        "stat_last10_avg": "Last 10 Avg",
        "stat_g1": "G1", "stat_g2": "G2", "stat_g3": "G3",
        "stat_g4": "G4", "stat_g5": "G5", "stat_g6": "G6",
        "stat_g7": "G7", "stat_g8": "G8", "stat_g9": "G9", "stat_g10": "G10",
        # Context
        "avg_minutes": "Avg Min",
        "game_script_mult": "Game Script Mult",
        "game_script_note": "Game Script Note",
    }
    clean = clean.rename(columns=rename)

    # Exclude voided rows from Tier sheets (after rename, void col is "Void Reason")
    void_col = "Void Reason" if "Void Reason" in clean.columns else None
    if void_col:
        clean_eligible = clean[clean[void_col].isna() | (clean[void_col] == "")].copy()
    else:
        clean_eligible = clean.copy()

    wb = Workbook()
    wb.remove(wb.active)
    write_sheet(wb, "ALL", clean)
    for tier in ["A", "B", "C", "D"]:
        subset = clean_eligible[clean_eligible["Tier"] == tier].copy()
        # Always write every Tier sheet (even if empty) so step9 never crashes
        # on a missing sheet_name
        write_sheet(wb, f"Tier {tier}", subset if len(subset) else clean_eligible.head(0))

    wb.save(xlsx_path)
    print(f"Clean XLSX saved -> {xlsx_path}")


def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="s7_soccer_ranked.xlsx")
    ap.add_argument("--sheet",  default="ALL")
    ap.add_argument("--output", default="step8_soccer_direction.csv")
    ap.add_argument("--xlsx",   default="step8_soccer_direction_clean.xlsx")
    ap.add_argument("--date",   default="", help="YYYY-MM-DD target date (default: today ET)")
    args = ap.parse_args()

    print(f"Loading: {args.input} (sheet={args.sheet})")
    df  = pd.read_excel(args.input, sheet_name=args.sheet, dtype=str).fillna("")

    if df.empty:
        print("ERROR [PropOracle-Soccer-S8] Empty input from S7 — aborting.")
        sys.exit(1)

    # ── Date filter: keep only target date's games ───────────────────────────
    import datetime, zoneinfo
    eastern = zoneinfo.ZoneInfo("America/New_York")
    target_str = (args.date.strip()[:10] if args.date
                  else datetime.datetime.now(tz=eastern).date().isoformat())
    if "start_time" in df.columns:
        before_filter = len(df)
        def _to_et_date(val):
            try:
                dt = pd.to_datetime(val)          # handles -04:00 offset natively
                if dt.tzinfo is None:
                    dt = dt.tz_localize("UTC")
                return dt.tz_convert(eastern).date().isoformat()
            except Exception:
                return ""
        df["_et_date"] = df["start_time"].apply(_to_et_date)
        mask = df["_et_date"] == target_str
        if not mask.any():
            ed = df["_et_date"].astype(str)
            ok = ed.str.len() >= 10
            future = ed[ok & (ed >= target_str)]
            if future.empty:
                fallback = ed[ok]
                pick = str(fallback.mode().iloc[0]) if len(fallback) else target_str
            else:
                pick = str(future.min())
            if pick != target_str:
                print(
                    f"[DateFilter] Kept 0/{before_filter} for {target_str} ET; "
                    f"using nearest slate date {pick} (>= target)"
                )
            mask = df["_et_date"] == pick
        df = df.loc[mask].drop(columns="_et_date")
        dropped = before_filter - len(df)
        print(f"[DateFilter] Kept {len(df)}/{before_filter} rows for slate ET (dropped {dropped} rows)")
    else:
        print("[DateFilter] WARNING: no start_time column — skipping date filter")

    out = df.copy()

    # Backfill opponent labels from game context when opp_team is missing.
    # Many soccer rows share pp_game_id; if each game has exactly two teams,
    # infer each row's opponent as the "other" team in that game.
    if "opp_team" in out.columns and "pp_game_id" in out.columns and "team" in out.columns:
        opp_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
        if opp_blank.any():
            game_team_map = (
                out.loc[:, ["pp_game_id", "team"]]
                .astype(str)
                .assign(team=lambda x: x["team"].str.strip(), pp_game_id=lambda x: x["pp_game_id"].str.strip())
                .groupby("pp_game_id")["team"]
                .apply(lambda s: sorted({t for t in s.tolist() if t and t.lower() not in ("nan", "none", "null")}))
                .to_dict()
            )
            inferred = []
            for _, r in out.loc[opp_blank, ["pp_game_id", "team"]].iterrows():
                gid = str(r.get("pp_game_id", "")).strip()
                team = str(r.get("team", "")).strip()
                teams = game_team_map.get(gid, [])
                if len(teams) == 2 and team in teams:
                    inferred.append(teams[0] if teams[1] == team else teams[1])
                else:
                    inferred.append("")
            out.loc[opp_blank, "opp_team"] = inferred
            # Keep non-empty placeholder for unresolved rows so downstream views are explicit.
            still_blank = out["opp_team"].astype(str).str.strip().isin(["", "nan", "None", "null"])
            out.loc[still_blank, "opp_team"] = "UNKNOWN_OPP"

    if "edge" not in out.columns:
        proj = pd.to_numeric(out.get("projection", ""), errors="coerce")
        line = pd.to_numeric(out.get("line",        ""), errors="coerce")
        out["edge"] = proj - line

    edge     = pd.to_numeric(out["edge"], errors="coerce")
    abs_edge = edge.abs()

    pick_type = out.get("pick_type", "Standard").astype(str).apply(_norm_pick_type)

    # BUG FIX: edge >= 0 is False for NaN, which wrongly assigns UNDER to rows
    # with no projection. Use explicit NaN check.
    has_edge  = edge.notna()
    model_dir = pd.Series(
        np.where(has_edge & (edge >= 0), "OVER",
                 np.where(has_edge & (edge < 0), "UNDER", "NO_EDGE")),
        index=out.index
    )
    out["model_dir"] = model_dir

    final_dir = model_dir.copy()
    reason    = pd.Series(
        np.where(has_edge & (abs_edge < 0.03), "MODEL_TIEBREAK_DIFF<0.03", ""),
        index=out.index
    )

    # Standard-only tie-break: when edge is effectively zero, use L5 side signal.
    # This preserves Goblin/Demon OVER-only behavior while allowing true Standard UNDERS.
    std_mask = pick_type.eq("Standard")
    tie_mask = std_mask & has_edge & (abs_edge < 0.03)
    if tie_mask.any():
        l5_over = pd.to_numeric(out.get("last5_over", np.nan), errors="coerce")
        l5_under = pd.to_numeric(out.get("last5_under", np.nan), errors="coerce")
        hr5 = pd.to_numeric(out.get("line_hit_rate_over_ou_5", np.nan), errors="coerce")

        under_from_l5 = tie_mask & l5_under.notna() & l5_over.notna() & (l5_under > l5_over)
        over_from_l5 = tie_mask & l5_under.notna() & l5_over.notna() & (l5_over > l5_under)
        unresolved = tie_mask & ~(under_from_l5 | over_from_l5)
        under_from_hr = unresolved & hr5.notna() & (hr5 < 0.50)
        over_from_hr = unresolved & ~(under_from_hr)

        final_dir = final_dir.where(~under_from_l5, "UNDER")
        final_dir = final_dir.where(~over_from_l5, "OVER")
        final_dir = final_dir.where(~under_from_hr, "UNDER")
        final_dir = final_dir.where(~over_from_hr, "OVER")

        reason = reason.where(~under_from_l5, "STANDARD_TIEBREAK_L5_UNDER")
        reason = reason.where(~over_from_l5, "STANDARD_TIEBREAK_L5_OVER")
        reason = reason.where(~under_from_hr, "STANDARD_TIEBREAK_HR5_UNDER")
        reason = reason.where(~over_from_hr, "STANDARD_TIEBREAK_HR5_OVER")

    # Additional Standard-only force from Last 5 Avg vs current line:
    # If we actually have stat_last5_avg and it is clearly UNDER the PP line,
    # flip to UNDER even if the edge model is near-zero.
    # (Helps cases where passes-attempted last5 stats exist but edge still ~0.)
    std_force = std_mask.copy()
    last5_avg = pd.to_numeric(out.get("stat_last5_avg", np.nan), errors="coerce")
    linev = pd.to_numeric(out.get("line", np.nan), errors="coerce")
    last5_known = last5_avg.notna() & linev.notna()
    force_under = std_force & last5_known & (last5_avg < (linev - 1e-9))
    force_over  = std_force & last5_known & (last5_avg > (linev + 1e-9))
    if force_under.any():
        final_dir = final_dir.where(~force_under, "UNDER")
        reason = reason.where(~force_under, "STANDARD_L5AVG_UNDER")
    if force_over.any():
        final_dir = final_dir.where(~force_over, "OVER")
        reason = reason.where(~force_over, "STANDARD_L5AVG_OVER")

    forced    = pick_type.isin(["Goblin", "Demon"])
    final_dir = final_dir.where(~forced, "OVER")
    reason    = reason.where(~forced, "FORCED_OVER_ONLY_GOB_DEM")

    out["final_bet_direction"] = final_dir
    out["final_dir_reason"]    = reason

    out.to_csv(args.output, index=False, encoding="utf-8-sig")
    print(f"Saved -> {args.output}")

    if out.empty:
        print("ERROR [PropOracle-Soccer-S8] Output is empty — aborting.")
        sys.exit(1)

    print("final_bet_direction counts:")
    print(pd.Series(out["final_bet_direction"]).value_counts().to_string())
    if "tier" in out.columns:
        print("tier counts:")
        print(out["tier"].value_counts().to_string())

    # Always use explicit --xlsx path (default: step8_soccer_direction_clean.xlsx)
    xlsx_path = args.xlsx if args.xlsx else "step8_soccer_direction_clean.xlsx"
    try:
        build_clean_xlsx(out, xlsx_path)
    except Exception as e:
        print(f"WARN build_clean_xlsx failed: {e}")
        print("   Writing raw fallback xlsx so combined pipeline can proceed...")
        try:
            with pd.ExcelWriter(xlsx_path, engine="openpyxl") as w:
                out.to_excel(w, sheet_name="ALL", index=False)
                for _tier in ["A", "B", "C", "D"]:
                    _mask = out.get("tier", pd.Series(dtype=str)) == _tier
                    _void = out.get("void_reason", pd.Series("", index=out.index)).fillna("")
                    _elig = out[_mask & (_void == "")].copy() if _mask.any() else out.head(0)
                    _elig.to_excel(w, sheet_name=f"Tier {_tier}", index=False)
            print(f"Fallback xlsx saved -> {xlsx_path}")
        except Exception as e2:
            print(f"ERROR Fallback xlsx also failed: {e2}")


if __name__ == "__main__":
    main()
