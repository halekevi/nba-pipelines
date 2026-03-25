#!/usr/bin/env python3
"""
Build ticket_eval_{date}.html (dated archive for /grades) and tickets_latest.html (main /tickets page).
Reads ticket JSON (or combined_slate_tickets_{date}.xlsx) and sport step8/graded workbooks,
matches legs to actuals, writes self-contained HTML.

Run after combined_slate_tickets.py --write-web (JSON only) so tickets_latest.html includes grades.
"""
from __future__ import annotations

import argparse
import html
import json
import re
from datetime import date, timedelta
from pathlib import Path
from typing import Any

import pandas as pd

REPO_ROOT = Path(__file__).resolve().parent
TEMPLATES_DIR = REPO_ROOT / "ui_runner" / "templates"

# Ticket source search order: dated JSON → dated xlsx (repo root) → tickets_latest.json
DATED_TICKET_JSON = "combined_slate_tickets_{date}.json"
FALLBACK_TICKET_JSON = TEMPLATES_DIR / "tickets_latest.json"

_XLSX_HDR_TO_LEG_FIELD: dict[str, str] = {
    "player": "player",
    "team": "team",
    "opp": "opp",
    "prop": "prop_type",
    "pick type": "pick_type",
    "line": "line",
    "dir": "direction",
    "edge": "edge",
    "hit rate": "hit_rate",
    "l5 avg": "l5_avg",
    "szn avg": "season_avg",
    "sport": "sport",
}

# Slate workbooks per sport bucket (first existing file wins within that bucket).
# Ticket legs with sport NBA1H / NBA1Q / WCBB must match rows from these files, not full-game NBA/CBB only.
SPORT_XLSX_CANDIDATES: dict[str, list[Path]] = {
    "NBA": [
        REPO_ROOT / "NBA" / "step8_all_direction_clean.xlsx",
        REPO_ROOT / "NBA" / "data" / "outputs" / "step8_all_direction_clean.xlsx",
    ],
    "NBA1H": [
        REPO_ROOT / "NBA" / "step8_nba1h_direction_clean.xlsx",
    ],
    "NBA1Q": [
        REPO_ROOT / "NBA" / "step8_nba1q_direction_clean.xlsx",
    ],
    "CBB": [
        REPO_ROOT / "CBB" / "step6_ranked_cbb.xlsx",
    ],
    "WCBB": [
        REPO_ROOT / "CBB" / "step6_ranked_wcbb.xlsx",
    ],
    "NHL": [
        REPO_ROOT / "NHL" / "step8_nhl_direction_clean.xlsx",
        REPO_ROOT / "NHL" / "outputs" / "step8_nhl_direction_clean.xlsx",
    ],
    "SOCCER": [
        REPO_ROOT / "Soccer" / "step8_soccer_direction_clean.xlsx",
        REPO_ROOT / "Soccer" / "outputs" / "step8_soccer_direction_clean.xlsx",
    ],
    "MLB": [
        REPO_ROOT / "MLB" / "step8_mlb_direction_clean.xlsx",
    ],
}

# Shell theme for tickets_latest.html / ticket_eval_* (tokens + background + nav match index.html).
_TICKETS_NAV_THEME_CSS = r'''/* ── CSS Variables (matches index.html) ── */
:root{
  --bg-main:#050505;
  --glass-bg:rgba(20,20,20,0.6);
  --glass-border:rgba(212,175,55,0.15);
  --text-primary:#ffffff;
  --accent-gold:#d4af37;
  --bg:#05050f;--bg2:rgba(255,255,255,0.03);--bg3:rgba(255,255,255,0.05);
  --border:rgba(255,255,255,0.08);--bd2:rgba(255,255,255,0.14);
  --text:rgba(255,255,255,0.92);--muted:rgba(255,255,255,0.45);--muted2:rgba(255,255,255,0.25);
  --accent:#d4af37;--cyan:#7fc7d9;
  --green:#9fd8a7;--amber:#c89a4a;--red:#d07a78;--purple:#8f7ab8;--blue:#7fc7d9;
  --glass:rgba(255,255,255,0.03);
  --glass-mid:rgba(255,255,255,0.055);
  --glass-hi:rgba(255,255,255,0.08);
  --glass-border:rgba(255,255,255,0.09);
  --glass-border-hi:rgba(255,255,255,0.16);
  --glass-shadow:0 8px 32px rgba(0,0,0,0.45), 0 1px 0 rgba(255,255,255,0.06) inset;
  --glass-shadow-lg:0 20px 60px rgba(0,0,0,0.55), 0 1px 0 rgba(255,255,255,0.07) inset;
  --blur:blur(22px) saturate(180%);
  --glass-border:rgba(212,175,55,0.15);
  --gold:#f0a500;
  --gold2:#d4a017;
  --pending:#666;
  --glass-bd:var(--glass-border);
}
[data-theme='light']{
  --bg-main:#f0f2f5;
  --glass-bg:rgba(255,255,255,0.7);
  --glass-border:rgba(255,255,255,0.4);
  --text-primary:#1d1d1f;
  --accent-gold:#c5a059;
  --bg:#f0f2f5;--bg2:rgba(0,0,0,0.03);--bg3:rgba(0,0,0,0.05);
  --border:rgba(0,0,0,0.10);--bd2:rgba(0,0,0,0.14);
  --text:rgba(0,0,0,0.88);--muted:rgba(0,0,0,0.45);--muted2:rgba(0,0,0,0.25);
  --accent:#c5a059;--cyan:#4a9ab5;
  --green:#3a7d44;--amber:#a07830;--red:#a04040;--purple:#6a5a98;
  --glass-border:rgba(0,0,0,0.12);
}

/* ── Reset ── */
*{box-sizing:border-box;margin:0;padding:0}

/* ── Body + deep space background (matches index.html exactly) ── */
body{
  font-family:'Inter',sans-serif;
  background:var(--bg-main);
  color:var(--text-primary);
  min-height:100vh;
  padding-top:16px;
  padding-bottom:80px;
  overflow-x:hidden;
}
body::before{
  content:'';position:fixed;inset:0;
  background:
    radial-gradient(ellipse at 12% 8%, rgba(127,199,217,0.18) 0%, transparent 38%),
    radial-gradient(ellipse at 88% 6%, rgba(212,175,55,0.13) 0%, transparent 36%),
    radial-gradient(ellipse at 50% 55%, rgba(183,168,255,0.07) 0%, transparent 50%),
    radial-gradient(ellipse at 20% 85%, rgba(212,175,55,0.06) 0%, transparent 35%),
    radial-gradient(ellipse at 80% 80%, rgba(127,199,217,0.08) 0%, transparent 32%),
    linear-gradient(180deg,#040404 0%,#090909 50%,#111111 100%);
  pointer-events:none;z-index:0;
}
body::after{
  content:'';position:fixed;inset:0;
  background-image:
    radial-gradient(1px 1px at 20% 30%, rgba(255,255,255,0.5) 0%, transparent 100%),
    radial-gradient(1px 1px at 60% 15%, rgba(255,255,255,0.4) 0%, transparent 100%),
    radial-gradient(1px 1px at 80% 60%, rgba(255,255,255,0.3) 0%, transparent 100%),
    radial-gradient(1px 1px at 40% 80%, rgba(255,255,255,0.35) 0%, transparent 100%),
    radial-gradient(1px 1px at 90% 35%, rgba(255,255,255,0.3) 0%, transparent 100%),
    linear-gradient(rgba(212,175,55,.012) 1px,transparent 1px),
    linear-gradient(90deg,rgba(212,175,55,.012) 1px,transparent 1px);
  background-size:auto,auto,auto,auto,auto,52px 52px,52px 52px;
  pointer-events:none;z-index:0;mix-blend-mode:screen;opacity:.6;
}
::-webkit-scrollbar{width:4px;height:4px}
::-webkit-scrollbar-track{background:rgba(255,255,255,0.02)}
::-webkit-scrollbar-thumb{background:rgba(255,255,255,0.12);border-radius:4px}

/* ── Nav (matches index.html exactly) ── */
.snav{position:sticky;top:1rem;z-index:200;background:var(--glass-bg);backdrop-filter:blur(40px) saturate(200%);-webkit-backdrop-filter:blur(40px) saturate(200%);border:1px solid var(--glass-border);box-shadow:0 1px 0 rgba(255,255,255,0.04),0 8px 32px rgba(0,0,0,0.4);padding:0 40px;display:flex;align-items:center;height:72px;gap:0;margin:1rem;border-radius:20px;position:relative;}
.nav-accent{position:absolute;top:0;left:0;right:0;height:2px;background:linear-gradient(90deg,transparent 0%,#c6ff00 15%,var(--cyan) 40%,var(--purple) 65%,#c6ff00 85%,transparent 100%);opacity:.7;animation:accentShift 6s ease-in-out infinite alternate;border-radius:20px 20px 0 0;}
@keyframes accentShift{from{opacity:.5;filter:hue-rotate(0deg);}to{opacity:.9;filter:hue-rotate(20deg);}}
.snav-brand{display:flex;align-items:center;gap:14px;text-decoration:none;margin-right:48px;flex-shrink:0;}
.hybrid-logo{display:flex;align-items:center;gap:10px;}
.logo-icon{width:182px;height:60px;object-fit:contain;display:block;filter:drop-shadow(0 0 6px rgba(212,175,55,0.45));}
.nav-brand{font-family:'Inter',sans-serif;font-size:22px;font-weight:700;letter-spacing:-0.5px;line-height:1;color:var(--text-primary);display:flex;align-items:baseline;}
.oracle-upper{text-transform:uppercase;font-weight:800;color:var(--accent-gold);}
.gold-shimmer{background:linear-gradient(to bottom,#d4af37,#f7ef8a);-webkit-background-clip:text;background-clip:text;color:transparent;}
.snav-links{display:flex;align-items:stretch;list-style:none;flex:1;gap:4px;}
.snav-links li a{display:flex;align-items:center;gap:10px;padding:0 20px;height:72px;font-family:'Inter',sans-serif;font-size:11px;font-weight:600;letter-spacing:1.5px;text-transform:uppercase;color:#7a9ac8;text-decoration:none;border-bottom:2px solid transparent;transition:color .2s,border-color .2s,background .2s;position:relative;}
.snav-links li a:hover{color:var(--text);background:rgba(255,255,255,.02);}
.snav-links li a.active{color:var(--accent);border-bottom-color:var(--accent);}
.snav-links li a.active::after{content:'';position:absolute;bottom:-1px;left:50%;transform:translateX(-50%);width:6px;height:6px;border-radius:50%;background:var(--accent);box-shadow:0 0 12px var(--accent),0 0 24px rgba(212,175,55,.5);}
.snav-links li:nth-child(1) a:hover{color:var(--cyan);}
.snav-links li:nth-child(2) a:hover{color:var(--accent);}
.snav-links li:nth-child(3) a:hover{color:var(--purple);}
.snav-links li:nth-child(4) a:hover{color:var(--green);}
.ni{width:26px;height:26px;border-radius:6px;display:flex;align-items:center;justify-content:center;flex-shrink:0;transition:transform .2s,box-shadow .2s;}
.snav-links li a:hover .ni{transform:scale(1.1) translateY(-1px);}
.ni-ctrl{background:rgba(127,199,217,.08);border:1px solid rgba(127,199,217,.2);color:var(--cyan);}
.ni-tick{background:rgba(183,168,255,.08);border:1px solid rgba(183,168,255,.2);color:var(--purple);}
.ni-grade{background:rgba(0,255,170,.08);border:1px solid rgba(0,255,170,.2);color:var(--green);}
.ni-pay{background:rgba(255,192,107,.08);border:1px solid rgba(255,192,107,.2);color:var(--amber);}
.snav-right{display:flex;align-items:center;gap:16px;margin-left:auto;}
.live-pill{font-family:'Inter',sans-serif;font-weight:600;font-size:10px;letter-spacing:1.5px;padding:6px 16px;border-radius:100px;border:1px solid rgba(46,204,113,.3);color:var(--green);background:rgba(46,204,113,.06);display:flex;align-items:center;gap:8px;}
.live-dot{width:6px;height:6px;border-radius:50%;background:var(--green);box-shadow:0 0 8px var(--green);animation:blink 2s ease-in-out infinite;}
@keyframes blink{0%,100%{opacity:1}50%{opacity:.2}}
.hamburger{display:none;flex-direction:column;gap:4px;cursor:pointer;padding:8px;border:1px solid rgba(255,255,255,0.09);border-radius:8px;background:rgba(255,255,255,0.04);backdrop-filter:blur(10px) saturate(180%);}
.hamburger span{display:block;width:18px;height:2px;background:var(--muted);border-radius:2px;transition:all .25s;}
.hamburger.open span:nth-child(1){transform:translateY(6px) rotate(45deg);}
.hamburger.open span:nth-child(2){opacity:0;}
.hamburger.open span:nth-child(3){transform:translateY(-6px) rotate(-45deg);}
.mobile-menu{display:none;position:fixed;top:88px;left:0;right:0;background:rgba(5,5,15,.98);backdrop-filter:blur(20px) saturate(180%);border-bottom:1px solid var(--border);z-index:199;padding:12px 0;}
.mobile-menu.open{display:block;}
.mobile-menu a{display:flex;align-items:center;gap:12px;padding:12px 24px;font-size:12px;letter-spacing:1.5px;text-transform:uppercase;color:var(--muted);text-decoration:none;border-left:3px solid transparent;transition:all .15s;}
.mobile-menu a:hover,.mobile-menu a.active{color:var(--accent);border-left-color:var(--accent);background:rgba(212,175,55,.04);}

/* ── Theme toggle ── */
.theme-toggle{display:flex;align-items:center;justify-content:center;width:36px;height:36px;border-radius:999px;background:rgba(255,255,255,0.07);border:1px solid rgba(255,255,255,0.12);cursor:pointer;font-size:16px;transition:background .2s,border-color .2s,transform .15s;flex-shrink:0;backdrop-filter:blur(10px) saturate(180%);-webkit-backdrop-filter:blur(10px) saturate(180%);}
.theme-toggle:hover{background:rgba(255,255,255,0.13);border-color:rgba(255,255,255,0.22);transform:scale(1.08);}
[data-theme='light'] .theme-toggle{background:rgba(0,0,0,0.06)!important;border-color:rgba(0,0,0,0.12)!important;}
[data-theme='light'] .theme-toggle:hover{background:rgba(0,0,0,0.10)!important;}

/* ── Light mode overrides for ticket eval content ── */
[data-theme='light'] body::before{
  background:
    radial-gradient(ellipse at 12% 8%, rgba(127,199,217,0.12) 0%, transparent 38%),
    radial-gradient(ellipse at 88% 6%, rgba(212,175,55,0.10) 0%, transparent 36%),
    linear-gradient(180deg,#f0f2f5 0%,#e8eaf0 100%);
}
[data-theme='light'] body::after{opacity:0.08;}
[data-theme='light'] .stats-bar{background:rgba(255,255,255,0.6);border-color:rgba(0,0,0,0.1);box-shadow:0 4px 24px rgba(0,0,0,0.08);}
[data-theme='light'] .ticket-card{background:rgba(255,255,255,0.65);border-color:rgba(0,0,0,0.1);box-shadow:0 4px 20px rgba(0,0,0,0.08);}
[data-theme='light'] .thdr{background:rgba(0,0,0,0.04);}
[data-theme='light'] .legrow{color:rgba(0,0,0,0.85);}
[data-theme='light'] .meta-muted{color:rgba(0,0,0,0.5);}
[data-theme='light'] .sum-val{color:#b8860b;}
[data-theme='light'] .sum-lab{color:rgba(0,0,0,0.5);}
[data-theme='light'] .sec-head{color:#b8860b;}
[data-theme='light'] .slate-kicker{color:rgba(0,0,0,0.45);}
[data-theme='light'] .thdr .tn{color:#b8860b;}
[data-theme='light'] .thdr .tg{color:rgba(0,0,0,0.5);}
[data-theme='light'] .payout{color:#1a6b7a;}
[data-theme='light'] .dir-over{color:#1a6b7a;}
[data-theme='light'] .dir-under{color:#b8860b;}
[data-theme='light'] .tier{color:#b8860b;border-color:rgba(0,0,0,0.12);background:rgba(0,0,0,0.04);}
[data-theme='light'] .pill{border-color:rgba(0,0,0,0.15)!important;}
[data-theme='light'] .banner.pend{color:rgba(0,0,0,0.4);border-color:rgba(0,0,0,0.12);}

/* Light: nav chrome + semantics (matches index.html — variables alone are not enough for .snav glass) */
[data-theme='light']{
  --accent:#4b6f96;
  --cyan:#2c7fb8;
  --green:#2e7f7a;
  --amber:#8f7396;
  --red:#b5687e;
  --purple:#6f84bf;
  --muted:rgba(23,43,64,0.52);
  --muted2:rgba(23,43,64,0.34);
  --text:rgba(14,36,58,0.90);
}
[data-theme='light'] .snav,
[data-theme='light'] nav{
  background:rgba(255,255,255,0.72)!important;
  border:1px solid rgba(0,0,0,0.08)!important;
  box-shadow:0 1px 0 rgba(255,255,255,0.9),0 4px 20px rgba(0,0,0,0.08)!important;
}
[data-theme='light'] .snav-links li a,
[data-theme='light'] .nav-links a{
  color:rgba(0,0,0,0.48)!important;
}
[data-theme='light'] .snav-links li a:hover,
[data-theme='light'] .nav-links a:hover{
  color:rgba(0,0,0,0.80)!important;
  background:rgba(0,0,0,0.04)!important;
}
[data-theme='light'] .snav-links li a.active,
[data-theme='light'] .nav-links a.active{
  color:var(--accent)!important;
  background:rgba(75,111,150,0.08)!important;
  border-color:rgba(75,111,150,0.25)!important;
}
[data-theme='light'] .live-pill{
  background:rgba(255,255,255,0.62)!important;
  border-color:rgba(26,140,69,0.30)!important;
  color:var(--green)!important;
}
[data-theme='light'] .hamburger{
  background:rgba(255,255,255,0.62)!important;
  border-color:rgba(0,0,0,0.10)!important;
}
[data-theme='light'] .hamburger span{background:rgba(0,0,0,0.55)!important;}
[data-theme='light'] .mobile-menu{
  background:rgba(245,248,255,0.96)!important;
  border-bottom-color:rgba(0,0,0,0.08)!important;
}
[data-theme='light'] .mobile-menu a{color:rgba(0,0,0,0.55)!important;}
[data-theme='light'] .mobile-menu a:hover,
[data-theme='light'] .mobile-menu a.active{
  color:var(--accent)!important;
  background:rgba(75,111,150,0.05)!important;
  border-left-color:var(--accent)!important;
}

/* Ticket page: collapsible sport / cross-sport buckets */
.ticket-bucket{margin-bottom:18px;border-radius:16px;border:1px solid var(--glass-bd);background:rgba(255,255,255,0.025);overflow:hidden;backdrop-filter:blur(14px) saturate(160%);-webkit-backdrop-filter:blur(14px) saturate(160%);}
.ticket-bucket > summary{cursor:pointer;list-style:none;display:flex;flex-wrap:wrap;align-items:center;justify-content:space-between;gap:10px 16px;padding:14px 18px;font-family:'Bebas Neue',sans-serif;font-size:clamp(22px,2.6vw,30px);letter-spacing:2px;color:var(--gold);user-select:none;}
.ticket-bucket > summary::-webkit-details-marker{display:none}
.ticket-bucket > summary::after{content:'▸';font-size:16px;opacity:.55;transition:transform .2s ease;color:var(--muted);flex-shrink:0;}
.ticket-bucket[open] > summary::after{transform:rotate(90deg)}
.ticket-bucket-meta{font-family:'Share Tech Mono',monospace;font-size:clamp(11px,1.1vw,13px);letter-spacing:.6px;font-weight:400;color:var(--muted);text-align:right;max-width:100%;}
.ticket-bucket-body{padding:4px 12px 16px;}
.ticket-bucket.sb-nba{border-left:4px solid rgba(240,165,0,.8);}
.ticket-bucket.sb-nba1h{border-left:4px solid rgba(255,155,86,.85);}
.ticket-bucket.sb-nba1q{border-left:4px solid rgba(255,214,102,.85);}
.ticket-bucket.sb-cbb{border-left:4px solid rgba(0,229,255,.65);}
.ticket-bucket.sb-wcbb{border-left:4px solid rgba(159,216,232,.75);}
.ticket-bucket.sb-nhl{border-left:4px solid rgba(196,165,255,.8);}
.ticket-bucket.sb-soccer{border-left:4px solid rgba(232,184,74,.8);}
.ticket-bucket.sb-mlb{border-left:4px solid rgba(255,154,154,.8);}
.ticket-bucket.sb-xsport{border-left:4px solid rgba(183,168,255,.85);}
.ticket-bucket.sb-default{border-left:4px solid rgba(255,255,255,.2);}
[data-theme='light'] .ticket-bucket{background:rgba(255,255,255,0.52);border-color:rgba(0,0,0,0.1);box-shadow:0 2px 16px rgba(0,0,0,0.06);}
[data-theme='light'] .ticket-bucket > summary{color:#b8860b;}
[data-theme='light'] .ticket-bucket-meta{color:rgba(0,0,0,0.45);}
[data-theme='light'] .ticket-bucket.sb-default{border-left-color:rgba(0,0,0,0.15);}

/* ── Mobile ── */
@media(max-width:768px){
  .snav{padding:0 16px;margin:.5rem;}
  .snav-links{display:none;}
  .hamburger{display:flex;}
  .snav-brand{margin-right:0;}
  .logo-icon{width:120px;height:40px;}
}
@media(max-width:900px){
  .legrow{grid-template-columns:52px 80px 1fr;gap:10px;padding:12px;font-size:14px;}
  .leg-extra{display:none;}
  .stats-bar{padding:14px 16px;}
  .sum-val{font-size:22px;}
}
'''

_TICKETS_NAV_HTML = """<nav class="snav">
  <div class="nav-accent"></div>
  <a class="snav-brand" href="/">
    <img src="/static/proporacle-logo-v3.png?v=20260320b" alt="PropORACLE" class="logo-icon">
    <span class="nav-brand">Prop<span class="oracle-upper gold-shimmer">ORACLE</span></span>
  </a>
  <ul class="snav-links">
    <li><a href="/">Home</a></li>
    <li><a href="/tickets" class="active">Tickets</a></li>
    <li><a href="/grades">Grades</a></li>
    <li><a href="/payout">Payouts</a></li>
  </ul>
  <div class="snav-right">
    <button class="theme-toggle" id="theme-toggle" onclick="toggleTheme()"
            title="Toggle light/dark mode" aria-label="Toggle theme">
      <span class="tt-moon">🌙</span>
      <span class="tt-sun" style="display:none">☀️</span>
    </button>
    <div class="live-pill"><div class="live-dot"></div>LIVE</div>
  </div>
  <button class="hamburger" id="hamburger">
    <span></span><span></span><span></span>
  </button>
</nav>
<div class="mobile-menu" id="mobile-menu">
  <a href="/">🏠 Home</a>
  <a href="/tickets" class="active">🎟 Tickets</a>
  <a href="/grades">✏️ Grades</a>
  <a href="/payout">💰 Payouts</a>
</div>"""

_TICKETS_THEME_JS = """<script>
(function(){
  const saved = localStorage.getItem('proporacle-theme') || 'dark';
  applyTheme(saved);
})();
function applyTheme(theme) {
  document.documentElement.setAttribute('data-theme', theme);
  localStorage.setItem('proporacle-theme', theme);
  localStorage.setItem('theme', theme);
  const moon = document.getElementById('theme-toggle')?.querySelector('.tt-moon');
  const sun  = document.getElementById('theme-toggle')?.querySelector('.tt-sun');
  if (moon) moon.style.display = theme === 'dark' ? '' : 'none';
  if (sun)  sun.style.display  = theme === 'dark' ? 'none' : '';
}
function toggleTheme() {
  const curr = document.documentElement.getAttribute('data-theme') || 'dark';
  applyTheme(curr === 'dark' ? 'light' : 'dark');
}
const ham = document.getElementById('hamburger');
const mob = document.getElementById('mobile-menu');
if (ham) ham.addEventListener('click', () => {
  ham.classList.toggle('open');
  mob.classList.toggle('open');
});
document.addEventListener('click', e => {
  if (ham && mob && !ham.contains(e.target) && !mob.contains(e.target)) {
    ham.classList.remove('open');
    mob.classList.remove('open');
  }
});
</script>"""


def _dated_candidates(date_str: str) -> dict[str, list[Path]]:
    """
    Returns a copy of SPORT_XLSX_CANDIDATES with dated archive paths prepended
    for each sport bucket. Dated paths follow the naming convention used by
    run_pipeline.ps1 archive step. Only paths that actually exist are prepended.
    """
    dated_dir = REPO_ROOT / "outputs" / date_str
    dated_map = {
        "NBA": dated_dir / f"step8_nba_direction_clean_{date_str}.xlsx",
        "NBA1H": dated_dir / f"step8_nba1h_direction_clean_{date_str}.xlsx",
        "NBA1Q": dated_dir / f"step8_nba1q_direction_clean_{date_str}.xlsx",
        "CBB": dated_dir / f"step6_ranked_cbb_{date_str}.xlsx",
        "WCBB": dated_dir / f"step6_ranked_wcbb_{date_str}.xlsx",
        "NHL": dated_dir / f"step8_nhl_direction_clean_{date_str}.xlsx",
        "SOCCER": dated_dir / f"step8_soccer_direction_clean_{date_str}.xlsx",
        "MLB": dated_dir / f"step8_mlb_direction_clean_{date_str}.xlsx",
    }
    result: dict[str, list[Path]] = {}
    for bucket, live_paths in SPORT_XLSX_CANDIDATES.items():
        dated_path = dated_map.get(bucket)
        if dated_path and dated_path.is_file():
            result[bucket] = [dated_path] + list(live_paths)
        else:
            result[bucket] = list(live_paths)
    return result


def _norm_header(s: str) -> str:
    return re.sub(r"\s+", " ", str(s).strip().lower())


def _canon_player(row: dict[str, Any]) -> str:
    for k in ("player", "athlete", "name"):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _canon_prop(row: dict[str, Any]) -> str:
    for k in (
        "prop_type",
        "prop type",
        "prop_type_norm",
        "prop_norm",
        "prop_label",
        "prop",
        "prop_display",
        "stat_type",
    ):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip()
    return ""


def _canon_direction(row: dict[str, Any]) -> str:
    for k in ("direction", "bet_direction", "final_bet_direction", "pick direction"):
        v = row.get(k)
        if v is not None and str(v).strip():
            return str(v).strip().upper()
    return ""


def _canon_line(row: dict[str, Any]) -> float | None:
    for k in ("line", "line_num"):
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        try:
            return float(v)
        except (TypeError, ValueError):
            continue
    return None


def _canon_actual(row: dict[str, Any]) -> float | None:
    for k in (
        "actual",
        "actual_value",
        "act",
        "result_value",
        "stat_actual",
        "final_stat",
        "box",
        "box_score",
        "game_stat",
        "stat",
        "final",
    ):
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        if isinstance(v, str) and not v.strip():
            continue
        try:
            return float(v)
        except (TypeError, ValueError):
            continue
    return None


def _cell_looks_like_grade_outcome(s: str) -> bool:
    """
    True if a workbook cell is probably HIT/MISS/etc., not a numeric game stat.
    Prevents columns named 'result' that hold 14.0 from forcing the wrong path.
    """
    u = str(s).strip().upper()
    if not u or u in (".", "-", "—"):
        return False
    if u in (
        "HIT",
        "WIN",
        "W",
        "MISS",
        "LOSS",
        "L",
        "VOID",
        "PUSH",
        "PENDING",
        "N/A",
        "NA",
        "TBD",
        "OPEN",
        "TRUE",
        "FALSE",
        "YES",
        "NO",
        "0",
        "1",
    ):
        return True
    if re.fullmatch(r"-?\d+\.?\d*", u):
        return False
    if len(u) <= 16 and re.fullmatch(r"[A-Z][A-Z0-9_/-]*", u):
        return True
    return False


def _canon_grade_raw(row: dict[str, Any]) -> str:
    for k in ("grade", "leg_result", "outcome", "result"):
        v = row.get(k)
        if v is None or (isinstance(v, float) and pd.isna(v)):
            continue
        s = str(v).strip().upper()
        if s and _cell_looks_like_grade_outcome(s):
            return s
    return ""


def _normalize_workbook_rows(path: Path) -> list[dict[str, Any]]:
    """Load all sheets; normalize headers to lowercase single-space keys."""
    xl = pd.ExcelFile(path)
    out: list[dict[str, Any]] = []
    for sheet in xl.sheet_names:
        df = pd.read_excel(path, sheet_name=sheet)
        if df.empty:
            continue
        df.columns = [_norm_header(c) for c in df.columns]
        out.extend(df.to_dict(orient="records"))
    return out


def _leg_grade(
    actual: float | None,
    line: float | None,
    direction: str,
    grade_col: str,
) -> str:
    g = (grade_col or "").strip().upper()
    if g in ("HIT", "WIN", "W", "1", "TRUE", "YES"):
        return "HIT"
    if g in ("MISS", "LOSS", "L", "0", "FALSE", "NO"):
        return "MISS"
    if g in ("VOID", "PUSH", "N/A", "NA"):
        return "VOID"
    if actual is None or line is None:
        return "UNGRADED"
    d = direction.upper()
    if d == "OVER" and actual >= line:
        return "HIT"
    if d == "UNDER" and actual <= line:
        return "HIT"
    return "MISS"


def _pick_type_tier(pick_type: str) -> str:
    p = (pick_type or "").strip().lower()
    if "goblin" in p:
        return "G"
    if "demon" in p:
        return "D"
    if "standard" in p:
        return "S"
    return (pick_type[:1].upper() if pick_type else "?")


def _sport_key(sport: str) -> str:
    """Normalize for display / CSS (keep variant labels visible)."""
    s = (sport or "").strip().upper().replace(" ", "")
    if s in ("SOC", "MLS", "EPL"):
        return "SOCCER"
    return s


def _ticket_group_bucket(group_name: str) -> str:
    """Bucket for collapsible UI: sport prefix from sheet title, or Cross-sport for XSPORT*."""
    n = (group_name or "").strip()
    if not n:
        return "Other"
    if n.upper().startswith("XSPORT"):
        return "Cross-sport"
    return n.split()[0] or "Other"


def _ticket_bucket_skin_class(bucket: str) -> str:
    if bucket == "Cross-sport":
        return "sb-xsport"
    sk = _sport_key(bucket)
    return {
        "NBA": "sb-nba",
        "NBA1H": "sb-nba1h",
        "NBA1Q": "sb-nba1q",
        "CBB": "sb-cbb",
        "WCBB": "sb-wcbb",
        "NHL": "sb-nhl",
        "SOCCER": "sb-soccer",
        "MLB": "sb-mlb",
    }.get(sk, "sb-default")


def _bucket_ticket_groups(groups: list[dict[str, Any]]) -> list[tuple[str, list[dict[str, Any]]]]:
    order: list[str] = []
    m: dict[str, list[dict[str, Any]]] = {}
    for g in groups:
        key = _ticket_group_bucket(str(g.get("group_name") or ""))
        if key not in m:
            m[key] = []
            order.append(key)
        m[key].append(g)
    return [(k, m[k]) for k in order]


def _leg_match_buckets(sport: str) -> list[str]:
    """
    Order matters: try variant-specific slate first, then parent sport fallback.
    """
    s = (sport or "").strip().upper().replace(" ", "").replace("-", "")
    if s in ("NBA1H", "NBA_1H"):
        return ["NBA1H", "NBA"]
    if s in ("NBA1Q", "NBA_1Q"):
        return ["NBA1Q", "NBA"]
    if s == "WCBB":
        return ["WCBB", "CBB"]
    if s in ("SOC", "MLS", "EPL"):
        return ["SOCCER"]
    if s in ("NBA", "WNBA"):
        return ["NBA"]
    if s == "CBB":
        return ["CBB"]
    if s == "NHL":
        return ["NHL"]
    if s == "SOCCER":
        return ["SOCCER"]
    if s == "MLB":
        return ["MLB"]
    return [s, "NBA", "CBB"]


def _ingest_workbook_rows_into_index(
    rows: list[dict[str, Any]],
    triple: dict[tuple[str, str, str], dict],
    pair_buckets: dict[tuple[str, str], list[dict]],
) -> None:
    for raw in rows:
        pl = _canon_player(raw).lower()
        pt = _canon_prop(raw).lower()
        dr = _canon_direction(raw)
        if not pl or not pt:
            continue
        row = {
            "player_lower": pl,
            "prop_lower": pt,
            "direction": dr,
            "line": _canon_line(raw),
            "actual": _canon_actual(raw),
            "grade_raw": _canon_grade_raw(raw),
        }
        key3 = (pl, pt, dr)
        triple[key3] = row
        pair_buckets.setdefault((pl, pt), []).append(row)


def _graded_xlsx_in_outputs_date(arg_date: str) -> list[Path]:
    """Graded slates dropped next to other daily artifacts: outputs/YYYY-MM-DD/*.xlsx."""
    d = REPO_ROOT / "outputs" / arg_date
    if not d.is_dir():
        return []
    found: set[Path] = set()
    for pat in ("graded_*.xlsx", "*_graded_*.xlsx"):
        for p in d.glob(pat):
            if not p.is_file():
                continue
            low = p.name.lower()
            # Avoid matching combined_tickets_graded_*.xlsx from *_graded_* glob (not a sport slate).
            if "combined_tickets_graded" in low:
                continue
            found.add(p)
    return sorted(found, key=lambda x: x.name.lower())


def _sport_buckets_for_graded_filename(path: Path) -> list[str]:
    """
    Map a graded workbook name to one or more SPORT_XLSX_CANDIDATES keys.
    Unknown names return [] (skipped).
    """
    n = path.name.lower()
    s = path.stem.lower()
    if "mlb" in n:
        return ["MLB"]
    if "nhl" in n:
        return ["NHL"]
    if "soccer" in n or s.startswith("soccer_graded"):
        return ["SOCCER"]
    if "wcbb" in n or "wcbb" in s:
        return ["WCBB"]
    if "cbb" in n or "cbb" in s or "ncaab" in n:
        return ["CBB"]
    if "nba1h" in n or "nba_1h" in n:
        return ["NBA1H"]
    if "nba1q" in n or "nba_1q" in n:
        return ["NBA1Q"]
    if "nba" in n:
        return ["NBA"]
    return []


def _merge_graded_workbooks_into_indices(
    indices: dict[str, tuple[dict[tuple[str, str, str], dict], dict[tuple[str, str], list[dict]]]],
    graded_paths: list[Path],
) -> int:
    """
    Overlay rows from dated graded exports so Actual / Result columns populate ticket eval.
    Returns number of workbook files successfully merged.
    """
    merged = 0
    for path in graded_paths:
        buckets = _sport_buckets_for_graded_filename(path)
        if not buckets:
            continue
        try:
            rows = _normalize_workbook_rows(path)
        except Exception:
            continue
        if not rows:
            continue
        for bkt in buckets:
            trip, pairs = indices.get(bkt, ({}, {}))
            _ingest_workbook_rows_into_index(rows, trip, pairs)
            indices[bkt] = (trip, pairs)
        merged += 1
    return merged


def _load_actuals_indices(
    sport_candidates: dict[str, list[Path]],
    arg_date: str | None = None,
) -> dict[str, tuple[dict[tuple[str, str, str], dict], dict[tuple[str, str], list[dict]]]]:
    """Per sport-bucket indices (NBA1H separate from NBA, etc.)."""
    out: dict[str, tuple[dict, dict]] = {}
    for bucket, paths in sport_candidates.items():
        triple: dict[tuple[str, str, str], dict] = {}
        pair_buckets: dict[tuple[str, str], list[dict]] = {}
        src = next((p for p in paths if p.is_file()), None)
        if not src:
            out[bucket] = (triple, pair_buckets)
            continue
        try:
            rows = _normalize_workbook_rows(src)
        except Exception:
            out[bucket] = (triple, pair_buckets)
            continue
        _ingest_workbook_rows_into_index(rows, triple, pair_buckets)
        out[bucket] = (triple, pair_buckets)

    if arg_date and re.match(r"^\d{4}-\d{2}-\d{2}$", arg_date):
        graded = _graded_xlsx_in_outputs_date(arg_date)
        _merge_graded_workbooks_into_indices(out, graded)

        graded_dir = REPO_ROOT / "outputs" / arg_date
        if graded_dir.is_dir():
            sport_to_bucket = {
                "nba": "NBA",
                "cbb": "CBB",
                "nhl": "NHL",
                "soccer": "SOCCER",
                "wcbb": "WCBB",
                "mlb": "MLB",
                "nba1h": "NBA1H",
                "nba1q": "NBA1Q",
            }
            for graded_file in sorted(graded_dir.glob(f"graded_*_{arg_date}.xlsx")):
                m = re.match(r"^graded_(.+)_(\d{4}-\d{2}-\d{2})$", graded_file.stem)
                if not m:
                    continue
                sport_tag = m.group(1).lower()
                bucket = sport_to_bucket.get(sport_tag)
                if bucket is None:
                    continue
                try:
                    xl = pd.ExcelFile(graded_file, engine="openpyxl")
                    sheet_priority = ["Box Raw", "Props", "Graded"]
                    sheet_to_use = next(
                        (s for s in sheet_priority if s in xl.sheet_names),
                        xl.sheet_names[0],
                    )
                    gdf = xl.parse(sheet_to_use)
                    gdf.columns = [str(c).lower().strip() for c in gdf.columns]
                    pcol = next((c for c in gdf.columns if c == "player"), None)
                    propcol = next(
                        (c for c in gdf.columns if c in ("prop", "prop_type", "stat")),
                        None,
                    )
                    dircol = next(
                        (c for c in gdf.columns if c in ("direction", "dir", "side")),
                        None,
                    )
                    resultcol = next(
                        (c for c in gdf.columns if c in ("result", "grade", "outcome", "grade_raw")),
                        None,
                    )
                    actualcol = next(
                        (
                            c
                            for c in gdf.columns
                            if c in ("actual", "actual_value", "result_value")
                        ),
                        None,
                    )
                    if pcol is None or propcol is None:
                        continue
                    trip, pairs = out.get(bucket, ({}, {}))
                    for _, grow in gdf.iterrows():
                        raw: dict[str, Any] = {
                            "player": grow[pcol],
                            "prop_type": grow[propcol],
                        }
                        if dircol and pd.notna(grow.get(dircol)):
                            raw["direction"] = grow[dircol]
                        if "line" in gdf.columns and pd.notna(grow.get("line")):
                            raw["line"] = grow["line"]
                        if actualcol and pd.notna(grow.get(actualcol)):
                            raw["actual"] = grow[actualcol]
                        if resultcol and pd.notna(grow.get(resultcol)):
                            raw["grade"] = grow[resultcol]
                        pl = _canon_player(raw).lower()
                        pt = _canon_prop(raw).lower()
                        dr = _canon_direction(raw)
                        if not pl or not pt:
                            continue
                        key3 = (pl, pt, dr)
                        if key3 in trip:
                            continue
                        row_out = {
                            "player_lower": pl,
                            "prop_lower": pt,
                            "direction": dr,
                            "line": _canon_line(raw),
                            "actual": _canon_actual(raw),
                            "grade_raw": _canon_grade_raw(raw),
                        }
                        trip[key3] = row_out
                        pairs.setdefault((pl, pt), []).append(row_out)
                    out[bucket] = (trip, pairs)
                except Exception:
                    continue
    return out


def _match_leg_in_index(
    leg: dict[str, Any],
    triple: dict[tuple[str, str, str], dict],
    pair_buckets: dict[tuple[str, str], list[dict]],
) -> dict | None:
    pl = str(leg.get("player") or "").strip().lower()
    pt = str(leg.get("prop_type") or "").strip().lower()
    dr = str(leg.get("direction") or "").strip().upper()
    if not pl or not pt:
        return None
    hit = triple.get((pl, pt, dr))
    if hit:
        return hit
    cands = pair_buckets.get((pl, pt))
    if not cands:
        return None
    for r in cands:
        if r["direction"] == dr:
            return r
    if len(cands) == 1:
        return cands[0]
    for r in cands:
        if r["direction"] == dr or not r["direction"]:
            return r
    return cands[0]


def _match_leg_to_row_multi(
    leg: dict[str, Any],
    indices: dict[str, tuple[dict, dict]],
) -> dict | None:
    for bkt in _leg_match_buckets(str(leg.get("sport") or "")):
        trip, pairs = indices.get(bkt, ({}, {}))
        row = _match_leg_in_index(leg, trip, pairs)
        if row:
            return row
    return None


def _graded_outputs_dir(arg_date: str) -> Path:
    return REPO_ROOT / "outputs" / arg_date


def _debug_list_outputs_graded(arg_date: str) -> list[Path]:
    d = _graded_outputs_dir(arg_date)
    if not d.is_dir():
        return []
    return sorted(d.glob("graded_*.xlsx"))


def _debug_sheet_headers(path: Path, max_sheets: int = 3) -> list[tuple[str, list[str]]]:
    """Per sheet: (sheet_name, normalized column names)."""
    out: list[tuple[str, list[str]]] = []
    try:
        xl = pd.ExcelFile(path)
    except Exception as e:
        return [("<<read_error>>", [str(e)])]
    for i, sh in enumerate(xl.sheet_names):
        if i >= max_sheets:
            out.append(("...", [f"(+{len(xl.sheet_names) - max_sheets} more sheets)"]))
            break
        try:
            df = pd.read_excel(path, sheet_name=sh, nrows=0)
            cols = [_norm_header(c) for c in df.columns]
        except Exception as e:
            cols = [f"<<{e}>>"]
        out.append((sh, cols))
    return out


def debug_report(
    arg_date: str,
    payload: dict[str, Any],
    tpath: Path,
    sport_candidates: dict[str, list[Path]],
) -> None:
    """Print why legs may not match (JSON date vs CLI, xlsx paths, headers, sample legs)."""
    print("\n=== build_ticket_eval.py --debug ===\n")
    print(f"CLI --date:     {arg_date}")
    print(f"Ticket source:  {tpath}")
    print(f"Payload \"date\": {payload.get('date')!r}")
    if str(payload.get("date") or "").strip() != arg_date:
        print(
            "  ! Mismatch: ticket payload date differs from --date; legs are still matched against"
            " STATIC pipeline workbooks (see below), not per-date outputs unless we add that."
        )
    out_dir = _graded_outputs_dir(arg_date)
    og = _debug_list_outputs_graded(arg_date)
    print(f"\noutputs/{arg_date}/ graded_*.xlsx:")
    if not out_dir.is_dir():
        print(f"  (folder missing: {out_dir})")
    elif not og:
        print("  (none found)")
    else:
        for p in og:
            print(f"  - {p.relative_to(REPO_ROOT)}")
    print("\nWorkbooks used for matching (first existing path per sport; NOT date-specific today):")
    for sport, paths in sport_candidates.items():
        src = next((p for p in paths if p.is_file()), None)
        if not src:
            print(f"  {sport}: (no file at any candidate path)")
            for p in paths:
                print(f"       tried: {p.relative_to(REPO_ROOT)}")
            continue
        print(f"  {sport}: {src.relative_to(REPO_ROOT)}")
        for sh, cols in _debug_sheet_headers(src):
            preview = cols[:24]
            extra = f" ...(+{len(cols) - 24})" if len(cols) > 24 else ""
            print(f"       sheet {sh!r}: {preview}{extra}")

    indices = _load_actuals_indices(sport_candidates, arg_date)
    gpaths = _graded_xlsx_in_outputs_date(arg_date)
    print(f"\noutputs/{arg_date}/ graded workbook(s) merged into indices:")
    if not gpaths:
        print("  (none — add graded_nba_{date}.xlsx, graded_nhl_{date}.xlsx, cbb_graded_{date}.xlsx, etc.)")
    else:
        for p in gpaths:
            bk = ", ".join(_sport_buckets_for_graded_filename(p)) or "?"
            print(f"  - {p.relative_to(REPO_ROOT)}  → buckets [{bk}]")
    total_triples = sum(len(t) for t, _ in indices.values())
    total_pairs = sum(len(p) for _, p in indices.values())
    print(f"\nIndex (all buckets): {total_triples:,} triple-keys, {total_pairs:,} player+prop buckets (sum per sport)")
    for bkt, (tr, pr) in indices.items():
        if tr or pr:
            print(f"  {bkt}: {len(tr):,} triples, {len(pr):,} pair-buckets")

    groups = payload.get("groups") or []
    legs_sample: list[dict[str, Any]] = []
    for g in groups:
        for t in g.get("tickets") or []:
            for leg in t.get("legs") or []:
                legs_sample.append(leg)
                if len(legs_sample) >= 8:
                    break
            if len(legs_sample) >= 8:
                break
        if len(legs_sample) >= 8:
            break

    print("\nSample legs (match against index above):")
    for i, leg in enumerate(legs_sample, 1):
        pl = str(leg.get("player") or "").strip().lower()
        pt = str(leg.get("prop_type") or "").strip().lower()
        dr = str(leg.get("direction") or "").strip().upper()
        row = _match_leg_to_row_multi(leg, indices)
        st = "MATCH" if row else "NO MATCH -> UNGRADED"
        sp = str(leg.get("sport") or "")
        bk = " → ".join(_leg_match_buckets(sp))
        print(f"  {i}. sport={sp!r} buckets=[{bk}] player={pl!r} prop_type={pt!r} direction={dr!r} -> {st}")
        if row:
            print(
                f"      actual={row.get('actual')!r} line={row.get('line')!r} "
                f"grade_raw={row.get('grade_raw')!r} dir_in_row={row.get('direction')!r}"
            )
    total = sum(len(t.get("legs") or []) for g in groups for t in g.get("tickets") or [])
    print(f"\nTotal legs in JSON: {total}")
    print(
        "\nNote: Base rows come from SPORT_XLSX_CANDIDATES (pre-game step8 slates)."
        f"\n      Same-day graded files under outputs/{arg_date}/ are merged on top when present."
    )
    print("=== end debug ===\n")


def find_ticket_json(arg_date: str) -> Path | None:
    """Resolve ticket file: dated JSON → dated xlsx at repo root → outputs/<date>/ → fallback."""
    p1 = REPO_ROOT / DATED_TICKET_JSON.format(date=arg_date)
    if p1.is_file():
        return p1
    px = REPO_ROOT / f"combined_slate_tickets_{arg_date}.xlsx"
    if px.is_file():
        return px
    # Daily pipeline writes combined tickets under outputs/YYYY-MM-DD/ (not always copied to root).
    out_dir = REPO_ROOT / "outputs" / arg_date
    p_out = out_dir / f"combined_slate_tickets_{arg_date}.xlsx"
    if p_out.is_file():
        return p_out
    p_out_strict = out_dir / f"combined_slate_tickets_{arg_date}.strict.xlsx"
    if p_out_strict.is_file():
        return p_out_strict
    if FALLBACK_TICKET_JSON.is_file():
        print(
            f"[WARN] No dated ticket file found for {arg_date} — falling back to "
            "tickets_latest.json (legs will not match this date's actual slate)",
            flush=True,
        )
        return FALLBACK_TICKET_JSON
    return None


def _player_initials(name: str) -> str:
    parts = str(name or "").strip().split()
    if not parts:
        return ""
    if len(parts) == 1:
        return parts[0][:1].upper()
    return (parts[0][:1] + parts[-1][:1]).upper()


def _clean_team_abbr(s: str) -> str:
    s = str(s or "").strip()
    if not s:
        return ""
    return re.sub(r"\s*\([^)]*\)\s*$", "", s).strip()


def _parse_ticket_banner(s: str) -> tuple[float, float, int]:
    m_no = re.search(r"Ticket\s*#?\s*(\d+)", s, re.I)
    ticket_no = int(m_no.group(1)) if m_no else 1
    m_pow = re.search(r"Power:\s*([\d.]+)\s*x", s, re.I)
    m_flex = re.search(r"Flex:\s*([\d.]+)\s*x", s, re.I)
    power = float(m_pow.group(1)) if m_pow else 0.0
    flex = float(m_flex.group(1)) if m_flex else 0.0
    return power, flex, ticket_no


def _ticket_header_colmap(row: tuple[Any, ...]) -> dict[int, str]:
    out: dict[int, str] = {}
    for i, cell in enumerate(row):
        key = _norm_header(cell)
        field = _XLSX_HDR_TO_LEG_FIELD.get(key)
        if field:
            out[i] = field
    return out


def _row_has_values(row: tuple[Any, ...]) -> bool:
    return any(str(c or "").strip() for c in row)


def _coerce_hit_rate_cell(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        f = float(v)
    else:
        s = str(v).strip().rstrip("%")
        try:
            f = float(s)
        except (TypeError, ValueError):
            return None
    if f > 1.0:
        f = f / 100.0
    return f


def _coerce_line_cell(v: Any) -> float | None:
    if v is None or (isinstance(v, str) and not v.strip()):
        return None
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return None


def _coerce_edge_cell(v: Any) -> float:
    if v is None or v == "":
        return 0.0
    if isinstance(v, (int, float)) and not (isinstance(v, float) and pd.isna(v)):
        return float(v)
    try:
        return float(str(v).strip())
    except (TypeError, ValueError):
        return 0.0


def _leg_from_xlsx_row(row: tuple[Any, ...], colmap: dict[int, str]) -> dict[str, Any] | None:
    leg: dict[str, Any] = {}
    for ci, field in colmap.items():
        if ci >= len(row):
            continue
        val = row[ci]
        if field == "player":
            leg["player"] = str(val or "").strip()
        elif field == "team":
            leg["team"] = _clean_team_abbr(str(val or ""))
        elif field == "opp":
            leg["opp"] = _clean_team_abbr(str(val or ""))
        elif field == "prop_type":
            leg["prop_type"] = str(val or "").strip()
        elif field == "pick_type":
            leg["pick_type"] = str(val or "").strip()
        elif field == "line":
            leg["line"] = _coerce_line_cell(val)
        elif field == "direction":
            leg["direction"] = str(val or "").strip().upper()
        elif field == "edge":
            leg["edge"] = _coerce_edge_cell(val)
        elif field == "hit_rate":
            leg["hit_rate"] = _coerce_hit_rate_cell(val)
        elif field == "l5_avg":
            x = _coerce_line_cell(val)
            leg["l5_avg"] = x
        elif field == "season_avg":
            x = _coerce_line_cell(val)
            leg["season_avg"] = x
        elif field == "sport":
            leg["sport"] = str(val or "").strip().upper()
    if not leg.get("player"):
        return None
    leg["initials"] = _player_initials(str(leg.get("player") or ""))
    return leg


def _skip_xlsx_ticket_sheet(sheet_name: str) -> bool:
    n = sheet_name.strip().lower()
    if n == "summary":
        return True
    if "slate" in n:
        return True
    return False


def _parse_ticket_sheet(ws: Any) -> list[dict[str, Any]]:
    current: dict[str, Any] | None = None
    colmap: dict[int, str] = {}
    expect_header = False
    tickets: list[dict[str, Any]] = []

    for row in ws.iter_rows(values_only=True):
        if row is None:
            continue
        r0 = row[0] if row else None
        s0 = str(r0 or "").strip()
        is_banner = (
            s0
            and "ticket #" in s0.lower()
            and ("power:" in s0.lower() or "flex:" in s0.lower())
        )
        if is_banner:
            if current is not None and current.get("legs"):
                tickets.append(current)
            pow_v, flex_v, tno = _parse_ticket_banner(s0)
            current = {
                "ticket_no": tno,
                "power_payout": pow_v,
                "flex_payout": flex_v,
                "legs": [],
            }
            expect_header = True
            continue

        if expect_header:
            colmap = _ticket_header_colmap(row)
            expect_header = False
            continue

        if current is None or not colmap:
            continue

        if not _row_has_values(row):
            continue

        leg = _leg_from_xlsx_row(row, colmap)
        if leg:
            current["legs"].append(leg)

    if current is not None and current.get("legs"):
        tickets.append(current)

    return tickets


def _load_tickets_from_xlsx(path: Path, arg_date: str) -> dict[str, Any]:
    try:
        from openpyxl import load_workbook
    except ImportError as e:
        raise ImportError(
            "openpyxl is required to read combined_slate_tickets_*.xlsx; "
            "install with: pip install openpyxl"
        ) from e

    groups: list[dict[str, Any]] = []
    wb = load_workbook(path, read_only=True, data_only=True)
    try:
        for sheet_name in wb.sheetnames:
            if _skip_xlsx_ticket_sheet(sheet_name):
                continue
            ws = wb[sheet_name]
            tix = _parse_ticket_sheet(ws)
            if tix:
                groups.append({"group_name": sheet_name, "tickets": tix})
    finally:
        wb.close()

    return {"date": arg_date, "groups": groups}


def _load_tickets(path: Path, arg_date: str) -> dict[str, Any]:
    if path.suffix.lower() == ".xlsx":
        return _load_tickets_from_xlsx(path, arg_date)
    with path.open(encoding="utf-8") as f:
        return json.load(f)


def _fmt_num(x: Any) -> str:
    if x is None:
        return "—"
    if isinstance(x, (int, float)):
        if isinstance(x, float) and x == int(x):
            return str(int(x))
        return f"{x:g}"
    return html.escape(str(x))


def _build_html(
    payload: dict[str, Any],
    arg_date: str,
    sport_candidates: dict[str, list[Path]],
) -> str:
    groups = payload.get("groups") or []
    indices = _load_actuals_indices(sport_candidates, arg_date)

    all_legs: list[tuple[dict, dict | None, str]] = []
    tickets_flat: list[dict] = []

    for g in groups:
        gname = str(g.get("group_name") or "Group")
        for t in g.get("tickets") or []:
            t["_group_name"] = gname
            tickets_flat.append(t)
            for leg in t.get("legs") or []:
                row = _match_leg_to_row_multi(leg, indices)
                line = leg.get("line")
                try:
                    line_f = float(line) if line is not None else None
                except (TypeError, ValueError):
                    line_f = None
                direction = str(leg.get("direction") or "").strip().upper()
                actual = row["actual"] if row else None
                graw = row["grade_raw"] if row else ""
                if row and row.get("line") is not None and line_f is None:
                    line_f = row["line"]
                grade = _leg_grade(actual, line_f, direction, graw)
                all_legs.append((leg, row, grade))

    total_legs = len(all_legs)
    hits = sum(1 for _, _, g in all_legs if g == "HIT")
    misses = sum(1 for _, _, g in all_legs if g == "MISS")
    voids = sum(1 for _, _, g in all_legs if g == "VOID")
    ungraded = sum(1 for _, _, g in all_legs if g == "UNGRADED")
    decided = hits + misses
    leg_pct = (100.0 * hits / decided) if decided else 0.0

    perfect = 0
    with_misses = 0
    for t in tickets_flat:
        legs = t.get("legs") or []
        gs = []
        for leg in legs:
            row = _match_leg_to_row_multi(leg, indices)
            try:
                lf = float(leg.get("line"))
            except (TypeError, ValueError):
                lf = None
            d = str(leg.get("direction") or "").strip().upper()
            act = row["actual"] if row else None
            gr = row["grade_raw"] if row else ""
            if row and row.get("line") is not None and lf is None:
                lf = row["line"]
            g = _leg_grade(act, lf, d, gr)
            gs.append(g)
        if not gs:
            continue
        if all(x == "HIT" for x in gs):
            perfect += 1
        if any(x == "MISS" for x in gs):
            with_misses += 1

    # ── HTML
    esc = html.escape
    json_date = esc(str(payload.get("date") or arg_date))

    sport_colors_css = """
.sport-nba{background:rgba(212,160,23,.12);color:#f0a500;border:1px solid rgba(212,160,23,.35);}
.sport-nba1h{background:rgba(255,155,86,.12);color:#ffb27d;border:1px solid rgba(255,155,86,.32);}
.sport-nba1q{background:rgba(255,214,102,.12);color:#ffd87a;border:1px solid rgba(255,214,102,.32);}
.sport-cbb{background:rgba(0,229,255,.10);color:#00e5ff;border:1px solid rgba(0,229,255,.32);}
.sport-wcbb{background:rgba(127,199,217,.10);color:#9fd8e8;border:1px solid rgba(127,199,217,.32);}
.sport-nhl{background:rgba(186,130,255,.12);color:#c4a5ff;border:1px solid rgba(186,130,255,.38);}
.sport-soccer{background:rgba(240,165,0,.10);color:#e8b84a;border:1px solid rgba(240,165,0,.34);}
.sport-mlb{background:rgba(255,121,121,.12);color:#ff9a9a;border:1px solid rgba(255,121,121,.32);}
.sport-default{background:rgba(255,255,255,.04);color:#888;border:1px solid rgba(255,255,255,.1);}
"""

    parts: list[str] = [
        "<!DOCTYPE html>",
        '<html lang="en" data-theme="dark">',
        "<head>",
        '<meta charset="UTF-8"/>',
        '<meta name="viewport" content="width=device-width, initial-scale=1.0"/>',
        f"<title>Ticket Eval — {json_date}</title>",
        '<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&family=Inter:wght@600;700;800&display=swap" rel="stylesheet"/>',
        "<style>",
        _TICKETS_NAV_THEME_CSS,
        "h1,h2,h3,h4,h5,h6{font-family:'Bebas Neue',sans-serif;letter-spacing:3px;}",
        ".bebas{font-family:'Bebas Neue',sans-serif;letter-spacing:3px;}",
        ".stats-bar{position:sticky;top:0;z-index:50;margin:0 auto 18px;width:100%;max-width:min(1520px,96vw);padding:18px clamp(16px,2.5vw,32px);"
        "background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);"
        "border:1px solid var(--glass-bd);border-radius:18px;box-shadow:0 8px 32px rgba(0,0,0,.35);}",
        ".sum-row{display:flex;flex-wrap:wrap;gap:18px 36px;align-items:center;justify-content:center;}",
        ".sum-item{display:flex;flex-direction:column;align-items:center;gap:4px;min-width:88px;}",
        ".sum-val{font-family:'Share Tech Mono',monospace;font-size:clamp(22px,2.6vw,30px);font-weight:700;color:var(--gold);text-shadow:0 0 20px rgba(240,165,0,.25);}",
        ".sum-val.green{color:var(--green);text-shadow:0 0 14px rgba(57,255,110,.35);}",
        ".sum-val.red{color:var(--red);text-shadow:0 0 14px rgba(255,77,77,.35);}",
        ".sum-val.pend{color:var(--pending);text-shadow:none;}",
        ".sum-val.void{color:var(--gold2);text-shadow:none;}",
        ".sum-val-sm{font-size:clamp(18px,2.1vw,24px)!important;}",
        ".sum-lab{font-family:'Bebas Neue',sans-serif;font-size:11px;letter-spacing:2.2px;color:var(--muted);text-align:center;line-height:1.2;max-width:11em;}",
        ".wrap{width:100%;max-width:min(1520px,96vw);margin:0 auto;padding:10px clamp(14px,2.5vw,32px) 0;}",
        ".sec{margin-top:36px;}",
        ".sec-head{font-family:'Bebas Neue',sans-serif;font-size:clamp(30px,3.2vw,40px);color:var(--gold);margin-bottom:8px;padding-bottom:14px;"
        "border-bottom:1px solid var(--glass-bd);letter-spacing:3px;text-shadow:0 0 24px rgba(240,165,0,.2);}",
        ".ticket-card{background:var(--glass);backdrop-filter:blur(20px) saturate(180%);-webkit-backdrop-filter:blur(20px) saturate(180%);"
        "border:1px solid var(--glass-bd);border-radius:14px;margin-bottom:22px;overflow:hidden;"
        "box-shadow:0 8px 32px rgba(0,0,0,.35);}",
        ".ticket-card.all-hit{background:rgba(57,255,110,0.06);border-color:rgba(57,255,110,.42);"
        "box-shadow:0 0 28px rgba(57,255,110,.14),0 8px 32px rgba(0,0,0,.3);}",
        ".ticket-card.card-missed{background:rgba(255,77,77,0.06);border:1px solid rgba(255,77,77,0.35);"
        "box-shadow:0 0 24px rgba(255,77,77,0.12),0 0 1px rgba(255,77,77,0.4),0 8px 32px rgba(0,0,0,.28);position:relative;}",
        ".ticket-card.card-missed::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;background:#ff4d4d;"
        "box-shadow:0 0 14px rgba(255,77,77,0.45);z-index:2;border-radius:14px 14px 0 0;pointer-events:none;}",
        ".thdr{display:flex;flex-wrap:wrap;gap:12px 20px;align-items:center;padding:18px clamp(14px,2vw,24px);border-bottom:1px solid var(--glass-bd);"
        "background:rgba(0,0,0,.18);backdrop-filter:blur(12px);}",
        ".thdr .tn{font-size:clamp(24px,2.8vw,32px);font-family:'Bebas Neue',sans-serif;letter-spacing:2px;color:var(--gold);}",
        ".thdr .tg{font-family:'Share Tech Mono',monospace;font-size:clamp(12px,1.35vw,15px);color:var(--muted);letter-spacing:0.5px;line-height:1.35;}",
        ".payout{font-family:'Share Tech Mono',monospace;font-size:clamp(13px,1.4vw,16px);color:var(--cyan);}",
        ".banner{font-family:'Bebas Neue',sans-serif;font-size:clamp(11px,1.2vw,13px);letter-spacing:2px;padding:8px 18px;border-radius:999px;font-weight:700;"
        "background:rgba(255,255,255,0.04);backdrop-filter:blur(20px);border:1px solid var(--glass-bd);}",
        ".banner.hit{color:var(--green);border-color:rgba(57,255,110,.45);box-shadow:0 0 16px rgba(57,255,110,.15);}",
        ".banner.miss{color:var(--red);border-color:rgba(255,77,77,.5);box-shadow:0 0 16px rgba(255,77,77,.12);}",
        ".banner.pend{color:var(--pending);border-color:rgba(255,255,255,.12);}",
        ".banner.void{color:var(--gold2);border-color:rgba(240,165,0,.35);}",
        "@keyframes missRowPulse{0%,100%{box-shadow:0 0 0 1px rgba(255,77,77,0.4),inset 0 0 20px rgba(255,77,77,0.06);}"
        "50%{box-shadow:0 0 0 2px rgba(255,77,77,0.65),0 0 22px rgba(255,77,77,0.18),inset 0 0 26px rgba(255,77,77,0.09);}}",
        ".legrow{font-family:'Share Tech Mono',monospace;display:grid;"
        "grid-template-columns:56px 92px minmax(120px,1fr) 44px minmax(240px,1.45fr) minmax(108px,1fr) minmax(96px,1fr) minmax(76px,.85fr);gap:12px;"
        "align-items:center;padding:14px clamp(14px,2vw,22px);font-size:clamp(13px,1.45vw,16px);line-height:1.35;"
        "border-bottom:1px solid rgba(255,255,255,.06);border-left:3px solid transparent;}",
        ".legrow:last-child{border-bottom:none;}",
        ".legrow.leg-hit{background:rgba(57,255,110,0.04);border-left-color:var(--green);}",
        ".legrow.leg-miss{background:rgba(255,77,77,0.10);border-left:4px solid #ff4d4d;"
        "box-shadow:0 0 0 1px rgba(255,77,77,0.4),inset 0 0 20px rgba(255,77,77,0.06);"
        "animation:missRowPulse 2.2s ease-in-out infinite;}",
        ".legrow.leg-miss .pl-miss{color:#ff4d4d;font-weight:700;"
        "text-shadow:0 0 16px rgba(255,77,77,0.8),0 0 32px rgba(255,77,77,0.4);}",
        ".legrow.leg-miss .pl-line{display:flex;align-items:center;flex-wrap:wrap;gap:8px;}",
        ".miss-tag{font-family:'Bebas Neue',sans-serif;display:inline-flex;align-items:center;"
        "background:rgba(255,77,77,0.15);border:1px solid #ff4d4d;color:#ff4d4d;font-size:9px;"
        "letter-spacing:2px;padding:2px 8px;border-radius:20px;line-height:1;vertical-align:middle;}",
        ".legrow.leg-miss .badge.miss{width:44px;height:44px;min-width:44px;border-radius:12px;display:flex;align-items:center;"
        "justify-content:center;font-size:clamp(22px,2.5vw,28px);line-height:1;background:rgba(255,77,77,0.25);"
        "border:2px solid #ff4d4d;box-shadow:0 0 12px rgba(255,77,77,0.6);color:#ff4d4d;text-shadow:none;}",
        ".legrow.leg-miss .leg-extra.val-miss{color:#ff4d4d;font-weight:700;}",
        ".legrow.leg-miss .miss-leg-cell{color:#ff5c5c!important;font-weight:700;}",
        ".legrow.leg-miss .leg-prop-col.miss-leg-cell > div:first-child{color:#ff7a7a!important;font-weight:800;}",
        ".legrow.leg-miss .leg-prop-col .meta-muted{color:rgba(255,170,170,.95)!important;font-weight:600;}",
        ".legrow.leg-miss .miss-leg-cell .dir-over,.legrow.leg-miss .miss-leg-cell .dir-under{color:#ffc9c9!important;font-weight:800;}",
        ".legrow.leg-miss > div:nth-child(2) .pill{box-shadow:0 0 0 1px rgba(255,90,90,.55),0 0 12px rgba(255,60,60,.2);}",
        ".legrow.leg-pend{background:transparent;border-left-color:transparent;}",
        ".legrow.leg-pend .pl-pend,.legrow.leg-pend .meta-muted{color:var(--pending)!important;}",
        ".legrow.leg-pend .pill{background:rgba(255,255,255,0.04)!important;border-color:rgba(255,255,255,0.1)!important;color:var(--pending)!important;}",
        ".legrow.leg-void{background:rgba(240,165,0,0.04);border-left-color:rgba(240,165,0,.55);}",
        ".legrow.leg-void .pl-void,.legrow.leg-void .meta-muted{color:var(--gold2)!important;}",
        ".legrow.leg-void .pill{background:rgba(240,165,0,0.07)!important;border-color:rgba(240,165,0,0.28)!important;color:var(--gold2)!important;}",
        ".badge{font-size:clamp(28px,3.2vw,36px);line-height:1;text-align:center;}",
        ".badge.hit{color:var(--green);text-shadow:0 0 14px rgba(57,255,110,.6);}",
        ".badge.miss{color:var(--red);text-shadow:0 0 14px rgba(255,77,77,.55);}",
        ".badge.pend{color:var(--pending);text-shadow:none;}",
        ".badge.void{color:var(--gold2);text-shadow:none;}",
        ".pill{font-family:'Bebas Neue',sans-serif;font-size:clamp(10px,1.1vw,12px);letter-spacing:1.2px;padding:5px 12px;border-radius:999px;text-transform:uppercase;}",
        sport_colors_css,
        ".tier{font-family:'Bebas Neue',sans-serif;width:32px;height:32px;border-radius:10px;display:flex;align-items:center;justify-content:center;"
        "font-weight:800;font-size:clamp(13px,1.4vw,15px);letter-spacing:0;background:rgba(255,255,255,0.05);color:var(--gold);"
        "border:1px solid var(--glass-bd);backdrop-filter:blur(12px);box-shadow:inset 0 1px 0 rgba(255,255,255,.06);}",
        ".pl-hit{color:var(--green);text-shadow:0 0 8px rgba(57,255,110,.4);}",
        ".pl-miss{color:var(--red);}",
        ".pl-pend{color:var(--pending);}",
        ".pl-void{color:var(--gold2);}",
        ".dir-over{color:var(--cyan);font-weight:700;}",
        ".dir-under{color:var(--gold);font-weight:700;}",
        ".meta-muted{font-family:'Share Tech Mono',monospace;color:var(--muted);font-size:clamp(11px,1.2vw,13px);margin-top:3px;}",
        ".slate-kicker{font-family:'Share Tech Mono',monospace;font-size:clamp(11px,1.2vw,13px);letter-spacing:3px;color:var(--muted);margin-bottom:10px;}",
        ".pl-hit,.pl-pend{font-size:1em;font-weight:600;}",
        "</style>",
        "</head>",
        "<body>",
        _TICKETS_NAV_HTML,
        '<div class="stats-bar">',
        '<div class="sum-row">',
        f'<div class="sum-item"><div class="sum-val">{leg_pct:.1f}%</div><div class="sum-lab">LEG HIT RATE</div></div>',
        f'<div class="sum-item"><div class="sum-val green">{hits}</div><div class="sum-lab">HITS</div></div>',
        f'<div class="sum-item"><div class="sum-val red">{misses}</div><div class="sum-lab">MISSES</div></div>',
        f'<div class="sum-item"><div class="sum-val void">{voids}</div><div class="sum-lab">VOID/PUSH</div></div>',
        f'<div class="sum-item"><div class="sum-val pend">{ungraded}</div><div class="sum-lab">UNGRADED</div></div>',
        f'<div class="sum-item"><div class="sum-val">{perfect}</div><div class="sum-lab">PERFECT TICKETS</div></div>',
        f'<div class="sum-item"><div class="sum-val">{with_misses}</div><div class="sum-lab">TIX W/ MISS</div></div>',
        f'<div class="sum-item"><div class="sum-val sum-val-sm">{total_legs}</div><div class="sum-lab">TOTAL LEGS</div></div>',
        "</div></div>",
        '<div class="wrap">',
        f'<p class="slate-kicker">SLATE DATE · {json_date}</p>',
        '<p class="meta-muted" style="margin:6px 0 14px;line-height:1.5">'
        "Each leg: <strong>Line</strong> + side · <strong>Actual</strong> (box-score stat; — until a graded file exists) · "
        f"<strong>Edge</strong> (model edge, not the result). Graded exports: <code>outputs/{json_date}/graded_*.xlsx</code>."
        "</p>",
    ]

    bucketed = _bucket_ticket_groups(groups)
    use_buckets = len(bucketed) > 1

    for bi, (bkey, grplist) in enumerate(bucketed):
        ntix = sum(len(x.get("tickets") or []) for x in grplist)
        nsec = len(grplist)
        if use_buckets:
            skin = _ticket_bucket_skin_class(bkey)
            open_attr = " open" if bi == 0 else ""
            parts.append(f'<details class="ticket-bucket {skin}"{open_attr}>')
            parts.append(
                f'<summary><span>{esc(bkey)}</span>'
                f'<span class="ticket-bucket-meta">{nsec} sections · {ntix} tickets</span></summary>'
            )
            parts.append('<div class="ticket-bucket-body">')

        for g in grplist:
            gname = str(g.get("group_name") or "Group")
            parts.append(f'<section class="sec"><h2 class="sec-head bebas">{esc(gname)}</h2>')
            for t in g.get("tickets") or []:
                tno = t.get("ticket_no", "?")
                pp = t.get("power_payout")
                fp = t.get("flex_payout")
                legs = t.get("legs") or []
                leg_grades: list[str] = []
                for leg in legs:
                    row = _match_leg_to_row_multi(leg, indices)
                    try:
                        lf = float(leg.get("line"))
                    except (TypeError, ValueError):
                        lf = None
                    d = str(leg.get("direction") or "").strip().upper()
                    act = row["actual"] if row else None
                    gr = row["grade_raw"] if row else ""
                    if row and row.get("line") is not None and lf is None:
                        lf = row["line"]
                    leg_grades.append(_leg_grade(act, lf, d, gr))

                h = leg_grades.count("HIT")
                m = leg_grades.count("MISS")
                pnd = leg_grades.count("UNGRADED")
                vct = leg_grades.count("VOID")
                n = len(leg_grades)

                if pnd > 0:
                    banner_cls, banner_txt = "pend", "UNGRADED"
                elif vct > 0 and m == 0:
                    banner_cls, banner_txt = "void", "VOID/PUSH"
                elif m == 0 and n > 0:
                    banner_cls, banner_txt = "hit", "ALL HIT"
                else:
                    banner_cls, banner_txt = "miss", f"MISSED {m}"

                card_cls = "ticket-card"
                if banner_txt == "ALL HIT":
                    card_cls += " all-hit"
                elif banner_cls == "miss":
                    card_cls += " card-missed"

                parts.append(f'<article class="{card_cls}">')
                parts.append('<div class="thdr">')
                parts.append(f'<span class="tn bebas">#{esc(str(tno))}</span>')
                parts.append(f'<span class="tg">{esc(gname)}</span>')
                parts.append(f'<span class="tg">{h}✓ {m}✗ / {n}</span>')
                parts.append(f'<span class="payout">PWR {_fmt_num(pp)}× · FLEX {_fmt_num(fp)}×</span>')
                parts.append(f'<span class="banner {banner_cls}">{esc(banner_txt)}</span>')
                parts.append("</div>")

                for leg, lg in zip(legs, leg_grades):
                    row = _match_leg_to_row_multi(leg, indices)
                    try:
                        lf = float(leg.get("line"))
                    except (TypeError, ValueError):
                        lf = None
                    d = str(leg.get("direction") or "").strip().upper()
                    act = row["actual"] if row else None
                    gr = row["grade_raw"] if row else ""
                    if row and row.get("line") is not None and lf is None:
                        lf = row["line"]

                    if lg == "HIT":
                        bcls, plcls = "hit", "pl-hit"
                    elif lg == "MISS":
                        bcls, plcls = "miss", "pl-miss"
                    elif lg == "VOID":
                        bcls, plcls = "void", "pl-void"
                    else:
                        bcls, plcls = "pend", "pl-pend"

                    sk = _sport_key(str(leg.get("sport") or ""))
                    sp_class = {
                        "NBA": "sport-nba",
                        "NBA1H": "sport-nba1h",
                        "NBA1Q": "sport-nba1q",
                        "CBB": "sport-cbb",
                        "WCBB": "sport-wcbb",
                        "NHL": "sport-nhl",
                        "SOCCER": "sport-soccer",
                        "MLB": "sport-mlb",
                    }.get(sk, "sport-default")

                    tier = _pick_type_tier(str(leg.get("pick_type") or ""))
                    team = esc(str(leg.get("team") or ""))
                    opp = esc(str(leg.get("opp") or ""))
                    ptype = esc(str(leg.get("prop_type") or ""))
                    player = esc(str(leg.get("player") or ""))
                    edge = leg.get("edge")
                    dir_cls = "dir-over" if d == "OVER" else "dir-under" if d == "UNDER" else ""

                    if lg == "HIT":
                        row_cls = "legrow leg-hit"
                    elif lg == "MISS":
                        row_cls = "legrow leg-miss"
                    elif lg == "VOID":
                        row_cls = "legrow leg-void"
                    else:
                        row_cls = "legrow leg-pend"
                    sym = "✓" if lg == "HIT" else "✗" if lg == "MISS" else "○" if lg == "VOID" else "·"

                    miss_cell = " miss-leg-cell" if lg == "MISS" else ""

                    if lg == "MISS":
                        pl_html = (
                            f'<div class="{plcls} pl-line{miss_cell}">'
                            f'<span class="pl-name">{player}</span>'
                            '<span class="miss-tag" aria-label="Missed leg">MISSED</span></div>'
                        )
                    else:
                        pl_html = f'<div class="{plcls}">{player}</div>'

                    if lg == "MISS":
                        act_div_cls = "leg-extra val-miss"
                    elif lg == "HIT":
                        act_div_cls = "leg-extra pl-hit"
                    elif lg == "VOID":
                        act_div_cls = "leg-extra pl-void"
                    else:
                        act_div_cls = "leg-extra pl-pend"

                    parts.append(f'<div class="{row_cls}">')
                    parts.append(f'<div class="badge {bcls}">{sym}</div>')
                    parts.append(f'<div><span class="pill {sp_class}">{esc(sk)}</span></div>')
                    parts.append(pl_html)
                    parts.append(f'<div class="tier{miss_cell}">{esc(tier)}</div>')
                    parts.append(
                        f'<div class="leg-prop-col{miss_cell}"><div>{ptype}</div>'
                        f'<div class="meta-muted">{team} vs {opp}</div></div>'
                    )
                    parts.append(
                        f'<div class="leg-extra{miss_cell}">{_fmt_num(lf)} <span class="{dir_cls}">{esc(d)}</span></div>'
                    )
                    parts.append(f'<div class="{act_div_cls}{miss_cell}">{_fmt_num(act)}</div>')
                    parts.append(f'<div class="leg-extra{miss_cell}">{_fmt_num(edge)}</div>')
                    parts.append("</div>")

                parts.append("</article>")
            parts.append("</section>")

        if use_buckets:
            parts.append("</div></details>")

    parts.append("</div>")
    parts.append(_TICKETS_THEME_JS)
    parts.append("</body></html>")
    return "\n".join(parts)


def main() -> int:
    ap = argparse.ArgumentParser(description="Build ticket_eval HTML for Grades UI.")
    ap.add_argument(
        "--date",
        default="",
        help="Slate date YYYY-MM-DD (default: yesterday local)",
    )
    ap.add_argument(
        "--debug",
        action="store_true",
        help="Print ticket JSON path, payload date, outputs/graded files, Excel headers, sample leg matches; then build.",
    )
    args = ap.parse_args()
    if args.date:
        arg_date = args.date.strip()
        if not re.match(r"^\d{4}-\d{2}-\d{2}$", arg_date):
            print("ERROR: --date must be YYYY-MM-DD")
            return 1
    else:
        arg_date = (date.today() - timedelta(days=1)).isoformat()

    sport_candidates = _dated_candidates(arg_date)

    tpath = find_ticket_json(arg_date)
    if not tpath:
        print(
            "ERROR: No ticket file found (combined_slate_tickets_{date}.json, "
            "combined_slate_tickets_{date}.xlsx, or ui_runner/templates/tickets_latest.json)."
        )
        return 1

    try:
        payload = _load_tickets(tpath, arg_date)
    except Exception as e:
        print(f"ERROR: Failed to read ticket file: {e}")
        return 1

    if args.debug:
        debug_report(arg_date, payload, tpath, sport_candidates)

    html_out = _build_html(payload, arg_date, sport_candidates)
    TEMPLATES_DIR.mkdir(parents=True, exist_ok=True)
    dated_name = f"ticket_eval_{arg_date}.html"
    out_dated = TEMPLATES_DIR / dated_name
    out_latest = TEMPLATES_DIR / "tickets_latest.html"
    try:
        out_dated.write_text(html_out, encoding="utf-8")
        out_latest.write_text(html_out, encoding="utf-8")
    except OSError as e:
        print(f"ERROR: Write failed: {e}")
        return 1

    print(f"Wrote {out_dated}")
    print(f"Wrote {out_latest} (main tickets page)")
    return 0


if __name__ == "__main__":
    raise SystemExit(main())
