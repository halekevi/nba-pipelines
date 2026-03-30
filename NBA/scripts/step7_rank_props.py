#!/usr/bin/env python3
"""
step7_rank_props.py  (VECTORIZED 2026-03-01)

PERF: All 12 .apply() calls and 2 row-by-row list comprehensions replaced with
      vectorized pandas/NumPy operations. Estimated 3-4x faster on 8,000+ row slates.
      Excel write engine switched from openpyxl → xlsxwriter (~5x faster write).

PATCH (2026-03-28):
- Hard edge gate now uses edge_adj_dr (direction-aware), not raw edge_adj.
  Previously UNDER plays with projection below line had edge_adj < 0 and were
  stripped of rank_score / tier despite strong under signals.

PATCH (2026-02-26):
- Fix edge_adj_dr to be direction-aware: UNDERs now get a positive edge
  contribution when projection < line.
- Support projection building for volume props (2PTA/2PTM, 3PTA/3PTM, FTA/FTM, FGA/FGM).
- Adds prop_norm aliases for volume props.

PATCH (2026-02-23):
- grading-informed reweight of scoring components.
"""

from __future__ import annotations

import argparse
import json
import re
import sys
from pathlib import Path
import numpy as np
import pandas as pd
import joblib

for _efe_anc in Path(__file__).resolve().parents:
    if (_efe_anc / "scripts" / "edge_feature_engineering.py").is_file():
        _efe_sd = str(_efe_anc / "scripts")
        if _efe_sd not in sys.path:
            sys.path.insert(0, _efe_sd)
        break
from edge_feature_engineering import apply_ticket_eligibility_voids, build_feature_vector  # noqa: E402

# UTF-8 safe Excel export
try:
    import xlsxwriter
    HAS_XLSXWRITER = True
except ImportError:
    HAS_XLSXWRITER = False

# -------------------- player consistency (data/cache/player_consistency.db) --------------------

import sys as _sys_pc


def _repo_root_pc() -> Path:
    here = Path(__file__).resolve()
    for anc in here.parents:
        if (anc / "scripts" / "build_player_consistency.py").is_file():
            return anc
    return here.parents[2]


def _load_bpc_pc():
    root = _repo_root_pc()
    sd = str(root / "scripts")
    if sd not in _sys_pc.path:
        _sys_pc.path.insert(0, sd)
    import build_player_consistency as bpc  # noqa: E402

    return bpc


_bpc_pc_mod = None


def _bpc_pc():
    global _bpc_pc_mod
    if _bpc_pc_mod is None:
        try:
            _bpc_pc_mod = _load_bpc_pc()
        except Exception:
            _bpc_pc_mod = False
    return _bpc_pc_mod


def _normalize_prop_type(raw: str) -> str:
    m = _bpc_pc()
    if not m:
        return str(raw or "").strip()
    return m._normalize_prop_type(str(raw), "NBA")


def _get_line_bucket(prop_type: str, line: float, sport: str) -> str:
    m = _bpc_pc()
    if not m:
        return "<5"
    try:
        ln = float(line)
    except (TypeError, ValueError):
        ln = 0.0
    return m.get_line_bucket(prop_type, ln, sport)


def _get_consistency_grade(player: str, sport: str, prop_type: str, direction: str, line: float) -> str:
    """
    Look up player consistency grade from player_consistency.db.
    Returns grade string S/A/B/C/D/F/?
    Returns '?' if DB missing or player not found.
    """
    import sqlite3

    repo_root = _repo_root_pc()
    db_path = repo_root / "data" / "cache" / "player_consistency.db"
    if not db_path.exists():
        return "?"

    prop_type = _normalize_prop_type(prop_type)
    direction = direction.upper().strip()
    bucket = _get_line_bucket(prop_type, line, sport)

    try:
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute(
            """
            SELECT grade, grade_locked, games_since_F
            FROM player_consistency
            WHERE player_name = ?
              AND sport = ?
              AND prop_type = ?
              AND direction = ?
              AND line_bucket = ?
        """,
            (player, sport, prop_type, direction, bucket),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            grade, locked, games_since_F = row
            return grade or "?"
        return "?"
    except Exception:
        return "?"


def _pc_grade_cache(sport: str) -> dict:
    import sqlite3

    dbp = _repo_root_pc() / "data" / "cache" / "player_consistency.db"
    if not dbp.is_file():
        return {}
    try:
        conn = sqlite3.connect(str(dbp))
        cur = conn.execute(
            "SELECT player_name, sport, prop_type, direction, line_bucket, grade FROM player_consistency WHERE sport = ?",
            (sport,),
        )
        d = {(a, b, c, d0, e): (g if g else "?") for a, b, c, d0, e, g in cur.fetchall()}
        conn.close()
        return d
    except Exception:
        return {}


def _apply_consistency_grade_scores(out: pd.DataFrame, sport: str) -> None:
    grade_multiplier = {
        "S": 1.25,
        "A": 1.15,
        "B": 1.05,
        "C": 1.00,
        "D": 0.80,
        "F": 0.00,
        "?": 0.95,
    }
    cache = _pc_grade_cache(sport)
    pc = next((c for c in ("player_norm", "player", "pp_player", "player_name") if c in out.columns), None)
    prop_col = "prop_norm" if "prop_norm" in out.columns else ("prop_type" if "prop_type" in out.columns else None)
    if pc is None or prop_col is None or "bet_direction" not in out.columns or "line" not in out.columns:
        out["consistency_grade"] = "?"
        out["consistency_multiplier"] = 0.95
        out["final_score"] = _to_num(out.get("final_score", pd.Series(np.nan, index=out.index))) * 0.95
        return

    players = out[pc].astype(str).str.strip()
    prop_raw = out[prop_col].astype(str)
    dirs = out["bet_direction"].astype(str).str.strip().str.upper()
    linev = _to_num(out["line"]).fillna(0.0)
    grades: list[str] = []
    for i in range(len(out)):
        ptype = _normalize_prop_type(prop_raw.iloc[i])
        try:
            ln = float(linev.iloc[i])
        except (TypeError, ValueError):
            ln = 0.0
        bkt = _get_line_bucket(ptype, ln, sport)
        g = cache.get((players.iloc[i], sport, ptype, dirs.iloc[i], bkt), "?")
        grades.append(g)
    gser = pd.Series(grades, index=out.index)
    mult = gser.map(lambda x: grade_multiplier.get(x, 0.95)).astype(float)
    out["consistency_grade"] = gser
    out["consistency_multiplier"] = mult
    out["final_score"] = _to_num(out["final_score"]).astype(float) * mult


# -------------------- helpers --------------------

def _to_num(s):
    return pd.to_numeric(s, errors="coerce")

def _norm_pick_type_series(s: pd.Series) -> pd.Series:
    t = s.astype(str).str.strip().str.lower()
    return np.where(t.str.contains("gob"), "Goblin",
           np.where(t.str.contains("dem"), "Demon", "Standard"))

# -------------------- weights --------------------

# Prop weights — calibrated from 9-day graded outcomes (2026-03-06 → 2026-03-14)
# Higher weight = model gives this prop more scoring influence.
# Fantasy pulled down (2026-03): 1.08 + 15% combo correction + high OVER prior stacked and
# over-ranked fantasy vs singles/combos in the slate.
_PROP_WEIGHTS = {
    "fantasy": 0.750,
    "pts": 1.000,
    "pr": 1.000,
    "reb": 1.000,
    "ra": 1.000,
    "pra": 1.000,
    "pa": 1.000,
    "ast": 1.000,
    "fg2a": 1.030,
    "fga": 1.030,
    "pf": 0.970,
    "personalfouls": 0.970,
    "tov": 1.000,
    "fgm": 0.950,
    "fg3a": 0.970,
    "3ptattempted": 0.970,
    "twopointersattempted": 0.980,
    "ftm": 0.940,
    "freethrowsmade": 0.940,
    "stocks": 0.920,
    "stl": 1.000,
    "fg2m": 0.950,
    "twopointersmade": 0.950,
    "fta": 1.000,
    "freethrowsattempted": 0.940,
    "fg3m": 0.900,
    "3ptmade": 0.900,
    "blk": 0.800,
}

# Hit rate priors — calibrated from 9-day graded data (2026-03-06 → 2026-03-14)
# Used in prop_hr_z scoring signal. Old values were based on season-long prior;
# these reflect actual pipeline output hit rates by prop type OVER direction.
_PROP_HR_PRIOR_OVER = {
    "fantasy": 0.560,
    "pts": 0.580,
    "pr": 0.565,
    "reb": 0.580,
    "ra": 0.479,
    "ast": 0.555,
    "fga": 0.510,
    "pra": 0.545,
    "pa": 0.550,
    "fgm": 0.510,
    "fg2m": 0.510,
    "twopointersmade": 0.510,
    "fg2a": 0.520,
    "twopointersattempted": 0.520,
    "tov": 0.423,
    "pf": 0.510,
    "personalfouls": 0.510,
    "fg3a": 0.490,
    "3ptattempted": 0.490,
    "stocks": 0.510,
    "stl": 0.530,
    "fg3m": 0.520,
    "3ptmade": 0.520,
    "ftm": 0.510,
    "freethrowsmade": 0.510,
    "fta": 0.442,
    "freethrowsattempted": 0.460,
    "blk": 0.356,
}

# UNDER overrides — calibrated from 9-day graded data.
# These are props where the UNDER signal is meaningfully different from (1 - OVER prior).
# Key insight: Steals/3PM/FTA/Blks+Stls UNDER are the best Standard signals in the dataset.
_PROP_HR_PRIOR_UNDER_OVERRIDE = {
    "stl": 0.667,
    "fg3m": 0.600,
    "3ptmade": 0.580,
    "stocks": 0.580,
    "fta": 0.559,
    "freethrowsattempted": 0.580,
    "ra": 0.521,
    "ftm": 0.545,
    "freethrowsmade": 0.545,
    "ast": 0.527,
    "tov": 0.578,
    "fga": 0.545,
    "fg2a": 0.545,
    "twopointersattempted": 0.480,
    "reb": 0.529,
    "pa": 0.550,
    "pts": 0.541,
    "pr": 0.541,
    "pra": 0.537,
    "fantasy": 0.330,
    "pf": 0.518,
    "personalfouls": 0.518,
}

_RELIABILITY_MAP = {
    "Standard": 1.00,  # baseline
    "Goblin":   1.08,  # was 1.06 — consistently outperforms, slight raise
    "Demon":    0.50,  # was 0.75 — 31.8% actual hit rate, needs to be near-invisible
}

def _repo_root_ml_nba() -> Path:
    return Path(__file__).resolve().parents[2]


_sd_ml_nba = str(_repo_root_ml_nba() / "scripts")
if _sd_ml_nba not in _sys_pc.path:
    _sys_pc.path.insert(0, _sd_ml_nba)
try:
    from ml_blend_weight import load_ml_blend_weight  # noqa: E402

    ML_BLEND_WEIGHT = float(load_ml_blend_weight(_repo_root_ml_nba(), "nba"))
except Exception:
    ML_BLEND_WEIGHT = 0.30

# -------------------- projection fallback --------------------

_PLAYER_PREFIX_BY_PROP = {
    "fga": "fga", "fgm": "fgm", "fg2a": "fg2a", "fg2m": "fg2m",
    "fg3a": "fg3a", "fg3m": "fg3m", "fta": "fta", "ftm": "ftm",
}

_COMBO_CORRECTIONS = {"pr": 1.05, "pa": 1.06, "ra": 1.08, "pra": 1.07, "fantasy": 0.92}

def _edge_transform_series(edge: pd.Series, cap: float = 3.0, power: float = 0.85) -> pd.Series:
    """Vectorized power-transform with sign preservation."""
    sign = np.sign(edge)
    clipped = np.clip(edge.abs(), 0, cap)
    return sign * (clipped ** power)

def _tier_from_score_series(score: pd.Series) -> pd.Series:
    return np.where(score >= 1.25, "A",
           np.where(score >= 0.75, "B",
           np.where(score >= 0.40, "C", "D")))


def _defense_tier_feature(out: pd.DataFrame) -> pd.Series:
    if "defense_tier" in out.columns:
        s = out["defense_tier"].astype(str).str.strip().str.lower()
        return pd.Series(
            np.where(
                s.str.contains("weak"),
                0,
                np.where(s.str.contains("avg|average|mid|med"), 1, np.where(s.str.contains("strong"), 2, 1)),
            ),
            index=out.index,
        ).astype(float)

    if "def_tier" in out.columns:
        s = out["def_tier"].astype(str).str.strip().str.lower()
        return pd.Series(
            np.where(
                s.str.contains("weak"),
                0,
                np.where(s.str.contains("avg|average|mid|med"), 1, np.where(s.str.contains("strong"), 2, 1)),
            ),
            index=out.index,
        ).astype(float)

    if "OVERALL_DEF_RANK" in out.columns:
        dr = _to_num(out["OVERALL_DEF_RANK"]).fillna(15.0)
        return pd.Series(np.where(dr <= 10, 2, np.where(dr <= 20, 1, 0)), index=out.index).astype(float)

    return pd.Series(1.0, index=out.index)


def _normalize_nba_prop_ml(raw: str) -> str:
    x = str(raw or "").strip().lower()
    x_compact = re.sub(r"\s+", "", x)
    if not x:
        return "unknown"
    if "pts+reb+ast" in x_compact or x_compact == "pra" or "pra" in x.split():
        return "pra"
    if "pts+asts" in x_compact or "ptsasts" in x_compact or "points+assists" in x_compact:
        return "pts_asts"
    if "pts+rebs" in x_compact or "ptsrebs" in x_compact:
        return "pts_rebs"
    if "rebs+asts" in x_compact or "rebsasts" in x_compact:
        return "rebs_asts"
    if "blks+stls" in x_compact or "blksstls" in x_compact:
        return "blks_stls"
    if "3-pt" in x or "3pt" in x_compact or "threes" in x or "fg3m" in x or "three" in x:
        return "threes"
    if "fantasy" in x:
        return "fantasy_score"
    if "free throw" in x or x.startswith("fta") or " ftm" in x:
        return "fta"
    if ("field goal" in x and "attempt" in x) or x == "fga" or "fg attempted" in x:
        return "fg_attempted"
    if "rebound" in x or x in ("reb", "rebs"):
        return "rebounds"
    if "assist" in x or x in ("ast", "asts"):
        return "assists"
    if "point" in x or x in ("pts", "pt"):
        return "points"
    if "steal" in x or x == "stl":
        return "steals"
    if "block" in x or x == "blk":
        return "blocks"
    if "turnover" in x or x == "tov" or x == "to":
        return "turnovers"
    return re.sub(r"[^a-z0-9]+", "_", x).strip("_") or "unknown"


def _nba_ml_defense_tier_4(out: pd.DataFrame) -> pd.Series:
    idx = out.index
    if "defense_tier" in out.columns:
        s = out["defense_tier"].astype(str).str.strip().str.lower()
        return pd.Series(
            np.where(
                s.str.contains("weak"),
                0,
                np.where(
                    s.str.contains("avg|average|mid|med"),
                    1,
                    np.where(
                        s.str.contains("good|solid|above"),
                        2,
                        np.where(s.str.contains("elite|strong"), 3, 1),
                    ),
                ),
            ),
            index=idx,
        ).astype(float)
    if "def_tier" in out.columns:
        s = out["def_tier"].astype(str).str.strip().str.lower()
        return pd.Series(
            np.where(
                s.str.contains("weak"),
                0,
                np.where(
                    s.str.contains("avg|average|mid|med"),
                    1,
                    np.where(
                        s.str.contains("good|solid|above"),
                        2,
                        np.where(s.str.contains("elite|strong"), 3, 1),
                    ),
                ),
            ),
            index=idx,
        ).astype(float)
    if "OVERALL_DEF_RANK" in out.columns:
        r = _to_num(out["OVERALL_DEF_RANK"]).fillna(15.0)
        return pd.Series(
            np.where(r <= 5, 3, np.where(r <= 10, 2, np.where(r <= 20, 1, 0))),
            index=idx,
        ).astype(float)
    return pd.Series(1.0, index=idx)


def _nba_ml_pick_col(out: pd.DataFrame, names: tuple[str, ...]) -> pd.Series:
    idx = out.index
    for n in names:
        if n in out.columns:
            return out[n]
    return pd.Series(np.nan, index=idx)


def _build_nba_ml_X(out: pd.DataFrame, model_features: list[str]) -> pd.DataFrame:
    idx = out.index
    pick_type_s = out.get("pick_type", pd.Series("Standard", index=idx)).astype(str).str.strip().str.lower()
    tier_num = pd.Series(
        np.where(pick_type_s.str.contains("gob"), 2, np.where(pick_type_s.str.contains("dem"), 0, 1)),
        index=idx,
    )
    dir_s = out.get("bet_direction", pd.Series("OVER", index=idx)).astype(str).str.upper().str.strip()
    direction_num = pd.Series(np.where(dir_s.eq("OVER"), 1, 0), index=idx)
    is_under = dir_s.eq("UNDER")

    hr5_raw = _to_num(_nba_ml_pick_col(out, ("line_hit_rate_over_ou_5", "line_hit_rate_over_5")))
    hr10_raw = _to_num(_nba_ml_pick_col(out, ("line_hit_rate", "line_hit_rate_over_ou_10")))
    hr20_raw = _to_num(_nba_ml_pick_col(out, ("line_hit_rate_over_ou_20", "line_hit_rate_over_20")))
    # Invert OVER hit rates for UNDER props so "high = good signal" is preserved
    hr5 = pd.Series(np.where(is_under, 1.0 - hr5_raw, hr5_raw), index=idx)
    hr10 = pd.Series(np.where(is_under, 1.0 - hr10_raw, hr10_raw), index=idx)
    hr20 = pd.Series(np.where(is_under, 1.0 - hr20_raw, hr20_raw), index=idx)

    def _scale_hit_pct(s: pd.Series) -> pd.Series:
        if s.notna().any() and s.dropna().median() > 1.0:
            return s / 100.0
        return s

    hr5, hr10, hr20 = _scale_hit_pct(hr5), _scale_hit_pct(hr10), _scale_hit_pct(hr20)
    hr5 = hr5.fillna(hr10).fillna(0.5)
    hr10 = hr10.fillna(0.5)
    hr20 = hr20.fillna(hr10).fillna(0.5)

    line = _to_num(_nba_ml_pick_col(out, ("line",))).fillna(0.0)
    minutes = _to_num(_nba_ml_pick_col(out, ("avg_minutes", "minutes"))).fillna(0.0)
    ha_raw = _nba_ml_pick_col(out, ("home_away", "home/away"))
    if ha_raw.notna().any():
        has = ha_raw.astype(str).str.strip().str.upper()
        home_away = pd.Series(np.where(has.str.startswith("H"), 1.0, 0.0), index=idx)
    else:
        home_away = pd.Series(0.5, index=idx)

    if "game_script_mult" in out.columns:
        gsm = _to_num(out["game_script_mult"]).fillna(1.0)
    else:
        gsm = pd.Series(1.0, index=idx)

    if "consistency_grade" in out.columns:
        cg_map = {"S": 5.0, "A": 4.0, "B": 3.0, "C": 2.0, "D": 1.0, "F": 0.0, "?": 2.0}
        cg = out["consistency_grade"].astype(str).str.strip().str.upper().map(lambda x: cg_map.get(x, 2.0))
        cg = _to_num(cg).fillna(2.0)
    else:
        cg = pd.Series(2.0, index=idx)

    base = pd.DataFrame(
        {
            "edge": _to_num(out.get("edge", pd.Series(np.nan, index=idx))).fillna(0.0),
            "hit_rate_l5": hr5,
            "hit_rate_l10": hr10,
            "hit_rate_l20": hr20,
            "line": line,
            "direction": _to_num(direction_num).fillna(0.0),
            "tier": _to_num(tier_num).fillna(1.0),
            "defense_tier": _nba_ml_defense_tier_4(out).fillna(1.0),
            "minutes": minutes,
            "home_away": home_away,
            "game_script_mult": gsm,
            "consistency_grade": cg,
        },
        index=idx,
    )
    prop_raw = out.get("prop_norm", out.get("prop_type", pd.Series("unknown", index=idx)))
    prop_norm = prop_raw.astype(str).map(_normalize_nba_prop_ml)
    dummies = pd.get_dummies(prop_norm, prefix="prop", dtype=float)
    X = pd.concat([base, dummies], axis=1)
    return X.reindex(columns=model_features, fill_value=0.0)


def _apply_ml_blend(out: pd.DataFrame, existing_score: pd.Series, source_hint: str = "") -> tuple[pd.Series, pd.Series, pd.Series]:
    root = Path(__file__).resolve().parents[2]
    source_key = str(source_hint).lower()
    model_keys = ["nba"]
    if "nba1h" in source_key:
        model_keys = ["nba1h", "nba"]
    elif "nba1q" in source_key:
        model_keys = ["nba1q", "nba"]

    model_path = None
    feat_path = None
    model_key_used = "nba"
    for mk in model_keys:
        mp = root / "models" / f"prop_model_{mk}.pkl"
        fp = root / "models" / f"prop_model_{mk}_features.json"
        if mp.exists() and fp.exists():
            model_path = mp
            feat_path = fp
            model_key_used = mk
            break

    if model_path is None or feat_path is None:
        print(f"⚠️  ML model not found for keys {model_keys} — skipping ML blend")
        return (
            pd.Series(np.nan, index=out.index),
            pd.Series(np.nan, index=out.index),
            existing_score.copy(),
        )

    try:
        model = joblib.load(model_path)
        model_features = json.loads(feat_path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"⚠️  Failed to load ML model/features: {e} — skipping ML blend")
        return (
            pd.Series(np.nan, index=out.index),
            pd.Series(np.nan, index=out.index),
            existing_score.copy(),
        )

    calibrator = None
    calib_path = root / "models" / f"prop_model_{model_key_used}_calibrator.pkl"
    try:
        if calib_path.exists():
            calibrator = joblib.load(calib_path)
    except Exception:
        calibrator = None

    try:
        from ml_blend_weight import load_ml_blend_weight as _load_nba_blend

        blend_w = float(_load_nba_blend(root, model_key_used))
    except Exception:
        blend_w = float(ML_BLEND_WEIGHT)

    try:
        X = _build_nba_ml_X(out, model_features)
        raw_prob = pd.Series(model.predict_proba(X)[:, 1], index=out.index, dtype=float)
        if calibrator is not None:
            try:
                if hasattr(calibrator, "predict_proba"):
                    cal_vals = calibrator.predict_proba(raw_prob.values.reshape(-1, 1))[:, 1]
                else:
                    cal_vals = calibrator.predict(raw_prob.values)
                ml_prob = pd.Series(cal_vals, index=out.index, dtype=float).clip(0.001, 0.999)
            except Exception:
                ml_prob = raw_prob
        else:
            ml_prob = raw_prob
    except Exception as e:
        print(f"⚠️  ML inference failed: {e} — skipping ML blend")
        return (
            pd.Series(np.nan, index=out.index),
            pd.Series(np.nan, index=out.index),
            existing_score.copy(),
        )

    ml_edge = ml_prob - 0.5
    final_score = (1.0 - blend_w) * existing_score + blend_w * ml_edge
    print(f"✅ NBA ML blend applied (model={model_key_used}, weight={blend_w:.2f})")
    return ml_prob, ml_edge, final_score

def _write_xlsx_openpyxl(output_path: str, out: pd.DataFrame, elig_mask: pd.Series) -> None:
    """Write XLSX with explicit UTF-8 encoding using openpyxl."""
    from openpyxl import Workbook
    from openpyxl.utils.dataframe import dataframe_to_rows
    
    wb = Workbook()
    wb.remove(wb.active)
    
    # Create both sheets with UTF-8 safe values
    for sheet_name, df_sheet in [("ALL", out), ("ELIGIBLE", out.loc[elig_mask])]:
        ws = wb.create_sheet(sheet_name)
        for r_idx, row in enumerate(dataframe_to_rows(df_sheet, index=False, header=True), 1):
            for c_idx, value in enumerate(row, 1):
                # Ensure value is properly UTF-8 encoded (especially for player names)
                if isinstance(value, str):
                    # Force string through UTF-8 encode/decode to ensure proper handling
                    value = value.encode('utf-8').decode('utf-8')
                elif pd.isna(value):
                    value = None
                ws.cell(row=r_idx, column=c_idx, value=value)
    
    # Set encoding in workbook properties
    wb.properties.encoding = 'UTF-8'
    wb.save(output_path)
    print(f"✅ Saved → {output_path} (openpyxl, UTF-8 encoded)")

# -------------------- main --------------------

def main() -> None:
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  default="step6_with_team_role_context.csv")
    ap.add_argument("--output", default="step7_ranked_props.xlsx")
    ap.add_argument(
        "--injuries-csv",
        default="",
        help="injuries_nba_*.csv from fetch_actuals (optional rank_score penalty)",
    )
    ap.add_argument(
        "--slate-date",
        default="",
        help="YYYY-MM-DD; if set with empty injuries-csv, load outputs/<date>/injuries_nba_<date>.csv",
    )
    args = ap.parse_args()

    print("[PropORACLE-step7_rank_props] Starting...")
    print(f"→ Loading: {args.input}")
    df = pd.read_csv(args.input, dtype=str, encoding="utf-8-sig", 
                     engine='python').fillna("")
    
    # Explicitly ensure all string columns are str type (not object with mixed types)
    for col in df.select_dtypes(include=['object']).columns:
        df[col] = df[col].astype(str)
    out = df.copy()
    _repo_usage = Path(__file__).resolve().parents[2]
    _sd_usage = str(_repo_usage / "scripts")
    if _sd_usage not in _sys_pc.path:
        _sys_pc.path.insert(0, _sd_usage)
    try:
        from usage_redistribution import apply_usage_redistribution  # noqa: E402
        run_date = (
            str(out["game_date"].dropna().iloc[0])[:10]
            if "game_date" in out.columns and not out["game_date"].dropna().empty
            else pd.Timestamp.today().strftime("%Y-%m-%d")
        )
        sport_for_usage = "NBA"
        _hint = str(args.input or "").lower()
        if "nba1q" in _hint:
            sport_for_usage = "NBA1Q"
        elif "nba1h" in _hint:
            sport_for_usage = "NBA1H"
        out = apply_usage_redistribution(out, sport=sport_for_usage, date=run_date, repo_root=str(_repo_usage))
    except Exception as e:
        print(f"⚠️  usage redistribution skipped: {e}")

    for col, default in [("line", ""), ("pick_type", "Standard"), ("prop_norm", "")]:
        if col not in out.columns:
            out[col] = default

    if "prop_norm" not in out.columns or out["prop_norm"].eq("").all():
        if "prop_type" in out.columns:
            out["prop_norm"] = out["prop_type"].astype(str).str.lower()

    # Normalize prop names
    _PROP_NORM_MAP = {
        "3-pt made": "fg3m", "3-pt attempted": "fg3a",
        "3pt made": "fg3m", "3pt attempted": "fg3a",
        "three pointers made": "fg3m", "three pointers attempted": "fg3a",
        "3-ptm": "fg3m", "3-pta": "fg3a", "3ptm": "fg3m", "3pta": "fg3a",
        "two pointers made": "fg2m", "two pointers attempted": "fg2a",
        "2 pointers made": "fg2m", "2 pointers attempted": "fg2a",
        "2pt made": "fg2m", "2pt attempted": "fg2a",
        "2-pt made": "fg2m", "2-pt attempted": "fg2a",
        "2-ptm": "fg2m", "2-pta": "fg2a", "2ptm": "fg2m", "2pta": "fg2a",
        "free throws made": "ftm", "free throws attempted": "fta",
        "freethrowsmade": "ftm", "freethrowsattempted": "fta",
        "ft made": "ftm", "ft attempted": "fta", "ftm": "ftm", "fta": "fta",
        "fg attempted": "fga", "fg made": "fgm",
        "field goals attempted": "fga", "field goals made": "fgm",
        "fga": "fga", "fgm": "fgm",
        "fg3a": "fg3a", "fg3m": "fg3m", "fg2a": "fg2a", "fg2m": "fg2m",
    }
    out["prop_norm"] = (out["prop_norm"].astype(str).str.lower().str.strip()
                        .map(lambda x: _PROP_NORM_MAP.get(x, x)))

    prop_norm_s = out["prop_norm"].astype(str).str.lower().str.strip()
    line_num    = _to_num(out["line"])
    pick_type_s = pd.Series(_norm_pick_type_series(out["pick_type"]), index=out.index)

    # ── VECTORIZED PROJECTION ─────────────────────────────────────────────────
    v5  = _to_num(out.get("stat_last5_avg",  ""))
    v10 = _to_num(out.get("stat_last10_avg", ""))
    vs  = _to_num(out.get("stat_season_avg", ""))

    # Weighted blend (50/30/20) with partial weight normalization
    w5 = np.where(v5.notna(),  0.50, 0.0)
    w10= np.where(v10.notna(), 0.30, 0.0)
    ws = np.where(vs.notna(),  0.20, 0.0)
    total_w = w5 + w10 + ws
    proj_raw = (
        v5.fillna(0)  * w5 +
        v10.fillna(0) * w10 +
        vs.fillna(0)  * ws
    )
    proj_raw = np.where(total_w > 0.1, proj_raw / total_w, np.nan)

    # Fallback for volume props: look for {prefix}_player_last5_avg etc.
    missing_proj = np.isnan(proj_raw)
    if missing_proj.any():
        for prop_key, prefix in _PLAYER_PREFIX_BY_PROP.items():
            mask = missing_proj & (prop_norm_s == prop_key)
            if not mask.any():
                continue
            for col_cand in [f"{prefix}_player_last5_avg", f"{prefix}_last5_avg"]:
                if col_cand in out.columns:
                    fb = _to_num(out[col_cand])
                    proj_raw = np.where(mask & fb.notna(), fb, proj_raw)
                    break

    # Combo/fantasy correction
    corr = prop_norm_s.map(lambda x: _COMBO_CORRECTIONS.get(x, 1.0)).values
    proj = pd.Series(proj_raw * corr, index=out.index)
    out["projection"] = proj
    if "usage_boost_proj" in out.columns:
        out["projection"] = _to_num(out["projection"]).fillna(0.0) + _to_num(out["usage_boost_proj"]).fillna(0.0)

    out["edge"]     = proj - line_num
    out["abs_edge"] = out["edge"].abs()
    # Normalized edge keeps cross-prop comparisons on the same scale
    # (e.g., Fantasy Score vs Points).
    line_safe = line_num.replace(0, np.nan)
    out["edge_norm"] = out["edge"] / line_safe

    # ── FORCED OVER / BET DIRECTION ───────────────────────────────────────────
    forced = pick_type_s.isin(["Goblin", "Demon"]).astype(int)
    out["forced_over_only"] = forced

    bet_dir = np.where(forced.eq(1), "OVER",
              np.where(_to_num(out["edge"]) >= 0, "OVER", "UNDER"))
    out["bet_direction"] = bet_dir

    # ── ELIGIBILITY ───────────────────────────────────────────────────────────
    miss       = line_num.isna() | proj.isna()
    # Goblin/Demon with negative edge: drop to audit sheet, exclude from scoring
    neg_forced = forced.eq(1) & (_to_num(out["edge"]) < 0)
    drop_mask  = neg_forced  # rows that go to DROPPED tab only

    eligible    = (~miss & ~drop_mask).astype(int)
    void_reason = pd.Series("", index=out.index)
    void_reason = void_reason.where(~miss,      "NO_PROJECTION_OR_LINE")
    void_reason = void_reason.where(~drop_mask, "DROPPED_NEG_EDGE_GOBDEM")

    # ── HARD BLOCKS: prop+direction combinations with <45% hit rate on Standard ──
    # Derived from 9-day calibration (2026-03-06 → 2026-03-14, 19,461 props).
    # These are blocked regardless of edge — the model has no predictive power here.
    _BLOCKED_STD_OVER = {
        "stl", "blk",               # Steals OVER 41.9%, Blocks OVER 38.3%
        "fta", "freethrowsattempted",# FT Attempted OVER 43.9%
        "stocks",                    # Blks+Stls OVER 47.5% (marginal, block for safety)
    }
    _BLOCKED_STD_UNDER = {
        "fantasy",                   # Fantasy Score UNDER 28.8% — worst prop in dataset
        "pts",                       # Points UNDER 49.2% — coin flip, not worth Standard slot
    }
    _BLOCKED_ANY_UNDER = {
        "reb",                       # Rebounds UNDER 49.5% overall (13.3% some days)
    }

    is_standard = pick_type_s == "Standard"
    is_over     = pd.Series(bet_dir, index=out.index) == "OVER"
    is_under    = ~is_over

    block_std_over  = is_standard & is_over  & prop_norm_s.isin(_BLOCKED_STD_OVER)
    block_std_under = is_standard & is_under & prop_norm_s.isin(_BLOCKED_STD_UNDER)
    block_any_under = is_under & prop_norm_s.isin(_BLOCKED_ANY_UNDER)
    hard_block      = block_std_over | block_std_under | block_any_under

    eligible    = np.where(hard_block, 0, eligible)
    void_reason = pd.Series(
        np.where(block_std_over,  "BLOCKED_STD_OVER_LOW_HR",
        np.where(block_std_under, "BLOCKED_STD_UNDER_LOW_HR",
        np.where(block_any_under, "BLOCKED_UNDER_LOW_HR",
        void_reason))),
        index=out.index
    )
    eligible    = pd.Series(eligible, index=out.index)

    out["eligible"]    = eligible
    out["void_reason"] = void_reason

    elig_mask = eligible.eq(1)

    # ── VECTORIZED EDGE TRANSFORM ─────────────────────────────────────────────
    out["edge_dr"] = _edge_transform_series(_to_num(out["edge_norm"]))

    # ── VECTORIZED LINE HIT RATE ──────────────────────────────────────────────
    # Direction-aware: pick the right column priority
    bet_is_under = pd.Series(bet_dir, index=out.index) == "UNDER"

    def _pick_first_valid(*col_names) -> pd.Series:
        result = pd.Series(np.nan, index=out.index)
        for col in col_names:
            if col in out.columns:
                v = _to_num(out[col])
                result = result.where(result.notna(), v)
        return result

    hr5_over  = _pick_first_valid("line_hit_rate_over_ou_5",  "line_hit_rate_over_5",  "last5_hit_rate")
    hr10_over = _pick_first_valid("line_hit_rate_over_ou_10", "line_hit_rate_over_10")
    hr5_under = _pick_first_valid("line_hit_rate_under_ou_5", "line_hit_rate_under_5")
    hr10_under= _pick_first_valid("line_hit_rate_under_ou_10","line_hit_rate_under_10")

    # Derived under from counts if direct column missing
    l5o = _to_num(out.get("last5_over",  ""))
    l5u = _to_num(out.get("last5_under", ""))
    denom_ou = (l5o + l5u).replace(0, np.nan)
    derived_under = l5u / denom_ou
    hr5_under = hr5_under.where(hr5_under.notna(), derived_under)

    # No push fallback (1 - over) when push==0
    l5p = _to_num(out.get("last5_push", ""))
    hr5_under = hr5_under.where(hr5_under.notna(),
        np.where(l5p.fillna(0) == 0, 1.0 - hr5_over, np.nan))

    hr5  = np.where(bet_is_under, hr5_under, hr5_over)
    hr10 = np.where(bet_is_under, hr10_under, hr10_over)
    hr5  = pd.Series(hr5,  index=out.index)
    hr10 = pd.Series(hr10, index=out.index)

    # Blend 5 and 10 game windows
    line_hit_rate = (
        np.where(hr5.notna() & hr10.notna(), hr5 * 0.50 + hr10 * 0.50,
        np.where(hr5.notna(),  hr5,
        np.where(hr10.notna(), hr10, np.nan)))
    )
    out["line_hit_rate"] = pd.Series(line_hit_rate, index=out.index)
    _lo = _to_num(hr5_over)
    _l10 = _to_num(hr10_over)
    if _lo.notna().any() and _lo.dropna().median() > 1.0:
        _lo = _lo / 100.0
        _l10 = _l10 / 100.0
    line_hit_over_only = (_lo * 0.50 + _l10 * 0.50).where(
        _lo.notna() & _l10.notna(), _lo.where(_lo.notna(), _l10)
    )
    out["composite_hit_rate"] = np.where(
        bet_is_under, 1.0 - line_hit_over_only, line_hit_over_only
    )
    out["composite_hit_rate"] = pd.to_numeric(out["composite_hit_rate"], errors="coerce")

    # ── VECTORIZED MINUTES CERTAINTY ──────────────────────────────────────────
    _MIN_TIER_MAP = {"HIGH": 1.00, "MEDIUM": 0.90, "LOW": 0.75}
    out["minutes_certainty"] = (
        out.get("minutes_tier", pd.Series("", index=out.index))
        .astype(str).str.upper()
        .map(lambda x: _MIN_TIER_MAP.get(x, 0.80))
    )

    # ── VECTORIZED PROP WEIGHT / RELIABILITY ─────────────────────────────────
    out["prop_weight"]      = prop_norm_s.map(lambda x: _PROP_WEIGHTS.get(x, 0.93))
    out["reliability_mult"] = pick_type_s.map(lambda x: _RELIABILITY_MAP.get(x, 0.97))

    # ── VECTORIZED DEF ADJUSTMENT ─────────────────────────────────────────────
    def_rank = _to_num(out.get("OVERALL_DEF_RANK", ""))
    def_adj  = ((def_rank - 15.0) / 15.0 * 0.06).fillna(0.0)
    out["def_adj"] = def_adj

    # ── GAME CONTEXT ADJUSTMENT (Step 6b: Vegas lines) ────────────────────────
    # ctx_adj: -0.08 low total on combo prop, -0.05 blowout risk, -0.15 both
    ctx_adj  = _to_num(out["ctx_adj"]).fillna(0.0)  if "ctx_adj"  in out.columns else pd.Series(0.0, index=out.index)
    out["ctx_adj"] = ctx_adj

    # ── SCHEDULE / REST ADJUSTMENT (Step 6c: B2B, rest days) ─────────────────
    # rest_adj: -0.10 B2B, 0.00 baseline (1-day rest), +0.02 two days, +0.04 three+
    rest_adj = _to_num(out["rest_adj"]).fillna(0.0) if "rest_adj" in out.columns else pd.Series(0.0, index=out.index)
    out["rest_adj"] = rest_adj

    # ── PACE SIGNAL ──────────────────────────────────────────────────────────
    # Derived from game_total (Step 6b). High total = fast pace = more possessions.
    # Neutral at 230pts, ±0.02 per 10pt deviation, capped ±0.04.
    # Direction-aware: fast pace helps OVER props, hurts UNDER props.
    if "game_total" in out.columns:
        pace_raw   = (_to_num(out["game_total"]).fillna(230.0) - 230.0) / 10.0 * 0.02
        pace_adj   = pace_raw.clip(-0.04, 0.04)
        pace_adj_dr = pd.Series(
            np.where(bet_is_under, -pace_adj, pace_adj), index=out.index
        )
    else:
        pace_adj_dr = pd.Series(0.0, index=out.index)
    out["pace_adj"] = pace_adj_dr

    # ── PROP-SPECIFIC OPP ALLOWANCE ───────────────────────────────────────────
    # intel_opp_vs_league_pct (Step 6e) measures how much more/less this
    # opponent gives up vs league avg for the specific stat being scored.
    # +8% on an AST prop = opponent gives up 8% more assists = stronger OVER signal.
    # Converted to a projection multiplier: ±0.02 per 5% deviation, capped ±0.06.
    # This is separate from the general intel_def_z weight (0.40) in the score —
    # it directly adjusts the projection so edge and hit rates benefit too.
    opp_pct_raw    = _to_num(out.get("intel_opp_vs_league_pct", pd.Series(np.nan, index=out.index))).fillna(0.0) / 100.0
    opp_prop_adj   = (opp_pct_raw / 0.05 * 0.02).clip(-0.06, 0.06)
    opp_prop_adj_dr = pd.Series(
        np.where(bet_is_under, -opp_prop_adj, opp_prop_adj), index=out.index
    )
    out["opp_prop_adj"] = opp_prop_adj_dr

    proj_base = _to_num(out["projection"])
    out["projection_adj"] = proj_base * (
        1.0 + def_adj + ctx_adj + rest_adj + pace_adj_dr + opp_prop_adj_dr
    )
    out["edge_adj"]       = out["projection_adj"] - line_num
    out["edge_adj_norm"]  = out["edge_adj"] / line_safe

    # ── VECTORIZED EDGE_ADJ_DR (direction-aware) ──────────────────────────────
    edge_adj_signed = np.where(bet_is_under, -_to_num(out["edge_adj_norm"]), _to_num(out["edge_adj_norm"]))
    out["edge_adj_dr"] = _edge_transform_series(pd.Series(edge_adj_signed, index=out.index))

    # ── VECTORIZED DEF RANK SIGNAL ────────────────────────────────────────────
    signal_raw = ((def_rank - 1.0) / 29.0 * 2.0 - 1.0)
    def_signal = np.where(bet_is_under, -signal_raw, signal_raw)
    out["def_rank_signal"] = pd.Series(def_signal, index=out.index)

    # ── VECTORIZED PROP HIT RATE PRIOR ───────────────────────────────────────
    base_prior = prop_norm_s.map(lambda x: _PROP_HR_PRIOR_OVER.get(x, 0.545))
    under_prior = prop_norm_s.map(
        lambda x: _PROP_HR_PRIOR_UNDER_OVERRIDE.get(x, 1.0 - _PROP_HR_PRIOR_OVER.get(x, 0.545))
    )
    out["prop_hr_prior"] = np.where(bet_is_under, under_prior, base_prior)

    # ── VECTORIZED AVG VS LINE ────────────────────────────────────────────────
    for col in ("stat_last5_avg", "stat_last10_avg", "stat_season_avg"):
        out[col + "_num"] = _to_num(out[col]) if col in out.columns else pd.Series(np.nan, index=out.index)

    def _avg_vs_line_vec(avg_col: str, w: float) -> pd.Series:
        v = _to_num(out[avg_col + "_num"]) if (avg_col + "_num") in out.columns else pd.Series(np.nan, index=out.index)
        raw = np.clip((v - line_safe) / line_safe, -1.0, 1.0)
        raw = np.where(bet_is_under, -raw, raw)
        return pd.Series(np.where(v.notna() & line_safe.notna(), raw * w, np.nan), index=out.index)

    avl5  = _avg_vs_line_vec("stat_last5_avg",  0.50)
    avl10 = _avg_vs_line_vec("stat_last10_avg", 0.30)
    avls  = _avg_vs_line_vec("stat_season_avg", 0.20)

    wt5  = np.where(_to_num(out.get("stat_last5_avg_num",  "")).notna() & line_safe.notna(), 0.50, 0.0)
    wt10 = np.where(_to_num(out.get("stat_last10_avg_num", "")).notna() & line_safe.notna(), 0.30, 0.0)
    wts  = np.where(_to_num(out.get("stat_season_avg_num", "")).notna() & line_safe.notna(), 0.20, 0.0)
    total_avl_w = pd.Series(wt5 + wt10 + wts, index=out.index)

    avg_vs_line = (avl5.fillna(0) + avl10.fillna(0) + avls.fillna(0))
    avg_vs_line = avg_vs_line.where(total_avl_w > 0.1, 0.0)
    out["avg_vs_line"] = avg_vs_line

    # ── Z-SCORE (direction-aware) ─────────────────────────────────────────────
    def zcol(s: pd.Series, direction_aware: bool = False) -> pd.Series:
        x = pd.to_numeric(s, errors="coerce")
        result = pd.Series(0.0, index=x.index)
        if direction_aware and "bet_direction" in out.columns:
            for direction in ("OVER", "UNDER"):
                dir_mask = elig_mask & (out["bet_direction"].astype(str).str.upper() == direction)
                if dir_mask.sum() < 2:
                    continue
                mu = x[dir_mask].mean()
                sd = x[dir_mask].std()
                if sd and not np.isnan(sd) and sd > 1e-9:
                    z_vals = (x[dir_mask] - mu) / sd
                    result.loc[dir_mask.index[dir_mask]] = z_vals.values
            return result
        mu = x[elig_mask].mean()
        sd = x[elig_mask].std()
        if sd and not np.isnan(sd) and sd > 1e-9:
            return (x - mu) / sd
        return result

    out["edge_z"]        = zcol(out["edge_norm"],      direction_aware=True)
    out["line_hit_z"]    = zcol(out["line_hit_rate"],   direction_aware=True)
    out["min_z"]         = zcol(out["minutes_certainty"])
    out["def_rank_z"]    = zcol(out["def_rank_signal"],  direction_aware=True)
    out["avg_vs_line_z"] = zcol(out["avg_vs_line"],      direction_aware=True)
    out["prop_hr_z"]     = zcol(out["prop_hr_prior"],    direction_aware=True)

    # ── Intel signals (from step6e) ───────────────────────────────────────────
    # intel_season_hit_rate: % of season games OVER this line (0-100 scale → normalise)
    intel_shr_raw  = _to_num(out.get("intel_season_hit_rate", pd.Series(np.nan, index=out.index))).fillna(50.0) / 100.0
    # intel_opp_vs_league_pct: how generous/tight this opponent is (+= give up more)
    intel_def_raw  = _to_num(out.get("intel_opp_vs_league_pct", pd.Series(np.nan, index=out.index))).fillna(0.0) / 100.0
    # intel_cv_pct: consistency — lower = better. Invert so high = consistent
    intel_cv_raw   = _to_num(out.get("intel_cv_pct", pd.Series(np.nan, index=out.index))).fillna(50.0)
    intel_cons_raw = (100.0 - intel_cv_raw.clip(0, 100)) / 100.0  # 0-1, higher=consistent

    out["intel_shr_z"]  = zcol(pd.Series(intel_shr_raw,  index=out.index), direction_aware=True)
    out["intel_def_z"]  = zcol(pd.Series(intel_def_raw,  index=out.index), direction_aware=True)
    out["intel_cons_z"] = zcol(pd.Series(intel_cons_raw, index=out.index))

    # ── FINAL SCORE ───────────────────────────────────────────────────────────
    # Edge > Rank > L5 hierarchy:
    # 1) edge_adj_dr is primary driver
    # 2) structural/context factors are secondary rank stabilizers
    # 3) L5/L10 style signals act only as bounded confidence modifiers

    b2b_penalty    = np.where(out.get("b2b_flag",    pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.20, 0.0)
    blowout_penalty= np.where(out.get("blowout_risk", pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.10, 0.0)
    low_total_pen  = np.where(out.get("low_total_flag", pd.Series(False, index=out.index)).astype(str).str.lower() == "true", -0.10, 0.0)

    _repo_inj = Path(__file__).resolve().parents[2]
    _sd_inj = _repo_inj / "scripts"
    if str(_sd_inj) not in _sys_pc.path:
        _sys_pc.path.insert(0, str(_sd_inj))
    from espn_injuries import auto_injuries_csv_from_outputs, penalty_series_for_slate  # noqa: E402

    _inj_path = str(args.injuries_csv or "").strip()
    if not _inj_path and str(getattr(args, "slate_date", "") or "").strip():
        _cand = auto_injuries_csv_from_outputs(_repo_inj, str(args.slate_date).strip(), "NBA")
        _inj_path = str(_cand) if _cand else ""
    _inj_pen = (
        penalty_series_for_slate(out, "player", "team", "NBA", _inj_path)
        if _inj_path
        else pd.Series(0.0, index=out.index)
    )

    # Pace z-score (direction-aware — already in out["pace_adj"] but score via z for scale)
    out["pace_z"] = zcol(out["pace_adj"], direction_aware=True)

    edge_core = _to_num(out["edge_adj_dr"]).fillna(0.0)
    rank_support = (
        _to_num(out["def_rank_z"]).fillna(0.0)   * 0.38
        + _to_num(out["prop_hr_z"]).fillna(0.0) * 0.30
        + _to_num(out["intel_def_z"]).fillna(0.0) * 0.20
        + _to_num(out["intel_cons_z"]).fillna(0.0) * 0.16
        + _to_num(out["pace_z"]).fillna(0.0) * 0.12
        + _to_num(out["min_z"]).fillna(0.0) * 0.10
        + pd.Series(b2b_penalty, index=out.index)
        + pd.Series(blowout_penalty, index=out.index)
        + pd.Series(low_total_pen, index=out.index)
        + _inj_pen.reindex(out.index).fillna(0.0)
    )

    # L5/L10 support is bounded and cannot dominate edge/rank.
    l5_support_signal = (
        _to_num(out["line_hit_z"]).fillna(0.0) * 0.60
        + _to_num(out["avg_vs_line_z"]).fillna(0.0) * 0.40
    )
    l5_support_mod = np.clip(0.08 * l5_support_signal, -0.12, 0.12)

    usage_bonus = np.clip(_to_num(out.get("usage_boost", pd.Series(0.0, index=out.index))).fillna(0.0) * 5.0, 0.0, 0.5)
    out["usage_bonus"] = usage_bonus
    base_raw = (edge_core * 1.20) + (rank_support * 0.35) + usage_bonus
    score_raw = base_raw * (1.0 + l5_support_mod)
    score_raw = (
        score_raw
        * _to_num(out["prop_weight"]).fillna(1.0)
        * _to_num(out["reliability_mult"]).fillna(1.0)
    )

    # Hard edge gate: keep rows where edge favors the play (same sign as edge_adj_dr).
    # Raw edge_adj = projection_adj - line is >0 for OVER-style math only; UNDER wins
    # when projection is below line, so edge_adj < 0 even for excellent unders.
    edge_gate = _to_num(out["edge_adj_dr"]).fillna(-999.0) > 0.0
    score_raw = score_raw.where(elig_mask & edge_gate, np.nan)

    out["l5_support_mod"] = pd.Series(l5_support_mod, index=out.index)
    out["rank_score_raw"] = score_raw
    out["rank_score"] = out["rank_score_raw"]
    out["ml_prob"], out["ml_edge"], out["final_score"] = _apply_ml_blend(out, out["rank_score_raw"], args.input)
    _apply_consistency_grade_scores(out, "NBA")

    # Game script risk adjustment
    from datetime import datetime, timezone

    _repo_gs = _repo_root_pc()
    _sd_gs = str(_repo_gs / "scripts")
    if _sd_gs not in _sys_pc.path:
        _sys_pc.path.insert(0, _sd_gs)
    from game_script_risk import get_game_script_multiplier  # noqa: E402

    _fallback_gd = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    if "start_time" in out.columns:
        _st = out["start_time"].astype(str).str.strip()
        _dp = _st.str[:10]
        _gd_series = _dp.where(_dp.str.match(r"^\d{4}-\d{2}-\d{2}$"), _fallback_gd)
    else:
        _gd_series = pd.Series([_fallback_gd] * len(out), index=out.index)
    _team_col = next((c for c in ("team", "Team") if c in out.columns), "team")
    _prop_gs = "prop_norm" if "prop_norm" in out.columns else "prop_type"
    _gmults: list[float] = []
    _gnotes: list[str] = []
    for _i in range(len(out)):
        _r = out.iloc[_i]
        _gd = str(_gd_series.iloc[_i])
        _tm = str(_r.get(_team_col, "") or "").strip()
        _pt = str(_r.get(_prop_gs, "") or "").strip()
        _gm, _gn = get_game_script_multiplier(_tm, "NBA", _pt, _gd)
        _gmults.append(round(float(_gm), 3))
        _gnotes.append(_gn)
    out["game_script_mult"] = _gmults
    out["game_script_note"] = _gnotes
    out["final_score"] = _to_num(out["final_score"]).astype(float) * pd.Series(_gmults, dtype=float).values
    out["rank_score_final"] = out["final_score"]
    # Keep downstream compatibility: rank_score remains the final ranking value.
    out["rank_score"] = out["rank_score_final"]
    out["tier"] = pd.Series(
        _tier_from_score_series(_to_num(out["rank_score"])), index=out.index
    )
    out.loc[~elig_mask, "tier"] = "D"
    out = build_feature_vector(out, "NBA")
    out = apply_ticket_eligibility_voids(out, "NBA")
    out = out.sort_values(by="final_score", ascending=False, na_position="last", kind="mergesort")

    # Split here — after all scoring/tier columns are populated
    dropped_df = out.loc[drop_mask].copy()
    out_active = out.loc[~drop_mask].copy()

    # ── WRITE XLSX (with explicit UTF-8 handling) ──────────────────────────────
    # Sheets:
    #   ALL        — all active rows (neg-edge Gob/Dem excluded)
    #   STANDARD   — Standard pick type only
    #   GOB_DEM    — Goblin + Demon (positive-edge only)
    #   ELIGIBLE   — active rows that passed scoring
    #   DROPPED    — neg-edge Goblin/Demon, for hit/miss audit only

    std_mask_active  = out_active["pick_type"].astype(str).str.strip().str.lower().str.contains("standard")
    gobdem_mask      = ~std_mask_active
    elig_mask_active = out_active["eligible"].eq(1)

    def _safe_excel_write(writer, df, sheet_name):
        if df.empty:
            pd.DataFrame(columns=df.columns).to_excel(writer, sheet_name=sheet_name, index=False)
        else:
            df.to_excel(writer, sheet_name=sheet_name, index=False)

    if HAS_XLSXWRITER:
        try:
            with pd.ExcelWriter(args.output, engine="xlsxwriter",
                               engine_kwargs={'options': {'strings_to_urls': False}}) as w:
                _safe_excel_write(w, out_active,                        "ALL")
                _safe_excel_write(w, out_active.loc[std_mask_active],   "STANDARD")
                _safe_excel_write(w, out_active.loc[gobdem_mask],       "GOB_DEM")
                _safe_excel_write(w, out_active.loc[elig_mask_active],  "ELIGIBLE")
                _safe_excel_write(w, dropped_df,                        "DROPPED")
            print(f"✅ Saved → {args.output} (xlsxwriter, UTF-8 encoded)")
        except Exception as e:
            print(f"⚠️  xlsxwriter failed: {e}, falling back to openpyxl")
            _write_xlsx_openpyxl(args.output, out_active, elig_mask_active)
    else:
        _write_xlsx_openpyxl(args.output, out_active, elig_mask_active)

    print(f"✅ Saved → {args.output}")
    print(f"ALL rows (active) : {len(out_active)}")
    print(f"STANDARD rows     : {int(std_mask_active.sum())}")
    print(f"GOB_DEM rows      : {int(gobdem_mask.sum())}")
    print(f"DROPPED rows      : {len(dropped_df)}  (neg-edge Gob/Dem, audit only)")
    print()
    print("Tier counts (ALL active):")
    print(out_active["tier"].value_counts().to_string())
    print()
    print("Ineligible reason breakdown (active):")
    vr = out_active.loc[~elig_mask_active, "void_reason"].value_counts()
    print(vr.to_string() if len(vr) else "(none)")
    print()
    print("Score percentiles (eligible):")
    rs = _to_num(out_active.loc[elig_mask_active, "rank_score"])
    print(rs.quantile([0.50, 0.70, 0.80, 0.85, 0.90, 0.95]).round(3).to_string())


if __name__ == "__main__":
    main()
