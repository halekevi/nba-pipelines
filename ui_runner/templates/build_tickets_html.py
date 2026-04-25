"""
build_tickets_html.py
=====================
DEPRECATED — do not use for production tickets UI.

The pipeline writes tickets_latest.json via combined_slate_tickets.py --write-web, then
scripts/build_ticket_eval.py emits graded ui_runner/templates/tickets_latest.html.

This script remains for one-off debugging from a combined_slate *.xlsx only.

Legacy: converts combined_slate_tickets_*.xlsx into HTML under ui_runner/templates/.

Usage:
    py -3.14 build_tickets_html.py
    py -3.14 build_tickets_html.py --date 2026-02-24
    py -3.14 build_tickets_html.py --input path\\to\\file.xlsx
"""

from __future__ import annotations

import argparse
import html as html_lib
import re
import sys
from datetime import datetime
from pathlib import Path
from typing import Any

# Add scripts directory to path for nav_renderer
sys.path.append(str(Path(__file__).resolve().parent.parent.parent / "scripts"))
try:
    from nav_renderer import render_static_nav
except ImportError:
    # Fallback for environments where scripts/ isn't easily reachable
    def render_static_nav(active_tab: str) -> str:
        return "<!-- nav_renderer failed to import -->"

try:
    import openpyxl
except ImportError:
    print("ERROR: openpyxl not installed. Run: pip install openpyxl")
    sys.exit(1)

# ── Paths ─────────────────────────────────────────────────────────────────────
SCRIPT_DIR  = Path(__file__).resolve().parent
OUTPUTS_DIR = SCRIPT_DIR / "outputs"
UI_DOCS_DIR = SCRIPT_DIR  # saves directly into templates/


# ── Helpers ───────────────────────────────────────────────────────────────────
def find_latest_tickets(date_str: str | None = None) -> Path:
    pattern = "combined_slate_tickets_*.xlsx"
    candidates = []
    for d in [SCRIPT_DIR] + sorted(OUTPUTS_DIR.glob("*"), reverse=True):
        if Path(d).is_dir():
            for f in Path(d).glob(pattern):
                if "TOP3" not in f.name:
                    candidates.append(f)
    if not candidates:
        raise FileNotFoundError(f"No combined_slate_tickets_*.xlsx found under {SCRIPT_DIR}")
    if date_str:
        matches = [c for c in candidates if date_str in c.name]
        if matches:
            return sorted(matches, key=lambda p: p.stat().st_mtime, reverse=True)[0]
    return sorted(candidates, key=lambda p: p.stat().st_mtime, reverse=True)[0]


def extract_date(path: Path) -> tuple[str, str]:
    m = re.search(r"(\d{4}-\d{2}-\d{2})", path.name)
    if m:
        try:
            d = datetime.strptime(m.group(1), "%Y-%m-%d")
            return d.strftime("%b %d, %Y").upper(), m.group(1)
        except ValueError:
            pass
    now = datetime.now()
    return now.strftime("%b %d, %Y").upper(), now.strftime("%Y-%m-%d")


def h(v: Any) -> str:
    return html_lib.escape(str(v) if v is not None else "")


def fmt(v: Any, dec: int = 2) -> str:
    try:
        return f"{float(v):.{dec}f}"
    except (TypeError, ValueError):
        return str(v) if v is not None else "—"


def pct(v: Any) -> str:
    try:
        f = float(v)
        return f"{f*100:.0f}%" if f <= 1.0 else f"{f:.0f}%"
    except (TypeError, ValueError):
        return str(v) if v is not None else "—"


def rate_color(v: Any) -> tuple[str, str]:
    try:
        f = float(v)
        if f <= 1.0: f *= 100
    except (TypeError, ValueError):
        return "rgba(255,255,255,0.92)", "var(--muted2)"
    if f >= 80:  return "#6ee7b7", "var(--green)"
    if f >= 65:  return "#6ee7b7", "var(--green)"
    if f >= 55:  return "#fcd34d", "var(--amber)"
    if f >= 50:  return "#93c5fd", "var(--blue)"
    return "#fca5a5", "var(--red)"


def rate_bar(v: Any) -> str:
    try:
        f = float(v)
        if f <= 1.0: f *= 100
    except (TypeError, ValueError):
        return f'<span class="mono muted">{h(v)}</span>'
    tc, bc = rate_color(v)
    return (f'<div class="rbar"><div class="rbar-bg">'
            f'<div class="rbar-fill" style="width:{min(f,100):.1f}%;background:{bc}"></div>'
            f'</div><span class="rbar-num" style="color:{tc}">{f:.0f}%</span></div>')


def tier_chip(t: Any) -> str:
    s = str(t).strip().upper() if t else ""
    cls = {"A":"chip-a","B":"chip-b","C":"chip-c","D":"chip-d"}.get(s, "chip-d")
    return f'<span class="chip {cls}">T{h(s)}</span>' if s else "—"


def pick_chip(p: Any) -> str:
    s = str(p).strip().lower() if p else ""
    if "goblin" in s:  return '<span class="chip chip-goblin">&#x1F47A; Goblin</span>'
    if "demon"  in s:  return '<span class="chip chip-demon">&#x1F608; Demon</span>'
    return '<span class="chip chip-std">&#x2B50; Std</span>'


def dir_chip(d: Any) -> str:
    s = str(d).strip().upper() if d else ""
    if s == "OVER":  return '<span class="chip chip-over">&#x25B2; OVER</span>'
    if s == "UNDER": return '<span class="chip chip-under">&#x25BC; UNDER</span>'
    return f'<span class="chip chip-d">{h(d)}</span>' if d else "—"


def def_chip(d: Any) -> str:
    s = str(d).strip().title() if d else ""
    cls = {"Elite":"chip-demon","Strong":"chip-c","Average":"chip-std",
           "Weak":"chip-a","Very Weak":"chip-a"}.get(s, "chip-d")
    return f'<span class="chip {cls}">{h(s)}</span>' if s else "—"


def sport_chip(s: Any) -> str:
    v = str(s).strip().upper() if s else ""
    if "NBA"    in v: return '<span class="chip chip-nba">NBA</span>'
    if "CBB"    in v: return '<span class="chip chip-cbb">CBB</span>'
    if "NHL"    in v: return '<span class="chip chip-nhl">NHL</span>'
    if "SOCCER" in v: return '<span class="chip chip-soc">SOC</span>'
    return f'<span class="chip chip-d">{h(v)}</span>'


# ── Sheet parsers ─────────────────────────────────────────────────────────────
def read_flat_sheet(ws) -> list[dict]:
    """Sheets like Full Slate / NBA Slate — first row is header."""
    rows = list(ws.iter_rows(values_only=True))
    if not rows: return []
    headers = [str(c).strip() if c else f"_c{i}" for i, c in enumerate(rows[0])]
    return [dict(zip(headers, r)) for r in rows[1:] if any(v is not None for v in r)]


def read_ticket_sheet(ws) -> list[dict]:
    """
    Ticket sheets pattern:
      Row N:   '  Ticket #1  · 3-Leg NBA Goblin · Power: 4.37x ...'  (title)
      Row N+1: '#', 'Player', 'Team', ...  (headers)
      Row N+2+: data rows
      Row M:   next ticket title...
    """
    rows = list(ws.iter_rows(values_only=True))
    tickets = []
    i = 0
    while i < len(rows):
        first = str(rows[i][0]).strip() if rows[i][0] else ""
        if first.startswith("Ticket"):
            title = first
            i += 1
            if i >= len(rows): break
            headers = [str(c).strip() if c else f"_c{j}" for j, c in enumerate(rows[i])]
            i += 1
            legs = []
            while i < len(rows):
                dr = rows[i]
                fd = str(dr[0]).strip() if dr[0] else ""
                if fd.startswith("Ticket"): break
                if any(v is not None for v in dr):
                    legs.append(dict(zip(headers, dr)))
                i += 1
            tickets.append({"title": title, "legs": legs})
        else:
            i += 1
    return tickets


def parse_ticket_title(title: str) -> dict:
    info = {"num":"","desc":"","power":"","flex":"","avg_hit":"","est_prob":"","avg_score":""}
    m = re.search(r"Ticket\s*#?(\d+)", title)
    if m: info["num"] = m.group(1)
    m = re.search(r"·\s*([^·]+Leg[^·]*)", title)
    if m: info["desc"] = m.group(1).strip()
    m = re.search(r"Power:\s*([\d.]+x)", title)
    if m: info["power"] = m.group(1)
    m = re.search(r"Flex:\s*([\d.]+x)", title)
    if m: info["flex"] = m.group(1)
    m = re.search(r"Avg Hit Rate:\s*([\d.]+%)", title)
    if m: info["avg_hit"] = m.group(1)
    m = re.search(r"Est Win Prob:\s*([\d.]+%)", title)
    if m: info["est_prob"] = m.group(1)
    m = re.search(r"Avg Rank Score:\s*([\d.]+)", title)
    if m: info["avg_score"] = m.group(1)
    return info


# ── Slate table ───────────────────────────────────────────────────────────────
def build_slate_table(rows: list[dict], limit: int = 500) -> str:
    if not rows:
        return '<div class="alert alert-amber"><div class="alert-title">No data.</div></div>'
    rows = rows[:limit]
    body = ""
    for r in rows:
        sport_val = str(r.get('Sport','')).lower()
        body += f"""<tr data-player="{h(r.get('Player','').lower())}" data-sport="{sport_val}" data-prop="{h(str(r.get('Prop','')).lower())}" data-team="{h(str(r.get('Team','')).lower())}">
          <td>{sport_chip(r.get('Sport'))}</td>
          <td>{tier_chip(r.get('Tier'))}</td>
          <td class="mono right">{fmt(r.get('Rank Score'))}</td>
          <td><strong>{h(r.get('Player',''))}</strong><div class="sub">{h(r.get('Team',''))} vs {h(r.get('Opp',''))}</div></td>
          <td class="mono">{h(r.get('Prop',''))}</td>
          <td>{pick_chip(r.get('Pick Type'))}</td>
          <td class="mono right">{fmt(r.get('Line'),1)}</td>
          <td>{dir_chip(r.get('Dir'))}</td>
          <td class="mono right pos">{fmt(r.get('Edge'),2)}</td>
          <td>{rate_bar(r.get('Hit Rate'))}</td>
          <td class="mono right">{fmt(r.get('L5 Avg'),1)}</td>
          <td>{def_chip(r.get('Def Tier'))}</td>
          <td class="mono muted small">{h(r.get('Game Time',''))}</td>
        </tr>"""
    return f"""<div class="table-wrap scrollx">
  <table>
    <thead><tr>
      <th></th><th>TIER</th><th class="right">SCORE</th><th>PLAYER</th>
      <th>PROP</th><th>TYPE</th><th class="right">LINE</th><th>DIR</th>
      <th class="right">EDGE</th><th>HIT RATE</th><th class="right">L5</th>
      <th>DEF</th><th>TIME</th>
    </tr></thead>
    <tbody>{body}</tbody>
  </table>
</div>"""


# ── Ticket card ───────────────────────────────────────────────────────────────
def build_ticket_card(ticket: dict) -> str:
    info = parse_ticket_title(ticket["title"])
    pills = ""
    if info["power"]:    pills += f'<span class="pill pill-green">&#x26A1; {h(info["power"])} Power</span>'
    if info["flex"]:     pills += f'<span class="pill pill-blue">&#x1F500; {h(info["flex"])} Flex</span>'
    if info["avg_hit"]:  pills += f'<span class="pill pill-amber">&#x1F3AF; {h(info["avg_hit"])} Hit</span>'
    if info["est_prob"]: pills += f'<span class="pill pill-purple">&#x1F4CA; {h(info["est_prob"])} Win</span>'

    legs_html = ""
    for leg in ticket["legs"]:
        legs_html += f"""<tr>
          <td><strong>{h(leg.get('Player',''))}</strong><div class="sub">{h(leg.get('Team',''))} vs {h(leg.get('Opp',''))}</div></td>
          <td class="mono">{h(leg.get('Prop',''))}</td>
          <td>{pick_chip(leg.get('Pick Type'))}</td>
          <td class="mono right">{fmt(leg.get('Line'),1)}</td>
          <td>{dir_chip(leg.get('Dir'))}</td>
          <td class="mono right pos">{fmt(leg.get('Edge'),2)}</td>
          <td>{rate_bar(leg.get('Hit Rate'))}</td>
          <td class="mono right muted">{fmt(leg.get('L5 Avg'),1)}</td>
          <td class="mono right muted">{fmt(leg.get('Rank Score'),2)}</td>
        </tr>"""

    # Mobile leg cards
    leg_cards_html = ""
    for leg in ticket["legs"]:
        hr_val = ""
        try:
            f = float(leg.get("Hit Rate", 0))
            hr_val = f"{f*100:.0f}%" if f <= 1.0 else f"{f:.0f}%"
        except Exception:
            hr_val = str(leg.get("Hit Rate","—"))
        lc, _ = rate_color(leg.get("Hit Rate"))
        leg_cards_html += f"""<div class="leg-card">
  <div class="leg-player">{h(leg.get('Player',''))}</div>
  <div class="leg-matchup">{h(leg.get('Team',''))} vs {h(leg.get('Opp',''))}</div>
  <div class="leg-row"><span class="leg-label">PROP</span><span class="leg-value">{h(leg.get('Prop',''))} · {fmt(leg.get('Line'),1)}</span></div>
  <div class="leg-row"><span class="leg-label">TYPE</span>{pick_chip(leg.get('Pick Type'))}</div>
  <div class="leg-row"><span class="leg-label">DIRECTION</span>{dir_chip(leg.get('Dir'))}</div>
  <div class="leg-row"><span class="leg-label">HIT RATE</span><span class="leg-value" style="color:{lc}">{hr_val}</span></div>
  <div class="leg-row"><span class="leg-label">EDGE</span><span class="leg-value pos">{fmt(leg.get('Edge'),2)}</span></div>
</div>"""

    return f"""<div class="ticket-card">
  <div class="ticket-header">
    <div class="ticket-num">#{h(info['num'])}</div>
    <div class="ticket-desc">{h(info['desc'])}</div>
    <div class="ticket-pills">{pills}</div>
  </div>
  <div class="table-wrap scrollx">
    <table>
      <thead><tr>
        <th>PLAYER</th><th>PROP</th><th>TYPE</th><th class="right">LINE</th>
        <th>DIR</th><th class="right">EDGE</th><th>HIT RATE</th>
        <th class="right">L5</th><th class="right">SCORE</th>
      </tr></thead>
      <tbody>{legs_html}</tbody>
    </table>
  </div>
  <div class="leg-cards">{leg_cards_html}</div>
</div>"""


def build_best_ticket_summary(tickets: list[dict]) -> str:
    """Build a highlighted 'Best Ticket' banner showing the #1 ticket from a group."""
    if not tickets:
        return ""
    best = tickets[0]
    info = parse_ticket_title(best["title"])
    legs = best["legs"]
    if not legs:
        return ""

    # compute avg hit rate display
    hrs = []
    for leg in legs:
        try:
            f = float(leg.get("Hit Rate", 0))
            hrs.append(f if f > 1 else f * 100)
        except (TypeError, ValueError):
            pass
    avg_hr = sum(hrs) / len(hrs) if hrs else 0
    hr_color = "#39ff6e" if avg_hr >= 75 else "#f0a500" if avg_hr >= 60 else "#ff4d4d"

    rows_html = ""
    for leg in legs:
        hr_val = ""
        try:
            f = float(leg.get("Hit Rate", 0))
            hr_val = f"{f*100:.0f}%" if f <= 1.0 else f"{f:.0f}%"
        except (TypeError, ValueError):
            hr_val = str(leg.get("Hit Rate", "—"))
        lc, _ = rate_color(leg.get("Hit Rate"))
        rows_html += f"""<div style="display:flex;align-items:center;gap:10px;padding:7px 0;border-bottom:1px solid rgba(255,255,255,.04);">
  <span style="font-size:12px;color:var(--text);flex:1;font-weight:600">{h(leg.get('Player',''))}</span>
  <span style="font-size:11px;color:var(--muted);font-family:'Inter',sans-serif">{h(leg.get('Prop',''))} {fmt(leg.get('Line'),1)}</span>
  {dir_chip(leg.get('Dir'))}
  <span style="font-size:11px;color:{lc};font-family:'Inter',sans-serif;min-width:38px;text-align:right">{hr_val}</span>
</div>"""

    power_str = f"⚡ {info['power']} Power" if info['power'] else ""
    flex_str  = f"🔄 {info['flex']} Flex"  if info['flex']  else ""

    return f"""<div style="background:linear-gradient(135deg,rgba(200,255,0,.06) 0%,rgba(0,229,255,.04) 100%);border:1px solid rgba(200,255,0,.2);border-left:3px solid var(--accent);border-radius:12px;padding:16px;margin-bottom:20px;">
  <div style="display:flex;align-items:center;gap:10px;margin-bottom:12px;flex-wrap:wrap;">
    <span style="font-family:'Bebas Neue',sans-serif;font-size:13px;letter-spacing:2px;color:var(--accent)">⭐ BEST TICKET</span>
    <span style="font-size:10px;color:var(--muted);letter-spacing:1px">{h(info['desc'])}</span>
    <span style="margin-left:auto;font-family:'Bebas Neue',sans-serif;font-size:22px;color:{hr_color}">{avg_hr:.0f}% HR</span>
  </div>
  <div style="margin-bottom:10px;">{rows_html}</div>
  <div style="display:flex;gap:8px;flex-wrap:wrap;margin-top:8px;">
    {"<span style='font-size:10px;padding:3px 10px;border-radius:20px;background:rgba(57,255,110,.1);color:var(--green);border:1px solid rgba(57,255,110,.2)'>" + power_str + "</span>" if power_str else ""}
    {"<span style='font-size:10px;padding:3px 10px;border-radius:20px;background:rgba(0,229,255,.1);color:var(--cyan);border:1px solid rgba(0,229,255,.2)'>" + flex_str + "</span>" if flex_str else ""}
  </div>
</div>"""


def build_ticket_group(tickets: list[dict], limit: int = 60) -> str:
    if not tickets:
        return '<div class="muted" style="padding:20px;font-family:\'DM Mono\',monospace;font-size:12px">No tickets in this group.</div>'
    cards = "".join(build_ticket_card(t) for t in tickets[:limit])
    note  = (f'<div class="muted small" style="padding:8px;font-family:\'DM Mono\',monospace">'
             f'Showing {min(len(tickets),limit)} of {len(tickets)} tickets</div>') if len(tickets) > limit else ""
    return cards + note


# ── Tab section builder ───────────────────────────────────────────────────────
def build_tab_section(wb, sheet_names: list[str], id_prefix: str) -> tuple[str, str]:
    """Group sheets by pick type, sub-tab by leg count."""
    # group by pick type: Goblin, Standard, Demon, Mix
    type_map: dict[str, dict[str, list]] = {}
    for name in sheet_names:
        if name not in wb.sheetnames: continue
        # extract pick type from sheet name
        # handles: "NBA Goblin 3-Leg", "MIX Standard 3-Leg", "COMBO Mix 3-Leg"
        m_leg  = re.search(r"(\d+)-Leg", name)
        m_type = re.search(r"\b(Goblin|Standard|Demon|Mix)\b", name, re.I)
        if not m_leg: continue
        leg  = m_leg.group(1)
        ptype = m_type.group(1).title() if m_type else "Other"
        tickets = read_ticket_sheet(wb[name])
        if not tickets: continue
        type_map.setdefault(ptype, {}).setdefault(leg, [])
        type_map[ptype][leg].extend(tickets)

    if not type_map:
        return "", ""

    type_order = ["Goblin", "Standard", "Demon", "Mix", "Other"]
    btns_html   = ""
    panels_html = ""
    first_type  = True

    for ptype in type_order:
        if ptype not in type_map: continue
        leg_data = type_map[ptype]
        type_id  = f"{id_prefix}-{ptype.lower()}"

        # Inner leg stabs
        inner_btns   = ""
        inner_panels = ""
        first_leg    = True
        for leg in sorted(leg_data.keys(), key=lambda x: int(x) if x.isdigit() else 99):
            lid     = f"{type_id}-{leg}"
            tickets = leg_data[leg]
            active  = "active" if first_leg else ""
            inner_btns   += f'<button class="stab {active}" onclick="switchStab(event,\'{lid}\')">{leg}-Leg <span class="count-badge">{len(tickets)}</span></button>'
            best_summary  = build_best_ticket_summary(tickets)
            inner_panels += f'<div id="{lid}" class="stab-panel {active}">{best_summary}{build_ticket_group(tickets)}</div>'
            first_leg = False

        total = sum(len(v) for v in leg_data.values())
        type_label = {"Goblin":"👺 Goblin","Standard":"⭐ Standard","Demon":"😈 Demon","Mix":"🔀 Mix"}.get(ptype, ptype)
        active_type = "active" if first_type else ""

        btns_html   += f'<button class="top-tab {active_type}" onclick="switchTop(event,\'{type_id}\')">{type_label} <span class="count-badge" style="margin-left:4px">{total}</span></button>'
        panels_html += f"""<div id="{type_id}" class="top-panel {active_type}">
  <div class="stab-bar" style="margin-bottom:16px">{inner_btns}</div>
  {inner_panels}
</div>"""
        first_type = False

    return btns_html, panels_html


# ── KPI cards ─────────────────────────────────────────────────────────────────
def build_kpi(rows: list[dict], total_tickets: int, display_date: str) -> str:
    total_props = len(rows)
    nba    = sum(1 for r in rows if str(r.get("Sport","")).upper()=="NBA")
    cbb    = sum(1 for r in rows if str(r.get("Sport","")).upper()=="CBB")
    nhl    = sum(1 for r in rows if str(r.get("Sport","")).upper()=="NHL")
    soccer = sum(1 for r in rows if str(r.get("Sport","")).upper()=="SOCCER")
    hrs  = []
    for r in rows:
        try:
            f = float(r.get("Hit Rate", 0))
            hrs.append(f if f > 1 else f*100)
        except (TypeError, ValueError):
            pass
    avg_hr = sum(hrs)/len(hrs) if hrs else 0
    tc, _  = rate_color(avg_hr/100)
    m = re.search(r"(\w+)\s+(\d+),\s+(\d+)", display_date)
    mon = m.group(1)[:3] if m else display_date[:6]
    day = m.group(2) if m else ""
    sport_parts = []
    if nba:    sport_parts.append(f"{nba:,} NBA")
    if cbb:    sport_parts.append(f"{cbb:,} CBB")
    if nhl:    sport_parts.append(f"{nhl:,} NHL")
    if soccer: sport_parts.append(f"{soccer:,} SOC")
    sport_sub = " · ".join(sport_parts) if sport_parts else "Multi-sport"

    return f"""<div class="stat-grid stat-grid-4">
  <div class="stat-card green">
    <div class="stat-label">TOTAL PROPS</div>
    <div class="stat-val" style="color:var(--green)">{total_props:,}</div>
    <div class="stat-sub">{sport_sub}</div>
  </div>
  <div class="stat-card blue">
    <div class="stat-label">TOTAL TICKETS</div>
    <div class="stat-val" style="color:var(--cyan)">{total_tickets:,}</div>
    <div class="stat-sub">All types &amp; leg counts</div>
  </div>
  <div class="stat-card amber">
    <div class="stat-label">AVG HIT RATE</div>
    <div class="stat-val" style="color:{tc}">{avg_hr:.1f}%</div>
    <div class="stat-sub">Full slate average</div>
  </div>
  <div class="stat-card purple">
    <div class="stat-label">SLATE DATE</div>
    <div class="stat-val" style="color:#c4b5fd;font-size:28px">{mon} {day}</div>
    <div class="stat-sub">2026</div>
  </div>
</div>"""


# ── CSS ───────────────────────────────────────────────────────────────────────
CSS = """
@import url('https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&display=swap');
:root{
  --bg:#05050f;--bg2:#0d0d1f;--bg3:#111128;--border:#1e1e3a;--bd2:#2a2a4a;
  --text:#e8e8f0;--muted:rgba(255,255,255,0.95);--muted2:rgba(255,255,255,0.88);
  --accent:#c8ff00;--cyan:#00e5ff;
  --green:#39ff6e;--amber:#f0a500;--red:#ff4d4d;--purple:#a78bfa;--blue:#00e5ff;
}
*{box-sizing:border-box;margin:0;padding:0}
html{scroll-behavior:smooth}
body{font-family:'Inter',sans-serif;background:var(--bg);color:var(--text);min-height:100vh;padding-bottom:100px;overflow-x:hidden;}
body::before{content:'';position:fixed;inset:0;background-image:linear-gradient(rgba(200,255,0,.03) 1px,transparent 1px),linear-gradient(90deg,rgba(200,255,0,.03) 1px,transparent 1px);background-size:40px 40px;animation:gridScroll 20s linear infinite;pointer-events:none;z-index:0;}
@keyframes gridScroll{from{background-position:0 0;}to{background-position:0 40px;}}
body::after{content:'';position:fixed;inset:0;background:repeating-linear-gradient(0deg,transparent,transparent 2px,rgba(0,0,0,.15) 2px,rgba(0,0,0,.15) 4px);pointer-events:none;z-index:0;}
::-webkit-scrollbar{width:4px;height:4px}::-webkit-scrollbar-track{background:var(--bg2)}::-webkit-scrollbar-thumb{background:var(--bd2);border-radius:4px}

/* ── UNIFIED NAV ── */
@import url('/static/css/site-nav-unified.css');
@import url('/static/css/nav-mobile-shared.css');
@import url('/static/css/mobile-content-width.css');
@import url('/static/css/site-nav-datetime.css');

/* ── PAGE HEADER ── */
.page-header{position:relative;z-index:1;padding:20px 20px 0;max-width:none;width:100%;margin:0 auto;box-sizing:border-box;}
.page-title{font-family:'Bebas Neue',sans-serif;font-size:clamp(28px,4vw,42px);letter-spacing:.08em;color:var(--accent);line-height:1;}
.page-subtitle{font-size:10px;color:var(--muted);letter-spacing:2.5px;margin-top:4px;}
.page-meta{display:flex;align-items:center;gap:10px;margin-top:10px;flex-wrap:wrap;}
.date-chip{font-size:10px;color:var(--muted);background:var(--bg2);border:1px solid var(--border);border-radius:6px;padding:4px 12px;letter-spacing:1px;}

/* ── SEARCH + FILTER BAR ── */
.filter-bar{position:sticky;top:58px;z-index:150;background:rgba(5,5,15,.95);backdrop-filter:blur(12px);border-bottom:1px solid var(--border);padding:10px 20px;max-width:100%;}
.filter-inner{max-width:none;width:100%;margin:0 auto;display:flex;gap:10px;align-items:center;flex-wrap:wrap;box-sizing:border-box;}
.search-box{display:flex;align-items:center;gap:8px;background:var(--bg2);border:1px solid var(--border);border-radius:8px;padding:7px 12px;flex:1;min-width:200px;max-width:360px;transition:border-color .2s;}
.search-box:focus-within{border-color:var(--accent);}
.search-box input{background:none;border:none;outline:none;color:var(--text);font-family:'Inter',sans-serif;font-size:12px;flex:1;min-width:0;}
.search-box input::placeholder{color:var(--muted2);}
.search-icon{color:var(--muted);font-size:12px;flex-shrink:0;}
.sport-filters{display:flex;gap:6px;flex-wrap:wrap;}
.sf-btn{font-size:10px;padding:5px 12px;border-radius:20px;border:1px solid var(--border);background:var(--bg2);color:var(--muted);cursor:pointer;font-family:'Inter',sans-serif;letter-spacing:1px;transition:all .15s;}
.sf-btn:hover{color:var(--text);border-color:var(--bd2);}
.sf-btn.active{border-color:currentColor;}
.sf-btn.sf-all.active{color:var(--text);border-color:var(--bd2);background:var(--bg3);}
.sf-btn.sf-nba.active{color:var(--accent);border-color:var(--accent);background:rgba(200,255,0,.06);}
.sf-btn.sf-cbb.active{color:var(--cyan);border-color:var(--cyan);background:rgba(0,229,255,.06);}
.sf-btn.sf-nhl.active{color:#5b9cf6;border-color:#5b9cf6;background:rgba(91,156,246,.06);}
.sf-btn.sf-soc.active{color:var(--green);border-color:var(--green);background:rgba(57,255,110,.06);}
.filter-count{font-size:10px;color:var(--muted);white-space:nowrap;margin-left:auto;}

/* ── LAYOUT ── */
.main{position:relative;z-index:1;max-width:none;width:100%;margin:0 auto;padding:20px;box-sizing:border-box;}

/* ── STAT CARDS ── */
.stat-grid{display:grid;gap:12px;margin-bottom:24px;}
.stat-grid-4{grid-template-columns:repeat(4,1fr);}
.stat-card{background:var(--bg2);border:1px solid var(--border);border-radius:12px;padding:14px 16px;position:relative;overflow:hidden;transition:border-color .2s;}
.stat-card:hover{border-color:var(--bd2);}
.stat-card::before{content:'';position:absolute;top:0;left:0;right:0;height:2px;}
.stat-card.green::before{background:linear-gradient(90deg,var(--green),transparent);}
.stat-card.blue::before{background:linear-gradient(90deg,var(--cyan),transparent);}
.stat-card.amber::before{background:linear-gradient(90deg,var(--amber),transparent);}
.stat-card.purple::before{background:linear-gradient(90deg,var(--purple),transparent);}
.stat-label{font-size:9px;color:var(--muted);letter-spacing:2.5px;margin-bottom:6px;}
.stat-val{font-family:'Bebas Neue',sans-serif;font-size:32px;letter-spacing:1px;line-height:1;}
.stat-sub{font-size:11px;color:var(--muted2);margin-top:4px;}

/* ── SPORT SECTION ── */
.sport-section{margin-bottom:32px;}
.sport-section[data-sport]{transition:opacity .2s;}
.sport-section.hidden{display:none;}
.sport-header{display:flex;align-items:center;gap:12px;margin-bottom:16px;cursor:pointer;user-select:none;}
.sport-label{font-family:'Bebas Neue',sans-serif;font-size:26px;letter-spacing:.08em;line-height:1;}
.sport-header-line{flex:1;height:1px;background:linear-gradient(90deg,var(--bd2),transparent);}
.sport-toggle{font-size:12px;color:var(--muted);transition:transform .2s;flex-shrink:0;}
.sport-section.collapsed .sport-toggle{transform:rotate(-90deg);}
.sport-body{overflow:hidden;transition:max-height .3s ease;}
.sport-section.collapsed .sport-body{max-height:0!important;}

/* ── TABS ── */
.top-tabs{display:flex;gap:0;border-bottom:1px solid var(--border);margin-bottom:20px;overflow-x:auto;-webkit-overflow-scrolling:touch;}
.top-tabs::-webkit-scrollbar{height:2px;}
.top-tab{font-size:10px;letter-spacing:1px;padding:8px 14px;cursor:pointer;color:var(--muted);border:none;background:none;border-bottom:2px solid transparent;transition:all .15s;display:flex;align-items:center;gap:5px;font-family:'Inter',sans-serif;white-space:nowrap;flex-shrink:0;}
.top-tab:hover{color:var(--text);}.top-tab.active{color:var(--accent);border-bottom-color:var(--accent);}
.top-panel{display:none;}.top-panel.active{display:block;}
.stab-bar{display:flex;flex-wrap:wrap;gap:5px;margin-bottom:14px;align-items:center;}
.stab{font-size:10px;letter-spacing:1px;padding:4px 12px;cursor:pointer;color:var(--muted2);border:1px solid var(--border);background:var(--bg2);border-radius:20px;transition:all .15s;display:flex;align-items:center;gap:4px;font-family:'Inter',sans-serif;}
.stab:hover{color:var(--text);border-color:var(--bd2);}.stab.active{color:var(--accent);border-color:var(--accent);background:rgba(200,255,0,.06);}
.stab-panel{display:none;}.stab-panel.active{display:block;}
.count-badge{background:var(--bg3);border-radius:10px;padding:1px 5px;font-size:9px;color:var(--muted2);}

/* ── TICKET CARDS ── */
.ticket-card{background:var(--bg2);border:1px solid var(--border);border-radius:12px;margin-bottom:10px;overflow:hidden;transition:transform .15s,box-shadow .15s,border-color .15s;}
.ticket-card:hover{transform:translateY(-2px);box-shadow:0 6px 24px rgba(200,255,0,.07);border-color:var(--bd2);}
.ticket-header{display:flex;align-items:center;gap:10px;padding:10px 14px;background:var(--bg3);border-bottom:1px solid var(--border);border-left:3px solid var(--accent);flex-wrap:wrap;}
.ticket-num{font-family:'Bebas Neue',sans-serif;font-size:18px;color:var(--accent);min-width:28px;}
.ticket-desc{font-size:10px;letter-spacing:1px;color:var(--text);flex:1;min-width:0;}
.ticket-pills{display:flex;gap:4px;flex-wrap:wrap;}
.pill{font-size:9px;padding:2px 8px;border-radius:20px;letter-spacing:.4px;white-space:nowrap;}
.pill-green{background:rgba(57,255,110,.1);color:var(--green);border:1px solid rgba(57,255,110,.2);}
.pill-blue{background:rgba(0,229,255,.1);color:var(--cyan);border:1px solid rgba(0,229,255,.2);}
.pill-amber{background:rgba(240,165,0,.1);color:var(--amber);border:1px solid rgba(240,165,0,.2);}
.pill-purple{background:rgba(167,139,250,.1);color:var(--purple);border:1px solid rgba(167,139,250,.2);}

/* ── TABLES (desktop) ── */
.table-wrap{background:var(--bg2);border:1px solid var(--border);border-radius:12px;overflow:hidden;margin-bottom:14px;}
.ticket-card .table-wrap{border:none;border-radius:0;margin-bottom:0;}
.scrollx{overflow-x:auto;-webkit-overflow-scrolling:touch;}
table{width:100%;border-collapse:collapse;font-size:12px;}
th{font-size:9px;letter-spacing:.08em;color:var(--accent);padding:8px 10px;text-align:left;background:rgba(200,255,0,.03);border-bottom:1px solid var(--border);white-space:nowrap;font-family:'Bebas Neue',sans-serif;}
th.right{text-align:right;}
td{padding:8px 10px;border-bottom:1px solid rgba(255,255,255,.03);vertical-align:middle;}
tr:last-child td{border-bottom:none;}
tr:hover td{background:rgba(200,255,0,.015);}
td.right{text-align:right;}td.mono{font-family:'Inter',sans-serif;font-size:11px;}
td.muted{color:var(--muted2);}td.small{font-size:10px;}

/* ── MOBILE TICKET CARDS (replace table on small screens) ── */
.leg-cards{display:none;padding:10px;}
.leg-card{background:var(--bg3);border:1px solid var(--border);border-radius:10px;padding:12px;margin-bottom:8px;}
.leg-card:last-child{margin-bottom:0;}
.leg-player{font-size:13px;font-weight:700;color:var(--text);margin-bottom:4px;}
.leg-matchup{font-size:10px;color:var(--muted);margin-bottom:8px;}
.leg-row{display:flex;justify-content:space-between;align-items:center;margin-bottom:4px;}
.leg-label{font-size:9px;color:var(--muted);letter-spacing:1.5px;}
.leg-value{font-size:11px;font-family:'Inter',sans-serif;}

/* ── RATE BAR ── */
.rbar{display:flex;align-items:center;gap:6px;min-width:80px;}
.rbar-bg{flex:1;height:4px;background:rgba(255,255,255,.06);border-radius:2px;overflow:hidden;}
.rbar-fill{height:100%;border-radius:2px;}
.rbar-num{font-size:10px;width:32px;text-align:right;flex-shrink:0;}

/* ── CHIPS ── */
.chip{display:inline-block;border-radius:5px;padding:2px 7px;font-size:9px;font-weight:700;letter-spacing:.5px;white-space:nowrap;font-family:'Inter',sans-serif;}
.chip-a{background:rgba(57,255,110,.1);color:var(--green);border:1px solid rgba(57,255,110,.25);}
.chip-b{background:rgba(0,229,255,.1);color:var(--cyan);border:1px solid rgba(0,229,255,.25);}
.chip-c{background:rgba(240,165,0,.1);color:var(--amber);border:1px solid rgba(240,165,0,.25);}
.chip-d{background:rgba(153,153,153,.08);color:rgba(255,255,255,0.92);border:1px solid rgba(153,153,153,.18);}
.chip-goblin{background:rgba(167,139,250,.1);color:var(--purple);border:1px solid rgba(167,139,250,.25);}
.chip-demon{background:rgba(255,77,77,.1);color:#ff8080;border:1px solid rgba(255,77,77,.25);}
.chip-std{background:rgba(0,229,255,.1);color:var(--cyan);border:1px solid rgba(0,229,255,.25);}
.chip-over{background:rgba(57,255,110,.1);color:var(--green);border:1px solid rgba(57,255,110,.25);}
.chip-under{background:rgba(240,165,0,.1);color:var(--amber);border:1px solid rgba(240,165,0,.25);}
/* sport chips */
.chip-nba{background:rgba(200,255,0,.12);color:var(--accent);border:1px solid rgba(200,255,0,.3);}
.chip-cbb{background:rgba(167,139,250,.12);color:var(--purple);border:1px solid rgba(167,139,250,.3);}
.chip-nhl{background:rgba(91,156,246,.12);color:#5b9cf6;border:1px solid rgba(91,156,246,.3);}
.chip-soc{background:rgba(57,255,110,.12);color:var(--green);border:1px solid rgba(57,255,110,.3);}

.pos{color:var(--green);font-weight:700;}.neg{color:var(--red);font-weight:700;}.muted{color:var(--muted2);}
.sub{font-size:10px;color:var(--muted2);margin-top:2px;}
.alert{border-radius:10px;padding:12px 16px;margin-bottom:16px;border:1px solid;font-size:12px;line-height:1.6;}
.alert-amber{background:rgba(240,165,0,.05);border-color:rgba(240,165,0,.2);}
.footer{font-size:9px;color:var(--muted2);text-align:center;margin-top:40px;letter-spacing:1.5px;}

/* ── BACK TO TOP ── */
.back-top{position:fixed;bottom:24px;right:20px;z-index:100;width:40px;height:40px;border-radius:10px;background:var(--bg2);border:1px solid var(--border);color:var(--muted);font-size:16px;cursor:pointer;display:flex;align-items:center;justify-content:center;transition:all .2s;opacity:0;pointer-events:none;}
.back-top.visible{opacity:1;pointer-events:auto;}
.back-top:hover{border-color:var(--accent);color:var(--accent);box-shadow:0 0 16px rgba(200,255,0,.15);}

/* ── RESPONSIVE ── */
@media(max-width:900px){
  .stat-grid-4{grid-template-columns:repeat(2,1fr);}
}
@media(max-width:640px){
  .snav-links{display:none;}
  .live-pill{display:none;}
  .hamburger{display:flex;}
  .snav{padding:0 16px;}
  .snav-name{font-size:18px;}
  .filter-bar{padding:8px 12px;}
  .filter-inner{gap:8px;}
  .search-box{max-width:100%;}
  .main{padding:12px;}
  .stat-grid-4{grid-template-columns:repeat(2,1fr);}
  .stat-val{font-size:26px;}
  /* on mobile: hide tables inside ticket cards, show leg-cards instead */
  .ticket-card .table-wrap{display:none;}
  .leg-cards{display:block;}
  .ticket-header{padding:8px 12px;}
  .ticket-num{font-size:16px;}
  .top-tab{padding:6px 10px;font-size:9px;}
  .stab{padding:3px 10px;font-size:9px;}
  .page-title{font-size:28px;}
}
@media(max-width:400px){
  .stat-grid-4{grid-template-columns:1fr 1fr;}
  .stat-val{font-size:22px;}
}
"""


# ── Full HTML ─────────────────────────────────────────────────────────────────
def build_html(xlsx_path: Path) -> str:
    print(f"  Loading: {xlsx_path.name}")
    wb = openpyxl.load_workbook(xlsx_path, read_only=True, data_only=True)
    print(f"  Sheets: {wb.sheetnames}")

    display_date, iso_date = extract_date(xlsx_path)
    generated = datetime.now().strftime("%Y-%m-%d %H:%M:%S")

    # Main slate rows
    full_slate   = read_flat_sheet(wb["Full Slate"])    if "Full Slate"   in wb.sheetnames else []
    nba_slate    = read_flat_sheet(wb["NBA Slate"])     if "NBA Slate"    in wb.sheetnames else []
    cbb_slate    = read_flat_sheet(wb["CBB Slate"])     if "CBB Slate"    in wb.sheetnames else []
    nhl_slate    = read_flat_sheet(wb["NHL Slate"])     if "NHL Slate"    in wb.sheetnames else []
    soccer_slate = read_flat_sheet(wb["Soccer Slate"])  if "Soccer Slate" in wb.sheetnames else []
    main_slate = full_slate or nba_slate
    print(f"  Slate rows: {len(main_slate)}")

    # Ticket sheets
    nba_sheets    = [sn for sn in wb.sheetnames if re.match(r"NBA (Goblin|Standard|Demon|Mix)", sn)]
    cbb_sheets    = [sn for sn in wb.sheetnames if re.match(r"CBB (Goblin|Standard|Demon|Mix)", sn)]
    nhl_sheets    = [sn for sn in wb.sheetnames if re.match(r"NHL (Goblin|Standard|Demon|Mix)", sn)]
    soccer_sheets = [sn for sn in wb.sheetnames if re.match(r"Soccer (Goblin|Standard|Demon|Mix)", sn)]
    combo_sheets  = [sn for sn in wb.sheetnames if re.match(r"COMBO ", sn)]
    mix_sheets    = [sn for sn in wb.sheetnames if re.match(r"MIX ", sn)]

    # Count all tickets for KPI
    total_tickets = 0
    for sheets in [nba_sheets, cbb_sheets, nhl_sheets, soccer_sheets, combo_sheets, mix_sheets]:
        for sname in sheets:
            if sname in wb.sheetnames:
                total_tickets += len(read_ticket_sheet(wb[sname]))
    print(f"  Total tickets: {total_tickets}")

    kpi = build_kpi(main_slate, total_tickets, display_date)

    # Slate tables
    t_all = build_slate_table(main_slate)
    t_nba = build_slate_table([r for r in main_slate if str(r.get("Sport","")).upper()=="NBA"])
    t_cbb = build_slate_table(cbb_slate or [r for r in main_slate if str(r.get("Sport","")).upper()=="CBB"])

    # Ticket section builder
    def ticket_section(label: str, color: str, sheets: list[str], prefix: str) -> str:
        btns, panels = build_tab_section(wb, sheets, prefix)
        if not btns: return ""
        return f"""<div id="tab-{prefix}" class="top-panel">
  <div class="sport-header" style="margin-top:4px">
    <div class="sport-label" style="color:{color}">{label}</div>
    <div class="sport-header-line"></div>
  </div>
  <div class="top-tabs" style="margin-bottom:20px">{btns}</div>
  {panels}
</div>"""

    nba_section    = ticket_section("NBA TICKETS",          "var(--accent)", nba_sheets,    "nba")
    cbb_section    = ticket_section("CBB TICKETS",          "var(--cyan)",   cbb_sheets,    "cbb")
    nhl_section    = ticket_section("NHL TICKETS",          "#5b9cf6",       nhl_sheets,    "nhl")
    soccer_section = ticket_section("SOCCER TICKETS",       "var(--green)",  soccer_sheets, "soccer")
    combo_section  = ticket_section("COMBO MIX TICKETS",   "var(--purple)", combo_sheets,  "combo")
    mix_section    = ticket_section("CROSS-SPORT TICKETS", "var(--amber)",  mix_sheets,    "mix")

    # count tickets per sport for nav badges
    def count_sheets(sheets_list): return sum(len(read_ticket_sheet(wb[sn])) for sn in sheets_list if sn in wb.sheetnames)
    nba_ct = count_sheets(nba_sheets); cbb_ct = count_sheets(cbb_sheets)
    nhl_ct = count_sheets(nhl_sheets); soc_ct = count_sheets(soccer_sheets)
    t_nba = build_slate_table([r for r in main_slate if str(r.get("Sport","")).upper()=="NBA"])
    t_cbb = build_slate_table(cbb_slate or [r for r in main_slate if str(r.get("Sport","")).upper()=="CBB"])
    t_nhl = build_slate_table(nhl_slate)
    t_soc = build_slate_table(soccer_slate)
    t_all = build_slate_table(main_slate)

    return f"""<!DOCTYPE html>
<html lang="en">
<head>
<meta charset="UTF-8"/>
<meta name="viewport" content="width=device-width, initial-scale=1.0"/>
<title>PropOracle — Tickets {display_date}</title>
<link href="https://fonts.googleapis.com/css2?family=Bebas+Neue&family=Share+Tech+Mono&display=swap" rel="stylesheet"/>
<style>{CSS}</style>
</head>
<body class="mobile-bundle">

{render_static_nav("tickets")}

<div class="page-header">
  <div class="page-title">PropOracle TICKETS</div>
  <div class="page-subtitle">MULTI-SPORT PROP INTELLIGENCE · NBA · CBB · NHL · SOCCER</div>
  <div class="page-meta">
    <span class="date-chip">📅 {display_date}</span>
    <span class="date-chip">⚡ {total_tickets:,} TICKETS</span>
    <span class="date-chip">📋 {len(main_slate):,} PROPS</span>
  </div>
</div>

<div class="filter-bar">
  <div class="filter-inner">
    <div class="search-box">
      <span class="search-icon">🔍</span>
      <input type="text" id="search-input" placeholder="Search player, prop, team..." autocomplete="off"/>
    </div>
    <div class="sport-filters">
      <button class="sf-btn sf-all active" data-sport="all">ALL</button>
      <button class="sf-btn sf-nba" data-sport="nba">🏀 NBA</button>
      <button class="sf-btn sf-cbb" data-sport="cbb">🎓 CBB</button>
      <button class="sf-btn sf-nhl" data-sport="nhl">🏒 NHL</button>
      <button class="sf-btn sf-soc" data-sport="soccer">⚽ SOC</button>
    </div>
    <span class="filter-count" id="filter-count"></span>
  </div>
</div>

<div class="main">

  {kpi}

  <div class="top-tabs" id="main-tabs">
    <button class="top-tab active" onclick="switchTop(event,'tab-slate')">📋 Slate</button>
    <button class="top-tab" onclick="switchTop(event,'tab-nba')" {'style="display:none"' if not nba_ct else ''}>🏀 NBA {f'<span class="count-badge">{nba_ct}</span>' if nba_ct else ''}</button>
    <button class="top-tab" onclick="switchTop(event,'tab-cbb')" {'style="display:none"' if not cbb_ct else ''}>🎓 CBB {f'<span class="count-badge">{cbb_ct}</span>' if cbb_ct else ''}</button>
    <button class="top-tab" onclick="switchTop(event,'tab-nhl')" {'style="display:none"' if not nhl_ct else ''}>🏒 NHL {f'<span class="count-badge">{nhl_ct}</span>' if nhl_ct else ''}</button>
    <button class="top-tab" onclick="switchTop(event,'tab-soccer')" {'style="display:none"' if not soc_ct else ''}>⚽ SOC {f'<span class="count-badge">{soc_ct}</span>' if soc_ct else ''}</button>
    <button class="top-tab" onclick="switchTop(event,'tab-combo')">🔀 Combo</button>
    <button class="top-tab" onclick="switchTop(event,'tab-mix')">🤝 Multi</button>
  </div>

  <div id="tab-slate" class="top-panel active">
    <div class="top-tabs">
      <button class="top-tab active" onclick="switchTop(event,'sl-all')">All ({len(main_slate):,})</button>
      {"<button class=\"top-tab\" onclick=\"switchTop(event,'sl-nba')\">🏀 NBA</button>" if nba_slate or any(str(r.get('Sport','')).upper()=='NBA' for r in main_slate) else ''}
      {"<button class=\"top-tab\" onclick=\"switchTop(event,'sl-cbb')\">🎓 CBB</button>" if cbb_slate else ''}
      {"<button class=\"top-tab\" onclick=\"switchTop(event,'sl-nhl')\">🏒 NHL</button>" if nhl_slate else ''}
      {"<button class=\"top-tab\" onclick=\"switchTop(event,'sl-soc')\">⚽ Soccer</button>" if soccer_slate else ''}
    </div>
    <div id="sl-all" class="top-panel active">{t_all}</div>
    <div id="sl-nba" class="top-panel">{t_nba}</div>
    <div id="sl-cbb" class="top-panel">{t_cbb}</div>
    <div id="sl-nhl" class="top-panel">{t_nhl}</div>
    <div id="sl-soc" class="top-panel">{t_soc}</div>
  </div>

  {nba_section}
  {cbb_section}
  {nhl_section}
  {soccer_section}
  {combo_section}
  {mix_section}

  <div class="footer">PropOracle &nbsp;·&nbsp; GENERATED {generated} &nbsp;·&nbsp; {h(xlsx_path.name)}</div>
</div>

<button class="back-top" id="back-top" onclick="window.scrollTo({{top:0,behavior:'smooth'}})">↑</button>

<script src="/static/js/site-nav-chrome.js"></script>
<script>

// Back to top
const backTop = document.getElementById('back-top');
window.addEventListener('scroll', () => {{
  backTop.classList.toggle('visible', window.scrollY > 400);
}});

// Search + sport filter
const searchInput = document.getElementById('search-input');
const sportBtns = document.querySelectorAll('.sf-btn');
let activeSport = 'all';

function applyFilters() {{
  const q = searchInput.value.toLowerCase().trim();
  let visCount = 0;
  document.querySelectorAll('tr[data-player]').forEach(row => {{
    const player = (row.dataset.player || '').toLowerCase();
    const sport  = (row.dataset.sport  || '').toLowerCase();
    const prop   = (row.dataset.prop   || '').toLowerCase();
    const team   = (row.dataset.team   || '').toLowerCase();
    const matchQ = !q || player.includes(q) || prop.includes(q) || team.includes(q);
    const matchS = activeSport === 'all' || sport === activeSport;
    const show = matchQ && matchS;
    row.style.display = show ? '' : 'none';
    if (show) visCount++;
  }});
  const cnt = document.getElementById('filter-count');
  if (cnt) cnt.textContent = q || activeSport !== 'all' ? visCount + ' results' : '';
}}

sportBtns.forEach(btn => {{
  btn.addEventListener('click', () => {{
    sportBtns.forEach(b => b.classList.remove('active'));
    btn.classList.add('active');
    activeSport = btn.dataset.sport;
    applyFilters();
  }});
}});

searchInput.addEventListener('input', applyFilters);
</script>
</body>
</html>"""


# ── Main ──────────────────────────────────────────────────────────────────────
def main() -> None:
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", type=str)
    parser.add_argument("--date",  type=str)
    parser.add_argument("--out",   type=str)
    args = parser.parse_args()

    if args.input:
        xlsx_path = Path(args.input).resolve()
        if not xlsx_path.exists():
            print(f"ERROR: Not found: {xlsx_path}"); sys.exit(1)
    else:
        xlsx_path = find_latest_tickets(args.date)
        print(f"  Auto-detected: {xlsx_path}")

    html = build_html(xlsx_path)
    out  = Path(args.out).resolve() if args.out else UI_DOCS_DIR / "tickets_latest.html"
    out.parent.mkdir(parents=True, exist_ok=True)
    out.write_text(html, encoding="utf-8")
    print(f"  Saved  -> {out}  ({len(html):,} bytes)")
    print("  Done.")

if __name__ == "__main__":
    main()
