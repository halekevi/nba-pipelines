from __future__ import annotations

# ──────────────────────────────────────────────────────────────────────────────
# Windows UTF-8 fix — MUST be at the very top
# ──────────────────────────────────────────────────────────────────────────────
import csv
import base64
import os
import sys

os.environ.setdefault("PYTHONUTF8", "1")
os.environ.setdefault("PYTHONIOENCODING", "utf-8")

if sys.platform == "win32":
    try:
        sys.stdout.reconfigure(encoding="utf-8", errors="replace")
        sys.stderr.reconfigure(encoding="utf-8", errors="replace")
    except AttributeError:
        pass

import gzip
import hashlib
import io
import json
import logging
import math
import re
import statistics
import sqlite3
import time
import uuid
import threading
import subprocess
import tempfile
import urllib.error
import urllib.request
import zipfile
from urllib.parse import quote
from dataclasses import dataclass, field
from pathlib import Path
from typing import Dict, List, Optional, Any, Sequence
from datetime import date, datetime, timedelta, timezone
from zoneinfo import ZoneInfo

from flask import (
    Flask,
    Response,
    abort,
    current_app,
    jsonify,
    make_response,
    redirect,
    render_template,
    request,
    send_file,
    send_from_directory,
)
from markupsafe import Markup

# ──────────────────────────────────────────────────────────────────────────────
# Paths
# ──────────────────────────────────────────────────────────────────────────────
BASE_DIR      = Path(__file__).resolve().parent.parent  # repo root (one level above ui_runner/)
if str(BASE_DIR) not in sys.path:
    sys.path.insert(0, str(BASE_DIR))  # monorepo bootstrap until `pip install -e .`

from utils.prop_reconcile import reconcile_props_history_dict
from utils.proporacle_data_root import grade_history_read_paths, persistent_data_dir
from scripts.payout_leg_resolver import PayoutLegResolver

UI_DIR        = Path(__file__).resolve().parent         # all UI assets live here (ui_runner/)
_cfg_json_env = (os.environ.get("PROPORACLE_COMMANDS_JSON") or "").strip()
if _cfg_json_env:
    _cfg_p = Path(_cfg_json_env).expanduser()
    CONFIG_PATH = _cfg_p.resolve() if _cfg_p.is_absolute() else (BASE_DIR / _cfg_p).resolve()
else:
    CONFIG_PATH = UI_DIR / "commands.json"
TEMPLATES_DIR = UI_DIR / "templates"
ARCHIVE_DIR   = TEMPLATES_DIR / "archive"
STATIC_DIR    = UI_DIR / "static"
# Bundled graded-prop exports for deploy hosts without data/cache/*_props_history.db (see scripts/export_grades_props_bundle.py).
GRADES_PROPS_EXPORT_DIR = UI_DIR / "data" / "grades_props"


def resolve_repo_root(config: Optional[dict] = None) -> Path:
    """
    Subprocess working-directory root and {REPO_ROOT} token.

    Precedence: PROPORACLE_REPO_ROOT env -> commands.json ``repo_root`` if set ->
    parent of ``ui_runner/`` (``BASE_DIR``). Pipelines should write only under
    that tree (e.g. ``outputs/``), not hard-coded drive letters.
    """
    raw = (os.environ.get("PROPORACLE_REPO_ROOT") or "").strip()
    if raw:
        return Path(raw).expanduser().resolve()
    if config and isinstance(config, dict):
        cfg = (config.get("repo_root") or "").strip()
        if cfg:
            return Path(cfg).expanduser().resolve()
    return BASE_DIR.resolve()


DATA_ROOT = persistent_data_dir(BASE_DIR)


def _sport_dir(sports_subdir: str) -> Path:
    """Prefer Sports/<subdir>/ (canonical pipeline layout); fall back to repo-root folder."""
    under_sports = BASE_DIR / "Sports" / sports_subdir
    legacy = BASE_DIR / sports_subdir
    return under_sports if under_sports.is_dir() else legacy


def _first_existing_file(candidates: list[Path]) -> Path:
    for p in candidates:
        if p.is_file():
            return p
    return candidates[0]


PAYOUT_SAMPLES_DIR = DATA_ROOT / "payout_samples"
PAYOUT_LOG_PATH = PAYOUT_SAMPLES_DIR / "payout_log_hand.csv"
PAYOUT_OBS_PATH = DATA_ROOT / "payout_observations.csv"
PAYOUT_LADDER_LOG_PATH = UI_DIR / "data" / "payout_ladder_log.csv"
PAYOUT_TICKET_LEGS_PATH = UI_DIR / "data" / "payout_ticket_legs.csv"
PAYOUT_LADDER_EXAMPLES_PATH = UI_DIR / "data" / "payout_ladder_examples.json"

# Pipeline output paths (used by status + slate endpoints)
NBA_DIR       = BASE_DIR / "NBA"
CBB_DIR       = BASE_DIR / "CBB"
NHL_DIR       = BASE_DIR / "NHL"
SOCCER_DIR    = _sport_dir("Soccer")
MLB_DIR       = _sport_dir("MLB")
NBA_FLAG      = NBA_DIR / "RUN_COMPLETE.flag"
# ALL sheet includes Standard Line (standard_line) from step8 export — use |Standard Line − Line| for goblin delta / payout ladder keys.
NBA_SLATE     = NBA_DIR / "step8_all_direction_clean.xlsx"
NBA1H_SLATE   = NBA_DIR / "step8_nba1h_direction_clean.xlsx"
NBA1Q_SLATE   = NBA_DIR / "step8_nba1q_direction_clean.xlsx"
NBA_TICKETS   = NBA_DIR / "best_tickets.xlsx"
NBA1H_TICKETS = NBA_DIR / "best_tickets_nba1h.xlsx"
NBA1Q_TICKETS = NBA_DIR / "best_tickets_nba1q.xlsx"
CBB_SLATE     = CBB_DIR / "step6_ranked_cbb.xlsx"
WCBB_SLATE    = CBB_DIR / "step6_ranked_wcbb.xlsx"
CFB_DIR       = _sport_dir("CFB")
CFB_SLATE     = _first_existing_file(
    [
        CFB_DIR / "step6_ranked_cfb.xlsx",
        BASE_DIR / "Sports" / "CFB" / "step6_ranked_cfb.xlsx",
    ]
)
# NHL pipeline writes under NHL/outputs/ (same as run_pipeline.ps1).
NHL_SLATE     = NHL_DIR / "outputs" / "step8_nhl_direction_clean.xlsx"
NHL_TICKETS   = NHL_DIR / "outputs" / "nhl_best_tickets.xlsx"
SOCCER_SLATE = _first_existing_file(
    [
        SOCCER_DIR / "outputs" / "step8_soccer_direction_clean.xlsx",
        SOCCER_DIR / "step8_soccer_direction_clean.xlsx",
    ]
)
SOCCER_TICKETS = SOCCER_DIR / "soccer_best_tickets.xlsx"
MLB_SLATE = _first_existing_file(
    [
        MLB_DIR / "step8_mlb_direction_clean.xlsx",
        MLB_DIR / "outputs" / "step8_mlb_direction_clean.xlsx",
        MLB_DIR / "data" / "outputs" / "step8_mlb_direction_clean.xlsx",
        MLB_DIR / "scripts" / "step8_mlb_direction_clean.xlsx",
    ]
)
MLB_TICKETS = _first_existing_file(
    [
        MLB_DIR / "mlb_best_tickets.xlsx",
        MLB_DIR / "outputs" / "mlb_best_tickets.xlsx",
        MLB_DIR / "scripts" / "mlb_best_tickets.xlsx",
    ]
)
TENNIS_DIR    = BASE_DIR / "Tennis"
# Same pattern as Soccer/MLB: run_daily.ps1 copies outputs → sport root for Railway.
TENNIS_SLATE  = TENNIS_DIR / "step8_tennis_direction_clean.xlsx"
WNBA_DIR = _sport_dir("WNBA")
WNBA_SLATE = _first_existing_file(
    [
        WNBA_DIR / "step8_wnba_direction_clean.xlsx",
        WNBA_DIR / "step8_wnba_direction.xlsx",
        WNBA_DIR / "data" / "outputs" / "step8_wnba_direction_clean.xlsx",
        BASE_DIR / "WNBA" / "data" / "outputs" / "step8_wnba_direction_clean.xlsx",
        BASE_DIR / "WNBA" / "step8_wnba_direction_clean.xlsx",
        BASE_DIR / "WNBA" / "step8_wnba_direction.xlsx",
    ]
)
NFL_DIR       = BASE_DIR / "NFL"
# NFL step8 target: same convention as NHL — sport folder + outputs/ (not repo-root outputs/).
# Pipeline should write: NFL/outputs/step8_nfl_direction_clean.xlsx
NFL_SLATE     = NFL_DIR / "outputs" / "step8_nfl_direction_clean.xlsx"
COMBINED_OUT  = BASE_DIR  # combined_slate_tickets_YYYY-MM-DD.xlsx may live here or under outputs/
OUTPUTS_ROOT  = BASE_DIR / "outputs"
DISABLED_SPORTS: set[str] = set()

# Always expose ui_runner/static — landing pages depend on /static/*.css (blank UI if unset).
try:
    STATIC_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    pass

app = Flask(
    __name__,
    template_folder=str(TEMPLATES_DIR),
    static_folder=str(STATIC_DIR),
)

try:
    DATA_ROOT.mkdir(parents=True, exist_ok=True)
    PAYOUT_SAMPLES_DIR.mkdir(parents=True, exist_ok=True)
except OSError:
    pass

# Optional: flask-compress if installed (requirements.txt does not include it).
try:
    from flask_compress import Compress as _FlaskCompress  # type: ignore[import-not-found]

    _FlaskCompress(app)
    _APP_USES_FLASK_COMPRESS = True
except ImportError:
    _APP_USES_FLASK_COMPRESS = False

# Visible on every response (curl -I); bump when you need to confirm Railway shipped new code.
_UI_BUILD_ID = os.environ.get("RAILWAY_GIT_COMMIT_SHA", "2026-05-15-slate-api")[:12] or "2026-05-15-slate-api"


def _deploy_git_sha_short() -> str:
    return (os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT") or "").strip()[:40]


# Capacitor DIY OTA: zip of repo `mobile/www` (must exist on the deploy host).
_MOBILE_WWW_DIR = BASE_DIR / "mobile" / "www"


def _mobile_www_bundle_fingerprint() -> tuple[str, bool]:
    """
    Short content fingerprint for mobile/www (stable order, skips dotfiles).
    Returns (hex_digest_prefix, dir_ok).
    """
    root = _MOBILE_WWW_DIR
    if not root.is_dir():
        return "", False
    files = [p for p in root.rglob("*") if p.is_file()]
    files.sort(key=lambda p: p.relative_to(root).as_posix())
    h = hashlib.sha256()
    for p in files:
        rel = p.relative_to(root).as_posix()
        if rel.split("/")[0].startswith("."):
            continue
        try:
            st = p.stat()
        except OSError:
            continue
        h.update(rel.encode("utf-8", errors="replace"))
        h.update(str(st.st_size).encode("ascii", errors="ignore"))
        h.update(str(int(st.st_mtime_ns)).encode("ascii", errors="ignore"))
        h.update(b"\n")
    return h.hexdigest()[:32], True


# ── Response compression + static caching ─────────────────────────────────────
_COMPRESSIBLE = ("text/", "application/json", "application/javascript")
_STATIC_EXTS  = (".css", ".js", ".png", ".jpg", ".jpeg", ".webp", ".gif", ".svg", ".ico", ".woff", ".woff2")

@app.after_request
def post_process_response(response):
    response.headers.setdefault("X-PropOracle-Build", _UI_BUILD_ID)
    if request.path.startswith("/api/mobile/bundle"):
        response.headers.setdefault("Access-Control-Allow-Origin", "*")
        response.headers.setdefault("Access-Control-Allow-Methods", "GET, OPTIONS")
        response.headers.setdefault("Access-Control-Allow-Headers", "*")
    # Cache static CSS/JS for 1 hour (images/fonts keep same policy)
    if request.path.startswith("/static/") and any(request.path.endswith(e) for e in _STATIC_EXTS):
        if "Cache-Control" not in response.headers:
            if request.path.endswith((".css", ".js")):
                response.headers["Cache-Control"] = "public, max-age=3600"
            else:
                response.headers["Cache-Control"] = "public, max-age=86400"

    # Gzip compress eligible text responses (skip if flask-compress handles it)
    if (
        _APP_USES_FLASK_COMPRESS
        or response.direct_passthrough
        or response.status_code < 200
        or response.status_code >= 300
        or "Content-Encoding" in response.headers
        or not any(t in response.content_type for t in _COMPRESSIBLE)
        or "gzip" not in request.headers.get("Accept-Encoding", "")
    ):
        return response
    data = response.get_data()
    if len(data) < 500:
        return response
    buf = io.BytesIO()
    with gzip.GzipFile(mode="wb", fileobj=buf, compresslevel=6) as gz:
        gz.write(data)
    compressed = buf.getvalue()
    if len(compressed) >= len(data):
        return response
    response.set_data(compressed)
    response.headers["Content-Encoding"] = "gzip"
    response.headers["Content-Length"]   = len(compressed)
    response.headers["Vary"]             = "Accept-Encoding"
    return response

# ──────────────────────────────────────────────────────────────────────────────
# Job Model
# ──────────────────────────────────────────────────────────────────────────────
@dataclass
class RunJob:
    job_id:      str
    label:       str
    started_at:  float         = field(default_factory=time.time)
    ended_at:    Optional[float] = None
    status:      str           = "RUNNING"   # RUNNING | OK | FAIL
    return_code: Optional[int] = None
    lines:       List[str]     = field(default_factory=list)
    # Step-level progress (populated for chain jobs)
    steps:       List[dict]    = field(default_factory=list)

JOBS: Dict[str, RunJob] = {}
LOCK = threading.Lock()


# ──────────────────────────────────────────────────────────────────────────────
# Helpers
# ──────────────────────────────────────────────────────────────────────────────
_commands_cfg_log = logging.getLogger("PropORACLE.commands_json")
# (cached dict, signature of CONFIG_PATH on disk) — signature change forces reload (add file, edit, mtime change).
_config_bundle: Optional[tuple[dict[str, Any], tuple[Any, ...]]] = None

# When commands.json is absent or invalid, UI/API still respond (pipeline buttons empty until fixed).
_DEFAULT_COMMANDS_CONFIG: dict[str, Any] = {
    "_repo_root_comment": "Add ui_runner/commands.json from the repo; repo_root '' uses auto base dir.",
    "repo_root": "",
    "pipelines": {},
}


def _commands_path_signature() -> tuple[Any, ...]:
    """Identity for cache invalidation: path state + mtime when the file exists."""
    if not CONFIG_PATH.exists():
        return ("missing", str(CONFIG_PATH))
    try:
        return ("ok", CONFIG_PATH.stat().st_mtime, str(CONFIG_PATH))
    except OSError as exc:
        return ("err", str(CONFIG_PATH), str(exc))


def load_config() -> dict:
    global _config_bundle
    sig = _commands_path_signature()
    if _config_bundle is not None:
        data, old_sig = _config_bundle
        if old_sig == sig:
            return data

    if not CONFIG_PATH.exists():
        _commands_cfg_log.warning("commands.json missing at %s — using empty pipeline config", CONFIG_PATH)
        out = dict(_DEFAULT_COMMANDS_CONFIG)
        out["_config_status"] = "missing_file"
        _config_bundle = (out, sig)
        return out
    try:
        raw = CONFIG_PATH.read_text(encoding="utf-8-sig")
        parsed = json.loads(raw)
        if not isinstance(parsed, dict):
            raise ValueError("commands.json must contain a JSON object at the root")
        _config_bundle = (parsed, sig)
        return parsed
    except Exception as exc:
        _commands_cfg_log.exception("commands.json invalid (%s)", CONFIG_PATH)
        out = dict(_DEFAULT_COMMANDS_CONFIG)
        out["_config_status"] = "invalid_json"
        out["_config_error"] = str(exc)
        _config_bundle = (out, sig)
        return out


_json_file_cache: dict[str, dict[str, Any]] = {}
_JSON_FILE_CACHE_LOCK = threading.Lock()
# Strong refs for large template JSON — same objects as read_json_cached entries (TTL still applies).
_LARGE_TEMPLATE_JSON_MEMORY: dict[str, Any] = {}
_LARGE_JSON_NAMES = frozenset({"tickets_latest.json", "slate_latest.json"})

# If set, fetch data from these URLs instead of baked-in files (avoids Docker layer cache).
# Optional on Railway — auto-defaults apply when RAILWAY_* is set (see below).
#   TICKETS_JSON_URL = https://raw.githubusercontent.com/halekevi/PropORACLE/main/ui_runner/templates/tickets_latest.json
#   SLATE_JSON_URL   = https://raw.githubusercontent.com/halekevi/PropORACLE/main/ui_runner/templates/slate_latest.json
#   PIPELINE_JSON_TTL_SEC = 45   # optional; on Railway default is 60 if unset (fresher slates)
#   GITHUB_JSON_FETCH_BUST = 0   # set to 0 to disable ?nocache= on raw GitHub fetches (not recommended)
#   DISABLE_AUTO_GITHUB_JSON = 1  # opt out of Railway → GitHub raw auto URLs
#   PROPORACLE_RAW_JSON_BASE = https://raw.githubusercontent.com/USER/REPO/BRANCH/ui_runner/templates
#   PROPORACLE_SLATE_DATE = 2026-03-28   # optional; prefer outputs/THIS_DATE/ step8_*_THIS_DATE.xlsx
#   SLATE_SPORT_SOURCE = auto|slate_latest|ticket_eval   # auto prefers slate_latest.json (full slate); ticket_eval = eval subset only
#   TICKET_EVAL_SLATE_JSON_URL = https://.../ticket_eval_slate_latest.json
#
# On Railway, baked-in .xlsx mtimes are often old while main-branch JSON is current. We default missing URLs to
# raw GitHub so /api/pipeline/status and slate endpoints stay fresh without manual env setup.
def _running_on_railway() -> bool:
    return bool(
        os.environ.get("RAILWAY_ENVIRONMENT")
        or os.environ.get("RAILWAY_SERVICE_ID")
        or os.environ.get("RAILWAY_PROJECT_ID")
        or os.environ.get("RAILWAY_PUBLIC_DOMAIN")
        or os.environ.get("RAILWAY_STATIC_URL")
    )


_JSON_BASE_DEFAULT = (
    "https://raw.githubusercontent.com/halekevi/PropORACLE/main/ui_runner/templates"
)

_TICKETS_JSON_URL = os.environ.get("TICKETS_JSON_URL", "").strip()
_SLATE_JSON_URL = os.environ.get("SLATE_JSON_URL", "").strip()
_TICKET_EVAL_SLATE_JSON_URL = os.environ.get("TICKET_EVAL_SLATE_JSON_URL", "").strip()
_TICKET_EV_JSON_URL = os.environ.get("TICKET_EV_JSON_URL", "").strip()

if not os.environ.get("DISABLE_AUTO_GITHUB_JSON", "").strip() and _running_on_railway():
    _base = os.environ.get("PROPORACLE_RAW_JSON_BASE", _JSON_BASE_DEFAULT).rstrip("/")
    _root_base = _base
    if _root_base.endswith("/ui_runner/templates"):
        _root_base = _root_base[: -len("/ui_runner/templates")]
    if not _SLATE_JSON_URL:
        _SLATE_JSON_URL = f"{_base}/slate_latest.json"
    if not _TICKETS_JSON_URL:
        _TICKETS_JSON_URL = f"{_base}/tickets_latest.json"
    if not _TICKET_EVAL_SLATE_JSON_URL:
        _TICKET_EVAL_SLATE_JSON_URL = f"{_base}/ticket_eval_slate_latest.json"
    if not _TICKET_EV_JSON_URL:
        _TICKET_EV_JSON_URL = f"{_root_base}/outputs/ticket_ev_latest.json"

_DATA_FILE_URL_MAP: dict[str, str] = {}
if _TICKETS_JSON_URL:
    _DATA_FILE_URL_MAP["tickets_latest.json"] = _TICKETS_JSON_URL
if _SLATE_JSON_URL:
    _DATA_FILE_URL_MAP["slate_latest.json"] = _SLATE_JSON_URL
if _TICKET_EVAL_SLATE_JSON_URL:
    _DATA_FILE_URL_MAP["ticket_eval_slate_latest.json"] = _TICKET_EVAL_SLATE_JSON_URL
if _TICKET_EV_JSON_URL:
    _DATA_FILE_URL_MAP["ticket_ev_latest.json"] = _TICKET_EV_JSON_URL


def _template_json_available(filename: str) -> bool:
    """True if JSON can be loaded from disk or from a configured remote URL (Railway)."""
    return (TEMPLATES_DIR / filename).exists() or bool(_DATA_FILE_URL_MAP.get(filename))


if _running_on_railway() and "PIPELINE_JSON_TTL_SEC" not in os.environ:
    _PIPELINE_JSON_TTL = 60.0
else:
    _PIPELINE_JSON_TTL = float(os.environ.get("PIPELINE_JSON_TTL_SEC", "300"))


def _github_raw_fetch_url(url: str) -> str:
    """
    raw.githubusercontent.com is often cached at the edge; client Cache-Control is ignored.
    Append a unique query param on each HTTP fetch so new pipeline commits are visible quickly.
    """
    if os.environ.get("GITHUB_JSON_FETCH_BUST", "1").strip() in ("0", "false", "no"):
        return url
    sep = "&" if "?" in url else "?"
    return f"{url}{sep}nocache={int(time.time() * 1000)}"


def _template_json_disk_mtime(name: str) -> float | None:
    """Return st_mtime for templates/<name>, or None if missing/unreadable."""
    p = TEMPLATES_DIR / name
    try:
        return p.stat().st_mtime
    except OSError:
        return None


def _explorer_json_gz_bust_token() -> str:
    """Vary gzip response keys when pipeline writes explorer JSON — avoids stale UI after publish."""
    parts: list[str] = []
    for fn in ("slate_latest.json", "ticket_eval_slate_latest.json", "tickets_latest.json"):
        mt = _template_json_disk_mtime(fn)
        parts.append(str(int(mt) if mt is not None else 0))
    return ":".join(parts)


def read_json_cached(path: Path, ttl: float | None = None) -> Any:
    """Load JSON from disk (or remote URL) with an in-process TTL."""
    if ttl is None:
        ttl = _PIPELINE_JSON_TTL
    key = str(path.resolve())
    now = time.time()
    url = _DATA_FILE_URL_MAP.get(path.name)

    disk_mtime: float | None = None
    if not url and path.exists():
        try:
            disk_mtime = path.stat().st_mtime
        except OSError:
            disk_mtime = None

    with _JSON_FILE_CACHE_LOCK:
        entry = _json_file_cache.get(key)
        if entry is not None and now - entry["ts"] <= ttl:
            if url:
                data = entry["data"]
                if path.name in _LARGE_JSON_NAMES:
                    _LARGE_TEMPLATE_JSON_MEMORY[path.name] = data
                return data
            if disk_mtime is not None and entry.get("mtime") == disk_mtime:
                data = entry["data"]
                if path.name in _LARGE_JSON_NAMES:
                    _LARGE_TEMPLATE_JSON_MEMORY[path.name] = data
                return data
            if disk_mtime is None and entry.get("mtime") is None:
                data = entry["data"]
                if path.name in _LARGE_JSON_NAMES:
                    _LARGE_TEMPLATE_JSON_MEMORY[path.name] = data
                return data

        if url:
            try:
                fetch_url = _github_raw_fetch_url(url)
                req = urllib.request.Request(
                    fetch_url,
                    headers={
                        "Cache-Control": "no-cache",
                        "Pragma": "no-cache",
                    },
                )
                with urllib.request.urlopen(req, timeout=25) as resp:
                    data = json.loads(resp.read().decode("utf-8"))
                _json_file_cache[key] = {"data": data, "ts": time.time(), "mtime": None}
                if path.name in _LARGE_JSON_NAMES:
                    _LARGE_TEMPLATE_JSON_MEMORY[path.name] = data
                return data
            except Exception:
                if not path.exists():
                    raise
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        _disk_mt = None
        try:
            _disk_mt = path.stat().st_mtime
        except OSError:
            _disk_mt = None
        _json_file_cache[key] = {"data": data, "ts": time.time(), "mtime": _disk_mt}
        if path.name in _LARGE_JSON_NAMES:
            _LARGE_TEMPLATE_JSON_MEMORY[path.name] = data
        return data


# ── Pre-serialized + pre-gzipped response cache ───────────────────────────────
# Avoids re-serializing and re-compressing large payloads on every request.
_gz_cache: dict[str, tuple[bytes, float]] = {}
_GZ_CACHE_LOCK = threading.Lock()


def _gz_json_response(key: str, build_fn, ttl: float = 300.0):
    """
    Call build_fn() once per TTL, serialize+gzip the result, serve from cache after.
    Handles both gzip-capable and plain clients.

    Build must run under the same lock as the cache check: with gthread, several
    threads could otherwise miss cache together and each load multi‑MB JSON → OOM
    (Railway 502 / worker SIGKILL).
    """
    now = time.time()
    with _GZ_CACHE_LOCK:
        entry = _gz_cache.get(key)
        if entry and now - entry[1] < ttl:
            gz_bytes = entry[0]
        else:
            data = build_fn()
            raw = json.dumps(data, separators=(",", ":")).encode("utf-8")
            buf = io.BytesIO()
            with gzip.GzipFile(mode="wb", fileobj=buf, compresslevel=6) as f:
                f.write(raw)
            gz_bytes = buf.getvalue()
            _gz_cache[key] = (gz_bytes, time.time())

    _JSON_CC = "no-cache, must-revalidate, max-age=0"
    if "gzip" in request.headers.get("Accept-Encoding", ""):
        resp = app.response_class(gz_bytes, status=200, mimetype="application/json")
        resp.headers["Content-Encoding"] = "gzip"
        resp.headers["Content-Length"]   = len(gz_bytes)
        resp.headers["Vary"]             = "Accept-Encoding"
        resp.headers["Cache-Control"]    = _JSON_CC
        resp.headers["Pragma"]           = "no-cache"
        return resp
    # Non-gzip client: decompress inline (rare — all modern browsers support gzip)
    with gzip.GzipFile(fileobj=io.BytesIO(gz_bytes)) as f:
        resp = app.response_class(f.read(), status=200, mimetype="application/json")
        resp.headers["Cache-Control"] = _JSON_CC
        resp.headers["Pragma"]        = "no-cache"
        return resp


def safe_tail(lines: List[str], max_lines: int = 2500) -> List[str]:
    return lines if len(lines) <= max_lines else lines[-max_lines:]


def _build_subprocess_env() -> dict:
    env = os.environ.copy()
    env["PYTHONUTF8"]                    = "1"
    env["PYTHONIOENCODING"]              = "utf-8"
    env["PYTHONLEGACYWINDOWSSTDIO"]      = ""
    env["PYTHONLEGACYWINDOWSFSENCODING"] = ""
    return env


def _maybe_wrap_powershell(cmd: List[str]) -> List[str]:
    if not cmd:
        return cmd
    exe = cmd[0].lower()
    if exe not in ("powershell", "powershell.exe", "pwsh", "pwsh.exe"):
        return cmd
    lower_args = [a.lower() for a in cmd]
    utf8_setup = (
        "[Console]::OutputEncoding = [System.Text.Encoding]::UTF8; "
        "$OutputEncoding = [System.Text.Encoding]::UTF8; "
        "chcp 65001 | Out-Null; "
    )
    if "-command" in lower_args:
        idx = lower_args.index("-command")
        cmd = list(cmd)
        cmd[idx + 1] = utf8_setup + cmd[idx + 1]
        return cmd
    if "-file" in lower_args:
        idx = lower_args.index("-file")
        script_path = cmd[idx + 1]
        extra_args  = cmd[idx + 2:]
        extra_str   = " ".join(f'"{a}"' for a in extra_args) if extra_args else ""
        ps_body = (
            utf8_setup +
            f"& '{script_path}' {extra_str}".strip()
        )
        pre_flags: List[str] = []
        i = 1
        while i < idx:
            pre_flags.append(cmd[i])
            i += 1
        return [cmd[0], "-NoProfile"] + pre_flags + ["-Command", ps_body]
    return cmd


def _auto_wrap_script_if_needed(cmd: List[str], workdir: Path) -> List[str]:
    if not cmd:
        return cmd
    first = cmd[0]
    first_lower = first.lower()
    if first_lower in ("powershell", "powershell.exe", "pwsh", "pwsh.exe"):
        return cmd
    if first_lower.endswith(".ps1"):
        script_path = Path(first)
        if not script_path.is_absolute():
            script_path = (workdir / script_path).resolve()
        try:
            rel = script_path.relative_to(workdir)
            script_for_ps = str(rel)
        except Exception:
            script_for_ps = str(script_path)
        return ["powershell", "-ExecutionPolicy", "Bypass", "-File", script_for_ps] + cmd[1:]
    return cmd


def _run_process(job: RunJob, cmd: List[str], workdir: Path) -> None:
    try:
        if not isinstance(cmd, list) or not all(isinstance(x, str) for x in cmd):
            raise ValueError(f"cmd must be List[str]. Got: {type(cmd)}")
        if not workdir.exists():
            raise FileNotFoundError(f"workdir does not exist: {workdir}")

        env      = _build_subprocess_env()
        cmd2     = _auto_wrap_script_if_needed(cmd, workdir)
        safe_cmd = _maybe_wrap_powershell(cmd2)

        proc = subprocess.Popen(
            safe_cmd,
            cwd=str(workdir),
            stdout=subprocess.PIPE,
            stderr=subprocess.STDOUT,
            text=True,
            encoding="utf-8",
            errors="replace",
            bufsize=1,
            env=env,
        )

        assert proc.stdout is not None
        for raw_line in proc.stdout:
            line = raw_line.rstrip("\r\n").replace("\x00", "")
            with LOCK:
                job.lines.append(line)
                job.lines = safe_tail(job.lines)

        rc = proc.wait()
        with LOCK:
            job.return_code = rc
            job.ended_at    = time.time()
            job.status      = "OK" if rc == 0 else "FAIL"

    except Exception as exc:
        with LOCK:
            job.lines.append(f"[ERROR] {type(exc).__name__}: {exc}")
            job.ended_at    = time.time()
            job.status      = "FAIL"
            job.return_code = -1


def start_job(label: str, cmd: List[str], workdir: Path) -> str:
    job_id = str(uuid.uuid4())
    job    = RunJob(job_id=job_id, label=label)
    with LOCK:
        JOBS[job_id] = job
    threading.Thread(target=_run_process, args=(job, cmd, workdir), daemon=True).start()
    return job_id


def resolve_command(config: dict, pipeline_name: str, command_id: str) -> Dict[str, Any]:
    pipelines = config.get("pipelines") or {}
    if pipeline_name not in pipelines:
        raise KeyError(f"Unknown pipeline '{pipeline_name}'. Available: {list(pipelines.keys())}")

    pipe          = pipelines[pipeline_name]
    commands_list = pipe.get("commands") or []
    cmds          = {c.get("id"): c for c in commands_list if isinstance(c, dict) and c.get("id")}

    if command_id not in cmds:
        raise KeyError(f"Unknown command_id '{command_id}' for pipeline '{pipeline_name}'.")

    c = cmds[command_id]
    if "cmd_chain" in c:
        chain_ids = c.get("cmd_chain") or []
        expanded  = []
        for x in chain_ids:
            if x not in cmds:
                raise KeyError(f"cmd_chain references missing command id '{x}'.")
            expanded.append(cmds[x])
        return {"type": "chain", "items": expanded, "label": c.get("label", command_id)}

    return {"type": "single", "item": c, "label": c.get("label", command_id)}


def subst_tokens(cmd: List[str], config: Optional[dict] = None) -> List[str]:
    today     = datetime.now().strftime("%Y-%m-%d")
    now_ts    = datetime.now().strftime("%Y-%m-%d_%H%M%S")
    repo_root = str(resolve_repo_root(config))
    out: List[str] = []
    for x in cmd:
        y = x.replace("{TODAY}", today).replace("{NOW}", now_ts).replace("{REPO_ROOT}", repo_root)
        out.append(y)
    return out



def _file_info(path: Path) -> dict:
    """Return size + modified time for a file, or None flags if missing."""
    if not path.exists():
        return {"exists": False, "modified": None, "size_kb": None}
    stat = path.stat()
    # Always UTC wall clock so browsers / JS can align with generated_at and US Eastern "slate day".
    return {
        "exists":   True,
        "modified": datetime.fromtimestamp(stat.st_mtime, tz=timezone.utc).strftime("%Y-%m-%d %H:%M:%S"),
        "size_kb":  round(stat.st_size / 1024, 1),
    }


def _mtime_ts(mod_str: str | None) -> float:
    if not mod_str or len(mod_str) < 19:
        return 0.0
    try:
        dt = datetime.strptime(mod_str[:19], "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
        return dt.timestamp()
    except ValueError:
        return 0.0


def _payload_timestamp_meta(payload: dict | None) -> tuple[float | None, str | None]:
    """Parse generated_at / date from slate or tickets JSON (authoritative build time on Railway)."""
    if not isinstance(payload, dict):
        return None, None
    ga = (payload.get("generated_at") or "").strip()
    if ga:
        core = ga.replace(" UTC", "").strip()
        prefix = core[:19].replace("T", " ")
        try:
            dt = datetime.strptime(prefix, "%Y-%m-%d %H:%M:%S").replace(tzinfo=timezone.utc)
            return dt.timestamp(), dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    ds = (payload.get("date") or "").strip()[:10]
    if len(ds) == 10:
        try:
            dt = datetime.strptime(ds, "%Y-%m-%d").replace(tzinfo=timezone.utc)
            return dt.timestamp(), dt.strftime("%Y-%m-%d %H:%M:%S")
        except ValueError:
            pass
    return None, None


def _fresher_meta(
    a: tuple[float | None, str | None],
    b: tuple[float | None, str | None],
) -> tuple[float | None, str | None]:
    """Use whichever payload (slate vs tickets) has the newer timestamp for status UI."""
    at, ad = a
    bt, bd = b
    av = float(at or 0.0)
    bv = float(bt or 0.0)
    if bv > av:
        return bt, bd
    return at, ad


def _slate_day_candidates(preferred_date: str | None) -> list[str]:
    """
    Newest-first YYYY-MM-DD list: optional tickets/slate date, env override, then rolling days.
    Used to find outputs/{date}/step8_*_{date}.xlsx style artifacts.
    """
    seen: set[str] = set()
    out: list[str] = []

    def _add(d: str) -> None:
        d = (d or "").strip()[:10]
        if len(d) == 10 and d not in seen:
            seen.add(d)
            out.append(d)

    _add(os.environ.get("PROPORACLE_SLATE_DATE", "").strip() or "")
    _add(preferred_date or "")
    et_today = datetime.now(ZoneInfo("America/New_York")).date()
    for i in range(14):
        _add((et_today - timedelta(days=i)).strftime("%Y-%m-%d"))
    return out


def _resolve_outputs_artifact(
    days: list[str],
    filename_fmt: str | Sequence[str],
    *legacy: Path,
) -> Path:
    """
    Prefer outputs/{{d}}/filename_fmt.format(d=d), then first existing legacy path.
    filename_fmt example: step8_nba_direction_clean_{d}.xlsx
    When a list is passed, try each pattern per day (first hit wins).
    """
    fmts: list[str]
    if isinstance(filename_fmt, str):
        fmts = [filename_fmt]
    else:
        fmts = list(filename_fmt)
    for d in days:
        for fmt in fmts:
            p = OUTPUTS_ROOT / d / fmt.format(d=d)
            if p.exists():
                return p
    for leg in legacy:
        if leg.exists():
            return leg
    if legacy:
        return legacy[0]
    d0 = days[0] if days else datetime.now(ZoneInfo("America/New_York")).date().strftime("%Y-%m-%d")
    return OUTPUTS_ROOT / d0 / fmts[0].format(d=d0)


def _count_slate_sport_rows(payload: dict) -> int:
    return sum(len(v or []) for v in (payload.get("sports") or {}).values())


def _selected_slate_sport_payload() -> dict:
    """
    Home slate explorer (/api/slate-sport).

    - slate_latest: full per-sport grids from the pipeline (use for explorer).
    - ticket_eval: small subset aligned with ticket HTML (use when forcing eval-only).
    - auto: prefer slate_latest whenever it has rows so every sport panel (NHL, 1H, …) fills;
      ticket_eval alone only lists a few sports/legs and left most panels empty.
    """
    src = os.environ.get("SLATE_SPORT_SOURCE", "auto").strip().lower()
    te_name = "ticket_eval_slate_latest.json"
    sl_name = "slate_latest.json"

    def _load(name: str) -> dict | None:
        if not _template_json_available(name):
            return None
        try:
            return read_json_cached(TEMPLATES_DIR / name)
        except Exception:
            return None

    def _apply_disabled_sports(payload: dict | None) -> dict | None:
        if not isinstance(payload, dict):
            return payload
        sports = payload.get("sports")
        if isinstance(sports, dict):
            payload = dict(payload)
            # Lowercase keys so UI (_slate_counts, index.html SPORTS) and pipeline agree (e.g. nfl not NFL).
            payload["sports"] = {
                str(k).strip().lower(): v
                for k, v in sports.items()
                if str(k).strip().lower() not in DISABLED_SPORTS
            }
        return payload

    if src == "slate_latest":
        d = _load(sl_name)
        if not d:
            raise ValueError("slate_latest.json unavailable")
        return _apply_disabled_sports(d)
    if src == "ticket_eval":
        d = _load(te_name)
        if not d:
            raise ValueError("ticket_eval_slate_latest.json unavailable")
        return _apply_disabled_sports(d)

    sl = _load(sl_name)
    if sl and _count_slate_sport_rows(sl) > 0:
        return _apply_disabled_sports(sl)
    te = _load(te_name)
    if te and _count_slate_sport_rows(te) > 0:
        return _apply_disabled_sports(te)
    if sl:
        return _apply_disabled_sports(sl)
    raise ValueError("no slate json available")


# Home slate explorer table only reads these keys (see templates/index.html renderSlateTable).
_SLATE_SPORT_UI_KEYS = frozenset(
    {
        "tier",
        "rank_score",
        "player",
        "team",
        "opp",
        "prop",
        "pick_type",
        "line",
        "dir",
        "edge",
        "abs_edge",
        "projection",
        "hit_rate",
        "l5_over",
        "l5_under",
        "l10_over",
        "l10_under",
        "season_avg",
        "ml_prob",
        "def_tier",
        "standard_line",
        "standard_projection",
        "opponent_def_rank",
        "image_url",
        "game_time",
        "sport",
    }
)


def _merged_combined_slim_rows(payload: dict) -> list[dict[str, Any]]:
    """All sports from slate_latest (or ticket_eval when selected), Full Slate–style merge for COMBINED."""
    sports = (payload or {}).get("sports") or {}
    out: list[dict[str, Any]] = []
    for k, v in sports.items():
        sk = str(k).strip().lower()
        if sk == "combined":
            continue
        if not isinstance(v, list):
            continue
        for r in v:
            if not isinstance(r, dict):
                continue
            slim = _slim_slate_sport_row(r)
            if "sport" not in slim:
                lab = str(r.get("sport") or "").strip().upper()
                slim["sport"] = lab or sk.upper()
            out.append(slim)

    def _rank(x: dict[str, Any]) -> float:
        v = x.get("rank_score")
        try:
            return float(v) if v is not None else float("-inf")
        except (TypeError, ValueError):
            return float("-inf")

    had_wnba = isinstance(sports.get("wnba"), list) and len(sports["wnba"]) > 0
    if not had_wnba:
        out.extend(_wnba_slate_rows_from_step8_fallback())
    out.sort(key=_rank, reverse=True)
    return _filter_invalid_demon_slate_rows(out)


def _wnba_slate_rows_from_step8_fallback() -> list[dict[str, Any]]:
    """
    slate_latest.json often omits WNBA after a standalone run (combined --write-web not re-run).
    If step8 exists on disk, build the same row shape the explorer expects.
    """
    import importlib.util

    cst_path = BASE_DIR / "scripts" / "combined_slate_tickets.py"
    if not cst_path.is_file():
        return []
    scripts_dir = str(BASE_DIR / "scripts")
    path_inserted = False
    if scripts_dir not in sys.path:
        sys.path.insert(0, scripts_dir)
        path_inserted = True
    try:
        try:
            spec = importlib.util.spec_from_file_location("combined_slate_tickets_wnba_fb", cst_path)
            if spec is None or spec.loader is None:
                return []
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
        except Exception:
            return []
        pref: str | None = None
        try:
            pld = _selected_slate_sport_payload()
            t = str((pld or {}).get("date") or "").strip()[:10]
            if len(t) == 10:
                pref = t
        except Exception:
            pass
        days = _slate_day_candidates(pref)
        xlsx_path = _resolve_outputs_artifact(
            days,
            [
                "wnba/step8_wnba_direction_clean.xlsx",
                "wnba/step8_wnba_direction.xlsx",
                "step8_wnba_direction_clean_{d}.xlsx",
                "step8_wnba_direction_{d}.xlsx",
            ],
            WNBA_SLATE,
            WNBA_DIR / "step8_wnba_direction_clean.xlsx",
            WNBA_DIR / "step8_wnba_direction.xlsx",
        )
        if not xlsx_path.is_file():
            return []
        df = mod.load_wnba(str(xlsx_path))
        raw_rows = mod.dataframe_to_slate_sport_rows(df)
        out_fb: list[dict[str, Any]] = []
        for r in raw_rows:
            if not isinstance(r, dict):
                continue
            rr = dict(r)
            rr["sport"] = "WNBA"
            out_fb.append(_slim_slate_sport_row(rr))
        return out_fb
    finally:
        if path_inserted and sys.path and sys.path[0] == scripts_dir:
            sys.path.pop(0)


def _slim_slate_sport_cell(key: str, v: Any) -> Any:
    """Normalize values for smaller JSON and stable sorting in the client."""
    if v is None:
        return None
    if isinstance(v, str):
        s = v.strip()
        return s if s else None
    if isinstance(v, bool):
        return v
    if isinstance(v, (int, float)):
        if isinstance(v, float) and (math.isnan(v) or math.isinf(v)):
            return None
        if key in ("edge", "rank_score", "abs_edge", "ml_prob"):
            return round(float(v), 4)
        if key == "hit_rate":
            return round(float(v), 6)
        if key == "opponent_def_rank":
            fv = float(v)
            if fv.is_integer():
                return int(fv)
            return round(fv, 4)
        if key in (
            "line",
            "l5_over",
            "l5_under",
            "l10_over",
            "l10_under",
            "projection",
            "standard_line",
            "season_avg",
        ):
            fv = float(v)
            if fv.is_integer():
                return int(fv)
            return round(fv, 3)
        return v
    return v


def _slim_slate_sport_row(r: dict) -> dict:
    """One slate row: only UI keys, omit nulls / blanks after coercion."""
    slim: dict[str, Any] = {}
    for kk in _SLATE_SPORT_UI_KEYS:
        if kk not in r:
            continue
        cv = _slim_slate_sport_cell(kk, r[kk])
        if cv is None:
            continue
        slim[kk] = cv
    return slim


def _filter_invalid_demon_slate_rows(rows: list[Any]) -> list[Any]:
    """Drop Demon rows that cannot be valid PP-style Demons (OVER + positive edge)."""
    if not isinstance(rows, list):
        return rows
    out: list[Any] = []
    for r in rows:
        if not isinstance(r, dict):
            out.append(r)
            continue
        pt = str(r.get("pick_type", "")).strip().lower()
        if pt != "demon":
            out.append(r)
            continue
        dr = str(r.get("dir") or r.get("direction") or "").strip().upper()
        try:
            ef = float(r.get("edge") or 0.0)
        except (TypeError, ValueError):
            ef = 0.0
        if dr != "OVER" or ef <= 0:
            continue
        out.append(r)
    return out


def _api_slate_pick_abs_edge(record: dict[str, Any]) -> float:
    """|edge| for /api/slate picks; prefer pipeline abs_edge when present."""
    ae = record.get("abs_edge")
    if ae is not None:
        try:
            fv = float(ae)
            if not (math.isnan(fv) or math.isinf(fv)):
                return round(fv, 4)
        except (TypeError, ValueError):
            pass
    try:
        ef = float(record.get("edge") or 0.0)
    except (TypeError, ValueError):
        ef = 0.0
    return round(abs(ef), 4)


def _slim_slate_sport_payload(payload: dict) -> dict:
    """Drop unused columns from /api/slate-sport to shrink the gzipped JSON payload."""
    if not isinstance(payload, dict):
        return payload
    sports = payload.get("sports")
    if not isinstance(sports, dict):
        return {
            "date": payload.get("date"),
            "generated_at": payload.get("generated_at"),
            "sports": sports,
        }
    slim_sports: dict[str, Any] = {}
    for k, rows in sports.items():
        if not isinstance(rows, list):
            slim_sports[k] = rows
            continue
        slim_rows: list[Any] = []
        for r in rows:
            if isinstance(r, dict):
                slim_rows.append(_slim_slate_sport_row(r))
            else:
                slim_rows.append(r)
        slim_sports[k] = _filter_invalid_demon_slate_rows(slim_rows)
    return {
        "date": payload.get("date"),
        "generated_at": payload.get("generated_at"),
        "sports": slim_sports,
    }


def _slate_counts() -> tuple[dict[str, int], dict]:
    """
    Return ({sport_key: row_count}, file_info for slate_latest.json on disk, if present).
    Counts follow the same source as /api/slate-sport when possible.
    """
    path = TEMPLATES_DIR / "slate_latest.json"
    disk_info = _file_info(path)
    try:
        payload = _selected_slate_sport_payload()
    except Exception:
        return {}, disk_info
    sports = payload.get("sports") or {}
    counts = {str(k).lower(): len(v or []) for k, v in sports.items()}
    return counts, disk_info


def _sport_slate_status(
    path: Path,
    sport_key: str,
    counts: dict[str, int],
    slate_disk_info: dict,
    json_ts: float | None,
    json_disp: str | None,
) -> dict:
    """
    Status row for Slate Explorer cards.

    json_disp is the unified web timestamp (card_disp: slate / tickets / ticket_eval).

    When this sport has rows in the live slate JSON (same source as the explorer UI), the
    timestamp should follow the merged slate/tickets JSON build time — not the step8 Excel
    mtime. On Railway, Excel shipped under outputs/ or from a later UTC day can be *newer*
    than the JSON pulled from GitHub, so cards showed e.g. 2026-04-04 00:43 while COMBINED
    still showed tickets_latest from 2026-04-03 (looked like a tomorrow slate).

    If the sport is absent from JSON (cnt == 0), fall back to Excel / disk like before.
    """
    key = sport_key.lower()
    cnt = int(counts.get(key, 0))
    if key == "cbb":
        cnt += int(counts.get("wcbb", 0))

    direct = _file_info(path)
    if cnt > 0 and json_disp:
        if direct.get("exists"):
            return {**direct, "modified": json_disp}
        return {
            "exists": True,
            "modified": json_disp,
            "size_kb": slate_disk_info.get("size_kb"),
        }

    if direct.get("exists"):
        if json_ts is not None and json_disp and json_ts > _mtime_ts(direct.get("modified")):
            return {**direct, "modified": json_disp}
        return direct
    if cnt > 0:
        return {
            "exists": True,
            "modified": json_disp or slate_disk_info.get("modified"),
            "size_kb": slate_disk_info.get("size_kb"),
        }
    return direct


# ──────────────────────────────────────────────────────────────────────────────
# Pages
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/")
def home():
    _pipe_cfg = load_config()
    resp = make_response(
        render_template(
            "index.html",
            config=_pipe_cfg,
            ui_build_id=_UI_BUILD_ID,
            deploy_git_sha=(os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT") or "")[:40],
        )
    )
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    return resp


@app.get("/ping")
def ping():
    """Lightweight health check. Use /ping?json=1 for deploy metadata (same as /api/build)."""
    if request.args.get("json") == "1":
        r = jsonify({
            "ok": True,
            "ui_build_id": _UI_BUILD_ID,
            "railway": _running_on_railway(),
            "slate_json_remote": bool(_DATA_FILE_URL_MAP.get("slate_latest.json")),
            "ticket_eval_slate_json_remote": bool(
                _DATA_FILE_URL_MAP.get("ticket_eval_slate_latest.json")
            ),
            "slate_sport_source": os.environ.get("SLATE_SPORT_SOURCE", "auto").strip() or "auto",
            "tickets_json_remote": bool(_DATA_FILE_URL_MAP.get("tickets_latest.json")),
            "deploy_git_sha": (os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT") or "")[:40],
        })
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return r
    return "OK", 200


@app.get("/api/build")
def api_build():
    """Confirm which propOracle UI revision is deployed (Railway / CI debugging)."""
    r = jsonify({
        "ui_build_id": _UI_BUILD_ID,
        "railway": _running_on_railway(),
        "slate_json_remote": bool(_DATA_FILE_URL_MAP.get("slate_latest.json")),
        "ticket_eval_slate_json_remote": bool(
            _DATA_FILE_URL_MAP.get("ticket_eval_slate_latest.json")
        ),
        "slate_sport_source": os.environ.get("SLATE_SPORT_SOURCE", "auto").strip() or "auto",
        "tickets_json_remote": bool(_DATA_FILE_URL_MAP.get("tickets_latest.json")),
        "deploy_git_sha": (os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT") or "")[:40],
    })
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return r


def _no_store_headers(resp):
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    resp.headers["Pragma"] = "no-cache"
    resp.headers["Expires"] = "0"
    return resp


def _tickets_built_slips_missing_html() -> Any:
    """HTML 404 when tickets_latest.json is unavailable (never serve graded tickets_latest.html here)."""
    body = """<!DOCTYPE html>
<html lang="en"><head><meta charset="utf-8"/><meta name="viewport" content="width=device-width,initial-scale=1"/>
<title>PropOracle — Built tickets</title>
<style>
body{font-family:system-ui,-apple-system,sans-serif;background:#050505;color:#e8e8f0;max-width:720px;margin:48px auto;padding:0 24px;line-height:1.55;}
h1{font-size:1.35rem;font-weight:700;margin-bottom:12px;}
p{margin:0 0 14px;}
a{color:#00e5ff;} code{background:#1a1a2e;padding:2px 7px;border-radius:4px;font-size:0.9em;}
</style></head><body>
<h1>Built slips not available</h1>
<p>This page shows <strong>today&rsquo;s generated slips</strong> from <code>tickets_latest.json</code>. The file was not found on disk and no remote JSON URL is configured (on Railway, <code>TICKETS_JSON_URL</code> defaults to raw GitHub when <code>RAILWAY_*</code> env is set).</p>
<p><strong>Graded</strong> results (actuals, hits/misses, ticket KPI bar) are under <a href="/grades">Grades</a> &rarr; Ticket evaluation — not here.</p>
<p>Run the combined slate script with <code>--write-web</code>, commit <code>ui_runner/templates/tickets_latest.json</code>, and redeploy.</p>
<p><a href="/">Home</a></p>
</body></html>"""
    r = make_response(body, 404)
    r.headers["Content-Type"] = "text/html; charset=utf-8"
    return r


@app.get("/tickets_latest.json")
def serve_tickets_latest_json():
    """Expose JSON at site root so the link on /tickets works (relative URLs break under /tickets)."""
    path = TEMPLATES_DIR / "tickets_latest.json"
    if not _template_json_available("tickets_latest.json"):
        abort(404)

    try:
        return _gz_json_response(
            "tickets-latest-json",
            lambda: read_json_cached(path),
            ttl=_PIPELINE_JSON_TTL,
        )
    except Exception:
        abort(404)


# ── Uniform-bucket tickets (built by scripts/build_uniform_tickets_artifacts.py) ─

_UNIFORM_DATE_RE = re.compile(r"^\d{4}-\d{2}-\d{2}$")


def _uniform_dates_from_disk() -> list[str]:
    out: list[str] = []
    for p in TEMPLATES_DIR.glob("uniform_tickets_*.json"):
        m = re.fullmatch(r"uniform_tickets_(\d{4}-\d{2}-\d{2})\.json", p.name)
        if m:
            out.append(m.group(1))
    return sorted(out, reverse=True)


@app.get("/api/uniform-tickets/dates")
def api_uniform_tickets_dates():
    """Return the list of slate dates for which uniform-bucket ticket JSON was published."""
    path = TEMPLATES_DIR / "uniform_tickets_dates.json"
    if path.exists():
        try:
            return _gz_json_response(
                "uniform-tickets-dates",
                lambda: read_json_cached(path),
                ttl=_PIPELINE_JSON_TTL,
            )
        except Exception:
            pass
    return jsonify({"dates": _uniform_dates_from_disk(), "source": "disk_scan"})


@app.get("/api/uniform-tickets/backtest")
def api_uniform_tickets_backtest():
    """Return the rolling per-(size, bucket) backtest summary."""
    path = TEMPLATES_DIR / "uniform_tickets_backtest.json"
    if not path.exists():
        return jsonify({"rows": []})
    try:
        return _gz_json_response(
            "uniform-tickets-backtest",
            lambda: read_json_cached(path),
            ttl=_PIPELINE_JSON_TTL,
        )
    except Exception:
        return jsonify({"rows": []})


@app.get("/api/uniform-tickets/latest")
def api_uniform_tickets_latest():
    path = TEMPLATES_DIR / "uniform_tickets_latest.json"
    if not path.exists():
        dates = _uniform_dates_from_disk()
        if not dates:
            abort(404)
        path = TEMPLATES_DIR / f"uniform_tickets_{dates[0]}.json"
        if not path.exists():
            abort(404)
    try:
        return _gz_json_response(
            "uniform-tickets-latest",
            lambda: read_json_cached(path),
            ttl=_PIPELINE_JSON_TTL,
        )
    except Exception:
        abort(404)


@app.get("/api/uniform-tickets/<date_str>")
def api_uniform_tickets_for_date(date_str: str):
    if not _UNIFORM_DATE_RE.fullmatch(date_str):
        abort(404)
    path = TEMPLATES_DIR / f"uniform_tickets_{date_str}.json"
    if not path.exists():
        abort(404)
    try:
        return _gz_json_response(
            f"uniform-tickets-{date_str}",
            lambda: read_json_cached(path),
            ttl=_PIPELINE_JSON_TTL,
        )
    except Exception:
        abort(404)


@app.get("/tickets")
def page_tickets():
    """
    Today's built ticket slips from tickets_latest.json (combined_slate_tickets --write-web).

    Graded legs, actuals, and hit/miss summaries live under Grades (/grades hub, or
    /grades/YYYY-MM-DD for ticket_eval_*.html from build_ticket_eval.py), not on this route.
    """
    import importlib.util

    json_path = TEMPLATES_DIR / "tickets_latest.json"
    has_json = _template_json_available("tickets_latest.json")

    def _render_slips_from_json() -> str | None:
        if not has_json:
            return None
        payload = read_json_cached(json_path)
        cst_path = BASE_DIR / "scripts" / "combined_slate_tickets.py"
        if not cst_path.exists():
            raise FileNotFoundError("scripts/combined_slate_tickets.py not in repo")
        scripts_dir = str(BASE_DIR / "scripts")
        path_inserted = False
        if scripts_dir not in sys.path:
            sys.path.insert(0, scripts_dir)
            path_inserted = True
        try:
            spec = importlib.util.spec_from_file_location("combined_slate_tickets", cst_path)
            if spec is None or spec.loader is None:
                raise RuntimeError("could not load combined_slate_tickets spec")
            mod = importlib.util.module_from_spec(spec)
            spec.loader.exec_module(mod)
            # tickets_latest.json is authoritative (EV gate applied only when combined is built with strict web mode).
            body, page_title = mod.render_tickets_body_html(
                payload,
                _non_ev_slips_removed=0,
            )
            return render_template(
                "tickets_built.html",
                tickets_body=Markup(body),
                page_title=page_title,
                ui_build_id=_UI_BUILD_ID,
                deploy_git_sha=(
                    os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT") or ""
                )[:40],
            )
        finally:
            if path_inserted and sys.path and sys.path[0] == scripts_dir:
                sys.path.pop(0)

    try:
        html = _render_slips_from_json()
        if html:
            return _no_store_headers(make_response(html))
    except Exception as e:
        current_app.logger.warning("/tickets: render from tickets_latest.json failed: %s", e)
        if has_json:
            return _no_store_headers(
                make_response(
                    (
                        "Could not render built slips from tickets_latest.json (see server log). "
                        f"Error: {e!s}"
                    ),
                    500,
                )
            )

    if not has_json:
        return _no_store_headers(_tickets_built_slips_missing_html())
    return _no_store_headers(
        make_response(
            "Could not render /tickets from tickets_latest.json. Check combined_slate_tickets.render_tickets_body_html "
            "and JSON shape.",
            500,
        )
    )


@app.get("/payout")
def page_payout():
    r = make_response(render_template("payout_calculator.html"))
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    r.headers["Expires"] = "0"
    return r


_PAYOUT_BASE_POWER = {2: 3.0, 3: 6.0, 4: 10.0, 5: 20.0, 6: 37.5}
_PAYOUT_LADDER_FIELDS = [
    "date",
    "n_legs",
    "leg_composition",
    "goblin_deltas",
    "demon_deltas",
    "power_payout_x",
    "flex_payout_x",
    "source",
    "notes",
]

_PAYOUT_TICKET_LEG_FIELDS = [
    "date",
    "ticket_id",
    "slip_type",
    "stake",
    "payout_to_win",
    "power_payout_x",
    "sport",
    "leg_slot",
    "player",
    "prop",
    "line",
    "direction",
    "pick_type",
    "source",
    "notes",
    "delta_quality",
    "matched_snapshot_path",
    "matched_standard_line",
    "delta_method",
    "delta",
]

_PAYOUT_LEG_RESOLVER: PayoutLegResolver | None = None


def _get_payout_leg_resolver() -> PayoutLegResolver:
    global _PAYOUT_LEG_RESOLVER
    if _PAYOUT_LEG_RESOLVER is None:
        _PAYOUT_LEG_RESOLVER = PayoutLegResolver(BASE_DIR)
    return _PAYOUT_LEG_RESOLVER


def _append_csv_with_schema_upgrade(path: Path, fieldnames: list[str], row: dict[str, Any]) -> None:
    path.parent.mkdir(parents=True, exist_ok=True)
    if not path.exists():
        with path.open("w", newline="", encoding="utf-8") as f:
            w = csv.DictWriter(f, fieldnames=fieldnames)
            w.writeheader()
            w.writerow({k: row.get(k, "") for k in fieldnames})
        return

    existing_rows: list[dict[str, Any]] = []
    existing_fields: list[str] = []
    with path.open("r", newline="", encoding="utf-8") as f:
        rdr = csv.DictReader(f)
        existing_fields = list(rdr.fieldnames or [])
        for r in rdr:
            existing_rows.append(dict(r))
    merged_fields = list(dict.fromkeys(existing_fields + fieldnames))
    existing_rows.append({k: row.get(k, "") for k in merged_fields})
    with path.open("w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=merged_fields)
        w.writeheader()
        for r in existing_rows:
            w.writerow({k: r.get(k, "") for k in merged_fields})


def _safe_float(v: Any, default: float = 0.0) -> float:
    try:
        f = float(v)
        return f if math.isfinite(f) else default
    except (TypeError, ValueError):
        return default


def _pick_projection_from_mapping(row: Dict[str, Any]) -> Optional[float]:
    """Model projection for UI charts (slate rows + ticket legs)."""
    for key in ("projection", "standard_projection", "Projection"):
        v = row.get(key)
        if v is None or v == "":
            continue
        try:
            f = float(v)
            if math.isfinite(f):
                return f
        except (TypeError, ValueError):
            continue
    return None


def _normalize_leg_pick_type(raw: Any) -> str:
    s = str(raw or "").strip().lower()
    if "gob" in s:
        return "Goblin"
    if "dem" in s:
        return "Demon"
    return "Standard"


def _ticket_prob_from_leg_hit(leg_hit_rate: float, n_legs: int) -> float:
    p = max(0.0, min(1.0, float(leg_hit_rate)))
    n = max(0, int(n_legs))
    return p**n


def predict_payout(legs: list[dict[str, Any]]) -> dict[str, Any]:
    """
    Estimate payout multiplier from leg composition (Standard / Goblin / Demon).
    """
    n = len(legs)
    if n < 2 or n > 6:
        raise ValueError("Ticket Composer supports 2 to 6 legs.")
    base_mult = float(_PAYOUT_BASE_POWER.get(n, 0.0))
    if base_mult <= 0:
        raise ValueError("Unsupported leg count for PrizePicks power baseline.")

    leg_breakdown: list[dict[str, Any]] = []
    composite_factor = 1.0
    for i, leg in enumerate(legs, start=1):
        pt = _normalize_leg_pick_type(leg.get("pick_type"))
        delta = abs(_safe_float(leg.get("delta"), 0.0))
        direction = str(leg.get("direction") or "OVER").strip().upper()
        if direction not in {"OVER", "UNDER"}:
            direction = "OVER"
        if pt == "Goblin":
            factor = max(0.2, 1.0 - 0.110 * delta)
        elif pt == "Demon":
            factor = max(1.0, 1.0 + 0.1782 * (delta ** 1.287))
        else:
            factor = 1.0
        composite_factor *= factor
        leg_breakdown.append(
            {
                "leg": i,
                "pick_type": pt,
                "direction": direction,
                "delta": round(delta, 4),
                "factor": round(factor, 4),
            }
        )

    estimated_multiplier = round(base_mult * composite_factor, 4)

    def _ev_for_leg_hit(p_leg: float) -> float:
        p_ticket = _ticket_prob_from_leg_hit(p_leg, n)
        # EV per $1 stake.
        return (p_ticket * (estimated_multiplier - 1.0)) - (1.0 - p_ticket)

    return {
        "estimated_multiplier": estimated_multiplier,
        "ev_at_65pct": round(_ev_for_leg_hit(0.65), 4),
        "ev_at_75pct": round(_ev_for_leg_hit(0.75), 4),
        "ev_at_85pct": round(_ev_for_leg_hit(0.85), 4),
        "leg_breakdown": leg_breakdown,
        "base_multiplier_standard": base_mult,
        "n_legs": n,
    }


def _extract_json_object(text: str) -> dict[str, Any]:
    if not text:
        raise ValueError("Gemini returned empty text.")
    s = text.strip()
    try:
        obj = json.loads(s)
        if isinstance(obj, dict):
            return obj
    except json.JSONDecodeError:
        pass
    m = re.search(r"\{[\s\S]*\}", s)
    if not m:
        raise ValueError("Could not find JSON object in Gemini response.")
    obj = json.loads(m.group(0))
    if not isinstance(obj, dict):
        raise ValueError("Gemini JSON payload is not an object.")
    return obj


def _has_missing_tier_delta(legs: list[dict[str, Any]]) -> bool:
    for leg in legs:
        pt = _normalize_leg_pick_type(leg.get("pick_type"))
        if pt not in {"Goblin", "Demon"}:
            continue
        v = leg.get("delta")
        if v is None:
            return True
        try:
            if float(v) <= 0:
                return True
        except (TypeError, ValueError):
            return True
    return False


def _norm_lookup_token(raw: Any) -> str:
    s = str(raw or "").strip().lower()
    s = re.sub(r"[^a-z0-9]+", " ", s)
    return re.sub(r"\s+", " ", s).strip()


def _safe_num_or_none(raw: Any) -> float | None:
    try:
        v = float(raw)
    except (TypeError, ValueError):
        return None
    if not math.isfinite(v):
        return None
    return v


def _lookup_standard_line_from_slate(
    leg: dict[str, Any],
    by_triplet: dict[tuple[str, str, str], float],
    by_pair: dict[tuple[str, str], float],
) -> float | None:
    player = _norm_lookup_token(leg.get("player"))
    prop = _norm_lookup_token(leg.get("prop_type") or leg.get("prop"))
    if not player or not prop:
        return None
    direction = str(leg.get("direction") or "").strip().upper()
    if direction in {"OVER", "UNDER"}:
        k3 = (player, prop, direction)
        if k3 in by_triplet:
            return by_triplet[k3]
    return by_pair.get((player, prop))


def _fill_missing_tier_deltas_from_slate(legs: list[dict[str, Any]]) -> None:
    if not _has_missing_tier_delta(legs):
        return
    try:
        payload = read_json_cached(TEMPLATES_DIR / "slate_latest.json")
    except Exception:
        return
    sports = (payload or {}).get("sports") or {}
    if not isinstance(sports, dict):
        return

    by_triplet: dict[tuple[str, str, str], float] = {}
    by_pair: dict[tuple[str, str], float] = {}
    for rows in sports.values():
        if not isinstance(rows, list):
            continue
        for row in rows:
            if not isinstance(row, dict):
                continue
            if _normalize_leg_pick_type(row.get("pick_type")) != "Standard":
                continue
            player = _norm_lookup_token(row.get("player"))
            prop = _norm_lookup_token(row.get("prop") or row.get("prop_type"))
            direction = str(row.get("dir") or row.get("direction") or "").strip().upper()
            line_val = _safe_num_or_none(row.get("line"))
            if not player or not prop or line_val is None:
                continue
            if direction in {"OVER", "UNDER"}:
                by_triplet.setdefault((player, prop, direction), line_val)
            by_pair.setdefault((player, prop), line_val)

    if not by_pair:
        return

    for leg in legs:
        pt = _normalize_leg_pick_type(leg.get("pick_type"))
        if pt not in {"Goblin", "Demon"}:
            continue
        cur_delta = _safe_num_or_none(leg.get("delta"))
        if cur_delta is not None and cur_delta > 0:
            continue
        played_line = _safe_num_or_none(leg.get("line"))
        if played_line is None:
            continue
        std_line = _lookup_standard_line_from_slate(leg, by_triplet, by_pair)
        if std_line is None:
            continue
        d = abs(std_line - played_line)
        if d > 0:
            leg["delta"] = round(d, 3)


_HIST_STD_LINE_CACHE: dict[str, Any] = {"built_at": 0.0, "triplet": {}, "pair": {}}


def _recent_combined_slate_paths(max_days: int = 10, max_files: int = 40) -> list[Path]:
    out_root = BASE_DIR / "outputs"
    if not out_root.is_dir():
        return []
    dated_dirs: list[tuple[date, Path]] = []
    for d in out_root.iterdir():
        if not d.is_dir():
            continue
        try:
            dd = datetime.strptime(d.name, "%Y-%m-%d").date()
        except ValueError:
            continue
        if (date.today() - dd).days < 0 or (date.today() - dd).days > max_days:
            continue
        dated_dirs.append((dd, d))
    dated_dirs.sort(key=lambda x: x[0], reverse=True)
    files: list[Path] = []
    seen: set[Path] = set()
    for _, folder in dated_dirs:
        for p in sorted(folder.glob("combined_slate_tickets_*.xlsx"), reverse=True):
            if p in seen:
                continue
            seen.add(p)
            files.append(p)
            if len(files) >= max_files:
                return files
    return files


def _historical_standard_line_index(ttl_s: float = 600.0) -> tuple[dict[tuple[str, str, str], float], dict[tuple[str, str], float]]:
    now = time.time()
    built_at = float(_HIST_STD_LINE_CACHE.get("built_at") or 0.0)
    if now - built_at < ttl_s:
        t = _HIST_STD_LINE_CACHE.get("triplet")
        p = _HIST_STD_LINE_CACHE.get("pair")
        if isinstance(t, dict) and isinstance(p, dict):
            return t, p

    trip_vals: dict[tuple[str, str, str], list[float]] = {}
    pair_vals: dict[tuple[str, str], list[float]] = {}
    for path in _recent_combined_slate_paths():
        try:
            import openpyxl  # type: ignore

            wb = openpyxl.load_workbook(str(path), data_only=True, read_only=True)
            if "Full Slate" not in wb.sheetnames:
                wb.close()
                continue
            ws = wb["Full Slate"]
            it = ws.iter_rows(values_only=True)
            hdr = [str(x).strip() if x is not None else "" for x in next(it)]
            idx = {h: i for i, h in enumerate(hdr)}
            if "Player" not in idx or "Prop" not in idx or "Pick Type" not in idx:
                wb.close()
                continue
            std_col = "Standard Line" if "Standard Line" in idx else "Line"
            if std_col not in idx:
                wb.close()
                continue
            for row in it:
                if _normalize_leg_pick_type(row[idx["Pick Type"]]) != "Standard":
                    continue
                player = _norm_lookup_token(row[idx["Player"]])
                prop = _norm_lookup_token(row[idx["Prop"]])
                direction = _norm_lookup_token(row[idx["Dir"]]).upper() if "Dir" in idx else ""
                std_line = _safe_num_or_none(row[idx[std_col]])
                if not player or not prop or std_line is None:
                    continue
                pair_vals.setdefault((player, prop), []).append(std_line)
                if direction in {"OVER", "UNDER"}:
                    trip_vals.setdefault((player, prop, direction), []).append(std_line)
            wb.close()
        except Exception:
            continue

    trip: dict[tuple[str, str, str], float] = {}
    pair: dict[tuple[str, str], float] = {}
    for k, vals in trip_vals.items():
        if vals:
            trip[k] = float(statistics.median(vals))
    for k, vals in pair_vals.items():
        if vals:
            pair[k] = float(statistics.median(vals))
    _HIST_STD_LINE_CACHE["built_at"] = now
    _HIST_STD_LINE_CACHE["triplet"] = trip
    _HIST_STD_LINE_CACHE["pair"] = pair
    return trip, pair


def _fill_missing_tier_deltas_from_history(legs: list[dict[str, Any]]) -> None:
    if not _has_missing_tier_delta(legs):
        return
    by_triplet, by_pair = _historical_standard_line_index()
    if not by_triplet and not by_pair:
        return
    for leg in legs:
        pt = _normalize_leg_pick_type(leg.get("pick_type"))
        if pt not in {"Goblin", "Demon"}:
            continue
        cur_delta = _safe_num_or_none(leg.get("delta"))
        if cur_delta is not None and cur_delta > 0:
            continue
        played_line = _safe_num_or_none(leg.get("line"))
        if played_line is None:
            continue
        std_line = _lookup_standard_line_from_slate(leg, by_triplet, by_pair)
        if std_line is None:
            continue
        d = abs(std_line - played_line)
        if d > 0:
            leg["delta"] = round(d, 3)


def _fill_missing_tier_deltas(
    model: Any,
    image_bytes: bytes,
    mime_type: str,
    legs: list[dict[str, Any]],
) -> None:
    if not _has_missing_tier_delta(legs):
        return
    legs_min = []
    for leg in legs:
        legs_min.append(
            {
                "player": str(leg.get("player") or "").strip(),
                "prop_type": str(leg.get("prop_type") or "").strip(),
                "line": leg.get("line"),
                "pick_type": str(leg.get("pick_type") or "").strip(),
                "direction": str(leg.get("direction") or "").strip(),
                "delta": leg.get("delta"),
            }
        )
    prompt = (
        "You are fixing missing PrizePicks tier deltas from a screenshot.\n"
        "Return ONLY valid JSON with this exact schema:\n"
        "{\"legs\":[{\"delta\":number|null,\"played_line\":number|null,\"standard_line\":number|null}]}\n"
        "One output leg per input leg, same order, no extra keys.\n"
        "Rules:\n"
        "- Standard legs must return null.\n"
        "- Goblin or Demon legs should return a positive numeric delta whenever readable.\n"
        "- Use visible line adjustment cues on the slip (line pairs, easier/harder text, badge values).\n"
        "- If you can read both lines, populate played_line and standard_line even if delta is null.\n"
        "- If truly unreadable for that leg, return null.\n"
        f"Input legs: {json.dumps(legs_min, ensure_ascii=True)}"
    )
    img_part = {"mime_type": mime_type or "image/png", "data": base64.b64encode(image_bytes).decode("utf-8")}
    try:
        resp = model.generate_content([prompt, img_part])
        txt = getattr(resp, "text", "") or ""
        obj = _extract_json_object(txt)
        out_legs = obj.get("legs")
        if not isinstance(out_legs, list):
            return
        for idx, leg in enumerate(legs):
            if idx >= len(out_legs):
                break
            row = out_legs[idx]
            if not isinstance(row, dict):
                continue
            pt = _normalize_leg_pick_type(leg.get("pick_type"))
            if pt not in {"Goblin", "Demon"}:
                leg["delta"] = None
                continue
            cur = leg.get("delta")
            curf = _safe_float(cur, 0.0)
            if cur is not None and curf > 0:
                continue
            val = row.get("delta")
            new_delta = abs(_safe_float(val, 0.0))
            if new_delta <= 0:
                played_line = _safe_num_or_none(row.get("played_line"))
                standard_line = _safe_num_or_none(row.get("standard_line"))
                if played_line is not None and standard_line is not None:
                    new_delta = abs(standard_line - played_line)
            if new_delta > 0:
                leg["delta"] = round(new_delta, 3)
    except Exception:
        return


def _gemini_extract_ticket_from_image(image_bytes: bytes, mime_type: str) -> dict[str, Any]:
    api_key = (
        (os.environ.get("GEMINI_API_KEY") or "").strip()
        or (os.environ.get("GOOGLE_API_KEY") or "").strip()
        or (os.environ.get("GOOGLE_GENERATIVE_AI_API_KEY") or "").strip()
    )
    if not api_key:
        raise RuntimeError(
            "Missing Gemini API key. Set GEMINI_API_KEY (preferred), GOOGLE_API_KEY, or "
            "GOOGLE_GENERATIVE_AI_API_KEY on the server."
        )
    try:
        import google.generativeai as genai  # type: ignore
    except Exception as exc:
        raise RuntimeError("google-generativeai is not installed in this environment.") from exc

    prompt = (
        "Extract PrizePicks ticket data. Return ONLY valid JSON (no markdown, no code fences).\n"
        "Schema:\n"
        "{"
        "\"n_legs\": int|null,"
        "\"legs\": ["
        "{\"player\":str,\"prop_type\":str,\"line\":number|null,\"pick_type\":\"Standard|Goblin|Demon|Unknown\",\"direction\":\"OVER|UNDER|UNKNOWN\",\"delta\":number|null}"
        "],"
        "\"power_payout_x\": number|null,"
        "\"flex_payout_x\": number|null,"
        "\"entry_amount\": number|null,"
        "\"notes\": str"
        "}\n"
        "General: If a scalar field is truly unreadable, use null (use \"Unknown\" only for pick_type/direction when needed).\n"
        "CRITICAL — legs[].delta (PrizePicks tier lines):\n"
        "- Standard legs: set \"delta\" to null (Standard does not use delta).\n"
        "- Goblin or Demon legs: set \"delta\" to a positive NUMBER (half-points allowed, e.g. 0.5, 1, 1.5, 2) = "
        "the absolute line movement in POINTS vs the Standard line for that same player+prop.\n"
        "  If the slip shows BOTH a standard/reference line and the played Goblin/Demon line, compute "
        "delta = round(abs(standard_line - played_line), 2) using the numbers on screen (same units as Points/Rebounds/etc.).\n"
        "  If only one line is visible but the UI states an adjustment (badge text, 'pts easier/harder', strikethrough pair, etc.), "
        "use that magnitude as delta.\n"
        "  Every Goblin/Demon leg must have its own delta — e.g. two Goblin legs ⇒ two separate numeric \"delta\" values, "
        "not null, whenever the screenshot supports it. Only use null for a Goblin/Demon leg if no adjustment magnitude can be determined at all.\n"
    )
    genai.configure(api_key=api_key)
    # gemini-1.5-flash was removed from the v1beta API (404). Use current stable Flash; override via GEMINI_VISION_MODEL if needed.
    model_name = (os.environ.get("GEMINI_VISION_MODEL") or "gemini-2.5-flash").strip()
    model = genai.GenerativeModel(model_name)
    img_part = {"mime_type": mime_type or "image/png", "data": base64.b64encode(image_bytes).decode("utf-8")}
    resp = model.generate_content([prompt, img_part])
    txt = getattr(resp, "text", "") or ""
    extracted = _extract_json_object(txt)
    legs = extracted.get("legs")
    if isinstance(legs, list):
        leg_dicts = [x for x in legs if isinstance(x, dict)]
        _fill_missing_tier_deltas(model, image_bytes, mime_type, leg_dicts)
        _fill_missing_tier_deltas_from_slate(leg_dicts)
        _fill_missing_tier_deltas_from_history(leg_dicts)
        extracted["legs"] = leg_dicts
    return extracted


def _composition_label_from_legs(legs: list[dict[str, Any]]) -> str:
    s = g = d = 0
    for leg in legs:
        pt = _normalize_leg_pick_type(leg.get("pick_type"))
        if pt == "Goblin":
            g += 1
        elif pt == "Demon":
            d += 1
        else:
            s += 1
    return f"{s}S+{g}G+{d}D"


def _read_payout_ladder_rows() -> list[dict[str, Any]]:
    p = PAYOUT_LADDER_LOG_PATH
    if not p.is_file():
        return []
    out: list[dict[str, Any]] = []
    try:
        with p.open("r", newline="", encoding="utf-8") as f:
            rdr = csv.DictReader(f)
            for row in rdr:
                if isinstance(row, dict):
                    out.append({str(k): (v if v is not None else "") for k, v in row.items()})
    except OSError:
        return []
    return out


@app.post("/payout/predict")
def api_payout_predict():
    body = request.get_json(silent=True) or {}
    legs = body.get("legs")
    if not isinstance(legs, list):
        return jsonify({"error": "legs must be an array"}), 400
    try:
        out = predict_payout([x for x in legs if isinstance(x, dict)])
    except ValueError as e:
        return jsonify({"error": str(e)}), 400
    r = jsonify(out)
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return r


@app.get("/payout/log")
def page_payout_log():
    return _grades_html_response("payout_log.html")


@app.post("/api/payout/log/extract")
def api_payout_log_extract():
    up = request.files.get("screenshot")
    if up is None:
        return jsonify({"error": "screenshot file is required"}), 400
    blob = up.read()
    if not blob:
        return jsonify({"error": "uploaded screenshot is empty"}), 400
    mime = str(up.mimetype or "image/png")
    try:
        extracted = _gemini_extract_ticket_from_image(blob, mime)
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500
    return jsonify({"ok": True, "extracted": extracted})


@app.post("/api/payout/log/save")
def api_payout_log_save():
    body = request.get_json(silent=True) or {}
    legs = body.get("legs") if isinstance(body.get("legs"), list) else []
    legs = [x for x in legs if isinstance(x, dict)]
    n_legs = int(_safe_float(body.get("n_legs"), float(len(legs))))
    if n_legs <= 0:
        n_legs = len(legs)
    if n_legs < 2 or n_legs > 6:
        return jsonify({"error": "n_legs must be between 2 and 6"}), 400

    comp = str(body.get("leg_composition") or _composition_label_from_legs(legs)).strip()
    goblin_deltas = [str(_safe_float(x.get("delta"), 0.0)) for x in legs if _normalize_leg_pick_type(x.get("pick_type")) == "Goblin"]
    demon_deltas = [str(_safe_float(x.get("delta"), 0.0)) for x in legs if _normalize_leg_pick_type(x.get("pick_type")) == "Demon"]
    row = {
        "date": str(body.get("date") or datetime.now().date().isoformat()).strip(),
        "n_legs": str(n_legs),
        "leg_composition": comp,
        "goblin_deltas": str(body.get("goblin_deltas") or ",".join(goblin_deltas)).strip(),
        "demon_deltas": str(body.get("demon_deltas") or ",".join(demon_deltas)).strip(),
        "power_payout_x": str(body.get("power_payout_x") or "").strip(),
        "flex_payout_x": str(body.get("flex_payout_x") or "").strip(),
        "source": str(body.get("source") or "screenshot").strip(),
        "notes": str(body.get("notes") or "").strip(),
    }
    PAYOUT_LADDER_LOG_PATH.parent.mkdir(parents=True, exist_ok=True)
    exists = PAYOUT_LADDER_LOG_PATH.is_file()
    with PAYOUT_LADDER_LOG_PATH.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=_PAYOUT_LADDER_FIELDS)
        if not exists:
            w.writeheader()
        w.writerow(row)
    # Per-leg payout lineage logging (additive/backward compatible).
    date_str = row["date"]
    ticket_id = str(body.get("ticket_id") or f"pp_{date_str}_{uuid.uuid4().hex[:8]}").strip()
    slip_type = str(body.get("slip_type") or "power").strip().lower()
    sport_hint = str(body.get("sport") or "NHL").strip() or "NHL"
    resolver = _get_payout_leg_resolver()
    resolved_legs: list[dict[str, Any]] = []
    for i, leg in enumerate(legs, start=1):
        leg_sport = str(leg.get("sport") or sport_hint).strip() or sport_hint
        line_val = _safe_num_or_none(leg.get("line"))
        res = resolver.resolve_leg(
            date=date_str,
            sport=leg_sport,
            player=str(leg.get("player") or ""),
            prop=str(leg.get("prop_type") or leg.get("prop") or ""),
            direction=str(leg.get("direction") or ""),
            played_line=line_val,
            pick_type=str(leg.get("pick_type") or "Standard"),
        )
        if line_val is not None and _normalize_leg_pick_type(leg.get("pick_type")) in {"Goblin", "Demon"}:
            cur_delta = _safe_num_or_none(leg.get("delta"))
            if (cur_delta is None or cur_delta <= 0) and _safe_num_or_none(res.get("delta")) is not None:
                leg["delta"] = res.get("delta")
        leg_row = {
            "date": date_str,
            "ticket_id": ticket_id,
            "slip_type": slip_type,
            "stake": str(body.get("entry_amount") or body.get("stake") or "").strip(),
            "payout_to_win": str(body.get("payout_to_win") or "").strip(),
            "power_payout_x": row.get("power_payout_x", ""),
            "sport": leg_sport,
            "leg_slot": str(i),
            "player": str(leg.get("player") or "").strip(),
            "prop": str(leg.get("prop_type") or leg.get("prop") or "").strip(),
            "line": "" if line_val is None else str(line_val),
            "direction": str(leg.get("direction") or "").strip().upper(),
            "pick_type": _normalize_leg_pick_type(leg.get("pick_type")),
            "source": row.get("source", ""),
            "notes": row.get("notes", ""),
            "delta_quality": str(res.get("delta_quality") or ""),
            "matched_snapshot_path": str(res.get("matched_snapshot_path") or ""),
            "matched_standard_line": (
                "" if _safe_num_or_none(res.get("matched_standard_line")) is None else str(res.get("matched_standard_line"))
            ),
            "delta_method": str(res.get("delta_method") or ""),
            "delta": "" if _safe_num_or_none(leg.get("delta")) is None else str(_safe_num_or_none(leg.get("delta"))),
        }
        _append_csv_with_schema_upgrade(PAYOUT_TICKET_LEGS_PATH, _PAYOUT_TICKET_LEG_FIELDS, leg_row)
        resolved_legs.append(leg_row)

    total = len(_read_payout_ladder_rows())
    return jsonify({"ok": True, "saved": row, "total_rows": total, "ticket_id": ticket_id, "legs_resolved": resolved_legs})


@app.get("/payout/ladder")
def page_payout_ladder():
    rows = _read_payout_ladder_rows()
    grouped: dict[tuple[int, str], list[dict[str, Any]]] = {}
    for r in rows:
        n = int(_safe_float(r.get("n_legs"), 0))
        comp = str(r.get("leg_composition") or "").strip()
        if n <= 0 or not comp:
            continue
        grouped.setdefault((n, comp), []).append(r)
    summary: list[dict[str, Any]] = []
    for (n, comp), recs in sorted(grouped.items(), key=lambda x: (x[0][0], x[0][1])):
        vals: list[float] = []
        for r in recs:
            p = _safe_float(r.get("power_payout_x"), float("nan"))
            f = _safe_float(r.get("flex_payout_x"), float("nan"))
            if math.isfinite(p) and p > 0:
                vals.append(p)
            if math.isfinite(f) and f > 0:
                vals.append(f)
        if vals:
            mn, mx, avg = min(vals), max(vals), statistics.mean(vals)
        else:
            mn = mx = avg = 0.0
        summary.append(
            {
                "n_legs": n,
                "leg_composition": comp,
                "samples": len(recs),
                "min_payout_x": round(mn, 4),
                "max_payout_x": round(mx, 4),
                "avg_payout_x": round(avg, 4),
                "is_sparse": len(recs) < 5,
            }
        )
    return _grades_html_response("payout_ladder.html", ladder_rows=summary, total_rows=len(rows))


@app.get("/payout/examples")
def page_payout_examples():
    p = PAYOUT_LADDER_EXAMPLES_PATH
    payload: dict[str, Any]
    if p.is_file():
        try:
            payload = json.loads(p.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            payload = {"generated_at": "", "examples": []}
    else:
        payload = {"generated_at": "", "examples": []}
    return _grades_html_response("payout_examples.html", payload=payload)


@app.get("/api/payout/rate-cards")
def api_payout_rate_cards():
    """JSON deck for /payout Rate cards tab (scripts/build_payout_rate_cards.py)."""
    path = BASE_DIR / "data" / "payout_rate_cards.json"
    if not path.is_file():
        return (
            jsonify(
                {
                    "schema_version": 1,
                    "error": "missing_file",
                    "message": "Run python scripts/build_payout_rate_cards.py from the repo root.",
                    "cards": [],
                }
            ),
            404,
        )
    try:
        data = json.loads(path.read_text(encoding="utf-8"))
    except (OSError, json.JSONDecodeError) as e:
        return jsonify({"schema_version": 1, "error": "read_failed", "message": str(e), "cards": []}), 500
    r = jsonify(data)
    r.headers["Cache-Control"] = "public, max-age=120"
    return r


@app.post("/api/payout/estimate-mult")
def api_payout_estimate_mult():
    """Estimated multiplier from shared Goblin/Demon curve (utils.goblin_demon_multiplier)."""
    from utils.goblin_demon_multiplier import leg_delta_pct, multiplier_summary

    body = request.get_json(silent=True) or {}
    mode_raw = str(body.get("mode", "power")).lower()
    mode = "flex" if "flex" in mode_raw else "power"

    raw_legs = body.get("legs")
    if not isinstance(raw_legs, list):
        return jsonify({"error": "legs array required"}), 400

    legs: List[dict[str, Any]] = []
    for L in raw_legs:
        if not isinstance(L, dict):
            continue
        typ = str(L.get("type", "")).lower()
        if typ == "goblin":
            pick = "Goblin"
        elif typ == "demon":
            pick = "Demon"
        else:
            pick = "Standard"
        dp = L.get("deltaPct")
        if dp is None and pick != "Standard":
            dp = leg_delta_pct(L.get("playedLine"), L.get("stdLine"))
        else:
            try:
                dp = float(dp) if dp is not None else None
            except (TypeError, ValueError):
                dp = None
        legs.append({"pick_type": pick, "delta_pct": dp})

    n = len(legs)
    if n < 2:
        return jsonify({"error": "at least 2 legs required"}), 400

    hits_raw = body.get("hits")
    try:
        hits = int(hits_raw) if hits_raw is not None and str(hits_raw).strip() != "" else None
    except (TypeError, ValueError):
        hits = None
    if mode == "flex" and hits is None:
        hits = n

    summ = multiplier_summary(legs, mode=mode, hits=hits)
    r = jsonify(summ)
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return r


def _graded_props_json_path_for_date(date_str: str) -> Path | None:
    fname = f"graded_props_{date_str}.json"
    for base in (TEMPLATES_DIR, ARCHIVE_DIR):
        if not base.is_dir():
            continue
        cand = base / fname
        if cand.is_file():
            return cand
    return None


GRADES_HTML_INITIAL_ROWS = 500


def _grades_group_props_into_sections(
    props_in: list[dict[str, Any]],
) -> list[tuple[str, list[dict[str, Any]]]]:
    pref = ("NBA", "NBA1H", "NBA1Q", "CBB", "WCBB", "NHL", "MLB", "SOCCER", "TENNIS")
    by_sport: dict[str, list[dict[str, Any]]] = {}
    for p in props_in:
        sk = str(p.get("sport") or "—").strip() or "—"
        by_sport.setdefault(sk, []).append(p)

    def sport_key(k: str) -> tuple:
        u = k.upper()
        if u in pref:
            return (0, pref.index(u), k)
        return (1, u, k)

    sport_keys = sorted(by_sport.keys(), key=sport_key)
    return [(k, by_sport[k]) for k in sport_keys]


def _grades_flat_rows_from_sections(
    by_sections: list[tuple[str, list[dict[str, Any]]]],
) -> list[tuple[str, dict[str, Any]]]:
    flat: list[tuple[str, dict[str, Any]]] = []
    for sk, rows in by_sections:
        for r in rows:
            flat.append((sk, r))
    return flat


def _grades_html_response(template: str, **kwargs: Any) -> Response:
    r = make_response(
        render_template(
            template,
            ui_build_id=_UI_BUILD_ID,
            deploy_git_sha=_deploy_git_sha_short(),
            **kwargs,
        )
    )
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    r.headers["Expires"] = "0"
    return r


def _send_grades_report_html(fname: str) -> Response | None:
    """Serve slate_eval_*.html or ticket_eval_*.html from templates or archive dir."""
    for base in (TEMPLATES_DIR, ARCHIVE_DIR):
        if base.exists() and (base / fname).is_file():
            response = send_from_directory(str(base), fname)
            response.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
            response.headers["Pragma"] = "no-cache"
            response.headers["Expires"] = "0"
            # Hub loads these same-origin in iframes; proxies must not force DENY (breaks mobile WebViews).
            response.headers.pop("X-Frame-Options", None)
            return response
    return None


@app.get("/grades/hub")
def page_grades_hub():
    """Bookmark alias: same experience as /grades (ticket evaluation hub)."""
    return redirect("/grades", code=302)


@app.get("/grades/props")
def page_grades_props_legacy_redirect():
    """Legacy URL: standalone raw-prop HTML was removed; use Grades → Prop Evaluation."""
    return redirect("/grades", code=302)


@app.get("/grades/props/<date_str>")
def page_grades_props_legacy_redirect_date(date_str: str):
    """Legacy URL per date; same redirect as /grades/props."""
    return redirect("/grades", code=302)


@app.get("/grades")
def page_grades():
    """Primary Grades hub: slate / ticket iframes, KPIs, per-sport views (indexGrades.html)."""
    return _grades_html_response("indexGrades.html")


@app.get("/grades/<date_str>")
def page_grades_ticket_date(date_str: str):
    """Ticket evaluation report for YYYY-MM-DD (same HTML as /grades/ticket_eval_<date>.html)."""
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_str):
        abort(404)
    r = _send_grades_report_html(f"ticket_eval_{date_str}.html")
    if r is not None:
        return r
    abort(404)


@app.route("/grades/slate_eval_<date>.html", methods=("GET", "HEAD"))
def serve_grade_report(date: str):
    """Serve individual slate_eval_YYYY-MM-DD.html files for the grades iframe."""
    fname = f"slate_eval_{date}.html"
    r = _send_grades_report_html(fname)
    if r is not None:
        return r
    abort(404)


@app.route("/grades/ticket_eval_<date>.html", methods=("GET", "HEAD"))
def serve_ticket_eval_report(date: str):
    """Serve individual ticket_eval_YYYY-MM-DD.html files for the ticket evaluation iframe."""
    fname = f"ticket_eval_{date}.html"
    r = _send_grades_report_html(fname)
    if r is not None:
        return r
    abort(404)


def _pick_type_from_tier_tier_el(tier_el) -> str:
    if tier_el is None:
        return "—"
    t = tier_el.get_text(strip=True).upper()
    return {"G": "Goblin", "S": "Standard", "D": "Demon"}.get(t, tier_el.get_text(strip=True) or "—")


def _sport_from_ticket_eval_pill(span) -> str:
    if span is None:
        return "—"
    classes = span.get("class") or []
    for c in classes:
        if isinstance(c, str) and c.startswith("sport-"):
            key = c[len("sport-") :].lower()
            return {
                "nba": "NBA",
                "cbb": "CBB",
                "nhl": "NHL",
                "mlb": "MLB",
                "soccer": "Soccer",
                "tennis": "Tennis",
                "wnba": "WNBA",
            }.get(key, key.upper())
    return (span.get_text(strip=True) or "—").upper()


def _player_name_from_ticket_legrow(row) -> str:
    for el in row.select("div.pl-hit, div.pl-void, div.pl-pend"):
        t = el.get_text(strip=True)
        if t:
            return t
    pn = row.select_one("span.pl-name")
    if pn is not None:
        t = pn.get_text(strip=True)
        if t:
            return t
    return ""


def _props_from_ticket_eval_html(date_q: str) -> list[dict[str, str]]:
    """
    When graded_props_*.json is missing, parse ticket_eval_*.html leg rows so the
    Prop Evaluation tab can still list ticket legs (subset of the full slate).
    """
    try:
        from bs4 import BeautifulSoup
    except ImportError:
        return []

    fname = f"ticket_eval_{date_q}.html"
    raw_path: Path | None = None
    for base in (TEMPLATES_DIR, ARCHIVE_DIR):
        if not base.exists():
            continue
        p = base / fname
        if p.is_file():
            raw_path = p
            break
    if raw_path is None:
        return []

    try:
        html = raw_path.read_text(encoding="utf-8", errors="replace")
    except OSError:
        return []

    soup = BeautifulSoup(html, "html.parser")
    props: list[dict[str, str]] = []
    def _first_num(s: str) -> float | None:
        m = re.search(r"-?\d+(?:\.\d+)?", s or "")
        if not m:
            return None
        try:
            return float(m.group(0))
        except Exception:
            return None

    for row in soup.select("div.legrow"):
        classes = " ".join(row.get("class") or [])
        if "leg-hit" in classes:
            result = "HIT"
        elif "leg-miss" in classes:
            result = "MISS"
        elif "leg-void" in classes:
            result = "VOID"
        elif "leg-pend" in classes:
            result = "—"
        else:
            continue

        player = _player_name_from_ticket_legrow(row)
        if not player:
            continue

        pill = row.select_one("span.pill")
        sport = _sport_from_ticket_eval_pill(pill)

        tier_el = row.select_one("div.tier")
        pick_type = _pick_type_from_tier_tier_el(tier_el)

        prop_txt = "—"
        team_txt = "—"
        pcol = row.select_one("div.leg-prop-col")
        if pcol is not None:
            kids = pcol.find_all("div", recursive=False)
            if kids:
                prop_txt = kids[0].get_text(strip=True) or "—"
            meta = pcol.select_one("div.meta-muted")
            if meta is not None:
                team_txt = meta.get_text(strip=True) or "—"

        direction = "—"
        line_txt = "—"
        actual_txt = "—"
        edge_txt = "—"
        margin_txt = "—"
        extras = row.select("div.leg-extra")
        if extras:
            first = extras[0]
            line_txt = first.get_text(" ", strip=True) or "—"
            if first.select_one(".dir-over") is not None:
                direction = "OVER"
            elif first.select_one(".dir-under") is not None:
                direction = "UNDER"
            if len(extras) > 1:
                actual_txt = extras[1].get_text(" ", strip=True) or "—"
            if len(extras) > 2:
                edge_txt = extras[2].get_text(" ", strip=True) or "—"
            line_num = _first_num(line_txt)
            actual_num = _first_num(actual_txt)
            if line_num is not None and actual_num is not None:
                margin = actual_num - line_num
                margin_txt = f"{margin:.2f}".rstrip("0").rstrip(".")

        props.append(
            {
                "sport": sport,
                "player": player,
                "team": team_txt,
                "prop": prop_txt,
                "line": line_txt,
                "direction": direction,
                "pick_type": pick_type,
                "actual_value": actual_txt,
                "edge": edge_txt,
                "margin": margin_txt,
                "result": result,
            }
        )

    return props


@app.get("/api/graded-props")
def api_graded_props():
    """JSON list of graded props for a slate date (from graded_props_YYYY-MM-DD.json)."""
    date_q = (request.args.get("date") or "").strip()
    if not date_q or not re.match(r"^\d{4}-\d{2}-\d{2}$", date_q):
        return jsonify(
            {"error": "missing_or_invalid_date", "detail": "Use ?date=YYYY-MM-DD"}
        ), 400
    fname = f"graded_props_{date_q}.json"
    path = TEMPLATES_DIR / fname
    if not path.exists() and ARCHIVE_DIR.exists():
        alt = ARCHIVE_DIR / fname
        if alt.exists():
            path = alt
    if not path.exists():
        legs = _props_from_ticket_eval_html(date_q)
        if legs:
            return jsonify(
                {
                    "date": date_q,
                    "count": len(legs),
                    "props": legs,
                    "source": "ticket_eval_html",
                    "note": "Ticket legs only (subset). Deploy graded_props JSON for full slate props.",
                }
            )
        return jsonify(
            {"date": date_q, "count": 0, "props": [], "missing": True}
        )
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
        if isinstance(data, dict):
            # Normalize stale graded_props bundles so Prop Evaluation does not show
            # VOID_* eligibility flags as active when actual+line reconcile to HIT/MISS.
            props_in = list(data.get("props") or [])
            props_out: list[dict[str, Any]] = []
            for p in props_in:
                if not isinstance(p, dict):
                    continue
                row = dict(p)
                if row.get("actual_value") in (None, "") and row.get("actual") not in (None, ""):
                    row["actual_value"] = row.get("actual")
                if row.get("direction") in (None, "") and row.get("dir") not in (None, ""):
                    row["direction"] = row.get("dir")
                if row.get("void_reason") in (None, "") and row.get("void_reason_grade") not in (None, ""):
                    row["void_reason"] = row.get("void_reason_grade")
                row = reconcile_props_history_dict(row)
                res_u = str(row.get("result") or "").strip().upper()
                if res_u in ("HIT", "MISS", "PUSH"):
                    # Keep legacy eligibility flags for archives, but don't surface as active voids
                    # once the row has a reconciled game result.
                    row["void_reason"] = ""
                props_out.append(row)
            out = dict(data)
            out["props"] = props_out
            out["count"] = len(props_out)
            return jsonify(out)
        return jsonify({"error": "invalid_shape", "detail": "expected object"}), 500
    except Exception as exc:
        return jsonify({"error": "read_failed", "detail": str(exc)}), 500


def _iter_props_history_db_paths():
    cache = BASE_DIR / "data" / "cache"
    if not cache.is_dir():
        return []
    return sorted(cache.glob("*_props_history.db"))


def _ensure_props_history_void_reason_column(conn: sqlite3.Connection) -> None:
    """Older props_history DBs lack void_reason; add it so SELECT lists stay valid."""
    try:
        cur = conn.execute("PRAGMA table_info(props_history)")
        cols = {r[1] for r in cur.fetchall()}
        if cols and "void_reason" not in cols:
            conn.execute("ALTER TABLE props_history ADD COLUMN void_reason TEXT")
            conn.commit()
    except Exception:
        pass


def _grades_props_payload(date_str: str) -> dict[str, Any]:
    """
    Per-prop graded rows for the Grades hub Prop Evaluation tab.

    Primary: scripts/step_archive.py → data/cache/{sport}_props_history.db (local, gitignored).
    Fallback: ui_runner/data/grades_props/YYYY-MM-DD.json (committed; for Railway / hosts without SQLite archives).
    """
    props: list[dict[str, Any]] = []
    for dbp in _iter_props_history_db_paths():
        conn: sqlite3.Connection | None = None
        try:
            conn = sqlite3.connect(str(dbp))
            _ensure_props_history_void_reason_column(conn)
            conn.row_factory = sqlite3.Row
            cur = conn.execute(
                """
                SELECT sport, grade_date, player_name, prop_type, line, direction,
                       actual_value, result, margin, opp_team, team, pick_type, tier, edge, ml_prob,
                       void_reason
                FROM props_history
                WHERE grade_date = ?
                ORDER BY sport, player_name, prop_type, direction
                """,
                (date_str,),
            )
            for row in cur.fetchall():
                item = {k: row[k] for k in row.keys()}
                props.append(item)
        except Exception:
            pass
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass

    from_bundle = False
    if not props and GRADES_PROPS_EXPORT_DIR.is_dir():
        bundle = GRADES_PROPS_EXPORT_DIR / f"{date_str}.json"
        if bundle.is_file():
            try:
                raw = json.loads(bundle.read_text(encoding="utf-8"))
                props = list(raw.get("props") or [])
                from_bundle = True
            except Exception:
                props = []

    props = [reconcile_props_history_dict(p) for p in props]

    props.sort(
        key=lambda r: (
            str(r.get("sport") or "").upper(),
            str(r.get("player_name") or "").lower(),
            str(r.get("prop_type") or "").lower(),
            str(r.get("direction") or "").upper(),
        )
    )
    n_total = len(props)
    n_hit = sum(1 for p in props if str(p.get("result") or "").upper() == "HIT")
    n_miss = sum(1 for p in props if str(p.get("result") or "").upper() == "MISS")
    cap = 8000
    truncated = n_total > cap
    if truncated:
        props = props[:cap]
    out: dict[str, Any] = {
        "date": date_str,
        "n": n_total,
        "n_returned": len(props),
        "n_hit": n_hit,
        "n_miss": n_miss,
        "n_other": max(0, n_total - n_hit - n_miss),
        "truncated": truncated,
        "props": props,
    }
    if from_bundle:
        out["from_bundle"] = True
    return out


def _grades_archive_dates_payload(max_dates: int = 40) -> dict[str, Any]:
    """
    Distinct grade_date values in local props_history DBs (newest first).
    Helps the Grades UI when the report date (often browser \"yesterday\") does not match
    archived slate dates (e.g. machine clock vs outputs/YYYY-MM-DD used when grading).
    """
    counts: dict[str, int] = {}
    for dbp in _iter_props_history_db_paths():
        conn: sqlite3.Connection | None = None
        try:
            conn = sqlite3.connect(str(dbp))
            cur = conn.execute(
                "SELECT grade_date, COUNT(*) FROM props_history GROUP BY grade_date"
            )
            for gd, n in cur.fetchall():
                if not gd:
                    continue
                key = str(gd).strip()[:10]
                if len(key) != 10 or key[4] != "-" or key[7] != "-":
                    continue
                counts[key] = counts.get(key, 0) + int(n)
        except Exception:
            pass
        finally:
            if conn is not None:
                try:
                    conn.close()
                except Exception:
                    pass
    # Bundled JSON exports (deploy)
    if GRADES_PROPS_EXPORT_DIR.is_dir():
        for p in sorted(GRADES_PROPS_EXPORT_DIR.glob("*.json")):
            mm = re.fullmatch(r"(\d{4}-\d{2}-\d{2})\.json", p.name)
            if not mm:
                continue
            key = mm.group(1)
            try:
                raw = json.loads(p.read_text(encoding="utf-8"))
                n = len(raw.get("props") or [])
            except Exception:
                n = 0
            if n <= 0:
                continue
            counts[key] = max(counts.get(key, 0), n)

    et_today = _eastern_today_ymd()
    non_future = [d for d in counts.keys() if d <= et_today]
    ordered_src = non_future if non_future else list(counts.keys())
    ordered = sorted(ordered_src, reverse=True)[:max_dates]
    return {
        "dates": ordered,
        "row_counts": {d: counts[d] for d in ordered},
    }


@app.get("/api/grades/props")
def api_grades_props():
    """Graded prop rows for a report date (props_history archives)."""
    raw = (request.args.get("date") or "").strip()
    date_str = raw[:10]
    if len(date_str) != 10 or date_str[4] != "-" or date_str[7] != "-":
        return jsonify({"error": "invalid date; use YYYY-MM-DD", "props": []}), 400
    try:
        y, m, d = int(date_str[0:4]), int(date_str[5:7]), int(date_str[8:10])
        datetime(y, m, d)
    except (TypeError, ValueError):
        return jsonify({"error": "invalid date", "props": []}), 400
    r = jsonify(_grades_props_payload(date_str))
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


@app.get("/api/grades/page-rows")
def api_grades_page_rows():
    """
    Paginated rows from graded_props_*.json on disk (sport ordering matches Grades hub).
    """
    raw = (request.args.get("date") or "").strip()
    date_str = raw[:10]
    if not re.fullmatch(r"\d{4}-\d{2}-\d{2}", date_str):
        return jsonify({"error": "invalid date; use YYYY-MM-DD", "items": [], "total": 0}), 400
    try:
        y, m, d = int(date_str[0:4]), int(date_str[5:7]), int(date_str[8:10])
        datetime(y, m, d)
    except (TypeError, ValueError):
        return jsonify({"error": "invalid date", "items": [], "total": 0}), 400
    path = _graded_props_json_path_for_date(date_str)
    if path is None:
        return jsonify({"error": "not found", "items": [], "total": 0}), 404
    try:
        data = json.loads(path.read_text(encoding="utf-8-sig"))
    except Exception as exc:
        return jsonify({"error": str(exc), "items": [], "total": 0}), 500
    if not isinstance(data, dict):
        return jsonify({"error": "invalid json shape", "items": [], "total": 0}), 500
    props_in = [p for p in (data.get("props") or []) if isinstance(p, dict)]
    by_sections = _grades_group_props_into_sections(props_in)
    flat = _grades_flat_rows_from_sections(by_sections)
    try:
        offset = max(0, int(request.args.get("offset", "0")))
    except (TypeError, ValueError):
        offset = 0
    try:
        limit = int(request.args.get("limit", str(GRADES_HTML_INITIAL_ROWS)))
    except (TypeError, ValueError):
        limit = GRADES_HTML_INITIAL_ROWS
    limit = max(1, min(limit, 2000))
    chunk = flat[offset : offset + limit]
    r = jsonify(
        {
            "date": date_str,
            "offset": offset,
            "total": len(flat),
            "next_offset": offset + len(chunk),
            "has_more": offset + len(chunk) < len(flat),
            "items": [{"sport": sk, "row": row} for sk, row in chunk],
        }
    )
    r.headers["Cache-Control"] = "no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


@app.get("/api/grades/archive_dates")
def api_grades_archive_dates():
    """Newest grade_date keys present in data/cache/*_props_history.db (for Prop Evaluation date hints)."""
    r = jsonify(_grades_archive_dates_payload())
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


_SLATE_EVAL_FN_RE = re.compile(r"^slate_eval_(\d{4}-\d{2}-\d{2})\.html$")
_TICKET_EVAL_FN_RE = re.compile(r"^ticket_eval_(\d{4}-\d{2}-\d{2})\.html$")


def _grade_report_dates_on_disk(which: str) -> list[str]:
    """
    List YYYY-MM-DD values for which static report HTML exists (templates/ or templates/archive/).
    which: 'slate' | 'ticket'
    """
    pat = _SLATE_EVAL_FN_RE if which == "slate" else _TICKET_EVAL_FN_RE
    found: set[str] = set()
    for base in (TEMPLATES_DIR, ARCHIVE_DIR):
        if not base.is_dir():
            continue
        try:
            for p in base.iterdir():
                if not p.is_file():
                    continue
                m = pat.match(p.name)
                if m:
                    found.add(m.group(1))
        except OSError:
            continue
    et_today = _eastern_today_ymd()
    non_future = [d for d in found if d <= et_today]
    return sorted(non_future if non_future else found)


@app.get("/api/grades/report_dates")
def api_grades_report_dates():
    """
    Dates that have slate_eval_*.html / ticket_eval_*.html on the server filesystem.

    The Grades hub uses this so date pills work even when many parallel HEAD requests
    fail through a CDN/proxy. Deploy must include committed files under ui_runner/templates/.
    """
    r = jsonify(
        {
            "slate_eval_dates": _grade_report_dates_on_disk("slate"),
            "ticket_eval_dates": _grade_report_dates_on_disk("ticket"),
        }
    )
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


@app.get("/api/grade-history")
def api_grade_history():
    """
    Daily ticket_eval summaries: predicted vs actual payout aggregates and recommendation buckets.
    Populated when ``scripts/build_ticket_eval.py`` runs (appends ``DATA_ROOT/grade_history.json``).
    """
    out = _read_grade_history_runs()
    r = jsonify(out)
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


# ──────────────────────────────────────────────────────────────────────────────
# NEW: Pipeline Status API
# Returns health of all pipeline outputs so UI can show green/red indicators
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/api/pipeline/status")
def api_pipeline_status():
    slate_counts, slate_disk_info = _slate_counts()

    slate_payload: dict | None = None
    tickets_payload: dict | None = None
    try:
        slate_payload = read_json_cached(TEMPLATES_DIR / "slate_latest.json")
    except Exception:
        pass
    try:
        tickets_payload = read_json_cached(TEMPLATES_DIR / "tickets_latest.json")
    except Exception:
        pass

    te_payload: dict | None = None
    try:
        if _template_json_available("ticket_eval_slate_latest.json"):
            te_payload = read_json_cached(TEMPLATES_DIR / "ticket_eval_slate_latest.json")
    except Exception:
        pass

    slate_js_ts, slate_js_disp = _payload_timestamp_meta(slate_payload)
    tik_js_ts, tik_js_disp = _payload_timestamp_meta(tickets_payload)
    te_js_ts, te_js_disp = _payload_timestamp_meta(te_payload)
    status_js_ts, status_js_disp = _fresher_meta(
        _fresher_meta((slate_js_ts, slate_js_disp), (tik_js_ts, tik_js_disp)),
        (te_js_ts, te_js_disp),
    )
    # One string for Slate Explorer cards: avoids COMBINED using only tickets while sports
    # used a merged fresher time, or status_js_disp None while tik_js_disp is set.
    card_disp = status_js_disp or tik_js_disp or slate_js_disp or te_js_disp
    has_web_tickets = bool(tickets_payload and tickets_payload.get("groups"))

    _pref_d = (tickets_payload or {}).get("date") or (slate_payload or {}).get("date")
    days = _slate_day_candidates(str(_pref_d)[:10] if _pref_d else None)

    nba_slate_p = _resolve_outputs_artifact(
        days, "step8_nba_direction_clean_{d}.xlsx", NBA_SLATE, NBA_DIR / "step8_nba_direction_clean.xlsx"
    )
    nba1h_slate_p = _resolve_outputs_artifact(days, "step8_nba1h_direction_clean_{d}.xlsx", NBA1H_SLATE)
    nba1q_slate_p = _resolve_outputs_artifact(days, "step8_nba1q_direction_clean_{d}.xlsx", NBA1Q_SLATE)
    cbb_slate_p = _resolve_outputs_artifact(days, "step6_ranked_cbb_{d}.xlsx", CBB_SLATE)
    wcbb_slate_p = _resolve_outputs_artifact(days, "step6_ranked_wcbb_{d}.xlsx", WCBB_SLATE)
    cfb_slate_p = _resolve_outputs_artifact(
        days,
        "cfb/step6_ranked_cfb.xlsx",
        CFB_SLATE,
        CFB_DIR / "step6_ranked_cfb.xlsx",
    )
    nhl_slate_p = _resolve_outputs_artifact(days, "step8_nhl_direction_clean_{d}.xlsx", NHL_SLATE)
    soccer_slate_p = _resolve_outputs_artifact(days, "step8_soccer_direction_clean_{d}.xlsx", SOCCER_SLATE)
    mlb_slate_p = _resolve_outputs_artifact(days, "step8_mlb_direction_clean_{d}.xlsx", MLB_SLATE)
    tennis_slate_p = _resolve_outputs_artifact(
        days,
        "step8_tennis_direction_clean_{d}.xlsx",
        TENNIS_SLATE,
        TENNIS_DIR / "outputs" / "step8_tennis_direction_clean.xlsx",
    )
    wnba_slate_p = _resolve_outputs_artifact(
        days,
        [
            "wnba/step8_wnba_direction_clean.xlsx",
            "wnba/step8_wnba_direction.xlsx",
            "step8_wnba_direction_clean_{d}.xlsx",
            "step8_wnba_direction_{d}.xlsx",
        ],
        WNBA_SLATE,
        WNBA_DIR / "step8_wnba_direction_clean.xlsx",
        WNBA_DIR / "step8_wnba_direction.xlsx",
    )
    nfl_slate_p = _resolve_outputs_artifact(
        days,
        "step8_nfl_direction_clean_{d}.xlsx",
        NFL_SLATE,
        NFL_DIR / "outputs" / "step8_nfl_direction_clean.xlsx",
    )
    udq_p = _resolve_outputs_artifact(days, "upstream_data_quality_{d}.csv")

    combined_candidates: list[Path] = []
    ui_data_dir = BASE_DIR / "ui_runner" / "data"
    for d in days:
        out_d = BASE_DIR / "outputs" / d
        combined_candidates.extend(BASE_DIR.glob(f"combined_slate_tickets_{d}*.xlsx"))
        combined_candidates.extend(BASE_DIR.glob(f"combined_slate_tickets_{d}*.json"))
        combined_candidates.extend(ui_data_dir.glob(f"combined_slate_tickets_{d}*.json"))
        if out_d.is_dir():
            combined_candidates.extend(out_d.glob(f"combined_slate_tickets_{d}*.xlsx"))
            combined_candidates.extend(out_d.glob(f"combined_slate_tickets_{d}*.json"))
    combined_path = (
        max(combined_candidates, key=lambda p: p.stat().st_mtime) if combined_candidates else None
    )
    combined_slate = _file_info(combined_path) if combined_path else {"exists": False}
    if combined_slate.get("exists"):
        if has_web_tickets and card_disp:
            combined_slate = {**combined_slate, "modified": card_disp}
        elif tik_js_ts and tik_js_disp and tik_js_ts > _mtime_ts(combined_slate.get("modified")):
            combined_slate = {**combined_slate, "modified": tik_js_disp}
    elif has_web_tickets and card_disp:
        approx_kb = round(len(json.dumps(tickets_payload)) / 1024, 1)
        combined_slate = {
            "exists": True,
            "modified": card_disp,
            "size_kb": approx_kb,
        }

    return jsonify({
        "nba": {
            "run_complete_flag": NBA_FLAG.exists(),
            "slate":   _sport_slate_status(nba_slate_p, "nba", slate_counts, slate_disk_info, status_js_ts, card_disp),
            "tickets": _file_info(NBA_TICKETS),
        },
        "nba1h": {
            "slate":   _sport_slate_status(nba1h_slate_p, "nba1h", slate_counts, slate_disk_info, status_js_ts, card_disp),
            "tickets": _file_info(NBA1H_TICKETS),
        },
        "nba1q": {
            "slate":   _sport_slate_status(nba1q_slate_p, "nba1q", slate_counts, slate_disk_info, status_js_ts, card_disp),
            "tickets": _file_info(NBA1Q_TICKETS),
        },
        "cbb": {
            "slate": _sport_slate_status(cbb_slate_p, "cbb", slate_counts, slate_disk_info, status_js_ts, card_disp),
        },
        "cfb": {
            "slate": _sport_slate_status(cfb_slate_p, "cfb", slate_counts, slate_disk_info, status_js_ts, card_disp),
        },
        "wcbb": {
            "slate": _sport_slate_status(wcbb_slate_p, "wcbb", slate_counts, slate_disk_info, status_js_ts, card_disp),
        },
        "nhl": {
            "slate":   _sport_slate_status(nhl_slate_p, "nhl", slate_counts, slate_disk_info, status_js_ts, card_disp),
            "tickets": _file_info(NHL_TICKETS),
        },
        "soccer": {
            "slate":   _sport_slate_status(soccer_slate_p, "soccer", slate_counts, slate_disk_info, status_js_ts, card_disp),
            "tickets": _file_info(SOCCER_TICKETS),
        },
        "mlb": {
            "slate":   _sport_slate_status(mlb_slate_p, "mlb", slate_counts, slate_disk_info, status_js_ts, card_disp),
            "tickets": _file_info(MLB_TICKETS),
        },
        "tennis": {
            "slate": _sport_slate_status(tennis_slate_p, "tennis", slate_counts, slate_disk_info, status_js_ts, card_disp),
        },
        "wnba": {
            "slate": _sport_slate_status(wnba_slate_p, "wnba", slate_counts, slate_disk_info, status_js_ts, card_disp),
        },
        "nfl": {
            "slate": _sport_slate_status(nfl_slate_p, "nfl", slate_counts, slate_disk_info, status_js_ts, card_disp),
        },
        "combined": {
            "slate": combined_slate,
        },
        "pipeline_outputs_date": days[0] if days else None,
        "upstream_data_quality": _file_info(udq_p),
        "as_of": datetime.now().strftime("%Y-%m-%d %H:%M:%S"),
        "slate_json_source": "remote" if bool(_DATA_FILE_URL_MAP.get("slate_latest.json")) else "disk",
        "ticket_eval_slate_source": "remote"
        if bool(_DATA_FILE_URL_MAP.get("ticket_eval_slate_latest.json"))
        else "disk",
        "slate_sport_source": os.environ.get("SLATE_SPORT_SOURCE", "auto").strip() or "auto",
        "tickets_json_source": "remote" if bool(_DATA_FILE_URL_MAP.get("tickets_latest.json")) else "disk",
        "railway_auto_github_json": bool(_running_on_railway() and not os.environ.get("DISABLE_AUTO_GITHUB_JSON", "").strip()),
        "deploy_git_sha": (os.environ.get("RAILWAY_GIT_COMMIT_SHA") or os.environ.get("GIT_COMMIT") or "")[:40],
    })


# ──────────────────────────────────────────────────────────────────────────────
# NEW: Active Job Status (quick poll for running job count)
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/api/pipeline/running")
def api_pipeline_running():
    with LOCK:
        running = [j for j in JOBS.values() if j.status == "RUNNING"]
    return jsonify({
        "running": len(running),
        "jobs": [{"job_id": j.job_id, "label": j.label, "started_at": j.started_at}
                 for j in running]
    })


# ──────────────────────────────────────────────────────────────────────────────
# API: Run Command
# ──────────────────────────────────────────────────────────────────────────────

# ── API: Config endpoint ──────────────────────────────────────────────────────
@app.get("/api/config")
def api_config():
    try:
        return jsonify(load_config())
    except Exception as exc:
        return jsonify({"error": str(exc)}), 500


@app.post("/api/run")
def api_run():
    data       = request.get_json(force=True) or {}
    pipeline   = data.get("pipeline")
    command_id = data.get("command_id")

    if not pipeline or not command_id:
        return jsonify({"error": "missing_pipeline_or_command_id"}), 400

    try:
        config      = load_config()
        repo_root   = resolve_repo_root(config)
        cmd_def     = resolve_command(config, pipeline, command_id)
        workdir_rel = (config["pipelines"][pipeline].get("workdir") or "").strip()
        workdir     = (repo_root / workdir_rel).resolve()
    except Exception as exc:
        return jsonify({"error": "config_or_command_error", "detail": str(exc)}), 400

    # ── Chain ──
    if cmd_def["type"] == "chain":
        parent_id = str(uuid.uuid4())
        parent    = RunJob(job_id=parent_id, label=cmd_def["label"])
        # Pre-populate step list for progress tracking
        parent.steps = [
            {"id": item.get("id"), "label": item.get("label", item.get("id")), "status": "PENDING"}
            for item in cmd_def["items"]
        ]
        with LOCK:
            JOBS[parent_id] = parent

        def chain_runner() -> None:
            ok = True
            for i, item in enumerate(cmd_def["items"]):
                label   = item.get("label") or item.get("id") or "STEP"
                raw_cmd = item.get("cmd")

                # Update step status to RUNNING
                with LOCK:
                    if i < len(parent.steps):
                        parent.steps[i]["status"] = "RUNNING"
                    parent.lines.append("")
                    parent.lines.append(f"=== {label} ===")

                if not isinstance(raw_cmd, list):
                    with LOCK:
                        parent.lines.append(f"[ERROR] Bad cmd for '{label}': expected list")
                        if i < len(parent.steps):
                            parent.steps[i]["status"] = "FAIL"
                    ok = False
                    break

                cmd   = subst_tokens(raw_cmd, config=config)
                child = RunJob(job_id=str(uuid.uuid4()), label=label)
                _run_process(child, cmd, workdir)

                with LOCK:
                    parent.lines.extend(child.lines)
                    parent.lines = safe_tail(parent.lines)
                    step_ok = (child.return_code == 0)
                    if i < len(parent.steps):
                        parent.steps[i]["status"] = "OK" if step_ok else "FAIL"
                    if not step_ok:
                        ok = False
                        parent.lines.append("[CHAIN] Stopping — step failed.")
                        # Mark remaining steps as SKIPPED
                        for j in range(i + 1, len(parent.steps)):
                            parent.steps[j]["status"] = "SKIPPED"
                        break

            with LOCK:
                parent.ended_at    = time.time()
                parent.status      = "OK" if ok else "FAIL"
                parent.return_code = 0 if ok else 1

        threading.Thread(target=chain_runner, daemon=True).start()
        return jsonify({"job_id": parent_id})

    # ── Single ──
    item    = cmd_def["item"]
    raw_cmd = item.get("cmd")
    if not isinstance(raw_cmd, list):
        return jsonify({"error": "bad_cmd_type", "detail": f"Expected list, got {type(raw_cmd)}"}), 400

    cmd    = subst_tokens(raw_cmd, config=config)
    job_id = start_job(cmd_def["label"], cmd, workdir)
    return jsonify({"job_id": job_id})


@app.post("/api/mobile/upload-data")
def api_mobile_upload_data():
    """
    Securely receive JSON data updates from local PC.
    Requires X-Mobile-Token header matching PROPORACLE_MOBILE_TOKEN env.
    """
    token = os.environ.get("PROPORACLE_MOBILE_TOKEN", "").strip()
    if not token:
        return jsonify({"error": "Server PROPORACLE_MOBILE_TOKEN not set"}), 503

    if request.headers.get("X-Mobile-Token") != token:
        return jsonify({"error": "Unauthorized"}), 401

    data = request.get_json(silent=True)
    if not data or "filename" not in data or "payload" not in data:
        return jsonify({"error": "Invalid payload"}), 400

    filename = data["filename"]
    # Restrict to known safe JSON types
    if not re.match(r"^[a-z0-9_.-]+\.json$", filename.lower()):
        return jsonify({"error": "Invalid filename"}), 400

    # Write to templates (which are gzipped/cached by app.py) or persistent root
    # For live mobile updates, we write to templates so /api/slate etc. see them immediately.
    target_path = TEMPLATES_DIR / filename
    try:
        target_path.write_text(json.dumps(data["payload"]), encoding="utf-8")
        # Invalidate cache
        with _JSON_FILE_CACHE_LOCK:
            _json_file_cache.pop(str(target_path.resolve()), None)
        return jsonify({"ok": True, "saved": filename})
    except Exception as e:
        return jsonify({"error": str(e)}), 500


@app.route("/api/mobile/bundle-version", methods=["OPTIONS"])
@app.route("/api/mobile/bundle.zip", methods=["OPTIONS"])
def api_mobile_bundle_options():
    r = make_response("", 204)
    r.headers["Access-Control-Allow-Origin"] = "*"
    r.headers["Access-Control-Allow-Methods"] = "GET, OPTIONS"
    r.headers["Access-Control-Allow-Headers"] = "*"
    return r


@app.get("/api/mobile/bundle-version")
def api_mobile_bundle_version():
    """Fingerprint of deployed `mobile/www` for Capacitor OTA (DIY Live Update)."""
    ver, ok = _mobile_www_bundle_fingerprint()
    if not ok:
        return jsonify({"ok": False, "error": "mobile_www_missing", "version": ""}), 404
    return jsonify(
        {
            "ok": True,
            "version": ver,
            "deploy_sha": _deploy_git_sha_short(),
        }
    )


@app.get("/api/mobile/bundle.zip")
def api_mobile_bundle_zip():
    """Zip of `mobile/www` for OTA install into Android app files dir (see OtaBundlePlugin)."""
    root = _MOBILE_WWW_DIR
    ver, ok = _mobile_www_bundle_fingerprint()
    if not ok:
        abort(404)
    cache_root = Path(tempfile.gettempdir()) / "proporacle_mobile_bundles"
    cache_root.mkdir(parents=True, exist_ok=True)
    zip_path = cache_root / f"mobile-www-{ver}.zip"
    if not zip_path.is_file():
        partial = zip_path.with_suffix(".zip.partial")
        try:
            with zipfile.ZipFile(partial, "w", compression=zipfile.ZIP_DEFLATED) as zf:
                files = [p for p in root.rglob("*") if p.is_file()]
                files.sort(key=lambda p: p.relative_to(root).as_posix())
                for p in files:
                    rel = p.relative_to(root).as_posix()
                    if rel.split("/")[0].startswith("."):
                        continue
                    zf.write(p, rel)
            partial.replace(zip_path)
        except Exception:
            try:
                if partial.exists():
                    partial.unlink()
            except OSError:
                pass
            raise
    resp = send_file(
        zip_path,
        mimetype="application/zip",
        as_attachment=True,
        download_name="proporacle-mobile-www.zip",
    )
    resp.headers["Cache-Control"] = "public, max-age=300"
    return resp


# ──────────────────────────────────────────────────────────────────────────────
# API: Job Status (with step progress for chain jobs)
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/api/job/<job_id>")
def api_job(job_id: str):
    with LOCK:
        j = JOBS.get(job_id)
        if not j:
            return jsonify({"error": "not_found"}), 404
        return jsonify({
            "job_id":      j.job_id,
            "label":       j.label,
            "status":      j.status,
            "return_code": j.return_code,
            "started_at":  j.started_at,
            "ended_at":    j.ended_at,
            "lines":       j.lines[-400:],
            "steps":       j.steps,   # NEW: step-level progress
            "elapsed_s":   round((j.ended_at or time.time()) - j.started_at, 1),
        })


@app.get("/api/jobs")
def api_jobs():
    with LOCK:
        out = [
            {
                "job_id":      j.job_id,
                "label":       j.label,
                "status":      j.status,
                "started_at":  j.started_at,
                "ended_at":    j.ended_at,
                "return_code": j.return_code,
                "elapsed_s":   round((j.ended_at or time.time()) - j.started_at, 1),
                "steps":       j.steps,
            }
            for j in JOBS.values()
        ]
    out.sort(key=lambda x: x["started_at"], reverse=True)
    return jsonify(out[:25])


# ──────────────────────────────────────────────────────────────────────────────
# Main
# ──────────────────────────────────────────────────────────────────────────────

@app.get("/api/tickets-latest")
def api_tickets_latest():
    """Full tickets payload (groups, filters, date) for debugging and clients."""
    json_path = TEMPLATES_DIR / "tickets_latest.json"
    if not _template_json_available("tickets_latest.json"):
        return jsonify({"error": "tickets_latest.json not found", "groups": []}), 404
    try:
        data = read_json_cached(json_path)
        # jsonify serializes each ticket's payout dict unchanged when present.
        resp = jsonify(data)
        resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        resp.headers["Pragma"] = "no-cache"
        return resp
    except Exception as e:
        return jsonify({"error": str(e), "groups": []}), 500


def _ticket_ev_summary_from_payload(data: dict | None) -> dict[str, Any]:
    """
    Aggregate empirical EV (tickets_latest.json payout blocks).
    Buckets are mutually exclusive: strong >1.5, ok >1.0, marginal >0.8, skip <=0.8.
    """
    date_str = str((data or {}).get("date") or "").strip()[:10] or None
    groups = list((data or {}).get("groups") or [])
    strong_c = ok_c = marg_c = skip_c = 0
    total_considered = 0
    best_ticket: dict[str, Any] | None = None
    best_ev = float("-inf")

    for g in groups:
        gname = str(g.get("group_name") or "")
        for t in g.get("tickets") or []:
            pay = t.get("payout")
            if not isinstance(pay, dict):
                continue
            try:
                ev = float(pay.get("ev"))
            except (TypeError, ValueError):
                continue
            if not math.isfinite(ev):
                continue
            total_considered += 1
            if ev > 1.5:
                strong_c += 1
            elif ev > 1.0:
                ok_c += 1
            elif ev > 0.8:
                marg_c += 1
            else:
                skip_c += 1
            if ev > best_ev:
                best_ev = ev
                try:
                    p_all = float(pay.get("p_all_win", 0))
                except (TypeError, ValueError):
                    p_all = 0.0
                try:
                    sp = float(
                        pay.get("sweep_payout", pay.get("first_place", 0))
                    )
                except (TypeError, ValueError):
                    sp = 0.0
                try:
                    mg = float(pay.get("payout", pay.get("min_guarantee", 0)))
                except (TypeError, ValueError):
                    mg = 0.0
                tno = t.get("ticket_no")
                nm = f"{gname} #{tno}".strip() if tno is not None else gname
                best_ticket = {
                    "name": nm,
                    "ev": round(ev, 4),
                    "recommendation": str(pay.get("recommendation") or ""),
                    "payout": round(mg, 4),
                    "sweep_payout": f"{sp:.1f}x",
                    "p_all_win": f"{p_all * 100:.1f}%",
                }

    return {
        "date": date_str,
        "total_tickets": total_considered,
        "strong_count": strong_c,
        "ok_count": ok_c,
        "marginal_count": marg_c,
        "skip_count": skip_c,
        "best_ticket": best_ticket,
    }


@app.get("/api/ticket-ev-summary")
def api_ticket_ev_summary():
    """Empirical EV summary from tickets_latest.json payout fields."""
    json_path = TEMPLATES_DIR / "tickets_latest.json"
    if not _template_json_available("tickets_latest.json"):
        return jsonify({"error": "tickets_latest.json not found"}), 404
    try:
        data = read_json_cached(json_path)
    except Exception as e:
        return jsonify({"error": str(e)}), 500
    out = _ticket_ev_summary_from_payload(data if isinstance(data, dict) else None)
    r = jsonify(out)
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


def _ticket_ev_path() -> Path:
    return OUTPUTS_ROOT / "ticket_ev_latest.json"


def _ev_top20_latest_path() -> Path:
    return TEMPLATES_DIR / "ev_top20_latest.json"


@app.get("/api/tickets-ev-top20")
def api_tickets_ev_top20():
    """
    Top EV slips: prefer templates/ev_top20_latest.json (build_ultimate_tickets.py),
    else outputs/ticket_ev_latest.json (legacy combined path).
    """
    ultimate_path = _ev_top20_latest_path()
    if ultimate_path.is_file():
        try:
            data = read_json_cached(ultimate_path)
        except Exception as e:
            return jsonify({"error": str(e), "tickets": []}), 500
        tickets_in = list((data or {}).get("tickets") or [])
        out = []
        for t in tickets_in:
            pw = t.get("p_win_pct")
            try:
                est = float(pw) / 100.0 if pw is not None and float(pw) > 1.0 else float(pw or 0)
            except (TypeError, ValueError):
                est = None
            legs = [str(x) for x in (t.get("legs") or [])]
            out.append(
                {
                    "ticket_no": t.get("rank"),
                    "n_legs": t.get("n_legs"),
                    "est_win_prob": est,
                    "power_payout": t.get("payout"),
                    "ev_power": t.get("ev"),
                    "sports": t.get("sports") or [],
                    "recommendation": str(t.get("recommendation") or "").upper(),
                    "correlation_flag": "",
                    "legs": legs,
                    "n_goblins": t.get("n_goblins"),
                    "n_demons": t.get("n_demons"),
                    "pick_types": t.get("pick_types"),
                    "min_guarantee": t.get("min_guarantee"),
                }
            )
        r = jsonify(
            {
                "date": (data or {}).get("date"),
                "generated_at": (data or {}).get("generated_at"),
                "group_name": f"ULTIMATE TOP20 ({(data or {}).get('mode') or 'balanced'})",
                "mode": (data or {}).get("mode"),
                "total_combos_evaluated": (data or {}).get("total_combos_evaluated"),
                "tickets": out,
            }
        )
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        r.headers["Pragma"] = "no-cache"
        return r

    path = _ticket_ev_path()
    if not (_template_json_available("ticket_ev_latest.json") or path.exists()):
        return jsonify({"error": "ticket_ev_latest.json not found", "tickets": []}), 404
    try:
        data = read_json_cached(path)
    except Exception as e:
        return jsonify({"error": str(e), "tickets": []}), 500

    groups = list((data or {}).get("groups") or [])
    if not groups:
        return jsonify(
            {
                "date": (data or {}).get("date"),
                "generated_at": (data or {}).get("generated_at"),
                "group_name": "",
                "tickets": [],
            }
        )

    grp = groups[0]
    tickets = list(grp.get("tickets") or [])
    tickets = sorted(
        tickets,
        key=lambda t: float(t.get("ev_power") or 0.0),
        reverse=True,
    )[:20]

    out = []
    for t in tickets:
        legs = [str((lg or {}).get("label") or "") for lg in (t.get("legs") or [])]
        out.append(
            {
                "ticket_no": t.get("ticket_no"),
                "n_legs": t.get("n_legs"),
                "est_win_prob": t.get("est_win_prob"),
                "power_payout": t.get("power_payout"),
                "ev_power": t.get("ev_power"),
                "sports": t.get("sports") or [],
                "recommendation": t.get("recommendation") or "",
                "correlation_flag": t.get("correlation_flag") or "",
                "legs": legs,
            }
        )

    r = jsonify(
        {
            "date": (data or {}).get("date"),
            "generated_at": (data or {}).get("generated_at"),
            "group_name": grp.get("group_name") or "TOP20 Exact EV",
            "tickets": out,
        }
    )
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


@app.get("/ev-top20")
def ev_top20_redirect():
    """Legacy bookmark fallback after EV TOP20 nav removal."""
    return redirect("/tickets", code=302)


@app.get("/api/slate/today-tickets")
def api_slate_today_tickets():
    """Tickets from tickets_latest.json for /payout import (pre-fills leg std/played lines)."""
    if not _template_json_available("tickets_latest.json"):
        return jsonify({"error": "tickets_latest.json not found", "tickets": []}), 404
    try:
        data = read_json_cached(TEMPLATES_DIR / "tickets_latest.json")
    except Exception as e:
        return jsonify({"error": str(e), "tickets": []}), 500
    et = _eastern_today_ymd()
    d_sl = str((data or {}).get("date") or "").strip()[:10]
    tickets_out: list[dict] = []
    for grp in (data or {}).get("groups") or []:
        sheet = str(grp.get("group_name") or "")
        for t in grp.get("tickets") or []:
            legs_in = t.get("legs") or []
            legs = []
            for lg in legs_in:
                std = lg.get("standard_line")
                played = lg.get("line")
                pt = str(lg.get("pick_type") or "Standard")
                dp = lg.get("delta_pct")
                legs.append(
                    {
                        "player": lg.get("player"),
                        "sport": lg.get("sport"),
                        "prop_type": lg.get("prop_type"),
                        "pick_type": pt,
                        "std_line": std,
                        "played_line": played,
                        "delta_pct": dp,
                        "direction": lg.get("direction"),
                        "ml_prob": lg.get("ml_prob"),
                    }
                )
            tickets_out.append(
                {
                    "ticket_id": f"{sheet}#{t.get('ticket_no')}",
                    "sheet": sheet,
                    "ticket_no": t.get("ticket_no"),
                    "n_legs": len(legs),
                    "legs": legs,
                    "est_mult": t.get("est_multiplier"),
                    "flat_mult": t.get("flat_multiplier"),
                    "combined_prob": t.get("combined_hit_prob_curve"),
                    "est_ev": t.get("est_ev"),
                    "payout": t.get("payout"),
                }
            )
    r = jsonify(
        {
            "date": d_sl,
            "eastern_today": et,
            "tickets": tickets_out,
        }
    )
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return r


@app.get("/api/slate-legs")
def api_slate_legs():
    """Flatten today's tickets_latest.json legs and include ticket-grouped options."""
    if not _template_json_available("tickets_latest.json"):
        return jsonify({"error": "tickets_latest.json not found", "legs": [], "tickets": []}), 404
    try:
        data = read_json_cached(TEMPLATES_DIR / "tickets_latest.json")
    except Exception as e:
        return jsonify({"error": str(e), "legs": [], "tickets": []}), 500

    groups = list((data or {}).get("groups") or [])
    legs: list[dict[str, Any]] = []
    tickets_out: list[dict[str, Any]] = []
    seen: set[str] = set()
    for grp in groups:
        group_name = str((grp or {}).get("group_name") or "")
        for t in (grp or {}).get("tickets") or []:
            tno = t.get("ticket_no")
            ticket_key = f"{group_name}#{tno}"
            t_legs: list[dict[str, Any]] = []
            for leg in (t or {}).get("legs") or []:
                item = {
                    "player": str(leg.get("player") or ""),
                    "sport": str(leg.get("sport") or ""),
                    "team": str(leg.get("team") or ""),
                    "opp": str(leg.get("opp") or ""),
                    "prop_type": str(leg.get("prop_type") or ""),
                    "pick_type": str(leg.get("pick_type") or ""),
                    "direction": str(leg.get("direction") or ""),
                    "line": leg.get("line"),
                    "standard_line": leg.get("standard_line") if leg.get("standard_line") is not None else leg.get("line"),
                    "hit_rate": leg.get("hit_rate"),
                    "ml_prob": leg.get("ml_prob"),
                    "group_name": group_name,
                    "ticket_no": tno,
                    "ticket_key": ticket_key,
                }
                t_legs.append(item)
                key = (
                    f"{item['player']}|{item['prop_type']}|{item['line']}|{item['direction']}"
                )
                if key not in seen:
                    seen.add(key)
                    legs.append(item)
            tickets_out.append(
                {
                    "group_name": group_name,
                    "ticket_no": tno,
                    "ticket_key": ticket_key,
                    "n_legs": len(t_legs),
                    "legs": t_legs,
                }
            )

    r = jsonify(
        {
            "legs": legs,
            "tickets": tickets_out,
            "built_at": str((data or {}).get("built_at") or (data or {}).get("generated_at") or ""),
            "date": str((data or {}).get("date") or ""),
        }
    )
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


@app.post("/api/payout/log-observation")
def api_payout_log_observation():
    """Append one row to payout_observations.csv (server-side curve learning; volume when mounted)."""
    body = request.get_json(silent=True) or {}
    csv_path = PAYOUT_OBS_PATH
    fieldnames = [
        "date",
        "n_legs",
        "slip_type",
        "combo_label",
        "base_mult",
        "est_mult",
        "actual_mult",
        "mult_delta",
        "stake",
        "actual_payout",
        "g_exp_used",
        "d_exp_used",
        "d_scale_used",
        "leg_details_json",
    ]
    try:
        from utils.goblin_demon_multiplier import load_params as _gd_load

        prm = _gd_load()
    except Exception:
        prm = {"G_EXP": 1.0, "D_EXP": 1.5, "D_SCALE": 3.0}

    def _f(name, default=""):
        v = body.get(name)
        return default if v is None else v

    try:
        actual_mult = float(_f("actual_mult", 0) or 0)
    except (TypeError, ValueError):
        return jsonify({"saved": False, "error": "invalid actual_mult"}), 400
    if actual_mult <= 0:
        return jsonify({"saved": False, "error": "actual_mult required"}), 400

    try:
        base_mult = float(_f("base_mult", 0) or 0)
        est_mult = float(_f("est_mult", 0) or 0)
    except (TypeError, ValueError):
        base_mult, est_mult = 0.0, 0.0
    mult_delta = round(actual_mult - est_mult, 4) if est_mult else ""

    leg_json = _f("leg_details_json", "")
    if isinstance(leg_json, (list, dict)):
        leg_json = json.dumps(leg_json, ensure_ascii=False)

    row = {
        "date": str(_f("date", "")).strip(),
        "n_legs": str(_f("n_legs", "")).strip(),
        "slip_type": str(_f("slip_type", "power")).strip().lower(),
        "combo_label": str(_f("combo_label", "")).strip(),
        "base_mult": base_mult,
        "est_mult": est_mult,
        "actual_mult": actual_mult,
        "mult_delta": mult_delta,
        "stake": str(_f("stake", "")).strip(),
        "actual_payout": str(_f("actual_payout", "")).strip(),
        "g_exp_used": prm.get("G_EXP", ""),
        "d_exp_used": prm.get("D_EXP", ""),
        "d_scale_used": prm.get("D_SCALE", ""),
        "leg_details_json": leg_json if isinstance(leg_json, str) else "",
    }

    csv_path.parent.mkdir(parents=True, exist_ok=True)
    file_exists = csv_path.exists()
    with csv_path.open("a", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=fieldnames)
        if not file_exists:
            w.writeheader()
        w.writerow(row)

    # Hand-logged payout rows used by manual screenshot calibration workflow.
    hand_csv_path = PAYOUT_LOG_PATH
    hand_fields = [
        "date",
        "group_name",
        "n_legs",
        "pick_types",
        "lines",
        "standard_lines",
        "actual_payout_multiplier",
        "slip_type",
        "result",
    ]
    hand_row = {
        "date": str(_f("date", "")).strip(),
        "group_name": str(_f("group_name", "")).strip(),
        "n_legs": str(_f("n_legs", "")).strip(),
        "pick_types": str(_f("pick_types", "")).strip(),
        "lines": str(_f("lines", "")).strip(),
        "standard_lines": str(_f("standard_lines", "")).strip(),
        "actual_payout_multiplier": actual_mult,
        "slip_type": str(_f("slip_type", "power")).strip().lower(),
        "result": str(_f("result", "pending")).strip().upper() or "PENDING",
    }
    hand_csv_path.parent.mkdir(parents=True, exist_ok=True)
    hand_exists = hand_csv_path.exists()
    with hand_csv_path.open("a", newline="", encoding="utf-8") as f2:
        w2 = csv.DictWriter(f2, fieldnames=hand_fields)
        if not hand_exists:
            w2.writeheader()
        w2.writerow(hand_row)

    n_lines = sum(1 for _ in csv_path.open("r", encoding="utf-8")) - 1
    hand_lines = sum(1 for _ in hand_csv_path.open("r", encoding="utf-8")) - 1
    warn = bool(est_mult and abs(float(actual_mult) - float(est_mult)) > 1.5)
    csv_row = ",".join(
        [
            hand_row["date"],
            hand_row["group_name"],
            hand_row["n_legs"],
            hand_row["pick_types"],
            hand_row["lines"],
            hand_row["standard_lines"],
            str(hand_row["actual_payout_multiplier"]),
            hand_row["slip_type"],
            hand_row["result"],
        ]
    )
    return jsonify(
        {
            "status": "ok",
            "saved": True,
            "total_obs": max(0, n_lines),
            "total_hand_obs": max(0, hand_lines),
            "mult_delta": mult_delta,
            "warning_large_delta": warn,
            "csv_row": csv_row,
            "message": "Row appended to payout_log_hand.csv (sync from Railway via PROPORACLE_PAYOUT_EXPORT_URL in run_daily if configured)",
        }
    )


@app.get("/api/payout/export-log-hand")
def api_payout_export_log_hand():
    """Download payout_log_hand.csv (manual / LOG THIS TICKET rows) from the persistent data root."""
    if PAYOUT_LOG_PATH.is_file():
        return send_file(
            str(PAYOUT_LOG_PATH),
            mimetype="text/csv",
            as_attachment=True,
            download_name="payout_log_hand.csv",
        )
    return jsonify({"error": "No payout_log_hand.csv yet"}), 404


@app.get("/api/payout/export-observations")
def api_payout_export_observations():
    """Download payout_observations.csv (full observation log) from the persistent data root."""
    if PAYOUT_OBS_PATH.is_file():
        return send_file(
            str(PAYOUT_OBS_PATH),
            mimetype="text/csv",
            as_attachment=True,
            download_name="payout_observations.csv",
        )
    return jsonify({"error": "No payout_observations.csv yet"}), 404


@app.get("/api/payout/observations")
def api_payout_observations():
    """Read server-side payout observations for /payout Patterns (payout_observations.csv)."""
    csv_path = PAYOUT_OBS_PATH
    if not csv_path.is_file():
        r = jsonify({"observations": [], "count": 0, "truncated": False})
        r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
        return r
    rows: list[dict[str, Any]] = []
    try:
        with csv_path.open("r", newline="", encoding="utf-8") as f:
            reader = csv.DictReader(f)
            for rec in reader:
                if isinstance(rec, dict):
                    rows.append({str(k): (v if v is not None else "") for k, v in rec.items()})
    except OSError as e:
        return jsonify({"error": str(e), "observations": [], "count": 0}), 500
    max_return = 800
    truncated = len(rows) > max_return
    out = rows[-max_return:] if truncated else rows
    resp = jsonify({"observations": out, "count": len(rows), "truncated": truncated})
    resp.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    return resp


@app.get("/api/payout/combo-table")
def api_payout_combo_table():
    """Static combo reference from outputs/combo_table_latest.json (run write_combo_table_latest.py)."""
    p = BASE_DIR / "outputs" / "combo_table_latest.json"
    if not p.is_file():
        return jsonify(
            {
                "error": "combo_table_latest.json missing — run scripts/write_combo_table_latest.py",
                "leg_counts": [],
                "combos": [],
            }
        ), 404
    try:
        data = json.loads(p.read_text(encoding="utf-8"))
    except Exception as e:
        return jsonify({"error": str(e), "leg_counts": [], "combos": []}), 500
    return jsonify(data)


def _eastern_today_ymd() -> str:
    """Calendar date in America/New_York (US slate day for NBA/CBB)."""
    return datetime.now(ZoneInfo("America/New_York")).date().strftime("%Y-%m-%d")


@app.get("/api/slate-display-date")
def api_slate_display_date():
    """
    YYYY-MM-DD for the home/nav slate chip.

    Prefer tickets_latest.json — it drives /tickets and /api/slate. Using max() across every
    JSON allowed ticket_eval or slate_latest to advertise a *later* day (e.g. 4/4) while the
    main slip was still 4/3.

    If tickets_latest is missing or has no date, use the newest candidate that is not after
    US Eastern \"today\"; if all are future (bad data), fall back to max(candidates).
    """
    candidates: list[str] = []
    tickets_date: str | None = None
    for name in ("tickets_latest.json", "slate_latest.json", "ticket_eval_slate_latest.json"):
        if not _template_json_available(name):
            continue
        try:
            data = read_json_cached(TEMPLATES_DIR / name)
            ds = str((data or {}).get("date") or "").strip()[:10]
            if len(ds) == 10 and ds[4] == "-" and ds[7] == "-":
                candidates.append(ds)
                if name == "tickets_latest.json":
                    tickets_date = ds
        except Exception:
            continue
    try:
        sp = _selected_slate_sport_payload()
        ds = str((sp or {}).get("date") or "").strip()[:10]
        if len(ds) == 10:
            candidates.append(ds)
    except Exception:
        pass

    et_today = _eastern_today_ymd()
    if tickets_date:
        best = tickets_date
    else:
        not_future = [c for c in candidates if c <= et_today]
        best = max(not_future) if not_future else (max(candidates) if candidates else None)
    r = jsonify({"date": best})
    r.headers["Cache-Control"] = "no-store, no-cache, must-revalidate, max-age=0"
    r.headers["Pragma"] = "no-cache"
    return r


def _side_hit_count_for_slate_picks(raw: object, n: int) -> int | None:
    """Match UI streakHits: rate in [0,1] or integer count in last n games."""
    if raw is None:
        return None
    try:
        x = float(raw)
    except (TypeError, ValueError):
        return None
    if x != x:  # NaN
        return None
    if x <= 1.0:
        return max(0, min(n, int(round(x * n))))
    return max(0, min(n, int(round(x))))


def _player_initials_from_name(player: object) -> str:
    s = str(player or "").strip()
    if not s:
        return ""
    parts = s.split()
    if len(parts) >= 2 and parts[0] and parts[1]:
        return (parts[0][0] + parts[1][0]).upper()
    return s[:2].upper()


def _extract_history_series(row: dict[str, Any]) -> tuple[list[float], list[float]]:
    """Return up-to-10 game actual/line history from known row key patterns."""
    actual_candidates = ("stat_g", "actual_g", "g")
    line_candidates = ("line_g", "prop_line_g")
    actual_vals: list[float] = []
    line_vals: list[float] = []

    for i in range(1, 11):
        aval: float | None = None
        for pref in actual_candidates:
            v = row.get(f"{pref}{i}")
            if v is None:
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            if math.isnan(fv) or math.isinf(fv):
                continue
            aval = fv
            break
        if aval is not None:
            actual_vals.append(aval)

        lval: float | None = None
        for pref in line_candidates:
            v = row.get(f"{pref}{i}")
            if v is None:
                continue
            try:
                fv = float(v)
            except (TypeError, ValueError):
                continue
            if math.isnan(fv) or math.isinf(fv):
                continue
            lval = fv
            break
        if lval is not None:
            line_vals.append(lval)

    return actual_vals, line_vals


def _normalize_prop_merge_key(raw: object) -> str:
    """Align with dashboard merge keys (sheet shortcodes vs ticket long names)."""
    t = (
        str(raw or "")
        .strip()
        .lower()
        .replace(" ", "")
        .replace("\u2013", "-")
        .replace("\u2014", "-")
    )
    canon = {
        "pra": "pts+rebs+asts",
        "pr": "pts+rebs",
        "pa": "pts+asts",
        "pts": "points",
        "ast": "assists",
        "reb": "rebounds",
        "points": "points",
        "rebounds": "rebounds",
        "assists": "assists",
        "pts+rebs": "pts+rebs",
        "pts+asts": "pts+asts",
        "pts+rebs+asts": "pts+rebs+asts",
        "points+rebounds": "pts+rebs",
        "points+assists": "pts+asts",
        "points+rebounds+assists": "pts+rebs+asts",
    }
    return canon.get(t, t)


def _normalize_player_merge_key(name: object) -> str:
    return " ".join(str(name or "").strip().lower().split())


def _normalize_dir_merge(raw: object) -> str:
    s = str(raw or "OVER").strip().upper()
    if s in ("O", "OV", "OVR"):
        return "OVER"
    if s in ("U", "UN", "UND"):
        return "UNDER"
    return s


def _line_tokens_from_record(row: dict[str, Any]) -> list[str]:
    toks: list[str] = []
    seen: set[str] = set()
    for key in ("line", "standard_line", "book_line", "prop_line"):
        v = row.get(key)
        if v is None or v == "":
            continue
        try:
            f = float(v)
            if math.isfinite(f):
                s = f"{f:.3f}"
                if s not in seen:
                    seen.add(s)
                    toks.append(s)
                continue
        except (TypeError, ValueError):
            pass
        s2 = str(v).strip()
        if s2 and s2 not in seen:
            seen.add(s2)
            toks.append(s2)
    return toks if toks else [""]


def _normalize_api_slate_sport_query(raw: str | None) -> str:
    """Uppercase sport key for ?sport= on /api/slate (case-insensitive input)."""
    return (raw or "").strip().upper()


def _apply_sport_filter_to_slate_payload(
    payload: dict[str, Any], sport_filter: str
) -> dict[str, Any]:
    if not sport_filter:
        return payload
    picks = payload.get("picks")
    if not isinstance(picks, list):
        return payload
    filtered = [
        p
        for p in picks
        if isinstance(p, dict)
        and str(p.get("sport") or "").strip().upper() == sport_filter
    ]
    out = {**payload, "picks": filtered}
    out["sport_filter"] = sport_filter
    return out


def _api_slate_pick_moat_fields(r: dict[str, Any]) -> dict[str, Any]:
    """Expose rank tier, ML prob, matchup, and book context on /api/slate picks."""

    def fnum(x: Any) -> float | None:
        if x is None or x == "":
            return None
        try:
            v = float(x)
            if math.isnan(v) or math.isinf(v):
                return None
            return v
        except (TypeError, ValueError):
            return None

    def fstr(x: Any) -> str | None:
        if x is None:
            return None
        s = str(x).strip()
        return s if s else None

    opp_raw = r.get("opp")
    if opp_raw is None or (isinstance(opp_raw, str) and not opp_raw.strip()):
        opp_raw = r.get("opp_team")

    def_rank: float | None = None
    for k in (
        "opponent_def_rank",
        "opp_def_rank",
        "OVERALL_DEF_RANK",
        "def_rank",
        "ncaa_rank",
    ):
        if k in r:
            def_rank = fnum(r.get(k))
            if def_rank is not None:
                break

    bcl = r.get("best_cross_line")
    book_line = fnum(bcl if bcl not in (None, "") else r.get("book_line"))
    line_plain = fnum(r.get("line"))
    prop_line = fnum(r.get("prop_line"))
    if prop_line is None:
        prop_line = line_plain

    gt = fstr(r.get("game_time") or r.get("event_start_time"))

    return {
        "team": fstr(r.get("team")),
        "opp": fstr(opp_raw),
        "ml_prob": fnum(r.get("ml_prob")),
        "tier": fstr(r.get("tier")),
        "rank": fnum(r.get("rank_score")),
        "def_tier": fstr(r.get("def_tier") or r.get("Def Tier")),
        "opponent_def_rank": def_rank,
        "book_line": book_line,
        "prop_line": prop_line,
        "game_time": gt,
        "image_url": fstr(r.get("image_url")),
        "injury_status": fstr(r.get("injury_status")),
        "pick_platform": fstr(r.get("pick_platform")),
        "leg_prob_used": fnum(r.get("leg_prob_used")),
        "leg_prob_source": fstr(r.get("leg_prob_source")),
    }


def _pick_scalar_history_fields(leg: dict[str, Any]) -> dict[str, Any]:
    """Expose per-game columns + standard_projection on /api/slate picks for client merge/charts."""
    out: dict[str, Any] = {}
    sp = leg.get("standard_projection")
    if sp is not None and sp != "":
        out["standard_projection"] = sp
    sl = leg.get("standard_line")
    if sl is not None and sl != "":
        out["standard_line"] = sl
    for i in range(1, 11):
        for suffix in (f"g{i}", f"stat_g{i}"):
            if suffix in leg and leg[suffix] is not None and leg[suffix] != "":
                out[suffix] = leg[suffix]
    return out


def _picks_payload_from_slate_latest() -> dict[str, Any] | None:
    """
    Home /api/slate fallback when tickets_latest.json is missing or has no leg rows.
    Uses the same slate_latest.json as /api/slate-sport (pipeline output).
    """
    if not _template_json_available("slate_latest.json"):
        return None
    try:
        data = read_json_cached(TEMPLATES_DIR / "slate_latest.json")
    except Exception:
        return None
    if not isinstance(data, dict):
        return None
    sports = data.get("sports") or {}
    if not isinstance(sports, dict):
        return None
    seen: set[tuple[Any, ...]] = set()
    picks: list[dict[str, Any]] = []
    for raw_key, rows in sports.items():
        lk = str(raw_key).strip().lower()
        if lk in DISABLED_SPORTS:
            continue
        if not isinstance(rows, list):
            continue
        sport_label = str(raw_key).strip().upper()
        for row in rows:
            if not isinstance(row, dict):
                continue
            player = row.get("player") or ""
            prop = row.get("prop") or row.get("prop_type") or ""
            dirv = row.get("dir") or row.get("direction") or "OVER"
            line = row.get("line")
            key = (player, prop, dirv, line)
            if key in seen:
                continue
            seen.add(key)
            l5_over = row.get("l5_over")
            l5_under = row.get("l5_under")
            if l5_under is None and l5_over is not None:
                ho = _side_hit_count_for_slate_picks(l5_over, 5)
                if ho is not None:
                    l5_under = 5 - ho
            l10_over = row.get("l10_over")
            l10_under = row.get("l10_under")
            if l10_under is None and l10_over is not None:
                ho = _side_hit_count_for_slate_picks(l10_over, 10)
                if ho is not None:
                    l10_under = 10 - ho
            try:
                hr = float(row.get("hit_rate") or 0.0)
            except (TypeError, ValueError):
                hr = 0.0
            try:
                edge = float(row.get("edge") or 0.0)
            except (TypeError, ValueError):
                edge = 0.0
            actual_series, line_series = _extract_history_series(row)
            pick_row: dict[str, Any] = {
                "sport": sport_label,
                "initials": _player_initials_from_name(player),
                "player": player,
                "prop": prop,
                "line": line,
                "pick": row.get("pick_type") or "Standard",
                "dir": str(dirv).strip().upper() or "OVER",
                "hit": round(hr * 100),
                "edge": edge,
                "projection": _pick_projection_from_mapping(row),
                "l5_over": l5_over,
                "l5_under": l5_under,
                "l10_over": l10_over,
                "l10_under": l10_under,
                "l5_avg": row.get("l5_avg"),
                "season_avg": row.get("season_avg") or row.get("szn_avg"),
                "actual_series": actual_series,
                "line_series": line_series,
            }
            pick_row.update(_api_slate_pick_moat_fields(row))
            picks.append(pick_row)
    if not picks:
        return None
    picks.sort(key=lambda p: _api_slate_pick_abs_edge(p), reverse=True)
    # Cap: full slate can be 10k+ rows; hero table + edges only need top props by |edge|.
    _max = 2500
    if len(picks) > _max:
        picks = picks[:_max]
    return {
        "picks": picks,
        "generated_at": data.get("generated_at"),
        "date": data.get("date"),
        "source": "slate_latest",
    }


# API: Slate picks — deduped legs from tickets_latest.json, else rows from slate_latest.json
@app.get("/api/slate")
def api_slate():
    tickets_path = TEMPLATES_DIR / "tickets_latest.json"
    sport_q = _normalize_api_slate_sport_query(request.args.get("sport"))

    def _build_picks():
        slate_history_map: dict[tuple[str, str, str, str], tuple[list[float], list[float]]] = {}
        if _template_json_available("slate_latest.json"):
            try:
                slate_data = read_json_cached(TEMPLATES_DIR / "slate_latest.json")
                sports = (slate_data or {}).get("sports") or {}
                if isinstance(sports, dict):
                    for raw_sport, rows in sports.items():
                        if not isinstance(rows, list):
                            continue
                        sport_norm = str(raw_sport or "").strip().upper()
                        for row in rows:
                            if not isinstance(row, dict):
                                continue
                            player_norm = _normalize_player_merge_key(row.get("player"))
                            prop_norm = _normalize_prop_merge_key(
                                row.get("prop") or row.get("prop_type")
                            )
                            dir_norm = _normalize_dir_merge(
                                row.get("dir") or row.get("direction")
                            )
                            actual_series, line_series = _extract_history_series(row)
                            if not actual_series:
                                continue
                            for lt in _line_tokens_from_record(row):
                                key = (sport_norm, player_norm, prop_norm, f"{dir_norm}|{lt}")
                                if key in slate_history_map:
                                    continue
                                slate_history_map[key] = (actual_series, line_series)
            except Exception:
                slate_history_map = {}

        if not _template_json_available("tickets_latest.json"):
            base: dict[str, Any] = {"picks": [], "generated_at": None, "date": None}
        else:
            data = read_json_cached(tickets_path)
            seen = set()
            picks: list[dict[str, Any]] = []
            for group in (data.get("groups") or []):
                for ticket in (group.get("tickets") or []):
                    for leg in (ticket.get("legs") or []):
                        key = (leg.get("player"), leg.get("prop_type"), leg.get("direction"), leg.get("line"))
                        if key in seen:
                            continue
                        seen.add(key)
                        l5_over = leg.get("l5_over")
                        l5_under = leg.get("l5_under")
                        if l5_under is None and l5_over is not None:
                            ho = _side_hit_count_for_slate_picks(l5_over, 5)
                            if ho is not None:
                                l5_under = 5 - ho
                        l10_over = leg.get("l10_over")
                        l10_under = leg.get("l10_under")
                        if l10_under is None and l10_over is not None:
                            ho = _side_hit_count_for_slate_picks(l10_over, 10)
                            if ho is not None:
                                l10_under = 10 - ho
                        actual_series, line_series = _extract_history_series(leg)
                        if not actual_series and slate_history_map:
                            sport_norm = str(leg.get("sport") or "").strip().upper()
                            player_norm = _normalize_player_merge_key(leg.get("player"))
                            prop_norm = _normalize_prop_merge_key(leg.get("prop_type"))
                            dir_norm = _normalize_dir_merge(leg.get("direction"))
                            for lt in _line_tokens_from_record(leg):
                                hist_key = (
                                    sport_norm,
                                    player_norm,
                                    prop_norm,
                                    f"{dir_norm}|{lt}",
                                )
                                backfill = slate_history_map.get(hist_key)
                                if backfill:
                                    actual_series, line_series = backfill
                                    break
                        abs_edge_leg = _api_slate_pick_abs_edge(leg)
                        pick_entry: dict[str, Any] = {
                            "sport": leg.get("sport", ""),
                            "initials": leg.get("initials", ""),
                            "player": leg.get("player", ""),
                            "prop": leg.get("prop_type", ""),
                            "line": leg.get("line", 0),
                            "pick": leg.get("pick_type", "Standard"),
                            "dir": leg.get("direction", "OVER"),
                            "hit": round((leg.get("hit_rate") or 0) * 100),
                            "edge": leg.get("edge") or 0,
                            "abs_edge": abs_edge_leg,
                            "projection": _pick_projection_from_mapping(leg),
                            "l5_over": l5_over,
                            "l5_under": l5_under,
                            "l10_over": l10_over,
                            "l10_under": l10_under,
                            "l5_avg": leg.get("l5_avg"),
                            "season_avg": leg.get("season_avg"),
                            "actual_series": actual_series,
                            "line_series": line_series,
                        }
                        pick_entry.update(_api_slate_pick_moat_fields(leg))
                        pick_entry.update(_pick_scalar_history_fields(leg))
                        picks.append(pick_entry)
            picks.sort(key=lambda p: _api_slate_pick_abs_edge(p), reverse=True)
            base = {
                "picks": picks,
                "generated_at": data.get("generated_at"),
                "date": data.get("date"),
            }
        if base.get("picks"):
            base["source"] = "tickets_latest"
            return base
        fb = _picks_payload_from_slate_latest()
        if fb:
            return fb
        base.setdefault("source", None)
        return base

    def _build_filtered():
        return _apply_sport_filter_to_slate_payload(_build_picks(), sport_q)

    try:
        return _gz_json_response(
            f"slate-picks-v3-tickets-or-slate:{sport_q or 'all'}:{_explorer_json_gz_bust_token()}",
            _build_filtered,
            ttl=_PIPELINE_JSON_TTL,
        )
    except Exception as e:
        return jsonify({"error": str(e), "picks": []}), 500


# ──────────────────────────────────────────────────────────────────────────────
# API: Full per-sport slate from combined Excel (openpyxl, no pandas needed)
# ──────────────────────────────────────────────────────────────────────────────
# ──────────────────────────────────────────────────────────────────────────────
# API: Full per-sport slate from slate_latest.json (written by pipeline)
# ──────────────────────────────────────────────────────────────────────────────
@app.get("/api/slate-sport")
def api_slate_sport():
    if not (
        _template_json_available("slate_latest.json")
        or _template_json_available("ticket_eval_slate_latest.json")
    ):
        return jsonify(
            {
                "error": "No slate JSON — run pipeline (slate_latest.json) or build_ticket_eval.py (ticket_eval_slate_latest.json)",
                "sports": {},
            }
        ), 404
    try:
        _selected_slate_sport_payload()
    except ValueError as e:
        return jsonify({"error": str(e), "sports": {}}), 404
    try:
        return _gz_json_response(
            f"slate-sport-slim-v2:{_explorer_json_gz_bust_token()}",
            lambda: _slim_slate_sport_payload(_selected_slate_sport_payload()),
            ttl=_PIPELINE_JSON_TTL,
        )
    except Exception as e:
        return jsonify({"error": str(e), "sports": {}}), 500


@app.get("/api/slate-sport/<sport>")
def api_slate_sport_single(sport: str):
    """
    Per-sport lazy slate endpoint for the home explorer.
    Uses the same source as /api/slate-sport (slate_latest.json or ticket_eval per SLATE_SPORT_SOURCE).
    sport=combined returns the merged Full Slate (all sports).
    """
    if not (
        _template_json_available("slate_latest.json")
        or _template_json_available("ticket_eval_slate_latest.json")
    ):
        return jsonify(
            {
                "error": "No slate JSON — run pipeline (slate_latest.json) or build_ticket_eval.py",
                "sport": str(sport or "").strip().lower(),
                "rows": [],
            }
        ), 404

    sport_key = str(sport or "").strip().lower()
    if not sport_key:
        return jsonify({"error": "missing sport key", "sport": sport_key, "rows": []}), 400

    try:
        _selected_slate_sport_payload()
    except ValueError as e:
        return jsonify({"error": str(e), "sport": sport_key, "rows": []}), 404

    def _build():
        payload = _selected_slate_sport_payload()
        sports = (payload or {}).get("sports") or {}
        if sport_key == "combined":
            slim_rows = _merged_combined_slim_rows(payload)
            return {
                "date": payload.get("date"),
                "generated_at": payload.get("generated_at"),
                "sport": sport_key,
                "rows": slim_rows,
            }
        rows = sports.get(sport_key) or []
        # Match existing UI behavior: CBB card combines cbb + wcbb rows.
        if sport_key == "cbb":
            rows = list(rows) + list(sports.get("wcbb") or [])
        if not isinstance(rows, list):
            rows = []
        slim_rows = [_slim_slate_sport_row(r) if isinstance(r, dict) else r for r in rows]
        if not slim_rows and sport_key == "wnba":
            slim_rows = _wnba_slate_rows_from_step8_fallback()
        slim_rows = _filter_invalid_demon_slate_rows(slim_rows)
        return {
            "date": payload.get("date"),
            "generated_at": payload.get("generated_at"),
            "sport": sport_key,
            "rows": slim_rows,
        }

    try:
        return _gz_json_response(
            f"slate-sport-single-v1:{sport_key}:{_explorer_json_gz_bust_token()}",
            _build,
            ttl=60.0,
        )
    except Exception as e:
        return jsonify({"error": str(e), "sport": sport_key, "rows": []}), 500


@app.get("/api/slate-excel")
def api_slate_excel():
    """Return all sheets from the combined Excel with non-blank columns only.

    Only sheets that exist in the workbook are returned; missing names (e.g. NFL Slate
    before the pipeline emits that tab) are skipped without error.
    """
    import openpyxl

    def _build():
        path, run_date = _find_combined_excel()
        if path is None:
            return {"error": "combined slate Excel not found", "sheets": {}, "date": None}
        sheet_map = {
            "Full Slate":  "combined",
            "NBA Slate":   "nba",
            "NBA1H Slate": "nba1h",
            "NBA1Q Slate": "nba1q",
            "NHL Slate":   "nhl",
            "Soccer Slate":"soccer",
            "Tennis Slate": "tennis",
            "WNBA Slate":  "wnba",
            "WCBB Slate":  "wcbb",
            "MLB Slate":   "mlb",
            "NFL Slate":   "nfl",
        }
        wb = openpyxl.load_workbook(str(path), read_only=True, data_only=True)
        result: dict[str, Any] = {}
        for sheet_name, key in sheet_map.items():
            if sheet_name not in wb.sheetnames:
                continue
            ws = wb[sheet_name]
            all_rows = list(ws.iter_rows(values_only=True))
            if not all_rows:
                continue
            headers = [str(h) if h is not None else "" for h in all_rows[0]]
            data_rows = all_rows[1:]
            # keep only columns with ≥1 non-null value in first 100 rows
            non_blank = [
                i for i, h in enumerate(headers)
                if h and any(r[i] is not None for r in data_rows[:100])
            ]
            columns = [headers[i] for i in non_blank]
            rows = []
            for r in data_rows:
                row = []
                for i in non_blank:
                    v = r[i]
                    row.append(str(v) if hasattr(v, "isoformat") else v)
                rows.append(row)
            result[key] = {"columns": columns, "rows": rows}
        wb.close()
        return {"sheets": result, "date": run_date}

    try:
        return _gz_json_response("slate-excel", _build, ttl=_PIPELINE_JSON_TTL)
    except Exception as e:
        return jsonify({"error": str(e), "sheets": {}}), 500


# ──────────────────────────────────────────────────────────────────────────────
# API: Full Slate — filterable rows from the "Full Slate" sheet of the combined
#      tickets xlsx.  Query params: sport=NBA, tier=A, pick_type=Goblin, dir=OVER
#      Returns: { date, columns, rows }  (rows are compact value arrays)
# ──────────────────────────────────────────────────────────────────────────────
@app.route("/api/full-slate")
def api_full_slate():
    # Re-use the same file-finder used by /api/slate-excel
    path, run_date = _find_combined_excel()
    if path is None:
        return jsonify({"error": "combined slate Excel not found", "date": None, "columns": [], "rows": []}), 404

    try:
        import pandas as pd
        df = pd.read_excel(str(path), sheet_name="Full Slate", dtype=object, engine="openpyxl")
    except Exception as e:
        return jsonify({"error": f"Could not read Full Slate sheet: {e}", "date": run_date, "columns": [], "rows": []}), 500

    # Normalise column names for filter matching (strip whitespace)
    df.columns = [str(c).strip() for c in df.columns]

    # Query-param filters — case-insensitive substring match on Sport/Tier/Pick Type/Direction
    sport    = request.args.get("sport", "").strip().upper()
    tier     = request.args.get("tier", "").strip().upper()
    pick_type= request.args.get("pick_type", "").strip().lower()
    direction= request.args.get("dir", "").strip().upper()

    col_map = {c.upper(): c for c in df.columns}
    def _filter(col_key: str, value: str):
        nonlocal df
        col = col_map.get(col_key)
        if col and value:
            df = df[df[col].astype(str).str.strip().str.upper() == value]

    _filter("SPORT",     sport)
    _filter("TIER",      tier)
    _filter("DIRECTION", direction)
    if pick_type:
        col = col_map.get("PICK TYPE")
        if col:
            df = df[df[col].astype(str).str.strip().str.lower() == pick_type]

    # Replace NaN with None for JSON serialisation
    df = df.where(df.notna(), other=None)

    columns = list(df.columns)
    rows = df.values.tolist()
    # Convert any remaining non-JSON-serialisable values (dates, etc.)
    import math
    def _clean(v):
        if v is None:
            return None
        if isinstance(v, float) and math.isnan(v):
            return None
        if hasattr(v, "isoformat"):
            return v.isoformat()
        return v
    rows = [[_clean(v) for v in row] for row in rows]

    resp = jsonify({"date": run_date, "columns": columns, "rows": rows})
    resp.headers["Cache-Control"] = "no-store"
    return resp


def _find_combined_excel() -> "tuple[Path | None, str | None]":
    """Return (path, date_str) for the most recent combined_slate_tickets_*.xlsx."""
    try:
        tj = read_json_cached(TEMPLATES_DIR / "tickets_latest.json")
        pref = str((tj or {}).get("date") or "")[:10]
    except Exception:
        pref = ""
    for d in _slate_day_candidates(pref if len(pref) == 10 else None):
        out_d = BASE_DIR / "outputs" / d
        candidates: list[Path] = [
            out_d / f"combined_slate_tickets_{d}.xlsx",
            COMBINED_OUT / f"combined_slate_tickets_{d}.xlsx",
        ]
        seen: set[Path] = set(candidates)
        if out_d.is_dir():
            for g in out_d.glob(f"combined_slate_tickets_{d}*.xlsx"):
                if g not in seen:
                    seen.add(g); candidates.append(g)
        for g in BASE_DIR.glob(f"combined_slate_tickets_{d}*.xlsx"):
            if g not in seen:
                seen.add(g); candidates.append(g)
        for p in candidates:
            if p.exists():
                return p, d
    return None, None


# ──────────────────────────────────────────────────────────────────────────────
# API: Screenshot → Google Gemini vision (server proxy).
#
# Get a free Gemini API key at: https://aistudio.google.com/apikey
# Set in Railway dashboard: "Google Screenshot API" = your key
# (fallback still supports GOOGLE_API_KEY for local env/.env usage)
# Set locally: $env:GOOGLE_API_KEY = "your key"
#
# Optional: $env:GEMINI_VISION_MODEL = "gemini-2.5-flash"   # override default model id
#
# Quota errors ("free_tier ... limit: 0" or RESOURCE_EXHAUSTED):
#   - Prefer a current model (default: gemini-2.5-flash-lite). Deprecated models
#     like gemini-2.0-flash may show zero free-tier quota.
#   - In Google Cloud Console, enable "Generative Language API" for the key's project.
#   - AI Studio → check usage: https://ai.google.dev/gemini-api/docs/rate-limits
# ──────────────────────────────────────────────────────────────────────────────
_DEFAULT_GEMINI_VISION_MODEL = "gemini-2.5-flash-lite"


@app.post("/api/vision/screenshot")
def api_vision_screenshot():
    key = (os.environ.get("GOOGLE_API_KEY") or os.environ.get("Google Screenshot API") or "").strip()
    print(f"[vision] key found: {bool(key)} len={len(key)}", flush=True)
    if not key:
        return jsonify({"error": "No API key configured"}), 503

    model_id = (os.environ.get("GEMINI_VISION_MODEL") or "").strip() or _DEFAULT_GEMINI_VISION_MODEL
    _mid_ok = set("abcdefghijklmnopqrstuvwxyzABCDEFGHIJKLMNOPQRSTUVWXYZ0123456789._-")
    if not model_id or not all(c in _mid_ok for c in model_id):
        model_id = _DEFAULT_GEMINI_VISION_MODEL

    payload = request.get_json(force=True, silent=True) or {}
    image_base64 = payload.get("image_base64")
    media_type = payload.get("media_type") or "image/jpeg"
    prompt = payload.get("prompt")

    if not image_base64 or not isinstance(image_base64, str):
        return jsonify({"error": "missing image_base64"}), 400
    if not prompt or not isinstance(prompt, str):
        return jsonify({"error": "missing prompt"}), 400
    if not isinstance(media_type, str):
        media_type = "image/jpeg"

    # Reinforce numeric extraction (client also sends a detailed prompt). Prevents lazy defaults
    # like 0.5 for Goblin lines when the model should read the slip pixels.
    _vision_line_hint = (
        "\n\nReminder: For every leg, copy the exact line numbers visible on the slip. "
        "Do not invent values (e.g. do not default Goblin lines to 0.5). "
        "When both market and played lines appear, put the market line in standard_line "
        "and the slip's played line in line."
    )
    prompt = str(prompt) + _vision_line_hint

    gemini_body = {
        "contents": [
            {
                "parts": [
                    {
                        "inline_data": {
                            "mime_type": media_type,
                            "data": image_base64,
                        }
                    },
                    {"text": prompt},
                ]
            }
        ]
    }

    url = (
        f"https://generativelanguage.googleapis.com/v1beta/models/{model_id}:generateContent"
        f"?key={quote(key, safe='')}"
    )
    req = urllib.request.Request(
        url,
        data=json.dumps(gemini_body).encode("utf-8"),
        method="POST",
        headers={"Content-Type": "application/json"},
    )
    try:
        with urllib.request.urlopen(req, timeout=120) as resp:
            raw = json.loads(resp.read().decode("utf-8"))
    except urllib.error.HTTPError as e:
        err_body = e.read().decode("utf-8")
        try:
            err_json = json.loads(err_body)
            return app.response_class(
                response=json.dumps(err_json),
                status=e.code,
                mimetype="application/json",
            )
        except json.JSONDecodeError:
            return jsonify({"error": err_body or e.reason}), e.code
    except urllib.error.URLError as e:
        return jsonify({"error": f"Upstream request failed: {e.reason}"}), 502
    except Exception as e:
        return jsonify({"error": str(e)}), 502

    text_chunks: List[str] = []
    for cand in raw.get("candidates") or []:
        content = cand.get("content") or {}
        for part in content.get("parts") or []:
            t = part.get("text")
            if isinstance(t, str) and t:
                text_chunks.append(t)
    merged = "".join(text_chunks).strip()
    if not merged:
        return jsonify({"error": "No text returned from Gemini (empty candidates)"}), 502

    normalized = {"content": [{"type": "text", "text": merged}]}
    return app.response_class(
        response=json.dumps(normalized),
        status=200,
        mimetype="application/json",
    )


# Order for Income page sport table (graded single-sport tickets / props only).
_SPORT_BREAKDOWN_ORDER = ("NBA", "CBB", "CFB", "WNBA", "MLB", "SOCCER", "TENNIS", "NHL", "NFL")


def _to_float(v: Any, default: float = 0.0) -> float:
    try:
        f = float(v)
        return f if math.isfinite(f) else default
    except (TypeError, ValueError):
        return default


def _to_int(v: Any, default: int = 0) -> int:
    try:
        return int(v)
    except (TypeError, ValueError):
        return default


def _grade_history_candidate_paths() -> list[Path]:
    """Ordered fallbacks for income history (shared with utils.proporacle_data_root.grade_history_read_paths)."""
    return grade_history_read_paths(BASE_DIR, templates_dir=TEMPLATES_DIR)


def _read_grade_history_runs() -> list[dict[str, Any]]:
    for path in _grade_history_candidate_paths():
        if not path.is_file():
            continue
        try:
            raw = json.loads(path.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        if isinstance(raw, list):
            return [x for x in raw if isinstance(x, dict)]
        if isinstance(raw, dict) and isinstance(raw.get("runs"), list):
            return [x for x in (raw.get("runs") or []) if isinstance(x, dict)]
    return []


def _load_grade_history_rows() -> list[dict[str, Any]]:
    runs = _read_grade_history_runs()

    rows: list[dict[str, Any]] = []
    for r in runs:
        if not isinstance(r, dict):
            continue
        d = str(r.get("date") or "").strip()[:10]
        if not d:
            continue
        n_tickets = max(0, _to_int(r.get("n_tickets"), 0))
        wins = max(0, _to_int(r.get("wins"), 0))
        guarantees = max(0, _to_int(r.get("guarantees"), 0))
        losses = max(0, _to_int(r.get("losses"), 0))
        decided = wins + guarantees + losses
        if decided == 0:
            decided = n_tickets
        paid = wins + guarantees
        net_per_10 = _to_float(r.get("net_per_10"), 0.0)
        net_dollars = round(net_per_10 * n_tickets, 2)
        roi_pct = _to_float(r.get("roi_pct"), 0.0)
        if n_tickets > 0 and abs(roi_pct) < 1e-9 and abs(net_dollars) > 1e-9:
            roi_pct = (net_dollars / (n_tickets * 10.0)) * 100.0
        day_win_rate = (paid / decided) if decided > 0 else None
        rows.append(
            {
                "date": d,
                "tickets": n_tickets,
                "wins": wins,
                "losses": losses,
                "guarantees": guarantees,
                "decided": decided,
                "paid": paid,
                "net_dollars": net_dollars,
                "roi_pct": round(roi_pct, 2),
                "win_rate": day_win_rate,
            }
        )

    rows.sort(key=lambda x: x["date"])
    return rows


def _parse_sports_tokens(raw: Any) -> list[str]:
    s = str(raw or "").upper()
    if not s:
        return []
    toks = [t for t in re.split(r"[^A-Z0-9]+", s) if t]
    seen: set[str] = set()
    out: list[str] = []
    for t in toks:
        if t not in seen:
            seen.add(t)
            out.append(t)
    return out


def _empty_sport_breakdown_rows() -> list[dict[str, Any]]:
    return [{"sport": s, "decided": 0, "paid": 0, "win_rate": None, "net_dollars": 0.0} for s in _SPORT_BREAKDOWN_ORDER]


def _normalize_sport_label(raw: Any) -> str:
    s = str(raw or "").strip().upper()
    if not s:
        return ""
    aliases = {
        "NCAAB": "CBB",
        "WCBB": "CBB",
        "NCAAF": "CFB",
        "NBA1Q": "NBA",
        "NBA1H": "NBA",
    }
    return aliases.get(s, s)


def _sport_breakdown_from_graded_props_json(stake_per_pick: float = 10.0) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, float]] = {
        s: {"decided": 0.0, "paid": 0.0, "net": 0.0} for s in _SPORT_BREAKDOWN_ORDER
    }
    files = sorted(TEMPLATES_DIR.glob("graded_props_*.json"))
    for fp in files:
        try:
            payload = json.loads(fp.read_text(encoding="utf-8"))
        except (OSError, json.JSONDecodeError):
            continue
        props = payload.get("props") if isinstance(payload, dict) else None
        if not isinstance(props, list):
            continue
        for row in props:
            if not isinstance(row, dict):
                continue
            sp = _normalize_sport_label(row.get("sport"))
            if sp not in stats:
                continue
            result = str(row.get("result") or "").strip().upper()
            if result in {"", "NO_ACTUAL", "PENDING", "VOID", "PUSH"}:
                continue
            is_hit = result == "HIT"
            is_miss = result == "MISS"
            if not is_hit and not is_miss:
                continue
            stats[sp]["decided"] += 1.0
            if is_hit:
                stats[sp]["paid"] += 1.0
                stats[sp]["net"] += stake_per_pick
            else:
                stats[sp]["net"] -= stake_per_pick

    out: list[dict[str, Any]] = []
    for sp in _SPORT_BREAKDOWN_ORDER:
        decided = int(stats[sp]["decided"])
        paid = int(stats[sp]["paid"])
        win_rate = (paid / decided) if decided > 0 else None
        out.append(
            {
                "sport": sp,
                "decided": decided,
                "paid": paid,
                "win_rate": win_rate,
                "net_dollars": round(float(stats[sp]["net"]), 2),
            }
        )
    return out


def _sport_breakdown_from_graded_workbooks(stake_per_ticket: float = 10.0) -> list[dict[str, Any]]:
    stats: dict[str, dict[str, float]] = {
        s: {"decided": 0.0, "paid": 0.0, "net": 0.0} for s in _SPORT_BREAKDOWN_ORDER
    }
    try:
        import pandas as pd
    except Exception:
        return _empty_sport_breakdown_rows()

    files = sorted(OUTPUTS_ROOT.glob("*/combined_tickets_graded_*.xlsx"))
    if not files:
        return _sport_breakdown_from_graded_props_json(stake_per_pick=stake_per_ticket)
    for fp in files:
        try:
            df = pd.read_excel(
                fp,
                sheet_name="TICKET_RESULTS",
                usecols=["sports", "payout_status", "applied_mult", "is_cash"],
            )
        except Exception:
            continue
        if df is None or len(df) == 0:
            continue
        for _, row in df.iterrows():
            toks = _parse_sports_tokens(row.get("sports"))
            if len(toks) != 1:
                # Skip cross-sport tickets to keep per-sport totals clean.
                continue
            sp = toks[0]
            if sp not in stats:
                continue
            status = str(row.get("payout_status") or "").strip().upper()
            if status in {"", "NO_ACTUAL", "PENDING"}:
                continue
            mult = _to_float(row.get("applied_mult"), float("nan"))
            if not math.isfinite(mult):
                continue
            is_cash = _to_int(row.get("is_cash"), 0) > 0 or mult > 0
            stats[sp]["decided"] += 1.0
            stats[sp]["paid"] += 1.0 if is_cash else 0.0
            stats[sp]["net"] += stake_per_ticket * (mult - 1.0)

    out: list[dict[str, Any]] = []
    for sp in _SPORT_BREAKDOWN_ORDER:
        decided = int(stats[sp]["decided"])
        paid = int(stats[sp]["paid"])
        win_rate = (paid / decided) if decided > 0 else None
        out.append(
            {
                "sport": sp,
                "decided": decided,
                "paid": paid,
                "win_rate": win_rate,
                "net_dollars": round(float(stats[sp]["net"]), 2),
            }
        )
    if any(r.get("decided", 0) > 0 for r in out):
        return out
    return _sport_breakdown_from_graded_props_json(stake_per_pick=stake_per_ticket)


def _current_streak_label(rows_asc: list[dict[str, Any]]) -> str:
    streak_sign = 0
    streak_len = 0
    for r in reversed(rows_asc):
        net = _to_float(r.get("net_dollars"), 0.0)
        sign = 1 if net > 0 else (-1 if net < 0 else 0)
        if sign == 0:
            if streak_len == 0:
                continue
            break
        if streak_sign == 0:
            streak_sign = sign
            streak_len = 1
        elif sign == streak_sign:
            streak_len += 1
        else:
            break
    if streak_len == 0:
        return "—"
    return f"{'W' if streak_sign > 0 else 'L'}{streak_len}"


@app.get("/income")
def page_income():
    rows_asc = _load_grade_history_rows()
    total_tickets = sum(_to_int(r.get("tickets"), 0) for r in rows_asc)
    total_decided = sum(_to_int(r.get("decided"), 0) for r in rows_asc)
    total_paid = sum(_to_int(r.get("paid"), 0) for r in rows_asc)
    total_net = round(sum(_to_float(r.get("net_dollars"), 0.0) for r in rows_asc), 2)
    win_rate = (total_paid / total_decided) if total_decided > 0 else None
    roi_pct = (total_net / (total_tickets * 10.0) * 100.0) if total_tickets > 0 else 0.0
    streak = _current_streak_label(rows_asc)

    cum = 0.0
    cum_points: list[dict[str, Any]] = []
    for r in rows_asc:
        cum += _to_float(r.get("net_dollars"), 0.0)
        cum_points.append({"date": r.get("date"), "cum_net": round(cum, 2)})

    rows_desc = list(reversed(rows_asc))
    sport_rows = _sport_breakdown_from_graded_workbooks(stake_per_ticket=10.0)
    return render_template(
        "dashboard_income.html",
        ui_build_id=_UI_BUILD_ID,
        summary={
            "total_tickets": total_tickets,
            "decided_tickets": total_decided,
            "paid_tickets": total_paid,
            "win_rate": win_rate,
            "net_pnl": total_net,
            "roi_pct": round(roi_pct, 2),
            "streak": streak,
        },
        daily_rows=rows_desc,
        chart_points=Markup(json.dumps(cum_points)),
        sport_rows=sport_rows,
    )


@app.get("/dashboard/income")
def dashboard_income_legacy_redirect():
    return redirect("/income", code=302)


if __name__ == "__main__":
    app.run(host="0.0.0.0", port=8787, debug=False)
