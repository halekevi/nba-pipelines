"""
Step 7 — Rank + Tier NHL Props
Scores each prop using a composite model and assigns A/B/C/D tier.

Scoring model:
  - composite_hit_rate (primary signal, weighted by stat stability)
  - defense tier (opponent difficulty)
  - scoring tier (player quality)
  - PP tier (power play usage context)
  - sample size confidence
  - home/road adjustment

Usage:
    py step7_rank_props_nhl.py --input step6_nhl_role_context.csv \
        --output step7_nhl_ranked.xlsx
"""

import argparse
import csv
import json
import re
import subprocess
import sys
from pathlib import Path
import joblib
import numpy as np
import pandas as pd
try:
    from tqdm import tqdm as _tqdm
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm", "--break-system-packages", "-q"])
    from tqdm import tqdm as _tqdm

# ── Player consistency (data/cache/player_consistency.db) ────────────────────


def _repo_root_pc_nhl() -> Path:
    here = Path(__file__).resolve()
    for anc in here.parents:
        if (anc / "scripts" / "build_player_consistency.py").is_file():
            return anc
    return here.parents[2]


def _load_bpc_pc_nhl():
    root = _repo_root_pc_nhl()
    sd = str(root / "scripts")
    if sd not in sys.path:
        sys.path.insert(0, sd)
    import build_player_consistency as bpc  # noqa: E402

    return bpc


_bpc_pc_nhl_mod = None


def _bpc_pc_nhl():
    global _bpc_pc_nhl_mod
    if _bpc_pc_nhl_mod is None:
        try:
            _bpc_pc_nhl_mod = _load_bpc_pc_nhl()
        except Exception:
            _bpc_pc_nhl_mod = False
    return _bpc_pc_nhl_mod


def _normalize_prop_type(raw: str) -> str:
    m = _bpc_pc_nhl()
    if not m:
        return str(raw or "").strip()
    return m._normalize_prop_type(str(raw), "NHL")


def _get_line_bucket(prop_type: str, line: float, sport: str) -> str:
    m = _bpc_pc_nhl()
    if not m:
        return "<5"
    try:
        ln = float(line)
    except (TypeError, ValueError):
        ln = 0.0
    return m.get_line_bucket(prop_type, ln, sport)


def _get_consistency_grade(player: str, sport: str, prop_type: str, direction: str, line: float) -> str:
    import sqlite3

    repo_root = _repo_root_pc_nhl()
    db_path = repo_root / "data" / "cache" / "player_consistency.db"
    if not db_path.exists():
        return "?"

    prop_type = _normalize_prop_type(prop_type)
    direction = direction.upper().strip()
    bucket = _get_line_bucket(prop_type, line, sport)

    try:
        conn = sqlite3.connect(str(db_path))
        cur = conn.cursor()
        cur.execute(
            """
            SELECT grade, grade_locked, games_since_F
            FROM player_consistency
            WHERE player_name = ?
              AND sport = ?
              AND prop_type = ?
              AND direction = ?
              AND line_bucket = ?
        """,
            (player, sport, prop_type, direction, bucket),
        )
        row = cur.fetchone()
        conn.close()
        if row:
            grade, locked, games_since_F = row
            return grade or "?"
        return "?"
    except Exception:
        return "?"


def _pc_grade_cache_nhl(sport: str) -> dict:
    import sqlite3

    dbp = _repo_root_pc_nhl() / "data" / "cache" / "player_consistency.db"
    if not dbp.is_file():
        return {}
    try:
        conn = sqlite3.connect(str(dbp))
        cur = conn.execute(
            "SELECT player_name, sport, prop_type, direction, line_bucket, grade FROM player_consistency WHERE sport = ?",
            (sport,),
        )
        d = {(a, b, c, d0, e): (g if g else "?") for a, b, c, d0, e, g in cur.fetchall()}
        conn.close()
        return d
    except Exception:
        return {}


def _apply_consistency_grade_scores_nhl(df: pd.DataFrame) -> None:
    sport = "NHL"
    grade_multiplier = {
        "S": 1.25,
        "A": 1.15,
        "B": 1.05,
        "C": 1.00,
        "D": 0.80,
        "F": 0.00,
        "?": 0.95,
    }
    cache = _pc_grade_cache_nhl(sport)
    pc = next((c for c in ("player_name", "player_norm", "player") if c in df.columns), None)
    prop_col = next((c for c in ("stat_norm", "prop_type") if c in df.columns), None)
    if "recommended_side" in df.columns:
        dir_col = "recommended_side"
    elif "bet_direction" in df.columns:
        dir_col = "bet_direction"
    else:
        dir_col = None
    line_col = next((c for c in ("line_score", "line") if c in df.columns), None)
    if pc is None or prop_col is None or dir_col is None or line_col is None:
        df["consistency_grade"] = "?"
        df["consistency_multiplier"] = 0.95
        df["final_score"] = _to_num(df.get("final_score", pd.Series(np.nan, index=df.index))) * 0.95
        return

    players = df[pc].astype(str).str.strip()
    prop_raw = df[prop_col].astype(str)
    dirs = df[dir_col].astype(str).str.strip().str.upper()
    linev = _to_num(df[line_col]).fillna(0.0)
    grades: list[str] = []
    for i in range(len(df)):
        ptype = _normalize_prop_type(prop_raw.iloc[i])
        try:
            ln = float(linev.iloc[i])
        except (TypeError, ValueError):
            ln = 0.0
        bkt = _get_line_bucket(ptype, ln, sport)
        g = cache.get((players.iloc[i], sport, ptype, dirs.iloc[i], bkt), "?")
        grades.append(g)
    gser = pd.Series(grades, index=df.index)
    mult = gser.map(lambda x: grade_multiplier.get(x, 0.95)).astype(float)
    df["consistency_grade"] = gser
    df["consistency_multiplier"] = mult
    df["final_score"] = _to_num(df["final_score"]).astype(float) * mult


# ── Head-to-Head (H2H) utility ────────────────────────────────────────────────
def _attach_h2h(df: "pd.DataFrame", cache_path: str, sport: str,
                player_col: str, opp_col: str, prop_col: str, line_col: str) -> "pd.DataFrame":
    """
    Attach H2H stats per row: how did this player perform vs this opponent
    historically in the boxscore cache?

    Adds columns:
      h2h_games      – number of H2H games found
      h2h_avg        – player's average stat value vs this opp
      h2h_over_rate  – fraction of those games where they hit OVER the current line
      h2h_last        – most recent game value vs this opp
    """
    import os, pandas as pd, numpy as np

    if not cache_path or not os.path.exists(cache_path):
        df["h2h_games"]     = 0
        df["h2h_avg"]       = np.nan
        df["h2h_over_rate"] = np.nan
        df["h2h_last"]      = np.nan
        return df

    try:
        cache = pd.read_csv(cache_path, low_memory=False)
    except Exception:
        df["h2h_games"]     = 0
        df["h2h_avg"]       = np.nan
        df["h2h_over_rate"] = np.nan
        df["h2h_last"]      = np.nan
        return df

    # Normalise cache columns: need player, opponent, stat_type, value, date
    cache.columns = [c.lower().strip() for c in cache.columns]

    # Detect player col in cache
    p_col  = next((c for c in ["player_norm","player_name","player","name"] if c in cache.columns), None)
    o_col  = next((c for c in ["opp_team","opp","opponent","opp_team_abbr"] if c in cache.columns), None)
    s_col  = next((c for c in ["stat_type","stat","prop_type","stat_norm"]   if c in cache.columns), None)
    v_col  = next((c for c in ["value","stat_value","actual","val"]          if c in cache.columns), None)
    d_col  = next((c for c in ["date","game_date","event_date"]              if c in cache.columns), None)

    if not all([p_col, o_col, v_col]):
        df["h2h_games"]     = 0
        df["h2h_avg"]       = np.nan
        df["h2h_over_rate"] = np.nan
        df["h2h_last"]      = np.nan
        return df

    cache[v_col] = pd.to_numeric(cache[v_col], errors="coerce")

    def _norm(x):
        return str(x).strip().lower() if x and str(x).strip() else ""

    # Build lookup: (player_norm, opp_norm) -> list of (value, date)
    lookup: dict = {}
    for row in cache.itertuples(index=False):
        pk = (_norm(getattr(row, p_col)), _norm(getattr(row, o_col)))
        v  = getattr(row, v_col)
        dt = getattr(row, d_col, "") if d_col else ""
        if pk not in lookup:
            lookup[pk] = []
        lookup[pk].append((v, str(dt)))

    h2h_games, h2h_avg, h2h_over, h2h_last = [], [], [], []

    for _, r in df.iterrows():
        player = _norm(r.get(player_col, ""))
        opp    = _norm(r.get(opp_col, ""))
        line   = r.get(line_col, None)
        try:
            line_f = float(line)
        except (TypeError, ValueError):
            line_f = None

        entries = lookup.get((player, opp), [])
        # sort by date desc, take up to 10
        try:
            entries_sorted = sorted(entries, key=lambda x: x[1], reverse=True)[:10]
        except Exception:
            entries_sorted = entries[:10]

        vals = [v for v, _ in entries_sorted if v is not None and not (isinstance(v, float) and v != v)]

        if not vals:
            h2h_games.append(0)
            h2h_avg.append(np.nan)
            h2h_over.append(np.nan)
            h2h_last.append(np.nan)
        else:
            avg = round(float(np.mean(vals)), 2)
            last = vals[0]
            over_rate = round(sum(1 for v in vals if line_f is not None and v > line_f) / len(vals), 3) if line_f else np.nan
            h2h_games.append(len(vals))
            h2h_avg.append(avg)
            h2h_over.append(over_rate)
            h2h_last.append(round(float(last), 2))

    df["h2h_games"]     = h2h_games
    df["h2h_avg"]       = h2h_avg
    df["h2h_over_rate"] = h2h_over
    df["h2h_last"]      = h2h_last
    return df
# ─────────────────────────────────────────────────────────────────────────────


# Stat stability weights (higher = more predictable/trackable)
STAT_STABILITY = {
    "shots_on_goal": 1.12,   # most consistent NHL prop — volume metric
    "saves":         1.10,   # consistent for starting goalies
    "assists":       1.00,
    "points":        1.00,
    "hits":          1.05,   # consistent for physical players
    "blocked_shots": 1.05,
    "goals":         0.82,   # highest variance in hockey
    "goals_allowed": 0.85,
    "fantasy_score": 0.88,
}

# Defense tier adjustments for OVER hit rate
DEF_TIER_BOOST = {
    "WEAK":    +0.04,    # easiest opp = small boost to over
    "AVERAGE": +0.00,
    "SOLID":   -0.02,
    "ELITE":   -0.05,    # hardest opp = penalize over
}

# Scoring tier adjustments
SCORING_TIER_BOOST = {
    "ELITE":    +0.05,
    "SECONDARY": +0.02,
    "DEPTH":     -0.02,
    "SHUTDOWN":  -0.04,
    "GOALIE":    +0.0,
    "UNKNOWN":   +0.0,
}

PP_TIER_BOOST = {
    "PP1_STAR":   +0.04,
    "PP_REGULAR": +0.01,
    "PP_OCC":     +0.0,
    "NO_PP":      -0.01,
    "N/A":        +0.0,
}

HOME_BOOST = 0.01
MIN_SAMPLE = 5  # Minimum games to be rankable


def _repo_root_ml_nhl() -> Path:
    return Path(__file__).resolve().parents[2]


_sd_ml_nhl = str(_repo_root_ml_nhl() / "scripts")
if _sd_ml_nhl not in sys.path:
    sys.path.insert(0, _sd_ml_nhl)
try:
    from ml_blend_weight import load_ml_blend_weight  # noqa: E402

    ML_BLEND_WEIGHT = float(load_ml_blend_weight(_repo_root_ml_nhl(), "nhl"))
except Exception:
    ML_BLEND_WEIGHT = 0.30


def install_openpyxl():
    subprocess.check_call([sys.executable, "-m", "pip", "install", "openpyxl", "--break-system-packages", "-q"])


def read_csv(path: str) -> list[dict]:
    with open(path, newline="", encoding="utf-8") as f:
        return list(csv.DictReader(f))


def safe_float(val, default=0.0) -> float:
    try:
        return float(val)
    except Exception:
        return default


def score_prop(row: dict) -> float:
    stat = row.get("stat_norm", "")
    composite = safe_float(row.get("composite_hit_rate"))
    if composite == 0.0 and row.get("composite_hit_rate", "") == "":
        # No hit rate data — return a neutral score so prop still appears with D tier
        return 0.0
    recommended_side = row.get("recommended_side", "OVER")

    # Flip composite to represent confidence in recommended side
    confidence = composite if recommended_side == "OVER" else 1 - composite

    # Stat stability weight
    stab = STAT_STABILITY.get(stat, 1.0)
    base_score = confidence * stab

    # Defense adjustment (applies to OVER on skater props; reverse for UNDER)
    def_tier = row.get("def_tier", "AVERAGE")
    def_adj = DEF_TIER_BOOST.get(def_tier, 0.0)
    if recommended_side == "UNDER":
        def_adj = -def_adj  # Under benefits from tougher defense

    # Scoring tier
    scoring_tier = row.get("scoring_tier", "DEPTH")
    score_adj = SCORING_TIER_BOOST.get(scoring_tier, 0.0)

    # PP tier
    pp_tier = row.get("pp_tier", "N/A")
    pp_adj = PP_TIER_BOOST.get(pp_tier, 0.0)

    # Home/road
    is_home = str(row.get("is_home", "0")) == "1"
    home_adj = HOME_BOOST if is_home else 0.0

    # Sample confidence (penalize small samples)
    sample = safe_float(row.get("sample_L10", 0))
    sample_conf = min(sample / 10.0, 1.0)

    total = (base_score + def_adj + score_adj + pp_adj + home_adj) * sample_conf
    return round(total, 5)


def assign_tier(score: float, sample: float) -> str:
    if sample < MIN_SAMPLE:
        return "D"
    if score >= 0.68:
        return "A"
    elif score >= 0.60:
        return "B"
    elif score >= 0.50:
        return "C"
    else:
        return "D"


def _to_num(s: pd.Series) -> pd.Series:
    return pd.to_numeric(s, errors="coerce")


def _first_present_nhl_ml(df: pd.DataFrame, options: list[str]) -> str | None:
    lookup = {str(c).lower(): c for c in df.columns}
    for c in options:
        if str(c).lower() in lookup:
            return lookup[str(c).lower()]
    return None


def _nhl_ml_pick_col(df: pd.DataFrame, names: tuple[str, ...]) -> pd.Series:
    idx = df.index
    for n in names:
        if n in df.columns:
            return df[n]
    return pd.Series(np.nan, index=idx)


def _nhl_ml_final_direction(df: pd.DataFrame) -> pd.Series:
    bt = _nhl_ml_pick_col(df, ("bet_direction",))
    dr = _nhl_ml_pick_col(df, ("direction",))
    rs = _nhl_ml_pick_col(df, ("recommended_side",))
    out = bt.fillna(dr).fillna(rs)
    out = out.astype(str).str.strip().str.upper()
    out = out.replace({"NAN": np.nan, "NONE": np.nan, "NULL": np.nan, "": np.nan})
    return out.fillna("OVER")


def _normalize_nhl_prop_ml(raw: str) -> str:
    x = str(raw or "").strip().lower()
    xc = re.sub(r"\s+", "", x)
    if not x:
        return "unknown"
    if ("goalie" in x and "save" in x) or "saves" in x:
        return "saves"
    if "faceoff" in x:
        return "faceoffs"
    if "blocked" in x:
        return "blocked_shots"
    if "shot" in x or "sog" in x or xc == "shots":
        return "shots"
    if "assist" in x:
        return "assists"
    if "goal" in x and "against" not in x:
        return "goals"
    if "point" in x and "goal" not in x:
        return "points"
    if "toi" in x or "time on ice" in x or "time_on" in xc:
        return "toi"
    if "hit" in x and "blocked" not in x:
        return "hits"
    return re.sub(r"[^a-z0-9]+", "_", x).strip("_") or "unknown"


def _nhl_ml_defense_tier_4(df: pd.DataFrame) -> pd.Series:
    dt = _first_present_nhl_ml(df, ["def_tier", "defense_tier", "DEF_TIER"])
    if dt:
        s = df[dt].astype(str).str.strip().str.lower()
        return pd.Series(
            np.where(
                s.str.contains("weak"),
                0,
                np.where(
                    s.str.contains("avg|average"),
                    1,
                    np.where(
                        s.str.contains("good|solid|above"),
                        2,
                        np.where(s.str.contains("elite|strong"), 3, 1),
                    ),
                ),
            ),
            index=df.index,
        ).astype(float)
    rk = _first_present_nhl_ml(df, ["def_rank", "opp_def_rank"])
    if rk:
        r = _to_num(df[rk]).fillna(16.0)
        return pd.Series(
            np.where(r <= 5, 3, np.where(r <= 10, 2, np.where(r <= 20, 1, 0))),
            index=df.index,
        ).astype(float)
    return pd.Series(1.0, index=df.index)


def _nhl_ml_pp_unit(df: pd.DataFrame) -> pd.Series:
    pp_col = _first_present_nhl_ml(df, ["pp_unit", "pp_tier", "PP Tier"])
    if not pp_col:
        return pd.Series(0.0, index=df.index)
    s = df[pp_col].astype(str).str.strip().str.upper()
    return pd.Series(
        np.where(s.str.contains("PP1"), 2.0, np.where(s.str.contains("PP2"), 1.0, 0.0)),
        index=df.index,
    )


def _nhl_ml_goalie_confirmed(df: pd.DataFrame) -> pd.Series:
    gc = _first_present_nhl_ml(df, ["goalie_confirmed", "Goalie Confirmed"])
    if not gc:
        return pd.Series(0.0, index=df.index)
    s = df[gc].astype(str).str.strip().str.lower()
    return pd.Series(
        np.where(s.isin(["1", "1.0", "true", "yes", "confirmed"]), 1.0, 0.0),
        index=df.index,
    )


def _nhl_ml_is_home(df: pd.DataFrame) -> pd.Series:
    hc = _first_present_nhl_ml(df, ["is_home", "home_away"])
    if not hc:
        return pd.Series(0.5, index=df.index)
    ih = df[hc].astype(str).str.strip()
    return pd.Series(
        np.where(ih.isin(["1", "1.0", "true", "True", "HOME", "home"]), 1.0, 0.0),
        index=df.index,
    )


def _build_nhl_ml_X(df: pd.DataFrame, model_features: list[str]) -> pd.DataFrame:
    idx = df.index
    hr5 = _to_num(
        _nhl_ml_pick_col(
            df,
            ("hit_rate_over_L5", "hit_rate_l5", "hr_L5", "hr_last5", "over_L5"),
        )
    )
    hr10 = _to_num(
        _nhl_ml_pick_col(
            df,
            (
                "composite_hit_rate",
                "hit_rate_over_L10",
                "hit_rate_l10",
                "hr_L10",
                "hr_last10",
            ),
        )
    )
    hr20 = _to_num(
        _nhl_ml_pick_col(df, ("hit_rate_over_L20", "hit_rate_l20", "hr_L20", "hr_last20"))
    )

    def _scale_hit_pct(s: pd.Series) -> pd.Series:
        if s.notna().any() and s.dropna().median() > 1.0:
            return s / 100.0
        return s

    hr5, hr10, hr20 = _scale_hit_pct(hr5), _scale_hit_pct(hr10), _scale_hit_pct(hr20)
    hr5 = hr5.fillna(hr10).fillna(0.5)
    hr10 = hr10.fillna(0.5)
    hr20 = hr20.fillna(hr10).fillna(0.5)

    line = _to_num(_nhl_ml_pick_col(df, ("line", "line_score", "Line"))).fillna(0.0)
    edge = _to_num(df.get("edge", pd.Series(np.nan, index=idx))).fillna(0.0)
    pick = df.get("pick_type", pd.Series("Standard", index=idx)).astype(str).str.lower()
    tier_num = pd.Series(
        np.where(pick.str.contains("gob"), 2, np.where(pick.str.contains("dem"), 0, 1)),
        index=idx,
    )
    final_direction = _nhl_ml_final_direction(df)
    dir_num = pd.Series(np.where(final_direction.eq("OVER"), 1, 0), index=idx)
    prop_raw = df.get("stat_norm", df.get("prop_type", pd.Series("unknown", index=idx)))
    prop_norm = prop_raw.astype(str).map(_normalize_nhl_prop_ml)
    dummies = pd.get_dummies(prop_norm, prefix="prop", dtype=float)

    base = pd.DataFrame(
        {
            "edge": edge,
            "hit_rate_l5": hr5,
            "hit_rate_l10": hr10,
            "hit_rate_l20": hr20,
            "line": line,
            "direction": _to_num(dir_num).fillna(0.0),
            "tier": _to_num(tier_num).fillna(1.0),
            "defense_tier": _nhl_ml_defense_tier_4(df).fillna(1.0),
            "pp_unit": _nhl_ml_pp_unit(df),
            "toi_minutes": _to_num(
                _nhl_ml_pick_col(
                    df,
                    ("toi_per_game_api", "toi_minutes", "toi_avg_L10", "Time On Ice"),
                )
            ).fillna(0.0),
            "goalie_confirmed": _nhl_ml_goalie_confirmed(df),
            "is_home": _nhl_ml_is_home(df),
        },
        index=idx,
    )
    return pd.concat([base, dummies], axis=1).reindex(columns=model_features, fill_value=0.0)


def _apply_ml_blend(df: pd.DataFrame) -> tuple[pd.Series, pd.Series, pd.Series]:
    root = Path(__file__).resolve().parents[2]
    model_path = root / "models" / "prop_model_nhl.pkl"
    feat_path = root / "models" / "prop_model_nhl_features.json"
    existing_score = _to_num(df.get("prop_score", pd.Series(np.nan, index=df.index))).fillna(0.0)
    if not (model_path.exists() and feat_path.exists()):
        print(f"⚠️  NHL ML model missing at {model_path} — skipping ML blend")
        return pd.Series(np.nan, index=df.index), pd.Series(np.nan, index=df.index), existing_score
    try:
        model = joblib.load(model_path)
        model_features = json.loads(feat_path.read_text(encoding="utf-8"))
    except Exception as e:
        print(f"⚠️  Failed loading NHL ML model: {e} — skipping ML blend")
        return pd.Series(np.nan, index=df.index), pd.Series(np.nan, index=df.index), existing_score

    calibrator = None
    calib_path = root / "models" / "prop_model_nhl_calibrator.pkl"
    try:
        if calib_path.exists():
            calibrator = joblib.load(calib_path)
    except Exception:
        calibrator = None

    try:
        from ml_blend_weight import load_ml_blend_weight as _load_nhl_blend

        blend_w = float(_load_nhl_blend(root, "nhl"))
    except Exception:
        blend_w = float(ML_BLEND_WEIGHT)

    try:
        X = _build_nhl_ml_X(df, model_features)
        raw_prob = pd.Series(model.predict_proba(X)[:, 1], index=df.index, dtype=float)
        if calibrator is not None:
            try:
                if hasattr(calibrator, "predict_proba"):
                    cal_vals = calibrator.predict_proba(raw_prob.values.reshape(-1, 1))[:, 1]
                else:
                    cal_vals = calibrator.predict(raw_prob.values)
                ml_prob = pd.Series(cal_vals, index=df.index, dtype=float).clip(0.001, 0.999)
            except Exception:
                ml_prob = raw_prob
        else:
            ml_prob = raw_prob
    except Exception as e:
        print(f"⚠️  NHL ML inference failed: {e} — skipping ML blend")
        return pd.Series(np.nan, index=df.index), pd.Series(np.nan, index=df.index), existing_score

    ml_edge = ml_prob - 0.5
    final_score = (1.0 - blend_w) * existing_score + blend_w * ml_edge
    print(f"✅ NHL ML blend applied (weight={blend_w:.2f})")
    return ml_prob, ml_edge, final_score


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="step6_nhl_role_context.csv")
    parser.add_argument("--output", default="step7_nhl_ranked.xlsx")
    parser.add_argument("--min-sample", type=int, default=MIN_SAMPLE)
    parser.add_argument("--cache", default="", help="Path to NHL boxscore cache CSV")
    parser.add_argument(
        "--injuries-csv",
        default="",
        help="injuries_nhl_*.csv (optional rank penalty from ESPN reports)",
    )
    parser.add_argument(
        "--slate-date",
        default="",
        help="YYYY-MM-DD; auto-load outputs/<date>/injuries_nhl_<date>.csv if injuries-csv omitted",
    )
    args = parser.parse_args()

    try:
        import openpyxl
    except ImportError:
        install_openpyxl()
        import openpyxl

    rows = read_csv(args.input)

    scored = []
    for row in _tqdm(rows, desc="  Scoring props", unit="prop"):
        prop_score = score_prop(row)
        row["prop_score"] = prop_score
        scored.append(row)
    scored_df = pd.DataFrame(scored)
    scored_df["ml_prob"], scored_df["ml_edge"], scored_df["final_score"] = _apply_ml_blend(scored_df)
    _apply_consistency_grade_scores_nhl(scored_df)

    # Game script risk adjustment
    from datetime import datetime, timezone

    _sd_gs = str(_repo_root_pc_nhl() / "scripts")
    if _sd_gs not in sys.path:
        sys.path.insert(0, _sd_gs)
    from game_script_risk import get_game_script_multiplier  # noqa: E402

    _fallback_gd = datetime.now(timezone.utc).strftime("%Y-%m-%d")
    _gmults: list[float] = []
    _gnotes: list[str] = []
    for _i in range(len(scored_df)):
        _r = scored_df.iloc[_i]
        _gs = str(_r.get("game_start", "") or "").strip()
        _gd = _gs[:10] if len(_gs) >= 10 and _gs[4] == "-" else _fallback_gd
        _tm = str(_r.get("team", "") or "").strip()
        _pt = str(_r.get("stat_norm", _r.get("stat_type", "")) or "").strip()
        _gm, _gn = get_game_script_multiplier(_tm, "NHL", _pt, _gd)
        _gmults.append(round(float(_gm), 3))
        _gnotes.append(_gn)
    scored_df["game_script_mult"] = _gmults
    scored_df["game_script_note"] = _gnotes
    scored_df["final_score"] = _to_num(scored_df["final_score"]).astype(float) * pd.Series(_gmults, dtype=float).values

    _nhl_sd_inj = str(_repo_root_pc_nhl() / "scripts")
    if _nhl_sd_inj not in sys.path:
        sys.path.insert(0, _nhl_sd_inj)
    from espn_injuries import auto_injuries_csv_from_outputs, penalty_series_for_slate  # noqa: E402

    _inj_nhl_path = str(args.injuries_csv or "").strip()
    if not _inj_nhl_path and str(getattr(args, "slate_date", "") or "").strip():
        _cnhl = auto_injuries_csv_from_outputs(_repo_root_pc_nhl(), str(args.slate_date).strip(), "NHL")
        _inj_nhl_path = str(_cnhl) if _cnhl else ""
    if _inj_nhl_path:
        _pc_nhl = next(
            (c for c in ("player_name", "player_norm", "player") if c in scored_df.columns),
            "player",
        )
        _pen_nhl = penalty_series_for_slate(scored_df, _pc_nhl, "team", "NHL", _inj_nhl_path)
        scored_df["final_score"] = _to_num(scored_df["final_score"]).astype(float) + _pen_nhl.values

    scored_df["prop_score"] = _to_num(scored_df["final_score"]).fillna(_to_num(scored_df["prop_score"]).fillna(0.0))
    scored_df["tier"] = scored_df.apply(
        lambda r: assign_tier(safe_float(r.get("prop_score", 0)), safe_float(r.get("sample_L10", 0))),
        axis=1,
    )
    scored = scored_df.to_dict("records")
    scored.sort(key=lambda x: -safe_float(x.get("prop_score", 0)))

    # ── Pass Demons through; only drop neg-edge Goblins to audit sheet ──────────
    def _norm_pt(x: str) -> str:
        t = (x or "").strip().lower()
        if "gob" in t: return "Goblin"
        if "dem" in t: return "Demon"
        return "Standard"

    active  = []
    dropped = []
    for row in scored:
        pt    = _norm_pt(row.get("pick_type", ""))
        edge  = safe_float(row.get("edge", 0))
        if pt == "Goblin" and edge < 0:
            row["void_reason"] = "DROPPED_NEG_EDGE_GOBLIN"
            dropped.append(row)
        else:
            # Demons pass through for data/tracking — excluded from tickets in combined_slate
            active.append(row)

    # Add rank only on active rows
    for i, row in enumerate(active):
        row["rank"] = i + 1
    for row in dropped:
        row["rank"] = ""

    print(f"  Active: {len(active)} | Dropped (neg-edge Goblin only): {len(dropped)}")

    # ── Head-to-Head stats ────────────────────────────────────────────────────
    if args.cache:
        import pandas as _pd
        df_h2h = _pd.DataFrame(active)
        player_col = next((c for c in ["player_name","player_norm","player"] if c in df_h2h.columns), "")
        opp_col    = next((c for c in ["opp_team","opp","pp_opp_team","opp_team_abbr"] if c in df_h2h.columns), "")
        line_col   = next((c for c in ["line_score","line"] if c in df_h2h.columns), "line_score")
        prop_col   = next((c for c in ["stat_norm","prop_type"] if c in df_h2h.columns), "stat_norm")
        if player_col and opp_col:
            df_h2h = _attach_h2h(df_h2h, args.cache, "nhl", player_col, opp_col, prop_col, line_col)
            active = df_h2h.to_dict("records")
            print(f"  H2H: {sum(1 for r in active if r.get('h2h_games',0) > 0)}/{len(active)} rows matched")

    # Write XLSX with multiple tabs
    wb = openpyxl.Workbook()

    # ── All Props tab ──────────────────────────────────────────────────────────
    ws_all = wb.active
    ws_all.title = "All Props"

    headers = list(active[0].keys()) if active else (list(dropped[0].keys()) if dropped else [])
    for col, h in enumerate(headers, 1):
        cell = ws_all.cell(row=1, column=col, value=h)
        cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
        cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")

    TIER_COLORS = {"A": "C6EFCE", "B": "FFEB9C", "C": "FFCCCC", "D": "E0E0E0"}
    for row in _tqdm(active, desc="  Writing All Props sheet", unit="row"):
        ws_all.append([row.get(h, "") for h in headers])
        last_row = ws_all.max_row
        tier_color = TIER_COLORS.get(row.get("tier", "D"), "FFFFFF")
        for col in range(1, len(headers) + 1):
            ws_all.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor=tier_color)

    # ── Skaters tab ────────────────────────────────────────────────────────────
    ws_sk = wb.create_sheet("Skaters")
    skaters = [r for r in active if r.get("player_role") == "SKATER"]
    if skaters:
        sk_headers = list(skaters[0].keys())
        for col, h in enumerate(sk_headers, 1):
            cell = ws_sk.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in skaters:
            ws_sk.append([row.get(h, "") for h in sk_headers])
            last_row = ws_sk.max_row
            tier_color = TIER_COLORS.get(row.get("tier", "D"), "FFFFFF")
            for col in range(1, len(sk_headers) + 1):
                ws_sk.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor=tier_color)

    # ── Goalies tab ────────────────────────────────────────────────────────────
    ws_g = wb.create_sheet("Goalies")
    goalies = [r for r in active if r.get("player_role") == "GOALIE"]
    if goalies:
        g_headers = list(goalies[0].keys())
        for col, h in enumerate(g_headers, 1):
            cell = ws_g.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="1F4E79")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in goalies:
            ws_g.append([row.get(h, "") for h in g_headers])
            last_row = ws_g.max_row
            tier_color = TIER_COLORS.get(row.get("tier", "D"), "FFFFFF")
            for col in range(1, len(g_headers) + 1):
                ws_g.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor=tier_color)

    # ── A-Tier only ────────────────────────────────────────────────────────────
    ws_a = wb.create_sheet("A-Tier Best")
    a_props = [r for r in active if r.get("tier") == "A"]
    if a_props:
        a_headers = list(a_props[0].keys())
        for col, h in enumerate(a_headers, 1):
            cell = ws_a.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="375623")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in a_props:
            ws_a.append([row.get(h, "") for h in a_headers])
            last_row = ws_a.max_row
            for col in range(1, len(a_headers) + 1):
                ws_a.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="C6EFCE")

    # ── DROPPED tab (neg-edge Goblin audit) ───────────────────────────────────
    ws_drop = wb.create_sheet("DROPPED")
    if dropped:
        d_headers = list(dropped[0].keys())
        for col, h in enumerate(d_headers, 1):
            cell = ws_drop.cell(row=1, column=col, value=h)
            cell.fill = openpyxl.styles.PatternFill("solid", fgColor="7B241C")
            cell.font = openpyxl.styles.Font(bold=True, color="FFFFFF")
        for row in dropped:
            ws_drop.append([row.get(h, "") for h in d_headers])
            last_row = ws_drop.max_row
            for col in range(1, len(d_headers) + 1):
                ws_drop.cell(last_row, col).fill = openpyxl.styles.PatternFill("solid", fgColor="FADBD8")

    # Autofit columns (approximate)
    for ws in [ws_all, ws_sk, ws_g, ws_a, ws_drop]:
        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            ws.column_dimensions[col[0].column_letter].width = min(max_len + 2, 35)

    wb.save(args.output)
    print(f"Saved ranked props -> {args.output}")

    # Summary
    tier_counts = {}
    for r in active:
        t = r.get("tier", "?")
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"Tier breakdown: {tier_counts}")
    print(f"Dropped (neg-edge Goblin only): {len(dropped)}")
    print(f"\nTop 10 props:")
    for r in active[:10]:
        print(f"  #{r['rank']} [{r['tier']}] {r['player_name']} {r['stat_norm']} "
              f"{r['line_score']} {r['recommended_side']} | score={r['prop_score']:.4f}")


if __name__ == "__main__":
    main()
