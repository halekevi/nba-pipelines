"""
Step 9 — Build PrizePicks NHL Tickets
Constructs 2-6 leg PrizePicks lineups from top-ranked props.

Strategy:
  - Power Play: 2-pick all-or-nothing (highest confidence)
  - Flex 3-pick: 3 legs with mix of skaters + optionally a goalie
  - Flex 4-pick: 4 legs from A+B tier
  - Feature 5-6 pick: longer shot high-upside ticket

Usage:
    py step9_build_tickets_nhl.py --input step8_nhl_direction_clean.xlsx \
        --output nhl_best_tickets.xlsx
"""

import argparse
import subprocess
import sys
from datetime import datetime
from itertools import combinations
try:
    from tqdm import tqdm as _tqdm
except ImportError:
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm", "--break-system-packages", "-q"])
    from tqdm import tqdm as _tqdm

MIN_HIT_RATE = 0.65      # minimum composite hit rate for any leg
MIN_TIER = {"A", "B"}    # only A and B tier props
MAX_PER_TEAM = 2         # max legs from same team
MAX_PER_PLAYER = 1       # no duplicate players in same ticket


def install(pkg):
    subprocess.check_call([sys.executable, "-m", "pip", "install", pkg, "--break-system-packages", "-q"])


def read_xlsx(path: str) -> list[dict]:
    try:
        import openpyxl
    except ImportError:
        install("openpyxl")
        import openpyxl
    wb = openpyxl.load_workbook(path)
    if "All Props" in wb.sheetnames:
        ws = wb["All Props"]
    else:
        ws = wb.active
    headers = [str(c.value or "") for c in ws[1]]
    rows = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        d = dict(zip(headers, row))
        rows.append(d)
    return rows


def safe_float(val, default=0.0) -> float:
    if val is None:
        return default
    s = str(val).strip().rstrip("%")
    if s == "":
        return default
    try:
        return float(s)
    except ValueError:
        return default


def safe_hit_rate(val, default=0.0) -> float:
    """Parse 0-1 hit rate; supports '73.3%', 73.3 (legacy scale), or 0.733."""
    if val is None:
        return default
    s = str(val).strip()
    had_pct = s.endswith("%")
    s = s.rstrip("%").strip()
    if s == "":
        return default
    try:
        f = float(s)
        if had_pct or f > 1.0:
            f = f / 100.0
        return f
    except ValueError:
        return default


def is_valid_leg(row: dict) -> bool:
    composite = safe_hit_rate(
        row.get(
            "Composite Hit Rate",
            row.get("composite_hit_rate", row.get("composite_hr", 0)),
        )
    )
    side = str(row.get("Recommended Side", row.get("recommended_side", "OVER"))).upper().strip()
    # composite_hit_rate is always the OVER rate — flip for UNDER bets
    direction_hr = composite if side == "OVER" else 1.0 - composite
    tier = str(row.get("Tier", row.get("tier", "D"))).strip()
    sample = safe_float(row.get("Sample L10", row.get("sample_L10", 0)))
    return direction_hr >= MIN_HIT_RATE and tier in MIN_TIER and sample >= 5


def ticket_diversity(legs: list[dict]) -> bool:
    """Check team/player diversity constraints."""
    teams = [str(l.get("Team", l.get("team", ""))).upper() for l in legs]
    players = [str(l.get("Player Name", l.get("player_name", ""))).lower() for l in legs]
    if len(set(players)) < len(players):
        return False  # duplicate player
    for t in set(teams):
        if teams.count(t) > MAX_PER_TEAM:
            return False
    return True


def leg_score(row: dict) -> float:
    composite = safe_hit_rate(
        row.get(
            "Composite Hit Rate",
            row.get("composite_hit_rate", row.get("composite_hr", 0)),
        )
    )
    side = str(row.get("Recommended Side", row.get("recommended_side", "OVER"))).upper().strip()
    direction_hr = composite if side == "OVER" else 1.0 - composite
    prop_score = safe_float(row.get("Prop Score", row.get("prop_score", 0)))
    return direction_hr * 0.6 + prop_score * 0.4


def format_leg(row: dict) -> str:
    name = row.get("Player Name", row.get("player_name", ""))
    stat = str(row.get("Stat Norm", row.get("stat_norm", ""))).upper().replace("_", " ")
    line = row.get("Line Score", row.get("line_score", ""))
    side = row.get("Recommended Side", row.get("recommended_side", "OVER"))
    opp = row.get("Opponent", row.get("opponent", ""))
    composite = safe_hit_rate(
        row.get(
            "Composite Hit Rate",
            row.get("composite_hit_rate", row.get("composite_hr", 0)),
        )
    )
    tier = row.get("Tier", row.get("tier", ""))
    return f"[{tier}] {name} {side} {line} {stat} vs {opp} (HR:{composite:.2f})"


def ticket_confidence(legs: list[dict]) -> float:
    """Combined geometric-mean-style confidence for a ticket."""
    prod = 1.0
    for leg in legs:
        c = safe_hit_rate(
            leg.get(
                "Composite Hit Rate",
                leg.get("composite_hit_rate", leg.get("composite_hr", 0)),
            )
        )
        side = str(leg.get("Recommended Side", leg.get("recommended_side", "OVER"))).upper().strip()
        direction_hr = c if side == "OVER" else 1.0 - c
        prod *= max(direction_hr, 0.001)  # guard against zero
    return round(prod, 5)


def build_tickets(valid_legs: list[dict]) -> dict:
    """Build ticket sets by size."""
    valid_legs.sort(key=lambda x: -leg_score(x))

    tickets = {
        "power_play_2": [],
        "flex_3": [],
        "flex_4": [],
        "goblin_5": [],
    }

    # Power Play (2-pick): top A-tier combos
    a_only = [l for l in valid_legs if str(l.get("Tier", l.get("tier", ""))).strip() == "A"]
    for combo in _tqdm(list(combinations(a_only[:20], 2)), desc="  Building 2-pick tickets", unit="combo", leave=False):
        legs = list(combo)
        if ticket_diversity(legs):
            conf = ticket_confidence(legs)
            tickets["power_play_2"].append((conf, legs))
    tickets["power_play_2"].sort(key=lambda x: -x[0])
    tickets["power_play_2"] = tickets["power_play_2"][:5]

    # Flex 3
    for combo in _tqdm(list(combinations(valid_legs[:25], 3)), desc="  Building 3-pick tickets", unit="combo", leave=False):
        legs = list(combo)
        if ticket_diversity(legs):
            conf = ticket_confidence(legs)
            tickets["flex_3"].append((conf, legs))
    tickets["flex_3"].sort(key=lambda x: -x[0])
    tickets["flex_3"] = tickets["flex_3"][:5]

    # Flex 4
    for combo in _tqdm(list(combinations(valid_legs[:20], 4)), desc="  Building 4-pick tickets", unit="combo", leave=False):
        legs = list(combo)
        if ticket_diversity(legs):
            conf = ticket_confidence(legs)
            tickets["flex_4"].append((conf, legs))
    tickets["flex_4"].sort(key=lambda x: -x[0])
    tickets["flex_4"] = tickets["flex_4"][:5]

    # Goblin 5 (5-pick flex)
    for combo in _tqdm(list(combinations(valid_legs[:18], 5)), desc="  Building 5-pick tickets", unit="combo", leave=False):
        legs = list(combo)
        if ticket_diversity(legs):
            conf = ticket_confidence(legs)
            tickets["goblin_5"].append((conf, legs))
    tickets["goblin_5"].sort(key=lambda x: -x[0])
    tickets["goblin_5"] = tickets["goblin_5"][:3]

    return tickets


def write_tickets_xlsx(tickets: dict, output_path: str):
    try:
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment
    except ImportError:
        install("openpyxl")
        import openpyxl
        from openpyxl.styles import Font, PatternFill, Alignment

    wb = openpyxl.Workbook()

    TICKET_COLORS = {
        "power_play_2": "1F4E79",
        "flex_3": "375623",
        "flex_4": "843C0C",
        "goblin_5": "5B2C8D",
    }
    TICKET_LABELS = {
        "power_play_2": "⚡ POWER PLAY 2-PICK (all-or-nothing)",
        "flex_3": "🏒 FLEX 3-PICK (win 2/3+)",
        "flex_4": "🎯 FLEX 4-PICK (win 3/4+)",
        "goblin_5": "👹 GOBLIN 5-PICK (high upside flex)",
    }

    first = True
    for ticket_type, ticket_list in tickets.items():
        if not ticket_list:
            continue

        if first:
            ws = wb.active
            ws.title = ticket_type.replace("_", " ").title()
            first = False
        else:
            ws = wb.create_sheet(ticket_type.replace("_", " ").title())

        color = TICKET_COLORS.get(ticket_type, "1F4E79")
        label = TICKET_LABELS.get(ticket_type, ticket_type)

        # Title row
        ws.append([label])
        ws.merge_cells(f"A1:H1")
        ws["A1"].font = Font(bold=True, color="FFFFFF", size=13)
        ws["A1"].fill = PatternFill("solid", fgColor=color)
        ws["A1"].alignment = Alignment(horizontal="center")

        ws.append([f"Generated: {datetime.utcnow().strftime('%Y-%m-%d %H:%M UTC')}"])
        ws.merge_cells(f"A2:H2")

        for ticket_num, (conf, legs) in enumerate(ticket_list, 1):
            ws.append([])
            ws.append([f"Ticket #{ticket_num}", f"Combined Confidence: {conf:.4f} ({conf*100:.1f}%)"])
            ws[ws.max_row][0].font = Font(bold=True, size=11)

            headers = ["Leg", "Player", "Team", "Opponent", "Stat", "Line", "Side", "Hit Rate", "Tier", "Trend"]
            ws.append(headers)
            for col_i, h in enumerate(headers, 1):
                cell = ws.cell(ws.max_row, col_i)
                cell.font = Font(bold=True, color="FFFFFF")
                cell.fill = PatternFill("solid", fgColor=color)
                cell.alignment = Alignment(horizontal="center")

            for leg_i, leg in enumerate(legs, 1):
                ws.append([
                    f"Leg {leg_i}",
                    leg.get("Player Name", leg.get("player_name", "")),
                    leg.get("Team", leg.get("team", "")),
                    leg.get("Opponent", leg.get("opponent", "")),
                    str(leg.get("Stat Norm", leg.get("stat_norm", ""))).upper().replace("_", " "),
                    leg.get("Line Score", leg.get("line_score", "")),
                    leg.get("Recommended Side", leg.get("recommended_side", "")),
                    safe_float(leg.get("Composite Hit Rate", leg.get("composite_hit_rate", 0))),
                    leg.get("Tier", leg.get("tier", "")),
                    leg.get("Trend", leg.get("trend", "")),
                ])
                last = ws.max_row
                side = str(leg.get("Recommended Side", leg.get("recommended_side", "OVER")))
                row_color = "E2EFDA" if side == "OVER" else "FCE4D6"
                for col_i in range(1, 11):
                    ws.cell(last, col_i).fill = PatternFill("solid", fgColor=row_color)

        for col in ws.columns:
            max_len = max((len(str(c.value or "")) for c in col), default=10)
            if hasattr(col[0], "column_letter"):
                ws.column_dimensions[col[0].column_letter].width = min(max_len + 3, 40)

    wb.save(output_path)
    print(f"Tickets saved -> {output_path}")


def main():
    parser = argparse.ArgumentParser()
    parser.add_argument("--input", default="step8_nhl_direction_clean.xlsx")
    parser.add_argument("--output", default="nhl_best_tickets.xlsx")
    parser.add_argument("--min-hit-rate", type=float, default=MIN_HIT_RATE)
    args = parser.parse_args()

    # Map S8 clean-XLSX column names to what S9 expects
    S8_MAP = {"player":"player_name","direction":"recommended_side","composite_hr":"composite_hit_rate","hr_L5":"hit_rate_over_L5","hr_L10":"hit_rate_over_L10","hr_L20":"hit_rate_over_L20","hr_season":"hit_rate_over_season","prop_type":"stat_norm","line":"line_score"}
    def norm(r):
        out = dict(r)
        for k,v in S8_MAP.items():
            if k in out and v not in out:
                out[v] = out[k]
        return out
    rows = [norm(r) for r in read_xlsx(args.input)]
    print(f"Loaded {len(rows)} props from {args.input}")

    valid_legs = [r for r in rows if is_valid_leg(r)]
    print(f"Valid legs (HR≥{args.min_hit_rate}, A/B tier, sample≥5): {len(valid_legs)}")

    if len(valid_legs) < 2:
        print("Not enough valid legs to build tickets. Run earlier steps first or lower --min-hit-rate.")
        return

    tickets = build_tickets(valid_legs)

    total = sum(len(v) for v in tickets.values())
    print(f"\nBuilt {total} tickets:")
    for ttype, tlist in tickets.items():
        print(f"  {ttype}: {len(tlist)} tickets")
        for conf, legs in tlist[:2]:
            print(f"    conf={conf:.4f}")
            for leg in legs:
                print(f"      {format_leg(leg)}")

    write_tickets_xlsx(tickets, args.output)


if __name__ == "__main__":
    main()
