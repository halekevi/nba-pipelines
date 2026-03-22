"""
Step 8 — Direction Context + Final Output (NHL)
Reads Step 7 ranked XLSX and produces a clean CSV + XLSX with human-readable
direction context and correct column mappings from the upstream pipeline.

Column mapping (Step 7 → Step 8 display):
  player_name          → player
  stat_norm            → prop_type  (normalized key)
  stat_type            → prop_display (original PrizePicks label)
  line_score           → line
  recommended_side     → direction
  opponent             → opp
  composite_hit_rate   → composite_hr
  hit_rate_over_L5     → hr_L5
  hit_rate_over_L10    → hr_L10
  hit_rate_over_L20    → hr_L20
  hit_rate_over_season → hr_season
  sample_L5/L10/L20    → sample_L5/L10/L20
  scoring_tier         → scoring_tier
  def_tier             → def_tier
  pp_tier              → pp_tier
  prop_score           → prop_score
  tier                 → tier
  rank                 → rank

Usage:
    py step8_add_direction_context_nhl.py --input step7_nhl_ranked.xlsx \
        --output step8_nhl_direction.csv
"""

import argparse
import csv
import openpyxl
from datetime import date, datetime, timezone
try:
    from tqdm import tqdm as _tqdm
except ImportError:
    import subprocess, sys
    subprocess.check_call([sys.executable, "-m", "pip", "install", "tqdm", "--break-system-packages", "-q"])
    from tqdm import tqdm as _tqdm
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


# ── Helpers ───────────────────────────────────────────────────────────────────

def safe_float(x, default=0.0) -> float:
    try:
        if x is None or (isinstance(x, str) and x.strip() == ""):
            return default
        return float(x)
    except Exception:
        return default


def hr_output(val):
    """Write hit rates as raw 0-1 floats (4 decimals) for XLSX/CSV; empty if missing."""
    if val is None:
        return ""
    if isinstance(val, str) and val.strip() == "":
        return ""
    try:
        return round(float(val), 4)
    except (TypeError, ValueError):
        return ""


def fmt_num(val, digits=2) -> str:
    f = safe_float(val)
    if f == 0.0 and (val is None or str(val).strip() == ""):
        return ""
    return f"{f:.{digits}f}"


# ── Column name resolution (handles upstream naming variants) ─────────────────

# Maps the canonical Step-8 field name → list of candidate column names in Step 7 output
# First match wins.
COLUMN_ALIASES = {
    "player":           ["player_name", "player"],
    "prop_display":     ["stat_type", "prop_display"],
    "prop_type":        ["stat_norm", "prop_type"],
    "line":             ["line_score", "line"],
    "pick_type":        ["pick_type"],
    "direction":        ["recommended_side", "bet_dir", "dir", "direction"],
    "opp":              ["opponent", "opp_team", "opp"],
    "composite_hr":     ["composite_hit_rate", "composite_hr"],
    "hr_L5":            ["hit_rate_over_L5",  "hr_last5",  "hr_L5"],
    "hr_L10":           ["hit_rate_over_L10", "hr_last10", "hr_L10"],
    "hr_L20":           ["hit_rate_over_L20", "hr_last20", "hr_L20"],
    "hr_season":        ["hit_rate_over_season", "hr_season"],
    "sample_L5":        ["sample_L5"],
    "over_L5":          ["over_L5"],
    "under_L5":         ["under_L5"],
    "over_L10":         ["over_L10"],
    "under_L10":        ["under_L10"],
    "sample_L10":       ["sample_L10"],
    "sample_L20":       ["sample_L20"],
    "sample_season":    ["sample_season"],
    "scoring_tier":     ["scoring_tier", "role_tier"],
    "def_tier":         ["def_tier"],
    "pp_tier":          ["pp_tier"],
    "edge":             ["edge"],
    "prop_score":       ["prop_score"],
    "tier":             ["tier"],
    "rank":             ["rank"],
    "team":             ["team"],
    "is_home":          ["is_home"],
    "player_role":      ["player_role"],
    "position_group":   ["position_group"],
    "opp_gaa":          ["opp_gaa"],
    "opp_saa":          ["opp_saa"],
    "opp_pk_pct":       ["opp_pk_pct"],
    "def_rank":         ["def_rank"],
    "avg_L5":           ["avg_L5", "stat_last5_avg"],
    "avg_L10":          ["avg_L10", "stat_last10_avg"],
    "avg_L20":          ["avg_L20"],
    "avg_season":       ["avg_season", "stat_season_avg"],
    "games_played":     ["games_played"],
    "pts_per_game":     ["pts_per_game"],
    "pp_pts_per_game":  ["pp_pts_per_game"],
    "toi_avg_L10":      ["toi_avg_L10"],
    "toi_per_game_api": ["toi_per_game_api"],
    "game_start":       ["game_start"],
    "game_script_mult": ["game_script_mult"],
    "game_script_note": ["game_script_note"],
}


def resolve(row: dict, canonical: str, available_cols: set) -> str:
    """Return the value from `row` for the first matching alias column."""
    candidates = COLUMN_ALIASES.get(canonical, [canonical])
    for c in candidates:
        if c in available_cols and row.get(c) not in (None, ""):
            return str(row[c])
    return ""


def get_last_n_raw(raw: dict, available_cols: set, n: int) -> str:
    """
    Find the last-N raw game value for the prop's stat type.
    Step 4 writes columns like last1_shots_on_goal, last2_goals, etc.
    We scan for any column matching last{n}_<anything> excluding fantasy_score.
    """
    prefix = f"last{n}_"
    for col in sorted(available_cols):
        if col.startswith(prefix) and "fantasy_score" not in col:
            v = raw.get(col)
            if v not in (None, ""):
                return str(v)
    return ""


# ── Build display row ─────────────────────────────────────────────────────────

def build_display_row(raw: dict, available_cols: set) -> dict:
    """Build a clean, fully-populated display row from a raw Step 7 row."""
    def r(key):
        return resolve(raw, key, available_cols)

    direction = r("direction") or "OVER"

    composite_hr = safe_float(r("composite_hr"))
    hr_L5        = safe_float(r("hr_L5"))
    hr_L10       = safe_float(r("hr_L10"))
    hr_L20       = safe_float(r("hr_L20"))
    hr_season    = safe_float(r("hr_season"))

    # Direction-adjust hit rates: if UNDER, flip all rates.
    # Use explicit None checks — 0.0 is a valid hit rate (falsy but meaningful).
    def _has_val(v) -> bool:
        return v is not None and str(v).strip() not in ("", "nan")

    if direction == "UNDER":
        composite_hr_adj = (1 - composite_hr) if _has_val(r("composite_hr")) else ""
        hr_L5_adj        = (1 - hr_L5)        if _has_val(r("hr_L5"))        else ""
        hr_L10_adj       = (1 - hr_L10)       if _has_val(r("hr_L10"))       else ""
        hr_L20_adj       = (1 - hr_L20)       if _has_val(r("hr_L20"))       else ""
        hr_season_adj    = (1 - hr_season)     if _has_val(r("hr_season"))    else ""
    else:
        composite_hr_adj = composite_hr if _has_val(r("composite_hr")) else ""
        hr_L5_adj        = hr_L5        if _has_val(r("hr_L5"))        else ""
        hr_L10_adj       = hr_L10       if _has_val(r("hr_L10"))       else ""
        hr_L20_adj       = hr_L20       if _has_val(r("hr_L20"))       else ""
        hr_season_adj    = hr_season    if _has_val(r("hr_season"))    else ""

    avg_L5     = safe_float(r("avg_L5"))
    avg_L10    = safe_float(r("avg_L10"))
    avg_L20    = safe_float(r("avg_L20"))
    avg_season = safe_float(r("avg_season"))
    line_val   = safe_float(r("line"))

    # Gap: rolling avg vs line (how far above/below)
    gap_L5  = round(avg_L5  - line_val, 3) if avg_L5  and line_val else ""
    gap_L10 = round(avg_L10 - line_val, 3) if avg_L10 and line_val else ""

    # Over/under counts — swap when direction is UNDER so columns reflect bet direction
    raw_over_L5  = safe_float(r("over_L5"))
    raw_under_L5 = safe_float(r("under_L5"))
    raw_over_L10  = safe_float(r("over_L10"))
    raw_under_L10 = safe_float(r("under_L10"))
    s5  = safe_float(r("sample_L5"))
    s10 = safe_float(r("sample_L10"))
    if direction == "UNDER":
        # UNDER bets win when value < line, so "wins" = under count
        win_L5  = int(raw_under_L5) if raw_under_L5 is not None else ""
        win_L10 = int(raw_under_L10) if raw_under_L10 is not None else ""
    else:
        win_L5  = int(raw_over_L5)  if raw_over_L5  is not None else ""
        win_L10 = int(raw_over_L10) if raw_over_L10 is not None else ""
    tot_L5  = int(s5)  if s5  is not None else ""
    tot_L10 = int(s10) if s10 is not None else ""
    hit_L5_display  = f"{win_L5}/{tot_L5}"   if win_L5  != "" and tot_L5  != "" else ""
    hit_L10_display = f"{win_L10}/{tot_L10}" if win_L10 != "" and tot_L10 != "" else ""

    return {
        "rank":             r("rank"),
        "tier":             r("tier"),
        "player":           r("player"),
        "team":             r("team"),
        "opp":              r("opp"),
        "is_home":          r("is_home"),
        "player_role":      r("player_role"),
        "position_group":   r("position_group"),
        "prop_display":     r("prop_display"),
        "prop_type":        r("prop_type"),
        "line":             fmt_num(r("line"), 1),
        "direction":        direction,
        # Hit rates (direction-adjusted) — raw 0-1 floats for step9 / Excel
        "composite_hr":     hr_output(composite_hr_adj),
        "hr_L5":            hr_output(hr_L5_adj),
        "hr_L10":           hr_output(hr_L10_adj),
        "hr_L20":           hr_output(hr_L20_adj),
        "hr_season":        hr_output(hr_season_adj),
        "sample_L5":        r("sample_L5"),
        "hit_L5":           hit_L5_display,
        "sample_L10":       r("sample_L10"),
        "hit_L10":          hit_L10_display,
        "sample_L20":       r("sample_L20"),
        "sample_season":    r("sample_season"),
        # Rolling averages
        "avg_L5":           fmt_num(r("avg_L5"), 2),
        "avg_L10":          fmt_num(r("avg_L10"), 2),
        "avg_L20":          fmt_num(r("avg_L20"), 2),
        "avg_season":       fmt_num(r("avg_season"), 2),
        "gap_vs_line_L5":   gap_L5,
        "gap_vs_line_L10":  gap_L10,
        # Projection: best available rolling avg (used by combined_slate_tickets as "Proj")
        "projection":       fmt_num(r("avg_L5"), 2) or fmt_num(r("avg_L10"), 2) or fmt_num(r("avg_L20"), 2) or "",
        # Raw over/under counts for combined slate L5 Over / L5 Under columns
        "over_L5_raw":      int(raw_over_L5)  if raw_over_L5  is not None else "",
        "under_L5_raw":     int(raw_under_L5) if raw_under_L5 is not None else "",
        "over_L10_raw":     int(raw_over_L10) if raw_over_L10 is not None else "",
        "under_L10_raw":    int(raw_under_L10) if raw_under_L10 is not None else "",
        # Last 3 raw game values (from Step 4 fix)
        "last1_raw":        get_last_n_raw(raw, available_cols, 1),
        "last2_raw":        get_last_n_raw(raw, available_cols, 2),
        "last3_raw":        get_last_n_raw(raw, available_cols, 3),
        "games_played":     r("games_played"),
        # Context
        "scoring_tier":     r("scoring_tier"),
        "pp_tier":          r("pp_tier"),
        "def_tier":         r("def_tier"),
        "def_rank":         r("def_rank"),
        "opp_gaa":          fmt_num(r("opp_gaa"), 3),
        "opp_saa":          fmt_num(r("opp_saa"), 3),
        "opp_pk_pct":       fmt_num(r("opp_pk_pct"), 3),
        "pts_per_game":     fmt_num(r("pts_per_game"), 3),
        "pp_pts_per_game":  fmt_num(r("pp_pts_per_game"), 3),
        "toi_avg_L10":      fmt_num(r("toi_avg_L10"), 2),
        "toi_per_game_api": fmt_num(r("toi_per_game_api"), 2),
        # Model scores
        "prop_score":       fmt_num(r("prop_score"), 5),
        "edge":             fmt_num(r("edge"), 4),
        "pick_type":        r("pick_type"),
        # Game info
        "game_start":       r("game_start"),
        "game_script_mult": (
            fmt_num(raw.get("game_script_mult"), 3)
            if raw.get("game_script_mult") not in (None, "")
            else ""
        ),
        "game_script_note": str(raw.get("game_script_note", "") or ""),
    }


# ── Writers ───────────────────────────────────────────────────────────────────

TIER_FILL = {
    "A": "C6EFCE",
    "B": "FFEB9C",
    "C": "FCE4D6",
    "D": "E0E0E0",
}
HEADER_FILL = "1F4E79"
HEADER_FONT = "FFFFFF"


def write_csv(rows: list[dict], path: str):
    if not rows:
        open(path, "w").close()
        return
    with open(path, "w", newline="", encoding="utf-8") as f:
        w = csv.DictWriter(f, fieldnames=list(rows[0].keys()))
        w.writeheader()
        w.writerows(rows)
    print(f"CSV saved -> {path}")


def _write_sheet(ws, rows: list[dict], headers: list[str]):
    """Write data + formatting to a worksheet."""
    for col_idx, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=col_idx, value=h)
        cell.font = Font(bold=True, color=HEADER_FONT)
        cell.fill = PatternFill("solid", fgColor=HEADER_FILL)
        cell.alignment = Alignment(horizontal="center")

    for row_data in rows:
        ws.append([row_data.get(h, "") for h in headers])
        last_row = ws.max_row
        tier = str(row_data.get("tier", "D")).upper()
        fill_color = TIER_FILL.get(tier, "FFFFFF")
        for col_idx in range(1, len(headers) + 1):
            ws.cell(last_row, col_idx).fill = PatternFill("solid", fgColor=fill_color)

    ws.freeze_panes = "A2"
    for col_idx, h in enumerate(headers, 1):
        max_len = len(str(h))
        for row_data in rows[:500]:
            v = row_data.get(h, "")
            max_len = max(max_len, len(str(v)) if v is not None else 0)
        ws.column_dimensions[get_column_letter(col_idx)].width = min(40, max(10, max_len + 2))


def write_xlsx(rows: list[dict], path: str):
    if not rows:
        openpyxl.Workbook().save(path)
        return

    headers = list(rows[0].keys())
    wb = openpyxl.Workbook()

    ws_all = wb.active
    ws_all.title = "All Props"
    _write_sheet(ws_all, rows, headers)

    skaters = [r for r in rows if str(r.get("player_role", "")).upper() == "SKATER"]
    if skaters:
        _write_sheet(wb.create_sheet("Skaters"), skaters, headers)

    goalies = [r for r in rows if str(r.get("player_role", "")).upper() == "GOALIE"]
    if goalies:
        _write_sheet(wb.create_sheet("Goalies"), goalies, headers)

    a_rows = [r for r in rows if str(r.get("tier", "")).upper() == "A"]
    if a_rows:
        _write_sheet(wb.create_sheet("A-Tier Best"), a_rows, headers)

    wb.save(path)
    print(f"XLSX saved -> {path}")


# ── Main ─────────────────────────────────────────────────────────────────────

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--input",  required=True, help="Step 7 ranked XLSX")
    ap.add_argument("--output", default="step8_nhl_direction.csv")
    args = ap.parse_args()

    wb = openpyxl.load_workbook(args.input, read_only=True, data_only=True)
    if "All Props" in wb.sheetnames:
        ws = wb["All Props"]
    else:
        ws = wb.active
        print(
            f"WARNING: 'All Props' sheet not found, reading active sheet: {ws.title}"
        )
    raw_headers = []
    raw_rows = []
    for r_i, row in enumerate(ws.iter_rows(values_only=True), start=1):
        if r_i == 1:
            raw_headers = [str(x).strip() if x is not None else f"col_{i}"
                           for i, x in enumerate(row)]
            continue
        r = {}
        for j, h in enumerate(raw_headers):
            r[h] = row[j] if j < len(row) else ""
        raw_rows.append(r)
    wb.close()

    available_cols = set(raw_headers)
    print(f"Loaded {len(raw_rows)} props from {args.input}")
    print(f"Upstream columns ({len(available_cols)}): {sorted(available_cols)}\n")

    display_rows = [build_display_row(r, available_cols) for r in _tqdm(raw_rows, desc="  Building display rows", unit="prop")]

    # ── Date filter: keep only today's games ──────────────────────────────────
    today_local = date.today()
    today_str   = today_local.isoformat()  # e.g. "2026-03-12"
    before_filter = len(display_rows)

    def _is_today(gs) -> bool:
        """
        Accept game_start as a datetime object, a timezone-aware string, or a
        plain date string.  Always compare against the *local* calendar date so
        that UTC-midnight NHL games (e.g. 2026-03-13T00:00:00+00:00 = March 12
        ET) are not accidentally dropped.
        """
        if gs is None or gs == "":
            return False
        try:
            # openpyxl may hand back a datetime object directly
            if isinstance(gs, datetime):
                if gs.tzinfo is not None:
                    local_date = gs.astimezone().date()
                else:
                    local_date = gs.date()
                return local_date == today_local
            # String path
            gs_str = str(gs).strip()
            if not gs_str:
                return False
            # Try full ISO parse (handles "+00:00" / "Z" suffixes)
            for fmt in (
                "%Y-%m-%dT%H:%M:%S%z",
                "%Y-%m-%d %H:%M:%S%z",
                "%Y-%m-%dT%H:%M:%S",
                "%Y-%m-%d %H:%M:%S",
            ):
                try:
                    dt = datetime.strptime(gs_str[:25], fmt)
                    if dt.tzinfo is not None:
                        return dt.astimezone().date() == today_local
                    return dt.date() == today_local
                except ValueError:
                    continue
            # Fallback: plain date prefix comparison
            return gs_str[:10] == today_str
        except Exception:
            return False

    display_rows = [r for r in display_rows if _is_today(r.get("game_start", ""))]
    dropped = before_filter - len(display_rows)
    print(f"[DateFilter] Kept {len(display_rows)}/{before_filter} rows for {today_str} (dropped {dropped} future/past rows)")

    display_rows.sort(
        key=lambda r: int(r.get("rank") or 9999)
        if str(r.get("rank", "")).isdigit() else 9999
    )

    if args.output.lower().endswith(".xlsx"):
        write_xlsx(display_rows, args.output)
        csv_out = args.output.replace(".xlsx", ".csv")
        write_csv(display_rows, csv_out)
    else:
        write_csv(display_rows, args.output)
        xlsx_out = args.output.replace(".csv", "_clean.xlsx")
        write_xlsx(display_rows, xlsx_out)

    # ── Console summary ───────────────────────────────────────────────────────
    tier_counts = {}
    for r in display_rows:
        t = str(r.get("tier", "?")).upper()
        tier_counts[t] = tier_counts.get(t, 0) + 1
    print(f"\nTier breakdown: {tier_counts}")

    for tier_label in ("A", "B"):
        tier_rows = [r for r in display_rows if str(r.get("tier", "")).upper() == tier_label]
        if not tier_rows:
            continue
        print(f"\n{'='*65}")
        print(f"{tier_label}-TIER PICKS ({len(tier_rows)} total)")
        print("="*65)
        for r in tier_rows[:25]:
            print(
                f"  [{tier_label}] #{str(r.get('rank','?')):>3}  "
                f"{str(r.get('player','')):25s}  "
                f"{str(r.get('direction','')):5s} {str(r.get('line','')):>5}  "
                f"{str(r.get('prop_display') or r.get('prop_type','')):20s}  "
                f"vs {str(r.get('opp','?')):4s}  |  "
                f"{str(r.get('scoring_tier','')):12s}  {str(r.get('def_tier','')):7s}"
            )
            print(
                f"       Composite HR: {str(r.get('composite_hr','')):>6}  |  "
                f"L5:{str(r.get('hr_L5','')):>6}  "
                f"L10:{str(r.get('hr_L10','')):>6}  "
                f"L20:{str(r.get('hr_L20','')):>6}  "
                f"Szn:{str(r.get('hr_season','')):>6}  |  "
                f"AvgL10:{str(r.get('avg_L10','')):>5}  "
                f"Line:{str(r.get('line','')):>5}  "
                f"Gap:{str(r.get('gap_vs_line_L10','')):>6}  |  "
                f"score={r.get('prop_score','')}"
            )


if __name__ == "__main__":
    main()
